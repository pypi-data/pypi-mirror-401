import os

import pandas as pd
from bclearer_core.constants.standard_constants import (
    DEFAULT_NULL_VALUE,
)
from bclearer_interop_services.excel_services.object_model.excel_workbooks import (
    ExcelWorkbooks,
)
from bclearer_interop_services.file_system_service.objects.files import (
    Files,
)
from bclearer_interop_services.file_system_service.objects.wrappers.path_wrappers import (
    PathWrappers,
)
from bclearer_orchestration_services.reporting_service.reporters.log_with_datetime_latest import (
    log_message,
)
from openpyxl import load_workbook
from pandas import read_excel

try:
    from xlrd import open_workbook
except ImportError:
    # xlrd is optional and may not be installed
    pass


class ExcelFacades:
    def __init__(self, file_path):

        self.file_path = PathWrappers(
            file_path
        )

        self.file_extension = (
            self.file_path.path.suffix
        )

        self.workbook = ExcelWorkbooks(
            self.file_path,
            self.file_extension,
        )

        # Dictionary to track cell edit history
        # Format: {(sheet_name, row, col): [{'value': value, 'timestamp': timestamp}, ...]}]
        self.edit_history = {}

    def read_cell(
        self,
        sheet_name: str,
        row_index: int,
        column_index: int,
    ):
        sheet = self.workbook.sheet(
            sheet_name,
        )

        cell = sheet.cell(
            row_index, column_index
        )

        return cell.value

    def read_sheet_to_dataframe(
        self,
        sheet_name: str,
        header_row_number: int = 1,
    ) -> pd.DataFrame:

        sheet = self.workbook.sheet(
            sheet_name,
        )

        # Convert the sheet rows into a list of lists (representing rows)
        sheet_dataframe = sheet.read_to_dataframe(
            header_row_number=header_row_number
        )

        return sheet_dataframe

    def write_cell(
        self,
        sheet_name,
        row_index,
        column_index,
        value,
    ):
        sheet = self.workbook.sheet(
            sheet_name,
        )
        sheet.cell(
            row_index, column_index
        ).value = value

    # DONE: Implemented edit history feature for cell updates
    def update_cell(
        self,
        sheet_name,
        row_index,
        column_index,
        value,
    ):
        # Get current value before updating
        sheet = self.workbook.sheet(
            sheet_name,
        )
        current_value = sheet.cell(
            row_index, column_index
        ).value

        # Create a key for this cell in the edit history
        cell_key = (
            sheet_name,
            row_index,
            column_index,
        )

        # Add current value to history if this is the first edit or value has changed
        if (
            cell_key
            not in self.edit_history
        ):
            self.edit_history[
                cell_key
            ] = []

        if current_value != value:
            import datetime

            # Record the previous value with timestamp
            self.edit_history[
                cell_key
            ].append(
                {
                    "value": current_value,
                    "timestamp": datetime.datetime.now().isoformat(),
                }
            )

        # Update the cell value
        self.write_cell(
            sheet_name,
            row_index,
            column_index,
            value,
        )

    def save(self, file_path=None):

        if file_path is None:
            file_path = (
                self.workbook.file_path
            )

        directory = os.path.dirname(
            file_path
        )

        if (
            directory
            and not os.path.exists(
                directory
            )
        ):
            os.makedirs(directory)

        self.workbook.save(file_path)

    def convert_to_csv(self, xlsx_path):

        print(
            f"Reading {self.workbook.file_path}"
        )

        # Extract workbook name without extension for use in CSV filenames
        workbook_name = (
            os.path.splitext(
                os.path.basename(
                    xlsx_path
                )
            )[0]
        )

        # DONE: Replaced single sheet conversion with iteration through all sheets
        # Now converts all sheets to individual CSV files with format <workbook_name>_<sheet_name>.csv
        try:
            # Get all sheet names
            excel_data = pd.ExcelFile(
                xlsx_path
            )
            sheet_names = (
                excel_data.sheet_names
            )

            for (
                sheet_name
            ) in sheet_names:
                # Create CSV filename with workbook and sheet names
                csv_filename = f"{workbook_name}_{sheet_name}.csv"
                csv_path = os.path.join(
                    os.path.dirname(
                        xlsx_path
                    ),
                    csv_filename,
                )

                if not os.path.isfile(
                    csv_path
                ):
                    # Read the specific sheet
                    df = pd.read_excel(
                        xlsx_path,
                        sheet_name=sheet_name,
                    )

                    # Export to CSV
                    df.to_csv(
                        csv_path,
                        index=False,
                    )
                    print(
                        f"Converted sheet '{sheet_name}' to {csv_filename}",
                    )
                else:
                    print(
                        f"CSV file {csv_filename} already exists.",
                    )

            print(
                f"All sheets successfully converted to CSV."
            )

        except Exception as e:
            print(
                f"Error converting sheets from {xlsx_path}: {str(e)}"
            )

    def summarise_sheet(
        self, sheet_name=None
    ):
        """
        Generate a summary of a specific sheet or all sheets in the workbook.

        Args:
            sheet_name: Optional name of the sheet to summarize. If None, summarizes all sheets.

        Returns:
            A pandas DataFrame with sheet summary information.
        """
        sheet_summary_df = (
            pd.DataFrame()
        )

        # Handle Excel files
        workbook = self.workbook.wb

        if sheet_name:
            # Summarize only the specified sheet
            sheet_names_list = [
                sheet_name
            ]
        else:
            # Summarize all sheets
            sheet_names_list = (
                self.workbook.wb.sheetnames
            )

        number_of_sheets = len(
            sheet_names_list
        )

        for (
            sheet_name
        ) in sheet_names_list:
            sheet = workbook[sheet_name]

            # Process the sheet
            sheet_data = sheet.values
            sheet_rows = sheet.max_row

            if sheet_rows > 0:
                column_list = []
                for cell in list(
                    sheet.iter_rows()
                )[0]:
                    column_list.append(
                        cell.value
                    )

                sheet_data_df = pd.DataFrame(
                    list(sheet.values),
                    columns=column_list,
                )

                sheet_data_df = sheet_data_df.dropna(
                    how="all",
                ).reset_index(
                    drop=True
                )

                dim = (
                    sheet_data_df.shape
                )
                number_of_rows = dim[0]
                number_of_columns = dim[
                    1
                ]

                excel_sheet_summary = pd.DataFrame(
                    {
                        "number_of_columns": [
                            number_of_columns
                        ],
                        "number_of_rows": [
                            number_of_rows
                        ],
                        "sheet_names": [
                            sheet_name
                        ],
                        "number_of_sheets": [
                            number_of_sheets
                        ],
                    }
                )

                print(
                    f"\nSheet {sheet_name} summary:\n{excel_sheet_summary}\n"
                )
                print(
                    f"Found columns: {column_list}\n"
                )

                sheet_summary_df = pd.concat(
                    [
                        sheet_summary_df,
                        excel_sheet_summary,
                    ]
                )
            else:
                # Handle empty sheets
                empty_summary = pd.DataFrame(
                    {
                        "number_of_columns": [
                            0
                        ],
                        "number_of_rows": [
                            0
                        ],
                        "sheet_names": [
                            sheet_name
                        ],
                        "number_of_sheets": [
                            number_of_sheets
                        ],
                    }
                )
                sheet_summary_df = pd.concat(
                    [
                        sheet_summary_df,
                        empty_summary,
                    ]
                )

        return sheet_summary_df

    @staticmethod
    def summarise_directory(
        directory_path_and_name,
        valid_file_extensions=None,
    ) -> pd.DataFrame:
        """
        Summarize all Excel and CSV files in a directory and its subdirectories.

        Args:
            directory_path_and_name: Path to the directory to scan
            valid_file_extensions: List of file extensions to include, defaults to ['.xlsx', '.xlsm', '.csv']

        Returns:
            A pandas DataFrame with summary information for all sheets in all files
        """
        if (
            valid_file_extensions
            is None
        ):
            valid_file_extensions = [
                ".xlsx",
                ".xlsm",
                ".csv",
            ]

        print(
            f"\n------Reading directory {directory_path_and_name}---------\n"
        )

        sheet_summary_report_schema = [
            "parent_directory_paths",
            "file_names",
            "number_of_sheets",
            "sheet_names",
            "number_of_columns",
            "number_of_rows",
        ]

        sheet_summary_report_dataframe = pd.DataFrame(
            columns=sheet_summary_report_schema
        )

        for (
            parent_directory_path,
            _,
            files,
        ) in os.walk(
            directory_path_and_name
        ):
            for file_name in files:
                _, file_extension = (
                    os.path.splitext(
                        file_name
                    )
                )

                # Skip CSV files as they're handled by delimited_text.csv_summarizer
                if (
                    file_extension.lower()
                    in valid_file_extensions
                    and file_extension.lower()
                    != ".csv"
                ):
                    print(
                        f"*********Summarising file {file_name} in {parent_directory_path}**********\n"
                    )

                    file_path = os.path.join(
                        parent_directory_path,
                        file_name,
                    )
                    try:
                        excel_facade = ExcelFacades(
                            file_path
                        )
                        sheet_summary = (
                            excel_facade.summarise_sheet()
                        )

                        # Add file information
                        sheet_summary[
                            "file_names"
                        ] = file_name
                        sheet_summary[
                            "parent_directory_paths"
                        ] = parent_directory_path

                        # Concatenate with master dataframe
                        sheet_summary_report_dataframe = pd.concat(
                            [
                                sheet_summary_report_dataframe,
                                sheet_summary,
                            ]
                        )
                    except (
                        Exception
                    ) as e:
                        print(
                            f"Error processing file {file_name}: {str(e)}"
                        )

        return sheet_summary_report_dataframe

    @staticmethod
    def extract_dataframe_from_excel_sheet(
        excel_file_path_and_name: str,
        excel_sheet_name: str,
    ):
        """
        Utility function to extract a dataframe from an Excel sheet.

        This function creates an ExcelFacades instance and uses it to read a sheet into a DataFrame.

        Args:
            excel_file_path_and_name: Path to the Excel file
            excel_sheet_name: Name of the sheet to extract

        Returns:
            pandas.DataFrame containing the sheet data
        """
        try:
            excel_facade = ExcelFacades(
                excel_file_path_and_name,
            )
            print(
                f"Successfully initialized ExcelFacade with file: {excel_file_path_and_name}",
            )

            excel_sheet_dataframe = excel_facade.read_sheet_to_dataframe(
                sheet_name=excel_sheet_name
            )

            return excel_sheet_dataframe

        except Exception as e:
            raise Exception(
                f"error reading sheet: {excel_sheet_name} in file file {excel_file_path_and_name} : {e}"
            )

    @staticmethod
    def convert_xlxs_to_dataframe_dictionary(
        xlsx_file: Files,
    ) -> dict:
        """
        Convert an Excel file to a dictionary of dataframes (one per sheet).

        Args:
            xlsx_file: A Files object representing the Excel file

        Returns:
            Dictionary mapping sheet names to pandas DataFrames
        """
        dataframe_dictionary = read_excel(
            xlsx_file.absolute_path_string,
            sheet_name=None,
            engine="openpyxl",
        )

        return dataframe_dictionary

    @staticmethod
    def convert_sheet_with_header_to_dataframe(
        file_name: str,
        sheet_name: str,
    ):
        """
        Convert an Excel sheet with a header row to a DataFrame.

        Args:
            file_name: Path to the Excel file
            sheet_name: Name of the sheet to convert

        Returns:
            pandas.DataFrame containing the sheet data, or DEFAULT_NULL_VALUE if the operation fails
        """
        try:
            dataframe = read_excel(
                io=file_name,
                dtype=object,
                sheet_name=sheet_name,
                header=0,
            )
            return dataframe
        except Exception as read_fail:
            log_message(
                message="Was not able to read from "
                + file_name
                + " because "
                + str(read_fail),
            )

            return DEFAULT_NULL_VALUE
