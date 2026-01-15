import uuid
from typing import Dict, List

import pandas as pd
from bclearer_interop_services.excel_services.object_model.excel_cell_coordinates import (
    CellCoordinates,
)
from bclearer_interop_services.excel_services.object_model.excel_cell_ranges import (
    Ranges,
)
from bclearer_interop_services.excel_services.object_model.excel_cells import (
    ExcelCells,
)
from bclearer_interop_services.excel_services.object_model.excel_columns import (
    ExcelColumns,
)
from bclearer_interop_services.excel_services.object_model.excel_objects import (
    ExcelObjects,
)
from bclearer_interop_services.excel_services.object_model.excel_rows import (
    ExcelRows,
)
from openpyxl.worksheet.worksheet import (
    Worksheet as OpenpyxlWorksheet,
)


class ExcelSheets(ExcelObjects):
    def __init__(
        self, sheet: OpenpyxlWorksheet
    ):
        self.cells: List[ExcelCells] = (
            []
        )
        self.rows: Dict[
            int, ExcelRows
        ] = {}
        self.columns: Dict[
            int, ExcelColumns
        ] = {}

        self.uuid = uuid.uuid4()

        self.sheet = sheet

        self.read_sheet_cells()

    def add_rows(self):
        """Adds all rows in the sheet that have been loaded into the Sheets object."""
        for row_index, row in enumerate(
            self.sheet.iter_rows(),
            start=1,
        ):
            if (
                row_index
                not in self.rows
            ):
                self.rows[row_index] = (
                    ExcelRows(
                        self.sheet,
                        row_index,
                    )
                )

    def add_columns(self):
        """Adds all columns in the sheet that have been loaded into the Sheets object."""
        for (
            row
        ) in self.sheet.iter_rows():
            for (
                column_index,
                _,
            ) in enumerate(
                row, start=1
            ):
                if (
                    column_index
                    not in self.columns
                ):
                    self.columns[
                        column_index
                    ] = ExcelColumns(
                        self.sheet,
                        column_index,
                    )

    def add_cells(self):
        """
        Reads all non-empty cells in the sheet into a list of Cells objects
        and also wraps them in a CellCoordinates object containing the Row and Column objects.
        """

        non_empty_cells = []
        for row_index, row in enumerate(
            self.sheet.iter_rows(),
            start=1,
        ):
            for (
                column_index,
                cell,
            ) in enumerate(
                row, start=1
            ):
                if (
                    cell.value
                    is not None
                ):
                    cell_coordinates = CellCoordinates(
                        self.rows[
                            row_index
                        ],
                        self.columns[
                            column_index
                        ],
                    )
                    non_empty_cells.append(
                        ExcelCells(
                            cell,
                            cell_coordinates,
                        )
                    )

        self.cells = non_empty_cells

    def read_sheet_cells(self):
        """
        High-level method to first add rows and columns, then populate non-empty cells.
        """
        self.add_rows()
        self.add_columns()
        self.add_cells()

    def cell(
        self, row: int, column: int
    ):
        # Try to find the cell in self.cells
        for cell in self.cells:
            if (
                cell.cell_coordinate.row.index
                == row
                and cell.cell_coordinate.column.index
                == column
            ):
                return cell
        # If not found, create new row and column objects if they don't exist
        if row not in self.rows:
            self.rows[row] = ExcelRows(
                self.sheet, row
            )
        if column not in self.columns:
            self.columns[column] = (
                ExcelColumns(
                    self.sheet, column
                )
            )
        # Get the Openpyxl cell object
        cell_obj = self.sheet.cell(
            row=row, column=column
        )
        # Create cell coordinates
        cell_coordinates = (
            CellCoordinates(
                self.rows[row],
                self.columns[column],
            )
        )
        # Create a new ExcelCells object
        excel_cell = ExcelCells(
            cell_obj, cell_coordinates
        )
        return excel_cell

    def row(self, index: int):

        return ExcelRows(
            self.sheet, index
        )

    def column(self, index: int):

        return ExcelColumns(
            self.sheet,
            index,
        )

    def range(
        self,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ):
        return Ranges(
            self.sheet,
            min_row,
            min_col,
            max_row,
            max_col,
        )

    def get_merged_ranges(
        self,
    ) -> Dict[str, Ranges]:
        """
        Returns a dictionary of all merged cell ranges in the sheet.
        The keys are the string representations of the ranges (e.g., "A1:B2").
        The values are instances of the Ranges class.
        """
        merged_ranges = {}
        for (
            merged_range
        ) in (
            self.sheet.merged_cells.ranges
        ):
            # Extract the coordinates of the merged range
            min_row = (
                merged_range.min_row
            )
            min_col = (
                merged_range.min_col
            )
            max_row = (
                merged_range.max_row
            )
            max_col = (
                merged_range.max_col
            )

            # Get the value from the top-left cell of the merged range
            cell_value = self.cell(
                row=min_row,
                column=min_col,
            ).value

            # Create a Ranges object for each merged range
            ranges_obj = Ranges(
                self.sheet,
                min_row,
                min_col,
                max_row,
                max_col,
                cell_value,
            )

            # Use the string representation of the range as the dictionary key
            merged_ranges[
                str(merged_range)
            ] = ranges_obj

        return merged_ranges

    def read_to_dataframe(
        self,
        header_row_number: int = 1,
    ) -> pd.DataFrame:
        # Create a dictionary to store cell values by their coordinates
        data_dict = {}

        data = self.__convert_cells_list_to_data_grid(
            data_dict
        )

        if not data:
            return (
                pd.DataFrame()
            )  # Return an empty DataFrame if there is no data

        # Convert to 0-based index for internal processing
        header_index = (
            header_row_number - 1
        )

        # Check for potential empty or merged rows
        while header_index < len(
            data,
        ) and not any(
            data[header_index],
        ):
            header_index += (
                1  # Skip empty rows
            )

        # Ensure we are within bounds
        if header_index >= len(data):
            raise ValueError(
                f"header_row_number {header_row_number} (adjusted to {header_index + 1}) is out of range.",
            )

        # Set headers from the identified row
        headers = data[header_index]

        # Exclude rows above the identified header row
        data = data[header_index + 1 :]

        # Create the DataFrame with the identified headers and data below
        return pd.DataFrame(
            data,
            columns=headers,
        )

    def __convert_cells_list_to_data_grid(
        self, data_dict
    ):
        for cell in self.cells:
            row_idx = (
                cell.cell_coordinate.row.index
            )
            col_idx = (
                cell.cell_coordinate.column.index
            )
            if row_idx not in data_dict:
                data_dict[row_idx] = {}
            data_dict[row_idx][
                col_idx
            ] = cell.value
        # Find the maximum row and column indices
        max_row_idx = max(
            data_dict.keys()
        )
        max_col_idx = max(
            max(row.keys())
            for row in data_dict.values()
        )
        # Create a list of lists to represent the DataFrame data
        data = [
            [
                data_dict.get(
                    row_idx, {}
                ).get(col_idx, None)
                for col_idx in range(
                    1, max_col_idx + 1
                )
            ]
            for row_idx in range(
                1, max_row_idx + 1
            )
        ]
        return data

    def save_dataframe(
        self,
        table: pd.DataFrame,
        full_filename: str,
        sheet_name: str = None,
    ):
        """
        Save a pandas DataFrame to an Excel file.

        Args:
            table: The pandas DataFrame to save
            full_filename: The path and filename to save to
            sheet_name: The name of the sheet (defaults to current sheet name if not provided)
        """
        if sheet_name is None:
            sheet_name = (
                self.sheet.title
            )

        writer = pd.ExcelWriter(
            path=full_filename,
            engine="xlsxwriter",
        )

        table.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
        )

        writer.close()


# DONE: Moved save_table_in_excel to ExcelSheets class as save_dataframe method
# Previous standalone function integrated as a class method for better organization
