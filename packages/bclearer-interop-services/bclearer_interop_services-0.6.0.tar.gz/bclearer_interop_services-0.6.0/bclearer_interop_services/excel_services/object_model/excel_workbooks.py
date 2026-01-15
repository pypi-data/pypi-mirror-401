from pathlib import Path
from typing import Dict

import pandas as pd
from bclearer_interop_services.excel_services.object_model.excel_objects import (
    ExcelObjects,
)
from bclearer_interop_services.excel_services.object_model.excel_sheets import (
    ExcelSheets,
)
from bclearer_interop_services.file_system_service.objects.wrappers.path_wrappers import (
    PathWrappers,
)
from openpyxl import (
    Workbook as OpenpyxlWorkbook,
)
from openpyxl import load_workbook


class ExcelWorkbooks(ExcelObjects):
    def __init__(
        self,
        file_path: (
            str | PathWrappers
        ) = None,
        file_extension: str = ".xlsx",
    ):
        if isinstance(file_path, str):

            self.file_path = (
                PathWrappers(file_path)
            )
        else:
            self.file_path = file_path

        self.sheets: Dict[
            str, ExcelSheets
        ] = {}

        self._set_excel_file_type(
            file_extension
        )

        self._load_excel_workbook()

    def sheet(
        self, sheet_name: str
    ) -> ExcelSheets:

        if sheet_name in self.sheets:
            return self.sheets[
                sheet_name
            ]
        raise ValueError(
            f"Sheet {sheet_name} does not exist",
        )

    def create_sheet(
        self, title: str
    ) -> ExcelSheets:
        if title in self.sheets:
            raise ValueError(
                f"Sheet {title} already exists",
            )

        openpyxl_sheet = (
            self.wb.create_sheet(
                title=title,
            )
        )
        sheet = ExcelSheets(
            openpyxl_sheet
        )
        self.sheets[title] = sheet

        return sheet

    def remove_sheet(
        self,
        sheet_name: str,
    ):
        if (
            sheet_name
            not in self.sheets
        ):
            raise ValueError(
                f"Sheet {sheet_name} does not exist",
            )

        if (
            sheet_name
            in self.wb.sheetnames
        ):
            del self.wb[sheet_name]
        del self.sheets[sheet_name]

    def _load_excel_workbook(self):
        if self.file_path.exists():
            if (
                self.excel_file_type
                == ".xlsx"
            ):
                self._load_xlsx(
                    self.file_path,
                )
            elif (
                self.excel_file_type
                == ".xls"
            ):
                self._load_xls(
                    self.file_path,
                )
            else:
                raise ValueError(
                    f"Unsupported file extension: {self.excel_file_type}",
                )
        else:
            self.wb = OpenpyxlWorkbook()

            self.sheets = {
                sheet.title: ExcelSheets(
                    sheet,
                )
                for sheet in self.wb.worksheets
            }

    def _set_excel_file_type(
        self, file_extension
    ):
        if self.file_path.exists():
            self.excel_file_type = (
                self.file_path.path.suffix.lower()
            )
        else:
            self.excel_file_type = (
                file_extension
            )

    def _load_xlsx(
        self, file_path: PathWrappers
    ):

        self.wb = load_workbook(
            file_path.path_string,
        )

        for sheet in self.wb.worksheets:

            self._remove_empty_rows(
                sheet,
            )

            self.sheets[sheet.title] = (
                ExcelSheets(sheet)
            )

    def _load_xls(
        self, file_path: PathWrappers
    ):

        xls = pd.ExcelFile(
            file_path.path_string,
            engine="xlrd",
        )

        for (
            sheet_name
        ) in xls.sheet_names:
            df = xls.parse(
                sheet_name,
                header=0,
            )
            self.sheets[sheet_name] = (
                self._convert_df_to_sheet(
                    df,
                    sheet_name,
                )
            )

    def _convert_df_to_sheet(
        self,
        dataframe,
        sheet_name,
    ):
        """Converts a pandas DataFrame to the internal Sheets object for .xls files."""
        openpyxl_sheet = OpenpyxlWorkbook().create_sheet(
            sheet_name,
        )

        openpyxl_sheet.append(
            dataframe.columns.tolist(),
        )

        for (
            r_idx,
            row,
        ) in dataframe.iterrows():
            openpyxl_sheet.append(
                row.tolist(),
            )
        return ExcelSheets(
            openpyxl_sheet
        )

    def _remove_empty_rows(self, sheet):
        """Removes empty rows where all cells are None from the openpyxl Worksheet object."""
        # Iterate over rows in reverse to avoid index shifting issues
        for row in sheet.iter_rows(
            values_only=True,
        ):
            # If the row is completely empty (all None values), remove the row
            if all(
                cell is None
                for cell in row
            ):
                # Get the index of the row to remove
                row_index = (
                    sheet.max_row
                )
                # Delete the row from the sheet
                sheet.delete_rows(
                    row_index,
                    1,
                )

    def save(
        self,
        file_path: str = None,
    ):
        if file_path is None:
            file_path = self.file_path

        if (
            self.excel_file_type
            == ".xlsx"
        ):
            self._save_xlsx(file_path)

        elif (
            self.excel_file_type
            == ".xls"
        ):
            self._save_xls(file_path)
        else:
            raise ValueError(
                f"Unsupported file extension: {self.excel_file_type}",
            )

    def _save_xlsx(self, file_path):
        self.wb.save(file_path)

    def _save_xls(self, file_path):
        with pd.ExcelWriter(
            file_path,
            engine="xlwt",
        ) as writer:
            for (
                sheet_name,
                sheet,
            ) in self.sheets.items():
                dataframe = pd.DataFrame(
                    [
                        [
                            cell.value
                            for cell in row
                        ]
                        for row in sheet.sheet.rows
                    ],
                )
                dataframe.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
