from typing import Literal

from bclearer_interop_services.excel_services.object_model.excel_cells import (
    ExcelCells,
)
from openpyxl.worksheet.worksheet import (
    Worksheet as OpenpyxlWorksheet,
)


class Ranges:
    def __init__(
        self,
        sheet: OpenpyxlWorksheet,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
        range_cell_value: str = None,
    ):
        self.sheet = sheet
        self.min_row = min_row
        self.min_col = min_col
        self.max_row = max_row
        self.max_col = max_col
        self._value = range_cell_value

    @property
    def value(self):
        return self._value

    @value.setter
    def value(self, value):

        self._value = value

    def __iter__(self):
        for row in self.sheet.iter_rows(
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_col,
            max_col=self.max_col,
            values_only=True,
        ):
            yield [
                ExcelCells(cell)
                for cell in row
            ]
