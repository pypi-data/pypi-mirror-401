"""Excel MCP Server with Pydantic-annotated tool parameters."""

import json
from typing import Annotated

from fastmcp import FastMCP
from pydantic import Field
from .models import ManagedWorkbook
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

from .formatting import format_sheet_table, format_range_table
from .errors import ToolError, ToolSuccess
from .helpers import (
    parse_range,
    get_sheet,
    translate_formula,
    extract_cell_data,
    get_cell_results,
    has_formula_errors,
    ensure_calculated,
    set_range_values,
    apply_style_to_range,
    add_conditional_formatting_rule_impl,
    add_data_validation_rule_impl,
)
from .sessions import sessions, with_workbook, with_workbook_mutation
from .schemas import (
    RenameSheetOperation,
    CreateSheetOperation,
    DeleteSheetOperation,
    DisplaySheetOperation,
    InsertRowsOperation,
    InsertColumnsOperation,
    DeleteRowsOperation,
    DeleteColumnsOperation,
    GetRangeDataOperation,
    SetRangeDataOperation,
    SetRangeOperation,
    AutoFillOperation,
    SetRangeStyleOperation,
    BatchStyleOperation,
    MergeCellsOperation,
    UnmergeCellsOperation,
    SearchCellsOperation,
    AddConditionalFormattingOperation,
    DeleteConditionalFormattingOperation,
    GetConditionalFormattingOperation,
    AddDataValidationOperation,
    DeleteDataValidationOperation,
    GetDataValidationOperation,
    RequestNewToolOperation,
)

mcp = FastMCP("Excel Manager")

# Authentication is handled via FastMCP Context in the @with_session and @with_workbook decorators.
# The X-User-ID header is extracted from the HTTP request to identify the session.


# --- Session Initialization (NOT exposed as MCP tool) ---
def init_session(session_id: str, filepath: str) -> str:
    """
    Initialize a session with a unique ID and load a workbook.

    This should be called before any other operations to set up the session context.

    Args:
        session_id: Unique identifier for this session (e.g., 'agent_1').
        filepath: Path to the Excel file to load.

    Returns:
        Success message with session ID and loaded file info.
    """

    state = sessions.get_session(session_id)
    state.workbook = ManagedWorkbook(
        wb=openpyxl.load_workbook(filepath),
        filepath=filepath,
        cached_values_wb=openpyxl.load_workbook(filepath, data_only=True),
    )
    return f"Initialized session '{session_id}' with workbook '{filepath}'"


# =============================================================================
# Workbook Management Tools
# =============================================================================


@mcp.tool()
@with_workbook
def get_workbook_info(mwb: ManagedWorkbook) -> str:
    """Get workbook metadata including sheet names and active sheet."""
    try:
        info = {
            "total_sheets": len(mwb.wb.sheetnames),
            "sheet_names": mwb.wb.sheetnames,
            "active_sheet": mwb.wb.active.title,
        }
        return json.dumps(info)
    except Exception as e:
        return ToolError(f"Error getting workbook info: {e}").to_json()


# =============================================================================
# Range Data Tools
# =============================================================================


@mcp.tool()
@with_workbook
async def get_range_data(
    mwb: ManagedWorkbook,
    operation: Annotated[
        GetRangeDataOperation, Field(description="Get range data operation parameters")
    ],
) -> str:
    """
    Get cell data for a range. Returns a formatted table by default.

    Note: Keep range size under 200 cells for best performance.
    """
    try:
        # Ensure formulas are calculated if workbook was modified
        await ensure_calculated(mwb)

        ranges = (
            operation.range_a1
            if isinstance(operation.range_a1, list)
            else operation.range_a1.split(",")
        )
        results = []

        for r in ranges:
            r = r.strip()
            sheet, cell_range = parse_range(mwb.wb, r)

            # Get corresponding cached sheet if available
            cached_sheet = None
            if mwb.cached_values_wb is not None:
                try:
                    cached_sheet = mwb.cached_values_wb[sheet.title]
                except KeyError:
                    pass

            if ":" in cell_range:
                cells = sheet[cell_range]
                cached_cells = cached_sheet[cell_range] if cached_sheet else None
                data = []
                for row_idx, row in enumerate(cells):
                    row_data = []
                    for col_idx, cell in enumerate(row):
                        cached_cell = (
                            cached_cells[row_idx][col_idx] if cached_cells else None
                        )
                        row_data.append(extract_cell_data(cell, cached_cell))
                    data.append(row_data)
            else:
                cell = sheet[cell_range]
                cached_cell = cached_sheet[cell_range] if cached_sheet else None
                data = [[extract_cell_data(cell, cached_cell)]]

            results.append({"range": r, "data": data})

        if operation.return_style_info:
            json_results = []
            for result in results:
                json_data = []
                for row in result["data"]:
                    for cell in row:
                        json_data.append(
                            {
                                "cell": cell["cell"],
                                "v": cell["value"],
                                "c": cell["calculated"],
                                "s": {
                                    "bg": cell["bg"],
                                    "bold": cell["bold"],
                                    "italic": cell["italic"],
                                },
                            }
                        )
                json_results.append({"range": result["range"], "data": json_data})
            return json.dumps(json_results)

        # Return Table as String
        output_parts = []
        for result in results:
            table_str = format_range_table(
                result["data"], result["range"], width=operation.width, max_rows=operation.max_rows
            )
            output_parts.append(table_str)

        return "\n".join(output_parts)
    except Exception as e:
        return ToolError(f"Error getting range data: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
async def set_range_data(
    mwb: ManagedWorkbook,
    operation: Annotated[
        SetRangeDataOperation, Field(description="Set range data operation parameters")
    ],
) -> str:
    """
    Set value(s) for cells in a range.

    Examples:
    - set_range_data({"range_a1": "A1", "value": "Hello"}) - single cell
    - set_range_data({"range_a1": "A1:C1", "value": ["X", "Y", "Z"]}) - row
    - set_range_data({"range_a1": "A1:A3", "value": [1, 2, 3]}) - column
    - set_range_data({"range_a1": "A1:B2", "value": [[1, 2], [3, 4]]}) - grid
    - set_range_data({"range_a1": "C1", "value": "=A1+B1"}) - formula

    Returns JSON with status and cells set. Formulas include 'calculated' value.
    """
    try:
        cells_set = set_range_values(mwb.wb, operation.range_a1, operation.value)
        cell_results = await get_cell_results(mwb, cells_set)
        if has_formula_errors(cell_results):
            return ToolError("Formula calculation error", cells=cell_results).to_json()
        return ToolSuccess(cells=cell_results).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


@mcp.tool()
@with_workbook_mutation
async def batch_set_range_data(
    mwb: ManagedWorkbook,
    operations: Annotated[
        list[SetRangeOperation], Field(description="List of range/value pairs to set")
    ],
) -> str:
    """
    Set multiple ranges in a single call. More efficient than multiple set_range_data calls.

    Example:
    batch_set_range_data([
        {"range_a1": "A1", "value": "Header"},
        {"range_a1": "B1:D1", "value": ["X", "Y", "Z"]},
        {"range_a1": "A2:A4", "value": [1, 2, 3]}
    ])

    Returns JSON with status and cells set. Formulas include 'calculated' value.
    """
    try:
        cells_set = []
        for op in operations:
            cells_set.extend(set_range_values(mwb.wb, op.range_a1, op.value))

        cell_results = await get_cell_results(mwb, cells_set)
        if has_formula_errors(cell_results):
            return ToolError("Formula calculation error", cells=cell_results).to_json()
        return ToolSuccess(cells=cell_results).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


# =============================================================================
# Auto Fill Tool
# =============================================================================


@mcp.tool()
@with_workbook_mutation
async def auto_fill(
    mwb: ManagedWorkbook,
    operation: Annotated[
        AutoFillOperation, Field(description="Auto fill operation parameters")
    ],
) -> str:
    """
    Auto fill data from source range to target range, adjusting formula references.

    Formulas are translated so relative references adjust (like Excel's fill handle).
    For example, filling =A1+B1 down creates =A2+B2, =A3+B3, etc.
    Absolute references ($A$1) are preserved.

    Examples:
        auto_fill({"source_range": "C1", "target_range": "C1:C3"})
        auto_fill({"source_range": "A1:A2", "target_range": "A1:A10"})
        auto_fill({"source_range": "A1:B1", "target_range": "A1:F1"})

    Returns JSON with status and cells filled. Formulas include 'calculated' value.
    """
    try:
        sheet, src_cells = parse_range(mwb.wb, operation.source_range)
        _, tgt_cells = parse_range(mwb.wb, operation.target_range)

        # Parse range boundaries with better error reporting
        try:
            min_col_src, min_row_src, max_col_src, max_row_src = range_boundaries(
                src_cells
            )
        except Exception as e:
            raise ValueError(
                f"Invalid source range: {operation.source_range!r} -> {src_cells!r}: {e}"
            )
        try:
            min_col_tgt, min_row_tgt, max_col_tgt, max_row_tgt = range_boundaries(
                tgt_cells
            )
        except Exception as e:
            raise ValueError(
                f"Invalid target range: {operation.target_range!r} -> {tgt_cells!r}: {e}"
            )

        # Collect source values with their positions
        src_values = []
        for r in range(min_row_src, max_row_src + 1):
            row_vals = []
            for c in range(min_col_src, max_col_src + 1):
                row_vals.append(sheet.cell(row=r, column=c).value)
            src_values.append(row_vals)

        # Check if source cells have content - fail early if empty
        # This handles race conditions when auto_fill runs before set_range_data completes
        all_empty = all(val is None or val == "" for row in src_values for val in row)
        if all_empty:
            return ToolError(
                f"Source range {operation.source_range} is empty. "
                "Ensure data is written to source cells before calling auto_fill. "
                "If you just called set_range_data, wait for it to complete first."
            ).to_json()

        src_rows = len(src_values)
        src_cols = len(src_values[0]) if src_values else 0
        cells_set = []  # Track (cell_ref, value) tuples for all cells

        # Fill target range by repeating source pattern with formula translation
        for r in range(min_row_tgt, max_row_tgt + 1):
            for c in range(min_col_tgt, max_col_tgt + 1):
                # Calculate which source cell to use (cycling through pattern)
                src_row_idx = (r - min_row_tgt) % src_rows
                src_col_idx = (c - min_col_tgt) % src_cols

                value = src_values[src_row_idx][src_col_idx]

                # If it's a formula, translate the references
                if isinstance(value, str) and value.startswith("="):
                    # Calculate offset from the source cell position to target cell position
                    src_row = min_row_src + src_row_idx
                    src_col = min_col_src + src_col_idx
                    row_offset = r - src_row
                    col_offset = c - src_col

                    value = translate_formula(value, row_offset, col_offset)

                cell_ref = f"{get_column_letter(c)}{r}"
                sheet.cell(row=r, column=c).value = value
                cells_set.append((cell_ref, value))

        # Get results for all cells (calculates formulas if any exist)
        cell_results = await get_cell_results(mwb, cells_set)
        if has_formula_errors(cell_results):
            return ToolError("Formula calculation error", cells=cell_results).to_json()
        return ToolSuccess(cells=cell_results).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


# =============================================================================
# Sheet Reading Tools
# =============================================================================


@mcp.tool()
@with_workbook
async def display_sheet(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DisplaySheetOperation | None,
        Field(description="Display sheet operation parameters"),
    ] = None,
) -> str:
    """Read sheet contents as a formatted table with column headers and row numbers."""
    try:
        # Use defaults if no operation provided
        if operation is None:
            operation = DisplaySheetOperation()

        # Ensure formulas are calculated if workbook was modified
        await ensure_calculated(mwb)

        sheet = get_sheet(mwb.wb, operation.sheet_name)

        # Get cached values sheet if available
        cached_sheet = None
        if mwb.cached_values_wb is not None:
            try:
                cached_sheet = mwb.cached_values_wb[sheet.title]
            except KeyError:
                pass

        return format_sheet_table(
            sheet=sheet,
            cached_sheet=cached_sheet,
            max_rows=operation.max_rows,
            total_rows=sheet.max_row,
            title=sheet.title,
            width=operation.width,
        )
    except Exception as e:
        return ToolError(f"Error reading sheet: {e}").to_json()


# =============================================================================
# Sheet Management Tools
# =============================================================================


@mcp.tool()
@with_workbook
def get_sheets(mwb: ManagedWorkbook) -> str:
    """Get all sheet names in the workbook. Returns JSON array of sheet names."""
    return json.dumps(mwb.wb.sheetnames)


@mcp.tool()
@with_workbook_mutation
def create_sheet(
    mwb: ManagedWorkbook,
    operation: Annotated[
        CreateSheetOperation, Field(description="Create sheet operation parameters")
    ],
) -> str:
    """Create one or more new worksheets."""
    try:
        for name in operation.sheet_names:
            mwb.wb.create_sheet(name)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error creating sheets: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def delete_sheet(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DeleteSheetOperation, Field(description="Delete sheet operation parameters")
    ],
) -> str:
    """Delete one or more worksheets by name."""
    try:
        for name in operation.sheet_names:
            if name in mwb.wb.sheetnames:
                del mwb.wb[name]
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error deleting sheets: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def rename_sheet(
    mwb: ManagedWorkbook,
    operation: Annotated[
        RenameSheetOperation, Field(description="Rename sheet operation parameters")
    ],
) -> str:
    """Rename a worksheet."""
    try:
        if operation.old_name in mwb.wb.sheetnames:
            mwb.wb[operation.old_name].title = operation.new_name
            return ToolSuccess().to_json()
        return ToolError(
            f"Sheet '{operation.old_name}' not found", code="SHEET_NOT_FOUND"
        ).to_json()
    except Exception as e:
        return ToolError(f"Error renaming sheet: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def batch_rename_sheets(
    mwb: ManagedWorkbook,
    operations: Annotated[
        list[RenameSheetOperation], Field(description="List of rename operations")
    ],
) -> str:
    """Rename multiple sheets in a single call."""
    try:
        for op in operations:
            if op.old_name in mwb.wb.sheetnames:
                mwb.wb[op.old_name].title = op.new_name
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error renaming sheets: {e}").to_json()


# =============================================================================
# Row/Column Tools
# =============================================================================


@mcp.tool()
@with_workbook_mutation
def insert_rows(
    mwb: ManagedWorkbook,
    operation: Annotated[
        InsertRowsOperation, Field(description="Insert row operation parameters")
    ],
) -> str:
    """Insert empty rows at the specified position."""
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        sheet.insert_rows(operation.position, amount=operation.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error inserting rows: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def insert_columns(
    mwb: ManagedWorkbook,
    operation: Annotated[
        InsertColumnsOperation, Field(description="Insert column operation parameters")
    ],
) -> str:
    """Insert empty columns at the specified position."""
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        sheet.insert_cols(operation.position, amount=operation.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error inserting columns: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def delete_rows(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DeleteRowsOperation, Field(description="Delete row operation parameters")
    ],
) -> str:
    """Delete rows starting at the specified position."""
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        sheet.delete_rows(operation.position, amount=operation.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error deleting rows: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def delete_columns(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DeleteColumnsOperation, Field(description="Delete column operation parameters")
    ],
) -> str:
    """Delete columns starting at the specified position."""
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        sheet.delete_cols(operation.position, amount=operation.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error deleting columns: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def batch_insert_rows(
    mwb: ManagedWorkbook,
    operations: Annotated[
        list[InsertRowsOperation], Field(description="List of insert row operations")
    ],
) -> str:
    """Insert rows at multiple positions in a single call."""
    try:
        for op in operations:
            sheet = get_sheet(mwb.wb, op.sheet_name)
            sheet.insert_rows(op.position, amount=op.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error inserting rows: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def batch_delete_rows(
    mwb: ManagedWorkbook,
    operations: Annotated[
        list[DeleteRowsOperation], Field(description="List of delete row operations")
    ],
) -> str:
    """Delete rows at multiple positions in a single call."""
    try:
        # Sort by position descending to avoid index shift issues
        sorted_ops = sorted(operations, key=lambda x: x.position, reverse=True)
        for op in sorted_ops:
            sheet = get_sheet(mwb.wb, op.sheet_name)
            sheet.delete_rows(op.position, amount=op.count)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error deleting rows: {e}").to_json()


# =============================================================================
# Style Tools
# =============================================================================


@mcp.tool()
@with_workbook_mutation
def set_range_style(
    mwb: ManagedWorkbook,
    operation: Annotated[
        SetRangeStyleOperation,
        Field(description="Set range style operation parameters"),
    ],
) -> str:
    """Set style properties for cells in a range."""
    try:
        apply_style_to_range(mwb.wb, operation.range_a1, operation)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error setting style: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def batch_set_styles(
    mwb: ManagedWorkbook,
    operations: Annotated[
        list[BatchStyleOperation],
        Field(description="List of range/style pairs to apply"),
    ],
) -> str:
    """
    Apply styles to multiple ranges in a single call.

    Example:
    batch_set_styles([
        {"range_a1": "A1:C1", "style": {"bold": true, "bg_color": "FFFF00"}},
        {"range_a1": "A2:A10", "style": {"italic": true, "h_align": "right"}}
    ])
    """
    try:
        for op in operations:
            apply_style_to_range(mwb.wb, op.range_a1, op.style)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error applying styles: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def set_merge(
    mwb: ManagedWorkbook,
    operation: Annotated[
        MergeCellsOperation, Field(description="Merge cells operation parameters")
    ],
) -> str:
    """Merge cells in a range. The value from the top-left cell is kept."""
    try:
        sheet, cell_range = parse_range(mwb.wb, operation.range_a1)
        sheet.merge_cells(cell_range)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error merging cells: {e}").to_json()


@mcp.tool()
@with_workbook_mutation
def unmerge(
    mwb: ManagedWorkbook,
    operation: Annotated[
        UnmergeCellsOperation, Field(description="Unmerge cells operation parameters")
    ],
) -> str:
    """Unmerge previously merged cells."""
    try:
        sheet, cell_range = parse_range(mwb.wb, operation.range_a1)
        sheet.unmerge_cells(cell_range)
        return ToolSuccess().to_json()
    except Exception as e:
        return ToolError(f"Error unmerging cells: {e}").to_json()


# =============================================================================
# Search Tools
# =============================================================================


@mcp.tool()
@with_workbook
def search_cells(
    mwb: ManagedWorkbook,
    operation: Annotated[
        SearchCellsOperation, Field(description="Search cells operation parameters")
    ],
) -> str:
    """Search for cells containing the keyword. Returns JSON with matching cells."""
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        results = []

        for row in sheet.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                found = False

                if operation.find_by == "value" and operation.keyword in val:
                    found = True
                elif (
                    operation.find_by == "formula"
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and operation.keyword in cell.value
                ):
                    found = True

                if found:
                    results.append(
                        {
                            "cell": cell.coordinate,
                            "value": val[:200],
                            "formula": cell.value
                            if isinstance(cell.value, str)
                            and cell.value.startswith("=")
                            else None,
                        }
                    )
                    if len(results) >= operation.max_results:
                        break
            if len(results) >= operation.max_results:
                break

        return json.dumps({"total": len(results), "results": results})
    except Exception as e:
        return ToolError(f"Error searching cells: {e}").to_json()


# =============================================================================
# Conditional Formatting Tools
# =============================================================================


@mcp.tool()
@with_workbook_mutation
def add_conditional_formatting_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        AddConditionalFormattingOperation,
        Field(description="Add conditional formatting operation parameters"),
    ],
) -> str:
    """
    Add one or more conditional formatting rules to the given sheet.

    Rule types supported:
    - highlightCell: Highlight cells based on value comparison (uses operator + value)
    - expression: Highlight cells based on formula (uses formula field)
    - colorScale: Apply color gradient based on cell values (uses points field)

    Example:
    add_conditional_formatting_rule({"sheet_name": "Sheet1", "rules": [
        {"rule_type": "highlightCell", "range": "A1:A10", "operator": "greaterThan", "value": 100,
         "style": {"bgColor": "#FF0000", "bold": true}},
        {"rule_type": "expression", "range": "B1:B10", "formula": "=B1>A1",
         "style": {"bgColor": "#00FF00"}}
    ]})

    Returns JSON with status and rule IDs.
    """
    return add_conditional_formatting_rule_impl(mwb, operation)


@mcp.tool()
@with_workbook_mutation
def set_conditional_formatting_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        AddConditionalFormattingOperation,
        Field(description="Set conditional formatting operation parameters"),
    ],
) -> str:
    """
    Set (add) conditional formatting rules for the given sheet.
    This is an alias for add_conditional_formatting_rule.

    Returns JSON with status and rule IDs.
    """
    return add_conditional_formatting_rule_impl(mwb, operation)


@mcp.tool()
@with_workbook_mutation
def delete_conditional_formatting_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DeleteConditionalFormattingOperation, Field(description="Delete conditional formatting operation parameters"),
    ],
) -> str:
    """
    Delete one or more conditional formatting rules by ID.

    Returns JSON with status and list of deleted rule IDs.
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        deleted = []

        for rid in operation.rule_ids:
            if rid in mwb.cf_rules_store:
                stored = mwb.cf_rules_store[rid]
                if stored["sheet"] != sheet.title:
                    continue

                rule_obj = stored["rule"]

                # Remove from sheet's conditional formatting
                try:
                    for cf_item in list(sheet.conditional_formatting):
                        if hasattr(cf_item, "cfRule") and rule_obj in cf_item.cfRule:
                            cf_item.cfRule.remove(rule_obj)
                    del mwb.cf_rules_store[rid]
                    deleted.append(rid)
                except Exception:
                    pass

        return ToolSuccess(deleted=deleted).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


@mcp.tool()
@with_workbook
def get_conditional_formatting_rules(
    mwb: ManagedWorkbook,
    operation: Annotated[
        GetConditionalFormattingOperation,
        Field(description="Get conditional formatting operation parameters"),
    ],
) -> str:
    """
    Get all conditional formatting rules for the given sheet.

    Returns JSON array of conditional formatting rules with type, ranges, and parameters.
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        rules = []

        for cf_range, cf_rules in sheet.conditional_formatting._cf_rules.items():
            # Extract the actual range string from the CellRange object
            range_str = (
                cf_range.coord
                if hasattr(cf_range, "coord")
                else str(cf_range).split()[-1].rstrip(">")
            )

            for rule in cf_rules:
                rule_info = {
                    "range": range_str,
                    "type": rule.type,
                    "priority": rule.priority,
                }

                # Add type-specific info
                if hasattr(rule, "operator") and rule.operator:
                    rule_info["operator"] = rule.operator
                if hasattr(rule, "formula") and rule.formula:
                    rule_info["formula"] = list(rule.formula)
                if hasattr(rule, "colorScale") and rule.colorScale:
                    cs = rule.colorScale
                    rule_info["colorScale"] = {
                        "cfvo": [
                            {"type": cfvo.type, "val": cfvo.val}
                            for cfvo in (cs.cfvo or [])
                        ],
                        "colors": [
                            c.rgb if hasattr(c, "rgb") else str(c)
                            for c in (cs.color or [])
                        ],
                    }
                if hasattr(rule, "dataBar") and rule.dataBar:
                    rule_info["dataBar"] = True
                if hasattr(rule, "iconSet") and rule.iconSet:
                    rule_info["iconSet"] = (
                        rule.iconSet.iconSet
                        if hasattr(rule.iconSet, "iconSet")
                        else True
                    )
                if hasattr(rule, "dxf") and rule.dxf:
                    dxf = rule.dxf
                    style = {}
                    if dxf.fill and dxf.fill.fgColor:
                        style["bgColor"] = (
                            dxf.fill.fgColor.rgb
                            if hasattr(dxf.fill.fgColor, "rgb")
                            else str(dxf.fill.fgColor)
                        )
                    if dxf.font:
                        if dxf.font.color:
                            style["fontColor"] = (
                                dxf.font.color.rgb
                                if hasattr(dxf.font.color, "rgb")
                                else str(dxf.font.color)
                            )
                        if dxf.font.bold:
                            style["bold"] = dxf.font.bold
                    if style:
                        rule_info["style"] = style

                rules.append(rule_info)

        return json.dumps(rules)
    except Exception as e:
        return ToolError(str(e)).to_json()


# =============================================================================
# Data Validation Tools
# =============================================================================


@mcp.tool()
@with_workbook_mutation
def add_data_validation_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        AddDataValidationOperation,
        Field(description="Add data validation operation parameters"),
    ],
) -> str:
    """
    Add one or more data validation rules to the specified sheet.
    Can be used for dropdowns, date pickers, checkboxes, and custom validation.

    Validation types:
    - list: Dropdown list (use 'source' for comma-separated values or cell range)
    - whole: Integer validation
    - decimal: Decimal number validation
    - date: Date validation
    - textLength: Text length validation
    - custom: Custom formula validation
    - checkbox: Checkbox (TRUE/FALSE dropdown)

    Example:
    add_data_validation_rule({"sheet_name": "Sheet1", "rules": [
        {"validation_type": "list", "range_a1": "A1:A10", "source": "Option1,Option2,Option3"},
        {"validation_type": "whole", "range_a1": "B1:B10", "operator": "between", "value1": 1, "value2": 100},
        {"validation_type": "checkbox", "range_a1": "C1:C10"}
    ]})

    Returns JSON with status and rule IDs.
    """
    return add_data_validation_rule_impl(mwb, operation)


@mcp.tool()
@with_workbook_mutation
def set_data_validation_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        AddDataValidationOperation,
        Field(description="Set data validation operation parameters"),
    ],
) -> str:
    """
    Set (add) data validation rules for the specified sheet.
    This is an alias for add_data_validation_rule.

    Returns JSON with status and rule IDs.
    """
    return add_data_validation_rule_impl(mwb, operation)


@mcp.tool()
@with_workbook_mutation
def delete_data_validation_rule(
    mwb: ManagedWorkbook,
    operation: Annotated[
        DeleteDataValidationOperation, Field(description="Delete data validation operation parameters"),
    ],
) -> str:
    """
    Delete one or more data validation rules by ID.

    Returns JSON with status and list of deleted rule IDs.
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        deleted = []

        for rid in operation.rule_ids:
            if rid in mwb.dv_rules_store:
                stored = mwb.dv_rules_store[rid]
                if stored["sheet"] != sheet.title:
                    continue

                dv = stored["dv"]
                try:
                    if dv in sheet.data_validations.dataValidation:
                        sheet.data_validations.dataValidation.remove(dv)
                        deleted.append(rid)
                        del mwb.dv_rules_store[rid]
                except (ValueError, AttributeError):
                    pass

        return ToolSuccess(deleted=deleted).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


@mcp.tool()
@with_workbook
def get_data_validation_rules(
    mwb: ManagedWorkbook,
    operation: Annotated[
        GetDataValidationOperation,
        Field(description="Get data validation operation parameters"),
    ],
) -> str:
    """
    Get all data validation rules for the specified sheet.

    Returns JSON array of validation rules with type, operator, formulas, and ranges.
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        dvs = []
        for dv in sheet.data_validations.dataValidation:
            dvs.append(
                {
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "sqref": str(dv.sqref),
                }
            )
        return json.dumps(dvs)
    except Exception as e:
        return ToolError(str(e)).to_json()


# =============================================================================
# Tool Request (for missing functionality)
# =============================================================================


@mcp.tool()
def request_new_tool(
    operation: Annotated[
        RequestNewToolOperation,
        Field(description="Request new tool operation parameters"),
    ],
) -> str:
    """
    Request a new tool when no existing combination of tools can accomplish the task.

    ONLY call this when you have determined that the current task CANNOT be completed
    with the available tools. This will log the request and return an error.
    """
    print("\n" + "=" * 60)
    print("NEW TOOL REQUEST")
    print("=" * 60)
    print(f"\nTool Name: {operation.tool_name}")
    print(f"\nDescription:\n  {operation.description}")
    print(f"\nParameters:\n  {operation.parameters}")
    print(f"\nReason (why existing tools are insufficient):\n  {operation.reason}")
    print("\n" + "=" * 60)

    # Return an error so the agent can continue and find an alternative approach
    return ToolError(
        f"Tool '{operation.tool_name}' is not available. "
        f"The requested functionality ({operation.description}) cannot be provided. "
        f"Please try to accomplish the task using the existing tools, "
        f"or acknowledge that this specific task cannot be completed.",
        code="TOOL_NOT_AVAILABLE",
    ).to_json()


def main():
    """Main entry point for the MCP server."""
    mcp.run()

if __name__ == "__main__":
    main()