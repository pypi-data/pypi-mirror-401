"""Central formatting utilities for Excel data display."""

from rich.console import Console
from rich.table import Table
from rich.text import Text
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula


def _render_to_string(table: Table, width: int = 300) -> str:
    """Render a rich table to plain text string (no ANSI codes).

    Args:
        table: Rich Table to render
        width: Maximum width for the output (default 300, cells will wrap if needed)
    """
    console = Console(record=True, width=width)
    console.print(table)
    return console.export_text()


def _format_cell_value(value, cached_value=None) -> str:
    """
    Format a cell value for display.

    For formula cells, shows both formula and calculated value: =A1+B1 [5]
    For regular cells, shows just the value.
    """
    if value is None:
        return ""

    # Handle ArrayFormula objects - extract the formula text
    if isinstance(value, ArrayFormula):
        value = value.text

    str_value = str(value)

    # If it's a formula and we have a calculated value, show both
    if isinstance(value, str) and value.startswith("=") and cached_value is not None:
        return f"{str_value} [{cached_value}]"

    return str_value


def format_sheet_table(
    sheet: Worksheet,
    cached_sheet: Worksheet | None = None,
    max_rows: int = 50,
    total_rows: int | None = None,
    title: str | None = None,
    width: int = 300,
) -> str:
    """
    Format sheet data as a rich table.

    Args:
        sheet: The worksheet with formulas
        cached_sheet: Optional worksheet loaded with data_only=True for calculated values
        max_rows: Maximum rows to display
        total_rows: Total rows in sheet (for truncation message)
        title: Optional table title
        width: Maximum width for the table output (cells wrap if needed)

    Returns:
        Formatted string representation of the table
    """
    # Use actual data extent, capped at max_rows
    actual_rows = min(sheet.max_row, max_rows) if sheet.max_row else 0
    rows = list(sheet.iter_rows(max_row=actual_rows))
    if not rows:
        return "Sheet is empty"

    table = Table(
        title=title,
        show_header=True,
        header_style="bold cyan",
        border_style="dim",
        row_styles=["", "dim"],
    )

    # Add row number column
    table.add_column("#", style="dim", justify="right", width=4)

    # Add data columns (A, B, C, ...)
    num_cols = max(len(row) for row in rows) if rows else 0
    for i in range(num_cols):
        table.add_column(get_column_letter(i + 1), justify="left")

    # Add rows (limited)
    for i, row in enumerate(rows[:max_rows]):
        str_row = []
        for j, cell in enumerate(row):
            cached_value = None
            if cached_sheet is not None:
                try:
                    cached_cell = cached_sheet.cell(row=i + 1, column=j + 1)
                    cached_value = cached_cell.value
                except Exception:
                    pass
            str_row.append(_format_cell_value(cell.value, cached_value))

        # Pad row if needed
        while len(str_row) < num_cols:
            str_row.append("")

        table.add_row(str(i + 1), *str_row)

    output = _render_to_string(table, width=width)

    if total_rows and total_rows > max_rows:
        output += f"\n... ({total_rows - max_rows} more rows)"

    return output


def _parse_range_bounds(range_str: str) -> tuple[int, int]:
    """
    Parse range string to get starting column and row numbers.

    Args:
        range_str: Range like 'A1', 'B2:D5', 'Sheet1!A1:C3'

    Returns:
        (start_col, start_row) as 1-based integers
    """
    # Remove sheet reference if present
    if "!" in range_str:
        range_str = range_str.split("!")[-1]

    # Get first cell reference (before colon if range)
    first_cell = range_str.split(":")[0].strip()

    # Parse column letters and row number
    col_str = ""
    row_str = ""
    for char in first_cell:
        if char.isalpha():
            col_str += char.upper()
        elif char.isdigit():
            row_str += char

    # Convert column letters to number (A=1, B=2, ..., AA=27)
    col_num = 0
    for char in col_str:
        col_num = col_num * 26 + (ord(char) - ord("A") + 1)

    row_num = int(row_str) if row_str else 1

    return col_num, row_num


def format_range_table(
    cells: list[list[dict]],
    range_str: str = "",
    width: int = 300,
    max_rows: int = 100,
) -> str:
    """
    Format range data as a rich table with styles.

    Args:
        cells: 2D list of cell dicts with 'value', 'calculated', 'bold', 'italic', 'bg' keys
        range_str: Optional range string for title (also used for column/row labels)
        width: Maximum width for the table output (cells wrap if needed)
        max_rows: Maximum rows to display (truncates with message if exceeded)

    Returns:
        Formatted string with styles applied
    """
    if not cells:
        return "(empty)"
    
    total_rows = len(cells)
    truncated = total_rows > max_rows
    cells = cells[:max_rows]  # Truncate to max_rows

    # Parse range to get starting column and row
    start_col, start_row = _parse_range_bounds(range_str) if range_str else (1, 1)

    table = Table(
        title=range_str if range_str else None,
        show_header=True,
        header_style="bold cyan",
        border_style="dim",
        row_styles=["", "dim"],
    )

    # Add row number column
    table.add_column("#", style="dim", justify="right", width=4)

    # Add data columns with proper column letters
    num_cols = max(len(row) for row in cells) if cells else 0
    for i in range(num_cols):
        table.add_column(get_column_letter(start_col + i), justify="left")

    # Add rows with proper row numbers
    for row_idx, row in enumerate(cells):
        styled_cells = []
        for cell in row:
            if cell is None:
                styled_cells.append("")
                continue

            value = cell.get("value")
            calculated = cell.get("calculated")

            # Format with formula [value] if applicable
            display_value = _format_cell_value(value, calculated)
            text = Text(display_value)

            # Apply styles
            if cell.get("bold"):
                text.stylize("bold")
            if cell.get("italic"):
                text.stylize("italic")
            if cell.get("bg"):
                bg_color = (
                    cell["bg"].lstrip("#") if isinstance(cell["bg"], str) else None
                )
                if bg_color and bg_color != "00000000":
                    try:
                        text.stylize(f"on #{bg_color[-6:]}")
                    except Exception:
                        pass

            styled_cells.append(text)

        # Pad row
        while len(styled_cells) < num_cols:
            styled_cells.append("")

        # Add row with proper row number
        table.add_row(str(start_row + row_idx), *styled_cells)

    output = _render_to_string(table, width=width)
    if truncated:
        output += f"\n... ({total_rows - max_rows} more rows, {total_rows} total)"
    return output
