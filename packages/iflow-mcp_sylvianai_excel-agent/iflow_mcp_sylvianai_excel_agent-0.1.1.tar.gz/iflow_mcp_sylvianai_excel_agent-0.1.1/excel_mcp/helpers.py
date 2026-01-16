"""Helper functions and utilities for the Excel MCP server."""

import asyncio
import logging
import os
import platform
import re
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, TYPE_CHECKING

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

from .errors import ToolError, ToolSuccess
from .sessions import sessions

if TYPE_CHECKING:
    from models import ManagedWorkbook
    from schemas import AddConditionalFormattingOperation, AddDataValidationOperation

logger = logging.getLogger(__name__)


# =============================================================================
# Range Parsing
# =============================================================================


def parse_range(wb: Workbook, range_ref: str) -> tuple:
    """
    Parse a range reference that may include a sheet name.

    Args:
        wb: Workbook instance
        range_ref: Range like "A1", "A1:B2", or "Sheet1!A1:B2"

    Returns:
        Tuple of (sheet, cell_range_string)

    Raises:
        ToolError: If sheet not found
    """
    if "!" in range_ref:
        sheet_name, cell_range = range_ref.split("!", 1)
        sheet_name = sheet_name.strip("'\"")
        if sheet_name not in wb.sheetnames:
            raise ToolError(f"Sheet '{sheet_name}' not found", code="SHEET_NOT_FOUND")
        return wb[sheet_name], cell_range
    return wb.active, range_ref


def get_sheet(wb: Workbook, sheet_name: str | None):
    """
    Get a worksheet by name, or the active sheet if name is None.

    Args:
        wb: Workbook instance
        sheet_name: Sheet name, or None for active sheet

    Returns:
        Worksheet object
    """
    return wb[sheet_name] if sheet_name else wb.active


# =============================================================================
# Formula Utilities
# =============================================================================


def is_array_formula(formula: str) -> bool:
    """
    Detect if a formula needs to be entered as an array formula (CSE).

    Array formulas use range operations that Excel would otherwise
    convert with implicit intersection (@), breaking the formula logic.

    Common patterns:
    - MATCH(1, (range=value)*(range=value), 0) - array AND logic
    - INDEX(..., MATCH(1, ...*..., 0)) - lookup with array matching

    Note: SUMPRODUCT handles arrays natively and doesn't need CSE.
    """
    if not formula.startswith("="):
        return False

    formula_upper = formula.upper()

    # SUMPRODUCT handles arrays natively - no CSE needed
    if formula_upper.startswith("=SUMPRODUCT"):
        return False

    # Pattern: MATCH(1, ...) with multiplication inside - array AND logic
    if "MATCH(1," in formula_upper or "MATCH( 1," in formula_upper:
        # Check if there's multiplication of range comparisons
        if re.search(
            r"\([^)]*\$?[A-Z]+\$?\d*:\$?[A-Z]+\$?\d*[^)]*\)\s*\*\s*\(", formula
        ):
            return True

    return False


def translate_formula(formula: str, row_offset: int, col_offset: int) -> str:
    """
    Translate cell references in a formula by the given row and column offsets.

    Handles:
    - Relative references (A1) - adjusted by offset
    - Absolute row ($A1) - column adjusted, row fixed
    - Absolute col (A$1) - row adjusted, column fixed
    - Fully absolute ($A$1) - not adjusted
    """
    # Pattern to match cell references: optional $ before column, column letters, optional $ before row, row number
    cell_ref_pattern = r"(\$?)([A-Z]+)(\$?)(\d+)"

    def replace_ref(match):
        col_absolute = match.group(1) == "$"
        col_letters = match.group(2)
        row_absolute = match.group(3) == "$"
        row_num = int(match.group(4))

        # Adjust column if not absolute
        if not col_absolute:
            col_idx = column_index_from_string(col_letters)
            col_idx += col_offset
            if col_idx < 1:
                col_idx = 1
            col_letters = get_column_letter(col_idx)

        # Adjust row if not absolute
        if not row_absolute:
            row_num += row_offset
            if row_num < 1:
                row_num = 1

        # Reconstruct the reference
        return f"{'$' if col_absolute else ''}{col_letters}{'$' if row_absolute else ''}{row_num}"

    return re.sub(cell_ref_pattern, replace_ref, formula, flags=re.IGNORECASE)


# =============================================================================
# Cell Value Utilities
# =============================================================================


def serialize_cell_value(value) -> str | int | float | bool | None:
    """Convert a cell value to a JSON-serializable type."""
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def is_formula_error(value) -> bool:
    """Check if a value is an Excel formula error (e.g., #NUM!, #VALUE!, #ERROR)."""
    if not isinstance(value, str):
        return False
    # Excel errors: #NUM!, #VALUE!, #REF!, #DIV/0!, #NAME?, #N/A, #NULL!, #ERROR
    return (
        value.startswith("#")
        and (value.endswith("!"))
        or value in ("#N/A", "#NAME?", "#ERROR")
    )


def has_formula_errors(cell_results: list[dict]) -> bool:
    """Check if any cell results contain formula calculation errors."""
    for cell in cell_results:
        calc_value = cell.get("calculated")
        if calc_value is not None and is_formula_error(calc_value):
            return True
    return False


def extract_cell_data(cell, cached_cell=None, with_style: bool = True) -> dict:
    """Extract cell value, calculated value, and style info."""
    value = cell.value

    # Handle ArrayFormula objects - extract the formula text
    if isinstance(value, ArrayFormula):
        value = value.text

    # Convert datetime to ISO string for JSON serialization
    value = serialize_cell_value(value)

    calculated = None

    # If this is a formula cell and we have a cached value, include it
    if isinstance(value, str) and value.startswith("=") and cached_cell is not None:
        calculated = serialize_cell_value(cached_cell.value)

    return {
        "value": value,
        "calculated": calculated,
        "cell": cell.coordinate,
        "bold": cell.font.bold if with_style else None,
        "italic": cell.font.italic if with_style else None,
        "bg": cell.fill.start_color.rgb
        if (with_style and cell.fill.patternType)
        else None,
    }


# =============================================================================
# Style Utilities
# =============================================================================


def parse_style_dict(style_dict: dict) -> tuple:
    """Parse a style dictionary into Font and PatternFill objects.

    Args:
        style_dict: Dict with keys like 'bold', 'italic', 'fgColor', 'bgColor'

    Returns:
        Tuple of (Font or None, PatternFill or None)
    """
    font_args = {}
    if "bold" in style_dict:
        font_args["bold"] = style_dict["bold"]
    if "italic" in style_dict:
        font_args["italic"] = style_dict["italic"]
    if "fgColor" in style_dict:
        font_args["color"] = style_dict["fgColor"].replace("#", "")

    style_font = Font(**font_args) if font_args else None

    style_fill = None
    if "bgColor" in style_dict:
        style_fill = PatternFill(
            start_color=style_dict["bgColor"].replace("#", ""),
            end_color=style_dict["bgColor"].replace("#", ""),
            fill_type="solid",
        )

    return style_font, style_fill


# =============================================================================
# Validation Type Mapping
# =============================================================================

# Map user-friendly validation type names to openpyxl types
VALIDATION_TYPE_MAP = {
    "text_length": "textLength",
    "custom_formula": "custom",
    "integer": "whole",
    "checkbox": "list",
}


def normalize_validation_type(validation_type: str, rule_dict: dict) -> str:
    """Normalize validation type and handle special cases like checkbox.

    Args:
        validation_type: User-provided validation type
        rule_dict: Rule dictionary (modified in-place for checkbox)

    Returns:
        Normalized validation type for openpyxl
    """
    if validation_type == "checkbox":
        rule_dict["source"] = "TRUE,FALSE"
    return VALIDATION_TYPE_MAP.get(validation_type, validation_type)


# =============================================================================
# Formula Calculation (via LibreOffice/Excel)
# =============================================================================


def _calculate_formulas_excel_mac(filepath: str, timeout: float = 60.0) -> bool:
    """Calculate formulas using Excel on macOS via AppleScript."""
    logger = logging.getLogger(__name__)

    filepath = str(Path(filepath).resolve())

    # AppleScript to open Excel, save (which triggers calculation), and close
    # Note: Explicit 'calculate' command doesn't work, but save triggers recalc
    applescript = f'''
    tell application "Microsoft Excel"
        open "{filepath}"
        delay 1
        set wb to active workbook
        save wb
        close wb saving no
    end tell
    '''

    try:
        result = subprocess.run(
            ["osascript", "-e", applescript],
            capture_output=True,
            text=True,
            timeout=timeout,
        )

        if result.returncode != 0:
            logger.warning(f"Excel (macOS) failed: {result.stderr}")
            return False

        logger.info("Excel (macOS) formula calculation succeeded")
        return True

    except subprocess.TimeoutExpired:
        logger.warning(f"Excel (macOS) timed out after {timeout}s")
        return False
    except Exception as e:
        logger.warning(f"Excel (macOS) calculation failed: {e}")
        return False


def _calculate_formulas_excel_win(filepath: str) -> bool:
    """Calculate formulas using Excel on Windows via win32com."""
    logger = logging.getLogger(__name__)

    try:
        from win32com.client import Dispatch
    except ImportError:
        logger.warning("win32com not available")
        return False

    filepath = str(Path(filepath).resolve())

    try:
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = False
        xlApp.ScreenUpdating = False

        try:
            xlBook = xlApp.Workbooks.Open(
                Filename=filepath, UpdateLinks=False, ReadOnly=False
            )
            xlBook.Save()
            xlBook.Close(SaveChanges=True)
            logger.info("Excel (Windows) formula calculation succeeded")
            return True
        finally:
            xlApp.Quit()

    except Exception as e:
        logger.warning(f"Excel (Windows) calculation failed: {e}")
        return False


def _calculate_formulas_libreoffice_sync(filepath: str, timeout: float = 120.0) -> bool:
    """
    Synchronous LibreOffice calculation - runs in thread pool.

    Uses DOUBLE two-step conversion (xlsx → ods → xlsx → ods → xlsx) to force
    LibreOffice to fully recalculate all formulas. LibreOffice headless mode
    doesn't always recalculate on the first pass, but running through the
    conversion twice ensures formulas are properly evaluated.
    """
    logger = logging.getLogger(__name__)

    # Try to find soffice - check PATH first, then common installation locations
    soffice_path = shutil.which("soffice") or shutil.which("libreoffice")

    if not soffice_path:
        # Check common installation paths
        common_paths = [
            "/usr/local/bin/soffice",  # macOS Homebrew / manual install
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS app bundle
            "/usr/bin/soffice",  # Linux
            "/usr/bin/libreoffice",  # Linux alternative
            "/snap/bin/libreoffice",  # Linux snap
        ]
        for path in common_paths:
            if Path(path).exists():
                soffice_path = path
                break

    if not soffice_path:
        logger.debug(
            "[LibreOffice] soffice/libreoffice not found in PATH or common locations"
        )
        return False

    filepath = str(Path(filepath).resolve())
    filename = Path(filepath).name
    filename_stem = Path(filepath).stem

    logger.debug(f"[LibreOffice] Using: {soffice_path}")
    logger.debug(f"[LibreOffice] Input file: {filepath}")

    # Use a temp directory for output to avoid overwrite issues
    with tempfile.TemporaryDirectory() as temp_dir:
        # Set HOME to temp dir to avoid LibreOffice hanging on profile issues
        env = os.environ.copy()
        env["HOME"] = temp_dir
        logger.debug(f"[LibreOffice] Using temp HOME: {temp_dir}")

        # Per-step timeout (4 steps total now)
        step_timeout = timeout / 4

        try:
            # Common args to suppress all popups/dialogs
            suppress_args = [
                "--headless",
                "--invisible",
                "--nologo",
                "--nofirststartwizard",
                "--norestore",
            ]

            # === FIRST PASS: xlsx -> ods -> xlsx ===
            # Step 1: Convert xlsx -> ods
            cmd1 = [
                soffice_path,
                *suppress_args,
                "--calc",
                "--convert-to",
                "ods",
                "--outdir",
                temp_dir,
                filepath,
            ]
            logger.debug(f"[LibreOffice] Pass 1, Step 1 (xlsx→ods): {' '.join(cmd1)}")

            result1 = subprocess.run(
                cmd1, capture_output=True, text=True, timeout=step_timeout, env=env
            )

            logger.debug(
                f"[LibreOffice] Pass 1, Step 1 exit code: {result1.returncode}"
            )
            if result1.returncode != 0:
                logger.warning(f"LibreOffice pass 1 step 1 failed: {result1.stderr}")
                return False

            ods_file = Path(temp_dir) / f"{filename_stem}.ods"
            if not ods_file.exists():
                logger.debug("[LibreOffice] Pass 1, Step 1 FAILED: ODS file not found")
                return False

            # Step 2: Convert ods -> xlsx
            cmd2 = [
                soffice_path,
                *suppress_args,
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                temp_dir,
                str(ods_file),
            ]
            logger.debug("[LibreOffice] Pass 1, Step 2 (ods→xlsx): converting...")

            result2 = subprocess.run(
                cmd2, capture_output=True, text=True, timeout=step_timeout, env=env
            )

            if result2.returncode != 0:
                logger.warning(f"LibreOffice pass 1 step 2 failed: {result2.stderr}")
                return False

            xlsx_file = Path(temp_dir) / filename
            if not xlsx_file.exists():
                logger.debug("[LibreOffice] Pass 1, Step 2 FAILED: XLSX file not found")
                return False

            logger.debug("[LibreOffice] Pass 1 complete")

            # Clean up intermediate ODS for second pass
            ods_file.unlink()

            # === SECOND PASS: xlsx -> ods -> xlsx (ensures recalculation) ===
            # Step 3: Convert xlsx -> ods again
            cmd3 = [
                soffice_path,
                *suppress_args,
                "--calc",
                "--convert-to",
                "ods",
                "--outdir",
                temp_dir,
                str(xlsx_file),
            ]
            logger.debug("[LibreOffice] Pass 2, Step 1 (xlsx→ods): converting...")

            result3 = subprocess.run(
                cmd3, capture_output=True, text=True, timeout=step_timeout, env=env
            )

            if result3.returncode != 0:
                logger.warning(f"LibreOffice pass 2 step 1 failed: {result3.stderr}")
                return False

            ods_file2 = Path(temp_dir) / f"{filename_stem}.ods"
            if not ods_file2.exists():
                logger.debug("[LibreOffice] Pass 2, Step 1 FAILED: ODS file not found")
                return False

            # Remove the xlsx so we can create fresh one
            xlsx_file.unlink()

            # Step 4: Convert ods -> xlsx final
            cmd4 = [
                soffice_path,
                *suppress_args,
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                temp_dir,
                str(ods_file2),
            ]
            logger.debug("[LibreOffice] Pass 2, Step 2 (ods→xlsx): converting...")

            result4 = subprocess.run(
                cmd4, capture_output=True, text=True, timeout=step_timeout, env=env
            )

            if result4.returncode != 0:
                logger.warning(f"LibreOffice pass 2 step 2 failed: {result4.stderr}")
                return False

            # Check for write errors in stderr
            if "Error:" in result4.stderr:
                logger.warning(f"LibreOffice error: {result4.stderr}")
                return False

            # Copy the calculated file back
            output_file = Path(temp_dir) / filename
            logger.debug(f"[LibreOffice] Looking for output: {output_file}")

            if output_file.exists():
                shutil.copy(output_file, filepath)
                logger.info(
                    "LibreOffice formula calculation succeeded (double two-step conversion)"
                )
                logger.debug(f"[LibreOffice] Success! Copied back to {filepath}")
                return True
            else:
                logger.warning(f"LibreOffice output file not found: {output_file}")
                logger.debug("[LibreOffice] FAILED: Output file not found")
                return False

        except subprocess.TimeoutExpired:
            logger.warning(f"LibreOffice timed out after {timeout}s")
            logger.debug(f"[LibreOffice] FAILED: Timed out after {timeout}s")
            return False
        except Exception as e:
            logger.warning(f"LibreOffice calculation failed: {e}")
            logger.debug(f"[LibreOffice] FAILED: Exception: {e}")
            return False


async def calculate_formulas_libreoffice(filepath: str, timeout: float = 120.0) -> bool:
    """Calculate formulas using LibreOffice headless (async-safe via thread pool)."""
    return await asyncio.to_thread(
        _calculate_formulas_libreoffice_sync, filepath, timeout
    )


async def calculate_formulas(filepath: str, timeout: float = 120.0) -> bool:
    """
    Calculate formulas by opening the file in LibreOffice/Excel and saving.

    Tries in order:
    1. Excel on macOS (native, best Excel compatibility)
    2. Excel on Windows (via win32com)
    3. LibreOffice headless (fallback, cross-platform)

    Returns True if successful, False if all methods failed.
    """
    logger = logging.getLogger(__name__)
    system = platform.system()

    # Try Excel first (native Excel formula compatibility)
    if system == "Darwin":  # macOS
        if _calculate_formulas_excel_mac(filepath, timeout):
            return True
    elif system == "Windows":
        if _calculate_formulas_excel_win(filepath):
            return True

    if await calculate_formulas_libreoffice(filepath, timeout):
        return True

    logger.warning("No formula calculation method available (Excel or LibreOffice)")
    return False


async def get_cell_results(
    mwb: "ManagedWorkbook", cells_set: list[tuple[str, Any]]
) -> list[dict]:
    """
    Get results for all cells that were set. Calculates formulas via LibreOffice if any exist.

    Args:
        mwb: ManagedWorkbook instance
        cells_set: List of (cell_ref, value) tuples, e.g. [("A1", "Hello"), ("C1", "=A1+B1")]

    Returns:
        List of dicts with 'cell', 'value', and 'calculated' (only for formulas).
        Example: [
            {"cell": "A1", "value": "Hello"},
            {"cell": "C1", "value": "=A1+B1", "calculated": 3}
        ]
    """
    if not cells_set:
        return []

    # Check if any formulas were set
    formula_cells = [
        (cell_ref, val)
        for cell_ref, val in cells_set
        if isinstance(val, str) and val.startswith("=")
    ]
    has_formulas = len(formula_cells) > 0

    cached_wb = None
    if has_formulas:
        filepath = mwb.filepath
        logger.debug(f"[get_cell_results] Saving workbook to: {filepath}")
        mwb.wb.save(filepath)

        # Calculate formulas using LibreOffice (async)
        logger.debug("[get_cell_results] Calling calculate_formulas...")
        await calculate_formulas(filepath)
        logger.debug("[get_cell_results] calculate_formulas completed")

        # Reload with data_only to get calculated values
        logger.debug("[get_cell_results] Loading workbook with data_only=True...")
        cached_wb = openpyxl.load_workbook(filepath, data_only=True)
        logger.debug(
            f"[get_cell_results] Loaded cached workbook, sheets: {cached_wb.sheetnames}"
        )

    results = []
    for cell_ref, value in cells_set:
        # Convert datetime to ISO string for JSON serialization
        if isinstance(value, datetime):
            value = value.isoformat()

        result = {"cell": cell_ref, "value": value}

        # Add calculated value for formulas
        if isinstance(value, str) and value.startswith("="):
            if cached_wb is not None:
                sheet, cell_range = parse_range(cached_wb, cell_ref)
                calc_value = sheet[cell_range].value
                # Convert datetime to ISO string
                if isinstance(calc_value, datetime):
                    calc_value = calc_value.isoformat()
                # If calculation returned None, mark as not calculated
                if calc_value is None:
                    result["calculated"] = "#ERROR"
                    # result["_calc_note"] = "Formula could not evaluated by LibreOffice (may use unsupported Excel functions)"
                else:
                    result["calculated"] = calc_value
                logger.debug(
                    f"[get_cell_results] {cell_ref}: formula={value[:50]}... -> calculated={calc_value}"
                )
            else:
                result["calculated"] = "#ERROR"
                # result["_calc_note"] = "Calculation failed"
                logger.debug(
                    f"[get_cell_results] {cell_ref}: formula={value[:50]}... -> no cached_wb available"
                )

        results.append(result)

    return results


# =============================================================================
# Workbook Save Operations
# =============================================================================


async def save_workbook_async(session_id: str) -> str:
    """Save the workbook and calculate formulas (async version)."""
    logger = logging.getLogger(__name__)

    try:
        state = sessions.get_session(session_id)
        if state.workbook is None:
            return "No workbook loaded in session"

        mwb = state.workbook
        filepath = mwb.filepath

        mwb.wb.save(filepath)
        logger.info(f"Saved workbook to {filepath}, calculating formulas...")

        # Calculate formulas with timeout (60s for complex formulas)
        calc_success = await calculate_formulas(filepath, timeout=120.0)
        if calc_success:
            logger.info("Formula calculation succeeded")
        else:
            logger.warning(
                "Formula calculation failed or timed out - values may not be computed"
            )

        # DON'T reload and resave - that would lose the calculated values!
        # The formulas library uppercases sheet names but that's acceptable.

        mwb.is_dirty = False
        return f"Saved workbook to {filepath}"
    except Exception as e:
        return ToolError(f"Error saving workbook: {e}").to_json()


def save_workbook(session_id: str) -> str:
    """Save the workbook and calculate formulas (sync wrapper for non-async contexts)."""
    try:
        asyncio.get_running_loop()  # Check if we're in an async context
        # If we're in an async context, we can't use asyncio.run
        # Fall back to sync-only save without calculation
        logger = logging.getLogger(__name__)
        state = sessions.get_session(session_id)
        if state.workbook is None:
            return "No workbook loaded in session"
        mwb = state.workbook
        filepath = mwb.filepath
        mwb.wb.save(filepath)
        mwb.is_dirty = False
        logger.warning(
            "save_workbook called from async context - formulas not calculated"
        )
        return f"Saved workbook to {filepath} (formulas not calculated)"
    except RuntimeError:
        # No running loop - we can use asyncio.run
        return asyncio.run(save_workbook_async(session_id))


async def ensure_calculated(mwb: "ManagedWorkbook") -> None:
    """
    Recalculate formulas if workbook was modified, then reload to get computed values.

    This ensures get_range_data and display_sheet show current formula results.
    Modifies the ManagedWorkbook in place.

    Args:
        mwb: ManagedWorkbook instance
    """
    if not mwb.is_dirty:
        return

    filepath = mwb.filepath

    try:
        mwb.wb.save(filepath)

        # Calculate formulas using LibreOffice (async)
        await calculate_formulas(filepath)

        # Close old workbooks
        if mwb.cached_values_wb is not None:
            try:
                mwb.cached_values_wb.close()
            except Exception as e:
                logger.debug(f"Error closing cached workbook: {e}")
        try:
            mwb.wb.close()
        except Exception as e:
            logger.debug(f"Error closing workbook: {e}")

        # Reload workbook with updated values
        mwb.wb = openpyxl.load_workbook(filepath)
        mwb.cached_values_wb = openpyxl.load_workbook(filepath, data_only=True)

        mwb.is_dirty = False
    except Exception as e:
        logger.error(f"Error during formula calculation: {e}")
        mwb.is_dirty = False


def set_cell_value(sheet, cell_ref: str, value) -> None:
    """Set a cell value, using ArrayFormula for array formulas."""
    if isinstance(value, str) and value.startswith("=") and is_array_formula(value):
        # Use ArrayFormula for array formulas to prevent @ insertion
        sheet[cell_ref] = ArrayFormula(ref=cell_ref, text=value)
    else:
        sheet[cell_ref].value = value


# =============================================================================
# Range Operation Helpers
# =============================================================================


def iter_range_cells(sheet, cell_range: str):
    """
    Iterate over all cells in a range (single cell or multi-cell range).

    Yields: cell objects from the sheet
    """
    if ":" in cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            yield from row
    else:
        yield sheet[cell_range]


def apply_style_to_range(wb: Workbook, range_a1: str, style) -> None:
    """
    Apply style properties to all cells in a range.

    Args:
        wb: Workbook instance
        range_a1: Range in A1 notation (can be comma-separated)
        style: Object with optional attributes: bold, italic, font_size, font_color, bg_color, h_align, v_align
    """
    ranges = range_a1.split(",")

    for r in ranges:
        r = r.strip()
        sheet, cell_range = parse_range(wb, r)

        for cell in iter_range_cells(sheet, cell_range):
            # Font properties
            if any(
                getattr(style, attr, None) is not None
                for attr in ["bold", "italic", "font_size", "font_color"]
            ):
                current = cell.font
                cell.font = Font(
                    name=current.name,
                    size=style.font_size
                    if getattr(style, "font_size", None) is not None
                    else current.size,
                    bold=style.bold
                    if getattr(style, "bold", None) is not None
                    else current.bold,
                    italic=style.italic
                    if getattr(style, "italic", None) is not None
                    else current.italic,
                    color=style.font_color.lstrip("#")
                    if getattr(style, "font_color", None) is not None
                    else current.color,
                    underline=current.underline,
                    strike=current.strike,
                )

            # Background color
            if getattr(style, "bg_color", None) is not None:
                color = style.bg_color.lstrip("#")
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            # Alignment
            if (
                getattr(style, "h_align", None) is not None
                or getattr(style, "v_align", None) is not None
            ):
                current = cell.alignment
                cell.alignment = Alignment(
                    horizontal=style.h_align
                    if getattr(style, "h_align", None) is not None
                    else current.horizontal,
                    vertical=style.v_align
                    if getattr(style, "v_align", None) is not None
                    else current.vertical,
                    wrap_text=current.wrap_text,
                    shrink_to_fit=current.shrink_to_fit,
                )


def set_range_values(wb: Workbook, range_a1: str, value) -> list[tuple[str, Any]]:
    """
    Set value(s) in a range. Handles single cells, rows, columns, and grids.

    Args:
        wb: Workbook instance
        range_a1: Range in A1 notation (can be comma-separated)
        value: Single value, 1D list, or 2D list

    Returns:
        List of (cell_ref, value) tuples for all cells that were set
    """
    cells_set = []
    ranges = range_a1.split(",")

    for r in ranges:
        r = r.strip()
        sheet, cell_range = parse_range(wb, r)

        if ":" in cell_range:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            rows = max_row - min_row + 1
            cols = max_col - min_col + 1

            if isinstance(value, list):
                if not value:
                    continue

                # Convert 1D to 2D if needed
                values_2d = value
                if not isinstance(value[0], list):
                    # Column vector: single column, multiple rows
                    if rows > 1 and cols == 1 and len(value) == rows:
                        values_2d = [[v] for v in value]
                    else:
                        # Row vector
                        values_2d = [value]

                for i, row_val in enumerate(values_2d):
                    r_idx = min_row + i
                    if r_idx > max_row:
                        break
                    for j, cell_val in enumerate(row_val):
                        c_idx = min_col + j
                        if c_idx > max_col:
                            break
                        cell_ref = f"{get_column_letter(c_idx)}{r_idx}"
                        set_cell_value(sheet, cell_ref, cell_val)
                        cells_set.append((cell_ref, cell_val))
            else:
                # Fill all cells with single value
                for row in sheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    for cell in row:
                        set_cell_value(sheet, cell.coordinate, value)
                        cells_set.append((cell.coordinate, value))
        else:
            # Single cell
            set_cell_value(sheet, cell_range, value)
            cells_set.append((cell_range, value))

    return cells_set


# =============================================================================
# Conditional Formatting and Data Validation Helpers
# =============================================================================


def add_conditional_formatting_rule_impl(
    mwb: "ManagedWorkbook",
    operation: "AddConditionalFormattingOperation",
) -> str:
    """
    Implementation for adding conditional formatting rules to a worksheet.

    Args:
        mwb: ManagedWorkbook instance containing the workbook and rule stores
        operation: AddConditionalFormattingOperation with sheet_name and rules

    Returns:
        JSON string with status and list of added rule IDs
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        added_ids: list[str] = []

        for rule_model in operation.rules:
            rule_dict = rule_model.model_dump()

            # ID generation
            rule_id: str = rule_dict.get("rule_id") or str(uuid.uuid4())

            # Range parsing
            target_range: str | None = rule_dict.get("range")
            ranges: list[str] = (
                [target_range] if target_range else rule_dict.get("ranges", [])
            )
            if isinstance(ranges, str):
                ranges = [ranges]
            if not ranges:
                continue

            # Parse style if present
            style_font: Font | None = None
            style_fill: PatternFill | None = None
            if "style" in rule_dict and isinstance(rule_dict["style"], dict):
                style_font, style_fill = parse_style_dict(rule_dict["style"])

            # Create rule based on type
            rule_type: str | None = rule_dict.get("rule_type")
            rule_obj: ColorScaleRule | CellIsRule | FormulaRule | None = None

            if rule_type == "highlightCell":
                op: str | None = rule_dict.get("operator")
                val = rule_dict.get("value")
                formula: list[str] = []
                if val is not None:
                    if isinstance(val, list):
                        formula = [str(v) for v in val]
                    else:
                        formula = [str(val)]
                elif "formula" in rule_dict:
                    formula = [rule_dict["formula"]]

                if op:
                    rule_obj = CellIsRule(
                        operator=op, formula=formula, fill=style_fill, font=style_font
                    )

            elif rule_type == "expression":
                formula = [rule_dict.get("formula", "")]
                rule_obj = FormulaRule(
                    formula=formula, fill=style_fill, font=style_font
                )

            elif rule_type == "colorScale":
                points: list[dict[str, Any]] = rule_dict.get("points", [])
                if len(points) >= 2:
                    start = points[0]
                    end = points[-1]
                    start_color: str = start.get("color", "#FF0000").replace("#", "")
                    end_color: str = end.get("color", "#00FF00").replace("#", "")
                    mid_color: str | None = None
                    if len(points) == 3:
                        mid_color = points[1].get("color", "#FFFF00").replace("#", "")

                    if mid_color:
                        rule_obj = ColorScaleRule(
                            start_type="min",
                            start_color=start_color,
                            mid_type="percentile",
                            mid_value=50,
                            mid_color=mid_color,
                            end_type="max",
                            end_color=end_color,
                        )
                    else:
                        rule_obj = ColorScaleRule(
                            start_type="min",
                            start_color=start_color,
                            end_type="max",
                            end_color=end_color,
                        )

            if rule_obj:
                for r in ranges:
                    sheet.conditional_formatting.add(r, rule_obj)

                # Store in workbook's rule store
                mwb.cf_rules_store[rule_id] = {
                    "sheet": operation.sheet_name,
                    "rule": rule_obj,
                    "ranges": ranges,
                }
                added_ids.append(rule_id)

        return ToolSuccess(ids=added_ids).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()


def add_data_validation_rule_impl(
    mwb: "ManagedWorkbook",
    operation: "AddDataValidationOperation",
) -> str:
    """
    Implementation for adding data validation rules to a worksheet.

    Args:
        mwb: ManagedWorkbook instance containing the workbook and rule stores
        operation: AddDataValidationOperation with sheet_name and rules

    Returns:
        JSON string with status and list of added rule IDs
    """
    try:
        sheet = get_sheet(mwb.wb, operation.sheet_name)
        added_ids: list[str] = []

        for rule_model in operation.rules:
            rule_dict = rule_model.model_dump()
            rule_id: str = rule_dict.get("rule_id") or str(uuid.uuid4())

            # Normalize validation type (handles aliases like 'checkbox' -> 'list')
            dv_type: str = normalize_validation_type(
                rule_dict.get("validation_type"), rule_dict
            )

            operator: str | None = rule_dict.get("operator")

            formula1: str | None = rule_dict.get("value1") or rule_dict.get("formula1")
            formula2: str | None = rule_dict.get("value2") or rule_dict.get("formula2")

            if rule_dict.get("source"):
                formula1 = rule_dict.get("source")
                # Ensure list source is quoted if literal comma-separated
                if (
                    "," in str(formula1)
                    and not str(formula1).startswith('"')
                    and not str(formula1).startswith("=")
                ):
                    formula1 = f'"{formula1}"'

            if rule_dict.get("custom_formula"):
                formula1 = rule_dict.get("custom_formula")

            prompt: str | None = rule_dict.get("input_message") or rule_dict.get(
                "prompt"
            )
            title: str | None = rule_dict.get("input_title") or rule_dict.get("title")

            dv = DataValidation(
                type=dv_type,
                operator=operator,
                formula1=formula1,
                formula2=formula2,
                allow_blank=rule_dict.get("ignore_blank", True),
                showErrorMessage=rule_dict.get("show_error_message", True),
                showInputMessage=rule_dict.get("show_input_message", True),
                prompt=prompt,
                promptTitle=title,
            )
            sheet.add_data_validation(dv)

            # Apply to ranges
            target_range: str | None = rule_dict.get("range_a1")
            ranges: list[str] = (
                [target_range] if target_range else rule_dict.get("ranges", [])
            )
            if isinstance(ranges, str):
                ranges = ranges.split(",")

            for r in ranges:
                if r:
                    dv.add(r.strip())

            mwb.dv_rules_store[rule_id] = {"sheet": operation.sheet_name, "dv": dv}
            added_ids.append(rule_id)

        return ToolSuccess(ids=added_ids).to_json()
    except Exception as e:
        return ToolError(str(e)).to_json()
