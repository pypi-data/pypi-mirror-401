from __future__ import annotations

from datetime import datetime
import json
import os

import numpy as np
import pandas as pd
import polars as pl
from tqdm import tqdm

from masster._version import get_version
from masster.spectrum import combine_peaks
from masster.study.defaults import export_mgf_defaults


def _get_mgf_df(self, **kwargs):
    """
    Generate MGF data as a Polars DataFrame.

    This is the core data generation function used by export_mgf().

    Parameters:
        **kwargs: Keyword arguments for export parameters. Same as export_mgf()
                 except return_data is not relevant here.

    Returns:
        pl.DataFrame: DataFrame with columns:
            - mgf_index: MGF index
            - title: MGF title string
            - feature_id: Consensus feature ID
            - feature_uid: Feature UID (from features_df for MS2, consensus_uid for MS1)
            - charge: Charge state
            - pepmass: Precursor m/z
            - rtinseconds: Retention time in seconds
            - mslevel: MS level
            - type: Spectrum type (e.g., "MS2")
            - energy: Collision energy (if available)
            - spec_len: Number of peaks in spectrum
            - spec_mz: List of spectrum m/z values
            - spec_int: List of spectrum intensity values
    """
    # parameters initialization
    params = export_mgf_defaults()
    for key, value in kwargs.items():
        if isinstance(value, export_mgf_defaults):
            params = value
            self.logger.debug("Using provided export_defaults parameters")
        elif hasattr(params, key):
            if params.set(key, value, validate=True):
                self.logger.debug(f"Updated parameter {key} = {value}")
            else:
                self.logger.warning(
                    f"Failed to set parameter {key} = {value} (validation failed)",
                )
        else:
            self.logger.debug(f"Unknown parameter {key} ignored")
    # end of parameter initialization

    # Get parameter values for use in the method
    selection = params.get("selection")
    split_energy = params.get("split_energy")
    merge = params.get("merge")
    mz_start = params.get("mz_start")
    mz_end = params.get("mz_end")
    rt_start = params.get("rt_start")
    rt_end = params.get("rt_end")
    centroid = params.get("centroid")
    inty_min = params.get("inty_min")
    deisotope = params.get("deisotope")

    if self.consensus_df is None:
        self.logger.error("No consensus map found. Please run merge() first.")
        return None

    # MS2 data is optional - we can generate MS1 data without it
    ms2_available = self.consensus_ms2 is not None and not self.consensus_ms2.is_empty()
    if not ms2_available:
        self.logger.info("No consensus MS2 data found. Generating MS1-only MGF data.")

    # Use polars for merge operations if we have MS2 data
    if ms2_available:
        consensus_ms2_merged = self.consensus_ms2

        # Join with samples_df to get sample_name
        if self.samples_df is not None and not self.samples_df.is_empty():
            consensus_ms2_merged = consensus_ms2_merged.join(
                self.samples_df.select(["sample_id", "sample_name"]),
                on="sample_id",
                how="left",
            )

        # Join with features_df to get feature_uid
        if self.features_df is not None and not self.features_df.is_empty():
            consensus_ms2_merged = consensus_ms2_merged.join(
                self.features_df.select(["feature_id", "feature_uid"]),
                on="feature_id",
                how="left",
            )

        # Join with consensus_df
        features = consensus_ms2_merged.join(
            self.consensus_df,
            on="consensus_id",
            how="left",
        )

        if features.is_empty():
            self.logger.warning("No MS2 features found.")
            grouped = {}  # Empty groupby result
        else:
            # Group by consensus_id for fast access
            grouped = features.group_by("consensus_id", maintain_order=True)
    else:
        grouped = {}  # No MS2 data available

    def filter_peaks(spec, inty_min=None):
        spec = spec.copy()
        length = len(spec.mz)
        mask = np.ones(length, dtype=bool)
        if inty_min is not None and inty_min > 0:
            mask = mask & (spec.inty >= inty_min)
        for attr in spec.__dict__:
            arr = getattr(spec, attr)
            if (
                isinstance(arr, list | np.ndarray)
                and hasattr(arr, "__len__")
                and len(arr) == length
            ):
                setattr(spec, attr, np.array(arr)[mask])
        return spec

    def safe_charge(charge_value):
        """Safely convert charge value to integer, handling NaN and None"""
        if charge_value is None or (
            isinstance(charge_value, float) and np.isnan(charge_value)
        ):
            charge = 1
        else:
            charge = round(charge_value)

        # Replace charge 0 with polarity-based charge
        if charge == 0:
            charge = 1

        # For negative polarity, ensure charge is negative
        if (
            hasattr(self.parameters, "polarity")
            and self.parameters.polarity is not None
        ):
            if self.parameters.polarity.lower() in ["negative", "neg"] and charge > 0:
                charge = -charge

        return charge

    def create_ion_dict(title, id, uid, mz, rt, charge, spect, mgf_id):
        """Create a dictionary representing an ion for the DataFrame."""
        if spect is None:
            return None

        # Prepare spectrum data
        spectrum_mz = (
            spect.mz.tolist() if hasattr(spect.mz, "tolist") else list(spect.mz)
        )
        spectrum_inty = (
            spect.inty.tolist() if hasattr(spect.inty, "tolist") else list(spect.inty)
        )

        # Determine MS level
        ms_level = spect.ms_level if spect.ms_level is not None else 1

        # Get energy if available
        energy = getattr(spect, "energy", None)

        # Determine spectrum type based on MS level
        spec_type = f"MS{ms_level}" if ms_level > 1 else "MS1"

        # Calculate spectrum length
        spec_len = len(spectrum_mz)

        return {
            "mgf_index": mgf_id,
            "title": title,
            "feature_id": id,
            "feature_uid": uid,
            "charge": charge,
            "pepmass": mz,
            "rtinseconds": rt,
            "mslevel": ms_level,
            "type": spec_type,
            "energy": energy,
            "spec_len": spec_len,
            "spec_mz": spectrum_mz,
            "spec_int": spectrum_inty,
        }

    # Collect all ion data
    ion_data = []
    skip = 0
    mgf_counter = 0
    self.logger.debug(
        f"Generating MGF data for {len(self.consensus_df)} consensus features...",
    )

    # First, generate MS1 spectra for all consensus features using isotope data
    self.logger.debug("Generating MS1 spectra from isotope data...")
    for row in self.consensus_df.iter_rows(named=True):
        # Apply filtering at individual feature level for MS1 data
        consensus_id = row["consensus_id"]
        consensus_mz = row["mz"]
        consensus_rt = row["rt"]
        consensus_inty_mean = row.get("inty_mean", 0)

        if mz_start is not None and consensus_mz < mz_start:
            continue
        if mz_end is not None and consensus_mz > mz_end:
            continue
        if rt_start is not None and consensus_rt < rt_start:
            continue
        if rt_end is not None and consensus_rt > rt_end:
            continue

        # Create MS1 spectrum using isotope data
        iso_data = row.get("iso", None)

        if iso_data is not None and len(iso_data) > 0:
            # Use isotope data for spectrum
            spectrum_mz = [float(peak[0]) for peak in iso_data]
            spectrum_inty = [float(peak[1]) for peak in iso_data]
        else:
            # Use consensus mz and inty_mean as single peak
            spectrum_mz = [float(consensus_mz)]
            spectrum_inty = [float(consensus_inty_mean)]

        # Apply intensity minimum filter if specified
        if inty_min is not None and inty_min > 0:
            filtered_pairs = [
                (mz, inty)
                for mz, inty in zip(spectrum_mz, spectrum_inty, strict=False)
                if inty >= inty_min
            ]
            if filtered_pairs:
                mz_tuple, inty_tuple = zip(*filtered_pairs, strict=False)
                spectrum_mz = list(mz_tuple)
                spectrum_inty = list(inty_tuple)
            else:
                # If all peaks are below threshold, skip this feature
                continue

        mgf_counter += 1

        # Create MS1 spectrum object to use with create_ion_dict
        class SimpleSpectrum:
            def __init__(self, mz_list, inty_list):
                self.mz = np.array(mz_list)
                self.inty = np.array(inty_list)
                self.ms_level = 1
                self.energy = None

        ms1_spectrum = SimpleSpectrum(spectrum_mz, spectrum_inty)

        # Use create_ion_dict to ensure consistent schema
        # For MS1 spectra from consensus features, use consensus_uid
        consensus_uid = row.get("consensus_uid", row["consensus_id"])
        ion_dict = create_ion_dict(
            f"id:{consensus_id}, rt:{consensus_rt:.2f}, mz:{consensus_mz:.4f}, MS1",
            row["consensus_id"],
            consensus_uid,
            consensus_mz,
            consensus_rt,
            safe_charge(row.get("charge_mean")),
            ms1_spectrum,
            mgf_counter,
        )

        if ion_dict is not None:
            ion_data.append(ion_dict)

    self.logger.debug(f"Generated {len(ion_data)} MS1 spectra from isotope data")

    # Now generate MS2 spectra if available
    if ms2_available and grouped:
        # Convert GroupBy to list for iteration with progress bar
        grouped_list: list = list(grouped) if not isinstance(grouped, dict) else []

        if grouped_list:
            self.logger.debug(
                f"Processing MS2 data for {len(grouped_list)} consensus features with MS2...",
            )
            tdqm_disable = self.log_level not in ["TRACE", "DEBUG", "INFO"]
            for _consensus_id, cons_ms2 in tqdm(
                grouped_list,
                total=len(grouped_list),
                desc=f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]} | INFO     | {self.log_label}Feature",
                disable=tdqm_disable,
            ):
                # Use the first row for feature-level info
                if cons_ms2.is_empty():
                    skip += 1
                    continue

                row = cons_ms2.row(0, named=True)
                if mz_start is not None and row["mz"] < mz_start:
                    continue
                if mz_end is not None and row["mz"] > mz_end:
                    continue
                if rt_start is not None and row["rt"] < rt_start:
                    continue
                if rt_end is not None and row["rt"] > rt_end:
                    continue

                if split_energy:
                    energies = cons_ms2["energy"].unique().to_list()
                    for e in energies:
                        cons_ms2_e = cons_ms2.filter(pl.col("energy") == e)
                        if selection == "best":
                            # Check if the filtered DataFrame is empty
                            if cons_ms2_e.is_empty():
                                continue
                            idx = cons_ms2_e["prec_inty"].arg_max()
                            cons_ms2_e_row = cons_ms2_e.row(idx, named=True)
                            spect = cons_ms2_e_row["spec"]
                            if spect is None:
                                skip += 1
                                continue
                            if centroid:
                                spect = spect.centroid()
                            if deisotope:
                                spect = spect.deisotope()
                            spect = filter_peaks(spect, inty_min=inty_min)
                            mgf_counter += 1
                            sample_name = cons_ms2_e_row.get(
                                "sample_name",
                                f"sample_{cons_ms2_e_row['sample_id']}",
                            )
                            feature_uid = cons_ms2_e_row.get(
                                "feature_uid",
                                cons_ms2_e_row["feature_id"],
                            )
                            ion_dict = create_ion_dict(
                                f"id:{cons_ms2_e_row['consensus_id']}, rt:{cons_ms2_e_row['rt']:.2f}, mz:{cons_ms2_e_row['mz']:.4f}, energy:{e}, sample_id:{cons_ms2_e_row['sample_id']}, scan_id:{cons_ms2_e_row['scan_id']}, {sample_name}",
                                cons_ms2_e_row["consensus_id"],
                                feature_uid,
                                cons_ms2_e_row["mz"],
                                cons_ms2_e_row["rt"],
                                safe_charge(cons_ms2_e_row["charge_mean"]),
                                spect,
                                mgf_counter,
                            )
                            if ion_dict is not None:
                                ion_data.append(ion_dict)
                        else:
                            for row_e in cons_ms2_e.iter_rows(named=True):
                                spect = row_e["spec"]
                                if spect is None:
                                    continue
                                if centroid:
                                    spect = spect.centroid()
                                if deisotope:
                                    spect = spect.deisotope()
                                spect = filter_peaks(spect, inty_min=inty_min)
                                mgf_counter += 1
                                sample_name = (
                                    row_e["sample_name"]
                                    if "sample_name" in cons_ms2_e.columns
                                    else f"sample_{row_e['sample_id']}"
                                )
                                feature_uid = row_e.get(
                                    "feature_uid",
                                    row_e["feature_id"],
                                )
                                ion_dict = create_ion_dict(
                                    f"id:{row_e['consensus_id']}, rt:{row_e['rt']:.2f}, mz:{row_e['mz']:.4f}, energy:{e}, sample_id:{row_e['sample_id']}, scanid:{row_e['scan_id']}, {sample_name}",
                                    row_e["consensus_id"],
                                    feature_uid,
                                    row_e["mz"],
                                    row_e["rt"],
                                    safe_charge(row_e["charge_mean"]),
                                    spect,
                                    mgf_counter,
                                )
                            if ion_dict is not None:
                                ion_data.append(ion_dict)
                elif selection == "best":
                    idx = cons_ms2["prec_inty"].arg_max()
                    cons_ms2_e_row = cons_ms2.row(idx, named=True)
                    spect = cons_ms2_e_row["spec"]
                    if spect is None:
                        continue
                    if centroid:
                        spect = spect.centroid()
                    if deisotope:
                        spect = spect.deisotope()
                    spect = filter_peaks(spect, inty_min=inty_min)
                    mgf_counter += 1
                    sample_name = cons_ms2_e_row.get(
                        "sample_name",
                        f"sample_{cons_ms2_e_row['sample_id']}",
                    )
                    feature_uid = cons_ms2_e_row.get(
                        "feature_uid",
                        cons_ms2_e_row["feature_id"],
                    )
                    ion_dict = create_ion_dict(
                        f"id:{cons_ms2_e_row['consensus_id']}, rt:{cons_ms2_e_row['rt']:.2f}, mz:{cons_ms2_e_row['mz']:.4f}, energy:{cons_ms2_e_row['energy']}, sample_id:{cons_ms2_e_row['sample_id']}, scan_id:{cons_ms2_e_row['scan_id']}, {sample_name}",
                        cons_ms2_e_row["consensus_id"],
                        feature_uid,
                        cons_ms2_e_row["mz"],
                        cons_ms2_e_row["rt"],
                        safe_charge(cons_ms2_e_row["charge_mean"]),
                        spect,
                        mgf_counter,
                    )
                    if ion_dict is not None:
                        ion_data.append(ion_dict)

                elif selection == "all":
                    if merge:
                        specs = [
                            row_e["spec"]
                            for row_e in cons_ms2.iter_rows(named=True)
                            if row_e["spec"] is not None
                        ]
                        if not specs:
                            continue
                        spect = combine_peaks(specs)
                        if centroid:
                            spect = spect.denoise()
                            spect = spect.centroid()
                        if deisotope:
                            spect = spect.deisotope()
                        spect = filter_peaks(spect, inty_min=inty_min)
                        mgf_counter += 1
                        feature_uid = row.get("feature_uid", row["feature_id"])
                        ion_dict = create_ion_dict(
                            f"id:{row['consensus_id']}, rt:{row['rt']:.2f}, mz:{row['mz']:.4f}, sample_id:{row['sample_id']}, scan_id:{row['scan_id']}",
                            row["consensus_id"],
                            feature_uid,
                            row["mz"],
                            row["rt"],
                            safe_charge(row["charge_mean"]),
                            spect,
                            mgf_counter,
                        )
                        if ion_dict is not None:
                            ion_data.append(ion_dict)
                    else:
                        for row_e in cons_ms2.iter_rows(named=True):
                            spect = row_e["spec"]
                            if spect is None:
                                continue
                            if centroid:
                                spect = spect.centroid()
                            if deisotope:
                                spect = spect.deisotope()
                            spect = filter_peaks(spect, inty_min=inty_min)
                            mgf_counter += 1
                            feature_uid = row_e.get("feature_uid", row_e["feature_id"])
                            ion_dict = create_ion_dict(
                                f"id:{row_e['consensus_id']}, rt:{row_e['rt']:.2f}, mz:{row_e['mz']:.4f}, energy:{row_e['energy']}, sample_id:{row_e['sample_id']}, scan_id:{row_e['scan_id']}",
                                row_e["consensus_id"],
                                feature_uid,
                                row_e["mz"],
                                row_e["rt"],
                                safe_charge(row_e["charge_mean"]),
                                spect,
                                mgf_counter,
                            )
                            if ion_dict is not None:
                                ion_data.append(ion_dict)
        else:
            self.logger.info(
                "Skipping MS2 data generation - no MS2 features in grouped data",
            )
    else:
        self.logger.info("Skipping MS2 data generation - no MS2 data available")

    self.logger.debug(f"Generated MGF data for {len(ion_data)} spectra (MS1 + MS2)")
    self.logger.debug(f"Skipped {skip} MS2 features due to missing data.")

    # Convert to Polars DataFrame
    if not ion_data:
        return pl.DataFrame()

    return pl.DataFrame(ion_data, infer_schema_length=None)


def export_mgf(self, **kwargs) -> None:
    """
    Export consensus features as MGF format for database searching.

    Parameters:
        **kwargs: Keyword arguments for export parameters. Can include:
            - An export_defaults instance to set all parameters at once
            - Individual parameter names and values (see export_defaults for details)

    Key Parameters:
        filename (str): Output MGF file name (default: "consensus.mgf").
        selection (str): "best" for first scan, "all" for every scan (default: "best").
        split_energy (bool): Process MS2 scans by unique energy (default: True).
        merge (bool): If selection="all", merge MS2 scans into one spectrum (default: False).
        mz_start (float): Minimum m/z for feature selection (default: None).
        mz_end (float): Maximum m/z for feature selection (default: None).
        rt_start (float): Minimum RT for feature selection (default: None).
        rt_end (float): Maximum RT for feature selection (default: None).
        centroid (bool): Apply centroiding to spectra (default: True).
        inty_min (float): Minimum intensity threshold (default: None).
        deisotope (bool): Apply deisotoping to spectra (default: True).
        verbose (bool): Enable verbose logging (default: False).
        precursor_trim (float): Precursor trimming value (default: -10).
        centroid_algo (str): Centroiding algorithm (default: "lmp").

    Returns:
        None: Writes MGF file to disk.
    """
    # Get mgf data as DataFrame
    from masster.study.export import _get_mgf_df

    mgf_data = _get_mgf_df(self, **kwargs)

    if mgf_data is None or len(mgf_data) == 0:
        self.logger.warning("No MGF data generated.")
        return

    # Get filename from parameters
    params = export_mgf_defaults()
    for key, value in kwargs.items():
        if isinstance(value, export_mgf_defaults):
            params = value
        elif hasattr(params, key):
            params.set(key, value, validate=True)

    filename = params.get("filename")

    # Prepare output path
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Write MGF file
    with open(filename, "w", encoding="utf-8") as f:
        for row in mgf_data.iter_rows(named=True):
            # Write BEGIN IONS
            f.write("BEGIN IONS\n")

            # Write metadata
            if row["mgf_index"] is not None:
                f.write(f"INDEX={row['mgf_index']}\n")
            f.write(f"TITLE={row['title']}\n")
            f.write(f"FEATURE_ID={row['feature_uid']}\n")
            # Format charge: positive as "1+", "2+", negative as "1-", "2-"
            charge = row["charge"]
            if charge < 0:
                charge_str = f"{abs(charge)}-"
            else:
                charge_str = f"{charge}+"
            f.write(f"CHARGE={charge_str}\n")
            f.write(f"PEPMASS={row['pepmass']}\n")
            f.write(f"RTINSECONDS={row['rtinseconds']}\n")
            f.write(f"MSLEVEL={row['mslevel']}\n")

            if row["energy"] is not None:
                # Always use absolute value for energy
                energy_val = abs(row["energy"])
                f.write(f"ENERGY={energy_val}\n")

            # Write spectrum data
            spectrum_mz = row["spec_mz"]
            spectrum_inty = row["spec_int"]
            for mz_val, inty in zip(spectrum_mz, spectrum_inty, strict=False):
                f.write(f"{mz_val:.5f} {inty:.0f}\n")

            # Write END IONS
            f.write("END IONS\n\n")

    # Log statistics from MGF data generation
    # Count spectra by type
    ms1_count = len(
        [row for row in mgf_data.iter_rows(named=True) if row["mslevel"] == 1],
    )
    ms2_count = len(mgf_data) - ms1_count

    if ms1_count > 0 and ms2_count > 0:
        self.logger.debug(f"Exported {ms1_count} MS1 and {ms2_count} MS2 spectra")
    elif ms1_count > 0:
        self.logger.debug(f"Exported {ms1_count} MS1 spectra")
    elif ms2_count > 0:
        self.logger.debug(f"Exported {ms2_count} MS2 spectra")

    self.logger.info(f"Exported {len(mgf_data)} spectra to {filename}")


def export_mztab(self, filename: str | None = None, include_mgf=True, **kwargs) -> None:
    """
    Export the study as a fully compliant mzTab-M file.

    Args:
        filename (str, optional): Path to the output mzTab-M file.
        title (str, optional): Human-readable title for the file.
        description (str, optional): Human-readable description.
        **kwargs: Additional metadata or export options.
    """

    def safe_str(value, default="null"):
        """Convert value to string, replacing empty strings with 'null'"""
        if value is None:
            return default
        str_val = str(value)
        return str_val if str_val.strip() != "" else default

    if filename is None:
        filename = "study.mztab"
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Get identification data if available
    top_id_data = None
    full_id_data = None
    try:
        # Import here to avoid circular imports
        from masster.study.id import get_id

        # Get full enriched identification data for SME section
        full_id_data = get_id(self)
        if full_id_data is not None and not full_id_data.is_empty():
            # Get top scoring identification for each consensus_id for SML section
            top_id_data = (
                full_id_data.group_by("consensus_id")
                .agg(pl.all().sort_by("score", descending=True).first())
                .sort("consensus_id")
            )
            # Keep raw id_data for backward compatibility (if needed elsewhere)
            self.id_df if hasattr(self, "id_df") and self.id_df is not None else None
        else:
            self.logger.info("No identification data available for mzTab export")
    except Exception as e:
        self.logger.debug(f"Could not retrieve identification data: {e}")
        top_id_data = None
        full_id_data = None

    # get mgf data only if requested
    mgf_data = None
    mgf_mapping: dict[str, list[int]] = {}
    if include_mgf:
        from masster.study.export import _get_mgf_df

        mgf_data = _get_mgf_df(self, **kwargs)
        # Create mapping from feature_id to MGF indexes
        if mgf_data is not None and len(mgf_data) > 0:
            for row in mgf_data.iter_rows(named=True):
                feature_id = row["feature_id"]
                mgf_index = row["mgf_index"]
                if feature_id not in mgf_mapping:
                    mgf_mapping[feature_id] = []
                mgf_mapping[feature_id].append(mgf_index)

    # --- Prepare MTD (metadata) section ---
    mtd_lines = []
    mtd_lines.append(
        f"COM\tfile generated by MASSter {get_version()} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    mtd_lines.append("\nMTD\tmzTab-version\t2.2.0-M")
    id = self.label if self.label else self.folder
    mtd_lines.append(f"MTD\tmzTab-id\t{id}")
    mtd_lines.append("")
    mtd_lines.append("MTD\tcv[1]-label\tMS")
    mtd_lines.append("MTD\tcv[1]-full_name\tPSI-MS controlled vocabulary")
    mtd_lines.append("MTD\tcv[1]-version\t4.1.199")
    mtd_lines.append(
        "MTD\tcv[1]-uri\thttps://raw.githubusercontent.com/HUPO-PSI/psi-ms-CV/master/psi-ms.obo",
    )
    mtd_lines.append("")
    mtd_lines.append(
        "MTD\tsmall_molecule-quantification_unit\t[MS, MS:1001844, MS1 feature area, ]",
    )
    mtd_lines.append(
        "MTD\tsmall_molecule_feature-quantification_unit\t[MS, MS:1001844, MS1 feature area, ]",
    )
    mtd_lines.append(
        "MTD\tsmall_molecule-identification_reliability\t[MS, MS:1002955, hr-ms compound identification confidence level, ]",
    )

    # Add identification confidence measures if identification data is available
    if full_id_data is not None:
        mtd_lines.append(
            "MTD\tid_confidence_measure[1]\t[MS, MS:1002888, small molecule confidence measure, ]",
        )
    else:
        mtd_lines.append(
            "MTD\tid_confidence_measure[1]\t[MS, MS:1002888, small molecule confidence measure, ]",
        )

    mtd_lines.append("")
    mtd_lines.append("MTD\tsoftware[1]\t[MS, MS:1003430, OpenMS, unknown]")
    mtd_lines.append(f"MTD\tsoftware[2]\t[MS, MS:1002878, MASSter, {get_version()}]")
    mtd_lines.append(
        "MTD\tquantification_method\t[MS, MS:1001834, LC-MS label-free quantitation analysis, ]",
    )
    mtd_lines.append("")

    # Database information - updated based on identification data
    if (
        full_id_data is not None
        and hasattr(self, "lib_df")
        and self.lib_df is not None
        and not self.lib_df.is_empty()
    ):
        mtd_lines.append('MTD\tdatabase[1]\t[, , "compound library", ]')
        mtd_lines.append("MTD\tdatabase[1]-prefix\tcmpd")
        mtd_lines.append("MTD\tdatabase[1]-version\tUnknown")
        mtd_lines.append("MTD\tdatabase[1]-uri\thttps://pubchem.ncbi.nlm.nih.gov/")
    else:
        mtd_lines.append('MTD\tdatabase[1]\t[, , "PubChem", ]')
        mtd_lines.append("MTD\tdatabase[1]-prefix\tCID")
        mtd_lines.append("MTD\tdatabase[1]-version\tUnknown")
        mtd_lines.append("MTD\tdatabase[1]-uri\thttps://pubchem.ncbi.nlm.nih.gov/")

    # Get abundance matrix to determine the number of assays needed
    abundance_matrix = self.get_consensus_matrix()

    # Get sample columns (excluding consensus_id)
    sample_columns = [col for col in abundance_matrix.columns if col != "consensus_id"]
    n_assays = len(sample_columns)

    # Load schema to identify protected columns
    schema_path = os.path.join(os.path.dirname(__file__), "study5_schema.json")
    protected_columns = set()
    if os.path.exists(schema_path):
        try:
            import json

            with open(schema_path) as f:
                schema = json.load(f)
                if "samples_df" in schema and "columns" in schema["samples_df"]:
                    protected_columns = set(schema["samples_df"]["columns"].keys())
        except Exception:
            # If schema loading fails, use default protected columns
            protected_columns = {
                "sample_id",
                "sample_name",
                "sample_path",
                "sample_source",
                "sample_type",
                "sample_group",
                "sample_batch",
                "sample_sequence",
                "sample_color",
                "num_features",
                "num_ms1",
                "num_ms2",
            }

    # Identify metadata columns (columns not in schema)
    all_sample_columns = set(self.samples_df.columns)
    metadata_columns = sorted(all_sample_columns - protected_columns)

    # Create a mapping from sample names to their metadata
    samples_pd = self.samples_df.to_pandas()
    sample_metadata = {}
    for _, row in samples_pd.iterrows():
        sample_name = str(row["sample_name"])
        description_parts = []

        # Add sample_id
        if "sample_id" in row and pd.notna(row["sample_id"]):
            description_parts.append(f"uid:{row['sample_id']}")

        # Add sample_name
        if "sample_name" in row and pd.notna(row["sample_name"]):
            description_parts.append(f"name:{row['sample_name']}")

        # Add sample_group
        if "sample_group" in row and pd.notna(row["sample_group"]):
            description_parts.append(f"group:{row['sample_group']}")

        # Add sample_batch
        if "sample_batch" in row and pd.notna(row["sample_batch"]):
            description_parts.append(f"batch:{row['sample_batch']}")

        # Add sample_sequence
        if "sample_sequence" in row and pd.notna(row["sample_sequence"]):
            description_parts.append(f"sequence:{row['sample_sequence']}")

        # Add all additional metadata columns
        for col in metadata_columns:
            if col in row and pd.notna(row[col]):
                # Convert to UTF8 string
                value = str(row[col])
                # Escape pipe characters in values to avoid conflicts
                value = value.replace("|", "\\|")
                description_parts.append(f"{col}:{value}")

        sample_metadata[sample_name] = " | ".join(description_parts)

    # Define samples, ms_runs, and assays based on the abundance matrix columns
    # Determine scan polarity based on study polarity
    study_polarity = getattr(self.parameters, "polarity", "positive")
    if study_polarity in ["negative", "neg"]:
        scan_polarity_cv = "[MS, MS:1000129, negative scan, ]"
    else:
        scan_polarity_cv = "[MS, MS:1000130, positive scan, ]"

    for i, sample_col in enumerate(sample_columns, 1):
        mtd_lines.append(f"\nMTD\tsample[{i}]\t{sample_col}")
        # Use the enriched description from sample_metadata
        description = sample_metadata.get(sample_col, sample_col)
        mtd_lines.append(f"MTD\tsample[{i}]-description\t{description}")
        mtd_lines.append(f"MTD\tms_run[{i}]-location\tfile://unknown")
        mtd_lines.append(f"MTD\tms_run[{i}]-scan_polarity\t{scan_polarity_cv}")
        mtd_lines.append(f"MTD\tassay[{i}]\tAssay_{i}")
        mtd_lines.append(f"MTD\tassay[{i}]-sample_ref\tsample[{i}]")
        mtd_lines.append(f"MTD\tassay[{i}]-ms_run_ref\tms_run[{i}]")
    mtd_lines.append("")
    mtd_lines.append("MTD\tstudy_variable[1]\tundefined")
    assay_refs = "|".join([f"assay[{i}]" for i in range(1, len(self.samples_df) + 1)])
    mtd_lines.append(f"MTD\tstudy_variable[1]-assay_refs\t{assay_refs}")
    with open(filename, "w", encoding="utf-8") as f:
        for line in mtd_lines:
            f.write(line + "\n")

    # --- SML (Small Molecule) table ---
    sml_lines = []
    sml_header = [
        "SMH",
        "SML_ID",
        "SMF_ID_REFS",
        "database_identifier",
        "chemical_formula",
        "smiles",
        "inchi",
        "chemical_name",
        "uri",
        "theoretical_neutral_mass",
        "adduct_ions",
        "reliability",
        "best_id_confidence_measure",
        "best_id_confidence_value",
        "opt_global_mgf_index",
    ]

    # round to int - handle both Polars and Pandas DataFrames
    if hasattr(abundance_matrix, "with_columns"):
        # Polars DataFrame
        numeric_cols = [
            col
            for col in abundance_matrix.columns
            if abundance_matrix[col].dtype.is_numeric()
        ]
        abundance_matrix = abundance_matrix.with_columns(
            [abundance_matrix[col].round(0) for col in numeric_cols],
        )
    else:
        # Pandas DataFrame
        abundance_matrix = abundance_matrix.round(0)

    # Use the n_assays already calculated from abundance matrix columns
    sml_header += [f"abundance_assay[{i}]" for i in range(1, n_assays + 1)]
    sml_header += [
        "abundance_study_variable[1]",
        "abundance_variation_study_variable[1]",
    ]
    sml_lines.append("\t".join(sml_header))

    # get adducts from consensus_df['adduct_top'] - use the top-ranked adduct directly
    adduct_list = []
    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        adduct = "null"
        # Use adduct_top if available, otherwise fall back to null
        if "adduct_top" in row and row["adduct_top"] is not None:
            adduct = str(row["adduct_top"])
            # Replace ? with H for better mzTab compatibility
            adduct = adduct.replace("?", "H")

        adduct_list.append(adduct)

    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        # Get identification information for this consensus_id if available
        consensus_id = row["consensus_id"]
        id_info = None
        if top_id_data is not None:
            id_matches = top_id_data.filter(pl.col("consensus_id") == consensus_id)
            if id_matches.height > 0:
                id_info = id_matches.row(0, named=True)

        # Populate identification fields
        database_identifier = "null"
        chemical_formula = "null"
        smiles_val = "null"
        inchi_val = "null"
        chemical_name = "null"
        best_id_confidence_measure = "null"
        best_id_confidence_value = "null"
        reliability = "4"  # Default: unknown compound
        theoretical_neutral_mass = (
            "null"  # Only set when we have database identification
        )

        if id_info:
            # Use cmpd_id as database identifier with prefix
            if id_info.get("cmpd_id") is not None:
                database_identifier = f"cmpd:{id_info['cmpd_id']}"

            # Chemical formula
            if id_info.get("formula") is not None and id_info["formula"] != "":
                chemical_formula = safe_str(id_info["formula"])

            # SMILES
            if id_info.get("smiles") is not None and id_info["smiles"] != "":
                smiles_val = safe_str(id_info["smiles"])

            # InChI
            if id_info.get("inchi") is not None and id_info["inchi"] != "":
                inchi_val = safe_str(id_info["inchi"])

            # Chemical name
            if id_info.get("name") is not None and id_info["name"] != "":
                chemical_name = safe_str(id_info["name"])

            # Theoretical neutral mass - only from identification data, not consensus_df
            if id_info.get("neutral_mass") is not None:
                theoretical_neutral_mass = safe_str(id_info["neutral_mass"])
            elif id_info.get("mass") is not None:
                theoretical_neutral_mass = safe_str(id_info["mass"])

            # Identification confidence
            if id_info.get("matcher") is not None:
                best_id_confidence_measure = f"[MS, MS:1002888, {id_info['matcher']}, ]"

            if id_info.get("score") is not None:
                best_id_confidence_value = safe_str(id_info["score"])

            # Set reliability based on identification quality
            # Using mzTab-M hr-ms identification levels: 2a=compound match, 2b=library spectrum match, 3=compound class, 4=unknown
            if id_info.get("score", 0) >= 0.8:
                reliability = "2a"  # High confidence compound match
            elif id_info.get("score", 0) >= 0.5:
                reliability = "2b"  # Moderate confidence match
            elif id_info.get("score", 0) >= 0.2:
                reliability = "3"  # Compound class level
            else:
                reliability = "4"  # Unknown compound

        # Get MGF indexes for this consensus feature
        mgf_indexes = mgf_mapping.get(row["consensus_id"], [])

        sml_row = [
            "SML",
            str(idx),
            str(idx),
            database_identifier,
            chemical_formula,
            smiles_val,
            inchi_val,
            chemical_name,
            safe_str(row.get("uri", "null")),
            theoretical_neutral_mass,  # Only set when database_identifier is not null
            adduct_list[idx - 1],
            reliability,
            best_id_confidence_measure,
            best_id_confidence_value,
            ",".join(map(str, mgf_indexes)) if mgf_indexes else "null",
        ]
        # Add abundance values for each assay
        consensus_id = row["consensus_id"]
        # Check if consensus_id exists in the abundance_matrix (Polars)
        filtered_matrix = abundance_matrix.filter(
            pl.col("consensus_id") == consensus_id,
        )
        if filtered_matrix.height > 0:
            # Get the first (and should be only) matching row
            abundance_row = filtered_matrix.row(0, named=True)
            # Extract values excluding the consensus_id column
            abundance_values = [
                abundance_row[col]
                for col in abundance_matrix.columns
                if col != "consensus_id"
            ]
            sml_row += [
                safe_str(val) if val is not None else "null" for val in abundance_values
            ]

            # Calculate study variable statistics
            non_null_values = [val for val in abundance_values if val is not None]
            if non_null_values:
                abundance_study_variable = sum(non_null_values) / len(non_null_values)
                abundance_variation_study_variable = (
                    (
                        sum(
                            (x - abundance_study_variable) ** 2 for x in non_null_values
                        )
                        / len(non_null_values)
                    )
                    ** 0.5
                    if len(non_null_values) > 1
                    else 0
                )
            else:
                abundance_study_variable = "null"
                abundance_variation_study_variable = "null"

            sml_row += [
                safe_str(abundance_study_variable),
                safe_str(abundance_variation_study_variable),
            ]
        else:
            sml_row += ["null"] * n_assays
            sml_row += ["null", "null"]  # Study variable columns
        sml_lines.append("\t".join(sml_row))
    with open(filename, "a", encoding="utf-8") as f:
        f.write("\n")
        for line in sml_lines:
            f.write(line + "\n")

    # --- SMF (Small Molecule Feature) table ---
    smf_lines = []
    smf_header = [
        "SFH",
        "SMF_ID",
        "SME_ID_REFS",
        "SME_ID_REF_ambiguity_code",
        "adduct_ion",
        "isotopomer",
        "exp_mass_to_charge",
        "charge",
        "retention_time_in_seconds",
        "retention_time_in_seconds_start",
        "retention_time_in_seconds_end",
    ]
    smf_header += [f"abundance_assay[{i}]" for i in range(1, n_assays + 1)]
    smf_header += [
        "abundance_study_variable[1]",
        "abundance_variation_study_variable[1]",
    ]
    smf_lines.append("\t".join(smf_header))

    # SMF table uses the same consensus features as SML, just different metadata
    for idx, row in enumerate(self.consensus_df.iter_rows(named=True), 1):
        # References to SME entries - each SMF can reference multiple SME entries for the same consensus_id
        SME_refs = "null"
        SME_ambiguity = "null"
        consensus_id = row["consensus_id"]

        if full_id_data is not None:
            # Find all SME entries for this consensus_id
            SME_matches = full_id_data.filter(pl.col("consensus_id") == consensus_id)
            if SME_matches.height > 0:
                # Generate SME IDs - we'll create a mapping in the SME section
                # For now, use a simple approach based on consensus_id and lib_id
                SME_ids = []
                for i, SME_row in enumerate(SME_matches.iter_rows(named=True)):
                    # Create a unique SME ID based on consensus_id and position
                    SME_id_base = (
                        consensus_id * 1000
                    )  # Ensure uniqueness across consensus features
                    SME_id = SME_id_base + i + 1
                    SME_ids.append(str(SME_id))

                if SME_ids:
                    SME_refs = "|".join(SME_ids)
                    # Set ambiguity code: 1=ambiguous identification, 2=multiple evidence same molecule, 3=both
                    if len(SME_ids) > 1:
                        # Check if all identifications point to the same compound
                        unique_cmpds = {
                            match["cmpd_id"]
                            for match in SME_matches.iter_rows(named=True)
                            if match.get("cmpd_id") is not None
                        }
                        if len(unique_cmpds) > 1:
                            SME_ambiguity = "1"  # Ambiguous identification
                        else:
                            SME_ambiguity = "2"  # Multiple evidence for same molecule
                    else:
                        SME_ambiguity = "null"

        # Format isotopomer according to mzTab-M specification
        iso_value = row.get("iso_mean", 0)
        if iso_value is not None and round(iso_value) != 0:
            isotopomer = f'[MS,MS:1002957,"isotopomer MS peak","+{round(iso_value)}"]'
        else:
            isotopomer = "null"

        smf_row = [
            "SMF",
            str(idx),
            SME_refs,
            SME_ambiguity,
            adduct_list[idx - 1],  # adduct_ion
            isotopomer,  # isotopomer formatted according to mzTab-M specification
            safe_str(row.get("mz", "null")),  # exp_mass_to_charge
            safe_str(
                f"{abs(row.get('adduct_charge_top', 1))}+"
                if row.get("adduct_charge_top", 1) >= 0
                else f"{abs(row.get('adduct_charge_top', 1))}-",
            )
            if row.get("adduct_charge_top") is not None
            and row.get("adduct_charge_top") != "null"
            else "null",  # Use top-ranked adduct charge with sign
            safe_str(row.get("rt", "null")),  # retention_time_in_seconds
            safe_str(row.get("retention_time_in_seconds_start", "null")),
            safe_str(row.get("retention_time_in_seconds_end", "null")),
        ]
        # Add abundance values for each assay - same as SML (Polars)
        consensus_id = row["consensus_id"]
        filtered_matrix = abundance_matrix.filter(
            pl.col("consensus_id") == consensus_id,
        )
        if filtered_matrix.height > 0:
            # Get the first (and should be only) matching row
            abundance_row = filtered_matrix.row(0, named=True)
            # Extract values excluding the consensus_id column
            abundance_values = [
                abundance_row[col]
                for col in abundance_matrix.columns
                if col != "consensus_id"
            ]
            abundance_strings = [
                safe_str(val) if val is not None else "null" for val in abundance_values
            ]
            smf_row += abundance_strings

            # Calculate study variable statistics (same as in SML section)
            non_null_values = [val for val in abundance_values if val is not None]
            if non_null_values:
                abundance_study_variable = sum(non_null_values) / len(non_null_values)
                abundance_variation_study_variable = (
                    (
                        sum(
                            (x - abundance_study_variable) ** 2 for x in non_null_values
                        )
                        / len(non_null_values)
                    )
                    ** 0.5
                    if len(non_null_values) > 1
                    else 0
                )
            else:
                abundance_study_variable = "null"
                abundance_variation_study_variable = "null"

            smf_row += [
                safe_str(abundance_study_variable),
                safe_str(abundance_variation_study_variable),
            ]
        else:
            smf_row += ["null"] * n_assays
            smf_row += ["null", "null"]  # Study variable columns
        smf_lines.append("\t".join(smf_row))
    with open(filename, "a", encoding="utf-8") as f:
        f.write("\n")
        for line in smf_lines:
            f.write(line + "\n")

    # --- SME (Small Molecule Evidence) table ---
    if full_id_data is not None and not full_id_data.is_empty():
        SME_lines = []
        # Add comment about spectra_ref being dummy placeholders
        SME_lines.append(
            "COM\tThe spectra_ref are dummy placeholders, as the annotation was based on aggregated data",
        )
        SME_header = [
            "SEH",
            "SME_ID",
            "evidence_input_id",
            "database_identifier",
            "chemical_formula",
            "smiles",
            "inchi",
            "chemical_name",
            "uri",
            "derivatized_form",
            "adduct_ion",
            "exp_mass_to_charge",
            "charge",
            "theoretical_mass_to_charge",
            "spectra_ref",
            "identification_method",
            "ms_level",
            "id_confidence_measure[1]",
            "rank",
        ]
        SME_lines.append("\t".join(SME_header))

        # Create SME entries for all identification results using enriched data
        for consensus_id in (
            self.consensus_df.select("consensus_id").to_series().unique()
        ):
            # Get consensus feature data for this consensus_id
            consensus_feature_data = self.consensus_df.filter(
                pl.col("consensus_id") == consensus_id,
            )
            if consensus_feature_data.height == 0:
                continue
            consensus_row = consensus_feature_data.row(0, named=True)

            # Get all identification results for this consensus feature from enriched data
            SME_matches = full_id_data.filter(pl.col("consensus_id") == consensus_id)

            if SME_matches.height > 0:
                # Sort by score descending to maintain rank order
                SME_matches = SME_matches.sort("score", descending=True)

                for i, SME_row in enumerate(SME_matches.iter_rows(named=True)):
                    # Generate unique SME_ID
                    SME_id_base = consensus_id * 1000
                    SME_id = SME_id_base + i + 1

                    # Create evidence input ID using consensus_id:mz:rt format
                    consensus_mz = consensus_row.get("mz", 0)
                    consensus_rt = consensus_row.get("rt", 0)
                    evidence_id = f"consensus_id={consensus_id}:mz={consensus_mz:.4f}:rt={consensus_rt:.2f}"

                    # Database identifier - use db_id if available, otherwise fallback to cmpd_id
                    db_id = "null"
                    if SME_row.get("db_id") is not None and SME_row["db_id"] != "":
                        db_id = safe_str(SME_row["db_id"])
                    elif SME_row.get("cmpd_id") is not None:
                        db_id = f"cmpd:{SME_row['cmpd_id']}"

                    # Get adduct information
                    adduct_ion = "null"
                    if SME_row.get("adduct") is not None and SME_row["adduct"] != "":
                        adduct_ion = safe_str(SME_row["adduct"])
                        # Replace ? with H for better mzTab compatibility
                        adduct_ion = adduct_ion.replace("?", "H")

                    # Spectra reference - reference to first ms_run with spectrum index 0
                    spectra_ref = "ms_run[1]:spectrum=0"

                    # Identification method
                    id_method = "[MS, MS:1002888, small molecule confidence measure, ]"
                    if SME_row.get("matcher") is not None:
                        id_method = f"[MS, MS:1002888, {SME_row['matcher']}, ]"

                    # MS level - check if ms1 exists in matched
                    if "ms1" in SME_row["matcher"].lower():
                        ms_level = "[MS, MS:1000511, ms level, 1]"
                    else:
                        ms_level = "[MS,MS:1000511, ms level, 2]"

                    # Experimental mass-to-charge from consensus feature
                    exp_mz = safe_str(consensus_mz)

                    # Theoretical mass-to-charge from lib_df
                    theoretical_mz = "null"
                    if (
                        SME_row.get("mz") is not None
                    ):  # This comes from lib_df via get_id() join
                        theoretical_mz = safe_str(SME_row["mz"])

                    SME_line = [
                        "SME",
                        str(SME_id),
                        evidence_id,
                        db_id,
                        safe_str(SME_row.get("formula", "null")),
                        safe_str(SME_row.get("smiles", "null")),
                        safe_str(SME_row.get("inchi", "null")),
                        safe_str(SME_row.get("name", "null")),
                        "null",  # uri - not available in current data
                        "null",  # derivatized_form
                        adduct_ion,
                        exp_mz,  # experimental m/z from consensus feature
                        safe_str(
                            f"{abs(consensus_row.get('adduct_charge_top', 1))}+"
                            if consensus_row.get("adduct_charge_top", 1) >= 0
                            else f"{abs(consensus_row.get('adduct_charge_top', 1))}-",
                        )
                        if consensus_row.get("adduct_charge_top") is not None
                        and consensus_row.get("adduct_charge_top") != "null"
                        else "1+",  # Use consensus feature's top adduct charge with sign
                        theoretical_mz,  # theoretical m/z from lib_df
                        spectra_ref,
                        id_method,
                        ms_level,
                        safe_str(SME_row.get("score", "null")),
                        str(i + 1),  # rank within this consensus feature
                    ]
                    SME_lines.append("\t".join(SME_line))

        # Write SME table
        with open(filename, "a", encoding="utf-8") as f:
            f.write("\n")
            for line in SME_lines:
                f.write(line + "\n")

    # --- MGF table ---
    if include_mgf and mgf_data is not None and len(mgf_data) > 0:
        mgf_lines = []
        # Header
        mgf_header = [
            "COM",
            "MGH",
            "mgf_id",
            "prec_id",
            "prec_rt",
            "prec_mz",
            "prec_int",
            "energy",
            "level",
            "title",
            "spec_tic",
            "spec_len",
            "spec_mz",
            "spec_int",
        ]
        mgf_lines.append("\t".join(mgf_header))

        # Data rows
        for row in mgf_data.iter_rows(named=True):
            # Calculate spectrum TIC (total ion current) from the spectrum data
            spectrum_mz = row["spec_mz"]
            spectrum_inty = row["spec_int"]
            spec_tic = sum(spectrum_inty) if spectrum_inty else 0
            spec_len = row["spec_len"] if row["spec_len"] is not None else 0

            # Format spectrum data as pipe-separated strings
            spec_mz_str = (
                "|".join([f"{mz:.4f}" for mz in spectrum_mz]) if spectrum_mz else ""
            )
            spec_int_str = (
                "|".join([f"{int(inty)}" for inty in spectrum_inty])
                if spectrum_inty
                else ""
            )

            mgf_row = [
                "COM",
                "MGF",
                str(row["mgf_index"]) if row["mgf_index"] is not None else "null",
                str(row["feature_id"]) if row["feature_id"] is not None else "null",
                f"{row['rtinseconds']:.2f}"
                if row["rtinseconds"] is not None
                else "null",
                f"{row['pepmass']:.4f}" if row["pepmass"] is not None else "null",
                "null",  # prec_int - not available in current data
                str(row["energy"]) if row["energy"] is not None else "null",
                str(row["mslevel"]) if row["mslevel"] is not None else "null",
                str(row["title"]) if row["title"] is not None else "null",
                f"{int(spec_tic)}" if spec_tic > 0 else "null",
                str(spec_len) if spec_len > 0 else "null",
                spec_mz_str if spec_mz_str else "null",
                spec_int_str if spec_int_str else "null",
            ]
            mgf_lines.append("\t".join(mgf_row))

        # Write MGF table
        with open(filename, "a", encoding="utf-8") as f:
            f.write("\n")
            for line in mgf_lines:
                f.write(line + "\n")

    self.logger.info(f"Exported mzTab-M to {filename}")


def export_excel(
    self,
    filename: str | None = None,
    worksheets: list[str] | None = None,
) -> None:
    """Export study data to multi-sheet Excel workbook.

    Generates comprehensive Excel file containing samples metadata, consensus
    features, identifications, matrices, and processing history.

    Args:
        filename (str | None): Output Excel file path. Defaults to "study.xlsx"
            in study folder. Defaults to None.
        worksheets (list[str] | None): Worksheet names to export. If None,
            exports all available. Valid names:
            - 'samples': Sample metadata and quality metrics
            - 'consensus': Consensus features table
            - 'identification': Identification results with annotations
            - 'gaps': Gaps matrix (filled vs detected features)
            - 'matrix': Consensus matrix (features × samples)
            - 'history': Processing history and parameters
            - 'diff': Differential analysis group definitions
            - 'diff data': Differential analysis results
            Defaults to None.

    Example:
        Export all data::

            >>> study = masster.Study("./study")
            >>> study.load()
            >>> study.export_excel()

        Selective worksheet export::

            >>> study.export_excel(
            ...     filename="results.xlsx",
            ...     worksheets=["samples", "consensus", "matrix"]
            ... )

        Export only identifications::

            >>> study.export_excel(
            ...     filename="ids.xlsx",
            ...     worksheets=["consensus", "identification"]
            ... )

        Custom output location::

            >>> study.export_excel("../reports/study_results.xlsx")

    Note:
        **Worksheet descriptions:**

        - samples: One row per sample with metadata, groups, batches
        - consensus: One row per consensus feature (mz, rt, intensity stats)
        - identification: Multiple rows per feature (all compound matches)
        - gaps: Binary matrix (rows=features, cols=samples, 1=filled, 0=detected)
        - matrix: Intensity matrix (rows=features, cols=samples)
        - history: Processing log with parameters and timestamps
        - diff: Comparison definitions (if differential analysis performed)
        - diff data: Statistical results (if differential analysis performed)

        **Performance considerations:**

        - Excel export can be slow for large studies (>10k features)
        - Matrix and gaps worksheets are slowest (~70% of export time)
        - For faster exports, use export_parquet() or export_csv()
        - Selective worksheet export reduces time significantly

        **Data availability:**

        - Worksheets only exported if corresponding data exists
        - identification: Requires identify() completion
        - gaps: Requires fill() completion
        - diff/diff data: Requires differential analysis
        - Missing worksheets logged as warnings

        **File format:**

        - Uses openpyxl engine for writing
        - Compatible with Excel 2007+ (.xlsx)
        - Preserves column types where possible
        - Large integers may display in scientific notation in Excel

        **Automatic filename:**

        If filename=None, creates "study.xlsx" in study.folder. Relative paths
        resolved against study.folder.

    Raises:
        ImportError: If openpyxl package not installed.
        ValueError: If invalid worksheet names provided.

    See Also:
        export_parquet: Faster export format for large studies.
        export_csv: Export individual tables as CSV files.
        get_consensus_matrix: Retrieve matrix programmatically.
        get_id: Retrieve identification results.
    """
    try:
        import openpyxl
    except ImportError:
        self.logger.error(
            "openpyxl package is required for Excel export. Install with: pip install openpyxl",
        )
        return

    # Set default filename
    if filename is None:
        filename = "study.xlsx"

    # Make filename absolute if not already
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Validate worksheets parameter
    valid_worksheets = {
        "samples",
        "consensus",
        "identification",
        "gaps",
        "matrix",
        "history",
        "diff",
        "diff data",
    }
    if worksheets is not None:
        invalid = set(worksheets) - valid_worksheets
        if invalid:
            self.logger.error(
                f"Invalid worksheet names: {invalid}. Valid names: {valid_worksheets}",
            )
            return
        worksheets_to_export = set(worksheets)
        self.logger.debug(f"Exporting selected worksheets: {worksheets_to_export}")
    else:
        worksheets_to_export = valid_worksheets
        self.logger.debug("Exporting all worksheets")

    # Prepare data for export in the desired order
    from collections import OrderedDict

    data_dict = OrderedDict()

    # 1. Samples dataframe (first worksheet)
    if (
        "samples" in worksheets_to_export
        and self.samples_df is not None
        and not self.samples_df.is_empty()
    ):
        self.logger.debug("Preparing samples worksheet...")
        samples_pandas = self.samples_df.to_pandas()
        data_dict["samples"] = samples_pandas
        self.logger.debug(f"Added samples worksheet with {len(samples_pandas)} rows")
    elif "samples" in worksheets_to_export:
        self.logger.warning("samples_df is empty or None, skipping worksheet")

    # 2. Consensus dataframe (renamed to 'consensus')
    if (
        "consensus" in worksheets_to_export
        and self.consensus_df is not None
        and not self.consensus_df.is_empty()
    ):
        self.logger.debug("Preparing consensus worksheet...")
        consensus_pandas = self.consensus_df.to_pandas()
        data_dict["consensus"] = consensus_pandas
        self.logger.debug(
            f"Added consensus worksheet with {len(consensus_pandas)} rows",
        )
    elif "consensus" in worksheets_to_export:
        self.logger.warning("consensus_df is empty or None, skipping worksheet")

    # 3. Identification results
    if "identification" in worksheets_to_export:
        try:
            from masster.study.id import get_id

            self.logger.debug("Preparing identification worksheet...")
            id_df = get_id(self)
            if id_df is not None and not id_df.is_empty():
                id_pandas = id_df.to_pandas()
                data_dict["identification"] = id_pandas
                self.logger.debug(
                    f"Added identification worksheet with {len(id_pandas)} rows",
                )
            else:
                self.logger.warning(
                    "get_id() returned empty data, skipping identification worksheet",
                )
        except Exception as e:
            self.logger.debug(
                f"Error getting identification data: {e}. Skipping identification worksheet.",
            )

    # 4. Gaps matrix (filled vs non-filled features)
    if "gaps" in worksheets_to_export:
        try:
            self.logger.debug("Preparing gaps worksheet...")
            gaps_df = self.get_gaps_matrix()
            if gaps_df is not None and not gaps_df.is_empty():
                gaps_pandas = gaps_df.to_pandas()
                data_dict["gaps"] = gaps_pandas
                self.logger.debug(
                    f"Added gaps worksheet with {len(gaps_pandas)} rows",
                )
            else:
                self.logger.warning(
                    "get_gaps_matrix() returned empty data, skipping gaps worksheet",
                )
        except Exception as e:
            self.logger.debug(
                f"Error getting gaps data: {e}. Skipping gaps worksheet.",
            )

    # 5. Consensus matrix
    if "matrix" in worksheets_to_export:
        try:
            self.logger.debug("Preparing matrix worksheet...")
            matrix_df = self.get_consensus_matrix()
            if matrix_df is not None and not matrix_df.is_empty():
                matrix_pandas = matrix_df.to_pandas()
                data_dict["matrix"] = matrix_pandas
                self.logger.debug(
                    f"Added matrix worksheet with {len(matrix_pandas)} rows",
                )
            else:
                self.logger.warning(
                    "get_consensus_matrix() returned empty data, skipping matrix worksheet",
                )
        except Exception as e:
            self.logger.error(f"Error getting consensus matrix: {e}")

    # 6. Differential analysis comparisons
    if (
        "diff" in worksheets_to_export
        and hasattr(self, "diff_df")
        and self.diff_df is not None
        and not self.diff_df.is_empty()
    ):
        self.logger.debug("Preparing diff worksheet...")
        try:
            import polars as pl

            diff_rows = []
            for row in self.diff_df.iter_rows(named=True):
                diff_id = row["diff_id"]
                samples_ids = row["samples"]
                refs_ids = row["refs"]
                pairing = row["pairing"]

                # Convert sample_id (UUID7) to sample names
                samples_names = []
                refs_names = []

                if self.samples_df is not None:
                    for sid in samples_ids:
                        # Match against sample_id column (UUID7 string)
                        match = self.samples_df.filter(pl.col("sample_id") == sid)
                        if not match.is_empty():
                            samples_names.append(str(match["sample_name"][0]))

                    for rid in refs_ids:
                        # Match against sample_id column (UUID7 string)
                        match = self.samples_df.filter(pl.col("sample_id") == rid)
                        if not match.is_empty():
                            refs_names.append(str(match["sample_name"][0]))

                diff_rows.append(
                    {
                        "diff_id": diff_id,
                        "number_of_samples": len(samples_ids),
                        "number_of_refs": len(refs_ids),
                        "list_of_samples": ", ".join(samples_names)
                        if samples_names
                        else ", ".join(samples_ids),
                        "list_of_refs": ", ".join(refs_names)
                        if refs_names
                        else ", ".join(refs_ids),
                        "pairing": pairing if pairing else "",
                    },
                )

            if diff_rows:
                diff_pandas = pd.DataFrame(diff_rows)
                data_dict["diff"] = diff_pandas
                self.logger.debug(f"Added diff worksheet with {len(diff_pandas)} rows")
        except Exception as e:
            self.logger.error(f"Error preparing diff worksheet: {e}. Skipping.")
    elif "diff" in worksheets_to_export:
        self.logger.debug("diff_df is empty or None, skipping diff worksheet")

    # 7. Differential analysis results
    if (
        "diff data" in worksheets_to_export
        and hasattr(self, "diff_df")
        and self.diff_df is not None
        and not self.diff_df.is_empty()
    ):
        self.logger.debug("Preparing 'diff data' worksheet...")
        try:
            # Get differential analysis results for all registered comparisons
            diff_results = self.get_diff()
            if diff_results is not None and not diff_results.is_empty():
                diff_data_pandas = diff_results.to_pandas()
                data_dict["diff data"] = diff_data_pandas
                self.logger.debug(
                    f"Added 'diff data' worksheet with {len(diff_data_pandas)} rows",
                )
            else:
                self.logger.warning(
                    "get_diff() returned empty data, skipping 'diff data' worksheet",
                )
        except Exception as e:
            self.logger.debug(f"Error preparing 'diff data' worksheet: {e}. Skipping.")
    elif "diff data" in worksheets_to_export:
        self.logger.debug("diff_df is empty or None, skipping 'diff data' worksheet")

    # 8. History (last worksheet)
    if "history" in worksheets_to_export:
        try:
            self.logger.debug("Preparing history worksheet...")
            if hasattr(self, "history") and self.history:
                import json

                # Flatten history dict into rows for Excel
                history_rows = []

                # Add study parameters as the first entries if available
                if hasattr(self, "parameters") and self.parameters is not None:
                    params_dict = self.parameters.to_dict()
                    for subkey, subvalue in params_dict.items():
                        history_rows.append(
                            {
                                "category": "parameters",
                                "parameter": subkey,
                                "value": json.dumps(subvalue)
                                if isinstance(subvalue, (dict, list))
                                else str(subvalue),
                            },
                        )

                # Add all history entries
                for key, value in self.history.items():
                    if isinstance(value, dict):
                        for subkey, subvalue in value.items():
                            history_rows.append(
                                {
                                    "category": key,
                                    "parameter": subkey,
                                    "value": json.dumps(subvalue)
                                    if isinstance(subvalue, (dict, list))
                                    else str(subvalue),
                                },
                            )
                    else:
                        history_rows.append(
                            {
                                "category": key,
                                "parameter": "",
                                "value": json.dumps(value)
                                if isinstance(value, (dict, list))
                                else str(value),
                            },
                        )

                if history_rows:
                    history_pandas = pd.DataFrame(history_rows)
                    data_dict["history"] = history_pandas
                    self.logger.debug(
                        f"Added history worksheet with {len(history_pandas)} rows",
                    )
                else:
                    self.logger.warning("history is empty, skipping worksheet")
            else:
                self.logger.warning("history is not available, skipping worksheet")
        except Exception as e:
            self.logger.debug(
                f"Error getting history data: {e}. Skipping history worksheet.",
            )

    # Check if we have any data to export
    if not data_dict:
        self.logger.error("No data available to export to Excel")
        return

    # Write to Excel file with progress logging
    try:
        total_worksheets = len(data_dict)
        self.logger.info(f"Writing {total_worksheets} worksheet(s) to Excel...")

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for idx, (sheet_name, data) in enumerate(data_dict.items(), 1):
                self.logger.debug(
                    f"Writing worksheet {idx}/{total_worksheets}: {sheet_name} ({data.shape[0]:,} rows x {data.shape[1]:,} cols)...",
                )
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.debug(
                    f"Written worksheet '{sheet_name}' with shape {data.shape}",
                )

        self.logger.info(f"Study exported to {filename}")

    except Exception as e:
        self.logger.error(f"Error writing Excel file: {e}")


def export_parquet(self, filename: str | None = None) -> None:
    """
    Export the study data to multiple Parquet files with different suffixes.

    The export creates separate Parquet files for each dataset:
    - <filename>_samples.parquet: Samples dataframe
    - <filename>_consensus.parquet: Consensus features dataframe
    - <filename>_identification.parquet: Identification results with library annotations
    - <filename>_matrix.parquet: Consensus matrix with samples as columns

    Args:
        filename (str, optional): Base name for the output files. Defaults to "study"
                                 in the study folder.
    """
    # Set default filename
    if filename is None:
        filename = "study"

    # Make filename absolute path if not already (without extension)
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    self.logger.debug(f"Exporting study to Parquet files with filename: {filename}")

    exported_files = []

    # 1. Samples dataframe
    if self.samples_df is not None and not self.samples_df.is_empty():
        samples_file = f"{filename}_samples.parquet"
        try:
            self.samples_df.write_parquet(samples_file)
            exported_files.append(samples_file)
            self.logger.debug(
                f"Exported samples to {samples_file} ({self.samples_df.height} rows)",
            )
        except Exception as e:
            self.logger.error(f"Error writing samples parquet file: {e}")
    else:
        self.logger.warning(
            "samples_df is empty or None, skipping samples parquet file",
        )

    # 2. Consensus dataframe
    if self.consensus_df is not None and not self.consensus_df.is_empty():
        consensus_file = f"{filename}_consensus.parquet"
        try:
            # Create a copy of consensus_df for parquet export
            consensus_export_df = self.consensus_df.clone()

            # Handle Object dtype columns that can't be serialized to parquet
            for col in consensus_export_df.columns:
                if consensus_export_df[col].dtype == pl.Object:
                    if col == "iso":
                        # Convert numpy arrays to string representation for parquet compatibility
                        # This preserves the data while making it parquet-serializable
                        consensus_export_df = consensus_export_df.with_columns(
                            [
                                pl.col("iso")
                                .map_elements(
                                    lambda x: str(x.tolist())
                                    if x is not None
                                    else None,
                                    return_dtype=pl.String,
                                )
                                .alias("iso"),
                            ],
                        )
                    else:
                        # For other Object columns, convert to string representation
                        consensus_export_df = consensus_export_df.with_columns(
                            [
                                pl.col(col)
                                .map_elements(
                                    lambda x: str(x) if x is not None else None,
                                    return_dtype=pl.String,
                                )
                                .alias(col),
                            ],
                        )

            consensus_export_df.write_parquet(consensus_file)
            exported_files.append(consensus_file)
            self.logger.debug(
                f"Exported consensus to {consensus_file} ({consensus_export_df.height} rows)",
            )
        except Exception as e:
            self.logger.error(f"Error writing consensus parquet file: {e}")
    else:
        self.logger.warning(
            "consensus_df is empty or None, skipping consensus parquet file",
        )

    # 3. Identification results
    try:
        from masster.study.id import get_id

        id_df = get_id(self)
        if id_df is not None and not id_df.is_empty():
            identification_file = f"{filename}_identification.parquet"
            try:
                id_df.write_parquet(identification_file)
                exported_files.append(identification_file)
                self.logger.debug(
                    f"Exported identification to {identification_file} ({id_df.height} rows)",
                )
            except Exception as e:
                self.logger.error(f"Error writing identification parquet file: {e}")
        else:
            self.logger.warning(
                "get_id() returned empty data, skipping identification parquet file",
            )
    except Exception as e:
        self.logger.warning(
            f"Error getting identification data: {e}. Skipping identification parquet file.",
        )

    # 4. Consensus matrix
    try:
        matrix_df = self.get_consensus_matrix()
        if matrix_df is not None and not matrix_df.is_empty():
            matrix_file = f"{filename}_matrix.parquet"
            try:
                matrix_df.write_parquet(matrix_file)
                exported_files.append(matrix_file)
                self.logger.debug(
                    f"Exported matrix to {matrix_file} ({matrix_df.height} rows)",
                )
            except Exception as e:
                self.logger.error(f"Error writing matrix parquet file: {e}")
        else:
            self.logger.warning(
                "get_consensus_matrix() returned empty data, skipping matrix parquet file",
            )
    except Exception as e:
        self.logger.error(f"Error getting consensus matrix: {e}")

    # Report results
    if exported_files:
        self.logger.info(f"Study exported to {len(exported_files)} Parquet files.")
    else:
        self.logger.error("No Parquet files were created - no data available to export")


def export_csv(self, filename="consensus.csv"):
    """
    Export the consensus features DataFrame to a CSV file with comprehensive information.

    This method exports the consensus features to a CSV format with comprehensive feature information
    such as m/z, RT, annotations, isotopic patterns, MS2 data, and intensity values for each sample.

    Parameters:
        filename (str): The path to the output CSV file. Defaults to 'consensus.csv'.

    Side Effects:
        Writes the exported data to the specified CSV file and logs the export operation.
    """
    if self.consensus_df is None:
        self.logger.warning(
            "No consensus features found. Cannot export consensus features.",
        )
        return

    # Make filename absolute if not already
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    df = self.consensus_df

    # Get consensus matrix for quantification across samples
    try:
        quant_matrix = self.get_consensus_matrix()
    except Exception as e:
        self.logger.error(f"Error getting consensus matrix: {e}")
        return

    # Evaluate the charge column
    if "charge_mean" in df.columns:
        charge_series = df.select(
            pl.when(pl.col("charge_mean") == 0)
            .then(1 if self.parameters.polarity == "positive" else -1)
            .otherwise(pl.col("charge_mean"))
            .alias("charge"),
        ).get_column("charge")
    else:
        charge_series = pl.Series(
            [1 if self.parameters.polarity == "positive" else -1] * len(df),
        )

    # Evaluate the group column (from adduct_group_top)
    # Features with adduct_group_top == 0 should each get a unique group index
    if "adduct_group_top" in df.columns:
        max_adduct_group = df.get_column("adduct_group_top").max()
        if max_adduct_group is None:
            max_adduct_group = 0

        group_series = df.select(
            pl.when(pl.col("adduct_group_top") == 0)
            .then(
                max_adduct_group
                + 1
                + pl.int_range(pl.len()).over(pl.col("adduct_group_top") == 0),
            )
            .otherwise(pl.col("adduct_group_top"))
            .alias("group"),
        ).get_column("group")
    else:
        group_series = pl.Series([None] * len(df))

    # Evaluate the annotation column (adduct + isotope info)
    if "adduct_top" in df.columns and "iso_mean" in df.columns:
        annotation_series = df.select(
            pl.when(pl.col("iso_mean") == 0)
            .then(pl.col("adduct_top").str.replace(r"\?", "H"))
            .otherwise(
                pl.col("adduct_top").str.replace(r"\?", "H")
                + " +"
                + pl.col("iso_mean").cast(pl.Int64).cast(pl.Utf8),
            )
            .alias("annotation"),
        ).get_column("annotation")
    elif "adduct_top" in df.columns:
        annotation_series = df.get_column("adduct_top").str.replace(r"\?", "H")
    else:
        annotation_series = pl.Series([""] * len(df))

    # Get sample columns from quant_matrix (excluding consensus_id)
    sample_columns = [col for col in quant_matrix.columns if col != "consensus_id"]

    # Create SLAW columns with appropriate mappings from consensus_df
    slaw_data = {
        "feature_id": df.get_column("consensus_uid")
        if "consensus_uid" in df.columns
        else pl.Series(range(1, len(df) + 1)),
        "mz": df.get_column("mz")
        if "mz" in df.columns
        else pl.Series([None] * len(df)),
        "rt": df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "group": group_series,
        "annotation": annotation_series,
        "neutral_mass": df.get_column("adduct_neutral_mass_top")
        if "adduct_neutral_mass_top" in df.columns
        else pl.Series([None] * len(df)),
        "charge": charge_series,
        "main_id": df.get_column("main_id")
        if "main_id" in df.columns
        else df.get_column("consensus_id")
        if "consensus_id" in df.columns
        else pl.Series(range(1, len(df) + 1)),
        "ion": df.get_column("adduct_top").str.replace(r"\?", "H")
        if "adduct_top" in df.columns
        else pl.Series([""] * len(df)),
        "iso": df.get_column("iso_mean").cast(pl.Int64)
        if "iso_mean" in df.columns
        else pl.Series([0] * len(df)),
        "clique": df.get_column("clique")
        if "clique" in df.columns
        else pl.Series([None] * len(df)),
        "num_detection": df.get_column("num_detection")
        if "num_detection" in df.columns
        else pl.Series([1] * len(df)),
        "total_detection": df.get_column("total_detection")
        if "total_detection" in df.columns
        else pl.Series([1] * len(df)),
        "mz_mean": df.get_column("mz")
        if "mz" in df.columns
        else pl.Series([None] * len(df)),
        "mz_min": df.get_column("mz_min")
        if "mz_min" in df.columns
        else df.get_column("mz")
        if "mz" in df.columns
        else pl.Series([None] * len(df)),
        "mz_max": df.get_column("mz_max")
        if "mz_max" in df.columns
        else df.get_column("mz")
        if "mz" in df.columns
        else pl.Series([None] * len(df)),
        "rt_mean": df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "rt_min": df.get_column("rt_min")
        if "rt_min" in df.columns
        else df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "rt_max": df.get_column("rt_max")
        if "rt_max" in df.columns
        else df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "rt_cor_mean": df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "rt_cor_min": df.get_column("rt_min")
        if "rt_min" in df.columns
        else df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "rt_cor_max": df.get_column("rt_max")
        if "rt_max" in df.columns
        else df.get_column("rt")
        if "rt" in df.columns
        else pl.Series([None] * len(df)),
        "height_mean": df.get_column("height_mean")
        if "height_mean" in df.columns
        else pl.Series([None] * len(df)),
        "height_min": df.get_column("height_min")
        if "height_min" in df.columns
        else pl.Series([None] * len(df)),
        "height_max": df.get_column("height_max")
        if "height_max" in df.columns
        else pl.Series([None] * len(df)),
        "intensity_mean": df.get_column("inty_mean")
        if "inty_mean" in df.columns
        else pl.Series([None] * len(df)),
        "intensity_min": df.get_column("inty_min")
        if "inty_min" in df.columns
        else pl.Series([None] * len(df)),
        "intensity_max": df.get_column("inty_max")
        if "inty_max" in df.columns
        else pl.Series([None] * len(df)),
        "SN_mean": df.get_column("sn_mean")
        if "sn_mean" in df.columns
        else pl.Series([None] * len(df)),
        "SN_min": df.get_column("sn_min")
        if "sn_min" in df.columns
        else pl.Series([None] * len(df)),
        "SN_max": df.get_column("sn_max")
        if "sn_max" in df.columns
        else pl.Series([None] * len(df)),
        "peakwidth_mean": df.get_column("fwhm_mean")
        if "fwhm_mean" in df.columns
        else pl.Series([None] * len(df)),
        "peakwidth_min": df.get_column("fwhm_min")
        if "fwhm_min" in df.columns
        else pl.Series([None] * len(df)),
        "peakwidth_max": df.get_column("fwhm_max")
        if "fwhm_max" in df.columns
        else pl.Series([None] * len(df)),
        "ms2_mgf_id": pl.Series([""] * len(df)),  # Not available in study
        "ms2_num_fused": pl.Series([None] * len(df)),  # Not available in study
        "ms2_source": pl.Series([""] * len(df)),  # Not available in study
        "isotopic_pattern_annot": pl.Series([""] * len(df)),  # Not available in study
        "isotopic_pattern_rel": pl.Series([""] * len(df)),  # Not available in study
        "isotopic_pattern_abs": pl.Series([""] * len(df)),  # Not available in study
    }

    # Add quantification columns for each sample
    for sample_col in sample_columns:
        quant_column_name = f"quant_{sample_col}"
        # Join with quant_matrix to get values for this sample
        sample_values = quant_matrix.join(
            df.select("consensus_id"),
            on="consensus_id",
            how="right",
        ).get_column(sample_col)
        slaw_data[quant_column_name] = sample_values

    # Create the polars DataFrame
    slaw_df = pl.DataFrame(slaw_data)

    # Convert to pandas for CSV export
    pandas_df = slaw_df.to_pandas()

    # Export to CSV with comma separator - only quote when necessary (QUOTE_MINIMAL)
    try:
        pandas_df.to_csv(
            filename,
            sep=",",
            index=False,
            quoting=0,
        )  # quoting=0 means QUOTE_MINIMAL
        self.logger.info(f"Features exported to {filename}")
        self.logger.debug(
            f"Exported {len(slaw_df)} features with {len(slaw_df.columns)} columns",
        )
    except PermissionError:
        self.logger.error(
            f"Permission denied: Cannot write to {filename}. The file may be open in another program. Please close it and try again.",
        )


def export_history(self, filename: str | None = None) -> None:
    """
    Export the processing history as a JSON file.

    The history dict contains all important processing steps that have been applied
    to the study, including parameters and timestamps.

    Args:
        filename (str, optional): Path to the output JSON file. Defaults to "history.json"
                                 in the study folder.
    """
    # Set default filename
    if filename is None:
        filename = "history.json"

    # Make filename absolute if not already
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Check if history exists and has content
    if not hasattr(self, "history") or self.history is None:
        self.logger.warning("No history available to export")
        return

    if not self.history:
        self.logger.warning("History is empty, nothing to export")
        return

    # Prepare history with parameters at the top
    export_dict = {}

    # Add study parameters as the first entry if available
    if hasattr(self, "parameters") and self.parameters is not None:
        export_dict["parameters"] = self.parameters.to_dict()

    # Add all history entries
    export_dict.update(self.history)

    # Write history to JSON file
    try:
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(export_dict, f, indent=2, ensure_ascii=False)

        self.logger.info(f"History exported to {filename}")
        self.logger.debug(f"Exported {len(export_dict)} history entries")
    except Exception as e:
        self.logger.error(f"Error writing history file: {e}")


def export_acquisition(self, filename: str | None = None) -> None:
    """
    Export the acquisition parameters (MS and LC methods) as a JSON file.

    This method exports only the acquisition-related entries from the history dict,
    specifically "acquisition_ms" and "acquisition_lc" if they exist. This is useful
    for documenting instrument parameters and method settings across samples in the study.

    Args:
        filename (str, optional): Path to the output JSON file. Defaults to "acquisition.json"
                                 in the study folder.

    Example:
        Basic export with default filename::

            >>> study = masster.Study()
            >>> study.load_study("data.study5")
            >>> study.export_acquisition()  # Creates acquisition.json

        Custom filename::

            >>> study.export_acquisition("method_params.json")

    Note:
        Only exports "acquisition_ms" and "acquisition_lc" from history if they exist.
        If neither exists, a warning is logged and no file is created.
    """
    # Set default filename
    if filename is None:
        filename = "acquisition.json"

    # Make filename absolute if not already
    if not os.path.isabs(filename):
        if self.folder is not None:
            filename = os.path.join(self.folder, filename)
        else:
            filename = os.path.join(os.getcwd(), filename)

    # Check if history exists and has content
    if not hasattr(self, "history") or self.history is None:
        self.logger.warning("No history available to export acquisition parameters")
        return

    if not self.history:
        self.logger.warning("History is empty, no acquisition parameters to export")
        return

    # Prepare acquisition dict with only MS and LC methods
    export_dict = {}

    # Add acquisition_ms if available
    if "acquisition_ms" in self.history:
        export_dict["acquisition_ms"] = self.history["acquisition_ms"]

    # Add acquisition_lc if available
    if "acquisition_lc" in self.history:
        export_dict["acquisition_lc"] = self.history["acquisition_lc"]

    # Check if we have anything to export
    if not export_dict:
        self.logger.warning(
            "No acquisition parameters (acquisition_ms or acquisition_lc) found in history",
        )
        return

    # Write acquisition parameters to JSON file
    try:
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(export_dict, f, indent=2, ensure_ascii=False)

        self.logger.info(f"Acquisition parameters exported to {filename}")
        self.logger.debug(f"Exported {len(export_dict)} acquisition entries")
    except Exception as e:
        self.logger.error(f"Error writing acquisition file: {e}")
