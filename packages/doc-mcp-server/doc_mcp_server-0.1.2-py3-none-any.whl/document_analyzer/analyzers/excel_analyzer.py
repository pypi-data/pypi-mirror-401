"""
Excel 文档分析器
支持复杂Excel(合并单元格、多层表头等)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Tuple, Optional
import os

from .base import (
    BaseAnalyzer,
    DocumentMeta,
    DocumentFormat,
    SectionInfo,
    FieldInfo
)


class ExcelAnalyzer(BaseAnalyzer):
    """Excel文档分析器"""

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        super().__init__(file_path)
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self.merged_cells_map = {}
        self._load_workbook()

    def _load_workbook(self):
        """加载Excel工作簿"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"文件不存在: {self.file_path}")

        self.wb = openpyxl.load_workbook(self.file_path, data_only=False)

        # 选择工作表
        if self.sheet_name:
            if self.sheet_name not in self.wb.sheetnames:
                raise ValueError(f"工作表不存在: {self.sheet_name}")
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.active

        # 构建合并单元格映射
        self._build_merged_cells_map()

    def _build_merged_cells_map(self):
        """构建合并单元格映射表"""
        self.merged_cells_map = {}
        for merged_range in self.ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            # 所有合并单元格都指向左上角的主单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    self.merged_cells_map[(row, col)] = (min_row, min_col, max_row, max_col)

    def _get_cell_value(self, row: int, col: int) -> Any:
        """获取单元格值(处理合并单元格)"""
        if (row, col) in self.merged_cells_map:
            main_row, main_col, _, _ = self.merged_cells_map[(row, col)]
            value = self.ws.cell(main_row, main_col).value
        else:
            value = self.ws.cell(row, col).value

        return str(value).strip() if value else ""

    def _get_cell_coordinate(self, row: int, col: int) -> str:
        """获取单元格坐标(Excel格式)"""
        return f"{get_column_letter(col)}{row}"

    def _detect_sections(self) -> List[SectionInfo]:
        """检测章节"""
        sections = []
        current_section = None

        for row in range(1, self.ws.max_row + 1):
            first_col_value = self._get_cell_value(row, 1)

            # 检测章节标题(包含"部分"、"信息概要"等关键词)
            is_section_title = any(keyword in first_col_value for keyword in
                                   ['部分', '信息概要', '明细', '附件', '记录'])

            if is_section_title and len(first_col_value) > 3:
                # 结束上一个章节
                if current_section:
                    current_section.end_row = row - 1
                    sections.append(current_section)

                # 开始新章节
                current_section = SectionInfo(
                    title=first_col_value,
                    start_row=row,
                    end_row=self.ws.max_row,  # 临时值
                    start_col=1,
                    end_col=self.ws.max_column
                )

        # 添加最后一个章节
        if current_section:
            sections.append(current_section)

        return sections

    def _detect_header_rows(self, section: SectionInfo) -> tuple:
        """
        检测章节的表头范围（改进版：允许表头中间有空行）
        返回: (header_start_row, header_end_row, data_start_row)

        策略：向下扫描，识别所有包含表头关键词的行，直到遇到明确的数据行
        """
        # 跳过标题行
        header_start = section.start_row
        first_row_value = self._get_cell_value(header_start, 1)
        if first_row_value and any(kw in first_row_value for kw in ['部分', '概要', '明细', '附件', '记录']):
            header_start += 1

        # 表头关键词
        header_keywords = [
            '账户', '余额', '金额', '日期', '机构', '笔数', '交易', '年份',
            '信贷', '贷款', '担保', '授信', '债务', '垫款', '欠税', '条数',
            '记录', '信息', '明细', '类型', '状态', '合计', '小计', '最大',
            '其中', '被追偿', '关注类', '不良类', '非信贷', '民事判决'
        ]

        # 扫描所有行，标记哪些是表头
        max_scan = min(section.start_row + 30, section.end_row)
        row_types = {}  # {row: 'header' | 'empty' | 'data'}

        for row in range(header_start, max_scan):
            non_empty = 0
            has_keyword = False
            pure_numbers = 0

            for col in range(1, min(self.ws.max_column + 1, 25)):
                value = self._get_cell_value(row, col)
                if value:
                    non_empty += 1
                    # 检查是否包含表头关键词
                    if any(kw in value for kw in header_keywords):
                        has_keyword = True
                    # 检查是否纯数字
                    elif value.replace('.', '').replace('-', '').replace(',', '').replace('%', '').strip().isdigit():
                        pure_numbers += 1

            # 分类这一行
            if non_empty == 0:
                row_types[row] = 'empty'
            elif has_keyword:
                row_types[row] = 'header'
            elif pure_numbers >= 2:  # 至少2个纯数字，判定为数据行
                row_types[row] = 'data'
            else:
                row_types[row] = 'unknown'

        # 找到表头的结束位置：最后一个'header'行
        last_header_row = header_start
        for row in range(header_start, max_scan):
            if row_types.get(row) == 'header':
                last_header_row = row

        # 找到数据的开始位置：第一个'data'行，或表头后的第一个非空行
        data_start = last_header_row + 1
        for row in range(last_header_row + 1, max_scan):
            row_type = row_types.get(row)
            if row_type == 'data':
                data_start = row
                break
            elif row_type not in ('empty', None):  # 非空且非数据，继续往下找
                data_start = row + 1

        header_end = last_header_row

        # 兜底：如果没找到明确的表头，使用默认值
        if header_end == header_start:
            section_size = section.end_row - section.start_row + 1
            default_rows = 7 if section_size > 20 else 5
            header_end = min(header_start + default_rows, section.end_row)
            data_start = header_end + 1

        return (header_start, header_end, data_start)

    def _merge_header_hierarchy(self, section: SectionInfo, header_start: int, header_end: int) -> Dict[int, str]:
        """
        合并多层表头，生成每列的完整字段路径（改进版：保留完整层级）
        返回: {col: full_path}

        策略：对于每一列的多个表头块，智能合并，保留层级结构
        """
        column_data = {}  # {col: [(row, value), ...]}

        # 收集每一列的所有非空单元格
        for col in range(1, self.ws.max_column + 1):
            cells = []
            for row in range(header_start, header_end + 1):
                value = self._get_cell_value(row, col)
                if value and value.strip():
                    clean_value = value.replace('其中：', '').replace('　', '').strip()
                    if len(clean_value) <= 30 and clean_value not in ['', '-', '/', '\\', '|']:
                        cells.append((row, clean_value))

            if cells:
                column_data[col] = cells

        # 合并每列的路径
        merged_paths = {}
        for col, cells in column_data.items():
            if not cells:
                continue

            # 按行号分组成块
            blocks = []
            current_block = []
            last_row = None

            for row, value in cells:
                if last_row is None or row - last_row <= 2:  # 间隔<=2行，属于同一块
                    current_block.append(value)
                else:
                    # 新块开始
                    if current_block:
                        blocks.append(current_block)
                    current_block = [value]

                last_row = row

            if current_block:
                blocks.append(current_block)

            # 合并策略：保留完整的层级结构
            if len(blocks) == 1:
                # 只有一个块，保留所有层级
                path_parts = blocks[0]
            elif len(blocks) == 2:
                # 两个块：合并第一个块和第二个块
                path_parts = blocks[0] + blocks[1]
            else:
                # 多个块：只用最后一个块（通常是最具体的）
                path_parts = blocks[-1]

            # 去重（保持顺序）
            unique_parts = []
            for part in path_parts:
                if not unique_parts or part != unique_parts[-1]:
                    unique_parts.append(part)

            # 放宽长度限制：最多保留4层（避免过长但不丢失关键信息）
            if len(unique_parts) > 4:
                # 保留第1层和最后3层
                unique_parts = [unique_parts[0]] + unique_parts[-3:]

            merged_paths[col] = '_'.join(unique_parts)

        return merged_paths

    def _extract_fields(self, sections: List[SectionInfo]) -> List[FieldInfo]:
        """提取字段信息（优化版：支持多层表头 + 去重）"""
        fields = []
        seen_keys = set()  # 用于去重

        for section in sections:
            # 检测表头范围
            header_start, header_end, data_start = self._detect_header_rows(section)

            # 合并多层表头
            column_paths = self._merge_header_hierarchy(section, header_start, header_end)

            # 为每列创建字段（去重）
            for col, full_path in column_paths.items():
                # 生成唯一键（章节名_完整路径）
                field_key = f"{section.title}_{full_path}"

                # 去重：如果已经添加过相同的字段键，跳过
                if field_key in seen_keys:
                    continue

                seen_keys.add(field_key)

                # 字段名是路径的最后一部分
                path_parts = full_path.split('_')
                field_name = path_parts[-1]

                # 坐标使用表头最后一行
                coord = self._get_cell_coordinate(header_end, col)

                field = FieldInfo(
                    key=field_key,
                    name=field_name,
                    full_path=full_path,
                    coord=coord,
                    row=header_end,  # 表头行
                    col=col,
                    section=section.title,
                    data_row=data_start  # 数据行
                )
                fields.append(field)

        return fields

    def _split_into_subsections(self, sections: List[SectionInfo]) -> List[SectionInfo]:
        """
        将大章节拆分为子章节（处理章节内有多组独立表格的情况）

        策略：识别章节内的多个表格区域，通过以下特征判断：
        1. 连续2行以上的空行 - 表格分隔
        2. 表头关键词的突然变化 - 新表格开始
        """
        refined_sections = []

        # 表头关键词列表
        header_keywords = [
            '账户', '余额', '金额', '日期', '机构', '笔数', '交易', '年份',
            '信贷', '贷款', '担保', '授信', '债务', '垫款', '欠税', '条数',
            '记录', '信息', '明细', '类型', '状态'
        ]

        for section in sections:
            # 扫描章节内部，找到表格分隔点
            subsection_starts = [section.start_row]  # 子章节起始行列表

            consecutive_empty = 0
            last_row_had_keywords = False

            # 跳过标题行
            scan_start = section.start_row
            first_row_value = self._get_cell_value(scan_start, 1)
            if first_row_value and any(kw in first_row_value for kw in ['部分', '概要', '明细', '附件']):
                scan_start += 1

            # 扫描章节内部
            for row in range(scan_start, section.end_row + 1):
                # 统计这一行的特征
                non_empty = 0
                has_keywords = False

                for col in range(1, min(self.ws.max_column + 1, 20)):
                    value = self._get_cell_value(row, col)
                    if value:
                        non_empty += 1
                        if any(kw in value for kw in header_keywords):
                            has_keywords = True

                # 判断是否是分隔点
                if non_empty == 0:
                    consecutive_empty += 1
                else:
                    # 如果之前有连续空行（>=2行），且当前行有表头关键词，则是新表格
                    if consecutive_empty >= 2 and has_keywords and last_row_had_keywords:
                        # 新子章节开始
                        subsection_starts.append(row)

                    consecutive_empty = 0
                    last_row_had_keywords = has_keywords

            # 创建子章节
            if len(subsection_starts) > 1:
                # 有多个子章节，拆分
                for i, start in enumerate(subsection_starts):
                    # 确定结束行
                    if i < len(subsection_starts) - 1:
                        end = subsection_starts[i + 1] - 1
                    else:
                        end = section.end_row

                    # 跳过末尾的空行
                    while end > start:
                        row_empty = True
                        for col in range(1, min(self.ws.max_column + 1, 10)):
                            if self._get_cell_value(end, col):
                                row_empty = False
                                break
                        if row_empty:
                            end -= 1
                        else:
                            break

                    # 生成子章节标题
                    if i == 0:
                        # 第一个子章节保留原标题
                        subsection_title = section.title
                    else:
                        # 后续子章节，尝试从第一行提取标题
                        first_value = self._get_cell_value(start, 1)
                        if first_value and len(first_value) > 3:
                            subsection_title = f"{section.title}_{first_value}"
                        else:
                            subsection_title = f"{section.title}_子表{i+1}"

                    subsection = SectionInfo(
                        title=subsection_title,
                        start_row=start,
                        end_row=end,
                        start_col=section.start_col,
                        end_col=section.end_col,
                        parent=section.title
                    )
                    refined_sections.append(subsection)
            else:
                # 没有子章节，保持原样
                refined_sections.append(section)

        return refined_sections

    def analyze(self) -> Dict[str, Any]:
        """分析Excel结构"""
        # 获取文档元数据
        file_size = os.path.getsize(self.file_path)
        meta = DocumentMeta(
            format=DocumentFormat.EXCEL,
            file_path=self.file_path,
            file_size=file_size,
            page_count=len(self.wb.sheetnames)
        )

        # 检测章节
        sections = self._detect_sections()

        # 拆分子章节
        sections = self._split_into_subsections(sections)

        # 提取字段
        fields = self._extract_fields(sections)
        meta.total_fields = len(fields)

        # 生成摘要
        summary = self._generate_summary(meta, sections, fields)

        return {
            'meta': meta,
            'sections': sections,
            'fields': fields,
            'summary': summary,
            'merged_cells_count': len(self.merged_cells_map)
        }

    def _generate_summary(
        self,
        meta: DocumentMeta,
        sections: List[SectionInfo],
        fields: List[FieldInfo]
    ) -> str:
        """生成结构摘要"""
        summary = f"""
Excel文档结构摘要
================
工作表: {self.ws.title}
总行数: {self.ws.max_row}
总列数: {self.ws.max_column}
章节数: {len(sections)}
字段数: {len(fields)}
合并单元格: {len(self.merged_cells_map)}

章节列表:
"""
        for i, section in enumerate(sections, 1):
            row_range = f"{section.start_row}-{section.end_row}"
            summary += f"  {i}. {section.title} (第{row_range}行)\n"

        return summary

    def get_field_value(self, field_key: str) -> Any:
        """获取字段值（从数据行读取）"""
        structure = self.get_structure()

        # 查找字段
        field = next((f for f in structure['fields'] if f.key == field_key), None)
        if not field:
            raise ValueError(f"字段不存在: {field_key}")

        # 从数据行读取值（而不是表头行）
        data_row = field.data_row if field.data_row else field.row + 1
        return self._get_cell_value(data_row, field.col)

    def set_field_value(self, field_key: str, value: Any):
        """设置字段值"""
        structure = self.get_structure()

        # 查找字段
        field = next((f for f in structure['fields'] if f.key == field_key), None)
        if not field:
            raise ValueError(f"字段不存在: {field_key}")

        # 写入值
        self.ws.cell(field.row, field.col, value)

    def get_section_data(self, section_name: str) -> Dict[str, Any]:
        """获取章节数据（从数据行读取）"""
        structure = self.get_structure()

        # 查找章节的所有字段
        section_fields = [f for f in structure['fields'] if f.section == section_name]

        if not section_fields:
            raise ValueError(f"章节不存在或无字段: {section_name}")

        data = {}
        for field in section_fields:
            # 使用完整路径作为键（避免重复）
            field_display_name = field.full_path if field.full_path else field.name

            # 从数据行读取值
            data_row = field.data_row if field.data_row else field.row + 1
            value = self._get_cell_value(data_row, field.col)

            data[field_display_name] = value

        return data

    def save(self, output_path: Optional[str] = None):
        """保存Excel文件"""
        save_path = output_path or self.file_path
        self.wb.save(save_path)

    def list_sections(self) -> List[str]:
        """列出所有章节名"""
        structure = self.get_structure()
        return [section.title for section in structure['sections']]

    def list_fields(self, section_name: Optional[str] = None) -> List[str]:
        """列出字段"""
        structure = self.get_structure()

        if section_name:
            return [f.key for f in structure['fields'] if f.section == section_name]
        return [f.key for f in structure['fields']]

    def export_structure(self, output_path: str, format: str = 'json'):
        """导出结构化文档"""
        structure = self.get_structure()

        if format == 'json':
            import json
            with open(output_path, 'w', encoding='utf-8') as f:
                # 转换dataclass为dict
                export_data = {
                    'meta': {
                        'format': structure['meta'].format.value,
                        'file_path': structure['meta'].file_path,
                        'file_size': structure['meta'].file_size,
                        'page_count': structure['meta'].page_count,
                        'total_fields': structure['meta'].total_fields
                    },
                    'sections': [
                        {
                            'title': s.title,
                            'start_row': s.start_row,
                            'end_row': s.end_row
                        } for s in structure['sections']
                    ],
                    'fields': [
                        {
                            'key': f.key,
                            'name': f.name,
                            'coord': f.coord,
                            'row': f.row,
                            'col': f.col,
                            'section': f.section
                        } for f in structure['fields']
                    ]
                }
                json.dump(export_data, f, ensure_ascii=False, indent=2)

        elif format == 'markdown':
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(structure['summary'])

    def __del__(self):
        """清理资源"""
        if self.wb:
            self.wb.close()
