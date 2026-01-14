from .common import Common
from .clickhouse import Clickhouse
from .ozon_reklama import OZONreklama
import requests
from datetime import datetime,timedelta
import clickhouse_connect
import pandas as pd
from transliterate import translit
import os
from dateutil import parser
import time
import hashlib
from io import StringIO
import json
from dateutil.relativedelta import relativedelta
import csv
import io
import openpyxl


class DISKbyPage:
    def __init__(self, bot_token:str = '', chats:str = '', message_type: str = '', subd: str = '',
                 host: str = '', port: str = '', username: str = '', password: str = '', database: str = '',
                 add_name: str = '', link:str = '', token: str  = '', start: str = '',  reports :str = ''):
        self.bot_token = bot_token
        self.chat_list = chats.replace(' ','').split(',')
        self.message_type  = message_type
        self.common = Common(self.bot_token, self.chat_list, self.message_type)
        self.link = link
        self.token = token
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.database = database
        self.subd = subd
        self.add_name = self.common.transliterate_key(add_name)
        self.now = datetime.now()
        self.today = datetime.now().date()
        self.start = start
        self.reports = reports
        self.backfill_days = 3
        self.platform = 'disk'
        self.err429 = False

    def download_file_from_yandex_disk(self,file_link):
        try:
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"OAuth {self.token}",
                "Accept": "application/json"
            }
            query_data = {"path": file_link.replace("\\", "/")}
            response = requests.get(
                "https://cloud-api.yandex.net/v1/disk/resources/download",
                headers=headers,
                params=query_data
            )
            if response.status_code not in [200, 201]:
                raise Exception(f"Ошибка при получении ссылки для скачивания: {response.status_code}")
            result = response.json()
            download_url = result.get("href")
            if not download_url:
                raise Exception("Ссылка для скачивания не найдена в ответе")
            file_response = requests.get(download_url, headers=headers)
            if file_response.status_code != 200:
                raise Exception(f"Ошибка при загрузке файла: {file_response.status_code}")
            file_content = file_response.content
            message = f'Платформа: DISK. Имя: {self.add_name}. Файл: {file_link}. Функция: download_file_from_yandex_disk. Результат: ОК.'
            self.common.log_func(self.bot_token, self.chat_list, message, 2)
            return file_content
        except Exception as e:
            message = f'Платформа: DISK. Имя: {self.add_name}. Файл: {file_link}. Функция: download_file_from_yandex_disk. Ошибка: {e}.'
            self.common.log_func(self.bot_token, self.chat_list, message, 3)
            raise


    def get_folder_contents(self,folder_link):
        try:
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"OAuth {self.token}",
                "Accept": "application/json"
            }
            query_data = {
                "path": folder_link.replace("\\", "/"),
                "limit": 1000000
            }
            response = requests.get(
                "https://cloud-api.yandex.net/v1/disk/resources",
                headers=headers,
                params=query_data
            )
            if response.status_code != 200:
                raise Exception(f"Ошибка при получении списка файлов: {response.status_code}")
            result = response.json()
            items = result.get("_embedded", {}).get("items", [])
            folder_contents = []
            for item in items:
                folder_contents.append(item.get("path").replace("disk:/", ""))
            message = f'Платформа: DISK. Имя: {self.add_name}. Файл: {folder_link}. Функция: get_folder_contents. Результат: ОК.'
            self.common.log_func(self.bot_token, self.chat_list, message, 2)
            return folder_contents
        except Exception as e:
            message = f'Платформа: DISK. Имя: {self.add_name}. Файл: {folder_link}. Функция: get_folder_contents. Ошибка: {e}.'
            self.common.log_func(self.bot_token, self.chat_list, message, 3)
            raise

    def parse_file(self,file_content: bytes):
        try:
            try:
                with io.BytesIO(file_content) as file:
                    wb = openpyxl.load_workbook(file, data_only=True)
                    result = {}
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        rows = list(ws.iter_rows(values_only=True))
                        if not rows:
                            continue
                        headers = rows[0]
                        headers = self.common.tuple_none_change(headers)
                        headers = tuple(map(self.common.transliterate_key, headers))
                        data = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
                        result[sheet] = data
                    return result  # Возвращает словарь с данными всех листов
            except Exception:
                pass
            try:
                decoded_content = file_content.decode('utf-8')
                with io.StringIO(decoded_content) as f:
                    reader = csv.DictReader(f)
                    new_fieldnames = [self.common.transliterate_key(header) for header in reader.fieldnames]
                    reader.fieldnames = new_fieldnames
                    return {'csv': list(reader)}
            except Exception as e:
                raise ValueError(str(e))
        except Exception as e:
            message = f'Платформа: DISK. Имя: {self.add_name}. Функция: parse_file. Ошибка: {e}.'
            self.common.log_func(self.bot_token, self.chat_list, message, 3)
            raise


    def get_file(self, date=''):
        try:
            file_content = self.download_file_from_yandex_disk(self.link)
            final_result = self.parse_file(file_content)
            message = f'Платформа: DISK. Имя: {self.add_name}. Функция: get_file. Результат: ОК'
            self.common.log_func(self.bot_token, self.chat_list, message, 1)
            return final_result
        except Exception as e:
            message = f'Платформа: DISK. Имя: {self.add_name}. Функция: get_file. Ошибка: {e}.'
            self.common.log_func(self.bot_token, self.chat_list, message, 3)
            return message


    def disk_collecting_report(self,  upload_table, data,   delay):
        try:
            self.clickhouse.test_clickhouse_connection()
            create_table_query_collect = f"""
                CREATE TABLE IF NOT EXISTS {self.platform}_collection_{self.add_name} (
                date Date, report String, collect Bool ) ENGINE = ReplacingMergeTree(collect) ORDER BY (report, date)"""
            self.clickhouse.ch_execute(create_table_query_collect)
            time.sleep(4)
            date = self.today.strftime('%Y-%m-%d')
            table_name = f'{self.platform}_{upload_table}_{self.add_name}'
            text_columns_set = self.clickhouse.ch_text_columns_set(table_name)
            refresh = f"TRUNCATE TABLE {table_name};"
            if not self.common.is_error(data):
                collect = True
                collection_data = pd.DataFrame(
                    {'date': pd.to_datetime([date], format='%Y-%m-%d'), 'report': [upload_table], 'collect': [collect]})
                if self.common.is_empty(data):
                    message = f'Платформа: {self.platform}. Имя: {self.add_name}. Таблица: {upload_table}. Дата: {date}. ПУСТОЙ ОТВЕТ!'
                    self.common.log_func(self.bot_token, self.chat_list, message, 2)
                if not self.common.is_empty(data):
                    self.clickhouse.create_alter_ch(data, table_name, 'timeStamp', '', 'MergeTree')
                    df = self.common.check_and_convert_types(data, '', '', text_columns_set)
                if self.clickhouse.ch_check(table_name):
                    self.clickhouse.ch_execute(refresh)
                if not self.common.is_empty(data):
                    self.clickhouse.ch_insert(df, table_name)
                self.clickhouse.ch_insert(collection_data, f'{self.platform}_collection_{self.add_name}')
            message = f'Платформа: {self.platform}. Имя: {self.add_name}. Таблица: {upload_table}. Функция: disk_collecting_report. Результат: ОК'
            self.common.log_func(self.bot_token, self.chat_list, message, 2)
        except Exception as e:
            message = f'Платформа: {self.platform}. Имя: {self.add_name}. Таблица: {upload_table}. Функция: disk_collecting_report. Ошибка сбора: {e}'
            self.common.log_func(self.bot_token, self.chat_list, message, 3)

    def collecting_manager(self):
        delay = 10
        report_list = self.reports.replace(' ', '').lower().split(',')
        for report in report_list:
            self.clickhouse = Clickhouse( self.bot_token, self.chat_list, self.message_type, self.host, self.port, self.username, self.password,
                                             self.database, self.start, self.add_name, self.err429, self.backfill_days, self.platform)
            if report == 'file':
                file_data = self.get_file()
                for key, value in file_data.items():
                    upload_table = self.common.transliterate_key(
                        self.link.replace(' ',"-").replace('.xlsx','_').replace('.xls','_').replace('.csv','_')
                        .replace(r'/','_').replace(r'\\','_') +key.strip()
                    ).replace('__',"_")
                    data = value
                    self.disk_collecting_report(upload_table,                        data,                        delay                    )
            if report == 'folder':
                files = self.get_folder_contents(self.link)
                for file in files:
                    file_data = self.parse_file(self.download_file_from_yandex_disk(file))
                    for key, value in file_data.items():
                        upload_table = self.common.transliterate_key(
                            file.replace(' ',"-").replace('.xlsx','_').replace('.xls','_').replace('.csv','_')
                            .                    replace(r'/','_').replace(r'\\','_') +key.strip()
                        ).replace('__',"_")
                        data = value
                        self.disk_collecting_report(upload_table,                        data,                        delay                    )
        self.common.send_logs_clear_anyway(self.bot_token, self.chat_list)


