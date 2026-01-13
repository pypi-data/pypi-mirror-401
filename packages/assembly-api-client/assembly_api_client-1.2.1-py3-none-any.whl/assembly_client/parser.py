"""Parser for Korean National Assembly API Excel specifications."""

from __future__ import annotations

import asyncio
import json
import logging
from dataclasses import asdict, dataclass
from pathlib import Path

import httpx
import openpyxl
import platformdirs

from .errors import SpecParseError

logger = logging.getLogger(__name__)


@dataclass
class APIParameter:
    """Represents a single API parameter."""

    name: str
    type: str
    required: bool
    description: str

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> APIParameter:
        return cls(**data)


@dataclass
class APISpec:
    """Represents a parsed API specification."""

    service_id: str
    endpoint: str
    endpoint_url: str
    basic_params: list[APIParameter]
    request_params: list[APIParameter]
    response_fields: list[APIParameter]

    def to_dict(self) -> dict:
        return {
            "service_id": self.service_id,
            "endpoint": self.endpoint,
            "endpoint_url": self.endpoint_url,
            "basic_params": [p.to_dict() for p in self.basic_params],
            "request_params": [p.to_dict() for p in self.request_params],
            "response_fields": [p.to_dict() for p in self.response_fields],
        }

    @classmethod
    def from_dict(cls, data: dict) -> APISpec:
        return cls(
            service_id=data["service_id"],
            endpoint=data["endpoint"],
            endpoint_url=data["endpoint_url"],
            basic_params=[APIParameter.from_dict(p) for p in data["basic_params"]],
            request_params=[APIParameter.from_dict(p) for p in data["request_params"]],
            response_fields=[APIParameter.from_dict(p) for p in data.get("response_fields", [])],
        )


class SpecParser:
    """Parser for Excel API specification files."""

    # Order of infSeq values to try when downloading specs
    DEFAULT_INF_SEQ_ORDER = [2, 1, 3, 4, 5]

    # Excel files (.xlsx) are ZIP archives with this magic number
    EXCEL_MAGIC_NUMBERS = [
        b"PK\x03\x04",  # Standard ZIP file (used by .xlsx)
        b"PK\x05\x06",  # Empty ZIP archive
        b"PK\x07\x08",  # Spanned ZIP archive
    ]

    def __init__(self, cache_dir: Path | None = None):
        """
        Initialize the spec parser.

        Args:
            cache_dir: Directory to cache parsed JSON specs.
                       If None, uses user cache directory (e.g., ~/.cache/assembly-api-client/specs).
        """
        if cache_dir is None:
            cache_base = Path(platformdirs.user_cache_dir("assembly-api-client"))
            self.cache_dir = cache_base / "specs"
        else:
            self.cache_dir = cache_dir

        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _is_valid_excel_file(self, content: bytes) -> bool:
        """
        Validate that the content is a valid Excel/ZIP file by checking magic numbers.
        """
        if len(content) < 4:
            return False
        return any(content[: len(magic)] == magic for magic in self.EXCEL_MAGIC_NUMBERS)

    def save_spec_json(self, spec: APISpec, output_dir: Path, filename: str | None = None) -> Path:
        """Save APISpec to a JSON file."""
        output_dir.mkdir(parents=True, exist_ok=True)
        name = filename or spec.service_id
        if not name.endswith(".json"):
            name += ".json"
        output_file = output_dir / name
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(spec.to_dict(), f, ensure_ascii=False, indent=2)
        return output_file

    async def _download_excel_bytes(self, service_id: str, inf_seq: int = 2) -> bytes:
        """
        Download Excel specification file content into memory.
        """
        url = f"https://open.assembly.go.kr/portal/data/openapi/downloadOpenApiSpec.do?infId={service_id}&infSeq={inf_seq}"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

        try:
            async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
                response = await client.get(url, headers=headers)
                try:
                    response.raise_for_status()
                except httpx.HTTPStatusError as exc:
                    raise SpecParseError(
                        f"Failed to download spec for {service_id}: {exc.response.status_code}"
                    ) from exc

                content = response.content
                if len(content) < 100:
                    raise SpecParseError(f"Downloaded content too small: {len(content)} bytes")

                if not self._is_valid_excel_file(content):
                    # Provide diagnostic information
                    preview = content[:200].decode("utf-8", errors="replace")
                    is_html = content.startswith(b"<!DOCTYPE") or content.startswith(b"<html")

                    error_msg = f"Downloaded content for {service_id} is not a valid Excel file."
                    if is_html:
                        error_msg += (
                            f"\n\nThe server returned an HTML page instead of an Excel file. "
                            f"This usually means:\n"
                            f"1. The service ID '{service_id}' is invalid or not found\n"
                            f"2. The infSeq parameter ({inf_seq}) is incorrect\n"
                            f"3. The public data portal's spec download endpoint has changed\n\n"
                            f"Content preview: {preview[:100]}..."
                        )
                    else:
                        error_msg += (
                            f"\n\nContent starts with: {content[:50]!r}\n"
                            f"Expected Excel magic numbers (ZIP format): PK\\x03\\x04\n"
                            f"This may indicate a temporary server error or API change."
                        )

                    raise SpecParseError(error_msg)

                logger.info(f"Downloaded spec for {service_id} ({len(content)} bytes)")
                return content

        except httpx.HTTPError as e:
            raise SpecParseError(f"Network error downloading spec for {service_id}: {e}") from e

    async def parse_spec(self, service_id: str, inf_seq: int = 2) -> APISpec:
        """
        Get API specification.
        1. Checks if JSON spec exists in cache.
        2. If not, downloads Excel to memory, parses it, saves JSON to cache, and returns it.

        Note: If download fails with the provided infSeq value, automatically retries with 
        other infSeq values (1-5) as a fallback.
        """
        json_file = self.cache_dir / f"{service_id}.json"

        # 1. Try Cache
        if json_file.exists():
            try:
                with open(json_file, encoding="utf-8") as f:
                    data = json.load(f)
                    logger.debug(f"Loaded spec for {service_id} from cache")
                    return APISpec.from_dict(data)
            except Exception as e:
                logger.warning(f"Failed to load cached spec for {service_id}, re-downloading: {e}")

        # 2. Download and Parse (Stream Processing with automatic fallback)
        excel_content = None
        
        # Try a range of infSeq values as some services use non-standard ones (1, 2, 3...)
        attempt_seqs = [inf_seq] + [s for s in self.DEFAULT_INF_SEQ_ORDER if s != inf_seq]
        last_error = None

        for seq in attempt_seqs:
            try:
                excel_content = await self._download_excel_bytes(service_id, seq)
                logger.info(f"Successfully downloaded spec for {service_id} using infSeq={seq}")
                break
            except SpecParseError as e:
                last_error = e
                logger.debug(f"Failed to download spec for {service_id} with infSeq={seq}: {e}")
                continue

        if excel_content is None:
            raise SpecParseError(
                f"Failed to download spec for {service_id} after trying infSeq values {attempt_seqs}.\n"
                f"Last error: {last_error}"
            )

        def _parse_sync(content: bytes):
            import re
            from io import BytesIO

            try:
                wb = openpyxl.load_workbook(BytesIO(content))
                ws = wb["Sheet1"]

                # Extract endpoint URL
                endpoint_url = self._extract_endpoint_url(ws)
                if not endpoint_url:
                    raise SpecParseError(f"Could not find endpoint URL in spec for {service_id}")

                endpoint = endpoint_url.split("/")[-1]

                # Extract parameters
                basic_params = []
                request_params = []
                response_fields = []

                in_basic_section = False
                in_request_section = False
                in_response_section = False

                for row in ws.iter_rows(min_row=1, values_only=True):
                    if not row or not any(row):
                        continue

                    first_cell = str(row[0]) if row[0] else ""

                    if "기본인자" in first_cell:
                        in_basic_section = True
                        in_request_section = False
                        in_response_section = False
                        continue
                    elif "요청인자" in first_cell:
                        in_basic_section = False
                        in_request_section = True
                        in_response_section = False
                        continue
                    elif "출력값" in first_cell or "출력명" in first_cell:
                        in_basic_section = False
                        in_request_section = False
                        in_response_section = True
                        continue

                    if (in_basic_section or in_request_section or in_response_section) and len(row) >= 3 and row[1]:
                        # For response fields, row[1] might be the field name or description depending on format
                        # Standard format: Name | Type | Description

                        type_str = str(row[1])

                        # Basic/Request params have "필수"/"선택" in type
                        is_param = "필수" in type_str or "선택" in type_str

                        if in_response_section:
                            # Heuristic to find the Field Name (English) and Description (Korean)
                            # Common formats:
                            # 1. Name | Description | Type
                            # 2. No | Name | Description | Type
                            # 3. No | Description | Name | Type

                            # Skip header rows
                            if (
                                "출력" in str(row[0])
                                or "설명" in str(row[0])
                                or "No" in str(row[0])
                                or "순번" in str(row[0])
                            ):
                                continue

                            # Find the column that looks like an English Key (uppercase, underscores)
                            field_name = ""
                            description = ""
                            found_key = False

                            for cell in row:
                                if not cell:
                                    continue
                                s = str(cell).strip()
                                # Check if it looks like an API Key (e.g. BILL_ID, AGE, etc)
                                # Must be mostly ASCII, maybe uppercase, no Korean
                                if re.match(r"^[A-Z0-9_]+$", s) and not re.search(r"[가-힣]", s):
                                    # Avoid numbers like "1", "1.0" unless they are the only thing?
                                    # But "1" is likely a sequence number.
                                    # Let's assume keys are at least 2 chars or contain letters?
                                    # Some keys might be "ID".
                                    if s.replace(".", "").isdigit():
                                        continue

                                    field_name = s
                                    found_key = True
                                    break

                            if found_key:
                                # Description is usually the cell with Korean
                                for cell in row:
                                    if not cell:
                                        continue
                                    s = str(cell).strip()
                                    if re.search(r"[가-힣]", s):
                                        description = s
                                        break

                                    # Fallback logic
                                    pass

                                # Skip standard error/info codes
                                if field_name in ["ERROR", "INFO", "CODE", "MESSAGE"]:
                                    continue

                                param = APIParameter(
                                    name=field_name,
                                    type="String",  # Default to String as type info is often messy
                                    required=False,
                                    description=description,
                                )
                                response_fields.append(param)

                        elif is_param:
                            param = APIParameter(
                                name=str(row[0]),
                                type=type_str,
                                required="필수" in type_str,
                                description=str(row[2]) if len(row) > 2 and row[2] else "",
                            )
                            if in_basic_section:
                                basic_params.append(param)
                            else:
                                request_params.append(param)

                spec = APISpec(
                    service_id=service_id,
                    endpoint=endpoint,
                    endpoint_url=endpoint_url,
                    basic_params=basic_params,
                    request_params=request_params,
                    response_fields=response_fields,
                )

                # Save to JSON cache
                self.save_spec_json(spec, self.cache_dir)
                return spec

            except Exception as e:
                raise SpecParseError(f"Failed to parse spec for {service_id}: {e}") from e

        return await asyncio.to_thread(_parse_sync, excel_content)

    def clear_cache(self, service_id: str | None = None) -> None:
        """
        Remove cached JSON spec.
        If service_id is provided, removes only that spec.
        If None, removes all specs in the cache directory.
        """
        if service_id:
            json_file = self.cache_dir / f"{service_id}.json"
            if json_file.exists():
                json_file.unlink()
                logger.debug(f"Cleared cache for {service_id}")
        else:
            for json_file in self.cache_dir.glob("*.json"):
                json_file.unlink()
            logger.debug("Cleared all cache")

    def _extract_endpoint_url(self, worksheet) -> str | None:
        """
        Extract endpoint URL from worksheet.

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            Endpoint URL or None if not found
        """
        for row in worksheet.iter_rows(min_row=1, max_row=50, max_col=1):
            cell = row[0]
            if cell.value and "요청주소" in str(cell.value):
                # Next row should contain the URL
                next_row_value = worksheet.cell(cell.row + 1, 1).value
                if next_row_value and "https://" in str(next_row_value):
                    url = str(next_row_value).strip().replace("- ", "")
                    return url
        return None


def load_service_map(cache_dir: Path) -> dict[str, str]:
    """Load service ID to Name mapping from cached master list."""
    service_map = {}
    master_file = cache_dir / "all_apis.json"

    if not master_file.exists():
        return {}

    try:
        with open(master_file, encoding="utf-8") as f:
            data = json.load(f)
            if "OPENSRVAPI" in data:
                for item in data["OPENSRVAPI"]:
                    if "row" in item:
                        for row in item["row"]:
                            inf_id = row.get("INF_ID")
                            inf_nm = row.get("INF_NM")
                            if inf_id and inf_nm:
                                service_map[inf_id] = inf_nm
    except Exception as e:
        logger.error(f"Failed to load master list {master_file}: {e}")

    return service_map


def load_service_metadata(cache_dir: Path) -> dict[str, dict[str, str]]:
    """Load comprehensive service metadata from cached master list.

    Returns:
        Dictionary mapping service_id to metadata dict with keys:
        - name: Service name (INF_NM)
        - description: Service description (INF_EXP)
        - category: Category (CATE_NM)
        - organization: Organization (ORG_NM)
        - endpoint: Service URL (SRV_URL)
    """
    service_metadata = {}
    master_file = cache_dir / "all_apis.json"

    if not master_file.exists():
        return {}

    try:
        with open(master_file, encoding="utf-8") as f:
            data = json.load(f)
            if "OPENSRVAPI" in data:
                for item in data["OPENSRVAPI"]:
                    if "row" in item:
                        for row in item["row"]:
                            inf_id = row.get("INF_ID")
                            if inf_id:
                                service_metadata[inf_id] = {
                                    "name": row.get("INF_NM", ""),
                                    "description": row.get("INF_EXP", ""),
                                    "category": row.get("CATE_NM", ""),
                                    "organization": row.get("ORG_NM", ""),
                                    "endpoint": row.get("SRV_URL", ""),
                                }
    except Exception as e:
        logger.error(f"Failed to load service metadata {master_file}: {e}")

    return service_metadata
