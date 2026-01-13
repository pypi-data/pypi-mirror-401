import os
import ast
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import ollama
import re

prompt_hdr = """
Write 2 test cases in natural English language to test the below Python 
function. Each test case should be less than 30 words and should start with 
words "Verify that"
Output should be in format
1. Verify that <test case>
2. Verify that <test case>

Python function -
""".strip()


def extract_functions_from_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        source = file.read()

    tree = ast.parse(source)
    functions = []

    class FunctionVisitor(ast.NodeVisitor):
        def __init__(self):
            self.class_name = None

        def visit_ClassDef(self, node):
            prev_class = self.class_name
            self.class_name = node.name
            self.generic_visit(node)
            self.class_name = prev_class

        def visit_FunctionDef(self, node):
            start_line = node.lineno - 1
            end_line = max(getattr(node, 'end_lineno', start_line + 1),
                           start_line + 1)
            func_body = ''.join(
                source.splitlines(keepends=True)[start_line:end_line])
            if not node.name.startswith("_"):
                functions.append((self.class_name, node.name, func_body))

    FunctionVisitor().visit(tree)
    return functions


def collect_py_files(folder_path):
    py_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.py'):
                py_files.append(os.path.join(root, file))
    return py_files


def parse_output(txt):
    ln = [x for x in txt.strip().splitlines() if 'verify that' in x.lower()]
    ln_1 = [re.sub(r'^\s*\d+\.', '', x) for x in ln]
    ln_2 = [x.strip() for x in ln_1 if len(x.strip()) > 15]
    return ln_2[:2]


def generate_test_cases(folder_path, excel_path, model, retry_attempts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Python Functions"

    headers = ["File Name", "Class Name", "Function Name", "Test Case"]
    ws.append(headers)

    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Set column widths
    column_widths = [20, 20, 20, 50]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  # 64 + 1 = 'A'

    # Get all .py files
    py_files = collect_py_files(folder_path)

    for file_path in py_files:
        print(f"Processing file {file_path}")
        try:
            functions = extract_functions_from_file(file_path)
            for class_name, func_name, func_body in functions:
                print(f"Getting test cases from model for function - "
                      f"{class_name}/{func_name}")
                rem_attempts = 2
                out_ar = []
                while rem_attempts:
                    response = ollama.chat(model=model, messages=[
                        {
                            'role': 'user',
                            'content': f'{prompt_hdr}\n{func_body}'
                        },
                    ])
                    rem_attempts -= 1
                    out_ar = parse_output(response['message']['content'])
                    if out_ar:
                        break
                    elif rem_attempts > 0:
                        print("Failed attempt. Retrying")
                    else:
                        print("Max retry attempts reached. Skipping this " +
                              "function")
                if out_ar:
                    for test_case in out_ar:
                        row = [os.path.relpath(file_path, folder_path),
                               class_name or "", func_name, test_case]
                        ws.append(row)
                else:
                    row = [os.path.relpath(file_path, folder_path),
                           class_name or "", func_name, 'Unable to get test ' +
                           'case from model']
                    ws.append(row)
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    # Enable word wrap for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Add auto filter
    max_row = ws.max_row
    max_col = ws.max_column
    ws.auto_filter.ref = f"A1:{chr(64 + max_col)}{max_row}"

    wb.save(excel_path)
