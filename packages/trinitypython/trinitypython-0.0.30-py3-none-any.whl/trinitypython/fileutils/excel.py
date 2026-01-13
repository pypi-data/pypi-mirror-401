from pathlib import Path
import openpyxl
from openpyxl.utils import column_index_from_string
import json
import re


class Excel:
    def __init__(self, fl):
        self.fl = fl
        self.flpth = Path(fl)
        self.workbook = openpyxl.load_workbook(self.fl, data_only=True)

    def extract_urls(self, refcols=None):
        if refcols is None:
            refcols = []
        urls = []
        colkeys = ["file_name", "sheet_name", "row_number", "column_name",
                   "type"]
        for col in refcols:
            colkeys.append("Ref " + col)
        colkeys.extend(["hyperlink_text", "hyperlink_URL"])
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    lnk = []
                    if cell.hyperlink:
                        hyperlink_info = {
                            'file_name': self.fl,
                            'sheet_name': sheet_name,
                            'row_number': cell.row,
                            'column_name': cell.column_letter,
                            'type': 'Explicit',
                            'hyperlink_text': cell.value,
                            'hyperlink_URL': cell.hyperlink.target
                        }
                        lnk.append(str(cell.hyperlink.target))
                        lnk.append(str(cell.value))
                        for col in refcols:
                            hyperlink_info["Ref " + col] = row[
                                column_index_from_string(col) - 1].value
                        urls.append(hyperlink_info)
                    for mtch in re.findall(r"(https:|http:|www\.)(\S*)",
                                           str(cell.value)):
                        if ''.join(mtch) in lnk:
                            continue
                        hyperlink_info = {
                            'file_name': self.fl,
                            'sheet_name': sheet_name,
                            'row_number': cell.row,
                            'column_name': cell.column_letter,
                            'type': 'Implicit',
                            'hyperlink_text': ''.join(mtch),
                            'hyperlink_URL': ''.join(mtch)
                        }
                        for col in refcols:
                            hyperlink_info["Ref " + col] = row[
                                column_index_from_string(col) - 1].value
                        urls.append(hyperlink_info)
                        lnk.append(''.join(mtch))

        return {'keys': colkeys, 'data': urls}
