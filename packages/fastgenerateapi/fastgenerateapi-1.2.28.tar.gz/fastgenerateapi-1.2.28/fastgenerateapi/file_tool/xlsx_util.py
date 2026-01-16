import abc
import importlib
import io
import operator
from collections.abc import Callable
from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import List, Optional, Dict, Type

import openpyxl
from fastapi import UploadFile, HTTPException
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from starlette._utils import is_async_callable
from tortoise import Model
from pydantic import ValidationError

from fastgenerateapi import BaseModel
from fastgenerateapi.api_view.mixin.dbmodel_mixin import DBModelMixin
from fastgenerateapi.api_view.mixin.response_mixin import ResponseMixin
from fastgenerateapi.api_view.mixin.tool_mixin import ToolMixin
from fastgenerateapi.file_tool.common import PushStatusInterface


class XlsxImporterInterface(PushStatusInterface):
    # 建议修改
    file: Optional[UploadFile] = None
    xlsx_headers: List[str] = None  # 导入首行校验
    fields: List[str] = None  # 二选一：每列对应的字段名
    title_field_dict: Dict[str, str] = None  # 二选一：标题转换为字段
    model_class: Optional[Type[Model]] = None
    create_schema: Optional[Type[BaseModel]] = None
    # 一般默认
    fields_handler: Callable = None
    field_handler_list: List[Callable] = None
    file_save_path = None
    sheet_name = None

    async def get_data_dict(self, sh: Worksheet) -> List[Dict[str, str]]:
        pass

    async def check_headers(self, row: tuple[Cell, ...]):
        pass

    async def handle_data(self, rows: List[Dict[str, str]]):
        pass

    async def upload_file(self, bytes_io: io.BytesIO):
        pass


class XlsxTemplateExporterInterface(abc.ABC):
    # 建议调整
    filename: Optional[str] = '导入模板.xlsx'
    sheet_name = "Sheet1"  # model_class._meta.table_description
    file_path = None  # 二选一：固定模板导出
    xlsx_headers: List[str] = None  # 二选一：自动生成模板导出
    # 一般默认
    model_class: Optional[Model] = None
    height = 18
    height_num = 1
    default_align = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=True,
        indent=0,
    )
    default_fill = PatternFill(start_color=Color(COLOR_INDEX[44]), end_color=Color(COLOR_INDEX[44]), fill_type='solid')
    thin_border = Border(top=Side('thin'), left=Side('thin'), right=Side('thin'), bottom=Side('thin'))
    col_width_list: List[int] = []

    def write_headers(self, sh: Worksheet) -> List[int]:
        """
        写入第一行信息
        :return:
        """
        pass

    def adaptive_format(self, sh: Worksheet):
        """
        自适应宽度
        :return:
        """
        pass


class XlsxExporterInterface(XlsxTemplateExporterInterface, PushStatusInterface):
    # 建议调整
    model_list: List[Model] = None  # 导出数据模型
    fields: List[str] = None  # 导出模型对应字段
    model_handler: Callable = None  # 数据列表模型处理方法
    model_handler_list: List[Callable] = None  # 单条数据模型处理方法
    fields_handler: Dict[str, Callable] = None  # 字段处理方法
    filename: Optional[str] = '导出文件.xlsx'  # 导出设定文件名
    # 一般默认
    index: Optional[bool] = True  # 是否添加序号
    file_save_path: Optional[str] = None  # 文件保存路径

    async def get_data(self) -> List[List[str]]:
        pass

    async def load_excel(self) -> Workbook:
        pass

    async def write_sheet(self, sh: Worksheet, data_list: List[List[str]]):
        pass

    async def upload_file(self, bytes_io: io.BytesIO):
        pass


class XlsxImporterImpl(XlsxImporterInterface):

    def __init__(self, file: Optional[UploadFile], *args, **kwargs):
        self.file = file
        self.args = args
        self.kwargs = kwargs

    # ------------------------------ 钩子函数 ------------------------------------
    async def handle_data(self, rows: List[Dict[str, str]]):
        if not self.model_class or not self.create_schema:
            raise NotImplementedError
        create_list = []
        global_dict = {}
        if self.fields_handler:
            if is_async_callable(self.fields_handler):
                rows = await self.fields_handler(rows, *self.args, **self.kwargs, global_dict=global_dict)
            else:
                rows = self.fields_handler(rows, *self.args, **self.kwargs, global_dict=global_dict)
        for row in rows:
            try:
                if self.create_schema:
                    create_dict = self.create_schema(**row).model_dump(exclude_unset=True)
                else:
                    create_dict = row
                if self.field_handler_list:
                    for field_handler in self.field_handler_list:
                        if is_async_callable(field_handler):
                            create_dict = await field_handler(create_dict, *self.args, **self.kwargs, global_dict=global_dict)
                        else:
                            create_dict = field_handler(create_dict, *self.args, **self.kwargs, global_dict=global_dict)
                if not create_dict:
                    continue
                create_obj = self.model_class(**create_dict)
            except ValidationError as e:
                alise_dict = ToolMixin.get_schema_alise_to_name(self.create_schema)
                _error_field = e.errors()[0].get('loc')[0]
                error_field = alise_dict.get(_error_field, _error_field)
                description = DBModelMixin.get_field_description(self.model_class, error_field)
                if not row.get(error_field):
                    return ResponseMixin.error(msg=f"第{row['_row_index']}行【{description}】不能为空")
                return ResponseMixin.error(msg=f"第{row['_row_index']}行【{description}】填写错误")
            create_list.append(create_obj)
        await self.model_class.bulk_create(create_list)
        return create_list

    async def upload_file(self, bytes_io: io.BytesIO):
        raise NotImplementedError

    async def push_status_progress(self):
        raise NotImplementedError

    async def push_status_success(self):
        raise NotImplementedError

    async def push_status_error(self, msg: Optional[str]):
        raise NotImplementedError

    # ------------------------------ 复用函数 ------------------------------------
    async def check_headers(self, row: tuple[Cell, ...]):
        if not self.xlsx_headers:
            raise NotImplementedError
        header_list = []
        add_field = not self.fields and self.title_field_dict
        field_list = []
        for col in row:
            header_list.append(col.value)
            if add_field:
                field_value = self.title_field_dict.get(col.value)
                if not field_value:
                    return ResponseMixin.error(msg=f"文件首行【{col.value}】不正确")
                field_list.append(field_value)
        if len(header_list) != len(self.xlsx_headers):
            return ResponseMixin.error(msg="文件首行长度校验错误")
        if not operator.eq(header_list, self.xlsx_headers):
            return ResponseMixin.error(msg="文件首行内容校验错误")
        if add_field and field_list:
            self.fields = field_list

        return

    async def get_data_dict(self, sh: Worksheet) -> List[Dict[str, str]]:
        data_list = []
        for row in range(2, sh.max_row + 1):
            row_data = await self.row_to_dict(sh[row], self.fields)
            if not row_data:
                continue
            row_data["_row_index"] = row
            data_list.append(row_data)

        return data_list

    @staticmethod
    async def row_to_dict(row, fields: List[str]) -> dict:
        row_data = {}
        for i in range(min(len(row), len(fields))):
            row_value = row[i].value
            if row_value not in [None, "", " "]:
                if isinstance(row_value, int):
                    row_value = str(row_value)
                row_data[fields[i]] = row_value
            else:
                row_data[fields[i]] = None
        if any(row_data.values()):
            return row_data
        return {}


class XlsxTemplateExporterImpl(XlsxTemplateExporterInterface):

    def write_headers(self, sh: Worksheet):
        """
        写入第一行信息
        :return:
        """
        sh.row_dimensions[str(1)].height = 26
        for col, header in enumerate(self.xlsx_headers, 1):
            sh.cell(1, col).value = header
            sh.cell(1, col).alignment = self.default_align
            sh.cell(1, col).fill = self.default_fill
            sh.cell(1, col).border = self.thin_border
            self.col_width_list.append(len(header.encode('gb18030')))

        return

    def adaptive_format(self, sh: Worksheet):
        """
        自适应宽度
        :param sh
        :return:
        """
        # 设置自适应列宽
        for i, col_max_len in enumerate(self.col_width_list, 1):
            # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
            max_width = col_max_len + 4
            if max_width > 256:
                max_width = 256
            sh.column_dimensions[get_column_letter(i)].width = max_width
        for y in range(2, self.height_num + 2):
            sh.row_dimensions[str(y)].height = self.height

        return


class XlsxExporterImpl(XlsxTemplateExporterImpl, XlsxExporterInterface):

    def __init__(self, *args, **kwargs):
        self.args = args
        self.kwargs = kwargs

    # ------------------------------ 钩子函数 ------------------------------------
    async def get_data(self) -> List[List[str]]:
        if not self.model_list or not self.fields:
            return [[]]
        if self.model_handler:
            if is_async_callable(self.model_handler):
                self.model_list = await self.model_handler(self.model_list, *self.args, **self.kwargs)
            else:
                self.model_list = self.model_handler(self.model_list, *self.args, **self.kwargs)
        data_list = []
        for model in self.model_list:
            if self.model_handler_list:
                for model_handler in self.model_handler_list:
                    if is_async_callable(model_handler):
                        model = await model_handler(model, *self.args, **self.kwargs)
                    else:
                        model = model_handler(model, *self.args, **self.kwargs)
            field_value_list = []
            for field in self.fields:
                info = getattr(model, field, "")
                if self.fields_handler:
                    handler = self.fields_handler.get(field)
                    if handler and hasattr(handler, "__call__"):
                        if is_async_callable(handler):
                            info = await handler(info, *self.args, **self.kwargs)
                        else:
                            info = handler(info, *self.args, **self.kwargs)
                field_value_list.append(info)
            data_list.append(field_value_list)
        return data_list

    async def upload_file(self, bytes_io: io.BytesIO):
        raise NotImplementedError

    async def push_status_progress(self):
        raise NotImplementedError

    async def push_status_success(self):
        raise NotImplementedError

    async def push_status_error(self, msg: Optional[str]):
        raise NotImplementedError

    # ------------------------------ 复用函数 ------------------------------------
    async def load_excel(self) -> Workbook:
        if self.file_path:
            wb = openpyxl.load_workbook(self.file_path)
        else:
            wb = Workbook()

        return wb

    async def write_sheet(self, sh: Worksheet, data_list: List[List[str]]):
        for row, datas in enumerate(data_list, 2):
            start_col = 1
            if self.index:
                self.write(sh, row, 1, str(row - 1))
                start_col += 1
            for col, data in enumerate(datas, start_col):
                self.write(sh, row, col, data)

        return

    def write(self, sh, row, col, value):
        if isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        sh.cell(row, col).value = value
        sh.cell(row, col).alignment = self.default_align
        if self.col_width_list[col - 1] < len(str(value).encode('gb18030')):
            self.col_width_list[col - 1] = len(str(value).encode('gb18030'))

        return


class XlsxUtil:

    @staticmethod
    async def start_import(importer: XlsxImporterInterface):
        with NamedTemporaryFile() as tmp2:
            tmp2.write(await importer.file.read())
            wb = openpyxl.load_workbook(tmp2, read_only=True, data_only=True)
            if importer.sheet_name:
                sh = wb[importer.sheet_name]
            else:
                sh = wb.active
            try:
                # 校验表头信息
                await importer.check_headers(sh[1])

                # 读取excel数据
                rows = await importer.get_data_dict(sh)
            finally:
                await importer.file.close()
                wb.close()
        if not rows:
            return ResponseMixin.error(msg="导入数据不能为空")
        # 处理完整数据
        await importer.handle_data(rows)

        return

    @staticmethod
    async def start_async_import(importer: XlsxImporterInterface):
        # 推送状态
        await importer.push_status_progress()
        try:
            with NamedTemporaryFile() as tmp2:
                tmp2.write(await importer.file.read())
                wb = openpyxl.load_workbook(tmp2, read_only=True, data_only=True)

                # 上传文件
                if importer.file_save_path:
                    wb.save(importer.file_save_path)
                else:
                    bytes_io = io.BytesIO()
                    wb.save(bytes_io)
                    bytes_io.seek(0)
                    await importer.upload_file(bytes_io)

                if importer.sheet_name:
                    sh = wb[importer.sheet_name]
                else:
                    sh = wb.active
                try:
                    # 校验表头信息
                    await importer.check_headers(sh[1])

                    # 读取excel数据
                    rows = await importer.get_data_dict(sh)
                finally:
                    await importer.file.close()
                    wb.close()

                if not rows:
                    return ResponseMixin.error(msg="导入数据不能为空")
                # 处理完整数据
                await importer.handle_data(rows)
        except HTTPException as e:
            await importer.push_status_error(e.detail)
            raise e
        except Exception as e:
            await importer.push_status_error(str(e))
            raise e

        await importer.push_status_success()

        return

    @staticmethod
    async def start_export_template(template_importer: XlsxTemplateExporterInterface):
        """
        导出excel文件模板
        :param template_importer: 导出模板实现类
        :return:
        """
        try:
            openpyxl = importlib.import_module("openpyxl")
        except Exception:
            return ResponseMixin.error(msg=f"please pip install openpyxl")
        if template_importer.file_path:
            wb = openpyxl.load_workbook(template_importer.file_path)
        else:
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = template_importer.sheet_name
            template_importer.write_headers(sh)
            template_importer.adaptive_format(sh)
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, filename=template_importer.filename, is_xlsx=True)

    @staticmethod
    async def start_export(exporter: XlsxExporterInterface):
        """
        导出excel文件
        :param exporter: 导出实现类
        :return:
        """
        data_list = await exporter.get_data()
        wb = await exporter.load_excel()
        sh = wb.active
        if exporter.sheet_name:
            sh.title = exporter.sheet_name
        exporter.write_headers(sh)
        await exporter.write_sheet(sh, data_list)
        exporter.adaptive_format(sh)

        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, filename=exporter.filename, is_xlsx=True)

    @staticmethod
    async def start_async_export(exporter: XlsxExporterInterface):
        """
        导出excel文件
        :param exporter: 导出实现类
        :return:
        """
        await exporter.push_status_progress()
        try:
            data_list = await exporter.get_data()
            wb = await exporter.load_excel()
            sh = wb.active
            if exporter.sheet_name:
                sh.title = exporter.sheet_name
            exporter.write_headers(sh)
            await exporter.write_sheet(sh, data_list)
            exporter.adaptive_format(sh)
            if exporter.file_save_path:
                wb.save(exporter.file_save_path)
            else:
                bytes_io = io.BytesIO()
                wb.save(bytes_io)
                bytes_io.seek(0)
                await exporter.upload_file(bytes_io)
        except HTTPException as e:
            await exporter.push_status_error(e.detail)
            raise e
        except Exception as e:
            await exporter.push_status_error(str(e))
            raise e

        await exporter.push_status_success()

        return
