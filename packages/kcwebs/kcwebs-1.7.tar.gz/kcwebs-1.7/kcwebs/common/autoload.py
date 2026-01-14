from kcws.common import *
from .. import config
import random,urllib,asyncio,websockets,smtplib,datetime,chardet,copy,multiprocessing,warnings,xlrd,threading
from kcwebs.utill.redis import redis as kcwsredis
from email.mime.text import MIMEText
from email.utils import formataddr
from mako.template import Template as kcwsTemplate
from mako.lookup import TemplateLookup
from kcwebs.utill.db import mysql as kcwsmysql
from kcwebs.utill.db import sqlite as kcwssqlite
from kcwebs.utill.cache import cache as kcwscache
from kcwebs.utill.db import mongodb as kcwsmongodb
from dateutil.relativedelta import relativedelta
from kcwebs.utill.http import Http
from kcwebs.utill.queues import Queues
from kcwebs.utill.db import model
from PIL import Image
import openpyxl
import pdf2docx
from pyquery import PyQuery as kcwebspq

redisobj=kcwsredis()
python_version=platform.python_version()
if python_version[0:3]!='3.8':
    print("\033[1;31;40m "+config.kcwebs['name']+"-"+config.kcwebs['version']+"依赖python3.8，与你现在的python"+python_version+"不兼容,推荐安装python3.8")
    exit()
class kcwebssocket:
    "websocket服务端"
    __clientlists={} #所有客户端绑定的对象

    __group={} #组
    __clientidgroup={}

    __uid={} #clientid绑定的uid
    __clientiduid={}

    def bindUid(self,clientid,uid):
        """将clientid与uid绑定，以便通过sendToUid(uid)发送数据

        clientid 客户端id

        uid uid与client_id是一对多的关系，系统允许一个uid下有多个client_id
        """
        try:
            self.__uid[uid]
        except KeyError:
            self.__uid[uid]=[]
        if clientid not in self.__uid[uid]:
            self.__uid[uid].append(clientid)
            self.__clientiduid[clientid]=uid
    def unbindUid(self,clientid,uid=0):
        """将clientid与uid解绑 当clientid下线（连接断开）时会自动与uid解绑，开发者无需在onClose事件调用unbindUid
        
        clientid 客户端id

        uid 数字或者字符串 非必传
        """
        if not uid:
            uid=self.getuid(clientid)
        if uid:
            try:
                self.__uid[uid]
            except KeyError:
                pass
            else:
                try:
                    self.__uid[uid].remove(clientid)
                except KeyError:
                    pass
                if not self.__uid[uid]:
                    del self.__uid[uid]
            try:
                del self.__clientiduid[clientid]
            except KeyError:
                pass
    def getuid(self,clientid):
        """通过clientid查询uid
        
        clientid 客户端id

        return uid
        """
        try:
            return self.__clientiduid[clientid]
        except KeyError:
            return 0
    def joinGroup(self,clientid,group):
        """将clientid加入某个组
        
        clientid 客户端id

        Group 组名 Group与client_id是一对多的关系，系统允许一个Group下有多个client_id
        """
        try:
            self.__group[group]
        except KeyError:
            self.__group[group]=[]
        if clientid not in self.__group[group]:
            self.__group[group].append(clientid)
            self.__clientidgroup[clientid]=group
    def getGroup(self,clientid):
        """通过clientid查询组名
        
        clientid 客户端id

        return 组名
        """
        try:
            return self.__clientidgroup[clientid]
        except KeyError:
            return ''
    def leaveGroup(self,clientid,group=''):
        """将clientid从某个组中删除
        
        clientid 客户端id

        group 组名
        """
        if not group:
            group=self.getGroup(clientid)
        if group:
            try:
                self.__group[group]
            except KeyError:
                pass
            else:
                try:
                    self.__group[group].remove(clientid)
                except KeyError:
                    pass
                if not self.__group[group]:
                    self.ungroup(group)
            try:
                del self.__clientidgroup[clientid]
            except KeyError:
                pass
    def ungroup(self,group):
        """解散分组。 解散分组后所有属于这个分组的用户的连接将被移出分组，此分组将不再存在，除非再次调用 joinGroup

        group 组名
        """
        clientidarr=self.__group[group]
        try:
            del self.__group[group]
        except KeyError:
            pass
        for clientid in clientidarr:
            try:
                del self.__clientidgroup[clientid]
            except KeyError:
                pass
    def getClientIdCountByGroup(self,group):
        """获取某分组当前在线成连接数（多少clientid在线）
        
        group 组名

        return int 返回一个数字
        """
        try:
            self.__group[group]
        except KeyError:
            return 0
        else:
            return len(self.__group[group])
    def getAllClientIdCount(self):
        """获取当前在线连接总数（多少client_id在线）
        
        return int 返回一个数字
        """
        return len(self.__clientlists.keys())
    def getGroupCount(self):
        """获取组数量
        
        return int 返回一个数字
        """
        return len(self.__group)
    def getGroupname(self):
        """获取组名称
        
        return list 返回一列表
        """
        return list(self.__group)
    
    async def sendToUid(self,uid,message):
        """向uid绑定的所有在线clientid发送数据

        uid uid可以是字符串、数字、或者包含uid的列表。如果为列表，则是给列表内所有uid发送数据

        message 要发送的数据（字符串类型）
        """
        if isinstance(uid,str):
            for clientid in self.__uid[uid]:
                try:
                    await self.send_client(clientid,message)
                except KeyError:
                    self.unbindUid(clientid,uid)
        elif isinstance(uid,list):
            for uids in uid:
                for clientid in self.__uid[uids]:
                    try:
                        await self.send_client(clientid,message)
                    except KeyError:
                        self.unbindUid(clientid,uids)
    async def sendToGroup(self,group,message,exclude_clientid=[]):
        """向某个分组的所有在线clientid发送数据。

        group 组名

        message 要发送的数据（字符串类型）

        exclude_clientid clientid组成的列表。exclude_clientid列表中指定的clientid将被排除在外，不会收到本次发的消息
        """
        try:
            self.__group[group]
        except KeyError:
            pass
        else:
            i=1
            for client in self.__group[group]:
                i+=1
                if exclude_clientid:
                    if client not in exclude_clientid:
                        try:
                            await self.send_client(client,message)
                        except KeyError:
                            self.leaveGroup(client,group)
                else:
                    try:
                        await self.send_client(client,message)
                    except KeyError:
                        self.leaveGroup(client,group)
    async def send_all(self,message):
        "给所有人发送消息，包括自己"
        for clientid in self.__clientlists.keys():
            await self.send_client(clientid,message)
    async def send_client(self,clientid,message):
        "给所指定客户端发送消息"
        websockets=self.__clientlists[clientid]
        try:
            await websockets.send(message)
        except:pass
    async def onConnect(self,clientid,params):
        "客户端发来连接时"
        if config.app['app_debug']:
            print("连接成功",clientid)
    async def onMessage(self,clientid,recv_text):
        "当客户端发来数据"
        await self.send_client(clientid,recv_text) #给当前用户发送消息
    async def onClose(self,clientid):
        "客户端与websocket的连接断开时触发"
        if config.app['app_debug']:
            print("onClose",clientid)
        await self.CloseSocket(clientid)
    async def CloseSocket(self,clientid):
        "关闭当前客户端socket"
        websockets=self.__clientlists[clientid]
        await websockets.close()
        del self.__clientlists[clientid]
        self.unbindUid(clientid)
        self.leaveGroup(clientid)
    async def __main2(self,clientid,websocket):
        "服务器端主逻辑"
        try:
            async for message in websocket:
                await self.onMessage(clientid, message)
        except:pass
        await self.onClose(clientid)
        await self.CloseSocket(clientid)
    async def __main1(self,clientid,path):
        t = urllib.parse.parse_qs(urllib.parse.urlparse(path).query)
        params={}
        for key in t.keys():
            params[key]=t[key][0]
        await self.onConnect(clientid, params)
    async def __main(self,websocket,path):
        "服务器端主逻辑"
        clientid=md5(str(random.random()))
        self.__clientlists[clientid]=websocket
        task1=asyncio.ensure_future(self.__main1(clientid,path))
        task2=asyncio.ensure_future(self.__main2(clientid,websocket))
        await task1
        await task2

        # task1=asyncio.create_task(self.__main1(clientid,path))
        # task2=asyncio.create_task(self.__main2(clientid,websocket,path))
        # await task1
        # await task2
    def start(self,ip='0.0.0.0',port='39020'):
        "启动websoeket服务"
        asyncio.set_event_loop(asyncio.new_event_loop()) # 防止出现RuntimeError
        asyncio.get_event_loop().run_until_complete(websockets.serve(self.__main,ip,port))
        asyncio.get_event_loop().run_forever()
# def start():
#     kwebsocket=kcwebssocket()
#     kwebsocket.start()
def img_is_con(image_path):
    """判断图片是否有内容"""
    img = Image.open(image_path)
    first_pixel = img.getpixel((0, 0))
    # 获取图片的宽度和高度
    width, height = img.size
    # 遍历图片的所有像素
    for x in range(width):
        for y in range(height):
            # 如果发现任何一个像素与第一个像素不同，则表示有内容
            if img.getpixel((x, y)) != first_pixel:
                return True
    # 如果所有像素都相同，则表示没有内容
    return False
def get_image_w_h(image_path):
    "获取图片宽高"
    with Image.open(image_path) as img:
        width, height = img.size
        return width, height
# from spire.pdf.common import *
from spire.pdf import *
def frgesregeeswfrgslkjhgfdertyu8765sdf44r56tg7uhik(pdfname,outname):
    # print('outname',outname)
    doc = PdfDocument()
    doc.LoadFromFile(pdfname)
    # print('outname',outname)
    doc.SaveToFile(outname, FileFormat.DOCX)
    doc.Close()
    # print('outname完成',outname)
def frgesregergslkjhgfdertyu876544r56tg7uhik(pdfname,outname):
    pdf = PdfDocument()
    # 加载PDF文件
    try:
        pdf.LoadFromFile(pdfname)
    except Exception as e:
        if config.app['app_debug']:
            print('pdf to xlsx-eee',e)
            print('pdf to xlsx-pdfname',pdfname)
        if 'Arg_NullReferenceException:' in str(e) or 'Font parsing exception:' in str(e):
            if os.name != 'nt':
                file_set_content(outname+'err','error：需要在windows系统中完成')
                return False
            else:
                return False
        elif 'Windows Error 0xe06d7363' in str(e):
            return False
        else:
            raise Exception(e)
    else:
        total_pages = pdf.Pages.Count # 计算总页数
        pdf.Close()
    if total_pages<=10: #pdf小于10 直接转换
        pdf = PdfDocument()
        pdf.LoadFromFile(pdfname) # 加载PDF文档
        convertOptions = XlsxLineLayoutOptions(True, True, False, True, False) # 创建 XlsxLineLayoutOptions 对象来指定转换选项
        pdf.ConvertOptions.SetPdfToXlsxOptions(convertOptions) # 设置转换选项
        try:
            pdf.SaveToFile(outname,FileFormat.XLSX) # 将PDF文档保存为Excel XLSX格式
        except Exception as e:
            if 'Arg_NullReferenceException:' in str(e) or 'Font parsing exception:' in str(e):
                if os.name != 'nt':
                    file_set_content(outname+'err','error：需要在windows系统中完成')
                    return False
                else:
                    if config.app['app_debug']:
                        print('pdf to xlsx',e)
                    return False
            elif 'Windows Error 0xe06d7363' in str(e):
                if config.app['app_debug']:
                    print('pdf to xlsx',e)
                return False
            else:
                raise Exception(e)
        pdf.Close()
    else:
        out_path="app/runtime/temp/"+md5(pdfname)
        try:
            split_pdf_by_page_count(input_file=pdfname,out_path=out_path,page_count=10)
            filearr=get_file(out_path)
            xlsxlists=[]
            for k in filearr:
                pdf = PdfDocument()
                pdf.LoadFromFile(k['path']) # 加载PDF文档
                convertOptions = XlsxLineLayoutOptions(True, True, False, True, False) # 创建 XlsxLineLayoutOptions 对象来指定转换选项
                pdf.ConvertOptions.SetPdfToXlsxOptions(convertOptions) # 设置转换选项
                pdf.SaveToFile(out_path+'/'+k['name']+'.xlsx', FileFormat.XLSX) # 将PDF文档保存为Excel XLSX格式
                pdf.Close()
                xlsxlists.append(out_path+'/'+k['name']+'.xlsx')
            workbook = openpyxl.Workbook()
            i=1
            for filename in xlsxlists:
                try:
                    wb = openpyxl.load_workbook(filename=filename)
                except Exception as e:
                    if 'Arg_NullReferenceException:' in str(e) or 'Font parsing exception:' in str(e):
                        if os.name != 'nt':
                            file_set_content(outname+'err','error：需要在windows系统中完成')
                            return False
                        else:
                            if config.app['app_debug']:
                                print('pdf to xlsx',e)
                            return False
                    elif 'Windows Error 0xe06d7363' in str(e):
                        if config.app['app_debug']:
                            print('pdf to xlsx',e)
                        return False
                    else:
                        raise Exception(e)
                else:
                    sheets =  wb.sheetnames
                    for sheet_name in sheets:
                        xls_sheet = wb[sheet_name]
                        xlsx_sheet = workbook.create_sheet(title='sheet'+str(i))
                        for row in xls_sheet.iter_rows():
                            for cell in row:
                                dst_cell = xlsx_sheet.cell(row=cell.row, column=cell.column)
                                dst_cell.value=copy.copy(cell.value)
                        i+=1
                try:
                    wb.close()
                except:pass
            workbook.save(outname)
            workbook.close()
        except Exception as e:
            if os.path.exists(out_path):
                shutil.rmtree(out_path)
            raise Exception(e)
        else:
            if os.path.exists(out_path):
                shutil.rmtree(out_path)
    return True
def lfjdsgjtr3535djdtdrgkrehsrsgtrssreiuiu43943uffh938raaqef(inputfile,outname):
    from spire import xls as spire_xls
    workbook = spire_xls.Workbook()
    workbook.LoadFromFile(inputfile)
    workbook.SaveToFile(outname, spire_xls.ExcelVersion.Version2016)
    workbook.Dispose()


def lfjdsgjtrdjdtdrgkgtrssreiuiu43943uffh938raaqef(xlsxfile,styles=True,imgpath=False,imgminsize=(10,10),outfiles=''):
    from spire import xls as spire_xls
    if imgpath:
        if imgpath[-1]=='/' or imgpath[-1]=='\\':
            imgpath=imgpath+md5(xlsxfile)+'/'
        else:
            imgpath=imgpath+'/'+md5(xlsxfile)+'/'
    try:
        work = openpyxl.load_workbook(filename=xlsxfile,data_only=True)
    except Exception as e:
        try:
            work.close()
        except:pass
        if 'list index out of range' in str(e) or 'Value does not match pattern' in str(e):
            work = xlrd.open_workbook(xlsxfile)
            sheet_names=work.sheet_names()
            work.release_resources()
            del work
        elif 'File is not a zip file' in str(e) or 'eaccf0860c6e23d9a1b3ad9140/c6ad59f6891fd0e3cf981e19f7af1802.xlsx' in xlsxfile:
            if outfiles:
                file_set_content(outfiles,'')
                return 
            else:
                raise Exception(e)
        elif 'File contains no valid workbook part' in str(e):
            if outfiles:
                file_set_content(outfiles,"error："+str(e))
                return
            else:
                raise Exception(e)
        else:
            print("openpyxl_1",e)
            raise Exception(e)
    else:
        sheet_names = work.sheetnames
        work.close()
    sheetarr=[]
    index=0
    for sheet in sheet_names:
        sheetarr.append({
            'sheet':sheet,
            'index':index
        })
        index+=1
    outfile=config.app['temp_folder']+'/'+md5(xlsxfile)
    workbook = spire_xls.Workbook()
    # workbook = xls.Workbook()
    try:
        workbook.LoadFromFile(xlsxfile)
    except Exception as e:
        print("LoadFromFile_e",e)
        if 'at Internal.Runtime.CompilerHelpers.ThrowHelpers.ThrowIndexOutOfRangeException' in str(e):
            if outfiles:
                file_set_content(outfiles,"error："+str(e))
                return
            else:
                raise Exception(e)
        elif 'Arg_NullReferenceException:   at sprfgv.sprq(String) + 0xa' in str(e):
            file_size = os.stat(xlsxfile).st_size
            if file_size<1024*10:
                if outfiles:
                    file_set_content(outfiles,'')
                    return
                else:
                    raise Exception(e)
            else:
                raise Exception(e)
        else:
            raise Exception(e)
    else:
        for item in sheetarr:
            # print('item',item)
            if 'hiddenSheet'==item['sheet'] or ('fd9f959560a0a6a51c32808907cac60f.xlsx' in xlsxfile or '52b0b52931a669832c93873c1740b1.xlsx' in xlsxfile) and item['index']==2:
                html=''
            else:
                # print('item',item,xlsxfile)
                sheet = workbook.Worksheets[item['index']]
                try:
                    sheet.SaveToHtml(outfile+"/"+str(item['index'])+".html")
                except Exception as e:
                    if os.path.exists(outfile):
                        shutil.rmtree(outfile)
                    print("SaveToHtml_e",e)
                    if 'Arg_NullReferenceException:   at sprd1q.spra(Stream, sprd33, String, HTMLOptions) + 0x3c0b' in str(e):
                        warnings.warn(str(e))
                        html=''
                    else:
                        if outfiles:
                            file_set_content(outfiles,"error："+str(e))
                            return
                        else:
                            raise Exception(e)
                else:
                    html=file_get_content(outfile+"/"+str(item['index'])+".html")
                # print('item_2',item,xlsxfile)
            if html and styles:
                htmls=''
                # print('item_3',item,xlsxfile)
                doc=kcwebspq(html.replace(' ',' ').replace('  ',' ').replace('\xa0',' ').replace('xmlns="http://www.w3.org/1999/xhtml"',''))
                doc('style').remove()
                doc('head').remove()
                doc('*').removeAttr('style')
                doc('*').removeAttr('class')
                # doc('td').attr('style','border: 1px solid #ccc')
                # doc('table').attr('style','border-collapse: collapse;')
                if doc("font").length:
                    for k in doc("font").items():
                        k.replaceWith(k.html())
                if doc("b").length:
                    for k in doc("b").items():
                        k.replaceWith(k.html())
                if doc("body").length:
                    for k in doc("body").items():
                        k.replaceWith(k.html())
                if doc("div").length:
                    for k in doc("div").items():
                        k.replaceWith(k.html())
                for k in doc('col').items():
                    if not k.text():
                        k.remove()
                # print('item_4',item,xlsxfile)
                table=doc('table')
                for k in doc('h2').items():
                    h2text=k.text()
                    if 'Evaluation' in h2text and 'Warning' in h2text and 'Spire.XLS' in h2text:
                        k.remove()
                for div in doc('div').items():
                    divs=div.text().replace(' ','').replace('\n','')
                    if not divs:
                        div.remove()
                # print('item_5',item,xlsxfile)
                for tables in table.items():
                    tr=tables.find("tr")
                    trindex=0
                    for trs in tr.items():
                        trtext=trs.text()
                        if 'Evaluation' in trtext and 'Warning' in trtext and 'Spire.PDF' in trtext:
                            trs.remove()
                        elif not trtext:
                            if trindex>=tr.length-1:
                                trs.remove()
                        trindex+=1
                # print('item_6',item,xlsxfile)
                for k in doc('table tr').items():
                    if '扫描全能王创建' == k.text().replace(' ','').replace('\n','').replace('&nbsp;','').replace(' ','').replace('  ','').replace('\xa0',''):
                        k.remove()
                # doc('#deletetr').remove()
                #删除空列 左到右
                table=doc('table')
                for items in table.items():
                    count=items.find('tr').eq(0).find('td').length
                    if not count:
                        count=0
                    count+=20
                    tempsl=0
                    for sl in range(count): #删除空列全部
                        tempsl+=1
                        if not items.find("tr td:nth-child("+str(tempsl)+")").text().replace(' ',''):
                            items.find("tr td:nth-child("+str(tempsl)+")").remove()
                            tempsl-=1
                        else:
                            break
                        
        
                htmls=doc.html()
                if not htmls:
                    htmls=''
            elif html:
                htmls=''
                doc=kcwebspq(html)
                # doc('style').remove()
                # doc('head').remove()
                for k in doc('p').items():
                    if '扫描全能王创建' == k.text().replace(' ','').replace('\n','').replace('&nbsp;','').replace(' ','').replace('  ','').replace('\xa0',''):
                        k.remove()
                if doc("body").length:
                    for k in doc("body").items():
                        k.replaceWith(k.html())
                htmls=doc.html()
                if not htmls:
                    htmls=''
            else:
                htmls=''
            item['html']=htmls
        workbook.Dispose()
        sheetarr1=[]
        if True:
            for k in sheetarr:
                if k['html']:
                    doc=kcwebspq("<div>"+k['html']+"</div>")
                    img=doc('img')
                    for kk in img.items():
                        src=kk.attr("src").replace('\\','/')
                        w,h=get_image_w_h(outfile+'/'+src)
                        if w==674 and h==98 or not img_is_con(outfile+'/'+src):
                            kk.remove()
                        elif w<imgminsize[0] or h<imgminsize[1]:
                            kk.remove()
                        elif not imgpath:
                            kk.attr("desc","不提取")
                            kk.attr("src","")
                        else:
                            kk.attr("name","local_image")
                            imgsrc=imgpath+src
                            tar=imgsrc.split('/')
                            directory=''
                            i=0
                            while i<len(tar)-1:
                                directory+=tar[i]+'/'
                                i+=1
                            if not os.path.exists(directory):
                                os.makedirs(directory)
                            shutil.move(outfile+'/'+src, imgsrc)
                            kk.attr("src",imgsrc)
                    k['html']=doc.html()
                    sheetarr1.append(k)
        else:
            for k in sheetarr:
                if k['html']:
                    doc=kcwebspq("<div>"+k['html']+"</div>")
                    doc('img').remove()
                    k['html']=doc.html()
                    sheetarr1.append(k)
        if os.path.exists(outfile):
            shutil.rmtree(outfile)
        if outfiles:
            file_set_content(outfiles,json_encode(sheetarr1))
        else:
            return sheetarr1
def rggestrsgrhklhtrdhbithjtiorjhiothposzfsgrgtsre(docfile,imgpath=False,imgminsize=(10,10),outfiles=''):
    """doc转html
    
    imgpath 是否保留图片 需要保存时传保存目录 以 / 结尾

    imgminsize 忽略 宽高 多少以下的图片

    return 返回格式 html 字符串
    """
    if imgpath:
        if imgpath[-1]=='/' or imgpath[-1]=='\\':
            imgpath=imgpath+md5(docfile)+'/'
        else:
            imgpath=imgpath+'/'+md5(docfile)+'/'
    outfile=config.app['temp_folder']+'/'+md5(docfile)
    from spire import doc as spire_doc
    try:
        document = spire_doc.Document()
        document.LoadFromFile(docfile)
        document.SaveToFile(outfile+'/temp.html', spire_doc.FileFormat.Html)
        document.Close()
    except Exception as e:
        if outfiles:
            file_set_content(outfiles,"error："+str(e))
            return
        else:
            raise Exception(e)
    html=file_get_content(outfile+'/temp.html')
    doc=kcwebspq(html.replace(' ',' ').replace('  ',' ').replace('\xa0',' ').replace('xmlns="http://www.w3.org/1999/xhtml"',''))
    doc('title').remove()
    doc('style').remove()
    doc('head').remove()
    doc('*').removeAttr('style')
    doc('*').removeAttr('class')
    if doc("span").length:
        for k in doc("span").items():
            k.replaceWith(k.html())
    if doc("td p").length:
        for k in doc("td p").items():
            k.replaceWith(k.html())
    if doc("body").length:
        for k in doc("body").items():
            k.replaceWith(k.html())
    
    html=doc.html()
    if not html:
        html=''
    html=html.replace('<p>Evaluation Warning: The document was created with Spire.Doc for Python.</p>','').replace('<div/>','')
    html=html.replace('<p><ins data-userid="0" data-username="徐建清" data-time="0001-01-01T00:00:00Z">Evaluation Warning: The document was created with Spire.Doc for Python.</ins></p>','')
    html=html.replace('Evaluation Warning: The document was created with Spire.Doc for Python.','')
    if True:
        if html:
            doc=kcwebspq("<div>"+html+"</div>")
            img=doc('img')
            for kk in img.items():
                src=kk.attr("src").replace('\\','/')
                w,h=get_image_w_h(outfile+'/'+src)
                if w==674 and h==98 or not img_is_con(outfile+'/'+src):
                    kk.remove()
                elif w<imgminsize[0] or h<imgminsize[1]:
                    kk.remove()
                elif not imgpath:
                    kk.attr("desc","不提取")
                    kk.attr("src","")
                else:
                    kk.attr("name","local_image")
                    imgsrc=imgpath+src
                    tar=imgsrc.split('/')
                    directory=''
                    i=0
                    while i<len(tar)-1:
                        directory+=tar[i]+'/'
                        i+=1
                    if not os.path.exists(directory):
                        os.makedirs(directory)
                    shutil.move(outfile+'/'+src, imgsrc)
                    kk.attr("src",imgsrc)
            html=doc.html()
    else:
        if html:
            doc=kcwebspq("<div>"+html+"</div>")
            doc('img').remove()
            for k in doc('p').items():
                t=k.text()
                if t:
                    t=t.replace(' ','').replace('  ','').replace('\xa0','').replace(' ','')
                if not t:
                    k.remove()
            if doc.text():
                html=doc.html()
            else:
                html=''
    shutil.rmtree(outfile)
    doc=kcwebspq(html)
    #删除空行
    doc('table').attr("border","1")
    table=doc('table')
    for tables in table.items():
        for trs in tables.find("tr").items():
            trstext=trs.text()
            if trstext:
                trstext=trstext.replace('\n','').replace(' ','').replace(' ','').replace('  ','').replace('\xa0','')
            tempimgsrcstr=trs.html()
            if tempimgsrcstr:
                tempimgsrcstr=tempimgsrcstr.replace(' ','')
            else:
                tempimgsrcstr=''
            if not trstext and '<imgsrc=' not in tempimgsrcstr:
                trs.remove()
    html=doc("*").html()
    if outfiles:
        file_set_content(outfiles,html)
    else:
        return html
def fesgrsgtrgtdrhbtdgrrgrgsgtsegr(pdfname,outname,stop_event=None):
    """pdf转docx
    
    pdfname pdf文件

    outname 转换后保存文件
    """
    folder_path=os.path.dirname(outname)
    if folder_path and not os.path.exists(folder_path):
        os.makedirs(folder_path, exist_ok=True)
    cv = pdf2docx.Converter(pdfname)
    cv.convert(outname, start=0, end=None)
    cv.close()
    return True
def tran_pdf_to_docx(pdfname,outname,timeout=None):
    """pdf转docx
    
    pdfname pdf文件

    outname 转换后保存文件

    timeout 转换超时 单位秒
    """
    if not timeout:
        return fesgrsgtrgtdrhbtdgrrgrgsgtsegr(pdfname,outname)
    else:
        thread = threading.Thread(target=fesgrsgtrgtdrhbtdgrrgrgsgtsegr,args=(pdfname,outname),daemon=True)
        thread.start()
        thread.join(timeout=timeout)
        if thread.is_alive():
            raise Exception('tran_pdf_to_docx timeout out')
        else:
            return True
def tran_pdf_to_xlsx_process(pdfname,outname,process=True):
    """pdf 转 xlsx
    
    pdfname pdf文件

    outname 转换后保存地址
    """
    if config.app['app_debug']:
        print('tran_pdf_to_xlsx_process',pdfname)
    if os.path.exists(outname):
        raise Exception('outname：'+outname+' 已存在')
    if process:
        p=multiprocessing.Process(target=frgesregergslkjhgfdertyu876544r56tg7uhik,args=(pdfname,outname),daemon=True)
        p.start()
        p.join()
        if os.path.exists(outname):
            sheetarr=file_get_content(outname+'err')
            if 'error：'==sheetarr[0:6]:
                os.remove(outname+'err')
                raise Exception(sheetarr)
            return True
        elif os.path.exists(outname+'err'):
            sheetarr=file_get_content(outname+'err')
            if 'error：'==sheetarr[0:6]:
                os.remove(outname+'err')
                raise Exception(sheetarr)
        else:
            return False
    else:
        frgesregergslkjhgfdertyu876544r56tg7uhik(pdfname,outname)
        sheetarr=file_get_content(outname+'err')
        if 'error：'==sheetarr[0:6]:
            os.remove(outname+'err')
            raise Exception(sheetarr)
        return True
# def tran_pdf_to_docx_process(pdfname,outname,process=True):
#     """pdf 转 docx
    
#     pdfname pdf文件

#     outname 转换后保存地址
#     """
#     if config.app['app_debug']:
#         print('tran_pdf_to_docx_process',pdfname)
#     if os.path.exists(outname):
#         raise Exception('outname：'+outname+' 已存在')
#     if process:
#         p=multiprocessing.Process(target=frgesregeeswfrgslkjhgfdertyu8765sdf44r56tg7uhik,args=(pdfname,outname),daemon=True)
#         p.start()
#         p.join()
#         if os.path.exists(outname):
#             return True
#         else:
#             return False
#     else:
#         frgesregeeswfrgslkjhgfdertyu8765sdf44r56tg7uhik(pdfname,outname)
#         return True

def tran_xlsx_to_html_process(xlsxfile,styles=True,imgpath=False,imgminsize=(10,10)):
    """xlsx 转 html
    
    xlsxfile pdf文件

    styles 是否处理样式

    imgpath 是否保留图片 需要保存时传保存目录 以 / 结尾

    imgminsize 忽略 宽高 多少以下的图片

    return 返回格式 [{'sheet':'','index':0,'','html':''}]
    """
    outfiles=config.app['temp_folder']+'/'+md5(xlsxfile)+'.html'
    if config.app['app_debug']:
        print('tran_xlsx_to_html_process',xlsxfile)

 
    p=multiprocessing.Process(target=lfjdsgjtrdjdtdrgkgtrssreiuiu43943uffh938raaqef,args=(xlsxfile,styles,imgpath,imgminsize,outfiles),daemon=True)
    p.start()
    p.join()
    if os.path.exists(outfiles):
        print('tran_xlsx_to_html_process_outfiles',outfiles)
        sheetarr=file_get_content(outfiles)
        os.remove(outfiles)
        if 'error：'==sheetarr[0:6]:
            raise Exception(sheetarr)
        sheetarr=json_decode(sheetarr)
        return sheetarr
    else:
        raise Exception('tran_xlsx_to_html_process失败')
def tran_doc_to_html_process(docfile,imgpath=False,imgminsize=(10,10)):
    """doc 转 html  （注意：linux 系统中可能会报Cannot found font installed on the system错误，可以把windows系统的字体安装到linux中，安装命令参考 “mkfontdir 字体目录”）
    
    docfile doc文件

    imgpath 是否保留图片 需要保存时传保存目录 以 / 结尾

    imgminsize 忽略 宽高 多少以下的图片

    return 返回格式 html 字符串
    """
    if config.app['app_debug']:
        print('tran_doc_to_html_process',docfile)
    outfiles=config.app['temp_folder']+'/'+md5(docfile)+'.html'
    p=multiprocessing.Process(target=rggestrsgrhklhtrdhbithjtiorjhiothposzfsgrgtsre,args=(docfile,imgpath,imgminsize,outfiles),daemon=True)
    p.start()
    p.join()
    if os.path.exists(outfiles):
        html=file_get_content(outfiles)
        os.remove(outfiles)
        if 'error：'==html[0:6]:
            raise Exception(html)
    else:
        raise Exception('tran_doc_to_html_process失败')
    return html

def esgsvesgrhghtsezgesgrezgseszgrgesresgrges(input_file,out_path,page_count=1):
    # 创建PdfDocument对象
    pdf = PdfDocument()
    # 加载PDF文件
    try:
        pdf.LoadFromFile(input_file)
    except Exception as e:
        raise Exception(e)
    else:
        total_pages = pdf.Pages.Count # 计算总页数
    if total_pages<page_count:
        if not os.path.isdir(out_path):
            os.makedirs(out_path)
        shutil.copy(input_file, out_path)
    else:
        # 按指定页数拆分PDF
        for i in range(0, total_pages, page_count):
            # 创建新的PdfDocument对象
            new_pdf = PdfDocument()
            # 计算当前要插入的页码范围
            start_page = i
            end_page = min(i + page_count - 1, total_pages - 1)  # 确保不超过总页数
            try:
                # 将当前页码范围的页面插入到新PDF中
                new_pdf.InsertPageRange(pdf, start_page, end_page)
            except:pass
            else:
                # 保存生成的文件
                new_pdf.SaveToFile(out_path+"/" + f"{start_page + 1}-{end_page + 1}页.pdf")
                # 关闭新创建的PdfDocument对象
                new_pdf.Close()
        pdf.Close()
def split_pdf_by_page_count(input_file,out_path,page_count=1):
    """ pdf文件拆分
    
    input_file 文件名

    out_path 输出目录

    page_count 按多页拆分
    """
    if config.app['app_debug']:
        print('split_pdf_by_page_count')
    if os.name == 'nt':
        process=False
    else:
        process=True
    if process:
        p=multiprocessing.Process(target=esgsvesgrhghtsezgesgrezgseszgrgesresgrges,args=(input_file,out_path,page_count),daemon=True)
        p.start()
        p.join()
        return True
    else:
        return esgsvesgrhghtsezgesgrezgseszgrgesresgrges(input_file,out_path,page_count)


def timestampToDate(times,format="%Y-%m-%d %H:%M:%S"):
    """时间戳转换时间

    times 10位时间戳

    format 日期格式 如%Y-%m-%d %H:%M:%S
    """
    timeArray = time.localtime(int(times))
    return time.strftime(format.encode('unicode-escape').decode(),timeArray).encode().decode('unicode-escape')
def send_mail(user,text="邮件内容",theme="邮件主题",recNick="收件人昵称"):
    """发送邮件

    参数 user：接收邮件的邮箱地址

    参数 text：邮件内容

    参数 theme：邮件主题

    参数 recNick：收件人昵称

    return Boolean类型
    """
    ret=True
    if not theme:
        theme=config.email['theme']
    if not recNick:
        recNick=config.email['recNick']
    try:
        msg=MIMEText(text,'plain','utf-8')
        msg['From']=formataddr([config.email['sendNick'],config.email['sender']]) 
        msg['To']=formataddr([recNick,user]) 
        msg['Subject']=theme

        server=smtplib.SMTP_SSL("smtp.qq.com", 465) 
        server.login(config.email['sender'], config.email['pwd']) 
        server.sendmail(config.email['sender'],[user,],msg.as_string())
        server.quit()
    except Exception:
        ret=False
    return ret
def Template(path,**context):
    "模板渲染引擎函数,使用配置的模板路径"
    return Templates(path,**context)
def Templates(path,**context):
    "模板渲染引擎函数，需要完整的模板目录文件"
    lookup = TemplateLookup(directories=[''])
    # body=''
    # with open(path, 'r',encoding='utf-8') as f:
    #     contents=f.read()
    #     t=kcwsTemplate(contents,lookup=lookup,module_directory=config.cache['path']+"/Template")
    #     body=t.render(**context)
    if (path[:1]=="/" or path[1:2]==":") and ('.html' in path or '.shtml' in path or '.htm' in path):
        t=kcwsTemplate(filename=path,module_directory=config.cache['path']+"/Template",lookup=lookup)
    else:
        t=lookup.get_template(path)
    body=t.render(**context)
    return body
def kcwsTemp(contents,**context):
    "模板渲染引擎函数，传字符串进来"
    lookup = TemplateLookup(directories=[''])
    t=kcwsTemplate(contents,lookup=lookup,module_directory=config.cache['path']+"/Template")
    body=t.render(**context)
    return body
def getfunction(strs,reload=False):
    """获取指定文件对象
    
    strs :app.index.common.autoload  获取app/index/common/目录下的autoload对象

    reload 是否重新加载已导入的模块（是否每次加载修改后的模块）
    """
    obj=importlib.import_module(strs)
    if reload:
        importlib.reload(obj)
    return obj
def M(table=None,confi=None):
    """数据库操作实例
    
    参数 table：表名

    参数 confi 数据库配置  可以传数据库名字符串
    """
    if confi:
        if confi['type']=='sqlite':
            return sqlite(table,confi)
        else:
            return mysql(table,confi)
    else:
        if config.database['type']=='sqlite':
            return sqlite(table)
        else:
            return mysql(table)
mysqldbobj=kcwsmysql.mysql()
def mysql(table=None,configss=None):
    """mysql数据库操作实例
    
    参数 table：表名

    参数 configss 数据库配置  可以传数据库名字符串
    """
    global mysqldbobj
    if not mysqldbobj:
        mysqldbobj=kcwsmysql.mysql()
    if table is None:
        return mysqldbobj
    elif configss:
        return mysqldbobj.connect(configss).table(table)
    else:
        return mysqldbobj.connect(config.database).table(table)
def sqlite(table=None,configss=None):
    """sqlite数据库操作实例
    
    参数 table：表名

    参数 configss 数据库配置  可以传数据库名字符串
    """
    dbs=kcwssqlite.sqlite()
    if table is None:
        return dbs
    elif configss:
        return dbs.connect(configss).table(table)
    else:
        return dbs.connect(config.sqlite).table(table)
def mongo(table=None,configss=None):
    """mongodb数据库操作实例
    
    参数 table：表名(mongodb数据库集合名)

    参数 configss mongodb数据库配置  可以传数据库名字符串
    """
    mObj=kcwsmongodb.mongo()
    if table is None:
        return mObj
    elif configss:
        return mObj.connect(configss).table(table)
    else:
        return mObj.connect(config.mongo).table(table)
def set_cache(name,values,expire="no"):
    """设置缓存

    参数 name：缓存名

    参数 values：缓存值

    参数 expire：缓存有效期 0表示永久  单位 秒
    
    return Boolean类型
    """
    return kcwscache.cache().set_cache(name,values,expire)
def get_cache(name):
    """获取缓存

    参数 name：缓存名

    return 或者的值
    """
    return kcwscache.cache().get_cache(name)
def del_cache(name):
    """删除缓存

    参数 name：缓存名

    return Boolean类型
    """
    return kcwscache.cache().del_cache(name)
def dateoperator(date,years=0,formats='%Y%m%d%H%M%S',months=0, days=0, hours=0, minutes=0,seconds=0,
                 leapdays=0, weeks=0, microseconds=0,
                 year=None, month=None, day=None, weekday=None,
                 yearday=None, nlyearday=None,
                 hour=None, minute=None, second=None, microsecond=None):
    """日期相加减计算
    date 2019-10-10
    formats 设置需要返回的时间格式 默认%Y%m%d%H%M%S
    
    years 大于0表示加年  反之减年
    months 大于0表示加月  反之减月
    days 大于0表示加日  反之减日

    return %Y%m%d%H%M%S
    """
    formatss='%Y%m%d%H%M%S'
    date=re.sub('[-年/月:：日 时分秒]','',date)
    if len(date) < 8:
        return None
    if len(date) < 14:
        s=14-len(date)
        i=0
        while i < s:
            date=date+"0"
            i=i+1
    d = datetime.datetime.strptime(date, formatss)
    strs=(d + relativedelta(years=years,months=months, days=days, hours=hours, minutes=minutes,seconds=seconds,
                 leapdays=leapdays, weeks=weeks, microseconds=microseconds,
                 year=year, month=month, day=day, weekday=weekday,
                 yearday=yearday, nlyearday=nlyearday,
                 hour=hour, minute=minute, second=second, microsecond=microsecond))
    strs=strs.strftime(formats)
    return strs
def get_kcwebs_folder():
    '获取当前框架目录'
    return os.path.split(os.path.realpath(__file__))[0][:-7] #当前框架目录
if not os.path.exists(get_kcwebs_folder()+"/pid/"):
    os.makedirs(get_kcwebs_folder()+"/pid/", exist_ok=True)
def kill_all_kcwebs_pid(types='all'):
    """结束kcwebps框架的所有子进程
    
    types 可选 all 和 kcwebs
    """
    if types=='all':
        kill_all_pid()
    folder=get_kcwebs_folder()+"/pid"
    lis=os.listdir(folder)
    for files in lis:
        if os.path.isfile(folder+"/"+files):
            f=open(folder+"/"+files,'r')
            pid=f.read()
            f.close()
            kill_pid(pid)
            os.remove(folder+"/"+files)
def is_index(params,index):
    """判断列表或字典里的索引是否存在

    params  列表或字典

    index   索引值

    return Boolean类型
    """
    try:
        params[index]
    except KeyError:
        return False
    except IndexError:
        return False
    else:
        return True

def list_to_tree(data, pk = 'id', pid = 'pid', child = 'lowerlist', root=0,childstatus=True):
    """列表转换tree
    
    data 要转换的列表

    pk 关联节点字段

    pid 父节点字段

    lowerlist 子节点列表

    root 主节点值

    childstatus 当子节点列表为空时是否需要显示子节点字段
    """
    arr = []
    for v in data:
        if v[pid] == root:
            kkkk=list_to_tree(data,pk,pid,child,v[pk],childstatus)
            if childstatus:
                v[child]=kkkk
            else:
                if kkkk:
                    v[child]=kkkk
            arr.append(v)
    return arr
def randoms(lens=6,types=1):
    """生成随机字符串
    
    lens 长度

    types 1数字 2字母 3字母加数字
    """
    strs="0123456789qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM,!@#$%^&*()_+=-;',./:<>?"
    if types==1:
        strs="0123456789"
    elif types==2:
        strs="qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM"
    elif types==3:
        strs="0123456789qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM"
    k=''
    i=0
    while i < lens:
        k+=random.choice(strs)
        i+=1
    return k
def file_set_content(filename,data,encoding="utf-8"):
    """写入文件内容
    
    filename 完整文件名

    data 要写入的内容

    encoding 保存编码
    """
    f=open(filename,'w',encoding=encoding)
    f.write(data)
    f.close()
    return True
def file_get_content(filename,cur_encoding='utf-8',encoding=False):
    """获取文件内容
    
    filename 完整文件名

    cur_encoding 指定编码获取文件内容

    encoding 是否返回文件编码  默认否
    """
    fileData=''
    if os.path.isfile(filename):
        if encoding:
            with open(filename, 'rb') as f:
                cur_encoding = chardet.detect(f.read())['encoding']
                # print('cur_encoding',cur_encoding)
        #用获取的编码读取该文件而不是python3默认的utf-8读取。
        with open(filename,encoding=cur_encoding) as file:
            fileData = file.read()
    if encoding:
        return fileData,cur_encoding
    else:
        return fileData
class kcwebssign:
    def getsign(params):
        "获取签名"
        if is_index(params,'sign'):
            del params['sign']
        content=kcwebssign.getSignContent(params)
        return md5(content)
    def exsignpra(params):
        "生成签名参数"
        params['time']=times()
        params['rands']=randoms()
        params['sign']=kcwebssign.getsign(params)
        return params
    def getSignContent(params):
        "字典排序"
        param={}
        for i in sorted (params) : 
            param[i]=params[i]
        i=0
        strs=""
        for k in param:
            if k:
                if isinstance(k,dict):
                    k=json_encode(k)
                    k=k.replace('"', '')
                    k=k.replace("'", '')
                if param[k]:
                    if i==0:
                        strs+=str(k)+"="+str(param[k])
                    else:
                        strs+="&"+str(k)+"="+str(param[k])
            i+=1
        return strs
kcwebszip=kcwszip
kcwebstar=kcwstar
class response:
    tpldata={} #公共模板变量
    def tpl(path=None,status='200 ok',response_cache=False,ETag=None,header={"Content-Type":"text/html; charset=utf-8"},**context):
        """返回模板内容
        
        path 文件地址

        status 响应状态码

        response_cache 是否启用浏览器缓存  响应状态码200 ok时有效

        ETag 缓存标识  响应状态码200 ok时有效

        header 响应头
        """
        for k in dir(response):
            if k not in ['download','json','pic','redirect','tpl','tpldata','video'] and k[-2:]!='__':
                try:
                    context[k]=response.__dict__[k]
                except KeyError:
                    pass
        context['config']=config
        headers=copy.deepcopy(header)
        getroutecomponent=globals.VAR.component
        if path:
            if (path[:1]=="/" or path[1:2]==":") and ('.' in path[-8:]):
                Temppath=path
            elif (path[:1]=="/"):
                Temppath=config.app['tpl_folder']+path+".html"
            else:
                Temppath=config.app['tpl_folder']+"/"+getroutecomponent[1]+"/controller/"+getroutecomponent[2]+"/tpl/"+path+".html"
        else:
            Temppath=config.app['tpl_folder']+"/"+getroutecomponent[1]+"/controller/"+getroutecomponent[2]+"/tpl/"+getroutecomponent[3]+"/"+getroutecomponent[4]+".html"
        # print('Temppath',Temppath)
        if status=='200 ok' and response_cache:
            if not ETag:
                ttt=''
                for k in context.keys():
                    ttt+=k+str(context[k])
                ETag=md5(Temppath+ttt+globals.HEADER.URL)
            try:
                HTTP_IF_NONE_MATCH=globals.HEADER.GET['HTTP_IF_NONE_MATCH']
            except:
                HTTP_IF_NONE_MATCH=None
            if HTTP_IF_NONE_MATCH and HTTP_IF_NONE_MATCH==ETag:
                status="304 Not Modified"
                body=''
            else:
                # if isinstance(response_cache,int) and response_cache>1:
                #     headers['response_cache']=str(response_cache)+" s"
                #     set_cache(ETag,1,response_cache)
                # else:
                #     headers['response_cache']="default"
                #     set_cache(ETag,1)
                body=Template(Temppath,**context)
            dateArray = datetime.datetime.utcfromtimestamp(times()-86400)
            otherStyleTime = dateArray.strftime('%a, %d %b %Y %H:%M:%S GMT')
            headers['Last-Modified']=otherStyleTime
            headers['ETag']=ETag
            return body,status,headers
        elif status:
            return Template(Temppath,tpldata=response.tpldata,**context),status,headers
        else:
            return Template(Temppath,tpldata=response.tpldata,**context),'200 ok',headers
    def json(res=[],status='200 ok',response_cache=False,ETag=None,header={"Content-Type":"application/json; charset=utf-8","Access-Control-Allow-Origin":"*"}):
        """响应json内容

        res  body内容

        status 响应状态码

        response_cache 是否启用浏览器缓存  响应状态码200 ok时有效

        ETag 缓存标识  响应状态码200 ok时有效

        header 响应头
        """
        headers=copy.deepcopy(header)
        if status=='200 ok' and response_cache:
            if not ETag:
                ETag=md5(globals.HEADER.URL)
            try:
                HTTP_IF_NONE_MATCH=globals.HEADER.GET['HTTP_IF_NONE_MATCH']
            except:
                HTTP_IF_NONE_MATCH=None
            if(HTTP_IF_NONE_MATCH and get_cache(ETag)):
                status="304 Not Modified"
                body=''
            else:
                if isinstance(response_cache,int) and response_cache>1:
                    set_cache(ETag,1,response_cache)
                    headers['response_cache']=str(response_cache)+" s"
                else:
                    set_cache(ETag,1)
                    headers['response_cache']="default"
                body=json_encode(res)
            dateArray = datetime.datetime.utcfromtimestamp(times()-86400)
            otherStyleTime = dateArray.strftime('%a, %d %b %Y %H:%M:%S GMT')
            headers['Last-Modified']=otherStyleTime
            headers['ETag']=ETag
            
        else:
            body=json_encode(res)
        return body,status,headers
    def pic(body,response_cache=True,ETag=None):
        """输出图片
        
        body 图片二进制内容或图片路径 建议使用图片路径

        response_cache 是否启用浏览器缓存  body使用图片路径时有效

        ETag 缓存标识
        """
        status='200 ok'
        header={"Cache-Control":"public, max-age=2592000"}
        if isinstance(body,str):
            if response_cache:
                if not ETag:
                    ETag=md5(body+globals.HEADER.URL)
                try:
                    HTTP_IF_NONE_MATCH=globals.HEADER.GET['HTTP_IF_NONE_MATCH']
                except:
                    HTTP_IF_NONE_MATCH=None
                if(HTTP_IF_NONE_MATCH and get_cache(ETag)):
                    status="304 Not Modified"
                    body=''
                else:
                    if isinstance(response_cache,int) and response_cache>1:
                        set_cache(ETag,1,response_cache)
                    else:
                        set_cache(ETag,1)
                    filename=body
                    f=open(filename,"rb")
                    body=f.read()
                    f.close()
                    kind = filetype.guess(filename)
                    try:
                        header['Content-Type']=kind.mime
                    except:
                        header['Content-Type']="image/png"
            else:
                filename=body
                f=open(filename,"rb")
                body=f.read()
                f.close()
                kind = filetype.guess(filename)
                try:
                    header['Content-Type']=kind.mime
                except:
                    header['Content-Type']="image/png"
            dateArray = datetime.datetime.utcfromtimestamp(times()-86400)
            otherStyleTime = dateArray.strftime('%a, %d %b %Y %H:%M:%S GMT')
            header['Last-Modified']=otherStyleTime
            header['ETag']=ETag
        else:
            header['Content-Type']="image/png"
        return body,status,header
    def video(body):
        """输出视频
        
        body 视频二进制内容或视频路径
        """
        status='200 ok'
        header={"Cache-Control":"public, max-age=2592000"}
        if isinstance(body,str):
            ETag=md5(body)
            try:
                HTTP_IF_NONE_MATCH=globals.HEADER.GET['HTTP_IF_NONE_MATCH']
            except:
                HTTP_IF_NONE_MATCH=None
            if(HTTP_IF_NONE_MATCH and get_cache(ETag)):
                header=get_cache(ETag)
                status="304 Not Modified"
                body=''
            else:
                filename=body
                f=open(filename,"rb")
                body=f.read()
                f.close()
                kind = filetype.guess(filename)
                try:
                    header['Content-Type']=kind.mime
                except:
                    header['Content-Type']="video/mp4"
                header['content-length']=str(len(body))
                set_cache(ETag,header,2592000)
            dateArray = datetime.datetime.utcfromtimestamp(times()-86400)
            otherStyleTime = dateArray.strftime('%a, %d %b %Y %H:%M:%S GMT')
            header['Last-Modified']=otherStyleTime
            header['ETag']=ETag
        else:
            header['Content-Type']="video/mp4"
        return body,status,header
    def audio(body):
        """输出音频
        
        body 音频二进制内容或音频路径
        """
        status='200 ok'
        header={"Cache-Control":"public, max-age=2592000"}
        if isinstance(body,str):
            ETag=md5(body)
            try:
                HTTP_IF_NONE_MATCH=globals.HEADER.GET['HTTP_IF_NONE_MATCH']
            except:
                HTTP_IF_NONE_MATCH=None
            if(HTTP_IF_NONE_MATCH and get_cache(ETag)):
                header=get_cache(ETag)
                status="304 Not Modified"
                body=''
            else:
                filename=body
                f=open(filename,"rb")
                body=f.read()
                f.close()
                kind = filetype.guess(filename)
                try:
                    header['Content-Type']=kind.mime
                except:
                    header['Content-Type']="audio/mpeg"
                header['content-length']=str(len(body))
                set_cache(ETag,header,2592000)
            dateArray = datetime.datetime.utcfromtimestamp(times()-86400)
            otherStyleTime = dateArray.strftime('%a, %d %b %Y %H:%M:%S GMT')
            header['Last-Modified']=otherStyleTime
            header['ETag']=ETag
        else:
            header['Content-Type']="audio/mpeg"
        return body,status,header
    def download(pathname):
        """下载文件
        
        pathname 文件路径
        """
        if os.path.isfile(pathname):
            f=open(pathname,"rb")
            body=f.read()
            f.close()
            kind = filetype.guess(pathname)
            try:
                return body,"200 ok",{"Content-Type":"application/"+kind.mime,"Accept-Ranges":"bytes"}
            except:
                return body,"200 ok",{"Content-Type":"application/text","Accept-Ranges":"bytes"}
        else:
            return Templates('E:\doc\python\kcwebps\kcwebs/tpl/err.html',title="文件不存在",content="文件不存在",imgsrc=config.domain['kcwebsimg']+"/icon/error.png",config=config)
    def redirect(url,status="302 Found",html='',header={"Content-Type":"application/html; charset=utf-8"}):
        """重定向

        参数 url 重定向地址 必须

        参数 status 响应码  可选

        参数 html body响应内容 可选

        参数 header 响应头  可选
        """
        header['Location']=url
        return html,status,header
    
if 'Linux' in get_sysinfo()['platform']:
    #添加自启命令
    if not os.path.isfile('/usr/bin/startkcwebs.sh'):
        open('/usr/bin/startkcwebs.sh', 'w').close()
        t=file_get_content("/etc/os-release")
        if 'centos' in t.lower():
            os.system("sed -i 's/bash startkcwebs.sh//g' /etc/rc.d/rc.local")
            os.system("echo 'bash startkcwebs.sh'  >> /etc/rc.d/rc.local")
            os.system('chmod 777 /etc/rc.d/rc.local')
            os.system('chmod 777 /usr/bin/startkcwebs.sh')
        elif 'ubuntu' in t.lower():
            if not os.path.isfile("/etc/rc.local"):
                file_set_content('/etc/rc.local','#!/bin/sh')
            os.system("sed -i 's/bash startkcwebs.sh//g' /etc/rc.local")
            os.system("echo 'bash startkcwebs.sh'  >> /etc/rc.local")
            os.system('sudo chmod +x /etc/rc.local')
            os.system('sudo chmod 777 /usr/bin/startkcwebs.sh')
def insert_system_up(cmd):
    """添加开机启动命令
    
    cmd 命令
    """
    if 'Linux' in get_sysinfo()['platform']:
        f=open("/usr/bin/startkcwebs.sh","a")
        f.write("\n"+cmd+"\n")
        f.close()
        return True
    else:
        raise Exception('暂不支持linux以外的系统')
def del_system_up(cmd,vague=False):
    """删除开机启动命令
    
    cmd 命令

    vague 是否模糊匹配 
    """
    if 'Linux' in get_sysinfo()['platform']:
        if vague:
            f = open("/usr/bin/startkcwebs.sh")
            con=''
            while True:
                line = f.readline()
                if not line:
                    break
                if cmd in line:
                    line=''
                con=con+line
            f.close()
            file_set_content("/usr/bin/startkcwebs.sh",con)
        else:
            content=file_get_content("/usr/bin/startkcwebs.sh")
            content=content.replace("\n"+cmd+"\n","")
            file_set_content("/usr/bin/startkcwebs.sh",content)
        return True
    else:
        raise Exception('暂不支持linux以外的系统')
def rfdsgeagesegdsfdsfdrrsebskcwebsafterrequest():
    #关闭数据库连接 不要修改改方法，也不要定义与该方法相同名字的方法
    try:
        mysqldbobj.close()
    except:pass
    try:
        mysqldbobj.close()
    except:pass