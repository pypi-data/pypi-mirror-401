"""Test hybrid engine + import_sheet functionality.

This test verifies that workbooks with import_sheet() can be saved using
engine="hybrid" which combines xlsxwriter for SheetNodes with openpyxl
for ImportedSheetNodes with post-merge.

Also tests that engine="xlsxwriter" correctly rejects import_sheet usage.
"""
# cspell:disable

from __future__ import annotations

import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest
import xlsxwriter
from openpyxl.styles import Alignment, Font, PatternFill

import xpyxl as x


def _create_template(path: Path) -> None:
    """Create a template workbook with specific properties for testing."""
    wb = xlsxwriter.Workbook(path)
    ws = wb.add_worksheet("Template")

    # Add content with a merge (styled)
    title_fmt = wb.add_format(
        {
            "bold": True,
            "font_color": "#123456",
            "bg_color": "#FFE699",
            "pattern": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )
    ws.merge_range("A1:C1", "Template Title", title_fmt)  # type: ignore[arg-type]

    # Add more content
    ws.write("A3", "Notes")  # type: ignore[arg-type]
    ws.write("A4", "Static content from template")  # type: ignore[arg-type]

    # Column-level fill format (Excel commonly stores this as column formatting,
    # not as individual cell styles for blank cells).
    col_fmt = wb.add_format({"bg_color": "#BFE3EF", "pattern": 1})
    # Fill a wide area via column formatting (should apply to all rows).
    ws.set_column("A:AJ", 8, col_fmt)  # type: ignore[arg-type]
    ws.set_column("A:A", 24, col_fmt)  # type: ignore[arg-type]
    ws.set_column("B:B", 16, col_fmt)  # type: ignore[arg-type]
    ws.set_column("F:F", None, col_fmt)  # type: ignore[arg-type]
    ws.write("F13", "hello")  # type: ignore[arg-type]
    ws.set_row(0, 28)

    # Row-level fill format (vertical band). This is how many templates paint
    # large empty regions without explicit cell styles.
    row_fmt = wb.add_format({"bg_color": "#BFE3EF", "pattern": 1})
    for r in range(1, 13):  # Excel rows 2..13
        ws.set_row(r, None, row_fmt)

    # Set freeze panes and auto filter
    ws.freeze_panes(2, 0)  # A3
    ws.autofilter(2, 0, 9, 2)  # A3:C10

    wb.close()


def _create_style_heavy_template(path: Path) -> None:
    """Create a template with many distinct alignments/styles.

    This catches cross-workbook style index copying bugs in openpyxl save
    (e.g. alignmentId out of range) when hybrid-saving with xlsxwriter.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Expected an active worksheet")
    ws.title = "Template"

    for i in range(1, 75):
        horiz = ["general", "left", "center", "right", "fill", "justify"][i % 6]
        vert = ["top", "center", "bottom", "justify"][i % 4]
        wrap = i % 2 == 0
        indent = i % 6
        rotation = (i * 15) % 90

        a = ws.cell(row=i, column=1, value=f"Row {i}")
        a.font = Font(bold=(i % 3 == 0), italic=(i % 5 == 0))
        rgb = f"FF{(i * 3) % 255:02X}{(i * 7) % 255:02X}{(i * 11) % 255:02X}"
        a.fill = PatternFill(fill_type="solid", start_color=rgb, end_color=rgb)
        a.alignment = Alignment(
            horizontal=horiz,
            vertical=vert,
            wrap_text=wrap,
            indent=indent,
            textRotation=rotation,
        )

        b = ws.cell(row=i, column=2, value=i)
        b.number_format = "#,##0"
        b.alignment = Alignment(horizontal="right", vertical=vert)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.row_dimensions[1].height = 20

    wb.save(path)


def test_hybrid_with_import_sheet() -> None:
    """Test that hybrid engine works with import_sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template
        _create_template(template_path)

        # Build workbook with imported sheet + generated sheets
        workbook = x.workbook()[
            x.sheet("Before")[
                x.row(style=[x.bold, x.bg_warning, x.text_white])[None],
                x.row()["A", 1],
            ],
            x.import_sheet(template_path, "Template", name="Imported"),
            x.sheet("After")[
                x.row(style=[x.bold])["Generated After"],
                x.row()["B", 2],
            ],
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output
        result_wb = openpyxl.load_workbook(output_path)

        # Ensure the produced file is round-trip writable by openpyxl (catches corruption)
        from io import BytesIO

        # Ensure the produced file is a valid zip container
        with zipfile.ZipFile(output_path) as zf:
            assert zf.testzip() is None

        roundtrip = BytesIO()
        result_wb.save(roundtrip)
        assert roundtrip.getvalue()

        # Check sheet order matches declaration order
        expected_order = ["Before", "Imported", "After"]
        actual_order = result_wb.sheetnames
        assert actual_order == expected_order, (
            f"Sheet order mismatch: expected {expected_order}, got {actual_order}"
        )

        # Check generated sheets have content
        before_ws = result_wb["Before"]
        assert before_ws["A1"].value is None
        assert (
            before_ws["A1"].fill is not None
            and before_ws["A1"].fill.fgColor.rgb == "FFB45309"
        )
        assert before_ws["A2"].value == "A"
        assert before_ws["B2"].value == 1

        after_ws = result_wb["After"]
        assert after_ws["A1"].value == "Generated After"
        assert after_ws["A2"].value == "B"
        assert after_ws["B2"].value == 2

        # Check imported sheet preserves properties
        imported_ws = result_wb["Imported"]
        assert imported_ws["A1"].value == "Template Title"
        assert imported_ws["A3"].value == "Notes"
        assert imported_ws["A4"].value == "Static content from template"
        assert imported_ws["F13"].value == "hello"

        # Check merge was preserved
        merged_ranges = list(imported_ws.merged_cells.ranges)
        assert len(merged_ranges) == 1
        assert str(merged_ranges[0]) == "A1:C1"

        # Check freeze panes
        assert imported_ws.freeze_panes == "A3"

        # Check auto filter
        assert imported_ws.auto_filter.ref == "A3:C10"

        # Check dimensions (with tolerance for float comparison)
        col_a_width = imported_ws.column_dimensions["A"].width
        assert col_a_width is not None and abs(col_a_width - 24) < 1, (
            f"Column A width should be ~24, got {col_a_width}"
        )

        row_1_height = imported_ws.row_dimensions[1].height
        assert row_1_height is not None and abs(row_1_height - 28) < 1, (
            f"Row 1 height should be ~28, got {row_1_height}"
        )

        print("✓ All assertions passed!")


def test_hybrid_only_imported_sheets() -> None:
    """Test hybrid engine when workbook has only imported sheets (no SheetNodes)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template
        _create_template(template_path)

        # Build workbook with only imported sheet
        workbook = x.workbook()[
            x.import_sheet(template_path, "Template", name="OnlyImported"),
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output
        result_wb = openpyxl.load_workbook(output_path)

        from io import BytesIO

        with zipfile.ZipFile(output_path) as zf:
            assert zf.testzip() is None

        roundtrip = BytesIO()
        result_wb.save(roundtrip)
        assert roundtrip.getvalue()
        assert result_wb.sheetnames == ["OnlyImported"]
        assert result_wb["OnlyImported"]["A1"].value == "Template Title"

        print("✓ Only-imported test passed!")


def test_xlsxwriter_no_imported_sheets() -> None:
    """Test that xlsxwriter works normally when no import_sheet is used."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        output_path = tmppath / "output.xlsx"

        # Build workbook with only generated sheets
        workbook = x.workbook()[
            x.sheet("Sheet1")[x.row()["Hello", "World"]],
            x.sheet("Sheet2")[x.row()["Foo", "Bar"]],
        ]

        # Save with xlsxwriter engine (should use standard path)
        workbook.save(output_path, engine="xlsxwriter")

        # Verify the output
        result_wb = openpyxl.load_workbook(output_path)

        from io import BytesIO

        with zipfile.ZipFile(output_path) as zf:
            assert zf.testzip() is None

        roundtrip = BytesIO()
        result_wb.save(roundtrip)
        assert roundtrip.getvalue()
        assert result_wb.sheetnames == ["Sheet1", "Sheet2"]
        assert result_wb["Sheet1"]["A1"].value == "Hello"
        assert result_wb["Sheet2"]["A1"].value == "Foo"

        print("✓ No-imported test passed!")


def test_xlsxwriter_rejects_import_sheet() -> None:
    """Test that xlsxwriter engine raises error when import_sheet is used."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template
        _create_template(template_path)

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.sheet("Gen")[x.row()["Data"]],
            x.import_sheet(template_path, "Template", name="Imported"),
        ]

        # Attempting to save with xlsxwriter should raise NotImplementedError
        with pytest.raises(NotImplementedError) as excinfo:
            workbook.save(output_path, engine="xlsxwriter")

        # Check error message mentions the alternative engines
        error_message = str(excinfo.value)
        assert "import_sheet" in error_message
        assert "xlsxwriter" in error_message
        assert "hybrid" in error_message or "openpyxl" in error_message

        print("✓ xlsxwriter rejection test passed!")


def test_hybrid_no_imported_sheets() -> None:
    """Test that hybrid engine works when no import_sheet is used.

    Even without imported sheets, the hybrid engine should work correctly,
    generating sheets via xlsxwriter and saving via openpyxl.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        output_path = tmppath / "output.xlsx"

        # Build workbook with only generated sheets
        workbook = x.workbook()[
            x.sheet("Sheet1")[x.row()["Hello", "World"]],
            x.sheet("Sheet2")[x.row()["Foo", "Bar"]],
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output
        result_wb = openpyxl.load_workbook(output_path)

        from io import BytesIO

        with zipfile.ZipFile(output_path) as zf:
            assert zf.testzip() is None

        roundtrip = BytesIO()
        result_wb.save(roundtrip)
        assert roundtrip.getvalue()
        assert result_wb.sheetnames == ["Sheet1", "Sheet2"]
        assert result_wb["Sheet1"]["A1"].value == "Hello"
        assert result_wb["Sheet2"]["A1"].value == "Foo"

        print("✓ Hybrid no-imported test passed!")


def test_hybrid_bytes_output() -> None:
    """Test hybrid engine returns bytes when target is None."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template.xlsx"

        # Create template
        _create_template(template_path)

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.sheet("Gen")[x.row()["Data"]],
            x.import_sheet(template_path, "Template"),
        ]

        # Save with hybrid engine, no target (returns bytes)
        result_bytes = workbook.save(engine="hybrid")

        assert result_bytes is not None
        assert isinstance(result_bytes, bytes)
        assert len(result_bytes) > 0

        # Verify the bytes are a valid xlsx
        from io import BytesIO

        result_wb = openpyxl.load_workbook(BytesIO(result_bytes))
        assert "Gen" in result_wb.sheetnames
        assert "Template" in result_wb.sheetnames

        print("✓ Bytes output test passed!")


def test_hybrid_with_style_heavy_import_sheet() -> None:
    """Regression: importing a style-heavy sheet must not break openpyxl save."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template.xlsx"
        output_path = tmppath / "output.xlsx"

        _create_style_heavy_template(template_path)

        workbook = x.workbook()[
            x.sheet("Gen")[x.row()["Data"]],
            x.import_sheet(template_path, "Template", name="Imported"),
        ]

        workbook.save(output_path, engine="hybrid")

        result_wb = openpyxl.load_workbook(output_path)

        from io import BytesIO

        with zipfile.ZipFile(output_path) as zf:
            assert zf.testzip() is None

        roundtrip = BytesIO()
        result_wb.save(roundtrip)
        assert roundtrip.getvalue()
        assert result_wb.sheetnames == ["Gen", "Imported"]
        assert result_wb["Imported"]["A1"].value == "Row 1"
        # Spot-check a couple of style properties survive the import.
        assert result_wb["Imported"]["A2"].alignment.wrap_text in (True, False)
        assert result_wb["Imported"]["B2"].number_format == "#,##0"


if __name__ == "__main__":
    pytest.main([__file__])
