"""Test that images and charts are preserved when importing sheets.

This test verifies that workbooks with images (photos/pictures) and charts
are correctly copied when using import_sheet() functionality.
"""

from __future__ import annotations

import tempfile
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

import xpyxl as x

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# Skip tests if Pillow is not available (required for image handling)
PIL_AVAILABLE = True
PILImage: Any = None
try:
    from PIL import Image as _PILImage

    PILImage = _PILImage
except ImportError:
    PIL_AVAILABLE = False


def _get_images(ws: object) -> list[Any]:
    """Get images from worksheet (works around type annotation issues)."""
    return list(getattr(ws, "_images", []))


def _get_charts(ws: object) -> list[Any]:
    """Get charts from worksheet (works around type annotation issues)."""
    return list(getattr(ws, "_charts", []))


def _create_test_image() -> BytesIO:
    """Create a simple test PNG image."""
    if not PIL_AVAILABLE or PILImage is None:
        pytest.skip("Pillow not available")
    pil_img = PILImage.new("RGB", (100, 100), color="red")
    img_bytes = BytesIO()
    pil_img.save(img_bytes, format="PNG")
    img_bytes.seek(0)
    return img_bytes


def _create_template_with_image(path: Path) -> None:
    """Create a template workbook containing an image."""
    img_bytes = _create_test_image()

    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "ImageSheet"
    ws["A1"] = "Sheet with image"

    # Add image at B2
    img = Image(img_bytes)
    img.anchor = "B2"
    ws.add_image(img)

    wb.save(path)


def _create_template_with_chart(path: Path) -> None:
    """Create a template workbook containing a chart."""
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "ChartSheet"

    # Add data for chart
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=f"Item {i}")
        ws.cell(row=i, column=2, value=i * 10)

    # Create bar chart
    chart = BarChart()
    chart.title = "Test Chart"
    data = Reference(ws, min_col=2, min_row=1, max_row=10)
    categories = Reference(ws, min_col=1, min_row=1, max_row=10)
    chart.add_data(data)
    chart.set_categories(categories)
    ws.add_chart(chart, "D1")

    wb.save(path)


def _create_template_with_both(path: Path) -> None:
    """Create a template workbook containing both image and chart."""
    img_bytes = _create_test_image()

    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "MixedSheet"
    ws["A1"] = "Sheet with image and chart"

    # Add image at B2
    img = Image(img_bytes)
    img.anchor = "B2"
    ws.add_image(img)

    # Add data for chart
    for i in range(1, 6):
        ws.cell(row=i + 10, column=1, value=f"Value {i}")
        ws.cell(row=i + 10, column=2, value=i * 5)

    # Create bar chart
    chart = BarChart()
    chart.title = "Sample Chart"
    data = Reference(ws, min_col=2, min_row=11, max_row=15)
    chart.add_data(data)
    ws.add_chart(chart, "D15")

    wb.save(path)


@pytest.mark.skipif(not PIL_AVAILABLE, reason="Pillow not installed")
def test_import_sheet_preserves_images() -> None:
    """Test that images are preserved when importing a sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template_with_image.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template with image
        _create_template_with_image(template_path)

        # Verify template has image
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb["ImageSheet"]
        assert len(_get_images(template_ws)) == 1, "Template should have 1 image"

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.sheet("Generated")[x.row()["Hello"]],
            x.import_sheet(template_path, "ImageSheet", name="ImportedImage"),
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output has the image
        result_wb = openpyxl.load_workbook(output_path)
        imported_ws = result_wb["ImportedImage"]

        assert len(_get_images(imported_ws)) == 1, "Imported sheet should have 1 image"
        assert imported_ws["A1"].value == "Sheet with image"  # type: ignore[union-attr]


@pytest.mark.skipif(not PIL_AVAILABLE, reason="Pillow not installed")
def test_import_sheet_preserves_charts() -> None:
    """Test that charts are preserved when importing a sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template_with_chart.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template with chart
        _create_template_with_chart(template_path)

        # Verify template has chart
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb["ChartSheet"]
        assert len(_get_charts(template_ws)) == 1, "Template should have 1 chart"

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.sheet("Generated")[x.row()["Hello"]],
            x.import_sheet(template_path, "ChartSheet", name="ImportedChart"),
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output has the chart
        result_wb = openpyxl.load_workbook(output_path)
        imported_ws = result_wb["ImportedChart"]

        assert len(_get_charts(imported_ws)) == 1, "Imported sheet should have 1 chart"
        # Verify data is preserved
        assert imported_ws["A1"].value == "Item 1"  # type: ignore[union-attr]
        assert imported_ws["B1"].value == 10  # type: ignore[union-attr]


@pytest.mark.skipif(not PIL_AVAILABLE, reason="Pillow not installed")
def test_import_sheet_preserves_images_and_charts() -> None:
    """Test that both images and charts are preserved when importing a sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template_with_both.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template with both image and chart
        _create_template_with_both(template_path)

        # Verify template has both
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb["MixedSheet"]
        assert len(_get_images(template_ws)) == 1, "Template should have 1 image"
        assert len(_get_charts(template_ws)) == 1, "Template should have 1 chart"

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.import_sheet(template_path, "MixedSheet", name="ImportedMixed"),
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output has both
        result_wb = openpyxl.load_workbook(output_path)
        imported_ws = result_wb["ImportedMixed"]

        assert len(_get_images(imported_ws)) == 1, "Imported sheet should have 1 image"
        assert len(_get_charts(imported_ws)) == 1, "Imported sheet should have 1 chart"
        assert imported_ws["A1"].value == "Sheet with image and chart"  # type: ignore[union-attr]


@pytest.mark.skipif(not PIL_AVAILABLE, reason="Pillow not installed")
def test_import_sheet_preserves_multiple_images() -> None:
    """Test that multiple images are preserved when importing a sheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template_multi_image.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template with multiple images
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        ws.title = "MultiImage"
        ws["A1"] = "Multiple images"

        # Add 3 images at different positions
        for _i, anchor in enumerate(["B2", "D2", "F2"]):
            img_bytes = _create_test_image()
            img = Image(img_bytes)
            img.anchor = anchor
            ws.add_image(img)

        wb.save(template_path)

        # Verify template has 3 images
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb["MultiImage"]
        assert len(_get_images(template_ws)) == 3, "Template should have 3 images"

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.import_sheet(template_path, "MultiImage", name="ImportedMulti"),
        ]

        # Save with hybrid engine
        workbook.save(output_path, engine="hybrid")

        # Verify the output has all 3 images
        result_wb = openpyxl.load_workbook(output_path)
        imported_ws = result_wb["ImportedMulti"]

        assert len(_get_images(imported_ws)) == 3, "Imported sheet should have 3 images"


@pytest.mark.skipif(not PIL_AVAILABLE, reason="Pillow not installed")
def test_import_sheet_with_openpyxl_engine() -> None:
    """Test that images are preserved when using openpyxl engine directly."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmppath = Path(tmpdir)
        template_path = tmppath / "template_with_image.xlsx"
        output_path = tmppath / "output.xlsx"

        # Create template with image
        _create_template_with_image(template_path)

        # Build workbook with imported sheet
        workbook = x.workbook()[
            x.import_sheet(template_path, "ImageSheet", name="ImportedImage"),
        ]

        # Save with openpyxl engine
        workbook.save(output_path, engine="openpyxl")

        # Verify the output has the image
        result_wb = openpyxl.load_workbook(output_path)
        imported_ws = result_wb["ImportedImage"]

        assert len(_get_images(imported_ws)) == 1, "Imported sheet should have 1 image"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
