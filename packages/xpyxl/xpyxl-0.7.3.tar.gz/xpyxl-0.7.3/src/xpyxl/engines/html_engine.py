"""HTML rendering engine implementation."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from html import escape
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from ..styles import BorderStyleName, normalize_hex
from .base import EffectiveStyle, Engine, SaveTarget

__all__ = ["HtmlEngine"]

DEFAULT_FONT_NAME = "Calibri"
DEFAULT_FONT_SIZE = 11.0
DEFAULT_TEXT_COLOR = normalize_hex("#000000")
DEFAULT_BORDER_COLOR = normalize_hex("#000000")

COLUMN_WIDTH_PX = 7.0
BASE_CELL_PADDING_PX = 6


@dataclass
class _CellData:
    value: object
    style: EffectiveStyle
    border_fallback_color: str


@dataclass
class _SheetData:
    name: str
    cells: dict[tuple[int, int], _CellData] = field(default_factory=dict)
    col_widths: dict[int, float] = field(default_factory=dict)
    row_heights: dict[int, float] = field(default_factory=dict)
    background_color: str | None = None
    background_max_row: int = 0
    background_max_col: int = 0
    max_row: int = 0
    max_col: int = 0


class HtmlEngine(Engine):
    """Rendering engine that outputs HTML with Tailwind CSS."""

    def __init__(self) -> None:
        super().__init__()
        self._sheets: list[_SheetData] = []
        self._current_sheet: _SheetData | None = None

    def create_sheet(self, name: str) -> None:
        sheet = _SheetData(name=name)
        self._sheets.append(sheet)
        self._current_sheet = sheet

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        sheet = self._require_sheet()
        sheet.cells[(row, col)] = _CellData(
            value=value,
            style=style,
            border_fallback_color=normalize_hex(border_fallback_color),
        )
        sheet.max_row = max(sheet.max_row, row)
        sheet.max_col = max(sheet.max_col, col)

    def set_column_width(self, col: int, width: float) -> None:
        sheet = self._require_sheet()
        sheet.col_widths[col] = width
        sheet.max_col = max(sheet.max_col, col)

    def set_row_height(self, row: int, height: float) -> None:
        sheet = self._require_sheet()
        sheet.row_heights[row] = height
        sheet.max_row = max(sheet.max_row, row)

    def fill_background(self, color: str, max_row: int, max_col: int) -> None:
        sheet = self._require_sheet()
        sheet.background_color = normalize_hex(color)
        sheet.background_max_row = max(sheet.background_max_row, max_row)
        sheet.background_max_col = max(sheet.background_max_col, max_col)

    def copy_sheet(
        self, source: SaveTarget | bytes | BinaryIO, sheet_name: str, dest_name: str
    ) -> None:
        source_wb = self._load_source_workbook(source)
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in source workbook")
        source_ws = source_wb[sheet_name]

        self.create_sheet(dest_name)
        sheet = self._require_sheet()

        for row in source_ws.iter_rows():
            for cell in row:
                if cell.value is None and not getattr(cell, "has_style", False):
                    continue
                if cell.row is None or cell.column is None:
                    continue
                effective = self._openpyxl_cell_to_effective_style(cell)
                sheet.cells[(cell.row, cell.column)] = _CellData(
                    value=cell.value,
                    style=effective,
                    border_fallback_color=DEFAULT_BORDER_COLOR,
                )
                sheet.max_row = max(sheet.max_row, cell.row)
                sheet.max_col = max(sheet.max_col, cell.column)

        for col_letter, dim in source_ws.column_dimensions.items():
            if dim.width:
                col_idx = _column_letter_to_index(col_letter)
                sheet.col_widths[col_idx] = dim.width
                sheet.max_col = max(sheet.max_col, col_idx)

        for row_idx, dim in source_ws.row_dimensions.items():
            if dim.height:
                sheet.row_heights[row_idx] = dim.height
                sheet.max_row = max(sheet.max_row, row_idx)

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        html = self._render_html()
        html_bytes = html.encode("utf-8")

        if target is None:
            return html_bytes
        if isinstance(target, (str, Path)):
            Path(target).write_bytes(html_bytes)
        else:
            target.write(html_bytes)
            if hasattr(target, "flush"):
                target.flush()
        return None

    def _render_html(self) -> str:
        tabs_html: list[str] = []
        sheets_html: list[str] = []

        for idx, sheet in enumerate(self._sheets):
            sheet_id = f"sheet-{idx}"
            is_active = idx == 0

            base_tab_classes = " ".join(
                [
                    "px-4",
                    "py-2",
                    "border",
                    "border-b-0",
                    "rounded-t",
                    "text-sm",
                    "font-medium",
                    "transition",
                    "duration-150",
                ]
            )
            active_tab_classes = "bg-white text-gray-900 border-gray-300"
            inactive_tab_classes = "bg-gray-200 text-gray-600 border-gray-200"
            tab_classes = (
                f"{base_tab_classes} "
                f"{active_tab_classes if is_active else inactive_tab_classes}"
            )
            tabs_html.append(
                """
                <button
                    type="button"
                    class="{tab_classes}"
                    data-sheet="{sheet_id}"
                >{label}</button>
                """.format(
                    tab_classes=tab_classes,
                    sheet_id=sheet_id,
                    label=escape(sheet.name),
                ).strip()
            )

            display = "block" if is_active else "none"
            bg_style = (
                f"background-color: {sheet.background_color};"
                if sheet.background_color
                else ""
            )
            table_html = self._render_sheet_table(sheet)
            sheets_html.append(
                """
                <div
                    id="{sheet_id}"
                    class="sheet-content"
                    style="display: {display}; {bg_style}"
                >{table}</div>
                """.format(
                    sheet_id=sheet_id,
                    display=display,
                    bg_style=bg_style,
                    table=table_html,
                ).strip()
            )

        tabs_markup = "".join(tabs_html)
        sheets_markup = "".join(sheets_html) if sheets_html else ""
        return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"UTF-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
  <title>Workbook</title>
  <script src=\"https://cdn.tailwindcss.com\"></script>
  <style>
    .sheet-content {{ min-height: 320px; overflow: auto; padding: 16px; }}
    table {{ border-collapse: collapse; }}
    td {{ padding: 2px 6px; vertical-align: top; }}
  </style>
</head>
<body class=\"bg-slate-100 p-6\">
  <div class=\"max-w-full\">
    <div class=\"border border-gray-300 rounded-lg bg-white shadow-sm\">
      <div class=\"flex gap-2 border-b border-gray-300 px-3 pt-3\">
        {tabs_markup}
      </div>
      <div class=\"p-2\">
        {sheets_markup}
      </div>
    </div>
  </div>
  <script>
    const tabs = Array.from(document.querySelectorAll(\"[data-sheet]\"));
    const sheets = Array.from(document.querySelectorAll(\".sheet-content\"));
    const activeClasses = [\"bg-white\", \"text-gray-900\", \"border-gray-300\"];
    const inactiveClasses = [\"bg-gray-200\", \"text-gray-600\", \"border-gray-200\"];

    const setActive = (sheetId, activeButton) => {{
      sheets.forEach((panel) => {{
        panel.style.display = panel.id === sheetId ? \"block\" : \"none\";
      }});
      tabs.forEach((button) => {{
        const isActive = button === activeButton;
        button.classList.remove(...activeClasses, ...inactiveClasses);
        button.classList.add(...(isActive ? activeClasses : inactiveClasses));
      }});
    }};

    if (tabs.length > 0) {{
      tabs.forEach((button) => {{
        button.addEventListener(\"click\", () => {{
          setActive(button.dataset.sheet, button);
        }});
      }});
      setActive(tabs[0].dataset.sheet, tabs[0]);
    }}
  </script>
</body>
</html>
"""

    def _render_sheet_table(self, sheet: _SheetData) -> str:
        max_row = max(sheet.max_row, sheet.background_max_row)
        max_col = max(sheet.max_col, sheet.background_max_col)
        max_row = max(max_row, max(sheet.row_heights.keys(), default=0))
        max_col = max(max_col, max(sheet.col_widths.keys(), default=0))

        if max_row == 0 or max_col == 0:
            return "<div class=\"text-gray-500\">Empty sheet</div>"

        colgroup = self._render_colgroup(max_col, sheet.col_widths)
        rows_html: list[str] = []
        for row in range(1, max_row + 1):
            cells_html: list[str] = []
            row_height = sheet.row_heights.get(row)
            for col in range(1, max_col + 1):
                cell_data = sheet.cells.get((row, col))
                if cell_data:
                    classes, inline_css = self._style_to_css(
                        cell_data.style, cell_data.border_fallback_color
                    )
                    value = _format_value(cell_data.value, cell_data.style.number_format)
                else:
                    classes = ""
                    inline_css = ""
                    value = ""
                cells_html.append(
                    "<td class=\"{classes}\" style=\"{style}\">{value}</td>".format(
                        classes=classes,
                        style=inline_css,
                        value=escape(value),
                    )
                )
            row_style = f"height: {row_height}pt;" if row_height else ""
            rows_html.append(
                "<tr style=\"{style}\">{cells}</tr>".format(
                    style=row_style, cells="".join(cells_html)
                )
            )

        return "<table class=\"text-sm\">{colgroup}{rows}</table>".format(
            colgroup=colgroup,
            rows="".join(rows_html),
        )

    def _render_colgroup(self, max_col: int, col_widths: dict[int, float]) -> str:
        if not col_widths:
            return ""
        cols: list[str] = []
        for col in range(1, max_col + 1):
            width = col_widths.get(col)
            if width:
                px_width = width * COLUMN_WIDTH_PX
                cols.append(f"<col style=\"width: {px_width:.1f}px;\">")
            else:
                cols.append("<col>")
        return "<colgroup>{}</colgroup>".format("".join(cols))

    def _style_to_css(
        self, style: EffectiveStyle, border_fallback_color: str
    ) -> tuple[str, str]:
        classes: list[str] = []
        css_parts: list[str] = []

        css_parts.append("font-family: '{}' , sans-serif".format(style.font_name))
        css_parts.append("font-size: {}pt".format(style.font_size))
        css_parts.append("color: {}".format(style.text_color))

        if style.fill_color:
            css_parts.append("background-color: {}".format(style.fill_color))

        if style.bold:
            classes.append("font-bold")
        if style.italic:
            classes.append("italic")

        align_map = {
            "left": "text-left",
            "center": "text-center",
            "right": "text-right",
        }
        if style.horizontal_align in align_map:
            classes.append(align_map[style.horizontal_align])

        valign_map = {
            "top": "align-top",
            "center": "align-middle",
            "bottom": "align-bottom",
        }
        if style.vertical_align in valign_map:
            classes.append(valign_map[style.vertical_align])

        if style.indent:
            padding = BASE_CELL_PADDING_PX + style.indent * 8
            css_parts.append(f"padding-left: {padding}px")

        if style.wrap_text:
            classes.append("whitespace-normal")
            classes.append("break-words")
        else:
            classes.append("whitespace-nowrap")

        if style.shrink_to_fit:
            classes.append("truncate")

        border_css = self._border_to_css(style, border_fallback_color)
        if border_css:
            css_parts.extend(border_css)

        return " ".join(classes).strip(), "; ".join(css_parts).strip()

    def _border_to_css(
        self, style: EffectiveStyle, border_fallback_color: str
    ) -> list[str]:
        if style.border == "none":
            return []
        if not style.border and not (
            style.border_top
            or style.border_bottom
            or style.border_left
            or style.border_right
        ):
            return []

        border_style: BorderStyleName = style.border or "thin"
        border_color = style.border_color or border_fallback_color
        width = _border_style_to_width(border_style)
        css_style = _border_style_to_css(border_style)

        parts: list[str] = []
        if style.border_top or style.border_bottom or style.border_left or style.border_right:
            if style.border_top:
                parts.append(f"border-top: {width} {css_style} {border_color}")
            if style.border_bottom:
                parts.append(f"border-bottom: {width} {css_style} {border_color}")
            if style.border_left:
                parts.append(f"border-left: {width} {css_style} {border_color}")
            if style.border_right:
                parts.append(f"border-right: {width} {css_style} {border_color}")
        else:
            parts.append(f"border: {width} {css_style} {border_color}")
        return parts

    def _require_sheet(self) -> _SheetData:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")
        return self._current_sheet

    def _load_source_workbook(self, source: SaveTarget | bytes | BinaryIO):
        if isinstance(source, (str, Path)):
            return load_workbook(filename=source, data_only=False)
        if isinstance(source, bytes):
            buffer: BinaryIO = BytesIO(source)
        else:
            buffer = source
            if hasattr(buffer, "seek"):
                try:
                    buffer.seek(0)
                except Exception:
                    pass
        return load_workbook(buffer, data_only=False)

    def _openpyxl_cell_to_effective_style(self, cell) -> EffectiveStyle:
        font = cell.font
        fill = cell.fill
        alignment = cell.alignment
        border = cell.border

        font_name = font.name if font and font.name else DEFAULT_FONT_NAME
        font_size = float(font.size) if font and font.size else DEFAULT_FONT_SIZE
        bold = bool(font.bold) if font else False
        italic = bool(font.italic) if font else False

        text_color = _openpyxl_color_to_hex(font.color) or DEFAULT_TEXT_COLOR

        fill_color = None
        if fill and getattr(fill, "fill_type", None) == "solid":
            fill_color = _openpyxl_color_to_hex(getattr(fill, "fgColor", None))

        horizontal_align = alignment.horizontal if alignment else None
        vertical_align = alignment.vertical if alignment else None
        indent = alignment.indent if alignment and alignment.indent else None
        wrap_text = bool(alignment.wrap_text) if alignment else False
        shrink_to_fit = bool(alignment.shrink_to_fit) if alignment else False

        border_style = None
        border_color = None
        border_top = False
        border_bottom = False
        border_left = False
        border_right = False

        for side_name in ("top", "bottom", "left", "right"):
            side = getattr(border, side_name, None)
            if not side or not side.style or side.style == "none":
                continue
            if border_style is None:
                border_style = side.style
            if side.color and not border_color:
                border_color = _openpyxl_color_to_hex(side.color)
            if side_name == "top":
                border_top = True
            elif side_name == "bottom":
                border_bottom = True
            elif side_name == "left":
                border_left = True
            elif side_name == "right":
                border_right = True

        return EffectiveStyle(
            font_name=font_name,
            font_size=font_size,
            bold=bold,
            italic=italic,
            text_color=text_color,
            fill_color=fill_color,
            horizontal_align=horizontal_align,
            vertical_align=vertical_align,
            indent=indent,
            wrap_text=wrap_text,
            shrink_to_fit=shrink_to_fit,
            auto_width=True,
            row_height=None,
            row_width=None,
            number_format=cell.number_format if cell.number_format != "General" else None,
            border=border_style,
            border_color=border_color,
            border_top=border_top,
            border_bottom=border_bottom,
            border_left=border_left,
            border_right=border_right,
        )


def _border_style_to_width(border: BorderStyleName) -> str:
    width_map = {
        "thin": "1px",
        "hair": "1px",
        "medium": "2px",
        "mediumDashed": "2px",
        "mediumDashDot": "2px",
        "mediumDashDotDot": "2px",
        "thick": "3px",
        "double": "3px",
    }
    return width_map.get(border, "1px")


def _border_style_to_css(border: BorderStyleName) -> str:
    style_map = {
        "thin": "solid",
        "medium": "solid",
        "thick": "solid",
        "hair": "solid",
        "dashed": "dashed",
        "mediumDashed": "dashed",
        "dotted": "dotted",
        "double": "double",
        "dashDot": "dashed",
        "mediumDashDot": "dashed",
        "dashDotDot": "dotted",
        "mediumDashDotDot": "dotted",
        "slantDashDot": "dashed",
    }
    return style_map.get(border, "solid")


def _openpyxl_color_to_hex(color, fallback: str | None = None) -> str | None:
    if color is None:
        return fallback
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6:
        return normalize_hex("#" + rgb[-6:])
    return fallback


def _format_value(value: object, number_format: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (date, datetime, time)):
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        return value.isoformat()
    if number_format and isinstance(value, (int, float)):
        if "%" in number_format:
            return f"{value * 100:.2f}%"
        if "$" in number_format:
            return f"${value:,.2f}"
        if "€" in number_format:
            return f"€{value:,.2f}"
        if "#,##0" in number_format:
            if ".00" in number_format or "0.00" in number_format:
                return f"{value:,.2f}"
            return f"{value:,.0f}"
    return str(value)


def _column_letter_to_index(letter: str) -> int:
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result
