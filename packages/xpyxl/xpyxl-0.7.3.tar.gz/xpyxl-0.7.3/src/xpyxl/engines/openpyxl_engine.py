"""OpenPyXL rendering engine implementation."""

from __future__ import annotations

import copy
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..styles import to_argb
from .base import EffectiveStyle, Engine, SaveTarget

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["OpenpyxlEngine"]


class OpenpyxlEngine(Engine):
    """Rendering engine using openpyxl."""

    def __init__(self) -> None:
        super().__init__()
        self._workbook = Workbook()
        # Remove default sheet created by openpyxl
        default_sheet = self._workbook.active
        if default_sheet is not None:
            self._workbook.remove(default_sheet)
        self._init_instance_vars()

    def _init_instance_vars(self) -> None:
        """Initialize common instance variables used by both __init__ and from_workbook."""
        self._current_sheet: Worksheet | None = None
        # Cache style objects to avoid duplicates
        self._style_cache: dict[tuple[Any, ...], dict[str, Any]] = {}
        # Cache color conversions
        self._color_cache: dict[str, str] = {}
        # Cache column letters
        self._column_letter_cache: dict[int, str] = {}

    @classmethod
    def from_workbook(cls, workbook: Workbook) -> "OpenpyxlEngine":
        """Create an engine instance wrapping an existing openpyxl Workbook.

        This allows reusing copy_sheet() and other engine methods against
        a workbook that was loaded or created externally (e.g., from xlsxwriter bytes).
        """
        engine = cls.__new__(cls)
        engine._workbook = workbook
        engine._init_instance_vars()
        return engine

    def create_sheet(self, name: str) -> None:
        self._current_sheet = self._workbook.create_sheet(title=name)

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        cell = self._current_sheet.cell(row=row, column=col)
        cell.value = value  # type: ignore[assignment]
        self._apply_style(cell, style, border_fallback_color)

    def _get_cached_color(self, color: str) -> str:
        """Get cached ARGB color or convert and cache it."""
        if color not in self._color_cache:
            self._color_cache[color] = to_argb(color)
        return self._color_cache[color]

    def _get_cached_styles(
        self, effective: EffectiveStyle, border_fallback_color: str
    ) -> dict[str, Any]:
        """Get or create cached style objects for the given style."""
        # Create a hashable key from style properties
        cache_key = (
            effective.font_name,
            effective.font_size,
            effective.bold,
            effective.italic,
            effective.text_color,
            effective.fill_color,
            effective.horizontal_align,
            effective.vertical_align,
            effective.indent,
            effective.wrap_text,
            effective.shrink_to_fit,
            effective.number_format,
            effective.border,
            effective.border_color or border_fallback_color
            if effective.border
            else None,
            effective.border_top,
            effective.border_bottom,
            effective.border_left,
            effective.border_right,
        )

        if cache_key in self._style_cache:
            return self._style_cache[cache_key]

        # Create style objects
        text_color_argb = self._get_cached_color(effective.text_color)
        font = Font(
            name=effective.font_name,
            size=effective.font_size,
            bold=effective.bold,
            italic=effective.italic,
            color=text_color_argb,
        )

        fill: PatternFill | None = None
        if effective.fill_color:
            fill_color_argb = self._get_cached_color(effective.fill_color)
            fill = PatternFill(
                fill_type="solid",
                start_color=fill_color_argb,
                end_color=fill_color_argb,
            )

        alignment: Alignment | None = None
        align_kwargs: dict[str, Any] = {}
        if effective.horizontal_align:
            align_kwargs["horizontal"] = effective.horizontal_align
        if effective.vertical_align:
            align_kwargs["vertical"] = effective.vertical_align
        if effective.indent is not None:
            align_kwargs["indent"] = effective.indent
        if effective.wrap_text:
            align_kwargs["wrap_text"] = True
        if effective.shrink_to_fit:
            align_kwargs["shrink_to_fit"] = True
        if align_kwargs:
            align_kwargs.setdefault("vertical", "bottom")
            alignment = Alignment(**align_kwargs)  # type: ignore[arg-type]
        elif effective.wrap_text or effective.shrink_to_fit:
            alignment = Alignment(
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        border: Border | None = None
        if effective.border:
            border_color = effective.border_color or border_fallback_color
            border_color_argb = self._get_cached_color(border_color)

            def build_side(enabled: bool) -> Side | None:
                if not enabled:
                    return None
                return Side(style=effective.border, color=border_color_argb)

            explicit = (
                effective.border_top
                or effective.border_bottom
                or effective.border_left
                or effective.border_right
            )
            if explicit:
                border = Border(
                    left=build_side(effective.border_left),
                    right=build_side(effective.border_right),
                    top=build_side(effective.border_top),
                    bottom=build_side(effective.border_bottom),
                )
            else:
                side = build_side(True)
                border = Border(left=side, right=side, top=side, bottom=side)

        styles = {
            "font": font,
            "fill": fill,
            "alignment": alignment,
            "border": border,
            "number_format": effective.number_format,
        }

        self._style_cache[cache_key] = styles
        return styles

    def _apply_style(
        self, cell: object, effective: EffectiveStyle, border_fallback_color: str
    ) -> None:
        """Apply style to an openpyxl cell."""
        styles = self._get_cached_styles(effective, border_fallback_color)

        cell.font = styles["font"]  # type: ignore[attr-defined]

        if styles["fill"]:
            cell.fill = styles["fill"]  # type: ignore[attr-defined]

        if styles["alignment"]:
            cell.alignment = styles["alignment"]  # type: ignore[attr-defined]
        elif cell.alignment is None and (  # type: ignore[attr-defined]
            effective.wrap_text or effective.shrink_to_fit
        ):
            # Handle edge case where alignment wasn't cached but wrap/shrink is needed
            cell.alignment = Alignment(  # type: ignore[attr-defined]
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        if styles["number_format"]:
            cell.number_format = styles["number_format"]  # type: ignore[attr-defined]

        if styles["border"]:
            cell.border = styles["border"]  # type: ignore[attr-defined]

    def set_column_width(self, col: int, width: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        if col not in self._column_letter_cache:
            self._column_letter_cache[col] = get_column_letter(col)
        letter = self._column_letter_cache[col]
        self._current_sheet.column_dimensions[letter].width = max(width, 8.0)

    def set_row_height(self, row: int, height: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        self._current_sheet.row_dimensions[row].height = height

    def fill_background(
        self,
        color: str,
        max_row: int,
        max_col: int,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        fill_color = self._get_cached_color(color)
        sheet_fill = PatternFill(
            fill_type="solid", start_color=fill_color, end_color=fill_color
        )
        # Reuse the same PatternFill object for all cells (openpyxl supports this)
        for row in self._current_sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col
        ):
            for cell in row:
                cell.fill = sheet_fill

    def _ensure_named_styles(self, source_wb: Workbook) -> None:
        """Ensure named styles used by imported sheets exist in the destination workbook."""
        try:
            source_styles = list(getattr(source_wb, "named_styles", []))
            dest_styles = list(getattr(self._workbook, "named_styles", []))
        except Exception:
            return

        existing = {getattr(style, "name", None) for style in dest_styles}
        for style in source_styles:
            name = getattr(style, "name", None)
            if not name or name in existing:
                continue
            try:
                self._workbook.add_named_style(copy.copy(style))
                existing.add(name)
            except Exception:
                # NamedStyle copying is best-effort; cell-level style objects
                # are still copied below.
                pass

    def _clone_sheet_contents(self, source_ws: Worksheet, target_ws: Worksheet) -> None:
        """Copy values and basic layout from a source sheet into a target sheet.

        This avoids copying workbook-scoped style indices (which can corrupt files
        cross-workbook) and instead copies resolved style objects.
        """
        # Basic sheet properties (safe, value-based)
        try:
            target_ws.sheet_properties = copy.copy(source_ws.sheet_properties)
        except Exception:
            pass
        try:
            target_ws.sheet_format = copy.copy(source_ws.sheet_format)
        except Exception:
            pass
        try:
            target_ws.page_margins = copy.copy(source_ws.page_margins)
        except Exception:
            pass
        try:
            target_ws.page_setup = copy.copy(source_ws.page_setup)
        except Exception:
            pass
        try:
            target_ws.print_options = copy.copy(source_ws.print_options)
        except Exception:
            pass
        try:
            target_ws.protection = copy.copy(source_ws.protection)
        except Exception:
            pass

        try:
            target_ws.freeze_panes = source_ws.freeze_panes
        except Exception:
            pass
        try:
            target_ws.auto_filter.ref = source_ws.auto_filter.ref
        except Exception:
            pass
        try:
            target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
        except Exception:
            pass

        # Values (including formulas) + cell-level style objects.
        for row in source_ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.row is None or cell.column is None:
                    continue
                target_cell = target_ws.cell(
                    row=cell.row, column=cell.column, value=cell.value
                )
                if getattr(cell, "has_style", False):
                    # Copy resolved style objects, not style indices.
                    target_cell.font = copy.copy(cell.font)  # pyright: ignore[reportAttributeAccessIssue]
                    target_cell.fill = copy.copy(cell.fill)  # pyright: ignore[reportAttributeAccessIssue]
                    target_cell.border = copy.copy(cell.border)  # pyright: ignore[reportAttributeAccessIssue]
                    target_cell.alignment = copy.copy(cell.alignment)  # pyright: ignore[reportAttributeAccessIssue]
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy.copy(cell.protection)  # pyright: ignore[reportAttributeAccessIssue]
                if getattr(cell, "hyperlink", None):
                    target_cell._hyperlink = copy.copy(cell._hyperlink)  # type: ignore[attr-defined]
                if cell.comment:
                    target_cell.comment = copy.copy(cell.comment)

        # Merge ranges
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # Data validations
        try:
            if source_ws.data_validations is not None:
                for dv in source_ws.data_validations.dataValidation:
                    cloned = copy.copy(dv)
                    cloned.ranges = list(dv.ranges)
                    target_ws.add_data_validation(cloned)
        except Exception:
            pass

        # Conditional formatting (best-effort)
        try:
            if source_ws.conditional_formatting:
                for key, rules in source_ws.conditional_formatting._cf_rules.items():  # type: ignore[attr-defined]
                    target_ws.conditional_formatting._cf_rules[key] = copy.deepcopy(  # type: ignore[attr-defined]
                        rules
                    )
        except Exception:
            pass

        def _copy_dimension_attrs(source_dim: object, dest_dim: object) -> None:
            # Copy common, non-style attributes; skip workbook-scoped style internals.
            for attr in (
                "width",
                "height",
                "hidden",
                "outlineLevel",
                "collapsed",
                "bestFit",
                "customWidth",
                "customHeight",
            ):
                try:
                    if hasattr(source_dim, attr):
                        value = getattr(source_dim, attr)
                        if value is not None:
                            setattr(dest_dim, attr, value)
                except Exception:
                    pass

        for column, dimension in source_ws.column_dimensions.items():
            _copy_dimension_attrs(dimension, target_ws.column_dimensions[column])

        for row, dimension in source_ws.row_dimensions.items():
            _copy_dimension_attrs(dimension, target_ws.row_dimensions[row])

        # Images (photos/pictures)
        self._copy_images(source_ws, target_ws)

        # Charts
        self._copy_charts(source_ws, target_ws)

    def _copy_images(self, source_ws: Worksheet, target_ws: Worksheet) -> None:
        """Copy all images from source worksheet to target worksheet."""
        try:
            from openpyxl.drawing.image import Image
        except ImportError:
            # Image support not available
            return

        for img in getattr(source_ws, "_images", []):
            try:
                # Get the image data from the ref (BytesIO containing the image data)
                ref = getattr(img, "ref", None)
                if ref is None:
                    continue

                # Seek to start if possible
                if hasattr(ref, "seek"):
                    try:
                        ref.seek(0)
                    except Exception:
                        pass

                # Create a new Image instance from the image data
                new_img = Image(ref)

                # Copy anchor (position in the sheet)
                anchor = getattr(img, "anchor", None)
                if anchor is not None:
                    new_img.anchor = anchor

                # Copy dimensions if set
                if getattr(img, "width", None) is not None:
                    new_img.width = img.width
                if getattr(img, "height", None) is not None:
                    new_img.height = img.height

                target_ws.add_image(new_img)
            except Exception:
                # Image copying is best-effort; skip failures
                pass

    def _copy_charts(self, source_ws: Worksheet, target_ws: Worksheet) -> None:
        """Copy all charts from source worksheet to target worksheet.

        When copying charts, we need to update the data references to point
        to the target sheet instead of the source sheet, since sheet names
        may differ between source and destination.
        """
        source_name = source_ws.title
        target_name = target_ws.title

        for chart in getattr(source_ws, "_charts", []):
            try:
                # Deep copy the chart to avoid sharing references
                new_chart = copy.deepcopy(chart)

                # Update sheet references in chart data
                self._update_chart_references(new_chart, source_name, target_name)

                target_ws.add_chart(new_chart)
            except Exception:
                # Chart copying is best-effort; skip failures
                pass

    def _update_chart_references(
        self, chart: object, old_sheet: str, new_sheet: str
    ) -> None:
        """Update all sheet references in a chart from old_sheet to new_sheet."""
        import re

        def update_ref(formula: str | None) -> str | None:
            if not formula:
                return formula

            # Escape special regex chars in old_sheet
            old_escaped = re.escape(old_sheet)

            # Pattern for quoted sheet name: 'Sheet Name'!
            quoted_pattern = f"'{old_escaped}'!"
            # Pattern for unquoted sheet name: SheetName!
            unquoted_pattern = f"{old_escaped}!"

            # New sheet name (quote if has spaces or special chars)
            if " " in new_sheet or any(c in new_sheet for c in "'![]:"):
                new_ref = f"'{new_sheet}'!"
            else:
                new_ref = f"{new_sheet}!"

            # Replace quoted version first
            result = re.sub(quoted_pattern, new_ref, formula)
            # Then try unquoted version if no change
            if result == formula:
                result = re.sub(unquoted_pattern, new_ref, formula)

            return result

        # Update series data references
        for series in getattr(chart, "series", []):
            # Value reference (numRef)
            if hasattr(series, "val") and series.val:
                num_ref = getattr(series.val, "numRef", None)
                if num_ref and hasattr(num_ref, "f"):
                    num_ref.f = update_ref(num_ref.f)

            # Category reference (can be numRef or strRef)
            if hasattr(series, "cat") and series.cat:
                cat_num_ref = getattr(series.cat, "numRef", None)
                if cat_num_ref and hasattr(cat_num_ref, "f"):
                    cat_num_ref.f = update_ref(cat_num_ref.f)

                cat_str_ref = getattr(series.cat, "strRef", None)
                if cat_str_ref and hasattr(cat_str_ref, "f"):
                    cat_str_ref.f = update_ref(cat_str_ref.f)

            # X values (for scatter/bubble charts)
            if hasattr(series, "xVal") and series.xVal:
                x_num_ref = getattr(series.xVal, "numRef", None)
                if x_num_ref and hasattr(x_num_ref, "f"):
                    x_num_ref.f = update_ref(x_num_ref.f)

            # Y values (for scatter/bubble charts)
            if hasattr(series, "yVal") and series.yVal:
                y_num_ref = getattr(series.yVal, "numRef", None)
                if y_num_ref and hasattr(y_num_ref, "f"):
                    y_num_ref.f = update_ref(y_num_ref.f)

    def _load_source_workbook(self, source: SaveTarget | bytes | BinaryIO) -> Workbook:
        if isinstance(source, (str, Path)):
            return load_workbook(
                filename=source,
                data_only=False,
            )

        buffer: BinaryIO
        if isinstance(source, bytes):
            buffer = BytesIO(source)
        else:
            buffer = source
            if hasattr(buffer, "seek"):
                try:
                    buffer.seek(0)
                except Exception:
                    pass

        return load_workbook(
            buffer,
            data_only=False,
        )

    def _ensure_sheet_name_available(self, name: str) -> None:
        if name in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' already exists in destination workbook")

    def copy_sheet(
        self, source: SaveTarget | bytes | BinaryIO, sheet_name: str, dest_name: str
    ) -> None:
        """Copy a sheet from an external workbook into the current workbook.

        Always copies via _clone_sheet_contents to ensure consistent behavior
        regardless of whether imported sheets come before or after generated sheets.
        """
        source_wb = self._load_source_workbook(source)

        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in source workbook")

        self._ensure_sheet_name_available(dest_name)
        self._ensure_named_styles(source_wb)

        source_ws = source_wb[sheet_name]
        target_ws = self._workbook.create_sheet(title=dest_name)
        self._clone_sheet_contents(source_ws, target_ws)

        self._current_sheet = target_ws

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        if target is None:
            buffer = BytesIO()
            self._workbook.save(buffer)
            return buffer.getvalue()

        if isinstance(target, (str, Path)):
            self._workbook.save(str(target))
        else:
            self._workbook.save(target)
        return None
