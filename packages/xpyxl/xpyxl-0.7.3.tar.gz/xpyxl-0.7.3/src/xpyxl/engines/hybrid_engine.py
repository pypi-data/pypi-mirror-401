"""Hybrid rendering engine combining xlsxwriter speed with openpyxl features.

This engine uses xlsxwriter for fast generation of new sheets and openpyxl
for importing existing sheets and performing the final save.
"""

from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

from openpyxl import Workbook as _OpenpyxlWorkbook
from openpyxl import load_workbook as _load_workbook

from .base import EffectiveStyle, Engine, SaveTarget
from .openpyxl_engine import OpenpyxlEngine
from .xlsxwriter_engine import XlsxWriterEngine

__all__ = ["HybridEngine"]


class HybridEngine(Engine):
    """Hybrid engine combining xlsxwriter performance with openpyxl features.

    This engine uses xlsxwriter internally for fast generation of new sheets,
    while supporting import_sheet via openpyxl. The final workbook is saved
    using openpyxl to merge generated and imported sheets.

    Use this engine when you need both:
    - Fast generation of large workbooks (xlsxwriter)
    - Ability to import sheets from existing workbooks (openpyxl)
    """

    def __init__(self) -> None:
        super().__init__()
        self._xlsx_engine = XlsxWriterEngine()
        # Deferred import operations: (source, source_sheet_name, dest_name)
        self._imports: list[tuple[SaveTarget | bytes | BinaryIO, str, str]] = []
        # Track sheet names in declaration order for reordering
        self._sheet_order: list[str] = []
        # Track if we have any generated sheets
        self._has_generated_sheets = False

    def create_sheet(self, name: str) -> None:
        """Create a new worksheet, delegating to xlsxwriter."""
        self._xlsx_engine.create_sheet(name)
        self._sheet_order.append(name)
        self._has_generated_sheets = True

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        """Write a cell value, delegating to xlsxwriter."""
        self._xlsx_engine.write_cell(row, col, value, style, border_fallback_color)

    def set_column_width(self, col: int, width: float) -> None:
        """Set column width, delegating to xlsxwriter."""
        self._xlsx_engine.set_column_width(col, width)

    def set_row_height(self, row: int, height: float) -> None:
        """Set row height, delegating to xlsxwriter."""
        self._xlsx_engine.set_row_height(row, height)

    def fill_background(
        self,
        color: str,
        max_row: int,
        max_col: int,
    ) -> None:
        """Fill background color, delegating to xlsxwriter."""
        self._xlsx_engine.fill_background(color, max_row, max_col)

    def copy_sheet(
        self, source: SaveTarget | bytes | BinaryIO, sheet_name: str, dest_name: str
    ) -> None:
        """Record a deferred import operation to be applied at save time.

        Unlike xlsxwriter which cannot import sheets, HybridEngine defers
        these operations and applies them using openpyxl during save().
        """
        self._imports.append((source, sheet_name, dest_name))
        self._sheet_order.append(dest_name)

    def _build_openpyxl_workbook_from_xlsx(self) -> _OpenpyxlWorkbook:
        """Build the base openpyxl workbook from xlsxwriter output.

        If there are generated sheets, saves xlsxwriter to bytes and loads
        them with openpyxl. If there are no generated sheets, creates an
        empty openpyxl workbook.
        """
        if self._has_generated_sheets:
            xlsx_bytes = self._xlsx_engine.save(None)
            assert xlsx_bytes is not None
            return _load_workbook(
                BytesIO(xlsx_bytes),
                data_only=False,
                rich_text=True,
            )
        else:
            # No generated sheets, create empty workbook
            workbook = _OpenpyxlWorkbook()
            default_sheet = workbook.active
            if default_sheet is not None:
                workbook.remove(default_sheet)
            return workbook

    def _reorder_sheets(self, workbook: _OpenpyxlWorkbook) -> None:
        """Reorder workbook sheets to match the declaration order.

        This ensures sheets appear in the same order as they were declared
        in the workbook definition, regardless of whether they are generated
        or imported.
        """
        expected_order = self._sheet_order

        # Access internal _sheets list (openpyxl doesn't expose a public reorder API).
        # Reorder in-place rather than replacing the list, to avoid breaking internal
        # workbook invariants that Excel is sensitive to.
        sheets = workbook._sheets  # type: ignore[attr-defined]
        title_to_index = {ws.title: i for i, ws in enumerate(sheets)}

        insert_at = 0
        for title in expected_order:
            idx = title_to_index.get(title)
            if idx is None:
                continue

            if idx != insert_at:
                ws = sheets.pop(idx)
                sheets.insert(insert_at, ws)

                # Update indices for the moved slice.
                start = min(insert_at, idx)
                end = max(insert_at, idx)
                for j in range(start, end + 1):
                    title_to_index[sheets[j].title] = j

            insert_at += 1

        # Ensure the active sheet index is valid after reordering.
        try:
            if sheets:
                workbook.active = 0
        except Exception:
            pass

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        """Finalize and save the workbook.

        This method:
        1. Phase A: Builds the base openpyxl workbook from xlsxwriter output.
        2. Phase B: Applies deferred import operations via OpenpyxlEngine.
        3. Phase C: Reorders sheets to match declaration order.
        4. Phase D: Saves the final workbook.
        """
        # Phase A: Build base workbook from xlsxwriter output
        merged_wb = self._build_openpyxl_workbook_from_xlsx()

        # Phase B: Apply imports via OpenpyxlEngine
        openpyxl_engine = OpenpyxlEngine.from_workbook(merged_wb)

        for source, sheet_name, dest_name in self._imports:
            openpyxl_engine.copy_sheet(source, sheet_name, dest_name)

        # Phase C: Reorder sheets to match declaration order
        self._reorder_sheets(merged_wb)

        # Phase D: Save the final workbook
        return openpyxl_engine.save(target)
