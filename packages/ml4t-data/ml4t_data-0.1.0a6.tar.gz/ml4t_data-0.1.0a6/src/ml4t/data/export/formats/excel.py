"""Excel export functionality."""

from __future__ import annotations

import gzip
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import polars as pl
import structlog

from ml4t.data.export.formats.base import BaseExporter, ExportConfig, ExportResult

logger = structlog.get_logger()

# Check if openpyxl or xlsxwriter is available
try:
    import xlsxwriter

    EXCEL_ENGINE = "xlsxwriter"
except ImportError:
    try:
        import openpyxl

        EXCEL_ENGINE = "openpyxl"
    except ImportError:
        EXCEL_ENGINE = None
        logger.warning("No Excel library found. Install xlsxwriter or openpyxl for Excel export.")


class ExcelExporter(BaseExporter):
    """Export data to Excel format."""

    def __init__(self, config: ExportConfig) -> None:
        """Initialize Excel exporter."""
        super().__init__(config)

        if EXCEL_ENGINE is None:
            raise ImportError(
                "Excel export requires xlsxwriter or openpyxl. Install with: pip install xlsxwriter"
            )

    def export(self, data: pl.DataFrame, symbol: str) -> ExportResult:
        """
        Export data to Excel file.

        Args:
            data: Data to export
            symbol: Symbol being exported

        Returns:
            Export result
        """
        start_time = time.time()

        try:
            # Apply transformations
            df = self._apply_transformations(data)

            # Determine output path
            output_file = self._get_output_path(symbol)
            self._ensure_output_dir()

            # Export to Excel
            self._export_to_excel(df, output_file, symbol)

            # Apply compression if configured
            if self.config.compression:
                output_file = self._compress_file(output_file)

            # Get file size
            file_size = output_file.stat().st_size

            duration = time.time() - start_time

            logger.info(
                f"Exported {symbol} to Excel",
                rows=len(df),
                file=str(output_file),
                size_mb=file_size / 1024 / 1024,
                duration=duration,
            )

            return ExportResult(
                success=True,
                output_path=output_file,
                rows_exported=len(df),
                file_size=file_size,
                duration_seconds=duration,
            )

        except Exception as e:
            error_msg = str(e)
            # Check if it's an import error for Excel libraries
            if "xlsxwriter" in error_msg.lower() and EXCEL_ENGINE == "openpyxl":
                error_msg = f"Excel export requires 'openpyxl' or 'xlsxwriter'.\nCurrent engine: {EXCEL_ENGINE}. Error: {error_msg}"
            logger.error(f"Failed to export {symbol} to Excel: {error_msg}")
            return ExportResult(
                success=False,
                output_path=self.config.output_path,
                rows_exported=0,
                file_size=0,
                duration_seconds=time.time() - start_time,
                error=error_msg,
            )

    def export_batch(self, datasets: dict[str, pl.DataFrame]) -> list[ExportResult]:
        """
        Export multiple datasets to a single Excel file with multiple sheets.

        Args:
            datasets: Dict of symbol -> data mappings

        Returns:
            List with single export result for the workbook
        """
        start_time = time.time()
        total_rows = 0

        try:
            # Determine output path
            output_file = self._get_batch_output_path()
            self._ensure_output_dir()

            if EXCEL_ENGINE == "xlsxwriter":
                self._export_batch_xlsxwriter(datasets, output_file)
            else:
                self._export_batch_polars(datasets, output_file)

            # Apply compression if configured
            if self.config.compression:
                output_file = self._compress_file(output_file)

            # Calculate totals
            for df in datasets.values():
                total_rows += len(df)

            file_size = output_file.stat().st_size
            duration = time.time() - start_time

            logger.info(
                f"Exported {len(datasets)} datasets to Excel",
                total_rows=total_rows,
                file=str(output_file),
                size_mb=file_size / 1024 / 1024,
                duration=duration,
            )

            return [
                ExportResult(
                    success=True,
                    output_path=output_file,
                    rows_exported=total_rows,
                    file_size=file_size,
                    duration_seconds=duration,
                )
            ]

        except Exception as e:
            logger.error(f"Failed to export batch to Excel: {e}")
            return [
                ExportResult(
                    success=False,
                    output_path=self.config.output_path,
                    rows_exported=0,
                    file_size=0,
                    duration_seconds=time.time() - start_time,
                    error=str(e),
                )
            ]

    def _get_output_path(self, symbol: str) -> Path:
        """Get output file path for a symbol."""
        if self.config.output_path.is_dir():
            return self.config.output_path / f"{symbol}.xlsx"
        return self.config.output_path

    def _get_batch_output_path(self) -> Path:
        """Get output file path for batch export."""
        if self.config.output_path.is_dir():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return self.config.output_path / f"export_{timestamp}.xlsx"
        return self.config.output_path

    def _export_to_excel(self, df: pl.DataFrame, output_file: Path, sheet_name: str) -> None:
        """
        Export single DataFrame to Excel.

        Args:
            df: Data to export
            output_file: Output file path
            sheet_name: Name of the Excel sheet
        """
        # Use Polars native Excel writer
        df.write_excel(
            output_file,
            worksheet=sheet_name[:31],  # Excel sheet name limit
            include_header=True,
            autofit=True,
            float_precision=2,
        )

    def _export_batch_xlsxwriter(
        self, datasets: dict[str, pl.DataFrame], output_file: Path
    ) -> None:
        """
        Export multiple DataFrames using xlsxwriter.

        Args:
            datasets: Dict of symbol -> data mappings
            output_file: Output file path
        """
        import xlsxwriter

        # Create workbook
        workbook = xlsxwriter.Workbook(str(output_file))

        # Add metadata sheet if configured
        if self.config.include_metadata:
            self._add_metadata_sheet(workbook, datasets)

        # Add data sheets
        for symbol, data in datasets.items():
            df = self._apply_transformations(data)
            worksheet = workbook.add_worksheet(symbol[:31])

            # Write headers
            for col_idx, col_name in enumerate(df.columns):
                worksheet.write(0, col_idx, col_name)

            # Write data
            for row_idx, row in enumerate(df.iter_rows(), start=1):
                for col_idx, value in enumerate(row):
                    # Handle datetime objects
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    worksheet.write(row_idx, col_idx, value)

        workbook.close()

    def _export_batch_polars(self, datasets: dict[str, pl.DataFrame], output_file: Path) -> None:
        """
        Export multiple DataFrames using openpyxl.

        Args:
            datasets: Dict of symbol -> data mappings
            output_file: Output file path
        """
        # Since polars doesn't have ExcelWriter, use openpyxl directly
        try:
            from openpyxl import Workbook

            wb = Workbook()
            # Remove default sheet
            if len(wb.worksheets) > 0:
                wb.remove(wb.active)

            for symbol, data in datasets.items():
                df = self._apply_transformations(data)
                # Create worksheet
                ws = wb.create_sheet(title=symbol[:31])

                # Write headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)

                # Write data
                for row_idx, row in enumerate(df.iter_rows(), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(output_file)
        except ImportError:
            # Fallback: write only the first dataset if openpyxl is not available
            if datasets:
                first_symbol = next(iter(datasets))
                df = self._apply_transformations(datasets[first_symbol])
                df.write_excel(str(output_file), worksheet=first_symbol[:31])

    def _add_metadata_sheet(self, workbook: Any, datasets: dict[str, pl.DataFrame]) -> None:
        """
        Add metadata sheet to workbook.

        Args:
            workbook: Excel workbook object
            datasets: Dict of symbol -> data mappings
        """
        metadata_sheet = workbook.add_worksheet("_metadata")

        # Write export info
        metadata_sheet.write(0, 0, "Export Metadata")
        metadata_sheet.write(1, 0, "Generated")
        metadata_sheet.write(1, 1, datetime.now().isoformat())
        metadata_sheet.write(2, 0, "Datasets")
        metadata_sheet.write(2, 1, len(datasets))

        # Write dataset info
        metadata_sheet.write(4, 0, "Symbol")
        metadata_sheet.write(4, 1, "Rows")
        metadata_sheet.write(4, 2, "Start Date")
        metadata_sheet.write(4, 3, "End Date")

        for idx, (symbol, data) in enumerate(datasets.items(), start=5):
            metadata_sheet.write(idx, 0, symbol)
            metadata_sheet.write(idx, 1, len(data))

            if "timestamp" in data.columns and len(data) > 0:
                metadata_sheet.write(idx, 2, str(data["timestamp"].min()))
                metadata_sheet.write(idx, 3, str(data["timestamp"].max()))

    def _compress_file(self, file_path: Path) -> Path:
        """
        Compress a file using configured compression method.

        Args:
            file_path: Path to file to compress

        Returns:
            Path to compressed file
        """
        if self.config.compression == "gzip":
            compressed_path = file_path.with_suffix(file_path.suffix + ".gz")

            with open(file_path, "rb") as f_in, gzip.open(compressed_path, "wb") as f_out:
                f_out.write(f_in.read())

            # Remove original file
            file_path.unlink()
            return compressed_path

        return file_path
