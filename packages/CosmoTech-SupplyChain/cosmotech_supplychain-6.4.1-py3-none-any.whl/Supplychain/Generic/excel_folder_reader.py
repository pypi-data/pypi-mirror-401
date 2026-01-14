import glob
import io
import json
import os

from openpyxl import load_workbook

from Supplychain.Generic.folder_io import FolderReader


class ExcelReader(FolderReader):

    def refresh(self):
        filename = (
            self.input_folder
            if self.input_folder.endswith(".xlsx")
            else glob.glob(os.path.join(self.input_folder, "*.xlsx"))[0]
        )
        self.files = dict()
        wb = load_workbook(filename, read_only=True, keep_vba=False,
                           data_only=True, keep_links=False, rich_text=False)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
                continue
            self.files[sheet_name] = list()
            headers = next(sheet.iter_rows(max_row=1, values_only=True))
            n = len(headers) - 1
            while n >= 0 and headers[n] is None:
                n -= 1
            if n < 0:
                continue
            max_col = n + 1
            n = 0
            while n < len(headers) and headers[n] is None:
                n += 1
            min_col = n + 1
            headers = headers[n:max_col]
            indices = tuple(i for i, h in enumerate(headers) if h is not None)
            reduce_row = len(indices) < len(headers)
            if reduce_row:
                headers = tuple(headers[i] for i in indices)

            for row in sheet.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True):
                if reduce_row:
                    row = tuple(row[i] for i in indices)
                if all(v is None for v in row):
                    continue
                new_row = {}
                for key, value in zip(headers, row):
                    if key in self.id_keys:
                        new_row[key] = None if value is None else str(value)
                        continue
                    try:
                        # Try to convert any json row to dict object
                        converted_value = json.load(io.StringIO(value))
                    except (json.decoder.JSONDecodeError, TypeError):
                        converted_value = value
                    if converted_value is not None or self.keep_nones:
                        new_row[key] = converted_value
                if new_row:
                    self.files[sheet_name].append(new_row)

    def __init__(self,
                 input_folder: str = "Input",
                 keep_nones: bool = True):

        FolderReader.__init__(self,
                              input_folder=input_folder,
                              keep_nones=keep_nones)

        self.refresh()
