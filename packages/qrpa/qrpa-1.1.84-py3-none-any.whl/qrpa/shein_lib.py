from .fun_file import read_dict_from_file, write_dict_to_file, read_dict_from_file_ex, write_dict_to_file_ex
from .fun_base import log, send_exception, md5_string, get_safe_value, NetWorkIdleTimeout
from .fun_web import fetch, fetch_shein, fetch_get, full_screen_shot, safe_goto
from .time_utils import TimeUtils
from .wxwork import WxWorkBot

from .shein_sqlite import insert_sales, get_last_week_sales, get_near_week_sales, get_near_month_sales, get_last_month_sales

import math
import time
import json, traceback
from datetime import datetime
from playwright.sync_api import Page

from .shein_daily_report_model import SheinStoreSalesDetailManager

class SheinLib:

    def __init__(self, config, bridge, web_page: Page, store_username, store_name):
        self.config = config
        self.bridge = bridge
        self.store_username = store_username
        self.store_name = store_name
        self.web_page = web_page
        self.dt = None
        self.dt_goods = None
        self.DictQueryTime = {}

        self.deal_auth()
        self.get_user()

    def close_modal(self):
        try:
            self.web_page.evaluate("""
                () => {
                    if (window.intervalId) {
                        clearInterval(window.intervalId);
                    }
                    window.intervalId = setInterval(() => {
                        document.querySelectorAll('.so-modal-show').forEach((el) => {
                            console.log('准备移除弹窗', el);
                            el.remove();
                        });
                    }, 3000);
                }
            """)
        except Exception as e:
            self.web_page.evaluate("""
                () => {
                    if (window.intervalId) {
                        clearInterval(window.intervalId);
                    }
                }
            """)
            log(e)
        pass

    def close_modal_once(self):
        try:
            self.web_page.evaluate("""
                () => {
                        document.querySelectorAll('.so-modal-show').forEach((el) => {
                            console.log('准备移除弹窗', el);
                            el.remove();
                        });
                }
            """)
        except Exception as e:
            log(e)
        pass

    # 处理鉴权
    def deal_auth(self):
        web_page = self.web_page

        # 等待页面稳定并处理导航
        for attempt in range(3):
            try:
                current_url = web_page.url
                log(f"尝试获取页面信息 - URL: {current_url}", self.store_username, self.store_name)

                # 检查是否在认证页面，如果是则直接跳转到目标页面
                if '/auth/SSLS' in current_url:
                    log("检测到SSLS认证页面，直接跳转到首页", self.store_username, self.store_name)
                    web_page.goto('https://sso.geiwohuo.com/#/home', wait_until='domcontentloaded', timeout=15000)
                    web_page.wait_for_timeout(3000)
                    current_url = web_page.url
                    log(f"跳转后URL: {current_url}", self.store_username, self.store_name)

                # 等待导航完成
                web_page.wait_for_load_state("domcontentloaded", timeout=6000)

                final_url = web_page.url
                final_title = web_page.title()
                log(f"页面稳定 - URL: {final_url}, 标题: {final_title}", self.store_username, self.store_name)
                break

            except Exception as e:
                log(f"第{attempt + 1}次等待页面稳定失败: {e}", self.store_username, self.store_name)
                if "crashed" in str(e) or "Target" in str(e):
                    log("页面稳定检查时崩溃，直接继续", self.store_username, self.store_name)
                    break
                elif "destroyed" in str(e) or "navigation" in str(e):
                    log("检测到导航中断，等待导航完成", self.store_username, self.store_name)
                    web_page.wait_for_timeout(4000)
                    continue
                elif attempt == 2:
                    log("页面稳定等待最终失败，继续执行", self.store_username, self.store_name)
                    break
                web_page.wait_for_timeout(2000)

        web_page.wait_for_timeout(2000)

        # 定义最大重试次数
        MAX_RETRIES = 5
        retries = 0
        wait_count = 0
        is_send = False

        while retries < MAX_RETRIES:
            try:
                retries += 1

                while not web_page.locator('//div[contains(text(),"商家后台")]').nth(1).is_visible():
                    try:
                        current_url = web_page.url
                        current_title = web_page.title()
                        log(f"循环检查 - URL: {current_url}, 标题: {current_title}", self.store_username, self.store_name)

                        # 如果在认证页面且出现问题，直接跳转
                        if '/auth/SSLS' in current_url:
                            log("在主循环中检测到SSLS认证页面，跳转到首页", self.store_username, self.store_name)
                            web_page.goto('https://sso.geiwohuo.com/#/home', wait_until='domcontentloaded', timeout=15000)
                            web_page.wait_for_timeout(3000)
                            continue

                    except Exception as status_error:
                        log(f"获取页面状态失败: {status_error}", self.store_username, self.store_name)
                        if "crashed" in str(status_error):
                            break

                    if web_page.locator('xpath=//div[text()="扫码登录"]').is_visible():
                        log('检查到扫码登录,切换至账号登录', self.store_username, self.store_name)
                        web_page.locator('xpath=//*[@id="container"]/div[2]/div[4]/img').click()

                    if web_page.locator('xpath=//div[@id="container" and @alita-name="gmpsso"]//button[@type="button" and @id]').nth(0).is_visible():
                        if 'https://sso.geiwohuo.com/#/home' not in web_page.url:
                            log("鉴权确定按钮可见 点击'确定'按钮", web_page.title(), web_page.url, self.store_username, self.store_name)
                            web_page.locator('xpath=//div[@id="container" and @alita-name="gmpsso"]//button[@type="button" and @id]').nth(0).click()
                            web_page.wait_for_timeout(5000)

                    while web_page.locator('//div[text()="手机号码验证"]').is_visible():
                        log(f'等待输入验证码: {wait_count}', self.store_username, self.store_name)
                        if not is_send:
                            is_send = True
                            img_path = full_screen_shot(web_page, self.config)
                            WxWorkBot(self.config.wxwork_bot_exception).send_img(img_path)
                            WxWorkBot(self.config.wxwork_bot_exception).send_text(f'{self.store_username},{self.store_name} 需要登录验证码')
                        time.sleep(5)
                        wait_count += 1

                    if web_page.locator('//div[contains(text(),"同意签署协议")]').count() > 0:
                        while web_page.locator('//div[contains(text(),"同意签署协议")]').count() == 0:
                            log('等待协议内容出现')
                            web_page.wait_for_timeout(1000)

                    if web_page.locator('//div[contains(text(),"同意签署协议")]').count() > 0:
                        log('检测到同意签署协议')
                        web_page.wait_for_timeout(1000)
                        log('点击同意复选框')
                        web_page.locator('//i[@class="so-checkinput-indicator so-checkinput-checkbox"]').click()
                        web_page.wait_for_timeout(1000)
                        log('点击同意按钮')
                        web_page.locator('//button[span[text()="同意"]]').click()

                    # //button[span[text()="登录"]]
                    if web_page.locator('//button[span[text()="登录"]]').is_visible() or web_page.locator('//input[@name="username"]').is_visible():
                        log("用户名输入框可见 等待5秒点击'登录'按钮", self.store_username, self.store_name)
                        web_page.wait_for_timeout(5000)
                        log('点击"登录"', self.store_username, self.store_name)
                        web_page.locator('//button[contains(@class,"login_btn")]').click()

                        log('再延时5秒', self.store_username, self.store_name)
                        web_page.wait_for_timeout(5000)

                    if web_page.locator('//span[contains(text(),"商品管理")]').nth(1).is_visible():
                        log('商品管理菜单可见 退出鉴权处理', self.store_username, self.store_name)
                        return

                    log('商家后台不可见', web_page.title(), web_page.url, self.store_username, self.store_name)
                    if 'https://sso.geiwohuo.com/#/home' in web_page.url:
                        web_page.wait_for_timeout(5000)
                        web_page.reload()

                    # while r'=/CN' in web_page.url:
                    #     safe_goto(web_page, 'https://sso.geiwohuo.com/#/home?q=0')
                    #
                    #     web_page.wait_for_timeout(5000)
                    #     if web_page.locator('//input[@name="username"]').is_visible():
                    #         log("用户名输入框可见 等待5秒点击'登录'按钮", self.store_username, self.store_name)
                    #         web_page.wait_for_timeout(5000)
                    #         log('点击"登录"', self.store_username, self.store_name)
                    #         web_page.locator('//button[contains(@class,"login_btn")]').click()
                    #
                    #         log('再延时5秒', self.store_username, self.store_name)
                    #         web_page.wait_for_timeout(5000)

                    web_page.wait_for_timeout(3000)

                    if 'https://sso.geiwohuo.com/#/home' in web_page.url:
                        if 'SHEIN全球商家中心' in web_page.title() or '后台首页' in web_page.title() or '商家后台' in web_page.title():
                            log(web_page.title(), '中断循环', self.store_username, self.store_name)
                            web_page.wait_for_timeout(5000)
                            break

                    if 'mrs.biz.sheincorp.cn' in web_page.url and '商家后台' in web_page.title():
                        try:
                            web_page.goto('https://sso.geiwohuo.com/#/home?q=1', wait_until='domcontentloaded', timeout=10000)
                            web_page.wait_for_timeout(3000)
                        except Exception as nav_error:
                            log(f"导航失败，尝试重新加载: {nav_error}", self.store_username, self.store_name)
                            web_page.reload(wait_until='domcontentloaded', timeout=10000)
                            web_page.wait_for_timeout(5000)

                    if web_page.locator('//h1[contains(text(),"鉴权")]').is_visible():
                        log('检测到鉴权 刷新页面', self.store_username, self.store_name)
                        web_page.reload()
                        web_page.wait_for_timeout(5000)
                        web_page.reload()
                        web_page.wait_for_timeout(5000)

                    if web_page.title() == 'SHEIN':
                        try:
                            web_page.goto('https://sso.geiwohuo.com/#/home?q=2', wait_until='domcontentloaded', timeout=10000)
                            web_page.wait_for_timeout(3000)
                        except Exception as nav_error:
                            log(f"导航失败，尝试重新加载: {nav_error}", self.store_username, self.store_name)
                            web_page.reload(wait_until='domcontentloaded', timeout=10000)
                            web_page.wait_for_timeout(5000)

                break
            except Exception as e:
                log(f"错误发生: {e}, 重试中...({self.store_username}, {self.store_name})")
                log(traceback.format_exc())

                # 收集崩溃时的详细信息
                try:
                    crash_url = web_page.url
                    crash_title = web_page.title()
                    log(f"崩溃时页面信息 - URL: {crash_url}, 标题: {crash_title}", self.store_username, self.store_name)

                    # 尝试截图保存崩溃现场
                    try:
                        screenshot_path = f"crash_screenshot_{self.store_username}_{int(time.time())}.png"
                        web_page.screenshot(path=screenshot_path)
                        log(f"已保存崩溃截图: {screenshot_path}", self.store_username, self.store_name)
                    except:
                        log("无法截取崩溃时的页面截图", self.store_username, self.store_name)

                except:
                    log("无法获取崩溃时的页面信息", self.store_username, self.store_name)

                # 检查特定类型的错误
                if any(keyword in str(e).lower() for keyword in ['memory', 'out of memory', 'oom']):
                    log("检测到内存相关崩溃", self.store_username, self.store_name)

                if "destroyed" in str(e) or "navigation" in str(e):
                    log("检测到导航中断，等待页面稳定后重试", self.store_username, self.store_name)
                    web_page.wait_for_timeout(5000)
                    continue

                if 'crashed' in str(e) or 'Target' in str(e):
                    log("检测到页面或目标崩溃，直接退出当前循环", self.store_username, self.store_name)
                    raise e
                retries += 1
                if retries >= MAX_RETRIES:
                    log(f"达到最大重试次数，停止尝试({self.store_username}, {self.store_name})")
                    break
                time.sleep(2)  # 错误时等待2秒后重试

        log('鉴权处理结束')
        # web_page.wait_for_load_state("load")
        # web_page.wait_for_load_state("networkidle")
        web_page.wait_for_timeout(3000)

    # 获取用户信息
    # {
    #     "code": "0",
    #     "msg" : "OK",
    #     "info": {
    #         "userName"        : "GS0365305", //这个是登录账号
    #         "userId"          : 2031373,
    #         "supplierId"      : 5230023,   //这个是商家ID
    #         "ulpName"         : null,
    #         "ulpEnName"       : null,
    #         "ulpEmplid"       : null,
    #         "isUlpLogin"      : 2,
    #         "timezone"        : "Asia/Shanghai",
    #         "timezoneName"    : "中国 北京 UTC+8",
    #         "switchNewMenu"   : 1,
    #         "supplierUserName": "GS0365305",
    #         "ssoTopNav"       : 1,
    #         "ssoHost"         : "https://sso.geiwohuo.com",
    #         "categoryId"      : 118275,
    #         "categoryOutId"   : 118275,
    #         "supplierSource"  : 10,
    #         "externalId"      : 5230023,
    #         "storeCode"       : 1785850489,     //这个是店铺ID
    #         "utcTimezone"     : "UTC+8",
    #         "merchantCode"    : "E169AMMFFW",   //这个是全球唯一编码
    #         "lv1CategoryId"   : 89887,
    #         "mainUserName"    : "GS0365305",
    #         "mainUserId"      : 2031373,
    #         "areaTimezone"    : "Asia/Shanghai",
    #         "lv1CategoryName" : "全托管-备CN-品类",
    #         "lv2CategoryName" : "全托管-备CN-品类-内睡服装",
    #         "schatId"         : 0
    #     },
    #     "bbl" : null
    # }
    def get_user(self, uuid=None):
        log(f'获取用户信息:{self.store_username} {self.store_name}')

        # 生成 uuid 参数，如果没有提供则使用时间戳
        if uuid is None:
            import time
            uuid = str(int(time.time() * 1000))

        url = f"https://sso.geiwohuo.com/sso-prefix/auth/getUser?uuid={uuid}"

        # 设置请求头，根据 Chrome 请求
        headers = {
            "gmpsso-language": "CN",
            "origin-url"     : "https://sso.geiwohuo.com/#/home/",
            "x-sso-scene"    : "gmpsso"
        }

        # 特定于此请求的配置
        fetch_config = {
            "credentials"   : "include",
            "referrer"      : "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info', {})
        cache_file = f'{self.config.auto_dir}/shein_user.json'
        info['store_username'] = self.store_username
        info['store_name'] = self.store_name
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])
        log(info)
        self.user_info = info
        return info

    # 获取供货商信息
    def get_supplier_data(self):
        self.web_page.goto('https://sso.geiwohuo.com/#/mws/seller/new-account-overview')
        self.web_page.wait_for_load_state('load')
        cache_file = f'{self.config.auto_dir}/shein/dict/supplier_data.json'
        info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 10)
        if info and len(info) > 0:
            return info

        log(f'正在获取 {self.store_name} 供货商信息')
        url = "https://sso.geiwohuo.com/mgs-api-prefix/supplierGrowth/querySupplierCommonData"
        payload = {}
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')

        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        return info

    # 获取钱包余额信息
    def get_wallet_balance(self, supplier_id):
        self.web_page.goto('https://sso.geiwohuo.com/#/mws/seller/new-account-overview')
        self.web_page.wait_for_load_state('load')

        cache_file = f'{self.config.auto_dir}/shein/cache/wallet_balance_{TimeUtils.today_date()}.json'
        info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 12)
        if info and len(info) > 0:
            return info

        log(f'正在获取 {self.store_name} 钱包余额')
        url = "https://sso.geiwohuo.com/mws/mwms/sso/balance/query"
        payload = {
            "reqSystemCode": "mws-front",
            "supplierId"   : supplier_id
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')

        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        return info

    def get_withdraw_list(self, supplier_id, year=0):
        self.web_page.goto('https://sso.geiwohuo.com/#/mws/seller/new-account-overview')
        self.web_page.wait_for_load_state("load")

        if year == 0:
            first_day, last_day = TimeUtils.get_last_month_range_time()
        else:
            first_day, last_day = TimeUtils.get_year_range_time(year)

        page_num = 1
        page_size = 200

        url = f"https://sso.geiwohuo.com/mws/mwms/sso/withdraw/transferRecordList"
        payload = {
            "reqSystemCode"  : "mws-front",
            "supplierId"     : supplier_id,
            "pageNum"        : page_num,
            "pageSize"       : page_size,
            "createTimeStart": first_day,
            "createTimeEnd"  : last_day,
            # "withdrawStatusList": [30]
        }
        log(payload)
        response_text = fetch(self.web_page, url, payload)
        log(response_text)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        # 检查是否存在 info 字段
        if 'info' not in response_text or response_text['info'] is None:
            log('⚠️ API 响应中没有 info 字段，返回空列表')
            return []

        withdraw_list = response_text['info']['list']
        total = response_text['info']['count']
        totalPage = math.ceil(total / page_size)

        cache_file = f'{self.config.auto_dir}/shein/cache/withdraw_list_{first_day}_{last_day}.json'
        withdraw_list_cache = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 12)
        if len(withdraw_list_cache) == int(total):
            log('返回缓存数据: ', len(withdraw_list_cache), total)
            return withdraw_list_cache

        for page in range(2, totalPage + 1):
            log(f'获取提现列表 第{page}/{totalPage}页')
            page_num = page
            payload['pageNum'] = page_num
            response_text = fetch(self.web_page, url, payload)
            withdraw_list += response_text['info']['list']
            time.sleep(0.1)

        write_dict_to_file_ex(cache_file, {self.store_username: withdraw_list}, [self.store_username])

        return withdraw_list

    # 获取提现成功记录列表（仅状态为30的记录）
    # 返回的 depositFlag 字段：1 表示需要验证密码，0 表示已验证
    def get_withdraw_success_list(self, supplier_id, start_time=None, end_time=None):
        """
        获取提现成功记录列表
        :param supplier_id: 供应商ID
        :param start_time: 开始时间戳（毫秒），默认为上月第一天
        :param end_time: 结束时间戳（毫秒），默认为上月最后一天
        :return: dict 包含 list(提现成功记录), count(总数), depositFlag(是否需要验证密码: 1需要 0不需要)
        """
        self.web_page.goto('https://sso.geiwohuo.com/#/mws/seller/new-account-overview')
        self.web_page.wait_for_load_state("load")

        # 默认时间范围：上个月
        if start_time is None or end_time is None:
            default_start, default_end = TimeUtils.get_last_month_range()
            from datetime import datetime
            if start_time is None:
                start_time = int(datetime.strptime(default_start, '%Y-%m-%d').timestamp() * 1000)
            if end_time is None:
                end_dt = datetime.strptime(default_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                end_time = int(end_dt.timestamp() * 1000)

        # 将时间戳转换为日期字符串用于缓存文件名
        from datetime import datetime
        start_date_str = datetime.fromtimestamp(start_time / 1000).strftime('%Y-%m-%d')
        end_date_str = datetime.fromtimestamp(end_time / 1000).strftime('%Y-%m-%d')

        # 缓存文件路径：按店铺和时间范围缓存文件
        import os
        cache_dir = f'{self.config.auto_dir}/shein/withdraw_success'
        os.makedirs(cache_dir, exist_ok=True)
        cache_file = f'{cache_dir}/withdraw_success_{self.store_username}_{start_date_str}_{end_date_str}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if cached_data and len(cached_data) > 0:
            log(f'返回缓存的提现成功记录: {self.store_name}')
            return cached_data

        page_num = 1
        page_size = 100

        url = "https://sso.geiwohuo.com/mws/mwms/sso/withdraw/transferRecordList"
        payload = {
            "reqSystemCode"     : "mws-front",
            "supplierId"        : supplier_id,
            "pageNum"           : page_num,
            "pageSize"          : page_size,
            "createTimeStart"   : start_time,
            "createTimeEnd"     : end_time,
            # "withdrawStatusList": [30]  # 30 = 提现成功
        }
        log(f'获取提现成功记录: {self.store_name}', payload)
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        if 'info' not in response_text or response_text['info'] is None:
            log('⚠️ API 响应中没有 info 字段，返回空结果')
            return {'list': [], 'count': 0, 'depositFlag': 1}

        info = response_text['info']
        withdraw_list = info.get('list', [])
        total = info.get('count', 0)
        deposit_flag = info.get('depositFlag', 1)  # 1需要验证密码 0不需要
        totalPage = math.ceil(total / page_size)

        # 分页获取所有数据
        for page in range(2, totalPage + 1):
            log(f'获取提现成功记录 第{page}/{totalPage}页')
            payload['pageNum'] = page
            response_text = fetch(self.web_page, url, payload)
            if response_text.get('info') and response_text['info'].get('list'):
                withdraw_list += response_text['info']['list']
            time.sleep(0.1)

        result = {
            'list'       : withdraw_list,
            'count'      : total,
            'depositFlag': deposit_flag,
            # 额外保存店铺信息，供后续导入数据库使用
            'store_username': self.store_username,
            'store_name': self.store_name,
            'supplier_id': supplier_id,
        }

        # 缓存结果
        write_dict_to_file(cache_file, result)

        return result

    # 获取钱包余额详情（包含可提现、不可提现、提现中金额等详细信息）
    def get_wallet_balance_detail(self, supplier_id):
        """
        获取钱包余额详细信息
        :param supplier_id: 供应商ID
        :return: dict 包含详细的钱包余额信息
        """
        self.web_page.goto('https://sso.geiwohuo.com/#/mws/seller/new-account-overview')
        self.web_page.wait_for_load_state('load')

        # 缓存文件路径：按店铺缓存文件（不使用缓存，每次都重新拉取）
        import os
        cache_dir = f'{self.config.auto_dir}/shein/wallet_balance'
        os.makedirs(cache_dir, exist_ok=True)
        cache_file = f'{cache_dir}/wallet_balance_detail_{self.store_username}_{TimeUtils.today_date()}.json'

        log(f'正在获取 {self.store_name} 钱包余额详情')
        url = "https://sso.geiwohuo.com/mws/mwms/sso/balance/query"
        payload = {
            "reqSystemCode": "mws-front",
            "supplierId"   : supplier_id
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 解析详细信息
        result = {
            'isFpAccount' : info.get('isFpAccount', False),
            'depositFlag' : info.get('depositFlag', 1),  # 1需要验证密码 0不需要
            'balanceList' : [],  # 余额列表
            'depositList' : []   # 保证金列表
        }

        # 解析余额详情
        for detail in info.get('detailResponseList', []):
            balance_item = {
                'currency'            : detail.get('currency', ''),
                'withdrawableAmount'  : detail.get('withdrawableAmount', '0'),      # 可提现金额
                'noWithdrawableAmount': detail.get('noWithdrawableAmount', '0'),    # 不可提现金额
                'withdrawingAmount'   : detail.get('withdrawingAmount', '0'),       # 提现中金额
                'lastUpdateTime'      : detail.get('lastUpdateTime'),
                'autoWithdrawState'   : detail.get('autoWithdrawState', 0),
                'canWithdraw'         : detail.get('canWithdraw', False),
                'noWithdrawReasons'   : [r.get('overlayReasonCode', '') for r in detail.get('noWithdrawReasonList', [])]
            }
            result['balanceList'].append(balance_item)

        # 解析保证金详情
        for deposit in info.get('depositDetailResponseList', []):
            deposit_item = {
                'currency'           : deposit.get('currency', ''),
                'depositAmountPaid'  : deposit.get('depositAmountPaid', '0'),    # 已缴保证金
                'depositAmountUnPaid': deposit.get('depositAmountUnPaid', '0'),  # 未缴保证金
                'lastUpdateTime'     : deposit.get('lastUpdateTime')
            }
            result['depositList'].append(deposit_item)

        # 额外保存店铺信息，供后续导入数据库使用
        result['store_username'] = self.store_username
        result['store_name'] = self.store_name
        result['supplier_id'] = supplier_id

        # 缓存结果
        write_dict_to_file(cache_file, result)

        return result

    # 获取质检报告pdf地址
    def get_qc_report_url(self, deliverCode, purchaseCode):
        log(f'获取质检报告:{deliverCode} {purchaseCode}')
        url = f"https://sso.geiwohuo.com/pfmp/returnPlan/queryQcReport"
        payload = {
            "deliverCode" : deliverCode,
            "purchaseCode": purchaseCode
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        qc_report_url = (response_text.get('info', {}).get('data') or [{'qcReportUrl': '质检报告生成中,请稍后查看'}])[0].get('qcReportUrl')
        log(qc_report_url)
        return qc_report_url

    # 获取稽查报表
    def get_inspect_report_url(self, returnOrderId):
        log(f'获取稽查报告:{returnOrderId}')
        url = f"https://sso.geiwohuo.com/pfmp/returnOrder/queryInspectReport"
        payload = {
            "returnOrderId": returnOrderId,
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        log(response_text)
        report_url = response_text.get('info', {}).get('reportUrl')
        return report_url

    def return_order_goods_detail(self, returnOrderId):
        log(f'获取退货商品明细: {returnOrderId}')
        url = f"https://sso.geiwohuo.com/pfmp/returnOrder/getReturnOrderGoodsDetail"
        payload = {
            "returnOrderId": returnOrderId,
            "page"         : 1,
            "perPage"      : 50
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']

        for item in list_item:
            # 遍历每个商品的详情列表
            for detail in item.get('details', []):
                # 在这里添加新字段
                supplier_sku = detail.get('supplierSku')
                erp_supplier_name = self.bridge.get_sku_supplier(supplier_sku, self.config.erp_source)
                log(self.config.erp_source, supplier_sku, erp_supplier_name)
                if erp_supplier_name != '-':
                    detail['erp_supplier_name'] = erp_supplier_name
                erp_cost_price = self.bridge.get_sku_cost(supplier_sku, self.config.erp_source)
                log(self.config.erp_source, supplier_sku, erp_cost_price)
                if erp_cost_price != '-':
                    detail['erp_cost_price'] = erp_cost_price

        log(list_item)
        cache_file = f'{self.config.auto_dir}/shein/cache/shein_return_order_goods_detail_{returnOrderId}.json'
        write_dict_to_file(cache_file, list_item)
        return list_item

    def get_return_order_box_detail(self, returnOrderId):
        log(f'获取退货包裹详情: {returnOrderId}')
        url = f"https://sso.geiwohuo.com/pfmp/returnOrder/getReturnOrderBoxDetail"
        payload = {
            "returnOrderId": returnOrderId,
            "page"         : 1,
            "perPage"      : 50
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']

        for item in list_item:
            # 遍历每个快递单的包裹列表
            for box in item.get('boxList', []):
                # 遍历每个包裹中的商品列表
                for good in box.get('goods', []):
                    # 遍历每个商品的详情列表（包含platformSku的层级）
                    for detail in good.get('details', []):
                        # 在这里添加新字段
                        # 示例：添加一个"status"字段，值为"processed"
                        supplier_sku = detail.get('supplierSku')
                        erp_supplier_name = self.bridge.get_sku_supplier(supplier_sku, self.config.erp_source)
                        log(self.config.erp_source, supplier_sku, erp_supplier_name)
                        if erp_supplier_name != '-':
                            detail['erp_supplier_name'] = erp_supplier_name
                        erp_cost_price = self.bridge.get_sku_cost(supplier_sku, self.config.erp_source)
                        log(self.config.erp_source, supplier_sku, erp_cost_price)
                        if erp_cost_price != '-':
                            detail['erp_cost_price'] = erp_cost_price

        log(list_item)
        cache_file = f'{self.config.auto_dir}/shein/cache/shein_return_order_box_detail_{returnOrderId}.json'
        write_dict_to_file(cache_file, list_item)
        return list_item

    def get_return_order_list(self, start_date, end_date, only_yesterday=1):
        log(f'获取退货列表: {self.store_username} {self.store_name} {start_date} {end_date}')

        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        url = f"https://sso.geiwohuo.com/pfmp/returnOrder/page"
        payload = {
            "returnOrderType": 1,  # 只查询退货
            "addTimeStart"   : f"{start_date} 00:00:00",
            "addTimeEnd"     : f"{end_date} 23:59:59",
            "page"           : page_num,
            "perPage"        : page_size
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取退供列表 第{page}/{totalPage}页 共{total}条记录')
            payload['page'] = page
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            list_item += spu_list_new
            time.sleep(0.1)

        all_list_item = []
        today_list_item = []
        # 过滤 退货出库时间 是昨天的
        for item in list_item:
            returnOrderId = item['id']
            item['store_username'] = self.store_username
            item['store_name'] = self.store_name
            item['store_manager'] = self.config.shein_store_manager.get(str(self.store_username).lower())

            item['qc_report_url'] = ''
            if int(item['returnScrapType']) in [1, 9]:
                purchaseCode = item['sellerOrderNo']
                delivery_code = item['sellerDeliveryNo']
                item['qc_report_url'] = self.get_qc_report_url(delivery_code, purchaseCode)

            item['report_url'] = ''
            if int(item['returnScrapType']) == 2:
                item['report_url'] = self.get_inspect_report_url(returnOrderId)

            item['return_goods_detail'] = self.return_order_goods_detail(returnOrderId)
            item['return_box_detail'] = []
            all_list_item.append(item)
            is_valid_yesterday = TimeUtils.is_yesterday(item['completeTime'], None) if item.get('completeTime') else False
            if is_valid_yesterday:
                today_list_item.append(item)

            # has_valid_package = item.get('hasPackage') == 1
            # if has_valid_package:
            #     return_box_detail = self.get_return_order_box_detail(returnOrderId)
            #     if len(return_box_detail) > 0:
            #         item['return_box_detail'] = return_box_detail
            #
            #         is_valid_yesterday = TimeUtils.is_yesterday(item['completeTime'], None) if item.get('completeTime') else False
            #         if is_valid_yesterday:
            #             today_list_item.append(item)

        cache_file = f'{self.config.auto_dir}/shein/cache/shein_return_order_list_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: today_list_item}, [self.store_username])

        cache_file = f'{self.config.auto_dir}/shein/cache/shein_return_order_list_{start_date}_{end_date}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: all_list_item}, [self.store_username])

        return list_item

    # 获取希音退供明细 和 台账明细一个接口
    def get_back_list(self, source='mb'):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/return_detail_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mils/changeDetail/page"
        payload = {
            "displayChangeTypeList": ["10"],
            "addTimeStart"         : f"{first_day} 00:00:00",
            "addTimeEnd"           : f"{last_day} 23:59:59",
            "pageNumber"           : page_num,
            "pageSize"             : page_size,
            "changeTypeIndex"      : "2"
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取台账明细列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        # cost_price =
        for item in list_item:
            supplierSku = item['supplierSku']
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, source)

        write_dict_to_file(cache_file, list_item)

        return list_item

    # 不结算列表
    def get_no_settlement_list(self, source='mb'):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/no_settlement_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mils/changeDetail/page"
        payload = {
            "addTimeStart"         : f"{first_day} 00:00:00",
            "addTimeEnd"           : f"{last_day} 23:59:59",
            "pageNumber"           : page_num,
            "pageSize"             : page_size,
            "changeTypeIndex"      : "2",
            "settleTypeList"       : ["1"],  # 不结算
            "displayChangeTypeList": ["6", "7", "9", "10", "11", "12", "13", "16", "18", "19"]  # 出库
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取台账明细列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        # cost_price =
        for item in list_item:
            supplierSku = item['supplierSku']
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, source)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_ledger_record(self, first_day, last_day):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        cache_file = f'{self.config.auto_dir}/shein/ledger/ledger_record_{self.store_username}_{first_day}_{last_day}.json'
        list_item_cache = read_dict_from_file(cache_file)

        url = f"https://sso.geiwohuo.com/mils/changeDetail/page"
        payload = {
            "displayChangeTypeList": ["6", "7", "9", "10", "11", "12", "13", "16", "18", "19", "21"],  # 出库
            "addTimeStart"         : f"{first_day} 00:00:00",
            "addTimeEnd"           : f"{last_day} 23:59:59",
            "pageNumber"           : page_num,
            "pageSize"             : page_size,
            "changeTypeIndex"      : "2"
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        if len(list_item_cache) == int(total):
            return list_item_cache

        for page in range(2, totalPage + 1):
            log(f'获取台账明细列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        for item in list_item:
            supplierSku = item['supplierSku']
            item['store_username'] = self.store_username
            item['store_name'] = self.store_name
            item['store_manager'] = self.config.shein_store_manager.get(str(self.store_username).lower())
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, self.config.erp_source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, self.config.erp_source)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_ledger_month_report(self, start_date, end_date):
        """
        获取台账月报数据

        Args:
            start_date: 开始日期（格式：YYYY-MM-DD）
            end_date: 结束日期（格式：YYYY-MM-DD）

        Returns:
            list: 月报数据列表
        """
        page_num = 1
        page_size = 50  # 月报数据通常不多，50条足够

        cache_file = f'{self.config.auto_dir}/shein/ledger_month/ledger_month_{self.store_username}_{start_date}_{end_date}.json'
        list_item_cache = read_dict_from_file(cache_file)

        url = "https://sso.geiwohuo.com/mils/report/month/list"
        payload = {
            "reportDateStart": start_date,
            "reportDateEnd"  : end_date,
            "pageNumber"     : page_num,
            "pageSize"       : page_size
        }

        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        # 如果缓存完整，直接返回
        if len(list_item_cache) == int(total):
            log(f'返回缓存数据: {len(list_item_cache)} 条')
            return list_item_cache

        # 分页获取剩余数据
        for page in range(2, totalPage + 1):
            log(f'获取台账月报 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch(self.web_page, url, payload)
            page_list = response_text['info']['data']['list']
            list_item += page_list
            time.sleep(0.1)

        # 添加店铺信息
        for item in list_item:
            item['store_username'] = self.store_username
            item['store_name'] = self.store_name
            item['store_manager'] = self.config.shein_store_manager.get(str(self.store_username).lower())

        write_dict_to_file(cache_file, list_item)

        log(f'获取台账月报成功，共 {len(list_item)} 条记录')
        return list_item

    def get_ledger_list(self, source='mb'):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/sales_detail_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mils/changeDetail/page"
        payload = {
            "displayChangeTypeList": ["6", "7", "9", "10", "11", "12", "13", "16", "18", "19"],  # 出库
            "addTimeStart"         : f"{first_day} 00:00:00",
            "addTimeEnd"           : f"{last_day} 23:59:59",
            "pageNumber"           : page_num,
            "pageSize"             : page_size,
            "changeTypeIndex"      : "2"
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取台账明细列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        # cost_price =
        for item in list_item:
            supplierSku = item['supplierSku']
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, source)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_shein_stock_list(self, source='mb'):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/stock_detail_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mils/report/month/detail/list"
        payload = {
            "reportDateStart": first_day,
            "reportDateEnd"  : last_day,
            "pageNumber"     : page_num,
            "pageSize"       : page_size,
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']['list']
        total = response_text['info']['data']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取库存结余明细列表 第{page}/{totalPage}页')
            page_num = page
            payload = {
                "reportDateStart": first_day,
                "reportDateEnd"  : last_day,
                "pageNumber"     : page_num,
                "pageSize"       : page_size,
            }
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        for item in list_item:
            supplierSku = item['supplierSku']
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, source)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def refresh_bridge_data_for_list(self, data_list, source='mb', sku_field='supplierSku'):
        """
        刷新列表中的bridge数据（成本价和SKU图片）
        
        Args:
            data_list: 需要刷新的数据列表
            source: ERP数据源，默认为'mb'
            sku_field: SKU字段名，默认为'supplierSku'
            
        Returns:
            刷新后的数据列表
        """
        log(f'开始刷新Bridge数据，共 {len(data_list)} 条记录', self.store_username, self.store_name)

        for index, item in enumerate(data_list):
            supplier_sku = item.get(sku_field)
            if supplier_sku:
                item['cost_price'] = self.bridge.get_sku_cost(supplier_sku, source)
                item['sku_img'] = self.bridge.get_sku_img(supplier_sku, source)

            # 每100条记录输出一次进度
            if (index + 1) % 100 == 0:
                log(f'刷新进度: {index + 1}/{len(data_list)}', self.store_username, self.store_name)

        log(f'Bridge数据刷新完成', self.store_username, self.store_name)
        return data_list

    def get_vssv_order_list(self):
        """
        获取VSSV订单列表

        Args:
            web_page: 页面对象
            store_username: 店铺账号
            store_name: 店铺名称

        Returns:
            list: 订单列表
        """
        page_num = 1
        page_size = 200
        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/vssv_order/vssv_order_list_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/vssv/order/page"
        payload = {
            "deductionStatus": "2",
            "beginTime"      : f"{first_day} 00:00:00",
            "endTime"        : f"{last_day} 23:59:59",
            "pageNumber"     : page_num,
            "pageSize"       : page_size
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
            raise
        list_item = response_text['info']['list']
        total = response_text['info']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取VSSV订单列表 第{page}/{totalPage}页')
            page_num = page
            payload = {
                "deductionStatus": "2",
                "beginTime"      : f"{first_day} 00:00:00",
                "endTime"        : f"{last_day} 23:59:59",
                "pageNumber"     : page_num,
                "pageSize"       : page_size
            }
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['list']
            list_item += spu_list_new
            time.sleep(0.1)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_replenish_list(self):
        page_num = 1
        page_size = 50
        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/replenish_list_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/gsfs/finance/selfReplenish/list"
        payload = {
            "page"        : page_num,
            "perPage"     : page_size,
            "tabType"     : 2,
            "addTimeStart": f"{first_day} 00:00:00",
            "addTimeEnd"  : f"{last_day} 23:59:59"
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取不扣款列表 第{page}/{totalPage}页')
            page_num = page
            payload = {
                "page"        : page_num,
                "perPage"     : page_size,
                "tabType"     : 2,
                "addTimeStart": f"{first_day} 00:00:00",
                "addTimeEnd"  : f"{last_day} 23:59:59"
            }
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            list_item += spu_list_new
            time.sleep(0.1)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_return_list(self):
        page_num = 1
        page_size = 200
        first_day, last_day = TimeUtils.get_last_month_range()

        cache_file = f'{self.config.auto_dir}/shein/cache/return_list_{self.store_username}_{first_day}_{last_day}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 24 * 20)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/pfmp/returnOrder/page"
        payload = {
            "addTimeStart"         : f"{first_day} 00:00:00",
            "addTimeEnd"           : f"{last_day} 23:59:59",
            "returnOrderStatusList": [4],
            "page"                 : page_num,
            "perPage"              : page_size
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取不扣款列表 第{page}/{totalPage}页')
            page_num = page
            payload = {
                "addTimeStart"         : f"{first_day} 00:00:00",
                "addTimeEnd"           : f"{last_day} 23:59:59",
                "returnOrderStatusList": [4],
                "page"                 : page_num,
                "perPage"              : page_size
            }
            response_text = fetch(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            list_item += spu_list_new
            time.sleep(0.1)

        write_dict_to_file(cache_file, list_item)

        return list_item

    def get_comment_list(self):
        cache_file = f'{self.config.auto_dir}/shein/dict/comment_list_{TimeUtils.today_date()}.json'
        comment_list = read_dict_from_file_ex(cache_file, self.store_username, 3600)
        if len(comment_list) > 0:
            return comment_list

        page_num = 1
        page_size = 50

        yesterday = TimeUtils.get_yesterday()

        url = f"https://sso.geiwohuo.com/gsp/goods/comment/list"
        payload = {
            "page"            : page_num,
            "perPage"         : page_size,
            "startCommentTime": f"{yesterday} 00:00:00",
            "commentEndTime"  : f"{yesterday} 23:59:59",
            "commentStarList" : ["3", "2", "1"]
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        comment_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取评价列表 第{page}/{totalPage}页')
            page_num = page
            payload['page'] = page_num
            response_text = fetch(self.web_page, url, payload)
            comment_list = response_text['info']['data']
            time.sleep(0.1)

        write_dict_to_file_ex(cache_file, {self.store_username: comment_list}, [self.store_username])
        return comment_list

    def get_last_month_outbound_amount(self):
        url = "https://sso.geiwohuo.com/mils/report/month/list"
        start, end = TimeUtils.get_current_year_range()
        payload = {
            "reportDateStart": start, "reportDateEnd": end, "pageNumber": 1, "pageSize": 50
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        lst = info.get('data', {}).get('list', [])
        if not lst:
            log(f'⚠️ {self.store_name} 最近一个月无出库记录，金额为0')
            return 0

        last_item = lst[-1]
        log(f'正在获取 {self.store_name} 最近一个月出库金额: {last_item["totalCustomerAmount"]}')
        return last_item['totalCustomerAmount']

    def query_attribute_multi(self, attribute_id_list):
        url = "https://sso.geiwohuo.com/spmp-api-prefix/spmp/attribute/query_attribute_multi"
        payload = {
            "attribute_id_list": attribute_id_list,
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        lst = info.get('data', {})
        return lst

    def get_product_attr(self, spu, attr_name):
        try:
            product_detail = self.get_product_detail(spu)
            product_type_id = product_detail.get('product_type_id')
            category_id = product_detail.get('category_id')

            if not product_type_id or not category_id:
                return None  # 或者根据需要返回一个默认值

            attribute_template = self.get_attribute_templates(spu, category_id, [product_type_id])
            attr_info = attribute_template.get('attribute_infos', [])

            # 查找材质属性映射，防止没有匹配项
            attr_item = next((item for item in attr_info if item.get('attribute_name') == attr_name), None)
            if not attr_item:
                return None  # 或者返回一个默认值

            attr_id = attr_item.get('attribute_id')

            # 拿到产品材质的属性值ID
            product_attribute_list = product_detail.get('product_attribute_list', [])
            attribute_value_id = next((item['attribute_value_id'] for item in product_attribute_list if item.get('attribute_id') == attr_id), None)
            if not attribute_value_id:
                return None  # 或者返回一个默认值

            # 获取属性值名称
            attr_value = next((item['attribute_value'] for item in attr_item.get('attribute_value_info_list', []) if item.get('attribute_value_id') == attribute_value_id), None)
            return attr_value  # 返回找到的属性值
        except Exception as e:
            log(f"Error occurred: {e}")
            send_exception()
            return None  # 或者返回一个默认值

    def get_attribute_templates(self, spu_name, category_id, product_type_id_list):
        log(f'正在获取 {spu_name} 商品属性模板')

        if not isinstance(product_type_id_list, list):
            raise '参数错误: product_type_id_list 需要是列表'

        cache_file = f'{self.config.auto_dir}/shein/attribute/attribute_template_{spu_name}.json'
        attr_list = read_dict_from_file(cache_file, 3600 * 24 * 7)
        if len(attr_list) > 0:
            return attr_list

        url = f"https://sso.geiwohuo.com/spmp-api-prefix/spmp/basic/query_attribute_templates"
        payload = {
            "category_id"         : category_id,
            "for_update"          : True,
            "product_type_id_list": product_type_id_list,
            "spu_name"            : spu_name
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')

        data = info.get('data')[0]
        write_dict_to_file(cache_file, data)
        return data

    def get_product_detail(self, spu_name, cache_interval=3600 * 24 * 7):
        cache_file = f'{self.config.auto_dir}/shein/product_detail/product_detail_{spu_name}.json'
        info = read_dict_from_file(cache_file, cache_interval)
        if info and len(info) > 0:
            return info

        log(f'正在获取 {spu_name} 商品详情')
        url = f"https://sso.geiwohuo.com/spmp-api-prefix/spmp/product/get_product_detail"
        payload = {
            "spu_name": spu_name
        }
        response_text = fetch(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')

        # 获取 area_attribute_id
        sample_sku_back_size = info.get('sample_sku_back_size', None)
        if sample_sku_back_size is not None:
            area_attribute_ids = [item['area_attribute_id'] for item in sample_sku_back_size.get('area_info_list', [])]
            attribute_multi = self.query_attribute_multi(area_attribute_ids)
            info["attribute_multi"] = attribute_multi

        write_dict_to_file(cache_file, info)
        return info

    def product_month_analysis(self, start_date, end_date):
        # 店铺信息包含 店铺名称 skc上架状态 skc商品层级 统计周期
        # 商品信息包含 SPU,SKC,商家SKC,质量等级,商品分类,上架日期，上架天数
        # SKU信息包含 商家SKU,属性集
        # 前9列均是skc维度,从SKU信息开始 后面是SKU维度
        excel_data = [
            ['店铺信息', '商品信息', 'SKC图片', '30天SKC曝光', '30天SKC点击率', '30天SKC转化率', '评论数', '差评率', '客单退货件数', 'SKU信息', 'SKU图片', 'SKU30天销量', '销售额', '核价', '成本', '30天利润', '30天利润率', 'skc']
        ]
        excel_data2 = [
            ['店铺信息', '商品信息', 'SKC图片', '日期', 'SKC销量', 'SKC曝光', 'SKC点击率', 'SKC转化率', 'skc']
        ]
        skc_list = self.get_bak_base_info()
        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku = read_dict_from_file(cache_file)
        cache_file = f'{self.config.auto_dir}/shein/quality_label/quality_label_{self.store_username}.json'
        dict_quality_label = read_dict_from_file(cache_file)

        cache_file_analysis = f'{self.config.auto_dir}/shein/product_analysis/skc_skc_analysis_{self.store_username}_{start_date}_{end_date}.json'
        dict_analysis = read_dict_from_file(cache_file_analysis)

        for skc_item in skc_list:
            categoryName = skc_item['categoryName']
            spu = skc_item['spu']  # SPU
            skc = skc_item['skc']  # SKC
            supplierCode = skc_item['supplierCode']  # 商家SKC
            skc_img = skc_item['picUrl']  # SKC图片
            shelfDate = skc_item['shelfDate']  # 上架日期
            shelfDays = skc_item['shelfDays']  # 上架天数
            shelfStatusName = skc_item['shelfStatus']['name']  # 上架状态
            quality_label = dict_quality_label.get(skc, {}).get('name', '')  # 质量等级
            if quality_label == '无判断':
                quality_label = ''

            if shelfStatusName == '待上架':
                log('商品未上架跳过:', skc)
                continue

            goods_level = skc_item.get('goodsLevel', {}).get('name', '-')
            if goods_level in ['自主停产', '退供款']:
                log(f'商品 {goods_level} 跳过:', skc)
                continue

            dict_sku_sales = self.get_skc_actual_sales_dict(skc, start_date, end_date)
            dict_skc_trend = self.get_skc_trend(spu, skc, start_date, end_date)

            # 检查这个 SKC 是否有任何 SKU 有销量（与 excel_data 保持一致）
            has_sales = False
            for sku_item in skc_item['skuList']:
                c30dSaleCnt = sku_item.get('c30dSaleCnt', 0)
                attr = sku_item.get('attr', '')
                if attr != '合计' and int(c30dSaleCnt) > 0:
                    has_sales = True
                    break

            # 只有当有趋势数据且有销量时才添加到 excel_data2（与 excel_data 保持一致）
            if dict_skc_trend and has_sales:
                for stat_date, dict_item in dict_skc_trend.items():
                    store_info = f'{self.store_username}\n{self.store_name}\n({shelfStatusName})\n{goods_level}\n{start_date}\n{end_date}'
                    product_info = f'SPU: {spu}\nSKC: {skc}\n商家SKC: {supplierCode}\n商品分类: {categoryName}\n上架日期: {shelfDate}\n上架天数: {shelfDays}\n质量等级: {quality_label}'

                    row_item2 = []
                    row_item2.append(store_info)
                    row_item2.append(product_info)
                    row_item2.append(skc_img)
                    row_item2.append(stat_date)
                    row_item2.append(dict_item.get('saleCnt', 0))
                    row_item2.append(dict_item.get('epsUvIdx', 0))
                    row_item2.append(dict_item.get('epsGdsCtrIdx', 0))
                    row_item2.append(dict_item.get('gdsPayCtrIdx', 0))
                    row_item2.append(skc if skc else '')  # 确保 skc 不为 None
                    excel_data2.append(row_item2)

            for sku_item in skc_item['skuList']:
                supplierSku = sku_item['supplierSku']  # 商家SKU
                attr = sku_item['attr']  # 属性集
                sku = sku_item['skuCode']  # SKU

                c30dSaleCnt = sku_item['c30dSaleCnt']  # 近30天销量
                if attr == '合计' or int(c30dSaleCnt) == 0:
                    log(f'跳过: {supplierSku},近30天销量: {c30dSaleCnt}')
                    continue

                price = dict_sku[sku]

                store_info = f'{self.store_username}\n{self.store_name}\n({shelfStatusName})\n{goods_level}\n{start_date}\n{end_date}'
                product_info = f'SPU: {spu}\nSKC: {skc}\n商家SKC: {supplierCode}\n商品分类: {categoryName}\n上架日期: {shelfDate}\n上架天数: {shelfDays}\n质量等级: {quality_label}'

                epsUvIdx = dict_analysis.get(skc, {}).get('epsUvIdx', 0)
                epsGdsCtrIdx = dict_analysis.get(skc, {}).get('epsGdsCtrIdx', 0)
                gdsPayCtrIdx = dict_analysis.get(skc, {}).get('gdsPayCtrIdx', 0)
                totalCommentCnt = dict_analysis.get(skc, {}).get('totalCommentCnt', 0)
                badCommentRate = dict_analysis.get(skc, {}).get('badCommentRate', 0)
                returnOrderCnt = dict_analysis.get(skc, {}).get('returnOrderCnt', 0)

                sku_info = f'平台SKU: {sku}\n商家SKU: {supplierSku}\n属性集: {attr}'
                sku_img = self.bridge.get_sku_img(supplierSku, 'mb')
                # cost_price = self.bridge.get_sku_cost(sku_item['supplierSku'], self.config.erp_source)
                cost_price = self.bridge.get_sku_cost(sku_item['supplierSku'], 'mb')

                row_item = []
                row_item.append(store_info)
                row_item.append(product_info)
                row_item.append(skc_img)
                row_item.append(epsUvIdx)
                row_item.append(epsGdsCtrIdx)
                row_item.append(gdsPayCtrIdx)
                row_item.append(totalCommentCnt)
                row_item.append(badCommentRate)
                row_item.append(returnOrderCnt)
                row_item.append(sku_info)
                row_item.append(sku_img)
                row_item.append(dict_sku_sales.get(sku, 0))
                row_item.append('')  # 销售额（公式计算）
                row_item.append(price)  # 核价
                row_item.append(cost_price)  # 成本
                row_item.append('')  # 30天利润（公式计算）
                row_item.append('')  # 30天利润率（公式计算）
                row_item.append(skc)
                excel_data.append(row_item)

        cache_file = f'{self.config.auto_dir}/shein/product_analysis/product_analysis_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: excel_data}, [self.store_username])

        cache_file = f'{self.config.auto_dir}/shein/product_analysis/product_analysis_2_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: excel_data2}, [self.store_username])
        return excel_data

    def get_product(self):
        excel_data = [
            ['店铺信息', '产品信息', 'SKC', '商家SKC', 'SKC图片', '商家SKU', '属性集', '近7天销量', '近30天销量', '核价', 'ERP成本价', '近7天利润', '近30天利润', '导出时间', 'SPU', 'SKC_FOR_STAT']
        ]
        skc_list = self.get_bak_base_info()
        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku = read_dict_from_file(cache_file)
        for skc_item in skc_list:
            categoryName = skc_item['categoryName']
            spu = skc_item['spu']
            skc = skc_item['skc']
            supplierCode = skc_item['supplierCode']
            skc_img = skc_item['picUrl']
            shelfDate = skc_item['shelfDate']
            shelfDays = skc_item['shelfDays']
            shelfStatusName = skc_item['shelfStatus']['name']
            # if shelfStatusName != '已下架':
            #     continue
            for sku_item in skc_item['skuList']:
                supplierSku = sku_item['supplierSku']
                attr = sku_item['attr']
                sku = sku_item['skuCode']
                c7dSaleCnt = sku_item['c7dSaleCnt']
                c30dSaleCnt = sku_item['c30dSaleCnt']
                if attr == '合计' or int(c30dSaleCnt) == 0:
                    log(f'跳过: {supplierSku},近30天销量: {c30dSaleCnt}')
                    continue

                price = dict_sku[sku]

                product_info = f'SPU: {spu}\n商品分类: {categoryName}\n上架日期: {shelfDate}\n上架天数: {shelfDays}\n上架状态: {shelfStatusName}'

                store_info = f'{self.store_username}\n{self.store_name}\n{self.config.shein_store_manager.get(self.store_username)}'

                row_item = []
                row_item.append(store_info)
                row_item.append(product_info)
                row_item.append(skc)
                row_item.append(supplierCode)
                row_item.append(skc_img)
                row_item.append(supplierSku)
                row_item.append(attr)
                row_item.append(c7dSaleCnt)
                row_item.append(c30dSaleCnt)
                row_item.append(price)
                row_item.append('')
                row_item.append('')
                row_item.append('')
                row_item.append(TimeUtils.current_datetime())
                row_item.append(spu)
                row_item.append(skc)
                excel_data.append(row_item)

        cache_file = f'{self.config.auto_dir}/shein/product/product_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: excel_data}, [self.store_username])
        return excel_data

    def generate_product_dict(self):
        pass
        dict_sku_to_skc = []
        dict_sku_not_found = []
        skc_list = self.get_bak_base_info()
        for skc_item in skc_list:
            skc_item['store_username'] = self.store_username
            skc_item['store_name'] = self.store_name
            skc_item['store_manager'] = self.config.shein_store_manager.get(str(self.store_username).lower())
            spu = skc_item['spu']
            skc = skc_item['skc']
            supplierCode = skc_item['supplierCode']

            shelf_status = skc_item.get('shelfStatus', {}).get('name', '-')
            if int(skc_item['shelfStatus']['value']) != 1:
                log('商品未上架跳过:', skc)
                continue

            goods_level = skc_item.get('goodsLevel', {}).get('name', '-')
            if goods_level in ['自主停产', '退供款']:
                log(f'商品{goods_level}跳过:', skc)
                continue

            # 倒序遍历 skuList，安全删除
            for i in range(len(skc_item['skuList']) - 1, -1, -1):
                sku_item = skc_item['skuList'][i]
                if sku_item['skuCode'] == '合计':
                    del skc_item['skuList'][i]  # 删除“合计”
                    continue

                cost_price = self.bridge.get_sku_cost(sku_item['supplierSku'], self.config.erp_source)
                if not isinstance(cost_price, (int, float)):
                    dict_sku_not_found.append([
                        self.store_username,
                        f'{self.store_name}',
                        self.config.shein_store_manager.get(str(self.store_username).lower()),
                        spu,
                        skc,
                        supplierCode,
                        sku_item['supplierSku'],
                        shelf_status,
                        goods_level,
                        '忆托未匹配到成本价,可能原因: 1.没填商家SKU,2.商家SKU没有绑定本地SKU,3.商家SKU填写错误'
                    ])
                elif cost_price == 0:
                    dict_sku_not_found.append([
                        self.store_username,
                        f'{self.store_name}',
                        self.config.shein_store_manager.get(str(self.store_username).lower()),
                        spu,
                        skc,
                        supplierCode,
                        sku_item['supplierSku'],
                        shelf_status,
                        goods_level,
                        '忆托未匹配到成本价为:0'
                    ])

                dict_sku_to_skc.append([
                    sku_item['supplierSku'],
                    supplierCode,
                ])

            cache_file = f'{self.config.auto_dir}/shein/dict/sku_not_found.json'
            write_dict_to_file_ex(cache_file, {self.store_username: dict_sku_not_found}, [self.store_username])

            cache_file = f'{self.config.auto_dir}/shein/dict/sku_to_skc.json'
            write_dict_to_file_ex(cache_file, {self.store_username: dict_sku_to_skc}, [self.store_username])

    # 存储商品库
    def store_product_info(self):
        # todo 商品详情 属性 规格 图片 重量 与 尺寸
        skc_list = self.get_bak_base_info()
        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku = read_dict_from_file(cache_file)
        dict_product_detail = []
        for skc_item in skc_list:
            skc_item['store_username'] = self.store_username
            skc_item['store_name'] = self.store_name
            skc_item['store_manager'] = self.config.shein_store_manager.get(str(self.store_username).lower())
            spu = skc_item['spu']
            if spu not in dict_product_detail:
                dict_product_detail.append(spu)
                material = self.get_product_attr(spu, '材质')
                log(material)  # 这一步是为了获取 spu 详情和属性

            # 倒序遍历 skuList，安全删除
            for i in range(len(skc_item['skuList']) - 1, -1, -1):
                sku_item = skc_item['skuList'][i]
                if sku_item['skuCode'] == '合计':
                    del skc_item['skuList'][i]  # 删除“合计”
                    continue
                sku_item['price'] = dict_sku[sku_item['skuCode']]
                cost_price = self.bridge.get_sku_cost(sku_item['supplierSku'], self.config.erp_source)
                sku_item['erp_cost_price'] = cost_price if isinstance(cost_price, (int, float)) else None
                sku_item['erp_supplier_name'] = self.bridge.get_sku_supplier(sku_item['supplierSku'], self.config.erp_source)
                stock = self.bridge.get_sku_stock(sku_item['supplierSku'], self.config.erp_source)
                sku_item['erp_stock'] = stock if isinstance(stock, (int, float)) else None

        cache_file = f'{self.config.auto_dir}/shein/product/skc_list_{self.store_username}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: skc_list}, [self.store_username])

        skc_file = f'{self.config.auto_dir}/shein/product/skc_list_file.json'
        write_dict_to_file_ex(skc_file, {self.store_username: cache_file}, [self.store_username])

        detail_file = f'{self.config.auto_dir}/shein/product/product_detail_file.json'
        write_dict_to_file_ex(detail_file, {self.store_username: dict_product_detail}, [self.store_username])

    def get_skc_diagnose_dict(self, start_date="", end_date=""):
        log(f'获取商品分析某个月的字典 {start_date} {end_date} {self.store_name} {self.store_username}')

        cache_file_analysis = f'{self.config.auto_dir}/shein/product_analysis/skc_skc_analysis_{self.store_username}_{start_date}_{end_date}.json'

        dict_analysis = read_dict_from_file(cache_file_analysis)
        if len(dict_analysis) > 0:
            return dict_analysis

        dt_goods = self.get_dt_time_goods()
        if not TimeUtils.is_yesterday_date(dt_goods, "%Y%m%d"):
            log("数据尚未更新: dt_goods:", dt_goods)
            return []

        url = "https://sso.geiwohuo.com/sbn/new_goods/get_skc_diagnose_list"
        page_num = 1
        page_size = 100
        payload = {
            "areaCd"     : "cn",
            "dt"         : dt_goods,
            "countrySite": [
                "shein-all"
            ],
            "startDate"  : start_date.replace('-', ""),
            "endDate"    : end_date.replace('-', ""),
            "pageNum"    : page_num,
            "pageSize"   : page_size,
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        spu_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取商品分析列表(最近上架的) 第{page}/{totalPage}页')
            payload.update({"pageNum": page})
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            spu_list += spu_list_new
            time.sleep(0.3)

        cache_file = f'{self.config.auto_dir}/shein/product_analysis/skc_dict_{self.store_username}_{start_date}_{end_date}.json'
        write_dict_to_file(cache_file, spu_list)

        for skc_item in spu_list:
            skc = skc_item['skc']
            skc_item['store_username'] = self.store_username
            skc_item['store_name'] = self.store_name
            dict_analysis[skc] = skc_item

        write_dict_to_file(cache_file_analysis, dict_analysis)

        return dict_analysis

    def get_skc_diagnose_list(self, shelf_date_begin="", shelf_date_end=""):
        log(f'获取商品分析列表(最近上架的或在售的) {shelf_date_begin} {shelf_date_end} {self.store_name} {self.store_username}')

        dt_goods = self.get_dt_time_goods()
        if not TimeUtils.is_yesterday_date(dt_goods, "%Y%m%d"):
            log("数据尚未更新: dt_goods:", dt_goods)
            return []

        yesterday = TimeUtils.get_past_nth_day(1, None, '%Y%m%d')

        url = "https://sso.geiwohuo.com/sbn/new_goods/get_skc_diagnose_list"
        page_num = 1
        page_size = 100
        payload = {
            "areaCd"     : "cn",
            "dt"         : dt_goods,
            "countrySite": [
                "shein-all"
            ],
            "startDate"  : yesterday,
            "endDate"    : yesterday,
            "pageNum"    : page_num,
            "pageSize"   : page_size,
            "onsaleFlag" : 1,
            # "localFrstSaleBeginDate": shelf_date_begin,
            # "localFrstSaleEndDate"  : shelf_date_end,
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        spu_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取商品分析列表(最近上架的) 第{page}/{totalPage}页')
            payload.update({"pageNum": page})
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            spu_list += spu_list_new
            time.sleep(0.3)

        cache_file = f'{self.config.auto_dir}/shein/product_analysis/skc_list_{self.store_username}.json'
        write_dict_to_file(cache_file, spu_list)

        cache_file = f'{self.config.auto_dir}/shein/dict/skc_shelf_date_{self.store_username}.json'
        dict_skc_shelf_date = read_dict_from_file(cache_file)

        # 活动信息
        # AB实验数据

        # 预先过滤掉不需要的商品状态
        log(f'过滤前商品数量: {len(spu_list)}')
        exclude_levels = ['退供款', '自主停产', '自主下架']
        spu_list = [item for item in spu_list if item['layerNm'] not in exclude_levels]
        log(f'过滤后剩余商品数量: {len(spu_list)}')

        for skc_item in spu_list:
            skc = skc_item['skc']
            skc_item['stat_date'] = datetime.strptime(yesterday, "%Y%m%d").strftime("%Y-%m-%d")
            skc_item['store_username'] = self.store_username
            skc_item['store_name'] = self.store_name
            skc_item['shelf_date'] = dict_skc_shelf_date[skc]
            ab_cache_file = f'{self.config.auto_dir}/shein/cache/ab_test_list_{skc}_{TimeUtils.today_date()}.json'
            skc_item['ab_test'] = read_dict_from_file(ab_cache_file)
            for prom_inf_ing in skc_item['promCampaign'].get('promInfIng') or []:
                prom_id = prom_inf_ing['promId']
                log('prom_id:', prom_id, len(prom_id))
                if len(prom_id) >= 11:
                    # 托管活动
                    prom_inf_ing['promDetail'] = self.get_skc_activity_price_info(skc, prom_id)
                elif len(prom_id) >= 8:
                    # 营销工具
                    prom_inf_ing['promDetail'] = self.query_goods_detail(prom_id)
                else:
                    # 营销活动
                    prom_inf_ing['promDetail'] = self.get_partake_activity_detail(prom_id, skc)

            for prom_inf_ready in skc_item['promCampaign'].get('promInfReady') or []:
                prom_id = prom_inf_ready['promId']
                log('prom_id:', prom_id, len(prom_id))
                if len(prom_id) >= 11:
                    prom_inf_ready['promDetail'] = self.get_skc_activity_price_info(skc, prom_id)
                elif len(prom_id) >= 8:
                    prom_inf_ready['promDetail'] = self.query_goods_detail(prom_id)
                else:
                    prom_inf_ready['promDetail'] = self.get_partake_activity_detail(prom_id, skc)

        cache_file = f'{self.config.auto_dir}/shein/product_analysis/skc_model_{self.store_username}_{TimeUtils.today_date()}.json'
        write_dict_to_file(cache_file, spu_list)

        return spu_list

    # 获取备货信息列表 最近35天上架的
    def get_latest_shelf_list(self, shelf_date_begin="", shelf_date_end=""):
        log(f'获取备货信息列表(最近上架的或已上架的) {shelf_date_begin} {shelf_date_end} {self.store_name} {self.store_username}')

        dict_skc_shelf_date = {}

        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        pageNumber = 1
        pageSize = 100
        payload = {
            "pageNumber"            : pageNumber,
            "pageSize"              : pageSize,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,  # 已上架
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : shelf_date_begin,
            "shelfDateEnd"          : shelf_date_end,
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['list']

        # skc_list = [item['skc'] for item in spu_list]
        # self.get_activity_label(skc_list)
        # self.get_preemption_list(skc_list)
        # self.get_sku_price_v2(skc_list)
        # self.get_stock_advice(skc_list)

        total = response_text['info']['count']
        totalPage = math.ceil(total / pageSize)
        for page in range(2, totalPage + 1):
            log(f'获取备货信息商品列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch_shein(self.web_page, url, payload)

            new_spu_list = response_text['info']['list']
            spu_list += new_spu_list

            # skc_list = [item['skc'] for item in new_spu_list]
            # self.get_activity_label(skc_list)
            # self.get_preemption_list(skc_list)
            # self.get_sku_price_v2(skc_list)
            # self.get_stock_advice(skc_list)

            time.sleep(0.3)

        # key = f'{self.store_username}'
        # cache_file = f'{self.config.auto_dir}/shein/cache/bak_info_list_{key}_{shelf_date_begin}_{shelf_date_end}.json'
        # write_dict_to_file_ex(cache_file, {key: spu_list}, [key])

        for skc_item in spu_list:
            skc = skc_item['skc']
            shelfDate = skc_item['shelfDate']
            dict_skc_shelf_date[skc] = shelfDate

        cache_file = f'{self.config.auto_dir}/shein/dict/skc_shelf_date_{self.store_username}.json'
        dict = read_dict_from_file(cache_file)
        dict.update(dict_skc_shelf_date)
        write_dict_to_file(cache_file, dict)

        return spu_list

    # 获取备货信息列表
    def get_bak_base_info(self):
        log(f'获取备货信息列表 {self.store_name} {self.store_username}')
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        pageNumber = 1
        pageSize = 100
        payload = {
            "pageNumber"            : pageNumber,
            "pageSize"              : pageSize,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [],
            "supplyStatus"          : "",
            "shelfStatus"           : "",
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['list']

        skc_list = [item['skc'] for item in spu_list]
        self.get_activity_label(skc_list)
        self.get_quality_label(skc_list)
        self.get_preemption_list(skc_list)
        self.get_sku_price_v2(skc_list)
        self.get_stock_advice(skc_list)

        total = response_text['info']['count']
        totalPage = math.ceil(total / pageSize)
        for page in range(2, totalPage + 1):
            log(f'获取备货信息商品列表 第{page}/{totalPage}页')
            payload['pageNumber'] = page
            response_text = fetch_shein(self.web_page, url, payload)

            new_spu_list = response_text['info']['list']
            spu_list += new_spu_list

            skc_list = [item['skc'] for item in new_spu_list]
            self.get_activity_label(skc_list)
            self.get_quality_label(skc_list)
            self.get_preemption_list(skc_list)
            self.get_sku_price_v2(skc_list)
            self.get_stock_advice(skc_list)

            time.sleep(0.3)

        key = f'{self.store_username}'
        cache_file = f'{self.config.auto_dir}/shein/cache/bak_info_list_{key}.json'
        write_dict_to_file_ex(cache_file, {key: spu_list}, [key])

        return spu_list

    def get_skc_week_sale_list(self, spu, skc, start_from=None):
        dict_skc = self.get_dict_skc_week_trend_v2(spu, skc, start_from)
        date_list = TimeUtils.get_past_7_days_list(start_from)
        saleCnt7d = 0
        sales_detail = []
        for date in date_list:
            saleCnt = get_safe_value(dict_skc.get(date, {}), 'saleCnt', 0)
            epsUvIdx = get_safe_value(dict_skc.get(date, {}), 'epsUvIdx', 0)

            saleCnt7d += saleCnt
            sales_detail.append(f'{date}({TimeUtils.get_weekday_name(date)}): {saleCnt}/{epsUvIdx}')

        sales_data = []
        for date in date_list:
            goodsUvIdx = get_safe_value(dict_skc.get(date, {}), 'goodsUvIdx', 0)  # 商详访客
            epsGdsCtrIdx = get_safe_value(dict_skc.get(date, {}), 'epsGdsCtrIdx', 0)  # 点击率

            payUvIdx = get_safe_value(dict_skc.get(date, {}), 'payUvIdx', 0)  # 支付人数
            gdsPayCtrIdx = get_safe_value(dict_skc.get(date, {}), 'gdsPayCtrIdx', 0)  # 转化率

            sales_data.append(f'{date}({TimeUtils.get_weekday_name(date)}): {epsGdsCtrIdx:.2%}({goodsUvIdx})/{gdsPayCtrIdx:.2%}({payUvIdx})')

        return sales_detail, sales_data, saleCnt7d

    def stat_new_product_to_bak(self):
        # 直接调用 get_skc_week_actual_sales 來获取周销 计算是否能转成备货款
        skc_list = self.get_bak_base_info()  # 这个地方 不要加已上架和正常供货参数 直接取所有的skc列表
        # 以算昨日7.2日为例 上架天数为29天(转换成上架日期),且过去7天销量达到类目备货标准和没有达到备货标准的skc数量
        # 1.计算某个skc的上架日期
        # 2.计算某个skc的基于某个日期的过去7天销量
        # 3.获取叶子类目的备货标准
        header = ['店铺账号', '店铺名称', '店长', '统计日期', 'SKC图片', '商品信息', '新品成功转备货款', '第4周SKC销量/SKC曝光', '第4周SKC点击率/SKC转化率', 'SKC', 'SPU']
        excel_data = []
        stat_date_list = TimeUtils.get_dates_from_first_of_month_to_yesterday()
        for stat_date in stat_date_list:
            # 计算 stat_date 这天 的上架日期 是 filter_shelf_date
            filter_shelf_date = TimeUtils.get_past_nth_day(29, stat_date)
            log(f'stat_date:{stat_date},filter_shelf_date:{filter_shelf_date}')
            # 筛选 上架日期是 filter_shelf_date 这天的skc有哪些
            filter_skc_list = [skc_item for skc_item in skc_list if skc_item['shelfDate'] == filter_shelf_date]
            # 再统计 这些skc 在 stat_date 这天的 前7天销量
            # 看看这个7天销量是否达到了类目的备货标准 统计 计数
            for skc_item in filter_skc_list:
                skc = skc_item['skc']
                spu = skc_item['spu']
                log(f'skc:{skc}, spu:{spu}')

                row_item = []
                row_item.append(self.store_username)

                status_cn = skc_item.get('shelfStatus').get('name')
                goods_level = skc_item['goodsLevel']['name']
                goods_label = [label["name"] for label in skc_item['goodsLabelList']]
                store_info = f'{self.store_name}\n({status_cn})\n{goods_level}\n{",".join(goods_label).strip()}\n{stat_date_list[-1]}\n{stat_date_list[0]}'
                row_item.append(store_info)
                store_manager = self.config.shein_store_manager.get(str(self.store_username).lower())
                row_item.append(store_manager)
                row_item.append(stat_date)
                row_item.append(skc_item['picUrl'])

                standard_value = (skc_item.get('stockStandard') or {}).get('value') or 0

                sale_detail, sale_rate, sale_num = self.get_skc_week_sale_list(spu, skc, stat_date)
                success = int(standard_value > 0 and sale_num >= standard_value)

                categoryName = skc_item['categoryName']
                shelfDate = skc_item['shelfDate']
                product_info = (
                    f'SPU: {spu}\n'
                    f'SKC: {skc}\n'
                    f'上架日期: {shelfDate}\n'
                    f'类目: {categoryName}\n'
                    f'备货标准/第4周销: {standard_value}/{sale_num}\n'
                )
                row_item.append(product_info)
                row_item.append(success)
                row_item.append("\n".join(sale_detail))
                row_item.append("\n".join(sale_rate))
                row_item.append(skc)
                row_item.append(spu)
                excel_data.append(row_item)

        cache_file = f'{self.config.auto_dir}/shein/dict/new_product_to_bak_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: [header] + excel_data}, [self.store_username])

    def get_funds_data_lz(self):
        log(f'正在获取 {self.store_name} 财务数据')
        url = "https://sso.geiwohuo.com/sso/homePage/dataOverview/v2/detail"
        payload = {
            "metaIndexIds": [
                298,
                67,
                70,
                72
            ],
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        num298 = 0  # 在途商品金额
        num67 = 0  # 在仓商品金额
        num70 = 0  # 待结算金额
        num72 = 0  # 可提现金额
        for item in info['list']:
            if item['metaIndexId'] == 298:
                num298 = item['count']
            if item['metaIndexId'] == 67:
                num67 = item['count']
            if item['metaIndexId'] == 70:
                num70 = item['count']
            if item['metaIndexId'] == 72:
                num72 = item['count']

        # 获取钱包余额中的不可提现金额和保证金
        no_withdrawable_amount = 0  # 不可提现金额
        deposit_amount_paid = 0  # 已缴保证金
        try:
            supplier_info = self.get_supplier_data()
            supplier_id = supplier_info.get('supplier_id')
            if supplier_id:
                wallet_info = self.get_wallet_balance(supplier_id)
                if wallet_info:
                    # 获取不可提现金额
                    detail_list = wallet_info.get('detailResponseList') or []
                    for detail in detail_list:
                        if detail.get('currency') == 'CNY':
                            no_withdrawable_amount = float(detail.get('noWithdrawableAmount') or 0)
                            break
                    # 获取已缴保证金
                    deposit_list = wallet_info.get('depositDetailResponseList') or []
                    for deposit in deposit_list:
                        if deposit.get('currency') == 'CNY':
                            deposit_amount_paid = float(deposit.get('depositAmountPaid') or 0)
                            break
        except Exception as e:
            log(f'获取钱包余额信息失败: {e}')

        outAmount = self.get_last_month_outbound_amount()
        store_manager = self.config.shein_store_manager.get(str(self.store_username).lower())

        # 获取用户信息（店铺ID、商家ID、全球唯一编码）
        store_code = ''  # 店铺ID
        supplier_id_str = ''  # 商家ID
        merchant_code = ''  # 全球唯一编码
        try:
            user_info = self.get_user()
            if user_info:
                store_code = str(user_info.get('storeCode', ''))
                supplier_id_str = str(user_info.get('supplierId', ''))
                merchant_code = user_info.get('merchantCode', '') or ''
        except Exception as e:
            log(f'获取用户信息失败: {e}')

        # 数据结构: [店铺名称, 店铺账号, 店长, 在途商品金额, 在仓商品金额, 待结算金额, 可提现金额, 不可提现金额, 保证金, 汇总, 销售出库金额, 新品上架数量, 成功转备货款, 成功率, 导出时间, 店铺ID, 商家ID, 全球唯一编码]
        NotifyItem = [f'{self.store_name}', self.store_username, store_manager, num298, num67, num70, num72, no_withdrawable_amount, deposit_amount_paid, '', outAmount, '', '', '', TimeUtils.current_datetime(), store_code, supplier_id_str, merchant_code]
        log(NotifyItem)
        cache_file = f'{self.config.auto_dir}/shein/cache/stat_fund_lz_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: NotifyItem}, [self.store_username])
        return NotifyItem

    def get_funds_data(self):
        log(f'正在获取 {self.store_name} 财务数据')
        url = "https://sso.geiwohuo.com/sso/homePage/dataOverview/v2/detail"
        payload = {
            "metaIndexIds": [
                298,
                67,
                70,
                72
            ],
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        num298 = 0  # 在途商品金额
        num67 = 0  # 在仓商品金额
        num70 = 0  # 待结算金额
        num72 = 0  # 可提现金额
        for item in info['list']:
            if item['metaIndexId'] == 298:
                num298 = item['count']
            if item['metaIndexId'] == 67:
                num67 = item['count']
            if item['metaIndexId'] == 70:
                num70 = item['count']
            if item['metaIndexId'] == 72:
                num72 = item['count']

        outAmount = self.get_last_month_outbound_amount()
        dict_store = read_dict_from_file(self.config.shein_store_alias)
        store_manager = dict_store.get(str(self.store_username).lower())
        NotifyItem = [f'{self.store_name}', self.store_username, store_manager, num298, num67, num70, num72, outAmount, '',
                      TimeUtils.current_datetime()]

        cache_file = f'{self.config.auto_dir}/shein/cache/stat_fund_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: NotifyItem}, [self.store_username])
        return NotifyItem

    def getQueryDate(self):
        query_time = self.DictQueryTime.get(self.store_username, None)
        if query_time is not None:
            log(f'从字典获取query_time: {query_time}')
            return query_time
        log('获取日期范围')
        url = "https://sso.geiwohuo.com/mgs-api-prefix/estimate/queryDateRange"
        payload = {}
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        query_time = response_text.get('info').get('quality_goods_query_time')
        self.DictQueryTime.update({self.store_username: query_time})
        log(f'query_time: {query_time}')
        return query_time

    def get_goods_quality_estimate_list(self, query_date):
        cache_file = f'{self.config.auto_dir}/shein/dict/googs_estimate_{query_date}.json'
        estimate_list = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 8)
        if len(estimate_list) > 0:
            return estimate_list

        page_num = 1
        page_size = 100

        url = f"https://sso.geiwohuo.com/mgs-api-prefix/estimate/queryNewQualityGoodsList"
        payload = {
            "page_no"   : page_num,
            "page_size" : page_size,
            "start_date": query_date,
            "end_date"  : query_date,
            "order_col" : "skc_sale_cnt_14d",
            "order_type": "desc"
        }
        response_text = fetch_shein(self.web_page, url, payload, {'lan': 'CN'})
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        estimate_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取质量评估列表 第{page}/{totalPage}页')
            page_num = page
            payload['page'] = page_num
            response_text = fetch_shein(self.web_page, url, payload)
            estimate_list = response_text['info']['data']
            time.sleep(0.1)

        write_dict_to_file_ex(cache_file, {self.store_username: estimate_list}, [self.store_username])
        return estimate_list

    # 已上架备货款A数量
    def get_product_bak_A_count(self):
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        payload = {
            "pageNumber"            : 1,
            "pageSize"              : 10,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [
                61,
                90
            ],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2,
            "goodsLevelFakeIdList"  : [
                3
            ]
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        count = info.get('count', 0)
        log('获取已上架备货款A数量', count)
        return count

    # 已上架备货款B数量
    def get_product_bak_B_count(self):
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        payload = {
            "pageNumber"            : 1,
            "pageSize"              : 10,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [
                62,
                227,
                12,
                230,
                80,
                58,
                224
            ],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : None,
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": None,
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2,
            "goodsLevelFakeIdList"  : [
                4
            ]
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        count = info.get('count', 0)
        log('获取已上架备货款B数量', count)
        return count

    # 已上架新款A数量
    def get_product_A_count(self):
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        payload = {
            "pageNumber"            : 1,
            "pageSize"              : 10,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [
                107
            ],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : None,
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": None,
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2,
            "goodsLevelFakeIdList"  : [
                2
            ]
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        count = info.get('count', 0)
        log('获取已上架新款A数量', count)
        return count

    # 本周已上架数量
    def get_week_shelf_product_count(self, start_date, end_date):
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        payload = {
            "pageNumber"            : 1,
            "pageSize"              : 10,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : start_date,
            "shelfDateEnd"          : end_date,
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        count = info.get('count', 0)
        log('获取本周上架数量')
        return count

    # 已上架数量
    def get_shelf_product_count(self):
        log('获取所有已上架数量')
        url = "https://sso.geiwohuo.com/spmp-api-prefix/spmp/product/list?page_num=1&page_size=10"
        payload = {
            "language"              : "zh-cn",
            "only_recommend_resell" : False,
            "only_spmb_copy_product": False,
            "search_abandon_product": False,
            "shelf_type"            : "ON_SHELF",
            "sort_type"             : 1
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info')
        customObj = info.get('meta').get('customObj')
        # 将数据转换成字典
        result = {item["shelf_status"]: item["count"] for item in customObj}
        return result

    def get_yesterday_upload_product_count(self, dt=None):
        url = "https://sso.geiwohuo.com/spmp-api-prefix/spmp/product/publish/record/page_list?page_num=1&page_size=100"
        payload = {
            "edit_type"                   : 0,
            "language"                    : "zh-cn",
            "only_current_month_recommend": False,
            "only_spmb_copy_product"      : False,
            "query_time_out"              : False,
            "search_diy_custom"           : False
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')
        count = 0
        for item in info.get('data', {}):
            if TimeUtils.is_yesterday(item['create_time'], dt):
                count += 1
        log('获取昨日已上传数量', count)
        return count

    def get_week_sales_stat_detail(self):
        dt = self.get_dt_time()
        yesterday = TimeUtils.get_yesterday(dt)
        date_7_days_ago = TimeUtils.get_past_nth_day(6, None, '%Y-%m-%d')
        log('-7', date_7_days_ago)
        date_1_days_ago = TimeUtils.get_past_nth_day(1, None, '%Y-%m-%d')
        log('-1', date_1_days_ago)

        url = "https://sso.geiwohuo.com/sbn/index/get_critical_indicator_curve_chart"
        payload = {
            "areaCd"     : "cn",
            "dt"         : dt,
            "countrySite": [
                "shein-all"
            ],
            "startDate"  : date_7_days_ago,
            "endDate"    : date_1_days_ago,
            "queryType"  : 1,
            "pageNum"    : 1,
            "pageSize"   : 100
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info', {})

        last_item = SheinStoreSalesDetailManager(self.config.database_url).get_records_as_dict(self.store_username, yesterday)
        log('last_item', last_item)
        day_item = info[-1]
        log(day_item)
        item = {}
        item["store_username"] = self.store_username
        item["store_name"] = self.store_name
        item["day"] = day_item["dataDate"]
        item["sales_num"] = day_item["saleCnt1d"] or 0
        item['sales_num_inc'] = item['sales_num'] - last_item.get('sales_num', 0)

        if int(self.user_info.get('lv1CategoryId')) == 216506:  # 自运营POP店
            log('gmv1d', day_item['gmv1d'])
            item['sales_amount'] = day_item['gmv1d'] if isinstance(day_item['gmv1d'], (int, float)) else 0
        else:
            item['sales_amount'] = day_item['dealAmt1d'] or 0

        log('sales_amount', item['sales_amount'])
        item['sales_amount_inc'] = item['sales_amount'] - float(last_item.get('sales_amount', 0))
        item['visitor_num'] = day_item['idxShopGoodsUv1d'] or 0
        item['visitor_num_inc'] = item['visitor_num'] - last_item.get('visitor_num', 0)
        item['bak_A_num'] = self.get_product_bak_A_count()
        item['bak_A_num_inc'] = item['bak_A_num'] - last_item.get('bak_A_num', 0)
        item['new_A_num'] = self.get_product_A_count()
        item['new_A_num_inc'] = item['new_A_num'] - last_item.get('new_A_num', 0)
        dictProduct = self.get_shelf_product_count()
        item['on_sales_product_num'] = dictProduct.get('ON_SHELF')
        item['on_sales_product_num_inc'] = item['on_sales_product_num'] - last_item.get('on_sales_product_num', 0)
        item['wait_shelf_product_num'] = dictProduct.get('WAIT_SHELF')
        item['wait_shelf_product_num_inc'] = item['wait_shelf_product_num'] - last_item.get('wait_shelf_product_num', 0)
        item['upload_product_num'] = self.get_yesterday_upload_product_count()
        item['upload_product_num_inc'] = item['upload_product_num'] - last_item.get('upload_product_num', 0)
        item['sold_out_product_num'] = dictProduct.get('SOLD_OUT')
        item['shelf_off_product_num'] = dictProduct.get('OUT_SHELF')

        SheinStoreSalesDetailManager(self.config.database_url).insert_data([item])

    def get_delivery_order_list(self, orderType=2):
        page_num = 1
        page_size = 200

        url = f"https://sso.geiwohuo.com/pfmp/order/list"
        payload = {}
        if orderType == 1:
            payload = {
                "orderType": orderType,
                "page"     : page_num,
                "perPage"  : page_size,
                "status"   : [2],
            }
        elif orderType == 2:
            payload = {
                "orderType" : orderType,
                "page"      : page_num,
                "perPage"   : page_size,
                "status"    : [2],
                "isJitOrder": 2
            }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        skc_list = [item['goods']['skcName'] for item in spu_list]
        self.get_activity_label(skc_list)

        for page in range(2, totalPage + 1):
            log(f'获取订单列表 第{page}/{totalPage}页')
            page_num = page
            if orderType == 1:
                payload = {
                    "orderType": orderType,
                    "page"     : page_num,
                    "perPage"  : page_size,
                    "status"   : [2],
                }
            elif orderType == 2:
                payload = {
                    "orderType" : orderType,
                    "page"      : page_num,
                    "perPage"   : page_size,
                    "status"    : [2],
                    "isJitOrder": 2
                }
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            skc_list = [item['goods']['skcName'] for item in spu_list_new]
            self.get_activity_label(skc_list)
            spu_list += spu_list_new
            time.sleep(0.3)

        if len(spu_list) == 0:
            log(f'无{["", "急采", "备货"][orderType]}发货单')
            return None

        write_to_excel = [
            # 0          1        2          3          4          5          6     7
            ['店铺名称', 'SKC图片', 'SKU图片', '商品信息', '下单/需求数量', '库存(模式/本地/在途/希音)', '成本价', '核价',
             '近7天SKU销量/SKC销量/SKC曝光', 'SKC点击率/SKC转化率', '自主参与活动', '最晚预约上门取件', '要求实际完成取件',
             'SKC', 'SKU']
        ]
        cache_file2 = f'{self.config.auto_dir}/shein/dict/skc_shelf_{self.store_username}.json'
        DictSkcShelf = read_dict_from_file(cache_file2)
        cache_file3 = f'{self.config.auto_dir}/shein/dict/skc_product_{self.store_username}.json'
        DictSkcProduct = read_dict_from_file(cache_file3)
        cache_file = f'{self.config.auto_dir}/shein/dict/activity_price_{self.store_name}.json'
        dictActivityPrice = read_dict_from_file(cache_file)
        cache_file4 = f'{self.config.auto_dir}/shein/dict/dict_sku_info_{self.store_username}.json'
        DictSkuInfo = read_dict_from_file(cache_file4)
        for spu_item in spu_list:
            skc = spu_item['goods']['skcName']
            skcCode = spu_item['goods']['supplierCode']
            skc_img = str(spu_item['goods']['imgPath'])
            orderNum = spu_item['sellerOrderNo']
            orderTime = spu_item['allocateTime']
            requestTakeParcelTime = spu_item['requestTakeParcelTime']
            suggestedReserveTime = spu_item['suggestedReserveTime']
            good_level = spu_item['goods']['goodsLevelName']

            self.get_skc_week_actual_sales(skc)

            spu = DictSkcProduct[skc]['spu_name']
            log('spu', spu)
            for sku_item in spu_item['detail']:
                needQuantity = sku_item['needQuantity']
                orderQuantity = sku_item['orderQuantity']
                # sku_img = sku_item['skuThumb']
                skuCode = sku_item['supplierSku']
                stock = self.bridge.get_sku_stock(skuCode, 'mb')
                cost_price = self.bridge.get_sku_cost(skuCode, 'mb')
                suffixZh = sku_item['suffixZh']
                sku = sku_item['skuCode']
                supplyPrice = sku_item['supplyPrice']
                sku_img = self.bridge.get_sku_img(skuCode, 'mb')
                sale_model = DictSkuInfo[skuCode][0]
                shein_stock = DictSkuInfo[skuCode][1]
                shelf_days = DictSkuInfo[skuCode][2]
                real_transit = DictSkuInfo[skuCode][3]
                stock_str = f'{sale_model}\n{stock}/{real_transit}/{shein_stock}'

                item = []
                item.append(f'{self.store_name}\n{good_level}')
                item.append(skc_img)
                item.append(sku_img)
                if cost_price == '-':
                    profit = '-'
                else:
                    profit = f'{float(supplyPrice) - float(cost_price):.2f}'
                # item.append(f'订单号: {orderNum}\n下单时间: {orderTime}\n最晚预约上门取件: {suggestedReserveTime}\nSKC: {skc}\nSKC货号: {skcCode}\nSKU货号: {skuCode}\n属性集: {suffixZh}\n上架时间: {DictSkcShelf[skc]}\n上架天数: {shelf_days}\n成本/核价/利润: ¥{cost_price}/¥{supplyPrice}/¥{profit}\n下单/需求数量: {orderQuantity}/{needQuantity}\n库存模式/本地/在途/希音: {sale_model}/{stock}/{real_transit}/{shein_stock}\n')
                item.append(f'订单号: {orderNum}\n下单时间: {orderTime}\n最晚预约上门取件: {suggestedReserveTime}\nSKC: {skc}\nSKC货号: {skcCode}\nSKU货号: {skuCode}\n属性集: {suffixZh}\n上架时间: {DictSkcShelf[skc]}\n上架天数: {shelf_days}\n成本/核价/利润: ¥{cost_price}/¥{supplyPrice}/¥{profit}')
                item.append(f'[{orderQuantity}/{needQuantity}]')
                item.append(stock_str)
                item.append(cost_price)
                item.append(supplyPrice)
                sale_num_list, sale_data_list = self.get_sku_week_sale_list(spu, skc, sku)
                item.append("\n".join(sale_num_list))
                item.append("\n".join(sale_data_list))
                item.append(self.get_skc_activity_label(skc, sku, dictActivityPrice))
                item.append(suggestedReserveTime)
                item.append(requestTakeParcelTime)
                item.append(skc)
                item.append(sku)
                write_to_excel.append(item)

        cache_file = f'{self.config.auto_dir}/shein/cache/jit_{TimeUtils.today_date()}_{orderType}_{TimeUtils.get_period()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: write_to_excel}, {self.store_username})

        return write_to_excel

    # 获取商品包含sku销量的列表
    def get_dict_sku_stock_detail(self):
        log(f'获取备货信息商品列表 做成字典')
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        pageNumber = 1
        pageSize = 100
        dictPayload = {
            "pageNumber"            : pageNumber,
            "pageSize"              : pageSize,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [],
            "supplyStatus"          : "",
            "shelfStatus"           : "",
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2
        }
        payload = dictPayload
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['list']

        total = response_text['info']['count']
        totalPage = math.ceil(total / pageSize)
        for page in range(2, totalPage + 1):
            log(f'获取备货信息商品列表 第{page}/{totalPage}页')
            dictPayload['pageNumber'] = page
            payload = dictPayload
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['list']
            spu_list += spu_list_new
            time.sleep(0.3)

        DictSkuInfo = {}
        for spu_info in spu_list:
            sale_model = spu_info.get('saleModel', {}).get('name') if spu_info.get('saleModel') else '-'
            shelfDays = spu_info['shelfDays']
            for sku_info in spu_info['skuList']:
                attr = sku_info['attr']
                if attr == '合计':
                    continue
                skuExtCode = str(sku_info['supplierSku'])
                shein_stock = sku_info['stock']

                transit = sku_info['transit']  # 在途
                real_transit = transit + sku_info['stayShelf'] - sku_info['transitSale']

                DictSkuInfo[skuExtCode] = [sale_model, shein_stock, shelfDays, real_transit]

        write_dict_to_file(f'{self.config.auto_dir}/shein/dict/dict_sku_info_{self.store_username}.json', DictSkuInfo)

        return DictSkuInfo

    def get_shop_notify_num(self):
        log(f'正在获取 {self.store_name} 通知数据')
        url = "https://sso.geiwohuo.com/sso/homePage/v4/detail"
        payload = {
            "metaIndexIds": [
                246,  # 急采-待发货
                245  # 备货-待发货
            ],
            "templateType": 0
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        info = response_text.get('info')

        cache_file = f'{self.config.auto_dir}/shein/notify/{self.store_name}_{TimeUtils.get_current_datetime()}.json'
        write_dict_to_file(cache_file, info)

        num245 = 0
        num246 = 0
        for item in info['list']:
            if item['metaIndexId'] == 245:
                num245 = item['count']
            if item['metaIndexId'] == 246:
                num246 = item['count']

        NotifyItem = [self.store_name, num246, num245]

        cache_file = f'{self.config.auto_dir}/shein/cache/jit_notify_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_username: NotifyItem}, {self.store_username})

        return info

    def get_activity_list(self):
        url = "https://sso.geiwohuo.com/mrs-api-prefix/mbrs/activity/get_activity_list?page_num=1&page_size=100"
        payload = {}
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        total = response_text.get('info', {}).get('total_count')
        excel_data = [[
            '店铺名称', '活动信息', '已报数量', '可报数量'
        ]]
        if total > 0:
            for item in response_text.get('info', {}).get('activity_detail_list'):
                activity_tag = item.get('text_tag_content')
                activity_name = item['activity_name']
                start_time = item['activity_start_zone_time']
                end_time = item['activity_end_zone_time']
                start_time2 = item['start_zone_time']
                end_time2 = item['end_zone_time']
                allow_goods_num = item.get('allow_goods_num')
                apply_goods_num = item.get('apply_goods_num')
                row_item = [
                    self.store_name,
                    f"活动名称: 【{activity_tag}】{activity_name}\n报名时间: {start_time}~{end_time}\n活动时间: {start_time2}~{end_time2}\n已报数量: {apply_goods_num}/{allow_goods_num}",
                    apply_goods_num,
                    allow_goods_num,
                ]
                excel_data.append(row_item)
                cache_file = f'{self.config.auto_dir}/shein/activity_list/activity_list_{TimeUtils.today_date()}.json'
                write_dict_to_file_ex(cache_file, {self.store_username: excel_data}, [self.store_username])

    def get_product_list(self):
        # self.web_page.goto('https://sso.geiwohuo.com/#/spmp/commdities/list')
        # self.web_page.wait_for_load_state("load")

        cache_file = f'{self.config.auto_dir}/shein/dict/product_list_{self.store_username}.json'
        DictSpuInfo = read_dict_from_file(cache_file, 3600)
        if len(DictSpuInfo) > 0:
            return DictSpuInfo

        page_num = 1
        page_size = 100

        url = f"https://sso.geiwohuo.com/spmp-api-prefix/spmp/product/list?page_num={page_num}&page_size={page_size}"
        payload = {
            "language"              : "zh-cn",
            "only_recommend_resell" : False,
            "only_spmb_copy_product": False,
            "search_abandon_product": False,
            "sort_type"             : 1
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page_num in range(2, totalPage + 1):
            log(f'获取商品列表 第{page_num}/{totalPage}页')
            url = f"https://sso.geiwohuo.com/spmp-api-prefix/spmp/product/list?page_num={page_num}&page_size={page_size}"
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            spu_list += spu_list_new
            time.sleep(0.3)

        DictSkcShelf = {}
        DictSkcProduct = {}
        DictSpuInfo = {}
        for spu_item in spu_list:
            spu = spu_item['spu_name']
            first_shelf_time = spu_item['first_shelf_time']
            for skc_item in spu_item['skc_info_list']:
                skc_name = skc_item['skc_name']
                DictSkcShelf[skc_name] = first_shelf_time
                DictSkcProduct[skc_name] = spu_item
            DictSpuInfo[spu] = spu_item

        cache_file2 = f'{self.config.auto_dir}/shein/dict/skc_shelf_{self.store_username}.json'
        write_dict_to_file(cache_file2, DictSkcShelf)
        cache_file3 = f'{self.config.auto_dir}/shein/dict/skc_product_{self.store_username}.json'
        write_dict_to_file(cache_file3, DictSkcProduct)

        write_dict_to_file(cache_file, DictSpuInfo)
        return DictSpuInfo

    def query_obm_activity_list(self):
        page_num = 1
        page_size = 100
        date_60_days_ago = TimeUtils.get_past_nth_day(59)
        cache_file = f'{self.config.auto_dir}/shein/cache/obm_activity_{self.store_name}_{date_60_days_ago}_{TimeUtils.today_date()}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 8)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mrs-api-prefix/promotion/obm/query_obm_activity_list"
        payload = {
            "insert_end_time"  : f"{TimeUtils.today_date()} 23:59:59",
            "insert_start_time": f"{date_60_days_ago} 00:00:00",
            "page_num"         : page_num,
            "page_size"        : page_size,
            "system"           : "mrs",
            "time_zone"        : "Asia/Shanghai",
            # "state": 3, # 活动开启中 不能用这个条件
            "type_id"          : 31  # 限时折扣
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取营销工具列表 第{page}/{totalPage}页')
            payload["page_num"] = page
            response_text = fetch_shein(self.web_page, url, payload)
            list_item += response_text['info']['data']
            time.sleep(0.1)

        write_dict_to_file(cache_file, list_item)
        return list_item

    def query_goods_detail(self, activity_id):
        # web_page.goto(f'https://sso.geiwohuo.com/#/mrs/tools/activity/obm-time-limit-info/{activity_id}')
        # web_page.wait_for_load_state('load')
        log(f'正在获取 {self.store_name} {activity_id} 营销工具商品详情')

        cache_file = f'{self.config.auto_dir}/shein/cache/query_goods_detail_{activity_id}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 8)
        if len(list_item) > 0:
            return list_item

        page_num = 1
        page_size = 100
        url = "https://sso.geiwohuo.com/mrs-api-prefix/promotion/simple_platform/query_goods_detail"
        payload = {
            "activity_id": activity_id,
            "page_num"   : page_num,
            "page_size"  : page_size
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取营销工具商品列表 第{page}/{totalPage}页')
            payload["page_num"] = page
            response_text = fetch_shein(self.web_page, url, payload)
            list_item += response_text['info']['data']
            time.sleep(0.1)

        log(list_item)
        write_dict_to_file(cache_file, list_item)
        return list_item

    def get_partake_activity_detail(self, activity_id, skc):
        log(f'正在获取营销活动报名记录详情 {self.store_name}')
        page_num = 1
        page_size = 100
        cache_file = f'{self.config.auto_dir}/shein/cache/platform_activity_{activity_id}_{skc}.json'
        list_item = read_dict_from_file(cache_file)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mrs-api-prefix/mbrs/activity/get_partake_activity_goods_list?page_num={page_num}&page_size={page_size}"
        payload = {
            "goods_audit_status": 1,
            "activity_id_list"  : [activity_id],
            "skc_list"          : [skc]
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']

        log(list_item)
        write_dict_to_file(cache_file, list_item)
        return list_item

    def get_partake_activity_goods_list(self):
        # self.web_page.goto(f'https://sso.geiwohuo.com/#/mbrs/marketing/list/1')
        # self.web_page.wait_for_load_state('load')
        log(f'正在获取 {self.store_name} 活动列表')
        page_num = 1
        page_size = 100
        date_60_days_ago = TimeUtils.get_past_nth_day(59)
        cache_file = f'{self.config.auto_dir}/shein/cache/platform_activity_{self.store_name}_{date_60_days_ago}_{TimeUtils.today_date()}.json'
        list_item = read_dict_from_file(cache_file, 3600 * 8)
        if len(list_item) > 0:
            return list_item

        url = f"https://sso.geiwohuo.com/mrs-api-prefix/mbrs/activity/get_partake_activity_goods_list?page_num={page_num}&page_size={page_size}"
        payload = {
            "goods_audit_status"    : 1,
            "insert_zone_time_end"  : f"{TimeUtils.today_date()} 23:59:59",
            "insert_zone_time_start": f"{date_60_days_ago} 00:00:00"
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取活动列表 第{page}/{totalPage}页')
            payload["page_num"] = page
            response_text = fetch_shein(self.web_page, url, payload)
            list_item += response_text['info']['data']
            time.sleep(0.1)

        write_dict_to_file(cache_file, list_item)
        return list_item

    def generate_activity_price_dict(self):
        cache_file = f'{self.config.auto_dir}/shein/dict/activity_price_{self.store_name}.json'
        dict_activity_price = {}
        activity_list = self.query_obm_activity_list()
        for activity in activity_list:
            activity_id = activity['activity_id']
            activity_name = activity['act_name']
            sub_type_id = activity['sub_type_id']  # 1.不限量 2.限量
            dateBegin = TimeUtils.convert_datetime_to_date(activity['start_time'])
            dateEnd = TimeUtils.convert_datetime_to_date(activity['end_time'])
            skc_list = self.query_goods_detail(activity_id)
            for skc_item in skc_list:
                attend_num_sum = skc_item['attend_num_sum']
                product_act_price = skc_item['product_act_price']  # 活动价
                if sub_type_id == 1:
                    attend_num_sum = '不限量'
                for sku_item in skc_item['sku_info_list']:
                    sku = sku_item['sku']  # 平台sku
                    product_act_price = sku_item['product_act_price'] if sku_item[
                        'product_act_price'] else product_act_price  # 活动价
                    key = f'{sku}_{dateBegin}_{dateEnd}_{activity_name}'
                    dict_activity_price[key] = [product_act_price, attend_num_sum]

        platform_activity_list = self.get_partake_activity_goods_list()
        for platform_activity in platform_activity_list:
            activity_name = platform_activity['activity_name']
            text_tag_content = platform_activity['text_tag_content']
            attend_num = platform_activity['attend_num']
            dateBegin = TimeUtils.convert_timestamp_to_date(platform_activity['start_time'])
            dateEnd = TimeUtils.convert_timestamp_to_date(platform_activity['end_time'])
            if text_tag_content != '新品':
                attend_num = '-'
            for sku_item in platform_activity['activity_sku_list']:
                sku = sku_item['sku_code']
                enroll_price = sku_item['enroll_display_str'][:-3]
                key = f'{sku}_{dateBegin}_{dateEnd}_{activity_name}'
                dict_activity_price[key] = [enroll_price, attend_num]

        write_dict_to_file(cache_file, dict_activity_price)

    def get_skc_actual_sales_dict(self, skc, first_day, last_day):
        cache_file = f'{self.config.auto_dir}/shein/cache/actual_sales_{skc}_{first_day}_{last_day}.json'
        if datetime.now().hour >= 9:
            DictSkuSales = read_dict_from_file(cache_file)
        else:
            DictSkuSales = read_dict_from_file(cache_file, 1800)
        if len(DictSkuSales) > 0:
            return DictSkuSales

        url = f"https://sso.geiwohuo.com/idms/sale-trend/detail"
        payload = {
            "skc"      : skc,
            "startDate": first_day,
            "endDate"  : last_day,
            "daysToAdd": 0
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            log(response_text)
            return {}
        info = response_text['info']
        for sale_item in info['actualSalesVolumeMap']:
            sku = sale_item['skuCode']
            if sku is not None:
                DictSkuSales[sku] = sale_item['actualSalesVolume']

        write_dict_to_file(cache_file, DictSkuSales)
        return DictSkuSales

    def get_skc_week_actual_sales(self, skc):
        first_day, last_day = TimeUtils.get_past_7_days_range()
        cache_file = f'{self.config.auto_dir}/shein/cache/{skc}_{first_day}_{last_day}.json'
        if datetime.now().hour >= 9:
            DictSkuSalesDate = read_dict_from_file(cache_file)
        else:
            DictSkuSalesDate = read_dict_from_file(cache_file, 1800)
        if len(DictSkuSalesDate) > 0:
            return DictSkuSalesDate

        url = f"https://sso.geiwohuo.com/idms/sale-trend/detail"
        payload = {
            "skc"      : skc,
            "startDate": first_day,
            "endDate"  : last_day,
            "daysToAdd": 0
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            log(response_text)
            return {}
        list_item = response_text['info']['salesVolumeDateVoList']
        for item in list_item:
            key = item['date']
            DictSkuSalesDate[key] = item['salesVolumeMap']
        list_item2 = response_text['info']['actualSalesVolumeMap']
        for item in list_item2:
            sku = item['skuCode']
            if sku is not None:
                DictSkuSalesDate[sku] = item['actualSalesVolume']

        write_dict_to_file(cache_file, DictSkuSalesDate)
        return DictSkuSalesDate

    def get_preemption_list(self, skc_list):
        url = f"https://sso.geiwohuo.com/idms/goods-skc/preemption-num"
        payload = skc_list
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        dict = response_text['info']

        cache_file = f'{self.config.auto_dir}/shein/preemption_num/preemption_num_{self.store_username}.json'
        dict_preemption_num = read_dict_from_file(cache_file)
        dict_preemption_num.update(dict)
        write_dict_to_file(cache_file, dict_preemption_num)

        return dict

    def get_quality_label(self, skc_list):
        url = f"https://sso.geiwohuo.com/idms/goods-skc/quality-label"
        payload = skc_list
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        dict = response_text['info']

        cache_file = f'{self.config.auto_dir}/shein/quality_label/quality_label_{self.store_username}.json'
        dict_label = read_dict_from_file(cache_file)
        dict_label.update(dict)
        write_dict_to_file(cache_file, dict_label)

        return dict

    def get_activity_label(self, skc_list):
        url = f"https://sso.geiwohuo.com/idms/goods-skc/activity-label"
        payload = skc_list
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        dict = response_text['info']

        cache_file = f'{self.config.auto_dir}/shein/activity_label/activity_label_{self.store_username}.json'
        dict_label = read_dict_from_file(cache_file)
        dict_label.update(dict)
        write_dict_to_file(cache_file, dict_label)

        return dict

    def get_sku_price_pop(self, spu):
        pass
        log(f'获取pop sku价格列表', spu)
        info = self.get_product_detail(spu)

        dict_sku_price_new = {}
        for skc_item in info['skc_list']:
            for sku_item in skc_item['sku_list']:
                sku = sku_item['sku_code']
                special_price = sku_item['price_info_list'][0]['special_price']
                dict_sku_price_new[sku] = special_price

        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku_price = read_dict_from_file(cache_file)
        dict_sku_price.update(dict_sku_price_new)
        write_dict_to_file(cache_file, dict_sku_price)

    def get_sku_price_v2(self, skc_list):
        log(f'获取sku价格列表', skc_list)
        url = "https://sso.geiwohuo.com/idms/goods-skc/price"
        response_text = fetch_shein(self.web_page, url, skc_list)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        dict = response_text['info']

        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku_price = read_dict_from_file(cache_file)
        dict_sku_price.update(dict)
        write_dict_to_file(cache_file, dict_sku_price)

        return dict

    def get_stock_advice(self, skc_list):
        log(f'获取sku库存建议列表', skc_list)
        url = f"https://sso.geiwohuo.com/idms/goods-skc/get-vmi-spot-advice"
        payload = skc_list
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        dict = response_text['info']

        cache_file = f'{self.config.auto_dir}/shein/vmi_spot_advice/spot_advice_{self.store_username}.json'
        dict_advice = read_dict_from_file(cache_file)
        dict_advice.update(dict)
        write_dict_to_file(cache_file, dict_advice)

        return dict

    def get_dt_time(self):
        if self.dt is not None:
            log(f'字典dt: {self.dt}')
            return self.dt
        log('获取非实时更新时间')
        url = "https://sso.geiwohuo.com/sbn/common/get_update_time"
        payload = {
            "pageCode": "Index",
            "areaCd"  : "cn"
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        self.dt = response_text.get('info').get('dt')
        log(f'dt: {self.dt}')
        return self.dt

    def get_dt_time_goods(self):
        if self.dt_goods is not None:
            log(f'字典dt_goods: {self.dt_goods}')
            return self.dt_goods
        log('获取非实时更新时间')
        url = "https://sso.geiwohuo.com/sbn/common/get_update_time"
        payload = {
            "pageCode": "GoodsPreviewNew",
            "areaCd"  : "cn"
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        self.dt_goods = response_text.get('info').get('dt')
        log(f'接口dt_goods: {self.dt_goods}')
        return self.dt_goods

    def get_dict_skc_week_trend(self):
        page_num = 1
        page_size = 100

        date_7_days_ago = TimeUtils.get_past_nth_day(7, None, '%Y%m%d')
        log('-7', date_7_days_ago)
        date_1_days_ago = TimeUtils.get_past_nth_day(1, None, '%Y%m%d')
        log('-1', date_1_days_ago)

        url = f"https://sso.geiwohuo.com/sbn/new_goods/get_skc_diagnose_list"
        payload = {
            "areaCd"     : "cn",
            "countrySite": [
                "shein-all"
            ],
            "startDate"  : date_7_days_ago,
            "endDate"    : date_1_days_ago,
            "pageNum"    : page_num,
            "pageSize"   : page_size
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        spu_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取商品列表 第{page}/{totalPage}页')
            page_num = page
            payload = {
                "areaCd"     : "cn",
                "countrySite": [
                    "shein-all"
                ],
                "startDate"  : date_7_days_ago,
                "endDate"    : date_1_days_ago,
                "pageNum"    : page_num,
                "pageSize"   : page_size
            }
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            spu_list += spu_list_new
            time.sleep(0.3)

        DictSkcWeekTrend = {}
        for spu_item in spu_list:
            skc = str(spu_item['skc'])
            DictSkcWeekTrend[skc] = spu_item

        log('len(DictSkcWeekTrend)', len(DictSkcWeekTrend))
        write_dict_to_file(f'{self.config.auto_dir}/shein/dict/dict_skc_week_trend_{self.store_username}.json', DictSkcWeekTrend)
        return DictSkcWeekTrend

    def get_skc_sales(self, skc, start_date, end_date):
        url = "https://sso.geiwohuo.com/idms/stockadvice/saleTrendDetail"
        payload = {
            "skc"      : skc,
            "startDate": start_date,
            "endDate"  : end_date
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        error_msg = response_text.get('msg')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        sales_list = response_text['info']['saleTrendDetailList']
        if sales_list:
            for skc_list in sales_list:
                date = skc_list['date']
                if date == '合计':
                    log('无销量skc: ', skc)
                    continue
                skc_sale = skc_list['skcSale']
                skc_order = skc_list['skcOrder']
                for sku_list in skc_list['skuSaleTrendDetailList']:
                    sku = sku_list['skuCode']
                    attr_name = sku_list['attributeName']
                    sku_sale = sku_list['skuSale']
                    sku_order = sku_list['skuOrder']
                    if sku_sale > 0:
                        insert_sales(skc, date, skc_sale, skc_order, sku, attr_name, sku_sale, sku_order)
        return sales_list

    # 获取商品包含sku销量的列表
    def get_product_sku_sales_list(self, source='mb'):
        log(f'获取销量列表')
        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        pageNumber = 1
        pageSize = 100
        dictPayload = {
            "pageNumber"            : pageNumber,
            "pageSize"              : pageSize,
            "supplierCodes"         : "",
            "skcs"                  : "",
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [],
            "supplyStatus"          : "",
            "shelfStatus"           : 1,
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2
        }
        payload = dictPayload
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['list']

        skc_list = [item['skc'] for item in spu_list]
        self.get_activity_label(skc_list)
        self.get_preemption_list(skc_list)
        self.get_sku_price_v2(skc_list)
        self.get_stock_advice(skc_list)

        total = response_text['info']['count']
        totalPage = math.ceil(total / pageSize)
        for page in range(2, totalPage + 1):
            log(f'获取SKU销量列表 第{page}/{totalPage}页')
            dictPayload['pageNumber'] = page
            payload = dictPayload
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['list']

            skc_list = [item['skc'] for item in spu_list_new]
            self.get_activity_label(skc_list)
            self.get_preemption_list(skc_list)
            self.get_sku_price_v2(skc_list)
            self.get_stock_advice(skc_list)

            spu_list += spu_list_new
            time.sleep(0.3)

        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku_price = read_dict_from_file(cache_file)

        cache_file2 = f'{self.config.auto_dir}/shein/dict/dict_skc_week_trend_{self.store_username}.json'
        DictSkcWeekTrend = read_dict_from_file(cache_file2)

        cache_file3 = f'{self.config.auto_dir}/shein/dict/product_list_{self.store_username}.json'
        DictSpuInfo = read_dict_from_file(cache_file3)

        cache_file = f'{self.config.auto_dir}/shein/dict/activity_price_{self.store_name}.json'
        dictActivityPrice = read_dict_from_file(cache_file)

        product_sku_list = [
            [
                '店铺名称', '商品信息', 'SKC图片', 'SKU图片', 'SKU', 'SKU货号', '在售天数', '库存(模式/本地/在途/希音)',
                '今天销量', '今日订单数',  # 9
                '远7天销量', '远7天订单数', '近7天销量', '近7天订单数', '周销增量', '远30天销量', '远30天订单数',
                '近30天销量', '近30天订单数', '月销增量', '总销量',  # 11
                '申报价', '成本价', '毛利润', '毛利率', '近7天利润', '近30天利润',  # 6
                'SPU', 'SKC', 'SKC货号', '近7天SKU销量/SKC销量/SKC曝光', 'SKC点击率/SKC转化率', '自主参与活动', '商品标题', '叶子类目',  # 5
                'SKC近7天销量', 'SKC近7天曝光人数', 'SKC近7天商详访客', 'SKC近7天点击率', 'SKC近7天支付人数',
                'SKC近7天支付率', 'SKC近7天评论数'
            ]
        ]

        date_60_days_ago = TimeUtils.get_past_nth_day(60, None, '%Y-%m-%d')
        log('-60', date_60_days_ago)
        date_7_days_ago = TimeUtils.get_past_nth_day(7, None, '%Y-%m-%d')
        log('-7', date_7_days_ago)
        date_1_days_ago = TimeUtils.get_past_nth_day(1, None, '%Y-%m-%d')
        log('-1', date_1_days_ago)

        count = 0
        for spu_info in spu_list:
            count += 1
            # if count > 10:
            #     break
            spu = spu_info['spu']
            skc = str(spu_info['skc'])
            # if not shein_db.exists_sales_1_days_ago(skc):
            #     log(f'未查到昨天销量: {skc}')
            self.get_skc_week_actual_sales(skc)
            self.get_skc_sales(skc, date_60_days_ago, date_1_days_ago)
            skcCode = spu_info['supplierCode']
            product_name = DictSpuInfo[spu]['product_name_en']
            category_name = spu_info['categoryName']
            shelfDays = spu_info['shelfDays']
            shelf_status = DictSpuInfo[spu]['shelf_status']
            dictStatus = {
                'WAIT_SHELF': "待上架",
                'ON_SHELF'  : "已上架",
                'SOLD_OUT'  : "已售罄",
                'OUT_SHELF' : "已下架"
            }
            status_cn = dictStatus[shelf_status]
            good_level = spu_info['goodsLevel']['name']
            sale_model = spu_info['saleModel']['name']

            # 过滤已经售罄
            if shelf_status == 'SOLD_OUT':
                log(f'过滤已售罄: {skc} {category_name} {product_name}')
                continue

            for sku_info in spu_info['skuList']:
                sku = sku_info['skuCode']
                skuExtCode = str(sku_info['supplierSku'])
                if sku == '合计':
                    continue

                # 获取基础数据
                stock = self.bridge.get_sku_stock(skuExtCode, source)
                cost_price = self.bridge.get_sku_cost(skuExtCode, source)
                sku_img = self.bridge.get_sku_img(skuExtCode, source)

                # 计算库存相关数据
                shein_stock = sku_info['stock']
                transit = sku_info['transit']  # 在途
                real_transit = transit + sku_info['stayShelf'] - sku_info['transitSale']

                # 获取销量数据
                week_sales = get_last_week_sales(sku)
                week_sales2 = get_near_week_sales(sku)
                month_sales = get_last_month_sales(sku)
                month_sales2 = get_near_month_sales(sku)

                # 获取SKC趋势数据
                key = str(skc)
                skc_trend = DictSkcWeekTrend.get(key, {})

                # 使用append组装数据
                sku_item = []

                # 店铺名称
                sku_item.append(f'{self.store_name}\n({status_cn})\n{good_level}\n{date_7_days_ago}\n{date_1_days_ago}')

                # 商品信息
                product_info = f"SPU: {spu}\nSKC: {skc}\nSKC货号: {skcCode}\n类目: {category_name}\n在售天数: {shelfDays}"
                sku_item.append(product_info)

                # SKC图片
                sku_item.append(spu_info['picUrl'])

                # SKU图片
                sku_item.append(sku_img)

                # SKU基本信息
                sku_item.append(sku)  # SKU
                sku_item.append(f"{sku_info['supplierSku']}")  # SKU货号
                sku_item.append(shelfDays)  # 在售天数

                # 库存信息
                sku_item.append(f'{sale_model}\n{stock}/{real_transit}/{shein_stock}')

                # 今日销量数据
                sku_item.append(sku_info['totalSaleVolume'])  # 今日销量
                sku_item.append(sku_info['orderCnt'])  # 今日订单数

                # 远7天销量数据
                sku_item.append(week_sales[0])  # 远7日销量
                sku_item.append(week_sales[1])  # 远7日订单数

                # 近7天销量数据
                sku_item.append(week_sales2[0])  # 近7日销量
                sku_item.append(week_sales2[1])  # 近7日订单数
                sku_item.append(week_sales2[1] - week_sales2[0])  # 周增销量

                # 远30天销量数据
                sku_item.append(month_sales[0])  # 远30日销量
                sku_item.append(month_sales[1])  # 远30日订单数

                # 近30天销量数据
                sku_item.append(month_sales2[0])  # 近30日销量
                sku_item.append(month_sales2[1])  # 近30日订单数
                sku_item.append(month_sales2[1] - month_sales2[0])  # 月增销量

                # 总销量
                sku_item.append('-')

                # 价格相关
                sku_item.append(dict_sku_price[sku])  # 申报价
                sku_item.append(cost_price)  # 成本价
                sku_item.append('')  # 毛利润
                sku_item.append('')  # 毛利率
                sku_item.append('')  # 近7天利润
                sku_item.append('')  # 近30天利润

                # 商品标识
                sku_item.append(spu)  # SPU
                sku_item.append(skc)  # SKC
                sku_item.append(spu_info['supplierCode'])  # SKC货号

                sale_num_list, sale_data_list = self.get_sku_week_sale_list(spu, skc, sku)
                sku_item.append("\n".join(sale_num_list))
                sku_item.append("\n".join(sale_data_list))
                sku_item.append(self.get_skc_activity_label(skc, sku, dictActivityPrice))

                sku_item.append(product_name)  # 商品标题
                sku_item.append(category_name)  # 叶子类目

                # SKC趋势数据
                sku_item.append(skc_trend.get('saleCnt', 0))  # SKC近7天销量
                sku_item.append(skc_trend.get('epsUvIdx', 0))  # SKC近7天曝光人数
                sku_item.append(skc_trend.get('goodsUv', 0))  # SKC近7天商详访客
                sku_item.append(skc_trend.get('epsGdsCtrIdx', 0))  # SKC近7天点击率
                sku_item.append(skc_trend.get('payUvIdx', 0))  # SKC近7天支付人数
                sku_item.append(skc_trend.get('gdsPayCtrIdx', 0))  # SKC近7天支付率
                sku_item.append(skc_trend.get('totalCommentCnt', 0))  # 评论数

                product_sku_list.append(sku_item)

        cache_file = f'{self.config.auto_dir}/shein/cache/week_sales_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_name: product_sku_list}, [self.store_name])

        return product_sku_list

    # 获取一个skc一段时间内的销售趋势（商品明细中的）
    def get_skc_trend(self, spu, skc, start_date, end_date):
        dt = self.get_dt_time_goods()

        # 将字符串转换为日期对象
        date1 = datetime.strptime(end_date, "%Y-%m-%d").date()
        date2 = datetime.strptime(dt, "%Y%m%d").date()
        if date1 > date2:
            log(f'get_skc_trend: dt:{dt} < end_date: {end_date}')

        cache_file = f'{self.config.auto_dir}/shein/dict/skc_trend_{skc}_{start_date}_{end_date}.json'
        DictSkc = read_dict_from_file(cache_file)
        if len(DictSkc) > 0:
            return DictSkc

        url = f"https://sso.geiwohuo.com/sbn/new_goods/get_skc_diagnose_trend"
        payload = {
            "areaCd"     : "cn",
            "countrySite": [
                "shein-all"
            ],
            "dt"         : dt,
            "endDate"    : end_date.replace('-', ''),
            "spu"        : [spu],
            "skc"        : [skc],
            "startDate"  : start_date.replace('-', ''),
        }
        response_text = fetch_shein(self.web_page, url, payload)
        log(response_text)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        data_list = response_text['info']
        DictSkc = {}
        for date_item in data_list:
            dataDate = date_item['dataDate']
            # epsUvIdx = date_item['epsUvIdx']
            # saleCnt = date_item['saleCnt']
            DictSkc[dataDate] = date_item

        log('len(DictSkc)', len(DictSkc))
        write_dict_to_file(cache_file, DictSkc)
        return DictSkc

    # 获取一个skc一周内的销售趋势（商品明细中的）
    def get_dict_skc_week_trend_v2(self, spu, skc, start_from=None):
        dt = self.get_dt_time()

        date_7_days_ago, date_1_days_ago = TimeUtils.get_past_7_days_range_format(start_from, '%Y%m%d')
        log(date_7_days_ago, date_1_days_ago, 'dt', dt)

        # 将字符串转换为日期对象
        date1 = datetime.strptime(date_1_days_ago, "%Y%m%d").date()
        date2 = datetime.strptime(dt, "%Y%m%d").date()
        if date1 > date2:
            send_exception(f'get_dict_skc_week_trend_v2: dt:{dt} < date_1_days_ago: {date_1_days_ago}')

        cache_file = f'{self.config.auto_dir}/shein/dict/dict_skc_week_trend_{skc}_{date_7_days_ago}_{date_1_days_ago}.json'
        if datetime.now().hour >= 9:
            DictSkc = read_dict_from_file(cache_file)
        else:
            DictSkc = read_dict_from_file(cache_file, 1800)
        if len(DictSkc) > 0:
            return DictSkc

        url = f"https://sso.geiwohuo.com/sbn/new_goods/get_skc_diagnose_trend"
        payload = {
            "areaCd"     : "cn",
            "countrySite": [
                "shein-all"
            ],
            "dt"         : dt,
            "endDate"    : date_1_days_ago,
            "spu"        : [spu],
            "skc"        : [skc],
            "startDate"  : date_7_days_ago,
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        data_list = response_text['info']
        DictSkc = {}
        for date_item in data_list:
            dataDate = date_item['dataDate']
            # epsUvIdx = date_item['epsUvIdx']
            # saleCnt = date_item['saleCnt']
            DictSkc[dataDate] = date_item

        log('len(DictSkc)', len(DictSkc))
        write_dict_to_file(cache_file, DictSkc)
        return DictSkc

    def get_sku_week_sale_list(self, spu, skc, sku):
        dict_skc = self.get_dict_skc_week_trend_v2(spu, skc)
        date_list = TimeUtils.get_past_7_days_list()
        first_day, last_day = TimeUtils.get_past_7_days_range()
        cache_file = f'{self.config.auto_dir}/shein/cache/{skc}_{first_day}_{last_day}.json'
        DictSkuSalesDate = read_dict_from_file(cache_file)
        sales_detail = []
        for date in date_list:
            sales_num = DictSkuSalesDate.get(date, {}).get(sku, {}).get("hisActualValue", 0)
            sales_num = sales_num if sales_num is not None else 0

            saleCnt = get_safe_value(dict_skc.get(date, {}), 'saleCnt', 0)
            epsUvIdx = get_safe_value(dict_skc.get(date, {}), 'epsUvIdx', 0)

            sales_detail.append(f'{date}({TimeUtils.get_weekday_name(date)}): {sales_num}/{saleCnt}/{epsUvIdx}')

        sales_data = []
        for date in date_list:
            goodsUvIdx = get_safe_value(dict_skc.get(date, {}), 'goodsUvIdx', 0)  # 商详访客
            epsGdsCtrIdx = get_safe_value(dict_skc.get(date, {}), 'epsGdsCtrIdx', 0)  # 点击率

            payUvIdx = get_safe_value(dict_skc.get(date, {}), 'payUvIdx', 0)  # 支付人数
            gdsPayCtrIdx = get_safe_value(dict_skc.get(date, {}), 'gdsPayCtrIdx', 0)  # 转化率

            sales_data.append(f'{date}({TimeUtils.get_weekday_name(date)}): {epsGdsCtrIdx:.2%}({goodsUvIdx})/{gdsPayCtrIdx:.2%}({payUvIdx})')

        return sales_detail, sales_data

    def get_activity_price(self, activity_dict, sku, activity_name, dateBegin, dateEnd):
        key = f'{sku}_{dateBegin}_{dateEnd}_{activity_name}'
        price_info = activity_dict.get(key, ['-', '-'])
        return f'活动价:¥{price_info[0]}, 活动库存:{price_info[1]}'

    def get_skc_activity_label(self, skc, sku, dict_activity_price=None):
        cache_file = f'{self.config.auto_dir}/shein/activity_label/activity_label_{self.store_username}.json'
        dict_label = read_dict_from_file(cache_file)
        operateLabelList = dict_label[skc]['operateLabelList']
        activityList = []
        activityList2 = []
        for item in operateLabelList:
            if item['name'] == '活动中':
                activityList.extend(item.get('activityList', []))
            if item['name'] == '即将开始':
                activityList2.extend(item.get('activityList', []))

        if activityList:
            activityLabel = '\n'.join([
                f'  [{act["date"]}]\n【{self.get_activity_price(dict_activity_price, sku, act["name"], act["dateBegin"], act["dateEnd"])}】{act["name"]}\n'
                for act in activityList])
        else:
            activityLabel = '无'
        if activityList2:
            activityLabel2 = '\n'.join([
                f'  [{act["date"]}]\n【{self.get_activity_price(dict_activity_price, sku, act["name"], act["dateBegin"], act["dateEnd"])}】{act["name"]}\n'
                for act in activityList2])
        else:
            activityLabel2 = '无'
        return f'活动中:\n{activityLabel}\n即将开始:\n{activityLabel2}'

    # 获取商品包含sku销量的列表
    # mode = 1.备货建议 2.已上架 3.昨日上架 4.昨日出单
    # 5.采购-缺货要补货      (有现货建议 建议采购为正 有销量)
    # 6.运营采购-滞销清库存   (无现货建议 建议采购为负 30天外 无销量)
    # 7.运营-新品上架需要优化 (无现货建议 建议采购为负 上架15天内)
    # 8.运营-潜在滞销款      (无现货建议 30天外 有销量)
    # 9.运营-潜力热销款      (有现货建议 30天内 有销量)
    # 10.运营-热销款         (有现货建议 30天外 有销量)
    def get_bak_advice(self, mode=1, skcs=None, source='mb'):
        log(f'获取备货信息商品列表 做成字典')
        if skcs == None or len(skcs) == 0:
            # if mode == 3:
            #     skcs = "sh2405133614611175"  # 这是一个不存在的skc
            # else:
            skcs = ""
        else:
            skcs = ",".join(skcs)

        url = "https://sso.geiwohuo.com/idms/goods-skc/list"
        pageNumber = 1
        pageSize = 100
        dictPayload = {
            "pageNumber"            : pageNumber,
            "pageSize"              : pageSize,
            "supplierCodes"         : "",
            "skcs"                  : skcs,
            "spu"                   : "",
            "c7dSaleCntBegin"       : "",
            "c7dSaleCntEnd"         : "",
            "goodsLevelIdList"      : [10, 107, 61, 90, 87, 237, 220, 219, 88, 75, 62, 227, 12, 230, 80, 58, 224, 97],
            "supplyStatus"          : "",
            "shelfStatus"           : "",
            "categoryIdList"        : [],
            "skcStockBegin"         : "",
            "skcStockEnd"           : "",
            "skuStockBegin"         : "",
            "skuStockEnd"           : "",
            "skcSaleDaysBegin"      : "",
            "skcSaleDaysEnd"        : "",
            "skuSaleDaysBegin"      : "",
            "skuSaleDaysEnd"        : "",
            "planUrgentCountBegin"  : "",
            "planUrgentCountEnd"    : "",
            "skcAvailableOrderBegin": "",
            "skcAvailableOrderEnd"  : "",
            "skuAvailableOrderBegin": "",
            "skuAvailableOrderEnd"  : "",
            "shelfDateBegin"        : "",
            "shelfDateEnd"          : "",
            "stockWarnStatusList"   : [],
            "labelFakeIdList"       : [],
            "sheinSaleByInventory"  : "",
            "tspIdList"             : [],
            "adviceStatus"          : [],
            "sortBy7dSaleCnt"       : 2,
            "goodsLevelFakeIdList"  : [1, 2, 3, 8, 14, 15, 4, 11]
        }
        payload = dictPayload
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        spu_list = response_text['info']['list']
        # if int(self.user_info.get('lv1CategoryId')) == 216506:  # 自运营POP店
        #     for spu_item in spu_list:
        #         spu = spu_item.get('spu')
        #         self.get_sku_price_pop(spu)

        skc_list = [item['skc'] for item in spu_list]
        self.get_activity_label(skc_list)
        self.get_preemption_list(skc_list)
        self.get_sku_price_v2(skc_list)
        self.get_stock_advice(skc_list)

        total = response_text['info']['count']
        totalPage = math.ceil(total / pageSize)
        for page in range(2, totalPage + 1):
            log(f'获取备货信息商品列表 第{page}/{totalPage}页')
            dictPayload['pageNumber'] = page
            payload = dictPayload
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['list']

            # if int(self.user_info.get('lv1CategoryId')) == 216506:  # 自运营POP店
            #     for spu_item in spu_list:
            #         spu = spu_item.get('spu')
            #         self.get_product_detail(spu)

            skc_list = [item['skc'] for item in spu_list_new]
            self.get_activity_label(skc_list)
            self.get_preemption_list(skc_list)
            self.get_sku_price_v2(skc_list)
            self.get_stock_advice(skc_list)

            spu_list += spu_list_new
            time.sleep(0.3)

        cache_file = f'{self.config.auto_dir}/shein/dict/activity_price_{self.store_name}.json'
        dictActivityPrice = read_dict_from_file(cache_file)
        # cache_file = f'{self.config.auto_dir}/shein/dict/product_list_{self.store_username}.json'
        # DictSpuInfo = read_dict_from_file(cache_file, 5)
        cache_file = f'{self.config.auto_dir}/shein/preemption_num/preemption_num_{self.store_username}.json'
        dict_preemption_num = read_dict_from_file(cache_file)
        cache_file = f'{self.config.auto_dir}/shein/vmi_spot_advice/spot_advice_{self.store_username}.json'
        dict_advice = read_dict_from_file(cache_file)
        cache_file = f'{self.config.auto_dir}/shein/sku_price/sku_price_{self.store_username}.json'
        dict_sku_price = read_dict_from_file(cache_file)
        date_list = TimeUtils.get_past_7_days_list()
        if mode in [2, 5, 6, 7, 8, 9, 10]:
            excel_data = [[
                '店铺名称', 'SKC图片', 'SKU图片', '商品信息', '建议现货数量', '现有库存数量', '已采购数量', '预测日销',
                '本地和采购可售天数', '生产天数', '建议采购', '产品起定量',
                '备货周期(天)', '备货建议', '近7天SKU销量/SKC销量/SKC曝光', 'SKC点击率/SKC转化率', '自主参与活动',
                'SKC',
                "SKU"
            ]]
        else:
            excel_data = [[
                '店铺名称', 'SKC图片', 'SKU图片', '商品信息', '备货建议', '近7天SKU销量/SKC销量/SKC曝光',
                'SKC点击率/SKC转化率', '自主参与活动', 'SKC', "SKU"
            ]]
        for spu_info in spu_list:
            spu = str(spu_info['spu'])
            skc = str(spu_info['skc'])

            status_cn = spu_info['shelfStatus']['name']
            if status_cn != '已上架':
                continue

            # shelf_status = DictSpuInfo.get(spu, {}).get('shelf_status', '')
            # if mode != 1:
            #     if shelf_status != 'ON_SHELF' and shelf_status != 'SOLD_OUT':
            #         log('跳过', skc, shelf_status)
            #         continue

            # if mode in [5, 6, 7, 8, 9, 10] and shelf_status == 'SOLD_OUT':
            #     continue
            #
            # dictStatus = {
            #     'WAIT_SHELF': "待上架",
            #     'ON_SHELF': "已上架",
            #     'SOLD_OUT': "已售罄",
            #     'OUT_SHELF': "已下架"
            # }
            # status_cn = dictStatus.get(shelf_status, '-')

            sale_model = spu_info['saleModel']['name']
            goods_level = spu_info['goodsLevel']['name']
            goods_label = [label["name"] for label in spu_info['goodsLabelList']]
            skc_img = spu_info['picUrl']
            shelfDate = spu_info['shelfDate']
            shelfDays = spu_info['shelfDays']
            categoryName = spu_info['categoryName']

            if mode in [3] and shelfDays != 1:
                continue

            DictSkuSalesDate = self.get_skc_week_actual_sales(skc)

            for sku_info in spu_info['skuList']:
                row_item = []
                attr = sku_info['attr']
                if attr == '合计':
                    continue
                predictDaySales = sku_info['predictDaySales']
                availableOrderCount = sku_info['availableOrderCount']
                if mode == 1:
                    if availableOrderCount is None or availableOrderCount <= 0:
                        log('跳过', skc, availableOrderCount)
                        continue

                row_item.append(f'{self.store_name}\n({status_cn})\n{goods_level}\n{",".join(goods_label)}')
                row_item.append(skc_img)
                sku = sku_info['skuCode']
                skuExtCode = str(sku_info['supplierSku'])
                sku_img = self.bridge.get_sku_img(skuExtCode, source)
                row_item.append(sku_img)

                transit = sku_info['transit']  # 在途

                stock = self.bridge.get_sku_stock(skuExtCode, source)
                cost_price = self.bridge.get_sku_cost(skuExtCode, source)

                supplyPrice = dict_sku_price[sku]
                shein_stock = sku_info['stock']
                if cost_price == '-':
                    profit = '-'
                else:
                    profit = f'{float(supplyPrice) - float(cost_price):.2f}'

                min_spot_advice = dict_advice.get(skc, {}).get(sku, {}).get('minSpotAdvice', 0)
                max_spot_advice = dict_advice.get(skc, {}).get(sku, {}).get('maxSpotAdvice', 0)
                stock_advice = f'{min_spot_advice}~{max_spot_advice}'
                log('stock_advice', stock_advice)
                # 建议现货数量
                advice_stock_number = round((min_spot_advice + max_spot_advice) / 4)

                # 有现货建议
                if mode in [5, 9, 10] and advice_stock_number == 0:
                    continue

                # 无现货建议
                if mode in [6, 7, 8] and advice_stock_number > 0:
                    continue

                stockSaleDays = sku_info['stockSaleDays']

                product_info = (
                    f'SPU: {spu}\n'
                    f'SKC: {skc}\n'
                    f'SKU货号: {skuExtCode}\n'
                    f'属性集: {attr}\n'
                    f'商品分类: {categoryName}\n'
                    f'上架日期: {shelfDate}\n'
                    f'上架天数: {shelfDays}\n'
                    f'库存可售天数/现货建议: {stockSaleDays}/{stock_advice}\n'
                )
                row_item.append(product_info)

                # 建议采购数量逻辑
                try:
                    # 尝试将字符串数字转换为 float，再转为 int（如有必要）
                    current_stock = float(stock)
                    advice_purchase_number = advice_stock_number - int(current_stock)

                    # 建议采购为正
                    if (mode == 5 and advice_purchase_number <= 0):
                        continue

                except (ValueError, TypeError):
                    # 无法转换为数值时
                    advice_purchase_number = '-'

                if mode in [2, 5, 6, 7, 8, 9, 10]:
                    row_item.append(advice_stock_number)
                    row_item.append(stock)

                    row_item.append(0)
                    row_item.append(predictDaySales)
                    row_item.append(0)
                    row_item.append(7)

                    row_item.append(advice_purchase_number)
                    row_item.append(0)  # 产品起定量
                    row_item.append(0)  # 备货周期(天)

                adviceOrderCount = sku_info['adviceOrderCount'] if sku_info['adviceOrderCount'] is not None else '-'
                if sku_info['autoOrderStatus'] is not None:
                    autoOrderStatus = ['-', '是', '否'][sku_info['autoOrderStatus']] if sku_info[
                                                                                            'adviceOrderCount'] is not None else '-'
                else:
                    autoOrderStatus = '-'
                orderCount = sku_info['orderCount']  # 已下单数
                c7dSaleCnt = sku_info['c7dSaleCnt']
                c30dSaleCnt = sku_info['c30dSaleCnt']
                orderCnt = sku_info['orderCnt']
                totalSaleVolume = sku_info['totalSaleVolume']
                planUrgentCount = sku_info['planUrgentCount']
                preemptionCount = dict_preemption_num[skc][sku]
                predictDaySales = sku_info['predictDaySales']
                goodsDate = sku_info['goodsDate']
                stockDays = sku_info['stockDays']

                real_transit = transit + sku_info['stayShelf'] - sku_info['transitSale']

                sales_info = (
                    f'近7天/30天销量: {c7dSaleCnt}/{c30dSaleCnt}\n'
                    f'当天销量/购买单数: {totalSaleVolume}/{orderCnt}\n'
                    f'预测日销/下单参数: {predictDaySales}/{goodsDate}+{stockDays}\n'
                    f'预占数/预计急采数: {preemptionCount}/{planUrgentCount}\n'
                    f'建议下单/已下单数: {adviceOrderCount}/{orderCount}\n'
                    f'拟下单数/自动下单: {availableOrderCount}/{autoOrderStatus}\n'
                    f'模式/本地/在途/希音: {sale_model[:2]}/{stock}/{real_transit}/{shein_stock}\n'
                    f'成本/核价/利润: ¥{cost_price}/¥{supplyPrice}/¥{profit}\n'
                )

                row_item.append(sales_info)

                flag_yesterday = 0
                sales7cn = 0
                for date in date_list:
                    sales_num = DictSkuSalesDate.get(date, {}).get(sku, {}).get("hisActualValue", 0)
                    sales_num = sales_num if sales_num is not None else 0
                    sales7cn += sales_num
                    if TimeUtils.is_yesterday_date(date) and sales_num == 0:
                        flag_yesterday = 1

                if mode == 4 and flag_yesterday:
                    continue

                # 过滤掉未建立马帮信息的
                if mode in [5, 6, 7, 8, 9, 10] and advice_purchase_number == '-':
                    continue

                # 建议采购为正
                if mode in [5] and advice_purchase_number < 0:
                    continue

                # 建议采购为负
                if mode in [6, 7] and advice_purchase_number >= 0:
                    continue

                # 30内
                if mode in [9] and shelfDays > 31:
                    continue

                # 15天内
                if mode in [7] and shelfDays > 15:
                    continue

                # 30外
                if mode in [6, 8, 10] and shelfDays < 31:
                    continue

                # 有销量
                if mode in [5, 8, 9, 10] and sales7cn == 0:
                    continue

                # 无销量
                if mode in [6] and sales7cn > 0:
                    continue

                sale_num_list, sale_data_list = self.get_sku_week_sale_list(spu, skc, sku)
                row_item.append("\n".join(sale_num_list))
                row_item.append("\n".join(sale_data_list))
                row_item.append(self.get_skc_activity_label(skc, sku, dictActivityPrice))
                row_item.append(skc)
                row_item.append(sku)
                excel_data.append(row_item)

        cache_file = f'{self.config.auto_dir}/shein/cache/bak_advice_{mode}_{TimeUtils.today_date()}.json'
        write_dict_to_file_ex(cache_file, {self.store_name: excel_data}, {self.store_name})

        cache_file = f'{self.config.auto_dir}/shein/cache/bak_advice_notify_{mode}_{TimeUtils.today_date()}.json'
        NotifyItem = [self.store_name, len(excel_data[1:])]
        write_dict_to_file_ex(cache_file, {self.store_name: NotifyItem}, {self.store_name})

        return excel_data

    def check_order_list(self, source, first_day, last_day):
        page_num = 1
        page_size = 200  # 列表最多返回200条数据 大了没有用

        cache_file = f'{self.config.auto_dir}/shein/cache/check_order_{first_day}_{last_day}.json'
        list_item_cache = read_dict_from_file_ex(cache_file, self.store_username)

        url = f"https://sso.geiwohuo.com/gsfs/finance/reportOrder/dualMode/checkOrderList/item/union"
        payload = {
            "page"              : page_num,
            "perPage"           : page_size,
            "detailAddTimeStart": f"{first_day} 00:00:00",
            "detailAddTimeEnd"  : f"{last_day} 23:59:59"
        }
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))
        list_item = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        log(self.store_name, self.store_username, total, len(list_item_cache))
        if int(total) == len(list_item_cache):
            log('总数与缓存数量相同 跳过剩余页抓取', total)
            return list_item_cache

        for page in range(2, totalPage + 1):
            log(f'获取收支明细列表 第{page}/{totalPage}页')
            payload['page'] = page
            response_text = fetch_shein(self.web_page, url, payload)
            spu_list_new = response_text['info']['data']
            list_item += spu_list_new
            time.sleep(0.1)

        for item in list_item:
            supplierSku = item['skuSn']
            item['cost_price'] = self.bridge.get_sku_cost(supplierSku, source)
            item['sku_img'] = self.bridge.get_sku_img(supplierSku, source)

        write_dict_to_file_ex(cache_file, {self.store_username: list_item}, [self.store_username])
        return list_item

    def get_ab_test_list(self, status=4, test_type=2):
        """
        获取AB测试列表
        
        Args:
            status: 测试状态，可选值：
                4: 进行中
            test_type: 测试类型，可选值：
                2: skc测试

        Returns:
            list: AB测试列表
        """
        log(f'获取AB测试列表: status={status}, test_type={test_type}')

        # 构建缓存文件名
        cache_key = f'{test_type}_{status}'
        cache_file = f'{self.config.auto_dir}/shein/cache/ab_test_list_{self.store_username}_{cache_key}.json'
        ab_test_list = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 12)
        if len(ab_test_list) > 0:
            log('返回缓存数据: ', len(ab_test_list))
            return ab_test_list

        page_num = 1
        page_size = 100

        url = f"https://sso.geiwohuo.com/spmc-api-prefix/spmp/image/ab_test/get_test_list?page_num={page_num}&page_size={page_size}"
        payload = {}

        # 添加可选参数
        if status is not None:
            payload["status"] = status
        if test_type is not None:
            payload["test_type"] = test_type

        log(payload)
        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        ab_test_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        totalPage = math.ceil(total / page_size)

        for page in range(2, totalPage + 1):
            log(f'获取AB测试列表 第{page}/{totalPage}页')
            page_num = page
            url = f"https://sso.geiwohuo.com/spmc-api-prefix/spmp/image/ab_test/get_test_list?page_num={page_num}&page_size={page_size}"
            response_text = fetch_shein(self.web_page, url, payload)
            ab_test_list += response_text['info']['data']
            time.sleep(0.1)

        write_dict_to_file_ex(cache_file, {self.store_username: ab_test_list}, [self.store_username])

        for test_list in ab_test_list:
            test_task_id = test_list['test_task_id']
            skc = test_list['spu_or_skc_name']
            test_list['experimental_data'] = self.get_ab_test_result(test_task_id)
            cache_file = f'{self.config.auto_dir}/shein/cache/ab_test_list_{skc}_{TimeUtils.today_date()}.json'
            write_dict_to_file(cache_file, test_list)

        return ab_test_list

    def get_ab_test_result(self, test_task_id):
        """
        获取AB测试实验结果
        
        Args:
            test_task_id: 测试任务ID
            
        Returns:
            dict: 实验结果数据，包含：
                - control_group_data: 对照组数据
                    - expose_uv: 曝光人数
                    - cart_uv: 加购人数
                    - other_click_uv: 其他点击人数
                    - goods_uv: 商详访客
                    - goods_cnt: 商品数量
                    - click_rate: 点击率
                    - cart_rate: 加购率
                    - conversion_rate: 转化率
                - experiment_group_data: 实验组数据（字段同对照组）
        """
        log(f'获取AB测试结果: test_task_id={test_task_id}')

        cache_file = f'{self.config.auto_dir}/shein/cache/ab_test_result_{test_task_id}.json'
        ab_test_result = read_dict_from_file(cache_file, 3600 * 12)
        if len(ab_test_result) > 0:
            log('返回缓存数据')
            return ab_test_result

        url = f"https://sso.geiwohuo.com/spmc-api-prefix/spmp/image/ab_test/compare_experimental_data"
        payload = {
            "test_task_id": test_task_id
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        ab_test_result = response_text['info']

        write_dict_to_file(cache_file, ab_test_result)

        return ab_test_result

    def download_finance_details(self, download_dir, start_date, end_date, output_file_name=None):
        """
        下载并处理财务收支明细文件
        
        功能说明：
        - 自动下载财务收支明细文件（可能是zip或xlsx格式）
        - 如果是zip文件（数据超过5000条），自动解压并合并多个Excel
        - 如果是xlsx文件（数据少于5000条），直接处理
        - 在Excel开头添加3列：店铺账号、店铺名称、店长
        - 自动识别需要保持为字符串的列（如业务单号等）
        
        Args:
            start_date: 开始日期，格式: YYYY-MM-DD
            end_date: 结束日期，格式: YYYY-MM-DD
            
        Returns:
            str: 处理后的Excel文件路径
        """
        import os
        import requests
        from datetime import datetime
        import zipfile
        import shutil
        import openpyxl
        from openpyxl import Workbook
        import pandas as pd

        log(f'开始下载财务收支明细: {start_date} ~ {end_date}', self.store_username, self.store_name)

        # 准备下载目录
        # download_dir = f'{self.config.auto_dir}/shein/finance_details'
        os.makedirs(download_dir, exist_ok=True)

        # 最终输出文件路径
        if output_file_name is None:
            output_file_name = f'finance_details_{self.store_username}_{start_date}_{end_date}.xlsx'
        output_file_path = os.path.join(download_dir, output_file_name)

        # 如果最终文件已存在，直接返回
        if os.path.exists(output_file_path):
            log(f'处理后的文件已存在，直接返回: {output_file_path}')
            return output_file_path

        # 第一步：查询当前已有的任务列表（用于后续对比）
        log('步骤1: 查询当前已有的任务列表')
        url = "https://sso.geiwohuo.com/sso/common/fileExport/list"
        query_start_time = TimeUtils.get_past_nth_day(1, None, '%Y-%m-%d')  # 查询最近1天的任务
        payload = {
            "page"           : 1,
            "perPage"        : 50,
            "fileStatusList" : [1],  # 1-已生成
            "createTimeStart": f"{query_start_time} 00:00:00",
            "createTimeEnd"  : f"{TimeUtils.today_date()} 23:59:59"
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        existing_task_ids = set()
        if str(error_code) == '0':
            data_list = response_text.get('info', {}).get('data', [])
            existing_task_ids = {item.get('id') for item in data_list if item.get('fileName') == '财务收支明细'}
            log(f'当前已有任务数量: {len(existing_task_ids)}')

        # 第二步：记录当前时间并创建导出任务
        log('步骤2: 创建导出任务')
        task_create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log(f'任务创建时间: {task_create_time}')

        url = "https://sso.geiwohuo.com/gsfs/common/file/export/financeDetailsItem"
        payload = {
            "type"              : 1,
            "mode"              : 2,
            "detailAddTimeStart": f"{start_date} 00:00:00",
            "detailAddTimeEnd"  : f"{end_date} 23:59:59"
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            # 检查是否是"暂无数据可导出"的情况，这是正常情况
            if str(error_code) == 'gsfs98008':
                log('暂无数据可导出，返回None')
                return None
            raise send_exception(f'创建导出任务失败: {json.dumps(response_text, ensure_ascii=False)}')

        log('导出任务创建成功，等待文件生成...')

        # 第三步：轮询查询任务状态，查找新创建的任务
        log('步骤3: 轮询查询任务状态')
        task_id = None
        file_extension = None
        max_retry = 60  # 最多查询60次（10分钟）
        retry_count = 0

        # 使用任务创建时间作为查询起始时间（向前推1分钟以避免时间误差）
        create_time_obj = datetime.strptime(task_create_time, '%Y-%m-%d %H:%M:%S')
        from datetime import timedelta
        query_time_obj = create_time_obj - timedelta(minutes=3)
        query_start_time_str = query_time_obj.strftime('%Y-%m-%d %H:%M:%S')

        while retry_count < max_retry:
            retry_count += 1
            time.sleep(10)  # 每10秒查询一次

            log(f'第{retry_count}次查询任务状态...', self.store_username, self.store_name)
            url = "https://sso.geiwohuo.com/sso/common/fileExport/list"
            payload = {
                "page"           : 1,
                "perPage"        : 50,
                "fileStatusList" : [1],  # 1-已生成
                "createTimeStart": query_start_time_str,
                "createTimeEnd"  : f"{TimeUtils.today_date()} 23:59:59"
            }

            log(payload)
            response_text = fetch_shein(self.web_page, url, payload)
            log(response_text)
            error_code = response_text.get('code')
            if str(error_code) != '0':
                log(f'查询任务列表失败: {response_text}')
                continue

            # 查找新出现的财务收支明细任务
            data_list = response_text.get('info', {}).get('data', [])
            for item in data_list:
                item_id = item.get('id')
                item_create_time = item.get('createTime')

                log(item_id, existing_task_ids)
                # 条件：1.文件名匹配 2.状态为已生成 3.不在之前的任务列表中 4.创建时间在任务创建时间之后
                if (item.get('fileName') == '财务收支明细' and item.get('fileStatus') == 1 and item_id not in existing_task_ids):
                    file_extension = item.get('fileExtension', 'xlsx')  # 获取文件扩展名，默认xlsx
                    log(f'找到新创建的任务: ID={item_id}, 创建时间={item_create_time}, 文件类型={file_extension}')
                    task_id = item_id
                    break

            if task_id:
                log(f'任务已完成，任务ID: {task_id}')
                break

        if not task_id:
            raise send_exception(f'导出任务超时，查询{max_retry}次后仍未完成')

        # 第四步：获取下载地址
        log('步骤4: 获取文件下载地址')
        url = f"https://sso.geiwohuo.com/sso/common/fileExport/getFileUrl?id={task_id}"
        headers = {
            "gmpsso-language": "CN",
            "origin-url"     : "https://sso.geiwohuo.com/#/download-management/list",
            "x-sso-scene"    : "gmpsso"
        }

        fetch_config = {
            "credentials"   : "include",
            "referrer"      : "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(f'获取下载地址失败: {json.dumps(response_text, ensure_ascii=False)}')

        download_url = response_text.get('info', {}).get('url')
        if not download_url:
            raise send_exception('下载地址为空')

        log(f'获取到下载地址: {download_url}')

        # 第五步：下载文件
        log(f'步骤5: 下载文件到本地（文件类型: {file_extension}）')

        # 下载临时文件
        temp_file_name = f'finance_details_temp_{self.store_username}_{start_date}_{end_date}.{file_extension}'
        temp_file_path = os.path.join(download_dir, temp_file_name)

        # 使用requests下载文件
        response = requests.get(download_url, stream=True)
        if response.status_code == 200:
            with open(temp_file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            log(f'文件下载成功: {temp_file_path}')
        else:
            raise send_exception(f'文件下载失败，状态码: {response.status_code}')

        # 第六步：处理文件并添加店铺信息列
        log('步骤6: 处理文件并添加店铺信息列')

        # 从文件名中提取store_username
        store_username = self.store_username

        # 自动识别需要保持为字符串的列（包含以下关键词的列保持为字符串）
        str_keywords = ['业务单号']

        all_data = []
        header = None
        dtype_dict = None

        if file_extension == 'zip':
            # 处理zip文件：解压并合并多个Excel
            log('文件类型为zip，开始解压和合并...')

            # 解压到临时目录
            extract_dir = os.path.join(download_dir, 'temp')
            os.makedirs(extract_dir, exist_ok=True)

            log(f'解压文件到: {extract_dir}')
            with zipfile.ZipFile(temp_file_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)

            # 查找所有excel文件
            excel_files = []
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        excel_files.append(os.path.join(root, file))

            log(f'找到 {len(excel_files)} 个excel文件')

            if len(excel_files) == 0:
                raise Exception('zip文件中未找到excel文件')

            # 读取并合并所有excel数据
            for idx, excel_file in enumerate(excel_files):
                log(f'读取文件 {idx + 1}/{len(excel_files)}: {os.path.basename(excel_file)}')

                try:
                    # 第一次读取时，确定需要保持为字符串的列
                    if idx == 0:
                        df_temp = pd.read_excel(excel_file, sheet_name=0, nrows=0)
                        all_columns = df_temp.columns.tolist()

                        # 自动识别字符串列
                        str_columns = []
                        for col in all_columns:
                            col_str = str(col)
                            if any(keyword in col_str for keyword in str_keywords):
                                str_columns.append(col)

                        if str_columns:
                            log(f'自动识别需要保持为字符串的列: {str_columns}')
                            dtype_dict = {col: str for col in str_columns}

                    # 使用pandas读取excel，指定特定列为字符串类型
                    df = pd.read_excel(excel_file, sheet_name=0, dtype=dtype_dict)
                    log(f'pandas读取成功，数据形状: {df.shape} (行数×列数)')

                    # 获取表头
                    if idx == 0:
                        header = df.columns.tolist()
                        log(f'表头: {header[:5]}... (显示前5列)')

                    # 获取数据
                    data_rows = df.values.tolist()
                    all_data.extend(data_rows)
                    log(f'第{idx + 1}个文件添加了 {len(data_rows)} 行数据')

                except Exception as e:
                    log(f'pandas读取失败: {e}，尝试使用openpyxl读取')
                    # 备用方案：使用openpyxl
                    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

                    if '财务收支明细' in wb.sheetnames:
                        ws = wb['财务收支明细']
                    else:
                        ws = wb.worksheets[0]

                    log(f'使用工作表: {ws.title}')
                    rows = list(ws.iter_rows(values_only=True))

                    if idx == 0 and len(rows) > 0:
                        header = list(rows[0])
                        all_data.extend(rows[1:])
                    elif len(rows) > 1:
                        all_data.extend(rows[1:])

                    wb.close()

            # 清理临时解压目录
            shutil.rmtree(extract_dir)
            log('临时解压目录已清理')

        else:
            # 处理单个xlsx文件
            log('文件类型为xlsx，直接读取...')

            try:
                # 确定需要保持为字符串的列
                df_temp = pd.read_excel(temp_file_path, sheet_name=0, nrows=0)
                all_columns = df_temp.columns.tolist()

                str_columns = []
                for col in all_columns:
                    col_str = str(col)
                    if any(keyword in col_str for keyword in str_keywords):
                        str_columns.append(col)

                if str_columns:
                    log(f'自动识别需要保持为字符串的列: {str_columns}')
                    dtype_dict = {col: str for col in str_columns}

                # 读取excel
                df = pd.read_excel(temp_file_path, sheet_name=0, dtype=dtype_dict)
                log(f'pandas读取成功，数据形状: {df.shape} (行数×列数)')

                header = df.columns.tolist()
                log(f'表头: {header[:5]}... (显示前5列)')

                all_data = df.values.tolist()
                log(f'读取了 {len(all_data)} 行数据')

            except Exception as e:
                log(f'pandas读取失败: {e}，尝试使用openpyxl读取')
                # 备用方案
                wb = openpyxl.load_workbook(temp_file_path, read_only=True, data_only=True)

                if '财务收支明细' in wb.sheetnames:
                    ws = wb['财务收支明细']
                else:
                    ws = wb.worksheets[0]

                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 0:
                    header = list(rows[0])
                    all_data = rows[1:]

                wb.close()

        log(f'合并完成，共 {len(all_data)} 行数据')

        # 在表头前添加3列
        new_header = ['店铺账号', '店铺名称', '店长'] + header

        # 在每行数据前添加3列
        new_data = []
        for row in all_data:
            # 将 tuple 转换为 list，并在前面添加3列
            new_row = [store_username, '', ''] + list(row)
            new_data.append(new_row)

        # 创建新的工作簿并写入数据
        log(f'写入合并后的excel: {output_file_path}')
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = '财务收支明细'

        # 写入表头
        ws_new.append(new_header)

        # 写入数据
        for row_data in new_data:
            ws_new.append(row_data)

        # 保存文件（显式转换为str以避免类型提示警告）
        wb_new.save(str(output_file_path))
        log(f'合并完成，文件已保存: {output_file_path}')

        # 删除临时下载文件
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
            log('临时下载文件已清理')

        return output_file_path

    def query_hosting_info_list(self):
        """
        查询店铺活动托管规则列表
        
        Returns:
            list: 托管规则列表，每个元素包含：
                - hosting_id: 托管规则ID
                - scene_type: 场景类型
                - state: 状态
                - hosting_name: 托管规则名称
                - hosting_tools_id: 托管工具ID
                - hosting_tools_state: 托管工具状态
                - time_zone: 时区
                - create_user: 创建用户
                - last_update_user: 最后更新用户
                - insert_time: 创建时间
                - last_update_time: 最后更新时间
                - exist_act_goods: 是否存在活动商品
        """
        log(f'正在获取 {self.store_name} 店铺活动托管规则列表')

        cache_file = f'{self.config.auto_dir}/shein/cache/hosting_info_list_{self.store_username}_{TimeUtils.today_date()}.json'
        hosting_list = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 8)
        if len(hosting_list) > 0:
            log('返回缓存数据')
            return hosting_list

        url = "https://sso.geiwohuo.com/mrs-api-prefix/promotion/hosting/query_hosting_info_list"
        payload = {}

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        hosting_list = response_text.get('info', [])
        log(f'获取到 {len(hosting_list)} 条托管规则')

        write_dict_to_file_ex(cache_file, {self.store_username: hosting_list}, [self.store_username])

        return hosting_list

    def query_hosting_activity_goods(self, hosting_id, goods_states=None):
        """
        查询托管活动参与的商品
        
        Args:
            hosting_id: 托管规则ID
            goods_states: 商品状态列表，默认为[1]（在售）
            
        Returns:
            list: 参与托管活动的商品列表，每个元素包含：
                - goods_state: 商品状态
                - skc_info_list: SKC信息列表
                    - skc_id: SKC ID
                    - skc_name: SKC名称
                    - goods_name: 商品名称
                    - image_url: 图片URL
                    - act_stock_num: 活动库存数量
                    - act_sales_num: 活动销量
                    - activity_info: 活动信息
                    - sku_info_list: SKU信息列表
        """
        if goods_states is None:
            goods_states = [1]

        log(f'正在获取 {self.store_name} 托管活动商品列表 hosting_id={hosting_id}')

        cache_file = f'{self.config.auto_dir}/shein/cache/hosting_activity_goods_{self.store_username}_{hosting_id}_{TimeUtils.today_date()}.json'
        goods_list = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 8)
        if len(goods_list) > 0:
            log('返回缓存数据')
            return goods_list

        page_num = 1
        page_size = 100

        url = "https://sso.geiwohuo.com/mrs-api-prefix/promotion/hosting/query_hosting_activity_goods"
        payload = {
            "goods_states": goods_states,
            "hosting_id"  : str(hosting_id),
            "page_num"    : page_num,
            "page_size"   : page_size
        }

        response_text = fetch_shein(self.web_page, url, payload)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        goods_list = response_text.get('info', [])
        if not goods_list:
            log('未获取到商品数据')
            return []

        # 获取第一页数据
        first_item = goods_list[0] if goods_list else {}
        skc_info_list = first_item.get('skc_info_list', {})
        all_data = skc_info_list.get('data', [])
        meta = skc_info_list.get('meta', {})
        total = meta.get('count', 0)

        log(f'第1页获取到 {len(all_data)} 条商品，总数: {total}')

        # 如果有多页，继续获取
        if total > page_size:
            totalPage = math.ceil(total / page_size)
            for page in range(2, totalPage + 1):
                log(f'获取托管活动商品列表 第{page}/{totalPage}页')
                payload['page_num'] = page
                response_text = fetch_shein(self.web_page, url, payload)
                error_code = response_text.get('code')
                if str(error_code) != '0':
                    log(f'获取第{page}页失败: {response_text}')
                    continue

                page_goods_list = response_text.get('info', [])
                if page_goods_list:
                    page_data = page_goods_list[0].get('skc_info_list', {}).get('data', [])
                    all_data.extend(page_data)
                    log(f'第{page}页获取到 {len(page_data)} 条商品')

                time.sleep(0.1)

        log(f'总共获取到 {len(all_data)} 条商品')

        # 保存缓存
        write_dict_to_file_ex(cache_file, {self.store_username: all_data}, [self.store_username])

        return all_data

    def get_skc_activity_price_info(self, skc, activity_id):
        """
        根据SKC和活动ID获取供货价、活动价和活动库存
        
        Args:
            skc: SKC名称
            activity_id: 活动ID（可以是字符串或整数）
            
        Returns:
            dict: 包含以下键值的字典，如果未找到则返回None：
                - sku_price: SKU供货价（取第一个SKU的价格）
                - act_sku_price: SKU活动价（取第一个SKU的活动价）
                - act_stock_num: 活动库存数量
                - skc_name: SKC名称
                - goods_name: 商品名称
                - activity_id: 活动ID
                - currency: 币种
                - image_url: 商品图片
        """
        log(f'获取SKC活动价格信息: skc={skc}, activity_id={activity_id}')

        # 转换activity_id为整数进行比较
        try:
            target_activity_id = int(activity_id)
        except (ValueError, TypeError):
            log(f'无效的activity_id: {activity_id}')
            return None

        # 缓存文件，使用skc和activity_id作为缓存key
        cache_file = f'{self.config.auto_dir}/shein/cache/skc_activity_price_{self.store_username}_{skc}_{activity_id}_{TimeUtils.today_date()}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 8)
        if cached_data:
            log('返回缓存的价格信息')
            return cached_data

        # 获取所有托管规则
        hosting_list = self.query_hosting_info_list()

        if not hosting_list:
            log('未找到任何托管规则')
            return None

        # 遍历所有托管规则，查找匹配的SKC和活动
        for hosting in hosting_list:
            hosting_id = hosting.get('hosting_id')
            if not hosting_id:
                continue

            log(f'查询托管规则: hosting_id={hosting_id}, hosting_name={hosting.get("hosting_name")}')

            # 获取该托管规则下的商品
            goods_list = self.query_hosting_activity_goods(hosting_id)

            # 在商品列表中查找匹配的SKC
            for goods_item in goods_list:
                skc_name = goods_item.get('skc_name', '')

                # 匹配SKC名称
                if skc_name != skc:
                    continue

                # 检查活动信息
                activity_info = goods_item.get('activity_info', {})
                goods_activity_id = activity_info.get('activity_id')

                # 匹配活动ID
                try:
                    if int(goods_activity_id) != target_activity_id:
                        continue
                except (ValueError, TypeError):
                    continue

                log(f'找到匹配的SKC: {skc_name}, activity_id={goods_activity_id}')

                # 提取活动库存
                act_stock_num = goods_item.get('act_stock_num', 0)

                # 获取第一个SKU的价格信息
                sku_info_list = goods_item.get('sku_info_list', [])
                if not sku_info_list:
                    log(f'SKC {skc_name} 没有SKU信息')
                    continue

                first_sku = sku_info_list[0]
                sku_price = first_sku.get('sku_price', 0)
                act_sku_price = first_sku.get('act_sku_price', 0)
                currency = first_sku.get('currency', 'CNY')

                # 构建返回结果
                result = {
                    'skc_name'     : skc_name,
                    'goods_name'   : goods_item.get('goods_name', ''),
                    'image_url'    : goods_item.get('image_url', ''),
                    'activity_id'  : goods_activity_id,
                    'act_stock_num': act_stock_num,
                    'sku_price'    : sku_price,
                    'act_sku_price': act_sku_price,
                    'currency'     : currency,
                    'start_time'   : activity_info.get('start_time', ''),
                    'end_time'     : activity_info.get('end_time', ''),
                    'time_zone'    : activity_info.get('time_zone', ''),
                }

                log(f'SKC供货价: {sku_price} {currency}, 活动价: {act_sku_price} {currency}, 活动库存: {act_stock_num}')

                # 保存缓存
                write_dict_to_file(cache_file, result)

                return result

        log(f'未找到匹配的SKC和活动: skc={skc}, activity_id={activity_id}')
        return None

    def get_announcement_list(self, page_size=100):
        """
        获取所有公告列表（自动翻页获取全部数据）

        Args:
            page_size: 每页数量，默认100

        Returns:
            dict: 包含公告列表和店铺信息
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 公告列表
                - total: 总数
        """
        log(f'获取公告列表: page_size={page_size}', self.store_username, self.store_name)

        cache_file = f'{self.config.auto_dir}/shein/cache/announcement_list_{self.store_username}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        page_number = 1
        url = 'https://sso.geiwohuo.com/ssls/announcement/getAnnoPageInfo'
        payload = {
            "pageSize": page_size,
            "pageNumber": page_number
        }
        headers = {"Accept-Language": "CN"}

        response_text = fetch_shein(self.web_page, url, payload, headers)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        announcement_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        total_page = math.ceil(total / page_size)

        for page in range(2, total_page + 1):
            log(f'获取公告列表 第{page}/{total_page}页', self.store_username, self.store_name)
            payload['pageNumber'] = page
            response_text = fetch_shein(self.web_page, url, payload, headers)
            if str(response_text.get('code')) == '0':
                announcement_list += response_text['info']['data']
            time.sleep(0.1)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': announcement_list,
            'total': total
        }

        write_dict_to_file(cache_file, result)
        log(f'获取公告列表完成，共{len(announcement_list)}条', self.store_username, self.store_name)

        return result

    def get_announcement_detail(self, announcement_id):
        """
        获取公告详情

        Args:
            announcement_id: 公告ID

        Returns:
            dict: 公告详情，包含店铺信息
                - store_username: 店铺账号
                - store_name: 店铺名称
                - info: 公告详情信息
        """
        log(f'获取公告详情: announcement_id={announcement_id}', self.store_username, self.store_name)

        cache_file = f'{self.config.auto_dir}/shein/cache/announcement_detail_{self.store_username}_{announcement_id}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据')
            return cached_data

        url = 'https://sso.geiwohuo.com/ssls/announcement/getAnnoDetailInfo'
        payload = {
            "announcementId": str(announcement_id)
        }
        headers = {"Accept-Language": "CN"}

        response_text = fetch_shein(self.web_page, url, payload, headers)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'info': response_text['info']
        }

        write_dict_to_file(cache_file, result)
        log(f'获取公告详情完成: {announcement_id}', self.store_username, self.store_name)

        return result

    def get_violation_penalty_list(self, page_size=100):
        """
        获取所有违规处罚通知列表（自动翻页获取全部数据）

        Args:
            page_size: 每页数量，默认100

        Returns:
            dict: 包含违规处罚通知列表和店铺信息
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 违规处罚通知列表
                - total: 总数
        """
        log(f'获取违规处罚通知列表: page_size={page_size}', self.store_username, self.store_name)

        cache_file = f'{self.config.auto_dir}/shein/cache/violation_penalty_list_{self.store_username}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        page_number = 1
        url = 'https://sso.geiwohuo.com/ssls/tab/getMsgPageInfo'
        payload = {
            "categoryIdList": [27, 28],
            "tabId": "1",
            "pageNumber": page_number,
            "pageSize": page_size
        }
        headers = {"Accept-Language": "CN"}

        response_text = fetch_shein(self.web_page, url, payload, headers)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        penalty_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        total_page = math.ceil(total / page_size)

        for page in range(2, total_page + 1):
            log(f'获取违规处罚通知列表 第{page}/{total_page}页', self.store_username, self.store_name)
            payload['pageNumber'] = page
            response_text = fetch_shein(self.web_page, url, payload, headers)
            if str(response_text.get('code')) == '0':
                penalty_list += response_text['info']['data']
            time.sleep(0.1)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': penalty_list,
            'total': total
        }

        write_dict_to_file(cache_file, result)
        log(f'获取违规处罚通知列表完成，共{len(penalty_list)}条', self.store_username, self.store_name)

        return result

    def get_violation_appeal_list(self, page_size=100):
        """
        获取所有违规申诉通知列表（自动翻页获取全部数据）

        Args:
            page_size: 每页数量，默认100

        Returns:
            dict: 包含违规申诉通知列表和店铺信息
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 违规申诉通知列表
                - total: 总数
        """
        log(f'获取违规申诉通知列表: page_size={page_size}', self.store_username, self.store_name)

        cache_file = f'{self.config.auto_dir}/shein/cache/violation_appeal_list_{self.store_username}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        page_number = 1
        url = 'https://sso.geiwohuo.com/ssls/tab/getMsgPageInfo'
        payload = {
            "categoryIdList": [29, 30],
            "tabId": "1",
            "pageNumber": page_number,
            "pageSize": page_size
        }
        headers = {"Accept-Language": "CN"}

        response_text = fetch_shein(self.web_page, url, payload, headers)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        appeal_list = response_text['info']['data']
        total = response_text['info']['meta']['count']
        total_page = math.ceil(total / page_size)

        for page in range(2, total_page + 1):
            log(f'获取违规申诉通知列表 第{page}/{total_page}页', self.store_username, self.store_name)
            payload['pageNumber'] = page
            response_text = fetch_shein(self.web_page, url, payload, headers)
            if str(response_text.get('code')) == '0':
                appeal_list += response_text['info']['data']
            time.sleep(0.1)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': appeal_list,
            'total': total
        }

        write_dict_to_file(cache_file, result)
        log(f'获取违规申诉通知列表完成，共{len(appeal_list)}条', self.store_username, self.store_name)

        return result

    def _get_default_start_time(self):
        """
        获取默认起始时间（前一日17:00:00）

        Returns:
            str: 格式化的时间字符串 "YYYY-MM-DD HH:MM:SS"
        """
        from datetime import timedelta
        yesterday = datetime.now() - timedelta(days=1)
        return yesterday.replace(hour=17, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')

    def get_announcement_list_with_detail(self, start_time=None):
        """
        获取指定时间后的公告列表（带详情）

        Args:
            start_time: 起始时间，默认前一日17:00:00，格式 "YYYY-MM-DD HH:MM:SS"

        Returns:
            dict: 包含带详情的公告列表
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 公告列表（每条记录增加 detail 字段）
                - total: 筛选后数量
        """
        if start_time is None:
            start_time = self._get_default_start_time()

        log(f'获取带详情的公告列表: start_time={start_time}', self.store_username, self.store_name)

        today_date = datetime.now().strftime('%Y-%m-%d')
        cache_file = f'{self.config.auto_dir}/shein/cache/announcement_list_with_detail_{self.store_username}_{today_date}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        all_announcements = self.get_announcement_list()
        filtered_list = []

        for item in all_announcements.get('data', []):
            item_time = item.get('startTime', '')
            if item_time >= start_time:
                announcement_id = item.get('announcementId')
                detail_data = self.get_announcement_detail(announcement_id)
                item['detail'] = detail_data.get('info', {})
                filtered_list.append(item)
                time.sleep(0.1)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': filtered_list,
            'total': len(filtered_list)
        }

        write_dict_to_file(cache_file, result)
        log(f'获取带详情的公告列表完成，共{len(filtered_list)}条', self.store_username, self.store_name)

        return result

    def get_violation_penalty_list_with_detail(self, start_time=None):
        """
        获取指定时间后的违规处罚通知列表

        Args:
            start_time: 起始时间，默认前一日17:00:00，格式 "YYYY-MM-DD HH:MM:SS"

        Returns:
            dict: 包含违规处罚通知列表
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 违规处罚通知列表
                - total: 筛选后数量
        """
        if start_time is None:
            start_time = self._get_default_start_time()

        log(f'获取违规处罚通知列表(时间过滤): start_time={start_time}', self.store_username, self.store_name)

        today_date = datetime.now().strftime('%Y-%m-%d')
        cache_file = f'{self.config.auto_dir}/shein/cache/violation_penalty_list_with_detail_{self.store_username}_{today_date}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        all_penalties = self.get_violation_penalty_list()
        filtered_list = []

        for item in all_penalties.get('data', []):
            item_time = item.get('addTime', '')
            if item_time >= start_time:
                filtered_list.append(item)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': filtered_list,
            'total': len(filtered_list)
        }

        write_dict_to_file(cache_file, result)
        log(f'获取违规处罚通知列表(时间过滤)完成，共{len(filtered_list)}条', self.store_username, self.store_name)

        return result

    def get_violation_appeal_list_with_detail(self, start_time=None):
        """
        获取指定时间后的违规申诉通知列表

        Args:
            start_time: 起始时间，默认前一日17:00:00，格式 "YYYY-MM-DD HH:MM:SS"

        Returns:
            dict: 包含违规申诉通知列表
                - store_username: 店铺账号
                - store_name: 店铺名称
                - data: 违规申诉通知列表
                - total: 筛选后数量
        """
        if start_time is None:
            start_time = self._get_default_start_time()

        log(f'获取违规申诉通知列表(时间过滤): start_time={start_time}', self.store_username, self.store_name)

        today_date = datetime.now().strftime('%Y-%m-%d')
        cache_file = f'{self.config.auto_dir}/shein/cache/violation_appeal_list_with_detail_{self.store_username}_{today_date}.json'
        cached_data = read_dict_from_file(cache_file, 3600 * 12)
        if len(cached_data) > 0:
            log('返回缓存数据: ', len(cached_data.get('data', [])))
            return cached_data

        all_appeals = self.get_violation_appeal_list()
        filtered_list = []

        for item in all_appeals.get('data', []):
            item_time = item.get('addTime', '')
            if item_time >= start_time:
                filtered_list.append(item)

        result = {
            'store_username': self.store_username,
            'store_name': self.store_name,
            'data': filtered_list,
            'total': len(filtered_list)
        }

        write_dict_to_file(cache_file, result)
        log(f'获取违规申诉通知列表(时间过滤)完成，共{len(filtered_list)}条', self.store_username, self.store_name)

        return result

    # ==================== 供应商信息 API ====================

    def check_supplier_verify_status(self):
        """
        检测供应商信息验证状态
        
        Returns:
            bool: True 表示需要验证，False 表示已验证
        """
        log(f'检测供应商信息验证状态: {self.store_username} {self.store_name}')

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/verify/statusCheck"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        is_alter = response_text.get('info', {}).get('isAlter', 0)
        need_verify = int(is_alter) == 1

        log(f'供应商信息验证状态: {"需要验证" if need_verify else "已验证"}', self.store_username, self.store_name)
        return need_verify

    def do_supplier_verify(self):
        """
        执行供应商信息验证
        先进入供应商信息页面，检测验证状态，如需验证则点击验证按钮
        
        Returns:
            bool: True 表示验证成功或无需验证，False 表示验证失败
        """
        log(f'执行供应商信息验证: {self.store_username} {self.store_name}')

        # 进入供应商信息页面
        self.web_page.goto('https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail')
        self.web_page.wait_for_load_state('load')
        self.web_page.wait_for_timeout(2000)

        # 检测验证状态
        need_verify = self.check_supplier_verify_status()

        if not need_verify:
            log(f'无需验证，已通过验证', self.store_username, self.store_name)
            return True

        # 需要验证，点击验证按钮
        try:
            verify_button = self.web_page.locator('//button[span[text()="进行验证"]]')
            if verify_button.is_visible():
                log(f'点击"进行验证"按钮', self.store_username, self.store_name)
                verify_button.click()
                self.web_page.wait_for_timeout(3000)

                # 等待验证完成，可能需要用户输入密码
                # 这里等待一段时间让用户完成验证
                max_wait = 60  # 最多等待60秒
                wait_count = 0
                while wait_count < max_wait:
                    # 再次检测验证状态
                    if not self.check_supplier_verify_status():
                        log(f'验证成功', self.store_username, self.store_name)
                        return True
                    log(f'等待验证完成... {wait_count}s', self.store_username, self.store_name)
                    self.web_page.wait_for_timeout(5000)
                    wait_count += 5

                log(f'验证超时', self.store_username, self.store_name)
                return False
            else:
                log(f'未找到"进行验证"按钮', self.store_username, self.store_name)
                return False
        except Exception as e:
            log(f'验证过程出错: {e}', self.store_username, self.store_name)
            return False

    def ensure_supplier_verified(self):
        """
        确保供应商信息已验证（获取供应商信息的前置条件）
        
        Returns:
            bool: True 表示已验证，False 表示验证失败
        """
        log(f'确保供应商信息已验证: {self.store_username} {self.store_name}')

        self.close_modal_once()

        # 进入供应商信息页面
        self.web_page.goto('https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail')
        self.web_page.wait_for_load_state('load')
        self.web_page.wait_for_timeout(2000)

        # 检测验证状态
        need_verify = self.check_supplier_verify_status()

        if not need_verify:
            log(f'已通过验证', self.store_username, self.store_name)
            return True

        # 需要验证，执行验证流程
        return self.do_supplier_verify()

    def get_basic_info(self, use_cache=True):
        """
        获取供应商基本信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: 供应商基本信息，包含联系人、联系方式、地址等
        """
        log(f'获取供应商基本信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/basic_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的供应商基本信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/basicInfo/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取供应商基本信息成功', self.store_username, self.store_name)
        return info


    def get_company_info(self, use_cache=True):
        """
        获取公司信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: 公司信息，包含营业执照、法人信息、实际控制人等
        """
        log(f'获取公司信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/company_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的公司信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/companyInfo/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取公司信息成功', self.store_username, self.store_name)
        return info


    def get_tax_info(self, use_cache=True):
        """
        获取税务信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: 税务信息，包含税号、税务地址等
        """
        log(f'获取税务信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/tax_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的税务信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/taxInfo/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取税务信息成功', self.store_username, self.store_name)
        return info


    def get_finance_info(self, use_cache=True):
        """
        获取财务信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: 财务信息，包含银行账户列表、结算信息等
        """
        log(f'获取财务信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/finance_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的财务信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/finance/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取财务信息成功', self.store_username, self.store_name)
        return info


    def get_epr_info(self, use_cache=True):
        """
        获取欧洲EPR资质信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: EPR资质信息，包含各国EPR注册号
        """
        log(f'获取欧洲EPR资质信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/epr_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的欧洲EPR资质信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/eprInfo/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取欧洲EPR资质信息成功', self.store_username, self.store_name)
        return info


    def get_store_info(self, use_cache=True):
        """
        获取店铺信息
        
        Args:
            use_cache: 是否使用缓存，默认True
        
        Returns:
            dict: 店铺信息，包含店铺名称、Logo、开通站点等
        """
        log(f'获取店铺信息: {self.store_username} {self.store_name}')

        # 检查缓存
        cache_file = f'{self.config.auto_dir}/shein/supplier_info/store_info_{self.store_username}.json'
        if use_cache:
            info = read_dict_from_file_ex(cache_file, self.store_username, 3600 * 24 * 365)  # 缓存1年
            if info and len(info) > 0:
                log(f'返回缓存的店铺信息', self.store_username, self.store_name)
                return info

        url = "https://sso.geiwohuo.com/mip-eur-api/supplier/store/detail"

        headers = {
            "content-type": "application/json;Charset=utf-8",
            "origin-url": "https://sso.geiwohuo.com/#/mip-eur/individual-center/new-my-file/detail",
            "x-req-zone-id": "Asia/Shanghai"
        }

        fetch_config = {
            "credentials": "include",
            "referrer": "https://sso.geiwohuo.com/",
            "referrerPolicy": "strict-origin-when-cross-origin"
        }

        response_text = fetch_get(self.web_page, url, headers, fetch_config)
        error_code = response_text.get('code')
        if str(error_code) != '0':
            raise send_exception(json.dumps(response_text, ensure_ascii=False))

        info = response_text.get('info', {})

        # 写入缓存
        write_dict_to_file_ex(cache_file, {self.store_username: info}, [self.store_username])

        log(f'获取店铺信息成功', self.store_username, self.store_name)
        return info


    def get_all_supplier_info(self, use_cache=True, auto_verify=True):
        """
        获取所有供应商信息
        
        Args:
            use_cache: 是否使用缓存，默认True
            auto_verify: 是否自动执行验证，默认True
        
        Returns:
            dict: 包含所有供应商信息的字典
                - basic_info: 基本信息
                - company_info: 公司信息
                - tax_info: 税务信息
                - finance_info: 财务信息
                - epr_info: 欧洲EPR资质信息
                - store_info: 店铺信息
        """
        log(f'获取所有供应商信息: {self.store_username} {self.store_name}')

        # 先执行验证（获取供应商信息的前置条件）
        if auto_verify:
            verified = self.ensure_supplier_verified()
            if not verified:
                log(f'供应商信息验证失败，无法获取信息', self.store_username, self.store_name)
                return {
                    'basic_info': None,
                    'company_info': None,
                    'tax_info': None,
                    'finance_info': None,
                    'epr_info': None,
                    'store_info': None
                }

        result = {
            'basic_info': None,
            'company_info': None,
            'tax_info': None,
            'finance_info': None,
            'epr_info': None,
            'store_info': None
        }

        # 获取基本信息
        try:
            result['basic_info'] = self.get_basic_info(use_cache)
        except Exception as e:
            log(f'获取基本信息失败: {e}', self.store_username, self.store_name)

        # 获取公司信息
        try:
            result['company_info'] = self.get_company_info(use_cache)
        except Exception as e:
            log(f'获取公司信息失败: {e}', self.store_username, self.store_name)

        # 获取税务信息
        try:
            result['tax_info'] = self.get_tax_info(use_cache)
        except Exception as e:
            log(f'获取税务信息失败: {e}', self.store_username, self.store_name)

        # 获取财务信息
        try:
            result['finance_info'] = self.get_finance_info(use_cache)
        except Exception as e:
            log(f'获取财务信息失败: {e}', self.store_username, self.store_name)

        # 获取欧洲EPR资质信息
        try:
            result['epr_info'] = self.get_epr_info(use_cache)
        except Exception as e:
            log(f'获取欧洲EPR资质信息失败: {e}', self.store_username, self.store_name)

        # 获取店铺信息
        try:
            result['store_info'] = self.get_store_info(use_cache)
        except Exception as e:
            log(f'获取店铺信息失败: {e}', self.store_username, self.store_name)

        log(f'获取所有供应商信息完成', self.store_username, self.store_name)
        return result
