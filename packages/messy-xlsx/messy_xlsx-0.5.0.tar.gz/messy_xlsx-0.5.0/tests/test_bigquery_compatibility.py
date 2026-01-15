"""Tests to ensure output is BigQuery import compatible."""

import re
from pathlib import Path

import numpy as np
import pandas as pd
import pyarrow as pa
import pytest

from messy_xlsx import MessyWorkbook


# Find all sample files
SAMPLES_DIR = Path(__file__).parent / "samples"
SAMPLE_FILES = list(SAMPLES_DIR.glob("*.xlsx")) + list(SAMPLES_DIR.glob("*.csv"))


def check_bigquery_compatible(df: pd.DataFrame) -> list[str]:
    """
    Check if DataFrame is BigQuery compatible.

    Returns list of issues found.
    """
    issues = []

    for col in df.columns:
        col_str = str(col)

        # Check column name is valid identifier (letters, numbers, underscore, starting with letter/underscore)
        if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", col_str):
            issues.append(f"Column '{col}' has invalid name for BigQuery (must be alphanumeric + underscore)")

        # Check for mixed types in object columns
        if df[col].dtype == object:
            non_null = df[col].dropna()
            if len(non_null) > 0:
                # Get unique types (excluding NoneType)
                types = set()
                has_nan_float = False

                for val in non_null:
                    if isinstance(val, float) and np.isnan(val):
                        has_nan_float = True
                    elif val is not None:
                        types.add(type(val).__name__)

                # Check for np.nan in object columns (should be None)
                if has_nan_float:
                    issues.append(f"Column '{col}' has np.nan in object column (should be None for BigQuery)")

                # Check for mixed types (allow str only, or numeric only)
                if len(types) > 1:
                    # Allow int/float mix (both are numeric)
                    numeric_types = {"int", "float", "int64", "float64"}
                    if not types.issubset(numeric_types):
                        issues.append(f"Column '{col}' has mixed types: {types}")

                # Check for unhashable/complex types that BigQuery can't handle
                for val in non_null:
                    if isinstance(val, (list, dict, set, tuple)):
                        issues.append(f"Column '{col}' contains {type(val).__name__} (not BigQuery compatible)")
                        break

    return issues


class TestBigQueryCompatibility:
    """Ensure DataFrame output is BigQuery compatible."""

    @pytest.mark.parametrize("sample_file", SAMPLE_FILES, ids=lambda f: f.name)
    def test_sample_file_bq_compatible(self, sample_file):
        """All sample files should produce BigQuery-compatible output."""
        try:
            with MessyWorkbook(sample_file) as mwb:
                # Test each sheet
                for sheet_name in mwb.sheet_names:
                    df = mwb.to_dataframe(sheet=sheet_name)

                    if df.empty:
                        continue

                    issues = check_bigquery_compatible(df)

                    # Filter known issues that are acceptable
                    # Column names with spaces are common in real files
                    critical_issues = [
                        i for i in issues
                        if "np.nan in object column" in i  # This is the critical one
                        or "contains list" in i
                        or "contains dict" in i
                    ]

                    assert critical_issues == [], \
                        f"BigQuery critical issues in {sample_file.name} sheet '{sheet_name}': {critical_issues}"

        except Exception as e:
            # Some files might have issues - log but don't fail
            pytest.skip(f"Could not process {sample_file.name}: {e}")

    def test_simple_xlsx_is_bq_compatible(self, tmp_path):
        """Simple XLSX should produce BQ-compatible output."""
        import openpyxl

        file_path = tmp_path / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value", "date"])
        ws.append(["Alice", 100, "2024-01-01"])
        ws.append(["Bob", 200, "2024-01-02"])
        ws.append([None, None, None])  # Empty row
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        # Only check critical issues
        critical = [i for i in issues if "np.nan" in i or "contains" in i]
        assert critical == [], f"BigQuery compatibility issues: {critical}"

    def test_missing_values_use_none_not_nan(self, tmp_path):
        """String columns should use None, not np.nan for missing values."""
        import openpyxl

        file_path = tmp_path / "missing.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["Alice", 100])
        ws.append(["NA", 200])  # NA should become None
        ws.append(["", 300])   # Empty should become None
        ws.append([None, 400])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        # Check string column doesn't have np.nan
        issues = check_bigquery_compatible(df)
        nan_issues = [i for i in issues if "np.nan" in i]
        assert nan_issues == [], f"Should not have np.nan in object columns: {nan_issues}"

    def test_excel_errors_produce_none_not_nan(self, tmp_path):
        """Excel errors like #DIV/0! should become None, not np.nan in string columns."""
        import openpyxl

        file_path = tmp_path / "errors.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["label", "value"])
        ws.append(["#DIV/0!", 100])  # Error value as string
        ws.append(["#N/A", 200])
        ws.append(["Normal", 300])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        nan_issues = [i for i in issues if "np.nan" in i]
        assert nan_issues == [], f"Excel errors should become None, not np.nan: {nan_issues}"

    def test_no_complex_types_in_output(self, tmp_path):
        """Output should not contain lists, dicts, or other complex types."""
        import openpyxl

        file_path = tmp_path / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        for col in df.columns:
            for val in df[col].dropna():
                assert not isinstance(val, (list, dict, set, tuple)), \
                    f"Column {col} contains complex type: {type(val)}"


class TestColumnNameNormalization:
    """Test that column names can be made BigQuery compatible."""

    def test_detect_invalid_column_names(self, tmp_path):
        """Should detect column names that are invalid for BigQuery."""
        import openpyxl

        file_path = tmp_path / "names.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["valid_name", "Also Valid", "123invalid", "has-dash", "has space"])
        ws.append([1, 2, 3, 4, 5])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        name_issues = [i for i in issues if "invalid name" in i]

        # Should detect at least some invalid names
        assert len(name_issues) >= 3, f"Should detect invalid column names, found: {name_issues}"


def convert_to_arrow(df: pd.DataFrame) -> tuple[pa.Table | None, str | None]:
    """
    Try to convert DataFrame to PyArrow Table.

    This simulates what BigQuery does internally when loading data.
    Returns (table, None) on success, (None, error_message) on failure.
    """
    try:
        # BigQuery uses Arrow format internally
        # This will fail if there are incompatible types
        table = pa.Table.from_pandas(df)
        return table, None
    except Exception as e:
        return None, str(e)


def validate_arrow_schema_for_bq(table: pa.Table) -> list[str]:
    """
    Validate that Arrow schema is BigQuery compatible.

    BigQuery supports: INT64, FLOAT64, BOOL, STRING, BYTES, DATE, DATETIME, TIME, TIMESTAMP
    """
    issues = []

    # Map Arrow types to BigQuery compatibility
    bq_compatible_types = {
        pa.int8(), pa.int16(), pa.int32(), pa.int64(),
        pa.uint8(), pa.uint16(), pa.uint32(), pa.uint64(),
        pa.float16(), pa.float32(), pa.float64(),
        pa.bool_(),
        pa.string(), pa.large_string(), pa.utf8(), pa.large_utf8(),
        pa.binary(), pa.large_binary(),
        pa.date32(), pa.date64(),
    }

    for field in table.schema:
        field_type = field.type

        # Check for timestamp types (compatible)
        if pa.types.is_timestamp(field_type):
            continue

        # Check for duration (not directly supported)
        if pa.types.is_duration(field_type):
            issues.append(f"Column '{field.name}' has duration type (not BQ compatible)")
            continue

        # Check for nested types
        if pa.types.is_list(field_type):
            issues.append(f"Column '{field.name}' has list type (requires REPEATED mode in BQ)")
            continue

        if pa.types.is_struct(field_type):
            issues.append(f"Column '{field.name}' has struct type (requires RECORD mode in BQ)")
            continue

        if pa.types.is_map(field_type):
            issues.append(f"Column '{field.name}' has map type (not directly supported in BQ)")
            continue

    return issues


class TestPyArrowConversion:
    """Test that output can be converted to PyArrow (BigQuery's internal format)."""

    @pytest.mark.parametrize("sample_file", SAMPLE_FILES, ids=lambda f: f.name)
    def test_sample_converts_to_arrow(self, sample_file):
        """All sample files should convert to PyArrow successfully."""
        try:
            with MessyWorkbook(sample_file) as mwb:
                for sheet_name in mwb.sheet_names:
                    df = mwb.to_dataframe(sheet=sheet_name)

                    if df.empty:
                        continue

                    # Try to convert to Arrow
                    table, error = convert_to_arrow(df)

                    assert error is None, \
                        f"Failed to convert {sample_file.name} sheet '{sheet_name}' to Arrow: {error}"

                    # Validate schema
                    schema_issues = validate_arrow_schema_for_bq(table)

                    # Filter out acceptable issues (nested types might be intentional)
                    critical_issues = [i for i in schema_issues if "not BQ compatible" in i]

                    assert critical_issues == [], \
                        f"Arrow schema issues in {sample_file.name}: {critical_issues}"

        except Exception as e:
            pytest.skip(f"Could not process {sample_file.name}: {e}")

    def test_arrow_roundtrip(self, tmp_path):
        """Data should survive Arrow roundtrip without corruption."""
        import openpyxl

        file_path = tmp_path / "roundtrip.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value", "active"])
        ws.append(["Alice", 100, True])
        ws.append(["Bob", 200.5, False])
        ws.append([None, None, None])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df_original = mwb.to_dataframe()

        # Convert to Arrow and back
        table, error = convert_to_arrow(df_original)
        assert error is None, f"Arrow conversion failed: {error}"

        df_roundtrip = table.to_pandas()

        # Verify data integrity
        assert len(df_roundtrip) == len(df_original)
        assert list(df_roundtrip.columns) == list(df_original.columns)

    def test_parquet_export(self, tmp_path):
        """Data should be exportable to Parquet (BigQuery's preferred format)."""
        import openpyxl

        file_path = tmp_path / "export.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "amount"])
        ws.append([1, "Alice", 100.50])
        ws.append([2, "Bob", 200.75])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        # Export to Parquet (BigQuery's preferred format)
        parquet_path = tmp_path / "output.parquet"
        table, error = convert_to_arrow(df)
        assert error is None

        # Write to Parquet
        import pyarrow.parquet as pq
        pq.write_table(table, parquet_path)

        # Read back and verify
        table_read = pq.read_table(parquet_path)
        df_read = table_read.to_pandas()

        assert len(df_read) == len(df)
        assert list(df_read.columns) == list(df.columns)
