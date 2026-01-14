 # -*- coding: utf-8 -*-
# @Time : 2024/11/13 10:00
# @Author : Zhang Shaodong
# @Email : zsd0830@163.com
# @File : universal_function.py
# @Info : 该脚本中包含通用的函数模块与核心参数

import os
import time
import json
import pandas as pd
import random
import pymysql
import psycopg2
from psycopg2.extras import execute_batch
import datetime
import bs4
import re
import sys
from openpyxl import load_workbook
import xlwings as xw

# 通用参数
#########################################################################
# 本地数据库信息
ip_database = {"localhost": "wom_db_local", "BA_USING": "BA_USING", "BI_READ": "bi"}  # 需修改本地数据库名
dic_connet = {"localhost": {"host": "localhost", "port": 3306, "user": "root", "password": "zwp650811", "database": "wom_db_local"},
              "BA_USING": {"host": "10.26.241.164", "port": 3306, "user": "BA_USING", "password": "BA_USING@2022", "database": "BA_USING"},
              "BI_READ": {"host": "10.26.241.164", "port": 3306, "user": "BI_READ", "password": "Bireader@1027", "database": "bi"}}
# 华为云 - gaussdb数据库信息 - 生产环境
gs_database = {"ba_data_develop": "bi"}
gs_connect = {"bi": {"host": "10.26.0.194", "port": 8000, "user": "ba_data_develop", "password": "Smart@2021", "database": "bi"}}
# 阿里云 - ali-holo-prod数据库信息 - 生产环境
ali_prod_database = {"BASIC$ba_data_read": "bi"}
ali_prod_connect = {"bi": {"host": "hgpostcn-cn-wuf413crl003-cn-hangzhou-vpc-st.hologres.aliyuncs.com",
                           "port": 80, "user": "BASIC$ba_data_read", "password": "jw3kHp!ehHw6e", "database": "bi"}}
# 阿里云 - ali-holo-test数据库信息 - 测试环境
ali_test_database = {"BASIC$ba_data_read": "bi"}
ali_test_connect = {"bi": {"host": "hgpostcn-cn-cfn40200a01z-cn-hangzhou-vpc-st.hologres.aliyuncs.com",
                           "port": 80, "user": "BASIC$ba_data_read", "password": "Smart@2021", "database": "bi"}}

# 青果的IP代理，目前已不使用
# proxy_dict = {"host": "tunnel2.qg.net", "port": "17104", "key": "66D2C356", "passwd": "8A0FDC1CC0D9"}
# 亿牛云的IP代理：
proxy_dict = {"host": "u11605.150.tp.16yun.cn", "port": "6239", "key": "16PSUKQD", "passwd": "007214"}

province_ls = ["北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽", "福建", "江西",
               "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西", "甘肃", "青海", "台湾", "内蒙古",
               "广西", "西藏", "宁夏", "新疆", "香港", "澳门"]
zhixia_list = ["北京", "天津", "上海", "重庆"]
month_dic = {"Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
             "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"}


# 通用功能
#########################################################################
# 功能：判断sth是否为NaN、None或者空字符串，NaN、None或者空字符串返回True，其他情况返回False
def is_nano(sth):
    if not sth:
        return True
    if isinstance(sth, float):
        if np.isnan(sth):
            return True
    return False

# 功能：解析代理proxy的信息
# 输入：proxy_info - 代理信息 - dict
# 输出：proxy_out - 代理地址字典 - dict
def get_proxy(proxy_info: dict):
    proxy = 'http://{}:{}@{}:{}'.format(proxy_info['key'], proxy_info['passwd'], proxy_info['host'], proxy_info['port'])
    return {"http": proxy, "https": proxy}

def get_proxy4other(proxy_info: dict, for_httpx=False, if_https=False):
    proxy = 'http://{}:{}@{}:{}'.format(proxy_info['key'], proxy_info['passwd'], proxy_info['host'], proxy_info['port'])
    return {f"http{'s' if if_https else ''}:{'//' if for_httpx else ''}": proxy}

def get_proxy4playwright(proxy_info: dict, for_httpx=False, if_https=False):
    proxy = 'http://{}:{}@{}:{}'.format(proxy_info['key'], proxy_info['passwd'], proxy_info['host'], proxy_info['port'])
    return {"http://": proxy, "https://": proxy}

# 给playwright使用的proxy必须做如下转化，暂时还没有使用，以后再根据实际使用开放
# def trans_proxy4playwright(proxy_info: dict):
#     proxy_trans = {
#         'server': 'http://'+proxy_info['host']+':'+proxy_info['port'],
#         'username': proxy_info['key'],
#         'password': proxy_info['passwd']
#     }
#     return proxy_trans



# 文件读取
#########################################################################

# 功能：读取json文件变成字典格式
# 输入：json_file - json文件的路径 - str
# 输出：data_dict - json文件转化成的字典 - dict
def read_json(json_file: str):
    with open(json_file, 'r', encoding='utf-8') as load_f:
        data_dict = json.load(load_f)
    return data_dict


# mysql数据库相关
#########################################################################

# 功能：尝试多次连接数据库
def get_cursor_times(config, times):
    cnt_conn = 0
    while cnt_conn < times:
        try:
            conn = pymysql.connect(**config)
            cursor = conn.cursor()
            break
        except:
            print(f'mysql数据库连接失败，现进行第{cnt_conn}次重新连接，总共{times}次重新连接的机会...')
            sleep_some_seconds(20, 21, 1)
            cnt_conn += 1
    return cursor, conn


# 将dataframe写进数据库
def df2mysql(db_config, db, insert_df, insert_tbl, cover_flag):
    if len(insert_df) != 0:
        insert_cursor, insert_conn = get_cursor_times(db_config[db], 10000)
        # 将dataframe中的Nan用None代替
        insert_df = insert_df.where(insert_df.notnull(), None)
        len_cols = insert_df.columns.size
        insert_sql = f"{'replace' if cover_flag else 'insert ignore'} into {insert_tbl} values ({', '.join(['%s']*len_cols)})"
        insert_data = (tuple(row) for _, row in insert_df.iterrows())
        try:
            insert_cursor.executemany(insert_sql, insert_data)
            insert_conn.commit()
            print(f"数据插入{insert_tbl}成功(๑•̀ㅂ•́)و✧")
        except Exception as e:
            print(e, f"数据插入{insert_tbl}失败：", insert_data)
        finally:
            insert_cursor.close()
            insert_conn.close()
    else:
        print(f'无数据需要插入{insert_tbl}ε=(´ο｀*)))')


# 功能：将dict格式的数据插入insert_db的insert_tbl中
# 输入样例
# insert_db = 'BA_USING'
# insert_tbl = 'autohome_policy_post'
# db_config = {"localhost": {'host': 'localhost', 'port': 3306, 'user': 'root', 'password': 'zwp650811', 'database': 'wom_db_local'},
#              "BA_USING": {'host': '10.26.241.164', 'port': 3306, 'user': 'BA_USING', 'password': 'BA_USING@2022', 'database': ip_database["BA_USING"]},
#              "BI_READ": {'host': '10.26.241.164', 'port': 3306, 'user': 'BI_READ', 'password': 'Bireader@1027', 'database': ip_database["BI_READ"]}}
def dict2mysql(db_config, db, insert_dict, insert_tbl, cover_flag):
    insert_cursor, insert_conn = get_cursor_times(db_config[db], 10000)
    keys = ", ".join(insert_dict.keys())
    values = ", ".join(['%s'] * len(insert_dict.keys()))
    if cover_flag:
        insert_sql = '''
                    INSERT INTO {table}({keys}) VALUES ({values}) ON DUPLICATE KEY UPDATE
                    '''.format(table=insert_tbl, keys=keys, values=values)
        update = ','.join([" {key} = values({key})".format(key=key) for key in insert_dict])
        insert_sql += update
    else:
        insert_sql = """
                    INSERT ignore INTO {table}({keys}) VALUES ({values})
                    """.format(table=insert_tbl, keys=keys, values=values)
    try:
        insert_cursor.execute(insert_sql, tuple(insert_dict.values()))
        insert_conn.commit()
        print(f"数据库插入结果数据成功(๑•̀ㅂ•́)و✧{str(insert_dict)[:100]}...", )
    except Exception as e:
        print(e, f"\n数据库表{insert_tbl}插入数据失败：", insert_dict)
    finally:
        insert_cursor.close()
        insert_conn.close()

# 功能：将dict格式的数据插入insert_db的insert_tbl中，并返回自增ID
# 输入样例
# insert_db = 'BA_USING'
# insert_tbl = 'autohome_policy_post'
# db_config = {"BA_USING": {'host': '10.26.241.164', 'port': 3306, 'user': 'BA_USING', 'password': 'BA_USING@2022',
#                           'database': "BA_USING"}}
def dict2mysql_with_id(db_config, db, insert_dict, insert_tbl, cover_flag):
    insert_cursor, insert_conn = get_cursor_times(db_config[db], 10000)
    keys = ", ".join(insert_dict.keys())
    values = ", ".join(['%s'] * len(insert_dict.keys()))
    if cover_flag:
        insert_sql = '''
                    INSERT INTO {table}({keys}) VALUES ({values}) ON DUPLICATE KEY UPDATE
                    '''.format(table=insert_tbl, keys=keys, values=values)
        update = ','.join([" {key} = values({key})".format(key=key) for key in insert_dict])
        insert_sql += update
    else:
        insert_sql = """
                    INSERT ignore INTO {table}({keys}) VALUES ({values})
                    """.format(table=insert_tbl, keys=keys, values=values)
    try:
        insert_cursor.execute(insert_sql, tuple(insert_dict.values()))
        insert_conn.commit()
        last_id = insert_cursor.lastrowid
        print("数据库插入结果数据成功(๑•̀ㅂ•́)و✧并返回自增ID")
        return True, last_id
    except Exception as e:
        print(e, f"\n数据库表{insert_tbl}插入数据失败：", insert_dict)
        return False, 0
    finally:
        insert_cursor.close()
        insert_conn.close()



# 功能：修正sql查询得到的数据，为其加入表头
def fix_sql_df(results, col_name_list):
    col_name = []
    for head in col_name_list:
        col_name.append(head[0])
    df = pd.DataFrame(list(results), columns=col_name)
    return df


# 功能：查询某段sql，并返回一个df（带表头）
def read_sql2df(db, db_config, sql):
    # 连接导入数据库
    cursor, conn = get_cursor_times(db_config[db], 10000)
    cursor.execute(sql)
    results = cursor.fetchall()  # 用于返回多条数据
    col_name_list = cursor.description
    cursor.close()
    conn.close()
    df = fix_sql_df(results, col_name_list)
    return df


# 功能：执行一段sql，但不返回数据
def process_sql(db, db_config, sql):
    cursor, conn = get_cursor_times(db_config[db], 10000)
    try:
        cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print("执行失败：", e)
    finally:
        cursor.close()
        conn.close()

# 功能：筛选出list中有但是mysql的某个表中没有的元素（高效）
def filter_list_by_mysql(db, db_config, ori_list, tbl_name, col_name):
    time_now_str = get_time_now()
    tmp_tbl_name = f'tmp_ids_{time_now_str}'
    sql_create = f"CREATE TEMPORARY TABLE {tmp_tbl_name} ({col_name} VARCHAR(100) PRIMARY KEY) ENGINE=MEMORY;"
    sql_insert = f"INSERT INTO {tmp_tbl_name} ({col_name}) VALUES (%s)"
    sql_diff = f"""
    SELECT t.{col_name}
    FROM {tmp_tbl_name} AS t
    LEFT JOIN {tbl_name} AS u USING ({col_name})
    WHERE u.{col_name} IS NULL;
    """
    cursor, conn = get_cursor_times(db_config[db], 10000)
    cursor.execute(sql_create)
    cursor.executemany(sql_insert, [(x,) for x in ori_list])  # 一次批量写
    cursor.execute(sql_diff)
    missing_item = [row[0] for row in cursor.fetchall()]  # 数据库里没有的元素
    cursor.close()
    conn.close()
    return missing_item

# pgsql数据库相关
#########################################################################

# 功能：尝试多次连接pg数据库
def get_cursor_times_pg(config, times):
    cnt_conn = 0
    while cnt_conn < times:
        try:
            conn = psycopg2.connect(user=config['user'], password=config['password'], host=config['host'],
                                    port=config['port'], database=config['database'])
            cursor = conn.cursor()
            break
        except Exception as e:
            print("错误信息:", e)
            print(f'pgsql数据库连接失败，现进行第{cnt_conn}次重新连接，总共{times}次重新连接的机会...')
            sleep_some_seconds(20, 21, 1)
            cnt_conn += 1
    return cursor, conn


# 将dataframe写进pg数据库
def df2pgsql(db_config, db, insert_schema, insert_df, insert_tbl, cover_flag):
    if len(insert_df) != 0:
        insert_cursor, insert_conn = get_cursor_times_pg(db_config[db], 10000)
        # 获取表格的主键
        select_sql = f'''
        SET search_path = {insert_schema};
        SELECT a.attname
        FROM pg_index i
        JOIN pg_attribute a ON a.attnum = ANY(i.indkey) AND a.attrelid = i.indrelid
        WHERE i.indrelid = '{insert_tbl}'::regclass AND i.indisprimary'''
        insert_cursor.execute(select_sql)
        primary_keys = [row[0] for row in insert_cursor.fetchall()]
        # 将dataframe中的Nan用None代替
        insert_df = insert_df.where(insert_df.notnull(), None)
        columns = ', '.join(insert_df.columns)
        if cover_flag:
            placeholders = ', '.join(['%s'] * len(insert_df.columns))
            update_set = ', '.join([f"{col} = EXCLUDED.{col}" for col in insert_df.columns])
            # 生成 ON CONFLICT 子句
            on_conflict_clause = f"ON CONFLICT ({', '.join(primary_keys)}) DO UPDATE SET {update_set}"
        else:
            on_conflict_clause = f"ON CONFLICT ({', '.join(primary_keys)}) DO NOTHING"
        insert_sql = f"""
        INSERT INTO {insert_schema}.{insert_tbl} ({columns}) VALUES ({placeholders})
        {on_conflict_clause};
        """
        insert_data = [tuple(x) for x in insert_df.to_numpy()]
        # insert_sql = f"{'replace' if cover_flag else 'insert ignore'} into {insert_schema}.{insert_tbl} values ({', '.join(['%s']*len_cols)})"
        # insert_data = (tuple(row) for _, row in insert_df.iterrows())
        try:
            execute_batch(insert_cursor, insert_sql, insert_data)
            insert_conn.commit()
            print(f"数据插入{insert_schema}.{insert_tbl}成功(๑•̀ㅂ•́)و✧")
        except Exception as e:
            print(e, f"数据插入{insert_tbl}失败：", insert_data)
        finally:
            insert_cursor.close()
            insert_conn.close()
    else:
        print(f'无数据需要插入{insert_tbl}ε=(´ο｀*)))')

# 功能：将dict格式的数据插入insert_db的insert_tbl中，该方式适用于阿里云PAI平台
def dict2pgsql_ali(db_config, db, insert_schema, insert_dict, insert_tbl, conflict_col, update_flag):
    insert_cursor, insert_conn = get_cursor_times_pg(db_config[db], 10000)
    keys_series = ", ".join(insert_dict.keys())
    values_series = ", ".join(['%s'] * len(insert_dict.keys()))
    excluded_str = ', '.join([f'EXCLUDED.{key_name}' for key_name in insert_dict.keys()])
    conflict_str = ', '.join(conflict_col)
    if update_flag:
        insert_sql = f'''
                    INSERT INTO {insert_schema}.{insert_tbl} ({keys_series}) VALUES ({values_series}) 
                    ON CONFLICT ({conflict_str}) DO UPDATE SET ({keys_series}) = ({excluded_str})
                    '''
    else:
        insert_sql = f'''
                    INSERT INTO {insert_schema}.{insert_tbl} ({keys_series}) VALUES ({values_series})
                    ON CONFLICT ({conflict_str}) DO NOTHING
                    '''
    insert_cursor.execute(insert_sql, tuple(insert_dict.values()))
    insert_conn.commit()



# 功能：将dict格式的数据插入insert_db的insert_tbl中，该方式适用于华为云MA平台
# 输入样例
# insert_db = 'BA_USING'
# insert_tbl = 'autohome_policy_post'
# db_config = {"localhost": {'host': 'localhost', 'port': 3306, 'user': 'root', 'password': 'zwp650811', 'database': 'wom_db_local'},
#              "BA_USING": {'host': '10.26.241.164', 'port': 3306, 'user': 'BA_USING', 'password': 'BA_USING@2022', 'database': ip_database["BA_USING"]},
#              "BI_READ": {'host': '10.26.241.164', 'port': 3306, 'user': 'BI_READ', 'password': 'Bireader@1027', 'database': ip_database["BI_READ"]}}
def dict2pgsql(db_config, db, insert_schema, insert_dict, insert_tbl, cover_flag):
    insert_cursor, insert_conn = get_cursor_times_pg(db_config[db], 10000)
    keys = ", ".join(insert_dict.keys())
    values = ", ".join(['%s'] * len(insert_dict.keys()))
    if cover_flag:
        insert_sql = f'''
                    INSERT INTO {insert_schema}.{insert_tbl}({keys}) VALUES ({values}) ON DUPLICATE KEY UPDATE
                    '''
        update = ','.join([" {key} = values({key})".format(key=key) for key in insert_dict])
        insert_sql += update
    else:
        insert_sql = f"""
                    INSERT ignore INTO {insert_schema}.{insert_tbl}({keys}) VALUES ({values})
                    """
    # insert_data = tuple(insert_dict.values())
    try:
        # execute_batch(insert_cursor, insert_sql, insert_data)
        insert_cursor.execute(insert_sql, list(insert_dict.values()))
        insert_conn.commit()
        print("数据库插入结果数据成功(๑•̀ㅂ•́)و✧")
    except Exception as e:
        print(e, f"数据库表{insert_tbl}插入数据失败：", insert_dict)
    finally:
        insert_cursor.close()
        insert_conn.close()


# 功能：查询某段sql，并返回一个df（带表头）
def read_sql2df_pg(db, db_config, pgsql):
    # 连接导入数据库
    cursor, conn = get_cursor_times_pg(db_config[db], 10000)
    df = pd.read_sql_query(pgsql, conn)
    cursor.close()
    conn.close()
    return df

# 功能：执行一段pgsql，但不返回数据
def process_sql_pg(db, db_config, sql):
    cursor, conn = get_cursor_times_pg(db_config[db], 10000)
    try:
        cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print("执行失败：", e)
    finally:
        cursor.close()
        conn.close()




# 时间处理的模块
#########################################################################

# 功能：获取当前时间，返回yyyymmddhhmmss格式的时间字符串
# 输入：-
# 输出：time_out - yyyymmddhhmmss格式的时间字符串 - str
def get_time_now():
    time_out = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    return time_out


# 功能：获取当前时间，返回 %Y-%m-%d %H:%M:%S 格式的时间字符串
# 输入：-
# 输出：datetime_out - %Y-%m-%d %H:%M:%S格式的时间字符串 - str
def get_datetime_now():
    datetime_out = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return datetime_out

# 功能：比较两个"yyyy-mm-dd"格式时间字符串的大小
# 输入：time1 - 时间1 - str，如'2022-01-01'
#      time2 - 时间2 - str，如'2022-01-01'
#      identifier - 比较符，可选5种：大于（>)、大于等于（>=）、等于（=）、小于等于（<=）、小于（<） - str
# 输出：True/False - 比较结果 - bool
def compare_time_v1(time1: str, time2: str, identifier: str):
    s_time = time.mktime(time.strptime(time1, '%Y-%m-%d'))
    e_time = time.mktime(time.strptime(time2, '%Y-%m-%d'))
    if identifier == '>':
        if int(s_time) - int(e_time) > 0:
            return True
        else:
            return False
    if identifier == '<':
        if int(s_time) - int(e_time) < 0:
            return True
        else:
            return False
    if identifier == '>=':
        if int(s_time) - int(e_time) >= 0:
            return True
        else:
            return False
    if identifier == '<=':
        if int(s_time) - int(e_time) <= 0:
            return True
        else:
            return False
    if identifier == '=':
        if int(s_time) - int(e_time) == 0:
            return True
        else:
            return False


# 功能：比较两个"yyyy-mm-dd hh:mm:ss"格式时间字符串的大小
# 输入：time1 - 时间1 - str，如'2022-01-01 00:00:00'
#      time2 - 时间2 - str，如'2022-01-01 00:00:00'
#      identifier - 比较符，可选5种：大于（>)、大于等于（>=）、等于（=）、小于等于（<=）、小于（<） - str
# 输出：True/False - 比较结果 - bool
def compare_time_v2(time1: str, time2: str, identifier: str):
    s_time = time.mktime(time.strptime(time1, '%Y-%m-%d %H:%M:%S'))
    e_time = time.mktime(time.strptime(time2, '%Y-%m-%d %H:%M:%S'))
    if identifier == '>':
        if int(s_time) - int(e_time) > 0:
            return True
        else:
            return False
    if identifier == '<':
        if int(s_time) - int(e_time) < 0:
            return True
        else:
            return False
    if identifier == '>=':
        if int(s_time) - int(e_time) >= 0:
            return True
        else:
            return False
    if identifier == '<=':
        if int(s_time) - int(e_time) <= 0:
            return True
        else:
            return False
    if identifier == '=':
        if int(s_time) - int(e_time) == 0:
            return True
        else:
            return False


# 功能：将自epoch以来经过的秒数字符串转化为"yyyy-mm-dd hh:mm:ss"格式
# 输入：stamp_string - 自epoch以来经过的秒数字符串 - str，如'1545925769'
# 输出：date_time - "yyyy-mm-dd hh:mm:ss"格式的时间字符串 - str
def stamp2date(stamp_string: str):
    time_array = time.localtime(int(stamp_string))
    date_time = time.strftime("%Y-%m-%d %H:%M:%S", time_array)
    return date_time


# 功能：随机sleep n - m 秒，k表示小数点后k位的时间
# 输入：m - 最短时间 - int
#      n - 最长时间 - int
#      k - 小数点后k位 - int
# 输出：True - 表示完成 - bool
def sleep_some_seconds(m: int, n: int, k: int):
    time.sleep(round(random.uniform(m, n), k))
    return True


# 功能：计算sec（秒数）与当前时间的时间间隔，并且输出指定格式的字符串形式
# 输入：sec - 最短时间 - int/str
#      date_format - 指定的输出格式 - str
# 输出：time_gap - 指定格式的时间间隔 - str
def change2date(sec, date_format: str):
    sec = int(sec)
    time_gap = time.time() - sec
    time_gap = time.strftime(date_format, time.localtime(time_gap))
    return time_gap


# 功能：将n天前/n小时前/n周前转化为%Y-%m-%d %H:%M:%S的字符串格式
# 输入：before - 时间表达形式 - str
# 输出：datetime_str - 格式化的时间字符串 - str
def before2datetime(before: str):
    before_str = re.findall(r"(\d{4}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2}:\d{1,2})", before)
    if before_str != []:
        datetime_str = before_str[0]
    else:
        if '前天' in before:
            sec = 2 * 86400
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '周前' in before:
            a = re.findall('(.*?)天前', before)
            sec = int(a[0]) * 7 * 86400
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '天前' in before:
            a = re.findall('(.*?)天前', before)
            sec = int(a[0]) * 86400
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '小时前' in before:
            a = re.findall('(.*?)小时前', before)
            sec = int(a[0]) * 3600
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '分钟前' in before:
            a = re.findall('(.*?)分钟前', before)
            sec = int(a[0]) * 60
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '秒钟前' in before:
            a = re.findall('(.*?)秒钟前', before)
            sec = int(a[0])
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        elif '刚刚' in before:
            sec = 0
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
        else:
            print('意外的时间表达方式！！！！！！return当前时间')
            sec = 0
            datetime_str = change2date(sec, date_format='%Y-%m-%d %H:%M:%S')
    return datetime_str


# 功能：将yyyymmdd-yyyymmdd格式的字符串拆分，并转化为前后两个时间的起始及结束datetime
# 输入：text - 两个时间的表达形式，yyyymmdd-yyyymmdd - str
# 输出：True/False - 若输入不符合格式要求，则直接返回False - bool
#      start_dt - 起始的dateframe字符串
#      end_dt - 结束的dateframe字符串
def split_yyyymmdd(text: str):
    date_list = text.split('-')
    start_text = date_list[0]
    end_text = date_list[1]
    if len(date_list) != 2 or len(start_text) != 8 or len(end_text) != 8 or int(start_text) > int(end_text):
        return False, '', ''
    else:
        start_dt = start_text[:4] + '-' + start_text[4:6] + '-' + start_text[6:8] + ' 00:00:00'
        end_dt = end_text[:4] + '-' + end_text[4:6] + '-' + end_text[6:8] + ' 23:59:59'
        return True, start_dt, end_dt


# 功能：根据当前时间，反馈现在是周几
# 输入：-
# 输出：weekday - 周几 - str
def get_day_of_week():
    day_of_week = datetime.datetime.now().weekday() + 1
    if day_of_week == 1:
        weekday = '周一'
    elif day_of_week == 2:
        weekday = '周二'
    elif day_of_week == 3:
        weekday = '周三'
    elif day_of_week == 4:
        weekday = '周四'
    elif day_of_week == 5:
        weekday = '周五'
    elif day_of_week == 6:
        weekday = '周六'
    else:
        weekday = '周日'
    return weekday


# 功能：根据输入的date格式，得到n_day前的date格式日期
# 输入：date_start - 起始的日期 - datetime.date
#      n_day - n天前 - int
# 输出：date_before - n_day天之前的date - datetime.date
def get_date_before(date_start, n_day):
    days = datetime.timedelta(days=n_day)
    date_before = date_start - days
    return date_before


# excel处理模块
#########################################################################

# v1版本，后因适配MA的私域分析代码而进行修改
# 功能：将一个df写入到excel的指定位置
# def wirte_df2excel(goal_sheet, start_row, start_col, df_input, align_style, header_include=False):
#     col_idx = start_col
#     # 如果表头也要写入，就把表头放在第一行
#     if header_include:
#         df_input.T.reset_index().rename(columns={'index': '列名'}).T
#     for row in df_input.iteritems():
#         input_list = row[1].apply(str).to_list()
#         row_idx = start_row
#         for ele in input_list:
#             try:
#                 goal_sheet.cell(row=row_idx, column=col_idx).value = ele
#                 goal_sheet.cell(row=row_idx, column=col_idx).alignment = align_style
#             except:
#                 print('无法写入：', ele)
#             row_idx += 1
#         col_idx += 1

# 功能：将一个df写入到excel的指定位置
def wirte_df2excel(goal_sheet, start_row, start_col, df_input, align_style, header_include=False):
    row_idx = start_row
    # 如果表头也要写入，就把表头放在第一行
    if header_include:
        df_input.T.reset_index().rename(columns={'index': '列名'}).T
    for index, row in df_input.iterrows():
        # input_list = row[1].apply(str).to_list()
        input_list = row.to_list()
        col_idx = start_col
        for ele in input_list:
            try:
                goal_sheet.cell(row=row_idx, column=col_idx).value = str(ele)
                goal_sheet.cell(row=row_idx, column=col_idx).alignment = align_style
            except:
                print('无法写入：', ele)
            col_idx += 1
        row_idx += 1

# 功能：将一个series写入到excel的指定位置
def write_series2excel(goal_sheet, start_row, start_col, series_input, align_style, value_only=False, top_n=10000000):
    cnt = 0
    for idx, val in series_input.items():
        if val is not None:
            # 如果要把索引也记录上
            if not value_only:
                goal_sheet.cell(row=start_row, column=start_col).value = idx
                goal_sheet.cell(row=start_row, column=start_col+1).value = val
                goal_sheet.cell(row=start_row, column=start_col).alignment = align_style
                goal_sheet.cell(row=start_row, column=start_col+1).alignment = align_style
            else:
                goal_sheet.cell(row=start_row, column=start_col).value = val
                goal_sheet.cell(row=start_row, column=start_col).alignment = align_style
        cnt += 1
        if cnt >= top_n:
            break
        start_row += 1


# 功能：将一个数值写入到excel的指定位置
def write_val2excel(goal_sheet, row_pos, col_pos, val, align_style):
    goal_sheet.cell(row=row_pos, column=col_pos).value = val
    goal_sheet.cell(row=row_pos, column=col_pos).alignment = align_style


# 自动化处理模块
#########################################################################

# 功能：自动发送指定邮件给指定的人
# 输入：f_origin - 宏文件的地址 - str
#      to_list - 发送的用户清单 - list
#      cc_list - 抄送的用户清单 - list
#      attached_list - 附件的地址list，支持最多两个附件 - list
#      subject_text - 邮件标题 - str
#      content_text - 正文内容 - str
# 输出：-
def send_mail(f_origin, to_list, cc_list, attached_list, subject_text, content_text):
    if len(attached_list) > 2:
        print('附件数过多，请缩减至2个及以下的附件。')
        sys.quit()
    to_string = ';'.join(to_list)
    cc_string = ';'.join(cc_list)
    # 更新xlsm的内容信息
    print('开始对xlsm的内容信息进行更新...')
    wb = load_workbook(f_origin, keep_vba=True)
    ws_1 = wb.get_sheet_by_name('发送清单')
    ws_1.cell(row=2, column=1).value = to_string
    ws_1.cell(row=2, column=2).value = cc_string
    n_idx = 3
    for attached in attached_list:
        ws_1.cell(row=2, column=n_idx).value = attached
        n_idx += 1
    ws_2 = wb.get_sheet_by_name('正文')
    ws_2.cell(row=1, column=2).value = subject_text
    ws_2.cell(row=2, column=2).value = content_text
    f_goal = f_origin[:-5] + f'_tmp_{get_time_now()}.xlsm'
    wb.save(f_goal)
    time.sleep(3)
    # 通过excel的宏进行发送
    app = xw.App(visible=False, add_book=False)
    app.screen_updating = False
    # 宏文件
    wb = app.books.open(f_goal)
    # 执行宏文件
    marco = wb.macro('Mygirl')  # 这里需要特别注意，当你的宏名是唯一的时候，不需要写模块名，但如果模块名和宏名重复，需要补全
    marco()
    time.sleep(3)  # 等待macro运行完毕，具体等待时长视宏运行的时间
    # 保存关闭退出
    wb.save()
    wb.close()
    app.quit()
    print('邮件发送成功(๑•̀ㅂ•́)و✧')
    os.remove(f_goal)

# 功能：根据color的色号，打印出带有颜色的文本，参考https://blog.csdn.net/XianZhe_/article/details/113075983
# 输入：text - 文本内容 - str
#      color_num - 色号，默认为31红色 - int，常见色号有31-红色、32-绿色、33-黄色
# 输出：-
def print_text_w_color(text, color_num=31):
    if isinstance(color_num, int):
        color = str(color_num)
    print(f"\033[1;{color}m{text}\033[0m")

# 功能：将关键词字符串转化为list，用于检索
def keyword2list(keyword_str):
    if keyword_str:
        return keyword_str.split('、')
    else:
        return []