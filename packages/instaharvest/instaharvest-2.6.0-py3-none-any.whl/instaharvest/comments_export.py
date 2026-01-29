"""
Instagram Scraper - Comments Export Utilities
Real-time export of comments to JSON and Excel

Features:
- Real-time JSON export
- Real-time Excel export with flat structure
- Nested comments support
- Progress tracking
- Automatic file management
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
import logging

from .config import ScraperConfig
from .comment_scraper import PostCommentsData
from .models import CommentData, CommentAuthor, Collaborator


class CommentsExporter:
    """
    Real-time comments exporter

    Supports:
    - JSON export (hierarchical structure with replies)
    - Excel export (flat structure for analysis)
    - Real-time writing
    - Multiple posts aggregation
    """

    def __init__(
        self,
        username: str,
        logger: Optional[logging.Logger] = None,
        config: Optional[ScraperConfig] = None,
        export_json: bool = True,
        export_excel: bool = True
    ):
        """
        Initialize comments exporter

        Args:
            username: Instagram username (for filename)
            logger: Logger instance
            config: ScraperConfig instance
            export_json: Enable JSON export
            export_excel: Enable Excel export
        """
        self.username = username
        self.logger = logger or logging.getLogger(__name__)
        self.config = config or ScraperConfig()
        self.export_json = export_json
        self.export_excel = export_excel

        # Storage for all comments data
        self.posts_data: List[PostCommentsData] = []
        self.excel_rows: List[Dict[str, Any]] = []

        # File paths
        self.json_file = Path(
            self.config.comments_json_filename_pattern.format(
                username=username,
                post_id='all'
            ).replace('_all', '')
        )
        self.excel_file = Path(
            self.config.comments_excel_filename_pattern.format(username=username)
        )

        # Initialize files
        if self.export_excel:
            self._create_excel_file()

        self.logger.info(f"Comments exporter initialized for @{username}")
        if self.export_json:
            self.logger.info(f"  JSON: {self.json_file}")
        if self.export_excel:
            self.logger.info(f"  Excel: {self.excel_file}")

    def _create_excel_file(self) -> None:
        """Create initial Excel file with headers"""
        try:
            df = pd.DataFrame(columns=self.config.comments_excel_columns)
            df.to_excel(self.excel_file, index=False, engine='openpyxl')
            self.logger.debug(f"Created Excel file: {self.excel_file}")
        except Exception as e:
            self.logger.error(f"Failed to create Excel file: {e}")
            raise

    def add_post_comments(self, post_comments: PostCommentsData) -> None:
        """
        Add comments from a single post

        Args:
            post_comments: PostCommentsData object with all comments
        """
        self.posts_data.append(post_comments)

        # Add to Excel rows (flat structure)
        for comment in post_comments.comments:
            self._add_comment_to_excel(post_comments, comment)

            # Add replies
            for reply in comment.replies:
                self._add_comment_to_excel(post_comments, reply)

        # Save to files
        if self.export_json:
            self._save_json()

        if self.export_excel:
            self._save_excel()

        self.logger.info(
            f"Added {post_comments.total_comments_scraped} comments, "
            f"{post_comments.total_replies_scraped} replies from {post_comments.post_url}"
        )

    def add_single_comment(
        self,
        post_url: str,
        post_id: str,
        comment: CommentData
    ) -> None:
        """
        Add a single comment (real-time mode)

        Args:
            post_url: Post URL
            post_id: Post ID
            comment: CommentData object
        """
        # Create minimal PostCommentsData for context
        post_data = PostCommentsData(
            post_url=post_url,
            post_id=post_id,
            total_comments_scraped=1,
            total_replies_scraped=len(comment.replies),
            comments=[comment]
        )

        # Add to Excel
        self._add_comment_to_excel(post_data, comment)
        for reply in comment.replies:
            self._add_comment_to_excel(post_data, reply)

        # Save Excel immediately (real-time)
        if self.export_excel:
            self._save_excel()

    def _add_comment_to_excel(
        self,
        post_data: PostCommentsData,
        comment: CommentData
    ) -> None:
        """Add a single comment to Excel rows"""
        author = comment.author if isinstance(comment.author, CommentAuthor) else CommentAuthor(**comment.author)

        row = {
            'Post URL': post_data.post_url,
            'Post ID': post_data.post_id,
            'Comment ID': comment.id,
            'Author Username': author.username,
            'Author Verified': author.is_verified,
            'Comment Text': comment.text,
            'Likes Count': comment.likes_count,
            'Reply Count': comment.reply_count,
            'Timestamp': comment.timestamp,
            'Timestamp ISO': comment.timestamp_iso,
            'Is Reply': comment.is_reply,
            'Parent Comment ID': comment.parent_id or '',
            'Comment URL': comment.permalink,
            'Scraped At': post_data.scraped_at
        }

        self.excel_rows.append(row)

    def _save_json(self) -> None:
        """Save all data to JSON file"""
        try:
            # Count total collaborators
            total_collaborators = sum(len(p.collaborators) for p in self.posts_data)

            data = {
                'username': self.username,
                'total_posts': len(self.posts_data),
                'total_comments': sum(p.total_comments_scraped for p in self.posts_data),
                'total_replies': sum(p.total_replies_scraped for p in self.posts_data),
                'total_collaborators': total_collaborators,
                'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'posts': [p.to_dict() for p in self.posts_data]
            }

            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

            self.logger.debug(f"Saved JSON: {self.json_file}")

        except Exception as e:
            self.logger.error(f"Failed to save JSON: {e}")

    def _save_excel(self) -> None:
        """Save all rows to Excel file"""
        try:
            df = pd.DataFrame(self.excel_rows, columns=self.config.comments_excel_columns)
            df.to_excel(self.excel_file, index=False, engine='openpyxl')
            self.logger.debug(f"Saved Excel: {self.excel_file}")
        except Exception as e:
            self.logger.error(f"Failed to save Excel: {e}")

    def finalize(self) -> None:
        """Finalize export (adjust columns, save final data)"""
        try:
            # Final save
            if self.export_json:
                self._save_json()

            if self.export_excel:
                self._save_excel()
                self._auto_adjust_columns()

            # Print summary
            total_comments = sum(p.total_comments_scraped for p in self.posts_data)
            total_replies = sum(p.total_replies_scraped for p in self.posts_data)
            total_collaborators = sum(len(p.collaborators) for p in self.posts_data)

            self.logger.info(f"\n{'='*60}")
            self.logger.info("COMMENTS EXPORT COMPLETE!")
            self.logger.info(f"Username: @{self.username}")
            self.logger.info(f"Posts processed: {len(self.posts_data)}")
            self.logger.info(f"Total comments: {total_comments}")
            self.logger.info(f"Total replies: {total_replies}")
            self.logger.info(f"Total collaborators: {total_collaborators}")
            self.logger.info(f"Total rows in Excel: {len(self.excel_rows)}")
            if self.export_json:
                self.logger.info(f"JSON file: {self.json_file}")
            if self.export_excel:
                self.logger.info(f"Excel file: {self.excel_file}")
            self.logger.info(f"{'='*60}")

        except Exception as e:
            self.logger.error(f"Finalize failed: {e}")

    def _auto_adjust_columns(self) -> None:
        """Auto-adjust Excel column widths"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(self.excel_file)
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.excel_file)
            self.logger.debug("Auto-adjusted Excel columns")

        except Exception as e:
            self.logger.warning(f"Failed to auto-adjust columns: {e}")

    def get_statistics(self) -> Dict[str, Any]:
        """Get export statistics"""
        return {
            'username': self.username,
            'total_posts': len(self.posts_data),
            'total_comments': sum(p.total_comments_scraped for p in self.posts_data),
            'total_replies': sum(p.total_replies_scraped for p in self.posts_data),
            'total_collaborators': sum(len(p.collaborators) for p in self.posts_data),
            'excel_rows': len(self.excel_rows),
            'json_file': str(self.json_file) if self.export_json else None,
            'excel_file': str(self.excel_file) if self.export_excel else None
        }


class RealTimeCommentsExporter:
    """
    Real-time comments exporter with callback support

    Saves comments as they are scraped, not waiting for batch completion
    """

    def __init__(
        self,
        username: str,
        logger: Optional[logging.Logger] = None,
        config: Optional[ScraperConfig] = None,
        batch_size: int = 5
    ):
        """
        Initialize real-time exporter

        Args:
            username: Instagram username
            logger: Logger instance
            config: ScraperConfig instance
            batch_size: Save to disk every N comments
        """
        self.username = username
        self.logger = logger or logging.getLogger(__name__)
        self.config = config or ScraperConfig()
        self.batch_size = batch_size

        self.comments_buffer: List[Dict[str, Any]] = []
        self.all_comments: List[Dict[str, Any]] = []
        self.post_id: str = ''
        self.post_url: str = ''

        # Files
        self.json_file = Path(
            self.config.comments_json_filename_pattern.format(
                username=username,
                post_id='realtime'
            )
        )
        self.excel_file = Path(
            self.config.comments_excel_filename_pattern.format(username=username)
        )

        self._create_excel_file()
        self.logger.info(f"Real-time comments exporter ready")

    def _create_excel_file(self) -> None:
        """Create Excel file with headers"""
        try:
            df = pd.DataFrame(columns=self.config.comments_excel_columns)
            df.to_excel(self.excel_file, index=False, engine='openpyxl')
        except Exception as e:
            self.logger.error(f"Failed to create Excel file: {e}")

    def set_current_post(self, post_url: str, post_id: str) -> None:
        """Set current post being scraped"""
        self.post_url = post_url
        self.post_id = post_id

    def on_comment_scraped(self, count: int, comment: CommentData) -> None:
        """
        Callback for when a comment is scraped

        This can be passed as progress_callback to CommentScraper.scrape()
        """
        # Convert to dict for storage
        comment_dict = comment.to_dict()
        comment_dict['post_url'] = self.post_url
        comment_dict['post_id'] = self.post_id

        self.comments_buffer.append(comment_dict)
        self.all_comments.append(comment_dict)

        # Check if should save batch
        if len(self.comments_buffer) >= self.batch_size:
            self._save_batch()

        self.logger.debug(
            f"[{count}] @{comment.author.username}: "
            f"{comment.text[:30]}... ({comment.likes_count} likes)"
        )

    def _save_batch(self) -> None:
        """Save current batch to files"""
        if not self.comments_buffer:
            return

        try:
            # Save to Excel (append mode)
            df = pd.DataFrame(self.all_comments)
            df.to_excel(self.excel_file, index=False, engine='openpyxl')

            # Save to JSON
            self._save_json()

            self.comments_buffer.clear()
            self.logger.debug(f"Saved batch, total: {len(self.all_comments)} comments")

        except Exception as e:
            self.logger.error(f"Failed to save batch: {e}")

    def _save_json(self) -> None:
        """Save all comments to JSON"""
        try:
            data = {
                'username': self.username,
                'total_comments': len(self.all_comments),
                'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'comments': self.all_comments
            }

            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

        except Exception as e:
            self.logger.error(f"Failed to save JSON: {e}")

    def finalize(self) -> None:
        """Save any remaining comments and finalize files"""
        if self.comments_buffer:
            self._save_batch()

        self.logger.info(
            f"Real-time export complete: {len(self.all_comments)} comments saved"
        )


def export_comments_to_json(
    comments_data: Union[PostCommentsData, List[PostCommentsData]],
    filename: str,
    logger: Optional[logging.Logger] = None
) -> bool:
    """
    Export comments to JSON file

    Args:
        comments_data: Single PostCommentsData or list
        filename: Output filename
        logger: Logger instance

    Returns:
        True if successful
    """
    logger = logger or logging.getLogger(__name__)

    try:
        # Handle single or multiple
        if isinstance(comments_data, PostCommentsData):
            data = comments_data.to_dict()
        else:
            data = {
                'total_posts': len(comments_data),
                'total_comments': sum(p.total_comments_scraped for p in comments_data),
                'total_replies': sum(p.total_replies_scraped for p in comments_data),
                'total_collaborators': sum(len(p.collaborators) for p in comments_data),
                'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'posts': [p.to_dict() for p in comments_data]
            }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info(f"Comments exported to JSON: {filename}")
        return True

    except Exception as e:
        logger.error(f"JSON export failed: {e}")
        return False


def export_comments_to_excel(
    comments_data: Union[PostCommentsData, List[PostCommentsData]],
    filename: str,
    config: Optional[ScraperConfig] = None,
    logger: Optional[logging.Logger] = None
) -> bool:
    """
    Export comments to Excel file

    Args:
        comments_data: Single PostCommentsData or list
        filename: Output filename
        config: ScraperConfig instance
        logger: Logger instance

    Returns:
        True if successful
    """
    logger = logger or logging.getLogger(__name__)
    config = config or ScraperConfig()

    try:
        rows = []

        # Handle single or multiple
        if isinstance(comments_data, PostCommentsData):
            posts_list = [comments_data]
        else:
            posts_list = comments_data

        # Flatten all comments
        for post_data in posts_list:
            for comment in post_data.get_all_comments_flat():
                author = comment.author if isinstance(comment.author, CommentAuthor) else CommentAuthor(**comment.author)

                row = {
                    'Post URL': post_data.post_url,
                    'Post ID': post_data.post_id,
                    'Comment ID': comment.id,
                    'Author Username': author.username,
                    'Author Verified': author.is_verified,
                    'Comment Text': comment.text,
                    'Likes Count': comment.likes_count,
                    'Reply Count': comment.reply_count,
                    'Timestamp': comment.timestamp,
                    'Timestamp ISO': comment.timestamp_iso,
                    'Is Reply': comment.is_reply,
                    'Parent Comment ID': comment.parent_id or '',
                    'Comment URL': comment.permalink,
                    'Scraped At': post_data.scraped_at
                }
                rows.append(row)

        # Create DataFrame and save
        df = pd.DataFrame(rows, columns=config.comments_excel_columns)
        df.to_excel(filename, index=False, engine='openpyxl')

        logger.info(f"Comments exported to Excel: {filename} ({len(rows)} rows)")
        return True

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return False
