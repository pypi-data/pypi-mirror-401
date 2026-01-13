# -*- coding: utf-8 -*-
"""
Data-loading utility functions for protograf
"""
# lib
import csv
import logging
import os

# third-party
from openpyxl import load_workbook
import requests
import xlrd

# local
from protograf.utils.tools import _lower, script_path, coordinate_to_tuple
from protograf.utils.messaging import feedback


log = logging.getLogger(__name__)


def json_strings_to_numbers(json_data: str | dict | list):
    """Iteratively convert JSON data into numbers, if possible.

    Doc Test:

    >>> json_strings_to_numbers({})
    {}
    >>> json_strings_to_numbers([])
    []
    >>> json_strings_to_numbers('[]')
    '[]'
    >>> json_strings_to_numbers('["a", "1", "2.3"]')
    '["a", "1", "2.3"]'
    >>> json_strings_to_numbers(["a", "1", "2.3", {"a": "0.123"}])
    ['a', 1, 2.3, {'a': 0.123}]
    """
    if isinstance(json_data, dict):
        for key, value in json_data.items():
            json_data[key] = json_strings_to_numbers(value)
    elif isinstance(json_data, list):
        for i, item in enumerate(json_data):
            json_data[i] = json_strings_to_numbers(item)
    elif isinstance(json_data, str):
        try:
            return int(json_data)
        except ValueError:
            try:
                return float(json_data)
            except ValueError:
                return json_data
    return json_data


def load_data(datasource=None, **kwargs):
    """
    Load data from a 'tabular' source (CSV, XLS, XSLX) into a dict
    """
    dataset = {}
    log.debug("Load data from a 'tabular' source (CSV, XLS, XSLX) %s", datasource)
    if datasource:
        filename, file_ext = os.path.splitext(datasource)
        if not filename:
            feedback("Unable to process a file without valid filename!", True)
        if file_ext and _lower(file_ext) == ".csv":
            headers = kwargs.get("headers", None)
            selected = kwargs.get("selected", None)
            dataset = open_csv(datasource, headers=headers, selected=selected)
        elif file_ext and _lower(file_ext) in [".xls", ".xlsx"]:
            headers = kwargs.get("headers", None)
            selected = kwargs.get("selected", None)
            sheet = kwargs.get("sheet", 0)
            sheetname = kwargs.get("sheetname", None)
            cells = kwargs.get("cells", None)
            dataset = open_excel(
                datasource,
                sheet=sheet,
                sheetname=sheetname,
                cells=cells,
                headers=headers,
            )
        else:
            feedback('Unable to process a file %s of type "%s"' % (filename, file_ext))
    return dataset


def load_googlesheet(sheet, **kwargs):
    """
    Read data from a Google Sheet into a dict

    Kwargs:

    - api_key (str): an API Key available from Google
    - spreadsheet_id (str): the unique ID (a mix of numbers and letters) which
      is randomly assigned by Google to a Google Sheet
    - sheet_name (str): the name of the tab in the Google Sheet storing the data;
      defaults to "Sheet1"
    """
    data_list = []
    spreadsheet_id = sheet
    api_key = kwargs.get("api_key", None)
    if not api_key:
        feedback('Cannot access a Google Sheet without an "api_key"', True)
    sheet_name = kwargs.get("name", None)
    if not sheet_name:
        feedback('Using default tab name of "Sheet1"', False, True)
        sheet_name = "Sheet1"
    log.debug("Load data from a Google Sheet %s", sheet)

    if sheet:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{sheet_name}?alt=json&key={api_key}"
        try:
            response = requests.get(url)
            response.raise_for_status()  # raise exception for HTTP errors
            raw_data = response.json()
            data_dict = json_strings_to_numbers(raw_data)
        except requests.exceptions.RequestException as err:
            feedback(f"Unable to load Google Sheet: {err}")
            return []

    if data_dict:
        _data_list = data_dict.get("values")
        if _data_list:
            keys = _data_list[0]  # get keys/names from first sub-list
            dict_list = []
            for row in _data_list[1:]:
                _dict = {}
                for idx, key in enumerate(keys):
                    # handles the "bug" that Sheet does not return all of a row
                    # if the last cell is "empty"
                    _dict[key] = row[idx] if idx < len(row) else ""
                dict_list.append(_dict)
            return dict_list

    return data_list


def open_csv(filename: str = None, headers: list = None, selected: list = None):
    """Read data from CSV file into a list of dictionaries

    Args:

    - filename (str): path to CSV file
    - headers (list): a list of strings to use instead of the first row
    - selected (list): a list of desired rows e.g. [2,4,7]
    """
    if not filename:
        feedback("A valid CSV filename must be supplied!")

    dict_list = []
    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find CSV "{filename}", including in {filepath}')

    try:
        csv_filename = _file_with_path or norm_filename
        if headers:
            reader = csv.DictReader(open(csv_filename), fieldnames=headers)
        else:
            reader = csv.DictReader(open(csv_filename))
        for key, item in enumerate(reader):
            if not selected:
                dict_list.append(item)
            else:
                if key + 1 in selected:
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open CSV "%s"' % csv_filename)
    return dict_list


def open_excel(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from an Excel file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    - headers (list): strings to use instead of the first row
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        result = "" if value is None else value
        return result

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')
    else:
        _file_with_path = norm_filename  # in same dir as script!

    if sheet is None and sheetname is None:
        feedback(
            f'Access to Excel file "{filename}" requires either the sheet number or the sheet name!',
            True,
        )

    _, file_ext = os.path.splitext(_file_with_path)
    if file_ext == ".xls":
        return open_xls(
            filename=filename,
            sheet=sheet,
            sheetname=sheetname,
            cells=cells,
            headers=headers,
        )
    elif file_ext == ".xlsx":
        return open_xlsx(
            filename=filename,
            sheet=sheet,
            sheetname=sheetname,
            cells=cells,
            headers=headers,
        )
    else:
        feedback(f'Cannot process data files with an extension of "{file_ext}".', True)


def open_xlsx(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from XLSX file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - headers (list): strings to use instead of the first row
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        result = "" if value is None else value
        return result

    def maximum_columns(worksheet):
        """Find the actual max column by iterating through cells.

        Notes:

            worksheet.max_column can have an inflated value if
            spreadsheet was created or modified by other software
        """
        actual_max_column = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if cell.column > actual_max_column:
                        actual_max_column = cell.column
        return actual_max_column

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    dict_list = []

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')

    try:
        excel_filename = _file_with_path or norm_filename
        book = load_workbook(excel_filename, data_only=True)
        if sheetname:
            sheet = book[sheetname]
        elif sheet is not None:
            sheet = int(sheet) - 1 if sheet > 0 else int(sheet)
            _sheetname = book.sheetnames[sheet]
            sheet = book[_sheetname]
        else:
            feedback(
                f'Access to Excel file "{filename}" requires either the sheet number or the sheet name!',
                True,
            )
        start = 1

        max_cols = maximum_columns(sheet)
        if not headers:
            keys = [
                sheet.cell(1, col_index).value for col_index in range(1, max_cols + 1)
            ]
            if None in keys:
                feedback(
                    f'Please assign headers to all columns in Excel file "{filename}"'
                    " and/or delete any empty columns.",
                    True,
                )
            start = 2
        else:
            start = 1
            keys = headers
        if len(keys) < max_cols:
            feedback(
                'Too few headers supplied for the existing columns in "%s"' % filename
            )
        else:
            dict_list = []
            if cells:
                _cells = cells.split(":")
                cell_range = sheet[_cells[0] : _cells[1]]
                for row_index, row in enumerate(cell_range):
                    item = {
                        keys[col_index]: cleaned(cell.value)
                        for col_index, cell in enumerate(row)
                    }
                    dict_list.append(item)
            else:
                for row_index in range(start, sheet.max_row + 1):
                    item = {
                        keys[col_index - 1]: cleaned(
                            sheet.cell(row=row_index, column=col_index).value
                        )
                        for col_index in range(1, max_cols + 1)
                    }
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open Excel "%s"' % excel_filename)
    except IndexError:
        feedback('Unable to open sheet "%s"' % (sheet or sheetname))
    except Exception:
        feedback('Unable to open or process sheet "%s"' % (sheet or sheetname))
    return dict_list


def open_xls(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from XLS file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - headers (list): strings to use instead of the first row
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        # if isinstance(value, str):
        #     return value.encode('utf8')
        return value

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    dict_list = []

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')

    try:
        excel_filename = _file_with_path or norm_filename
        book = xlrd.open_workbook(excel_filename)
        if sheet:
            sheet = sheet - 1
            sheet = book.sheet_by_index(sheet)
        elif sheetname:
            sheet = book.sheet_by_name(sheetname)
        else:
            sheet = book.sheet_by_index(0)
        start = 1
        if not headers:
            keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
        else:
            start = 0
            keys = headers
        if len(keys) < sheet.ncols:
            feedback(
                'Too few headers supplied for the existing columns in "%s"' % filename
            )
        else:
            dict_list = []
            if cells:
                _cells = cells.split(":")
                _topleft = coordinate_to_tuple(_cells[0], True)  # (col, row)
                _btmrite = coordinate_to_tuple(_cells[1], True)  # (col, row)
                for row_index in range(_topleft[1], _btmrite[1] + 1):
                    item = {
                        keys[col_index]: cleaned(sheet.cell(row_index, col_index).value)
                        for col_index in range(_topleft[0], _btmrite[0] + 1)
                    }
                    dict_list.append(item)
            else:
                for row_index in range(start, sheet.nrows):
                    item = {
                        keys[col_index]: cleaned(sheet.cell(row_index, col_index).value)
                        for col_index in range(sheet.ncols)
                    }
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open Excel "%s"' % excel_filename)
    except IndexError:
        feedback('Unable to open sheet "%s"' % (sheet or sheetname))
    except xlrd.biffh.XLRDError:
        feedback('Unable to open sheet "%s"' % sheetname)
    return dict_list
