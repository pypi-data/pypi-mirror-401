"""
Write and read Excel tables.
"""
from __future__ import annotations

import logging
import os
import re
from copy import copy
from enum import Enum, Flag
from pathlib import Path
from typing import TYPE_CHECKING, Any, overload

from openpyxl import DEFUSEDXML, Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from typing import Literal, TypeVar
    T = TypeVar('T')

NO_FORMULA = object()

_logger = logging.getLogger(__name__)



def is_excel_path(path: str|os.PathLike, *, accept_table_suffix = False):
    """
    Return True if path is a str or Path object ending with `.xlsx` (case-insensitive),
    or, if `accept_table_suffix` is set to `True`, return true if path ends with `.xlsx:Table` where Table is the name of a table.

    Return False otherwise.
    """
    if isinstance(path, os.PathLike):
        path = str(path)
    elif not isinstance(path, str):
        return False
    
    return True if re.search(r'\.xlsx(?:\:[\w\d\{\}_]+)?$' if accept_table_suffix else r'\.xlsx$', path, re.IGNORECASE) else False


def split_excel_path(path: str|os.PathLike, *, default_table_name: str|None = None) -> tuple[Path|None,str|None]:
    """
    Return tuple (workbook_path, table_name) from an Excel path (using `default_table_name` if the Excel path does not end with `:Table`).
    
    Return tuple (None, None) if this is not an Excel path.
    """
    if isinstance(path, os.PathLike):
        path = str(path)
    elif not isinstance(path, str):
        raise ValueError(f'Invalid path type: {type(path)}')
    
    m = re.match(r'^(.+\.xlsx)(?:\:([\w\d\{\}_]*))?$', path, re.IGNORECASE)
    if not m:
        return (None, None)
    
    return (Path(m[1]), m[2] if m[2] else default_table_name)


class ExcelWorkbook:
    _instances: dict[str,ExcelWorkbook] = {}
    """ Workbooks per canonical path """
    
    _defusedxml_alert_emitted = False

    def __init__(self, path: str|os.PathLike, *, formula = NO_FORMULA, save_as: str|os.PathLike|None = None):
        """
        - `formula`: value to return if a cell is a formula. If not set, return the formula itself (starting with '=').
        - `save_as`: save at this path instead of the original path.
        """
        if not DEFUSEDXML and not self.__class__._defusedxml_alert_emitted:
            _logger.warning("Install package `defusedxml` (in addition to `openpyxl`) to guard against quadratic blowup or billion laughs xml attacks")
            self.__class__._defusedxml_alert_emitted = True

        if not isinstance(path, (str,os.PathLike)):
            raise TypeError(f"path: {type(path).__name__}")
        self.path = path

        self.needs_save = False
        self._formula = formula
        self._save_as = save_as
        
        self._pyxl_workbook: Workbook|None = None
        self._tables: dict[str,ExcelTable] = {}

    @classmethod
    def get_instance(cls, path: str|os.PathLike, *, formula = NO_FORMULA, save_as: str|os.PathLike|None = None) -> ExcelWorkbook:
        canonical_path = os.path.realpath(path)
        if not canonical_path in cls._instances:
            cls._instances[canonical_path] = ExcelWorkbook(path, formula=formula, save_as=save_as)
        return cls._instances[canonical_path]

    def __enter__(self):
        return self

    def __exit__(self, exc_type = None, exc_value = None, exc_traceback = None):
        self.close()

    @property
    def pyxl_workbook(self):
        if self._pyxl_workbook is None:

            if os.path.exists(self.path):
                _logger.debug("Load excel workbook %s", self.path)
                with open(self.path, 'rb') as fp:
                    self._pyxl_workbook = load_workbook(fp) #NOTE: using keep_vba=True causes issues, example on openpyxl 3.1.2: load and save samples/empty-table.xlsx without modification corrupts the file
                self._create_next_table_in_active_sheet = False
            else:
                _logger.debug("Create excel workbook for %s", self.path)
                self._pyxl_workbook = Workbook()
                self._create_next_table_in_active_sheet = True

        return self._pyxl_workbook

    @classmethod
    def close_all(cls):
        for instance in cls._instances.values():
            instance.close()

    def close(self):
        if not self.needs_save:
            _logger.debug("Excel workbook %s not modified", self.path)
        else:
            self.save()

    def save(self, *, save_as: str|os.PathLike|None = None):
        if not self._pyxl_workbook:
            return
        
        for table in self._tables.values():
            table.redefine()

        if save_as is None:
            save_as = self._save_as

        if save_as is False:
            _logger.debug("Would save Excel workbook %s", self.path)
        else:
            if save_as is None:
                save_as = self.path

            _logger.debug("Save Excel workbook %s", save_as)
            with open(save_as, 'wb') as fp:
                self.pyxl_workbook.save(fp)
        
        self.needs_save = False

    @overload
    def get_table(self, name: str|None = None) -> ExcelTable:
        ...

    @overload
    def get_table(self, name: str|None, default: T) -> ExcelTable|T:
        ...

    def get_table(self, name: str|None = None, default: T|Literal['__raise__'] = '__raise__') -> ExcelTable|T:
        if name is None: # Search the name of the table (if there is only one table in the file)
            for sheet_name in self.pyxl_workbook.sheetnames:
                pyxl_worksheet: Worksheet = self.pyxl_workbook[sheet_name]
                for a_name in pyxl_worksheet.tables:
                    if name is not None:
                        raise ValueError(f"Several tables found in workbook \"{self.path}\"")
                    else:
                        name = a_name

        if name is not None:
            if name in self._tables:
                return self._tables[name]
            
            for sheet_name in self.pyxl_workbook.sheetnames:
                pyxl_worksheet: Worksheet = self.pyxl_workbook[sheet_name]
                if name in pyxl_worksheet.tables:
                    pyxl_table = pyxl_worksheet.tables[name]
                    self._tables[name] = ExcelTable(pyxl_table, pyxl_worksheet, self)
                    return self._tables[name]

        # Case when no table is found            
        if default == '__raise__':
            raise KeyError(f"No table found with name \"{name}\" in workbook \"{self.path}\"")
        else:
            return default
        
    def create_table(self, name: str, no_headers: bool = False) -> ExcelTable:
        for sheet_name in self.pyxl_workbook.sheetnames:
            pyxl_worksheet: Worksheet = self.pyxl_workbook[sheet_name]
            if name in pyxl_worksheet.tables:
                raise ValueError(f"Table {name} already exist")
    
        self.needs_save = True
        
        if self._create_next_table_in_active_sheet:
            if not self.pyxl_workbook.active:
                raise ValueError("No active sheet")
            pyxl_worksheet: Worksheet = self.pyxl_workbook.active
            pyxl_worksheet.title = name
            self._create_next_table_in_active_sheet = False
        else:
            pyxl_worksheet: Worksheet = self.pyxl_workbook.create_sheet(title=name)

        self._tables[name] = ExcelTable(name, pyxl_worksheet, self, no_headers=no_headers)
        return self._tables[name]

    def get_or_create_table(self, name: str, no_headers: bool = False) -> tuple[ExcelTable, bool]:
        """
        Return (table, created).
        """
        table = self.get_table(name, default=None)
        if table is None:
            table = self.create_table(name, no_headers)
            return table, True
        else:
            return table, False

    def get_global_named_values(self, name: str) -> list:
        defn = self.pyxl_workbook.defined_names[name]
        values = []
        
        for title, coord in defn.destinations:
            worksheet = self.pyxl_workbook[title]
            cell = worksheet[coord]
            value = cell.value # type: ignore
            values.append(value)

        return values

    def get_global_named_value(self, name: str) -> Any:
        values = self.get_global_named_values(name)
        
        if len(values) > 1:
            raise ValueError(f"More than one cell")
        elif len(values) == 0:
            raise ValueError(f"Global name not found")
        else:
            return values[0]


class ExcelTable:
    min_row_index: int
    """ 0-base index of the first data row (excluding headers). """
    
    min_col_index: int
    """ 0-base index of the first column. """

    row_count: int
    """ Number of data rows (excluding headers). """

    def __init__(self, pyxl_table: Table|str, pyxl_worksheet: Worksheet, workbook: ExcelWorkbook, no_headers: bool|None = None):
        self.columns: list[str] = []
        self._column_indexes: dict[str, int]|None = None
        
        self.workbook = workbook
        self.pyxl_worksheet = pyxl_worksheet

        self.pyxl_table: Table|None
        self.name: str|None

        if isinstance(pyxl_table, str):
            self.pyxl_table = None
            self.name = pyxl_table
            if no_headers:
                self.has_headers = False
                self.min_row_index = 0
            else:
                self.has_headers = True
                self.min_row_index = 1
            self.min_col_index = 0
            self.row_count = 0

        elif isinstance(pyxl_table, Table):
            self.pyxl_table = pyxl_table
            self.name = pyxl_table.name # type: ignore

            self.min_col_index, col_count, min_row_index, self.row_count = self._get_table_data_area()
            self.min_row_index = min_row_index or 0

            self.has_headers = pyxl_table.headerRowCount is not None and pyxl_table.headerRowCount > 0
            
            # NOTE: if no headers, default names are returned, example in French: ['Colonne1', 'Colonne2']
            self.columns = [name for name in self.pyxl_table.column_names] 
            if len(self.columns) != col_count:
                raise ValueError(f'Invalid columns length ({len(self.columns)}, expected {col_count}): {self.columns}')
            
            if self.row_count == 1 and self.is_row_empty(0):
                self.row_count = 0
        else:
            raise ValueError(f"Invalid type for pyxl_table: {type(pyxl_table).__name__}")

        self._column_formats: list[dict[str,Any]]|None = None
    
    def get_column_index(self, name: str) -> int:
        if self._column_indexes is None:
            self._column_indexes = {n: i for i, n in enumerate(self.columns)}
        return self._column_indexes[name]
    
    def _get_table_data_area(self) -> tuple[int, int, int, int]:
        """
        Return current 0-based (min_col_index, col_count, min_row_index, row_count) of the current pyxl table.
        """
        if not self.pyxl_table:
            raise ValueError("Table not built")
        
        min_col, min_row, max_col, max_row = range_boundaries(self.pyxl_table.ref) # NOTE: 1-based indexes
        if not isinstance(min_col, int) or not isinstance(min_row, int) or not isinstance(max_col, int) or not isinstance(max_row, int):
            raise ValueError("range_boundaries() did not return all integer values")
        
        min_col_index = min_col - 1
        col_count = max_col - min_col_index

        if self.pyxl_table.headerRowCount == 0:
            min_row_index = min_row - 1
        elif self.pyxl_table.headerRowCount == 1:
            min_row_index = min_row
        else:
            raise ValueError(f'Invalid headerRowCount: {self.pyxl_table.headerRowCount}')
        row_count = max_row - min_row_index

        return min_col_index, col_count, min_row_index, row_count
    
    def _get_table_data_boundaries(self, initial = False):
        """
        Return 1-based (min_col, min_row, max_col, max_row) data range for the current pyxl table.
        """
        if initial:
            min_col_index, col_count, min_row_index, row_count = self._get_table_data_area()
        else:
            min_col_index, col_count, min_row_index, row_count = self.min_col_index, self.col_count, self.min_row_index, self.row_count

        min_col = min_col_index + 1
        min_row = min_row_index + 1
        max_col = min_col + col_count - 1
        max_row = min_row + row_count - 1
        return min_col, min_row, max_col, max_row
    
    @property
    def col_count(self) -> int:
        return len(self.columns)
    
    @property
    def ref(self) -> str:
        if self.col_count == 0:
            raise ValueError(f"Cannot get table ref: table does not contain any column")
        if self.row_count == 0:
            raise ValueError(f"Cannot get table ref: table does not contain any row")
                
        return f"{get_column_letter(self.min_col_index + 1)}{self.min_row_index - (1 if self.has_headers else 0) + 1}:{get_column_letter(self.min_col_index + self.col_count)}{self.min_row_index + self.row_count}"

    def get_row(self, index: int, *, readonly: bool = False) -> ExcelRow:
        """
        Get row at the given 0-base index.
        """
        if index == -1:
            if self.row_count == 0:
                raise ValueError(f"Cannot get last row: table does not contain any row")
            index = self.row_count - 1
        elif index < 0:
            raise ValueError(f"Invalid row index: {index}")
        
        if index >= self.row_count:
            raise ValueError(f"Cannot get row at index {index}: table contains {self.row_count} rows")

        return ExcelRow(self, index, readonly=readonly)
    
    def is_row_empty(self, row_index: int):
        for col_index in range(0, self.col_count):
            cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
            if cell.value is not None:
                return False
            
        return True
    
    def is_column_empty(self, col_index: int):
        for row_index in range(0, self.row_count):
            cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
            if cell.value is not None:
                return False
            
        return True
    
    def insert_row(self) -> ExcelRow:
        self.workbook.needs_save = True
        row_index = self.row_count

        if not self.is_row_empty(row_index):
            raise ValueError(f'Cannot insert row: row at index {row_index} is not empty')

        self.row_count += 1

        # erase old styles and apply column format
        for col_index in range(0, self.col_count):
            self.erase_cell(row_index, col_index)
            
        return self.get_row(row_index)
    
    def add_column(self, name: str):
        self.workbook.needs_save = True
        if not name:
            raise ValueError(f'Name cannot be empty')
        
        name = str(name)
        if name in self.columns:
            raise ValueError(f'Column name already used: {name}')
            
        col_index = self.col_count

        if not self.is_column_empty(col_index):
            raise ValueError(f'Cannot insert column: column {col_index} is not empty')
        
        self.columns.append(name) # implies self.col_count += 1
        self._column_indexes = None # invalidate columns

        if self.has_headers:
            cell = self.pyxl_worksheet.cell(self.min_row_index, self.min_col_index + 1 + col_index)
            cell.value = name # type: ignore

        # erase old styles and apply column format
        for row_index in range(0, self.row_count):
            self.erase_cell(row_index, col_index)

    def truncate(self):
        self.workbook.needs_save = True
        prev_row_count = self.row_count
        self.row_count = 0

        for row_index in range(0, prev_row_count):
            for col_index in range(0, self.col_count):
                self.erase_cell(row_index, col_index, allow_outside=True)

    def redefine(self):
        self.workbook.needs_save = True

        # table cannot be empty (must have at least one blank row)
        if self.row_count == 0:
            self.insert_row()

        new_ref = self.ref
        if self.pyxl_table is not None and new_ref == self.pyxl_table.ref:
            return
        
        _logger.debug("Define table %s: %s => %s", self.name, self.pyxl_table.ref if self.pyxl_table is not None else None, new_ref)

        newcolumns = []

        for i in range(0, self.col_count):
            if self.has_headers:
                name = str(self.pyxl_worksheet.cell(self.min_row_index, self.min_col_index + 1 + i).value)
            else:
                name = self.columns[i] if i < len(self.columns) else ''

            newcolumn = TableColumn(id=i+1, name=name)
            newcolumns.append(newcolumn)

            if self.pyxl_table is not None and i < len(self.pyxl_table.tableColumns):
                prevcolumn: TableColumn = self.pyxl_table.tableColumns[i]
                newcolumn.dataCellStyle = prevcolumn.dataCellStyle
                newcolumn.dataDxfId = prevcolumn.dataDxfId # refers to workbook._differential_styles
                newcolumn.calculatedColumnFormula = prevcolumn.calculatedColumnFormula

        if self.pyxl_table is not None:
            self._adapt_worksheet_object_ranges()

        newtable_kwargs = {
            'name': self.name,
            'displayName': self.name,
            'ref': new_ref,
            'tableColumns': newcolumns,
            'headerRowCount': 1 if self.has_headers else 0,
        }

        if self.pyxl_table is not None:
            newtable_kwargs['autoFilter'] = self.pyxl_table.autoFilter
            newtable_kwargs['sortState'] = self.pyxl_table.sortState
            newtable_kwargs['tableStyleInfo'] = self.pyxl_table.tableStyleInfo
        else:
            newtable_kwargs['autoFilter'] = AutoFilter()
            newtable_kwargs['tableStyleInfo'] = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        newtable = Table(**newtable_kwargs)

        self.pyxl_table = newtable
        
        if self.name in self.pyxl_worksheet.tables:
            del self.pyxl_worksheet.tables[self.name]
        self.pyxl_worksheet.add_table(newtable)

    def _adapt_worksheet_object_ranges(self):
        """
        Adapt table data validation and conditional formatting.
        """
        prv_min_col, prv_min_row, prv_max_col, prv_max_row = self._get_table_data_boundaries(initial=True)
        _, new_min_row, _, new_max_row = self._get_table_data_boundaries()
        if new_min_row == prv_min_row and new_max_row == prv_max_row:
            return # The lines did not change: nothing to adapt
        
        def ranges_in_table(ranges: MultiCellRange):
            range: CellRange
            for range in ranges:
                if not (range.min_row == prv_min_row and range.max_row == prv_max_row and range.min_col >= prv_min_col and range.max_col <= prv_max_col):
                    return False
            return True
        
        def update_ranges(ranges: MultiCellRange):
            prev_ranges_str = str(ranges)

            range: CellRange
            for range in ranges:
                range.min_row = new_min_row
                range.max_row = new_max_row
                
            _logger.debug("Adapted data validation or conditional formatting ranges: %s => %s", prev_ranges_str, ranges)

        dv: DataValidation
        for dv in self.pyxl_worksheet.data_validations.dataValidation:
            if ranges_in_table(dv.ranges):
                update_ranges(dv.ranges)
        
        cf: ConditionalFormatting
        for cf in list(self.pyxl_worksheet.conditional_formatting):
            del self.pyxl_worksheet.conditional_formatting[cf.sqref]
            if ranges_in_table(cf.cells):
                update_ranges(cf.cells)
            for rule in cf.rules:
                self.pyxl_worksheet.conditional_formatting.add(str(cf.cells), rule)

    def get_value(self, row_index: int, col_index: int) -> Any:
        self._check_indexes(row_index, col_index)
        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
        value = cell.value
        if self.workbook._formula is not NO_FORMULA and isinstance(value, str) and value.startswith('='):
            return self.workbook._formula
        else:
            return value
    
    def set_value(self, row_index: int, col_index: int, value: Any):
        """
        Set the value of the cell located at the given 0-base indices, and apply the default formatting and formulas of the corresponding table column.

        Any value (including `None`) overrides default column formulas. If you want to use the default column formula, use `erase_cell` method instead.
        """
        self._check_indexes(row_index, col_index)
        self.workbook.needs_save = True

        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
        if isinstance(cell, MergedCell):
            raise ValueError("Cannot assign value of merged cell")

        if isinstance(value, (Enum,Flag)):
            value = value.value

        self._apply_column_format(cell, newline=isinstance(value, str) and '\n' in value)

        try:
            cell.value = value
        except ValueError as err:
            if str(err).startswith('Cannot convert'):
                cell.value = str(value)
            else:
                raise
        
    def _check_indexes(self, row_index: int, col_index: int):
        if row_index == -1:
            row_index = self.row_count - 1
        if col_index == -1:
            col_index = self.col_count - 1

        if row_index < 0 or row_index >= self.row_count:
            raise ValueError(f"Invalid row index: {row_index} (row count: {self.row_count})")
        if col_index < 0 or col_index >= self.col_count:
            raise ValueError(f"Invalid row index: {col_index} (row count: {self.col_count})")

    def erase_cell(self, row_index: int, col_index: int, allow_outside: bool = False):
        """
        Erase the value of the cell located at the given 0-base indices, and apply the default formatting and formulas of the corresponding table column.
        
        If `allow_outside` is set, the cell may be located outside of the table. In this case, no formatting or formula is applied.
        """
        if not allow_outside:
            self._check_indexes(row_index, col_index)

        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
        if isinstance(cell, MergedCell):
            raise ValueError("Cannot assign value of merged cell")
        
        self.workbook.needs_save = True
        cell.style = 'Normal'
        cell.value = None

        if not allow_outside or (row_index < self.row_count and col_index < self.col_count):
            self._apply_column_format(cell)
    
    def _apply_column_format(self, cell: Cell, *, newline = False):
        if self._column_formats is None:
            self._column_formats = self._build_column_formats()
        
        index = (cell.col_idx - 1) - self.min_col_index
        if index >= len(self._column_formats):
            return
        
        fmt = self._column_formats[index]

        if 'formula' in fmt:
            formula = fmt['formula']
            if isinstance(formula, ArrayFormula):
                _logger.warning(f"Array formula ignored")
            else:
                cell.value = formula

        if 'style' in fmt:
            cell.style = fmt['style']

        for fmt_key, fmt_value in fmt.items():
            if fmt_key in ['formula', 'style']:
                continue
            setattr(cell, fmt_key, fmt_value)

        if newline:
            if not cell.alignment:
                cell.alignment = Alignment(wrap_text=True)
            elif not cell.alignment.wrap_text:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment # type: ignore

    def _build_column_formats(self) -> list[dict[str,Any]]:
        if not self.pyxl_table:
            return []
                
        column: TableColumn
        fmt_list = []
        for column in self.pyxl_table.tableColumns:
            fmt: dict[str,Any] = {}
            fmt_list.append(fmt)

            # Read dataCellStyle
            if column.dataCellStyle:
                fmt['style'] = column.dataCellStyle
            
            # Read dxf
            if column.dataDxfId is not None:
                dxf: DifferentialStyle = self.workbook.pyxl_workbook._differential_styles[column.dataDxfId] # type: ignore

                if dxf.numFmt:
                    fmt['number_format'] = dxf.numFmt.formatCode
                else:
                    if not 'style' in fmt:
                        fmt['number_format'] = self._DEFAULT_NUMBER_FORMAT

                fmt['alignment'] = dxf.alignment if dxf.alignment else self._DEFAULT_ALIGNMENT
                fmt['border'] = dxf.border if dxf.border else self._DEFAULT_BORDER
                fmt['font'] = dxf.font if dxf.font else self._DEFAULT_FONT
                fmt['protection'] = dxf.protection if dxf.protection else self._DEFAULT_PROTECTION

                if dxf.fill and dxf.fill.fill_type is not None: # type: ignore
                    fmt['fill'] = PatternFill(fill_type=dxf.fill.fill_type, bgColor=dxf.fill.fgColor, fgColor=dxf.fill.bgColor) # type: ignore # NOTE: fgcolor and bgcolor are inversed in DifferentialStyle
                else:
                    fmt['fill'] = self._DEFAULT_FILL

            # Read formula
            if column.calculatedColumnFormula:
                formula = column.calculatedColumnFormula
                if formula.array:
                    fmt['formula'] = ArrayFormula(formula.attr_text)
                else:
                    fmt['formula'] = '=' + formula.attr_text
            
        return fmt_list

    _DEFAULT_NUMBER_FORMAT = 'General'
    _DEFAULT_FILL = PatternFill(fill_type=None)
    _DEFAULT_ALIGNMENT = None # openpyxl.styles.alignment.Alignment
    _DEFAULT_BORDER = None # openpyxl.styles.alignment.Border
    _DEFAULT_FONT = None # openpyxl.styles.fonts.Font
    _DEFAULT_PROTECTION = None # openpyxl.styles.protection.Protection

    def __iter__(self):
        return self.Iterator(self)
    
    def iter_rows(self, readonly: bool = False):
        return self.Iterator(self, readonly=readonly)

    class Iterator:
        def __init__(self, table: ExcelTable, readonly: bool = False):
            self.next_index = 0
            self.table = table
            self.readonly = readonly

        def __iter__(self):
            return self

        def __next__(self):
            if self.next_index >= self.table.row_count:
                raise StopIteration()
            
            row = self.table.get_row(self.next_index, readonly=self.readonly)
            self.next_index += 1
            return row


class ExcelRow:
    def __init__(self, table: ExcelTable, index: int, *, readonly: bool = False):
        self.table = table
        self.index = index
        self.readonly = readonly

        self._values: list[Any]|None = None
        self._must_refresh: dict[int,bool]|None = None

    @property
    def values(self) -> list[Any]:
        if self._values is None:
            self._values = list(self.table.get_value(self.index, col_index) for col_index in range(0, self.table.col_count))
            self._must_refresh = None

        elif self._must_refresh:
            for col_index in range(0, self.table.col_count):
                if self._must_refresh[col_index]:
                    self._values[col_index] = self.table.get_value(self.index, col_index)
            self._must_refresh = None
        
        return self._values
    
    def __getitem__(self, key: int|str):
        if not isinstance(key, int):
            key = self.table.get_column_index(key)
            
        if self._values is None:
            self._values = [ self._UNSET_VALUE ] * self.table.col_count
            self._must_refresh = {index: True for index in range(0, self.table.col_count)}

            value = self.table.get_value(self.index, key)
            self._values[key] = value
            self._must_refresh[key] = False
            return value

        elif self._must_refresh and self._must_refresh[key]:
            value = self.table.get_value(self.index, key)
            self._values[key] = value
            self._must_refresh[key] = False
            return value

        else:
            return self._values[key]
    
    def __setitem__(self, key: int|str, value):
        if not isinstance(key, int):
            try:
                key = self.table.get_column_index(key)
            except KeyError:
                self.table.add_column(key)
                key = self.table.get_column_index(key)

        if self._values is None:
            self._values = [ self._UNSET_VALUE ] * self.table.col_count
            self._must_refresh = {index: True for index in range(0, self.table.col_count)}
        else:
            if self._must_refresh is None:
                self._must_refresh = {}
            for index in range(len(self._values), key+1):
                self._values.append(self._UNSET_VALUE)
                self._must_refresh[index] = True

        self._values[key] = value
        if self._must_refresh is not None:
            self._must_refresh[key] = False
        
        if not self.readonly:
            self.table.set_value(self.index, key, value)

    _UNSET_VALUE = object()
