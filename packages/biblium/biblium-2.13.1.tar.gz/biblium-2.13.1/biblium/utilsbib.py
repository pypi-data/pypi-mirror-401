from __future__ import annotations

# --- Standard library ---
import os
import re
import math
import datetime
from functools import reduce
from collections import Counter
from itertools import chain, combinations
from typing import Any, Callable, Dict, List, Optional, Tuple, Union, Iterable, Hashable, Literal, Mapping
import warnings
import difflib

# --- Data handling (always needed) ---
import pandas as pd
import numpy as np

# --- Progress bars ---
try:
    from tqdm.auto import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    def tqdm(iterable, *args, **kwargs):
        """Fallback when tqdm is not installed."""
        return iterable

# =============================================================================
# DEFERRED IMPORTS - Loaded on first use for faster initial import
# =============================================================================

# Caches for lazy-loaded modules
_sklearn_cache = {}
_scipy_cache = {}
_nltk_cache = {}
_nx_module = None
_rapidfuzz_module = None
_openpyxl_styles = None


def _get_sklearn_module(name: str):
    """Get a sklearn submodule, caching the result."""
    if name not in _sklearn_cache:
        import importlib
        _sklearn_cache[name] = importlib.import_module(f"sklearn.{name}")
    return _sklearn_cache[name]


def _get_scipy_module(name: str):
    """Get a scipy submodule, caching the result."""
    if name not in _scipy_cache:
        import importlib
        _scipy_cache[name] = importlib.import_module(f"scipy.{name}")
    return _scipy_cache[name]


def _get_nltk():
    """Get nltk module."""
    if "nltk" not in _nltk_cache:
        import nltk
        _nltk_cache["nltk"] = nltk
    return _nltk_cache["nltk"]


def _get_networkx():
    """Get networkx module."""
    global _nx_module
    if _nx_module is None:
        import networkx as nx
        _nx_module = nx
    return _nx_module


def _get_rapidfuzz():
    """Get rapidfuzz module."""
    global _rapidfuzz_module
    if _rapidfuzz_module is None:
        from rapidfuzz import fuzz
        _rapidfuzz_module = fuzz
    return _rapidfuzz_module


def _get_openpyxl_styles():
    """Get openpyxl.styles module."""
    global _openpyxl_styles
    if _openpyxl_styles is None:
        from openpyxl import styles as openpyxl_styles
        _openpyxl_styles = openpyxl_styles
    return _openpyxl_styles


# Convenience functions for commonly used classes
def _get_TfidfVectorizer():
    """Get TfidfVectorizer class from sklearn (lazy import)."""
    return _get_sklearn_module("feature_extraction.text").TfidfVectorizer


def _get_CountVectorizer():
    """Get CountVectorizer class from sklearn (lazy import)."""
    return _get_sklearn_module("feature_extraction.text").CountVectorizer


def _get_TfidfTransformer():
    """Get TfidfTransformer class from sklearn (lazy import)."""
    return _get_sklearn_module("feature_extraction.text").TfidfTransformer


def _get_KMeans():
    """Get KMeans class from sklearn (lazy import)."""
    return _get_sklearn_module("cluster").KMeans


def _get_entropy():
    """Get entropy function from scipy.stats (lazy import)."""
    return _get_scipy_module("stats").entropy


def _get_rankdata():
    """Get rankdata function from scipy.stats (lazy import)."""
    return _get_scipy_module("stats").rankdata


# =============================================================================
# MODULE-LEVEL IMPORTS (Required for backward compatibility)
# These are loaded at import time but cached for performance
# =============================================================================

# sklearn imports (deferred via property-like access in functions)
# Note: Functions that need these should call _get_sklearn_module()

# scipy imports (deferred)
# Note: Functions that need these should call _get_scipy_module()

# networkx (commonly used, loaded on first network function call)
# Note: Functions should call _get_networkx() or use 'nx = _get_networkx()'


# =============================================================================
# OPTIONAL DEPENDENCIES
# =============================================================================

def _get_igraph():
    """Get igraph if available."""
    try:
        import igraph as ig
        return ig
    except ImportError:
        return None

def _get_community_louvain():
    """Get python-louvain if available."""
    try:
        import community as community_louvain
        return community_louvain
    except ImportError:
        return None

def _get_prince():
    """Get prince (correspondence analysis) if available."""
    try:
        import prince
        return prince
    except ImportError:
        return None

def _get_umap():
    """Get UMAP if available."""
    try:
        import umap
        return umap
    except ImportError:
        return None

def _get_spacy():
    """Get spaCy if available."""
    try:
        import spacy
        return spacy
    except ImportError:
        return None


# Backward compatibility: These are accessed by other modules
# We'll lazy-load them on first access
ig = None  # Will be set on first _get_igraph() call
community_louvain = None  # Will be set on first _get_community_louvain() call
prince = None  # Will be set on first _get_prince() call
umap = None  # Will be set on first _get_umap() call


# Suppress the Windows + MKL KMeans memory-leak warning
warnings.filterwarnings(
    "ignore",
    message="KMeans is known to have a memory leak on Windows with MKL",
    category=UserWarning,
)

# --- OpenAI (if needed) ---
try:
    import openai
except ImportError:
    openai = None


# =============================================================================
# COMPATIBILITY LAYER: Import heavy modules only when first accessed
# =============================================================================
# These imports are deferred to speed up initial module load.
# They will be loaded on first use by the functions that need them.

class _DeferredImports:
    """Container for deferred imports that loads modules on first access."""
    
    _entropy = None
    _rankdata = None
    _skew = None
    _kurtosis = None
    _normaltest = None
    _ks_2samp = None
    _f_oneway = None
    _kruskal = None
    _shapiro = None
    _fisher_exact = None
    _chi2_contingency = None
    _pearsonr = None
    _pdist = None
    _linkage = None
    _fcluster = None
    _csr_matrix = None
    _multipletests = None
    _PatternFill = None
    _TfidfVectorizer = None
    _CountVectorizer = None
    _TfidfTransformer = None
    _ENGLISH_STOP_WORDS = None
    _silhouette_score = None
    _r2_score = None
    _mean_squared_error = None
    _pairwise_distances = None
    _StandardScaler = None
    _OneHotEncoder = None
    _KMeans = None
    _AgglomerativeClustering = None
    _SpectralClustering = None
    _SpectralCoclustering = None
    _LDA = None
    _NMF = None
    _TruncatedSVD = None
    _WordNetLemmatizer = None
    _word_tokenize = None
    _stopwords = None
    _SentimentIntensityAnalyzer = None
    _fuzz = None
    _nx = None
    
    @classmethod
    def load_scipy_stats(cls):
        if cls._entropy is None:
            from scipy.stats import (
                rankdata, entropy, skew, kurtosis, normaltest, ks_2samp,
                f_oneway, kruskal, shapiro, fisher_exact, chi2_contingency, pearsonr
            )
            cls._entropy = entropy
            cls._rankdata = rankdata
            cls._skew = skew
            cls._kurtosis = kurtosis
            cls._normaltest = normaltest
            cls._ks_2samp = ks_2samp
            cls._f_oneway = f_oneway
            cls._kruskal = kruskal
            cls._shapiro = shapiro
            cls._fisher_exact = fisher_exact
            cls._chi2_contingency = chi2_contingency
            cls._pearsonr = pearsonr
    
    @classmethod
    def load_scipy_cluster(cls):
        if cls._linkage is None:
            from scipy.cluster.hierarchy import linkage, fcluster
            from scipy.spatial.distance import pdist
            cls._linkage = linkage
            cls._fcluster = fcluster
            cls._pdist = pdist
    
    @classmethod
    def load_scipy_sparse(cls):
        if cls._csr_matrix is None:
            from scipy.sparse import csr_matrix
            cls._csr_matrix = csr_matrix
    
    @classmethod
    def load_sklearn_text(cls):
        if cls._TfidfVectorizer is None:
            from sklearn.feature_extraction.text import (
                CountVectorizer, TfidfVectorizer, TfidfTransformer, ENGLISH_STOP_WORDS
            )
            cls._TfidfVectorizer = TfidfVectorizer
            cls._CountVectorizer = CountVectorizer
            cls._TfidfTransformer = TfidfTransformer
            cls._ENGLISH_STOP_WORDS = ENGLISH_STOP_WORDS
    
    @classmethod
    def load_sklearn_metrics(cls):
        if cls._silhouette_score is None:
            from sklearn.metrics import silhouette_score, r2_score, mean_squared_error, pairwise_distances
            cls._silhouette_score = silhouette_score
            cls._r2_score = r2_score
            cls._mean_squared_error = mean_squared_error
            cls._pairwise_distances = pairwise_distances
    
    @classmethod
    def load_sklearn_cluster(cls):
        if cls._KMeans is None:
            from sklearn.cluster import KMeans, AgglomerativeClustering, SpectralClustering, SpectralCoclustering
            cls._KMeans = KMeans
            cls._AgglomerativeClustering = AgglomerativeClustering
            cls._SpectralClustering = SpectralClustering
            cls._SpectralCoclustering = SpectralCoclustering
    
    @classmethod
    def load_sklearn_preprocessing(cls):
        if cls._StandardScaler is None:
            from sklearn.preprocessing import StandardScaler, OneHotEncoder
            cls._StandardScaler = StandardScaler
            cls._OneHotEncoder = OneHotEncoder
    
    @classmethod
    def load_sklearn_decomposition(cls):
        if cls._LDA is None:
            from sklearn.decomposition import LatentDirichletAllocation as LDA, NMF, TruncatedSVD
            cls._LDA = LDA
            cls._NMF = NMF
            cls._TruncatedSVD = TruncatedSVD
    
    @classmethod
    def load_nltk(cls):
        if cls._WordNetLemmatizer is None:
            import nltk
            from nltk.stem import WordNetLemmatizer
            from nltk.tokenize import word_tokenize
            from nltk.corpus import stopwords
            from nltk.sentiment import SentimentIntensityAnalyzer
            cls._WordNetLemmatizer = WordNetLemmatizer
            cls._word_tokenize = word_tokenize
            cls._stopwords = stopwords
            cls._SentimentIntensityAnalyzer = SentimentIntensityAnalyzer
    
    @classmethod
    def load_rapidfuzz(cls):
        if cls._fuzz is None:
            from rapidfuzz import fuzz
            cls._fuzz = fuzz
    
    @classmethod
    def load_networkx(cls):
        if cls._nx is None:
            import networkx as nx
            cls._nx = nx
    
    @classmethod
    def load_openpyxl(cls):
        if cls._PatternFill is None:
            from openpyxl.styles import PatternFill
            cls._PatternFill = PatternFill
    
    @classmethod
    def load_statsmodels(cls):
        if cls._multipletests is None:
            from statsmodels.stats.multitest import multipletests
            cls._multipletests = multipletests


# Create deferred import accessor
_deferred = _DeferredImports

# Compatibility properties - access these to trigger lazy loading
def entropy(*args, **kwargs):
    """Lazy-loaded scipy.stats.entropy. Computes Shannon entropy."""
    _deferred.load_scipy_stats()
    return _deferred._entropy(*args, **kwargs)


def rankdata(*args, **kwargs):
    """Lazy-loaded scipy.stats.rankdata. Assigns ranks to data."""
    _deferred.load_scipy_stats()
    return _deferred._rankdata(*args, **kwargs)


def skew(*args, **kwargs):
    """Lazy-loaded scipy.stats.skew. Computes skewness of data."""
    _deferred.load_scipy_stats()
    return _deferred._skew(*args, **kwargs)


def kurtosis(*args, **kwargs):
    """Lazy-loaded scipy.stats.kurtosis. Computes kurtosis of data."""
    _deferred.load_scipy_stats()
    return _deferred._kurtosis(*args, **kwargs)


def normaltest(*args, **kwargs):
    """Lazy-loaded scipy.stats.normaltest. Tests for normality."""
    _deferred.load_scipy_stats()
    return _deferred._normaltest(*args, **kwargs)


def ks_2samp(*args, **kwargs):
    """Lazy-loaded scipy.stats.ks_2samp. Kolmogorov-Smirnov two-sample test."""
    _deferred.load_scipy_stats()
    return _deferred._ks_2samp(*args, **kwargs)


def f_oneway(*args, **kwargs):
    """Lazy-loaded scipy.stats.f_oneway. One-way ANOVA test."""
    _deferred.load_scipy_stats()
    return _deferred._f_oneway(*args, **kwargs)


def kruskal(*args, **kwargs):
    """Lazy-loaded scipy.stats.kruskal. Kruskal-Wallis H-test."""
    _deferred.load_scipy_stats()
    return _deferred._kruskal(*args, **kwargs)


def shapiro(*args, **kwargs):
    """Lazy-loaded scipy.stats.shapiro. Shapiro-Wilk normality test."""
    _deferred.load_scipy_stats()
    return _deferred._shapiro(*args, **kwargs)


def fisher_exact(*args, **kwargs):
    """Lazy-loaded scipy.stats.fisher_exact. Fisher's exact test."""
    _deferred.load_scipy_stats()
    return _deferred._fisher_exact(*args, **kwargs)


def chi2_contingency(*args, **kwargs):
    """Lazy-loaded scipy.stats.chi2_contingency. Chi-square test."""
    _deferred.load_scipy_stats()
    return _deferred._chi2_contingency(*args, **kwargs)


def pearsonr(*args, **kwargs):
    """Lazy-loaded scipy.stats.pearsonr. Pearson correlation coefficient."""
    _deferred.load_scipy_stats()
    return _deferred._pearsonr(*args, **kwargs)


def pdist(*args, **kwargs):
    """Lazy-loaded scipy.spatial.distance.pdist. Pairwise distances."""
    _deferred.load_scipy_cluster()
    return _deferred._pdist(*args, **kwargs)


def linkage(*args, **kwargs):
    """Lazy-loaded scipy.cluster.hierarchy.linkage. Hierarchical clustering."""
    _deferred.load_scipy_cluster()
    return _deferred._linkage(*args, **kwargs)


def fcluster(*args, **kwargs):
    """Lazy-loaded scipy.cluster.hierarchy.fcluster. Cluster assignment."""
    _deferred.load_scipy_cluster()
    return _deferred._fcluster(*args, **kwargs)


def csr_matrix(*args, **kwargs):
    """Lazy-loaded scipy.sparse.csr_matrix. Compressed sparse row matrix."""
    _deferred.load_scipy_sparse()
    return _deferred._csr_matrix(*args, **kwargs)


def multipletests(*args, **kwargs):
    """Lazy-loaded statsmodels.stats.multitest.multipletests. Multiple testing correction."""
    _deferred.load_statsmodels()
    return _deferred._multipletests(*args, **kwargs)


def fuzz_ratio(*args, **kwargs):
    """Lazy-loaded rapidfuzz.fuzz.ratio. Fuzzy string matching ratio."""
    _deferred.load_rapidfuzz()
    return _deferred._fuzz.ratio(*args, **kwargs)


# For classes, we need a different approach - lazy class accessors
class TfidfVectorizer:
    """Lazy-loaded TfidfVectorizer."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_text()
        return _deferred._TfidfVectorizer(*args, **kwargs)

class CountVectorizer:
    """Lazy-loaded CountVectorizer."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_text()
        return _deferred._CountVectorizer(*args, **kwargs)

class TfidfTransformer:
    """Lazy-loaded TfidfTransformer."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_text()
        return _deferred._TfidfTransformer(*args, **kwargs)

class KMeans:
    """Lazy-loaded KMeans."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_cluster()
        return _deferred._KMeans(*args, **kwargs)

class AgglomerativeClustering:
    """Lazy-loaded AgglomerativeClustering."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_cluster()
        return _deferred._AgglomerativeClustering(*args, **kwargs)

class SpectralClustering:
    """Lazy-loaded SpectralClustering."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_cluster()
        return _deferred._SpectralClustering(*args, **kwargs)

class SpectralCoclustering:
    """Lazy-loaded SpectralCoclustering."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_cluster()
        return _deferred._SpectralCoclustering(*args, **kwargs)

class StandardScaler:
    """Lazy-loaded StandardScaler."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_preprocessing()
        return _deferred._StandardScaler(*args, **kwargs)

class OneHotEncoder:
    """Lazy-loaded OneHotEncoder."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_preprocessing()
        return _deferred._OneHotEncoder(*args, **kwargs)

class WordNetLemmatizer:
    """Lazy-loaded WordNetLemmatizer."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_nltk()
        return _deferred._WordNetLemmatizer(*args, **kwargs)

class SentimentIntensityAnalyzer:
    """Lazy-loaded SentimentIntensityAnalyzer."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_nltk()
        return _deferred._SentimentIntensityAnalyzer(*args, **kwargs)

def word_tokenize(*args, **kwargs):
    """Lazy-loaded NLTK word_tokenize. Tokenizes text into words."""
    _deferred.load_nltk()
    return _deferred._word_tokenize(*args, **kwargs)


def get_stopwords():
    """Get NLTK stopwords corpus (lazy-loaded)."""
    _deferred.load_nltk()
    return _deferred._stopwords

# For networkx, create a module-like accessor
class _NetworkXModule:
    """Lazy networkx module accessor."""
    def __getattr__(self, name):
        _deferred.load_networkx()
        return getattr(_deferred._nx, name)

nx = _NetworkXModule()

# fuzz accessor
class _FuzzModule:
    """Lazy rapidfuzz.fuzz accessor."""
    def __getattr__(self, name):
        _deferred.load_rapidfuzz()
        return getattr(_deferred._fuzz, name)

fuzz = _FuzzModule()

# PatternFill accessor
def PatternFill(*args, **kwargs):
    """Lazy-loaded openpyxl PatternFill for Excel cell styling."""
    _deferred.load_openpyxl()
    return _deferred._PatternFill(*args, **kwargs)


# ENGLISH_STOP_WORDS accessor
def get_english_stop_words():
    """Get sklearn's English stop words set (lazy-loaded)."""
    _deferred.load_sklearn_text()
    return _deferred._ENGLISH_STOP_WORDS


# For backward compatibility, create module-level reference
ENGLISH_STOP_WORDS = None  # Will be populated on first use


def _ensure_english_stop_words():
    """Ensure ENGLISH_STOP_WORDS is loaded and return it."""
    global ENGLISH_STOP_WORDS
    if ENGLISH_STOP_WORDS is None:
        _deferred.load_sklearn_text()
        ENGLISH_STOP_WORDS = _deferred._ENGLISH_STOP_WORDS
    return ENGLISH_STOP_WORDS


# sklearn decomposition classes
class LDA:
    """Lazy-loaded LatentDirichletAllocation."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_decomposition()
        return _deferred._LDA(*args, **kwargs)

LatentDirichletAllocation = LDA

class NMF:
    """Lazy-loaded NMF."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_decomposition()
        return _deferred._NMF(*args, **kwargs)

class TruncatedSVD:
    """Lazy-loaded TruncatedSVD."""
    def __new__(cls, *args, **kwargs):
        _deferred.load_sklearn_decomposition()
        return _deferred._TruncatedSVD(*args, **kwargs)


# sklearn metrics functions
def silhouette_score(*args, **kwargs):
    """Lazy-loaded sklearn silhouette_score for clustering evaluation."""
    _deferred.load_sklearn_metrics()
    return _deferred._silhouette_score(*args, **kwargs)


def r2_score(*args, **kwargs):
    """Lazy-loaded sklearn r2_score for regression evaluation."""
    _deferred.load_sklearn_metrics()
    return _deferred._r2_score(*args, **kwargs)


def mean_squared_error(*args, **kwargs):
    """Lazy-loaded sklearn mean_squared_error for regression evaluation."""
    _deferred.load_sklearn_metrics()
    return _deferred._mean_squared_error(*args, **kwargs)


def pairwise_distances(*args, **kwargs):
    """Lazy-loaded sklearn pairwise_distances for distance computation."""
    _deferred.load_sklearn_metrics()
    return _deferred._pairwise_distances(*args, **kwargs)

"""General-purpose utility helpers used throughout the Biblium codebase (small, frequently reused functions)."""
# general
fd = os.path.dirname(__file__)

"""Helpers for path and folder manipulation (creating result directories, resolving project paths, etc.)."""
# folder manipulation

def make_folder(
    folder,
):
    """Create a folder if it doesn't exist."""
    if not os.path.exists(folder):
        os.makedirs(folder)


def make_folders(
    folders,
):
    """Create multiple folders if they don't exist."""
    for folder in folders:
        make_folder(folder)


# =============================================================================
# PROGRESS BAR UTILITIES
# =============================================================================

def progress_bar(
    iterable,
    desc: str = None,
    total: int = None,
    disable: bool = False,
    leave: bool = True,
    unit: str = "it",
    **kwargs
):
    """
    Wrap an iterable with a progress bar.
    
    Parameters
    ----------
    iterable : iterable
        The iterable to wrap.
    desc : str
        Description to show.
    total : int
        Total number of iterations (auto-detected if possible).
    disable : bool
        Whether to disable the progress bar.
    leave : bool
        Whether to leave the progress bar after completion.
    unit : str
        Unit name for iterations.
    **kwargs
        Additional arguments to tqdm.
    
    Returns
    -------
    iterable
        Wrapped iterable with progress bar (or original if tqdm unavailable).
    """
    if not TQDM_AVAILABLE or disable:
        return iterable
    
    return tqdm(
        iterable,
        desc=desc,
        total=total,
        leave=leave,
        unit=unit,
        **kwargs
    )


def progress_apply(
    df: pd.DataFrame,
    func: Callable,
    axis: int = 0,
    desc: str = "Processing",
    disable: bool = False,
    **kwargs
) -> pd.Series:
    """
    Apply a function to DataFrame with progress bar.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to process.
    func : callable
        Function to apply.
    axis : int
        Axis to apply function (0=rows, 1=columns).
    desc : str
        Progress bar description.
    disable : bool
        Whether to disable progress bar.
    **kwargs
        Additional arguments to apply().
    
    Returns
    -------
    pd.Series or pd.DataFrame
        Result of apply operation.
    """
    if TQDM_AVAILABLE and not disable:
        tqdm.pandas(desc=desc)
        return df.progress_apply(func, axis=axis, **kwargs)
    else:
        return df.apply(func, axis=axis, **kwargs)


"""String manipulation utilities for cleaning, normalizing and transforming text labels and fields."""
# misc string manipulation fucttions

"""Generic Python object utilities (safe access, flattening structures, type checks, etc.)."""
# misc object manipulationg functions

def rename_attributes(
    obj,
    rename_dict,
):
    """Rename attributes of an object according to a dictionary mapping."""
    for old_attr, new_attr in rename_dict.items():
        if hasattr(obj, old_attr):
            setattr(obj, new_attr, getattr(obj, old_attr))
            delattr(obj, old_attr)


def first_existing(
    df,
    cols,
):
    """Return the first column name from cols that exists in df, or None."""
    for c in cols:
        if c in df.columns:
            return c
    return None

"""Regular-expression helpers for building inclusion/exclusion patterns and reusable regex objects."""
# misc re functions


def build_exclusion_regex(
    words: Iterable[str],
    *,
    whole_word: bool = True,
    case_insensitive: bool = True,
) -> str:
    """
    Build a regex that MATCHES texts that do NOT contain any of the given words.
    Suitable for passing to `regex_filter` to EXCLUDE items.

    Parameters
    ----------
    words : Iterable[str]
        Words/phrases to exclude.
    whole_word : bool, default True
        If True, wrap each term with word boundaries \\b…\\b.
    case_insensitive : bool, default True
        If True, prefix the pattern with (?i).

    Returns
    -------
    str
        Regex string, e.g. r"(?i)^(?!.*\\b(?:term1|term2)\\b).*".
    """
    terms = [re.escape(w) for w in words if w]
    if not terms:
        return "^.*$"  # nothing to exclude → match all
    if whole_word:
        terms = [fr"\b{t}\b" for t in terms]
    alt = "|".join(terms)
    pat = fr"^(?!.*(?:{alt})).*"
    return f"(?i){pat}" if case_insensitive else pat

def build_inclusion_regex(
    words: Iterable[str],
    *,
    whole_word: bool = True,
    case_insensitive: bool = True,
    require_all: bool = False,
) -> str:
    """
    Build a regex that MATCHES texts that contain the given words.
    Use `require_all=False` to keep items containing ANY term,
    or `require_all=True` to keep items containing ALL terms.

    Parameters
    ----------
    words : Iterable[str]
        Words/phrases to include.
    whole_word : bool, default True
        If True, wrap each term with word boundaries \\b…\\b.
    case_insensitive : bool, default True
        If True, prefix the pattern with (?i).
    require_all : bool, default False
        If True, text must contain every term (uses lookaheads).

    Returns
    -------
    str
        Regex string, e.g. r"(?i).*(?:term1|term2).*" or with lookaheads for ALL.
    """
    terms = [re.escape(w) for w in words if w]
    if not terms:
        return "^.*$"  # nothing specified → match all
    if whole_word:
        terms = [fr"\b{t}\b" for t in terms]
    if require_all:
        lookaheads = "".join(fr"(?=.*{t})" for t in terms)
        pat = fr"^{lookaheads}.*"
    else:
        alt = "|".join(terms)
        pat = fr".*(?:{alt}).*"
    return f"(?i){pat}" if case_insensitive else pat

"""Convenience functions for mapping values between different coding schemes or lookup tables."""
# mapping

def reconstruct_mapping(
    mapping_df: pd.DataFrame,
    alias_df: Optional[pd.DataFrame] = None,
) -> Dict[str, Dict[str, Any]]:
    """
    Reconstruct a dict-of-dicts mapping from a wide "mapping" DataFrame, and
    extend it with alias keys from a wide "alias" DataFrame.

    Expected input
    --------------
    mapping_df : wide table where rows are first-level keys (canonical items) and
                 columns are second-level keys (attributes). The row index should
                 be the canonical item names. If the index was written to Excel and
                 re-read without specifying index_col, a leading "item" column is OK.
    alias_df   : wide table where each column name is a canonical item (matching
                 a row label in mapping_df) and each column lists synonym strings
                 (one per row, empty/NaN cells allowed).

    Behavior
    --------
    1) Builds the canonical dict:
       {item: {attribute: value, ...}, ...}
       NaN/empty string cells are omitted from the inner dicts.
    2) Extends the dict with alias keys so that for each synonym S under column C:
       mapping[S] is the SAME object as mapping[C] (no deep copy).

    Returns
    -------
    dict[str, dict[str, Any]]
        The reconstructed and alias-extended mapping.

    Examples
    --------
    >>> # From DataFrames:
    >>> mapping = reconstruct_mapping(mapping_df, alias_df)
    >>>
    >>> # From Excel file with sheets "mapping" and "alias":
    >>> # mapping = reconstruct_mapping_from_excel("mapping.xlsx")
    """
    # Normalize index
    if "item" in mapping_df.columns:
        mapping_df = mapping_df.set_index("item")

    # Canonical dict (drop NaN/empty cells)
    canonical: Dict[str, Dict[str, Any]] = {}
    for item, row in mapping_df.iterrows():
        row_dict = {}
        for k, v in row.items():
            if pd.isna(v):
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            row_dict[str(k)] = v
        canonical[str(item)] = row_dict

    # No alias sheet -> done
    if alias_df is None or alias_df.empty:
        return canonical

    # Build {synonym -> canonical_item} map from wide alias sheet
    syn_to_canonical: Dict[str, str] = {}
    for canonical_item in alias_df.columns:
        # Only consider aliases for canonical items that exist
        if canonical_item not in canonical:
            continue
        col = alias_df[canonical_item]
        for v in col.dropna().astype(str):
            s = v.strip()
            if s:
                # keep first occurrence if duplicates appear across columns
                syn_to_canonical.setdefault(s, canonical_item)

    # Extend mapping with alias keys (reusing the same inner dict object)
    extended = dict(canonical)
    for syn, canon_key in syn_to_canonical.items():
        extended[syn] = canonical[canon_key]

    return extended

def reconstruct_mapping_from_excel(
    path: str,
) -> Dict[str, Dict[str, Any]]:
    """
    Convenience wrapper that reads an Excel file with sheets "mapping" and "alias"
    and returns the alias-extended mapping dict.

    Parameters
    ----------
    path : str
        Path to the .xlsx file.

    Returns
    -------
    dict[str, dict[str, Any]]
        The reconstructed mapping extended with alias keys.
    """
    sheets = pd.read_excel(path, sheet_name=None)
    mapping_df = sheets.get("mapping")
    if mapping_df is None:
        raise ValueError('Sheet "mapping" not found.')
    alias_df = sheets.get("alias", pd.DataFrame())
    return reconstruct_mapping(mapping_df, alias_df)

"""Helpers for adding derived document labels and flags to bibliographic DataFrames."""
# add document labels

def add_document_labels_abbrev(
    df,
    authors_col: str = None,
    source_col: str = None,
    year_col: str = None,
    title_col: str = None,
):
    """
    Adds 'Document Short Label' and 'Document Label' using 'Abbreviated Source Title' instead of 'Source title'.

    'Document Short Label': 'FirstAuthor, AbbreviatedSourceTitle (Year)'
    'Document Label': Adds first three words of title after the short label.

    Parameters:
        df (pd.DataFrame): Input dataframe.
        authors_col (str): Column name for authors. Auto-detected if None.
        source_col (str): Column name for abbreviated source title. Auto-detected if None.
        year_col (str): Column name for year. Auto-detected if None.
        title_col (str): Column name for title. Auto-detected if None.

    Returns:
        pd.DataFrame: DataFrame with updated label columns.
    """
    def extract_first_author(
        authors,
    ):
        if pd.isna(authors) or not isinstance(authors, str):
            return ""
        return authors.split(";")[0].strip()

    def extract_first_three_words(
        title,
    ):
        if pd.isna(title) or not isinstance(title, str):
            return ""
        return " ".join(title.strip().split()[:3])

    df = df.copy()

    # Auto-detect column names if not provided
    if authors_col is None:
        authors_col = first_existing(df, ["Authors", "Authors or Inventors", "Author", "AU"])
    if source_col is None:
        source_col = first_existing(df, ["Abbreviated Source Title", "Journal ISO Abbreviation", "Source title", "Source", "SO"])
    if year_col is None:
        year_col = first_existing(df, ["Year", "Publication Year", "PY"])
    if title_col is None:
        title_col = first_existing(df, ["Title", "TI", "Document Title"])

    # Handle missing columns gracefully
    if authors_col and authors_col in df.columns:
        first_authors = df[authors_col].fillna("").apply(extract_first_author)
    else:
        first_authors = pd.Series([""] * len(df), index=df.index)
    
    if source_col and source_col in df.columns:
        source_titles = df[source_col].fillna("")
    else:
        source_titles = pd.Series([""] * len(df), index=df.index)
    
    if year_col and year_col in df.columns:
        years = df[year_col].fillna("").astype(str).replace("nan", "")
    else:
        years = pd.Series([""] * len(df), index=df.index)
    
    if title_col and title_col in df.columns:
        titles = df[title_col].fillna("").apply(extract_first_three_words)
    else:
        titles = pd.Series([""] * len(df), index=df.index)

    df["Document Short Label"] = first_authors + ", " + source_titles + " (" + years + ")"
    df["Document Label"] = df["Document Short Label"] + ": " + titles

    return df

"""Functions for abbreviating long strings (for example journal titles or labels) according to custom rules."""
# Abbreviation of strings

def abbreviate_words(
    text,
    n = 4,
    letters_only = True,
    skip_len_n_minus_1 = True,
    skip_len_n_plus_1 = False,
    min_length = 25,
):
    """
    Abbreviate each word to its first n letters and add '.' only when truncated.

    Parameters
    ----------
    text : str or NaN
        Input text to abbreviate.
    n : int, default=4
        Number of letters to keep before abbreviation.
    letters_only : bool, default=True
        Strip non-letters before abbreviating.
    skip_len_n_minus_1 : bool, default=True
        Leave words of length n-1 unchanged.
    skip_len_n_plus_1 : bool, default=False
        Leave words of length n+1 unchanged if True.
    min_length : int, default=25
        If the entire text string has fewer characters than this,
        it is returned unchanged (no abbreviation).

    Returns
    -------
    str
        Abbreviated text, or unchanged if conditions apply.
    """
    if pd.isna(text):
        return text

    text_str = str(text)
    if len(text_str) < min_length:
        return text_str

    out = []
    for token in text_str.split():
        w = re.sub(r"[^A-Za-z]+", "", token) if letters_only else token
        if not w:
            continue
        L = len(w)
        if (skip_len_n_minus_1 and L == n - 1) or L <= n or (skip_len_n_plus_1 and L == n + 1):
            out.append(w)           # no dot; word stays the same
        else:
            out.append(w[:n] + ".") # truncated -> add dot
    return " ".join(out)

"""General-purpose pandas DataFrame utilities, including merging, reshaping and column tweaks."""
# misc manipulations with dataframes

def merge_on_key(
    df1,
    df2,
    key_column,
):
    """
    Merge two pandas DataFrames on a common key column.

    The resulting DataFrame will:
    - Contain only rows where the key column values exist in both df1 and df2 (intersection).
    - Include all columns from both DataFrames (column union).
    - Preserve the column order and values from df1 for overlapping columns.

    Parameters:
        df1 (pd.DataFrame): The first DataFrame (typically larger, with fewer columns).
        df2 (pd.DataFrame): The second DataFrame (subset of rows, with additional columns).
        key_column (str): The name of the column to align and merge on.

    Returns:
        pd.DataFrame: A merged DataFrame with intersected rows and unioned columns.
    """
    # Step 1: Find intersection of keys
    common_keys = df1[key_column].isin(df2[key_column])
    df1_common = df1[common_keys]
    df2_common = df2[df2[key_column].isin(df1[key_column])]

    # Step 2: Avoid duplicate columns (except key)
    df2_extra_cols = [col for col in df2_common.columns if col not in df1.columns or col == key_column]

    # Step 3: Merge on the key column
    merged = pd.merge(
        df1_common,
        df2_common[df2_extra_cols],
        on=key_column,
        how="inner",
        suffixes=("", "_df2")  # Prevent name clash but keep df1 priority
    )

    return merged

def combine_item_dataframes(
    df_list,
    df_names = None,
):
    """
    Combines a list of dataframes with "Item", "Number of documents",
    "Fractional number of documents", and "Proportion of documents"
    into a single dataframe where each row corresponds to one input dataframe.
    Columns represent item-metric pairs. Missing values in populated rows
    are filled with 0. Completely empty dataframes result in rows filled with NaN.

    Parameters:
        df_list (list of pd.DataFrame): List of input dataframes.
        df_names (list of str, optional): Row index names for the result.

    Returns:
        pd.DataFrame: Combined dataframe with one row per input dataframe.
    """
    combined_rows = []
    all_columns = set()

    # First pass: process non-empty frames to collect all possible column names
    temp_rows = []
    for df in df_list:
        if not df.empty:
            df = df[["Item", "Number of documents", "Fractional number of documents", "Proportion of documents"]].copy()
            df_row = (
                df.set_index("Item")
                  .stack()
                  .rename_axis(["Item", "Metric"])
                  .reset_index(name="Value")
                  .assign(Col=lambda d: d["Item"] + " [" + d["Metric"] + "]")
                  .set_index("Col")["Value"]
            )
            all_columns.update(df_row.index)
            temp_rows.append(df_row)
        else:
            temp_rows.append(None)

    # Second pass: build aligned rows
    all_columns = sorted(all_columns)
    for df_row in temp_rows:
        if df_row is None:
            combined_rows.append(pd.Series(index=all_columns, dtype=float))  # All NaNs
        else:
            combined_rows.append(df_row.reindex(all_columns, fill_value=0))

    result = pd.DataFrame(combined_rows)

    if df_names:
        result.index = df_names

    return result

"""Utilities for handling missing values, including NA detection, cleaning and basic imputation."""
# missings

def check_missing_values(
    df,
    columns = None,
):
    """
    Returns the number and proportion of missing values for each column in a given list of columns in a pandas DataFrame.

    Parameters:
        df (pandas.DataFrame): The pandas DataFrame to check for missing values.
        columns (list): A list of column names to check for missing values.

    Returns:
        pandas.DataFrame: A DataFrame containing the number and proportion of missing values for each column in the given list of columns, as well as a new column indicating the quality of each column based on the number of missing values.
    """
    if columns is None:
        columns = df.columns
    else:
        columns = [c for c in columns if c in df.columns]

    missing_values_list = []
    missing_d = {}

    for column in columns:
        missing_count = df[column].isna().sum()
        proportion_missing = missing_count / len(df)
        missing_d[column] = df[column].isna()

        # Set the missing value quality based on the proportion of missing values
        if proportion_missing == 0:
            quality = "Excellent (0%)"
        elif proportion_missing < 0.1:
            quality = "Good (<10%)"
        elif proportion_missing < 0.5:
            quality = "Fair (10-50%)"
        elif proportion_missing < 0.9:
            quality = "Poor (50-90%)"
        else:
            quality = "Bad (>90%)"

        missing_values_list.append({
            "Column": column,
            "Missing Values": missing_count,
            "Proportion": proportion_missing,
            "Missing Value Quality": quality
        })

    # Convert the list of dictionaries to a DataFrame
    missing_values_df = pd.DataFrame(missing_values_list)

    # Sort the DataFrame by the proportion of missing values in ascending order
    missing_values_df = missing_values_df.sort_values("Proportion")

    return missing_values_df, missing_d

"""Helpers for subsetting and filtering DataFrames and Series based on flexible conditions."""
# filtering

def filter_dataframe(
    df: pd.DataFrame,
    filters: dict = {},
    bradford_filter: str = 'all',
) -> pd.DataFrame:
    """
    Filters a DataFrame based on multiple criteria per column, with optional Bradford's Law zone filtering.

    Supported filter keys per column:
    - "regex_include": list of regex patterns (OR-matched)
    - "regex_exclude": list of regex patterns (OR-matched)
    - "include": list of exact values to include
    - "exclude": list of exact values to exclude
    - "min": minimum value (for numeric/date columns)
    - "max": maximum value (for numeric/date columns)

    Parameters:
    - df (pd.DataFrame): The input DataFrame to filter.
    - filters (dict): Dictionary of column-wise filtering rules. Default is {} (no filtering).
    - bradford_filter (str): One of {"all", "core", "core+zone2"} to control filtering by Bradford zones.

    Returns:
    - pd.DataFrame: Filtered DataFrame with index reset.

    Example:
    --------
    To filter only Articles from core sources published after 2000 with at least 1 citation:

    >>> filters = {
    ...     "Document Type": {"include": ["Article"]},
    ...     "Year": {"min": 2000},
    ...     "Cited by": {"min": 1}
    ... }
    >>> filtered_df = filter_dataframe(df, filters, bradford_filter="core")
    """
    mask = pd.Series(True, index=df.index)

    for col, criteria in filters.items():
        if col not in df.columns:
            continue

        col_data = df[col].astype(str) if df[col].dtype == object else df[col]

        if "regex_include" in criteria:
            pattern = "|".join(criteria["regex_include"])
            mask &= col_data.str.contains(pattern, na=False, regex=True)

        if "regex_exclude" in criteria:
            pattern = "|".join(criteria["regex_exclude"])
            mask &= ~col_data.str.contains(pattern, na=False, regex=True)

        if "include" in criteria:
            mask &= col_data.isin(criteria["include"])

        if "exclude" in criteria:
            mask &= ~col_data.isin(criteria["exclude"])

        if "min" in criteria:
            mask &= pd.to_numeric(col_data, errors="coerce") >= criteria["min"]

        if "max" in criteria:
            mask &= pd.to_numeric(col_data, errors="coerce") <= criteria["max"]

    filtered_df = df[mask].copy()

    # Apply Bradford filtering if needed
    if bradford_filter in {"core", "core+zone2"}:
        if "Source title" not in filtered_df.columns:
            raise ValueError('"Source title" column is required for Bradford filtering.')

        source_counts = filtered_df["Source title"].value_counts()
        total_sources = len(source_counts)
        third = math.ceil(total_sources / 3)

        core_sources = source_counts.index[:third]
        zone2_sources = source_counts.index[third:2 * third]

        if bradford_filter == "core":
            allowed_sources = set(core_sources)
        elif bradford_filter == "core+zone2":
            allowed_sources = set(core_sources).union(set(zone2_sources))

        filtered_df = filtered_df[filtered_df["Source title"].isin(allowed_sources)]

    return filtered_df.reset_index(drop=True)

"""Citation-related helpers for normalizing, counting and transforming citation data."""
# misc functions - citations

def compute_average_citations_per_year(
    df,
    year_col = 'Year',
    citations_col = 'Cited by',
):
    """
    Compute the average number of citations per document for each year.

    Parameters:
        df (pd.DataFrame): DataFrame containing bibliometric records.
        year_col (str): Name of the column indicating the publication year.
        citations_col (str): Name of the column indicating the number of citations.

    Returns:
        pd.DataFrame: A DataFrame with "Year", "Number of Documents",
                      "Total Citations", and "Average Citations per Document".
    """
    grouped = df.groupby(year_col).agg(
        {"{}".format(citations_col): ["count", "sum"]}
    )
    grouped.columns = ["Number of Documents", "Total Citations"]
    grouped = grouped.reset_index()

    grouped["Average Citations per Document"] = (
        grouped["Total Citations"] / grouped["Number of Documents"]
    )

    return grouped

"""Functions for parsing, cleaning and restructuring author names and author-related fields."""
# Authors manipulation

def extract_author_mappings(
    df,
    column,
):
    """
    Extracts ID-to-author and author-to-ID mappings from a dataframe column.

    Parameters:
        df (pd.DataFrame): The dataframe containing the author strings.
        column (str): The column name where author strings are stored.

    Returns:
        tuple: Two dictionaries (id_to_author, author_to_id).
    """
    id_to_author = {}
    author_to_id = {}

    for author_str in df[column].dropna():
        for entry in author_str.split("; "):
            if not entry.strip():
                continue
            try:
                name, id_with_parens = entry.rsplit(" (", 1)
                author_id = id_with_parens.rstrip(")")
                id_to_author[author_id] = name
                author_to_id[name] = author_id
            except ValueError:
                continue

    return id_to_author, author_to_id

def split_author_id(
    df,
):
    """
    Splits a column named 'Author ID' in the DataFrame into two new columns:
    'Author' and 'ID'.

    The function assumes that each entry in the 'Author ID' column is a string
    in the format 'Author Name (ID)'.

    Parameters:
        df (pd.DataFrame): A pandas DataFrame with a column named 'Author ID'.

    Returns:
        pd.DataFrame: The same DataFrame with two additional columns:
                      - 'Author': the name of the author.
                      - 'ID': the ID extracted from parentheses as a string.
    """
    df["Author"] = df["Author(s) ID"].map(lambda x: x.split(" (")[0])
    df["ID"] = df["Author(s) ID"].str.extract(r'\(([^()]*)\)\s*$', expand=False)
    return df

def openalex_build_author_id_name_dict(
    df: pd.DataFrame,
    id_col: str = 'Author(s) ID',
    name_col: str = 'Authors 2',
    *,
    sep: str = '|',
) -> dict:
    """
    Build a dictionary mapping OpenAlex author IDs to author names.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing author IDs and names in OpenAlex format.
    id_col : str, default="Author(s) ID"
        Column name containing "|"–separated author IDs (e.g., "https://openalex.org/A5074473734|https://openalex.org/A5100777365").
    name_col : str, default="Authors 2"
        Column name containing "|"–separated author names corresponding to the IDs.
    sep : str, default="|"
        Delimiter used to separate multiple authors within a cell.

    Returns
    -------
    dict
        Dictionary mapping each author ID (string) to the corresponding author name (string).

    Notes
    -----
    - Handles missing values or mismatched lengths gracefully (pairs truncated to the shorter list).
    - Duplicate IDs keep the first encountered name.
    - Empty or malformed rows are skipped.

    Examples
    --------
    >>> data = {
    ...     "Author(s) ID": [
    ...         "https://openalex.org/A5074473734|https://openalex.org/A5100777365",
    ...         "https://openalex.org/A5087878724"
    ...     ],
    ...     "Authors 2": [
    ...         "Naveen Donthu|Satish Kumar",
    ...         "Henry Small"
    ...     ]
    ... }
    >>> df = pd.DataFrame(data)
    >>> openalex_build_author_id_name_dict(df)
    {'https://openalex.org/A5074473734': 'Naveen Donthu',
     'https://openalex.org/A5100777365': 'Satish Kumar',
     'https://openalex.org/A5087878724': 'Henry Small'}
    """
    id_name_dict = {}

    for id_raw, name_raw in zip(df[id_col], df[name_col]):
        if pd.isna(id_raw) or pd.isna(name_raw):
            continue

        ids = str(id_raw).split(sep)
        names = str(name_raw).split(sep)

        # Align lengths safely
        n = min(len(ids), len(names))
        pairs = zip(ids[:n], names[:n])

        for author_id, author_name in pairs:
            author_id = author_id.strip()
            author_name = author_name.strip()
            if author_id and author_id not in id_name_dict:
                id_name_dict[author_id] = author_name

    return id_name_dict

"""Functions for computing collaboration indices and related co-authorship measures."""
# Collaboration Index

def collaboration_index(
    df: pd.DataFrame,
    author_col: str = 'Author(s) ID',
    sep: str = ';',
) -> float:
    """
    Compute the Collaboration Index (CI) for a set of articles where authors are stored
    as a single string per row, separated by a delimiter.

    The Collaboration Index is defined as:
        CI = (Total number of author-instances in multi-authored articles)
             / (Total number of multi-authored articles)

    Only articles with more than one author are considered.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing article records, with one column listing authors as a string.
    author_col : str, default "authors"
        Name of the column in `df` where each row is a delimiter-separated string of author names.
    sep : str, default ";"
        Delimiter used to separate author names in the string.

    Returns
    -------
    float
        The Collaboration Index (average co-authors per multi-authored article).
        Returns 0.0 if there are no multi-authored articles.
    """
    # Parse the author strings into lists
    parsed = df[author_col].astype(str).map(lambda s: [a.strip() for a in s.split(sep) if a.strip()])

    # Filter to multi-authored articles
    multi = parsed[parsed.map(len) > 1]
    num_multi_articles = len(multi)
    if num_multi_articles == 0:
        return 0.0

    # Sum total authors across those articles
    total_authors = multi.map(len).sum()

    # Compute Collaboration Index
    return total_authors / num_multi_articles

"""Helpers for grouping and merging scientific fields or subject categories into broader areas."""
# merging sciences

def enrich_bibliometric_data(
    biblio_df,
    asjc_map_df,
    asjc_meta_df,
):
    """
    Enrich a bibliometric dataframe with ASJC codes and their metadata (Field, Area, Science).

    Parameters:
        biblio_df (pd.DataFrame): Bibliometric data with column "Source title".
        asjc_map_df (pd.DataFrame): Journal-to-code mapping with "Source title" and "All Science Journal Classification Codes (ASJC)".
        asjc_meta_df (pd.DataFrame): ASJC metadata with "code", "Field", "Area", and "Science".

    Returns:
        pd.DataFrame: Enriched dataframe with additional columns for ASJC code, Field, Area, and Science.
    """
    if "Science" in biblio_df.columns:
        print("Sciences already in the dataset")
        return biblio_df
    
    # Find source title column - could be different names in different databases
    source_col = first_existing(biblio_df, [
        "Source title", "Source Title", "source title",
        "Journal", "journal", "Source", "SO",
        "primary_location.source.display_name"
    ])
    
    if source_col is None:
        print("Warning: No source title column found. Skipping ASJC enrichment.")
        return biblio_df

    # Rename first column of asjc_map_df to "Source title"
    asjc_map_df.columns = ["Source title"] + list(asjc_map_df.columns[1:])

    # Use the found column name
    biblio_df["source title orig"] = biblio_df[source_col]
    # Normalize titles for matching
    biblio_df["Source title"] = biblio_df[source_col].astype(str).str.strip().str.lower()
    asjc_map_df["Source title"] = asjc_map_df["Source title"].astype(str).str.strip().str.lower()

    # Merge on source title only
    merged = biblio_df.merge(asjc_map_df, how="left", on="Source title")

    # Clean and split codes
    merged["All Science Journal Classification Codes (ASJC)"] = merged["All Science Journal Classification Codes (ASJC)"].fillna("").str.strip(";")
    exploded = merged.copy()
    exploded = exploded.assign(code=exploded["All Science Journal Classification Codes (ASJC)"].str.split(";")).explode("code")
    exploded["code"] = exploded["code"].str.strip()

    # Ensure consistent dtypes for merging
    asjc_meta_df["code"] = asjc_meta_df["code"].astype(str)
    exploded["code"] = exploded["code"].astype(str)

    # Merge with ASJC metadata
    enriched = exploded.merge(asjc_meta_df, on="code", how="left")

    # Optionally re-aggregate codes and metadata into lists or semicolon-separated strings
    agg = enriched.groupby("Source title", as_index=False).agg({
        "code": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Field": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Area": lambda x: "; ".join(sorted(set(x.dropna()))),
        "Science": lambda x: "; ".join(sorted(set(x.dropna())))
    })

    # Merge aggregated results back into original dataframe
    result = biblio_df.merge(agg, on="Source title", how="left")

    result["Source title"] = biblio_df["source title orig"]
    result = result.drop(columns=["source title orig"])
    biblio_df = biblio_df.drop(columns=["source title orig"])

    return result

"""Country-related helpers, such as standardizing country names and extracting country information."""
# countries

fd = os.path.dirname(__file__)

# Lazy loading for country data
_countries_loaded = False
_df_countries = None
_domain_dct = None
_c_off_dct = None
_code_dct = None
_code_dct_r = None
_country_iso3_dct = None
_continent_dct = None
_code_to_coords = None
_l_countries = None
_eu_countries = None


def _load_countries_data():
    """Load country data from Excel file (called once on first access)."""
    global _countries_loaded, _df_countries, _domain_dct, _c_off_dct, _code_dct
    global _code_dct_r, _country_iso3_dct, _continent_dct, _code_to_coords
    global _l_countries, _eu_countries
    
    if _countries_loaded:
        return
    
    _df_countries = pd.read_excel(os.path.join(fd, "additional files", "countries.xlsx"))
    _domain_dct = _df_countries.set_index("Internet domain").to_dict()["Name"]
    _c_off_dct = _df_countries.set_index("Official name").to_dict()["Name"]
    _code_dct = _df_countries.set_index("Name").to_dict()["Code"]
    _code_dct_r = _df_countries.set_index("Code").to_dict()["Name"]
    _country_iso3_dct = _df_countries.set_index("Name").to_dict()["ISO-3"]
    _continent_dct = _df_countries.set_index("Name").to_dict()["Continent"]
    
    df_countries_un_iso = _df_countries.drop_duplicates(subset="ISO-3")
    _code_to_coords = df_countries_un_iso[["ISO-3", "latitude", "longitude"]].set_index("ISO-3")[["latitude", "longitude"]].to_dict(orient="index")
    
    _l_countries = list(_df_countries["Name"])
    _eu_countries = list(_df_countries[_df_countries["EU"] == 1]["Name"])
    
    _countries_loaded = True


# Property-like accessors for backward compatibility
class _CountryDataAccessor:
    """Lazy accessor for country data."""
    
    @property
    def df_countries(self):
        _load_countries_data()
        return _df_countries
    
    @property  
    def domain_dct(self):
        _load_countries_data()
        return _domain_dct
    
    @property
    def c_off_dct(self):
        _load_countries_data()
        return _c_off_dct
    
    @property
    def code_dct(self):
        _load_countries_data()
        return _code_dct
    
    @property
    def code_dct_r(self):
        _load_countries_data()
        return _code_dct_r
    
    @property
    def country_iso3_dct(self):
        _load_countries_data()
        return _country_iso3_dct
    
    @property
    def continent_dct(self):
        _load_countries_data()
        return _continent_dct
    
    @property
    def code_to_coords(self):
        _load_countries_data()
        return _code_to_coords
    
    @property
    def l_countries(self):
        _load_countries_data()
        return _l_countries
    
    @property
    def eu_countries(self):
        _load_countries_data()
        return _eu_countries


_country_data = _CountryDataAccessor()

# Backward compatibility: These will trigger lazy loading on first access
def _get_df_countries():
    """Get DataFrame with country data (lazy-loaded)."""
    _load_countries_data()
    return _df_countries


def _get_domain_dct():
    """Get domain to country mapping dictionary (lazy-loaded)."""
    _load_countries_data()
    return _domain_dct


def _get_c_off_dct():
    """Get country official names dictionary (lazy-loaded)."""
    _load_countries_data()
    return _c_off_dct


def _get_code_dct():
    """Get country name to code dictionary (lazy-loaded)."""
    _load_countries_data()
    return _code_dct


def _get_code_dct_r():
    """Get country code to name dictionary (lazy-loaded)."""
    _load_countries_data()
    return _code_dct_r


def _get_country_iso3_dct():
    """Get country to ISO3 code dictionary (lazy-loaded)."""
    _load_countries_data()
    return _country_iso3_dct


def _get_continent_dct():
    """Get country to continent dictionary (lazy-loaded)."""
    _load_countries_data()
    return _continent_dct


def _get_code_to_coords():
    """Get country code to coordinates dictionary (lazy-loaded)."""
    _load_countries_data()
    return _code_to_coords


def _get_l_countries():
    """Get list of country names (lazy-loaded)."""
    _load_countries_data()
    return _l_countries


def _get_eu_countries():
    """Get list of EU country names (lazy-loaded)."""
    _load_countries_data()
    return _eu_countries


# For direct module-level access (backward compatibility)
# Import from utilsbib_modules.countries which loads data at module import time
try:
    from biblium.utilsbib_modules.countries import (
        df_countries,
        domain_dct,
        c_off_dct,
        code_dct,
        code_dct_r,
        country_iso3_dct,
        continent_dct,
        code_to_coords,
        l_countries,
        eu_countries,
    )
except ImportError:
    # Fallback - use lazy loading
    df_countries = None
    domain_dct = None
    c_off_dct = None
    code_dct = None
    code_dct_r = None
    country_iso3_dct = None
    continent_dct = None
    code_to_coords = None
    l_countries = None
    eu_countries = None

def correct_country_name(
    s,
):
    """
    Return the corrected country name based on known lists and mappings.

    Parameters:
        s (str): Input country name.

    Returns:
        str: Corrected country name if recognized, empty string otherwise.
    """
    _load_countries_data()  # Ensure data is loaded
    if not isinstance(s, str):
        return ""
    if s in _l_countries:
        return s
    return _c_off_dct.get(s, "")

def split_ca(
    s,
):
    """
    Split a Scopus corresponding author string into name, affiliation, and country.

    Parameters:
        s (str): Raw Scopus corresponding author string.

    Returns:
        tuple: (corresponding author, affiliation, country) or (np.nan, np.nan, np.nan) if parsing fails.
    """
    try:
        ca, long_aff = s.split("; ", 1)
        parts = long_aff.split(", ")
        return ca, parts[0], parts[-1]
    except Exception:
        return np.nan, np.nan, np.nan

def parse_mail(
    s,
):
    """
    Attempt to extract the country based on the email domain.

    Parameters:
        s (str): Full string that may contain an email.

    Returns:
        str or np.nan: Country inferred from email domain or np.nan if not found.
    """
    if "@" in s:
        domain = s.split("@")[1].split(" ")[0].split(".")[-1]
        return domain_dct.get(domain, np.nan)
    return np.nan

def get_ca_country_scopus(
    s,
    l_countries = l_countries,
):
    """
    Extract the country of the corresponding author from a Scopus entry.

    Parameters:
        s (str): Scopus corresponding author string.
        l_countries (list): List of recognized country names.

    Returns:
        str or np.nan: Extracted or inferred country name.
    """
    ca, aff, country = split_ca(s)

    if country not in l_countries:
        if isinstance(country, str):
            matches = [c for c in l_countries if c in country]
            if len(matches) == 1:
                country = matches[0]
            else:
                country = parse_mail(s)
        elif isinstance(s, str):
            matches = [c for c in l_countries if c in s]
            if len(matches) == 1:
                country = matches[0]
            else:
                country = parse_mail(s)

    return country

def get_ca_country_wos(
    s,
    l_countries = l_countries,
):
    """
    Extract the corresponding author's country from a WoS entry string.

    Parameters:
        s (str): Raw string from WoS corresponding author field.
        l_countries (list): List of valid country names.

    Returns:
        str or np.nan: Extracted country name or np.nan if not recognized.
    """
    if not isinstance(s, str):
        return np.nan

    if "USA" in s:
        return "United States"

    uk_terms = ["England", "Scotland", "Wales", "Northern Ireland", "Great Britain"]
    if any(term in s for term in uk_terms):
        return "United Kingdom"

    country = s.split(", ")[-1].replace(".", "")
    if country not in l_countries:
        try:
            country = c_off_dct.get(country, np.nan)
        except Exception:
            print(f"Unrecognized country: {country}")
            return np.nan
        if country not in l_countries:
            return np.nan

    return country

def get_ca_country(
    s,
    db,
    l_countries = l_countries,
):
    """
    Determine the country of the corresponding author based on the source database.

    Parameters:
        s (str): Raw corresponding author string.
        db (str): Name of the database ('scopus' or 'wos').
        l_countries (list): List of valid country names.

    Returns:
        str or np.nan: Extracted or inferred country name.
    """
    db = db.lower()
    if db == "scopus":
        return get_ca_country_scopus(s, l_countries=l_countries)
    elif db == "wos":
        return get_ca_country_wos(s, l_countries=l_countries)
    return np.nan

import pandas as pd

def openalex_add_corresponding_country(
    df: pd.DataFrame,
    countries_col: str = 'authorships.countries',
    is_corr_col: str = 'authorships.is_corresponding',
    new_col: str = 'CA Country',
    *,
    sep: str = '|',
    strict: bool = False,
    code_dct_r: dict | None = None,
) -> pd.DataFrame:
    """
    Append a column named "CA Country" with the corresponding author's country
    to an OpenAlex-style DataFrame and optionally map it through a dictionary.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing OpenAlex-style authorship information.
    countries_col : str, default="authorships.countries"
        Column containing "|"–separated ISO country codes per document.
    is_corr_col : str, default="authorships.is_corresponding"
        Column containing "|"–separated boolean flags aligned with `countries_col`.
    new_col : str, default="CA Country"
        Name of the new column to add (default fixed to "CA Country").
    sep : str, default="|"
        Delimiter used in encoded strings.
    strict : bool, default=False
        If True, raise ValueError when multiple distinct corresponding countries
        are found in one record; otherwise take the first in left-to-right order.
    code_dct_r : dict or None, optional
        Dictionary to map ISO country codes to readable country names.
        If None, no mapping is applied.

    Returns
    -------
    pandas.DataFrame
        The same DataFrame with an additional column `CA Country` containing the
        corresponding author's country (mapped if `code_dct_r` provided) or <NA> if none.

    Notes
    -----
    - At most one corresponding country is expected per document.
    - Handles NaN, empty strings, or mismatched lengths gracefully.
    - The function modifies and returns the same DataFrame (not a copy).

    """
    def to_list(
        x,
    ):
        if pd.isna(x):
            return []
        s = str(x)
        return s.split(sep) if s != "" else []

    def to_bools(
        x,
    ):
        if pd.isna(x):
            return []
        s = str(x)
        vals = s.split(sep) if s != "" else []
        out = []
        for v in vals:
            t = str(v).strip().lower()
            out.append(t in ("true", "1", "yes", "y"))
        return out

    result = []
    for countries_raw, flags_raw in zip(df[countries_col], df[is_corr_col]):
        countries = to_list(countries_raw)
        flags = to_bools(flags_raw)

        if len(flags) < len(countries):
            flags += [False] * (len(countries) - len(flags))

        corr = [c for c, f in zip(countries, flags) if f and c.strip()]

        if not corr:
            result.append(pd.NA)
        elif len(corr) == 1 or not strict:
            result.append(corr[0])
        else:
            raise ValueError(f"Multiple corresponding countries: {corr}")

    df[new_col] = pd.Series(result, index=df.index, dtype="string")

    # Map to readable country names if dictionary provided
    if code_dct_r is not None:
        df[new_col] = df[new_col].map(code_dct_r).astype("string")

    return df

def add_ca_country_df(
    df,
    db,
):
    """
    Add a 'CA Country' column to a DataFrame based on corresponding author information.

    Parameters:
        df (pd.DataFrame): DataFrame with corresponding author information.
        db (str): Database name ('scopus', 'wos', 'oa' supported).

    Returns:
        pd.DataFrame: DataFrame with added 'CA Country' column if applicable.
    """
    if db.lower() == "scopus" and "Correspondence Address" in df.columns:
        df["CA Country"] = df["Correspondence Address"].map(get_ca_country_scopus)
    elif db.lower() in ["open alex", "openalex", "oa"]:
        country_column = [c for c in ["Countries of Authors", "authorships.countries"] if c in df.columns][0]
        df = openalex_add_corresponding_country(df, countries_col=country_column, code_dct_r=code_dct_r)
    elif db.lower() == "wos":
        # For WoS, use Reprint Address or Correspondence Address
        addr_col = first_existing(df, ["Reprint Address", "Correspondence Address"])
        if addr_col:
            df["CA Country"] = df[addr_col].map(get_ca_country_wos)
    else:
        print("Not supported yet")
    return df

"""Convenience functions for retrieving all distinct countries appearing in the dataset."""
# get all countries

def extract_countries_from_affiliations(
    df,
    aff_column = 'Affiliations',
    return_matrix = True,
):
    """
    Extracts valid countries from the affiliations column and computes collaboration metrics.

    This function processes the specified affiliations column to:
    - Extract country names from the last comma-separated segment of each affiliation.
    - Validate and normalize each country using the globally defined correct_country_name() function.
    - Add three new columns:
        - "Countries of Authors Multiple": all valid country names found (can repeat), joined by "; ".
        - "Countries of Authors": unique list of valid country names, joined by "; ".
        - "Countries Count": number of unique valid countries per record.
    - Compute a symmetric country collaboration matrix (co-authorships across countries).

    Parameters:
    df (pd.DataFrame): Input DataFrame containing the affiliation data.
    aff_column (str): Name of the column with affiliation strings (default: "Affiliations").
    return_matrix (bool): Whether to compute and return the country collaboration matrix (default: True).

    Returns:
    tuple: (Updated DataFrame, Collaboration matrix as a symmetric DataFrame or empty DataFrame if not computed)
    """

    multiple_countries = []
    unique_countries = []
    country_counts = []

    matrix_df = pd.DataFrame()

    if aff_column not in df.columns:
        return df, matrix_df

    for affil in df[aff_column].fillna(""):
        entries = [entry.strip() for entry in affil.split(";")]
        countries_raw = [entry.split(",")[-1].strip() for entry in entries if "," in entry]
        countries_checked = [correct_country_name(c) for c in countries_raw]
        valid_countries = [c for c in countries_checked if c]
        unique_set = sorted(set(valid_countries))

        multiple_countries.append("; ".join(valid_countries))
        unique_countries.append("; ".join(unique_set))
        country_counts.append(len(unique_set))

    df = df.copy()
    df["Countries of Authors Multiple"] = multiple_countries
    df["Countries of Authors"] = unique_countries
    df["Countries Count"] = country_counts

    if return_matrix:
        matrix_counter = Counter()
        for country_str in unique_countries:
            countries = [c.strip() for c in country_str.split(";") if c.strip()]
            if len(countries) > 1:
                for pair in combinations(sorted(set(countries)), 2):
                    matrix_counter[pair] += 1

        if matrix_counter:
            all_countries = sorted(set(c for pair in matrix_counter for c in pair))
            matrix_df = pd.DataFrame(0, index=all_countries, columns=all_countries)
            for (c1, c2), count in matrix_counter.items():
                matrix_df.loc[c1, c2] = count
                matrix_df.loc[c2, c1] = count

    return df, matrix_df


def extract_countries_from_wos_addresses(
    df,
    addr_column='Correspondence Address',
    return_matrix=True,
):
    """
    Extract countries from WoS-style address fields and compute collaboration matrix.
    
    WoS addresses have format like:
    "[Author1; Author2] University, Department, City, Country. [Author3] Other Univ, City, Country."
    
    Parameters:
        df (pd.DataFrame): Input DataFrame with WoS address data.
        addr_column (str): Column name containing address info.
        return_matrix (bool): Whether to compute collaboration matrix.
        
    Returns:
        tuple: (Updated DataFrame, Collaboration matrix DataFrame)
    """
    multiple_countries = []
    unique_countries = []
    country_counts = []
    matrix_df = pd.DataFrame()
    
    if addr_column not in df.columns:
        return df, matrix_df
    
    for addr in df[addr_column].fillna(""):
        # Split by periods to get individual address blocks
        # Each block ends with country name
        blocks = [b.strip() for b in str(addr).split(".") if b.strip()]
        
        countries_found = []
        for block in blocks:
            # Skip blocks that don't look like addresses
            if "," not in block:
                continue
            
            # Get the last comma-separated part (should be country)
            parts = block.split(",")
            if len(parts) < 2:
                continue
                
            country_raw = parts[-1].strip()
            
            # Clean up common WoS quirks
            # Remove trailing brackets and numbers
            country_raw = country_raw.rstrip("]").strip()
            
            # Handle special cases
            if "USA" in country_raw or country_raw.startswith("DC ") or country_raw.startswith("NY "):
                country_raw = "United States"
            elif any(uk in country_raw for uk in ["England", "Scotland", "Wales", "N Ireland"]):
                country_raw = "United Kingdom"
            elif "Peoples R China" in country_raw or "PR China" in country_raw:
                country_raw = "China"
            
            # Try to correct/validate the country name
            country = correct_country_name(country_raw)
            if country:
                countries_found.append(country)
        
        unique_set = sorted(set(countries_found))
        multiple_countries.append("; ".join(countries_found))
        unique_countries.append("; ".join(unique_set))
        country_counts.append(len(unique_set))
    
    df = df.copy()
    df["Countries of Authors Multiple"] = multiple_countries
    df["Countries of Authors"] = unique_countries
    df["Countries Count"] = country_counts
    
    if return_matrix:
        matrix_counter = Counter()
        for country_str in unique_countries:
            countries = [c.strip() for c in country_str.split(";") if c.strip()]
            if len(countries) > 1:
                for pair in combinations(sorted(set(countries)), 2):
                    matrix_counter[pair] += 1
        
        if matrix_counter:
            all_countries = sorted(set(c for pair in matrix_counter for c in pair))
            matrix_df = pd.DataFrame(0, index=all_countries, columns=all_countries)
            for (c1, c2), count in matrix_counter.items():
                matrix_df.loc[c1, c2] = count
                matrix_df.loc[c2, c1] = count
    
    return df, matrix_df


import pandas as pd

def openalex_map_country_codes(
    df: pd.DataFrame,
    country_col: str = 'Countries of Authors',
    code_dict: dict | None = None,
    *,
    sep: str = '|',
) -> pd.DataFrame:
    """
    Map country abbreviations in an OpenAlex-style column to full country names,
    while preserving the original column in a backup column.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame containing a column with country codes separated by '|'.
    country_col : str, default="authorships.countries"
        Column containing ISO country abbreviations (e.g., "US|IN|MY").
    code_dict : dict, optional
        Dictionary mapping country codes (e.g., "US") to full country names (e.g., "United States").
        If None, the function simply leaves the column unchanged.
    sep : str, default="|"
        Delimiter used between multiple country codes in one cell.

    Returns
    -------
    pandas.DataFrame
        The same DataFrame with:
        - A backup column named `"{country_col}.orig"` containing the original values.
        - The original `country_col` replaced by its mapped full names.

    Notes
    -----
    - Handles NaN or empty entries gracefully.
    - Keeps the original code if not found in `code_dict`.
    - Preserves order of countries as in the original field.
    - Modifies and returns the same DataFrame.

    """
    if country_col not in df:
        raise ValueError(f"Column '{country_col}' not found in DataFrame.")

    # Backup original
    backup_col = f"{country_col}.orig"
    df[backup_col] = df[country_col].astype("string")

    # If no mapping provided, keep as is
    if code_dict is None:
        return df

    mapped_values = []
    for val in df[country_col]:
        if pd.isna(val) or not str(val).strip():
            mapped_values.append(pd.NA)
            continue

        codes = [v.strip() for v in str(val).split(sep) if v.strip()]
        mapped = [code_dict.get(code, code) for code in codes]
        mapped_values.append(sep.join(mapped))

    df[country_col] = pd.Series(mapped_values, index=df.index, dtype="string")
    return df

def compute_openalex_country_collaboration_matrix(
    df: pd.DataFrame,
    column: str = 'authorships.countries',
    sep: str = '|',
) -> pd.DataFrame:
    """
    Compute a symmetric country collaboration matrix from an OpenAlex dataframe column.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the OpenAlex dataset.
    column : str, default='authorships.countries'
        Column listing country codes separated by a delimiter.
    sep : str, default='|'
        Delimiter separating country codes within each cell.

    Returns
    -------
    pd.DataFrame
        Symmetric DataFrame of country collaborations.
        Diagonal values represent total documents per country.
    """
    country_lists = df[column].dropna().apply(lambda x: [c.strip() for c in str(x).split(sep) if c.strip()])

    pairs = []
    for countries in country_lists:
        unique_countries = sorted(set(countries))
        pairs.extend(combinations(unique_countries, 2))
        pairs.extend([(c, c) for c in unique_countries])

    matrix = pd.DataFrame(pairs, columns=["Country1", "Country2"])
    collaboration = matrix.value_counts().unstack(fill_value=0)

    collaboration = collaboration.add(collaboration.T, fill_value=0).astype(int)
    return collaboration

"""Utilities for parsing, cleaning and analysing affiliation strings and institutional data."""
# affiliations

import ast

def openalex_extract_affiliations(
    df: pd.DataFrame,
    affil_col: str = 'authorships.affiliations',
    new_col: str = 'Affiliations',
    *,
    sep: str = '|',
) -> pd.DataFrame:
    """
    Extract raw affiliation strings from OpenAlex-style 'authorships.affiliations' column.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing a column with serialized affiliation dictionaries.
    affil_col : str, default="authorships.affiliations"
        Column containing affiliation data in string form, where entries are separated by '|'
        and each entry represents a Python-like dict with keys such as 'raw_affiliation_string'.
    new_col : str, default="Affiliations"
        Name of the new column to create, containing extracted affiliation strings separated by '|'.
    sep : str, default="|"
        Separator used between affiliation records within a cell.

    Returns
    -------
    pandas.DataFrame
        The same DataFrame with an additional column `Affiliations`
        where all extracted `raw_affiliation_string` values are concatenated using '|'.

    Notes
    -----
    - Handles missing, malformed, or empty values gracefully.
    - Ignores entries that cannot be parsed into dictionaries.
    - Multiple affiliations within a document are joined by '|'.
    - The function modifies and returns the same DataFrame.

    """
    extracted = []

    for cell in df[affil_col]:
        if pd.isna(cell) or not str(cell).strip():
            extracted.append(pd.NA)
            continue

        affiliations = []
        for part in str(cell).split(sep):
            part = part.strip()
            if not part:
                continue
            try:
                # Safely evaluate each dictionary-like string
                affil_dict = ast.literal_eval(part)
                affil_raw = affil_dict.get("raw_affiliation_string", "").strip()
                if affil_raw:
                    affiliations.append(affil_raw)
            except (ValueError, SyntaxError):
                # Skip malformed parts
                continue

        if affiliations:
            extracted.append(sep.join(affiliations))
        else:
            extracted.append(pd.NA)

    df[new_col] = pd.Series(extracted, index=df.index, dtype="string")
    return df

"""Helpers for working with URL and link fields, including normalization, extraction and simple validation."""
# links

def build_links_from_matrix(
    matrix_df,
    min_weight = 1,
):
    """
    Constructs a DataFrame of collaboration links from a symmetric matrix.

    Parameters:
    matrix_df (pd.DataFrame): Symmetric collaboration matrix.
    min_weight (int): Minimum weight to include a link (default = 1).

    Returns:
    pd.DataFrame: DataFrame with columns: "source", "target", "weight".
    """
    if matrix_df.empty:
        return pd.DataFrame(columns=["source", "target", "weight"])

    links = []
    for i, source in enumerate(matrix_df.index):
        for j, target in enumerate(matrix_df.columns):
            if j <= i:
                continue  # Only upper triangle to avoid duplicates
            weight = matrix_df.iloc[i, j]
            if weight >= min_weight:
                links.append((source, target, weight))

    return pd.DataFrame(links, columns=["source", "target", "weight"])

"""Functions for identifying and working with top-cited documents, authors or sources."""
# top cited

def select_global_top_cited_documents(
    df: pd.DataFrame,
    top_n: int = 10,
    cols: list | None = None,
    filters: dict | None = None,
    cite_col: str = 'Cited by',
    include_ties: bool = False,
) -> pd.DataFrame:
    """
    Return globally top-cited documents, sorted and trimmed correctly.
    """
    # Helper to find first available column from candidates
    def _find_col(candidates, available):
        for c in candidates:
            if c in available:
                return c
        return None
    
    available_cols = df.columns.tolist()
    
    if cols is None:
        # Build default cols list using available columns
        author_col = _find_col(["Authors", "Authors or Inventors", "Author full names"], available_cols)
        title_col = _find_col(["Title"], available_cols)
        source_col = _find_col(["Source title"], available_cols)
        year_col = _find_col(["Year"], available_cols)
        doctype_col = _find_col(["Document Type"], available_cols)
        
        cols = [c for c in [author_col, title_col, source_col, year_col, doctype_col] if c is not None]

    data = df.copy()
    if filters:
        for k, cond in filters.items():
            data = data[data[k].apply(cond)]

    data[cite_col] = pd.to_numeric(data[cite_col], errors="coerce").fillna(0)
    sort_cols = [cite_col]
    ascending = [False]
    if "Year" in data.columns:
        data["Year"] = pd.to_numeric(data["Year"], errors="coerce")
        sort_cols.append("Year")
        ascending.append(True)
    if "Title" in data.columns:
        sort_cols.append("Title")
        ascending.append(True)

    data = data.sort_values(sort_cols, ascending=ascending)

    if include_ties and top_n < len(data):
        cutoff = data[cite_col].iloc[top_n - 1]
        data = data[data[cite_col] >= cutoff]
    else:
        data = data.head(top_n)

    missing = [c for c in cols if c not in data.columns]
    if missing:
        raise KeyError(f"Missing columns: {missing}. Available: {list(data.columns)}")

    out_cols = list(dict.fromkeys(cols + [cite_col]))
    return data[out_cols].reset_index(drop=True)

def select_local_top_cited_documents(
    df,
    top_n = 10,
    cols = None,
    filters = None,
    title_col = 'Title',
    ref_col = 'References',
    cite_col = 'Cited by',
):
    """
    Select locally top-cited documents based on how often their title appears in other documents' References.

    Parameters:
    - df: input DataFrame
    - top_n: number of top entries (default: 10)
    - cols: columns to return (default: ["Authors", "Title", "Source title", "Year", "Document Type"])
    - filters: dict of column: condition (e.g. {"Year": lambda x: x >= 2015})
    - title_col: name of the column containing document titles
    - ref_col: name of the column containing references
    - cite_col: name of the column with global citations (renamed to "Global citations")

    Returns:
    - DataFrame with top locally cited documents, including local and global citation counts
    """
    if cols is None:
        cols = ["Authors", "Title", "Source title", "Year", "Document Type"]

    if filters:
        for k, cond in filters.items():
            df = df[df[k].apply(cond)]

    titles = df[title_col].dropna().unique()
    title_counts = dict.fromkeys(titles, 0)

    # Count how many times each title appears in the References column
    for ref in df[ref_col].dropna():
        for t in titles:
            if t in ref:
                title_counts[t] += 1

    # Create a local citation column
    df = df.copy()
    df["Local citations"] = df[title_col].map(title_counts).fillna(0).astype(int)
    df["Global citations"] = df[cite_col]

    # Sort by local citations and handle ties
    df_sorted = df.sort_values(by="Local citations", ascending=False)
    cutoff = df_sorted["Local citations"].iloc[top_n - 1] if top_n < len(df_sorted) else -1
    df_top = df_sorted[df_sorted["Local citations"] >= cutoff]

    out_cols = list(dict.fromkeys(cols + ["Local citations", "Global citations"]))
    return df_top[out_cols].reset_index(drop=True)

def ensure_citations_per_year(
    df: pd.DataFrame,
    year_col: str = 'Year',
    cite_col: str = 'Cited by',
    current_year: int | None = None,
    out_col: str = 'Citations per year',
    overwrite: bool = False,
) -> pd.DataFrame:
    """
    Compute normalized citation rate and store it in `out_col`.

    Formula:
        citations_per_year = (Cited by) / (current_year - Year + 1), age clipped to >= 1.
    """
    if (out_col in df.columns) and not overwrite:
        return df.copy()

    out = df.copy()
    y = pd.to_numeric(out[year_col], errors="coerce")
    if current_year is None:
        current_year = int(y.max())
    age = (current_year - y + 1).clip(lower=1)
    cites = pd.to_numeric(out[cite_col], errors="coerce").fillna(0)
    out[out_col] = cites / age
    return out

def select_top_cited_normalized_per_year(
    df: pd.DataFrame,
    top_n: int = 10,
    cols: list = None,
    filters: dict = None,
    year_col: str = 'Year',
    cite_col: str = 'Cited by',
    current_year: int = None,
) -> pd.DataFrame:
    """
    Select top-cited documents normalized by citation rate per year since publication.

    Normalization formula:
        citations_per_year = (Cited by) / (current_year - Year + 1)

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing publication year and citation count columns.
    top_n : int, default=10
        Number of top entries to return.
    cols : list, optional
        Columns to include in the output.
        Default: ["Authors", "Title", "Source title", "Year", "Document Type"]
    filters : dict, optional
        Dictionary of filtering conditions (e.g. {"Year": lambda x: x >= 2015}).
    year_col : str, default="Year"
        Name of the year column.
    cite_col : str, default="Cited by"
        Name of the citation count column.
    current_year : int, optional
        Year used for normalization. If None, uses the maximum year in the dataset.

    Returns
    -------
    pd.DataFrame
        DataFrame containing top documents ranked by normalized citation rate per year.
    """

    if cols is None:
        cols = ["Authors", "Title", "Source title", "Year", "Document Type"]

    if filters:
        for k, cond in filters.items():
            df = df[df[k].apply(cond)]

    if current_year is None:
        current_year = df[year_col].max()

    df = df.copy()
    df["Citations per year"] = df[cite_col] / (current_year - df[year_col] + 1)

    df_sorted = df.sort_values(by="Citations per year", ascending=False)
    cutoff = df_sorted["Citations per year"].iloc[top_n - 1] if top_n < len(df_sorted) else -1
    df_top = df_sorted[df_sorted["Citations per year"] >= cutoff]

    out_cols = list(dict.fromkeys(cols + [cite_col, "Citations per year"]))
    return df_top[out_cols].reset_index(drop=True)

def top_openalex_local_cited_documents(
    df: pd.DataFrame,
    refs_col: str = 'References',
    id_col: str = 'References',
    sep: str = '|',
    top_n: int = 10,
    cols: Optional[List[str]] = None,
) -> pd.DataFrame:
    """
    Compute Top-N locally cited OpenAlex works and attach selected columns from `df`.

    A "local" cited work is any reference that also exists in the same `df`, matched via `id_col`.
    This simplified version only:
      1) counts local citations,
      2) merges those counts with a (deduplicated) subset of `df` using normalized OpenAlex IDs,
      3) appends a canonical "OpenAlex URL".

    Parameters
    ----------
    df : pandas.DataFrame
        Input table with:
        - `id_col`: each row's OpenAlex work link/ID (e.g., "https://openalex.org/W123..." or "W123...").
        - `refs_col`: the reference list per document (string with delimiter or iterable).
    refs_col : str, default "References"
        Column containing references.
    id_col : str, default "Reference"
        Column containing each document's own OpenAlex link/ID.
    sep : str, default "; "
        Delimiter if `refs_col` is stored as a single string.
    top_n : int, default 10
        Number of locally cited documents to return (<=0 returns all).
    cols : list[str] or None, default None
        Extra columns to include from `df`. These are added to the built-in defaults
        ["Authors", "Title", "Source title", "Year", "Document Type"] and only kept if present.

    Returns
    -------
    pandas.DataFrame
        Columns:
        - "OpenAlex ID": normalized "W\\d+" identifier
        - "Local Citations": local in-collection citation count
        - "OpenAlex URL": canonical "https://openalex.org/W..."
        - Selected columns from `df` (first occurrence per local work)
        Sorted by "Local Citations" (desc), then "OpenAlex ID" (asc).
    """
    # --- helpers --------------------------------------------------------------
    wid_re = re.compile(r"(W\d+)")
    def norm_id(
        x,
    ) -> Optional[str]:
        if not isinstance(x, str):
            return None
        m = wid_re.search(x)
        return m.group(1) if m else None

    def to_list(
        val,
    ) -> List[str]:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return []
        if isinstance(val, str):
            return [s for s in (v.strip() for v in val.split(sep)) if s]
        if isinstance(val, Iterable):
            return [str(v) for v in val]
        return [str(val)]

    def to_url(
        wid: str,
    ) -> str:
        return f"https://openalex.org/{wid}"

    # --- normalize local IDs --------------------------------------------------
    df_ids = df[id_col].map(norm_id)
    local_id_set = set(df_ids.dropna().astype(str))

    # --- count local citations ------------------------------------------------
    counter = Counter()
    for refs in df[refs_col]:
        for wid in (norm_id(x) for x in to_list(refs)):
            if wid and wid in local_id_set:
                counter[wid] += 1

    # Empty result early-exit
    if not counter:
        # Prepare predictable schema (counts + URL + selected cols)
        default_cols = ["Authors", "Title", "Source title", "Year", "Document Type", "Cited by", "Percent Rank Cited by"]
        want_all = list(dict.fromkeys(default_cols + (cols or [])))  # deduplicate, keep order
        keep = [c for c in want_all if c in df.columns]
        out = pd.DataFrame(columns=["OpenAlex ID", "Local Citations", "OpenAlex URL"] + keep)
        return out

    counts_df = (
        pd.DataFrame(counter.items(), columns=["OpenAlex ID", "Local Citations"])
        .sort_values(["Local Citations", "OpenAlex ID"], ascending=[False, True])
        .reset_index(drop=True)
    )
    if top_n and top_n > 0:
        counts_df = counts_df.head(top_n)

    # --- prepare columns to attach from df -----------------------------------
    default_cols = ["Authors", "Title", "Source title", "Year", "Document Type", "Cited by", "Percent Rank Cited by"]
    want_all = list(dict.fromkeys(default_cols + (cols or [])))  # deduplicate while preserving order
    keep_cols = [c for c in want_all if c in df.columns]

    # Build a one-row-per-local-work auxiliary table from df
    # Handle the case where id_col might already be "OpenAlex ID"
    if id_col == "OpenAlex ID":
        # id_col is already "OpenAlex ID", so we just need to normalize it
        aux_cols = ["OpenAlex ID"] + keep_cols if keep_cols else ["OpenAlex ID"]
        aux = df[aux_cols].copy()
        aux["OpenAlex ID"] = df_ids  # overwrite with normalized IDs
    else:
        # id_col is different (e.g., "unique-id"), include both then drop original
        aux_cols = [id_col] + keep_cols if keep_cols else [id_col]
        aux = df[aux_cols].copy()
        aux["OpenAlex ID"] = df_ids  # add normalized IDs as new column
        aux = aux.drop(columns=[id_col])  # drop the original id column
    
    aux = aux.dropna(subset=["OpenAlex ID"]).drop_duplicates(subset=["OpenAlex ID"], keep="first")

    # Merge counts with local records and add URL
    out = counts_df.merge(aux, on="OpenAlex ID", how="left")
    out["OpenAlex URL"] = out["OpenAlex ID"].map(to_url)

    # Reorder columns: counts + URL + selected cols
    ordered = ["OpenAlex ID", "Local Citations", "OpenAlex URL"] + keep_cols
    return out[ordered]

"""Helpers for handling language codes, translation hints and language-related metadata."""
# language translation

# Try to load language dictionary, but handle missing/malformed files gracefully
try:
    _lang_dict_path = os.path.join(fd, "additional files", "language dictionary.xlsx")
    lang_dict_df = pd.read_excel(_lang_dict_path)
    # Check if required columns exist
    if "term" not in lang_dict_df.columns:
        lang_dict_df = pd.DataFrame({"term": [], "en": []})  # Empty but valid structure
except Exception:
    lang_dict_df = pd.DataFrame({"term": [], "en": []})  # Empty but valid structure


def ldf(
    x,
    lang_dict_df = lang_dict_df,
    l = 'en',
):
    """
    Translate a term using the language dictionary.
    
    Parameters
    ----------
    x : str
        Term to translate.
    lang_dict_df : pd.DataFrame
        Language dictionary DataFrame with 'term' column and language columns.
    l : str
        Target language code (default 'en').
        
    Returns
    -------
    str
        Translated term, or original if not found.
    """
    # Handle missing/empty dictionary gracefully
    if lang_dict_df is None or lang_dict_df.empty or "term" not in lang_dict_df.columns:
        return x
    if l not in lang_dict_df.columns:
        return x
    try:
        lang_dict = lang_dict_df.set_index("term").to_dict()[l]
        return lang_dict.get(x, x)
    except Exception:
        return x

"""Core descriptive-statistics routines operating on bibliometric data frames and series."""
# descriptive statistics

def compute_descriptives(
    df: pd.DataFrame,
    column: str,
    col_type: Literal['numeric', 'categorical', 'text', 'list', 'auto'] = 'auto',
    stopwords: Optional[Union[Iterable[str], set]] = None,
    list_separators: str = '; ',
    top_n: int = 10,
    as_frame: bool = False,
    citation_thresholds: Optional[List[int]] = None,
    extra_stats: bool = False,
) -> Union[Dict[str, Any], pd.DataFrame]:
    """
    Compute descriptive statistics for a DataFrame column with robust typing and edge-case handling.

    Key features
    ------------
    - Optional skewness, kurtosis, and normality test via `extra_stats` (disabled by default).
    - Guards SciPy normality test: if sample size < 8, returns NaNs (no exception).
    - Adds bibliometric citation thresholds Ck and "Ck [%]" **only** for citation-count columns
      (e.g., "Cited by"); skipped for other numeric columns like Year, Pages, etc.
      Defaults to C1, C5, C10, C20, C50, C100, C1000 (configurable via `citation_thresholds`).
    - "auto" type inference for convenience; supports "numeric", "categorical",
      "text", and "list" columns.
    - Optional `as_frame=True` returns a compact one-row DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    column : str
        Name of the column to analyze.
    col_type : {"numeric","categorical","text","list","auto"}, default "auto"
        Column interpretation; "auto" uses a light heuristic.
    stopwords : Iterable[str] or set or None, default None
        Custom stopwords for text analysis; defaults to ENGLISH_STOP_WORDS.
    list_separators : str, default r";"
        Regex for splitting multi-value "list" cells (e.g., r"[;,]").
    top_n : int, default 10
        Number of top categories/words/items reported.
    as_frame : bool, default False
        If True, return a one-row DataFrame with flattened keys.
    citation_thresholds : list[int] or None, default None
        Thresholds k for which to compute Ck and "Ck [%]". Defaults to [1, 5, 10, 20, 50, 100, 1000].
    extra_stats : bool, default False
        If True, compute skewness, kurtosis, and the normality test.

    Returns
    -------
    dict or pd.DataFrame
        Descriptive summary (or a one-row DataFrame if `as_frame=True`).
    """
    if column not in df.columns:
        # Column doesn't exist - return empty result
        return pd.DataFrame([{}]) if as_frame else {}

    series = df[column]
    result: Dict[str, Any] = {}

    non_missing = series.dropna()
    n_non_missing = int(non_missing.shape[0])
    n_missing = int(series.isna().sum())
    result["Number of documents"] = n_non_missing
    result["Missing values"] = {
        "count": n_missing,
        "percent": round((n_missing / len(series)) * 100, 2) if len(series) else 0.0,
    }

    if n_non_missing == 0:
        return pd.DataFrame([result]) if as_frame else result

    # ---- Type inference ("auto") ----
    inferred = col_type
    if col_type == "auto":
        coerced = pd.to_numeric(non_missing, errors="coerce")
        looks_numeric = (coerced.notna().mean() >= 0.7) and (coerced.dropna().nunique() >= 3)
        has_list_sep = non_missing.astype(str).str.contains(list_separators).any()
        token_counts = non_missing.astype(str).str.findall(r"\b\w+\b").map(len)
        avg_tokens = float(token_counts.mean()) if len(token_counts) else 0.0

        if looks_numeric:
            inferred = "numeric"
        elif has_list_sep:
            inferred = "list"
        elif series.dtype == "object" and avg_tokens >= 3:
            inferred = "text"
        else:
            inferred = "categorical"

    # Helper: decide if the numeric column is a citation-count field
    def _is_citation_column(
        name: str,
    ) -> bool:
        name_l = name.strip().lower()
        candidates = {
            "cited by", "cited_by", "citations", "times cited", "times_cited",
            "citedby", "n_citations", "citation count", "citation_count"
        }
        if name_l in candidates:
            return True
        return any(tok in name_l for tok in ["cited", "citation"])

    # ---- Numeric ----
    if inferred == "numeric":
        values = pd.to_numeric(non_missing, errors="coerce").dropna()
        if values.empty:
            # No valid numeric values - just return basic stats already computed
            return pd.DataFrame([result]) if as_frame else result

        percentiles = {f"{p}%": float(np.nanpercentile(values, p)) for p in (10, 25, 50, 75, 90)}
        result.update(
            {
                "Mean": float(values.mean()),
                "Median": float(values.median()),
                "Mode": float(values.mode().iloc[0]) if not values.mode().empty else None,
                "Min": float(values.min()),
                "Max": float(values.max()),
                "Range": float(values.max() - values.min()),
                "Standard deviation": float(values.std(ddof=1)),
                "Percentiles": percentiles,
            }
        )

        # Bibliometric thresholds Ck and "Ck [%]" — ONLY for citation columns
        if _is_citation_column(column):
            thresholds = citation_thresholds if citation_thresholds is not None else [1, 5, 10, 20, 50, 100, 1000]
            n_total = int(values.shape[0])
            for k in thresholds:
                n_ge = int((values >= k).sum())
                perc = round(100 * n_ge / n_total, 2) if n_total else 0.0
                result[f"C{k}"] = n_ge
                result[f"C{k} [%]"] = perc

        # Extra stats (optional)
        if extra_stats:
            result["Skewness"] = float(skew(values, bias=False, nan_policy="omit"))
            result["Kurtosis"] = float(kurtosis(values, bias=False, nan_policy="omit"))

            if len(values) >= 8:
                try:
                    stat, p = normaltest(values, nan_policy="omit")
                    result["Normality test (D’Agostino–Pearson)"] = {
                        "statistic": float(stat),
                        "p-value": float(p),
                        "normal": bool(p >= 0.05),
                    }
                except Exception:
                    result["Normality test (D’Agostino–Pearson)"] = {
                        "statistic": np.nan,
                        "p-value": np.nan,
                        "normal": np.nan,
                    }
            else:
                result["Normality test (D’Agostino–Pearson)"] = {
                    "statistic": np.nan,
                    "p-value": np.nan,
                    "normal": np.nan,
                }

    # ---- Categorical ----
    elif inferred == "categorical":
        freqs = non_missing.value_counts(dropna=False)
        total = int(freqs.sum())
        top_items = freqs.head(top_n)
        result.update(
            {
                "Number of unique values": int(non_missing.nunique(dropna=False)),
                "Top categories": [
                    {"value": idx, "count": int(count), "percent": round(100 * count / total, 2) if total else 0.0}
                    for idx, count in top_items.items()
                ],
            }
        )

    # ---- Text ----
    elif inferred == "text":
        # Ensure stopwords is a set, never None
        if stopwords is not None:
            sw = set(stopwords)
        elif ENGLISH_STOP_WORDS is not None:
            sw = set(ENGLISH_STOP_WORDS) if not isinstance(ENGLISH_STOP_WORDS, set) else ENGLISH_STOP_WORDS
        else:
            # Fallback minimal stopwords if nothing else available
            sw = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 
                  'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 
                  'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'shall', 'can',
                  'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they'}
        
        tokens_per_row = non_missing.astype(str).str.findall(r"\b\w+\b")
        # Handle None values from findall - convert to list and filter
        tokens_list = []
        for row in tokens_per_row.tolist():
            if row is None:
                tokens_list.append([])
            elif isinstance(row, list):
                tokens_list.append([w for w in row if w is not None])
            else:
                tokens_list.append([])
        
        word_lengths = pd.Series([len(row) for row in tokens_list])
        all_words = []
        for row in tokens_list:
            if row:
                for w in row:
                    if w and isinstance(w, str) and w.lower() not in sw:
                        all_words.append(w.lower())
        total_words = len(all_words)
        top_words = Counter(all_words).most_common(top_n)

        result.update(
            {
                "Length (in words)": {
                    "mean": float(word_lengths.mean()),
                    "median": float(word_lengths.median()),
                    "min": int(word_lengths.min()),
                    "max": int(word_lengths.max()),
                    "std": float(word_lengths.std(ddof=1)) if len(word_lengths) > 1 else 0.0,
                },
                "Total unique words": int(len(set(all_words))),
                "Top words": [
                    {
                        "word": word,
                        "count": int(count),
                        "percent": round(100 * count / total_words, 2) if total_words else 0.0,
                    }
                    for word, count in top_words
                ],
            }
        )

    # ---- List (multi-value cells) ----
    elif inferred == "list":
        list_separators = ["|", ";", ","]
        pattern = "|".join(re.escape(sep) for sep in list_separators)

        parsed = non_missing.astype(str).apply(
            lambda x: [i.strip().lower() for i in re.split(pattern, x) if i.strip()]
        )
        lengths = parsed.map(len)
        all_items = [item for sub in parsed for item in sub]
        total_items = len(all_items)
        item_counts = Counter(all_items).most_common(top_n)

        result.update(
            {
                "List length": {
                    "mean": float(lengths.mean()),
                    "median": float(lengths.median()),
                    "min": int(lengths.min()),
                    "max": int(lengths.max()),
                    "std": float(lengths.std(ddof=1)) if len(lengths) > 1 else 0.0,
                },
                "Total unique elements": int(len(set(all_items))),
                "Top items": [
                    {
                        "item": item,
                        "count": int(count),
                        "percent": round(100 * count / total_items, 2) if total_items else 0.0,
                    }
                    for item, count in item_counts
                ],
            }
        )

    # If col_type is not recognized, we simply don't add extra statistics
    # (the basic Number of documents and Missing values are already included)

    if not as_frame:
        return result

    # Flatten nested dict for a one-row DataFrame
    def _flatten(
        d: Dict[str, Any],
        prefix: str = '',
    ) -> Dict[str, Any]:
        flat = {}
        for k, v in d.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                flat.update(_flatten(v, key))
            else:
                flat[key] = v
        return flat

    flat = _flatten(result)
    return pd.DataFrame([flat])

def flatten_descriptives(
    name,
    summary,
):
    """
    Flatten a nested dictionary into a list of (Variable, Indicator, Value) tuples.

    - Lists of dicts (e.g., Top 10 items) are collapsed into a single string.
    - Items are separated by ';\n' for readability in Excel cells.
    """
    rows = []
    for key, value in summary.items():
        if isinstance(value, dict):
            for subkey, subval in value.items():
                label = f"{key} - {subkey}"
                rows.append((name, label, subval))
        elif isinstance(value, list):
            if all(isinstance(item, dict) and "count" in item for item in value):
                collapsed = ";\n".join(
                    f"{item.get('value') or item.get('word') or item.get('item')} ({item['count']})"
                    for item in value
                )
                rows.append((name, key, collapsed))
            else:
                rows.append((name, key, "\n".join(str(x) for x in value)))
        else:
            rows.append((name, key, value))
    return rows

def save_descriptives_to_excel(
    dataframes_with_sheets,
    excel_path,
    freeze_top_row: bool = False,
) -> None:
    """
    Save one or more descriptive summary DataFrames to a styled Excel file.

    Parameters
    ----------
    dataframes_with_sheets : list[tuple[pd.DataFrame, str]]
        List of (DataFrame, sheet_name) pairs.
    excel_path : str
        Path to save the Excel file (existing file will be overwritten).
    freeze_top_row : bool, default=False
        Whether to freeze the top row on each sheet.
    """
    from pathlib import Path
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        for df, raw_sheet_name in dataframes_with_sheets:
            if df is None or df.empty:
                continue

            # Ensure valid, unique sheet name (Excel limit 31 chars)
            base_name = str(raw_sheet_name)[:31] or "Sheet1"
            sheet_name = base_name
            counter = 1
            while sheet_name in used_sheet_names:
                suffix = f" ({counter})"
                sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"
                counter += 1
            used_sheet_names.add(sheet_name)

            # Write DataFrame
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
            sheet = writer.sheets[sheet_name]

            # Bold and center-align header row
            for col_idx in range(1, df.shape[1] + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Merge and center-align cells in "Variable" column (if it exists)
            if "Variable" in df.columns:
                var_col_idx = df.columns.get_loc("Variable") + 1
                current_row = 2
                for var, group_df in df.groupby("Variable", sort=False):
                    count = len(group_df)
                    if count > 1:
                        sheet.merge_cells(
                            start_row=current_row,
                            start_column=var_col_idx,
                            end_row=current_row + count - 1,
                            end_column=var_col_idx,
                        )
                    cell = sheet.cell(row=current_row, column=var_col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    current_row += count

            # Freeze header row if requested
            if freeze_top_row:
                sheet.freeze_panes = "A2"

            # Autofit column widths
            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 100)
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Wrap text in all data cells and keep existing alignment settings
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    alignment = cell.alignment or Alignment()
                    cell.alignment = Alignment(
                        horizontal=alignment.horizontal,
                        vertical=alignment.vertical,
                        wrap_text=True,
                    )


def compute_descriptive_statistics(
    df,
    columns_with_types,
    stopwords = None,
    extra_stats = False,
    sep = '; ',
):
    """
    Compute descriptive bibliometric statistics for multiple columns.

    Parameters:
        df (pd.DataFrame): Input DataFrame.
        columns_with_types (list): List of (column_name, column_type) pairs.
        stopwords (set or list): Custom stopwords for text analysis.

    Returns:
        pd.DataFrame: Summary table with columns: Variable, Indicator, Value.
    """
    all_rows = []

    for column, col_type in columns_with_types:
        summary = compute_descriptives(df, column, col_type, stopwords, extra_stats=extra_stats, list_separators=sep)
        flat = flatten_descriptives(column, summary)
        all_rows.extend(flat)

    return pd.DataFrame(all_rows, columns=["Variable", "Indicator", "Value"])

"""Counting helpers for occurrences of items such as authors, keywords or sources in bibliographic data."""
# counting

def _balance_closing_parenthesis(s: str) -> str:
    """
    Balance parentheses in a keyword.

    If there are more '(' than ')', append a ')' at the end.
    Mirrors the behaviour used in ``count_occurrences`` so that
    keywords like 'factor (programming language' become
    'factor (programming language)'.
    """
    if not isinstance(s, str):
        s = str(s)
    if s.count("(") > s.count(")"):
        return s + ")"
    return s


import pandas as pd


def count_occurrences(
    df: pd.DataFrame,
    column_name: str,
    count_type: str = "single",
    ngram_range: tuple[int, int] = (1, 1),
    item_column_name: str = "Item",
    rename_dict: dict | None = None,
    translated_column_name: str = "Translated Item",
    sep: str = "; ",
    token_pattern: str = r"(?u)[^\s]+",
) -> pd.DataFrame:
    """
    Process a DataFrame column and return a DataFrame with counts,
    proportions, and ranks.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    column_name : str
        Name of the column to process.
    count_type : {"single", "list", "text"}, default "single"
        Type of processing:

        - "single": each cell is a single item.
        - "list": each cell is a delimited list of items (separated by ``sep``).
        - "text": free text; n-grams are extracted with CountVectorizer.
    ngram_range : tuple, default (1, 1)
        N-gram range for text processing (passed to CountVectorizer).
    item_column_name : str, default "Item"
        Name of the item column in the output.
    rename_dict : dict or None, default None
        If provided, adds a column with translated / renamed items,
        mapped from ``item_column_name``.
    translated_column_name : str, default "Translated Item"
        Name of the translated column if ``rename_dict`` is provided.
    sep : str, default "; "
        Separator used for "list" count_type. Treated as a literal,
        so values like "|" are safe. Sequences like "")|" or "); "
        are treated as ")" + separator.
    token_pattern : str, default r"(?u)[^\\s]+"
        Regex pattern for tokenization in the "text" mode. The default
        keeps all non-whitespace characters together, so parentheses
        are preserved.

    Returns
    -------
    pd.DataFrame
        Processed counts sorted in descending order of
        "Number of documents", with additional columns:

        - "Rank"        (1 = most frequent)
        - "Percentrank" (0–1, top item = 1)
    """

    def _balance_closing_parenthesis(s: str) -> str:
        """If there are more '(' than ')', append ')' to balance."""
        if s.count("(") > s.count(")"):
            return s + ")"
        return s

    def _add_rank_columns(df_out: pd.DataFrame) -> pd.DataFrame:
        """Add Rank and Percentrank columns based on Number of documents."""
        if df_out.empty:
            df_out["Rank"] = []
            df_out["Percentrank"] = []
            return df_out

        df_out = df_out.sort_values(
            by="Number of documents",
            ascending=False,
        ).reset_index(drop=True)

        n = len(df_out)
        df_out["Rank"] = np.arange(1, n + 1)
        if n == 1:
            df_out["Percentrank"] = 1.0
        else:
            # Top item => 1.0, lowest => 0.0
            df_out["Percentrank"] = (n - df_out["Rank"]) / (n - 1)
        return df_out

    total_rows = len(df)

    # Remove NaNs and trim outer whitespace
    data = df[column_name].dropna().astype(str).str.strip()
    # Remove empty values
    data = data[data != ""]

    if count_type == "single":
        counts = Counter(data)

    elif count_type == "list":
        # Treat sep as literal, so "|" and "; " etc. are safe.
        split_series = data.str.split(sep, regex=False)

        # Convert to list-of-lists and clean each item:
        #   - strip whitespace only (do not touch parentheses)
        #   - balance closing parentheses if needed
        split_series = split_series.dropna().tolist()
        cleaned_lists: list[list[str]] = []
        for items in split_series:
            cleaned: list[str] = []
            for item in items:
                s = item.strip()
                if not s:
                    continue
                s = _balance_closing_parenthesis(s)
                cleaned.append(s)
            cleaned_lists.append(cleaned)

        # Flatten, preserving parentheses
        flattened = list(chain.from_iterable(cleaned_lists))
        counts = Counter(flattened)

        fractional_counts: Counter = Counter()
        for items in cleaned_lists:
            unique_items = set(items)
            weight = 1 / len(unique_items) if unique_items else 0
            for item in unique_items:
                fractional_counts[item] += weight

    elif count_type == "text":
        if data.empty:
            columns = [
                item_column_name,
                "Number of documents",
                "Proportion of documents",
                "Percentage of documents",
                "Number of occurrences",
                "Rank",
                "Percentrank",
            ]
            if rename_dict is not None:
                columns.insert(1, translated_column_name)
            return pd.DataFrame(columns=columns)

        vectorizer = CountVectorizer(
            ngram_range=ngram_range,
            token_pattern=token_pattern,
        )
        term_matrix = vectorizer.fit_transform(data)
        terms = vectorizer.get_feature_names_out()

        doc_counts = (term_matrix > 0).sum(axis=0).A1
        total_counts = term_matrix.sum(axis=0).A1

        result_df = pd.DataFrame(
            {
                item_column_name: terms,
                "Number of documents": doc_counts,
                "Proportion of documents": doc_counts / total_rows,
                "Percentage of documents": (doc_counts / total_rows) * 100,
                "Number of occurrences": total_counts,
            }
        )

        if rename_dict is not None:
            result_df.insert(
                1,
                translated_column_name,
                result_df[item_column_name].map(rename_dict).fillna(""),
            )

        # Add Rank & Percentrank (by Number of documents)
        result_df = _add_rank_columns(result_df)
        return result_df

    else:
        raise ValueError('Invalid count_type. Choose from "single", "list", or "text".')

    # "single" and "list" branches share this aggregation
    result_data: dict[str, list] = {
        item_column_name: list(counts.keys()),
        "Number of documents": list(counts.values()),
        "Proportion of documents": [v / total_rows for v in counts.values()],
        "Percentage of documents": [(v / total_rows) * 100 for v in counts.values()],
    }

    if count_type == "list":
        result_data["Fractional number of documents"] = [
            fractional_counts[item] for item in counts.keys()
        ]

    result_df = pd.DataFrame(result_data)

    if rename_dict is not None:
        result_df.insert(
            1,
            translated_column_name,
            result_df[item_column_name].map(rename_dict).fillna(""),
        )

    # Add Rank & Percentrank (by Number of documents)
    result_df = _add_rank_columns(result_df)
    return result_df





"""Time-series helpers for analysing scientific production, such as documents or citations per year."""
# Scientific production

def get_scientific_production(
    df,
    relative_counts=True,
    cumulative=True,
    predict_last_year=True,
    percent_change=True,
):
    """
    Computes the annual scientific production statistics from a dataset containing publication years and citation counts.
    """
    # --- 1. ROBUST DATA CLEANING ---
    # Work on a copy to avoid SettingWithCopyWarning
    df = df.copy()

    # Fix the "Year" column:
    # pd.to_numeric handles strings like "2021", "2021.0", and turns "N/A" into NaN
    df["Year"] = pd.to_numeric(df["Year"], errors='coerce')
    
    # Drop rows where Year is NaN (invalid data)
    df = df.dropna(subset=["Year"])
    
    # Now it is safe to convert to integer
    df["Year"] = df["Year"].astype(int)

    # Clean "Cited by" column to ensure sums work
    if "Cited by" in df.columns:
        df["Cited by"] = pd.to_numeric(df["Cited by"], errors='coerce').fillna(0)
    else:
        df["Cited by"] = 0

    # Safety check: If dataframe is empty after cleaning, return empty result
    if df.empty:
        return pd.DataFrame()

    # --- 2. CORE LOGIC ---

    # Create a complete range of years
    all_years = pd.Series(range(df["Year"].min(), df["Year"].max() + 1), name="Year")

    # Aggregate counts and citations
    production = df.groupby("Year").agg(
        Documents=("Year", "count"),
        Total_Citations=("Cited by", "sum")
    ).reset_index()

    # Merge with the full range of years and fill missing values with 0
    production = all_years.to_frame().merge(production, on="Year", how="left").fillna(0)

    # Convert back to integer where applicable
    production["Documents"] = production["Documents"].astype(int)
    production["Total_Citations"] = production["Total_Citations"].astype(int)

    # Compute relative counts if needed
    if relative_counts:
        total_docs = production["Documents"].sum()
        production["Proportion Documents"] = production["Documents"] / total_docs if total_docs > 0 else 0
        production["Percentage Documents"] = production["Proportion Documents"] * 100

    # Compute cumulative values if needed
    if cumulative:
        production["Cumulative Documents"] = production["Documents"].cumsum()
        production["Cumulative Citations"] = production["Total_Citations"].cumsum()

        if relative_counts:
            # Re-calculate total_docs to be safe
            total_docs = production["Documents"].sum() 
            production["Cumulative Proportion Documents"] = production["Cumulative Documents"] / total_docs if total_docs > 0 else 0
            production["Cumulative Percentage Documents"] = production["Cumulative Proportion Documents"] * 100

    # Compute percentage change if needed
    if percent_change:
        production["Percentage Change Documents"] = production["Documents"].pct_change() * 100
        production["Percentage Change Citations"] = production["Total_Citations"].pct_change() * 100

    # Predict last year if applicable
    if predict_last_year and len(production) >= 2:
        current_year = datetime.datetime.now().year
        
        # Only predict if the data actually extends to the current year
        if production["Year"].max() == current_year:
            previous_years = production[production["Year"] < current_year]
            
            # Avoid crash if previous_years is empty
            if not previous_years.empty:
                avg_growth_docs = previous_years["Documents"].pct_change().mean()
                avg_growth_citations = previous_years["Total_Citations"].pct_change().mean()

                # Calculate Predicted Documents
                if np.isfinite(avg_growth_docs):
                    pred_docs = production["Documents"].iloc[-2] * (1 + avg_growth_docs)
                else:
                    pred_docs = production["Documents"].iloc[-2]
                
                production.loc[production["Year"] == current_year, "Predicted Documents"] = pred_docs

                # Calculate Predicted Citations
                if np.isfinite(avg_growth_citations):
                    pred_cits = production["Total_Citations"].iloc[-2] * (1 + avg_growth_citations)
                else:
                    pred_cits = production["Total_Citations"].iloc[-2]
                
                production.loc[production["Year"] == current_year, "Predicted Citations"] = pred_cits

                # Fill NaNs in prediction columns with actual values and convert to int
                production["Predicted Documents"] = production["Predicted Documents"].fillna(production["Documents"]).astype(int)
                production["Predicted Citations"] = production["Predicted Citations"].fillna(production["Total_Citations"]).astype(int)

    # Rename columns for final output
    production = production.rename(columns={
        "Documents": "Number of Documents",
        "Total_Citations": "Total Citations",
        "Predicted Documents": "Predicted Number of Documents",
        "Predicted Citations": "Predicted Total Citations"
    })

    return production

def summarize_publication_timeseries(
    production_df,
    exclude_last_year_for_growth: bool = True,
):
    """
    Compute and format summary statistics from a time series production_df.

    Parameters:
        production_df (pd.DataFrame): Must contain:
            "Year", "Number of Documents", "Total Citations",
            "Percentage Change Documents"
        exclude_last_year_for_growth (bool, optional): If True, the last
            year in the series is excluded when computing geometric mean
            growth rates (all windows). This is useful when the last year
            is incomplete. For windowed averages (3/5/10 years), the
            indicator label is written as "Average Growth (YYYY–YYYY)".
            If a growth average cannot be computed, it is omitted from
            the output.

    Returns:
        pd.DataFrame: Formatted DataFrame with columns: Variable, Indicator, Value
    """
    df = production_df.copy()

    if df.empty:
        return pd.DataFrame(
            [("Time series analysis", "Timespan", "No data")],
            columns=["Variable", "Indicator", "Value"],
        )

    if "Year" not in df.columns:
        raise ValueError("production_df must contain column \"Year\".")

    # Ensure numeric years and sort
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    df = df.dropna(subset=["Year"]).sort_values("Year").reset_index(drop=True)

    if df.empty:
        return pd.DataFrame(
            [("Time series analysis", "Timespan", "No valid years")],
            columns=["Variable", "Indicator", "Value"],
        )

    n_years = len(df)

    # Timespan
    timespan = f"{int(df['Year'].min())}–{int(df['Year'].max())}"

    # Most productive year
    if "Number of Documents" in df and not df["Number of Documents"].isna().all():
        max_idx = df["Number of Documents"].idxmax()
        max_docs_row = df.loc[max_idx]
        most_productive = (
            f"{int(max_docs_row['Year'])} "
            f"({int(max_docs_row['Number of Documents'])} documents)"
        )
    else:
        most_productive = "N/A"

    # Helper for highest / lowest growth (documents) – always uses full series
    def _growth_stats(series):
        series = pd.to_numeric(series, errors="coerce").replace(
            [np.inf, -np.inf],
            np.nan,
        )
        valid = series.dropna()
        if valid.empty:
            return None, None
        max_idx = valid.idxmax()
        min_idx = valid.idxmin()
        max_row = df.loc[max_idx]
        min_row = df.loc[min_idx]
        highest = f"{int(max_row['Year'])} ({round(valid.loc[max_idx], 2)}%)"
        lowest = f"{int(min_row['Year'])} ({round(valid.loc[min_idx], 2)}%)"
        return highest, lowest

    if "Percentage Change Documents" in df:
        highest_growth, lowest_growth = _growth_stats(df["Percentage Change Documents"])
    else:
        highest_growth, lowest_growth = None, None

    def _or_na(value):
        return "N/A" if value is None else value

    highest_growth = _or_na(highest_growth)
    lowest_growth = _or_na(lowest_growth)

    # Geometric mean helpers
    def geometric_mean_growth(series):
        series = pd.to_numeric(series, errors="coerce").replace(
            [np.inf, -np.inf],
            np.nan,
        )
        series = series.dropna()
        if len(series) == 0:
            return None
        rates = 1 + series / 100
        gmean = np.prod(rates) ** (1 / len(rates)) - 1
        return round(gmean * 100, 2)

    def _fmt_pct(value):
        if value is None:
            return None
        if isinstance(value, (float, np.floating)) and np.isnan(value):
            return None
        return f"{value}%"

    # Base series for growth and a potentially truncated view excluding last year
    if "Percentage Change Documents" in df:
        if exclude_last_year_for_growth and n_years > 1:
            df_growth = df.iloc[:-1].copy()
        else:
            df_growth = df.copy()
        pct_series = df_growth["Percentage Change Documents"]
    else:
        df_growth = None
        pct_series = None

    # All-years geometric mean (over df_growth)
    if pct_series is not None and not pct_series.dropna().empty:
        g_all = geometric_mean_growth(pct_series)
        avg_growth_all = _fmt_pct(g_all)
    else:
        g_all = None
        avg_growth_all = None

    # Windowed geometric means (3/5/10 years) based on df_growth
    g_3 = g_5 = g_10 = None
    label_3 = label_5 = label_10 = None

    def _range_label(df_slice):
        if df_slice is None or df_slice.empty:
            return None
        start_year = int(df_slice["Year"].iloc[0])
        end_year = int(df_slice["Year"].iloc[-1])
        return f"{start_year}–{end_year}"

    if df_growth is not None and not df_growth.empty:
        if len(df_growth) >= 3:
            tail_3 = df_growth.tail(3)
            g_3 = geometric_mean_growth(tail_3["Percentage Change Documents"])
            label_3 = _range_label(tail_3)
        if len(df_growth) >= 5:
            tail_5 = df_growth.tail(5)
            g_5 = geometric_mean_growth(tail_5["Percentage Change Documents"])
            label_5 = _range_label(tail_5)
        if len(df_growth) >= 10:
            tail_10 = df_growth.tail(10)
            g_10 = geometric_mean_growth(tail_10["Percentage Change Documents"])
            label_10 = _range_label(tail_10)

    indicator_3 = indicator_5 = indicator_10 = None

    if g_3 is not None:
        indicator_3 = (
            f"Average Growth ({label_3})"
            if exclude_last_year_for_growth and label_3 is not None
            else "Average Growth (Last 3 Years)"
        )
        avg_growth_3 = _fmt_pct(g_3)
    else:
        avg_growth_3 = None

    if g_5 is not None:
        indicator_5 = (
            f"Average Growth ({label_5})"
            if exclude_last_year_for_growth and label_5 is not None
            else "Average Growth (Last 5 Years)"
        )
        avg_growth_5 = _fmt_pct(g_5)
    else:
        avg_growth_5 = None

    if g_10 is not None:
        indicator_10 = (
            f"Average Growth ({label_10})"
            if exclude_last_year_for_growth and label_10 is not None
            else "Average Growth (Last 10 Years)"
        )
        avg_growth_10 = _fmt_pct(g_10)
    else:
        avg_growth_10 = None

    # Most influential years (unchanged – uses full df)
    if "Total Citations" in df and "Number of Documents" in df:
        df["Citations per Document"] = df["Total Citations"] / df["Number of Documents"]
    else:
        df["Citations per Document"] = np.nan

    def influential(df_slice):
        if df_slice.empty:
            return None
        cpdoc = pd.to_numeric(
            df_slice["Citations per Document"],
            errors="coerce",
        ).replace([np.inf, -np.inf], np.nan)
        valid = cpdoc.dropna()
        if valid.empty:
            return None
        idx = valid.idxmax()
        row = df_slice.loc[idx]
        return (
            f"{int(row['Year'])} "
            f"({round(valid.loc[idx], 2)} citations/doc)"
        )

    influential_all = influential(df)
    influential_3 = influential(df.tail(3)) if n_years >= 3 else None
    influential_5 = influential(df.tail(5)) if n_years >= 5 else None
    influential_10 = influential(df.tail(10)) if n_years >= 10 else None

    influential_all = _or_na(influential_all)

    # Collect rows
    rows = [
        ("Time series analysis", "Timespan", timespan),
        ("Time series analysis", "Most Productive Year", most_productive),
        ("Time series analysis", "Highest Growth", highest_growth),
        ("Time series analysis", "Lowest Growth", lowest_growth),
    ]

    # Add average growth rows only if they are computable (not None / NaN)
    if avg_growth_all is not None:
        rows.append(
            ("Time series analysis", "Average Growth (All Years)", avg_growth_all),
        )

    if avg_growth_3 is not None and indicator_3 is not None:
        rows.append(
            ("Time series analysis", indicator_3, avg_growth_3),
        )

    if avg_growth_5 is not None and indicator_5 is not None:
        rows.append(
            ("Time series analysis", indicator_5, avg_growth_5),
        )

    if avg_growth_10 is not None and indicator_10 is not None:
        rows.append(
            ("Time series analysis", indicator_10, avg_growth_10),
        )

    rows.append(
        ("Time series analysis", "Most Influential Year", influential_all),
    )
    if influential_3 is not None:
        rows.append(
            (
                "Time series analysis",
                "Most Influential (Last 3 Years)",
                influential_3,
            ),
        )
    if influential_5 is not None:
        rows.append(
            (
                "Time series analysis",
                "Most Influential (Last 5 Years)",
                influential_5,
            ),
        )
    if influential_10 is not None:
        rows.append(
            (
                "Time series analysis",
                "Most Influential (Last 10 Years)",
                influential_10,
            ),
        )

    return pd.DataFrame(rows, columns=["Variable", "Indicator", "Value"])



"""More general time-series utilities for summarising trends over arbitrary metrics."""
# More general timeseries

def compute_time_series_by_year(
    df: pd.DataFrame,
    year_col: str = 'Year',
    agg: Union[str, Callable[[pd.Series], float]] = 'mean',
    include_proportion: bool = False,
    include_percentage: bool = False,
    include_cum_percentage: bool = False,
    year_start: Optional[int] = None,
    year_end: Optional[int] = None,
) -> pd.DataFrame:
    """
    Build a time-series table of numerical columns aggregated by year, with optional
    proportion/percentage metrics, and zero-filled gaps in the year range.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame that contains a numeric year column and other numeric columns.
    year_col : str, default "Year"
        Name of the column containing years (integer-like).
    agg : str or callable, default "mean"
        Aggregation applied to each numeric column per year. Accepts any pandas .agg
        reduction understood by GroupBy (e.g., "mean", "sum", "median", np.mean, etc.).
        Note: a single aggregation must be provided (not a list/dict of multiple aggs).
    include_proportion : bool, default False
        If True, for each numeric variable X add a "Proportion X" column where each
        year's value is X_year / sum(X over the full year range). When the column sum
        is 0, the proportion is set to 0 to avoid division by zero.
    include_percentage : bool, default False
        If True, also include "Percentage X" = 100 * Proportion X.
    include_cum_percentage : bool, default False
        If True, include "Cumulative Percentage X", computed over years in ascending
        order as the cumulative sum of "Percentage X".
    year_start : int, optional
        If provided, force the start of the output year range to this year. Otherwise,
        the minimum observed year in df[year_col] is used.
    year_end : int, optional
        If provided, force the end of the output year range to this year. Otherwise,
        the maximum observed year in df[year_col] is used.

    Returns
    -------
    pd.DataFrame
        A DataFrame indexed by year with one column per numeric variable aggregated
        by the chosen function. Missing years within the range [year_start, year_end]
        (or inferred min..max) are present and filled with 0. Optional columns for
        "Proportion X", "Percentage X", and "Cumulative Percentage X" are appended
        per variable when requested.

    Notes
    -----
    - Proportion/percentage metrics are computed per column across the *entire*
      output year range (i.e., each year's value divided by the column's total over
      the full range). If you need per-year shares across variables instead, we can
      add an option to normalize by row totals.
    - Non-numeric columns (besides the year column) are ignored.
    - Years with NaN or non-integer values in year_col are dropped.

    Examples
    --------
    >>> ts = compute_time_series_by_year(df, agg="sum",
    ...                                  include_proportion=True,
    ...                                  include_percentage=True,
    ...                                  include_cum_percentage=True)
    """
    if year_col not in df.columns:
        raise KeyError(f"Column '{year_col}' not found in DataFrame.")

    # Clean and prepare the year column
    years = pd.to_numeric(df[year_col], errors="coerce")
    valid = years.notna()
    if not valid.any():
        raise ValueError("No valid numeric years found in the DataFrame.")
    df_clean = df.loc[valid].copy()
    df_clean[year_col] = years[valid].astype(int)

    # Determine numeric columns to aggregate (exclude the year column)
    num_cols = df_clean.select_dtypes(include="number").columns.tolist()
    if year_col in num_cols:
        num_cols.remove(year_col)
    if not num_cols:
        raise ValueError("No numeric columns to aggregate besides the year column.")

    # Group by year and aggregate
    grouped = (
        df_clean.groupby(year_col, sort=True)[num_cols]
        .agg(agg)
    )

    # Establish full year range and reindex (fill missing years with 0)
    min_year = grouped.index.min() if year_start is None else int(year_start)
    max_year = grouped.index.max() if year_end is None else int(year_end)
    if min_year > max_year:
        raise ValueError("year_start must be <= year_end.")
    full_year_index = pd.Index(range(min_year, max_year + 1), name=year_col)

    ts = grouped.reindex(full_year_index, fill_value=0).copy()

    # Optional derived metrics per column
    def _safe_div(
        numer: pd.Series,
        denom: float,
    ) -> pd.Series:
        if denom == 0:
            # All zeros -> return zeros to avoid NaN/inf
            return pd.Series(0.0, index=numer.index)
        return numer / denom

    added_cols: list[Hashable] = []
    if include_proportion or include_percentage or include_cum_percentage:
        for col in num_cols:
            total = float(ts[col].sum())
            if include_proportion:
                prop_name = f"Proportion {col}"
                ts[prop_name] = _safe_div(ts[col], total)
                added_cols.append(prop_name)
            if include_percentage:
                # Ensure proportion exists (compute if needed)
                if f"Proportion {col}" not in ts.columns:
                    ts[f"Proportion {col}"] = _safe_div(ts[col], total)
                    added_cols.append(f"Proportion {col}")
                pct_name = f"Percentage {col}"
                ts[pct_name] = 100.0 * ts[f"Proportion {col}"]
                added_cols.append(pct_name)
            if include_cum_percentage:
                # Ensure percentage exists (compute if needed)
                pct_col = f"Percentage {col}"
                if pct_col not in ts.columns:
                    ts[f"Proportion {col}"] = _safe_div(ts[col], total)
                    ts[pct_col] = 100.0 * ts[f"Proportion {col}"]
                    added_cols.extend([f"Proportion {col}", pct_col])
                cumpct_name = f"Cumulative Percentage {col}"
                ts[cumpct_name] = ts[pct_col].cumsum()
                added_cols.append(cumpct_name)

    # Ensure column order: original aggregated metrics first, then any derived ones
    ordered_cols = num_cols + [c for c in ts.columns if c not in num_cols]
    ts = ts.loc[:, ordered_cols]

    return ts

def compute_item_time_stats(
    df: pd.DataFrame,
    item_col: str,
    year_col: str = 'Year',
    cite_col: str = 'Cited by',
    *,
    list_mode: bool = True,
    list_separators: str = ';',
    strip_items: bool = True,
    items: Optional[Iterable[str]] = None,
    top_n: int = 10,
    min_docs_per_item: int = 1,
    include_overall_cols: bool = True,
) -> pd.DataFrame:
    """
    Build a tidy per-item, per-year table with docs, citations, and h-index.

    What it computes
    ----------------
    For every (Item, Year):
      - n_docs: number of documents that year
      - total_citations: sum of citations of those docs
      - citations_per_article: total_citations / n_docs
    Also, per item (attached to each row when include_overall_cols=True):
      - item_total_docs, item_total_citations, item_h_index, first_year, last_year

    Parameters
    ----------
    df : pd.DataFrame
        Input table with at least item_col, year_col, and (optionally) cite_col.
    item_col : str
        Column with item names; can be multi-valued like "A; B; C".
    year_col : str, default "Year"
        Publication year (numeric).
    cite_col : str, default "Cited by"
        Citation count per document; created with zeros if missing.
    list_mode : bool, default True
        Split and explode item_col using list_separators (regex).
    list_separators : str, default r";"
        Regex separators for multi-valued item_col.
    strip_items : bool, default True
        Strip whitespace around split tokens.
    items : Optional[Iterable[str]], default None
        Explicit items to retain; if None, take top_n by document count.
    top_n : int, default 10
        Number of top items by documents if items is None.
    min_docs_per_item : int, default 1
        Filter out items with fewer total docs.
    include_overall_cols : bool, default True
        Attach per-item descriptors to each (Item, Year) row.

    Returns
    -------
    pd.DataFrame
        Columns always include:
          "Item", "Year", "n_docs", "total_citations", "citations_per_article"
        and, when include_overall_cols=True, also:
          "item_total_docs", "item_total_citations", "item_h_index", "first_year", "last_year"
    """
    # Local fallback if an h_index util isn't already in scope
    def h_index(
        cites: Iterable[float],
    ) -> int:
        arr = sorted((int(c) for c in cites if pd.notna(c)), reverse=True)
        h = 0
        for i, c in enumerate(arr, start=1):
            if c >= i:
                h = i
            else:
                break
        return h

    if df.empty:
        cols = ["Item", "Year", "n_docs", "total_citations", "citations_per_article"]
        if include_overall_cols:
            cols += ["item_total_docs", "item_total_citations", "item_h_index", "first_year", "last_year"]
        return pd.DataFrame(columns=cols)

    work = df.copy()

    # Ensure valid numeric year and citations
    work = work[pd.to_numeric(work[year_col], errors="coerce").notna()].copy()
    work[year_col] = work[year_col].astype(int)

    if cite_col not in work.columns:
        work[cite_col] = 0
    work[cite_col] = pd.to_numeric(work[cite_col], errors="coerce").fillna(0).astype(int)

    # Split/explode multi-valued items if requested
    if list_mode:
        tokens = work[item_col].fillna("").astype(str).str.split(list_separators, regex=True)
        if strip_items:
            tokens = tokens.apply(lambda xs: [x.strip() for x in xs if x and x.strip() != ""])
        work = work.drop(columns=[item_col]).join(pd.DataFrame({item_col: tokens})).explode(item_col, ignore_index=True)
        work = work[work[item_col].notna() & (work[item_col] != "")]
    else:
        work = work[work[item_col].notna() & (work[item_col].astype(str).str.strip() != "")]
        if strip_items:
            work[item_col] = work[item_col].astype(str).str.strip()

    # Per-item overall descriptors
    overall = (
        work.groupby(item_col, dropna=False)
            .agg(
                item_total_docs=(year_col, "size"),
                item_total_citations=(cite_col, "sum"),
                first_year=(year_col, "min"),
                last_year=(year_col, "max"),
            )
            .reset_index()
            .rename(columns={item_col: "Item"})
    )

    cit_lists = (
        work.groupby(item_col)[cite_col]
            .apply(list)
            .reset_index()
            .rename(columns={item_col: "Item", cite_col: "_cit"})
    )
    cit_lists["_h"] = cit_lists["_cit"].apply(h_index)
    overall = overall.merge(cit_lists[["Item", "_h"]], on="Item", how="left").rename(columns={"_h": "item_h_index"})

    # Apply min-docs filter
    overall = overall.loc[overall["item_total_docs"] >= int(min_docs_per_item)]

    # Decide which items to keep
    if items is None:
        keep_items = set(
            overall.sort_values(
                ["item_total_docs", "item_total_citations"],
                ascending=[False, False]
            ).head(int(top_n))["Item"]
        )
    else:
        keep_items = set(items)

    work = work[work[item_col].isin(keep_items)].copy()
    work.rename(columns={item_col: "Item"}, inplace=True)

    # Per-year aggregates
    per_year = (
        work.groupby(["Item", year_col])
            .agg(n_docs=(year_col, "size"), total_citations=(cite_col, "sum"))
            .reset_index()
            .rename(columns={year_col: "Year"})
    )
    per_year["citations_per_article"] = per_year["total_citations"] / per_year["n_docs"]

    if include_overall_cols:
        per_year = per_year.merge(overall, on="Item", how="left")

    return per_year.sort_values(["Item", "Year"]).reset_index(drop=True)

"""Utilities for cleaning, normalizing and aggregating author and index keywords."""
# Keywords processing

def merge_keywords_columns(
    df: pd.DataFrame,
    author_col: str = 'Author Keywords',
    index_col: str = 'Index Keywords',
    sep = '; ',
) -> pd.Series:
    """
    Merges two keyword columns (default: "Author Keywords" and "Index Keywords") in a DataFrame.
    Removes duplicates and returns a new Series with combined keywords separated by sep, default "; ".

    Parameters:
        df (pd.DataFrame): The DataFrame containing the keyword columns.
        author_col (str): Name of the column with author keywords. Default is "Author Keywords".
        index_col (str): Name of the column with index keywords. Default is "Index Keywords".

    Returns:
        pd.Series: A new Series with merged and deduplicated keywords.
    """
    # Check which columns exist
    has_author = author_col in df.columns
    has_index = index_col in df.columns
    
    if not has_author and not has_index:
        # Return empty series if neither column exists
        return pd.Series([''] * len(df), index=df.index)
    
    def merge_keywords(row):
        ak = []
        ik = []
        if has_author and pd.notnull(row.get(author_col)):
            ak = str(row[author_col]).split(sep) if row[author_col] else []
        if has_index and pd.notnull(row.get(index_col)):
            ik = str(row[index_col]).split(sep) if row[index_col] else []
        merged = sorted(set([k.strip() for k in ak + ik if k.strip()]))
        return sep.join(merged)

    return df.apply(merge_keywords, axis=1)

def merge_text_columns(
    df: pd.DataFrame,
    title_col: str = 'Title',
    abstract_col: str = 'Abstract',
    author_col: str = 'Author Keywords',
    index_col: str = 'Index Keywords',
    combined_col: str = 'Combined Text',
    sep: str = ';',
) -> pd.DataFrame:
    """
    Builds a single text field by concatenating Title, Abstract, and merged keywords,
    removes punctuation and non-alphanumeric characters, and collapses repeated runs of the same character.

    Parameters:
        df (pd.DataFrame): Input DataFrame.
        title_col (str): Column name for titles. Default is "Title".
        abstract_col (str): Column name for abstracts. Default is "Abstract".
        author_col (str): Column name for author keywords. Default is "Author Keywords".
        index_col (str): Column name for index keywords. Default is "Index Keywords".
        combined_col (str): Name of the new combined-text column. Default is "Combined Text".

    Returns:
        pd.DataFrame: DataFrame with a new column `combined_col` containing the cleaned, merged text.
    """
    # 1) Merge keywords into a Series
    keywords_series = merge_keywords_columns(df, author_col=author_col, index_col=index_col, sep=sep)

    # Check which columns actually exist
    has_title = title_col and title_col in df.columns
    has_abstract = abstract_col and abstract_col in df.columns

    # 2) Helper to merge fields and clean text
    def merge_fields(
        row,
        keywords,
    ):
        title = ""
        abstract = ""
        if has_title:
            title = row[title_col] if pd.notnull(row[title_col]) else ""
        if has_abstract:
            abstract = row[abstract_col] if pd.notnull(row[abstract_col]) else ""
        kw = keywords[row.name]
        parts = [title.strip(), abstract.strip(), kw]
        raw = " ".join(part for part in parts if part)
        # remove any non-word, non-space characters
        cleaned = re.sub(r"[^\w\s]", "", raw)
        # collapse any character repeated more than twice to a single instance
        cleaned = re.sub(r"(\w)\1{2,}", r"\1", cleaned)
        return cleaned

    # 3) Apply and assign
    df[combined_col] = df.apply(lambda r: merge_fields(r, keywords_series), axis=1)
    return df

"""Small helpers for ensuring that required NLTK resources are downloaded and available."""
# Download necessary nltk data


def ensure_wordnet() -> None:
    """
    Ensure that the NLTK WordNet corpus is available.

    Downloads it quietly only if it is missing.
    """
    import nltk
    from nltk.data import find
    try:
        find("corpora/wordnet")
    except LookupError:
        nltk.download("wordnet", quiet=True)


import unicodedata

def preprocess_keywords(
    df: pd.DataFrame,
    column: str,
    exclude_list: list | pd.DataFrame | None = None,
    synonyms: pd.DataFrame | dict | None = None,
    lemmatize: bool = False,
    sep: str = '; ',
    *,
    normalize_compounds: str | None = 'space',
    fold_accents: bool = False,
    alt_separators: list[str] | None = None,
    sort_keywords: bool = True,
    strip_punctuation: bool = True,
    normalize_underscores: bool = True,
) -> pd.DataFrame:
    """
    Clean and harmonize a keywords column, with smart handling of hyphen/space variants.

    The function lowercases, optionally removes unwanted keywords, applies synonym mapping,
    lemmatizes, and **normalizes obvious hyphen/space variants** such as "policy-based" ⇄ "policy based".
    By default, compound terms are canonicalized to use a **space** (e.g., "policy based").
    You can change this with `normalize_compounds="hyphen"` or disable with `None`.

    Steps:
      1. Validate input column.
      2. Build canonicalization helper that:
         - unifies hyphen-like characters (‐-–—−) to a single space or hyphen per `normalize_compounds`,
         - optionally converts underscores to spaces,
         - optionally strips accents,
         - trims stray punctuation at the ends,
         - collapses internal whitespace.
      3. Canonicalize `exclude_list` and `synonyms` so "policy-based" matches "policy based".
      4. For each keyword phrase:
         - lowercase → canonicalize → synonym-replace → optional per-word lemmatize → exclude.
      5. De-duplicate (order-preserving) and optionally sort.
      6. Write result to a new column named "Processed {column}".

    Args:
        df: Input DataFrame.
        column: Name of the column containing keywords.
        exclude_list: List or DataFrame (first column) of keywords to remove.
        synonyms: Either
            - DataFrame with two columns [old, new], or
            - dict mapping {new_keyword: [old_keyword1, old_keyword2, ...]}.
          Matching is done after canonicalization and lowercase.
        lemmatize: If True, lemmatize each word within a (possibly multiword) keyword.
        sep: Separator used in the keyword column (default: "; ").
        normalize_compounds: {"space", "hyphen", None}. Default "space".
            - "space": "policy-based" → "policy based"
            - "hyphen": "policy based" → "policy-based"
            - None: leave compounds unchanged.
        fold_accents: If True, strip diacritics (e.g., "café" → "cafe").
        alt_separators: Additional separators to accept when splitting (e.g., [",", "|"]).
        sort_keywords: If True, sort unique keywords inside each cell; else keep first-seen order.
        strip_punctuation: If True, strip leading/trailing punctuation like quotes, commas, periods.
        normalize_underscores: If True, convert "_" runs to a single space.

    Returns:
        DataFrame with an added column "Processed {column}".

    Notes:
        - Default canonical form uses **spaces** for compounds: this maximizes readability and
          matches most bibliographic keyword conventions where keywords may be multiword phrases.
        - If NLTK's WordNet is not available, `lemmatize=True` is silently skipped.
    """
    # --- guardrails ------------------------------------------------------------
    if column not in df.columns:
        print(f"Column \"{column}\" not found in the DataFrame.")
        return df

    # Build lemmatizer if requested and available
    lemmatizer = None
    if lemmatize and WordNetLemmatizer is not None:
        lemmatizer = WordNetLemmatizer()

    # --- helpers ---------------------------------------------------------------
    hyphen_chars = r"\-\u2010\u2011\u2012\u2013\u2014\u2212"  # -, ‐, -, ‒, – , — , −
    end_punct = " .,:;/'\"()[]{}"
    splitters = [re.escape(sep)]
    if alt_separators:
        splitters.extend([re.escape(s) for s in alt_separators])
    split_pattern = re.compile(r"\s*(?:%s)\s*" % "|".join(splitters))

    def _fold_accents(
        s: str,
    ) -> str:
        if not fold_accents:
            return s
        nfkd = unicodedata.normalize("NFKD", s)
        return "".join(ch for ch in nfkd if not unicodedata.combining(ch))

    def _normalize_compounds(
        s: str,
    ) -> str:
        if normalize_compounds not in {"space", "hyphen", None}:
            raise ValueError("normalize_compounds must be one of {\"space\", \"hyphen\", None}")
        if normalize_compounds is None:
            return s
        repl = " " if normalize_compounds == "space" else "-"
        s = re.sub(f"[{hyphen_chars}]+", repl, s)
        if normalize_underscores:
            s = re.sub(r"_+", " " if normalize_compounds == "space" else "-", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _strip_punct(
        s: str,
    ) -> str:
        return s.strip(end_punct) if strip_punctuation else s

    def _canon(
        s: str,
    ) -> str:
        s = s.lower().strip()
        s = _fold_accents(s)
        s = _normalize_compounds(s)
        s = _strip_punct(s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _lemmatize_phrase(
        s: str,
    ) -> str:
        if not lemmatizer:
            return s
        words = s.split()
        return " ".join(lemmatizer.lemmatize(w) for w in words)

    # --- canonicalize synonyms/exclusions -------------------------------------
    synonym_dict: dict[str, str] = {}
    if isinstance(synonyms, pd.DataFrame) and not synonyms.empty:
        old_col = synonyms.iloc[:, 0].astype(str).map(_canon)
        new_col = synonyms.iloc[:, 1].astype(str).map(_canon)
        synonym_dict = dict(zip(old_col, new_col))
    elif isinstance(synonyms, dict):
        tmp = {}
        for new_term, old_list in synonyms.items():
            new_c = _canon(str(new_term))
            for old_term in old_list:
                tmp[_canon(str(old_term))] = new_c
        synonym_dict = tmp

    if isinstance(exclude_list, pd.DataFrame) and not exclude_list.empty:
        exclude_set = set(exclude_list.iloc[:, 0].astype(str).map(_canon))
    elif isinstance(exclude_list, list):
        exclude_set = { _canon(str(w)) for w in exclude_list }
    else:
        exclude_set = set()

    # --- main row processor ----------------------------------------------------
    def process_row(
        keyword_str,
    ):
        if pd.isna(keyword_str) or str(keyword_str).strip() == "":
            return ""

        # Split on primary + alternative separators
        raw_tokens = [t for t in split_pattern.split(str(keyword_str)) if t != ""]
        processed = []
        for token in raw_tokens:
            t = _canon(token)

            # synonym replacement on the whole phrase
            t = synonym_dict.get(t, t)

            # optional lemmatization (per word inside the phrase)
            if lemmatizer:
                t = _lemmatize_phrase(t)

            # exclusion
            if t and t not in exclude_set:
                processed.append(t)

        if not processed:
            return ""

        # deduplicate
        if sort_keywords:
            unique_tokens = sorted(set(processed))
        else:
            unique_tokens = list(dict.fromkeys(processed))
        return sep.join(unique_tokens)

    # --- write output ----------------------------------------------------------
    new_column = f"Processed {column}"
    df[new_column] = df[column].apply(process_row)
    return df

"""Text-processing routines focused on abstracts, including tokenization, cleaning and normalization."""
# abstract processing

stopwords_file = os.path.join(fd, "additional files", "stopwords.xlsx")

def _ensure_nltk():
    """Ensure required NLTK resources are available without redundant downloads."""
    import nltk
    resources = [
        ("punkt", "tokenizers/punkt"),
        ("wordnet", "corpora/wordnet"),
        ("stopwords", "corpora/stopwords"),
    ]
    for pkg, res in resources:
        try:
            nltk.data.find(res)
        except LookupError:
            nltk.download(pkg, quiet=True)

def process_text_column(
    df: pd.DataFrame,
    column_name: str,
    stopwords_file: Optional[str] = None,
    lang: str = "english",
    remove_numbers: bool = True,
    remove_two_letter_words: bool = True,
    extra_stopwords: Optional[Iterable[str]] = None,
    exclude_specific_stopwords: Optional[Iterable[str]] = None,
) -> pd.DataFrame:
    """
    Process a text column by tokenizing, lowercasing, lemmatizing, and removing stopwords.

    All stopwords—whether loaded from file, taken from the "specific" sheet by
    category, or passed via `extra_stopwords`—are normalized (lowercased +
    lemmatized) so they match inflected/cased variants in text.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the text column.
    column_name : str
        Name of the column to process.
    stopwords_file : str, optional
        Path to an Excel file with stopwords. Expected layout:

        - Sheet "general": base stopwords (possibly multilingual; if a column
          named `lang` exists it is used, otherwise the first column).
        - Sheet "specific": optional project-specific stopwords. It is treated
          as a table with at least two columns:
              1. category
              2. word

          Only rows whose category is listed in `exclude_specific_stopwords`
          are used as additional stopwords.

        If "general" is missing, the first sheet is used as base stopwords.
        If `stopwords_file` has a single sheet (legacy), that sheet is used as
        base stopwords and the "specific" logic is skipped.
    lang : str, default "english"
        Language key for stopwords (used for both NLTK and Excel column
        selection in the "general" sheet).
    remove_numbers : bool, default True
        If True, drops tokens that are purely numeric or contain any digit.
    remove_two_letter_words : bool, default True
        If True, drops alphabetic tokens of exactly two letters.
    extra_stopwords : iterable of str, optional
        Additional stopwords to merge in-memory. They will be lowercased and
        lemmatized and are always used, independently of the "specific" sheet.
    exclude_specific_stopwords : iterable of str, optional
        Names of categories from the "specific" sheet that should be treated
        as additional stopwords. Category matching is done case-insensitively
        after stripping.

        Example category names (no slashes or quotes):

        - "Scholarly boilerplate"
        - "Methods and analysis"
        - "Publication document type terms"
        - "Scope and qualifiers"
        - "Bibliometrics specific terms"
        - "Generic science vocabulary"
        - "Section labels and boilerplate"

        If None or empty, the "specific" sheet is ignored.

    Returns
    -------
    pd.DataFrame
        The DataFrame with a new column named "Processed {column_name}".
    """
    _ensure_nltk()
    lemmatizer = WordNetLemmatizer()

    def _extract_sw_from_df(sw_df: pd.DataFrame) -> list[str]:
        if lang in sw_df.columns:
            col = sw_df[lang]
        else:
            col = sw_df[sw_df.columns[0]]
        return (
            col.dropna()
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )

    # ---- load base stopwords from "general" (and optionally "specific")
    base_sw: list[str] = []

    if stopwords_file:
        sheets = pd.read_excel(stopwords_file, sheet_name=None)

        if isinstance(sheets, dict):
            # General sheet
            if "general" in sheets:
                general_df = sheets["general"]
            else:
                # Fallback: first sheet in the workbook
                general_df = next(iter(sheets.values()))
            base_sw = _extract_sw_from_df(general_df)

            # Optional "specific" sheet with (category, word) table
            categories_to_use = {
                c.strip().lower()
                for c in (exclude_specific_stopwords or [])
                if str(c).strip()
            }
            if categories_to_use and "specific" in sheets:
                specific_df = sheets["specific"]
                if specific_df.shape[1] >= 2:
                    cat_col = specific_df.columns[0]
                    word_col = specific_df.columns[1]
                    tmp = specific_df[[cat_col, word_col]].dropna()

                    tmp[cat_col] = tmp[cat_col].astype(str).str.strip()
                    tmp[word_col] = (
                        tmp[word_col]
                        .astype(str)
                        .str.strip()
                        .str.lower()
                    )

                    mask = tmp[cat_col].str.strip().str.lower().isin(categories_to_use)
                    specific_sw = tmp.loc[mask, word_col].tolist()
                    base_sw.extend(specific_sw)
        else:
            # Legacy: single-sheet workbook
            base_sw = _extract_sw_from_df(sheets)
    else:
        try:
            base_sw = [w.lower() for w in stopwords.words(lang)]
        except OSError:
            base_sw = [w.lower() for w in stopwords.words("english")]

    # ---- merge extra stopwords (logic unchanged)
    extra_sw: list[str] = []
    if extra_stopwords:
        extra_sw = [str(w).strip().lower() for w in extra_stopwords if str(w).strip()]

    # ---- normalize all stopwords: lowercase + lemma (to catch plurals/variants)
    def _normalize_stopword(w: str) -> set[str]:
        lw = w.lower().strip()
        return {lw, lemmatizer.lemmatize(lw)}

    stopwords_set: set[str] = set()
    for w in base_sw:
        stopwords_set |= _normalize_stopword(w)
    for w in extra_sw:
        stopwords_set |= _normalize_stopword(w)

    def _drop_for_numbers(token: str) -> bool:
        """Return True if token should be dropped due to number rules."""
        if not remove_numbers:
            return False
        return token.isdigit() or any(ch.isdigit() for ch in token)

    def clean_text(text):
        """Tokenize, lowercase, lemmatize, stopword-filter, and return space-joined string."""
        if pd.isna(text):
            return None

        tokens = word_tokenize(str(text))
        processed: list[str] = []

        for tok in tokens:
            t = tok.lower().strip()
            if not t:
                continue

            lemma = lemmatizer.lemmatize(t)

            # stopword check uses lemma+lowercase set
            if lemma in stopwords_set or t in stopwords_set:
                continue

            if _drop_for_numbers(t):
                continue

            if remove_two_letter_words and lemma.isalpha() and len(lemma) == 2:
                continue

            # keep word-like tokens
            if lemma.isalnum() or lemma.isalpha():
                processed.append(lemma)

        return " ".join(processed) if processed else None

    df[f"Processed {column_name}"] = df[column_name].apply(clean_text)
    return df



"""Helpers for combining topic-related text fields (abstracts, titles, keywords) into unified representations."""
# Combining topical variables: abstract, title and keywords

def build_combined_text(
    df: pd.DataFrame,
    include_index_keywords: bool = False,
    base_name: str = 'Combined Text',
) -> pd.DataFrame:
    """
    Create and append a merged text column to the dataframe from Title, Abstract,
    Author Keywords, and (optionally) Index Keywords.

    Rules
    -----
    - Prefer processed columns when available:
        Title: ["Processed Title", "Title"]
        Abstract: ["Processed Abstract", "Abstract"]
        Author Keywords: ["Processed Author Keywords", "Author Keywords"]
        Index Keywords (optional): ["Processed Index Keywords", "Index Keywords"]
    - Output column name:
        * Base name = "Combined Text".
        * If Title OR Abstract used a processed variant, prefix with "Processed ".
          → "Processed Combined Text".
    - Missing columns are silently skipped; if none are present, an empty column is created.
    - Values are joined with a single space; repeated whitespace is collapsed and trimmed.

    Parameters
    ----------
    df : pd.DataFrame
        Input data.
    include_index_keywords : bool, default False
        Include Index Keywords if present.
    base_name : str, default "Combined Text"
        Base name; "Processed " is auto-prefixed when processed Title/Abstract is used.

    Returns
    -------
    pd.DataFrame
        The dataframe with the new combined text column appended.
    """

    title_col = first_existing(df, ["Processed Title", "Title"])
    abstr_col = first_existing(df, ["Processed Abstract", "Abstract"])
    auth_kw_col = first_existing(df, ["Processed Author Keywords", "Author Keywords"])
    index_col = None
    if include_index_keywords:
        index_col = first_existing(df, ["Processed Index Keywords", "Index Keywords"])

    processed_used = (title_col == "Processed Title") or (abstr_col == "Processed Abstract")
    out_col = f'{"Processed " if processed_used else ""}{base_name}'

    cols_to_use = [c for c in [title_col, abstr_col, auth_kw_col, index_col] if c is not None]

    if cols_to_use:
        df[out_col] = (
            df[cols_to_use]
            .fillna("")
            .astype(str)
            .agg(" ".join, axis=1)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )
    else:
        df[out_col] = ""

    return df

def preprocess_keywords_and_text(
    df,
    db="scopus",
    exclude_list_kw=None,
    synonyms_kw=None,
    lemmatize_kw=False,
    combine_with_index_keywords=False,
    stopwords_file=None,
    lang_of_docs="en",
    extra_stopwords=None,
    specific_stopword_categories=None,
    remove_numbers=True,
    remove_two_letter_words=True,
    normalize_compounds='space',
    fold_accents=False,
    alt_separators=None,
    sort_keywords=True,
    strip_punctuation=True,
    normalize_underscores=True,
):
    """
    Process keywords and text fields in a bibliographic dataframe.
    
    This function extends the input dataframe with processed keyword columns,
    processed text columns (title/abstract), combined text, and processed
    combined text.
    
    Parameters
    ----------
    df : pandas.DataFrame
        Input bibliographic dataframe.
    db : str, default "scopus"
        Database identifier (e.g., "scopus", "oa", "open alex").
    exclude_list_kw : list of str or None, optional
        Keywords to exclude during processing.
    synonyms_kw : dict or None, optional
        Mapping of synonyms for keyword normalization.
    lemmatize_kw : bool, default False
        Whether to lemmatize keywords.
    combine_with_index_keywords : bool, default False
        Whether to include index keywords when building combined text.
    stopwords_file : str or path-like, optional
        Path to stopwords Excel file. If None, uses default stopwords_file.
    lang_of_docs : str, default "en"
        Language of documents for stopword selection.
    extra_stopwords : iterable of str or None, optional
        Additional custom stopwords to apply.
    specific_stopword_categories : iterable of str or None, optional
        Specific stopword categories from the stopwords file to apply.
    remove_numbers : bool, default True
        If True, numeric tokens are removed from text.
    remove_two_letter_words : bool, default True
        If True, tokens of length two are removed from text.
    normalize_compounds : str, default 'space'
        How to normalize compound keywords.
    fold_accents : bool, default False
        Whether to remove accents from keywords.
    alt_separators : list or None, optional
        Alternative separators for keyword splitting.
    sort_keywords : bool, default True
        Whether to sort keywords alphabetically.
    strip_punctuation : bool, default True
        Whether to strip punctuation from keywords.
    normalize_underscores : bool, default True
        Whether to normalize underscores in keywords.
    
    Returns
    -------
    pandas.DataFrame
        Extended dataframe with additional processed columns:
        - Processed Author Keywords
        - Processed Index Keywords
        - Processed Author and Index Keywords (if applicable)
        - Combined Text
        - Processed Title
        - Processed Abstract
    
    Notes
    -----
    The function modifies a copy of the input dataframe and returns it.
    Original dataframe remains unchanged.
    """
    df = df.copy()
    
    # Determine default separator based on database
    default_separator = {"scopus": "; ", "open alex": "|", "oa": "|"}.get(
        db.lower(), "; "
    )
    
    # Use default stopwords file if not provided
    if stopwords_file is None:
        stopwords_file = stopwords_file  # uses module-level default
    
    # Add database-specific stopwords
    extra_sw = list(extra_stopwords) if extra_stopwords is not None else []
    if db.lower() == "scopus":
        extra_sw += ["elsevier", "reserved"]
    
    # Process Author Keywords if available
    if "Author Keywords" in df.columns:
        df = preprocess_keywords(
            df,
            column="Author Keywords",
            exclude_list=exclude_list_kw,
            synonyms=synonyms_kw,
            lemmatize=lemmatize_kw,
            sep=default_separator,
            normalize_compounds=normalize_compounds,
            fold_accents=fold_accents,
            alt_separators=alt_separators,
            sort_keywords=sort_keywords,
            strip_punctuation=strip_punctuation,
            normalize_underscores=normalize_underscores,
        )
    
    # Process Index Keywords if available
    if "Index Keywords" in df.columns:
        df = preprocess_keywords(
            df,
            column="Index Keywords",
            exclude_list=exclude_list_kw,
            synonyms=synonyms_kw,
            lemmatize=lemmatize_kw,
            sep=default_separator,
            normalize_compounds=normalize_compounds,
            fold_accents=fold_accents,
            alt_separators=alt_separators,
            sort_keywords=sort_keywords,
            strip_punctuation=strip_punctuation,
            normalize_underscores=normalize_underscores,
        )
    
    # Process Author and Index Keywords if available
    if "Author and Index Keywords" in df.columns:
        df = preprocess_keywords(
            df,
            column="Author and Index Keywords",
            exclude_list=exclude_list_kw,
            synonyms=synonyms_kw,
            lemmatize=lemmatize_kw,
            sep=default_separator,
            normalize_compounds=normalize_compounds,
            fold_accents=fold_accents,
            alt_separators=alt_separators,
            sort_keywords=sort_keywords,
            strip_punctuation=strip_punctuation,
            normalize_underscores=normalize_underscores,
        )
    
    # Process Abstract if available
    if "Abstract" in df.columns:
        df = process_text_column(
            df,
            "Abstract",
            stopwords_file=stopwords_file,
            lang=lang_of_docs,
            remove_numbers=remove_numbers,
            remove_two_letter_words=remove_two_letter_words,
            extra_stopwords=extra_sw,
            exclude_specific_stopwords=specific_stopword_categories,
        )
    
    # Process Title if available
    if "Title" in df.columns:
        df = process_text_column(
            df,
            "Title",
            stopwords_file=stopwords_file,
            lang=lang_of_docs,
            remove_numbers=remove_numbers,
            remove_two_letter_words=remove_two_letter_words,
            extra_stopwords=extra_sw,
            exclude_specific_stopwords=specific_stopword_categories,
        )
    
    # Build combined text
    df = build_combined_text(
        df,
        include_index_keywords=combine_with_index_keywords,
    )
    
    return df


"""Utilities for defining and managing high-level 'concepts' used in Biblium analyses."""
# concept

def add_concept_indicators(
    df: pd.DataFrame,
    concept_df: pd.DataFrame,
    text_col: str,
) -> pd.DataFrame:
    """
    Add binary indicator columns to `df` for each concept defined in `concept_df`,
    supporting:
      1) Token-level wildcard at the END of a token with `*` (e.g., "gov*" matches "government", "governance").
      2) Multi-word phrases that must appear in order (e.g., "collaborative governance").
      3) Combination of both (e.g., "good gov*" matches "good government"/"good governance").

    Notes
    -----
    - Wildcard is ONLY recognized at the END of a token. A token is any sequence split by whitespace.
    - Each token is matched case-insensitively with a word boundary at the token start.
      For wildcard tokens, the trailing part may continue within the same word.
      For non-wildcard tokens, a trailing word boundary is enforced.
    - Tokens in a phrase must appear consecutively with one or more whitespace characters between them.
    - Empty/NaN entries in `concept_df` are ignored.

    Parameters
    ----------
    df : pd.DataFrame
        Main dataframe containing a column with text.
    concept_df : pd.DataFrame
        DataFrame where each column represents a concept. Each cell contains a word or phrase
        describing that concept (e.g., "governance", "collaborative governance", "gov*").
    text_col : str
        Name of the column in `df` that contains the text to search.

    Returns
    -------
    pd.DataFrame
        A copy of `df` extended with one binary (0/1) column per concept (same column names as in `concept_df`).
    """
    def token_pattern(
        token: str,
    ) -> str:
        """Build a regex for a single token, honoring trailing '*' wildcard."""
        token = token.strip()
        if not token:
            return ""
        if token.endswith("*"):
            stem = re.escape(token[:-1])
            # Start-of-word boundary, then any word continuation (letters/digits/_)
            return rf"\b{stem}\w*"
        # Exact word with boundaries
        return rf"\b{re.escape(token)}\b"

    def phrase_to_regex(
        phrase: str,
    ) -> str:
        """
        Convert a phrase like 'good gov*' into a regex that matches tokens in order with whitespace between.
        """
        # Split on whitespace, build per-token patterns, and join with \s+ to enforce adjacency in text.
        tokens = [t for t in phrase.strip().split() if t]
        if not tokens:
            return ""
        parts = [token_pattern(t) for t in tokens]
        return r"\s+".join(parts)

    # Work on a copy; normalize text column to string
    out = df.copy()
    out[text_col] = out[text_col].fillna("").astype(str)

    for concept in concept_df.columns:
        # Collect all phrases/words for this concept
        phrases: List[str] = (
            concept_df[concept]
            .dropna()
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )
        # Skip empty concepts
        phrases = [p for p in phrases if p]
        if not phrases:
            continue

        # Build combined regex: any of the phrase patterns
        pattern_list = [phrase_to_regex(p) for p in phrases]
        # Remove any empty patterns after processing
        pattern_list = [p for p in pattern_list if p]
        if not pattern_list:
            continue

        combined = "(?:" + "|".join(pattern_list) + ")"
        regex = re.compile(combined, flags=re.IGNORECASE)

        # Vectorized contains → 0/1
        out[concept] = out[text_col].str.contains(regex, na=False).astype(int)

    return out

"""Wrappers and helpers for topic modelling (LDA, NMF, LSA) on bibliographic text fields."""
# topic modelling

def determine_optimal_topics(
    texts,
    max_topics = 10,
    language = 'english',
):
    """Determine the optimal number of topics using perplexity scores.

    Args:
        texts (list of str): Preprocessed text documents.
        max_topics (int): Maximum number of topics to evaluate.
        language (str): Language stop words to use (default: English).

    Returns:
        int: Optimal number of topics.
    """
    vectorizer = CountVectorizer(stop_words=language)
    doc_term_matrix = vectorizer.fit_transform(texts)
    perplexities = []
    topic_range = range(2, max_topics + 1)
    for n_topics in topic_range:
        lda = LatentDirichletAllocation(n_components=n_topics, random_state=42)
        lda.fit(doc_term_matrix)
        perplexities.append(lda.perplexity(doc_term_matrix))
    optimal_topics = topic_range[np.argmin(perplexities)]
    return optimal_topics

def topic_modeling(
    df: pd.DataFrame,
    text_column: str,
    model_type: str = 'LDA',
    n_topics: Optional[int] = None,
    max_topics: int = 10,
    max_features: int = 5000,
    stop_words: str | list[str] | None = 'english',
    *,
    include_doc_topic_weights: bool = True,
    topic_prefix: str = 'Topic',
    weight_suffix: str = 'Weight',
    normalize: Literal['row_sum', 'softmax', 'none'] = 'row_sum',
    lsa_nonneg: Literal['abs', 'square', 'none'] = 'abs',
    empty_weight_as_nan: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Perform topic modeling (LDA, NMF, or LSA) with robust handling of missing/empty text,
    and optionally append per-document topic weights as N extra columns.

    Behavior
    --------
    - Keeps dataframe length; empty texts get Topic=<NA>.
    - Adds columns: f"{topic_prefix} {i} {weight_suffix}" for i=1..n_topics when
      include_doc_topic_weights=True. Values are:
        * LDA: model.transform() probabilities (row-normalized by default).
        * NMF: nonnegative activations, row-normalized by default.
        * LSA: document-topic projections, transformed via `lsa_nonneg` (default |·|)
          then row-normalized by default.
    - If a row is empty, its weight columns are NaN (or 0.0 if empty_weight_as_nan=False).

    Returns
    -------
    (df_out, topics_df)
        df_out   : original df + "Topic" (+ N weight columns if requested).
        topics_df: long table with columns ["Topic", "Term", "Weight"] (top-10 terms/topic).
    """
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found in df.")

    texts = df[text_column].fillna("").astype(str).str.strip()
    mask = texts != ""
    df_out = df.copy()

    # Early return if nothing to model
    if mask.sum() == 0:
        df_out["Topic"] = pd.Series([pd.NA] * len(df_out), index=df_out.index, dtype="Float64")
        return df_out, pd.DataFrame(columns=["Topic", "Term", "Weight"])

    # Vectorizer by model
    mt = model_type.upper()
    if mt in {"LDA", "LSA"}:
        vectorizer = CountVectorizer(max_features=max_features, stop_words=stop_words)
    elif mt == "NMF":
        vectorizer = TfidfVectorizer(max_features=max_features, stop_words=stop_words)
    else:
        raise ValueError('model_type must be one of: "LDA", "NMF", "LSA".')

    # Vectorize valid rows
    try:
        X = vectorizer.fit_transform(texts[mask])
    except ValueError:
        df_out["Topic"] = pd.Series([pd.NA] * len(df_out), index=df_out.index, dtype="Float64")
        return df_out, pd.DataFrame(columns=["Topic", "Term", "Weight"])

    n_samples, n_features = X.shape
    if n_features == 0:
        df_out["Topic"] = pd.Series([pd.NA] * len(df_out), index=df_out.index, dtype="Float64")
        return df_out, pd.DataFrame(columns=["Topic", "Term", "Weight"])

    # Choose n_topics safely
    if n_topics is None:
        if mt == "LSA":
            n_topics = max(1, min(max_topics, max(1, n_features - 1)))
        else:
            n_topics = max(1, min(max_topics, n_features, n_samples))
    n_topics = int(n_topics)

    # Fit model
    if mt == "LDA":
        model = LDA(n_components=n_topics, random_state=42)
    elif mt == "NMF":
        model = NMF(n_components=n_topics, init="nndsvda", random_state=42, max_iter=400)
    else:  # LSA
        if n_features < 2:
            df_out["Topic"] = pd.Series([pd.NA] * len(df_out), index=df_out.index, dtype="Float64")
            return df_out, pd.DataFrame(columns=["Topic", "Term", "Weight"])
        model = TruncatedSVD(n_components=n_topics, random_state=42)

    model.fit(X)

    # Topic -> top terms
    feature_names = vectorizer.get_feature_names_out()
    topic_rows: list[dict] = []
    for i, comp in enumerate(model.components_):
        top_idx = np.argsort(comp)[-10:][::-1]
        for idx in top_idx:
            topic_rows.append({"Topic": f"Topic {i+1}", "Term": feature_names[idx], "Weight": float(comp[idx])})
    topics_df = pd.DataFrame(topic_rows, columns=["Topic", "Term", "Weight"])

    # Doc -> topic weights (gamma)
    gamma = model.transform(X)  # shape: (n_valid_docs, n_topics)

    # Make LSA nonnegative if requested
    if mt == "LSA":
        if lsa_nonneg == "abs":
            gamma = np.abs(gamma)
        elif lsa_nonneg == "square":
            gamma = np.square(gamma)
        # "none": leave as-is (can be negative)

    # Normalize rows if requested
    if normalize in {"row_sum", "softmax"}:
        if normalize == "softmax":
            gamma = np.exp(gamma - gamma.max(axis=1, keepdims=True))
        row_sums = gamma.sum(axis=1, keepdims=True)
        # Avoid div-by-zero
        gamma = np.divide(gamma, row_sums, out=np.zeros_like(gamma), where=row_sums > 0)

    # Hard assignment
    assignments = gamma.argmax(axis=1) + 1
    topic_series = pd.Series(np.nan, index=df_out.index, dtype="Float64")
    topic_series.loc[texts[mask].index] = assignments
    df_out["Topic"] = topic_series

    # Optional: append per-doc topic weight columns
    if include_doc_topic_weights:
        fill_val = np.nan if empty_weight_as_nan else 0.0
        W = np.full((len(df_out), n_topics), fill_val, dtype=float)
        W[mask.to_numpy(), :] = gamma  # align by boolean mask order

        # Create columns
        for k in range(n_topics):
            colname = f"{topic_prefix} {k+1} {weight_suffix}".strip()
            df_out[colname] = W[:, k]

    return df_out, topics_df


def topic_modeling_extended(
    df: pd.DataFrame,
    text_column: str,
    model_type: str = 'LDA',
    n_topics: Optional[int] = None,
    max_topics: int = 15,
    max_features: int = 5000,
    stop_words: str | list[str] | None = 'english',
    ngram_range: Tuple[int, int] = (1, 2),
    min_df: int = 2,
    max_df: float = 0.95,
    auto_select_topics: bool = True,
    coherence_method: str = 'umass',
) -> Dict[str, Any]:
    """
    Extended topic modeling with coherence scores and additional metrics.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe.
    text_column : str
        Column containing text to model.
    model_type : str, default="LDA"
        Model type: "LDA", "NMF", or "LSA".
    n_topics : int or None
        Number of topics. If None and auto_select_topics=True, auto-selects.
    max_topics : int, default=15
        Maximum topics to try for auto-selection.
    max_features : int, default=5000
        Maximum vocabulary size.
    stop_words : str or list, default="english"
        Stop words to remove.
    ngram_range : tuple, default=(1, 2)
        N-gram range for vectorization.
    min_df : int, default=2
        Minimum document frequency for terms.
    max_df : float, default=0.95
        Maximum document frequency for terms.
    auto_select_topics : bool, default=True
        Whether to auto-select optimal n_topics if n_topics is None.
    coherence_method : str, default="umass"
        Coherence score method: "umass" or "c_v".
    
    Returns
    -------
    dict with keys:
        - df_out: DataFrame with Topic assignments and weights
        - topics_df: Topic-term weights
        - model: Fitted model
        - vectorizer: Fitted vectorizer
        - feature_names: Vocabulary
        - doc_topic_matrix: Document-topic distribution
        - topic_term_matrix: Topic-term distribution
        - coherence_scores: Dict of coherence scores by n_topics (if auto-selected)
        - topic_coherence: Per-topic coherence scores
        - topic_stats: DataFrame with topic statistics
        - optimal_n_topics: Selected number of topics
    """
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.decomposition import LatentDirichletAllocation, NMF, TruncatedSVD
    
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found.")
    
    texts = df[text_column].fillna("").astype(str).str.strip()
    mask = texts != ""
    
    if mask.sum() == 0:
        return {
            "df_out": df.copy(),
            "topics_df": pd.DataFrame(columns=["Topic", "Term", "Weight"]),
            "model": None,
            "vectorizer": None,
            "feature_names": [],
            "doc_topic_matrix": None,
            "topic_term_matrix": None,
            "coherence_scores": {},
            "topic_coherence": {},
            "topic_stats": pd.DataFrame(),
            "optimal_n_topics": 0,
        }
    
    # Setup vectorizer
    mt = model_type.upper()
    if mt in {"LDA", "LSA"}:
        vectorizer = CountVectorizer(
            max_features=max_features, 
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    else:
        vectorizer = TfidfVectorizer(
            max_features=max_features, 
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    
    try:
        X = vectorizer.fit_transform(texts[mask])
    except ValueError:
        return {
            "df_out": df.copy(),
            "topics_df": pd.DataFrame(columns=["Topic", "Term", "Weight"]),
            "model": None,
            "vectorizer": vectorizer,
            "feature_names": [],
            "doc_topic_matrix": None,
            "topic_term_matrix": None,
            "coherence_scores": {},
            "topic_coherence": {},
            "topic_stats": pd.DataFrame(),
            "optimal_n_topics": 0,
        }
    
    feature_names = vectorizer.get_feature_names_out()
    n_samples, n_features = X.shape
    
    # Auto-select optimal number of topics
    coherence_scores = {}
    if n_topics is None and auto_select_topics:
        max_k = min(max_topics, max(2, n_features - 1), n_samples)
        
        for k in range(2, max_k + 1):
            if mt == "LDA":
                model_k = LatentDirichletAllocation(n_components=k, random_state=42, max_iter=10)
            elif mt == "NMF":
                model_k = NMF(n_components=k, init="nndsvda", random_state=42, max_iter=100)
            else:
                model_k = TruncatedSVD(n_components=k, random_state=42)
            
            model_k.fit(X)
            
            # Compute coherence (UMass approximation)
            coherence = _compute_topic_coherence(model_k.components_, X, feature_names)
            coherence_scores[k] = np.mean(coherence)
        
        # Select k with highest coherence
        if coherence_scores:
            n_topics = max(coherence_scores, key=coherence_scores.get)
        else:
            n_topics = min(5, max_k)
    elif n_topics is None:
        n_topics = min(5, n_features - 1, n_samples)
    
    n_topics = int(n_topics)
    
    # Fit final model
    if mt == "LDA":
        model = LatentDirichletAllocation(n_components=n_topics, random_state=42, max_iter=50)
    elif mt == "NMF":
        model = NMF(n_components=n_topics, init="nndsvda", random_state=42, max_iter=400)
    else:
        model = TruncatedSVD(n_components=n_topics, random_state=42)
    
    model.fit(X)
    
    # Get matrices
    doc_topic_matrix = model.transform(X)
    topic_term_matrix = model.components_
    
    # Make LSA non-negative
    if mt == "LSA":
        doc_topic_matrix = np.abs(doc_topic_matrix)
        topic_term_matrix = np.abs(topic_term_matrix)
    
    # Normalize document-topic to probabilities
    row_sums = doc_topic_matrix.sum(axis=1, keepdims=True)
    doc_topic_matrix_norm = np.divide(
        doc_topic_matrix, row_sums, 
        out=np.zeros_like(doc_topic_matrix), 
        where=row_sums > 0
    )
    
    # Build topics_df
    topic_rows = []
    for i in range(n_topics):
        top_idx = np.argsort(topic_term_matrix[i])[-20:][::-1]
        for idx in top_idx:
            topic_rows.append({
                "Topic": f"Topic {i+1}",
                "Topic_ID": i + 1,
                "Term": feature_names[idx],
                "Weight": float(topic_term_matrix[i, idx])
            })
    topics_df = pd.DataFrame(topic_rows)
    
    # Build df_out
    df_out = df.copy()
    
    # Hard assignment
    assignments = doc_topic_matrix_norm.argmax(axis=1) + 1
    topic_series = pd.Series(np.nan, index=df_out.index, dtype="Float64")
    topic_series.loc[texts[mask].index] = assignments
    df_out["Topic"] = topic_series
    
    # Add per-topic weights
    W = np.full((len(df_out), n_topics), np.nan, dtype=float)
    W[mask.to_numpy(), :] = doc_topic_matrix_norm
    for k in range(n_topics):
        df_out[f"Topic {k+1} Weight"] = W[:, k]
    
    # Compute per-topic coherence
    topic_coherence = _compute_topic_coherence(topic_term_matrix, X, feature_names)
    
    # Compute topic statistics (pass df_out which has Topic assignments)
    topic_stats = _compute_topic_statistics(
        doc_topic_matrix_norm, 
        topic_term_matrix, 
        feature_names,
        n_topics,
        df=df_out,  # Pass original df for citation/year stats
    )
    
    return {
        "df_out": df_out,
        "topics_df": topics_df,
        "model": model,
        "vectorizer": vectorizer,
        "feature_names": feature_names,
        "doc_topic_matrix": doc_topic_matrix_norm,
        "topic_term_matrix": topic_term_matrix,
        "coherence_scores": coherence_scores,
        "topic_coherence": {f"Topic {i+1}": topic_coherence[i] for i in range(n_topics)},
        "topic_stats": topic_stats,
        "optimal_n_topics": n_topics,
    }


def _compute_topic_coherence(
    topic_term_matrix: np.ndarray,
    doc_term_matrix,
    feature_names: np.ndarray,
    top_n: int = 10,
) -> List[float]:
    """
    Compute UMass coherence for each topic.
    
    UMass coherence measures how often top words co-occur in documents.
    Higher (less negative) values indicate more coherent topics.
    """
    from scipy import sparse
    
    n_topics = topic_term_matrix.shape[0]
    coherence = []
    
    # Convert to dense if needed for co-occurrence computation
    if sparse.issparse(doc_term_matrix):
        dtm = doc_term_matrix.toarray()
    else:
        dtm = doc_term_matrix
    
    # Binary document-term matrix
    dtm_binary = (dtm > 0).astype(float)
    
    # Word co-occurrence counts
    word_doc_counts = dtm_binary.sum(axis=0)
    cooc_matrix = dtm_binary.T @ dtm_binary
    
    for topic_idx in range(n_topics):
        top_word_idx = np.argsort(topic_term_matrix[topic_idx])[-top_n:][::-1]
        
        score = 0.0
        count = 0
        for i, w1 in enumerate(top_word_idx[1:], 1):
            for w2 in top_word_idx[:i]:
                d_w1 = word_doc_counts[0, w1] if hasattr(word_doc_counts, 'shape') and len(word_doc_counts.shape) > 1 else word_doc_counts[w1]
                d_w1_w2 = cooc_matrix[w1, w2]
                
                if d_w1 > 0:
                    score += np.log((d_w1_w2 + 1) / d_w1)
                    count += 1
        
        coherence.append(score / count if count > 0 else 0.0)
    
    return coherence


def _compute_topic_statistics(
    doc_topic_matrix: np.ndarray,
    topic_term_matrix: np.ndarray,
    feature_names: np.ndarray,
    n_topics: int,
    df: Optional[pd.DataFrame] = None,
    citation_col: Optional[str] = None,
    year_col: Optional[str] = None,
) -> pd.DataFrame:
    """
    Compute statistics for each topic including citation and year metrics.
    
    Parameters
    ----------
    doc_topic_matrix : np.ndarray
        Document-topic distribution matrix.
    topic_term_matrix : np.ndarray
        Topic-term distribution matrix.
    feature_names : np.ndarray
        Vocabulary.
    n_topics : int
        Number of topics.
    df : pd.DataFrame, optional
        Original dataframe with Topic, citation and year columns.
    citation_col : str, optional
        Name of citation column.
    year_col : str, optional
        Name of year column.
    
    Returns
    -------
    pd.DataFrame
        Topic statistics.
    """
    stats = []
    
    # Get topic assignments from doc_topic_matrix
    topic_assignments = doc_topic_matrix.argmax(axis=1)
    
    # Try to find citation and year columns if df provided
    citations = None
    years = None
    df_topics = None
    
    if df is not None and "Topic" in df.columns:
        # Use df's Topic column which is aligned with the full dataframe
        df_topics = df["Topic"].values
        
        # Find citation column
        if citation_col and citation_col in df.columns:
            citations = pd.to_numeric(df[citation_col], errors='coerce').values
        else:
            for col in ['Cited by', 'Times Cited', 'Citation Count', 'Citations', 'TC', 'Cited By']:
                if col in df.columns:
                    citations = pd.to_numeric(df[col], errors='coerce').values
                    break
        
        # Find year column
        if year_col and year_col in df.columns:
            years = pd.to_numeric(df[year_col], errors='coerce').values
        else:
            for col in ['Year', 'Publication Year', 'PY']:
                if col in df.columns:
                    years = pd.to_numeric(df[col], errors='coerce').values
                    break
    
    for i in range(n_topics):
        # Document statistics from doc_topic_matrix
        topic_weights = doc_topic_matrix[:, i]
        dominant_docs_matrix = (topic_assignments == i).sum()
        
        # Term statistics
        top_term_idx = np.argsort(topic_term_matrix[i])[-5:][::-1]
        top_terms = [feature_names[idx] for idx in top_term_idx]
        
        stat_dict = {
            "Topic": f"Topic {i+1}",
            "Topic_ID": i + 1,
            "Dominant Documents": int(dominant_docs_matrix),
            "Document Share (%)": round(float(dominant_docs_matrix / len(doc_topic_matrix) * 100), 2) if len(doc_topic_matrix) > 0 else 0.0,
            "Avg Weight": round(float(topic_weights.mean()), 4),
            "Max Weight": round(float(topic_weights.max()), 4),
        }
        
        # Add citation and year statistics using df's Topic column
        if df_topics is not None:
            # Create mask for this topic using df's Topic column
            # Handle NaN values by creating a proper boolean mask
            try:
                topic_mask = np.asarray(df_topics == (i + 1), dtype=bool)
            except (ValueError, TypeError):
                # If comparison fails (e.g., due to NaN), create mask manually
                topic_mask = np.zeros(len(df_topics), dtype=bool)
                for idx, val in enumerate(df_topics):
                    try:
                        if val == (i + 1):
                            topic_mask[idx] = True
                    except:
                        pass
            
            # Add citation statistics if available
            if citations is not None and len(citations) == len(topic_mask):
                try:
                    topic_citations = citations[topic_mask]
                    valid_citations = topic_citations[~np.isnan(topic_citations)]
                    
                    stat_dict["Total Citations"] = int(np.nansum(valid_citations)) if len(valid_citations) > 0 else 0
                    stat_dict["Avg Citations"] = round(float(np.nanmean(valid_citations)), 2) if len(valid_citations) > 0 else 0.0
                    stat_dict["Max Citations"] = int(np.nanmax(valid_citations)) if len(valid_citations) > 0 else 0
                except (IndexError, TypeError, ValueError):
                    # Skip citation stats if indexing fails
                    pass
            
            # Add year statistics if available
            if years is not None and len(years) == len(topic_mask):
                try:
                    topic_years = years[topic_mask]
                    valid_years = topic_years[~np.isnan(topic_years)]
                    
                    stat_dict["Avg Year"] = round(float(np.nanmean(valid_years)), 2) if len(valid_years) > 0 else 0.0
                    stat_dict["Min Year"] = int(np.nanmin(valid_years)) if len(valid_years) > 0 else 0
                    stat_dict["Max Year"] = int(np.nanmax(valid_years)) if len(valid_years) > 0 else 0
                except (IndexError, TypeError, ValueError):
                    # Skip year stats if indexing fails
                    pass
        
        stat_dict["Top Terms"] = ", ".join(top_terms)
        
        stats.append(stat_dict)
    
    return pd.DataFrame(stats)


def compute_topic_similarity(
    topic_term_matrix: np.ndarray,
    method: str = 'cosine',
) -> pd.DataFrame:
    """
    Compute pairwise similarity between topics.
    
    Parameters
    ----------
    topic_term_matrix : np.ndarray
        Topic-term distribution matrix (n_topics x n_terms).
    method : str, default='cosine'
        Similarity method: 'cosine', 'jaccard', or 'js' (Jensen-Shannon).
    
    Returns
    -------
    pd.DataFrame
        Similarity matrix as DataFrame.
    """
    from sklearn.metrics.pairwise import cosine_similarity
    from scipy.spatial.distance import jensenshannon
    
    n_topics = topic_term_matrix.shape[0]
    
    if method == 'cosine':
        sim_matrix = cosine_similarity(topic_term_matrix)
    elif method == 'jaccard':
        # Binarize and compute Jaccard
        binary = (topic_term_matrix > np.median(topic_term_matrix)).astype(float)
        sim_matrix = np.zeros((n_topics, n_topics))
        for i in range(n_topics):
            for j in range(n_topics):
                intersection = (binary[i] * binary[j]).sum()
                union = ((binary[i] + binary[j]) > 0).sum()
                sim_matrix[i, j] = intersection / union if union > 0 else 0
    elif method == 'js':
        # Jensen-Shannon divergence (convert to similarity)
        # Normalize rows to probability distributions
        probs = topic_term_matrix / topic_term_matrix.sum(axis=1, keepdims=True)
        sim_matrix = np.zeros((n_topics, n_topics))
        for i in range(n_topics):
            for j in range(n_topics):
                js_dist = jensenshannon(probs[i], probs[j])
                sim_matrix[i, j] = 1 - js_dist
    else:
        raise ValueError(f"Unknown method: {method}")
    
    labels = [f"Topic {i+1}" for i in range(n_topics)]
    return pd.DataFrame(sim_matrix, index=labels, columns=labels)


def compute_topic_trends(
    df: pd.DataFrame,
    year_column: str = 'Year',
    normalize: bool = True,
) -> pd.DataFrame:
    """
    Compute topic prevalence over time.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with Topic assignment and weight columns.
    year_column : str, default='Year'
        Column containing year information.
    normalize : bool, default=True
        Whether to normalize to proportions within each year.
    
    Returns
    -------
    pd.DataFrame
        Topic prevalence by year.
    """
    if year_column not in df.columns or 'Topic' not in df.columns:
        return pd.DataFrame()
    
    # Find weight columns
    weight_cols = [c for c in df.columns if c.startswith('Topic ') and c.endswith(' Weight')]
    
    if weight_cols:
        # Use weights
        result = df.groupby(year_column)[weight_cols].sum()
        result.columns = [c.replace(' Weight', '') for c in result.columns]
    else:
        # Use hard assignments
        result = pd.crosstab(df[year_column], df['Topic'])
        result.columns = [f"Topic {int(c)}" if isinstance(c, (int, float)) else str(c) for c in result.columns]
    
    if normalize:
        result = result.div(result.sum(axis=1), axis=0)
    
    return result


# =============================================================================
# SEQUENTIAL TOPIC MODELING (STM)
# =============================================================================

def sequential_topic_modeling(
    df: pd.DataFrame,
    text_column: str,
    time_column: str = 'Year',
    n_topics: int = 5,
    time_slices: Optional[List[int]] = None,
    model_type: str = 'LDA',
    max_features: int = 5000,
    stop_words: str | list[str] | None = 'english',
    min_df: int = 2,
    max_df: float = 0.95,
    ngram_range: Tuple[int, int] = (1, 1),
    progress_callback: Optional[callable] = None,
) -> Dict[str, Any]:
    """
    Sequential Topic Modeling - fit separate topic models per time period.
    
    This approach fits independent topic models to each time slice,
    then aligns topics across periods using similarity matching.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with text and time columns.
    text_column : str
        Column containing text to model.
    time_column : str, default='Year'
        Column containing time information (year).
    n_topics : int, default=5
        Number of topics per time slice.
    time_slices : list of int, optional
        Custom time slice boundaries. If None, uses each unique year.
    model_type : str, default='LDA'
        Model type: 'LDA', 'NMF', or 'LSA'.
    max_features : int, default=5000
        Maximum vocabulary size.
    stop_words : str or list, default='english'
        Stop words to remove.
    min_df : int, default=2
        Minimum document frequency.
    max_df : float, default=0.95
        Maximum document frequency.
    ngram_range : tuple, default=(1, 1)
        N-gram range for vectorization.
    progress_callback : callable, optional
        Function to call with progress updates.
    
    Returns
    -------
    dict with keys:
        - df_out: DataFrame with topic assignments per period
        - period_models: dict of {period: model}
        - period_topics: dict of {period: topics_df}
        - topic_evolution: DataFrame showing topic alignment across periods
        - topic_prevalence: DataFrame of topic prevalence over time
        - vocabulary: shared vocabulary across all periods
        - period_stats: DataFrame with period-level statistics
    """
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.decomposition import LatentDirichletAllocation, NMF, TruncatedSVD
    from sklearn.metrics.pairwise import cosine_similarity
    
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found.")
    if time_column not in df.columns:
        raise KeyError(f"Column '{time_column}' not found.")
    
    df_work = df.copy()
    df_work[time_column] = pd.to_numeric(df_work[time_column], errors='coerce')
    df_work = df_work.dropna(subset=[time_column])
    df_work[time_column] = df_work[time_column].astype(int)
    
    # Determine time periods
    if time_slices is None:
        periods = sorted(df_work[time_column].unique())
    else:
        periods = sorted(time_slices)
    
    if len(periods) < 2:
        raise ValueError("Need at least 2 time periods for sequential modeling.")
    
    # Setup vectorizer
    mt = model_type.upper()
    if mt in {"LDA", "LSA"}:
        vectorizer = CountVectorizer(
            max_features=max_features,
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    else:
        vectorizer = TfidfVectorizer(
            max_features=max_features,
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    
    # Fit vectorizer on all text to get shared vocabulary
    all_texts = df_work[text_column].fillna("").astype(str)
    all_texts = all_texts[all_texts.str.strip() != ""]
    
    try:
        vectorizer.fit(all_texts)
    except ValueError:
        return {
            "df_out": df.copy(),
            "period_models": {},
            "period_topics": {},
            "topic_evolution": pd.DataFrame(),
            "topic_prevalence": pd.DataFrame(),
            "vocabulary": [],
            "period_stats": pd.DataFrame(),
        }
    
    vocabulary = vectorizer.get_feature_names_out()
    
    # Fit models per period
    period_models = {}
    period_topics = {}
    period_doc_topics = {}
    period_topic_terms = {}
    period_stats_list = []
    
    for idx, period in enumerate(periods):
        if progress_callback:
            progress_callback(f"Processing period {period} ({idx+1}/{len(periods)})")
        
        period_mask = df_work[time_column] == period
        period_df = df_work[period_mask]
        
        if len(period_df) < n_topics:
            continue
        
        period_texts = period_df[text_column].fillna("").astype(str)
        valid_mask = period_texts.str.strip() != ""
        
        if valid_mask.sum() < n_topics:
            continue
        
        # Transform with shared vocabulary
        X = vectorizer.transform(period_texts[valid_mask])
        
        # Fit model
        if mt == "LDA":
            model = LatentDirichletAllocation(n_components=n_topics, random_state=42, max_iter=30)
        elif mt == "NMF":
            model = NMF(n_components=n_topics, init="nndsvda", random_state=42, max_iter=200)
        else:
            model = TruncatedSVD(n_components=n_topics, random_state=42)
        
        model.fit(X)
        
        # Get document-topic distribution
        doc_topics = model.transform(X)
        if mt == "LSA":
            doc_topics = np.abs(doc_topics)
        
        # Normalize
        row_sums = doc_topics.sum(axis=1, keepdims=True)
        doc_topics = np.divide(doc_topics, row_sums, out=np.zeros_like(doc_topics), where=row_sums > 0)
        
        period_models[period] = model
        period_doc_topics[period] = (period_df.index[valid_mask], doc_topics)
        period_topic_terms[period] = model.components_
        
        # Build topics_df for this period
        topic_rows = []
        for i in range(n_topics):
            top_idx = np.argsort(model.components_[i])[-10:][::-1]
            for j in top_idx:
                topic_rows.append({
                    "Period": period,
                    "Topic": f"Topic {i+1}",
                    "Topic_ID": i + 1,
                    "Term": vocabulary[j],
                    "Weight": float(model.components_[i, j]),
                })
        period_topics[period] = pd.DataFrame(topic_rows)
        
        # Period statistics
        period_stats_list.append({
            "Period": period,
            "Documents": len(period_df),
            "Valid Documents": valid_mask.sum(),
            "Topics": n_topics,
        })
    
    period_stats = pd.DataFrame(period_stats_list)
    
    # Align topics across periods using cosine similarity
    topic_evolution = _align_topics_across_periods(
        period_topic_terms, periods, n_topics, vocabulary
    )
    
    # Build df_out with topic assignments
    df_out = df.copy()
    df_out["STM_Topic"] = np.nan
    df_out["STM_Period"] = np.nan
    
    for period, (indices, doc_topics) in period_doc_topics.items():
        assignments = doc_topics.argmax(axis=1) + 1
        df_out.loc[indices, "STM_Topic"] = assignments
        df_out.loc[indices, "STM_Period"] = period
        
        # Add weight columns
        for k in range(n_topics):
            col_name = f"STM_Topic_{k+1}_Weight"
            if col_name not in df_out.columns:
                df_out[col_name] = np.nan
            df_out.loc[indices, col_name] = doc_topics[:, k]
    
    # Compute topic prevalence over time
    topic_prevalence = _compute_sequential_prevalence(period_doc_topics, periods, n_topics)
    
    return {
        "df_out": df_out,
        "period_models": period_models,
        "period_topics": period_topics,
        "topic_evolution": topic_evolution,
        "topic_prevalence": topic_prevalence,
        "vocabulary": vocabulary,
        "period_stats": period_stats,
        "n_topics": n_topics,
        "periods": periods,
    }


def _align_topics_across_periods(
    period_topic_terms: Dict[int, np.ndarray],
    periods: List[int],
    n_topics: int,
    vocabulary: np.ndarray,
) -> pd.DataFrame:
    """Align topics across time periods using cosine similarity."""
    from sklearn.metrics.pairwise import cosine_similarity
    
    if len(periods) < 2:
        return pd.DataFrame()
    
    evolution_rows = []
    
    # Use first period as reference
    ref_period = periods[0]
    if ref_period not in period_topic_terms:
        return pd.DataFrame()
    
    ref_topics = period_topic_terms[ref_period]
    
    # Track topic mapping
    topic_mapping = {ref_period: list(range(n_topics))}
    
    for i, period in enumerate(periods[1:], 1):
        if period not in period_topic_terms:
            continue
        
        curr_topics = period_topic_terms[period]
        prev_period = periods[i-1]
        
        if prev_period not in period_topic_terms:
            continue
        
        prev_topics = period_topic_terms[prev_period]
        
        # Compute similarity matrix
        sim_matrix = cosine_similarity(prev_topics, curr_topics)
        
        # Greedy matching (Hungarian algorithm would be better but this is simpler)
        mapping = []
        used = set()
        for prev_idx in range(n_topics):
            best_curr = -1
            best_sim = -1
            for curr_idx in range(n_topics):
                if curr_idx not in used and sim_matrix[prev_idx, curr_idx] > best_sim:
                    best_sim = sim_matrix[prev_idx, curr_idx]
                    best_curr = curr_idx
            if best_curr >= 0:
                mapping.append(best_curr)
                used.add(best_curr)
                
                evolution_rows.append({
                    "From_Period": prev_period,
                    "To_Period": period,
                    "From_Topic": f"Topic {prev_idx+1}",
                    "To_Topic": f"Topic {best_curr+1}",
                    "Similarity": round(best_sim, 4),
                })
        
        topic_mapping[period] = mapping
    
    return pd.DataFrame(evolution_rows)


def _compute_sequential_prevalence(
    period_doc_topics: Dict[int, Tuple],
    periods: List[int],
    n_topics: int,
) -> pd.DataFrame:
    """Compute topic prevalence for each period."""
    prevalence_data = []
    
    for period in periods:
        if period not in period_doc_topics:
            continue
        
        indices, doc_topics = period_doc_topics[period]
        
        # Average topic weight across documents
        avg_weights = doc_topics.mean(axis=0)
        
        for k in range(n_topics):
            prevalence_data.append({
                "Period": period,
                "Topic": f"Topic {k+1}",
                "Topic_ID": k + 1,
                "Prevalence": round(float(avg_weights[k]), 4),
                "Document_Count": int((doc_topics.argmax(axis=1) == k).sum()),
            })
    
    return pd.DataFrame(prevalence_data)


# =============================================================================
# DYNAMIC TOPIC MODELING (DTM)
# =============================================================================

def dynamic_topic_modeling(
    df: pd.DataFrame,
    text_column: str,
    time_column: str = 'Year',
    n_topics: int = 5,
    n_time_slices: int = 5,
    model_type: str = 'LDA',
    max_features: int = 5000,
    stop_words: str | list[str] | None = 'english',
    min_df: int = 2,
    max_df: float = 0.95,
    ngram_range: Tuple[int, int] = (1, 1),
    chain_variance: float = 0.1,
    progress_callback: Optional[callable] = None,
) -> Dict[str, Any]:
    """
    Dynamic Topic Modeling - track topic evolution with temporal smoothing.
    
    This implementation uses a sliding window approach with topic alignment
    and temporal smoothing to track how topics evolve over time.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with text and time columns.
    text_column : str
        Column containing text to model.
    time_column : str, default='Year'
        Column containing time information.
    n_topics : int, default=5
        Number of topics.
    n_time_slices : int, default=5
        Number of time slices to create.
    model_type : str, default='LDA'
        Model type: 'LDA', 'NMF', or 'LSA'.
    max_features : int, default=5000
        Maximum vocabulary size.
    stop_words : str or list, default='english'
        Stop words to remove.
    min_df : int, default=2
        Minimum document frequency.
    max_df : float, default=0.95
        Maximum document frequency.
    ngram_range : tuple, default=(1, 1)
        N-gram range for vectorization.
    chain_variance : float, default=0.1
        Controls smoothness of topic evolution (lower = smoother).
    progress_callback : callable, optional
        Function to call with progress updates.
    
    Returns
    -------
    dict with keys:
        - df_out: DataFrame with topic assignments
        - topic_word_evolution: dict of {topic_id: DataFrame} showing word weights over time
        - topic_prevalence_evolution: DataFrame of topic prevalence over time
        - time_slice_info: DataFrame with time slice boundaries
        - global_topics: DataFrame with overall topic-word weights
        - coherence_over_time: dict of {time_slice: coherence_scores}
        - n_topics: number of topics
        - n_time_slices: number of time slices
    """
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.decomposition import LatentDirichletAllocation, NMF, TruncatedSVD
    
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found.")
    if time_column not in df.columns:
        raise KeyError(f"Column '{time_column}' not found.")
    
    df_work = df.copy()
    df_work[time_column] = pd.to_numeric(df_work[time_column], errors='coerce')
    df_work = df_work.dropna(subset=[time_column])
    
    if len(df_work) == 0:
        return _empty_dtm_result(df)
    
    # Create time slices
    min_year = int(df_work[time_column].min())
    max_year = int(df_work[time_column].max())
    
    if max_year - min_year < n_time_slices - 1:
        n_time_slices = max(2, max_year - min_year + 1)
    
    slice_boundaries = np.linspace(min_year, max_year + 1, n_time_slices + 1).astype(int)
    time_slices = []
    
    for i in range(len(slice_boundaries) - 1):
        start = slice_boundaries[i]
        end = slice_boundaries[i + 1] - 1 if i < len(slice_boundaries) - 2 else slice_boundaries[i + 1]
        time_slices.append((start, end))
    
    time_slice_info = pd.DataFrame([
        {"Slice": i + 1, "Start": s, "End": e, "Label": f"{s}-{e}"}
        for i, (s, e) in enumerate(time_slices)
    ])
    
    # Setup vectorizer
    mt = model_type.upper()
    if mt in {"LDA", "LSA"}:
        vectorizer = CountVectorizer(
            max_features=max_features,
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    else:
        vectorizer = TfidfVectorizer(
            max_features=max_features,
            stop_words=stop_words,
            ngram_range=ngram_range,
            min_df=min_df,
            max_df=max_df,
        )
    
    # Fit vectorizer on all text
    all_texts = df_work[text_column].fillna("").astype(str)
    valid_text_mask = all_texts.str.strip() != ""
    
    try:
        vectorizer.fit(all_texts[valid_text_mask])
    except ValueError:
        return _empty_dtm_result(df)
    
    vocabulary = vectorizer.get_feature_names_out()
    n_vocab = len(vocabulary)
    
    # Initialize topic-word distributions with prior
    topic_word_prior = np.ones((n_topics, n_vocab)) / n_vocab
    
    # Track evolution
    topic_word_history = []  # List of (time_slice, topic_word_matrix)
    topic_prevalence_data = []
    coherence_over_time = {}
    slice_doc_topics = {}
    
    prev_topic_word = topic_word_prior.copy()
    
    for slice_idx, (start_year, end_year) in enumerate(time_slices):
        if progress_callback:
            progress_callback(f"Processing slice {slice_idx+1}/{len(time_slices)}: {start_year}-{end_year}")
        
        # Get documents for this time slice
        slice_mask = (df_work[time_column] >= start_year) & (df_work[time_column] <= end_year)
        slice_df = df_work[slice_mask]
        
        if len(slice_df) < n_topics:
            # Use previous topic-word distribution
            topic_word_history.append((slice_idx, prev_topic_word.copy()))
            continue
        
        slice_texts = slice_df[text_column].fillna("").astype(str)
        valid_mask = slice_texts.str.strip() != ""
        
        if valid_mask.sum() < n_topics:
            topic_word_history.append((slice_idx, prev_topic_word.copy()))
            continue
        
        # Transform texts
        X = vectorizer.transform(slice_texts[valid_mask])
        
        # Fit model with warm start from previous slice
        if mt == "LDA":
            # Initialize with smoothed previous topic-word
            init_components = (1 - chain_variance) * prev_topic_word + chain_variance * topic_word_prior
            init_components = init_components / init_components.sum(axis=1, keepdims=True)
            
            model = LatentDirichletAllocation(
                n_components=n_topics,
                random_state=42,
                max_iter=30,
            )
            model.fit(X)
            
            # Apply temporal smoothing
            raw_components = model.components_
            smoothed = (1 - chain_variance) * prev_topic_word + chain_variance * raw_components
            model.components_ = smoothed
            
        elif mt == "NMF":
            model = NMF(n_components=n_topics, init="nndsvda", random_state=42, max_iter=200)
            model.fit(X)
            
            # Apply temporal smoothing
            raw_components = model.components_
            smoothed = (1 - chain_variance) * prev_topic_word + chain_variance * raw_components
            model.components_ = smoothed
        else:
            model = TruncatedSVD(n_components=n_topics, random_state=42)
            model.fit(X)
            raw_components = np.abs(model.components_)
            smoothed = (1 - chain_variance) * prev_topic_word + chain_variance * raw_components
            model.components_ = smoothed
        
        # Store topic-word distribution
        topic_word_history.append((slice_idx, model.components_.copy()))
        prev_topic_word = model.components_.copy()
        
        # Get document-topic distribution
        doc_topics = model.transform(X)
        if mt == "LSA":
            doc_topics = np.abs(doc_topics)
        
        # Normalize
        row_sums = doc_topics.sum(axis=1, keepdims=True)
        doc_topics = np.divide(doc_topics, row_sums, out=np.zeros_like(doc_topics), where=row_sums > 0)
        
        slice_doc_topics[slice_idx] = (slice_df.index[valid_mask], doc_topics)
        
        # Compute prevalence
        avg_weights = doc_topics.mean(axis=0)
        for k in range(n_topics):
            topic_prevalence_data.append({
                "Time_Slice": slice_idx + 1,
                "Time_Label": f"{start_year}-{end_year}",
                "Topic": f"Topic {k+1}",
                "Topic_ID": k + 1,
                "Prevalence": round(float(avg_weights[k]), 4),
                "Document_Count": int((doc_topics.argmax(axis=1) == k).sum()),
            })
        
        # Compute coherence
        coherence = _compute_topic_coherence(model.components_, X, vocabulary)
        coherence_over_time[slice_idx] = {f"Topic {k+1}": coherence[k] for k in range(n_topics)}
    
    # Build topic-word evolution DataFrames
    topic_word_evolution = {}
    for k in range(n_topics):
        evo_data = []
        for slice_idx, topic_word in topic_word_history:
            slice_info = time_slices[slice_idx]
            top_idx = np.argsort(topic_word[k])[-15:][::-1]
            for idx in top_idx:
                evo_data.append({
                    "Time_Slice": slice_idx + 1,
                    "Time_Label": f"{slice_info[0]}-{slice_info[1]}",
                    "Term": vocabulary[idx],
                    "Weight": round(float(topic_word[k, idx]), 4),
                })
        topic_word_evolution[f"Topic {k+1}"] = pd.DataFrame(evo_data)
    
    # Build global topics (average across time)
    if topic_word_history:
        avg_topic_word = np.mean([tw for _, tw in topic_word_history], axis=0)
        global_topic_rows = []
        for k in range(n_topics):
            top_idx = np.argsort(avg_topic_word[k])[-15:][::-1]
            for idx in top_idx:
                global_topic_rows.append({
                    "Topic": f"Topic {k+1}",
                    "Topic_ID": k + 1,
                    "Term": vocabulary[idx],
                    "Weight": round(float(avg_topic_word[k, idx]), 4),
                })
        global_topics = pd.DataFrame(global_topic_rows)
    else:
        global_topics = pd.DataFrame()
    
    # Build df_out
    df_out = df.copy()
    df_out["DTM_Topic"] = np.nan
    df_out["DTM_Time_Slice"] = np.nan
    
    for slice_idx, (indices, doc_topics) in slice_doc_topics.items():
        assignments = doc_topics.argmax(axis=1) + 1
        df_out.loc[indices, "DTM_Topic"] = assignments
        df_out.loc[indices, "DTM_Time_Slice"] = slice_idx + 1
        
        for k in range(n_topics):
            col_name = f"DTM_Topic_{k+1}_Weight"
            if col_name not in df_out.columns:
                df_out[col_name] = np.nan
            df_out.loc[indices, col_name] = doc_topics[:, k]
    
    return {
        "df_out": df_out,
        "topic_word_evolution": topic_word_evolution,
        "topic_prevalence_evolution": pd.DataFrame(topic_prevalence_data),
        "time_slice_info": time_slice_info,
        "global_topics": global_topics,
        "coherence_over_time": coherence_over_time,
        "n_topics": n_topics,
        "n_time_slices": len(time_slices),
        "vocabulary": vocabulary,
    }


def _empty_dtm_result(df: pd.DataFrame) -> Dict[str, Any]:
    """Return empty DTM result structure."""
    return {
        "df_out": df.copy(),
        "topic_word_evolution": {},
        "topic_prevalence_evolution": pd.DataFrame(),
        "time_slice_info": pd.DataFrame(),
        "global_topics": pd.DataFrame(),
        "coherence_over_time": {},
        "n_topics": 0,
        "n_time_slices": 0,
        "vocabulary": [],
    }


def compute_term_evolution(
    df: pd.DataFrame,
    text_column: str,
    time_column: str = 'Year',
    top_n_terms: int = 20,
    min_df: int = 2,
    max_df: float = 0.95,
    stop_words: str | list | None = 'english',
    ngram_range: Tuple[int, int] = (1, 1),
    normalize: bool = True,
) -> pd.DataFrame:
    """
    Compute term frequency evolution over time.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe.
    text_column : str
        Column containing text.
    time_column : str, default='Year'
        Column containing time information.
    top_n_terms : int, default=20
        Number of top terms to track.
    min_df : int, default=2
        Minimum document frequency.
    max_df : float, default=0.95
        Maximum document frequency.
    stop_words : str or list, default='english'
        Stop words to remove.
    ngram_range : tuple, default=(1, 1)
        N-gram range.
    normalize : bool, default=True
        Normalize frequencies within each time period.
    
    Returns
    -------
    pd.DataFrame
        Term evolution with columns: Period, Term, Frequency, Rank
    """
    from sklearn.feature_extraction.text import CountVectorizer
    
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found.")
    if time_column not in df.columns:
        raise KeyError(f"Column '{time_column}' not found.")
    
    df_work = df.copy()
    df_work[time_column] = pd.to_numeric(df_work[time_column], errors='coerce')
    df_work = df_work.dropna(subset=[time_column, text_column])
    df_work[time_column] = df_work[time_column].astype(int)
    
    periods = sorted(df_work[time_column].unique())
    
    if len(periods) == 0:
        return pd.DataFrame()
    
    # Fit vectorizer on all text to get consistent vocabulary
    all_texts = df_work[text_column].fillna("").astype(str)
    valid_mask = all_texts.str.strip() != ""
    
    vectorizer = CountVectorizer(
        stop_words=stop_words,
        ngram_range=ngram_range,
        min_df=min_df,
        max_df=max_df,
    )
    
    try:
        vectorizer.fit(all_texts[valid_mask])
    except ValueError:
        return pd.DataFrame()
    
    vocabulary = vectorizer.get_feature_names_out()
    
    # Compute term frequencies per period
    evolution_data = []
    
    # Get overall top terms
    X_all = vectorizer.transform(all_texts[valid_mask])
    term_freqs_all = np.array(X_all.sum(axis=0)).flatten()
    top_term_indices = np.argsort(term_freqs_all)[-top_n_terms:][::-1]
    top_terms = [vocabulary[i] for i in top_term_indices]
    
    for period in periods:
        period_mask = df_work[time_column] == period
        period_texts = df_work.loc[period_mask, text_column].fillna("").astype(str)
        period_valid = period_texts.str.strip() != ""
        
        if period_valid.sum() == 0:
            continue
        
        X_period = vectorizer.transform(period_texts[period_valid])
        term_freqs = np.array(X_period.sum(axis=0)).flatten()
        
        if normalize:
            total = term_freqs.sum()
            if total > 0:
                term_freqs = term_freqs / total
        
        # Get frequencies for top terms
        for rank, term in enumerate(top_terms, 1):
            term_idx = list(vocabulary).index(term)
            freq = float(term_freqs[term_idx])
            
            evolution_data.append({
                "Period": period,
                "Term": term,
                "Frequency": round(freq, 6) if normalize else int(freq),
                "Rank": rank,
            })
    
    return pd.DataFrame(evolution_data)


def compute_term_trends(
    df: pd.DataFrame,
    text_column: str,
    time_column: str = 'Year',
    terms: Optional[List[str]] = None,
    top_n_terms: int = 10,
    min_df: int = 2,
    max_df: float = 0.95,
    stop_words: str | list | None = 'english',
    normalize: bool = True,
) -> pd.DataFrame:
    """
    Compute trends for specific terms over time.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe.
    text_column : str
        Column containing text.
    time_column : str, default='Year'
        Column containing time information.
    terms : list of str, optional
        Specific terms to track. If None, uses top terms.
    top_n_terms : int, default=10
        Number of top terms if terms not specified.
    min_df : int, default=2
        Minimum document frequency.
    max_df : float, default=0.95
        Maximum document frequency.
    stop_words : str or list, default='english'
        Stop words to remove.
    normalize : bool, default=True
        Normalize to proportion of documents containing term.
    
    Returns
    -------
    pd.DataFrame
        Pivoted DataFrame with periods as index, terms as columns.
    """
    from sklearn.feature_extraction.text import CountVectorizer
    
    if text_column not in df.columns:
        raise KeyError(f"Column '{text_column}' not found.")
    if time_column not in df.columns:
        raise KeyError(f"Column '{time_column}' not found.")
    
    df_work = df.copy()
    df_work[time_column] = pd.to_numeric(df_work[time_column], errors='coerce')
    df_work = df_work.dropna(subset=[time_column, text_column])
    df_work[time_column] = df_work[time_column].astype(int)
    
    all_texts = df_work[text_column].fillna("").astype(str)
    valid_mask = all_texts.str.strip() != ""
    
    vectorizer = CountVectorizer(
        stop_words=stop_words,
        min_df=min_df,
        max_df=max_df,
    )
    
    try:
        vectorizer.fit(all_texts[valid_mask])
    except ValueError:
        return pd.DataFrame()
    
    vocabulary = list(vectorizer.get_feature_names_out())
    
    # Determine terms to track
    if terms is None:
        X_all = vectorizer.transform(all_texts[valid_mask])
        term_freqs = np.array(X_all.sum(axis=0)).flatten()
        top_indices = np.argsort(term_freqs)[-top_n_terms:][::-1]
        terms = [vocabulary[i] for i in top_indices]
    else:
        # Filter to terms in vocabulary
        terms = [t for t in terms if t in vocabulary]
    
    if not terms:
        return pd.DataFrame()
    
    periods = sorted(df_work[time_column].unique())
    
    # Build result
    result_data = {term: [] for term in terms}
    result_data["Period"] = []
    
    for period in periods:
        period_mask = df_work[time_column] == period
        period_texts = df_work.loc[period_mask, text_column].fillna("").astype(str)
        period_valid = period_texts.str.strip() != ""
        n_docs = period_valid.sum()
        
        if n_docs == 0:
            continue
        
        X_period = vectorizer.transform(period_texts[period_valid])
        
        result_data["Period"].append(period)
        
        for term in terms:
            term_idx = vocabulary.index(term)
            if normalize:
                # Proportion of documents containing term
                doc_count = (X_period[:, term_idx].toarray() > 0).sum()
                freq = doc_count / n_docs
            else:
                freq = X_period[:, term_idx].sum()
            
            result_data[term].append(round(float(freq), 4))
    
    result_df = pd.DataFrame(result_data)
    result_df = result_df.set_index("Period")
    
    return result_df


"""Utilities related to Semantic interdisciplinarity analysis in the Biblium project."""
# semantic interdisciplinarity analysis

"""Utilities related to Optional import with fallback in the Biblium project."""
# Optional import with fallback
try:
    from sentence_transformers import SentenceTransformer
    _model_available = True
    _model = SentenceTransformer("all-MiniLM-L6-v2")
except ImportError:
    _model_available = False
    _model = None

def semantic_interdisciplinarity(
    text: str,
    mode: str = 'keywords',
    sep: str = '; ',
) -> float:
    """
    Compute semantic interdisciplinarity based on average pairwise cosine distance
    using sentence-transformers embeddings.

    Parameters:
        text (str): Keywords separated by sep (default "; "), or free-form text.
        mode (str): Either 'keywords' or 'text'.

    Returns:
        float: Mean pairwise cosine distance between embedded concepts.
               Returns np.nan if input is missing or insufficient.

    Raises:
        ImportError: If sentence-transformers is not installed.
    """
    if not _model_available:
        raise ImportError(
            "The 'sentence-transformers' package is required for semantic_interdisciplinarity().\n"
            "Install it with: pip install sentence-transformers"
        )

    if not isinstance(text, str) or not text.strip():
        return np.nan  # Handle None, NaN, or empty string

    if mode == "keywords":
        tokens = [t.strip() for t in text.split(sep) if t.strip()]
    elif mode == "text":
        tokens = text.split()
    else:
        raise ValueError("mode must be 'keywords' or 'text'")

    if len(tokens) < 2:
        return np.nan  # Not enough concepts to compare

    embeddings = _model.encode(tokens, convert_to_numpy=True)
    normed = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)

    distances = [1 - np.dot(normed[i], normed[j])
                 for i, j in combinations(range(len(normed)), 2)]

    return float(np.mean(distances)) if distances else np.nan

"""Utilities related to Sentiment analysis in the Biblium project."""
# sentiment analysis

def analyze_sentiment(
    df,
    text_column,
    sentiment_threshold = 0.05,
    top_words = 10,
):
    """
    Performs sentiment analysis on a given DataFrame.

    Parameters:
    - df (pd.DataFrame): Input DataFrame with text data.
    - text_column (str): Name of the column containing the text to analyze.
    - sentiment_threshold (float): Threshold for sentiment classification (default 0.05).
    - top_words (int): Number of most common words to return for each sentiment category.

    Returns:
    - df (pd.DataFrame): Updated DataFrame with sentiment scores and categories.
    - stats_df (pd.DataFrame): DataFrame containing sentiment statistics and common words.
    """
    # Import required modules
    try:
        import nltk
        from nltk.sentiment import SentimentIntensityAnalyzer
    except ImportError:
        raise ImportError("nltk is required for sentiment analysis. Install with: pip install nltk")

    # Download lexicon and initialize sentiment analyzer
    try:
        nltk.download("vader_lexicon", quiet=True)
    except Exception:
        pass  # May already be downloaded or offline
    
    sia = SentimentIntensityAnalyzer()

    # Compute sentiment scores
    df["Sentiment Score"] = df[text_column].apply(lambda x: sia.polarity_scores(str(x))["compound"])

    # Assign sentiment categories
    df["Sentiment Category"] = df["Sentiment Score"].apply(
        lambda x: "Positive" if x > sentiment_threshold else "Negative" if x < -sentiment_threshold else "Neutral"
    )

    # Compute statistics
    stats = {
        "Mean Sentiment Score": df["Sentiment Score"].mean(),
        "Median Sentiment Score": df["Sentiment Score"].median(),
        "Standard Deviation": df["Sentiment Score"].std()
    }

    sentiment_distribution = df["Sentiment Category"].value_counts(normalize=True) * 100
    stats.update(sentiment_distribution.to_dict())

    # Word Frequency Analysis for each sentiment category
    word_stats = {}
    for category in ["Positive", "Neutral", "Negative"]:
        words = " ".join(df[df["Sentiment Category"] == category][text_column].astype(str)).lower().split()
        common_words = Counter(words).most_common(top_words)
        word_stats[f"Top Words ({category} Documents)"] = "; ".join([word for word, _ in common_words])

    # Convert stats and word frequencies to DataFrame
    stats_dict = {**stats, **word_stats}
    stats_df = pd.DataFrame(list(stats_dict.items()), columns=["Metric", "Value"])

    return df, stats_df

"""Definitions and helpers for bibliometric performance indicators such as citation counts and h-index variants."""
# Performance indicators

"""Performance-indicator helpers that work on simple Python lists or one-dimensional arrays."""
# --- Performance Indicators: List-based functions ---


# --- HELPER: Universal Cleaner ---
def _clean_list(data):
    """
    Helper to strictly filter a list to valid integers only.
    Handles: strings ("10"), floats (10.0), NaNs, and ignores garbage ("N/A").
    """
    if data is None:
        return []
        
    clean = []
    for x in data:
        try:
            # Skip None or NaN
            if x is None:
                continue
            if isinstance(x, float) and np.isnan(x):
                continue
            
            # Convert to float first to handle strings like "10.0", then to int
            val = int(float(x))
            clean.append(val)
        except (ValueError, TypeError):
            continue
    return clean

# --- CORE INDICATOR FUNCTIONS ---

def h_index(citations, alpha=1):
    """
    Compute the generalized h-index with scaling factor alpha.
    """
    citations = _clean_list(citations)
    return sum(c >= alpha * (i + 1) for i, c in enumerate(sorted(citations, reverse=True)))

def g_index(citations):
    """
    Compute the g-index based on citation distribution.
    """
    citations = sorted(_clean_list(citations), reverse=True)
    if not citations:
        return 0
        
    cumulative = np.cumsum(citations)
    thresholds = np.arange(1, len(citations) + 1) ** 2
    differences = cumulative - thresholds
    failing = [i for i, x in enumerate(differences) if x < 0]
    return failing[0] if failing else len(citations)

def hg_index(citations):
    """
    Compute the HG-index as the geometric mean of h and g indices.
    """
    # h_index and g_index already clean the data internally
    return np.sqrt(h_index(citations) * g_index(citations))

def c_index(citations, thresholds=[1, 5, 10, 20, 50, 100, 1000]):
    """
    Compute the C-index for a set of citation thresholds.
    Returns a dictionary of counts.
    """
    citations = np.array(_clean_list(citations))
    if len(citations) == 0:
        return {c: 0 for c in thresholds}
        
    return {c: int(np.sum(citations >= c)) for c in thresholds}

def tapered_h_index(citations):
    """
    Compute the tapered h-index (a more nuanced variant of h-index).
    """
    citations = _clean_list(citations)
    h = 0.0
    for i, cit in enumerate(sorted(citations, reverse=True)):
        # logic adjusted to work on sorted array properly or unsorted map? 
        # Tapered h usually requires sorting descending.
        # Assuming original logic intended unsorted map based on paper ID? 
        # Usually h-indexes work on ranked lists. I will sort descending to be safe.
        k = min(i + 1, cit)
        h += k / (2 * i + 1)
        h += sum(1 / (2 * i + 1) for _ in range(i + 1, cit)) # fixed variable shadowing
    return h

def chi_index(citations):
    """
    Compute the chi-index, a robust citation indicator.
    """
    citations = sorted(_clean_list(citations), reverse=True)
    if len(citations) == 0:
        return np.nan
    return np.sqrt(max((i + 1) * c for i, c in enumerate(citations)))

# --- EXPERIMENTAL INDICATORS ---

def a_index(citations):
    """
    Compute the a-index: the average number of citations in the h-core.
    """
    citations = _clean_list(citations)
    h = h_index(citations)
    if h == 0:
        return 0.0
    return np.mean(sorted(citations, reverse=True)[:h])

def r_index(citations):
    """
    Compute the r-index: square root of total citations in the h-core.
    """
    citations = _clean_list(citations)
    h = h_index(citations)
    if h == 0:
        return 0.0
    return np.sqrt(sum(sorted(citations, reverse=True)[:h]))

def h2_index(citations):
    """
    Compute the h(2)-index: largest number h2 such that h2 papers have at least (h2)^2 citations.
    """
    citations = sorted(_clean_list(citations), reverse=True)
    return sum(c >= (i + 1) ** 2 for i, c in enumerate(citations))

def w_index(citations):
    """
    Compute the w-index: number of papers with at least 10 × rank citations.
    """
    citations = sorted(_clean_list(citations), reverse=True)
    return sum(c >= 10 * (i + 1) for i, c in enumerate(citations))

def t_index(citations):
    """
    Compute the t-index: square root of the sum of square roots of citation counts.
    """
    citations = _clean_list(citations)
    return np.sqrt(sum(np.sqrt(c) for c in citations if c >= 0))

def pi_index(citations):
    """
    Compute the pi-index: total citations for the top √N publications.
    """
    citations = sorted(_clean_list(citations), reverse=True)
    n = len(citations)
    if n == 0:
        return 0
    top_k = int(np.sqrt(n))
    return sum(citations[:top_k])

def gini_index(citations):
    """
    Compute the Gini coefficient: measures inequality in the citation distribution.
    """
    citations = sorted(_clean_list(citations)) # Gini requires ascending sort
    n = len(citations)
    if n == 0:
        return 0.0
    
    citations = np.array(citations)
    cum_cit = np.cumsum(citations)
    
    if cum_cit[-1] == 0:
        return 0.0
        
    return (n + 1 - 2 * np.sum(cum_cit) / cum_cit[-1]) / n

# --- DATAFRAME WRAPPERS ---

def _get_clean_col(df, col_name):
    """Helper to extract a clean numeric series from DataFrame."""
    if col_name not in df.columns:
        return pd.Series(dtype=float)
    return pd.to_numeric(df[col_name], errors='coerce').dropna()

def total_citations(df):
    """
    Return total citations from the 'Cited by' column.
    """
    return _get_clean_col(df, "Cited by").sum()

def count_cited_documents(df):
    """
    Return the number of documents that have been cited (non-zero entries).
    """
    clean_series = _get_clean_col(df, "Cited by")
    return clean_series.gt(0).sum()


def average_year(df):
    """
    Compute the average publication year from a DataFrame.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Year' column.
        
    Returns
    -------
    float or None
        Mean year, or None if no valid years.
    """
    clean_years = _get_clean_col(df, "Year")
    if clean_years.empty:
        return None
    return clean_years.mean()

def percentile_year(df, quantile=0.5):
    """
    Return year at given percentile (e.g., median with quantile=0.5).
    """
    clean_years = _get_clean_col(df, "Year")
    if clean_years.empty:
        return None
    return clean_years.quantile(q=quantile)

def first_publication_year(df):
    """
    Return the earliest publication year.
    """
    clean_years = _get_clean_col(df, "Year")
    if clean_years.empty:
        return None
    return clean_years.min()

def last_publication_year(df):
    """
    Return the latest publication year.
    """
    clean_years = _get_clean_col(df, "Year")
    if clean_years.empty:
        return None
    return clean_years.max()


def h_index_df(df, alpha=1):
    """
    Compute h-index from a DataFrame's citation column.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Cited by' column.
    alpha : float, optional
        Scaling factor for generalized h-index.
        
    Returns
    -------
    int
        H-index value.
    """
    return h_index(_get_clean_col(df, "Cited by").tolist(), alpha)


def g_index_df(df):
    """
    Compute g-index from a DataFrame's citation column.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Cited by' column.
        
    Returns
    -------
    int
        G-index value.
    """
    return g_index(_get_clean_col(df, "Cited by").tolist())


def hg_index_df(df):
    """
    Compute HG-index (geometric mean of h and g) from a DataFrame.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Cited by' column.
        
    Returns
    -------
    float
        HG-index value.
    """
    return hg_index(_get_clean_col(df, "Cited by").tolist())


def c_index_df(df, thresholds=[1, 5, 10, 20, 50, 100, 1000]):
    """
    Compute citation thresholds from a DataFrame.
    """
    citations = _get_clean_col(df, "Cited by").tolist()
    total_docs = len(citations)

    result = {}
    for t in thresholds:
        count = sum(c >= t for c in citations)
        result[f"{t}"] = count
        result[f"{t} [%]"] = round(100 * count / total_docs, 2) if total_docs > 0 else 0.0

    return result


def tapered_h_index_df(df):
    """
    Compute tapered h-index from a DataFrame's citation column.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Cited by' column.
        
    Returns
    -------
    float
        Tapered h-index value.
    """
    return tapered_h_index(_get_clean_col(df, "Cited by").tolist())


def chi_index_df(df):
    """
    Compute chi-index from a DataFrame's citation column.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with 'Cited by' column.
        
    Returns
    -------
    float
        Chi-index value.
    """
    return chi_index(_get_clean_col(df, "Cited by").tolist())

"""Utilities related to Aggregator Functions --- in the Biblium project."""
# --- Aggregator Functions ---

def get_performance_indicators(
    df,
    name = None,
    mode = 'core',
    name_col = 'Name',
    diversity_measure = entropy,
):
    """
    Compute a set of bibliometric performance indicators from a DataFrame.

    Parameters:
        df (pd.DataFrame): Input data containing at least 'Cited by' and optionally 'Year' and 'Cited <field>'.
        name (str, optional): Optional identifier (e.g., author name) to include in results.
        mode (str): One of "core", "extended", or "full" to determine indicator depth.
        name_col (str): Name of the column representing the identifier.
        diversity_measure (function): Function used to compute interdisciplinarity.

    Returns:
        list of tuples: List of (indicator_name, value) pairs.
    """
    indicators = [(name_col, name)] if name else []

    # Core indicators
    indicators += [
        ("Number of documents", len(df)),
        ("Total citations", total_citations(df)),
        ("H-index", h_index_df(df)),
        ("Average year", average_year(df))
    ]

    # Extended indicators
    if mode in ["extended", "full"]:
        g_idx = g_index_df(df)
        c_idx = c_index_df(df)

        indicators += [
            ("G-index", g_idx),
            *[(f"C{c}", c_idx[c]) for c in c_idx],
            ("First year", first_publication_year(df)),
            ("Q1 year", percentile_year(df, 0.25)),
            ("Median year", percentile_year(df, 0.5)),
            ("Q3 year", percentile_year(df, 0.75)),
            ("Last year", last_publication_year(df))
        ]

    # Full mode includes advanced indicators
    if mode == "full" and "Cited by" in df.columns:
        citations = df["Cited by"].tolist()

        indicators += [
            ("Number of cited documents", count_cited_documents(df)),
            ("A-index", a_index(citations)),
            ("R-index", r_index(citations)),
            ("H(2)-index", h2_index(citations)),
            ("W-index", w_index(citations)),
            ("T-index", t_index(citations)),
            ("Pi-index", pi_index(citations)),
            ("Gini index", gini_index(citations)),
            ("HG-index", hg_index(citations)),
            ("Chi-index", chi_index(citations)),
           # ("Tapered H-index", tapered_h_index(citations))
        ]

        # Interdisciplinarity
        fields = [c for c in df.columns if isinstance(c, str) and "Cited" in c and c != "Cited by"]
        if fields:
            # Convert to numeric before summing (handles string columns)
            cited_df = df[fields].apply(pd.to_numeric, errors="coerce").fillna(0)
            cited_fields = cited_df.sum()
            diversity = diversity_measure(cited_fields)
            if diversity_measure == entropy:
                diversity = 0.0 if len(fields) <= 1 or not np.isfinite(diversity) else diversity / np.log(len(fields))
            indicators += [("Interdisciplinarity", diversity)] + list(zip(cited_fields.index, cited_fields))

    ci = None
    if "Author(s) ID" in df.columns:
        ci = collaboration_index(df, "Author(s) ID")
    elif "Author full names" in df.columns:
        try:
            ci = collaboration_index(df, "Author full names")
        except:
            pass
    elif "Authors or Inventors" in df.columns:
        try:
            ci = collaboration_index(df, "Authors or Inventors")
        except:
            pass
    elif "Authors" in df.columns:
        try:
            ci = collaboration_index(df, "Authors")
        except:
            pass
    if ci is not None:
        indicators += [("Collaboration index", ci)]

    rng = last_publication_year(df) - first_publication_year(df) + 1

    indicators += [("Documents per active year", len(df) / rng),
                   ("Citations per active year", total_citations(df) / rng)]

    return indicators

def get_specific_indicators(
    df,
    name = None,
    mode = 'core',
    name_col = 'Name',
    **kwargs,
):
    """
    Compute sentiment-based specific indicators from abstracts.
    
    Parameters
    ----------
    df : pd.DataFrame
        Bibliographic data with abstracts.
    name : str, optional
        Name label for the indicator set.
    mode : str
        Indicator mode ('core' or 'extended').
    name_col : str
        Column name for the name indicator.
    **kwargs
        Additional arguments.
        
    Returns
    -------
    list
        List of (indicator_name, value) tuples.
    """
    indicators = [(name_col, name)] if name else []
    try:
        text_column = next((col for col in ["Processed Abstract", "Abstract"] if col in df.columns), None)
        df, stats_df = analyze_sentiment(df, text_column, sentiment_threshold=0.05, top_words=10)

        indicators += [("Mean sentiment score", np.mean(df["Sentiment Score"])),
        ("Stdev of sentiment score", np.std(df["Sentiment Score"])),
        ("Highest sentiment score", np.max(df["Sentiment Score"])),
        ("Lowest sentiment score", np.min(df["Sentiment Score"])),
         ]
    except:
        pass

    try:
        top5 = df.nlargest(5, "Sentiment Score")["Title"]
        bottom5 = df.nsmallest(5, "Sentiment Score")["Title"]

        # Join into newline-separated strings
        top5_titles = "\n".join(top5.tolist())
        bottom5_titles = "\n".join(bottom5.tolist())

        indicators += [("5 titles of documents with highest sentiment score\n(based on abstracts)", top5_titles),
                       ("5 titles of documents with lowest sentiment score\n(based on abstracts)", bottom5_titles)]
    except:
        pass

    return indicators

"""Helpers for selecting and preparing DataFrames for subsequent performance analyses."""
# selection of dataframe for further performance analysis

def match_items_and_compute_binary_indicators(
    df,
    col,
    items_of_interest,
    value_type = 'string',
    separator = '; ',
    indicators = True,
    missing_as_zero = True,
):
    """
    Match items of interest in a specified dataframe column and optionally compute binary indicators.

    Notes:
        For 'text' matching, values are converted to lowercase to enable substring matching.
        Full-column lowercase transformation is avoided to prevent issues (e.g. with journal names).

    Parameters:
        df (pd.DataFrame): Input dataframe.
        col (str): Column in which to search for matches.
        items_of_interest (list): Items to search for.
        value_type (str): Type of values in column: 'string', 'list', or 'text'.
        separator (str): Separator for splitting list-type entries (used if value_type is 'list').
        indicators (bool): Whether to compute binary indicator columns.
        missing_as_zero (bool): If True, missing indicator values are replaced with 0.

    Returns:
        match_indices (dict): Dictionary mapping each item to a list of matched row indices.
        indicators_dict (dict): Dictionary of indicator DataFrames (empty if indicators=False).
    """

    match_indices = {item: [] for item in items_of_interest}

    for idx, val in df[col].items():
        if pd.isna(val):
            continue
        val_str = str(val).lower()
        if value_type == "string":
            if val in items_of_interest:
                match_indices[val].append(idx)
        elif value_type == "list":
            parts = [v.strip() for v in val.split(separator)]
            for item in items_of_interest:
                if item in parts:
                    match_indices[item].append(idx)
        elif value_type == "text":
            for item in items_of_interest:
                if item.lower() in val_str:
                    match_indices[item].append(idx)

    indicators_dict = {}
    if not indicators:
        return match_indices, indicators_dict

    indicator_01 = pd.DataFrame(index=df.index, columns=items_of_interest, dtype="float")
    # Identify rows with missing values in col
    missing_rows = df[col].isna()

    for item in items_of_interest:
        # Initialize mask: 1.0 for match, 0.0 for no match
        mask = df.index.isin(match_indices[item]).astype(float)
        # Set to NaN for rows with missing values in the relevant column
        mask[missing_rows] = np.nan
        indicator_01[item] = mask
        if missing_as_zero:
            indicator_01[item] = indicator_01[item].fillna(0)
    indicators_dict["binary"] = indicator_01

    return match_indices, indicators_dict

def select_documents(
    df,
    col,
    items_of_interest=None,
    exclude_items=None,
    top_items_df=None,
    top_items_col=None,
    top_items_criterion="Number of documents",
    top_n=20,
    regex_include=None,
    regex_exclude=None,
    indicators=False,
    missing_as_zero=False,
    separator="; ",
    value_type="string",
    text_norm="tfidf",
):
    """
    Select documents containing given entities and optionally compute
    document-level indicator matrices.

    This function is designed to work consistently with ``count_occurrences``:
    any normalization applied there (notably balancing of parentheses in
    list-like values) is mirrored here so that entities with unbalanced
    parentheses (e.g. "Value (missing") are correctly matched and their
    statistics are computed.

    Parameters
    ----------
    df : pd.DataFrame
        Document-level data.
    col : str
        Name of the column in ``df`` containing the entity values.
    items_of_interest : list of str or None, default None
        Explicit list of entity names to select. If None, they are
        derived from ``top_items_df`` and ``top_items_col``.
    exclude_items : list of str or None, default None
        Entity names to exclude.
    top_items_df : pd.DataFrame or None, default None
        Table of entities and their counts (e.g. from ``count_occurrences``).
        Used only when ``items_of_interest`` is None.
    top_items_col : str or None, default None
        Column in ``top_items_df`` holding the entity labels.
    top_items_criterion : str, default "Number of documents"
        Column in ``top_items_df`` used to rank entities when deriving
        the top-N items.
    top_n : int, default 20
        Number of top entities to select when deriving ``items_of_interest``
        from ``top_items_df`` and no regex filters are used.
    regex_include, regex_exclude : str or None
        Optional regex filters applied to the label column of
        ``top_items_df`` when deriving ``items_of_interest``.
    indicators : bool, default False
        If True, also return a dictionary of indicator matrices, such as:
        - "binary" (0/1 presence),
        - "fractional" (for list values),
        - "count"/"tfidf"/"df-icf"/"mtf-idf" (for text values).
    missing_as_zero : bool, default False
        If True, missing indicators are filled with 0.
    separator : str, default "; "
        Separator for list-like values when ``value_type="list"``.
    value_type : {"string", "list", "text"}, default "string"
        How to interpret the values in ``df[col]``.
    text_norm : {"tfidf", "df-icf", "mtf-idf", None}, default "tfidf"
        Normalization scheme for text indicators when ``value_type="text"``.

    Returns
    -------
    match_indices : dict[str, list[int]]
        Mapping from each entity name to a list of row indices in ``df``
        that contain the entity.
    indicators_dict : dict[str, pd.DataFrame]
        Dictionary of indicator DataFrames. Empty if ``indicators=False``.
    """
    if value_type not in {"string", "list", "text"}:
        raise ValueError('value_type must be one of "string", "list", or "text"')
    if text_norm not in {"tfidf", "df-icf", "mtf-idf", None}:
        raise ValueError('text_norm must be one of "tfidf", "df-icf", "mtf-idf", or None')

    def _canonicalize_item(s: str) -> str:
        """Strip whitespace and balance unclosed '('. Matches count_occurrences."""
        s = str(s).strip()
        if not s:
            return ""
        if s.count("(") > s.count(")"):
            s = s + ")"
        return s

    items_of_interest = items_of_interest or []
    exclude_items = exclude_items or []

    # ------------------------------------------------------------------
    # Derive items_of_interest from top_items_df if not provided
    # ------------------------------------------------------------------
    if not items_of_interest:
        if top_items_df is None or top_items_col is None:
            raise ValueError(
                "top_items_df and top_items_col are required when items_of_interest is not provided"
            )

        filtered_df = top_items_df.dropna(subset=[top_items_col])

        if regex_include:
            filtered_df = filtered_df[
                filtered_df[top_items_col]
                .astype(str)
                .str.contains(regex_include, regex=True, na=False)
            ]
        if regex_exclude:
            filtered_df = filtered_df[
                ~filtered_df[top_items_col]
                .astype(str)
                .str.contains(regex_exclude, regex=True, na=False)
            ]

        if regex_include or regex_exclude:
            items_of_interest = filtered_df[top_items_col].astype(str).tolist()
        else:
            df_sorted = filtered_df.sort_values(
                by=top_items_criterion,
                ascending=False,
            )
            if len(df_sorted) <= top_n:
                items_of_interest = df_sorted[top_items_col].astype(str).tolist()
            else:
                cutoff = df_sorted.iloc[top_n - 1][top_items_criterion]
                selected = df_sorted[df_sorted[top_items_criterion] >= cutoff]
                items_of_interest = selected[top_items_col].astype(str).tolist()

    # ------------------------------------------------------------------
    # Canonicalize items and exclusions (important for parentheses)
    # ------------------------------------------------------------------
    items_of_interest = [_canonicalize_item(x) for x in items_of_interest]
    exclude_items = [_canonicalize_item(x) for x in exclude_items]

    items_of_interest = sorted(set(items_of_interest) - set(exclude_items))
    item_set = set(items_of_interest)

    # Initialize match_indices
    match_indices: dict[str, list[int]] = {item: [] for item in items_of_interest}

    # ------------------------------------------------------------------
    # Match documents to entities
    # ------------------------------------------------------------------
    if value_type == "string":
        for idx, val in df[col].items():
            if pd.isna(val):
                continue
            val_canon = _canonicalize_item(val)
            if val_canon in item_set:
                match_indices[val_canon].append(idx)

    elif value_type == "list":
        for idx, val in df[col].items():
            if pd.isna(val):
                continue
            raw_parts = str(val).split(separator)
            parts = []
            for p in raw_parts:
                s = _canonicalize_item(p)
                if s:
                    parts.append(s)
            if not parts:
                continue
            part_set = set(parts)
            for item in items_of_interest:
                if item in part_set:
                    match_indices[item].append(idx)

    elif value_type == "text":
        # Case-insensitive substring search
        lower_items = {item: item.lower() for item in items_of_interest}
        for idx, val in df[col].items():
            if pd.isna(val):
                continue
            val_str = str(val).lower()
            for item, item_lower in lower_items.items():
                if item_lower in val_str:
                    match_indices[item].append(idx)

    # ------------------------------------------------------------------
    # Indicators
    # ------------------------------------------------------------------
    indicators_dict: dict[str, pd.DataFrame] = {}
    if not indicators:
        return match_indices, indicators_dict

    # Binary 0/1 indicators (all value_types)
    indicator_01 = pd.DataFrame(index=df.index, columns=items_of_interest, dtype="float")
    for item in items_of_interest:
        idxs = match_indices.get(item, [])
        indicator_01[item] = df.index.isin(idxs).astype(float)
        if missing_as_zero:
            indicator_01[item] = indicator_01[item].fillna(0.0)
    indicators_dict["binary"] = indicator_01

    # Fractional indicators for list-type values
    if value_type == "list":
        indicator_frac = pd.DataFrame(0.0, index=df.index, columns=items_of_interest)
        for idx, val in df[col].items():
            if pd.isna(val):
                if missing_as_zero:
                    indicator_frac.loc[idx] = 0.0
                continue
            raw_parts = str(val).split(separator)
            parts = []
            for p in raw_parts:
                s = _canonicalize_item(p)
                if s:
                    parts.append(s)
            total = len(parts)
            if total == 0:
                if missing_as_zero:
                    indicator_frac.loc[idx] = 0.0
                continue
            for p in parts:
                if p in item_set:
                    indicator_frac.at[idx, p] += 1.0 / total
        indicators_dict["fractional"] = indicator_frac

    # Text-based counts and normalizations
    if value_type == "text":
        count_df = pd.DataFrame(0.0, index=df.index, columns=items_of_interest)
        lower_items = {item: item.lower() for item in items_of_interest}

        for idx, val in df[col].items():
            if pd.isna(val):
                if missing_as_zero:
                    count_df.loc[idx] = 0.0
                continue
            val_str = str(val).lower()
            for item, item_lower in lower_items.items():
                count_df.at[idx, item] = val_str.count(item_lower)

        indicators_dict["count"] = count_df

        if text_norm == "tfidf":
            tfidf = TfidfTransformer()
            tfidf_data = tfidf.fit_transform(count_df.fillna(0.0))
            tfidf_df = pd.DataFrame(
                tfidf_data.toarray(),
                index=df.index,
                columns=items_of_interest,
            )
            indicators_dict["tfidf"] = tfidf_df

        elif text_norm == "df-icf":
            df_vec = (count_df > 0).sum(axis=0)
            icf = np.log((1 + len(count_df)) / (1 + df_vec)).values
            df_icf_df = count_df.multiply(icf, axis=1)
            indicators_dict["df-icf"] = df_icf_df

        elif text_norm == "mtf-idf":
            mtf = count_df.div(
                count_df.max(axis=1).replace(0, np.nan),
                axis=0,
            )
            df_vec = (count_df > 0).sum(axis=0)
            idf = np.log((1 + len(count_df)) / (1 + df_vec)).values
            mtf_idf_df = mtf.multiply(idf, axis=1)
            indicators_dict["mtf-idf"] = mtf_idf_df

    return match_indices, indicators_dict


def get_entity_stats(
    df,
    entity_col,
    entity_label,
    items_of_interest=None,
    exclude_items=None,
    top_n=20,
    counts_df=None,
    count_method=None,
    regex_include=None,
    regex_exclude=None,
    value_type="string",
    indicators=False,
    missing_as_zero=False,
    mode="full",
    sep="; ",
    openalex_add=False,
    translation=None,
    max_items=0,
):
    """
    Compute performance indicators for selected entities.

    Intended pipeline
    -----------------
    1) Obtain a counts table (counts_df or count_method).
    2) Derive items_of_interest from that table (unless provided), 
       applying limits based on top_n (default) or max_items (for regex).
    3) Use select_documents to find documents containing these entities.
    4) Compute performance indicators per entity.
    5) Merge counts and stats on the *same* label column.

    This guarantees that any entity that is counted is also evaluated in
    stats, and the same logic works for the full dataset and for group
    subsets.

    Parameters
    ----------
    df : pd.DataFrame
        Document-level data.
    entity_col : str
        Column in df containing the entity values (string or list-like).
    entity_label : str
        Name of the entity label column in the stats output (e.g. "Keyword",
        "Author", "Affiliation").
    items_of_interest : list of str or None, default None
        If provided, these are used directly (Mode 2).
    exclude_items : list of str or None, default None
        Entities to exclude.
    top_n : int, default 20
        Number of top entities to select from counts when no regex is used (Mode 1).
    counts_df : pd.DataFrame or None, default None
        Precomputed counts table.
    count_method : callable or None, default None
        Function returning a counts_df when called with no arguments.
    regex_include, regex_exclude : str or None
        Regex filters applied to the label column of counts_df to choose
        entities (Mode 3).
    value_type : {"string", "list", "text"}, default "string"
        How to interpret entity_col (passed to select_documents).
    indicators : bool, default False
        If True, also return a document-level indicator DataFrame.
    missing_as_zero : bool, default False
        Passed through to select_documents.
    mode : str, default "full"
        Passed to get_performance_indicators.
    sep : str, default "; "
        Separator for list-like entity columns (passed to select_documents).
    openalex_add : bool, default False
        If True, attempt to merge OpenAlex metadata on entity_label.
    translation : dict or pd.DataFrame or None, default None
        Optional translation mapping for the final entity_label column.
    max_items : int, default 0
        If positive, this is the maximum number of entities selected 
        when using regex filters (Mode 3). NOTE: This parameter is NOT 
        used to limit the number of documents analysed, but the number of entities.

    Returns
    -------
    stats_df : pd.DataFrame
        Aggregated indicators per entity, with counts merged in (if
        counts_df is provided or derived).
    indicator_df : pd.DataFrame or None
        Document-level indicators if indicators=True, else None.
    """
    import pandas as pd

    # ------------------------------------------------------------------
    # Pre-step: Document truncation (kept if you want to limit the base DF)
    # ------------------------------------------------------------------
    # NOTE: This implementation assumes max_items is intended for ENTITY LIMIT.
    # If you still need to limit the documents in DF for speed, you should rename
    # the max_items parameter used below (e.g., entity_limit) to avoid confusion.
    # We will remove the document truncation logic to follow the last request's spirit.

    # ------------------------------------------------------------------
    # 1. Prepare counts_df if needed
    # ------------------------------------------------------------------
    if counts_df is None and items_of_interest is None and count_method is not None:
        counts_df = count_method()

    label_col = None
    if counts_df is not None and not counts_df.empty:
        counts_df = counts_df.copy()

        # Prefer entity_label column if available; otherwise use first column
        if entity_label in counts_df.columns:
            label_col = entity_label
        else:
            label_col = counts_df.columns[0]

        # Sort counts by Number of documents if present
        if "Number of documents" in counts_df.columns:
            counts_df = counts_df.sort_values("Number of documents", ascending=False)

    # ------------------------------------------------------------------
    # 2. Derive items_of_interest from counts_df (if not given explicitly)
    # ------------------------------------------------------------------
    if items_of_interest is None:
        if counts_df is None or label_col is None:
            raise ValueError(
                "Either items_of_interest must be provided, or counts_df / "
                "count_method must supply a counts table."
            )

        filtered_df = counts_df.dropna(subset=[label_col])

        # Apply regex filters first, if present
        is_regex_mode = (regex_include is not None) or (regex_exclude is not None)
        
        if regex_include:
            filtered_df = filtered_df[
                filtered_df[label_col]
                .astype(str)
                .str.contains(regex_include, regex=True, na=False)
            ]
        if regex_exclude:
            filtered_df = filtered_df[
                ~filtered_df[label_col]
                .astype(str)
                .str.contains(regex_exclude, regex=True, na=False)
            ]

        # NEW LOGIC for applying limits based on mode
        if is_regex_mode:
            # Mode 3: Regex is used. Limit to max_items (if > 0)
            # If max_items is 0 (default), all filtered items are taken.
            limit = max_items if max_items > 0 else len(filtered_df)
            items_of_interest = filtered_df[label_col].astype(str).tolist()[:limit]
        
        else:
            # Mode 1: No regex. Limit to top_n (default 20)
            items_of_interest = filtered_df[label_col].astype(str).tolist()[:top_n]

    # Mode 2: If items_of_interest was provided explicitly, this block is skipped,
    # and all provided items are used.

    # ------------------------------------------------------------------
    # 3. Select documents containing these entities
    # ------------------------------------------------------------------
    selected_entities, indicator_df = select_documents(
        df,
        entity_col,
        value_type=value_type,
        items_of_interest=items_of_interest,
        exclude_items=exclude_items,
        top_items_df=None,
        top_items_col=None,
        top_n=top_n, # NOTE: top_n is redundant here but kept for API consistency
        regex_include=None,
        regex_exclude=None,
        indicators=indicators,
        missing_as_zero=missing_as_zero,
        separator=sep,
    )

    # ------------------------------------------------------------------
    # 4. Compute performance indicators per entity
    # ------------------------------------------------------------------
    stats_list = []
    for name, idx_list in selected_entities.items():
        if not idx_list:
            continue

        # IMPORTANT: use .loc because select_documents returns index labels
        entity_df = df.loc[idx_list]

        metrics = get_performance_indicators(
            entity_df,
            name=name,
            mode=mode,
            name_col=entity_label,
        )
        stats_list.append(dict(metrics))

    if not stats_list:
        # No entities matched; return empty frame with entity_label col
        empty = pd.DataFrame(columns=[entity_label])
        return empty, (indicator_df if indicators else None)

    stats_df = pd.DataFrame(stats_list)

    # Clean missing / empty labels defensively
    if entity_label in stats_df.columns:
        stats_df = stats_df.dropna(subset=[entity_label])
        stats_df = stats_df[
            stats_df[entity_label].astype(str).str.strip() != ""
        ]

    # ------------------------------------------------------------------
    # 5. Merge counts_df (if available) with stats on the same label
    # ------------------------------------------------------------------
    if counts_df is not None and label_col is not None and not counts_df.empty:
        # Ensure the counts label column is named entity_label
        if label_col != entity_label:
            counts_df = counts_df.rename(columns={label_col: entity_label})

        # Inner join: keep only entities for which we actually computed stats
        stats_df = pd.merge(
            counts_df,
            stats_df,
            how="inner",
            on=entity_label,
            suffixes=("", "_stats"),
        )

        if "Number of documents" in stats_df.columns:
            stats_df = stats_df.sort_values(
                "Number of documents",
                ascending=False,
            )

    # Clean again after merge
    if entity_label in stats_df.columns:
        stats_df = stats_df.dropna(subset=[entity_label])
        stats_df = stats_df[
            stats_df[entity_label].astype(str).str.strip() != ""
        ]

    # ------------------------------------------------------------------
    # 6. Optional: merge OpenAlex metadata
    # ------------------------------------------------------------------
    if openalex_add and not stats_df.empty:
        try:
            # Assuming merge_openalex_metadata and related functions are defined elsewhere
            stats_df = merge_openalex_metadata(
                stats_df,
                ref_col=entity_label,
                min_docs=0,
            )
        except Exception:
            # Fail silently if metadata cannot be merged
            pass

    # ------------------------------------------------------------------
    # 7. Optional: translate entity_label (label only, not the key logic)
    # ------------------------------------------------------------------
    if translation is not None and entity_label in stats_df.columns:
        # Build mapping
        if isinstance(translation, dict):
            mapping = {str(k): str(v) for k, v in translation.items()}
        elif isinstance(translation, pd.DataFrame):
            if translation.shape[1] < 2:
                raise ValueError(
                    "translation DataFrame must have at least two columns (key, value)."
                )
            kcol, vcol = translation.columns[:2]
            mapping = dict(
                zip(
                    translation[kcol].astype(str),
                    translation[vcol].astype(str),
                )
            )
        else:
            raise TypeError(
                "translation must be a dict, a pandas DataFrame, or None."
            )

        # Unique rename for original label
        original_col = f"{entity_label} (original)"
        i = 2
        while original_col in stats_df.columns:
            original_col = f"{entity_label} (original {i})"
            i += 1
        stats_df = stats_df.rename(columns={entity_label: original_col})

        # Create translated label under the original name
        base_series = stats_df[original_col].astype(str)
        translated = base_series.map(mapping).fillna(base_series)
        stats_df[entity_label] = translated

        # Optional tidy order: translated label, original label, then the rest
        front = [entity_label, original_col]
        rest = [c for c in stats_df.columns if c not in front]
        stats_df = stats_df[front + rest]

    return stats_df, (indicator_df if indicators else None)

def get_all_performances(
    df,
    name_col,
    items,
    search_mode = 'exact',
    mode = 'core',
):
    """
    Get performance indicators for multiple items in a DataFrame.

    Parameters:
        df (pd.DataFrame): Full dataset.
        name_col (str): Column used to identify individuals/items.
        items (list): List of names/items to retrieve.
        search_mode (str): 'exact' or 'substring'.
        mode (str): One of 'core', 'extended', 'full'.

    Returns:
        pd.DataFrame: Performance summary.
    """
    results = []
    for item in items:
        if search_mode == "exact":
            subset = df[df[name_col] == item]
        elif search_mode == "substring":
            subset = df[df[name_col].astype(str).str.contains(item)]
        else:
            raise ValueError("search_mode must be 'exact' or 'substring'")
        perf = get_performance_indicators(subset, item, mode=mode, name_col=name_col)
        results.append(perf)
    return pd.DataFrame([dict(p) for p in results])

"""Utilities for exporting tables and results to Excel workbooks with basic formatting."""
# Excel saving


def to_excel_fancy(
    data,
    f_name='styled_output.xlsx',
    sheet_names=None,
    top_n=3,
    bottom_n=3,
    top_color='99FF99',
    bottom_color='FF9999',
    autofit=True,
    conditional_formatting=True,
):
    """
    Save one or multiple DataFrames to an Excel file with optional formatting.
    Automatically handles MultiIndex columns by enabling the index export.
    """

    # Ensure data is a list of DataFrames
    if isinstance(data, pd.DataFrame):
        data = [data]

    # Default sheet names if none provided
    if sheet_names is None or not sheet_names:
        sheet_names = [f"Sheet{i+1}" for i in range(len(data))]

    # Ensure correct sheet name length
    if len(sheet_names) != len(data):
        raise ValueError("Number of sheet names must match number of DataFrames.")

    # Define colors
    top_fill = PatternFill(start_color=top_color, end_color=top_color, fill_type="solid")
    bottom_fill = PatternFill(start_color=bottom_color, end_color=bottom_color, fill_type="solid")

    with pd.ExcelWriter(f_name, engine="openpyxl") as writer:
        for df, sheet_name in zip(data, sheet_names):
            
            # -----------------------------------------------------------
            # FIX: Detect MultiIndex columns
            # -----------------------------------------------------------
            is_multiindex = isinstance(df.columns, pd.MultiIndex)
            
            # If MultiIndex, we MUST save the index to allow hierarchical headers 
            # and to preserve the Entity labels (which are in the index).
            save_index = is_multiindex
            
            df.to_excel(writer, sheet_name=sheet_name, index=save_index)
            sheet = writer.sheets[sheet_name]

            # -----------------------------------------------------------
            # Autofit column width
            # -----------------------------------------------------------
            if autofit:
                for col in sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter  # Get column letter
                    for cell in col:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                # Truncate massive strings to avoid Excel errors
                                if len(val_str) > 100: 
                                    val_str = val_str[:100]
                                max_length = max(max_length, len(val_str))
                        except:
                            pass
                    # Add a little padding
                    sheet.column_dimensions[col_letter].width = max_length + 2

            # -----------------------------------------------------------
            # Conditional Formatting
            # -----------------------------------------------------------
            if conditional_formatting:
                # IMPORTANT: If we saved the index, the data columns in Excel 
                # form shifted to the right by the number of index levels.
                col_offset = df.index.nlevels if save_index else 0
                
                # Loop through the DataFrame columns (this skips the index automatically)
                for i, col_name in enumerate(df.columns):
                    
                    # Calculate the actual column index in the Excel sheet (1-based)
                    excel_col_idx = (i + 1) + col_offset

                    if pd.api.types.is_numeric_dtype(df[col_name]):
                        # Determine where data starts (Row 2 for flat, Row 3+ for MultiIndex)
                        # MultiIndex header takes up n_levels + 1 rows usually
                        header_rows = df.columns.nlevels if is_multiindex else 1
                        start_row = header_rows + 1

                        values = [
                            cell for row in sheet.iter_rows(
                                min_col=excel_col_idx, 
                                max_col=excel_col_idx, 
                                min_row=start_row
                            ) 
                            for cell in row
                        ]

                        if values:
                            numeric_values = [
                                (idx, float(cell.value)) 
                                for idx, cell in enumerate(values) 
                                if isinstance(cell.value, (int, float))
                            ]
                            
                            if not numeric_values:
                                continue
                                
                            val_list = [val for _, val in numeric_values]

                            # Get rankings
                            ranks_min = rankdata(val_list, method='min')
                            ranks_max = rankdata(val_list, method='max')

                            min_rank_threshold = bottom_n
                            max_rank_threshold = len(ranks_max) - top_n + 1

                            # Apply formatting
                            for (idx, val), rank_min, rank_max in zip(numeric_values, ranks_min, ranks_max):
                                if rank_min <= min_rank_threshold:
                                    values[idx].fill = bottom_fill
                                if rank_max >= max_rank_threshold:
                                    values[idx].fill = top_fill

    print(f"Saved to {f_name}")

"""Helpers for parsing and managing reference lists and cited-works metadata."""
# References management

def parse_reference(
    ref,
    excluded_sources = None,
):
    """
    Parse a bibliographic reference string into a DataFrame with structured fields.

    Parameters:
        ref (str): The reference string to parse.
        excluded_sources (set, optional): A set of known non-source terms. Defaults to empty set.

    Returns:
        pd.DataFrame: A single-row DataFrame with columns:
                      ['Authors', 'Title', 'Source', 'Volume', 'Pages', 'Year']
    """
    if excluded_sources is None:
        excluded_sources = set()

    # Extract year
    year_match = re.search(r"\((\d{4})\)", ref)
    year = year_match.group(1) if year_match else None
    ref_wo_year = re.sub(r"\(\d{4}\)", "", ref).strip()

    # Extract authors from start
    author_match = re.match(r"^((?:[^,]+?\.,\s?)+)", ref_wo_year)
    authors = author_match.group(1).strip().rstrip(",") if author_match else None

    # Get remaining parts
    remaining = ref_wo_year[len(authors):].lstrip(", ").strip() if authors else ref_wo_year
    parts = [p.strip() for p in remaining.split(",") if p.strip()]

    pages = volume = source = title = None

    # Right-to-left: extract pages and volume
    while parts:
        part = parts[-1]
        if re.search(r"(pp\.\s*)?\d{1,5}-\d{1,5}", part):
            pages = re.search(r"\d{1,5}-\d{1,5}", part).group(0)
            parts.pop()
        elif re.match(r"^\d{1,4}$", part):
            volume = parts.pop()
        else:
            break

    # Source: last valid non-excluded part
    for i in reversed(range(len(parts))):
        candidate = parts[i]
        if candidate not in excluded_sources:
            source = candidate
            parts = parts[:i]
            break

    # Title: whatever remains
    if parts:
        title = ", ".join(parts)

    return pd.DataFrame([{
        "Authors": authors,
        "Title": title,
        "Source": source,
        "Volume": volume,
        "Pages": pages,
        "Year": year
    }])

def parse_references(
    ref_blob,
    excluded_sources = None,
):
    """
    Parse a semicolon-separated string of bibliographic references.

    Parameters:
        ref_blob (str): A long reference string with multiple references separated by semicolons.
        excluded_sources (set, optional): Terms to exclude from source detection.

    Returns:
        pd.DataFrame: Combined DataFrame with parsed information for each reference.
    """
    refs = [r.strip() for r in ref_blob.split(";") if r.strip()]
    parsed_dfs = [parse_reference(ref, excluded_sources) for ref in refs]
    return pd.concat(parsed_dfs, ignore_index=True)

def parse_references_dataframe(
    df,
    excluded_sources = None,
):
    """
    Parse a DataFrame with a 'References' column into structured bibliographic records.

    Parameters:
        df (pd.DataFrame): Input DataFrame with a 'References' column and optional 'Doc ID'.
        excluded_sources (set, optional): Set of known non-source terms to exclude. Defaults to empty set.

    Returns:
        pd.DataFrame: Parsed references with structured columns and 'Document source' as the first column.

    Raises:
        ValueError: If the 'References' column is missing.
    """
    if "References" not in df.columns:
        raise ValueError("Input DataFrame must contain a 'References' column.")

    if excluded_sources is None:
        excluded_sources = set()

    parsed_records = []

    for idx, row in df.iterrows():
        doc_id = row["Doc ID"] if "Doc ID" in df.columns else f"Row_{idx}"
        references = row["References"]

        if pd.isna(references):
            continue

        for ref in references.split(";"):
            ref = ref.strip()
            if ref:
                parsed = parse_references(ref, excluded_sources)
                parsed.insert(0, "Document source", doc_id)  # Make it the first column
                parsed_records.append(parsed)

    if parsed_records:
        return pd.concat(parsed_records, ignore_index=True)
    else:
        return pd.DataFrame(columns=["Document source", "Authors", "Title", "Source", "Volume", "Pages", "Year"])

def summarize_parsed_references(
    df_references,
):
    """
    Summarizes a structured references DataFrame into a compact descriptive format.

    Parameters:
        df_references (pd.DataFrame): Input parsed references with columns:
            - 'Document source'
            - 'Authors'
            - 'Source'
            - 'Year'

    Returns:
        pd.DataFrame: Summary with columns ['Variable', 'Item', 'Value']
    """
    records = []
    var_name = "references stats"

    # Clean and cast Year
    year_series = pd.to_numeric(df_references["Year"], errors="coerce").dropna().astype(int)

    # Distinct sources
    records.append((var_name, "Number of distinct sources", df_references["Source"].dropna().nunique()))

    # Top 10 sources
    top_sources = df_references["Source"].dropna().value_counts().head(10)
    sources_str = "\n".join([f"{s} ({c})" for s, c in top_sources.items()])
    records.append((var_name, "Top 10 cited sources", sources_str))

    # Year stats
    if not year_series.empty:
        records.append((var_name, "Average year of references", round(year_series.mean(), 2)))
        records.append((var_name, "Median year of references", year_series.median()))
        records.append((var_name, "Q1 (25%)", year_series.quantile(0.25)))
        records.append((var_name, "Q3 (75%)", year_series.quantile(0.75)))
        records.append((var_name, "Time span of references", f"{year_series.min()}–{year_series.max()}"))
    else:
        records.extend([
            (var_name, "Average year of references", None),
            (var_name, "Median year of references", None),
            (var_name, "Q1 (25%)", None),
            (var_name, "Q3 (75%)", None),
            (var_name, "Time span of references", None),
        ])

    # Authors - clean before splitting
    cleaned_authors = (
        df_references["Authors"]
        .dropna()
        .str.replace(r"\bet al\.?\b", "", case=False, regex=True)
        .str.replace(r"\s+", " ", regex=True)
    )

    author_list = (
        cleaned_authors
        .str.split(",")
        .explode()
        .str.strip()
        .replace(r"^\.?$", pd.NA, regex=True)
        .dropna()
    )

    records.append((var_name, "Number of distinct cited authors", author_list.nunique()))
    top_authors = author_list.value_counts().head(10)
    authors_str = "\n".join([f"{a} ({c})" for a, c in top_authors.items()])
    records.append((var_name, "Top 10 cited authors", authors_str))

    # References per document
    refs_per_doc = df_references["Document source"].value_counts()
    records.append((var_name, "Average number of references per document", round(refs_per_doc.mean(), 2)))

    return pd.DataFrame(records, columns=["Variable", "Item", "Value"])

def extract_cited_sciences(
    df,
    asjc_map_df,
    asjc_meta_df,
):
    """
    Processes the 'References' column of a DataFrame to extract, enrich, and aggregate science field occurrences
    based on parsed source titles. Returns a DataFrame with counts of cited science fields.

    Parameters:
        df (pd.DataFrame): Input DataFrame with a 'References' column containing reference strings.
        asjc_map_df (pd.DataFrame): Mapping table for ASJC codes.
        asjc_meta_df (pd.DataFrame): Metadata for ASJC codes.

    Returns:
        pd.DataFrame: Aggregated DataFrame with cited science field counts, column names prefixed with 'Cited '.
    """
    all_counts = []

    for _, row in df.iterrows():
        references = row.get("References")

        if pd.notna(references):
            parsed_refs = parse_references(references)
            parsed_refs = parsed_refs.rename(columns={"Source": "Source title"})

            enriched_refs = enrich_bibliometric_data(parsed_refs, asjc_map_df, asjc_meta_df)
            science_counts = count_occurrences(enriched_refs, "Science", count_type="list", sep="; ")
        else:
            science_counts = pd.DataFrame()

        all_counts.append(science_counts)

    aggregated_counts = combine_item_dataframes(all_counts)
    aggregated_counts = aggregated_counts.rename(columns=lambda col: f"Cited {col}")

    return aggregated_counts

def compute_interdisciplinarity_entropy(
    df: pd.DataFrame,
    counting_types: list[str],
) -> pd.DataFrame:
    """
    Computes interdisciplinarity (Shannon entropy) for each row based on column subsets
    defined by different counting types. Adds one column per type with entropy scores.

    Parameters:
    -----------
    df : pd.DataFrame
        The input dataframe where column names contain counting type in square brackets.
        Example: 'Physics [Number of documents]', 'Biology [Proportion of documents]', etc.

    counting_types : list of str
        A list of counting types to compute entropy for. Each string should match the
        text inside brackets exactly, e.g. "Number of documents".

    Returns:
    --------
    pd.DataFrame
        The input dataframe with additional columns:
        'Interdisciplinarity {counting_type}' for each type in counting_types.
        Rows with no data for a given type will have NaN in the corresponding column.
    """

    for ctype in counting_types:
        # Select columns that match the counting type
        pattern = re.compile(fr'\[{re.escape(ctype)}\]')
        selected_cols = [col for col in df.columns if pattern.search(col)]

        if not selected_cols:
            continue

        # Subset and normalize row-wise to get probability distributions
        data_subset = df[selected_cols].fillna(0)
        row_sums = data_subset.sum(axis=1)

        # Normalize only non-zero rows
        prob_dist = data_subset.div(row_sums, axis=0).where(row_sums != 0)

        # Compute entropy: rows with all zero will remain NaN
        entropy_values = prob_dist.apply(lambda row: entropy(row.dropna(), base=2) if row.notna().any() else np.nan, axis=1)

        df[f'Interdisciplinarity {ctype}'] = entropy_values

    df["Interdisciplinarity"] = df["Interdisciplinarity Number of documents"]
    return df

"""Functions for working with OpenAlex references and metadata, including ID extraction and mapping."""
# OpenAlex references managenent

from biblium.readbib import harvest_openalex_links_to_df, _extract_openalex_key

def merge_openalex_metadata(
    stats_df: pd.DataFrame,
    ref_col: str = 'Reference',
    count_col: str = 'Number of documents',
    min_docs: int = 10,
    mailto: Optional[str] = None,
    drop_all_empty_cols: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Merge bibliometric statistics with OpenAlex metadata.

    Parameters
    ----------
    stats_df : pandas.DataFrame
        Source dataframe with columns `ref_col` and `count_col`.
    ref_col : str, default "Reference"
        Column holding OpenAlex links or IDs (works).
    count_col : str, default "Number of documents"
        Numeric column used to filter rows.
    min_docs : int, default 10
        Keep rows where `count_col` is greater than or equal to this threshold.
    mailto : str, optional
        Contact email for OpenAlex polite pool requests.
    drop_all_empty_cols : bool, default True
        If True, drop columns that are entirely empty after merging.

    Returns
    -------
    merged_df : pandas.DataFrame
        Filtered input rows enriched with OpenAlex metadata on matching keys.
    meta_df : pandas.DataFrame
        The harvested OpenAlex metadata table used in the merge.
    """
    # Validate required columns
    missing = [c for c in (ref_col, count_col) if c not in stats_df.columns]
    if missing:
        raise KeyError(f"Missing required columns: {missing}")

    # 1) Filter by threshold (>= min_docs)
    filt = stats_df[count_col] >= min_docs
    df_filt = stats_df.loc[filt].copy()

    if df_filt.empty:
        return df_filt.copy(), pd.DataFrame()

    # 2) Safely extract OpenAlex keys
    def _safe_key(
        x,
    ):
        try:
            return _extract_openalex_key(str(x))[1]
        except Exception:
            return None

    df_filt["_openalex_key"] = df_filt[ref_col].map(_safe_key)
    df_filt = df_filt.dropna(subset=["_openalex_key"])

    if df_filt.empty:
        return df_filt.copy(), pd.DataFrame()

    # 3) Harvest OpenAlex metadata
    refs = df_filt[ref_col].tolist()
    meta_df = harvest_openalex_links_to_df(refs, mailto=mailto)

    # 4) Merge metadata
    merged_df = df_filt.merge(
        meta_df,
        left_on="_openalex_key",
        right_on="openalex_key",
        how="left",
        suffixes=("", "_meta"),
    )

    # 5) Drop entirely empty columns if requested
    if drop_all_empty_cols:
        merged_df = merged_df.dropna(axis=1, how="all")

    return merged_df

"""Utilities for building and storing relations between concepts, such as co-occurrence or contingency tables."""
# Relations between concepts

def compute_relation_matrix(
    df1: pd.DataFrame,
    df2: pd.DataFrame = None,
    normalization: bool = False,
    pmi: bool = False,
    tfidf: bool = False,
    network: bool = False,
    eps: float = 1e-09,
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], 'Union[nx.Graph, None]', pd.DataFrame, pd.DataFrame]:
    """
    Compute a co-occurrence or relation matrix from one or two document-item matrices,
    with optional TF-IDF weighting, PMI transformation, several normalization metrics,
    Fisher's exact test p-values, and optional network output.

    Parameters
    ----------
    df1 : pd.DataFrame
        Document-item matrix 1 (documents as rows, items as columns).
    df2 : pd.DataFrame, optional
        Document-item matrix 2. If None, computes co-occurrence within df1.
    normalization : bool, optional
        If True, compute all supported normalizations (see below).
    pmi : bool, optional
        If True, apply Pointwise Mutual Information (only when df2 is None or df1).
    tfidf : bool, optional
        If True, apply TF-IDF weighting to df1 (and df2 if provided).
    network : bool, optional
        If True, return the result as a NetworkX graph in addition to the matrix.
    eps : float, optional
        Small constant to avoid division or log of zero.

    Returns
    -------
    relation : pd.DataFrame
        Raw relation matrix (counts or weighted counts).
    normalized_matrices : dict[str, pd.DataFrame]
        Dictionary of all computed normalizations (each as DataFrame).
    G : nx.Graph | None
        Optional NetworkX graph if network=True, else None.
    all_measures_df : pd.DataFrame
        Wide-format DataFrame with all normalizations, counts, and conditional proportions.
        Indexed by (row item, column item).
    all_measures_df_T : pd.DataFrame
        Wide-format DataFrame with all normalizations, counts, and conditional proportions.
        Indexed by (col item, row item).

    Notes
    -----
    - All normalizations ('association', 'inclusion', 'salton', 'jaccard', 'equivalence', 'yule_q', 'fisher_p')
      are computed for both square and rectangular matrices. For rectangular (non-square) matrices, Jaccard and Equivalence measure co-occurrence ratio rather than strict similarity.
    - Yule's Q and Fisher's p-value assume binary input; results for non-binary matrices may not be interpretable.
    - The all_measures_df combines all normalizations, raw counts, and conditional proportions (row/col-based) for easy comparison and analysis.
    """

    if tfidf:
        transformer = TfidfTransformer()
        df1 = pd.DataFrame(
            transformer.fit_transform(csr_matrix(df1.values)).toarray(),
            index=df1.index,
            columns=df1.columns
        )
        if df2 is not None:
            df2 = pd.DataFrame(
                transformer.fit_transform(csr_matrix(df2.values)).toarray(),
                index=df2.index,
                columns=df2.columns
            )

    if df2 is None:
        df2 = df1

    relation = df1.T @ df2

    if pmi:
        if not df1.equals(df2):
            raise ValueError("PMI can only be applied to co-occurrence matrices (df2 must be None or df1).")
        total = relation.values.sum()
        Pi = np.diag(relation).astype(float) / total
        Pij = relation / total
        Pi = np.maximum(Pi, eps)
        Pij = np.maximum(Pij, eps)
        PMI = np.log(Pij / (Pi[:, None] * Pi[None, :]))
        PMI[PMI < 0] = 0
        relation = pd.DataFrame(PMI, index=df1.columns, columns=df2.columns)

    normalized_matrices = {}

    if normalization:
        row_sums = np.array(df1.sum(axis=0), dtype=float)
        col_sums = np.array(df2.sum(axis=0), dtype=float)
        R = relation.values

        def safe_df(
            matrix: np.ndarray,
        ) -> pd.DataFrame:
            return pd.DataFrame(matrix, index=df1.columns, columns=df2.columns)

        try:
            normalized_matrices["association"] = safe_df(R / (row_sums[:, None] * col_sums[None, :] + eps))
        except Exception:
            pass
        try:
            normalized_matrices["inclusion"] = safe_df(R / (np.minimum(row_sums[:, None], col_sums[None, :]) + eps))
        except Exception:
            pass
        try:
            normalized_matrices["salton"] = safe_df(R / (np.sqrt(row_sums[:, None] * col_sums[None, :]) + eps))
        except Exception:
            pass
        try:
            n_docs = df1.shape[0]
            a = R
            b = row_sums[:, None] - R
            c = col_sums[None, :] - R
            d = n_docs - (a + b + c)
            denom = (a * d + b * c + eps)
            yule_q = (a * d - b * c) / denom
            normalized_matrices["yule_q"] = safe_df(yule_q)
        except Exception:
            pass
        try:
            fisher_p = np.zeros_like(R, dtype=float)
            for i in range(R.shape[0]):
                for j in range(R.shape[1]):
                    table = np.array([
                        [a[i, j], b[i, j]],
                        [c[i, j], d[i, j]]
                    ])
                    try:
                        _, p = fisher_exact(table, alternative='two-sided')
                    except Exception:
                        p = 1.0
                    fisher_p[i, j] = p
            normalized_matrices["fisher_p"] = safe_df(fisher_p)
        except Exception:
            pass
        try:
            denom = row_sums[:, None] + col_sums[None, :] - R
            normalized_matrices["jaccard"] = safe_df(R / (denom + eps))
        except Exception:
            pass
        try:
            normalized_matrices["equivalence"] = safe_df((R ** 2) / (row_sums[:, None] * col_sums[None, :] + eps))
        except Exception:
            pass
        try:
            count_df = safe_df(R)
            prop_given_row_df = safe_df(R / (row_sums[:, None] + eps))
            prop_given_col_df = safe_df(R / (col_sums[None, :] + eps))
            stacked_dict = normalized_matrices.copy()
            stacked_dict["count"] = count_df
            stacked_dict["prop_given_row"] = prop_given_row_df
            stacked_dict["prop_given_col"] = prop_given_col_df
            measures_list = []
            for measure, df in stacked_dict.items():
                df_ = df.stack().rename(measure)
                measures_list.append(df_)
            all_measures_df = pd.concat(measures_list, axis=1)
            all_measures_df = all_measures_df.reset_index()
            all_measures_df = all_measures_df.set_index(list(all_measures_df.columns[:2]))
        except Exception:
            all_measures_df = None
    else:
        try:
            R = relation.values
            row_sums = np.array(df1.sum(axis=0), dtype=float)
            col_sums = np.array(df2.sum(axis=0), dtype=float)
            def safe_df(
                matrix: np.ndarray,
            ) -> pd.DataFrame:
                return pd.DataFrame(matrix, index=df1.columns, columns=df2.columns)
            count_df = safe_df(R)
            prop_given_row_df = safe_df(R / (row_sums[:, None] + eps))
            prop_given_col_df = safe_df(R / (col_sums[None, :] + eps))
            stacked_dict = {"count": count_df, "prop_given_row": prop_given_row_df, "prop_given_col": prop_given_col_df}
            measures_list = []
            for measure, df in stacked_dict.items():
                df_ = df.stack().rename(measure)
                measures_list.append(df_)
            all_measures_df = pd.concat(measures_list, axis=1)
            all_measures_df = all_measures_df.reset_index()
            all_measures_df = all_measures_df.set_index(list(all_measures_df.columns[:2]))
        except Exception:
            all_measures_df = None

    G = None
    if network:
        G = nx.from_pandas_adjacency(relation) if df1.equals(df2) else nx.from_pandas_adjacency(relation, create_using=nx.DiGraph)

    try:
        if all_measures_df is not None:
            transposed_dict = {}
            for key, df in stacked_dict.items():
                transposed_dict[key] = df.T
            measures_list_T = []
            for measure, df in transposed_dict.items():
                df_ = df.stack().rename(measure)
                measures_list_T.append(df_)
            all_measures_df_T = pd.concat(measures_list_T, axis=1)
            all_measures_df_T = all_measures_df_T.reset_index()
            all_measures_df_T = all_measures_df_T.set_index(list(all_measures_df_T.columns[:2]))
        else:
            all_measures_df_T = None
    except Exception:
        all_measures_df_T = None
    except Exception:
        all_measures_df_T = None

    return relation, normalized_matrices, G, all_measures_df, all_measures_df_T

"""Analysis routines for concept relations, including association strengths, residuals and ranking."""
# Analysis of these relations

def remove_zero_margins(
    df_relation,
):
    """
    Removes rows and columns from the contingency table that have zero marginal sums.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.

    Returns:
        pd.DataFrame: Cleaned table with only non-zero-sum rows and columns.
    """
    df_clean = df_relation.loc[df_relation.sum(axis=1) > 0, df_relation.sum(axis=0) > 0]
    return df_clean

def compute_diversity_metrics(
    df_relation: pd.DataFrame,
    clean_zeros: bool = True,
) -> dict:
    """
    Compute a variety of diversity metrics for a count-based matrix relating two concepts,
    applied both row-wise (axis=1) and column-wise (axis=0).

    Parameters:
        df_relation (pd.DataFrame): A matrix of counts (e.g., Authors × Sources or similar)
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        dict: A dictionary with keys "row_metrics" and "column_metrics",
              each containing a DataFrame of diversity scores.
    """

    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    def shannon_entropy(
        counts,
    ):
        """
        Shannon entropy H = -Σ p * log2(p), robust to pandas ExtensionDtypes/NA.
        """
        # counts is a pandas Series (row/column). Convert to a clean float64 ndarray.
        if hasattr(counts, "to_numpy"):
            vals = counts.to_numpy(dtype="float64", na_value=0.0)
        else:
            vals = np.asarray(counts, dtype="float64")
            vals[~np.isfinite(vals)] = 0.0

        total = vals.sum()
        if total <= 0:
            return 0.0

        p = vals / total
        p = p[p > 0.0]  # avoid log2(0)
        return float(-(p * np.log2(p)).sum())

    def normalized_shannon(
        counts: np.ndarray,
    ) -> float:
        if counts.sum() == 0:
            return 0.0
        H = shannon_entropy(counts)
        return H / np.log2(len(counts)) if len(counts) > 1 else 0.0

    def gini_coefficient(
        counts: np.ndarray,
    ) -> float:
        if counts.sum() == 0:
            return 0.0
        sorted_counts = np.sort(counts)
        n = len(counts)
        cum_x = np.cumsum(sorted_counts)
        return (n + 1 - 2 * np.sum(cum_x) / cum_x[-1]) / n

    def herfindahl_index(
        counts: np.ndarray,
    ) -> float:
        if counts.sum() == 0:
            return 0.0
        proportions = counts / counts.sum()
        return np.sum(proportions**2)

    def simpson_index(
        counts: np.ndarray,
    ) -> float:
        if counts.sum() == 0:
            return 0.0
        proportions = counts / counts.sum()
        return 1 - np.sum(proportions**2)

    def richness(
        counts: np.ndarray,
    ) -> int:
        return np.count_nonzero(counts)

    metric_functions = {
        "Shannon": shannon_entropy,
        "Normalized Shannon": normalized_shannon,
        "Gini": gini_coefficient,
        "HHI": herfindahl_index,
        "Simpson": simpson_index,
        "Richness": richness
    }

    row_metrics = pd.DataFrame(index=df_relation.index)
    column_metrics = pd.DataFrame(index=df_relation.columns)

    for name, func in metric_functions.items():
        row_metrics[name] = df_relation.apply(func, axis=1)
        column_metrics[name] = df_relation.apply(func, axis=0)

    return {
        "row_metrics": row_metrics,
        "column_metrics": column_metrics
    }

def analyze_bipartite_relation(
    df_relation: pd.DataFrame,
    stats: list = None,
    clean_zeros: bool = True,
) -> dict:
    """
    Perform bipartite network analysis on a relation matrix, including projections and metrics.

    Parameters:
        df_relation (pd.DataFrame): Count matrix with rows and columns as distinct concepts.
        stats (list): List of node-level stats to compute on projections.
                      Supported: "degree", "strength", "betweenness", "closeness",
                                 "eigenvector", "pagerank", "clustering", "triangle_count".
                      If None, all are computed.
        clean_zeros (bool): Whether to remove rows and columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "bipartite_graph": B,
            "row_projection": G_row,
            "column_projection": G_col,
            "row_stats": pd.DataFrame,
            "column_stats": pd.DataFrame,
            "bipartite_global": dict,
            "row_global": dict,
            "column_global": dict
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    B = nx.Graph()
    row_nodes = df_relation.index.tolist()
    col_nodes = df_relation.columns.tolist()

    # Add nodes
    B.add_nodes_from(row_nodes, bipartite=0)
    B.add_nodes_from(col_nodes, bipartite=1)

    # Add weighted edges
    for row in row_nodes:
        for col, weight in df_relation.loc[row][df_relation.loc[row] > 0].items():
            B.add_edge(row, col, weight=weight)

    # Projections - handle edge cases where one side has too few nodes
    G_row = None
    G_col = None
    
    # Need at least 1 node on the opposite side for projection to work
    # NetworkX requires: len(nodes_to_project) < len(B)
    if len(col_nodes) >= 1 and len(row_nodes) < len(B.nodes()):
        try:
            G_row = nx.bipartite.weighted_projected_graph(B, row_nodes)
        except Exception as e:
            warnings.warn(f"Could not compute row projection: {e}")
            G_row = nx.Graph()
            G_row.add_nodes_from(row_nodes)
    else:
        G_row = nx.Graph()
        G_row.add_nodes_from(row_nodes)
        
    if len(row_nodes) >= 1 and len(col_nodes) < len(B.nodes()):
        try:
            G_col = nx.bipartite.weighted_projected_graph(B, col_nodes)
        except Exception as e:
            warnings.warn(f"Could not compute column projection: {e}")
            G_col = nx.Graph()
            G_col.add_nodes_from(col_nodes)
    else:
        G_col = nx.Graph()
        G_col.add_nodes_from(col_nodes)

    if stats is None:
        stats = [
            "degree", "strength", "betweenness", "closeness",
            "eigenvector", "pagerank", "clustering", "triangle_count"
        ]

    def compute_node_stats(
        G,
    ):
        data = {}
        if "degree" in stats:
            data["Degree"] = dict(G.degree())
        if "strength" in stats:
            data["Strength"] = dict(G.degree(weight="weight"))
        if "betweenness" in stats:
            data["Betweenness"] = nx.betweenness_centrality(G, weight="weight")
        if "closeness" in stats:
            data["Closeness"] = nx.closeness_centrality(G)
        if "eigenvector" in stats:
            try:
                data["Eigenvector"] = nx.eigenvector_centrality(G, weight="weight", max_iter=500)
            except nx.PowerIterationFailedConvergence:
                data["Eigenvector"] = {node: np.nan for node in G.nodes()}
        if "pagerank" in stats:
            data["PageRank"] = nx.pagerank(G, weight="weight")
        if "clustering" in stats:
            data["Clustering"] = nx.clustering(G, weight="weight")
        if "triangle_count" in stats:
            data["Triangles"] = nx.triangles(G)
        return pd.DataFrame(data)

    def compute_global_stats(
        G,
        weighted: bool = False,
        clustering: bool = False,
    ):
        degrees = dict(G.degree())
        strengths = dict(G.degree(weight="weight")) if weighted else None
        largest_cc = max(nx.connected_components(G), key=len) if not nx.is_connected(G) else G.nodes
        return {
            "Nodes": G.number_of_nodes(),
            "Edges": G.number_of_edges(),
            "Density": nx.density(G),
            "AvgDegree": np.mean(list(degrees.values())) if degrees else 0,
            "AvgStrength": np.mean(list(strengths.values())) if strengths else 0,
            "Components": nx.number_connected_components(G),
            "LargestComponentSize": len(largest_cc),
            "AvgClustering": nx.average_clustering(G, weight="weight") if clustering else None
        }

    return {
        "bipartite_graph": B,
        "row_projection": G_row,
        "column_projection": G_col,
        "row_stats": compute_node_stats(G_row),
        "column_stats": compute_node_stats(G_col),
        "bipartite_global": compute_global_stats(B),
        "row_global": compute_global_stats(G_row, weighted=True, clustering=True),
        "column_global": compute_global_stats(G_col, weighted=True, clustering=True)
    }

"""Clustering helpers that operate on relationship matrices between concepts or entities."""
# clustering on relationship matrix

def cluster_relation_matrix(
    df_relation: pd.DataFrame,
    method: str = 'kmeans',
    axis: int = 0,
    scale: bool = True,
    k_range: tuple = (2, 10),
    n_clusters: int = None,
    linkage_method: str = 'ward',
    return_scores: bool = False,
    clean_zeros: bool = True,
) -> dict:
    """
    Cluster rows or columns of a relation matrix using various clustering methods.

    Parameters:
        df_relation (pd.DataFrame): Matrix of counts between two concepts.
        method (str): Clustering method: "kmeans", "hierarchical", "spectral".
        axis (int): 0 for columns, 1 for rows.
        scale (bool): Whether to standardize data (recommended for KMeans/Spectral).
        k_range (tuple): Range for automatic KMeans cluster selection (if n_clusters not given).
        n_clusters (int): Number of clusters to use (if known; overrides k_range).
        linkage_method (str): Linkage method for hierarchical clustering.
        return_scores (bool): Whether to return silhouette scores (only for KMeans).
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "clusters": pd.Series of cluster labels,
            "n_clusters": number of clusters,
            "silhouette_scores": dict of silhouette scores (if applicable)
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    if axis == 1:
        X = df_relation.values
        labels = df_relation.index
    else:
        X = df_relation.T.values
        labels = df_relation.columns

    if scale and method in ["kmeans", "spectral"]:
        X = StandardScaler().fit_transform(X)

    scores = {}
    clusters = None
    best_k = None

    if method == "kmeans":
        if n_clusters is None:
            for k in range(k_range[0], k_range[1] + 1):
                model = KMeans(n_clusters=k, random_state=42)
                labels_k = model.fit_predict(X)
                if len(np.unique(labels_k)) > 1:
                    scores[k] = silhouette_score(X, labels_k)
            best_k = max(scores, key=scores.get)
            model = KMeans(n_clusters=best_k, random_state=42)
        else:
            best_k = n_clusters
            model = KMeans(n_clusters=best_k, random_state=42)
        final_labels = model.fit_predict(X)
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    elif method == "hierarchical":
        dist = pdist(X)
        Z = linkage(dist, method=linkage_method)
        best_k = n_clusters or 5
        final_labels = fcluster(Z, best_k, criterion="maxclust")
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    elif method == "spectral":
        best_k = n_clusters or 5
        model = SpectralClustering(n_clusters=best_k, affinity="nearest_neighbors", random_state=42)
        final_labels = model.fit_predict(X)
        clusters = pd.Series(final_labels, index=labels, name="Cluster")

    else:
        raise ValueError(f"Unsupported method: {method}")

    result = {
        "clusters": clusters,
        "n_clusters": best_k
    }

    if method == "kmeans" and return_scores:
        result["silhouette_scores"] = scores

    return result

def bicluster_relation_matrix(
    df_relation: pd.DataFrame,
    n_clusters: int = 5,
    scale: bool = True,
    clean_zeros: bool = True,
) -> dict:
    """
    Perform biclustering on a relation matrix using Spectral Co-clustering.

    Parameters:
        df_relation (pd.DataFrame): Matrix of counts between two concepts.
        n_clusters (int): Number of biclusters to form.
        scale (bool): Whether to standardize the matrix before clustering.
        clean_zeros (bool): Whether to remove rows and columns with zero marginal sums. Default is True.

    Returns:
        dict: {
            "model": fitted SpectralCoclustering model,
            "row_clusters": pd.Series with cluster labels for rows,
            "column_clusters": pd.Series with cluster labels for columns
        }
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    n_rows, n_cols = df_relation.shape
    
    # Check if matrix is too small for biclustering
    if n_rows < 2 or n_cols < 2:
        warnings.warn(f"Matrix too small for biclustering ({n_rows}x{n_cols}). Returning empty results.")
        return {
            "model": None,
            "row_clusters": pd.Series(dtype=int),
            "column_clusters": pd.Series(dtype=int)
        }
    
    # Adjust n_clusters if needed (must be <= min(n_rows, n_cols))
    max_clusters = min(n_rows, n_cols)
    actual_clusters = min(n_clusters, max_clusters)
    
    if actual_clusters < n_clusters:
        warnings.warn(f"Reduced n_clusters from {n_clusters} to {actual_clusters} due to matrix size ({n_rows}x{n_cols}).")
    
    if actual_clusters < 2:
        warnings.warn(f"Not enough data for biclustering (need at least 2 clusters, got {actual_clusters}).")
        return {
            "model": None,
            "row_clusters": pd.Series(0, index=df_relation.index, name="RowCluster"),
            "column_clusters": pd.Series(0, index=df_relation.columns, name="ColumnCluster")
        }

    X = df_relation.values
    if scale:
        X = StandardScaler().fit_transform(X)

    model = SpectralCoclustering(n_clusters=actual_clusters, random_state=42)
    model.fit(X)

    row_clusters = pd.Series(model.row_labels_, index=df_relation.index, name="RowCluster")
    col_clusters = pd.Series(model.column_labels_, index=df_relation.columns, name="ColumnCluster")

    return {
        "model": model,
        "row_clusters": row_clusters,
        "column_clusters": col_clusters
    }


def cluster_entities(
    df: pd.DataFrame,
    entity_column: str,
    *,
    method: str = 'kmeans',
    n_clusters: Optional[int] = None,
    k_range: range = range(2, 11),
    features: str = 'cooccurrence',
    feature_column: Optional[str] = None,
    sep: str = '; ',
    min_occurrence: int = 2,
    scorer: str = 'silhouette',
    linkage_method: str = 'ward',
    random_state: int = 42,
) -> Dict[str, Any]:
    """
    Cluster entities (authors, sources, keywords, etc.) based on their co-occurrence patterns.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input bibliometric dataframe.
    entity_column : str
        Column containing the entities to cluster (e.g., "Authors", "Author Keywords", "Source title").
    method : str, default="kmeans"
        Clustering method: "kmeans", "hierarchical", "spectral".
    n_clusters : int or None, default=None
        Number of clusters. If None:
        - kmeans: auto-select using silhouette score over k_range
        - hierarchical/spectral: defaults to 5
    k_range : range, default=range(2, 11)
        Range of k values to try for auto-selection (kmeans only).
    features : str, default="cooccurrence"
        Feature type for clustering:
        - "cooccurrence": entities co-occurring in same documents
        - "documents": which documents contain each entity (binary)
        - "citation": co-citation patterns (requires citation data)
    feature_column : str or None, default=None
        For "cooccurrence", optionally specify another column to measure co-occurrence with.
        If None, uses self-cooccurrence within entity_column.
    sep : str, default="; "
        Separator for splitting multi-value fields.
    min_occurrence : int, default=2
        Minimum occurrences for an entity to be included.
    scorer : str, default="silhouette"
        Scoring metric for auto k-selection: "silhouette" or "calinski".
    linkage_method : str, default="ward"
        Linkage method for hierarchical clustering.
    random_state : int, default=42
        Random seed for reproducibility.
    
    Returns
    -------
    dict with keys:
        - "clusters_df": DataFrame with entity names and cluster assignments
        - "n_clusters": number of clusters used
        - "silhouette_score": overall silhouette score (if applicable)
        - "cluster_sizes": Series with cluster sizes
        - "feature_matrix": the feature matrix used for clustering
        - "top_entities_by_cluster": dict mapping cluster_id to top entities
    """
    from scipy import sparse
    from sklearn.preprocessing import normalize
    
    # Extract entities from column
    def extract_entities(series, separator):
        """Extract all entities from a multi-value column."""
        entities = []
        for val in series.dropna():
            if isinstance(val, str):
                entities.extend([x.strip() for x in val.split(separator) if x.strip()])
            elif isinstance(val, (list, tuple)):
                entities.extend([str(x).strip() for x in val if str(x).strip()])
        return entities
    
    if entity_column not in df.columns:
        raise ValueError(f"Column '{entity_column}' not found in dataframe.")
    
    # Count entity occurrences
    all_entities = extract_entities(df[entity_column], sep)
    entity_counts = Counter(all_entities)
    
    # Filter by minimum occurrence
    valid_entities = [e for e, c in entity_counts.items() if c >= min_occurrence]
    
    if len(valid_entities) < 3:
        raise ValueError(f"Not enough entities with >= {min_occurrence} occurrences. Found {len(valid_entities)}.")
    
    # Build feature matrix based on feature type
    if features == "cooccurrence":
        # Build co-occurrence matrix
        cooc_col = feature_column if feature_column and feature_column in df.columns else entity_column
        
        # Create entity-to-documents mapping
        entity_docs = {e: set() for e in valid_entities}
        for doc_idx, row in df.iterrows():
            val = row.get(entity_column)
            if pd.isna(val):
                continue
            if isinstance(val, str):
                doc_entities = [x.strip() for x in val.split(sep) if x.strip()]
            elif isinstance(val, (list, tuple)):
                doc_entities = [str(x).strip() for x in val if str(x).strip()]
            else:
                continue
            for e in doc_entities:
                if e in entity_docs:
                    entity_docs[e].add(doc_idx)
        
        # Build co-occurrence matrix
        n_entities = len(valid_entities)
        entity_to_idx = {e: i for i, e in enumerate(valid_entities)}
        
        # Use sparse matrix for efficiency
        from scipy.sparse import lil_matrix
        cooc_matrix = lil_matrix((n_entities, n_entities), dtype=np.float32)
        
        for doc_idx, row in df.iterrows():
            val = row.get(cooc_col)
            if pd.isna(val):
                continue
            if isinstance(val, str):
                doc_entities = [x.strip() for x in val.split(sep) if x.strip()]
            elif isinstance(val, (list, tuple)):
                doc_entities = [str(x).strip() for x in val if str(x).strip()]
            else:
                continue
            
            # Filter to valid entities
            doc_valid = [e for e in doc_entities if e in entity_to_idx]
            
            # Add co-occurrences
            for i, e1 in enumerate(doc_valid):
                for e2 in doc_valid[i:]:
                    idx1, idx2 = entity_to_idx[e1], entity_to_idx[e2]
                    cooc_matrix[idx1, idx2] += 1
                    if idx1 != idx2:
                        cooc_matrix[idx2, idx1] += 1
        
        feature_matrix = cooc_matrix.tocsr()
        
    elif features == "documents":
        # Binary document-entity matrix
        from sklearn.preprocessing import MultiLabelBinarizer
        
        # Get entities per document
        doc_entities_list = []
        for _, row in df.iterrows():
            val = row.get(entity_column)
            if pd.isna(val):
                doc_entities_list.append([])
                continue
            if isinstance(val, str):
                ents = [x.strip() for x in val.split(sep) if x.strip() and x.strip() in valid_entities]
            elif isinstance(val, (list, tuple)):
                ents = [str(x).strip() for x in val if str(x).strip() and str(x).strip() in valid_entities]
            else:
                ents = []
            doc_entities_list.append(ents)
        
        mlb = MultiLabelBinarizer(classes=valid_entities, sparse_output=True)
        doc_entity_matrix = mlb.fit_transform(doc_entities_list)
        
        # Transpose to get entity-document matrix
        feature_matrix = doc_entity_matrix.T.tocsr()
        
    else:
        raise ValueError(f"Unknown features type: {features}. Use 'cooccurrence' or 'documents'.")
    
    # Normalize features
    feature_matrix_norm = normalize(feature_matrix, norm='l2', axis=1)
    
    # Auto-select k for kmeans
    def auto_select_k(X, k_range, scorer_name):
        from sklearn.cluster import KMeans
        from sklearn.metrics import silhouette_score, calinski_harabasz_score
        
        best_k, best_score = None, -np.inf
        X_dense = X.toarray() if sparse.issparse(X) else X
        
        for k in k_range:
            if k >= X.shape[0]:
                continue
            try:
                km = KMeans(n_clusters=k, n_init='auto', random_state=random_state)
                labels = km.fit_predict(X_dense)
                if len(np.unique(labels)) < 2:
                    continue
                if scorer_name == "calinski":
                    score = calinski_harabasz_score(X_dense, labels)
                else:
                    score = silhouette_score(X_dense, labels)
                if score > best_score:
                    best_k, best_score = k, score
            except:
                continue
        return best_k or min(5, len(valid_entities) - 1)
    
    # Perform clustering
    X = feature_matrix_norm.toarray() if sparse.issparse(feature_matrix_norm) else feature_matrix_norm
    
    if method == "kmeans":
        from sklearn.cluster import KMeans
        
        if n_clusters is None:
            n_clusters = auto_select_k(feature_matrix_norm, k_range, scorer)
        
        model = KMeans(n_clusters=n_clusters, n_init='auto', random_state=random_state)
        labels = model.fit_predict(X)
        
    elif method == "hierarchical":
        from sklearn.cluster import AgglomerativeClustering
        
        if n_clusters is None:
            n_clusters = min(5, len(valid_entities) - 1)
        
        model = AgglomerativeClustering(n_clusters=n_clusters, linkage=linkage_method)
        labels = model.fit_predict(X)
        
    elif method == "spectral":
        from sklearn.cluster import SpectralClustering
        
        if n_clusters is None:
            n_clusters = min(5, len(valid_entities) - 1)
        
        model = SpectralClustering(
            n_clusters=n_clusters, 
            affinity='nearest_neighbors',
            random_state=random_state
        )
        labels = model.fit_predict(X)
        
    else:
        raise ValueError(f"Unknown method: {method}. Use 'kmeans', 'hierarchical', or 'spectral'.")
    
    # Create results DataFrame
    clusters_df = pd.DataFrame({
        'Entity': valid_entities,
        'Cluster': labels + 1,  # 1-based
        'Occurrences': [entity_counts[e] for e in valid_entities]
    }).sort_values(['Cluster', 'Occurrences'], ascending=[True, False])
    
    # Calculate silhouette score
    try:
        from sklearn.metrics import silhouette_score
        sil_score = silhouette_score(X, labels)
    except:
        sil_score = None
    
    # Cluster sizes
    cluster_sizes = clusters_df['Cluster'].value_counts().sort_index()
    
    # Top entities by cluster
    top_by_cluster = {}
    for cluster_id in sorted(clusters_df['Cluster'].unique()):
        cluster_entities = clusters_df[clusters_df['Cluster'] == cluster_id].nlargest(10, 'Occurrences')
        top_by_cluster[cluster_id] = cluster_entities['Entity'].tolist()
    
    # Create feature matrix DataFrame for export
    feature_df = pd.DataFrame(
        feature_matrix.toarray() if sparse.issparse(feature_matrix) else feature_matrix,
        index=valid_entities,
        columns=valid_entities if features == "cooccurrence" else [f"Doc_{i}" for i in range(feature_matrix.shape[1])]
    )
    
    return {
        "clusters_df": clusters_df,
        "n_clusters": n_clusters,
        "silhouette_score": sil_score,
        "cluster_sizes": cluster_sizes,
        "feature_matrix": feature_df,
        "top_entities_by_cluster": top_by_cluster,
        "method": method,
        "entity_column": entity_column,
    }


def compute_correspondence_analysis(
    df_relation,
    n_components = 2,
    clean_zeros = True,
):
    """
    Robust CA with stable dtype handling and version-agnostic inertia.

    Why this patch?
    ----------------
    Some `prince` versions do not expose `CA.explained_inertia_`, causing:
        AttributeError: 'CA' object has no attribute 'explained_inertia_'
    We (a) coerce the input to float64 to satisfy SciPy, and
    (b) compute explained inertia ourselves via the standard CA SVD if
        the attribute is missing.

    Parameters
    ----------
    df_relation : pd.DataFrame
        Nonnegative contingency table (rows × columns).
    n_components : int
        Number of CA dimensions to return.
    clean_zeros : bool
        Drop all-zero rows/cols before fitting (recommended).

    Returns
    -------
    row_coords_full : pd.DataFrame
        Row coordinates for all original rows (zeros where dropped).
    col_coords_full : pd.DataFrame
        Column coordinates for all original columns (zeros where dropped).
    explained_inertia : list[float]
        Fraction of inertia explained per component (length ≥ n_components).
    """
    import numpy as np
    import pandas as pd
    import prince

    # 1) Force numeric float64 with NA→0
    X = (
        df_relation.copy()
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0.0)
        .astype("float64")
    )

    # 2) Optionally remove all-zero margins
    if clean_zeros:
        row_mask = X.sum(axis=1) > 0
        col_mask = X.sum(axis=0) > 0
        X_clean = X.loc[row_mask, col_mask]
    else:
        X_clean = X

    # Degenerate case: nothing left to analyze
    if X_clean.shape[0] == 0 or X_clean.shape[1] == 0:
        dims = [f"Dim {i+1}" for i in range(n_components)]
        return (
            pd.DataFrame(0.0, index=X.index, columns=dims),
            pd.DataFrame(0.0, index=X.columns, columns=dims),
            [0.0] * n_components,
        )

    # Adjust n_components based on matrix size (must be >= 1 and < min(rows, cols))
    max_components = min(X_clean.shape[0], X_clean.shape[1]) - 1
    if max_components < 1:
        # Matrix too small for CA
        dims = [f"Dim {i+1}" for i in range(n_components)]
        return (
            pd.DataFrame(0.0, index=X.index, columns=dims),
            pd.DataFrame(0.0, index=X.columns, columns=dims),
            [0.0] * n_components,
        )
    actual_components = min(n_components, max_components)

    # 3) Fit CA on the cleaned matrix
    ca = prince.CA(n_components=actual_components, n_iter=10, engine="sklearn", random_state=42).fit(X_clean)

    row_coords = ca.row_coordinates(X_clean)
    col_coords = ca.column_coordinates(X_clean)

    # 4) Inertia: try attribute; if missing, compute from SVD definition
    inertia = None
    try:
        # Newer prince versions
        inertia_attr = getattr(ca, "explained_inertia_", None)
        if inertia_attr is not None:
            inertia = list(np.asarray(inertia_attr, dtype="float64"))
    except Exception:
        inertia = None

    if inertia is None:
        # Compute explained inertia from correspondence matrix SVD
        # P = X / grand_total
        grand_total = float(X_clean.values.sum())
        if grand_total <= 0:
            inertia = [0.0] * n_components
        else:
            P = X_clean / grand_total
            r = P.sum(axis=1).to_numpy(dtype="float64")
            c = P.sum(axis=0).to_numpy(dtype="float64")
            # Safe since all-zero margins were removed
            Z = (P.values - np.outer(r, c)) / (np.sqrt(r)[:, None] * np.sqrt(c)[None, :])
            s = np.linalg.svd(Z, compute_uv=False)
            eig = s**2
            denom = float(eig.sum()) if eig.size else 0.0
            if denom > 0.0:
                inertia = (eig / denom).tolist()
            else:
                inertia = [0.0] * n_components

    # Pad inertia to requested n_components (even if we computed fewer)
    if inertia:
        inertia = inertia[:n_components]
        # Pad with zeros if we got fewer components than requested
        while len(inertia) < n_components:
            inertia.append(0.0)
    else:
        inertia = [0.0] * n_components

    # 5) Re-expand to original indices (zeros for dropped rows/cols)
    # Use requested n_components for output dimensions (pad with zeros if needed)
    dims = [f"Dim {i}" for i in range(n_components)]
    row_full = pd.DataFrame(0.0, index=X.index, columns=dims)
    col_full = pd.DataFrame(0.0, index=X.columns, columns=dims)
    
    # Copy computed coordinates (may be fewer than n_components)
    actual_dims = row_coords.shape[1] if row_coords is not None else 0
    if actual_dims > 0:
        for i in range(min(actual_dims, n_components)):
            # Use .loc with both row and column indexers to avoid chained assignment
            row_full.loc[row_coords.index, dims[i]] = row_coords.iloc[:, i].values
            col_full.loc[col_coords.index, dims[i]] = col_coords.iloc[:, i].values

    return row_full, col_full, inertia

def extract_sorted_residual_pairs(
    df_relation,
    clean_zeros = True,
):
    """
    Compute standardized Pearson residuals, expected counts, chi², and DoF
    in a way that is robust to pandas extension/object dtypes.

    Fixes
    -----
    Avoids:
        TypeError: loop of ufunc does not support argument 0 of type float which has no callable sqrt method
    by converting inputs to plain float64 ndarrays before using NumPy ufuncs.

    Parameters
    ----------
    df_relation : pd.DataFrame
        Contingency/association table (nonnegative).
    clean_zeros : bool, default True
        If True, drop all-zero rows/columns for computations, then re-expand
        results back to original shape (filling with 0 where dropped).

    Returns
    -------
    residuals_df : pd.DataFrame
        Standardized Pearson residuals (same index/columns as df_relation).
    sorted_resid : list[tuple]
        List of (row_label, col_label, residual, observed, expected), sorted by
        |residual| descending, based on the non-all-zero submatrix used.
    expected_df : pd.DataFrame
        Expected counts under independence (same index/columns as df_relation).
    chi2_stat : float
        Pearson chi-square statistic (on the cleaned submatrix).
    dof : int
        Degrees of freedom (r-1)*(c-1) for the cleaned submatrix.
    """
    import numpy as np
    import pandas as pd

    # 1) Coerce to numeric float64 (NA→0) to make NumPy ufuncs happy
    X = (
        df_relation.copy()
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0.0)
        .astype("float64")
    )

    # 2) Optionally remove all-zero margins for stable stats
    if clean_zeros:
        row_mask = X.sum(axis=1) > 0
        col_mask = X.sum(axis=0) > 0
        Xc = X.loc[row_mask, col_mask]
    else:
        row_mask = pd.Series(True, index=X.index)
        col_mask = pd.Series(True, index=X.columns)
        Xc = X

    # Degenerate case: nothing to compute
    if Xc.shape[0] == 0 or Xc.shape[1] == 0:
        zeros_resid = pd.DataFrame(0.0, index=X.index, columns=X.columns)
        zeros_exp = zeros_resid.copy()
        return zeros_resid, [], zeros_exp, 0.0, 0

    # 3) Observed/expected using ndarrays to avoid object dtypes
    obs = Xc.to_numpy(dtype="float64", na_value=0.0)
    row_tot = obs.sum(axis=1, keepdims=True)
    col_tot = obs.sum(axis=0, keepdims=True)
    grand = float(obs.sum())

    if grand <= 0.0:
        zeros_resid = pd.DataFrame(0.0, index=X.index, columns=X.columns)
        zeros_exp = zeros_resid.copy()
        return zeros_resid, [], zeros_exp, 0.0, 0

    exp = (row_tot @ col_tot) / grand  # outer product / N

    # 4) Standardized Pearson residuals
    with np.errstate(divide="ignore", invalid="ignore"):
        denom = np.sqrt(exp)
        resid = np.divide(obs - exp, denom, out=np.zeros_like(obs), where=denom > 0)

    # 5) Chi-square statistic and DoF on cleaned matrix
    with np.errstate(divide="ignore", invalid="ignore"):
        chi2_terms = np.divide((obs - exp) ** 2, exp, out=np.zeros_like(obs), where=exp > 0)
    chi2_stat = float(chi2_terms.sum())
    dof = max((obs.shape[0] - 1) * (obs.shape[1] - 1), 0)

    # 6) Wrap to DataFrames in cleaned space
    resid_df_c = pd.DataFrame(resid, index=Xc.index, columns=Xc.columns)
    exp_df_c = pd.DataFrame(exp, index=Xc.index, columns=Xc.columns)

    # 7) Re-expand to original shape (fill dropped rows/cols with 0)
    resid_full = pd.DataFrame(0.0, index=X.index, columns=X.columns)
    exp_full = pd.DataFrame(0.0, index=X.index, columns=X.columns)
    resid_full.loc[resid_df_c.index, resid_df_c.columns] = resid_df_c
    exp_full.loc[exp_df_c.index, exp_df_c.columns] = exp_df_c

    # 8) Sorted residual pairs from the cleaned submatrix
    #    (keep it lean; consumer can truncate if needed)
    sorted_resid = sorted(
        (
            (ri, cj, float(resid_df_c.loc[ri, cj]), float(Xc.loc[ri, cj]), float(exp_df_c.loc[ri, cj]))
            for ri in resid_df_c.index
            for cj in resid_df_c.columns
        ),
        key=lambda t: abs(t[2]),
        reverse=True,
    )

    return resid_full, sorted_resid, exp_full, chi2_stat, dof

def compute_svd_statistics(
    df_relation,
    n_components = 2,
    clean_zeros = True,
):
    """
    Applies Truncated SVD to a normalized contingency table.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        n_components (int): Number of SVD components to retain.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums. Default is True.

    Returns:
        row_projection (pd.DataFrame): Row projections onto components.
        singular_values (np.ndarray): Singular values of components.
        explained_variance (np.ndarray): Explained variance ratio per component.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    # Check if matrix is too small for SVD
    n_rows, n_cols = df_relation.shape
    if n_rows < 2 or n_cols < 2:
        # Return empty/minimal results for matrices that are too small
        import warnings
        warnings.warn(f"Matrix too small for SVD ({n_rows}x{n_cols}). Returning empty projections.")
        row_projection = pd.DataFrame(
            np.zeros((n_rows, min(n_components, max(1, min(n_rows, n_cols) - 1)))),
            index=df_relation.index,
            columns=[f'Comp {i+1}' for i in range(min(n_components, max(1, min(n_rows, n_cols) - 1)))]
        )
        return row_projection, np.array([]), np.array([])

    # Adjust n_components if necessary
    max_components = min(n_rows, n_cols) - 1
    actual_components = min(n_components, max_components)
    if actual_components < 1:
        actual_components = 1
    
    # Normalize (relative frequency)
    normed = df_relation / df_relation.values.sum()

    # Apply SVD
    svd_model = TruncatedSVD(n_components=actual_components, random_state=42)
    row_proj = svd_model.fit_transform(normed)

    # Format output
    row_projection = pd.DataFrame(row_proj, index=df_relation.index,
                                  columns=[f'Comp {i+1}' for i in range(actual_components)])
    singular_values = svd_model.singular_values_
    explained_variance = svd_model.explained_variance_ratio_

    return row_projection, singular_values, explained_variance

def compute_log_ratio(
    df_relation,
    clean_zeros: bool = True,
):
    """
    Computes log(observed / expected) values from a contingency table.

    Parameters:
        df_relation (pd.DataFrame): Contingency table.
        clean_zeros (bool): Whether to remove rows/columns with zero marginal sums.

    Returns:
        log_ratio_df (pd.DataFrame): Matrix of log(observed / expected) values.
        expected_df (pd.DataFrame): Expected frequency matrix under independence.
        sorted_log_ratios (pd.DataFrame): Flattened DataFrame with ['Row', 'Column', 'LogRatio'],
                                          sorted by LogRatio descending.
    """
    if clean_zeros:
        df_relation = remove_zero_margins(df_relation)

    observed = df_relation.values
    row_totals = observed.sum(axis=1, keepdims=True)
    col_totals = observed.sum(axis=0, keepdims=True)
    total = observed.sum()
    expected = row_totals @ col_totals / total

    log_ratio = np.log((observed + 1e-6) / (expected + 1e-6))
    log_ratio_df = pd.DataFrame(log_ratio, index=df_relation.index, columns=df_relation.columns)
    expected_df = pd.DataFrame(expected, index=df_relation.index, columns=df_relation.columns)

    # Flatten and sort the log-ratio values
    flattened = [
        (row_label, col_label, log_ratio[i, j])
        for i, row_label in enumerate(df_relation.index)
        for j, col_label in enumerate(df_relation.columns)
    ]
    sorted_log_ratios = pd.DataFrame(flattened, columns=['Row', 'Column', 'LogRatio'])
    sorted_log_ratios = sorted_log_ratios.sort_values(by='LogRatio', ascending=False).reset_index(drop=True)

    return log_ratio_df, expected_df, sorted_log_ratios

"""Legacy relation-utility functions imported from relate_utils.py and used in network and association analyses."""
# relate_utils.py

from dataclasses import dataclass

@dataclass
class Relation:
    """
    Container for the relation between two concepts with optional analyses attached.

    Attributes
    ----------
    concept1, concept2 : str
        Concept identifiers.
    rm : pd.DataFrame | None
        Relation matrix (rows from concept1, cols from concept2).
    Other attributes
        Populated based on `include_stats` (diversity, bipartite, cluster, etc.).
    """
    concept1: str
    concept2: str
    rm: Optional[pd.DataFrame] = None

    # Diversity
    diversity_row_metrics: Optional[pd.DataFrame] = None
    diversity_column_metrics: Optional[pd.DataFrame] = None

    # Bipartite
    bipartite_graph: Any = None
    bipartite_row_projection: Any = None
    bipartite_column_projection: Any = None
    bipartite_row_stats: Optional[pd.DataFrame] = None
    bipartite_column_stats: Optional[pd.DataFrame] = None
    bipartite_global_stats: Optional[Dict[str, Any]] = None
    bipartite_row_global: Optional[Dict[str, Any]] = None
    bipartite_column_global: Optional[Dict[str, Any]] = None

    # Clustering
    clusters: Optional[pd.Series] = None
    n_clusters: Optional[int] = None
    silhouette_scores: Optional[pd.Series] = None

    # Biclustering
    bicluster_model: Any = None
    bicluster_row_clusters: Optional[pd.Series] = None
    bicluster_column_clusters: Optional[pd.Series] = None

    # Correspondence analysis
    ca_row_coords: Optional[pd.DataFrame] = None
    ca_col_coords: Optional[pd.DataFrame] = None
    ca_explained_inertia: Optional[Iterable[float]] = None

    # Chi-square
    chi2_residuals_df: Optional[pd.DataFrame] = None
    chi2_sorted_residuals: Optional[pd.DataFrame] = None
    chi2_expected_df: Optional[pd.DataFrame] = None
    chi2_chi2_stat: Optional[float] = None
    chi2_dof: Optional[int] = None

    # SVD
    svd_row_projection: Optional[pd.DataFrame] = None
    svd_singular_values: Optional[Iterable[float]] = None
    svd_explained_variance: Optional[Iterable[float]] = None

    # Log-ratio
    log_ratio_df: Optional[pd.DataFrame] = None
    log_ratio_expected_df: Optional[pd.DataFrame] = None
    log_ratio_sorted_log_ratios: Optional[pd.DataFrame] = None

    def set_matrix(
        self,
        rm: pd.DataFrame,
    ) -> None:
        "Attach the relation matrix."
        self.rm = rm

    def link_to_self(
        self,
        obj: Any,
        prefix: str = 'Relations_',
    ) -> None:
        """
        Attach public attributes of this Relation onto `obj` as
        f\"{prefix}{concept1}_{concept2}_{attr}\", replacing spaces with underscores.
        """
        c1 = self.concept1.replace(" ", "_")
        c2 = self.concept2.replace(" ", "_")
        for att, val in self.__dict__.items():
            if att.startswith("_"):
                continue
            setattr(obj, f"{prefix}{c1}_{c2}_{att}", val)

def _resolve_matrix(
    concept: str,
    known_matrices: Mapping[str, pd.DataFrame] | None,
    custom_matrices: Mapping[str, pd.DataFrame] | None,
) -> pd.DataFrame:
    """
    Return the binary matrix for `concept` from `known_matrices` or `custom_matrices`.
    Raises ValueError if not found.
    """
    if known_matrices and concept in known_matrices:
        return known_matrices[concept]
    if custom_matrices and concept in custom_matrices:
        return custom_matrices[concept]
    raise ValueError(f"No binary matrix available for concept: {concept}")

def relate_concepts_general(
    concept1: str,
    concept2: str,
    *,
    known_matrices: Mapping[str, pd.DataFrame] | None = None,
    custom_matrices: Mapping[str, pd.DataFrame] | None = None,
    rename_maps: Mapping[str, Mapping[str, str]] | None = None,
    include_stats: Iterable[str] = ('diversity', 'bipartite network', 'bicluster', 'correspondence', 'chi2', 'svd', 'log-ratio'),
    clean_zeros: bool = True,
    relation_matrices: Optional[Dict[str, Dict[str, pd.DataFrame]]] = None,
    relations: Optional[Dict[str, Dict[str, Relation]]] = None,
    to_self: bool = False,
    self_obj: Any = None,
    link_prefix: str = 'Relations_',
    **kwargs,
) -> Tuple[Relation, pd.DataFrame]:
    """
    Compute a relation matrix for (concept1, concept2), run selected analyses,
    and optionally store/link results—without needing a class-bound `self`.

    Parameters
    ----------
    concept1, concept2 : str
        Keys to pick matrices from `known_matrices` or `custom_matrices`.
    known_matrices, custom_matrices : Mapping[str, DataFrame] | None
        Sources of binary matrices.
    rename_maps : Mapping[str, Mapping[str, str]] | None
        Optional renamers per concept. Applied to rm.index when matching concept1,
        and to rm.columns when matching concept2.
    include_stats : Iterable[str]
        Any of {"diversity","bipartite network","cluster","bicluster",
        "correspondence","chi2","svd","log-ratio"}.
    clean_zeros : bool
        Drop all-zero rows/cols in rm before analyses.
    relation_matrices, relations : dict or None
        If given, they are populated to mirror your original side-effects.
    to_self : bool
        If True, attach results to `self_obj` with `link_prefix`.
    self_obj : Any
        Target for attribute linking.
    link_prefix : str
        Name prefix for attributes attached via `link_to_self`.
    **kwargs
        Forwarded to specific utils (e.g., clustering params).

    Returns
    -------
    (Relation, pd.DataFrame)
        The Relation object and the processed relation matrix.
    """
    # 1) Resolve inputs
    df1 = _resolve_matrix(concept1, known_matrices, custom_matrices)
    df2 = _resolve_matrix(concept2, known_matrices, custom_matrices)

    # 2) Relation matrix
    rm = compute_relation_matrix(df1, df2)[0]

    # 3) Optional renaming
    if rename_maps:
        if concept1 in rename_maps:
            rm = rm.rename(index=rename_maps[concept1])
        if concept2 in rename_maps:
            rm = rm.rename(columns=rename_maps[concept2])

    # 4) Optional cleanup
    if clean_zeros:
        rm = rm.loc[rm.sum(axis=1) != 0, rm.sum(axis=0) != 0]

    # 5) Relation container
    R = Relation(concept1, concept2)
    R.set_matrix(rm)

    # 6) Analyses
    wanted = {s.lower() for s in include_stats}
    if "diversity" in wanted:
        m = compute_diversity_metrics(rm)
        R.diversity_row_metrics = m.get("row_metrics")
        R.diversity_column_metrics = m.get("column_metrics")

    if "bipartite network" in wanted:
        bp = analyze_bipartite_relation(rm, **kwargs)
        R.bipartite_graph = bp.get("bipartite_graph")
        R.bipartite_row_projection = bp.get("row_projection")
        R.bipartite_column_projection = bp.get("column_projection")
        R.bipartite_row_stats = bp.get("row_stats")
        R.bipartite_column_stats = bp.get("column_stats")
        R.bipartite_global_stats = bp.get("bipartite_global")
        R.bipartite_row_global = bp.get("row_global")
        R.bipartite_column_global = bp.get("column_global")

    if "cluster" in wanted:
        res = cluster_relation_matrix(rm, **kwargs)
        R.clusters = res.get("clusters")
        R.n_clusters = res.get("n_clusters")
        if "silhouette_scores" in res:
            R.silhouette_scores = res["silhouette_scores"]

    if "bicluster" in wanted:
        res = bicluster_relation_matrix(rm, **kwargs)
        R.bicluster_model = res.get("model")
        R.bicluster_row_clusters = res.get("row_clusters")
        R.bicluster_column_clusters = res.get("column_clusters")

    if "correspondence" in wanted:
        row_c, col_c, inertia = compute_correspondence_analysis(rm)
        R.ca_row_coords, R.ca_col_coords, R.ca_explained_inertia = row_c, col_c, inertia

    if "chi2" in wanted:
        resid, sorted_resid, expected, chi2_stat, dof = extract_sorted_residual_pairs(rm)
        R.chi2_residuals_df = resid
        R.chi2_sorted_residuals = sorted_resid
        R.chi2_expected_df = expected
        R.chi2_chi2_stat = chi2_stat
        R.chi2_dof = dof

    if "svd" in wanted:
        row_proj, svals, exvar = compute_svd_statistics(rm)
        R.svd_row_projection, R.svd_singular_values, R.svd_explained_variance = row_proj, svals, exvar

    if "log-ratio" in wanted:
        lrd, expd, sorted_lr = compute_log_ratio(rm)
        R.log_ratio_df, R.log_ratio_expected_df, R.log_ratio_sorted_log_ratios = lrd, expd, sorted_lr

    # 7) Optional storage
    if relation_matrices is not None:
        relation_matrices.setdefault(concept1, {})[concept2] = rm
        relation_matrices.setdefault(concept2, {})[concept1] = rm.T

    if relations is not None:
        relations.setdefault(concept1, {})[concept2] = R
        relations.setdefault(concept2, {})[concept1] = R

    # 8) Optional link-to-self
    if to_self and self_obj is not None:
        R.link_to_self(self_obj, prefix=link_prefix)

    return R, rm

"""Network-analysis utilities for building graphs, computing centralities and basic graph statistics."""
# Netwrok analysis

def export_graph_formats(
    G: nx.Graph,
    filename_base: str,
    output_dir: str = '.',
) -> None:
    """
    Export a NetworkX graph to GraphML, GEXF, and Pajek NET formats.

    Parameters:
        G (nx.Graph): The graph to export.
        filename_base (str): Base filename without extension.
        output_dir (str): Directory to save files (default is current directory).

    Returns:
        None
    """
    os.makedirs(output_dir, exist_ok=True)

    path_graphml = os.path.join(output_dir, f"{filename_base}.graphml")
    path_gexf = os.path.join(output_dir, f"{filename_base}.gexf")
    path_net = os.path.join(output_dir, f"{filename_base}.net")

    nx.write_graphml(G, path_graphml)
    nx.write_gexf(G, path_gexf)
    nx.write_pajek(G, path_net)

    print(f"Graphs exported to:\n- {path_graphml}\n- {path_gexf}\n- {path_net}")

import networkx as nx

def keep_large_components(
    G: nx.Graph,
    min_size: int = 4,
    *,
    mode: str = 'weak',
) -> nx.Graph:
    """
    Return a copy of G induced by nodes that belong to components
    of size >= min_size.

    For undirected graphs uses nx.connected_components.
    For directed graphs uses:
      - nx.weakly_connected_components if mode == "weak" (default),
      - nx.strongly_connected_components if mode == "strong".

    Parameters
    ----------
    G : nx.Graph
        The input graph (Graph/DiGraph/…).
    min_size : int, default 4
        Minimum component size to keep.
    mode : {"weak","strong"}, default "weak"
        How to treat directed graphs.

    Returns
    -------
    nx.Graph
        Subgraph copy with small components removed.
    """
    if G.is_directed():
        comps = (nx.weakly_connected_components(G) if mode != "strong"
                 else nx.strongly_connected_components(G))
    else:
        comps = nx.connected_components(G)

    keep = [n for c in comps if len(c) >= min_size for n in c]
    return G.subgraph(keep).copy()

def set_layer_by_component(
    G: nx.Graph,
    *,
    mode: str = 'weak',
) -> dict:
    """
    Assign the same "layer" to all nodes in the same connected component.

    Behavior
    --------
    - Undirected: uses nx.connected_components.
    - Directed:
        * mode="weak"  → nx.weakly_connected_components (default)
        * mode="strong" → nx.strongly_connected_components

    Returns
    -------
    dict
        {node: layer_id} mapping also set on the graph as the "layer" attribute.
    """
    if G.is_directed():
        comps = (nx.weakly_connected_components(G) if mode != "strong"
                 else nx.strongly_connected_components(G))
    else:
        comps = nx.connected_components(G)

    comps = sorted(comps, key=lambda c: -len(c))  # stable order: largest first
    layer_map = {n: i for i, comp in enumerate(comps) for n in comp}
    nx.set_node_attributes(G, layer_map, "layer")
    return layer_map

def _resolve_positions(
    G,
    layout,
    **layout_kwargs,
):
    """
    Return a position dict for drawing.

    Accepts:
    - str: layout name ("spring","circular","kamada_kawai","shell","multipartite",...)
    - callable: a layout function (e.g., nx.multipartite_layout)
    - dict: a precomputed {node: (x, y)} mapping (returned as-is)
    """
    if isinstance(layout, dict):
        return layout  # already positions

    layouts = {
        "spring": nx.spring_layout,
        "circular": nx.circular_layout,
        "kamada_kawai": nx.kamada_kawai_layout,
        "kk": nx.kamada_kawai_layout,
        "shell": nx.shell_layout,
        "spectral": nx.spectral_layout,
        "random": nx.random_layout,
        "multipartite": nx.multipartite_layout,
    }

"""Utilities related to Functions for group analysis in the Biblium project."""
# functions for group analysis


def generate_group_matrix(
    df: pd.DataFrame,
    group_desc: Union[int, str, pd.DataFrame, Dict[str, str], Dict[str, List[str]], List[str]],
    *,
    force_type: Optional[Literal["column", "multiitem", "regex", "binary", "year", "concept"]] = None,
    text_column: Optional[str] = None,
    regex_flags: int = re.IGNORECASE,
    concept_whole_word: bool = False,
    sep: str = "; ",
    top_n: Optional[int] = None,
    include_items: Optional[List[str]] = None,
    exclude_items: Optional[List[str]] = None,
    year_column: Optional[str] = None,
    year_range: Optional[Tuple[int, int]] = None,
    cutpoints: Optional[List[int]] = None,
    n_periods: Optional[int] = None,
    cut_labels: Optional[List[str]] = None,
    binary_as_int: bool = False,
    invert_matrix: bool = False,
) -> pd.DataFrame:
    """
    Generate a document × group binary matrix from a flexible descriptor.

    What you can pass via `group_desc`
    ----------------------------------
    1) Concept dataframe (auto or force_type="concept"):
       * Columns = group names; cells are strings/phrases (supports "*" wildcard).
       * Requires `text_column` to search in.
       * Use `concept_whole_word=True` to wrap each term with "\\b...\\b".

    2) Binary dataframe (auto or force_type="binary"):
       * Same index as `df`, columns are group names, values are bool or {0,1}.

    3) List of column names (list of strings):
       * Each column name must exist in `df` and contain binary values (bool or {0,1}).
       * Returns a matrix with these columns as groups.

    4) Dict (auto-detected):
       * {group -> pattern} or {group -> [terms]}.
       * Terms may include "*" wildcard. Requires `text_column`.

    5) Column name in `df` (string):
       * force_type="column": one-hot single-valued categorical.
       * force_type="multiitem": treat as list-wise (see below).
       * Auto mode:
           - If column is list-wise (native lists/tuples/sets, stringified lists, or
             strings that split on literal `sep` into ≥2 non-empty tokens) → one-hot multi-item.
           - Else, if the column is "year-ish" (≥80% of non-null numeric values within
             [1900, 2100]) → one-hot years or binned years (cutpoints/n_periods).
           - Else → one-hot categorical.

    6) Positive integer n:
       * Random overlapping groups:
         - For each document and each group, membership is assigned independently
           at random (≈50% chance).
         - Groups can overlap; a document may belong to multiple groups or none.
       * The function enforces that there are no completely empty rows or columns
         by adding at least one membership where needed.
       * Returns columns named "Random 1", ..., "Random n".

    Shared utilities
    ----------------
    * `year_column` + `year_range=(start, end)` pre-filters rows before grouping.
    * `include_items` / `exclude_items` filter result columns by name.
    * `top_n` keeps top-N most frequent groups (multi-item only).
    * `invert_matrix=True` flips membership; `binary_as_int=True` returns 0/1.

    Returns
    -------
    pd.DataFrame
        Binary matrix (bool or 0/1) aligned to the filtered `df.index`.
    """
    import ast  # local import for safe literal_eval

    # ---------------------------
    # Helpers
    # ---------------------------
    def _finalize(
        B: pd.DataFrame,
    ) -> pd.DataFrame:
        B = B.astype(bool)
        if invert_matrix:
            B = ~B
        return B.astype(int) if binary_as_int else B

    def _ensure_text_column_present() -> None:
        if text_column is None or text_column not in work_df.columns:
            raise ValueError('When matching concepts/regex/dicts you must set "text_column" to a valid df column.')

    def _is_binary_dataframe(
        X: pd.DataFrame,
    ) -> bool:
        if X.empty:
            return True
        for c in X.columns:
            s = X[c]
            if pd.api.types.is_bool_dtype(s):
                continue
            if pd.api.types.is_numeric_dtype(s):
                vals = pd.unique(s.dropna())
                if len(vals):
                    arr = pd.to_numeric(pd.Series(vals), errors="coerce")
                    if not np.isin(arr, [0.0, 1.0]).all():
                        return False
            else:
                return False
        return True

    def _wildcardize(
        token: str,
    ) -> str:
        # Escape regex, then expand "*" → ".*"
        return re.escape(token).replace("\\*", ".*")

    def _to_non_capturing(
        pat: str,
    ) -> str:
        # Turn bare "(" into "(?:" to avoid unintended capture groups
        return re.sub(r"\((?!\?)", "(?:", pat)

    def _build_alt_pattern(
        terms: Iterable[str],
        whole_word: bool,
    ) -> str:
        alts: List[str] = []
        for t in terms:
            if t is None:
                continue
            s = str(t).strip()
            if not s:
                continue
            p = _wildcardize(s)
            if whole_word:
                p = rf"\b{p}\b"
            alts.append(p)
        return "|".join(alts) if alts else r"(?!x)x"  # never-matches sentinel

    def _filter_columns(
        B: pd.DataFrame,
    ) -> pd.DataFrame:
        cols = list(B.columns)
        if include_items is not None:
            if not isinstance(include_items, (list, tuple)):
                raise ValueError('"include_items" must be a list of column names.')
            cols = [c for c in include_items if c in B.columns]
        if exclude_items is not None:
            if not isinstance(exclude_items, (list, tuple)):
                raise ValueError('"exclude_items" must be a list of column names.')
            cols = [c for c in cols if c not in set(exclude_items)]
        return B.loc[:, cols]

    def _is_yearish(
        numeric: pd.Series,
    ) -> bool:
        """
        Robust year detector: True if ≥80% of non-null values lie in [1900, 2100].
        Always returns a plain bool (never pandas.NA), avoiding ambiguous truth values.
        """
        valid = numeric.notna()
        if not bool(valid.any()):
            return False
        frac = numeric[valid].between(1900, 2100).mean()
        return bool(frac > 0.8)

    # ----- List-wise detection & tokenization (sep is treated literally) -----
    _listlike_types = (list, tuple, set)

    def _split_on_sep(
        s: str,
    ) -> List[str]:
        """
        Split on the literal `sep` and return non-empty, stripped tokens.
        Returns [] if `s` is empty or contains only separators/whitespace.
        """
        if not isinstance(s, str):
            s = str(s) if s is not None else ""
        s = s.strip()
        if not s:
            return []
        if not sep:
            return [s]
        parts = [t.strip() for t in s.split(sep)]
        tokens = [t for t in parts if t != ""]
        return tokens

    def _tokens_from_cell(
        x: Any,
    ) -> List[str]:
        """
        Convert a cell to a list of tokens.

        Handles native lists/tuples/sets, stringified lists, sep-delimited strings,
        and singletons. Guarantees no empty tokens and respects `sep` literally.
        """
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return []
        if isinstance(x, _listlike_types):
            return [str(t).strip() for t in x if str(t).strip() != ""]
        s = str(x)
        if not s.strip():
            return []
        # Stringified lists like '["a", "b"]'
        if re.match(r"^\s*\[.*\]\s*$", s):
            try:
                parsed = ast.literal_eval(s)
                if isinstance(parsed, _listlike_types):
                    return [str(t).strip() for t in parsed if str(t).strip() != ""]
            except Exception:
                # Fall through to literal split
                pass
        # sep-delimited strings (literal, not regex)
        toks = _split_on_sep(s)
        if toks:
            return toks
        # Fallback: single token if not only separators
        return [s.strip()] if s.strip() != "" else []

    def _explode_multiitem(
        series: pd.Series,
    ) -> pd.DataFrame:
        """
        Explode a possibly mixed-type series (lists / stringified lists / sep-strings / singletons)
        into a one-hot matrix via explode + get_dummies.
        """
        exploded = (
            series.dropna()
            .map(_tokens_from_cell)
            .explode()
            .dropna()
            .map(lambda z: str(z).strip())
        )
        exploded = exploded[exploded != ""]
        if exploded.empty:
            return pd.DataFrame(index=series.index)
        B = pd.get_dummies(exploded, dtype=bool).groupby(level=0).max()
        return B

    def _is_multiitem_series(
        series: pd.Series,
    ) -> bool:
        """
        Heuristics for list-wise columns:
          * any native list/tuple/set
          * any stringified list like "[...]"
          * any string that, when split on literal `sep`, yields ≥2 non-empty tokens

        Critically, strings that are exactly the separator (or only separators)
        do NOT count as list-wise.
        """
        sample = series.head(5000)
        has_native = sample.map(lambda v: isinstance(v, _listlike_types)).any()

        try:
            as_str = sample.dropna().astype(str)
        except Exception:
            as_str = sample.dropna().map(str)

        has_list_str = as_str.str.match(r"^\s*\[.*\]\s*$").any()
        has_multi_split = bool(sep) and as_str.map(lambda s: len(_split_on_sep(s)) >= 2).any()
        return bool(has_native or has_list_str or has_multi_split)

    # ---------------------------
    # Optional pre-filtering by (year_column, year_range)
    # ---------------------------
    work_df = df
    if year_column is not None and year_range is not None:
        if year_column not in work_df.columns:
            raise ValueError(f"`year_column='{year_column}'` not found in df.")
        y = pd.to_numeric(work_df[year_column], errors="coerce")
        start, end = year_range
        work_df = work_df.loc[y.between(start, end)]
        if work_df.empty:
            return pd.DataFrame(index=work_df.index)

    # ---------------------------
    # List of binary column names
    # ---------------------------
    if isinstance(group_desc, list) and not isinstance(group_desc, str):
        if not group_desc:
            raise ValueError("group_desc list cannot be empty.")
        
        # Check all items are strings
        if not all(isinstance(col, str) for col in group_desc):
            raise ValueError("When group_desc is a list, all items must be column names (strings).")
        
        # Check all columns exist
        missing_cols = [col for col in group_desc if col not in work_df.columns]
        if missing_cols:
            raise ValueError(f"Columns not found in df: {missing_cols}")
        
        # Extract the subset
        B = work_df[group_desc].copy()
        
        # Validate that all columns are binary
        if not _is_binary_dataframe(B):
            raise ValueError("All columns in group_desc list must be binary (bool or 0/1).")
        
        B = _filter_columns(B)
        return _finalize(B.reindex(index=work_df.index, fill_value=False))

    # ---------------------------
    # Random groups: integer group_desc (overlapping)
    # ---------------------------
    if isinstance(group_desc, int) and not isinstance(group_desc, bool):
        n_groups = int(group_desc)
        if n_groups <= 0:
            raise ValueError('When "group_desc" is an integer, it must be a positive number of groups.')
        if work_df.empty:
            return _finalize(pd.DataFrame(index=work_df.index))

        col_names = [f"Random {i + 1}" for i in range(n_groups)]

        # Each document × group membership is random (≈50% chance)
        rand_membership = np.random.rand(len(work_df.index), n_groups) < 0.5
        B = pd.DataFrame(rand_membership, index=work_df.index, columns=col_names)

        # Ensure no completely empty rows: give such rows one random group
        empty_rows = ~B.any(axis=1)
        if empty_rows.any():
            choices = np.random.randint(n_groups, size=empty_rows.sum())
            for idx, g_idx in zip(B.index[empty_rows], choices):
                B.loc[idx, col_names[g_idx]] = True

        # Ensure no completely empty columns: give such cols one random document
        empty_cols = ~B.any(axis=0)
        if empty_cols.any():
            for col in B.columns[empty_cols]:
                row_pos = np.random.randint(len(B.index))
                B.iloc[row_pos, B.columns.get_loc(col)] = True

        B = _filter_columns(B)
        return _finalize(B)

    # ---------------------------
    # Forced-type short-circuits
    # ---------------------------
    if force_type == "binary":
        if not isinstance(group_desc, pd.DataFrame):
            raise TypeError('force_type="binary" expects a binary dataframe as `group_desc`.')
        return _finalize(group_desc.reindex(index=work_df.index, fill_value=False))

    if force_type == "column":
        if not isinstance(group_desc, str):
            raise TypeError('force_type="column" expects a column name (str).')
        col = group_desc
        if col not in work_df.columns:
            raise ValueError(f'Column "{col}" not found in df.')
        S = work_df[col].astype("string")
        B = pd.get_dummies(S, dtype=bool)
        return _finalize(B.reindex(index=work_df.index, fill_value=False))

    if force_type == "multiitem":
        if not isinstance(group_desc, str):
            raise TypeError('force_type="multiitem" expects a column name (str).')
        col = group_desc
        if col not in work_df.columns:
            raise ValueError(f'Column "{col}" not found in df.')
        B = _explode_multiitem(work_df[col])
        if not B.empty and top_n is not None and top_n < B.shape[1]:
            top_cols = B.sum(axis=0).sort_values(ascending=False).head(top_n).index
            B = B.loc[:, top_cols]
        B = _filter_columns(B)
        return _finalize(B.reindex(index=work_df.index, fill_value=False))

    if force_type == "regex":
        if not isinstance(group_desc, dict):
            raise TypeError('force_type="regex" expects a dict: {group: pattern or terms}.')
        _ensure_text_column_present()
        text = work_df[text_column].fillna("")
        groups = list(group_desc.keys())
        B = pd.DataFrame(False, index=work_df.index, columns=groups)
        for g, pat_or_terms in group_desc.items():
            if isinstance(pat_or_terms, (list, tuple, set)):
                pat = _build_alt_pattern(pat_or_terms, concept_whole_word)
            else:
                pat = str(pat_or_terms)
                # Allow "*" expansion and optional whole-word wrapping for convenience
                pat = _build_alt_pattern([pat], concept_whole_word)
            B[g] = text.str.contains(_to_non_capturing(pat), flags=regex_flags, regex=True)
        return _finalize(B)

    if force_type == "concept":
        if not isinstance(group_desc, pd.DataFrame):
            raise TypeError('force_type="concept" expects a concept dataframe.')
        _ensure_text_column_present()
        concept_df = group_desc.copy()
        for c in concept_df.columns:
            concept_df[c] = concept_df[c].map(
                lambda x: None if pd.isna(x) else (str(x).strip() or None)
            )
        patterns = {
            c: _build_alt_pattern(concept_df[c].dropna(), concept_whole_word)
            for c in concept_df.columns
        }
        text = work_df[text_column].fillna("")
        B = pd.DataFrame(False, index=work_df.index, columns=list(patterns.keys()))
        for g, pat in patterns.items():
            B[g] = text.str.contains(_to_non_capturing(pat), flags=regex_flags, regex=True)
        return _finalize(B)

    if force_type == "year":
        if not isinstance(group_desc, str):
            raise TypeError('force_type="year" expects a column name (str).')
        col = group_desc
        if col not in work_df.columns:
            raise ValueError(f'Column "{col}" not found in df.')
        numeric = pd.to_numeric(work_df[col], errors="coerce")
        yy = numeric.astype("Int64")
        if cutpoints is not None or n_periods is not None:
            valid = yy.dropna().astype(int)
            if valid.empty:
                return _finalize(pd.DataFrame(index=work_df.index))
            if cutpoints is not None:
                bins = [-np.inf] + list(cutpoints) + [np.inf]
            else:
                mn, mx = int(valid.min()), int(valid.max())
                if n_periods is None or n_periods < 1:
                    raise ValueError("n_periods must be a positive integer.")
                bins = np.linspace(mn, mx + 1, n_periods + 1).astype(int).tolist()
            labels = (
                cut_labels
                if cut_labels
                else [f"{bins[i]}–{bins[i + 1] - 1}" for i in range(len(bins) - 1)]
            )
            if len(labels) != len(bins) - 1:
                raise ValueError("Length of cut_labels must equal number of bins minus one.")
            cats = pd.cut(valid, bins=bins, labels=labels, right=False, include_lowest=True)
            B = pd.get_dummies(cats, dtype=bool)
            return _finalize(B.reindex(index=work_df.index, fill_value=False))
        else:
            B = pd.get_dummies(yy, dtype=bool)
            return _finalize(B.reindex(index=work_df.index, fill_value=False))

    # ---------------------------
    # Auto-detection path
    # ---------------------------

    # A) DataFrame: binary or concept
    if isinstance(group_desc, pd.DataFrame):
        G = group_desc
        if _is_binary_dataframe(G):
            return _finalize(G.reindex(index=work_df.index, fill_value=False))
        _ensure_text_column_present()
        concept_df = G.copy()
        for c in concept_df.columns:
            concept_df[c] = concept_df[c].map(
                lambda x: None if pd.isna(x) else (str(x).strip() or None)
            )
        patterns = {
            c: _build_alt_pattern(concept_df[c].dropna(), concept_whole_word)
            for c in concept_df.columns
        }
        text = work_df[text_column].fillna("")
        B = pd.DataFrame(False, index=work_df.index, columns=list(patterns.keys()))
        for g, pat in patterns.items():
            B[g] = text.str.contains(_to_non_capturing(pat), flags=regex_flags, regex=True)
        return _finalize(B)

    # B) Dict: {group -> pattern or terms}
    if isinstance(group_desc, dict):
        _ensure_text_column_present()
        groups = list(group_desc.keys())
        text = work_df[text_column].fillna("")
        B = pd.DataFrame(False, index=work_df.index, columns=groups)
        for g, matcher in group_desc.items():
            if isinstance(matcher, (list, tuple, set)):
                pat = _build_alt_pattern(list(matcher), concept_whole_word)
            else:
                pat = _build_alt_pattern([str(matcher)], concept_whole_word)
            B[g] = text.str.contains(_to_non_capturing(pat), flags=regex_flags, regex=True)
        return _finalize(B)

    # C) String: name of a column in df (with rich list-wise detection)
    if isinstance(group_desc, str):
        col = group_desc
        if col not in work_df.columns:
            raise ValueError(f'Column "{col}" not found in df.')

        S_raw = work_df[col]

        # Auto-detect list-wise columns first
        if _is_multiitem_series(S_raw):
            B = _explode_multiitem(S_raw)
            if not B.empty and top_n is not None and top_n < B.shape[1]:
                top_cols = B.sum(axis=0).sort_values(ascending=False).head(top_n).index
                B = B.loc[:, top_cols]
            B = _filter_columns(B)
            return _finalize(B.reindex(index=work_df.index, fill_value=False))

        # Not list-wise → maybe a year column?
        S = S_raw.astype("string")
        numeric = pd.to_numeric(S, errors="coerce")
        if _is_yearish(numeric):
            yy = numeric.astype("Int64")
            if cutpoints is not None or n_periods is not None:
                valid = yy.dropna().astype(int)
                if valid.empty:
                    return _finalize(pd.DataFrame(index=work_df.index))
                if cutpoints is not None:
                    bins = [-np.inf] + list(cutpoints) + [np.inf]
                else:
                    mn, mx = int(valid.min()), int(valid.max())
                    if n_periods is None or n_periods < 1:
                        raise ValueError("n_periods must be a positive integer.")
                    bins = np.linspace(mn, mx + 1, n_periods + 1).astype(int).tolist()
                labels = (
                    cut_labels
                    if cut_labels
                    else [f"{bins[i]}–{bins[i + 1] - 1}" for i in range(len(bins) - 1)]
                )
                if len(labels) != len(bins) - 1:
                    raise ValueError("Length of cut_labels must equal number of bins minus one.")
                cats = pd.cut(
                    valid,
                    bins=bins,
                    labels=labels,
                    right=False,
                    include_lowest=True,
                )
                B = pd.get_dummies(cats, dtype=bool)
                return _finalize(B.reindex(index=work_df.index, fill_value=False))
            else:
                B = pd.get_dummies(yy, dtype=bool)
                return _finalize(B.reindex(index=work_df.index, fill_value=False))

        # Fallback: single-valued categorical one-hot
        B = pd.get_dummies(S, dtype=bool)
        return _finalize(B.reindex(index=work_df.index, fill_value=False))

    # ---------------------------
    # No match
    # ---------------------------
    raise ValueError('Unable to interpret "group_desc". Consider setting "force_type" explicitly.')


def merge_group_performances(
    group_dfs: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """
    Merge performance dataframes from different groups by aligning on 'Variable' and 'Indicator' columns.

    Args:
        group_dfs (dict[str, pd.DataFrame]): Dictionary where keys are group names and values are
                                             corresponding performance dataframes with columns
                                             ['Variable', 'Indicator', 'Value'].

    Returns:
        pd.DataFrame: Merged dataframe with 'Variable' and 'Indicator' as index columns and one column
                      for each group's performance values.
    """
    merged_df = None
    for group_name, df in group_dfs.items():
        temp_df = df.copy()
        temp_df = temp_df.rename(columns={'Value': group_name})
        if merged_df is None:
            merged_df = temp_df
        else:
            merged_df = pd.merge(merged_df, temp_df, on=['Variable', 'Indicator'], how='outer')
    return merged_df

def count_occurrences_across_groups(
    groups,
    group_matrix,
    count_func_name,
    merge_type='all items',
    **kwargs,
):
    """
    Count item occurrences (e.g., keywords, authors, sources) across multiple groups 
    and merge the results into a single DataFrame with renamed columns indicating the group.
    Also computes combined statistics across all groups.
    
    Args:
        groups (dict): Mapping of group names to group objects.
        group_matrix (pd.DataFrame): DataFrame with group names as columns (order defines merge order).
        count_func_name (str): Name of the counting method to call on each group.
        merge_type (str): Either "all items" (outer join) or "shared items" (inner join).
        **kwargs: Additional keyword arguments passed to the counting function.
    
    Returns:
        pd.DataFrame: Merged DataFrame with group-labeled columns and combined statistics.
                     Missing items in a group get 0 for counts/proportions/percentages,
                     but NaN for rank and percentrank.
    """
    how_map = {"all items": "outer", "shared items": "inner"}
    how = how_map.get(merge_type)
    if how is None:
        raise ValueError('merge_type must be "all items" or "shared items"')
    
    if group_matrix is None or len(getattr(group_matrix, "columns", [])) == 0:
        raise ValueError("group_matrix must have at least one column (group).")
    
    # Run the counting function per group and suffix non-key columns with the group name
    dfs = []
    group_names = list(group_matrix.columns)
    
    for g in group_names:
        if g not in groups:
            raise KeyError(f'Group "{g}" not found in the provided groups mapping.')
        if not hasattr(groups[g], count_func_name):
            raise AttributeError(f'Group "{g}" has no method "{count_func_name}".')
        
        df = getattr(groups[g], count_func_name)(**kwargs)
        if not isinstance(df, pd.DataFrame) or df.shape[1] < 2:
            raise ValueError(f'Counting method for group "{g}" must return a DataFrame with at least 2 columns.')
        
        key_col = df.columns[0]
        df = df.rename(columns={c: f"{c} ({g})" for c in df.columns[1:]})
        dfs.append(df)
    
    # Ensure all key columns match
    first_key = dfs[0].columns[0]
    if any(d.columns[0] != first_key for d in dfs[1:]):
        raise ValueError("All group DataFrames must share the same first (key) column.")
    
    # Merge all group dataframes
    merged = reduce(lambda l, r: pd.merge(l, r, on=first_key, how=how), dfs)
    
    # Identify rank columns (to keep as NaN) vs count columns (to fill with 0)
    rank_cols = [col for col in merged.columns if isinstance(col, str) and 'rank' in col.lower()]
    non_rank_cols = [col for col in merged.columns[1:] if col not in rank_cols]
    
    # Fill NaN: 0 for counts/proportions/percentages, leave NaN for ranks
    for col in non_rank_cols:
        merged[col] = merged[col].fillna(0)
    
    # Safe numeric conversion per column
    for col in merged.columns[1:]:
        try:
            merged[col] = pd.to_numeric(merged[col])
        except (ValueError, TypeError):
            pass
    
    # Compute combined statistics
    # Find the base column names (without group suffix)
    base_cols = set()
    for col in merged.columns[1:]:
        if '(' in col and ')' in col:
            base_name = col.split('(')[0].strip()
            base_cols.add(base_name)
    
    # Collect all new combined columns in a dictionary first to avoid fragmentation
    new_cols = {}
    
    # For each base metric, compute combined stats
    for base_col in base_cols:
        if 'rank' in base_col.lower():
            # Don't compute combined ranks directly - will compute after counts
            continue
        
        # Get all columns for this metric across groups
        metric_cols = [col for col in merged.columns if col.startswith(f"{base_col} (")]
        
        if metric_cols:
            if 'OCC' in base_col or 'Num' in base_col or base_col.startswith('N'):
                # For counts: sum across groups
                new_cols[f"{base_col} (Combined)"] = merged[metric_cols].sum(axis=1)
            elif 'Proportion' in base_col or 'Percentage' in base_col or '%' in base_col:
                # For proportions/percentages: recalculate based on combined counts
                # Will be recalculated after we have combined counts
                pass
    
    # Recalculate proportions and percentages based on combined counts
    count_col_name = None
    for base_col in base_cols:
        if 'OCC' in base_col or 'Num' in base_col or (base_col.startswith('N') and 'rank' not in base_col.lower()):
            count_col_name = f"{base_col} (Combined)"
            break
    
    if count_col_name and count_col_name in new_cols:
        total_combined = new_cols[count_col_name].sum()
        
        for base_col in base_cols:
            if 'Proportion' in base_col:
                new_cols[f"{base_col} (Combined)"] = new_cols[count_col_name] / total_combined if total_combined > 0 else 0
            elif 'Percentage' in base_col or '%' in base_col:
                new_cols[f"{base_col} (Combined)"] = (new_cols[count_col_name] / total_combined * 100) if total_combined > 0 else 0
        
        # Compute combined ranks based on combined counts
        count_series = new_cols[count_col_name]
        ranks = count_series.rank(method='first', ascending=False).astype(int)
        new_cols['Rank (Combined)'] = ranks
        
        # Compute combined percentrank
        max_rank = len(merged)
        new_cols['PercentRank (Combined)'] = ((max_rank - ranks + 1) / max_rank * 100) if max_rank > 0 else 0
    
    # Add all new columns at once using pd.concat (avoids fragmentation warning)
    if new_cols:
        new_cols_df = pd.DataFrame(new_cols, index=merged.index)
        merged = pd.concat([merged, new_cols_df], axis=1)
    
    # Reorder columns: key, then all metrics for group1, then group2, ..., then combined
    key_col = merged.columns[0]
    ordered_cols = [key_col]
    
    # Add columns for each group in order
    for g in group_names:
        group_cols = [col for col in merged.columns if f"({g})" in col]
        ordered_cols.extend(sorted(group_cols))
    
    # Add combined columns
    combined_cols = [col for col in merged.columns if "(Combined)" in col]
    ordered_cols.extend(sorted(combined_cols))
    
    merged = merged[ordered_cols]
    
    return merged

def compute_group_intersections(
    group_matrix: pd.DataFrame,
    include_ids: bool = False,
    id_column: pd.Series = None,
) -> pd.DataFrame:
    """
    Computes all unique intersections from a binary group membership matrix.

    Parameters:
    - group_matrix: pd.DataFrame
        Binary DataFrame with rows as items (e.g., documents) and columns as group names (0/1).
    - include_ids: bool
        If True, includes a column listing the item IDs from `id_column` (Series aligned with group_matrix) or the index if not provided.
    - id_column: pd.Series or None
        Optional Series providing IDs for the items (same index as group_matrix). If None, uses the index.

    Returns:
    - pd.DataFrame with columns:
        - 'Groups': tuple of intersecting groups
        - 'Size': number of items in the intersection
        - 'ID' (optional): list of item IDs in the intersection
    """
    group_cols = group_matrix.columns.tolist()
    results = []

    for mask, df_subset in group_matrix.groupby(group_cols):
        active_groups = tuple(col for col, flag in zip(group_cols, mask) if flag == 1)
        if not active_groups:
            continue
        if include_ids:
            ids = id_column.loc[df_subset.index].tolist() if id_column is not None else df_subset.index.tolist()
            row = {"Groups": active_groups, "Size": len(df_subset), "ID": ids}
        else:
            row = {"Groups": active_groups, "Size": len(df_subset)}
        results.append(row)

    return pd.DataFrame(results).sort_values(by="Size", ascending=False).reset_index(drop=True)

def compute_group_similarity_matrices(
    group_matrix: pd.DataFrame,
    methods: list = ['jaccard'],
) -> dict:
    """
    Computes group × group similarity matrices for given methods.

    Parameters:
    - group_matrix: pd.DataFrame
        Binary matrix (0/1), rows = items, columns = group names.
    - methods: list of str
        List of methods to compute. Supported: 'jaccard', 'count', 'sokal-michener', 'simple-matching', 'rogers-tanimoto'.

    Returns:
    - dict: keys are method names, values are DataFrames (group × group similarity matrices)
    """

    supported_methods = {
        "jaccard": "Jaccard Index",
        "count": "Shared Items",
        "sokal-michener": "Sokal-Michener",
        "simple-matching": "Simple Matching",
        "rogers-tanimoto": "Rogers-Tanimoto"
    }

    binary = group_matrix.astype(bool).to_numpy()
    groups = group_matrix.columns.tolist()
    matrices = {}

    for method in methods:
        if method not in supported_methods:
            print(f"Warning: Unsupported method '{method}', skipping.")
            continue

        if method == "count":
            mat = pd.DataFrame(index=groups, columns=groups, dtype=float)
            for i, g1 in enumerate(groups):
                set1 = set(group_matrix.index[group_matrix[g1] == 1])
                for j, g2 in enumerate(groups):
                    set2 = set(group_matrix.index[group_matrix[g2] == 1])
                    mat.loc[g1, g2] = len(set1 & set2)
        else:
            dist = pairwise_distances(binary.T, metric=method)
            sim = 1 - dist
            mat = pd.DataFrame(sim, index=groups, columns=groups)

        matrices[method] = mat

    return matrices

def compare_continuous_by_binary_groups(
    df,
    numerical_cols,
    group_matrix,
    output_format = 'long',
):
    """
    Compare continuous variables across groups defined by a binary group matrix using both parametric and
    non-parametric statistical tests, while computing descriptive statistics and handling missing values.

    Parameters:
    df (pd.DataFrame): Original dataframe containing continuous numerical data.
    numerical_cols (list of str): List of column names in df that contain continuous numerical variables.
    group_matrix (pd.DataFrame): Binary matrix (same number of rows as df) where each column defines a group
                                  (1 = in group, 0 = not in group).
    output_format (str): "long" (default) for detailed per-group stats; "wide" for matrix-style summary of stats and p-values.

    Returns:
    pd.DataFrame: DataFrame containing descriptive statistics and p-values for each variable and group.
    """
    results = []

    for col in numerical_cols:
        group_data = []
        group_names = []
        group_stats = {}

        for group in group_matrix.columns:
            valid_mask = group_matrix[group] == 1
            data = df.loc[valid_mask, col]
            
            # Convert to numeric and filter finite values
            data = pd.to_numeric(data, errors='coerce')
            data = data.dropna()
            data = data[np.isfinite(data.values)]

            if len(data) >= 3:
                group_data.append(data)
                group_names.append(group)

                group_stats[group] = {
                    "Mean": data.mean(),
                    "SD": data.std(),
                    "Median": data.median(),
                    "IQR": data.quantile(0.75) - data.quantile(0.25),
                    "Shapiro p": shapiro(data)[1] if len(data) <= 5000 else np.nan,
                    "N": len(data)
                }

        if len(group_data) >= 2:
            try:
                param_p = f_oneway(*group_data)[1]
            except ValueError:
                param_p = np.nan

            try:
                nonparam_p = kruskal(*group_data)[1]
            except ValueError:
                nonparam_p = np.nan

            for group in group_names:
                stats = group_stats[group]
                results.append({
                    "Variable": col,
                    "Group": group,
                    "Mean": stats["Mean"],
                    "SD": stats["SD"],
                    "Median": stats["Median"],
                    "IQR": stats["IQR"],
                    "Shapiro p": stats["Shapiro p"],
                    "N": stats["N"],
                    "Parametric p (ANOVA)": param_p,
                    "Non-parametric p (Kruskal)": nonparam_p
                })

    long_df = pd.DataFrame(results)

    if output_format == "wide":
        # Pivot group-level descriptives into wide format
        descriptives = long_df.pivot(index="Variable", columns="Group", values=["Mean", "SD", "Median", "IQR", "Shapiro p", "N"])
        descriptives.columns = [f"{stat} ({grp})" for stat, grp in descriptives.columns]

        # Add p-values to wide format
        p_values = long_df[["Variable", "Parametric p (ANOVA)", "Non-parametric p (Kruskal)"]].drop_duplicates().set_index("Variable")
        wide_df = pd.concat([descriptives, p_values], axis=1).reset_index()
        return wide_df
    else:
        return long_df

"""Utilities related to Time series analysis in the Biblium project."""
# Time series analysis

def aggregate_bibliometrics_by_group_and_year(
    df,
    binary_df,
    group_columns = None,
    year_column = 'Year',
    metrics = ['Cited by'],
    additional_binary_dfs = None,
    normalize = True,
    include_cumulative = False,
    include_percentage = False,
    group_selection_top_n = None,
    group_selection_include_regex = None,
    group_selection_exclude_regex = None,
    aggfunc = 'sum',
    return_format = 'wide',
):
    """
    Aggregate bibliometric indicators by group and year using one or more binary dataframes.

    Parameters
    ----------
    df : pd.DataFrame
        The main dataframe with bibliographic data, including a "Year" column.
    binary_df : pd.DataFrame
        A binary dataframe where each column represents a group of interest.
    group_columns : list of str, optional
        Subset of binary_df columns to include in the analysis. If None, use all (after optional selection filters).
    year_column : str, default "Year"
        Name of the column in df indicating the year of publication.
    metrics : list of str
        Column names in df or additional binary dataframes to aggregate.
    additional_binary_dfs : dict of {str: pd.DataFrame}, optional
        Dictionary of binary dataframes to be used for metric calculation, where key is a metric name.
    normalize : bool, default True
        If True, include both raw and normalized metrics by document count.
    include_cumulative : bool, default False
        If True, include cumulative values.
    include_percentage : bool, default False
        If True, include percentage of yearly totals for each metric.
    group_selection_top_n : int, optional
        If specified, use only the top N columns in binary_df with the most total counts.
    group_selection_include_regex : str, optional
        A regex pattern to include only matching columns.
    group_selection_exclude_regex : str, optional
        A regex pattern to exclude matching columns.
    aggfunc : str or callable, default "sum"
        Aggregation function to use for metrics (e.g., "sum", "mean", etc.).
    return_format : str, default "wide"
        Format of the result: "long" for tidy format, "wide" for pivoted format, or "both" to return a dictionary.

    Returns
    -------
    pd.DataFrame or dict of DataFrames
        A dataframe with aggregated results by group and year or both formats.
    """

    all_columns = binary_df.columns

    if group_selection_include_regex:
        all_columns = [col for col in all_columns if re.search(group_selection_include_regex, col)]

    if group_selection_exclude_regex:
        all_columns = [col for col in all_columns if not re.search(group_selection_exclude_regex, col)]

    if group_selection_top_n is not None:
        top_cols = binary_df[all_columns].sum().sort_values(ascending=False).head(group_selection_top_n).index.tolist()
        all_columns = top_cols

    if group_columns is None:
        group_columns = all_columns

    results_wide = []
    results_long = []

    full_year_range = pd.Index(sorted(df[year_column].dropna().unique()))

    for group in group_columns:
        mask = binary_df[group] == 1
        sub_df = df.loc[mask].copy()
        sub_df["__group_marker__"] = 1

        base = sub_df[[year_column, "__group_marker__"]].copy()
        base = base.groupby(year_column).count().rename(columns={"__group_marker__": "Number of documents"})
        base = base.reindex(full_year_range, fill_value=0)

        metric_data = base.copy()

        for metric in metrics:
            is_additional = additional_binary_dfs and metric in additional_binary_dfs

            if metric in df.columns:
                values = sub_df.groupby(year_column)[metric].agg(aggfunc).reindex(full_year_range, fill_value=0)
            elif is_additional:
                bin_df = additional_binary_dfs[metric]
                values = bin_df.loc[mask].groupby(df[year_column]).sum().sum(axis=1).reindex(full_year_range, fill_value=0)
            else:
                continue

            metric_data[metric] = values

            if normalize and metric in metric_data.columns:
                norm_col = f"{metric} (Per Document)"
                metric_data[norm_col] = metric_data[metric] / metric_data["Number of documents"].replace(0, pd.NA)

                if is_additional:
                    percent_col = f"{metric} (Percentage of Group)"
                    metric_data[percent_col] = metric_data[metric] / metric_data["Number of documents"].replace(0, pd.NA) * 100

        if include_cumulative:
            for metric in metric_data.columns:
                metric_data[f"{metric} (Cumulative)"] = metric_data[metric].cumsum()

        if include_percentage:
            total_per_year = metric_data.sum(axis=0, numeric_only=True)
            for metric in metrics:
                if metric in metric_data.columns:
                    metric_data[f"{metric} (Percentage)"] = metric_data[metric] / total_per_year[metric] * 100

        metric_data["Group"] = group
        metric_data = metric_data.reset_index().rename(columns={"index": year_column})

        wide_df = metric_data.copy()
        results_wide.append(wide_df)

        long_df = wide_df.melt(id_vars=[year_column, "Group"], var_name="Metric", value_name="Value")
        results_long.append(long_df)

    wide_result = pd.concat(results_wide, ignore_index=True)
    long_result = pd.concat(results_long, ignore_index=True)

    if return_format == "long":
        return long_result
    elif return_format == "wide":
        return wide_result
    elif return_format == "both":
        return {"wide": wide_result, "long": long_result}

"""Utilities related to Group analysis in the Biblium project."""
# Group analysis

def get_scientific_production_by_group0(
    df,
    group_matrix,
    relative_counts = True,
    cumulative = True,
    predict_last_year = True,
    percent_change = True,
    output_format = 'both',
    rename_wide_columns = True,
):
    """
    Computes the annual scientific production statistics separately for each group and returns the results in long, wide, or both formats.

    Parameters:
    df (pd.DataFrame): DataFrame containing at least "Year" and "Cited by" columns.
    group_matrix (pd.DataFrame): Binary DataFrame where columns are group names and rows align with df rows.
    relative_counts (bool, optional): Whether to compute relative proportions and percentages of documents. Default is True.
    cumulative (bool, optional): Whether to compute cumulative document and citation counts. Default is True.
    predict_last_year (bool, optional): Whether to predict the current year's document and citation counts. Default is True.
    percent_change (bool, optional): Whether to compute year-over-year percentage change in documents and citations. Default is True.
    output_format (str, optional): Format of the result: "long", "wide", or "both". Default is "both".
    rename_wide_columns (bool, optional): Whether to rename wide-format columns to include group names. Default is True.

    Returns:
    pd.DataFrame or dict: A single DataFrame in long or wide format, or a dictionary with both if output_format="both".
    """
    group_dfs = []
    all_years = pd.Series(range(df["Year"].min(), df["Year"].max() + 1), name="Year")

    for group in group_matrix.columns:
        group_rows = group_matrix[group] == 1
        if group_rows.any():
            group_df = df[group_rows].copy()
            production = get_scientific_production(
                group_df,
                relative_counts=relative_counts,
                cumulative=cumulative,
                predict_last_year=predict_last_year,
                percent_change=percent_change
            )
            production = all_years.to_frame().merge(production, on="Year", how="left").fillna(0)
            production["Group"] = group
            group_dfs.append(production)

    if not group_dfs:
        return pd.DataFrame()

    long_df = pd.concat(group_dfs, ignore_index=True)

    if output_format == "long":
        return long_df

    elif output_format == "wide":
        id_vars = ["Year", "Group"]
        value_vars = [col for col in long_df.columns if col not in id_vars]
        wide_df = long_df.pivot(index="Year", columns="Group", values=value_vars)
        wide_df = wide_df.fillna(0)

        if rename_wide_columns:
            wide_df.columns = [f"{val} (" + f"{grp})" for val, grp in wide_df.columns]
            wide_df.columns.name = None

        return wide_df

    elif output_format == "both":
        id_vars = ["Year", "Group"]
        value_vars = [col for col in long_df.columns if col not in id_vars]
        wide_df = long_df.pivot(index="Year", columns="Group", values=value_vars)
        wide_df = wide_df.fillna(0)

        if rename_wide_columns:
            wide_df.columns = [f"{val} (" + f"{grp})" for val, grp in wide_df.columns]
            wide_df.columns.name = None

        return {"long": long_df, "wide": wide_df}

    else:
        raise ValueError("output_format must be one of: \"long\", \"wide\", \"both\"")

"""Utilities related to Bibliographic laws in the Biblium project."""
# Bibliographic laws

def compute_lotka_distribution(
    df,
    author_col = 'Authors',
    separator = '; ',
):
    """
    Compute author productivity distribution and expected values under Lotka's law.

    Parameters:
        df (pd.DataFrame): DataFrame containing author information.
        author_col (str): Column name where author names are listed.
        separator (str): Separator between multiple authors in a single cell.

    Returns:
        pd.DataFrame: A DataFrame with observed and expected author productivity.
    """
    # Flatten list of authors across all papers
    all_authors = df[author_col].dropna().str.split(separator).explode()

    # Count number of publications per author
    author_counts = Counter(all_authors)

    # Count how many authors published n papers
    productivity = Counter(author_counts.values())

    # Convert to DataFrame
    lotka_df = pd.DataFrame(sorted(productivity.items()), columns=["n_pubs", "n_authors"])

    # Normalize and apply Lotka's expected law: expected ~ C / n^2
    C = lotka_df["n_authors"].iloc[0]  # authors with 1 publication
    lotka_df["expected_n_authors"] = C / (lotka_df["n_pubs"] ** 2)

    return lotka_df

def evaluate_lotka_fit(
    lotka_df,
):
    """
    Compute fit statistics comparing observed and expected values under Lotka's Law.

    Parameters:
        lotka_df (pd.DataFrame): Output of compute_lotka_distribution.

    Returns:
        dict: Dictionary with R2, RMSE, and KS statistic and p-value.
    """
    observed = lotka_df["n_authors"]
    expected = lotka_df["expected_n_authors"]

    r2 = r2_score(observed, expected)
    rmse = np.sqrt(mean_squared_error(observed, expected))
    ks_stat, ks_pvalue = ks_2samp(observed, expected)

    metrics = {
        "R2": r2,
        "RMSE": rmse,
        "KS_statistic": ks_stat,
        "KS_pvalue": ks_pvalue
    }
    return pd.DataFrame(list(metrics.items()),
                      columns=["Measure", "Value"])

def compute_bradford_distribution(
    df,
    source_col = 'Source title',
    zone_count = 3,
    lowercase = False,
):
    """
    Compute Bradford's Law distribution by dividing sources into zones.
    """
    sources = df[source_col].dropna()
    if lowercase:
        sources = sources.str.lower()

    source_counts = sources.value_counts().reset_index()
    source_counts.columns = ["Source", "Document_Count"]

    source_counts["Cumulative_Documents"] = source_counts["Document_Count"].cumsum()
    source_counts["Cumulative_Percentage"] = source_counts["Cumulative_Documents"] / source_counts["Document_Count"].sum() * 100

    total_documents = source_counts["Document_Count"].sum()
    documents_per_zone = total_documents / zone_count
    zones = []
    current_zone = 1
    documents_in_current_zone = 0

    for count in source_counts["Document_Count"]:
        documents_in_current_zone += count
        zones.append(current_zone)
        if documents_in_current_zone >= documents_per_zone and current_zone < zone_count:
            current_zone += 1
            documents_in_current_zone = 0

    source_counts["Zone"] = zones

    return source_counts


def evaluate_bradford_fit(
    source_counts,
    zone_count = 3,
):
    """
    Evaluate how well data fits Bradford's Law.
    
    Computes deviation from expected equal distribution
    of documents across zones.
    
    Parameters
    ----------
    source_counts : pd.DataFrame
        Source counts with 'Zone' and 'Document_Count' columns.
    zone_count : int
        Number of zones to evaluate.
        
    Returns
    -------
    pd.DataFrame
        Statistics including documents per zone, expected,
        deviations, and mean deviation.
    """
    zone_stats = source_counts.groupby("Zone")["Document_Count"].sum()
    total_documents = source_counts["Document_Count"].sum()
    expected_per_zone = total_documents / zone_count
    deviations = (zone_stats - expected_per_zone).abs() / expected_per_zone

    return pd.DataFrame({"Documents per Zone": zone_stats.to_dict(),
        "Expected Documents per Zone": expected_per_zone,
        "Deviation per Zone": deviations.to_dict(),
        "Mean Deviation": deviations.mean()})

def compute_zipf_distribution_from_counts(
    df,
    word_col = 0,
    count_col = 1,
):
    """
    Compute Zipf's Law distribution given a DataFrame with word/item counts.
    """
    if isinstance(word_col, int):
        word_col = df.columns[word_col]
    if isinstance(count_col, int):
        count_col = df.columns[count_col]

    zipf_df = df[[word_col, count_col]].copy()
    zipf_df.columns = ["Word", "Frequency"]
    zipf_df = zipf_df.sort_values(by="Frequency", ascending=False).reset_index(drop=True)
    zipf_df["Rank"] = np.arange(1, len(zipf_df) + 1)
    return zipf_df

def evaluate_zipf_fit(
    zipf_df,
):
    """
    Evaluate Zipf's Law fit statistics.
    """
    log_rank = np.log(zipf_df["Rank"])
    log_freq = np.log(zipf_df["Frequency"])

    r2 = r2_score(log_freq, -log_rank)
    rmse = np.sqrt(mean_squared_error(log_freq, -log_rank))
    ks_stat, ks_pvalue = ks_2samp(log_freq, -log_rank)

    return {
        "R2": r2,
        "RMSE": rmse,
        "KS_statistic": ks_stat,
        "KS_pvalue": ks_pvalue
    }

def evaluate_prices_law(
    author_counts,
):
    """
    Evaluate Price's Law: check if square root of authors produces 50% of the documents.

    Parameters:
        author_counts (pd.Series): Series with authors as index and number of documents as values.

    Returns:
        dict: Core size, actual proportion produced by core, ideal square root size.
    """
    sorted_counts = author_counts.sort_values(ascending=False)
    total_docs = sorted_counts.sum()
    total_authors = len(sorted_counts)
    ideal_core_size = int(np.sqrt(total_authors))
    actual_core_docs = sorted_counts.iloc[:ideal_core_size].sum()
    actual_proportion = actual_core_docs / total_docs

    return {
        "Total Authors": total_authors,
        "Ideal Core Size (sqrt N)": ideal_core_size,
        "Core Documents": actual_core_docs,
        "Proportion from Core": actual_proportion
    }

def evaluate_pareto_principle(
    counts,
    top_percentage = 20,
    outcome_percentage = 80,
):
    """
    Evaluate Pareto Principle (default 80/20 rule).

    Parameters:
        counts (pd.Series): Series of items and their counts.
        top_percentage (float): Percentage of top items (default 20).
        outcome_percentage (float): Expected percentage of outcomes (default 80).

    Returns:
        dict: Top items needed, actual outcome proportion, comparison with expected.
    """
    sorted_counts = counts.sort_values(ascending=False)
    total = sorted_counts.sum()
    top_n = int(np.ceil(top_percentage / 100 * len(sorted_counts)))
    top_sum = sorted_counts.iloc[:top_n].sum()
    actual_percentage = top_sum / total * 100

    return {
        "Total Items": len(sorted_counts),
        "Top Items Needed": top_n,
        "Actual Outcome %": actual_percentage,
        "Expected Outcome %": outcome_percentage
    }

"""Utilities related to Association group analysis in the Biblium project."""
# Association group analysis

def compute_binary_associations(
    groups_df: pd.DataFrame,
    items_df: pd.DataFrame,
    association_measures: Optional[List[str]] = None,
    p_adjust_method: str = 'fdr_bh',
    min_count: int = 1,
    min_jaccard: float = 0.0,
    significance_level: float = 0.05,
    treat_na_as_zero: bool = False,
    output_format: str = 'long',
    filters: Optional[dict] = None,
) -> pd.DataFrame:
    """
    Compute binary association measures between groups and items.
    
    Calculates various association metrics (Jaccard, Dice, Phi, etc.)
    between binary group membership and item occurrence matrices.
    
    Parameters
    ----------
    groups_df : pd.DataFrame
        Binary matrix of group memberships (rows=docs, cols=groups).
    items_df : pd.DataFrame
        Binary matrix of item occurrences (rows=docs, cols=items).
    association_measures : list, optional
        List of measures to compute. Defaults to common measures.
    p_adjust_method : str
        Multiple testing correction method ('fdr_bh', 'bonferroni', etc.).
    min_count : int
        Minimum co-occurrence count to include.
    min_jaccard : float
        Minimum Jaccard index to include.
    significance_level : float
        Significance threshold for statistical tests.
    treat_na_as_zero : bool
        If True, treat NA values as 0.
    output_format : str
        Output format: 'long' or 'wide'.
    filters : dict, optional
        Additional filters for results.
        
    Returns
    -------
    pd.DataFrame
        Association measures between groups and items.
    """
    if association_measures is None:
        association_measures = [
            "Jaccard", "Sokal-Michener", "Dice", "Yule's Q", "Phi",
            "Odds Ratio", "Kulczynski", "Ochiai", "Cosine",
            "Conditional on Group (a/b)", "Conditional on Feature (a/c)",
            "Chi2 p", "Relative Risk", "Cramer's V"
        ]

    def calculate_measures(
        a,
        b,
        c,
        d,
    ):
        n = a + b + c + d
        measures = {}
        if "Jaccard" in association_measures:
            measures["Jaccard"] = a / (a + b + c) if (a + b + c) else np.nan
        if "Sokal-Michener" in association_measures:
            measures["Sokal-Michener"] = (a + d) / n if n else np.nan
        if "Dice" in association_measures:
            measures["Dice"] = 2 * a / (2 * a + b + c) if (2 * a + b + c) else np.nan
        if "Yule's Q" in association_measures:
            measures["Yule's Q"] = (a * d - b * c) / (a * d + b * c) if (a * d + b * c) else np.nan
        if "Phi" in association_measures:
            denom = np.sqrt((a + b) * (a + c) * (b + d) * (c + d))
            measures["Phi"] = (a * d - b * c) / denom if denom else np.nan
        if "Odds Ratio" in association_measures:
            measures["Odds Ratio"] = (a * d / (b * c)) if b * c != 0 else np.nan
        if "Kulczynski" in association_measures:
            measures["Kulczynski"] = 0.5 * (a / (a + b) + a / (a + c)) if (a + b) and (a + c) else np.nan
        if "Ochiai" in association_measures:
            measures["Ochiai"] = a / np.sqrt((a + b) * (a + c)) if (a + b) and (a + c) else np.nan
        if "Cosine" in association_measures:
            measures["Cosine"] = a / np.sqrt((a + b) * (a + c)) if (a + b) and (a + c) else np.nan
        if "Conditional on Group (a/b)" in association_measures:
            measures["Conditional on Group (a/b)"] = a / b if b else np.nan
        if "Conditional on Feature (a/c)" in association_measures:
            measures["Conditional on Feature (a/c)"] = a / c if c else np.nan
        if "Relative Risk" in association_measures:
            measures["Relative Risk"] = (a / (a + b)) / (c / (c + d)) if (a + b) and (c + d) else np.nan
        if "Cramer's V" in association_measures:
            chi2 = (a * d - b * c) ** 2 * n / ((a + b) * (c + d) * (a + c) * (b + d)) if n else np.nan
            measures["Cramer's V"] = np.sqrt(chi2 / n) if n else np.nan
        return measures

    results = []
    fisher_pvals = []

    for group in groups_df.columns:
        for item in items_df.columns:
            x = groups_df[group]
            y = items_df[item]

            pair = pd.concat([x, y], axis=1)
            if not treat_na_as_zero:
                pair = pair.dropna()
            else:
                pair = pair.fillna(0)

            x_vals = pair.iloc[:, 0].astype(int).values
            y_vals = pair.iloc[:, 1].astype(int).values

            a = np.sum((x_vals == 1) & (y_vals == 1))
            b = np.sum((x_vals == 1) & (y_vals == 0))
            c = np.sum((x_vals == 0) & (y_vals == 1))
            d = np.sum((x_vals == 0) & (y_vals == 0))

            if a < min_count:
                continue

            row = {
                "Group": group,
                "Item": item,
                "a": a, "b": b, "c": c, "d": d
            }

            row.update(calculate_measures(a, b, c, d))

            try:
                _, fisher_p = fisher_exact([[a, b], [c, d]])
            except:
                fisher_p = np.nan
            row["Fisher p"] = fisher_p
            fisher_pvals.append(fisher_p)

            try:
                _, chi2_p, _, _ = chi2_contingency([[a, b], [c, d]], correction=False)
            except:
                chi2_p = np.nan
            row["Chi2 p"] = chi2_p

            results.append(row)

    df = pd.DataFrame(results)

    if not df.empty:
        df["P-adj"] = np.nan
        mask = df["Fisher p"].notnull()
        adjusted = multipletests(df.loc[mask, "Fisher p"].values, method=p_adjust_method)
        df.loc[mask, "P-adj"] = adjusted[1]
        df["Significant"] = df["P-adj"] < significance_level

        if filters:
            for key, val in filters.items():
                if key in df.columns:
                    df = df[df[key] >= val]

        if "Jaccard" in df.columns:
            df = df[df["Jaccard"] >= min_jaccard]

        if output_format == "wide":
            df = df.pivot(index="Group", columns="Item")

    return df

"""Utilities related to Additional normalization methods for a given symetric dataframe in the Biblium project."""
# additional normalization methods for a given symetric dataframe

def normalize_symmetric_matrix(
    matrix_df,
    method = 'jaccard',
):
    """
    Normalizes a symmetric co-occurrence matrix using the specified method.

    Supported methods:
    - "jaccard": Jaccard similarity
    - "cosine": Cosine similarity
    - "row": Row-wise proportion (normalized by row sums)
    - "none": No normalization, returns a copy

    Parameters:
    matrix_df (pd.DataFrame): Symmetric matrix (e.g. country collaboration).
    method (str): Normalization method ("jaccard", "cosine", "row", "none").

    Returns:
    pd.DataFrame: Normalized matrix with same index and columns.
    """
    if matrix_df.empty:
        return pd.DataFrame()

    if method == "none":
        return matrix_df.copy()

    if method == "row":
        row_sums = matrix_df.sum(axis=1).replace(0, 1)
        return matrix_df.div(row_sums, axis=0)

    if method == "cosine":
        from sklearn.metrics.pairwise import cosine_similarity
        similarity = cosine_similarity(matrix_df)
        return pd.DataFrame(similarity, index=matrix_df.index, columns=matrix_df.columns)

    if method == "jaccard":
        index = matrix_df.index
        row_sums = matrix_df.sum(axis=1)
        jaccard_df = pd.DataFrame(0.0, index=index, columns=index)

        for i in index:
            for j in index:
                if i == j:
                    jaccard_df.loc[i, j] = 1.0
                else:
                    numerator = matrix_df.loc[i, j]
                    denominator = row_sums[i] + row_sums[j] - numerator
                    jaccard_df.loc[i, j] = numerator / denominator if denominator > 0 else 0.0

        return jaccard_df

"""Helpers for exploratory factor analysis and related matrix decompositions."""
# Factor analysis

def build_document_term_matrix(
    df: pd.DataFrame,
    field: str = 'Author Keywords',
    method: str = 'count',
    min_df: int = 2,
    ngram_range: tuple = (1, 1),
    use_lemmatization: bool = False,
    pos_filter: list = None,
) -> pd.DataFrame:
    """
    Build a document-term matrix with optional TF-IDF, n-grams, and lemmatization/POS filtering.
    """
    texts = df[field].fillna("")
    processed_texts = []
    for doc in texts:
        if use_lemmatization and nlp is not None:
            tokens = []
            for token in nlp(doc):
                if pos_filter and token.pos_ not in pos_filter:
                    continue
                tokens.append(token.lemma_)
            processed_texts.append(" ".join(tokens))
        else:
            processed_texts.append(doc)

    Vectorizer = TfidfVectorizer if method == "tfidf" else CountVectorizer
    vectorizer = Vectorizer(
        token_pattern=r"(?u)\b\w+\b",
        min_df=min_df,
        ngram_range=ngram_range
    )
    X = vectorizer.fit_transform(processed_texts)
    return pd.DataFrame(
        X.toarray(),
        index=df.index,
        columns=vectorizer.get_feature_names_out()
    )

def suggest_k(
    term_coords: np.ndarray,
    min_k: int = 2,
    max_k: int = 10,
) -> dict:
    """
    Suggest optimal number of clusters using silhouette scores for KMeans.

    Examples
    --------
    >>> result = conceptual_structure_analysis(df)
    >>> suggest_k(result['term_embeddings'], min_k=2, max_k=8)
    """
    scores = {}
    max_k = min(max_k, len(term_coords) - 1)
    for k in range(min_k, max_k + 1):
        labels = KMeans(n_clusters=k, random_state=0).fit_predict(term_coords)
        scores[k] = silhouette_score(term_coords, labels)
    return scores

def conceptual_structure_analysis(
    df: pd.DataFrame,
    field: str = "Author Keywords",
    dr_method: str = "MCA",
    cluster_method: str = "kmeans",
    n_clusters: int = 5,
    n_terms: int = 100,
    n_components: int = 2,
    dtm_method: str = "count",
    term_selection: str = "frequency",
    y: np.ndarray | None = None,
    min_df: int = 2,
    ngram_range: tuple = (1, 1),
    use_lemmatization: bool = False,
    pos_filter: list | None = None,
    include_terms: list | None = None,
    exclude_terms: list | None = None,
    term_regex: str | None = None,
    compute_metrics: bool = False,
    excel_path: str | None = None,
    keyword_separator: str = "; ",
) -> dict:
    """
    Perform conceptual structure analysis with flexible dimensionality
    reduction, clustering, term selection, and advanced term filtering.

    The function is aware of different text field types (keywords, abstracts,
    titles) and handles them as follows:

    * Keyword fields (e.g. "Author Keywords", "Index Keywords" and their
      processed variants) treat complete keywords/phrases as atomic terms.
      Keywords are split only by ``keyword_separator`` (default "; ").
      Parentheses are preserved and unbalanced "(" are fixed by appending
      a ")" (same trick as in ``count_occurrences``).
    * Free-text fields (e.g. "Abstract", "Title") use standard
      scikit-learn tokenisation. If processed variants exist (e.g.
      "Processed Abstract", "Processed Title"), they are preferred.

    Optionally, the main tabular results are written to an Excel file with
    multiple sheets.

    Parameters
    ----------
    df : pd.DataFrame
        Source data.
    field : str, default "Author Keywords"
        Column used for text analysis. The function resolves and prefers
        processed variants when available (for example "Processed Author
        Keywords", "Processed Abstract").
    dr_method : {"MCA", "CA", "MDS", "PCA", "LSA", "t-SNE", "UMAP", "NMF", "LDA"}, \
            default "MCA"
        Dimensionality reduction method.
    cluster_method : {"kmeans", "agglomerative", "dbscan", "spectral", "louvain"}, \
            default "kmeans"
        Clustering method.
    n_clusters : int, default 5
        Number of clusters (where applicable).
    n_terms : int, default 100
        Target number of terms to select.
    n_components : int, default 2
        Number of dimensions for embeddings.
    dtm_method : {"count", "tfidf"}, default "count"
        Document–term matrix method.
    term_selection : {"frequency", "chi2", "mutual_info"}, default "frequency"
        Term selection method.
    y : np.ndarray or None, default None
        Target labels for supervised term selection.
    min_df : int, default 2
        Minimum document frequency for term inclusion.
    ngram_range : tuple, default (1, 1)
        N-gram range (only used for non-keyword free-text fields).
    use_lemmatization : bool, default False
        Placeholder flag for lemmatisation on free-text fields.
    pos_filter : list or None, default None
        Placeholder for POS-based filtering for free-text fields.
    include_terms, exclude_terms, term_regex :
        Advanced term filtering arguments.
    compute_metrics : bool, default False
        If True, compute clustering quality metrics.
    excel_path : str or None, default None
        If provided, write tables to an Excel file with sheets:
        "terms", "clusters_terms", "dtm_top", "documents", "metrics",
        "config".
    keyword_separator : str, default "; "
        Separator used to split keyword fields into individual keywords.

    Returns
    -------
    dict
        Result dictionary with embeddings, labels, tables, metrics and
        ``suggested_k``.
    """
    import re
    import numpy as np
    import pandas as pd
    import networkx as nx
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.feature_selection import SelectKBest, chi2, mutual_info_classif
    from sklearn.decomposition import TruncatedSVD, NMF, LatentDirichletAllocation
    from sklearn.manifold import MDS, TSNE
    from sklearn.cluster import (
        KMeans,
        AgglomerativeClustering,
        DBSCAN,
        SpectralClustering,
    )
    from sklearn.metrics import silhouette_score, davies_bouldin_score

    try:
        import prince
    except ImportError:
        prince = None


    def _infer_field_type(col_name: str) -> str:
        """Infer logical field type from column name."""
        lname = col_name.lower()
        if "author keywords" in lname:
            return "author_keywords"
        if "index keywords" in lname:
            return "index_keywords"
        if "keywords" in lname:
            return "keywords"
        if "abstract" in lname:
            return "abstract"
        if "title" in lname:
            return "title"
        return "generic"

    def _resolve_effective_field(df_obj: pd.DataFrame, field_name: str) -> tuple[str, str]:
        """
        Resolve actual column for a logical field name.

        Tries to match case-insensitively and prefers processed variants
        (e.g. "Processed Author Keywords", "Processed Abstract").
        """
        if field_name in df_obj.columns:
            col = field_name
        else:
            matches = [c for c in df_obj.columns if isinstance(c, str) and c.lower() == field_name.lower()]
            col = matches[0] if matches else field_name

        field_type = _infer_field_type(col)

        if field_type in {"author_keywords", "index_keywords", "keywords"}:
            if field_type == "author_keywords":
                base = "Author Keywords"
            elif field_type == "index_keywords":
                base = "Index Keywords"
            else:
                base = "Keywords"
            preferred = f"Processed {base}"
            if preferred in df_obj.columns:
                col = preferred
            else:
                candidates = [
                    c
                    for c in df_obj.columns
                    if isinstance(c, str) and "processed" in c.lower() and "keyword" in c.lower()
                ]
                if candidates:
                    col = candidates[0]
        elif field_type in {"abstract", "title"}:
            base = "Abstract" if field_type == "abstract" else "Title"
            preferred = f"Processed {base}"
            if preferred in df_obj.columns:
                col = preferred
        else:
            preferred = f"Processed {field_name}"
            if preferred in df_obj.columns:
                col = preferred

        field_type = _infer_field_type(col)
        if col not in df_obj.columns:
            raise ValueError(
                f"Column {col!r} (resolved from {field_name!r}) not found in DataFrame."
            )
        return col, field_type

    def _build_document_term_matrix(
        df_obj: pd.DataFrame,
        column: str,
        method: str,
        min_df_val: int,
        ngram_range_val: tuple,
        use_lemmatization_val: bool,
        pos_filter_val: list | None,
        field_type_val: str,
        kw_sep: str,
    ) -> pd.DataFrame:
        """
        Build a document–term matrix.

        Keyword fields:
            * Keep complete keywords/phrases as atomic tokens.
            * Split only by kw_sep.
            * Balance closing parentheses as in ``count_occurrences``.

        Free-text fields:
            * Use standard scikit-learn tokenisation.
        """
        is_keyword_field = field_type_val in {
            "author_keywords",
            "index_keywords",
            "keywords",
        }

        if is_keyword_field:
            series = df_obj[column].fillna("")

            def _to_terms(v) -> list[str]:
                if isinstance(v, (list, tuple, set)):
                    tokens = list(v)
                else:
                    tokens = str(v).split(kw_sep)

                terms: list[str] = []
                for item in tokens:
                    s = str(item).strip()
                    if not s:
                        continue
                    s = _balance_closing_parenthesis(s)
                    terms.append(s)
                return terms

            term_lists = series.apply(_to_terms)

            if method == "count":
                vectorizer = CountVectorizer(
                    tokenizer=lambda x: x,
                    preprocessor=lambda x: x,
                    token_pattern=None,
                    min_df=min_df_val,
                    lowercase=False,
                )
            elif method == "tfidf":
                vectorizer = TfidfVectorizer(
                    tokenizer=lambda x: x,
                    preprocessor=lambda x: x,
                    token_pattern=None,
                    min_df=min_df_val,
                    lowercase=False,
                )
            else:
                raise ValueError(f"Unknown DTM method: {method}")

            X = vectorizer.fit_transform(term_lists)
            columns = [t.strip() for t in vectorizer.get_feature_names_out()]
            return pd.DataFrame(X.toarray(), columns=columns)

        # Free-text fields
        if method == "count":
            vectorizer = CountVectorizer(
                min_df=min_df_val,
                ngram_range=ngram_range_val,
            )
        elif method == "tfidf":
            vectorizer = TfidfVectorizer(
                min_df=min_df_val,
                ngram_range=ngram_range_val,
            )
        else:
            raise ValueError(f"Unknown DTM method: {method}")

        docs = df_obj[column].fillna("").astype(str).values
        X = vectorizer.fit_transform(docs)
        columns = vectorizer.get_feature_names_out()
        return pd.DataFrame(X.toarray(), columns=columns)

    def _get_clusterer(method: str, k: int):
        if method == "kmeans":
            return KMeans(n_clusters=k, random_state=0)
        if method == "agglomerative":
            return AgglomerativeClustering(n_clusters=k)
        if method == "dbscan":
            return DBSCAN()
        if method == "spectral":
            return SpectralClustering(
                n_clusters=k,
                assign_labels="discretize",
                random_state=0,
            )
        raise ValueError(f"Unknown cluster method: {method}")

    # ------------------------------------------------------------------
    # Step 1: resolve field and build DTM
    # ------------------------------------------------------------------
    effective_field, field_type = _resolve_effective_field(df, field)

    dtm = _build_document_term_matrix(
        df_obj=df,
        column=effective_field,
        method=dtm_method,
        min_df_val=min_df,
        ngram_range_val=ngram_range,
        use_lemmatization_val=use_lemmatization,
        pos_filter_val=pos_filter,
        field_type_val=field_type,
        kw_sep=keyword_separator,
    )

    if dtm.shape[1] == 0:
        raise ValueError(
            "Document–term matrix has no terms. Check your field, separator, and preprocessing."
        )

    # ------------------------------------------------------------------
    # Step 2: term selection
    # ------------------------------------------------------------------
    if term_selection == "frequency":
        freqs = dtm.sum(axis=0).values
        top_idx = np.argsort(freqs)[::-1][:n_terms]
        selected_terms = list(dtm.columns[top_idx])
    elif term_selection in {"chi2", "mutual_info"}:
        if y is None:
            raise ValueError("Argument 'y' must be provided for supervised term selection.")
        k = min(n_terms, dtm.shape[1])
        score_func = chi2 if term_selection == "chi2" else mutual_info_classif
        selector = SelectKBest(score_func=score_func, k=k)
        selector.fit(dtm, y)
        selected_terms = list(dtm.columns[selector.get_support()])
    else:
        raise ValueError(f"Unknown term_selection: {term_selection}")

    if include_terms:
        for term in include_terms:
            if term in dtm.columns and term not in selected_terms:
                selected_terms.append(term)

    if exclude_terms:
        selected_terms = [t for t in selected_terms if t not in exclude_terms]

    if term_regex:
        pattern = re.compile(term_regex)
        selected_terms = [t for t in selected_terms if pattern.search(t)]

    # Deduplicate & truncate
    seen: set[str] = set()
    unique_terms: list[str] = []
    for t in selected_terms:
        if t not in seen:
            unique_terms.append(t)
            seen.add(t)
    selected_terms = unique_terms[:n_terms]

    if not selected_terms:
        raise ValueError(
            "No terms left after filtering; relax filters or adjust 'n_terms'."
        )

    dtm_top = dtm[selected_terms]

    # Cleaned terms for *output / plotting*
    clean_terms = [_balance_closing_parenthesis(t) for t in dtm_top.columns]

    # ------------------------------------------------------------------
    # Step 3: dimensionality reduction
    # ------------------------------------------------------------------
    doc_coords = None
    metrics: dict[str, float] = {}

    if dr_method in {"MCA", "CA"}:
        if prince is None:
            raise ImportError("prince is required for dr_method='MCA' or 'CA'.")

        # Handle NaN and infinite values before fitting
        dtm_top = dtm_top.fillna(0)
        dtm_top = dtm_top.replace([np.inf, -np.inf], 0)
        
        # For MCA, need to ensure data is suitable (no constant columns, no NaN after processing)
        # Remove columns with zero variance (constant columns cause issues in MCA)
        col_std = dtm_top.std()
        non_constant_cols = col_std[col_std > 0].index.tolist()
        if len(non_constant_cols) < 2:
            raise ValueError("Not enough variable columns for dimensionality reduction")
        dtm_top = dtm_top[non_constant_cols]
        
        # Convert to float and ensure clean data
        dtm_top = dtm_top.astype(float)
        dtm_top = pd.DataFrame(
            np.nan_to_num(dtm_top.values, nan=0.0, posinf=0.0, neginf=0.0),
            columns=dtm_top.columns,
            index=dtm_top.index
        )

        # Remove rows and columns with zero sums (cause divide by zero in CA/MCA)
        row_sums = dtm_top.sum(axis=1)
        col_sums = dtm_top.sum(axis=0)
        valid_rows = row_sums > 0
        valid_cols = col_sums > 0
        
        if valid_rows.sum() < 2 or valid_cols.sum() < 2:
            raise ValueError(
                f"Not enough non-zero rows ({valid_rows.sum()}) or columns ({valid_cols.sum()}) "
                f"for MCA/CA. Try lowering min_df or increasing n_terms."
            )
        
        dtm_clean = dtm_top.loc[valid_rows, valid_cols].copy()
        
        # Adjust n_components if needed
        max_components = min(dtm_clean.shape[0], dtm_clean.shape[1]) - 1
        actual_components = min(n_components, max_components)
        if actual_components < 1:
            raise ValueError(
                f"Matrix too small for dimensionality reduction ({dtm_clean.shape}). "
                f"Try increasing n_terms or lowering min_df."
            )
        
        if dr_method == "MCA":
            model = prince.MCA(n_components=actual_components)
        else:
            model = prince.CA(n_components=actual_components)
        
        try:
            model = model.fit(dtm_clean)
        except Exception as e:
            raise ValueError(f"MCA/CA fitting failed: {e}. Try different parameters.")
        
        prince_terms = model.column_coordinates(dtm_clean)
        
        # Use dtm_clean for subsequent operations
        dtm_top = dtm_clean

        term_coords_list: list[np.ndarray] = []
        for term in dtm_top.columns:
            if term in prince_terms.index:
                coords = prince_terms.loc[term].values
            else:
                matches = [idx for idx in prince_terms.index if str(idx).startswith(str(term))]
                if matches:
                    coords = prince_terms.loc[matches[0]].values
                else:
                    coords = np.zeros(n_components, dtype=float)
            term_coords_list.append(coords)

        term_coords = np.vstack(term_coords_list)
        doc_coords = model.row_coordinates(dtm_top).values

    elif dr_method == "MDS":
        corr = np.corrcoef(dtm_top.T.values)
        corr = np.nan_to_num(corr, nan=0.0)
        np.fill_diagonal(corr, 1.0)
        dist = 1.0 - corr
        mds = MDS(
            n_components=n_components,
            dissimilarity="precomputed",
            random_state=0,
        )
        term_coords = mds.fit_transform(dist)
        if hasattr(mds, "stress_"):
            metrics["stress"] = float(mds.stress_)

    elif dr_method in {"PCA", "LSA"}:
        svd = TruncatedSVD(n_components=n_components, random_state=0)
        doc_coords = svd.fit_transform(dtm_top)
        term_coords = svd.components_.T

    elif dr_method == "t-SNE":
        tsne = TSNE(n_components=n_components, random_state=0)
        term_coords = tsne.fit_transform(dtm_top.T.values)
        if hasattr(tsne, "kl_divergence_"):
            metrics["kl_divergence"] = float(tsne.kl_divergence_)

    elif dr_method == "UMAP":
        try:
            import umap
        except ImportError as exc:
            raise ImportError("umap-learn is required for dr_method='UMAP'.") from exc
        reducer = umap.UMAP(n_components=n_components, random_state=0)
        term_coords = reducer.fit_transform(dtm_top.T.values)

    elif dr_method == "NMF":
        nmf = NMF(n_components=n_components, random_state=0)
        doc_coords = nmf.fit_transform(dtm_top)
        term_coords = nmf.components_.T

    elif dr_method == "LDA":
        lda = LatentDirichletAllocation(n_components=n_components, random_state=0)
        doc_coords = lda.fit_transform(dtm_top)
        term_coords = lda.components_.T

    else:
        raise ValueError(f"Unknown dr_method: {dr_method}")

    # ------------------------------------------------------------------
    # Step 4: clustering of terms
    # ------------------------------------------------------------------
    if cluster_method == "louvain":
        try:
            import community as community_louvain
        except ImportError as exc:
            raise ImportError(
                "python-louvain is required for cluster_method='louvain'."
            ) from exc

        cooccur = (dtm_top.T.dot(dtm_top) > 0).astype(int)
        cooccur_df = pd.DataFrame(
            cooccur,
            index=dtm_top.columns,
            columns=dtm_top.columns,
        )
        G = nx.from_pandas_adjacency(cooccur_df)
        partition = community_louvain.best_partition(G)
        term_labels = np.array([partition.get(term, -1) for term in dtm_top.columns])
        metrics["inertia"] = None
    else:
        clusterer_terms = _get_clusterer(cluster_method, n_clusters)
        term_labels = clusterer_terms.fit_predict(term_coords)
        if hasattr(clusterer_terms, "inertia_"):
            metrics["inertia"] = float(clusterer_terms.inertia_)

    # ------------------------------------------------------------------
    # Step 5: optional document clustering
    # ------------------------------------------------------------------
    doc_labels = None
    if doc_coords is not None and cluster_method != "louvain":
        clusterer_docs = _get_clusterer(cluster_method, n_clusters)
        doc_labels = clusterer_docs.fit_predict(doc_coords)

    # ------------------------------------------------------------------
    # Step 6: diagnostics
    # ------------------------------------------------------------------
    if compute_metrics:
        unique_labels = np.unique(term_labels)
        if len(unique_labels) > 1 and len(term_labels) > len(unique_labels):
            try:
                metrics["silhouette"] = float(
                    silhouette_score(term_coords, term_labels)
                )
            except Exception:
                pass
            try:
                metrics["davies_bouldin"] = float(
                    davies_bouldin_score(term_coords, term_labels)
                )
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Step 7: assemble result tables (using CLEANED term labels)
    # ------------------------------------------------------------------
    dim_cols = [f"dim{i + 1}" for i in range(n_components)]
    terms_df = pd.DataFrame(term_coords, columns=dim_cols)
    terms_df.insert(0, "term", clean_terms)
    terms_df.insert(1, "label", term_labels)

    label_terms_dict = (
        terms_df.groupby("label")["term"].apply(list).to_dict()
    )

    cluster_dict: dict[str, list] = {}
    for label, terms_list in sorted(label_terms_dict.items(), key=lambda kv: kv[0]):
        cluster_name = f"Cluster{label}"
        cluster_dict[cluster_name] = terms_list

    if cluster_dict:
        max_len = max(len(v) for v in cluster_dict.values())
        for key in cluster_dict:
            cluster_dict[key] = cluster_dict[key] + [None] * (
                max_len - len(cluster_dict[key])
            )
        clusters_df = pd.DataFrame(cluster_dict).fillna("")
    else:
        clusters_df = pd.DataFrame()

    # ------------------------------------------------------------------
    # Step 8: optional Excel export
    # ------------------------------------------------------------------
    if excel_path is not None:
        with pd.ExcelWriter(excel_path) as writer:
            terms_df.to_excel(writer, sheet_name="terms", index=False)
            clusters_df.to_excel(writer, sheet_name="clusters_terms", index=False)
            dtm_top.to_excel(writer, sheet_name="dtm_top", index=False)

            if doc_coords is not None:
                doc_df = pd.DataFrame(doc_coords, columns=dim_cols)
                if doc_labels is not None:
                    doc_df.insert(0, "cluster", doc_labels)
                doc_df.to_excel(writer, sheet_name="documents", index=False)

            if metrics:
                metrics_df = pd.DataFrame(
                    [{"metric": k, "value": v} for k, v in metrics.items()]
                )
                metrics_df.to_excel(writer, sheet_name="metrics", index=False)

            config_df = pd.DataFrame(
                {
                    "field": [field],
                    "effective_field": [effective_field],
                    "dr_method": [dr_method],
                    "cluster_method": [cluster_method],
                    "n_clusters": [n_clusters],
                    "n_terms": [len(dtm_top.columns)],
                    "n_components": [n_components],
                    "dtm_method": [dtm_method],
                    "term_selection": [term_selection],
                    "min_df": [min_df],
                    "ngram_min": [ngram_range[0]],
                    "ngram_max": [ngram_range[1]],
                    "keyword_separator": [keyword_separator],
                }
            )
            config_df.to_excel(writer, sheet_name="config", index=False)

    # ------------------------------------------------------------------
    # Step 9: final result
    # ------------------------------------------------------------------
    result = {
        "term_embeddings": term_coords,
        "terms": clean_terms,  # <--- CLEANED labels, used by dendrogram
        "term_labels": term_labels,
        "terms_df": terms_df,
        "label_terms_dict": label_terms_dict,
        "clusters_terms_df": clusters_df,
        "doc_embeddings": doc_coords,
        "doc_labels": doc_labels,
        "metrics": metrics,
        "suggested_k": suggest_k(term_coords),
    }
    return result


def words_by_cluster(
    term_embeddings: np.ndarray,
    terms: list,
    labels: np.ndarray,
) -> pd.DataFrame:
    """
    Return a DataFrame with columns:
      - word: term string (original or fallback "cluster_<id>")
      - Dim1, Dim2: first two embedding components
      - cluster: cluster label

    Automatically fills missing terms if `terms` is shorter than embeddings.
    """
    emb = np.asarray(term_embeddings)
    # Ensure at least 2D
    if emb.ndim == 1:
        emb = np.vstack((emb, np.zeros_like(emb))).T

    n_pts = emb.shape[0]
    # Auto-fill missing term labels
    if len(terms) < n_pts:
        fallback = [f"cluster_{lbl}" for lbl in labels[len(terms):]]
        terms = list(terms) + fallback

    # Validate lengths
    if len(terms) != n_pts or len(labels) != n_pts:
        raise ValueError(
            f"words_by_cluster: embeddings ({n_pts}) must match len(terms) ({len(terms)}) and len(labels) ({len(labels)})"
        )

    return pd.DataFrame({
        'word':    terms,
        'Dim1':    emb[:, 0],
        'Dim2':    emb[:, 1],
        'cluster': labels
    })

def documents_per_cluster(
    df: pd.DataFrame,
    doc_embeddings: np.ndarray,
    doc_labels: np.ndarray,
    tc_field: str = 'Cited by',
) -> pd.DataFrame:
    """
    Return a DataFrame with columns:
      - Documents: from df['Title']
      - dim1, dim2: first two document embedding components
      - contrib: relative contribution (squared distance / total)
      - tc_field: values from df[tc_field] if available
      - Cluster: cluster label

    Parameters
    ----------
    tc_field : str
        Name of the column in df to use for citation counts (e.g. 'Cited by').

    Examples
    --------
    >>> result = conceptual_structure_analysis(df)
    >>> df_docs = documents_per_cluster(
    ...     df, result['doc_embeddings'], result['doc_labels'], tc_field='Cited by'
    >>> )
    """
    emb = np.asarray(doc_embeddings)
    # Ensure at least 2D
    if emb.ndim == 1:
        emb = np.vstack((emb, np.zeros_like(emb))).T
    # Compute contribution as squared distance / sum of all
    squared = emb[:, 0]**2 + emb[:, 1]**2
    contrib = squared / np.sum(squared) if np.sum(squared) != 0 else np.zeros_like(squared)
    # Lookup citation counts
    if tc_field in df.columns:
        tc_vals = df[tc_field].values
    else:
        tc_vals = np.full(len(df), np.nan)
    df_out = pd.DataFrame({
        'Documents': df['Title'].values,
        'dim1': emb[:, 0],
        'dim2': emb[:, 1],
        'contrib': contrib,
        tc_field: tc_vals,
        'Cluster': doc_labels
    }, index=df.index)
    return df_out

"""Document-clustering utilities built on various methods and similarity measures."""
# clustering of the documents

def vectorize_text(
    texts: list[str],
    max_features: int = 1000,
    ngram_range: tuple[int, int] = (1, 2),
) -> csr_matrix:
    """
    Transform a list of text documents into a TF-IDF feature matrix.

    :param texts: List of text strings (e.g., titles, abstracts, keywords).
    :param max_features: Maximum number of features (vocabulary size).
    :param ngram_range: The lower and upper boundary of the n-grams to be extracted.
    :return: TF-IDF feature matrix (sparse).
    """
    vectorizer = TfidfVectorizer(max_features=max_features, ngram_range=ngram_range)
    return vectorizer.fit_transform(texts)

def find_optimal_k(
    X: csr_matrix,
    k_range: range = range(2, 11),
) -> int:
    """
    Determine the optimal number of clusters for KMeans using silhouette score.

    :param X: Feature matrix.
    :param k_range: Range of k values to search.
    :return: k value with highest silhouette score.
    """
    best_k = k_range.start
    best_score = -1
    for k in k_range:
        labels = KMeans(n_clusters=k, random_state=0).fit_predict(X)
        score = silhouette_score(X, labels)
        if score > best_score:
            best_k, best_score = k, score
    return best_k

def cluster_kmeans(
    X: csr_matrix,
    n_clusters: int,
) -> np.ndarray:
    """
    Apply KMeans clustering to data.

    :param X: Feature matrix.
    :param n_clusters: Number of clusters.
    :return: Array of cluster labels.
    """
    model = KMeans(n_clusters=n_clusters, random_state=0)
    return model.fit_predict(X)

def cluster_hierarchical(
    X: csr_matrix,
    n_clusters: int,
    linkage: str = 'ward',
) -> np.ndarray:
    """
    Apply agglomerative (hierarchical) clustering to data.

    :param X: Feature matrix.
    :param n_clusters: Number of clusters.
    :param linkage: Linkage criterion ("ward", "complete", "average", "single").
    :return: Array of cluster labels.
    """
    model = AgglomerativeClustering(n_clusters=n_clusters, linkage=linkage)
    return model.fit_predict(X.to_numpy())

def build_coupling_network(
    refs: list[list[str]],
) -> csr_matrix:
    """
    Build a bibliographic coupling graph and return adjacency matrix.

    :param refs: List where each element is the list of reference IDs for a document.
    :return: Sparse adjacency matrix of coupling weights.
    """
    # Create mapping from ref ID to docs
    ref_to_docs: dict[str, list[int]] = {}
    for doc_idx, doc_refs in enumerate(refs):
        for r in doc_refs:
            ref_to_docs.setdefault(r, []).append(doc_idx)
    # Accumulate coupling counts
    n = len(refs)
    row, col, data = [], [], []
    for docs in ref_to_docs.values():
        for i in range(len(docs)):
            for j in range(i + 1, len(docs)):
                row += [docs[i], docs[j]]
                col += [docs[j], docs[i]]
                data += [1, 1]
    return csr_matrix((data, (row, col)), shape=(n, n))

def cluster_by_coupling(
    coupling_mat: csr_matrix,
    n_clusters: int,
) -> np.ndarray:
    """
    Cluster documents based on bibliographic coupling using KMeans on embedding of the coupling graph.

    :param coupling_mat: Adjacency matrix of coupling weights.
    :param n_clusters: Number of clusters.
    :return: Array of cluster labels.
    """

    eigenvalues, eigenvectors = spla.eigs(coupling_mat, k=n_clusters + 1, which="SR")
    L = eigenvectors.real[:, 1:]
    return KMeans(n_clusters=n_clusters, random_state=0).fit_predict(L)

def save_cluster_results(
    df: pd.DataFrame,
    labels: np.ndarray,
    prefix: str = '',
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Add cluster labels to df and create a binary membership matrix.

    :param df: Original DataFrame of documents.
    :param labels: 1D array of cluster labels per document.
    :param prefix: Optional prefix for new columns.
    :return: Tuple of (annotated_df, binary_df).
    """
    df_out = df.copy()
    col_name = f"{prefix}cluster_label"
    df_out[col_name] = labels
    # One-hot encode labels
    encoder = OneHotEncoder(sparse_output=False, categories="auto")
    onehot = encoder.fit_transform(labels.reshape(-1, 1))
    binary_df = pd.DataFrame(
        onehot,
        columns=[f"{prefix}cluster_{int(c)}" for c in encoder.categories_[0]]
    )
    return df_out, binary_df, col_name

def cluster_documents(
    df: pd.DataFrame,
    *,
    text_field: str = 'Abstract',
    method: str = 'kmeans',
    n_clusters: Optional[int] = None,
    k_range: range = range(2, 11),
    coupling_fields: Optional[list[str] | str] = None,
    sep: Optional[str] = None,
    scorer: str = 'silhouette',
    vectorizer: str = 'tfidf',
    vectorize_fn: Optional[callable] = None,
    random_state: int = 42,
    **vectorize_kwargs,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Cluster documents via vector-space ("kmeans" / "hierarchical") or list-based coupling ("coupling").

    Parameters
    ----------
    df : pd.DataFrame
        Input frame with documents (index preserved). Must be non-empty.
    text_field : str, default="Abstract"
        Column to vectorize for vector-space methods.
    method : {"kmeans","hierarchical","coupling"}, default="kmeans"
        Clustering strategy.
    n_clusters : int | None, default=None
        If None:
          - "kmeans": auto-select over `k_range` using `scorer`.
          - "hierarchical": must be provided (raises if None).
          - "coupling": uses sqrt(n_docs) bounded to [2, 20].
    k_range : range, default=range(2, 11)
        Candidate k values for "kmeans" auto selection.
    coupling_fields : list[str] | str | None, default=None
        Field(s) containing list-like or delimited strings for coupling. If None, tries ["References"].
    sep : str | None, default=None
        Delimiter for splitting string cells in `coupling_fields` (defaults to "; ").
    scorer : {"silhouette","calinski"}, default="silhouette"
        Selection metric for kmeans.
    vectorizer : {"tfidf","count"}, default="tfidf"
        Built-in vectorizer choice if `vectorize_fn` is not provided.
    vectorize_fn : callable | None, default=None
        Custom vectorization function; should return a scipy sparse or numpy array.
    random_state : int, default=42
        Random seed used where applicable.
    **vectorize_kwargs
        Extra kwargs forwarded to the vectorizer or `vectorize_fn`.

    Returns
    -------
    df_out : pd.DataFrame
        Copy of `df` with an added 1-based cluster label column named f"{method}_cluster".
    matrix_df : pd.DataFrame
        - Vector methods: document–term matrix (sparse DataFrame).
        - Coupling: document–document similarity (sparse DataFrame).

    Notes
    -----
    - Robust to missing values; fills text with "" and list-like fields with [].
    - Uses cosine silhouette on sparse inputs for stability.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        raise ValueError("df must be a non-empty DataFrame.")

    from scipy import sparse  # Import sparse for issparse checks
    
    sep = "; " if sep is None else sep
    method = method.lower().strip()

    # ---- helpers ----
    def _to_list_cell(
        x,
    ) -> list[str]:
        if isinstance(x, list | tuple | set):
            return [str(i).strip() for i in x if str(i).strip()]
        if isinstance(x, str):
            return [i.strip() for i in x.split(sep) if i.strip()]
        return []

    def _kmeans_auto_k(
        X,
        _k_range,
        _scorer: str,
    ) -> int:
        from sklearn.cluster import KMeans
        from sklearn.metrics import silhouette_score, calinski_harabasz_score
        from sklearn.preprocessing import normalize as _normalize
        from scipy import sparse

        Xn = _normalize(X, norm="l2", copy=False) if sparse.issparse(X) else X
        best_k, best_score = None, -np.inf
        for k in _k_range:
            if k < 2 or k >= Xn.shape[0]:
                continue
            km = KMeans(n_clusters=k, n_init="auto", random_state=random_state)
            labels = km.fit_predict(Xn)
            if len(np.unique(labels)) < 2:
                continue
            try:
                if _scorer == "calinski":
                    score = calinski_harabasz_score(Xn.toarray() if sparse.issparse(Xn) else Xn, labels)
                else:
                    metric = "cosine" if sparse.issparse(Xn) else "euclidean"
                    score = silhouette_score(Xn, labels, metric=metric)
            except Exception:
                score = -np.inf
            if score > best_score:
                best_k, best_score = k, score
        return best_k or 2

    # ---- vector-space path ----
    if method in {"kmeans", "hierarchical"}:
        if text_field not in df.columns:
            raise ValueError(f"Column '{text_field}' not found in df.")
        if vectorize_fn is not None:
            X = vectorize_fn(df[text_field].fillna(""), **vectorize_kwargs)
            feat_names = getattr(X, "get_feature_names_out", lambda: None)() or np.arange(X.shape[1]).astype(str)
        else:
            from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer

            vec_cls = TfidfVectorizer if vectorizer.lower() == "tfidf" else CountVectorizer
            vec = vec_cls(**vectorize_kwargs)
            X = vec.fit_transform(df[text_field].fillna(""))
            feat_names = vec.get_feature_names_out()

        if method == "kmeans":
            from sklearn.cluster import KMeans
            from sklearn.preprocessing import normalize

            if n_clusters is None:
                n_clusters = _kmeans_auto_k(X, k_range, scorer)
            Xn = normalize(X, norm="l2", copy=False) if sparse.issparse(X) else X
            model = KMeans(n_clusters=int(n_clusters), n_init="auto", random_state=random_state)
            labels = model.fit_predict(Xn)
        else:
            if n_clusters is None:
                raise ValueError("For 'hierarchical', n_clusters must be specified.")
            from sklearn.cluster import AgglomerativeClustering

            X_dense = X.toarray() if sparse.issparse(X) else X
            model = AgglomerativeClustering(n_clusters=int(n_clusters), linkage="ward")
            labels = model.fit_predict(X_dense)

        try:
            matrix_df = pd.DataFrame.sparse.from_spmatrix(X, index=df.index, columns=feat_names)
        except Exception:
            matrix_df = pd.DataFrame(X.toarray(), index=df.index, columns=feat_names)

    # ---- coupling path ----
    elif method == "coupling":
        fields: list[str]
        if isinstance(coupling_fields, str):
            fields = [coupling_fields]
        elif coupling_fields:
            fields = list(coupling_fields)
        else:
            fields = ["References"] if "References" in df.columns else []
        if not fields:
            raise ValueError("No coupling fields provided and 'References' not found in df.")

        missing = [c for c in fields if c not in df.columns]
        if missing:
            raise ValueError(f"Missing coupling column(s): {missing}")

        from sklearn.preprocessing import MultiLabelBinarizer

        B = None  # incidence matrix (csr)
        for col in fields:
            lists = df[col].apply(_to_list_cell)
            mlb = MultiLabelBinarizer(sparse_output=True)
            Xi = mlb.fit_transform(lists)
            Xi = sparse.csr_matrix(Xi)
            B = Xi if B is None else sparse.hstack([B, Xi], format="csr")

        if B is None or B.shape[1] == 0:
            raise ValueError("Coupling produced an empty incidence matrix.")

        S = (B @ B.T).astype(np.float32).tocsr()
        S.setdiag(0)
        S.eliminate_zeros()

        if n_clusters is None:
            n_docs = S.shape[0]
            n_clusters = max(2, min(20, int(round(n_docs ** 0.5))))

        if S.shape[0] < 2:
            labels = np.zeros(S.shape[0], dtype=int)
        else:
            try:
                from sklearn.cluster import SpectralClustering

                model = SpectralClustering(
                    n_clusters=int(n_clusters),
                    affinity="precomputed",
                    assign_labels="kmeans",
                    random_state=random_state,
                    n_init="auto",
                )
                labels = model.fit_predict(S)
            except Exception:
                from sklearn.cluster import AgglomerativeClustering

                Sd = S.toarray()
                row_max = Sd.max(axis=1, keepdims=True)
                row_max[row_max == 0] = 1.0
                A = Sd / row_max
                D = 1.0 - A
                model = AgglomerativeClustering(n_clusters=int(n_clusters), linkage="average", metric="precomputed")
                labels = model.fit_predict(D)

        matrix_df = pd.DataFrame.sparse.from_spmatrix(S, index=df.index, columns=df.index)

    else:
        raise ValueError(f"Unknown method: {method!r}. Choose from 'kmeans', 'hierarchical', 'coupling'.")

    # 1-based labels, friendly column name
    col = f"{method}_cluster"
    df_out = df.copy()
    df_out[col] = np.asarray(labels, dtype=int) + 1
    return df_out, matrix_df

from dataclasses import is_dataclass, asdict
from pandas.api.types import is_numeric_dtype

def _to_dataframe(
    value,
):
    """
    Convert a value to a pandas.DataFrame (index preserved when present).
    """
    if isinstance(value, pd.DataFrame):
        return value
    if isinstance(value, pd.Series):
        return value.to_frame()
    if isinstance(value, (np.ndarray, list, tuple)):
        try:
            return pd.DataFrame(value)
        except Exception:
            return pd.DataFrame({"value": [value]})
    if is_dataclass(value):
        return _to_dataframe(asdict(value))
    if isinstance(value, dict):
        try:
            return pd.DataFrame(value)
        except Exception:
            rows = [{"key": k, "value": v} for k, v in value.items()]
            return pd.DataFrame(rows)
    if hasattr(value, "__dict__"):
        try:
            return _to_dataframe(vars(value))
        except Exception:
            return pd.DataFrame({"value": [str(value)]})
    return pd.DataFrame({"value": [value]})

def _safe_sheet(
    name: str,
) -> str:
    """
    Produce a valid Excel sheet name (<=31 chars; no []:*?/\\).
    """
    s = str(name).strip().replace(" ", "_")
    for ch in r"[]:*?/\\":  # careful: raw string + escaped backslash
        s = s.replace(ch, "_")
    return s[:31] or "Sheet"

def _iter_public_attrs(
    obj,
):
    """
    Yield (name, value) for public, non-callable attributes, ignoring dunders.
    """
    for k in dir(obj):
        if k.startswith("_"):
            continue
        try:
            v = getattr(obj, k)
        except Exception:
            continue
        if callable(v):
            continue
        yield k, v

def _is_primitive_leaf(
    x,
) -> bool:
    """
    Return True if x should be treated as a leaf (no further descent).
    """
    primitive_types = (pd.DataFrame, pd.Series, np.ndarray, list, tuple, dict, str, int, float, bool, type(None))
    if isinstance(x, primitive_types):
        return True
    # Prevent descending into modules/classes/functions and common problematic types
    import types
    return isinstance(x, (types.ModuleType, types.FunctionType, types.MethodType, type))

def _flatten_attrs(
    obj,
    prefix = '',
    seen = None,
    depth = 0,
    max_depth = 6,
):
    """
    Recursively (and safely) flatten attributes for export.

    Yields
    ------
    (sheet_name, value)

    Notes
    -----
    - Avoids infinite recursion via an object-id 'seen' set.
    - Limits nesting with 'max_depth'.
    - Treats many types as leaves to prevent deep/linked traversal.
    """
    if seen is None:
        seen = set()
    oid = id(obj)
    if oid in seen:
        return
    seen.add(oid)

    if depth > max_depth or _is_primitive_leaf(obj):
        yield prefix or "data", obj
        return

    if is_dataclass(obj):
        payload = asdict(obj)
        for k, v in payload.items():
            name = f"{prefix}__{k}" if prefix else k
            if _is_primitive_leaf(v):
                yield name, v
            else:
                yield from _flatten_attrs(v, name, seen, depth + 1, max_depth)
        return

    if isinstance(obj, dict):
        for k, v in obj.items():
            name = f"{prefix}__{k}" if prefix else str(k)
            if _is_primitive_leaf(v):
                yield name, v
            else:
                yield from _flatten_attrs(v, name, seen, depth + 1, max_depth)
        return

    # Generic object with attributes
    for k, v in _iter_public_attrs(obj):
        name = f"{prefix}__{k}" if prefix else k
        if _is_primitive_leaf(v):
            yield name, v
        else:
            yield from _flatten_attrs(v, name, seen, depth + 1, max_depth)

def _has_numeric(
    df: pd.DataFrame,
) -> bool:
    """
    True if DataFrame has at least one numeric column (bool counts as numeric).
    """
    if df is None or df.empty:
        return False
    return any(is_numeric_dtype(dt) for dt in df.dtypes)

def _save_associations_xlsx(
    assoc,
    save_root: str,
):
    """
    Save all (nested) attributes of an associations object to '{save_root}.xlsx'.

    Rules
    -----
    - Writes index (index=True).
    - Skips sheets with no numeric columns.
    - Prevents infinite recursion with object-id tracking and depth limit.
    """
    xlsx_path = f"{save_root}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as xlw:
        added = set()
        for name, value in _flatten_attrs(assoc):
            df = _to_dataframe(value)
            if not _has_numeric(df):
                continue
            sheet = _safe_sheet(name)
            # ensure unique sheet names after truncation
            suffix = 1
            while sheet in added:
                tail = f"_{suffix}"
                sheet = _safe_sheet(name[: max(0, 31 - len(tail))] + tail)
                suffix += 1
            df.to_excel(xlw, sheet_name=sheet, index=True)
            added.add(sheet)
    return xlsx_path

"""Spectrogram-style helpers for visualising reference or citation distributions over time."""
# spectroscopy

def extract_cited_years(
    reference_text,
    include_pre_1900 = False,
):
    """
    Extracts all 4-digit years from a reference string. Includes 17xx and 18xx if enabled.

    Args:
        reference_text (str): String containing reference text.
        include_pre_1900 (bool): If True, includes years before 1900.

    Returns:
        list of str: List of extracted years.
    """
    if include_pre_1900:
        pattern = r"\b(17\d{2}|18\d{2}|19\d{2}|20\d{2})\b"
    else:
        pattern = r"\b(19\d{2}|20\d{2})\b"
    return re.findall(pattern, reference_text or "")

def compute_reference_spectrogram_scopus(
    df,
    reference_column = 'References',
    include_pre_1900 = False,
):
    """
    Computes the distribution of cited years from a DataFrame column containing references.

    Args:
        df (pd.DataFrame): DataFrame with a column of reference text.
        reference_column (str): Name of the column with references.
        include_pre_1900 (bool): Whether to include years before 1900.

    Returns:
        pd.DataFrame: DataFrame with cited years and their citation counts.
    """
    current_year = datetime.datetime.now().year
    all_years = df[reference_column].fillna("").apply(lambda x: extract_cited_years(x, include_pre_1900))
    all_years_flat = list(chain.from_iterable(all_years))
    all_years_flat = [int(y) for y in all_years_flat if int(y) <= current_year]
    year_counts = Counter(all_years_flat)
    result_df = pd.DataFrame.from_dict(year_counts, orient="index", columns=["Citations"])
    result_df.index.name = "Cited Year"
    result_df = result_df.sort_index()
    return result_df

def compute_reference_spectrogram_openalex(
    references_counts_df: pd.DataFrame,
    ref_col: str = 'Reference',
    count_col: str = 'Number of documents',
    min_docs: int = 5,
    mailto: Optional[str] = None,
    include_pre_1900: bool = False,
) -> pd.DataFrame:
    """
    Compute a cited-year spectrogram for an OpenAlex dataset.

    Workflow:
    1) Filter `references_counts_df` to references with at least `min_docs` in `count_col`.
    2) Harvest OpenAlex metadata for the surviving references via `harvest_openalex_links_to_df`.
    3) Use the `publication_year` from the harvested metadata.
    4) Aggregate citation weight by year, where weight is the original `count_col` per reference.

    Assumptions:
    - `harvest_openalex_links_to_df` returns a DataFrame that includes the original input key as
      a column named exactly `ref_col` (default "Reference") and a column "Year".
      If your harvester uses a different column name for the input key, adapt `ref_col` or
      modify the harvester to echo the input as `ref_col`.

    Parameters
    ----------
    references_counts_df : pandas.DataFrame
        Must contain columns `ref_col` (OpenAlex work links/IDs) and `count_col`.
    ref_col : str, default "Reference"
        Column holding OpenAlex links/IDs (works).
    count_col : str, default "Number of documents"
        Column with the number of local documents citing each reference.
    min_docs : int, default 5
        Keep references with value >= `min_docs`.
    mailto : str, optional
        Contact email for OpenAlex polite pool, forwarded to the harvester.
    include_pre_1900 : bool, default False
        If False, discard years < 1900.

    Returns
    -------
    pandas.DataFrame
        Index "Cited Year", single column "Citations" (sum of `count_col` by publication year),
        sorted by year ascending.
    """
    import datetime
    import pandas as pd

    # --- Validate inputs --------------------------------------------------------
    needed = {ref_col, count_col}
    missing = needed - set(references_counts_df.columns)
    if missing:
        raise KeyError(f"Missing required columns in references_counts_df: {sorted(missing)}")

    # --- Filter references by threshold ----------------------------------------
    filt_df = references_counts_df.loc[references_counts_df[count_col] >= min_docs, [ref_col, count_col]].copy()
    if filt_df.empty:
        return pd.DataFrame(columns=["Citations"]).rename_axis("Cited Year")

    # --- Harvest OpenAlex metadata ---------------------------------------------
    links = filt_df[ref_col].dropna().astype(str).tolist()
    from biblium import readbib
    meta_df = readbib.harvest_openalex_links_to_df(links, mailto=mailto)
    meta_df = meta_df.rename(columns={"input": ref_col})

    # Ensure required columns exist in harvested metadata
    if "Year" not in meta_df.columns:
        raise KeyError('Harvested metadata is missing required column "Year".')
    if ref_col not in meta_df.columns:
        # If your harvester does not echo the input link/ID as `ref_col`, raise a clear error
        raise KeyError(f'Harvested metadata is missing the input key column "{ref_col}". '
                       "Modify the harvester to echo inputs, or update ref_col accordingly.")

    # --- Merge counts with publication years -----------------------------------
    merged = (
        meta_df[[ref_col, "Year"]]
        .merge(filt_df, on=ref_col, how="inner")
        .dropna(subset=["Year"])
        .copy()
    )

    # --- Clean and constrain years ---------------------------------------------
    current_year = datetime.datetime.now().year
    merged["Year"] = pd.to_numeric(merged["Year"], errors="coerce").astype("Int64")
    merged = merged.dropna(subset=["Year"])
    merged = merged[(merged["Year"] <= current_year)]
    if not include_pre_1900:
        merged = merged[merged["Year"] >= 1900]
    if merged.empty:
        return pd.DataFrame(columns=["Citations"]).rename_axis("Cited Year")

    # --- Aggregate: weight by local document counts ----------------------------
    spectrogram_df = (
        merged.groupby("Year", as_index=True)[count_col]
        .sum()
        .rename("Citations")
        .to_frame()
        .sort_index()
    )
    spectrogram_df.index.name = "Cited Year"
    return spectrogram_df

def compute_reference_correlation(
    df,
    reference_column = 'References',
    year_column = 'Year',
    include_pre_1900 = False,
):
    """
    Computes correlation between document year and:
    1. Mean of reference years per document
    2. Each individual reference year occurrence

    Returns:
        dict: Correlation results and associated data
    """
    current_year = datetime.datetime.now().year
    ref_years = df[reference_column].fillna("").apply(lambda x: extract_cited_years(x, include_pre_1900))
    doc_years = df[year_column].values

    mean_refs = []
    repeated_pairs = []
    for doc_year, years in zip(doc_years, ref_years):
        ref_years_filtered = [int(y) for y in years if int(y) <= current_year and int(y) <= doc_year]
        if ref_years_filtered:
            mean_ref = np.mean(ref_years_filtered)
            mean_refs.append((doc_year, mean_ref))
            repeated_pairs.extend([(doc_year, y) for y in ref_years_filtered])

    mean_refs_df = pd.DataFrame(mean_refs, columns=["Document Year", "Mean Reference Year"])
    repeated_df = pd.DataFrame(repeated_pairs, columns=["Document Year", "Reference Year"])

    mean_corr = pearsonr(mean_refs_df["Document Year"], mean_refs_df["Mean Reference Year"])
    ref_corr = pearsonr(repeated_df["Document Year"], repeated_df["Reference Year"])

    return {
        "mean_reference_df": mean_refs_df,
        "repeated_year_df": repeated_df,
        "mean_reference_corr": mean_corr,
        "reference_year_corr": ref_corr
    }

"""Helpers for analysing scientific production across predefined document groups."""
# scientific production groups

def get_scientific_production_by_group(
    df: pd.DataFrame,
    group_matrix: pd.DataFrame | np.ndarray,
    group_names: list[str] = None,
    output_format: str = 'wide',
    **production_kwargs,
) -> pd.DataFrame:
    """
    Compute annual scientific production stats for each group and merge into a single DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain columns "Year" and "Cited by".
    group_matrix : pd.DataFrame or 2D array of shape (n_docs, n_groups)
        If DataFrame, its index must match df.index and its columns are group names.
    group_names : list of str, optional
        If group_matrix is an array, names of groups (in column order).
    output_format : str, default "wide"
        "wide" for a pivoted DataFrame with Year as a column and metrics grouped by group,
        "long" for a tidy DataFrame with columns [Group, Year, ...metrics...].
    **production_kwargs
        Passed to get_scientific_production (relative_counts, cumulative, etc.).

    Returns
    -------
    pd.DataFrame
        Merged production statistics in the specified format.
    """
    # Make a deep copy of df to avoid SettingWithCopy warnings
    df = df.copy()
    # Use .loc to safely assign
    df.loc[:, "Year"] = df.loc[:, "Year"].astype(int)

    # Prepare group DataFrame
    if not isinstance(group_matrix, pd.DataFrame):
        if group_names is None:
            raise ValueError("Please supply group_names when group_matrix is an array")
        group_df = pd.DataFrame(group_matrix, index=df.index, columns=group_names)
    else:
        group_df = group_matrix.copy()

    # Compute per-group stats, index by Year
    results = {}
    for grp in group_df.columns:
        mask = group_df[grp] == 1
        subset = df.loc[mask].copy()
        prod = (
            get_scientific_production(subset, **production_kwargs)
            if not subset.empty
            else get_scientific_production(df.iloc[0:0], **production_kwargs)
        )
        prod = prod.set_index("Year")
        results[grp] = prod

    # Concatenate into MultiIndex: [Group, Year]
    merged = pd.concat(results, names=["Group", "Year"])

    if output_format == "wide":
        # Pivot groups into wide format
        wide = merged.unstack(level="Group")
        # Flatten column index and adjust metric names
        cols = []
        for metric, grp in wide.columns:
            m = metric.replace("Number of Documents", "Number of documents")
            cols.append(f"{m} {grp}")
        wide.columns = cols
        # fill missing years/groups with zeros
        wide = wide.fillna(0)
        # Flatten Year index into a column
        wide = wide.reset_index()
        return wide

    elif output_format == "long":
        # Reset index; missing group-year combos may be omitted
        long = merged.reset_index()
        # adjust number of documents label
        long = long.rename(columns={"Number of Documents": "Number of documents"})
        # fill numeric NaNs with zeros
        num_cols = long.select_dtypes(include=[np.number]).columns
        long.loc[:, num_cols] = long.loc[:, num_cols].fillna(0)
        return long

    else:
        raise ValueError("output_format must be 'long' or 'wide'")

import pandas as pd
import numpy as np
from typing import Any, Optional, Union


# Column name mappings between different naming conventions
COUNT_COLUMN_ALIASES = {
    # Standard name -> list of possible aliases
    "Number of documents": ["Number of documents", "N OCC", "OCC", "Count", "Docs"],
    "Fraction of documents": ["Fraction of documents", "Proportion", "Proportion of documents", "Fraction"],
    "Percentage of documents": ["Percentage of documents", "Percentage", "%", "Percent"],
    "Rank": ["Rank"],
    "Percentrank of documents": ["Percentrank of documents", "PercentRank", "Percentrank", "Percent Rank"],
}

# Reverse mapping: alias -> standard name
ALIAS_TO_STANDARD = {}
for standard, aliases in COUNT_COLUMN_ALIASES.items():
    for alias in aliases:
        ALIAS_TO_STANDARD[alias] = standard
        ALIAS_TO_STANDARD[alias.lower()] = standard


def _standardize_count_columns(df: pd.DataFrame, entity_label: str) -> pd.DataFrame:
    """
    Standardize column names in a counts DataFrame to use consistent naming.
    
    Maps various naming conventions to standard names:
    - Proportion -> Fraction of documents
    - Percentage -> Percentage of documents
    - PercentRank -> Percentrank of documents
    - etc.
    """
    if df.empty:
        return df
    
    df = df.copy()
    rename_map = {}
    
    for col in df.columns:
        if col == entity_label:
            continue
        
        # Check if this column name (or lowercase version) has a standard mapping
        std_name = ALIAS_TO_STANDARD.get(col) or ALIAS_TO_STANDARD.get(col.lower())
        
        if std_name and std_name != col and std_name not in df.columns:
            rename_map[col] = std_name
    
    if rename_map:
        df = df.rename(columns=rename_map)
    
    return df


def _generate_counts_from_column(
    df: pd.DataFrame,
    entity_col: str,
    entity_label: str,
    value_type: str = "string",
    sep: str = "; ",
) -> pd.DataFrame:
    """
    Generate a counts DataFrame from an entity column.
    
    This is a fallback when no counts_df or count_method is provided.
    
    Parameters
    ----------
    df : pd.DataFrame
        Document-level data.
    entity_col : str
        Column containing entity values.
    entity_label : str
        Name for the entity column in output.
    value_type : str
        "string", "list", or "text".
    sep : str
        Separator for list-like columns (can be regex pattern).
    
    Returns
    -------
    pd.DataFrame
        Counts table with entity_label and standard count columns.
    """
    if entity_col not in df.columns:
        return pd.DataFrame(columns=[entity_label, "Number of documents"])
    
    series = df[entity_col].dropna()
    
    if series.empty:
        return pd.DataFrame(columns=[entity_label, "Number of documents"])
    
    if value_type == "list":
        # Split string values by separator and explode
        if series.dtype == object:
            # Handle string representation of lists - use regex split
            exploded = series.astype(str).str.split(sep, regex=True).explode()
        else:
            # Already a list-like
            exploded = series.explode()
        
        # Clean up values
        exploded = exploded.astype(str).str.strip()
        exploded = exploded[exploded != ""]
        exploded = exploded[exploded.str.lower() != "nan"]
        
        if exploded.empty:
            return pd.DataFrame(columns=[entity_label, "Number of documents"])
        
        # Count occurrences (number of documents containing each entity)
        # Use nunique on index to count documents, not total occurrences
        reset_df = exploded.reset_index(); value_col = reset_df.columns[1]; doc_counts = reset_df.groupby(value_col)['index'].nunique()
        counts = doc_counts.reset_index()
        counts.columns = [entity_label, "Number of documents"]
        
    elif value_type == "string":
        # Direct value counts - each row has one value
        clean_series = series.astype(str).str.strip()
        clean_series = clean_series[clean_series != ""]
        clean_series = clean_series[clean_series.str.lower() != "nan"]
        
        if clean_series.empty:
            return pd.DataFrame(columns=[entity_label, "Number of documents"])
        
        counts = clean_series.value_counts().reset_index()
        counts.columns = [entity_label, "Number of documents"]
        
    else:  # text or other
        # For text, we typically need external count_method (ngrams, etc.)
        # Fallback to treating as string
        clean_series = series.astype(str).str.strip()
        clean_series = clean_series[clean_series != ""]
        clean_series = clean_series[clean_series.str.lower() != "nan"]
        
        if clean_series.empty:
            return pd.DataFrame(columns=[entity_label, "Number of documents"])
        
        counts = clean_series.value_counts().reset_index()
        counts.columns = [entity_label, "Number of documents"]
    
    if counts.empty:
        return pd.DataFrame(columns=[entity_label, "Number of documents"])
    
    # Filter out empty/nan values
    counts = counts[counts[entity_label].astype(str).str.strip() != ""]
    counts = counts[counts[entity_label].astype(str).str.lower() != "nan"]
    
    # Sort by count descending
    counts = counts.sort_values("Number of documents", ascending=False).reset_index(drop=True)
    
    # Add derived metrics with STANDARD names
    total_docs = len(df)
    counts["Fraction of documents"] = counts["Number of documents"] / total_docs if total_docs > 0 else 0.0
    counts["Percentage of documents"] = counts["Fraction of documents"] * 100
    counts["Rank"] = counts["Number of documents"].rank(ascending=False, method="min")
    counts["Percentrank of documents"] = counts["Number of documents"].rank(pct=True)
    
    return counts


def group_entity_stats(
    df: pd.DataFrame,
    group_matrix: Any,
    entity_col: str,
    entity_label: str,
    *,
    items_of_interest: Optional[list] = None,
    exclude_items: Optional[list] = None,
    top_n: int = 20,
    counts_df: Optional[pd.DataFrame] = None,
    count_method: Optional[Any] = None,
    group_counts: Optional[dict[str, pd.DataFrame]] = None,
    regex_include: Optional[str] = None,
    regex_exclude: Optional[str] = None,
    value_type: str = "string",
    sep: str = "; ",
    indicators: bool = False,
    missing_as_zero: bool = False,
    mode: str = "full",
    output_format: str = "wide",
    **extra_kwargs,
):
    """
    Compute performance indicators for entities across multiple groups.
    
    Parameters
    ----------
    df : pd.DataFrame
        Document-level data.
    group_matrix : pd.DataFrame, pd.Series, or array-like
        Boolean matrix indicating group membership (rows=documents, columns=groups).
    entity_col : str
        Column in df containing the entity values.
    entity_label : str
        Name of the entity label column in the stats output.
    items_of_interest : list, optional
        Specific entities to analyze.
    exclude_items : list, optional
        Entities to exclude.
    top_n : int, default 20
        Number of top entities to select.
    counts_df : pd.DataFrame, optional
        Global pre-computed counts table.
    count_method : callable, optional
        Function returning a counts_df when called.
    group_counts : dict[str, pd.DataFrame], optional
        Pre-computed counts per group. Keys are group names, values are DataFrames.
        Column names are automatically standardized (e.g., "Proportion" -> "Fraction of documents").
    regex_include, regex_exclude : str, optional
        Regex filters for entity selection.
    value_type : str, default "string"
        How to interpret entity_col ("string", "list", "text").
    sep : str, default "; "
        Separator for list-like entity columns.
    indicators : bool, default False
        If True, also return document-level indicators.
    missing_as_zero : bool, default False
        Passed through to select_documents.
    mode : str, default "full"
        Passed to get_performance_indicators.
    output_format : str, default "wide"
        Output format: "wide", "long", "pivot", or "matrix".
    **extra_kwargs
        Additional arguments (openalex_add, translation, max_items).
    
    Returns
    -------
    stats_df : pd.DataFrame or dict
        Aggregated indicators per entity.
    indicator_df : dict or None
        Document-level indicators if indicators=True.
    """
    # ------------------------------------------------------------------
    # 0. Normalize and align group_matrix to df
    # ------------------------------------------------------------------
    if isinstance(group_matrix, pd.DataFrame):
        gm = group_matrix.copy()
    elif isinstance(group_matrix, pd.Series):
        gm = group_matrix.to_frame()
    else:
        gm = pd.DataFrame(group_matrix)

    if not gm.index.equals(df.index):
        gm = gm.reindex(df.index)

    if gm.shape[0] != len(df):
        raise ValueError("group_matrix must have the same number of rows as df.")

    gm.columns = [str(c) for c in gm.columns]

    if gm.empty:
        if output_format in {"wide", "long"}:
            return pd.DataFrame(), None
        return {}, None

    # ------------------------------------------------------------------
    # 1. Handle special extra kwargs
    # ------------------------------------------------------------------
    openalex_add = bool(extra_kwargs.pop("openalex_add", False))
    translation = extra_kwargs.pop("translation", None)
    max_items = extra_kwargs.get("max_items", 0)

    # ------------------------------------------------------------------
    # 1b. Standardize and generate counts_df if not provided
    # ------------------------------------------------------------------
    if counts_df is not None:
        counts_df = _standardize_count_columns(counts_df, entity_label)
    
    if counts_df is None and count_method is None and items_of_interest is None:
        # Generate counts from the data
        counts_df = _generate_counts_from_column(
            df, entity_col, entity_label, value_type=value_type, sep=sep
        )

    # ------------------------------------------------------------------
    # 2. Define Universe (Items of Interest)
    # ------------------------------------------------------------------
    if items_of_interest is None:
        global_stats, _ = get_entity_stats(
            df,
            entity_col,
            entity_label,
            items_of_interest=None,
            exclude_items=exclude_items,
            top_n=top_n,
            counts_df=counts_df,
            count_method=count_method,
            regex_include=regex_include,
            regex_exclude=regex_exclude,
            value_type=value_type,
            indicators=False,
            missing_as_zero=missing_as_zero,
            mode=mode,
            sep=sep,
            openalex_add=openalex_add,
            translation=translation,
            max_items=max_items,
        )
        if global_stats.empty or entity_label not in global_stats.columns:
            if output_format in {"wide", "long"}:
                return pd.DataFrame(), None
            return {}, None
        items_list = global_stats[entity_label].astype(str).tolist()
    else:
        items_list = list(dict.fromkeys(items_of_interest))
        if not items_list:
            if output_format in {"wide", "long"}:
                return pd.DataFrame(), None
            return {}, None

        global_stats, _ = get_entity_stats(
            df,
            entity_col,
            entity_label,
            items_of_interest=items_list,
            exclude_items=exclude_items,
            top_n=top_n,
            counts_df=counts_df,
            count_method=count_method,
            regex_include=regex_include,
            regex_exclude=regex_exclude,
            value_type=value_type,
            indicators=False,
            missing_as_zero=missing_as_zero,
            mode=mode,
            sep=sep,
            openalex_add=openalex_add,
            translation=translation,
            max_items=max_items,
        )

    items_list = sorted(list(set(items_list)), key=lambda x: str(x).lower())
    base_entities = pd.Index(items_list, name=entity_label)

    # Define metric columns (use STANDARD names)
    counting_cols = [
        "Number of documents",
        "Fraction of documents",
        "Percentage of documents",
        "Rank",
        "Percentrank of documents",
    ]
    
    if global_stats.empty:
        metric_cols = counting_cols.copy()
    else:
        # Standardize global_stats column names
        global_stats = _standardize_count_columns(global_stats, entity_label)
        metric_cols = [c for c in global_stats.columns if c != entity_label]

    metric_cols = [c.replace("_stats", "") for c in metric_cols]

    # Ensure counting columns are in the list
    for col in counting_cols:
        if col not in metric_cols:
            metric_cols.append(col)

    metric_cols = list(dict.fromkeys(metric_cols))

    # ------------------------------------------------------------------
    # 3. PRE-COMPUTE COUNTS per Group (or use provided group_counts)
    # ------------------------------------------------------------------
    group_precomputed_counts = {}

    for group_name in gm.columns:
        mask = gm[group_name].astype(bool)
        
        # Check if pre-computed counts were provided for this group
        if group_counts is not None and group_name in group_counts:
            pre_counts = group_counts[group_name].copy()
            
            # Ensure entity_label column exists
            if entity_label not in pre_counts.columns:
                first_col = pre_counts.columns[0]
                pre_counts = pre_counts.rename(columns={first_col: entity_label})
            
            # STANDARDIZE column names
            pre_counts = _standardize_count_columns(pre_counts, entity_label)
            
            # Filter to items of interest
            pre_counts = pre_counts[
                pre_counts[entity_label].astype(str).isin(items_list)
            ]
            
            group_precomputed_counts[group_name] = pre_counts
            
        elif mask.any():
            subset_df = df.loc[mask[mask].index]

            # Generate counts for this subset
            subset_counts = _generate_counts_from_column(
                subset_df, entity_col, entity_label, value_type=value_type, sep=sep
            )
            
            # Filter to items of interest
            if not subset_counts.empty and entity_label in subset_counts.columns:
                subset_counts = subset_counts[
                    subset_counts[entity_label].astype(str).isin(items_list)
                ]

            # Get stats with the generated counts
            g_counts, _ = get_entity_stats(
                subset_df,
                entity_col,
                entity_label,
                items_of_interest=items_list,
                exclude_items=exclude_items,
                top_n=top_n,
                counts_df=subset_counts,
                count_method=None,
                regex_include=None,
                regex_exclude=None,
                value_type=value_type,
                indicators=False,
                missing_as_zero=missing_as_zero,
                mode=mode,
                sep=sep,
                openalex_add=False,
                translation=translation,
                max_items=max_items,
            )

            if not g_counts.empty and entity_label in g_counts.columns:
                # Standardize column names
                g_counts = _standardize_count_columns(g_counts, entity_label)
                
                # Clean Column Names for _stats suffix
                if "Number of documents_stats" in g_counts.columns:
                    if "Number of documents" in g_counts.columns:
                        g_counts = g_counts.drop(columns=["Number of documents"])
                    g_counts = g_counts.rename(
                        columns={"Number of documents_stats": "Number of documents"}
                    )

                # Drop Duplicates
                g_counts = g_counts.loc[:, ~g_counts.columns.duplicated()]
                g_counts = g_counts.drop_duplicates(subset=[entity_label])

                # Ensure derived metrics exist (recalculate for this subset)
                if "Number of documents" in g_counts.columns:
                    n_total = len(subset_df)
                    g_counts["Number of documents"] = g_counts[
                        "Number of documents"
                    ].fillna(0)

                    # Always recalculate these for the subset
                    g_counts["Fraction of documents"] = (
                        g_counts["Number of documents"] / n_total
                        if n_total > 0
                        else 0.0
                    )

                    g_counts["Percentage of documents"] = (
                        g_counts["Fraction of documents"] * 100
                    )

                    g_counts["Rank"] = g_counts["Number of documents"].rank(
                        ascending=False, method="min"
                    )

                    g_counts["Percentrank of documents"] = g_counts[
                        "Number of documents"
                    ].rank(pct=True)

                group_precomputed_counts[group_name] = g_counts
            else:
                # Use the subset_counts directly if get_entity_stats returned empty
                if not subset_counts.empty:
                    group_precomputed_counts[group_name] = subset_counts
                else:
                    group_precomputed_counts[group_name] = pd.DataFrame(
                        columns=[entity_label]
                    )
        else:
            group_precomputed_counts[group_name] = pd.DataFrame(
                columns=[entity_label]
            )

    # ------------------------------------------------------------------
    # 4. Compute Full Stats & Merge
    # ------------------------------------------------------------------
    group_frames: dict[str, pd.DataFrame] = {}
    indicator_dict: Optional[dict] = {} if indicators else None

    for group_name in gm.columns:
        mask = gm[group_name].astype(bool)
        pre_counts_df = group_precomputed_counts[group_name]

        # A. Run standard stats
        if mask.any():
            subset_df = df.loc[mask[mask].index]

            # Use pre_counts_df if available, otherwise generate
            if pre_counts_df.empty or entity_label not in pre_counts_df.columns:
                subset_counts = _generate_counts_from_column(
                    subset_df, entity_col, entity_label, value_type=value_type, sep=sep
                )
            else:
                subset_counts = pre_counts_df

            stats_g, ind_g = get_entity_stats(
                subset_df,
                entity_col,
                entity_label,
                items_of_interest=items_list,
                exclude_items=exclude_items,
                top_n=top_n,
                counts_df=subset_counts,
                count_method=None,
                regex_include=None,
                regex_exclude=None,
                value_type=value_type,
                indicators=indicators,
                missing_as_zero=missing_as_zero,
                mode=mode,
                sep=sep,
                openalex_add=openalex_add,
                translation=translation,
                max_items=max_items,
            )

            if indicators and ind_g is not None:
                indicator_dict[group_name] = ind_g

            if not stats_g.empty and entity_label in stats_g.columns:
                # Standardize column names
                stats_g = _standardize_count_columns(stats_g, entity_label)
                
                if "Number of documents_stats" in stats_g.columns:
                    if "Number of documents" in stats_g.columns:
                        stats_g = stats_g.drop(columns=["Number of documents"])
                    stats_g = stats_g.rename(
                        columns={"Number of documents_stats": "Number of documents"}
                    )

                stats_g = stats_g.loc[:, ~stats_g.columns.duplicated()]
                stats_g = stats_g.drop_duplicates(subset=[entity_label])
                stats_g = stats_g.set_index(entity_label)

                if stats_g.index.duplicated().any():
                    stats_g = stats_g[~stats_g.index.duplicated(keep="first")]

                # Ensure metric_cols exist in stats_g before reindexing
                available_cols = [c for c in metric_cols if c in stats_g.columns]
                missing_cols = [c for c in metric_cols if c not in stats_g.columns]
                
                gf = stats_g.reindex(index=base_entities, columns=available_cols)
                
                # Add missing columns with NaN
                for col in missing_cols:
                    gf[col] = np.nan
                
                # Reorder to match metric_cols
                gf = gf[metric_cols]
            else:
                gf = pd.DataFrame(index=base_entities, columns=metric_cols, dtype=float)
        else:
            gf = pd.DataFrame(index=base_entities, columns=metric_cols, dtype=float)

        # B. OVERWRITE Counting Columns with Trusted Pre-Computed Values
        if not pre_counts_df.empty and entity_label in pre_counts_df.columns:
            p_counts = pre_counts_df.drop_duplicates(subset=[entity_label]).set_index(
                entity_label
            )
            if p_counts.index.duplicated().any():
                p_counts = p_counts[~p_counts.index.duplicated(keep="first")]

            aligned_counts = p_counts.reindex(base_entities)

            # Overwrite ALL counting columns from pre-computed values
            for col in counting_cols:
                if col in aligned_counts.columns and col in gf.columns:
                    gf[col] = aligned_counts[col]

        # C. Targeted Fillna (Only for counting cols, but NOT rank columns)
        for col in counting_cols:
            if col in gf.columns:
                if 'rank' not in col.lower():
                    gf[col] = gf[col].fillna(0.0)

        group_frames[group_name] = gf

    # ------------------------------------------------------------------
    # 5. Assemble output
    # ------------------------------------------------------------------
    if not metric_cols:
        return (
            (pd.DataFrame(), indicator_dict)
            if output_format in {"wide", "long"}
            else ({}, indicator_dict)
        )

    if output_format == "wide":
        pieces = []
        for group_name, gf in group_frames.items():
            gf_group = gf.copy()
            gf_group.columns = [f"{group_name} - {c}" for c in gf_group.columns]
            pieces.append(gf_group)

        if not pieces:
            return pd.DataFrame(), indicator_dict

        wide_df = pd.concat(pieces, axis=1)
        wide_df.index.name = entity_label
        wide_df = wide_df.reset_index()
        return wide_df, indicator_dict

    if output_format == "long":
        long_rows = []
        for group_name, gf in group_frames.items():
            tmp = gf.copy()
            tmp[entity_label] = tmp.index
            tmp.insert(0, "Group", group_name)
            long_rows.append(tmp.reset_index(drop=True))

        if not long_rows:
            return pd.DataFrame(), indicator_dict

        long_df = pd.concat(long_rows, ignore_index=True)
        cols = ["Group", entity_label] + [
            c for c in long_df.columns if c not in ("Group", entity_label)
        ]
        return long_df[cols], indicator_dict

    pivot_dict = {}
    for metric in metric_cols:
        data = {}
        for group_name, gf in group_frames.items():
            if metric in gf.columns:
                data[group_name] = gf[metric]
            else:
                data[group_name] = pd.Series(np.nan, index=base_entities)
        metric_df = pd.DataFrame(data, index=base_entities)
        metric_df.index.name = entity_label
        pivot_dict[metric] = metric_df

    if output_format == "pivot":
        return pivot_dict, indicator_dict

    if output_format == "matrix":
        matrix_dict = {m: df_m.to_numpy() for m, df_m in pivot_dict.items()}
        return matrix_dict, indicator_dict

    raise ValueError(f"Unknown output_format {output_format!r}")


def extract_group_counts_from_merged(
    merged_df: pd.DataFrame,
    entity_label: str,
    group_names: list[str],
    column_mapping: Optional[dict[str, str]] = None,
) -> dict[str, pd.DataFrame]:
    """
    Extract per-group counts DataFrames from a merged DataFrame 
    (as produced by count_occurrences_across_groups).
    
    Parameters
    ----------
    merged_df : pd.DataFrame
        The merged DataFrame from count_occurrences_across_groups.
    entity_label : str
        The name of the entity column (first column).
    group_names : list[str]
        List of group names to extract.
    column_mapping : dict, optional
        Mapping from merged column base names to standard names.
        Default mapping:
            "Number of documents" or "N OCC" -> "Number of documents"
            "Proportion" or "Fraction" -> "Fraction of documents"
            "Percentage" or "%" -> "Percentage of documents"
            "Rank" -> "Rank"
            "PercentRank" or "Percentrank" -> "Percentrank of documents"
    
    Returns
    -------
    dict[str, pd.DataFrame]
        Dictionary mapping group names to their counts DataFrames.
    
    Example
    -------
    >>> merged = count_occurrences_across_groups(groups, gm, "count_keywords")
    >>> group_counts = extract_group_counts_from_merged(merged, "Keyword", ["Group1", "Group2"])
    >>> stats, _ = group_entity_stats(df, gm, "keywords", "Keyword", group_counts=group_counts)
    """
    if column_mapping is None:
        column_mapping = {
            "Number of documents": "Number of documents",
            "N OCC": "Number of documents",
            "OCC": "Number of documents",
            "Proportion": "Fraction of documents",
            "Fraction": "Fraction of documents",
            "Percentage": "Percentage of documents",
            "%": "Percentage of documents",
            "Rank": "Rank",
            "PercentRank": "Percentrank of documents",
            "Percentrank": "Percentrank of documents",
        }
    
    # Get entity column (first column)
    entity_col = merged_df.columns[0]
    
    group_counts = {}
    
    for group_name in group_names:
        # Find columns for this group
        group_cols = [col for col in merged_df.columns if f"({group_name})" in col]
        
        if not group_cols:
            continue
        
        # Extract subset
        group_df = merged_df[[entity_col] + group_cols].copy()
        
        # Rename columns to standard names
        rename_map = {entity_col: entity_label}
        for col in group_cols:
            # Extract base name (before the group suffix)
            base_name = col.replace(f" ({group_name})", "").strip()
            
            # Find matching standard name
            standard_name = None
            for pattern, std_name in column_mapping.items():
                if pattern.lower() in base_name.lower():
                    standard_name = std_name
                    break
            
            if standard_name:
                rename_map[col] = standard_name
            else:
                # Keep original base name
                rename_map[col] = base_name
        
        group_df = group_df.rename(columns=rename_map)
        
        # Remove duplicate columns (keep first)
        group_df = group_df.loc[:, ~group_df.columns.duplicated()]
        
        group_counts[group_name] = group_df
    
    return group_counts
    
"""Utilities for comparing local versus global frequency distributions of items."""
# comparison of two frequency distributions (global, local)

def compare_counts(
    series_full: pd.Series,
    series_subset: pd.Series,
    top_n: int = None,
) -> pd.DataFrame:
    """
    Compare item counts between a reference (full) and a subset dataset.

    Parameters
    ----------
    series_full : pd.Series
        Item counts from the full dataset.
    series_subset : pd.Series
        Item counts from the subset dataset.
    top_n : int, optional
        If specified, return only the top_n items by absolute percentage point difference.

    Returns
    -------
    pd.DataFrame
        DataFrame with counts, proportions, percentage point difference, fold-change, and flag.
    """
    df = pd.DataFrame({
        "Count_Full": series_full,
        "Count_Sub": series_subset
    }).fillna(0)

    df["Prop_Full"] = df["Count_Full"] / df["Count_Full"].sum()
    df["Prop_Sub"] = df["Count_Sub"] / df["Count_Sub"].sum()
    df["PP_Diff"] = 100 * (df["Prop_Sub"] - df["Prop_Full"])

    with np.errstate(divide="ignore", invalid="ignore"):
        df["Rel_Diff"] = np.where(df["Prop_Full"] > 0, df["Prop_Sub"] / df["Prop_Full"], np.nan)

    df["Interesting"] = (df["PP_Diff"].abs() > 2) | (df["Rel_Diff"] > 2) | (df["Rel_Diff"] < 0.5)

    if top_n is not None:
        df = df.reindex(df["PP_Diff"].abs().sort_values(ascending=False).index).head(top_n)

    return df

"""Helpers for interacting with large language models within Biblium workflows."""
# LLM

"""Small helpers for configuring OpenAI API keys in a safe and overridable way."""
# Set OpenAI API key for authentication
"""User-specific OpenAI configuration notes for private, paid API keys that should not be shared."""
# Tale je moj - ne uporabljaj, ker ga plačujem
#API keys should be configured via GUI or environment variables

from huggingface_hub import InferenceClient
from huggingface_hub.utils import BadRequestError

"""Comments and helpers around safe default LLM model presets for hosted or shared environments."""
# You can still keep your MODEL_PRESETS elsewhere; these are just safe fallbacks for hosted use.
SAFE_FALLBACK_CHAT_MODELS: List[str] = [
    "google/gemma-2-2b-it",                 # small, widely served on HF router
    "TinyLlama/TinyLlama-1.1B-Chat-v1.0",   # very small, almost always available
]
SAFE_FALLBACK_SUMMARIZER: str = "facebook/bart-large-cnn"

def _extract_chat_text(
    resp: Any,
) -> str:
    """
    Extract plain text from Hugging Face chat_completion response.
    """
    try:
        return (resp.choices[0].message["content"] or "").strip()
    except Exception:
        try:
            return (resp.choices[0].message.content or "").strip()
        except Exception:
            return str(resp).strip()

def _align_for_chat(
    gen_kwargs: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Normalize kwargs for chat_completion: map max_new_tokens -> max_tokens and
    drop text-only args.
    """
    g = dict(gen_kwargs or {})
    if "max_tokens" not in g and "max_new_tokens" in g:
        g["max_tokens"] = g.pop("max_new_tokens")
    g.pop("return_full_text", None)
    return g

def _align_for_text(
    gen_kwargs: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Normalize kwargs for text_generation: map max_tokens -> max_new_tokens and
    drop chat-only args.
    """
    g = dict(gen_kwargs or {})
    if "max_new_tokens" not in g and "max_tokens" in g:
        g["max_new_tokens"] = g.pop("max_tokens")
    for k in ("logprobs", "top_logprobs"):
        g.pop(k, None)
    return g

def _try_chat_once(
    client: InferenceClient,
    prompt: str,
    gen_kwargs: Dict[str, Any],
) -> str:
    """
    Call chat_completion once with aligned kwargs.
    """
    chat_kwargs = _align_for_chat(gen_kwargs)
    chat_kwargs.setdefault("max_tokens", 256)
    chat_kwargs.setdefault("temperature", 0.2)
    resp = client.chat_completion(messages=[{"role": "user", "content": prompt}], **chat_kwargs)
    return _extract_chat_text(resp)

def _try_text_once(
    client: InferenceClient,
    prompt: str,
    gen_kwargs: Dict[str, Any],
) -> str:
    """
    Call text_generation once with aligned kwargs.
    """
    text_kwargs = _align_for_text(gen_kwargs)
    text_kwargs.setdefault("max_new_tokens", 256)
    text_kwargs.setdefault("temperature", 0.2)
    out = client.text_generation(prompt, **text_kwargs)
    return out.strip() if isinstance(out, str) else str(out).strip()

def invoke_llm(
    prompt: str,
    model: str,
    provider: str = 'huggingface',
    hf_token: Optional[str] = None,
    gen_kwargs: Optional[Dict[str, Any]] = None,
    *,
    mode: str = 'chat',
    hf_provider: Optional[str] = None,
    fallback_models: Optional[Iterable[str]] = None,
) -> str:
    """
    Invoke an LLM via Hugging Face InferenceClient with robust fallbacks.

    Parameters
    ----------
    prompt : str
        Input text.
    model : str
        Primary model id (you can pass your alias earlier and resolve before calling).
    provider : str, default="huggingface"
        "huggingface" for general chat/text; "huggingface_summarization"/"bert" handled below.
    hf_token : str | None
        Your HF token. Pass it from your test code; this function does not read env vars.
    gen_kwargs : dict | None
        Generation kwargs. For chat use "max_tokens"; for text use "max_new_tokens".
    mode : str, {"chat","text","auto"}, default="chat"
        Execution mode. "chat" avoids task mismatch for chat-only models.
    hf_provider : str | None
        Force a specific Inference Provider (e.g., "featherless-ai"). Leave None to let the router choose.
    fallback_models : Iterable[str] | None
        Models to try if the primary model is not supported by your enabled providers.
        If None, uses SAFE_FALLBACK_CHAT_MODELS when mode in {"chat","auto"}.

    Returns
    -------
    str
        Generated text.

    Notes
    -----
    - If you see "model_not_supported", it means that model is not available on any provider
      you have enabled. This function will try `fallback_models` automatically.
    """
    if provider not in {"huggingface", "bert", "huggingface_summarization"}:
        raise ValueError(f"Unsupported provider: {provider}")
    if not hf_token:
        raise RuntimeError("Pass your HF token via hf_token=... from your test code.")

    gen_kwargs = gen_kwargs or {}
    fallbacks = list(fallback_models or (SAFE_FALLBACK_CHAT_MODELS if mode in {"chat", "auto"} else []))

    def _client(
        m: str,
    ) -> InferenceClient:
        return InferenceClient(model=m, token=hf_token, provider=hf_provider)

    # --- Summarization branch (explicit) -------------------------------------
    if provider in {"huggingface_summarization", "bert"}:
        # Summarization endpoint typically works with BART/T5 on HF's own infra.
        client = _client(model or SAFE_FALLBACK_SUMMARIZER)
        params = {
            "max_length": gen_kwargs.get("max_length", 256),
            "min_length": gen_kwargs.get("min_length", 30),
            "temperature": gen_kwargs.get("temperature", 0.0),
            "do_sample": gen_kwargs.get("do_sample", False),
        }
        resp = client.summarization(prompt, **params)
        if isinstance(resp, dict):
            for k in ("summary_text", "generated_text", "text"):
                if isinstance(resp.get(k), str):
                    return resp[k].strip()
        if isinstance(resp, list) and resp and isinstance(resp[0], dict):
            for k in ("summary_text", "generated_text", "text"):
                if isinstance(resp[0].get(k), str):
                    return resp[0][k].strip()
        return str(resp).strip()

    # --- Chat/Text branch -----------------------------------------------------
    candidates = [model] + fallbacks
    last_err: Optional[Exception] = None

    for m in candidates:
        try:
            cl = _client(m)
            if mode == "text":
                return _try_text_once(cl, prompt, gen_kwargs)
            # Prefer chat
            return _try_chat_once(cl, prompt, gen_kwargs)
        except BadRequestError as e:
            # Typical when the model is not supported by enabled providers.
            msg = getattr(e, "response", None)
            last_err = e
            continue
        except Exception as e:
            last_err = e
            # If chat failed and mode == "auto", try text_generation for the same model
            if mode == "auto":
                try:
                    cl = _client(m)
                    return _try_text_once(cl, prompt, gen_kwargs)
                except Exception as e2:
                    last_err = e2
                    continue
            continue

    raise RuntimeError(
        "All candidate models failed. Likely the primary model is not supported by your enabled providers. "
        "Enable a provider for that model in your Hugging Face account settings, or pass `fallback_models=` "
        "with models available to your providers."
    ) from last_err

"""Convenience wrappers around LLM calls that avoid relying on global default configuration."""
# ---- Convenience wrappers (no global defaults referenced) --------------------

def llm_summarize_abstracts(
    abstracts: List[str],
    llm_fn: callable = invoke_llm,
    model: Optional[str] = None,
    provider: str = 'huggingface',
    prompt_template: str = 'Summarize the following abstracts into a concise, coherent synthesis (3–5 sentences):\n{abstracts}\n',
    hf_token: Optional[str] = None,
    gen_kwargs: Optional[Dict[str, Any]] = None,
    **invoke_kwargs: Any,
) -> str:
    """
    Summarize multiple abstracts with a hosted open model and automatic fallbacks.

    Parameters
    ----------
    abstracts : list[str]
        Abstract texts to summarize.
    llm_fn : callable, default=invoke_llm
        Backend invoker.
    model : str | None, default=None
        Primary model id. If None, uses the first fallback chat model.
    provider : str, default="huggingface"
        Provider switch ("huggingface" recommended).
    prompt_template : str
        Prompt template with "{abstracts}" placeholder.
    hf_token : str | None
        Your HF token (pass from test code).
    gen_kwargs : dict | None
        Generation kwargs.
    **invoke_kwargs : Any
        Extra kwargs to `invoke_llm` (e.g., mode="chat", hf_provider="...").

    Returns
    -------
    str
        A short synthesis of the abstracts.
    """
    joined = "\n---\n".join(abstracts or [""])
    prompt = prompt_template.format(abstracts=joined)
    # Choose model: prefer caller's model, else first safe fallback
    chosen = model or SAFE_FALLBACK_CHAT_MODELS[0]
    return llm_fn(
        prompt=prompt,
        model=chosen,
        provider=provider,
        hf_token=hf_token,
        gen_kwargs=gen_kwargs,
        mode=invoke_kwargs.pop("mode", "chat"),
        fallback_models=invoke_kwargs.pop("fallback_models", SAFE_FALLBACK_CHAT_MODELS[1:]),
        **invoke_kwargs,
    )

def llm_describe_table(
    table: Any,
    llm_fn: callable = invoke_llm,
    model: Optional[str] = None,
    provider: str = 'huggingface',
    prompt_template: str = 'You are given the following dataframe. Describe its main information and performance highlights in one tight paragraph. For most interesting cases use numbers:\n{table_md}\n',
    hf_token: Optional[str] = None,
    gen_kwargs: Optional[Dict[str, Any]] = None,
    **invoke_kwargs: Any,
) -> str:
    """
    Describe a table with a hosted model and automatic fallbacks.

    Parameters
    ----------
    table : Any
        pandas.DataFrame or any object convertible to string (Markdown preferred).
    llm_fn : callable, default=invoke_llm
        Backend invoker.
    model : str | None, default=None
        Primary model id. If None, uses the first fallback chat model.
    provider : str, default="huggingface"
        Provider switch.
    prompt_template : str
        Prompt template with "{table_md}" placeholder.
    hf_token : str | None
        Your HF token (pass from test code).
    gen_kwargs : dict | None
        Generation kwargs.
    **invoke_kwargs : Any
        Extra kwargs to `invoke_llm`.

    Returns
    -------
    str
        A concise description of the table.
    """
    try:
        import pandas as pd  # type: ignore
        if hasattr(pd, "DataFrame") and isinstance(table, pd.DataFrame):
            table_md = table.to_markdown(index=False)
        else:
            table_md = str(table)
    except Exception:
        table_md = str(table)

    chosen = model or SAFE_FALLBACK_CHAT_MODELS[0]
    prompt = prompt_template.format(table_md=table_md)
    return llm_fn(
        prompt=prompt,
        model=chosen,
        provider=provider,
        hf_token=hf_token,
        gen_kwargs=gen_kwargs,
        mode=invoke_kwargs.pop("mode", "chat"),
        fallback_models=invoke_kwargs.pop("fallback_models", SAFE_FALLBACK_CHAT_MODELS[1:]),
        **invoke_kwargs,
    )

"""Helpers for building and analysing document-level citation networks."""
# Citation network of documents

from thefuzz import fuzz

def normalize_text(text: str) -> str:
    """Normalize text for comparison: lowercase, remove punctuation, collapse whitespace."""
    text = text.lower()
    text = re.sub(r"[\W_]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_title_from_reference(ref: str) -> Optional[str]:
    """
    Extract the likely title portion from a Scopus reference string.
    Scopus format: "Author(s), Title, Journal, Volume, Issue, Pages, (Year)"
    The title is typically the second comma-separated segment.
    """
    # Split by comma and try to identify title
    parts = [p.strip() for p in ref.split(",")]
    
    if len(parts) >= 2:
        # Title is usually the second part (after author)
        # But sometimes author has multiple parts, so we look for the longest
        # non-numeric segment that's not the first one
        candidates = []
        for i, part in enumerate(parts[1:], start=1):
            # Skip if it looks like: volume, pages, year, or very short
            if re.match(r"^\d+$", part):  # Just numbers (volume/issue)
                continue
            if re.match(r"^pp?\.\s*\d+", part):  # Page numbers
                continue
            if re.match(r"^\(\d{4}\)$", part):  # Year in parentheses
                continue
            if len(part) < 10:  # Too short to be a title
                continue
            candidates.append((i, part))
        
        if candidates:
            # Return the first viable candidate (usually the title)
            return candidates[0][1]
    
    return None


def build_citation_network(
    df: pd.DataFrame,
    title_col: str = "Title",
    ref_col: str = "References",
    id_col: str = "EID",
    threshold: int = 80,
    use_token_set: bool = True,
    largest_only: bool = True,
    verbose: bool = False,
    return_stats: bool = False,
) -> tuple[nx.DiGraph, dict[str, list[str]]]:
    """
    Build a citation network from Scopus data.
    
    An edge from node A to node B means: document A cites document B.
    
    Args:
        df: DataFrame with title, references, and ID columns.
        title_col: Column name for document titles.
        ref_col: Column name for references (semicolon-separated).
        id_col: Column name for document IDs (e.g., EID, DOI).
        threshold: Fuzzy-match score threshold (0-100). Default 80.
        use_token_set: Use token_set_ratio (better for word reordering) vs partial_ratio.
        largest_only: If True, return only the largest weakly connected component.
        verbose: Print progress information.
        return_stats: If True, return 3 values (G, unmatched, stats); otherwise 2.
    
    Returns:
        G: Directed graph with document IDs as node labels.
        unmatched: Dict mapping doc ID to list of unmatched reference strings.
        stats: (only if return_stats=True) Dictionary with network statistics.
    """
    # Validate columns
    required_cols = [title_col, ref_col, id_col]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}. Available: {list(df.columns)}")
    
    # Reset index to ensure integer indexing works
    df = df.reset_index(drop=True)
    
    # Prepare normalized titles and create lookup
    titles = df[title_col].tolist()
    doc_ids = df[id_col].tolist()
    norm_titles = [normalize_text(str(t)) if pd.notna(t) else "" for t in titles]
    
    # Create exact match lookup for speed
    title_to_idx = {}
    for idx, nt in enumerate(norm_titles):
        if nt and nt not in title_to_idx:  # Keep first occurrence
            title_to_idx[nt] = idx
    
    # Choose fuzzy matching function
    match_func = fuzz.token_set_ratio if use_token_set else fuzz.partial_ratio
    
    # Initialize graph with all documents as nodes
    G = nx.DiGraph()
    for idx, doc_id in enumerate(doc_ids):
        G.add_node(doc_id, title=titles[idx], index=idx)
    
    unmatched: dict[str, list[str]] = {}
    total_refs = 0
    matched_refs = 0
    
    if verbose:
        print(f"Processing {len(df)} documents...")
    
    for idx, row in df.iterrows():
        refs = row[ref_col]
        source_id = row[id_col]
        
        if not isinstance(refs, str) or pd.isna(refs):
            continue
        
        # Split references (Scopus uses semicolon separator)
        ref_list = [r.strip() for r in refs.split(";") if r.strip()]
        local_unmatched: list[str] = []
        
        for ref in ref_list:
            total_refs += 1
            matched = False
            
            # Try to extract title from reference
            extracted_title = extract_title_from_reference(ref)
            
            # Texts to try matching against
            search_texts = []
            if extracted_title:
                search_texts.append(normalize_text(extracted_title))
            search_texts.append(normalize_text(ref))  # Also try full reference
            
            for search_text in search_texts:
                if matched:
                    break
                    
                # First try exact match on extracted/normalized title
                if search_text in title_to_idx:
                    tgt_idx = title_to_idx[search_text]
                    tgt_id = doc_ids[tgt_idx]
                    if tgt_id != source_id:  # Avoid self-citation
                        G.add_edge(source_id, tgt_id)
                        matched = True
                        matched_refs += 1
                        break
                
                # Fuzzy match
                best_score, best_idx = 0, None
                for j, nt in enumerate(norm_titles):
                    if not nt or j == idx:  # Skip empty or self
                        continue
                    score = match_func(search_text, nt)
                    if score > best_score:
                        best_score, best_idx = score, j
                
                if best_score >= threshold and best_idx is not None:
                    tgt_id = doc_ids[best_idx]
                    G.add_edge(source_id, tgt_id)
                    matched = True
                    matched_refs += 1
                    break
            
            if not matched:
                local_unmatched.append(ref)
        
        if local_unmatched:
            unmatched[source_id] = local_unmatched
        
        if verbose and (idx + 1) % 50 == 0:
            print(f"  Processed {idx + 1}/{len(df)} documents...")
    
    # Remove self-loops (shouldn't exist but just in case)
    G.remove_edges_from(nx.selfloop_edges(G))
    
    # Calculate stats before pruning
    stats = {
        "total_documents": len(df),
        "total_references_processed": total_refs,
        "matched_references": matched_refs,
        "match_rate": matched_refs / total_refs if total_refs > 0 else 0,
        "nodes_before_pruning": G.number_of_nodes(),
        "edges_before_pruning": G.number_of_edges(),
    }
    
    # Remove isolated nodes (no citations in or out within dataset)
    isolates = list(nx.isolates(G))
    G.remove_nodes_from(isolates)
    
    stats["nodes_after_removing_isolates"] = G.number_of_nodes()
    stats["edges_after_removing_isolates"] = G.number_of_edges()
    
    # Optionally keep only largest component
    if largest_only and G.number_of_nodes() > 0:
        comps = list(nx.weakly_connected_components(G))
        stats["num_components"] = len(comps)
        stats["component_sizes"] = sorted([len(c) for c in comps], reverse=True)[:10]
        
        largest = max(comps, key=len)
        G = G.subgraph(largest).copy()
    
    stats["final_nodes"] = G.number_of_nodes()
    stats["final_edges"] = G.number_of_edges()
    
    if verbose:
        print(f"\nNetwork Statistics:")
        print(f"  Total references processed: {total_refs}")
        print(f"  Matched within dataset: {matched_refs} ({stats['match_rate']:.1%})")
        print(f"  Final network: {stats['final_nodes']} nodes, {stats['final_edges']} edges")
    
    # Filter unmatched to only include nodes still in graph
    remaining_nodes = set(G.nodes())
    unmatched = {k: v for k, v in unmatched.items() if k in remaining_nodes}
    
    if return_stats:
        return G, unmatched, stats
    return G, unmatched


def analyze_network(G: nx.DiGraph) -> dict:
    """
    Compute common citation network metrics.
    """
    if G.number_of_nodes() == 0:
        return {"error": "Empty network"}
    
    metrics = {
        "nodes": G.number_of_nodes(),
        "edges": G.number_of_edges(),
        "density": nx.density(G),
    }
    
    # Degree statistics
    in_degrees = dict(G.in_degree())
    out_degrees = dict(G.out_degree())
    
    metrics["avg_in_degree"] = sum(in_degrees.values()) / len(in_degrees)
    metrics["avg_out_degree"] = sum(out_degrees.values()) / len(out_degrees)
    metrics["max_in_degree"] = max(in_degrees.values())
    metrics["max_out_degree"] = max(out_degrees.values())
    
    # Most cited documents (highest in-degree)
    metrics["most_cited"] = sorted(in_degrees.items(), key=lambda x: -x[1])[:10]
    
    # Documents citing most others (highest out-degree)  
    metrics["most_citing"] = sorted(out_degrees.items(), key=lambda x: -x[1])[:10]
    
    # Try PageRank
    try:
        pagerank = nx.pagerank(G)
        metrics["top_pagerank"] = sorted(pagerank.items(), key=lambda x: -x[1])[:10]
    except:
        metrics["top_pagerank"] = None
    
    return metrics

from pathlib import Path


def build_openalex_citation_network(
    df: pd.DataFrame,
    *,
    id_col: str = 'unique-id',
    refs_col: str = 'referenced_works',
    title_col: str = 'title',
    year_col: str = 'publication_year',
    citations_col: str = 'cited_by_count',
    sep: str = '|',
    normalize: Literal['short', 'url', 'auto'] = 'short',
    keep_external: bool = False,
    drop_self_loops: bool = True,
    deduplicate: bool = True,
    keep_largest_component: bool = True,
    path_base: Optional[str] = None,
    save_formats: Iterable[Literal['csv', 'graphml', 'pajek']] = ('csv',),
    verbose: bool = False,
) -> nx.DiGraph:
    """
    Build a directed citation network (citing → cited) from OpenAlex IDs.
    
    This function uses exact ID matching from OpenAlex's referenced_works field,
    which contains direct links to cited works. No fuzzy matching is needed.

    Parameters
    ----------
    df : pd.DataFrame
        Table with OpenAlex work IDs and pipe-separated references.
    id_col : str, default 'unique-id'
        Column name for document IDs.
    refs_col : str, default 'referenced_works'
        Column name for references (pipe-separated OpenAlex IDs).
    title_col : str, default 'title'
        Column name for document titles.
    year_col : str, default 'publication_year'
        Column name for publication year.
    citations_col : str, default 'cited_by_count'
        Column name for citation counts.
    sep : str, default "|"
        Delimiter in `refs_col`.
    normalize : {"short","url","auto"}, default "short"
        ID normalization strategy:
        - "short": Use W123456789 format
        - "url": Use https://openalex.org/W123456789 format
        - "auto": Keep original format
    keep_external : bool, default False
        Keep citations to works outside the dataset (dangling nodes).
    drop_self_loops : bool, default True
        Remove self-citations.
    deduplicate : bool, default True
        Collapse repeated (citing, cited) pairs.
    keep_largest_component : bool, default True
        Keep only the largest *weakly* connected component.
    path_base : str or None, default None
        Base filepath (no extension) for saving outputs.
    save_formats : {"csv","graphml","pajek"}, default ("csv",)
        Formats to write if `path_base` is given.
    verbose : bool, default False
        Print progress information.

    Returns
    -------
    nx.DiGraph
        Directed graph of citations with node attributes:
        - title: Document title
        - year: Publication year
        - citations: Citation count
    """
    url_prefix = "https://openalex.org/"
    tail_pat = re.compile(r"(W\d+)$")

    def to_short(s: str) -> str:
        m = tail_pat.search(s)
        return m.group(1) if m else s

    def to_url(s: str) -> str:
        if s.startswith(url_prefix):
            return s
        m = tail_pat.search(s)
        return f"{url_prefix}{m.group(1)}" if m else s

    def normalize_id(s: str) -> str:
        if normalize == "short":
            return to_short(s)
        if normalize == "url":
            return to_url(s)
        return s if s.startswith(url_prefix) or tail_pat.search(s) else s

    # Find columns with fallbacks
    def find_col(df, options):
        for opt in options:
            if opt in df.columns:
                return opt
        return None
    
    actual_id_col = find_col(df, [id_col, 'id', 'unique-id', 'ids.openalex', 'work_id'])
    actual_refs_col = find_col(df, [refs_col, 'referenced_works', 'references'])
    actual_title_col = find_col(df, [title_col, 'title', 'Title', 'display_name'])
    actual_year_col = find_col(df, [year_col, 'publication_year', 'Year', 'year'])
    actual_cite_col = find_col(df, [citations_col, 'cited_by_count', 'Cited by', 'citations'])
    
    if actual_id_col is None:
        raise ValueError(f"No ID column found. Tried: {id_col}, id, unique-id, ids.openalex")
    if actual_refs_col is None:
        raise ValueError(f"No references column found. Tried: {refs_col}, referenced_works, references")
    
    if verbose:
        print(f"OpenAlex Citation Network Builder")
        print(f"  ID column: {actual_id_col}")
        print(f"  References column: {actual_refs_col}")
        print(f"  Documents: {len(df)}")

    # Normalize IDs
    ids = df[actual_id_col].dropna().astype(str).map(str.strip)
    if normalize != "auto":
        ids = ids.map(normalize_id)
    id_set = set(ids)

    # Build ID to row data mapping for node attributes
    id_to_data = {}
    for idx, row in df.iterrows():
        raw_id = row.get(actual_id_col)
        if pd.isna(raw_id):
            continue
        node_id = normalize_id(str(raw_id).strip()) if normalize != "auto" else str(raw_id).strip()
        
        title = row.get(actual_title_col, '') if actual_title_col else ''
        year = row.get(actual_year_col, 2000) if actual_year_col else 2000
        citations = row.get(actual_cite_col, 0) if actual_cite_col else 0
        
        id_to_data[node_id] = {
            'title': str(title)[:100] if pd.notna(title) else node_id,
            'year': int(year) if pd.notna(year) and year else 2000,
            'citations': int(citations) if pd.notna(citations) else 0,
        }

    # Process references
    refs = (
        df[[actual_id_col, actual_refs_col]]
        .copy()
        .assign(
            **{
                actual_id_col: df[actual_id_col].astype(str).str.strip(),
                actual_refs_col: df[actual_refs_col]
                .fillna("")
                .astype(str)
                .map(lambda x: [r.strip() for r in x.split(sep) if r.strip()]),
            }
        )
    )

    if normalize != "auto":
        refs[actual_id_col] = refs[actual_id_col].map(normalize_id)
        refs[actual_refs_col] = refs[actual_refs_col].map(lambda lst: [normalize_id(r) for r in lst])

    edges = (
        refs.explode(actual_refs_col, ignore_index=True)
        .dropna(subset=[actual_refs_col])
        .rename(columns={actual_id_col: "citing", actual_refs_col: "cited"})
    )
    
    total_refs = len(edges)

    if not keep_external:
        edges = edges[edges["cited"].isin(id_set)]
    
    internal_refs = len(edges)

    if drop_self_loops:
        edges = edges[edges["citing"] != edges["cited"]]

    if deduplicate:
        edges = edges.drop_duplicates(["citing", "cited"])

    # Build graph with node attributes
    G = nx.DiGraph()
    
    for node_id in id_set:
        data = id_to_data.get(node_id, {'title': node_id, 'year': 2000, 'citations': 0})
        G.add_node(node_id, **data)
    
    G.add_edges_from(edges.itertuples(index=False, name=None))

    if verbose:
        print(f"  Total references: {total_refs}")
        print(f"  Internal references (within corpus): {internal_refs}")
        print(f"  Edges after dedup: {G.number_of_edges()}")

    # Keep only the largest weakly connected component (if requested)
    if keep_largest_component and G.number_of_nodes() > 0:
        if G.number_of_edges() == 0:
            # If no edges, keep all nodes
            pass
        else:
            largest = max(nx.weakly_connected_components(G), key=len)
            G = G.subgraph(largest).copy()

    if verbose:
        print(f"  Final network: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

    # Optional save
    if path_base is not None:
        base = Path(path_base)
        base.parent.mkdir(parents=True, exist_ok=True)

        if "csv" in save_formats:
            edf = pd.DataFrame(G.edges(), columns=["citing", "cited"])
            edf.to_csv(base.with_suffix(".csv"), index=False)

        if "graphml" in save_formats:
            nx.write_graphml(G, base.with_suffix(".graphml"))

        if "pajek" in save_formats:
            nx.write_pajek(G, base.with_suffix(".net"))

    return G

def compute_main_path(
    G: nx.DiGraph,
) -> list[str]:
    """
    Compute the main citation path. Condenses cycles to DAG.

    Args:
        G: Directed graph with Doc ID labels.

    Returns:
        List of Doc ID labels on the critical path.
    """
    cG = nx.condensation(G)
    comp_path = nx.dag_longest_path(cG)
    main_path = []
    for comp in comp_path:
        members = sorted(cG.nodes[comp]["members"])
        main_path.append(members[0])
    return main_path

"""Functions for constructing historiographs and time-ordered citation structures."""
# historiograph

def extract_reference_titles(
    ref_string,
):
    """Extract potential titles from raw reference strings using pattern-based heuristics."""
    if not isinstance(ref_string, str):
        return []
    refs = [s.strip() for s in ref_string.split(";") if s.strip()]
    cleaned = []
    for ref in refs:
        # Remove author names and years: keep the middle (likely title) segment
        parts = [part.strip() for part in ref.split(",")]
        # Heuristic: remove authors, year, source; keep middle part
        if len(parts) >= 3:
            candidate = parts[2]  # usually the 3rd component has the title or journal name
        else:
            candidate = parts[0]
        cleaned.append(candidate)
    return cleaned

def approximate_match(
    ref_title,
    titles,
    cutoff = 0.85,
):
    """Find the closest match to a reference title among known titles."""
    if not isinstance(ref_title, str):
        return None
    matches = difflib.get_close_matches(ref_title.lower(), titles, n=1, cutoff=cutoff)
    return matches[0] if matches else None

def build_historiograph(
    df,
    title_col = 'Title',
    year_col = 'Year',
    refs_col = 'References',
    cutoff = 0.85,
    label_col = 'Document Short Label',
    weight_col = 'Cited by',
    save_path = None,
):
    """Construct a directed citation graph using approximate title matching and optional node relabeling."""
    G = nx.DiGraph()

    title_map = {title.lower(): title for title in df[title_col].dropna().unique()}

    label_map = {}
    for _, row in df.iterrows():
        title = row.get(title_col)
        label = row.get(label_col) if label_col and pd.notna(row.get(label_col)) else title
        weight = row.get(weight_col, 0) if weight_col in row else 0
        if pd.notna(title) and pd.notna(row.get(year_col)):
            label_map[title.lower()] = label
            G.add_node(label, year=row[year_col], title=label, **{weight_col: weight})

    for _, row in df.iterrows():
        citing_title = row.get(title_col)
        citing_label = label_map.get(str(citing_title).lower()) if citing_title and pd.notna(citing_title) else None
        if not citing_label:
            continue

        cited_titles = extract_reference_titles(row.get(refs_col))

        for ref_title in cited_titles:
            match_key = approximate_match(ref_title, list(title_map.keys()), cutoff=cutoff)
            if match_key:
                matched_title = title_map[match_key]
                cited_label = label_map.get(matched_title.lower())
                if cited_label and cited_label != citing_label:
                    G.add_edge(cited_label, citing_label)

    if save_path:
        save_network(G, save_path)

    return G


def build_historiograph_auto(
    df,
    db: str = None,
    cutoff: float = 0.85,
    min_citations: int = 1,
    separator: str = "; ",
):
    """
    Build a historiograph (citation network) with automatic column detection for different databases.
    
    Parameters
    ----------
    df : pd.DataFrame
        Bibliographic DataFrame.
    db : str, optional
        Database type: 'scopus', 'wos', 'oa' (OpenAlex), 'pubmed', etc.
        If None, will attempt auto-detection.
    cutoff : float, default 0.85
        Similarity threshold for title matching (0-1).
    min_citations : int, default 1
        Minimum citations to include a document.
    separator : str, default "; "
        Separator for multi-value fields.
        
    Returns
    -------
    nx.DiGraph
        Directed citation graph with year, citations, and title attributes on nodes.
    """
    nx = _get_networkx()
    
    # Column mappings for different databases
    COLUMN_MAPPINGS = {
        'scopus': {
            'title': ['Title', 'title'],
            'year': ['Year', 'year', 'PY'],
            'refs': ['References', 'references', 'Cited References'],
            'citations': ['Cited by', 'cited_by', 'Times Cited'],
            'authors': ['Authors', 'authors', 'AU'],
            'doi': ['DOI', 'doi'],
        },
        'wos': {
            'title': ['Title', 'TI', 'Article Title'],
            'year': ['Year', 'PY', 'Publication Year'],
            'refs': ['Cited References', 'CR', 'References'],
            'citations': ['Times Cited', 'TC', 'Cited by', 'Times Cited, All Databases'],
            'authors': ['Authors', 'AU', 'Author Full Names', 'AF'],
            'doi': ['DOI', 'DI'],
        },
        'oa': {
            'title': ['title', 'Title', 'display_name'],
            'year': ['publication_year', 'Year', 'year'],
            'refs': ['referenced_works', 'References', 'references'],
            'citations': ['cited_by_count', 'Cited by', 'citations'],
            'authors': ['authorships', 'Authors', 'authors'],
            'doi': ['doi', 'DOI'],
        },
        'pubmed': {
            'title': ['Title', 'TI', 'ArticleTitle'],
            'year': ['Year', 'PY', 'PubDate'],
            'refs': ['References', 'RN'],
            'citations': ['Cited by', 'Citations'],
            'authors': ['Authors', 'AU', 'FAU'],
            'doi': ['DOI', 'AID'],
        },
    }
    
    def find_column(df, col_options):
        """Find first matching column from options."""
        for col in col_options:
            if col in df.columns:
                return col
        return None
    
    # Auto-detect database if not specified
    if db is None:
        # Check for database-specific columns
        if 'Cited by' in df.columns and 'Source title' in df.columns:
            db = 'scopus'
        elif 'Times Cited' in df.columns or 'WoS Categories' in df.columns:
            db = 'wos'
        elif 'cited_by_count' in df.columns or 'publication_year' in df.columns:
            db = 'oa'
        else:
            db = 'scopus'  # Default fallback
    
    # Get column mapping
    mapping = COLUMN_MAPPINGS.get(db, COLUMN_MAPPINGS['scopus'])
    
    # Find columns
    title_col = find_column(df, mapping['title'])
    year_col = find_column(df, mapping['year'])
    refs_col = find_column(df, mapping['refs'])
    cite_col = find_column(df, mapping['citations'])
    authors_col = find_column(df, mapping['authors'])
    
    if not title_col:
        raise ValueError(f"No title column found. Available columns: {list(df.columns)[:10]}")
    if not year_col:
        raise ValueError(f"No year column found. Available columns: {list(df.columns)[:10]}")
    if not refs_col:
        raise ValueError(f"No references column found. Available columns: {list(df.columns)[:10]}")
    
    # Build the graph
    G = nx.DiGraph()
    
    # Create lookup structures
    title_to_data = {}  # lowercase title -> data dict
    label_to_node = {}  # short label -> ensure uniqueness
    
    for idx, row in df.iterrows():
        title = row.get(title_col)
        if pd.isna(title):
            continue
            
        title_str = str(title).strip()
        if len(title_str) < 5:  # Skip very short titles
            continue
            
        title_lower = title_str.lower()
        
        year = row.get(year_col, 0)
        if pd.isna(year):
            continue
        try:
            year = int(float(year))
        except:
            continue
        
        if year < 1900 or year > 2100:  # Skip invalid years
            continue
            
        citations = row.get(cite_col, 0) if cite_col else 0
        if pd.isna(citations):
            citations = 0
        try:
            citations = int(float(citations))
        except:
            citations = 0
        
        # Filter by min citations (0 means include all)
        if min_citations > 0 and citations < min_citations:
            continue
        
        # Create short label
        authors = row.get(authors_col, "") if authors_col else ""
        if pd.isna(authors):
            authors = ""
        
        # Parse first author more robustly
        author_str = str(authors)
        first_author = author_str.split(separator)[0].split(",")[0].split(";")[0].strip()[:20]
        if not first_author:
            first_author = "Unknown"
        
        # Ensure unique labels
        short_label = f"{first_author} ({year})"
        if short_label in label_to_node:
            # Add suffix if duplicate
            counter = 2
            while f"{short_label}_{counter}" in label_to_node:
                counter += 1
            short_label = f"{short_label}_{counter}"
        
        label_to_node[short_label] = title_lower
        
        title_to_data[title_lower] = {
            'title': title_str,
            'year': year,
            'citations': citations,
            'label': short_label,
        }
        
        # Add node
        G.add_node(short_label, 
                   title=title_str[:100],
                   year=year, 
                   citations=citations,
                   full_title=title_str)
    
    if len(G.nodes()) == 0:
        raise ValueError(f"No valid documents found. Check year and citation columns.")
    
    # Build edges based on reference matching
    if refs_col:
        titles_list = list(title_to_data.keys())
        
        # Also create a list of title keywords for partial matching
        title_keywords = {}
        for t in titles_list:
            # Extract significant words (4+ chars)
            words = [w for w in t.split() if len(w) >= 4 and w.isalpha()]
            if len(words) >= 2:
                title_keywords[t] = set(words[:5])  # First 5 significant words
        
        edges_added = 0
        
        for idx, row in df.iterrows():
            citing_title = row.get(title_col)
            if pd.isna(citing_title):
                continue
            
            citing_lower = str(citing_title).lower().strip()
            citing_data = title_to_data.get(citing_lower)
            if not citing_data:
                continue
            
            citing_label = citing_data['label']
            
            refs = row.get(refs_col)
            if pd.isna(refs):
                continue
            
            # Parse references
            ref_str = str(refs)
            ref_list = ref_str.split(separator)
            
            for ref in ref_list:
                ref = ref.strip()
                if len(ref) < 20:  # Skip very short references
                    continue
                
                ref_lower = ref.lower()
                
                # Method 1: Try direct fuzzy matching on reference parts
                ref_parts = [p.strip() for p in ref.replace(";", ",").split(",")]
                
                matched = False
                for part in ref_parts:
                    if len(part) < 15:  # Skip short parts
                        continue
                    
                    # Try fuzzy matching
                    matches = difflib.get_close_matches(part.lower(), titles_list, n=1, cutoff=cutoff)
                    if matches:
                        matched_title = matches[0]
                        matched_data = title_to_data.get(matched_title)
                        if matched_data and matched_data['label'] != citing_label:
                            # Edge from cited -> citing (knowledge flow)
                            if matched_data['year'] <= citing_data['year']:
                                G.add_edge(matched_data['label'], citing_label)
                                edges_added += 1
                                matched = True
                                break
                
                # Method 2: Keyword-based matching if no match found
                if not matched and title_keywords:
                    ref_words = set(w.lower() for w in ref_lower.split() if len(w) >= 4 and w.isalpha())
                    if len(ref_words) >= 2:
                        best_match = None
                        best_overlap = 0
                        for title, keywords in title_keywords.items():
                            overlap = len(ref_words & keywords)
                            if overlap >= 3 and overlap > best_overlap:  # At least 3 matching keywords
                                best_overlap = overlap
                                best_match = title
                        
                        if best_match:
                            matched_data = title_to_data.get(best_match)
                            if matched_data and matched_data['label'] != citing_label:
                                if matched_data['year'] <= citing_data['year']:
                                    G.add_edge(matched_data['label'], citing_label)
                                    edges_added += 1
    
    # If no edges found, keep all nodes but warn
    if len(G.edges()) == 0:
        # Don't remove nodes - just return the graph with no edges
        # This allows user to see their documents even if no internal citations
        pass
    else:
        # Remove isolated nodes only if we have edges
        isolated = list(nx.isolates(G))
        G.remove_nodes_from(isolated)
    
    return G


def build_thematic_map(
    df: pd.DataFrame,
    field: str = "keywords",
    field_col: str = None,
    min_occurrences: int = 2,
    top_n: int = 50,
    separator: str = "; ",
    synonyms: dict = None,
    db: str = None,
) -> dict:
    """
    Build a thematic map (strategic diagram) based on Callon's centrality-density analysis.
    
    The thematic map plots research themes in a 2D space where:
    - X-axis (Centrality): measures external links to other clusters (theme importance)
    - Y-axis (Density): measures internal cohesion within clusters (theme development)
    
    Quadrants:
    - Motor themes (high centrality, high density): well-developed, central topics
    - Basic themes (low centrality, high density): transversal, general topics
    - Emerging/Declining themes (low centrality, low density): peripheral topics
    - Niche themes (high centrality, low density): specialized but important
    
    Parameters
    ----------
    df : pd.DataFrame
        Bibliographic DataFrame.
    field : str, default "keywords"
        Field to analyze: "keywords", "authors", "sources", "countries".
    field_col : str, optional
        Specific column name. If None, auto-detected based on field type.
    min_occurrences : int, default 2
        Minimum occurrences to include an item.
    top_n : int, default 50
        Maximum number of items to include.
    separator : str, default "; "
        Separator for multi-value fields.
    synonyms : dict, optional
        Dictionary mapping items to their canonical form (e.g., {"ai": "artificial intelligence"}).
    db : str, optional
        Database type for column detection.
        
    Returns
    -------
    dict
        Dictionary containing:
        - 'clusters_df': DataFrame with cluster statistics
        - 'graph': NetworkX graph of co-occurrences
        - 'partition': Node-to-cluster mapping
        - 'items_df': DataFrame with item statistics
    """
    nx = _get_networkx()
    
    # Column mappings for different field types and databases
    FIELD_COLUMNS = {
        'keywords': ['Author Keywords', 'Processed Author Keywords', 'Keywords', 'DE', 
                     'author_keywords', 'keywords'],
        'index_keywords': ['Index Keywords', 'ID', 'Processed Index Keywords', 'index_keywords'],
        'authors': ['Authors', 'AU', 'Author Full Names', 'AF', 'authors'],
        'sources': ['Source title', 'Source Title', 'SO', 'Journal', 'source_title', 'journal'],
        'countries': ['Countries', 'Country', 'CU', 'countries', 'Affiliations Countries'],
        'institutions': ['Affiliations', 'Institution', 'C1', 'affiliations'],
    }
    
    # Find the column
    if field_col and field_col in df.columns:
        col = field_col
    else:
        col = None
        col_options = FIELD_COLUMNS.get(field.lower(), [])
        for c in col_options:
            if c in df.columns:
                col = c
                break
        
        if col is None:
            # Try case-insensitive match
            for c in df.columns:
                if field.lower() in c.lower():
                    col = c
                    break
    
    if col is None:
        raise ValueError(f"Could not find column for field '{field}'. Available: {list(df.columns)[:10]}")
    
    # Apply synonyms function
    def apply_synonym(item):
        if synonyms and item.lower() in synonyms:
            return synonyms[item.lower()]
        return item
    
    # Extract and count items
    items_series = df[col].dropna().astype(str)
    all_items = []
    
    for val in items_series:
        parts = [p.strip() for p in val.split(separator) if p.strip()]
        if synonyms:
            parts = [apply_synonym(p) for p in parts]
        all_items.extend(parts)
    
    item_counts = pd.Series(all_items).value_counts()
    
    # Filter by min occurrences and take top N
    valid_items = item_counts[item_counts >= min_occurrences].head(top_n)
    valid_item_set = set(valid_items.index)
    
    if len(valid_item_set) < 3:
        raise ValueError(f"Not enough items with min {min_occurrences} occurrences. Found {len(valid_item_set)}.")
    
    # Build co-occurrence graph
    G = nx.Graph()
    
    # Add nodes with occurrence counts
    for item, count in valid_items.items():
        G.add_node(item, occurrences=int(count))
    
    # Build edges (co-occurrence)
    for val in items_series:
        parts = [p.strip() for p in val.split(separator) if p.strip()]
        if synonyms:
            parts = [apply_synonym(p) for p in parts]
        parts = [p for p in parts if p in valid_item_set]
        
        # Add edges for co-occurring items
        for i, item1 in enumerate(parts):
            for item2 in parts[i+1:]:
                if item1 != item2:
                    if G.has_edge(item1, item2):
                        G[item1][item2]["weight"] += 1
                    else:
                        G.add_edge(item1, item2, weight=1)
    
    # Remove isolated nodes
    isolated = list(nx.isolates(G))
    G.remove_nodes_from(isolated)
    
    if len(G.nodes()) < 3:
        raise ValueError("Not enough connected nodes for clustering after removing isolates.")
    
    # Community detection
    try:
        import community as community_louvain
        partition = community_louvain.best_partition(G, weight='weight')
    except ImportError:
        # Fallback to greedy modularity
        communities = list(nx.algorithms.community.greedy_modularity_communities(G, weight='weight'))
        partition = {}
        for i, comm in enumerate(communities):
            for node in comm:
                partition[node] = i
    
    # Group nodes by cluster
    clusters = {}
    for node, cluster_id in partition.items():
        if cluster_id not in clusters:
            clusters[cluster_id] = []
        clusters[cluster_id].append(node)
    
    # Calculate metrics for each cluster
    cluster_data = []
    
    for cluster_id, nodes in clusters.items():
        if len(nodes) == 0:
            continue
        
        # Subgraph for this cluster
        subG = G.subgraph(nodes)
        
        # Density: internal weighted links / possible links
        n_nodes = len(nodes)
        internal_weight = sum(d.get('weight', 1) for _, _, d in subG.edges(data=True))
        possible_edges = n_nodes * (n_nodes - 1) / 2 if n_nodes > 1 else 1
        density = internal_weight / possible_edges if possible_edges > 0 else 0
        
        # Centrality: external weighted links to other clusters / number of nodes
        external_weight = 0
        for node in nodes:
            for neighbor in G.neighbors(node):
                if neighbor not in nodes:
                    external_weight += G[node][neighbor].get("weight", 1)
        centrality = external_weight / n_nodes if n_nodes > 0 else 0
        
        # Total occurrences (documents) in cluster
        total_occurrences = sum(G.nodes[n].get("occurrences", 1) for n in nodes)
        
        # Get top keywords in cluster by occurrences
        sorted_nodes = sorted(nodes, key=lambda n: G.nodes[n].get("occurrences", 0), reverse=True)
        top_keywords = sorted_nodes[:5]
        label = ", ".join(top_keywords[:3])
        
        # Callon centrality (rank centrality)
        callon_centrality = centrality
        
        # Callon density 
        callon_density = density
        
        cluster_data.append({
            "Cluster": cluster_id,
            "Label": label,
            "Items": n_nodes,
            "Occurrences": total_occurrences,
            "Centrality": round(centrality, 4),
            "Density": round(density, 4),
            "Top_Items": "; ".join(top_keywords),
            "All_Items": "; ".join(sorted_nodes),
        })
    
    clusters_df = pd.DataFrame(cluster_data)
    
    # Determine quadrant for each cluster based on medians
    if len(clusters_df) > 0:
        centrality_median = clusters_df["Centrality"].median()
        density_median = clusters_df["Density"].median()
        
        def get_quadrant(row):
            high_c = row["Centrality"] >= centrality_median
            high_d = row["Density"] >= density_median
            if high_c and high_d:
                return "Motor"
            elif not high_c and high_d:
                return "Basic"
            elif not high_c and not high_d:
                return "Emerging/Declining"
            else:
                return "Niche"
        
        clusters_df["Quadrant"] = clusters_df.apply(get_quadrant, axis=1)
        clusters_df["Centrality_Median"] = centrality_median
        clusters_df["Density_Median"] = density_median
    
    # Create items DataFrame with their cluster assignments
    items_data = []
    for node in G.nodes():
        items_data.append({
            "Item": node,
            "Occurrences": G.nodes[node].get("occurrences", 0),
            "Degree": G.degree(node),
            "Weighted_Degree": G.degree(node, weight='weight'),
            "Cluster": partition.get(node, -1),
        })
    
    items_df = pd.DataFrame(items_data)
    items_df = items_df.sort_values(["Cluster", "Occurrences"], ascending=[True, False])
    
    return {
        "clusters_df": clusters_df,
        "items_df": items_df,
        "graph": G,
        "partition": partition,
        "field": field,
        "field_column": col,
    }


"""Additional network-analysis helpers for plotting, filtering and describing graphs."""
# Network analysis functions

def louvain_partition(
    G,
    resolution = 1.0,
    randomize = False,
    random_state = None,
):
    """
    Detect communities using the Louvain method.

    Parameters
    ----------
    G : networkx.Graph
    resolution : float, optional
    randomize : bool, optional
    random_state : int or RandomState, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if community_louvain is None:
        raise ImportError("python-louvain package is required for Louvain partition")
    return community_louvain.best_partition(
        G, resolution=resolution, randomize=randomize, random_state=random_state
    )

def greedy_modularity_partition(
    G,
    weight = 'weight',
):
    """
    Detect communities by greedy modularity maximization.

    Parameters
    ----------
    G : networkx.Graph
    weight : str or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.greedy_modularity_communities(G, weight=weight)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def label_propagation_partition(
    G,
    weight = None,
    seed = None,
):
    """
    Detect communities via asynchronous label propagation.

    Parameters
    ----------
    G : networkx.Graph
    weight : str or None, optional
    seed : int or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.asyn_lpa_communities(G, weight=weight, seed=seed)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def girvan_newman_partition(
    G,
    n_communities = 2,
):
    """
    Detect communities using the Girvan-Newman algorithm.

    Parameters
    ----------
    G : networkx.Graph
    n_communities : int, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if n_communities < 2:
        raise ValueError("n_communities must be >= 2")
    comp_gen = nx.algorithms.community.girvan_newman(G)
    for _ in range(n_communities - 2):
        next(comp_gen)
    communities = next(comp_gen)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def k_clique_partition(
    G,
    k = 3,
):
    """
    Detect communities using k-clique percolation.

    Parameters
    ----------
    G : networkx.Graph
    k : int, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    communities = nx.algorithms.community.k_clique_communities(G, k)
    return {node: cid for cid, comm in enumerate(communities) for node in comm}

def kernighan_lin_partition(
    G,
    max_iter = 10,
):
    """
    Bisect graph using Kernighan-Lin algorithm.

    Parameters
    ----------
    G : networkx.Graph
    max_iter : int, optional

    Returns
    -------
    dict
        Node-to-partition mapping {0,1}.
    """
    sets = nx.algorithms.community.kernighan_lin_bisection(G, max_iter=max_iter)
    partition = {node: 0 for node in sets[0]}
    partition.update({node: 1 for node in sets[1]})
    return partition

def edge_betweenness_partition(
    G,
    n_communities = 2,
):
    """
    Alias for Girvan-Newman via edge betweenness.
    """
    return girvan_newman_partition(G, n_communities)

def walktrap_partition(
    G,
    steps = 4,
    weights = None,
):
    """
    Detect communities using the Walktrap algorithm via igraph.

    Parameters
    ----------
    G : networkx.Graph
    steps : int, optional
    weights : str or None, optional

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    g = ig.Graph.TupleList(G.edges(data=bool(weights)), directed=False,
                            edge_attrs=[weights] if weights else [])
    wc = g.community_walktrap(weights=weights, steps=steps).as_clustering()
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def infomap_partition(
    G,
    edge_weights = 'weight',
):
    """
    Detect communities using the Infomap algorithm.
    """
    g = ig.Graph.TupleList(G.edges(data=edge_weights), directed=False,
                            edge_attrs=[edge_weights])
    wc = g.community_infomap(edge_weights)
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def leading_eigenvector_partition(
    G,
):
    """
    Detect communities using leading eigenvector method.
    """
    g = ig.Graph.TupleList(G.edges(), directed=False)
    wc = g.community_leading_eigenvector()
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def leiden_partition(
    G,
    resolution_parameter = 1.0,
):
    """
    Detect communities using the Leiden algorithm.
    """
    g = ig.Graph.TupleList(G.edges(), directed=False)
    wc = g.community_leiden(resolution_parameter=resolution_parameter)
    return {v['name']: cid for cid, cluster in enumerate(wc) for v in g.vs if v.index in cluster}

def spinglass_partition(
    G,
    weights = None,
    start_temp = 1.0,
    stop_temp = 0.01,
    cool_fact = 0.99,
    spins = 2,
):
    """
    Detect communities using the Spinglass algorithm via igraph.

    Parameters
    ----------
    G : networkx.Graph
        The graph to partition.
    weights : str or None, optional
        Edge attribute name for weights.
    start_temp : float, optional
        Starting temperature for the spin glass model.
    stop_temp : float, optional
        Stopping temperature for the model.
    cool_fact : float, optional
        Cooling factor between iterations.
    spins : int, optional
        Number of spins (must be >= 2).

    Returns
    -------
    dict
        Node-to-community mapping.
    """
    if spins < 2:
        spins = 2
    g = ig.Graph.TupleList(G.edges(data=bool(weights)), directed=False,
                            edge_attrs=[weights] if weights else [])
    wc = g.community_spinglass(weights=weights, start_temp=start_temp,
                                stop_temp=stop_temp, cool_fact=cool_fact,
                                spins=spins)
    return {v['name']: wc.membership[v.index] for v in g.vs}

def add_partitions(
    G,
    louvain_kwargs = None,
    greedy_kwargs = None,
    label_kwargs = None,
    girvan_kwargs = None,
    k_clique_kwargs = None,
    kernighan_kwargs = None,
    walktrap_kwargs = None,
    edge_betweenness_kwargs = None,
    infomap_kwargs = None,
    leading_kwargs = None,
    leiden_kwargs = None,
    spinglass_kwargs = None,
):
    """
    Compute and attach multiple community partitions to G.

    Returns
    -------
    dict
        Mapping method names to partition dicts.
    """
    configs = {
        "walktrap": (walktrap_partition, walktrap_kwargs or {}),
        "edge_betweenness": (edge_betweenness_partition, edge_betweenness_kwargs or {}),
        "infomap": (infomap_partition, infomap_kwargs or {}),
        "leading_eigenvector": (leading_eigenvector_partition, leading_kwargs or {}),
        "leiden": (leiden_partition, leiden_kwargs or {}),
        "spinglass": (spinglass_partition, spinglass_kwargs or {}),
        "louvain": (louvain_partition, louvain_kwargs or {}),
        "greedy_modularity": (greedy_modularity_partition, greedy_kwargs or {}),
        "label_propagation": (label_propagation_partition, label_kwargs or {}),
        "girvan_newman": (girvan_newman_partition, girvan_kwargs or {}),
        "k_clique": (k_clique_partition, k_clique_kwargs or {}),
        "kernighan_lin": (kernighan_lin_partition, kernighan_kwargs or {}),
    }
    results = {}
    for name, (func, kwargs) in configs.items():
        try:
            partition = func(G, **kwargs)
            results[name] = partition
            for node, cid in partition.items():
                G.nodes[node][f"partition_{name}"] = cid
        except:
            pass
    return results

def add_vectors_from_dataframe(
    G,
    df,
    node_col,
    vector_cols,
):
    """
    Add vector attributes to nodes from a DataFrame.

    Parameters
    ----------
    G : networkx.Graph
    df : pandas.DataFrame
    node_col : str
        Column name for node IDs.
    vector_cols : list of str
        DataFrame columns to use as node attributes.
    """
    for _, row in df.iterrows():
        node = row[node_col]
        if node in G:
            for col in vector_cols:
                G.nodes[node][col] = row[col]
    return G

def nodes_to_dataframe(
    G,
):
    """
    Export all node attributes to a pandas DataFrame, including partitions and vector values.

    Parameters
    ----------
    G : networkx.Graph
        The graph whose node attributes are to be exported.

    Returns
    -------
    pandas.DataFrame
        DataFrame with one row per node and one column per node attribute.
    """
    nodes = list(G.nodes(data=True))
    keys = set()
    for _, attrs in nodes:
        keys.update(attrs.keys())
    columns = ["node"] + sorted(keys)
    records = []
    for n, attrs in nodes:
        rec = {"node": n}
        for k in keys:
            rec[k] = attrs.get(k)
        records.append(rec)
    return pd.DataFrame(records, columns=columns)

def save_network(
    G,
    filename,
    formats=None,
    vector_cols=None,
    partition_attr="partition",
):
    """
    Save a NetworkX graph in various formats, including Pajek (.net, .clu, .vec),
    GraphML, and GEXF.

    Parameters
    ----------
    G : networkx.Graph
        The graph to save.
    filename : str
        Path/prefix for saving files (no extension).
    formats : list[str], optional
        Formats to save ("pajek", "graphml", "gexf"). Defaults to all.
    vector_cols : list[str], optional
        Node attributes to export as .vec files.
    partition_attr : str, optional
        Node attribute prefix to use for .clu partition file(s).
        All node attributes whose names start with this prefix are exported
        as separate .clu files.
    """
    if formats is None:
        formats = ["pajek", "graphml", "gexf"]

    # --- Pajek export (.net, .clu, .vec) ------------------------------------
    if "pajek" in formats:
        # Copy G and convert all node/edge attributes to strings for Pajek.
        # This avoids "Non-string attribute" warnings from NetworkX's Pajek writer.
        G_pajek = G.copy()

        # Node attributes -> strings
        for _, attrs in G_pajek.nodes(data=True):
            for key, value in list(attrs.items()):
                if value is None:
                    # Drop None-valued attributes
                    del attrs[key]
                elif not isinstance(value, str):
                    attrs[key] = str(value)

        # Edge attributes -> strings
        for _, _, attrs in G_pajek.edges(data=True):
            for key, value in list(attrs.items()):
                if value is None:
                    del attrs[key]
                elif not isinstance(value, str):
                    attrs[key] = str(value)

        # Main Pajek network file
        nx.write_pajek(G_pajek, f"{filename}.net")

        # Partition .clu files (use ORIGINAL G so cluster IDs stay numeric)
        if any(
            any(k.startswith(partition_attr) for k in d)
            for _, d in G.nodes(data=True)
        ):
            partition_attrs = set()
            for _, d in G.nodes(data=True):
                partition_attrs.update(
                    k for k in d if k.startswith(partition_attr)
                )

            # Save each partition to its own .clu file
            for attr in sorted(partition_attrs):  # sorted for consistent order
                with open(f"{filename}_{attr}.clu", "w", encoding="utf-8") as f:
                    f.write(f"*Vertices {G.number_of_nodes()}\n")
                    for n in G.nodes():
                        # Pajek expects integers, but str(int) is fine in text file
                        f.write(f"{G.nodes[n].get(attr, 1)}\n")

        # Vector .vec files (again, use ORIGINAL G)
        if vector_cols:
            for vec_col in vector_cols:
                with open(f"{filename}_{vec_col}.vec", "w", encoding="utf-8") as f:
                    f.write(f"*Vertices {G.number_of_nodes()}\n")
                    for n in G.nodes():
                        f.write(f"{G.nodes[n].get(vec_col, 0)}\n")

    # --- GraphML -------------------------------------------------------------
    if "graphml" in formats:
        nx.write_graphml(G, f"{filename}.graphml")

    # --- GEXF ----------------------------------------------------------------
    if "gexf" in formats:
        nx.write_gexf(G, f"{filename}.gexf")


def save_to_pajek(
    G,
    basename,
    partition_attrs = None,
    vector_attrs = None,
):
    """
    Export graph and node attributes to Pajek files.

    Parameters
    ----------
    G : networkx.Graph
    basename : str
        Base filename for .net, .clu, .vec files.
    partition_attrs : str or list of str, optional
    vector_attrs : list of str, optional
    """
    nx.write_pajek(G, f"{basename}.net")
    nodes = list(G.nodes())
    if partition_attrs:
        attrs = [partition_attrs] if isinstance(partition_attrs, str) else partition_attrs
        for attr in attrs:
            fname = f"{basename}.clu" if len(attrs) == 1 else f"{basename}_{attr}.clu"
            with open(fname, "w") as f:
                f.write(f"*Vertices {len(nodes)}\n")
                for n in nodes:
                    f.write(str(G.nodes[n].get(attr, 0)) + "\n")
    if vector_attrs:
        with open(f"{basename}.vec", "w") as f:
            f.write(f"*Vertices {len(nodes)}\n")
            for n in nodes:
                vals = [str(G.nodes[n].get(v, 0)) for v in vector_attrs]
                f.write(" ".join(vals) + "\n")

def save_graph(
    G,
    path,
    file_format = 'pajek',
    partition_attrs = None,
    vector_attrs = None,
):
    """
    Save graph in various formats (pajek, graphml, gexf, gml, adjlist, edgelist).
    """
    fmt = file_format.lower()
    basename, _ = os.path.splitext(path)
    if fmt == "pajek":
        save_to_pajek(G, basename, partition_attrs, vector_attrs)
    elif fmt == "graphml":
        nx.write_graphml(G, path)
    elif fmt == "gexf":
        nx.write_gexf(G, path)
    elif fmt == "gml":
        nx.write_gml(G, path)
    elif fmt == "adjlist":
        nx.write_adjlist(G, path)
    elif fmt == "edgelist":
        nx.write_edgelist(G, path)
    else:
        raise ValueError(f"Unsupported format: {file_format}")

def compute_basic_stats(
    G,
):
    """
    Compute basic statistics of the graph.

    Parameters
    ----------
    G : networkx.Graph

    Returns
    -------
    dict
        Contains num_nodes, num_edges, density, avg_degree, avg_clustering, num_connected_components, largest_cc_size, diameter, avg_shortest_path_length.
    """
    n = G.number_of_nodes()
    m = G.number_of_edges()
    density = nx.density(G)
    degrees = [d for _, d in G.degree()]
    avg_degree = sum(degrees) / n if n else 0
    avg_clustering = nx.average_clustering(G)
    cc = list(nx.connected_components(G))
    num_cc = len(cc)
    largest_cc = max(cc, key=len) if cc else set()
    sub = G.subgraph(largest_cc)
    try:
        diameter = nx.diameter(sub)
        avg_sp = nx.average_shortest_path_length(sub)
    except (nx.NetworkXError, nx.NetworkXNoPath):
        diameter = None
        avg_sp = None
    return {
        "num_nodes": n,
        "num_edges": m,
        "density": density,
        "avg_degree": avg_degree,
        "avg_clustering": avg_clustering,
        "num_connected_components": num_cc,
        "largest_cc_size": len(largest_cc),
        "diameter": diameter,
        "avg_shortest_path_length": avg_sp
    }

def compute_centralities(
    G,
):
    """
    Compute centrality measures for the graph.

    Parameters
    ----------
    G : networkx.Graph

    Returns
    -------
    dict
        degree, closeness, betweenness, eigenvector centralities.
    """
    deg = nx.degree_centrality(G)
    clo = nx.closeness_centrality(G)
    bet = nx.betweenness_centrality(G)
    try:
        eig = nx.eigenvector_centrality(G, max_iter=1000)
    except nx.NetworkXError:
        eig = {}
    return {
        "degree_centrality": deg,
        "closeness_centrality": clo,
        "betweenness_centrality": bet,
        "eigenvector_centrality": eig
    }

def compute_cluster_metrics(
    G,
    partition_attr,
):
    """
    Compute density and average degree centrality for each cluster.

    Parameters
    ----------
    G : networkx.Graph
    partition_attr : str
        Node attribute name (with or without "partition_" prefix).

    Returns
    -------
    dict
        Mapping cluster ID to dict with keys:
        - "density": internal edge density of the cluster subgraph
        - "avg_degree_centrality": average degree centrality of nodes in the cluster
        - "size": number of nodes in the cluster
    """
    key = partition_attr if partition_attr.startswith("partition_") else f"partition_{partition_attr}"
    cent = nx.degree_centrality(G)
    clusters = {}
    for n, attrs in G.nodes(data=True):
        cid = attrs.get(key)
        clusters.setdefault(cid, []).append(n)
    metrics = {}
    for cid, nodes in clusters.items():
        sub = G.subgraph(nodes)
        density = nx.density(sub)
        avg_cent = sum(cent.get(n, 0) for n in nodes) / len(nodes) if nodes else 0
        metrics[cid] = {
            "density": density,
            "avg_degree_centrality": avg_cent,
            "size": len(nodes)
        }
    return metrics

def export_cluster_dataframe(
    G,
    partition_attr,
):
    """
    Export clusters to a pandas DataFrame.

    Parameters
    ----------
    G : networkx.Graph
    partition_attr : str
        Node attribute name (with or without "partition_" prefix).

    Returns
    -------
    pandas.DataFrame
        DataFrame with columns: cluster, items (all nodes sorted by descending degree centrality),
        density, centrality
    """
    key = partition_attr if partition_attr.startswith("partition_") else f"partition_{partition_attr}"
    metrics = compute_cluster_metrics(G, key)
    deg_c = nx.degree_centrality(G)
    rows = []
    for cid, m in metrics.items():
        nodes = [n for n, d in G.nodes(data=True) if d.get(key) == cid]
        sorted_nodes = sorted(nodes, key=lambda n: deg_c.get(n, 0), reverse=True)
        items_str = ";".join(str(n) for n in sorted_nodes)
        rows.append({
            "cluster": cid,
            "items": items_str,
            "density": m["density"],
            "centrality": m["avg_degree_centrality"]
        })
    return pd.DataFrame(rows)


"""
This part provides computation classes and functions for identifying 
"Sleeping Beauties" in academic literature - papers that remain dormant 
for years before experiencing a sudden surge in recognition.

Key Concepts:
- Sleeping Beauty: A paper with a long dormant period followed by awakening
- Prince: The paper that "awakens" the Sleeping Beauty by citing it
- Storyteller: Authors who repeatedly cite/promote sleeping beauties

Based on the Beauty Coefficient (B) methodology by van Raan (2004) and 
subsequent refinements by Ke et al. (2015).
"""


def kleinberg_burst_detection(offsets: list[float], s: float = 2.0, gamma: float = 1.0) -> list[tuple[float, float, float]]:
    """
    Implements Kleinberg's burst detection algorithm (2-state model).
    
    Parameters
    ----------
    offsets : list[float]
        A list of time points (e.g., years) when the event occurred. 
        Must be sorted. Duplicate years are allowed (handled via epsilon).
    s : float, default 2.0
        The scaling factor for the burst state (how much faster is the burst 
        rate compared to the baseline).
    gamma : float, default 1.0
        The cost of transitioning to a higher state. Higher gamma makes it 
        harder to start a burst (reduces false positives).
        
    Returns
    -------
    list[tuple[float, float, float]]
        A list of bursts, where each tuple is (start_time, end_time, weight).
    """
    if not offsets or len(offsets) < 2:
        return []

    # 1. Preprocessing: Handle simultaneous events (zero gaps)
    # We add a tiny epsilon to separate identical timestamps slightly
    offsets = np.sort(offsets)
    distinct_offsets = np.unique(offsets)
    if len(distinct_offsets) == 1:
        return [] # Cannot detect bursts if all events are at the exact same time
        
    # Total time span
    T = offsets[-1] - offsets[0]
    n = len(offsets)
    
    # Average gap (baseline rate estimate)
    g_hat = T / n if T > 0 else 1.0
    
    # 2. Define Rates for States
    # State 0 (Baseline): alpha_0 = 1 / g_hat
    # State 1 (Burst):    alpha_1 = s / g_hat
    alpha = np.array([1.0 / g_hat, s / g_hat])
    
    # 3. Calculate Gaps
    # gaps[i] is the time between event i and i+1
    gaps = np.diff(offsets)
    # Ensure no exact zeros for log calculation (add tiny noise if needed, or handle mathematically)
    gaps = np.maximum(gaps, 1e-5) 
    
    # 4. Viterbi Algorithm (Dynamic Programming)
    k = 2 # Number of states
    m = len(gaps) # Number of transitions
    
    # Cost matrix: C[t][state] is min cost to reach state at time t
    C = np.full((m + 1, k), np.inf)
    C[0, 0] = 0 # Start in baseline state
    
    # Backpointers to reconstruct path
    prev = np.zeros((m + 1, k), dtype=int)
    
    # Transition cost function
    # tau(i, j) cost to move from state i to j
    def tau(i, j):
        if i >= j: return 0
        return (j - i) * gamma * np.log(n)
    
    for t in range(m):
        x = gaps[t]
        for j in range(k): # Current state
            # Emission cost: -ln( f_j(x) ) = -ln( alpha_j * exp(-alpha_j * x) )
            # = -ln(alpha_j) + alpha_j * x
            emission_cost = -np.log(alpha[j]) + alpha[j] * x
            
            # Find best previous state
            for i in range(k): # Previous state
                cost = C[t, i] + tau(i, j) + emission_cost
                if cost < C[t+1, j]:
                    C[t+1, j] = cost
                    prev[t+1, j] = i
                    
    # 5. Backtrack to find state sequence
    states = np.zeros(m + 1, dtype=int)
    states[m] = np.argmin(C[m])
    
    for t in range(m, 0, -1):
        states[t-1] = prev[t, states[t]]
        
    # 6. Extract Bursts (periods where state == 1)
    bursts = []
    in_burst = False
    start_t = 0
    
    # Calculate weight (measure of burst intensity)
    # Weight is roughly the difference in cost between being in burst vs baseline
    
    current_burst_weight = 0.0
    
    for t in range(m):
        if states[t+1] == 1:
            if not in_burst:
                in_burst = True
                start_t = offsets[t]
                current_burst_weight = 0.0
            
            # Add weight: (Base Cost - Burst Cost)
            # This is simplified; rigorous weight calc involves full cost diff
            r0 = -np.log(alpha[0]) + alpha[0] * gaps[t]
            r1 = -np.log(alpha[1]) + alpha[1] * gaps[t]
            current_burst_weight += (r0 - r1)
            
        else:
            if in_burst:
                in_burst = False
                end_t = offsets[t]
                bursts.append({
                    "start": start_t,
                    "end": end_t,
                    "weight": current_burst_weight
                })
                
    # Handle case where burst ends at the very end
    if in_burst:
         bursts.append({
            "start": start_t,
            "end": offsets[-1],
            "weight": current_burst_weight
        })
         
    return bursts


from dataclasses import dataclass

"""
Sleeping Beauty Detection for OpenAlex Data
Fixed to work with pipe-separated citation data in 'Citations by Year' column
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Tuple, Optional
from dataclasses import dataclass


@dataclass
class SleepingBeautyResult:
    """Results for a single paper's Sleeping Beauty analysis"""
    paper_id: str
    title: str
    publication_year: int
    total_citations: int
    beauty_coefficient: float
    awakening_year: Optional[int]
    sleep_duration: Optional[int]
    max_citation_year: int
    max_citations_in_year: int
    citation_history: Dict[int, int]
    awakening_intensity: float


def parse_citation_history(years_str: str, citations_str: str) -> Dict[int, int]:
    """
    Parse pipe-separated citation history into a dictionary.
    
    Args:
        years_str: Pipe-separated years (e.g., "2025|2024|2023")
        citations_str: Pipe-separated citations (e.g., "2042|2916|2244")
    
    Returns:
        Dictionary mapping year -> citation count
    """
    if pd.isna(years_str) or pd.isna(citations_str):
        return {}
    
    try:
        years = [int(y.strip()) for y in str(years_str).split('|')]
        citations = [int(c.strip()) for c in str(citations_str).split('|')]
        
        if len(years) != len(citations):
            return {}
        
        return dict(zip(years, citations))
    except (ValueError, AttributeError):
        return {}


def calculate_beauty_coefficient(
    citation_history: Dict[int, int],
    pub_year: int,
    current_year: int
) -> Tuple[float, Optional[int], Optional[int]]:
    """
    Calculate the Beauty Coefficient (B) for a paper.
    
    The Beauty Coefficient measures how much a paper's citation pattern
    deviates from linear growth, indicating delayed recognition.
    
    Method:
    1. Fit a linear model to cumulative citations over time
    2. Find the maximum deviation of actual citations from the linear trend
    3. B = max_deviation / sqrt(total_citations)
    4. Identify awakening year as the year of maximum acceleration
    
    Args:
        citation_history: Year -> citation count mapping
        pub_year: Publication year
        current_year: Current year for analysis
    
    Returns:
        Tuple of (beauty_coefficient, awakening_year, sleep_duration)
    """
    if not citation_history or len(citation_history) < 3:
        return 0.0, None, None
    
    # Sort by year
    sorted_years = sorted(citation_history.keys())
    sorted_citations = [citation_history[y] for y in sorted_years]
    
    # Calculate cumulative citations
    cumulative = np.cumsum(sorted_citations)
    
    # Normalize time (0 to 1)
    time_normalized = np.array([(y - pub_year) / (current_year - pub_year) 
                                 for y in sorted_years])
    
    # Fit linear model to cumulative citations
    if len(time_normalized) < 2:
        return 0.0, None, None
    
    # Linear regression: cumulative = a * time + b
    A = np.vstack([time_normalized, np.ones(len(time_normalized))]).T
    try:
        m, c = np.linalg.lstsq(A, cumulative, rcond=None)[0]
    except:
        return 0.0, None, None
    
    # Calculate deviations from linear trend
    linear_trend = m * time_normalized + c
    deviations = cumulative - linear_trend
    
    # Find maximum positive deviation (awakening point)
    max_deviation = np.max(deviations)
    awakening_idx = np.argmax(deviations)
    
    if max_deviation <= 0:
        return 0.0, None, None
    
    awakening_year = sorted_years[awakening_idx]
    
    # Calculate sleep duration (years from publication to awakening)
    sleep_duration = awakening_year - pub_year
    
    # Beauty coefficient: normalized by square root of total citations
    # to account for papers with different citation scales
    total_citations = cumulative[-1]
    beauty_coefficient = max_deviation / np.sqrt(total_citations) if total_citations > 0 else 0
    
    return beauty_coefficient, awakening_year, sleep_duration


def calculate_awakening_intensity(
    citation_history: Dict[int, int],
    awakening_year: Optional[int]
) -> float:
    """
    Calculate the awakening intensity ratio.
    
    Intensity = (citations during awakening period) / (citations during sleep period)
    
    Args:
        citation_history: Year -> citation count mapping
        awakening_year: Year when awakening began
    
    Returns:
        Awakening intensity ratio (higher = more dramatic awakening)
    """
    if not awakening_year or not citation_history:
        return 0.0
    
    sorted_years = sorted(citation_history.keys())
    
    # Split into sleep and awakening periods
    sleep_citations = sum(citation_history[y] for y in sorted_years if y < awakening_year)
    awakening_citations = sum(citation_history[y] for y in sorted_years if y >= awakening_year)
    
    if sleep_citations == 0:
        return float('inf') if awakening_citations > 0 else 0.0
    
    return awakening_citations / sleep_citations


def extract_sleeping_beauties(
    df: pd.DataFrame,
    min_beauty_coefficient: float = 50.0,
    min_sleep_years: int = 5,
    min_total_citations: int = 50,
    min_awakening_intensity: float = 2.0,
    current_year: int = 2025
) -> pd.DataFrame:
    """
    Extract Sleeping Beauties from an OpenAlex dataset.
    
    A Sleeping Beauty is a paper that:
    1. Has a high Beauty Coefficient (large deviation from linear citation growth)
    2. Remained dormant for several years before awakening
    3. Eventually achieved significant citation counts
    4. Had a dramatic awakening (sharp increase in citations)
    
    Args:
        df: OpenAlex DataFrame with citation data
        min_beauty_coefficient: Minimum B coefficient to qualify
        min_sleep_years: Minimum years of dormancy required
        min_total_citations: Minimum total citations required
        min_awakening_intensity: Minimum awakening intensity ratio
        current_year: Current year for analysis
    
    Returns:
        DataFrame with Sleeping Beauty papers and their metrics
    """
    print("Extracting Sleeping Beauties...")
    print(f"   Criteria: B >= {min_beauty_coefficient}, Sleep >= {min_sleep_years} years, "
          f"Citations >= {min_total_citations}, Intensity >= {min_awakening_intensity}")
    
    results = []
    
    for idx, row in df.iterrows():
        # Skip if insufficient citation data
        # FIXED: Use correct column name "Citations by Year"
        if pd.isna(row.get("counts_by_year.year")) or pd.isna(row.get("Citations by Year")):
            continue
        
        citation_history = parse_citation_history(
            row["counts_by_year.year"],
            row["Citations by Year"]  # FIXED: Changed from "counts_by_year.cited_by_count"
        )
        
        if not citation_history:
            continue
        
        pub_year = int(row["Year"])
        
        # FIXED: Use "Cited by" column for total citations
        total_citations = int(row["Cited by"]) if pd.notna(row.get("Cited by")) else 0
        
        # Calculate beauty coefficient
        beauty, awakening_year, sleep_duration = calculate_beauty_coefficient(
            citation_history, pub_year, current_year
        )
        
        # Calculate awakening intensity
        intensity = calculate_awakening_intensity(citation_history, awakening_year)
        
        # Find peak citation info
        max_citations = max(citation_history.values()) if citation_history else 0
        max_year = max(citation_history.keys(), key=lambda y: citation_history[y]) if citation_history else pub_year
        
        # FIXED: Use correct ID column
        paper_id = row.get("OpenAlex ID", row.get("unique-id", f"unknown_{idx}"))
        title = row.get("Title", row.get("Title 2", "Unknown"))
        
        result = SleepingBeautyResult(
            paper_id=paper_id,
            title=title,
            publication_year=pub_year,
            total_citations=total_citations,
            beauty_coefficient=beauty,
            awakening_year=awakening_year,
            sleep_duration=sleep_duration,
            max_citation_year=max_year,
            max_citations_in_year=max_citations,
            citation_history=citation_history,
            awakening_intensity=intensity
        )
        results.append(result)
    
    # Convert to DataFrame
    results_df = pd.DataFrame([vars(r) for r in results])
    
    if results_df.empty:
        print("   Found 0 Sleeping Beauties")
        return results_df
    
    # Filter for Sleeping Beauties
    mask = (
        (results_df["beauty_coefficient"] >= min_beauty_coefficient) &
        (results_df["total_citations"] >= min_total_citations)
    )
    
    # Only filter by sleep_duration if min_sleep_years > 0
    if min_sleep_years > 0:
        mask &= (
            (results_df["sleep_duration"].notna()) &
            (results_df["sleep_duration"] >= min_sleep_years)
        )
    
    if min_awakening_intensity > 0:
        mask &= (
            (results_df["awakening_intensity"].notna()) &
            (results_df["awakening_intensity"] >= min_awakening_intensity)
        )
    
    sleeping_beauties = results_df[mask].sort_values("beauty_coefficient", ascending=False)
    
    print(f"   Found {len(sleeping_beauties)} Sleeping Beauties")
    
    return sleeping_beauties


def extract_all_papers_with_metrics(
    df: pd.DataFrame,
    current_year: int = 2025
) -> pd.DataFrame:
    """
    Calculate Sleeping Beauty metrics for ALL papers in the dataset.
    
    Useful for exploration and custom filtering.
    
    Args:
        df: OpenAlex DataFrame with citation data
        current_year: Current year for analysis
    
    Returns:
        DataFrame with all papers and their SB metrics
    """
    return extract_sleeping_beauties(
        df,
        min_beauty_coefficient=0,
        min_sleep_years=0,
        min_total_citations=0,
        min_awakening_intensity=0,
        current_year=current_year
    )


def identify_potential_princes(
    df: pd.DataFrame,
    sleeping_beauty_id: str,
    awakening_year: int
) -> List[Dict[str, Any]]:
    """
    Identify potential "Princes" - papers that may have awakened a Sleeping Beauty.
    
    A Prince is typically a highly influential paper published around the 
    awakening year that cites the Sleeping Beauty, bringing it renewed attention.
    
    Note: This function provides candidates based on the dataset. Full Prince
    identification requires access to citing papers data from OpenAlex API.
    
    Args:
        df: OpenAlex DataFrame
        sleeping_beauty_id: OpenAlex ID of the Sleeping Beauty
        awakening_year: Year when the SB was awakened
    
    Returns:
        List of potential Prince candidates (papers published near awakening 
        year with high citation counts that reference the SB)
    """
    princes = []
    
    # Look for papers published around awakening year with high impact
    # that might reference the sleeping beauty
    window = 3  # years before/after awakening
    
    candidates = df[
        (df["Year"] >= awakening_year - window) &
        (df["Year"] <= awakening_year + window) &
        (df["Cited by"] > 100)  # FIXED: Use "Cited by" column
    ].copy()
    
    for idx, row in candidates.iterrows():
        # Check if this paper references the sleeping beauty
        referenced = row.get("referenced_works", "")
        if pd.notna(referenced) and sleeping_beauty_id in str(referenced):
            paper_id = row.get("OpenAlex ID", row.get("unique-id", f"unknown_{idx}"))
            title = row.get("Title", row.get("Title 2", "Unknown"))
            
            princes.append({
                "prince_id": paper_id,
                "title": title,
                "Year": row["Year"],
                "citations": row["Cited by"],
                "relationship": "cites_sleeping_beauty"
            })
    
    return sorted(princes, key=lambda x: x["citations"], reverse=True)


def identify_storytellers(
    df: pd.DataFrame,
    sleeping_beauties_df: pd.DataFrame,
    min_sb_citations: int = 2
) -> pd.DataFrame:
    """
    Identify "Storytellers" - authors who repeatedly cite Sleeping Beauties.
    
    Storytellers are researchers who recognize the value of dormant papers
    and help bring them back to attention through their citations.
    
    Args:
        df: Full OpenAlex DataFrame
        sleeping_beauties_df: DataFrame of identified Sleeping Beauties
        min_sb_citations: Minimum number of SBs an author must cite
    
    Returns:
        DataFrame of Storytellers with their citation patterns
    """
    sb_ids = set(sleeping_beauties_df["paper_id"].tolist())
    author_sb_citations = {}
    
    for idx, row in df.iterrows():
        # Get author info - FIXED: Use correct column names
        author_id = row.get("Author(s) ID", "")
        author_name = row.get("Authors 2", "")
        
        if pd.isna(author_id):
            continue
            
        # Handle multiple authors (pipe-delimited)
        author_ids = str(author_id).split('|')
        author_names = str(author_name).split('|') if pd.notna(author_name) else ["Unknown"] * len(author_ids)
        
        # Check references
        referenced = row.get("referenced_works", "")
        if pd.isna(referenced):
            continue
            
        refs = str(referenced).split("|")
        sb_refs = [r for r in refs if r in sb_ids]
        
        if sb_refs:
            paper_id = row.get("OpenAlex ID", row.get("unique-id", f"unknown_{idx}"))
            for aid, aname in zip(author_ids, author_names):
                if aid not in author_sb_citations:
                    author_sb_citations[aid] = {
                        "author_id": aid,
                        "author_name": aname,
                        "sb_papers_cited": set(),
                        "citing_papers": []
                    }
                author_sb_citations[aid]["sb_papers_cited"].update(sb_refs)
                author_sb_citations[aid]["citing_papers"].append(paper_id)
    
    # Convert to DataFrame
    storytellers = []
    for aid, data in author_sb_citations.items():
        if len(data["sb_papers_cited"]) >= min_sb_citations:
            storytellers.append({
                "author_id": data["author_id"],
                "author_name": data["author_name"],
                "num_sb_cited": len(data["sb_papers_cited"]),
                "num_citing_papers": len(data["citing_papers"]),
                "sb_papers_cited": list(data["sb_papers_cited"])
            })
    
    if not storytellers:
        return pd.DataFrame()
    
    return pd.DataFrame(storytellers).sort_values("num_sb_cited", ascending=False)



"""Meta-utilities that help organise, label and route analysis outputs such as tables, plots and reports."""
# utility functions for helping me orginaizing the outputs


def list_pngs_recursive_to_excel(
    folder: str,
    excel_name: str = 'png_files.xlsx',
) -> pd.DataFrame:
    """
    Recursively list all PNG files in `folder`, store them in a pandas DataFrame,
    and save the DataFrame to an Excel file.

    Columns:
    - subfolder : path of the subfolder relative to `folder` (empty string for files directly in `folder`)
    - filename  : file name with extension

    Parameters
    ----------
    folder : str
        Path to the root folder to search.
    excel_name : str, optional
        Name of the output Excel file (saved inside `folder`), by default "png_files.xlsx".

    Returns
    -------
    pd.DataFrame
        DataFrame with columns ['subfolder', 'filename'].
    """
    folder_path = Path(folder)

    if not folder_path.is_dir():
        raise ValueError(f"{folder!r} is not a valid directory")

    subfolders = []
    filenames = []

    # Recursively search for PNG files (case-insensitive)
    for f in folder_path.rglob("*"):
        if f.is_file() and f.suffix.lower() == ".png":
            rel = f.relative_to(folder_path)
            subfolder = str(rel.parent) if str(rel.parent) != "." else ""
            subfolders.append(subfolder)
            filenames.append(rel.name)  # includes extension

    df = pd.DataFrame({
        "subfolder": subfolders,
        "filename": filenames,
    })

    # Save to Excel in the root folder
    output_path = folder_path / excel_name
    df.to_excel(output_path, index=False)

    return df

def list_df_attributes_to_excel(
    obj,
    excel_path: str | Path = 'df_attributes.xlsx',
) -> pd.DataFrame:
    """
    Find all attributes of `obj` that are pandas DataFrames, return them in a
    DataFrame, and save the result to an Excel file.

    Columns:
    - attribute : attribute name on the object
    - n_rows    : number of rows in the DataFrame
    - n_cols    : number of columns in the DataFrame

    Parameters
    ----------
    obj : any
        The object whose attributes will be inspected.
    excel_path : str or Path, optional
        Path to the Excel file to save (default: "df_attributes.xlsx" in CWD).

    Returns
    -------
    pd.DataFrame
        DataFrame with information about DataFrame attributes.
    """
    records = []

    for name in dir(obj):
        # skip dunder/private attributes
        if name.startswith("_"):
            continue

        try:
            value = getattr(obj, name)
        except AttributeError:
            # some attributes may not be accessible
            continue

        if isinstance(value, pd.DataFrame):
            records.append(
                {
                    "attribute": name,
                    "n_rows": value.shape[0],
                    "n_cols": value.shape[1],
                }
            )

    df = pd.DataFrame(records).sort_values("attribute").reset_index(drop=True)

    excel_path = Path(excel_path)
    df.to_excel(excel_path, index=False)

    return df


# =============================================================================
# CO-OCCURRENCE MATRIX FUNCTIONS
# =============================================================================

def compute_cooccurrence_matrix(
    df: pd.DataFrame,
    column: str,
    top_n: int = 20,
    separator: str = "; ",
    items_of_interest: list = None,
    normalize: bool = False,
) -> pd.DataFrame:
    """
    Compute a co-occurrence matrix for items in a column.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    column : str
        Column containing multi-value items (e.g., "Author Keywords").
    top_n : int, default 20
        Number of top items to include (by frequency).
    separator : str, default "; "
        Separator used in multi-value fields.
    items_of_interest : list, optional
        Specific items to include. If None, uses top_n most frequent.
    normalize : bool, default False
        If True, normalize by diagonal (Jaccard-like).
        
    Returns
    -------
    pd.DataFrame
        Square co-occurrence matrix.
    """
    if column not in df.columns:
        raise ValueError(f"Column '{column}' not found in DataFrame")
    
    # Get items for each document
    docs_items = []
    for val in df[column].dropna():
        if isinstance(val, str):
            items = [x.strip() for x in val.split(separator) if x.strip()]
        elif isinstance(val, list):
            items = [str(x).strip() for x in val if x]
        else:
            items = []
        docs_items.append(items)
    
    # Count item frequencies
    from collections import Counter
    all_items = [item for doc in docs_items for item in doc]
    item_counts = Counter(all_items)
    
    # Select items
    if items_of_interest:
        selected_items = [i for i in items_of_interest if i in item_counts]
    else:
        selected_items = [item for item, _ in item_counts.most_common(top_n)]
    
    if not selected_items:
        return pd.DataFrame()
    
    # Create co-occurrence matrix
    n_items = len(selected_items)
    item_to_idx = {item: i for i, item in enumerate(selected_items)}
    cooc = np.zeros((n_items, n_items), dtype=int)
    
    for doc_items in docs_items:
        doc_selected = [i for i in doc_items if i in item_to_idx]
        for i, item1 in enumerate(doc_selected):
            idx1 = item_to_idx[item1]
            cooc[idx1, idx1] += 1  # Diagonal = frequency
            for item2 in doc_selected[i+1:]:
                idx2 = item_to_idx[item2]
                cooc[idx1, idx2] += 1
                cooc[idx2, idx1] += 1
    
    matrix = pd.DataFrame(cooc, index=selected_items, columns=selected_items)
    
    if normalize:
        # Jaccard-like normalization: cooc[i,j] / (diag[i] + diag[j] - cooc[i,j])
        diag = np.diag(cooc)
        for i in range(n_items):
            for j in range(n_items):
                if i != j:
                    denom = diag[i] + diag[j] - cooc[i, j]
                    if denom > 0:
                        matrix.iloc[i, j] = cooc[i, j] / denom
    
    return matrix


def build_links_from_matrix(
    matrix: pd.DataFrame,
    min_weight: int = 1,
    exclude_diagonal: bool = True,
) -> pd.DataFrame:
    """
    Convert a co-occurrence matrix to a links/edges DataFrame.
    
    Parameters
    ----------
    matrix : pd.DataFrame
        Square co-occurrence matrix.
    min_weight : int, default 1
        Minimum weight to include an edge.
    exclude_diagonal : bool, default True
        Exclude self-loops.
        
    Returns
    -------
    pd.DataFrame
        DataFrame with columns: source, target, weight.
    """
    links = []
    items = matrix.index.tolist()
    
    for i, item1 in enumerate(items):
        start = i + 1 if exclude_diagonal else i
        for j in range(start, len(items)):
            item2 = items[j]
            weight = matrix.iloc[i, j]
            if weight >= min_weight:
                links.append({
                    "source": item1,
                    "target": item2,
                    "weight": weight,
                })
    
    return pd.DataFrame(links)


# =============================================================================
# DATA/CODE REPOSITORY LINK EXTRACTION
# =============================================================================

# Known repository URL patterns
REPOSITORY_PATTERNS = {
    "github": {
        "pattern": r'github\.com/[\w\-\.]+/[\w\-\.]+',
        "name": "GitHub",
        "type": "code"
    },
    "gitlab": {
        "pattern": r'gitlab\.com/[\w\-\.]+/[\w\-\.]+',
        "name": "GitLab", 
        "type": "code"
    },
    "bitbucket": {
        "pattern": r'bitbucket\.org/[\w\-\.]+/[\w\-\.]+',
        "name": "Bitbucket",
        "type": "code"
    },
    "zenodo": {
        "pattern": r'zenodo\.org/record[s]?/\d+',
        "name": "Zenodo",
        "type": "data"
    },
    "figshare": {
        "pattern": r'figshare\.com/[\w/\-\.]+/\d+',
        "name": "Figshare",
        "type": "data"
    },
    "dryad": {
        "pattern": r'datadryad\.org/[\w/\-\.]+',
        "name": "Dryad",
        "type": "data"
    },
    "osf": {
        "pattern": r'osf\.io/[\w]+',
        "name": "OSF",
        "type": "data"
    },
    "dataverse": {
        "pattern": r'dataverse\.[\w\.]+/[\w/\-\?=]+',
        "name": "Dataverse",
        "type": "data"
    },
    "mendeley": {
        "pattern": r'data\.mendeley\.com/datasets/[\w]+',
        "name": "Mendeley Data",
        "type": "data"
    },
    "kaggle": {
        "pattern": r'kaggle\.com/datasets/[\w\-]+/[\w\-]+',
        "name": "Kaggle",
        "type": "data"
    },
    "huggingface": {
        "pattern": r'huggingface\.co/[\w\-]+/[\w\-]+',
        "name": "Hugging Face",
        "type": "code"
    },
    "arxiv_code": {
        "pattern": r'arxiv\.org/abs/[\d\.]+',
        "name": "arXiv",
        "type": "preprint"
    },
    "codeocean": {
        "pattern": r'codeocean\.com/capsule/[\w\-]+',
        "name": "Code Ocean",
        "type": "code"
    },
    "sourceforge": {
        "pattern": r'sourceforge\.net/projects/[\w\-]+',
        "name": "SourceForge",
        "type": "code"
    },
}


def extract_repository_links_from_text(
    text: str,
    include_types: Optional[List[str]] = None,
) -> List[Dict[str, str]]:
    """
    Extract repository/data links from text using regex patterns.
    
    Parameters
    ----------
    text : str
        Text to search (abstract, full text, data availability statement).
    include_types : list of str, optional
        Types to include: "code", "data", "preprint". If None, include all.
    
    Returns
    -------
    list of dict
        Each dict has: url, repository, type
    """
    if not text or not isinstance(text, str):
        return []
    
    found_links = []
    seen_urls = set()
    
    for repo_key, repo_info in REPOSITORY_PATTERNS.items():
        # Filter by type if specified
        if include_types and repo_info["type"] not in include_types:
            continue
        
        pattern = repo_info["pattern"]
        matches = re.findall(pattern, text, re.IGNORECASE)
        
        for match in matches:
            # Normalize URL
            url = match.lower()
            if not url.startswith("http"):
                url = "https://" + url
            
            if url not in seen_urls:
                seen_urls.add(url)
                found_links.append({
                    "url": url,
                    "repository": repo_info["name"],
                    "type": repo_info["type"],
                })
    
    return found_links


def extract_repository_links_from_df(
    df: pd.DataFrame,
    text_columns: Optional[List[str]] = None,
    include_types: Optional[List[str]] = None,
    add_columns: bool = True,
) -> pd.DataFrame:
    """
    Extract repository links from a DataFrame's text columns.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with text columns.
    text_columns : list of str, optional
        Columns to search. Default: ["Abstract", "Title", "Data Availability"].
    include_types : list of str, optional
        Types to include: "code", "data", "preprint".
    add_columns : bool, default=True
        If True, add columns to df. If False, return separate results df.
    
    Returns
    -------
    pd.DataFrame
        If add_columns=True: original df with new columns added.
        If add_columns=False: separate df with extracted links.
    """
    if text_columns is None:
        # Default columns to search
        text_columns = [
            "Abstract", "Processed Abstract", 
            "Title", "Data Availability",
            "Notes", "Full Text"
        ]
    
    # Filter to existing columns
    text_columns = [c for c in text_columns if c in df.columns]
    
    if not text_columns:
        if add_columns:
            df["Repository Links"] = ""
            df["Has Code Link"] = False
            df["Has Data Link"] = False
            return df
        else:
            return pd.DataFrame(columns=["Doc ID", "URL", "Repository", "Type"])
    
    results = []
    all_links_by_doc = []
    
    for idx, row in df.iterrows():
        # Combine text from all columns
        combined_text = " ".join(
            str(row.get(col, "")) for col in text_columns
            if pd.notna(row.get(col))
        )
        
        links = extract_repository_links_from_text(combined_text, include_types)
        
        doc_id = row.get("Doc ID", idx)
        
        # Store individual links
        for link in links:
            results.append({
                "Doc ID": doc_id,
                "URL": link["url"],
                "Repository": link["repository"],
                "Type": link["type"],
            })
        
        # Store aggregated info for this document
        all_links_by_doc.append({
            "idx": idx,
            "links": links,
            "has_code": any(l["type"] == "code" for l in links),
            "has_data": any(l["type"] == "data" for l in links),
            "link_str": "; ".join(l["url"] for l in links),
        })
    
    if add_columns:
        df = df.copy()
        df["Repository Links"] = [d["link_str"] for d in all_links_by_doc]
        df["Has Code Link"] = [d["has_code"] for d in all_links_by_doc]
        df["Has Data Link"] = [d["has_data"] for d in all_links_by_doc]
        df["N Repository Links"] = [len(d["links"]) for d in all_links_by_doc]
        return df
    else:
        return pd.DataFrame(results)


def fetch_datacite_links(
    dois: List[str],
    timeout: int = 10,
    max_requests: int = 100,
) -> Dict[str, List[Dict]]:
    """
    Fetch related datasets from DataCite API for given DOIs.
    
    Parameters
    ----------
    dois : list of str
        List of DOIs to lookup.
    timeout : int, default=10
        Request timeout in seconds.
    max_requests : int, default=100
        Maximum number of API requests to make.
    
    Returns
    -------
    dict
        Mapping from DOI to list of related datasets.
    """
    import requests
    
    results = {}
    
    # DataCite API base URL
    base_url = "https://api.datacite.org/dois"
    
    for i, doi in enumerate(dois[:max_requests]):
        if not doi or pd.isna(doi):
            continue
        
        # Clean DOI
        doi_clean = str(doi).strip()
        if doi_clean.startswith("https://doi.org/"):
            doi_clean = doi_clean.replace("https://doi.org/", "")
        elif doi_clean.startswith("http://doi.org/"):
            doi_clean = doi_clean.replace("http://doi.org/", "")
        
        try:
            # Query DataCite for this DOI
            response = requests.get(
                f"{base_url}/{doi_clean}",
                timeout=timeout,
                headers={"Accept": "application/json"}
            )
            
            if response.status_code == 200:
                data = response.json().get("data", {})
                attributes = data.get("attributes", {})
                
                # Check if this is a dataset
                resource_type = attributes.get("types", {}).get("resourceTypeGeneral", "")
                
                related = []
                
                # Get related identifiers
                for rel in attributes.get("relatedIdentifiers", []):
                    related.append({
                        "identifier": rel.get("relatedIdentifier"),
                        "type": rel.get("relatedIdentifierType"),
                        "relation": rel.get("relationType"),
                    })
                
                if related or resource_type == "Dataset":
                    results[doi] = {
                        "is_dataset": resource_type == "Dataset",
                        "title": attributes.get("titles", [{}])[0].get("title", ""),
                        "related": related,
                    }
            
        except Exception as e:
            # Silently continue on errors
            pass
    
    return results


def fetch_paperswithcode_links(
    dois: Optional[List[str]] = None,
    arxiv_ids: Optional[List[str]] = None,
    titles: Optional[List[str]] = None,
    timeout: int = 10,
    max_requests: int = 50,
) -> Dict[str, Dict]:
    """
    Fetch code repository links from Papers With Code API.
    
    Parameters
    ----------
    dois : list of str, optional
        List of DOIs to lookup.
    arxiv_ids : list of str, optional
        List of arXiv IDs to lookup.
    titles : list of str, optional
        List of paper titles to search.
    timeout : int, default=10
        Request timeout in seconds.
    max_requests : int, default=50
        Maximum number of API requests.
    
    Returns
    -------
    dict
        Mapping from identifier to code repository info.
    """
    import requests
    
    results = {}
    request_count = 0
    
    base_url = "https://paperswithcode.com/api/v1"
    
    # Search by arXiv ID (most reliable)
    if arxiv_ids:
        for arxiv_id in arxiv_ids:
            if request_count >= max_requests:
                break
            if not arxiv_id or pd.isna(arxiv_id):
                continue
            
            # Clean arXiv ID
            arxiv_clean = str(arxiv_id).strip()
            if "arxiv.org" in arxiv_clean:
                arxiv_clean = arxiv_clean.split("/")[-1]
            
            try:
                response = requests.get(
                    f"{base_url}/papers/",
                    params={"arxiv_id": arxiv_clean},
                    timeout=timeout
                )
                request_count += 1
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get("results"):
                        paper = data["results"][0]
                        paper_id = paper.get("id")
                        
                        # Get repositories for this paper
                        repo_response = requests.get(
                            f"{base_url}/papers/{paper_id}/repositories/",
                            timeout=timeout
                        )
                        request_count += 1
                        
                        if repo_response.status_code == 200:
                            repos = repo_response.json().get("results", [])
                            if repos:
                                results[arxiv_id] = {
                                    "paper_title": paper.get("title"),
                                    "repositories": [
                                        {
                                            "url": r.get("url"),
                                            "stars": r.get("stars"),
                                            "is_official": r.get("is_official"),
                                        }
                                        for r in repos
                                    ]
                                }
            except Exception:
                pass
    
    # Search by title (less reliable but works for non-arXiv papers)
    if titles and request_count < max_requests:
        for title in titles:
            if request_count >= max_requests:
                break
            if not title or pd.isna(title) or len(str(title)) < 10:
                continue
            
            try:
                response = requests.get(
                    f"{base_url}/papers/",
                    params={"q": str(title)[:100]},
                    timeout=timeout
                )
                request_count += 1
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get("results"):
                        # Take first result if title matches closely
                        paper = data["results"][0]
                        paper_title = paper.get("title", "").lower()
                        if title.lower()[:50] in paper_title or paper_title in title.lower():
                            paper_id = paper.get("id")
                            
                            repo_response = requests.get(
                                f"{base_url}/papers/{paper_id}/repositories/",
                                timeout=timeout
                            )
                            request_count += 1
                            
                            if repo_response.status_code == 200:
                                repos = repo_response.json().get("results", [])
                                if repos:
                                    results[title] = {
                                        "paper_title": paper.get("title"),
                                        "repositories": [
                                            {
                                                "url": r.get("url"),
                                                "stars": r.get("stars"),
                                                "is_official": r.get("is_official"),
                                            }
                                            for r in repos
                                        ]
                                    }
            except Exception:
                pass
    
    return results


def enrich_with_repository_links(
    df: pd.DataFrame,
    use_text_mining: bool = True,
    use_datacite: bool = False,
    use_paperswithcode: bool = False,
    doi_column: str = "DOI",
    title_column: str = "Title",
    text_columns: Optional[List[str]] = None,
    max_api_requests: int = 50,
    progress_callback: Optional[callable] = None,
) -> Tuple[pd.DataFrame, Dict]:
    """
    Enrich a DataFrame with repository links using multiple methods.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with publications.
    use_text_mining : bool, default=True
        Extract links from text columns.
    use_datacite : bool, default=False
        Query DataCite API for related datasets.
    use_paperswithcode : bool, default=False
        Query Papers With Code API for code links.
    doi_column : str, default="DOI"
        Column containing DOIs.
    title_column : str, default="Title"
        Column containing titles.
    text_columns : list of str, optional
        Columns to search for text mining.
    max_api_requests : int, default=50
        Maximum API requests per source.
    progress_callback : callable, optional
        Function to call with progress updates.
    
    Returns
    -------
    tuple of (pd.DataFrame, dict)
        Enriched dataframe and summary statistics.
    """
    df = df.copy()
    stats = {
        "total_docs": len(df),
        "docs_with_code_links": 0,
        "docs_with_data_links": 0,
        "total_links_found": 0,
        "sources_used": [],
    }
    
    # Initialize columns
    df["Repository Links"] = ""
    df["Has Code Link"] = False
    df["Has Data Link"] = False
    df["Code Repositories"] = ""
    df["Data Repositories"] = ""
    
    all_links = {idx: [] for idx in df.index}
    
    # Method 1: Text mining
    if use_text_mining:
        if progress_callback:
            progress_callback("Extracting links from text...")
        
        stats["sources_used"].append("text_mining")
        
        for idx, row in df.iterrows():
            # Combine text columns
            if text_columns is None:
                text_columns = ["Abstract", "Title", "Notes"]
            
            combined_text = " ".join(
                str(row.get(col, "")) for col in text_columns
                if col in df.columns and pd.notna(row.get(col))
            )
            
            links = extract_repository_links_from_text(combined_text)
            all_links[idx].extend(links)
    
    # Method 2: DataCite API
    if use_datacite and doi_column in df.columns:
        if progress_callback:
            progress_callback("Querying DataCite API...")
        
        stats["sources_used"].append("datacite")
        
        dois = df[doi_column].dropna().tolist()
        datacite_results = fetch_datacite_links(dois, max_requests=max_api_requests)
        
        for idx, row in df.iterrows():
            doi = row.get(doi_column)
            if doi and doi in datacite_results:
                result = datacite_results[doi]
                if result.get("is_dataset"):
                    all_links[idx].append({
                        "url": f"https://doi.org/{doi}",
                        "repository": "DataCite",
                        "type": "data",
                    })
    
    # Method 3: Papers With Code API
    if use_paperswithcode and title_column in df.columns:
        if progress_callback:
            progress_callback("Querying Papers With Code API...")
        
        stats["sources_used"].append("paperswithcode")
        
        titles = df[title_column].dropna().tolist()
        pwc_results = fetch_paperswithcode_links(titles=titles, max_requests=max_api_requests)
        
        # Create title to index mapping
        title_to_idx = {}
        for idx, row in df.iterrows():
            title = row.get(title_column)
            if title:
                title_to_idx[title] = idx
        
        for title, result in pwc_results.items():
            if title in title_to_idx:
                idx = title_to_idx[title]
                for repo in result.get("repositories", []):
                    all_links[idx].append({
                        "url": repo.get("url", ""),
                        "repository": "Papers With Code",
                        "type": "code",
                        "stars": repo.get("stars"),
                        "is_official": repo.get("is_official"),
                    })
    
    # Aggregate results
    if progress_callback:
        progress_callback("Aggregating results...")
    
    for idx in df.index:
        links = all_links[idx]
        
        # Deduplicate by URL
        seen_urls = set()
        unique_links = []
        for link in links:
            url = link.get("url", "").lower()
            if url and url not in seen_urls:
                seen_urls.add(url)
                unique_links.append(link)
        
        if unique_links:
            code_links = [l for l in unique_links if l["type"] == "code"]
            data_links = [l for l in unique_links if l["type"] == "data"]
            
            df.at[idx, "Repository Links"] = "; ".join(l["url"] for l in unique_links)
            df.at[idx, "Has Code Link"] = len(code_links) > 0
            df.at[idx, "Has Data Link"] = len(data_links) > 0
            df.at[idx, "Code Repositories"] = "; ".join(l["url"] for l in code_links)
            df.at[idx, "Data Repositories"] = "; ".join(l["url"] for l in data_links)
            df.at[idx, "N Repository Links"] = len(unique_links)
            
            stats["total_links_found"] += len(unique_links)
            if code_links:
                stats["docs_with_code_links"] += 1
            if data_links:
                stats["docs_with_data_links"] += 1
    
    return df, stats


# Import growth model functions from utilsbib_modules
try:
    from biblium.utilsbib_modules.time_series import (
        fit_growth_model,
        fit_life_cycle_model,
    )
except ImportError:
    pass

