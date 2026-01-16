# -*- coding: utf-8 -*-
"""
Excel I/O utilities - reading, writing, and formatting Excel files.

This module contains:
- to_excel_fancy: Write DataFrames with conditional formatting
- save_descriptives_to_excel: Save multiple descriptive tables
- Excel reading helpers
"""

from __future__ import annotations

import os
from typing import Dict, List, Optional, Union, Any
import pandas as pd
import numpy as np
from scipy.stats import rankdata
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter


# Modern color constants
MODERN_HEADER_BG = "2C3E50"      # Dark blue-gray
MODERN_HEADER_FG = "FFFFFF"      # White
MODERN_ALT_ROW = "F8F9FA"        # Light gray
MODERN_ACCENT = "3498DB"         # Blue
MODERN_SUCCESS = "27AE60"        # Green
MODERN_DANGER = "E74C3C"         # Red
MODERN_BORDER = "DEE2E6"         # Light border


def to_excel_fancy(
    data: Union[pd.DataFrame, List[pd.DataFrame]],
    f_name: str = "styled_output.xlsx",
    sheet_names: Optional[List[str]] = None,
    top_n: int = 3,
    bottom_n: int = 3,
    top_color: str = None,
    bottom_color: str = None,
    autofit: bool = True,
    conditional_formatting: bool = True,
    # New modern styling options
    modern_style: bool = True,
    zebra_rows: bool = True,
    data_bars: bool = False,
    freeze_header: bool = True,
) -> None:
    """
    Save one or multiple DataFrames to an Excel file with optional formatting.
    
    Automatically handles MultiIndex columns by enabling the index export.
    
    Parameters
    ----------
    data : DataFrame or list of DataFrames
        Data to save.
    f_name : str
        Output file path.
    sheet_names : list of str, optional
        Sheet names for each DataFrame.
    top_n : int
        Number of top values to highlight in green.
    bottom_n : int
        Number of bottom values to highlight in red.
    top_color : str, optional
        Hex color for top values. Defaults to modern green.
    bottom_color : str, optional
        Hex color for bottom values. Defaults to modern red.
    autofit : bool
        If True, auto-adjust column widths.
    conditional_formatting : bool
        If True, apply color formatting to numeric columns.
    modern_style : bool
        If True, apply modern table styling (colored headers, zebra rows).
    zebra_rows : bool
        If True, apply alternating row colors.
    data_bars : bool
        If True, add data bars to numeric columns.
    freeze_header : bool
        If True, freeze the header row.
    """
    # Set default colors based on style
    if top_color is None:
        top_color = MODERN_SUCCESS if modern_style else "99FF99"
    if bottom_color is None:
        bottom_color = MODERN_DANGER if modern_style else "FF9999"
    
    # Ensure data is a list of DataFrames
    if isinstance(data, pd.DataFrame):
        data = [data]

    # Default sheet names if none provided
    if sheet_names is None or not sheet_names:
        sheet_names = [f"Sheet{i+1}" for i in range(len(data))]

    # Ensure correct sheet name length
    if len(sheet_names) != len(data):
        raise ValueError("Number of sheet names must match number of DataFrames.")

    # Define fill styles
    top_fill = PatternFill(start_color=top_color, end_color=top_color, fill_type="solid")
    bottom_fill = PatternFill(start_color=bottom_color, end_color=bottom_color, fill_type="solid")
    
    # Modern styles
    header_fill = PatternFill(start_color=MODERN_HEADER_BG, end_color=MODERN_HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=MODERN_HEADER_FG)
    alt_row_fill = PatternFill(start_color=MODERN_ALT_ROW, end_color=MODERN_ALT_ROW, fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(bottom=Side(style="thin", color=MODERN_BORDER))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    with pd.ExcelWriter(f_name, engine="openpyxl") as writer:
        for df, sheet_name in zip(data, sheet_names):
            
            # Detect MultiIndex columns
            is_multiindex = isinstance(df.columns, pd.MultiIndex)
            
            # If MultiIndex, we MUST save the index to allow hierarchical headers
            save_index = is_multiindex
            
            df.to_excel(writer, sheet_name=sheet_name, index=save_index)
            sheet = writer.sheets[sheet_name]
            
            # Get dimensions
            header_rows = df.columns.nlevels if is_multiindex else 1
            col_offset = df.index.nlevels if save_index else 0
            n_data_rows = len(df)
            n_cols = len(df.columns) + col_offset
            
            # Freeze header row
            if freeze_header:
                sheet.freeze_panes = sheet.cell(row=header_rows + 1, column=1)
            
            # Apply modern header styling
            if modern_style:
                for row_idx in range(1, header_rows + 1):
                    for col_idx in range(1, n_cols + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
            
            # Apply zebra rows and data styling
            if modern_style or zebra_rows:
                for row_idx in range(header_rows + 1, header_rows + n_data_rows + 1):
                    is_alt = (row_idx - header_rows) % 2 == 0
                    for col_idx in range(1, n_cols + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if modern_style:
                            cell.font = data_font
                            cell.border = thin_border
                            cell.alignment = left_align
                        if zebra_rows and is_alt:
                            # Don't overwrite top/bottom formatting
                            if cell.fill.start_color.rgb in (None, "00000000", "FFFFFFFF"):
                                cell.fill = alt_row_fill

            # Autofit column width
            if autofit:
                for col in sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                if len(val_str) > 50:
                                    val_str = val_str[:50]
                                max_length = max(max_length, len(val_str))
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width

            # Add data bars to numeric columns
            if data_bars and n_data_rows > 1:
                for i, col_name in enumerate(df.columns):
                    if pd.api.types.is_numeric_dtype(df[col_name]):
                        col_idx = i + 1 + col_offset
                        col_letter = get_column_letter(col_idx)
                        
                        data_bar_rule = DataBarRule(
                            start_type="min",
                            end_type="max",
                            color=MODERN_ACCENT,
                            showValue=True,
                            minLength=None,
                            maxLength=None
                        )
                        
                        range_str = f"{col_letter}{header_rows + 1}:{col_letter}{header_rows + n_data_rows}"
                        try:
                            sheet.conditional_formatting.add(range_str, data_bar_rule)
                        except Exception:
                            pass

            # Conditional Formatting (top/bottom highlighting)
            if conditional_formatting:
                for i, col_name in enumerate(df.columns):
                    excel_col_idx = (i + 1) + col_offset

                    if pd.api.types.is_numeric_dtype(df[col_name]):
                        start_row = header_rows + 1

                        values = [
                            cell for row in sheet.iter_rows(
                                min_col=excel_col_idx,
                                max_col=excel_col_idx,
                                min_row=start_row
                            )
                            for cell in row
                        ]

                        if values:
                            numeric_values = [
                                (idx, float(cell.value))
                                for idx, cell in enumerate(values)
                                if isinstance(cell.value, (int, float))
                            ]
                            
                            if not numeric_values:
                                continue
                                
                            val_list = [val for _, val in numeric_values]
                            ranks_min = rankdata(val_list, method="min")
                            ranks_max = rankdata(val_list, method="max")

                            min_rank_threshold = bottom_n
                            max_rank_threshold = len(ranks_max) - top_n + 1

                            for (idx, val), rank_min, rank_max in zip(numeric_values, ranks_min, ranks_max):
                                if rank_min <= min_rank_threshold:
                                    values[idx].fill = bottom_fill
                                if rank_max >= max_rank_threshold:
                                    values[idx].fill = top_fill

    print(f"Saved to {f_name}")


def save_descriptives_to_excel(
    descriptives_list: List[tuple],
    f_name: str = "descriptives.xlsx",
    autofit: bool = True,
    modern_style: bool = True,
) -> None:
    """
    Save multiple descriptive tables to an Excel file.
    
    Parameters
    ----------
    descriptives_list : list of tuples
        Each tuple is (DataFrame, sheet_name).
    f_name : str
        Output file path.
    autofit : bool
        If True, auto-adjust column widths.
    modern_style : bool
        If True, apply modern styling (colored headers, zebra rows).
    """
    # Modern styles
    header_fill = PatternFill(start_color=MODERN_HEADER_BG, end_color=MODERN_HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=MODERN_HEADER_FG)
    alt_row_fill = PatternFill(start_color=MODERN_ALT_ROW, end_color=MODERN_ALT_ROW, fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(bottom=Side(style="thin", color=MODERN_BORDER))
    center_align = Alignment(horizontal="center", vertical="center")
    
    with pd.ExcelWriter(f_name, engine="openpyxl") as writer:
        for df, sheet_name in descriptives_list:
            if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                continue
            
            # Truncate sheet name to Excel limit
            safe_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
            
            df.to_excel(writer, sheet_name=safe_name, index=True)
            sheet = writer.sheets[safe_name]
            
            n_header_rows = 1
            n_data_rows = len(df)
            n_cols = len(df.columns) + df.index.nlevels
            
            # Apply modern styling
            if modern_style:
                # Style header row
                for col_idx in range(1, n_cols + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # Style data rows with zebra striping
                for row_idx in range(2, n_data_rows + 2):
                    is_alt = (row_idx - 1) % 2 == 0
                    for col_idx in range(1, n_cols + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        if is_alt:
                            cell.fill = alt_row_fill
                
                # Freeze header
                sheet.freeze_panes = sheet.cell(row=2, column=1)
            
            if autofit:
                for col in sheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                if len(val_str) > 50:
                                    val_str = val_str[:50]
                                max_length = max(max_length, len(val_str))
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width


def save_plot(
    path: str,
    dpi: int = 600,
    formats: List[str] = None,
    tight: bool = True,
) -> None:
    """
    Save the current matplotlib figure to file(s).
    
    Parameters
    ----------
    path : str
        Base path (without extension).
    dpi : int
        Resolution for raster formats.
    formats : list of str, optional
        File formats to save. Defaults to ["png", "svg", "pdf"].
    tight : bool
        If True, use tight bounding box.
    """
    import matplotlib.pyplot as plt
    
    if formats is None:
        formats = ["png", "svg", "pdf"]
    
    bbox = "tight" if tight else None
    
    for fmt in formats:
        full_path = f"{path}.{fmt}"
        plt.savefig(full_path, dpi=dpi, bbox_inches=bbox, format=fmt)
