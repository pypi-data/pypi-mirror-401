"""
This module contains classes to manipulate file system elements, 
like files, zipfiles and directories.
"""
from __future__ import annotations
from typing import Union

import os
import re
import copy
import shutil
import datetime
import pandas
import json
import zipfile
import geopandas

import openpyxl
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from dbfread import DBF
import dbf
import pickle
import dill

from io import BytesIO
from aiosmb.commons.interfaces.directory import SMBDirectory
from aiosmb.commons.interfaces.file import SMBFile

from smb.SMBConnection import SMBConnection
from smb.base import SharedFile

from datablender.base import getDirectoryElementName
from datablender.base.request import AsyncRequest
from datablender.base.fileServer import FileServer
    
class File:
    """Represents a single file.
    
    It contains informations like size, modification time, extension, Excel sheets.
    The file can be created, deleted, moved or renamed.
    The file can also be read. If it contains data, it return a dataframe. If it's a shapefile, the geometry column is a geojson.

    Attributes:
    ----------
        name (str): File name
        directory_name (str): Directory name
        modification_time (datetime): 
        size (int):
        sheets
        file_extension
      
    Methods:
    ----------
        touch (self) -> None: 
        rename (self) -> None: 
        move (self) -> None: 
        delete (self) -> None: 
        read (self) -> None: 
    """
    def __init__(
        self,
        directory_name:str=None,
        file_name:str=None,
        path:str=None,
        content=None,
        file_server:SMBConnection = None,
        is_temporary:bool = False
    ):
        """File class.

        Args:
            directory_name (str): Directory name
            file_name (str): File name
            path (str, optional): Path to file. Defaults to None.
            content (_type_, optional): File content. Defaults to None.
        """
        self.directory_name,self.name = getDirectoryElementName(
            path,
            directory_name,
            file_name
        )
        self.content = content
        self.file_server = file_server
        self.is_temporary = is_temporary

        self.attributes:SharedFile = (
            self.file_server.getAttributes('data',self.path)
            if self.exists and self.file_server and not self.is_temporary else None
        )

    def __str__(self):
        return 'File {}/{}'.format(
            self.directory_name,
            self.name
        )

    @property
    def path(self) -> str:
        """Set path of the file
        """
        return os.path.join(
            self.directory_name,
            self.name
        )

    @property
    def exists(self):
        """Set path of the file
        """
        
        if self.file_server and not self.is_temporary:
            return self.name in [
                getattr(f,'filename') for f in self.file_server.listPath('data', self.directory_name)
            ]
            
        return os.path.isfile(self.path)
    
    @property
    def extension(self):
        """Get the file extension
        """
        return os.path.splitext(self.path)[-1].lower()[1:]
    
    @property
    def name_(self):
        """Get the name without the extension
        """
        return os.path.splitext(self.name)[0].lower()

    @property
    def size(self):
        """Get the file size
        """
        return self.attributes.file_size if self.attributes else os.stat(self.path).st_size
    @property
    def modification_time(self):
        """Get the modification date of the file
        """
        return datetime.datetime.fromtimestamp(
            self.attributes.create_time if self.attributes else os.path.getmtime(self.path)
        )

    def touch(self):
        """Create file
        """
        if not self.exists:
            self.write()
        return self

    def rename(
        self,
        new_file_name:str
    ):
        """Rename file
        Args:
            new_file_name (str): New file name
        """
        old_path = copy.deepcopy(self.path)
        
        self.name = new_file_name + (
            '.'+self.extension
        ) if not '.' in new_file_name else None
        
        if self.file_server and not self.is_temporary:
            self.file_server.rename('data',old_path, self.path)
        else:
            os.rename(old_path, self.path)

    def move(
        self,
        new_directory_name:str
    ) -> None:
        """Change the file location
        Args:
            new_directory_name (str): New directory name
        """
        old_path = copy.deepcopy(self.path)
        self.directory_name = new_directory_name
        shutil.move(self.path, old_path)

    def delete(self):
        """Delete file
        """
        os.unlink(self.path)

    def read(
        self,
        is_data:bool=True,
        *args,
        **kwargs
    ) -> File:
        """Read file.
        Returns:
            File: self.
        """
        
        if self.extension in ['xlsx','xls','xlsm']:
            if 'sheet_name' in kwargs:
                self.content = pandas.read_excel(self.path,*args,**kwargs)
            elif 'sheets' in kwargs:
                sheets = kwargs.pop('sheets')

                self.content = pandas.concat([
                    self.readExcel(
                        *args,
                        **{
                            **kwargs,
                            **({'sheet_name':sheet} if isinstance(sheet,str) else sheet)
                        }
                    ) for sheet in sheets
                ])
            
            else:
                excel_content = openpyxl.load_workbook(
                    self.path,
                    data_only=True,
                    read_only=True,
                    keep_vba=False,
                    keep_links=False
                )
                self.content = pandas.read_excel(
                    self.path,
                    sheet_name=excel_content.sheetnames[0],
                    *args,
                    **kwargs
                )

        elif self.extension in ['csv','zip','txt'] and is_data:
            if self.file_server and not self.is_temporary:
                s_buf = BytesIO()
                self.file_server.retrieveFile('data',self.path,s_buf)
                s_buf.seek(0)
            
            self.content = pandas.read_csv(
                s_buf if self.file_server and not self.is_temporary else self.path,
                *args,
                **kwargs
            )

        elif self.extension =='dbf':

            table = DBF(
                self.path,
                ignore_missing_memofile=True,
                encoding = kwargs.get('encoding')
            )

            if 'nrows' in kwargs and kwargs.get('nrows') == 0:
                self.content = pandas.DataFrame(columns = table.field_names)
            else:
                self.content = pandas.DataFrame(iter(table))

        elif self.extension =='shp':
            rows = kwargs.pop('nrows',None)
            self.content = geopandas.read_file(
                self.path,
                rows = rows,
                *args,
                **kwargs
            )

        elif self.extension in ['xml','kml']:

            if is_data:
                self.content = pandas.read_xml(self.path,*args,**kwargs)

            else:
                markup_reader = kwargs.pop('markup_reader')

                with open(self.path, "r",*args,**kwargs) as file_content:

                    if markup_reader == 'soup':
                        raw_content = file_content.read()
                        self.content = BeautifulSoup(raw_content, self.extension)

                    else:
                        self.content = ET.XML(file_content)

        elif self.extension == 'sav':
            self.cont = pandas.read_spss(self.path)
        
        elif self.extension == 'pkl':
            
            with open(self.path, 'rb') as input:
                self.content = pickle.load(input)

        else:
            kwargs.pop('nrows',None)
            with open(self.path, "r",*args,**kwargs) as file_content:
  
                if self.extension  == 'sql':
                    self.content = file_content.read()

                elif self.extension == 'json':
                    self.content = json.load(file_content)

        return self

    def readExcel(self,*args,**kwargs) -> pandas.DataFrame:
        
        frame = pandas.read_excel(
            self.path,*args,**kwargs
        )
        if all([column[-2:] == '.1' for column in frame.columns]):
            frame.columns = [column[:-2] for column in frame.columns]

        return frame

    def write(self,content:Union[pandas.DataFrame]=None,*args,**kwargs):

        content = content if content is not None else self.content

        if self.extension == 'csv':
            if self.file_server and not self.is_temporary:
                s_buf = BytesIO()
                content.to_csv(
                    s_buf,
                    index=False
                )
                s_buf.seek(0)
                self.file_server.storeFile('data',self.path,s_buf)
            
            else:
                content.to_csv(
                    self.path,
                    index=False
                )

        elif self.extension == 'dbf':
            
            data = content.reset_index(drop=True)

            struct = {}

            # Create structure from dataframe dtypes
            for tup in list(zip(data.dtypes.index, data.dtypes)):

                dataBis = data[tup[0]][pandas.notna(data[tup[0]])]

                if tup[1].kind == 'i':
                    precision = 0
                    total = max(len(str(max(list(dataBis)))) + 1, precision + 2)
                    struct[tup[0]] = 'N (' + str(total) + ', ' + str(precision) + ')'

                elif tup[1].kind == 'f':
                
                    precision = max(dataBis.astype(str).apply(lambda x: len(x.split('.')[1])))
                    total = max(max(dataBis.astype(str).apply(lambda x: len(x.split('.')[0]))) + precision + 1, precision + 2)

                    struct[tup[0]] = 'N (' + str(total) + ', ' + str(precision) + ')'

                elif tup[1].kind == 'M':
                
                    struct[tup[0]] = 'D'

                else:
                    length = max(dataBis.astype(str).apply(lambda x: len(x)))
                    data[tup[0]] = data[tup[0]].astype(str)
                    struct[tup[0]] = 'C (' + str(length) + ')'

                    if length > 255 or tup[0] == 'GEOMET' or tup[0] == 'COMMENT':
                        struct[tup[0]] = 'M'

            cols=[]

            #verify that col names are legit (they should be)
            for col in data.columns:
                #name cannot contain special characters
                name = re.sub(re.compile("[^A-Za-z0-9]"), '_', col)
                #name cannot start with a number
                name = '_'+name if name[0].isdigit() else name
                #name is restricted to 10 digits
                name = name[:10]

                if name != col:
                    cols.append(name)
                    struct[name] = struct.pop(col)
                else:
                    cols.append(col)

            data.columns = cols

            #replace all NaN with None so the dbf librairy can work
            data = data.astype(object).where(data.notnull(), None)

            #drop columns that are not in struct
            data = data.drop(columns=[c for c in data.columns if c not in struct.keys()])

            #and now build the struc as expected by the dbf librairy
            struct = [f"{col} {struct[col]}" for col in data.columns]

            #create and fill the table
            with dbf.Table(
                self.path,
                on_disk=True,
                dbf_type='db3',
                codepage='cp1252',
                memo_size=512,
                field_specs=struct
            ) as table:

                tabled_dic = data.to_dict(orient='index')

                for key in tabled_dic.keys():
                    table.append(tabled_dic[key])

        elif self.extension  in['geojson','json']:
            file = open(self.path, "w",*args,**kwargs)
            json.dump(content, file, indent=4)
            file.close()
        
        elif self.extension in ['sql','py']:
            file = open(self.path, "w",*args,**kwargs)
            file.write(content)
            file.close()
            
        elif self.extension == 'pkl':
            with open(self.path, 'wb') as output:
                dill.dump(content, output)
            
        else:
            if self.file_server and not self.is_temporary:
                s_buf = BytesIO(content)
                s_buf.seek(0)
                self.file_server.storeFile('data',self.path,s_buf)

            else:
                file = open(self.path, "wb",*args,**kwargs)
                file.write(content)
                file.close()

class ZipFile_(zipfile.ZipFile):
    """Represents a zip file

    Attributes:
    ----------
        Name (str): Name of the zip file

    Methods:
    ----------
        Methods
    """
    def __init__(
        self,
        name:str = None,
        directory_name:str = None,
        path:str = None,
        file_server:SMBConnection = None,
        is_temporary:bool = False
    ):
        """Initiate the zip file

        Args:
            name (str): Name of the zip file
            directory_name (str): Name of the directory
            path (str, optional): _description_. Defaults to None.
        """
        
        self.directory_name,self.name = getDirectoryElementName(
            path,
            directory_name,
            name
        )
        
        self.file_server = file_server
        self.is_temporary = is_temporary

        self.attributes:SharedFile = (
            self.file_server.getAttributes('data',self.path)
            if self.file_server and not self.is_temporary else None
        )

        if self.file_server:
            s_buf = BytesIO()
            self.file_server.retrieveFile('data',self.path,s_buf)
            
        super(ZipFile_,self).__init__(
            s_buf if self.file_server else self.path
        )

    @property
    def path(self):
        """Set path of the zip file
        """
        return os.path.join(
            self.directory_name,
            self.name
        )
    
    @property
    def name_(self):
        """Get the name without the extension
        """
        return os.path.splitext(self.name)[0].lower()

    @property
    def size(self):
        """Get the file size
        """
        return self.attributes.file_size if self.attributes else os.stat(self.path).st_size

class AsyncFile:
    """Represents a single file.
    
    It contains informations like size, modification time, extension, Excel sheets.
    The file can be created, deleted, moved or renamed.
    The file can also be read. If it contains data, it return a dataframe. If it's a shapefile, the geometry column is a geojson.

    Attributes:
    ----------
        name (str): File name
        directory_name (str): Directory name
        modification_time (datetime): 
        size (int):
        sheets
        file_extension
      
    Methods:
    ----------
        touch (self) -> None: 
        rename (self) -> None: 
        move (self) -> None: 
        delete (self) -> None: 
        read (self) -> None: 
    """
    def __init__(
        self,
        directory_name:str=None,
        file_name:str=None,
        path:str=None,
        content=None,
        file_server:FileServer = None,
        is_temporary:bool = False,
        share_name:str = 'Data'
    ):
        """File class.

        Args:
            directory_name (str): Directory name
            file_name (str): File name
            path (str, optional): Path to file. Defaults to None.
            content (_type_, optional): File content. Defaults to None.
        """
        self.directory_name,self.name = getDirectoryElementName(
            path,
            directory_name,
            file_name
        )
        self.content = content
        self.file_server = file_server
        self.share_name = share_name
        self.is_temporary = is_temporary
        self.parent_directory:SMBDirectory = None
        self.file_object:SMBFile = None

    async def initiate(self) -> AsyncFile:
        if self.file_server:
            if self.parent_directory is None:
                self.parent_directory = await self.file_server.getParentDirectory(
                    self.share_name,
                    self.path
                )
            if self.exists:
                self.file_object = self.parent_directory.files[self.name]

        return self

    def __str__(self):
        return 'File {}/{}'.format(
            self.directory_name,
            self.name
        )

    @property
    def path(self) -> str:
        """Set path of the file
        """
        return os.path.join(
            self.directory_name,
            self.name
        )

    @property
    def exists(self):
        """Set path of the file
        """
    
        if self.file_server:
            return self.name in self.parent_directory.files.keys()
        
        return os.path.isfile(self.path)
    
    @property
    def extension(self):
        """Get the file extension
        """
        return os.path.splitext(self.path)[-1].lower()[1:]
    
    @property
    def name_(self):
        """Get the name without the extension
        """
        return os.path.splitext(self.name)[0].lower()

    @property
    def size(self):
        """Get the file size
        """
        return self.file_object.size if self.file_object else os.stat(self.path).st_size
    
    @property
    def modification_time(self):
        """Get the modification date of the file
        """
        return  self.file_object.last_write_time if self.file_object else datetime.datetime.fromtimestamp(
            os.path.getmtime(self.path)
        )

    def touch(self):
        """Create file
        """
        if not self.exists:
            self.write()
        return self

    def rename(
        self,
        new_file_name:str
    ):
        """Rename file
        Args:
            new_file_name (str): New file name
        """
        old_path = copy.deepcopy(self.path)
        
        self.name = new_file_name + (
            '.'+self.extension
        ) if not '.' in new_file_name else None
        
        if self.file_server and not self.is_temporary:
            self.file_server.rename('data',old_path, self.path)
        else:
            os.rename(old_path, self.path)

    def move(
        self,
        new_directory_name:str
    ) -> None:
        """Change the file location
        Args:
            new_directory_name (str): New directory name
        """
        old_path = copy.deepcopy(self.path)
        self.directory_name = new_directory_name
        shutil.move(self.path, old_path)

    def delete(self):
        """Delete file
        """
        os.unlink(self.path)

    async def read(
        self,
        is_data:bool=True,
        *args,
        **kwargs
    ) -> File:
        """Read file.
        Returns:
            File: self.
        """
        
        if self.extension in ['xlsx','xls','xlsm']:
            if 'sheet_name' in kwargs:
                self.content = pandas.read_excel(self.path,*args,**kwargs)
            elif 'sheets' in kwargs:
                sheets = kwargs.pop('sheets')

                self.content = pandas.concat([
                    self.readExcel(
                        *args,
                        **{
                            **kwargs,
                            **({'sheet_name':sheet} if isinstance(sheet,str) else sheet)
                        }
                    ) for sheet in sheets
                ])
            
            else:
                excel_content = openpyxl.load_workbook(
                    self.path,
                    data_only=True,
                    read_only=True,
                    keep_vba=False,
                    keep_links=False
                )
                self.content = pandas.read_excel(
                    self.path,
                    sheet_name=excel_content.sheetnames[0],
                    *args,
                    **kwargs
                )

        elif self.extension in ['csv','zip','txt'] and is_data:
            if self.file_server:
                data = await self.file_server.readFile(self.name,self.path)

            self.content = pandas.read_csv(
                data if self.file_server else self.path,
                *args,
                **kwargs
            )

        elif self.extension =='dbf':

            table = DBF(
                self.path,
                ignore_missing_memofile=True,
                encoding = kwargs.get('encoding')
            )

            if 'nrows' in kwargs and kwargs.get('nrows') == 0:
                self.content = pandas.DataFrame(columns = table.field_names)
            else:
                self.content = pandas.DataFrame(iter(table))

        elif self.extension =='shp':
            rows = kwargs.pop('nrows',None)
            self.content = geopandas.read_file(
                self.path,
                rows = rows,
                *args,
                **kwargs
            )

        elif self.extension in ['xml','kml']:

            if is_data:
                self.content = pandas.read_xml(self.path,*args,**kwargs)

            else:
                markup_reader = kwargs.pop('markup_reader')

                with open(self.path, "r",*args,**kwargs) as file_content:

                    if markup_reader == 'soup':
                        raw_content = file_content.read()
                        self.content = BeautifulSoup(raw_content, self.extension)

                    else:
                        self.content = ET.XML(file_content)

        elif self.extension == 'sav':
            self.cont = pandas.read_spss(self.path)
        
        elif self.extension == 'pkl':
            
            with open(self.path, 'rb') as input:
                self.content = pickle.load(input)

        else:
            kwargs.pop('nrows',None)
            with open(self.path, "r",*args,**kwargs) as file_content:
  
                if self.extension  == 'sql':
                    self.content = file_content.read()

                elif self.extension == 'json':
                    self.content = json.load(file_content)

        return self

    def readExcel(self,*args,**kwargs) -> pandas.DataFrame:
        
        frame = pandas.read_excel(
            self.path,*args,**kwargs
        )
        if all([column[-2:] == '.1' for column in frame.columns]):
            frame.columns = [column[:-2] for column in frame.columns]

        return frame

    async def write(self,content:Union[pandas.DataFrame]=None,*args,**kwargs):

        content = content if content is not None else self.content

        if self.extension == 'csv':
            if self.file_server and not self.is_temporary:
                s_buf = BytesIO()
                content.to_csv(
                    s_buf,
                    index=False
                )
                s_buf.seek(0)
                self.file_server.storeFile('data',self.path,s_buf)
            
            else:
                content.to_csv(
                    self.path,
                    index=False
                )

        elif self.extension == 'dbf':
            
            data = content.reset_index(drop=True)

            struct = {}

            # Create structure from dataframe dtypes
            for tup in list(zip(data.dtypes.index, data.dtypes)):

                dataBis = data[tup[0]][pandas.notna(data[tup[0]])]

                if tup[1].kind == 'i':
                    precision = 0
                    total = max(len(str(max(list(dataBis)))) + 1, precision + 2)
                    struct[tup[0]] = 'N (' + str(total) + ', ' + str(precision) + ')'

                elif tup[1].kind == 'f':
                
                    precision = max(dataBis.astype(str).apply(lambda x: len(x.split('.')[1])))
                    total = max(max(dataBis.astype(str).apply(lambda x: len(x.split('.')[0]))) + precision + 1, precision + 2)

                    struct[tup[0]] = 'N (' + str(total) + ', ' + str(precision) + ')'

                elif tup[1].kind == 'M':
                
                    struct[tup[0]] = 'D'

                else:
                    length = max(dataBis.astype(str).apply(lambda x: len(x)))
                    data[tup[0]] = data[tup[0]].astype(str)
                    struct[tup[0]] = 'C (' + str(length) + ')'

                    if length > 255 or tup[0] == 'GEOMET' or tup[0] == 'COMMENT':
                        struct[tup[0]] = 'M'

            cols=[]

            #verify that col names are legit (they should be)
            for col in data.columns:
                #name cannot contain special characters
                name = re.sub(re.compile("[^A-Za-z0-9]"), '_', col)
                #name cannot start with a number
                name = '_'+name if name[0].isdigit() else name
                #name is restricted to 10 digits
                name = name[:10]

                if name != col:
                    cols.append(name)
                    struct[name] = struct.pop(col)
                else:
                    cols.append(col)

            data.columns = cols

            #replace all NaN with None so the dbf librairy can work
            data = data.astype(object).where(data.notnull(), None)

            #drop columns that are not in struct
            data = data.drop(columns=[c for c in data.columns if c not in struct.keys()])

            #and now build the struc as expected by the dbf librairy
            struct = [f"{col} {struct[col]}" for col in data.columns]

            #create and fill the table
            with dbf.Table(
                self.path,
                on_disk=True,
                dbf_type='db3',
                codepage='cp1252',
                memo_size=512,
                field_specs=struct
            ) as table:

                tabled_dic = data.to_dict(orient='index')

                for key in tabled_dic.keys():
                    table.append(tabled_dic[key])

        elif self.extension  in['geojson','json']:
            if 'FILE_SERVER_GEOGRAPHY_PORT' in os.environ:
                request = AsyncRequest(
                    host = os.getenv('FILE_SERVER_GEOGRAPHY_HOST','localhost'),
                    port = os.getenv('FILE_SERVER_GEOGRAPHY_PORT')
                )
                await request.setSession(
                    verify_ssl=False,
                    limit=2,
                    force_close=True
                )

                try:
                    await request.session.post(
                        '',
                        json = {
                            'path':self.path,
                            'data':content
                        }
                    )
                except Exception as e:
                    print(e)


                await request.close()


            else :
                file = open(self.path, "w",*args,**kwargs)
                json.dump(content, file, indent=4)
                file.close()
        
        elif self.extension in ['sql','py']:
            file = open(self.path, "w",*args,**kwargs)
            file.write(content)
            file.close()
            
        elif self.extension == 'pkl':
            with open(self.path, 'wb') as output:
                dill.dump(content, output)
            
        else:
            if self.file_server and not self.is_temporary:
                s_buf = BytesIO(content)
                s_buf.seek(0)
                await self.file_server.writeFile(self.name,s_buf,self.path)

            else:
                file = open(self.path, "wb",*args,**kwargs)
                file.write(content)
                file.close()

        if not self.file_object:
            await self.initiate()

    async def getParentDirectory(self) -> None:
        self.parent_directory = await self.file_server.getParentDirectory(
            self.share_name,
            self.path
        )

class AsyncZipFile(zipfile.ZipFile):
    """Represents a zip file

    Attributes:
    ----------
        Name (str): Name of the zip file

    Methods:
    ----------
        Methods
    """
    def __init__(
        self,
        name:str = None,
        directory_name:str = None,
        path:str = None,
        file_server:FileServer = None,
        is_temporary:bool = False,
        share_name:str = 'Data'
    ):
        """Initiate the zip file

        Args:
            name (str): Name of the zip file
            directory_name (str): Name of the directory
            path (str, optional): _description_. Defaults to None.
        """
        
        self.directory_name,self.name = getDirectoryElementName(
            path,
            directory_name,
            name
        )
        
        self.file_server = file_server
        self.share_name = share_name
        self.is_temporary = is_temporary

        self.parent_directory:SMBDirectory = None
        self.file_object:SMBFile = None

    async def initiate(self) -> AsyncZipFile:
        if self.file_server:
            if self.parent_directory is None:
                self.parent_directory = await self.file_server.getParentDirectory(
                    self.share_name,
                    self.path
                )
            
            self.file_object = self.parent_directory.files[self.name]
            
        super(AsyncZipFile,self).__init__(
            await self.file_server.readFile(
                self.name,
                self.path,
                self.parent_directory
            ) if self.file_server else self.path
        )

    @property
    def path(self):
        """Set path of the zip file
        """
        return os.path.normpath(os.path.join(
            self.directory_name,
            self.name
        ))
    
    @property
    def name_(self):
        """Get the name without the extension
        """
        return os.path.splitext(self.name)[0].lower()

    @property
    def size(self):
        """Get the file size
        """
        return self.file_object.size if self.file_server else os.stat(self.path).st_size

