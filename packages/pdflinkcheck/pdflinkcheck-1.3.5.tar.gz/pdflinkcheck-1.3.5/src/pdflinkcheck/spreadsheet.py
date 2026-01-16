#!/usr/bin/env python3
# SPDX-License-Identifier: MIT

from __future__ import annotations
import sys
import re
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from pdflinkcheck.io import PDFLINKCHECK_HOME, get_friendly_path, get_unique_unix_time

# ----------------- Helper Functions -----------------

def is_temp_pdf(pdf_name: str) -> bool:
    """Detect likely temp or unstable PDF filenames."""
    temp_patterns = [r'^tmp', r'^~', r'^[0-9a-fA-F]{8,}-', r'^temp']
    return any(re.match(pat, pdf_name, re.IGNORECASE) for pat in temp_patterns)

def sanitize_excel_text(text: str) -> str:
    """Remove characters illegal for Excel cells."""
    if not isinstance(text, str):
        text = str(text)
    return ''.join(c for c in text if c.isprintable() or c in '\t')

def convert_goto_link(link: Dict, pdf_path: str) -> str:
    """
    Convert internal GoTo link into a clickable file:// URL.
    Excel will recognize this as a hyperlink.
    """
    page_index = link.get('destination_page')
    if page_index is None:
        return f'file://{Path(pdf_path).resolve()}'

    human_page = page_index + 1
    full_path = Path(pdf_path).resolve()
    return f'file://{full_path}#page={human_page}'

def prepare_links_by_type(report: Dict, pdf_path: str = None) -> Dict[str, List[Dict]]:
    """Prepare links grouped by type for separate Excel shee6ts."""
    pdf_name = report['metadata']['file_overview']['pdf_name']
    if pdf_path is None:
        pdf_path = report["metadata"]["file_overview"]["source_path"]
    if not pdf_path:
        raise RuntimeError("source_path missing from report metadata")

    if is_temp_pdf(pdf_name):
        raise ValueError(f"PDF filename '{pdf_name}' looks like a temporary or unstable file. Provide a stable filename.")

    grouped_links = {'Internal GoTo': [], 'External URI': [], 'Other': []}

    all_links = (
        report['data'].get('internal_links', []) +
        report['data'].get('external_links', []) +
        report['data'].get('other_links', [])
    )

    for link in all_links:
        link_type = link.get('type', 'Unknown')
        anchor_text = link.get('link_text', 'N/A')
        anchor_text = sanitize_excel_text(anchor_text)

        if link_type in ('Internal (GoTo/Dest)', 'Internal (Resolved Action)'):
            url = convert_goto_link(link, pdf_path)
            grouped_links['Internal GoTo'].append({
                'page': link.get('page', 'N/A'),
                'anchor_text': anchor_text,
                'hyperlink': url
            })
        elif link_type == 'External (URI)':
            url = link.get('url') or link.get('remote_file') or link.get('target') or ''
            grouped_links['External URI'].append({
                'page': link.get('page', 'N/A'),
                'anchor_text': anchor_text,
                'hyperlink': url
            })
        else:
            url = link.get('url') or link.get('remote_file') or link.get('target') or ''
            grouped_links['Other'].append({
                'page': link.get('page', 'N/A'),
                'anchor_text': anchor_text,
                'hyperlink': url
            })

    return grouped_links

def _export_links_to_xlsx(grouped_links: Dict[str, List[Dict]], output_file: Path):
    """
    Export grouped links into separate sheets in an XLSX workbook.
    Accepts a pre-constructed Path object for the output file.
    """
    from openpyxl import Workbook

    wb = Workbook()

    for sheet_name, links in grouped_links.items():
        ws = wb.create_sheet(sheet_name)
        # Heading row
        headers = ['Page', 'Anchor Text', 'Hyperlink']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for link in links:
            page = sanitize_excel_text(link['page'])
            anchor = sanitize_excel_text(link['anchor_text'])
            hyperlink = link['hyperlink']

            ws.append([page, anchor, hyperlink])
            ws.cell(row=ws.max_row, column=3).hyperlink = hyperlink
            ws.cell(row=ws.max_row, column=3).style = 'Hyperlink'

        # Auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(output_file)
    print(f"XLSX exported successfully to {get_friendly_path(output_file)}")

def export_report_links_to_xlsx(report: Dict, output_dir: Path = None) -> Path:
    """
    Takes a report dictionary (from run_report_meat) and exports
    grouped clickable links to XLSX. Returns the XLSX path.
    """
    output_dir = output_dir or PDFLINKCHECK_HOME

    # 1. Group and process links
    grouped_links = prepare_links_by_type(report)

    # 2. Construct unique output file name with Unix timestamp
    pdf_name = report['metadata']['file_overview']['pdf_name']
    pdf_stem = Path(pdf_name).stem
    timestamp = get_unique_unix_time()
    output_file = output_dir / f"{pdf_stem}_{timestamp}_report.xlsx"

    # 3. Write XLSX
    _export_links_to_xlsx(grouped_links, output_file)

    return output_file

# ----------------- Main / Proof-of-Concept -----------------

def main(pdf_path: str = None):
    from pdflinkcheck.report import run_report_meat, get_first_pdf_in_cwd
    if pdf_path is None:
        pdf_path = get_first_pdf_in_cwd()
        if not pdf_path:
            print("No PDF found in current directory.")
            sys.exit(1)

    report = run_report_meat(pdf_path=pdf_path, pdf_library = "auto", print_bool=True, concise_print=True)

    export_report_links_to_xlsx(report)

if __name__ == "__main__":
    main()
