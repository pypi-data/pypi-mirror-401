"""ExcelReader class for extracting data from Excel files."""

import warnings
import zipfile
from pathlib import Path

from openpyxl import load_workbook, workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException


class ExcelReader:
    """A class to read and extract data from Excel files (.xlsx, .xlsm, .xlsb).

    This class provides methods to load workbooks, extract data from sheets,
    tables, and named ranges, with robust error handling.
    """

    def __init__(self, file_path: Path | str) -> None:
        """Initialize the ExcelReader with the path to the Excel file.

        Args:
            file_path: Path to the Excel file, specified as a Path object or string.

        Raises:
            ImportError: If the file does not exist, is not a valid Excel file.
        """
        if isinstance(file_path, str):
            file_path = Path(file_path)
        if not file_path.is_file():
            msg = f"File {file_path} does not exist or is not a valid file."
            raise ImportError(msg)

        # Check extension to see if it's an Excel file
        if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xlsb"}:
            msg = f"File {file_path} is not a valid Excel file."
            raise ImportError(msg)

        self.file_path = file_path

    def load_excel_workbook(self, *, data_only: bool = True, read_only: bool = False) -> workbook.Workbook:
        """Load an Excel workbook with robust error handling.

        Args:
            data_only: Whether to return cell values (not formulas).
            read_only: Use openpyxl's read-only mode for large files.

        Returns:
            workbook (workbook.Workbook): Workbook object.

        Raises:
            ImportError: If the file cannot be loaded due to various reasons.
        """
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        try:
            wb = load_workbook(
                filename=self.file_path, data_only=data_only, read_only=read_only
            )
        except FileNotFoundError:
            msg = f"File not found: {self.file_path}"
        except PermissionError:
            msg = f"Permission denied: {self.file_path}"
        except InvalidFileException:
            msg = f"Invalid Excel file format: {self.file_path}"
        except zipfile.BadZipFile:
            msg = f"Corrupt Excel file (not a valid ZIP): {self.file_path}"
        except OSError as e:
            msg = f"Unexpected error loading Excel file {self.file_path}: {e}"
        else:
            if wb is None:
                msg = f"Failed to load workbook from {self.file_path}. The file may be corrupted or not a valid Excel file."
            else:
                return wb

        raise ImportError(msg)

    def extract_data(self, source_name: str, source_type: str) -> list[dict]:
        """Extract data from an Excel file based on the source type and name.

        Expected the specified source type to be either:
        - An entire worksheet with the header in the first row (sheet)
        - A named range (range)
        - An Excel table (table)

        Args:
            source_name: Name of the sheet, table, or range to extract.
            source_type: Type of source ('sheet', 'table', or 'range').

        Returns:
            Data extracted as a list of dictionaries.

        Raises:
            ImportError: If the source type is invalid or if there are issues extracting data.
        """
        if source_type == "sheet":
            return self.extract_from_sheet(source_name)
        if source_type == "table":
            return self.extract_from_table(source_name)
        if source_type == "range":
            return self.extract_from_range(source_name)

        msg = f"Invalid source type '{source_type}'. Must be 'sheet', 'table', or 'range'."
        raise ImportError(msg)

    def extract_from_sheet(self, sheet_name: str) -> list[dict]:
        """Extract a sheet from an Excel file and return it as a list of dictionaries.

        Args:
            sheet_name: Name of the sheet to extract.

        Returns:
            A list containing the sheet data as dictionaries.

        Raises:
            ImportError: If the sheet cannot be loaded or if there are issues with the file.
        """
        try:
            wb = self.load_excel_workbook(data_only=True, read_only=True)

            if sheet_name not in wb.sheetnames:
                msg = f"Sheet '{sheet_name}' not found in Excel file {self.file_path}."
                raise ImportError(msg)  # noqa: TRY301

            ws = wb[sheet_name]

            # Get all data from the sheet
            data = []
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                return data

            # First row is header
            header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

            # Convert remaining rows to dictionaries
            for row in rows[1:]:
                # Skip empty rows
                if all(cell is None or not str(cell).strip() for cell in row):
                    continue

                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(header):
                        row_dict[header[i]] = cell
                data.append(row_dict)

        except ImportError:
            raise
        except Exception as e:
            msg = f"Error loading sheet '{sheet_name}' from Excel file {self.file_path}: {e}"
            raise ImportError(msg) from e

        return data

    def extract_from_table(self, table_name: str) -> list[dict]:
        """Extract a table from an Excel file and return it as a list of dictionaries.

        Args:
            table_name: Name of the table to extract.

        Returns:
            A list containing the table data as dictionaries.

        Raises:
            ImportError: If the table cannot be loaded or if there are issues with the file.
        """
        wb = self.load_excel_workbook(data_only=True, read_only=False)
        found = False
        selected_sheet = None
        table_range = None

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if table_name in ws.tables:
                table = ws.tables[table_name]
                table_range = table.ref  # e.g., "B3:F12"
                selected_sheet = sheet_name
                found = True
                break

        if not found:
            msg = f"Table '{table_name}' not found in any worksheet of the Excel file {self.file_path}."
            raise ImportError(msg)

        try:
            min_col, min_row, max_col, max_row = range_boundaries(table_range)  # pyright: ignore[reportArgumentType]

            if min_col is None or max_col is None or max_row is None or min_row is None:
                msg = f"Invalid table range boundaries for table '{table_name}' in file {self.file_path}."
                raise ImportError(msg)  # noqa: TRY301

            ws = wb[selected_sheet]  # pyright: ignore[reportArgumentType]

            # Extract data from the table range
            data = []
            rows = list(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )

            if not rows:
                return data

            # First row is header
            header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

            # Convert remaining rows to dictionaries
            for row in rows[1:]:
                # Skip empty rows
                if all(cell is None or not str(cell).strip() for cell in row):
                    continue

                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(header):
                        row_dict[header[i]] = cell
                data.append(row_dict)

        except Exception as e:
            msg = f"Error loading table '{table_name}' from Excel file {self.file_path}: {e}"
            raise ImportError(msg) from e

        return data

    def extract_from_range(self, range_name: str) -> list[dict]:  # noqa: PLR0914
        """Extract a named range from an Excel file and return it as a list of dictionaries.

        Args:
            range_name: Name of the range to extract.

        Returns:
            A list containing the range data as dictionaries.

        Raises:
            ImportError: If the range cannot be loaded or if there are issues with the file.
        """
        wb = self.load_excel_workbook(data_only=True, read_only=True)
        found = False
        sheet_name = None
        ref = None

        for defined_name in wb.defined_names:
            if defined_name == range_name:
                dn = wb.defined_names[defined_name]
                destinations = list(dn.destinations)
                if len(destinations) != 1:
                    msg = f"Range '{range_name}' refers to multiple areas, which is unsupported."
                    raise ImportError(msg)

                sheet_name, ref = destinations[0]
                found = True
                break

        if not found:
            msg = f"Range '{range_name}' not found in any worksheet of the Excel file {self.file_path}."
            raise ImportError(msg)

        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)  # pyright: ignore[reportArgumentType]

            if min_col is None or max_col is None or max_row is None or min_row is None:
                msg = f"Invalid range boundaries for range '{range_name}' in file {self.file_path}."
                raise ImportError(msg)  # noqa: TRY301

            ws = wb[sheet_name]  # pyright: ignore[reportArgumentType]

            # Extract data from the named range
            data = []
            rows = list(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )

            if not rows:
                return data

            # First row is header
            header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

            # Convert remaining rows to dictionaries
            for row in rows[1:]:
                # Skip empty rows
                if all(cell is None or not str(cell).strip() for cell in row):
                    continue

                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(header):
                        row_dict[header[i]] = cell
                data.append(row_dict)

        except Exception as e:
            msg = f"Error loading data range '{range_name}' from Excel file {self.file_path}: {e}"
            raise ImportError(msg) from e

        return data
