from FAIRLinked.QBWorkflow.utility import NAMESPACE_MAP, HEADER_NAMESPACE, HEADER_BASE_URI
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

def generate_namespace_excel(excel_file_path):
    """
    Generates an Excel file listing default namespaces and their corresponding base URIs.
    The Excel file is styled with borders, colored headers, and centered text for better readability.

    Args:
        excel_file_path (str): The path where the generated Excel file will be saved.

    Returns:
        None: The function saves an Excel file containing the namespace-URI mappings to the provided file path.
    """
    
    try:
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define header fill colors
        header_fill_namespace = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
        header_fill_base_uri = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # Light blue

        # Define header font
        header_font = Font(bold=True)

        # Define alignment with wrap_text for long URIs
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define column headers
        headers = [HEADER_NAMESPACE, HEADER_BASE_URI]

        # Write headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Apply different fill colors for each header
            if col_num == 1:
                cell.fill = header_fill_namespace
                # Shorten the Namespace column width
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
            elif col_num == 2:
                cell.fill = header_fill_base_uri
                # Increase the Base URI column width
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 50
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Write data starting from the second row
        for row_num, (namespace, uri) in enumerate(NAMESPACE_MAP.items(), start=2):
            # Namespace column
            cell = ws.cell(row=row_num, column=1, value=namespace)
            cell.alignment = center_alignment
            cell.border = thin_border

            # Base URI column
            cell = ws.cell(row=row_num, column=2, value=uri)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Adjust row heights for wrapping text
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 40  # Adjust as needed for your data

        # Save the workbook
        wb.save(excel_file_path)
        print(f"Excel file '{excel_file_path}' has been generated with default namespaces.")
    except Exception as e:
        print(f"An unexpected error occurred while generating the Excel file: {e}")
