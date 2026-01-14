from FAIRLinked.QBWorkflow.utility import NAMESPACE_MAP
import openpyxl

def parse_excel_to_namespace_map(excel_file_path):
    """
    Parses the Excel file containing namespaces and base URIs.
    Updates and returns the namespace map.

    Args:
        excel_file_path (str): The path to the Excel file to parse.

    Returns:
        dict: Updated namespace map.
    """
    # Start with the default namespace map
    namespace_map = NAMESPACE_MAP.copy()

    try:
        # Load the workbook and select the active worksheet
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # Get headers from the first row
        headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

        # Check if required columns are present
        if "Namespace you are using" not in headers or "Base URI" not in headers:
            raise KeyError("Missing expected columns in the Excel file.")

        # Map header names to column indices
        header_indices = {header: idx for idx, header in enumerate(headers)}

        # Prepare a reverse mapping to detect conflicts
        uri_to_namespace = {v: k for k, v in namespace_map.items()}

        # Iterate over the rows starting from the second row
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Safely get namespace and base URI values
            namespace = row[header_indices["Namespace you are using"]]
            base_uri = row[header_indices["Base URI"]]

            if namespace and base_uri:
                namespace = str(namespace).strip().lower()
                base_uri = str(base_uri).strip()

                # Check for conflicting URIs
                if base_uri in uri_to_namespace and uri_to_namespace[base_uri] != namespace:
                    raise ValueError(f"Conflict detected: URI '{base_uri}' is already mapped to namespace '{uri_to_namespace[base_uri]}'.")
                namespace_map[namespace] = base_uri
                uri_to_namespace[base_uri] = namespace

    except FileNotFoundError:
        print(f"The file '{excel_file_path}' was not found.")
    except KeyError as e:
        print(f"Missing expected column in the Excel file: {e}")
    except ValueError as e:
        print(e)
    except Exception as e:
        print(f"An unexpected error occurred while parsing the Excel file: {e}")

    return namespace_map
