import openpyxl
from openpyxl.styles import (
    PatternFill, Alignment, Font, Border, Side
)
from FAIRLinked.QBWorkflow.utility import (
    CATEGORY_COLORS, ALT_LABEL_INSTR, 
    UNIT_INSTR, IS_MEASURE_INSTR, EXISTING_URI_INSTR, 
    EXPERIMENT_ID_INSTR, NO_TERMS_MSG, TEMPLATE_GENERATED_MSG
)

def generate_data_xlsx_template(children_terms, output_xlsx_file_path):
    """
    Generates an Excel template for data collection with various features such as merged cells, category headers, 
    colors, formatting, and borders. If no categories are provided, a 'Miscellaneous' category is generated.

    Args:
        children_terms (dict): A dictionary where keys represent ontology categories (e.g., 'mds:tool') and values 
                               are lists of terms (e.g., ['InstrumentId', 'InstrumentName']). If categories are empty, 
                               a default 'Miscellaneous' category is created.
        output_xlsx_file_path (str): The path where the generated Excel file should be saved.

    Returns:
        None: The function creates and saves an Excel file with the specified formatting at the provided path.
    """

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row indices based on 1-indexing
    row_category_headers = 1
    row_static_instructions_start = 2
    row_static_instructions_end = 5
    row_variable_names = 6
    row_data_start = 7

    # Add the instruction in Row 1, Column B
    cell = ws.cell(row=row_category_headers, column=2, value=EXPERIMENT_ID_INSTR)
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.border = thin_border
    ws.column_dimensions['B'].width = 50

    # Static instructions in Column A
    instruction_colors = ["FFF2CC", "E2EFDA", "D9E1F2", "FCE4D6"]  # Light colors for each instruction
    static_instructions = [
        ALT_LABEL_INSTR,
        UNIT_INSTR,
        IS_MEASURE_INSTR,
        EXISTING_URI_INSTR
    ]
    for idx, instruction in enumerate(static_instructions, start=row_static_instructions_start):
        cell = ws.cell(row=idx, column=1, value=instruction)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        fill_color = instruction_colors[idx - row_static_instructions_start]
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.border = thin_border
        ws.row_dimensions[idx].height = 60  # Adjust row height for wrapping
    ws.column_dimensions['A'].width = 80  # To accommodate long instructions

    # Column B: ExperimentId header (Row 6)
    cell = ws.cell(row=row_variable_names, column=2, value="ExperimentId")
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.border = thin_border
    ws.column_dimensions['B'].width = 20

    # Determine which categories have terms
    categories_with_terms = {category: terms for category, terms in children_terms.items() if terms}

    if not categories_with_terms:
        # No categories have terms, create a 'Miscellaneous' category
        print(NO_TERMS_MSG)
        categories_with_terms = {'Miscellaneous': ['var1', 'var2', 'var3']}
        category_colors = {'Miscellaneous': 'CCCCCC'}  # Light gray
    else:
        # Use the existing CATEGORY_COLORS for the categories present
        category_colors = {category: CATEGORY_COLORS.get(category, "FFFFFF") for category in categories_with_terms.keys()}

    # Starting from Column C
    col_idx = 3  # Column C

    for category, terms in categories_with_terms.items():
        # Add debug print
        # print(f"Processing category: {category}")
        # print(f"Available colors: {CATEGORY_COLORS}")
        
        start_col = col_idx
        # Write variable names
        for term in terms:
            cell = ws.cell(row=row_variable_names, column=col_idx, value=term)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(term) + 2, 15)
            col_idx += 1
        end_col = col_idx - 1

        # Merge category headers (Row 1)
        ws.merge_cells(start_row=row_category_headers, start_column=start_col, end_row=row_category_headers, end_column=end_col)
        category_name = category.replace("mds:", "").capitalize()
        category_cell = ws.cell(row=row_category_headers, column=start_col, value=category_name)

        # Apply formatting to category header
        fill_color = category_colors.get(category, "FFFFFF")
        # print(f"Using color {fill_color} for category {category}")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        category_cell.fill = fill
        category_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        category_cell.font = Font(bold=True)

        # Apply borders to each cell in the merged range of category header
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_category_headers, column=col)
            cell.border = thin_border

        # Apply fill and borders to variable name cells and data cells
        for col in range(start_col, end_col + 1):
            # Variable name cells
            cell = ws.cell(row=row_variable_names, column=col)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = thin_border
            # Data cells
            for row in range(row_data_start, row_data_start + 2):  # Adjust range as needed
                data_cell = ws.cell(row=row, column=col)
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
                data_cell.border = thin_border

    # Add the "Data (if any)" column
    cell = ws.cell(row=row_category_headers, column=col_idx, value="Data (if any) Supply the name of your data file")
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30

    # FileName header in Row 6
    cell = ws.cell(row=row_variable_names, column=col_idx, value="FileName")
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.border = thin_border

    # Apply alignment and borders to FileName data cells
    for row in range(row_data_start, row_data_start + 2):  # Adjust range as needed
        data_cell = ws.cell(row=row, column=col_idx)
        data_cell.alignment = Alignment(horizontal='center', vertical='center')
        data_cell.border = thin_border

    # Sample data rows
    ws.cell(row=row_data_start, column=2, value="1")  # ExperimentId
    ws.cell(row=row_data_start, column=col_idx, value="Image_1.jpg")  # FileName
    ws.cell(row=row_data_start + 1, column=2, value="2")  # Next ExperimentId

    # Apply borders and alignment to ExperimentId data cells
    for row in range(row_data_start, row_data_start + 2):
        cell = ws.cell(row=row, column=2)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Center align all cells and apply borders
    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            if cell.row > row_variable_names:  # Data cells
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Save the workbook
    wb.save(output_xlsx_file_path)
    print(TEMPLATE_GENERATED_MSG.format(output_xlsx_file_path))
