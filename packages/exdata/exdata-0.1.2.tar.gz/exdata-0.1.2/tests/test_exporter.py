from collections.abc import Iterator
from io import BytesIO
from math import floor
from pathlib import Path
from tempfile import NamedTemporaryFile

import pytest
from openpyxl.reader.excel import load_workbook

from exdata import XlsxExporter
from exdata.exceptions import FormatNotFoundError
from exdata.struct import Cell, Column, Format, RichValue, Row, Sheet
from tests import make_workbook


def test_one_cell() -> None:
    wb = make_workbook(data=[Row([Cell('Test')])])

    assert wb['Test']['A1'].value == 'Test'


def test_multiple_cells() -> None:
    wb = make_workbook(data=[Row([Cell('Test 1'), Cell('Test 2')])])

    assert wb['Test']['A1'].value == 'Test 1'
    assert wb['Test']['B1'].value == 'Test 2'


def test_cells_format() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell(value='', format='format1'),
                    Cell(value='', format='format2'),
                ]
            )
        ],
        formats={
            'format1': {
                'bg_color': '#ff00ff',
                'align': 'center',
                'valign': 'center',
                'border': 1,
                'bottom': 2,
                'text_wrap': True,
            },
            'format2': {'bg_color': '#0000ff'},
        },
    )
    cell = wb['Test']['A1']
    assert cell.fill.fgColor.rgb == 'FFFF00FF'
    assert cell.alignment.horizontal == 'center'
    assert cell.border.top.style == 'thin'
    assert cell.border.left.style == 'thin'
    assert cell.border.right.style == 'thin'
    assert cell.border.bottom.style == 'medium'
    assert cell.alignment.wrapText is True

    assert wb['Test']['B1'].fill.fgColor.rgb == 'FF0000FF'


def test_format_inheritance() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell('Test 1'),
                    Cell('Test 2', format='cell'),
                ],
                format='row',
            ),
        ],
        sheet_format='sheet',
        formats={
            'sheet': {'align': 'center'},
            'row': {'bg_color': '#ff00ff'},
            'cell': {'bg_color': '#0000ff'},
        },
    )

    assert wb['Test']['A1'].alignment.horizontal == 'center'
    assert wb['Test']['A1'].fill.fgColor.rgb == 'FFFF00FF'
    assert wb['Test']['B1'].alignment.horizontal == 'center'
    assert wb['Test']['B1'].fill.fgColor.rgb == 'FF0000FF'


def test_cell_offset() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell('Test 1', x_offset=1, y_offset=1),
                    Cell('Test 2', x_offset=2, y_offset=2),
                ]
            )
        ]
    )

    assert wb['Test']['B2'].value == 'Test 1'
    assert wb['Test']['E3'].value == 'Test 2'


def test_cell_rows() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell('Test 1', rows=3),
                    Cell('Test 2', rows=3),
                ]
            )
        ]
    )

    mc1, mc2 = wb['Test'].merged_cells.sorted()

    assert wb['Test']['A1'].value == 'Test 1'
    assert mc1.coord == 'A1:A3'

    assert wb['Test']['B1'].value == 'Test 2'
    assert mc2.coord == 'B1:B3'


def test_cell_columns() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell('Test 1', columns=2),
                    Cell('Test 2', columns=3),
                ]
            )
        ]
    )

    mc1, mc2 = wb['Test'].merged_cells.sorted()

    assert wb['Test']['A1'].value == 'Test 1'
    assert mc1.coord == 'A1:B1'

    assert wb['Test']['C1'].value == 'Test 2'
    assert mc2.coord == 'C1:E1'


def test_multiple_rows() -> None:
    wb = make_workbook(
        data=[
            Row([Cell('Test 1')]),
            Row([Cell('Test 2')]),
        ]
    )

    assert wb['Test']['A1'].value == 'Test 1'
    assert wb['Test']['A2'].value == 'Test 2'


def test_column() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Column(
                        [
                            Cell('Test 1'),
                            Cell('Test 2'),
                        ]
                    )
                ]
            )
        ]
    )

    assert wb['Test']['A1'].value == 'Test 1'
    assert wb['Test']['A2'].value == 'Test 2'


def test_multiple_columns() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Column(
                        [
                            Cell('Test 1'),
                            Cell('Test 2'),
                        ]
                    ),
                    Column(
                        [
                            Cell('Test 3'),
                            Cell('Test 4'),
                        ]
                    ),
                ]
            )
        ]
    )

    assert wb['Test']['A1'].value == 'Test 1'
    assert wb['Test']['A2'].value == 'Test 2'
    assert wb['Test']['B1'].value == 'Test 3'
    assert wb['Test']['B2'].value == 'Test 4'


@pytest.mark.parametrize('cell_format', [None, 'cell_format'])
def test_rich_value(cell_format: str | None) -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell(
                        value=RichValue(['Test', Format('bold'), ' bold']),
                        format=cell_format,
                    )
                ]
            )
        ],
        formats={
            'bold': {'bold': True},
            'cell_format': {'italic': True},
        },
    )

    cell = wb['Test']['A1']
    assert cell.value.as_list() == ['Test', ' bold']
    assert cell.value[1].font.bold is True
    assert cell.font.i is (cell_format is not None)


def test_multiple_sheets() -> None:
    exporter = XlsxExporter(
        sheets=[
            Sheet(
                data=[Row([Cell('Test 1')])],
                name='Test 1',
            ),
            Sheet(
                data=[Row([Cell('Test 2')])],
                name='Test 2',
            ),
        ],
    )
    wb = load_workbook(exporter.export())

    assert wb['Test 1']['A1'].value == 'Test 1'
    assert wb['Test 2']['A1'].value == 'Test 2'


def test_generators() -> None:
    def generate_column_data(row_id, col_id) -> Iterator[Cell]:
        yield Cell(f'Test {row_id}.{col_id}')

    def generate_row_data(row_id) -> Iterator[Column]:
        for col_id in range(1, 4):
            yield Column(generate_column_data(row_id, col_id))

    def generate_sheet_data() -> Iterator[Row]:
        for row_id in range(1, 4):
            yield Row(generate_row_data(row_id))

    wb = make_workbook(
        data=generate_sheet_data(),
    )

    for i in range(1, 4):
        for j in range(1, 4):
            assert wb['Test'].cell(row=i, column=j).value == f'Test {i}.{j}'


def test_row_height() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Column(
                        [
                            Cell('Test 1'),
                            Cell('Test 2'),
                            Cell('Test 3'),
                        ]
                    )
                ],
                heights=[50, 40],
            )
        ]
    )
    assert wb['Test'].row_dimensions[1].ht == 50
    assert wb['Test'].row_dimensions[2].ht == 40


def test_column_width() -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell('Test 1'),
                    Cell('Test 2'),
                    Cell('Test 3'),
                ],
            )
        ],
        column_width={1: 50, 2: 40},
    )
    assert floor(wb['Test'].column_dimensions['B'].width) == 50
    assert floor(wb['Test'].column_dimensions['C'].width) == 40


@pytest.mark.parametrize(
    ('max_width', 'expected'),
    [(None, None), (10, 1.42578125), (100, 14.28515625)],
)
def test_auto_width(max_width: int | None, expected: float | None) -> None:
    wb = make_workbook(
        data=[
            Row(
                [
                    Cell(
                        'Lorem ipsum dolor sit amet, consectetur adipiscing '
                        'elit. In tincidunt vehicula purus et varius. Vivamus '
                        'id tincidunt ipsum. Etiam mattis felis tincidunt, '
                        'finibus ex id, dignissim nibh. Morbi lacinia, tortor '
                        'sed sollicitudin rhoncus, leo enim venenatis mauris, '
                        'tristique dictum libero massa iaculis sapien.'
                    ),
                ],
            )
        ],
        autofit_max_width=max_width,
    )
    width = getattr(wb['Test'].column_dimensions.get('A'), 'width', None)
    assert width == expected


def test_export_to_buffer() -> None:
    buffer = BytesIO()
    exporter = XlsxExporter(
        sheets=[
            Sheet(
                data=[Row([Cell('Test')])],
                name='Test',
            ),
        ],
        buffer=buffer,
    )
    exporter.export()
    wb = load_workbook(buffer)
    buffer.close()
    assert wb['Test']['A1'].value == 'Test'


def test_export_to_file_obj() -> None:
    with NamedTemporaryFile() as f:
        exporter = XlsxExporter(
            sheets=[
                Sheet(
                    data=[Row([Cell('Test')])],
                    name='Test',
                ),
            ],
            buffer=f,
        )
        wb = load_workbook(exporter.export())
    assert wb['Test']['A1'].value == 'Test'


def test_export_to_filepath() -> None:
    with NamedTemporaryFile() as f:
        exporter = XlsxExporter(
            sheets=[
                Sheet(
                    data=[Row([Cell('Test')])],
                    name='Test',
                ),
            ],
            buffer=f.name,
        )
        exporter.export()
        wb = load_workbook(f)

    assert wb['Test']['A1'].value == 'Test'


def test_export_to_path() -> None:
    with NamedTemporaryFile() as f:
        exporter = XlsxExporter(
            sheets=[
                Sheet(
                    data=[Row([Cell('Test')])],
                    name='Test',
                ),
            ],
            buffer=Path(f.name),
        )
        exporter.export()
        wb = load_workbook(f)

    assert wb['Test']['A1'].value == 'Test'


def test_format_not_found() -> None:
    name = 'non_existent_format'
    message = (
        f'Format `{name}` not found. Make sure it is defined in `formats`.'
    )
    with pytest.raises(FormatNotFoundError, match=message):
        make_workbook(
            data=[Row([Cell('Test', format='non_existent_format')])],
        )
