from openpyxl import load_workbook


def excel_autosize_column(file_path: str):
    """Auto size column for all sheets in excel file

    Args:
        file_path (str): Path to excel file

    Raises:
        e: Exception when loading workbook
    """
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        raise e

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name (e.g., 'A')
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2  # Adjust as needed for padding
            ws.column_dimensions[column].width = adjusted_width

    # Save the modified workbook

    wb.save(file_path)
