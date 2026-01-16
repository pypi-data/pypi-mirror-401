import logging

import os

import pandas as pd

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

BOOTSTRAP_SHEET = "sheet1"


def export_data_excel(file_path: str, sheet_name: str, df: pd.DataFrame):

    if not os.path.exists(file_path):
        with pd.ExcelWriter(path=file_path, engine="openpyxl") as writer:
            empty_df = pd.DataFrame()

            empty_df.to_excel(writer, sheet_name=BOOTSTRAP_SHEET, index=False)

    try:
        with pd.ExcelWriter(
            path=file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:

            df.to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        logger.warning(e)

    workbook = load_workbook(file_path)

    if BOOTSTRAP_SHEET in workbook.sheetnames:
        del workbook[BOOTSTRAP_SHEET]
        workbook.save(file_path)
