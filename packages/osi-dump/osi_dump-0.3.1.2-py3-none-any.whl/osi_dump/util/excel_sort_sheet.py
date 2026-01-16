import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def excel_sort_sheet(file_path):
    """
    Sorts sheets in an Excel workbook alphabetically and saves the workbook.

    Args:
        file_path (str): The path to the Excel file to be sorted.
    """
    try:
        # Load the existing workbook
        workbook = load_workbook(file_path)

        # Get the current sheet names
        sheet_names = workbook.sheetnames

        # Sort sheet names alphabetically
        sorted_sheet_names = sorted(sheet_names)

        # Reorder sheets in the workbook
        for index, sheet_name in enumerate(sorted_sheet_names):
            sheet = workbook[sheet_name]
            workbook._sheets.remove(sheet)  # Remove the sheet from its current position
            workbook._sheets.insert(index, sheet)  # Insert it at the new position

        # Save the workbook with the sheets reordered
        workbook.save(file_path)

    except Exception as e:
        logger.warning(f"Error in sorting sheet: {e}")
