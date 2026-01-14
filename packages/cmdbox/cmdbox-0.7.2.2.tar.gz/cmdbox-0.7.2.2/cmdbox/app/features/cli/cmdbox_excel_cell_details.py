from cmdbox.app import common, filer
from cmdbox.app.commons import convert, redis_client
from cmdbox.app.features.cli import excel_base
from cmdbox.app.options import Options
from datetime import datetime, timedelta, time
from pathlib import Path
from typing import Dict, Any, List, Tuple
import argparse
import logging
import json


class ExcelCellDetails(excel_base.ExcelBase):
    def get_cmd(self):
        """
        この機能のコマンドを返します

        Returns:
            str: コマンド
        """
        return 'cell_details'
    
    def get_option(self):
        """
        この機能のオプションを返します

        Returns:
            Dict[str, Any]: オプション
        """
        opt = super().get_option()
        opt['description_ja'] = "データフォルダ配下のExcelファイルの指定したセルの詳細情報を取得します。"
        opt['description_en'] = "Get the details of the specified cell in the Excel file under the data folder."
        opt['choice'] += [
            dict(opt="formula_data_only", type=Options.T_BOOL, default=False, required=True, multi=False, hide=False, choice=[True, False],
                 description_ja="数式データのみを参照するかどうかを指定します。このオプションはキャッシュされたデータが存在する場合に有効です。",
                 description_en="Specify whether to get only formula data. This option is valid if cached data exists."),
            dict(opt="sheet_name", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を取得するシートの名前を指定します。省略した場合、最初のシートが使用されます。",
                 description_en="Specify the sheet name to get the cell value.If omitted, the first sheet will be used."),
            dict(opt="cell_name", type=Options.T_STR, default=None, required=False, multi=True, hide=False, choice=None,
                 description_ja="セルの値を取得するセルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the cell name to get the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="cell_top_left", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を取得する左上セルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the top-left cell name to get the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="cell_bottom_right", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を取得する右下セルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the bottom-right cell name to get the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="output_detail_format", type=Options.T_STR, default='json', required=False, multi=False, hide=False, choice=['json', 'text'],
                 description_ja="出力フォーマットを指定します。例えば、`json`、`text`。",
                 description_en="Specify the output format. For example, `json`, `text`."),
        ]
        return opt

    def chk_args(self, args:argparse.Namespace, tm:float, pf:List[Dict[str, float]]=[]) -> Tuple[bool, str, Any]:
        """
        引数のチェックを行います

        Args:
            args (argparse.Namespace): 引数

        Returns:
            Tuple[bool, str]: チェック結果, メッセージ
        """
        if args.svname is None:
            msg = dict(warn=f"Please specify the --svname option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.scope is None:
            msg = dict(warn=f"Please specify the --scope option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.svpath is None:
            msg = dict(warn=f"Please specify the --svpath option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        return self.RESP_SUCCESS, None, None

    def excel_proc(self, abspath:Path, args:argparse.Namespace, logger:logging.Logger, tm:float, pf:List[Dict[str, float]]=[]) -> Dict[str, Any]:
        """
        Excel処理のベース

        Args:
            abspath (Path): Excelファイルの絶対パス
            args (argparse.Namespace): 引数
            logger (logging.Logger): ロガー
            tm (float): 処理時間
            pf (List[Dict[str, float]]): パフォーマンス情報

        Returns:
            Dict[str, Any]: 結果
        """
        res = self.get_cell_details(abspath, args.formula_data_only, args.sheet_name,
                                    args.cell_name, args.cell_top_left, args.cell_bottom_right,
                                    args.output_detail_format, logger)
        return res

    def get_svparam(self, args:argparse.Namespace) -> List[str]:
        """
        サーバーに送信するパラメーターを返します

        Args:
            args (argparse.Namespace): 引数

        Returns:
            List[str]: サーバーに送信するパラメーター
        """
        cell_name = json.dumps(args.cell_name, default=common.default_json_enc) if args.cell_name is not None else '[]'
        ret = [convert.str2b64str(str(args.svpath)), str(args.formula_data_only), convert.str2b64str(str(args.sheet_name)),
               convert.str2b64str(cell_name), convert.str2b64str(args.cell_top_left), convert.str2b64str(args.cell_bottom_right),
               convert.str2b64str(args.output_detail_format)]
        return ret

    def is_cluster_redirect(self):
        """
        クラスター宛のメッセージの場合、メッセージを転送するかどうかを返します

        Returns:
            bool: メッセージを転送する場合はTrue
        """
        return False

    def svrun(self, data_dir:Path, logger:logging.Logger, redis_cli:redis_client.RedisClient, msg:List[str],
              sessions:Dict[str, Dict[str, Any]]) -> int:
        """
        この機能のサーバー側の実行を行います

        Args:
            data_dir (Path): データディレクトリ
            logger (logging.Logger): ロガー
            redis_cli (redis_client.RedisClient): Redisクライアント
            msg (List[str]): 受信メッセージ
            sessions (Dict[str, Dict[str, Any]]): セッション情報
        
        Returns:
            int: 終了コード
        """
        svpath = convert.b64str2str(msg[2])
        formula_data_only = msg[3]=='True'
        sheet_name = convert.b64str2str(msg[4])
        sheet_name = None if sheet_name=='None' else sheet_name
        cell_name = json.loads(convert.b64str2str(msg[5]))
        cell_top_left = convert.b64str2str(msg[6])
        cell_bottom_right = convert.b64str2str(msg[7])
        output_detail_format = convert.b64str2str(msg[8])

        try:
            f = filer.Filer(data_dir, logger)
            chk, abspath, res = f._file_exists(svpath)
            if not chk:
                logger.warning(f"File not found. {svpath}")
                redis_cli.rpush(msg[1], res)
                return self.RESP_WARN
            res = self.get_cell_details(abspath, formula_data_only, sheet_name, cell_name, cell_top_left, cell_bottom_right,
                                        output_detail_format, logger)
            redis_cli.rpush(msg[1], res)
        except Exception as e:
            logger.warning(f"Failed to cell details: {e}", exc_info=True)
            redis_cli.rpush(msg[1], dict(warn=f"Failed to cell details: {e}"))
            return self.RESP_WARN
        return self.RESP_SUCCESS

    def get_cell_details(self, filepath:str, formula_data_only:bool, sheet_name:str, cell_name:List[str], cell_top_left:str, cell_bottom_right:str,
                          output_detail_format:str, logger:logging.Logger) -> Dict[str, Any]:
        """
        指定したワークブックの単一セルの値、データ型、スタイル、コメント、数式、ハイパーリンクなどの詳細を取得します。

        Args:
            filepath (str): ワークブックのパス
            formula_data_only (bool): 数式データのみを参照するかどうか。このオプションはキャッシュされたデータが存在する場合に有効です。
            sheet_name (str): 詳細情報を取得するシートの名前
            cell_name (List[str]): 詳細情報を取得するセルの名前のリスト。例えば、`A1`、`B2`、`R5987`。
            cell_top_left (str): 詳細情報を取得する範囲の左上セルの名前。例えば、`A1`、`B2`、`R5987`。
            cell_bottom_right (str): 詳細情報を取得する範囲の右下セルの名前。例えば、`A1`、`B2`、`R5987`。
            output_detail_format (str): 出力フォーマット。`json`または`text`。
            logger (logging.Logger): ロガー
        Returns:
            dict: セルの詳細情報
        """
        wb:Workbook = None
        try:
            from openpyxl.cell import Cell
            from openpyxl.workbook.workbook import Workbook
            from openpyxl.worksheet.worksheet import Worksheet
            from openpyxl.utils.datetime import from_excel
            import openpyxl

            wb:Workbook = openpyxl.load_workbook(filename=filepath, read_only=True, data_only=formula_data_only)
            if sheet_name not in wb.sheetnames:
                if len(wb.sheetnames) <= 0:
                    msg = dict(warn=f"There is no worksheet. filepath: {filepath}")
                    logger.warning(f"There is no worksheet. filepath: {filepath}")
                    return msg
                sheet_name = wb.sheetnames[0]

            def _proc(cellinfos:Dict[str,Any], sheet:Worksheet, cn:str):
                cell:Cell = sheet[cn]
                celltxt = f"Cell: {cn}\n" \
                        + f"  Value: {cell.value}\n" \
                        + f"  Data type: {self.OPENPYXL_TYPE_TO_STRING.get(cell.data_type, cell.data_type)}\n" \
                        + f"  Number format: {cell.number_format}\n"
                cellinfo = {
                    "Cell": cn,
                    "Value": cell.value,
                    "Data_type": self.OPENPYXL_TYPE_TO_STRING.get(cell.data_type, cell.data_type),
                    "Number_format": cell.number_format,
                }
                cellinfos[cn] = cellinfo

                if hasattr(cell, "style"):
                    cellinfo["Style"] = cell.style
                    celltxt += f"  Style: {cell.style}\n"

                # 日時データ
                if cell.is_date:
                    date_value: datetime | timedelta | time | None = from_excel(cell.value)
                    if isinstance(date_value, (datetime, time)):
                        cellinfo["Value_as_Date"] = date_value.isoformat()
                        celltxt += f"  Value as Date: {date_value.isoformat()}\n"
                    elif isinstance(date_value, timedelta):
                        cellinfo["Value_as_Time_Interval"] = date_value.total_seconds()
                        celltxt += f"  Value as Time Interval: {date_value.total_seconds()}\n"

                # 数式
                if cell.data_type == "f":
                    cellinfo["Formula"] = cell.value
                    celltxt += f"  Formula: {cell.value}\n"

                # ハイパーリンク
                if hasattr(cell, "hyperlink") and cell.hyperlink:
                    cellinfo["Hyperlink_Text"] = cell.hyperlink
                    celltxt += f"  Hyperlink: {cell.hyperlink}\n"
                    if hasattr(cell.hyperlink, "target"):
                        cellinfo["Hyperlink_Target"] = cell.hyperlink.target
                        celltxt += f"  Hyperlink Target: {cell.hyperlink.target}\n"
                    if hasattr(cell.hyperlink, "tooltip"):
                        cellinfo["Hyperlink_Tooltip"] = cell.hyperlink.tooltip
                        celltxt += f"  Hyperlink Tooltip: {cell.hyperlink.tooltip}\n"

                # コメント
                if hasattr(cell, "comment") and cell.comment:
                    cellinfo["Comment"] = cell.comment.text
                    cellinfo["Comment_Author"] = cell.comment.author
                    celltxt += f"  Comment: {cell.comment.text}\n"
                    celltxt += f"  Comment Author: {cell.comment.author}\n"

                # フォント
                font = cell.font
                cellinfo["Font_Name"] = font.name
                cellinfo["Font_Size"] = font.size
                cellinfo["Font_Bold"] = font.bold
                cellinfo["Font_Italic"] = font.italic
                cellinfo["Font_Underline"] = font.underline
                celltxt += f"  Font Name: {font.name}\n"
                celltxt += f"  Font Size: {font.size}\n"
                celltxt += f"  Font Bold: {font.bold}\n"
                celltxt += f"  Font Italic: {font.italic}\n"
                celltxt += f"  Font Underline: {font.underline}\n"
                if font.color:
                    if hasattr(font.color, "rgb") and font.color.rgb:
                        cellinfo["Font_Color_RGB"] = str(font.color.rgb)
                        celltxt += f"  Font Color (RGB): {font.color.rgb}\n"
                    elif hasattr(font.color, "theme") and font.color.theme is not None:
                        cellinfo["Font_Color_Theme"] = str(font.color.theme)
                        celltxt += f"  Font Color (Theme): {font.color.theme}\n"
                    else:
                        cellinfo["Font_Color"] = str(font.color)
                        celltxt += f"  Font Color: {font.color}\n"
                else:
                    cellinfo["Font_Color"] = "Default"
                    celltxt += f"  Font Color: Default\n"

                fill = cell.fill
                if hasattr(fill, "patternType") and fill.patternType:
                    cellinfo["Fill_Pattern_Type"] = fill.patternType
                    celltxt += f"  Fill Pattern Type: {fill.patternType}\n"

                    if hasattr(fill, "fgColor") and fill.fgColor:
                        if hasattr(fill.fgColor, "rgb") and fill.fgColor.rgb:
                            cellinfo["Fill_Foreground_Color_RGB"] = str(fill.fgColor.rgb)
                            celltxt += f"  Fill Foreground Color (RGB): {fill.fgColor.rgb}\n"
                        elif hasattr(fill.fgColor, "theme") and fill.fgColor.theme is not None:
                            cellinfo["Fill_Foreground_Color_Theme"] = str(fill.fgColor.theme)
                            celltxt += f"  Fill Foreground Color (Theme): {fill.fgColor.theme}\n"
                        else:
                            cellinfo["Fill_Foreground_Color"] = str(fill.fgColor)
                            celltxt += f"  Fill Foreground Color: {fill.fgColor}\n"

                    if hasattr(fill, "bgColor") and fill.bgColor:
                        if hasattr(fill.bgColor, "rgb") and fill.bgColor.rgb:
                            cellinfo["Fill_Background_Color_RGB"] = str(fill.bgColor.rgb)
                            celltxt += f"  Fill Background Color (RGB): {fill.bgColor.rgb}\n"
                        elif hasattr(fill.bgColor, "theme") and fill.bgColor.theme is not None:
                            cellinfo["Fill_Background_Color_Theme"] = str(fill.bgColor.theme)
                            celltxt += f"  Fill Background Color (Theme): {fill.bgColor.theme}\n"
                        else:
                            cellinfo["Fill_Background_Color"] = str(fill.bgColor)
                            celltxt += f"  Fill Background Color: {fill.bgColor}\n"
                else:
                    cellinfo["Fill"] = "No fill pattern"
                    celltxt += f"  Fill: No fill pattern\n"

                # 配置
                if hasattr(cell, "alignment") and cell.alignment:
                    alignment = cell.alignment
                    cellinfo["Horizontal_Alignment"] = alignment.horizontal
                    cellinfo["Vertical_Alignment"] = alignment.vertical
                    cellinfo["Text_Rotation"] = alignment.textRotation
                    cellinfo["Wrap_Text"] = alignment.wrapText
                    cellinfo["Indent"] = alignment.indent
                    cellinfo["Shrink_to_Fit"] = alignment.shrinkToFit
                    celltxt += f"  Horizontal Alignment: {alignment.horizontal}\n"
                    celltxt += f"  Vertical Alignment: {alignment.vertical}\n"
                    celltxt += f"  Text Rotation: {alignment.textRotation}\n"
                    celltxt += f"  Wrap Text: {alignment.wrapText}\n"
                    celltxt += f"  Indent: {alignment.indent}\n"
                    celltxt += f"  Shrink to Fit: {alignment.shrinkToFit}\n"

                # 罫線
                if hasattr(cell, "border") and cell.border:
                    if cell.border:
                        border_sides = {
                            "left": cell.border.left,
                            "right": cell.border.right,
                            "top": cell.border.top,
                            "bottom": cell.border.bottom,
                            "diagonal": cell.border.diagonal,
                        }

                        for side_name, side in border_sides.items():
                            b_name = f"Border {side_name.capitalize()}"
                            celltxt += f"  {b_name}:\n"
                            if side and side.style:
                                cellinfo[f"{b_name}_Style"] = side.style
                                celltxt += f"    Style: {side.style}\n"
                                if side.color:
                                    if hasattr(side.color, "rgb") and side.color.rgb:
                                        cellinfo[f"{b_name}_Color_RGB"] = str(side.color.rgb)
                                        celltxt += f"    Color (RGB): {side.color.rgb}\n"
                                    elif hasattr(side.color, "theme") and side.color.theme is not None:
                                        cellinfo[f"{b_name}_Color_Theme"] = str(side.color.theme)
                                        celltxt += f"    Color (Theme): {side.color.theme}\n"
                                    else:
                                        cellinfo[f"{b_name}_Color"] = str(side.color)
                                        celltxt += f"    Color: {side.color}\n"

                # セルの保護
                if hasattr(cell, "protection") and cell.protection:
                    if cell.protection:
                        cellinfo["Is_Cell_Locked"] = cell.protection.locked
                        cellinfo["Is_Cell_Hidden"] = cell.protection.hidden
                        celltxt += f"  Is Cell Locked: {cell.protection.locked}\n"
                        celltxt += f"  Is Cell Hidden: {cell.protection.hidden}\n"

                # 条件付き書式
                if hasattr(cell, "conditional_formatting") and cell.conditional_formatting:
                    cf_rules = []
                    for rule in sheet.conditional_formatting:
                        if f"P{cell.row}" in rule.cells.ranges:
                            cf_rules.append(rule)

                    if cf_rules:
                        cellinfo["Conditional_Formatting_Rules"] = []
                        for i, rule in enumerate(cf_rules):
                            rule_info = {"Rule_Index": i + 1, "SubRules": []}
                            cellinfo["Conditional_Formatting_Rules"].append(rule_info)
                            celltxt += f"  Conditional Formatting Rule {i + 1}:\n"
                            for subrule in rule.rules:
                                subrule_info = {"Type": type(subrule).__name__}
                                rule_info["SubRules"].append(subrule_info)
                                celltxt += f"    SubRule Type: {type(subrule).__name__}\n"
                                if hasattr(subrule, "formula"):
                                    subrule_info["Formula"] = subrule.formula
                                    celltxt += f"      Formula: {subrule.formula}\n"
                                if hasattr(subrule, "operator"):
                                    subrule_info["Operator"] = subrule.operator
                                    celltxt += f"      Operator: {subrule.operator}\n"
                                if hasattr(subrule, "dxf") and subrule.dxf:
                                    subrule_info["Differential_Style"] = {}
                                    celltxt += f"      Differential Style:\n"
                                    if hasattr(subrule.dxf, "font") and subrule.dxf.font:
                                        subrule_info["Differential_Style"]["Font"] = subrule.dxf.font
                                        celltxt += f"        Font: {subrule.dxf.font}\n"
                                    if hasattr(subrule.dxf, "fill") and subrule.dxf.fill:
                                        subrule_info["Differential_Style"]["Fill"] = subrule.dxf.fill
                                        celltxt += f"        Fill: {subrule.dxf.fill}\n"
                                    if hasattr(subrule.dxf, "border") and subrule.dxf.border:
                                        subrule_info["Differential_Style"]["Border"] = subrule.dxf.border
                                        celltxt += f"        Border: {subrule.dxf.border}\n"

                # merged cells
                if hasattr(sheet, "merged_cells") and sheet.merged_cells:
                    celltxt += f"  Merged Cells:\n"
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cellinfo["Cell_is_part_of_merged_range"] = {
                                "Merge_starts_at": {"min_row": merged_range.min_row, "min_col": merged_range.min_col},
                                "Merge_ends_at": {"max_row": merged_range.max_row, "max_col": merged_range.max_col}
                            }
                            celltxt += f"    Merge starts at: (Row: {merged_range.min_row}, Column: {merged_range.min_col})\n"
                            celltxt += f"    Merge ends at: (Row: {merged_range.max_row}, Column: {merged_range.max_col})\n"
                            break

                return celltxt

            cellinfos = {}
            celltxt = ""
            sheet:Worksheet = wb[sheet_name]
            if cell_name is not None and len(cell_name) > 0:
                for cn in cell_name:
                    celltxt += _proc(cellinfos, sheet, cn)

            if cell_top_left is not None and cell_bottom_right is not None:
                range_str = ":".join(sorted([cell_top_left, cell_bottom_right]))
                cell_range = sheet[range_str]
                for row in cell_range:
                    for cell in row:
                        if hasattr(cell, 'coordinate'):
                            celltxt += _proc(cellinfos, sheet, cell.coordinate)

            if (cell_name is None or len(cell_name) <= 0) \
                and cell_top_left is None and cell_bottom_right is None:
                for row in sheet.iter_rows():
                    for cell in row:
                        if hasattr(cell, 'coordinate'):
                            celltxt += _proc(cellinfos, sheet, cell.coordinate)

            res = dict(success={sheet_name:cellinfos if output_detail_format=='json' else celltxt})
            return res
        except Exception as e:
            msg = dict(warn=f"Failed to cell details: {e}")
            logger.warning(f"Failed to cell details: {e}", exc_info=True)
            return msg
        finally:
            if wb is not None:
                wb.close()
