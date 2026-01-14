from cmdbox.app import common, filer
from cmdbox.app.commons import convert, redis_client
from cmdbox.app.features.cli import excel_base
from cmdbox.app.options import Options
from pathlib import Path
from typing import Dict, Any, List, Tuple
import argparse
import logging
import json
import re


class ExcelCellSearch(excel_base.ExcelBase):
    def get_cmd(self):
        """
        この機能のコマンドを返します

        Returns:
            str: コマンド
        """
        return 'cell_search'
    
    def get_option(self):
        """
        この機能のオプションを返します

        Returns:
            Dict[str, Any]: オプション
        """
        opt = super().get_option()
        opt['description_ja'] = "データフォルダ配下のExcelファイルの指定したセルの値を検索します。"
        opt['description_en'] = "Searches for the value in the specified cell of an Excel file located in the data folder."
        opt['choice'] += [
            dict(opt="formula_data_only", type=Options.T_BOOL, default=False, required=True, multi=False, hide=False, choice=[True, False],
                 description_ja="数式データのみを参照するかどうかを指定します。このオプションはキャッシュされたデータが存在する場合に有効です。",
                 description_en="Specify whether to get only formula data. This option is valid if cached data exists."),
            dict(opt="sheet_name", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を取得するシートの名前を指定します。省略した場合、すべてのシートが使用されます。",
                 description_en="Specify the sheet name to get the cell value.If omitted, all sheets will be used."),
            dict(opt="cell_name", type=Options.T_STR, default=None, required=False, multi=True, hide=False, choice=None,
                 description_ja="セルの値を検索するセルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the cell name to search for the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="cell_top_left", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を検索する左上セルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the top-left cell name to search for the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="cell_bottom_right", type=Options.T_STR, default=None, required=False, multi=False, hide=False, choice=None,
                 description_ja="セルの値を検索する右下セルの名前を指定します。例えば、`A1`、`B2`、`R5987`。",
                 description_en="Specify the bottom-right cell name to search for the cell value. For example, `A1`, `B2`, `R5987`."),
            dict(opt="match_type", type=Options.T_STR, default="partial", required=True, multi=False, hide=False, choice=['full', 'partial', 'regex'],
                 description_ja="検索するセルの値に対するマッチ方法を指定します。`full`: 完全一致、`partial`: 部分一致、`regex`: 正規表現。",
                 description_en="Specifies the matching method for the value in the search cell. `full`: Exact match, `partial`: Partial match, `regex`: Regular expression."),
            dict(opt="search_value", type=Options.T_STR, default=None, required=True, multi=False, hide=False, choice=None,
                 description_ja="検索するセルの値を指定します。指定方法は `match_type` によって異なります。",
                 description_en="Specify the value to search for in the cell. The method of specification depends on `match_type`."),
            dict(opt="output_cell_format", type=Options.T_STR, default='json', required=False, multi=False, hide=False, choice=['json', 'csv', 'md', 'html'],
                 description_ja="出力フォーマットを指定します。例えば、`json`、`csv`、 `md`、 `html`。",
                 description_en="Specify the output format. For example, `json`, `csv`、 `md`、 `html`."),
        ]
        return opt

    def chk_args(self, args:argparse.Namespace, tm:float, pf:List[Dict[str, float]]=[]) -> Tuple[bool, str, Any]:
        """
        引数のチェックを行います

        Args:
            args (argparse.Namespace): 引数

        Returns:
            Tuple[bool, str]: チェック結果, メッセージ
        """
        if args.svname is None:
            msg = dict(warn=f"Please specify the --svname option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.scope is None:
            msg = dict(warn=f"Please specify the --scope option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.svpath is None:
            msg = dict(warn=f"Please specify the --svpath option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.match_type is None:
            msg = dict(warn=f"Please specify the --match_type option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        if args.search_value is None:
            msg = dict(warn=f"Please specify the --search_value option.")
            common.print_format(msg, args.format, tm, args.output_json, args.output_json_append, pf=pf)
            return self.RESP_WARN, msg, None
        return self.RESP_SUCCESS, None, None

    def excel_proc(self, abspath:Path, args:argparse.Namespace, logger:logging.Logger, tm:float, pf:List[Dict[str, float]]=[]) -> Dict[str, Any]:
        """
        Excel処理のベース

        Args:
            abspath (Path): Excelファイルの絶対パス
            args (argparse.Namespace): 引数
            logger (logging.Logger): ロガー
            tm (float): 処理時間
            pf (List[Dict[str, float]]): パフォーマンス情報

        Returns:
            Dict[str, Any]: 結果
        """
        res_json = self.cell_search(abspath, args.formula_data_only, args.sheet_name,
                                        args.cell_name, args.cell_top_left, args.cell_bottom_right,
                                        args.match_type, args.search_value, args.output_cell_format, logger)
        return res_json

    def get_svparam(self, args:argparse.Namespace) -> List[str]:
        """
        サーバーに送信するパラメーターを返します

        Args:
            args (argparse.Namespace): 引数

        Returns:
            List[str]: サーバーに送信するパラメーター
        """
        cell_name = json.dumps(args.cell_name, default=common.default_json_enc) if args.cell_name is not None else '[]'
        ret = [convert.str2b64str(str(args.svpath)), str(args.formula_data_only), convert.str2b64str(str(args.sheet_name)),
               convert.str2b64str(cell_name), convert.str2b64str(args.cell_top_left), convert.str2b64str(args.cell_bottom_right),
               convert.str2b64str(args.match_type), convert.str2b64str(args.search_value), convert.str2b64str(args.output_cell_format)]
        return ret

    def is_cluster_redirect(self):
        """
        クラスター宛のメッセージの場合、メッセージを転送するかどうかを返します

        Returns:
            bool: メッセージを転送する場合はTrue
        """
        return False

    def svrun(self, data_dir:Path, logger:logging.Logger, redis_cli:redis_client.RedisClient, msg:List[str],
              sessions:Dict[str, Dict[str, Any]]) -> int:
        """
        この機能のサーバー側の実行を行います

        Args:
            data_dir (Path): データディレクトリ
            logger (logging.Logger): ロガー
            redis_cli (redis_client.RedisClient): Redisクライアント
            msg (List[str]): 受信メッセージ
            sessions (Dict[str, Dict[str, Any]]): セッション情報
        
        Returns:
            int: 終了コード
        """
        svpath = convert.b64str2str(msg[2])
        formula_data_only = msg[3]=='True'
        sheet_name = convert.b64str2str(msg[4])
        sheet_name = None if sheet_name=='None' else sheet_name
        cell_name = json.loads(convert.b64str2str(msg[5]))
        cell_top_left = convert.b64str2str(msg[6])
        cell_bottom_right = convert.b64str2str(msg[7])
        match_type = convert.b64str2str(msg[8])
        search_value = convert.b64str2str(msg[9])
        output_cell_format = convert.b64str2str(msg[10])

        try:
            f = filer.Filer(data_dir, logger)
            chk, abspath, res = f._file_exists(svpath)
            if not chk:
                logger.warning(f"File not found. {svpath}")
                redis_cli.rpush(msg[1], res)
                return self.RESP_WARN
            res = self.cell_search(abspath, formula_data_only, sheet_name, cell_name, cell_top_left, cell_bottom_right,
                                   match_type, search_value, output_cell_format, logger)
            redis_cli.rpush(msg[1], res)
        except Exception as e:
            logger.warning(f"Failed to cell search: {e}", exc_info=True)
            redis_cli.rpush(msg[1], dict(warn=f"Failed to cell search: {e}"))
            return self.RESP_WARN
        return self.RESP_SUCCESS

    def cell_search(self, filepath:str, formula_data_only:bool, sheet_name:str, cell_name:List[str], cell_top_left:str, cell_bottom_right:str,
                        match_type:str, search_value:str, output_cell_format:str, logger:logging.Logger) -> Dict[str, Any]:
        """
        指定したワークブックのセルの値を検索します。

        Args:
            filepath (str): ワークブックのパス
            formula_data_only (bool): 数式データのみを参照するかどうか。このオプションはキャッシュされたデータが存在する場合に有効です。
            sheet_name (str): 詳細情報を取得するシートの名前
            cell_name (List[str]): 詳細情報を取得するセルの名前のリスト。例えば、`A1`、`B2`、`R5987`。
            cell_top_left (str): 詳細情報を取得する左上セルの名前。例えば、`A1`、`B2`、`R5987`。
            cell_bottom_right (str): 詳細情報を取得する右下セルの名前。 例えば、`A1`、`B2`、`R5987`。
            match_type (str): 検索するセルの値に対するマッチ方法。`full`: 完全一致、`partial`: 部分一致、`regex`: 正規表現。
            search_value (str): 検索するセルの値。
            output_cell_format (str): 出力フォーマット。例えば、`json`、`csv`、 `md`、 `html`。
            logger (logging.Logger): ロガー
        Returns:
            dict: セルの詳細情報
        """
        wb:Workbook = None
        try:
            from openpyxl.cell import Cell
            from openpyxl.workbook.workbook import Workbook
            from openpyxl.worksheet.worksheet import Worksheet
            import openpyxl

            wb:Workbook = openpyxl.load_workbook(filename=filepath, read_only=True, data_only=formula_data_only)
            if sheet_name is not None and sheet_name not in wb.sheetnames:
                msg = dict(warn=f"Sheet '{sheet_name}' does not exist in the workbook. filepath: {filepath}")
                logger.warning(f"Sheet '{sheet_name}' does not exist in the workbook. filepath: {filepath}")
                return msg

            def match_func(value:str) -> bool:
                if match_type == 'full':
                    return value == search_value
                elif match_type == 'partial':
                    return search_value in str(value)
                elif match_type == 'regex':
                    return re.search(search_value, str(value)) is not None
                return False

            def _proc(cellinfos:Dict[str, Any], sheet:Worksheet):
                cellinfo = {}
                cellinfos[sheet.title] = cellinfo
                celltxt = ""
                if cell_top_left is not None and cell_bottom_right is not None:
                    range_str = ":".join(sorted([cell_top_left, cell_bottom_right]))
                    cell_range = sheet[range_str]
                    for row in cell_range:
                        for cell in row:
                            if hasattr(cell, 'coordinate') and match_func(cell.value):
                                celltxt += self.format_cell(output_cell_format, celltxt, cell.value, logger)
                                cellinfo[cell.coordinate] = cell.value
                        celltxt = self.format_newline(output_cell_format, celltxt, logger)
                    celltxt = self.format_table(output_cell_format, celltxt, logger)

                if cell_name is not None and len(cell_name) > 0:
                    for cn in cell_name:
                        cell:Cell = sheet[cn]
                        if match_func(cell.value):
                            celltxt += self.format_cell(output_cell_format, celltxt, cell.value, logger)
                            cellinfo[cn] = cell.value
                    celltxt = self.format_newline(output_cell_format, celltxt, logger)
                    celltxt = self.format_table(output_cell_format, celltxt, logger)

                if (cell_name is None or len(cell_name) <= 0) \
                    and cell_top_left is None and cell_bottom_right is None:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if hasattr(cell, 'coordinate') and match_func(cell.value):
                                celltxt += self.format_cell(output_cell_format, celltxt, cell.value, logger)
                                cellinfo[cell.coordinate] = cell.value
                    celltxt = self.format_newline(output_cell_format, celltxt, logger)
                    celltxt = self.format_table(output_cell_format, celltxt, logger)
                return celltxt

            cellinfos = {}
            if sheet_name is not None:
                sheet:Worksheet = wb[sheet_name]
                celltxt = _proc(cellinfos, sheet)
                if output_cell_format!='json':
                    cellinfos = {sheet_name: celltxt}
            else:
                for sn in wb.sheetnames:
                    sheet:Worksheet = wb[sn]
                    celltxt = _proc(cellinfos, sheet)
                    if output_cell_format!='json':
                        cellinfos[sn] = celltxt

            res = dict(success=cellinfos)
            return res
        except Exception as e:
            msg = dict(warn=f"Failed to search cell search: {e}")
            logger.warning(f"Failed to search cell search: {e}", exc_info=True)
            return msg
        finally:
            if wb is not None:
                wb.close()
