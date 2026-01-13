"""
Document Capability - PDF, Excel, CSV, and document manipulation.

Provides capabilities for:
- PDF operations (read, create, merge, extract text/images)
- Excel/spreadsheet operations (read, write, manipulate)
- CSV operations (read, write, transform)
- Document conversion
"""

import asyncio
import csv
import io
import json
import os
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any

from .base import BaseCapability, CapabilityResult


class DocumentFormat(str, Enum):
    """Supported document formats."""

    PDF = "pdf"
    EXCEL = "excel"
    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"
    JSON = "json"
    MARKDOWN = "markdown"
    HTML = "html"
    TXT = "txt"


@dataclass
class DocumentConfig:
    """Configuration for document capability."""

    # PDF settings
    pdf_dpi: int = 150
    pdf_extract_images: bool = False

    # Excel settings
    excel_default_sheet: str = "Sheet1"

    # CSV settings
    csv_delimiter: str = ","
    csv_encoding: str = "utf-8"
    csv_quotechar: str = '"'

    # General settings
    temp_dir: str | None = None
    max_file_size_mb: int = 100


class DocumentCapability(BaseCapability):
    """
    Document processing capability for PDF, Excel, CSV, and more.

    Provides operations for:
    - Reading and extracting content from documents
    - Creating and writing documents
    - Converting between formats
    - Merging and splitting documents

    Example:
        capability = DocumentCapability()

        # Read PDF
        result = await capability.pdf_read("report.pdf")
        print(result.output["text"])

        # Read Excel
        result = await capability.excel_read("data.xlsx", sheet="Sales")
        print(result.output["data"])

        # Write CSV
        result = await capability.csv_write(
            "output.csv",
            data=[{"name": "Alice", "age": 30}]
        )
    """

    name = "document"
    description = "PDF, Excel, CSV, and document manipulation"

    def __init__(self, config: DocumentConfig | None = None):
        """Initialize document capability."""
        self.config = config or DocumentConfig()
        self._check_availability()

    def _check_availability(self) -> None:
        """Check which document libraries are available."""
        self._pypdf_available = False
        self._openpyxl_available = False
        self._pandas_available = False
        self._pdfplumber_available = False
        self._markdown_available = False

        try:
            import pypdf  # noqa: F401

            self._pypdf_available = True
        except ImportError:
            pass

        try:
            import openpyxl  # noqa: F401

            self._openpyxl_available = True
        except ImportError:
            pass

        try:
            import pandas  # noqa: F401

            self._pandas_available = True
        except ImportError:
            pass

        try:
            import pdfplumber  # noqa: F401

            self._pdfplumber_available = True
        except ImportError:
            pass

        try:
            import markdown  # noqa: F401

            self._markdown_available = True
        except ImportError:
            pass

    @property
    def is_available(self) -> bool:
        """Check if basic document operations are available."""
        # CSV is always available (stdlib)
        return True

    async def execute(
        self,
        operation: str,
        **kwargs: Any,
    ) -> CapabilityResult:
        """
        Execute a document operation.

        Args:
            operation: Operation to perform
            **kwargs: Operation-specific parameters

        Returns:
            CapabilityResult with operation output
        """
        operations = {
            # PDF
            "pdf_read": self._pdf_read,
            "pdf_create": self._pdf_create,
            "pdf_merge": self._pdf_merge,
            "pdf_split": self._pdf_split,
            "pdf_extract_images": self._pdf_extract_images,
            # Excel
            "excel_read": self._excel_read,
            "excel_write": self._excel_write,
            "excel_get_sheets": self._excel_get_sheets,
            # CSV
            "csv_read": self._csv_read,
            "csv_write": self._csv_write,
            # Conversion
            "convert": self._convert,
            # Markdown
            "markdown_to_html": self._markdown_to_html,
        }

        if operation not in operations:
            return CapabilityResult(
                success=False,
                output={"error": f"Unknown operation: {operation}"},
                error=f"Supported operations: {list(operations.keys())}",
            )

        try:
            result = await operations[operation](**kwargs)
            return CapabilityResult(success=True, output=result)
        except Exception as e:
            return CapabilityResult(
                success=False, output={"error": str(e)}, error=str(e)
            )

    # =========================================================================
    # PDF Operations
    # =========================================================================

    async def _pdf_read(
        self,
        path: str,
        pages: list[int] | None = None,
        extract_tables: bool = False,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Read text and metadata from PDF."""
        if self._pdfplumber_available:
            return await self._pdf_read_pdfplumber(path, pages, extract_tables)
        elif self._pypdf_available:
            return await self._pdf_read_pypdf(path, pages)
        else:
            raise RuntimeError(
                "No PDF library available. "
                "Install with: pip install pypdf or pip install pdfplumber"
            )

    async def _pdf_read_pypdf(
        self, path: str, pages: list[int] | None = None
    ) -> dict[str, Any]:
        """Read PDF using pypdf."""
        import pypdf

        loop = asyncio.get_event_loop()

        def read_pdf():
            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                metadata = reader.metadata
                text_pages = []

                page_indices = pages if pages else range(len(reader.pages))
                for i in page_indices:
                    if 0 <= i < len(reader.pages):
                        page = reader.pages[i]
                        text_pages.append(
                            {"page": i + 1, "text": page.extract_text() or ""}
                        )

                return {
                    "text": "\n\n".join(p["text"] for p in text_pages),
                    "pages": text_pages,
                    "page_count": len(reader.pages),
                    "metadata": {
                        "title": metadata.title if metadata else None,
                        "author": metadata.author if metadata else None,
                        "subject": metadata.subject if metadata else None,
                        "creator": metadata.creator if metadata else None,
                    },
                }

        return await loop.run_in_executor(None, read_pdf)

    async def _pdf_read_pdfplumber(
        self, path: str, pages: list[int] | None = None, extract_tables: bool = False
    ) -> dict[str, Any]:
        """Read PDF using pdfplumber (better table extraction)."""
        import pdfplumber

        loop = asyncio.get_event_loop()

        def read_pdf():
            with pdfplumber.open(path) as pdf:
                text_pages = []
                tables = []

                page_indices = pages if pages else range(len(pdf.pages))
                for i in page_indices:
                    if 0 <= i < len(pdf.pages):
                        page = pdf.pages[i]
                        text_pages.append(
                            {"page": i + 1, "text": page.extract_text() or ""}
                        )

                        if extract_tables:
                            page_tables = page.extract_tables()
                            for table in page_tables:
                                tables.append({"page": i + 1, "data": table})

                result = {
                    "text": "\n\n".join(p["text"] for p in text_pages),
                    "pages": text_pages,
                    "page_count": len(pdf.pages),
                    "metadata": pdf.metadata,
                }

                if extract_tables:
                    result["tables"] = tables

                return result

        return await loop.run_in_executor(None, read_pdf)

    async def _pdf_create(
        self,
        path: str,
        content: str | list[str],
        title: str | None = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Create a simple PDF from text content."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            raise RuntimeError(
                "reportlab not installed. Install with: pip install reportlab"
            )

        loop = asyncio.get_event_loop()

        def create_pdf():
            c = canvas.Canvas(path, pagesize=letter)
            width, height = letter

            if title:
                c.setTitle(title)

            # Simple text rendering
            if isinstance(content, str):
                pages = [content]
            else:
                pages = content

            for page_content in pages:
                y = height - 50
                for line in page_content.split("\n"):
                    if y < 50:
                        c.showPage()
                        y = height - 50
                    c.drawString(50, y, line[:100])  # Truncate long lines
                    y -= 15
                c.showPage()

            c.save()
            return {"success": True, "path": path, "pages": len(pages)}

        return await loop.run_in_executor(None, create_pdf)

    async def _pdf_merge(
        self, paths: list[str], output: str, **kwargs: Any
    ) -> dict[str, Any]:
        """Merge multiple PDFs into one."""
        if not self._pypdf_available:
            raise RuntimeError("pypdf not installed. Install with: pip install pypdf")

        import pypdf

        loop = asyncio.get_event_loop()

        def merge_pdfs():
            merger = pypdf.PdfMerger()
            for pdf_path in paths:
                merger.append(pdf_path)
            merger.write(output)
            merger.close()
            return {"success": True, "path": output, "merged_count": len(paths)}

        return await loop.run_in_executor(None, merge_pdfs)

    async def _pdf_split(
        self, path: str, output_dir: str, pages_per_file: int = 1, **kwargs: Any
    ) -> dict[str, Any]:
        """Split PDF into multiple files."""
        if not self._pypdf_available:
            raise RuntimeError("pypdf not installed. Install with: pip install pypdf")

        import pypdf

        loop = asyncio.get_event_loop()

        def split_pdf():
            os.makedirs(output_dir, exist_ok=True)
            output_files = []

            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                total_pages = len(reader.pages)

                for start in range(0, total_pages, pages_per_file):
                    end = min(start + pages_per_file, total_pages)
                    writer = pypdf.PdfWriter()

                    for i in range(start, end):
                        writer.add_page(reader.pages[i])

                    output_path = os.path.join(
                        output_dir, f"split_{start + 1}-{end}.pdf"
                    )
                    with open(output_path, "wb") as out:
                        writer.write(out)
                    output_files.append(output_path)

            return {
                "success": True,
                "files": output_files,
                "total_pages": total_pages,
            }

        return await loop.run_in_executor(None, split_pdf)

    async def _pdf_extract_images(
        self, path: str, output_dir: str, **kwargs: Any
    ) -> dict[str, Any]:
        """Extract images from PDF."""
        if not self._pypdf_available:
            raise RuntimeError("pypdf not installed. Install with: pip install pypdf")

        import pypdf

        loop = asyncio.get_event_loop()

        def extract_images():
            os.makedirs(output_dir, exist_ok=True)
            extracted = []

            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                for page_num, page in enumerate(reader.pages):
                    for img_num, image in enumerate(page.images):
                        output_path = os.path.join(
                            output_dir, f"page{page_num + 1}_img{img_num + 1}.png"
                        )
                        with open(output_path, "wb") as img_file:
                            img_file.write(image.data)
                        extracted.append(output_path)

            return {"success": True, "images": extracted, "count": len(extracted)}

        return await loop.run_in_executor(None, extract_images)

    # =========================================================================
    # Excel Operations
    # =========================================================================

    async def _excel_read(
        self,
        path: str,
        sheet: str | None = None,
        header_row: int = 0,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Read data from Excel file."""
        if self._pandas_available:
            return await self._excel_read_pandas(path, sheet, header_row)
        elif self._openpyxl_available:
            return await self._excel_read_openpyxl(path, sheet)
        else:
            raise RuntimeError(
                "No Excel library available. "
                "Install with: pip install openpyxl or pip install pandas openpyxl"
            )

    async def _excel_read_pandas(
        self, path: str, sheet: str | None, header_row: int
    ) -> dict[str, Any]:
        """Read Excel using pandas."""
        import pandas as pd

        loop = asyncio.get_event_loop()

        def read_excel():
            df = pd.read_excel(path, sheet_name=sheet, header=header_row)
            return {
                "data": df.to_dict(orient="records"),
                "columns": list(df.columns),
                "rows": len(df),
                "sheet": sheet or "Sheet1",
            }

        return await loop.run_in_executor(None, read_excel)

    async def _excel_read_openpyxl(
        self, path: str, sheet: str | None
    ) -> dict[str, Any]:
        """Read Excel using openpyxl."""
        import openpyxl

        loop = asyncio.get_event_loop()

        def read_excel():
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if sheet else wb.active

            data = []
            headers = []

            for row_num, row in enumerate(ws.iter_rows(values_only=True)):
                if row_num == 0:
                    headers = [str(cell) if cell else f"col_{i}" for i, cell in enumerate(row)]
                else:
                    data.append(dict(zip(headers, row)))

            wb.close()
            return {
                "data": data,
                "columns": headers,
                "rows": len(data),
                "sheet": ws.title,
            }

        return await loop.run_in_executor(None, read_excel)

    async def _excel_write(
        self,
        path: str,
        data: list[dict[str, Any]],
        sheet: str | None = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Write data to Excel file."""
        if self._pandas_available:
            import pandas as pd

            loop = asyncio.get_event_loop()

            def write_excel():
                df = pd.DataFrame(data)
                df.to_excel(
                    path, sheet_name=sheet or "Sheet1", index=False
                )
                return {"success": True, "path": path, "rows": len(data)}

            return await loop.run_in_executor(None, write_excel)
        elif self._openpyxl_available:
            import openpyxl

            loop = asyncio.get_event_loop()

            def write_excel():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet or "Sheet1"

                if data:
                    # Write headers
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                    # Write data
                    for row_num, row_data in enumerate(data, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row_num, column=col, value=row_data.get(header))

                wb.save(path)
                return {"success": True, "path": path, "rows": len(data)}

            return await loop.run_in_executor(None, write_excel)
        else:
            raise RuntimeError(
                "No Excel library available. "
                "Install with: pip install openpyxl"
            )

    async def _excel_get_sheets(self, path: str, **kwargs: Any) -> dict[str, Any]:
        """Get list of sheets in Excel file."""
        if not self._openpyxl_available:
            raise RuntimeError(
                "openpyxl not installed. Install with: pip install openpyxl"
            )

        import openpyxl

        loop = asyncio.get_event_loop()

        def get_sheets():
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return {"sheets": sheets, "count": len(sheets)}

        return await loop.run_in_executor(None, get_sheets)

    # =========================================================================
    # CSV Operations
    # =========================================================================

    async def _csv_read(
        self,
        path: str,
        delimiter: str | None = None,
        encoding: str | None = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Read data from CSV file."""
        delimiter = delimiter or self.config.csv_delimiter
        encoding = encoding or self.config.csv_encoding

        loop = asyncio.get_event_loop()

        def read_csv():
            data = []
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = reader.fieldnames or []
                for row in reader:
                    data.append(dict(row))

            return {
                "data": data,
                "columns": list(headers),
                "rows": len(data),
            }

        return await loop.run_in_executor(None, read_csv)

    async def _csv_write(
        self,
        path: str,
        data: list[dict[str, Any]],
        delimiter: str | None = None,
        encoding: str | None = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Write data to CSV file."""
        delimiter = delimiter or self.config.csv_delimiter
        encoding = encoding or self.config.csv_encoding

        loop = asyncio.get_event_loop()

        def write_csv():
            if not data:
                with open(path, "w", encoding=encoding, newline="") as f:
                    f.write("")
                return {"success": True, "path": path, "rows": 0}

            headers = list(data[0].keys())
            with open(path, "w", encoding=encoding, newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(data)

            return {"success": True, "path": path, "rows": len(data)}

        return await loop.run_in_executor(None, write_csv)

    # =========================================================================
    # Conversion
    # =========================================================================

    async def _convert(
        self,
        input_path: str,
        output_path: str,
        from_format: str | None = None,
        to_format: str | None = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Convert between document formats."""
        # Detect formats from extensions if not provided
        if not from_format:
            from_format = Path(input_path).suffix.lstrip(".").lower()
        if not to_format:
            to_format = Path(output_path).suffix.lstrip(".").lower()

        # CSV to JSON
        if from_format == "csv" and to_format == "json":
            result = await self._csv_read(input_path)
            with open(output_path, "w") as f:
                json.dump(result["data"], f, indent=2)
            return {"success": True, "from": from_format, "to": to_format}

        # JSON to CSV
        if from_format == "json" and to_format == "csv":
            with open(input_path, "r") as f:
                data = json.load(f)
            if isinstance(data, list):
                await self._csv_write(output_path, data)
            else:
                await self._csv_write(output_path, [data])
            return {"success": True, "from": from_format, "to": to_format}

        # Excel to CSV
        if from_format in ("xlsx", "xls", "excel") and to_format == "csv":
            result = await self._excel_read(input_path)
            await self._csv_write(output_path, result["data"])
            return {"success": True, "from": from_format, "to": to_format}

        # CSV to Excel
        if from_format == "csv" and to_format in ("xlsx", "excel"):
            result = await self._csv_read(input_path)
            await self._excel_write(output_path, result["data"])
            return {"success": True, "from": from_format, "to": to_format}

        # Markdown to HTML
        if from_format in ("md", "markdown") and to_format == "html":
            with open(input_path, "r") as f:
                md_content = f.read()
            result = await self._markdown_to_html(content=md_content)
            with open(output_path, "w") as f:
                f.write(result["html"])
            return {"success": True, "from": from_format, "to": to_format}

        raise ValueError(f"Conversion from {from_format} to {to_format} not supported")

    async def _markdown_to_html(
        self, content: str, **kwargs: Any
    ) -> dict[str, Any]:
        """Convert Markdown to HTML."""
        if self._markdown_available:
            import markdown

            html = markdown.markdown(
                content, extensions=["tables", "fenced_code", "codehilite"]
            )
            return {"html": html}
        else:
            # Basic conversion without library
            import re

            html = content
            # Headers
            html = re.sub(r"^### (.+)$", r"<h3>\1</h3>", html, flags=re.MULTILINE)
            html = re.sub(r"^## (.+)$", r"<h2>\1</h2>", html, flags=re.MULTILINE)
            html = re.sub(r"^# (.+)$", r"<h1>\1</h1>", html, flags=re.MULTILINE)
            # Bold and italic
            html = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", html)
            html = re.sub(r"\*(.+?)\*", r"<em>\1</em>", html)
            # Code
            html = re.sub(r"`(.+?)`", r"<code>\1</code>", html)
            # Paragraphs
            html = re.sub(r"\n\n", r"</p><p>", html)
            html = f"<p>{html}</p>"

            return {"html": html}

    # =========================================================================
    # Convenience Methods
    # =========================================================================

    async def pdf_read(
        self, path: str, pages: list[int] | None = None, extract_tables: bool = False
    ) -> CapabilityResult:
        """Read text and metadata from PDF."""
        return await self.execute(
            "pdf_read", path=path, pages=pages, extract_tables=extract_tables
        )

    async def pdf_create(
        self, path: str, content: str | list[str], title: str | None = None
    ) -> CapabilityResult:
        """Create a PDF from text content."""
        return await self.execute("pdf_create", path=path, content=content, title=title)

    async def pdf_merge(self, paths: list[str], output: str) -> CapabilityResult:
        """Merge multiple PDFs into one."""
        return await self.execute("pdf_merge", paths=paths, output=output)

    async def pdf_split(
        self, path: str, output_dir: str, pages_per_file: int = 1
    ) -> CapabilityResult:
        """Split PDF into multiple files."""
        return await self.execute(
            "pdf_split", path=path, output_dir=output_dir, pages_per_file=pages_per_file
        )

    async def excel_read(
        self, path: str, sheet: str | None = None, header_row: int = 0
    ) -> CapabilityResult:
        """Read data from Excel file."""
        return await self.execute(
            "excel_read", path=path, sheet=sheet, header_row=header_row
        )

    async def excel_write(
        self, path: str, data: list[dict[str, Any]], sheet: str | None = None
    ) -> CapabilityResult:
        """Write data to Excel file."""
        return await self.execute("excel_write", path=path, data=data, sheet=sheet)

    async def csv_read(
        self, path: str, delimiter: str | None = None, encoding: str | None = None
    ) -> CapabilityResult:
        """Read data from CSV file."""
        return await self.execute(
            "csv_read", path=path, delimiter=delimiter, encoding=encoding
        )

    async def csv_write(
        self,
        path: str,
        data: list[dict[str, Any]],
        delimiter: str | None = None,
        encoding: str | None = None,
    ) -> CapabilityResult:
        """Write data to CSV file."""
        return await self.execute(
            "csv_write", path=path, data=data, delimiter=delimiter, encoding=encoding
        )

    async def convert(
        self,
        input_path: str,
        output_path: str,
        from_format: str | None = None,
        to_format: str | None = None,
    ) -> CapabilityResult:
        """Convert between document formats."""
        return await self.execute(
            "convert",
            input_path=input_path,
            output_path=output_path,
            from_format=from_format,
            to_format=to_format,
        )

    async def markdown_to_html(self, content: str) -> CapabilityResult:
        """Convert Markdown to HTML."""
        return await self.execute("markdown_to_html", content=content)

    async def run(self, operation: str, **kwargs: Any) -> CapabilityResult:
        """Run a document operation (alias for execute)."""
        return await self.execute(operation, **kwargs)
