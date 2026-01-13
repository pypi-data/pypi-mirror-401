from io import BytesIO
from os import PathLike
from typing import BinaryIO
import openpyxl
from pathlib import Path

from openpyxl.workbook import Workbook


class ExcelHandler:

    @classmethod
    def read_from_file(cls, file_path: str | PathLike, fields_map: dict, work_sheet_name: str = "", row_offset: int = 0) -> list:
        """
        从excel文件读取excel内容
        :param file_path: 文件路径
        :param fields_map: excel表头 -> 字段名映射
        :param work_sheet_name: excel工作表名称
        :param row_offset: 表头所在行
        :return:
        """
        if isinstance(file_path, str):
            file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(file_path)
        return cls._read_excel(file_path, fields_map, work_sheet_name, row_offset)

    @classmethod
    def read_from_object(cls, file_object: BinaryIO, fields_map: dict, work_sheet_name: str = "", row_offset: int = 0) -> list:
        """
        从文件对象读取excel内容
        :param file_object: 文件对象
        :param fields_map: excel表头 -> 字段名映射
        :param work_sheet_name: excel工作表名称
        :param row_offset: 表头所在行
        :return:
        """
        return cls._read_excel(file_object, fields_map, work_sheet_name, row_offset)

    @staticmethod
    def _read_excel(file_object: BinaryIO | PathLike, fields_map: dict, work_sheet_name, row_offset: int = 0):
        wb = openpyxl.load_workbook(file_object,read_only=True)
        if not work_sheet_name:
            ws = wb.active
        else:
            ws = wb[work_sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= row_offset:
            return []

        headers = rows[row_offset]
        # 只保留出现在 fields_map 中的表头及其索引
        selected_columns: list[tuple[int, str]] = []
        for idx, h in enumerate(headers):
            if h in fields_map:
                selected_columns.append((idx, fields_map[h]))

        result: list[dict] = []
        for row in rows[row_offset + 1:]:
            row_dict: dict = {}
            for idx, mapped_name in selected_columns:
                value = row[idx] if idx < len(row) else None
                row_dict[mapped_name] = value
            result.append(row_dict)
        return result

    @classmethod
    def write_to_file(cls, data: list[dict], file_path: str | PathLike, reverse_fields_map: dict) -> bool:
        """
        将数据写入excel文件
        :param data: 数据列表
        :param file_path: 文件路径
        :param reverse_fields_map: 字段名 -> excel表头映射
        :return:
        """
        if isinstance(file_path, str):
            file_path = Path(file_path)
        if file_path.exists():
            raise FileExistsError(file_path)
        wb = cls._write_excel(data, reverse_fields_map)
        wb.save(file_path)
        if file_path.exists():
            return True
        return False

    @classmethod
    def write_to_object(cls, data: list[dict], reverse_fields_map: dict) -> bytes:
        """
        将数据写入文件对象
        :param data: 数据列表
        :param reverse_fields_map: 字段名 -> excel表头映射
        :return:
        """
        buffer = BytesIO()
        wb = cls._write_excel(data, reverse_fields_map)
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _write_excel(data: list[dict], reverse_fields_map: dict) -> Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active

        # 用反向映射生成表头
        headers = [reverse_fields_map.get(k, k) for k in data[0]]
        ws.append(headers)

        for item in data:
            ws.append([item.get(k, "") for k in data[0].keys()])

        return wb