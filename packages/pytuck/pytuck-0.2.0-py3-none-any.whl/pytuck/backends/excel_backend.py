"""
Pytuck Excel存储引擎

使用单个Excel工作簿（.xlsx），每个表一个工作表，可视化友好
"""

import json
import os
import base64
from typing import Any, Dict, TYPE_CHECKING
from datetime import datetime
from .base import StorageBackend
from ..common.exceptions import SerializationError
from .versions import get_format_version

from ..common.options import ExcelBackendOptions

if TYPE_CHECKING:
    from ..core.storage import Table
    from openpyxl import Workbook


class ExcelBackend(StorageBackend):
    """Excel format storage engine (requires openpyxl)"""

    ENGINE_NAME = 'excel'
    REQUIRED_DEPENDENCIES = ['openpyxl']
    FORMAT_VERSION = get_format_version('excel')

    def __init__(self, file_path: str, options: ExcelBackendOptions):
        """
        初始化 Excel 后端

        Args:
            file_path: Excel 文件路径
            options: Excel 后端配置选项
        """
        super().__init__(file_path, options)

    def save(self, tables: Dict[str, 'Table']) -> None:
        """保存所有表数据到Excel工作簿"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        temp_path = self.file_path + '.tmp'
        try:
            wb = Workbook()
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 创建元数据工作表
            metadata_sheet = wb.create_sheet('_metadata', 0)
            metadata_sheet.append(['Key', 'Value'])
            metadata_sheet.append(['format_version', self.FORMAT_VERSION])
            metadata_sheet.append(['timestamp', datetime.now().isoformat()])
            metadata_sheet.append(['table_count', len(tables)])

            # 创建统一的表结构工作表 _pytuck_tables
            tables_sheet = wb.create_sheet('_pytuck_tables', 1)
            tables_sheet.append(['table_name', 'primary_key', 'next_id', 'comment', 'columns'])
            for table_name, table in tables.items():
                columns_json = json.dumps([
                    {
                        'name': col.name,
                        'type': col.col_type.__name__,
                        'nullable': col.nullable,
                        'primary_key': col.primary_key,
                        'index': col.index,
                        'comment': col.comment
                    }
                    for col in table.columns.values()
                ])
                tables_sheet.append([table_name, table.primary_key, table.next_id, table.comment or '', columns_json])

            # 为每个表创建数据工作表
            for table_name, table in tables.items():
                self._save_table_to_workbook(wb, table_name, table)

            # 原子性保存
            wb.save(temp_path)

            if os.path.exists(self.file_path):
                os.remove(self.file_path)
            os.rename(temp_path, self.file_path)

        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise SerializationError(f"Failed to save Excel file: {e}")

    def load(self) -> Dict[str, 'Table']:
        """从Excel工作簿加载所有表数据"""
        if not self.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        try:
            wb = load_workbook(self.file_path)

            # 从 _pytuck_tables 工作表读取所有表的 schema
            tables_schema: Dict[str, Dict[str, Any]] = {}
            if '_pytuck_tables' in wb.sheetnames:
                tables_sheet = wb['_pytuck_tables']
                rows = list(tables_sheet.iter_rows(min_row=2, values_only=True))
                for row in rows:
                    if row[0]:  # table_name 不为空
                        table_name = row[0]
                        tables_schema[table_name] = {
                            'primary_key': row[1],
                            'next_id': int(row[2]) if row[2] else 1,
                            'comment': row[3] if row[3] else None,
                            'columns': json.loads(row[4]) if row[4] else []
                        }

            # 获取所有数据表名（排除元数据表）
            table_names = [
                name for name in wb.sheetnames
                if not name.startswith('_')
            ]

            tables = {}
            for table_name in table_names:
                schema = tables_schema.get(table_name, {})
                table = self._load_table_from_workbook(wb, table_name, schema)
                tables[table_name] = table

            return tables

        except Exception as e:
            raise SerializationError(f"Failed to load Excel file: {e}")

    def exists(self) -> bool:
        """检查文件是否存在"""
        return os.path.exists(self.file_path)

    def delete(self) -> None:
        """删除文件"""
        if self.exists():
            os.remove(self.file_path)

    def _save_table_to_workbook(self, wb: 'Workbook', table_name: str, table: 'Table') -> None:
        """保存单个表的数据到工作簿"""
        # 数据工作表
        data_sheet = wb.create_sheet(table_name)

        # 写入表头
        columns = list(table.columns.keys())
        data_sheet.append(columns)

        # 写入数据行
        for record in table.data.values():
            row = []
            for col_name in columns:
                value = record.get(col_name)
                # 处理特殊类型
                if isinstance(value, bytes):
                    value = base64.b64encode(value).decode('ascii')
                elif value is None:
                    value = ''
                elif isinstance(value, bool):
                    # Excel会将bool自动转换，这里用字符串明确表示
                    value = 'TRUE' if value else 'FALSE'
                row.append(value)
            data_sheet.append(row)

    def _load_table_from_workbook(
        self, wb: 'Workbook', table_name: str, schema: Dict[str, Any]
    ) -> 'Table':
        """从工作簿加载单个表"""
        from ..core.storage import Table
        from ..core.orm import Column

        primary_key = schema.get('primary_key', 'id')
        next_id = schema.get('next_id', 1)
        table_comment = schema.get('comment')
        columns_data = schema.get('columns', [])

        # 重建列
        columns = []
        type_map = {
            'int': int,
            'str': str,
            'float': float,
            'bool': bool,
            'bytes': bytes
        }

        for col_data in columns_data:
            col_type = type_map.get(col_data['type'], str)
            column = Column(
                col_data['name'],
                col_type,
                nullable=col_data['nullable'],
                primary_key=col_data['primary_key'],
                index=col_data.get('index', False),
                comment=col_data.get('comment')
            )
            columns.append(column)

        # 创建表
        table = Table(table_name, columns, primary_key, comment=table_comment)
        table.next_id = next_id

        # 读取数据
        data_sheet = wb[table_name]
        rows = list(data_sheet.iter_rows(values_only=True))

        if len(rows) > 1:
            headers = rows[0]
            for row_data in rows[1:]:
                record = {}
                for col_name, value in zip(headers, row_data):
                    if col_name not in table.columns:
                        continue

                    # 反序列化特殊类型
                    column = table.columns[col_name]

                    if value == '' or value is None:
                        value = None
                    elif column.col_type == bytes and value:
                        value = base64.b64decode(value)
                    elif column.col_type == bool:
                        # 处理Excel的bool值
                        if isinstance(value, bool):
                            pass  # 保持原样
                        elif isinstance(value, str):
                            value = (value.upper() == 'TRUE')
                        else:
                            value = bool(value)
                    elif column.col_type == int and value is not None:
                        value = int(value)
                    elif column.col_type == float and value is not None:
                        value = float(value)

                    record[col_name] = value

                pk = record[primary_key]
                table.data[pk] = record

        # 重建索引
        for col_name, column in table.columns.items():
            if column.index:
                if col_name in table.indexes:
                    del table.indexes[col_name]
                table.build_index(col_name)

        return table

    def get_metadata(self) -> Dict[str, Any]:
        """获取元数据"""
        if not self.exists():
            return {}

        try:
            file_size = os.path.getsize(self.file_path)
            modified_time = os.path.getmtime(self.file_path)

            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True)

            metadata = {
                'engine': 'excel',
                'file_size': file_size,
                'modified': modified_time
            }

            # 尝试读取元数据工作表
            if '_metadata' in wb.sheetnames:
                sheet = wb['_metadata']
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        metadata[row[0]] = row[1]

            wb.close()
            return metadata

        except:
            return {}
