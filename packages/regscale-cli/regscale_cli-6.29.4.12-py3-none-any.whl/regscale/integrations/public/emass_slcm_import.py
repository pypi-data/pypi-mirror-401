#!/usr/bin/python
"""Script to parse a .xlsx file and load the inventory into RegScale as assets"""

import os
from typing import Optional
from urllib.parse import urljoin

import click
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.progress import track

from regscale.core.app.logz import create_logger
from regscale.models import (
    SecurityPlan,
    Catalog,
    ControlImplementation,
    ControlImplementationStatus,
    CustomField,
    CustomFieldsData,
)

NOT_APPLICABLE = "Not Applicable"
logger = create_logger()


class ControlImportMeta:
    def __init__(self, control_id: str, import_row: int):
        self.control_id = control_id
        self.import_row = import_row
        self.implementation_id = 0
        self.parent_id = 0


def import_emass_slcm_file(file_name: click.Path, regscale_id: int, catalogue_id: int, tenant_id: int) -> None:
    """
    This function imports a SLCM export created by an eMASS system.

    :param click.Path file_name:
    :param int regscale_id:
    :param int catalogue_id:
    :param int tenant_id:
    :return None:
    :rtype None:
    """
    if not verify_sp_id(regscale_id):
        return logger.error("Not a valid Security Plan ID. Aborting!")
    if not verify_cat_id(catalogue_id):
        return logger.error("Not a valid Catalog ID. Aborting!")
    if not os.path.isfile(file_name):
        return logger.error("No file name/path provided to import. Aborting!")
    custom_fields = check_for_custom_fields(tenant_id)
    wb_obj = load_workbook(file_name)
    ws = wb_obj.active
    control_list = build_list_of_controls(ws)
    control_list = get_or_initialize_control_implementation(control_list, regscale_id, catalogue_id)
    identify_controls_to_remove(control_list, regscale_id)
    for cur_control_index in track(
        range(len(control_list)), description="Loading SLCM Control Implementations updates..."
    ):
        cur_control = control_list[cur_control_index]
        process_row(ws, cur_control, regscale_id, custom_fields)


def build_list_of_controls(ws: Worksheet) -> list:
    """
    Builds a list of controls to import, so any missing can be created.

    :param Worksheet ws:
    :return list:
    :rtype list:
    """
    control_list = []
    row = 7
    control_id = ws.cell(row=row, column=1).value
    while control_id is not None:
        control_list.append(ControlImportMeta(control_id, row))
        row += 1
        control_id = ws.cell(row=row, column=1).value
    return control_list


def identify_controls_to_remove(control_list: list, regscale_id: int) -> None:
    """
    Identifies controls that are in the security plan but not in the import, and designate them for deletion.

    :param list control_list:
    :param int regscale_id:
    :return None:
    :rtype None:
    """
    current_implementations = get_all_current_control_implementation(regscale_id)
    if current_implementations is not None:
        for cur_control_index in track(
            range(len(current_implementations)), description="Identifying SLCM Control Implementations to remove..."
        ):
            cur_implementation = current_implementations[cur_control_index]
            verify_control_implementation_to_remove(cur_implementation, control_list)


def verify_control_implementation_to_remove(control: dict, control_list: list) -> None:
    """
    Verifies whether a control implementation is in the import, and if not, deletes it from the SP.

    :param dict control:
    :param list control_list:
    :return None:
    :rtype None:
    """
    control_id = control["controlName"]
    if find_control_in_metadata_list(control_list, control_id) is None:
        remove_control_implementation(control)


def remove_control_implementation(control: dict) -> None:
    """
    Deletes the control implementation from the security plan.

    :param dict control:
    :return None:
    :rtype None:
    """
    from regscale.core.app.api import Api

    api = Api()
    control_id = control["id"]
    conmon_url = urljoin(api.config["domain"], f"/api/controlImplementation/{control_id}")
    conmon_response = api.delete(conmon_url)
    if not conmon_response.ok:
        logger.warning("Unable to delete control implementation %s", control_id)


def get_or_initialize_control_implementation(control_list: list, regscale_id: int, catalogue_id: int) -> list:
    """
    Retrieves all control implementations associated with the security plan.

    :param list control_list:
    :param int regscale_id:
    :param int catalogue_id:
    :return list:
    :rtype list:
    """
    current_implementations = get_all_current_control_implementation(regscale_id)
    if current_implementations is not None:
        control_list = match_control_rows(control_list, current_implementations)
    return create_control_implementations(control_list, regscale_id, catalogue_id)


def get_all_current_control_implementation(regscale_id: int) -> Optional[list]:
    """
    Retrieves the current set of control implementations associated with the security plan.

    :param int regscale_id:
    :return list:
    :rtype list:
    """
    from regscale.core.app.api import Api

    api = Api()
    # Using the REST API as the Security Control name is dropped by the model
    conmon_url = urljoin(api.config["domain"], f"/api/controlImplementation/getAllByPlan/{regscale_id}")
    conmon_response = api.get(conmon_url)
    if conmon_response.ok:
        try:
            cil_json = conmon_response.json()
            if len(cil_json) > 0:
                return cil_json
        except Exception:
            return None
    return None


def create_control_implementations(control_list: list, regscale_id: int, catalogue_id: int) -> list:
    """
    Retrieves all security controls associated with the catalog.

    :param list control_list:
    :param int regscale_id:
    :param int catalogue_id:
    :return list:
    :rtype list:
    """
    from regscale.core.app.api import Api

    api = Api()
    conmon_url = urljoin(api.config["domain"], f"/api/SecurityControls/getList/{catalogue_id}")
    conmon_response = api.get(conmon_url)
    if conmon_response.ok:
        try:
            scl_json = conmon_response.json()
            if len(scl_json) > 0:
                control_list = match_control_parent_rows(control_list, scl_json, regscale_id)
        except Exception:
            return control_list
    return control_list


def match_control_parent_rows(control_list: list, sec_control_list: list, regscale_id: int) -> list:
    """
    Matches the control IDs from the SLCM import and the list of controls associated with the security plan.

    :param list control_list:
    :param list sec_control_list:
    :param int regscale_id:
    :param int catalogue_id:
    :return list:
    :rtype list:
    """
    rtn_list = []
    for cur_control_index in track(range(len(control_list)), description="Creating missing control implementations..."):
        cur_meta = control_list[cur_control_index]
        if cur_meta.implementation_id == 0:
            sec_ctl = find_control_in_list(sec_control_list, cur_meta.control_id, "controlId")
            if sec_ctl is not None:
                cur_meta.parent_id = sec_ctl["id"]
                cur_meta.implementation_id = create_control_implementation(sec_ctl, regscale_id)
        rtn_list.append(cur_meta)
    return rtn_list


def create_control_implementation(sec_control: dict, regscale_id: int) -> int:
    """
    Creates a control implementation that is in the import but currently doesn't exist in the security plan.

    :param sec_control:
    :param regscale_id:
    :return:
    """
    from regscale.core.app.api import Api

    config = Api().config
    updated_control_imp = sec_control
    ctrl_imp = ControlImplementation.create_new_control_implementation(
        sec_control, regscale_id, "securityplans", ControlImplementationStatus.NotImplemented, config["userId"]
    )
    try:
        if isinstance(ctrl_imp, dict):
            updated_control_imp = ctrl_imp
        return save_new_control_implementation(updated_control_imp)
    except Exception as ex:
        logger.error(f"Unable to save new control implementation - {ex}")
    return 0


def save_new_control_implementation(control_imp: dict) -> int:
    """
    Saves a new control implementation

    :param dict control_imp:
    :return int:
    :rtype int:
    """
    from regscale.core.app.api import Api

    api = Api()
    # Using the REST API as the model throws errors when some non-required data elements aren't populated
    conmon_url = urljoin(api.config["domain"], "/api/controlImplementation/")
    conmon_response = api.post(conmon_url, json=control_imp)
    if conmon_response.ok:
        scl_json = conmon_response.json()
        return scl_json["id"]
    else:
        logger.error(conmon_response)
    return 0


def check_for_custom_fields(tenant_id: int) -> list:
    """
    Retrieves the custom fields that some SLCM values are populated into.

    :param int tenant_id:
    :return list:
    :rtype list:
    """
    from regscale.core.app.api import Api

    api = Api()
    cfl = CustomField.get_list_by_module_id(15)
    if len(cfl) == 0:
        conmon_url = urljoin(api.config["domain"], f"/api/customFields/module/{tenant_id}/15")
        conmon_response = api.get(conmon_url)
        if conmon_response.ok:
            cfl_json = conmon_response.json()
            return cfl_json
        else:
            logger.error(conmon_response)
        return []
    cfl_list = []
    for cur_cfl in cfl:
        cur_json = cur_cfl.dict()
        cfl_list.append(cur_json)
    return cfl_list


def match_control_rows(control_list: list, control_imp_list: list) -> list:
    """
    Matches the controls in the import with those that exist in the SP.

    :param list control_list:
    :param list control_imp_list:
    :return list:
    :rtype list:
    """
    rtn_list = []
    for cur_meta in control_list:
        ctl_imp = find_control_in_list(control_imp_list, cur_meta.control_id, "controlName")
        if ctl_imp is not None:
            cur_meta.implementation_id = ctl_imp["id"]
            cur_meta.parent_id = ctl_imp["controlID"]
        rtn_list.append(cur_meta)
    return rtn_list


def find_control_in_list(control_imp_list: list, control_id: str, data_field: str) -> Optional[dict]:
    """
    Finds the control from the import in the list of existing controls

    :param list control_imp_list:
    :param str control_id:
    :param str data_field:
    :return dict:
    :rtype dict:
    """
    for cur_imp in control_imp_list:
        if cur_imp[data_field].upper() == control_id.upper():
            return cur_imp
    return None


def find_control_in_metadata_list(control_imp_list: list, control_id: str) -> Optional[dict]:
    """
    Finds the control from the import in the list of existing controls

    :param list control_imp_list:
    :param str control_id:
    :param str data_field:
    :return dict:
    :rtype dict:
    """
    for cur_imp in control_imp_list:
        if cur_imp.control_id.upper() == control_id.upper():
            return cur_imp
    return None


def process_row(ws: Worksheet, control: ControlImportMeta, regscale_id: int, custom_fields: list) -> None:
    """
    Processes the current row in the import.

    :param Worksheet ws:
    :param ControlImportMeta control:
    :param int regscale_id:
    :param list custom_fields:
    :return None:
    :rtype None:
    """
    control_imp = get_control_implementation(control.implementation_id)
    if control_imp is None:
        return
    # Column B (2) - Control Title - Security Control (Parent Control) Title (should be from Catalog)
    # Column C (3) - Control Information - Security Control Description (should be from Catalog)
    # Column D (4) - Compliance Status - Last Assessment Result
    status = ws.cell(row=control.import_row, column=4).value
    control_imp.lastAssessmentResult = "Fail"
    if status == "Compliant":
        control_imp.lastAssessmentResult = "Pass"
    if status == NOT_APPLICABLE:
        control_imp.status = NOT_APPLICABLE
    save_control_implementation_update(control_imp)
    process_implementation(ws, control, control_imp, custom_fields)
    process_slcm(ws, control, control_imp, custom_fields)
    process_risk_assessment(ws, control, regscale_id, control_imp)


def save_control_implementation_update(control_imp: ControlImplementation) -> None:
    """
    Saves the current control implementation

    :param ControlImplementation control_imp:
    :return None:
    :rtype None:
    """
    from regscale.core.app.api import Api

    api = Api()
    try:
        control_imp.save(False)
    except Exception:
        control_imp_payload = control_imp.dict()
        conmon_url = urljoin(api.config["domain"], f"/api/controlImplementation/{control_imp.id}")
        conmon_response = api.put(conmon_url, json=control_imp_payload)
        if conmon_response.ok:
            scl_json = conmon_response.json()
            if scl_json is None:
                logger.error("Could not update")
        else:
            logger.error(conmon_response)


def get_control_implementation(control_id: int) -> ControlImplementation:
    """
    Retrieves the current control implementation

    :param int control_id:
    :return ControlImplementation:
    :rtype ControlImplementation:
    """
    # Using the model throws errors about missing data elements
    from regscale.core.app.api import Api

    api = Api()
    conmon_url = urljoin(api.config["domain"], f"/api/controlImplementation/{control_id}")
    conmon_response = api.get(conmon_url)
    if conmon_response.ok:
        try:
            ctl_imp_json = conmon_response.json()
            return ControlImplementation(**ctl_imp_json)
        except Exception:
            return None
    return None


def process_implementation(
    ws: Worksheet, control: ControlImportMeta, control_imp: ControlImplementation, custom_fields: list
) -> None:
    """
    Processes the implementation columns in the import.

    :param Worksheet ws:
    :param ControlImportMeta control:
    :param ControlImplementation control_imp:
    :param list custom_fields:
    :return None:
    :rtype None:
    """
    # Column E (5) - Implementation Status - Control and Compliance Status
    control_imp.status = slcm_status(ws.cell(row=control.import_row, column=5).value)
    # Column F (6) - Common Control Provider - Responsibility - Custom Field (Common Control Provider)
    # control_imp.responsibility = slcm_responsibility(ws.cell(row=control.import_row, column=6).value)
    set_custom_field_data(
        custom_fields, control_imp.id, "Common Control Provider", ws.cell(row=control.import_row, column=6).value
    )
    # Column G (7) - Security Control Designation - Responsibility
    control_imp.responsibility = slcm_scdesignation(
        ws.cell(row=control.import_row, column=7).value, control_imp.responsibility
    )
    # Column H (8) - Test Method - Control Test Plans - Ignore - Should load the default tests and CCIs
    # Column I (9) - N/A Justification - Exclusion Justification
    control_imp.exclusionJustification = ws.cell(row=control.import_row, column=9).value
    # Column J (10) - Estimated Completion Date - Planned Implementation Date
    control_imp.plannedImplementationDate = ws.cell(row=control.import_row, column=10).value
    # Column K (11) - Implementation Narrative - Steps to Implement
    control_imp.stepsToImplement = ws.cell(row=control.import_row, column=11).value
    # Column L (12) - Responsible Entities - Customer/Cloud Implementation
    entities = ws.cell(row=control.import_row, column=12).value
    if control_imp.responsibility == "Customer":
        control_imp.customerImplementation = entities
    elif control_imp.responsibility == "Provider":
        control_imp.cloudImplementation = entities
    else:
        control_imp.customerImplementation = entities
        control_imp.cloudImplementation = entities
    # Column M (13) - Application Layer - Custom Field (Application Layer)
    set_custom_field_data(
        custom_fields, control_imp.id, "Application Layer", ws.cell(row=control.import_row, column=13).value
    )
    # Column N (14) - Database Layer - Custom Field (Database Layer)
    set_custom_field_data(
        custom_fields, control_imp.id, "Database Layer", ws.cell(row=control.import_row, column=14).value
    )
    # Column O (15) - Operating System Layer - Custom Field (Operating System Layer)
    set_custom_field_data(
        custom_fields, control_imp.id, "Operating System Layer", ws.cell(row=control.import_row, column=15).value
    )
    save_control_implementation_update(control_imp)


def set_custom_field_data(custom_fields: list, control_id: int, field_name: str, field_value: str) -> None:
    """
    Creates a custom field data object

    :param list custom_fields:
    :param int control_id:
    :param str field_name:
    :param str field_value:
    :return None:
    :rtype None:
    """
    from regscale.core.app.api import Api

    api = Api()
    if (field_value is None) or (len(field_value) == 0):
        return
    custom_field_rec = find_custom_field(custom_fields, field_name)
    if custom_field_rec is None:
        return
    cfdata = CustomFieldsData()
    cfdata.moduleId = custom_field_rec["moduleId"]
    cfdata.fieldId = custom_field_rec["id"]
    cfdata.fieldName = custom_field_rec["fieldName"]
    cfdata.fieldDataType = custom_field_rec["fieldDataType"]
    cfdata.fieldValue = field_value
    cfdata.parentId = control_id
    cfdata_payload = cfdata.dict()
    conmon_url = urljoin(api.config["domain"], "/api/customFieldsData/")
    conmon_response = api.post(conmon_url, json=cfdata_payload)
    if not conmon_response.ok:
        logger.error(conmon_response)


def find_custom_field(custom_fields: list, field_name: str) -> Optional[dict]:
    """
    Find the appropriate custom field in the list of existing custom fields

    :param list custom_fields:
    :param str field_name:
    :return dict:
    :rtype dict:
    """
    for cur_custom_field in custom_fields:
        if cur_custom_field["fieldName"] == field_name:
            return cur_custom_field
    return None


def slcm_status(status: str) -> str:
    """
    Convert the status from the import to the appropriate RS value

    :param str status:
    :return str:
    :rtype str:
    """
    if status == "Implemented":
        return "Fully Implemented"
    if (status == "Not Implemented") or (status == "Planned") or (status == "Inherited"):
        return status
    return NOT_APPLICABLE


def slcm_responsibility(responsibility: str) -> str:
    """
    Convert the responsibility from the import to the appropriate RS value

    :param str responsibility:
    :return str:
    :rtype str:
    """
    if responsibility == "DoD":
        return "Customer"
    if responsibility == "Component":
        return "Provider"
    if responsibility == "Enclave":
        return "Shared"


def slcm_scdesignation(scdesignation: str, responsibility: str) -> str:
    """
    Convert the SLCM designation from the import to the appropriate RS value

    :param str scdesignation:
    :param str responsibility:
    :return str:
    :rtype str:
    """
    if scdesignation == "Hybrid":
        return "Hybrid"
    return responsibility


def process_slcm(
    ws: Worksheet, control: ControlImportMeta, control_imp: ControlImplementation, custom_fields: list
) -> None:
    """
    Process the SLCM columns of the import

    :param Worksheet ws:
    :param ControlImportMeta control:
    :param ControlImplementation control_imp:
    :param list custom_fields:
    :return None:
    :rtype None:
    """
    # Column Q (17) - Criticality - Ignore - Leaving blank in the export
    # Column R (18) - Frequency
    control_imp.assessmentFrequency = slcm_frequency_calculation(ws.cell(row=control.import_row, column=18).value)
    # Column S (19) - Method - Custom Field (Method)
    set_custom_field_data(custom_fields, control_imp.id, "Method", ws.cell(row=control.import_row, column=19).value)
    # Column T (20) - Reporting - Ignore - Statement built based on whether control has CCI's and ConMon
    # Column U (21) - Tracking - Ignore - Statement built based on whether control has CCI's
    # Column V (22) - SLCM Comments - Assessments
    save_control_implementation_update(control_imp)


def slcm_frequency_calculation(frequency: str) -> int:
    """
    Convert the SLCM frequency to a RS frequency in days.

    :param str frequency:
    :return int:
    :rtype int:
    """
    if frequency == "Every Three Years":
        return 365 * 3
    if frequency == "Every Two Years":
        return 365 * 2
    if frequency == "Annually":
        return 365
    if frequency == "Semi-annually":
        return 163
    if frequency == "Quarterly":
        return 90
    if frequency == "Monthly":
        return 30
    if frequency == "Weekly":
        return 7
    return 1


def process_risk_assessment(
    ws: Worksheet, control: ControlImportMeta, regscale_id: int, control_imp: ControlImplementation
) -> None:
    """
    Process the Risk Assessment columns of the import - These are calculated values in the export,
    and thus are not included in the import (at this time).

    :param Worksheet ws:
    :param ControlImportMeta control:
    :param int regscale_id:
    :param ControlImplementation control_imp:
    :return None:
    :rtype None:
    """
    # This whole section goes to Mitigations and Risks - Ignore - These don't have the key information needed to create
    # Column X (24) - Severity - Risk - Consequence
    # Column Y (25) - Relevance of Threat - Risk - Probability
    # Column Z (26) - Likelihood - Calculated from Consequence and Probability - Leave for now
    # Column AA (27) - Impact - Risk - Spread across all Risk fields
    # Column AB (28) - Residual Risk Level - Calculated - Leave for now
    # Column AC (29) - Vulnerability Summary - Risk - RiskStatement
    # Column AD (30) - Mitigations - Risk - Mitigation
    # Column AE (31) - Impact Description - Risk - ImpactDescription
    # Column AF (32) - Recommendations - RiskTreatments (tied to Risk.id)
    if (ws is None) or (control is None) or (regscale_id <= 0) or (control_imp is None):
        logger.warning("Processing Risk Assessment - No data to import")


def verify_sp_id(regscale_id: int) -> bool:
    """
    Verify the provided SP ID is valid

    :param int regscale_id:
    :return bool:
    :rtype bool:
    """
    from regscale.core.app.api import Api

    api = Api()
    sp = SecurityPlan.get_object(regscale_id)
    if sp is None:
        logger.warning("Model didn't return any security plan, trying again with the REST API")
        conmon_url = urljoin(api.config["domain"], f"/api/securityplans/{regscale_id}")
        conmon_response = api.get(conmon_url)
        if conmon_response.ok:
            try:
                sec_plan = conmon_response.json()
                if sec_plan is None:
                    return False
            except Exception:
                return False
    return True


def verify_cat_id(catalogue_id: int) -> bool:
    """
    Verify the provided catalog ID is valid

    :param int catalogue_id:
    :return bool:
    :rtype bool:
    """
    from regscale.core.app.api import Api

    api = Api()
    sp = Catalog.get_object(catalogue_id)
    if sp is None:
        logger.warning("Model didn't return any catalog, trying again with the REST API")
        conmon_url = urljoin(api.config["domain"], f"/api/catalog/{catalogue_id}")
        conmon_response = api.get(conmon_url)
        if conmon_response.ok:
            try:
                sec_plan = conmon_response.json()
                if sec_plan is None:
                    return False
            except Exception:
                return False
    return True
