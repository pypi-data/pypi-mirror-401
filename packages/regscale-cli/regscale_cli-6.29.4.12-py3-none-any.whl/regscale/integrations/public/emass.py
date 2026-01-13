#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""eMASS integration to the CLI to allow support for eMASS documents"""

# standard python imports
import os
from typing import Any

import click
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from pathlib import Path

from regscale.core.app.api import Api
from regscale.integrations.public.emass_slcm_import import import_emass_slcm_file
from regscale.core.app.utils.app_utils import (
    check_file_path,
    create_progress_object,
    error_and_exit,
    get_current_datetime,
    get_file_type,
    reformat_str_date,
)
from regscale.models import regscale_id

SKIP_ROWS: int = 7
COLUMNS = ["M", "N", "O", "P"]


@click.group()
def emass():
    """Performs bulk processing of eMASS files (Upload trusted data only)."""


@emass.command("populate_controls")
@click.option(
    "--file_name",
    type=click.Path(exists=True, dir_okay=False, file_okay=True),
    required=True,
    prompt="Enter the full file path of the eMASS controls document.",
    help="Enter the full file path of the eMASS controls document to populate with RegScale data.",
)
@regscale_id(help="Enter the desired SSP ID # from RegScale.")
def populate_workbook(file_name: click.Path, regscale_id: int) -> None:
    """
    [BETA] Populate controls from a System Security Plan in RegScale into an eMASS formatted excel workbook.
    """
    populate_emass_workbook(file_name=file_name, regscale_id=regscale_id)


@emass.command("import_slcm")
@click.option(
    "--file_name",
    type=click.Path(exists=True, dir_okay=False, file_okay=True),
    required=True,
    prompt="Enter the full file path of the eMASS controls document.",
    help="Enter the full file path of the eMASS controls document to populate with RegScale data.",
)
@regscale_id(help="Enter the desired SSP ID # from RegScale.")
@click.option(
    "--catalogue_id",
    "-c",
    type=click.INT,
    help="The RegScale ID # of the catalogue to use for controls in the profile.",
    required=True,
)
@click.option(
    "--tenant_id",
    "-t",
    type=click.INT,
    help="The RegScale ID # of the tenant to use for this security plan.",
    required=True,
)
def import_slcm(file_name: click.Path, regscale_id: int, catalogue_id: int, tenant_id: int) -> None:
    """
    [BETA] Populate controls from a System Security Plan in RegScale into an eMASS formatted excel workbook.
    """
    import_emass_slcm_file(file_name=file_name, regscale_id=regscale_id, catalogue_id=catalogue_id, tenant_id=tenant_id)


def fetch_template_from_blob() -> None:
    """
    Fetch a template for the eMASS controls document

    :rtype: None
    """
    api = Api()

    # check if the artifacts folder exists
    check_file_path("artifacts")

    # get the template from the API
    template = api.get(
        url="https://regscaleblob.blob.core.windows.net/blob/eMASS_Control_Template.xlsx",
        headers={},
    )

    # write the template to a file
    with open(f".{os.sep}artifacts{os.sep}eMASS_Template.xlsx", "wb") as f:
        f.write(template.content)
    api.logger.info(f"Template saved to .{os.sep}artifacts{os.sep}eMASS_Template.xlsx")


def populate_emass_workbook(file_name: Path, regscale_id: int) -> None:
    """
    Function to populate an eMASS workbook with control assessments from RegScale

    :param Path file_name: Path to the eMASS control workbook
    :param int regscale_id: ID of the SSP in RegScale to get the controls & assessments from
    :rtype: None
    """
    # make sure the user gave a path to an Excel workbook
    if get_file_type(file_name) not in [".xlsx", ".xls"]:
        error_and_exit("Please provide a file path to an Excel workbook in .xlsx or .xls format.")

    # convert file_name to a Path object
    file_name = Path(file_name)

    # initialize the Application and API classes
    api = Api()

    # populate the controls in the Excel workbook
    output_name = populate_assessment_results(file_name=file_name, ssp_id=regscale_id, api=api)
    api.logger.info("Please open %s and verify the data before uploading into eMASS.", output_name)


def map_ccis(file_data_dict: dict, file_name: str) -> dict:
    """
    Function to map each cci and its row number in the eMASS workbook

    :param dict file_data_dict: Dictionary of an Excel file column
    :param str file_name: Name of the file file_data_dict is from
    :return: dictionary of ccis and their row numbers
    :rtype: dict
    """
    # convert the control names to match RegScale control names
    try:
        formatted_ccis = {
            # create a dictionary with the key as the control name and the value as the row number
            # cci has a prefix of CCI- and must be a 6-digit number
            f"CCI-{val:06}": {
                "cci": f"CCI-{val:06}",
                "row": key + SKIP_ROWS,
            }
            for key, val in file_data_dict["CCI"].items()
        }
    except KeyError:
        error_and_exit(
            f"{file_name} doesn't match the expected eMASS format.\nPlease view an example "
            "template here: https://regscale.readme.io/docs/emass-beta#template"
        )

    return formatted_ccis or {}


def fetch_assessments_and_controls(ssp_id: int, api: Api) -> list:
    """
    Fetch assessments and controls from RegScale

    :param int ssp_id: SSP ID from RegScale
    :param Api api: API Object
    :return: List of controls with assessments
    :rtype: list
    """
    # create the GraphQL query
    query = f"""
            query {{
              controls:controlImplementations(
                take: 50
                skip: 0
                where: {{
                  parentId: {{ eq: {ssp_id} }}
              parentModule: {{ eq: "securityplans" }}
              assessments: {{ any: true }}
            }}
          ) {{
            items {{
              id
              control {{
                controlId
                cci {{
                  name
                  description
                }}
              }}
              assessments {{
                id
                actualFinish
                assessmentResult
                summaryOfResults
                leadAssessor {{
                  firstName
                  lastName
                }}
              }}
            }}
            totalCount
            pageInfo {{
              hasNextPage
            }}
          }}
        }}
        """

    # get the data from GraphQL
    response = api.graph(query=query)

    # try to get the items from the GraphQL response
    try:
        controls = response["controls"]["items"]
    except KeyError:
        controls = []

    total_controls = api.get(
        f"{api.config['domain']}/api/controlImplementation/getCountByParent/{ssp_id}/securityplans"
    )

    if not total_controls.ok:
        error_and_exit(f"Received unexpected response: {total_controls.status_code}\n{total_controls.text}")

    if controls:
        api.logger.info(
            "Received %s/%s controls with Assessments. Total control count for SSP #%s in RegScale: %s.",
            len(controls),
            response["controls"]["totalCount"],
            ssp_id,
            total_controls.text,
        )
        for control in controls:
            try:
                control["ccis"] = [cci["name"] for cci in control["control"]["cci"]]
            except (KeyError, TypeError):
                control["ccis"] = []
    else:
        error_and_exit(
            "The RegScale SSP provided has no assessments associated with the controls. "
            + "Please add assessments to the controls and try again."
        )
    return controls


def populate_assessment_results(file_name: Path, ssp_id: int, api: Api) -> Path:
    """
    Populate assessment results from a System Security Plan in RegScale into an eMASS formatted excel workbook

    :param Path file_name: path to the Excel workbook to populate with assessments from SSP
    :param int ssp_id: ID for a System Security Plan from RegScale
    :param Api api: API Object
    :return: Path to output file
    :rtype: Path
    """
    import pandas as pd  # Optimize import performance

    author = "RegScale CLI"
    job_progress = create_progress_object()
    logger = api.logger
    controls = fetch_assessments_and_controls(ssp_id=ssp_id, api=api)
    # load the Excel file in pandas to find row # to update the data
    file_data = pd.read_excel(file_name, skiprows=SKIP_ROWS - 2)

    # load the workbook using openpyxl to retain worksheet styling
    wb = load_workbook(file_name)

    # set the sheet to the first sheet in the provided workbook
    sheet = wb.active

    # convert to a dictionary
    file_data_dict = file_data.to_dict()

    # format the controls
    cci_mappings = map_ccis(file_data_dict=file_data_dict, file_name=file_name.name)

    # create variable to count number of rows updated and skipped
    update_counter: int = 0
    skipped_counter: int = 0

    # create a dictionary of all ccis and their assessments
    regscale_cci_assessments = {
        cci: {"assessment": ctrl["assessments"][0]}
        for ctrl in controls
        for cci in ctrl.get("ccis", [])
        if "ccis" in ctrl
    }

    # create comment & fill attribute for columns with missing data
    comment = Comment(
        text=f"SSP #{ssp_id} doesn't contain an assessment associated with this control.",
        author=author,
        height=150,
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    with job_progress:
        populating_controls = job_progress.add_task(
            f"[#21a5bb]Analyzing controls in {file_name.name}...",
            total=len(cci_mappings),
        )
        # iterate through the ccis
        for cci in cci_mappings.values():
            row_number = cci["row"]
            cci_id = cci["cci"]

            # see if the cci has an assessment by comparing it to our GraphQL query results
            if cci_id not in regscale_cci_assessments or not regscale_cci_assessments[cci_id].get("assessment"):
                # increment the skip counter
                skipped_counter += 1

                # highlight and add a comment
                for column in COLUMNS:
                    sheet[f"{column}{row_number}"].comment = comment
                    sheet[f"{column}{row_number}"].fill = yellow_fill
                job_progress.update(populating_controls, advance=1)
                continue
            # get the assessment for the cci
            assessment = regscale_cci_assessments[cci_id]["assessment"]

            map_finish_date(
                assessment=assessment,
                sheet=sheet,
                row_number=row_number,
                author=author,
                yellow_fill=yellow_fill,
            )

            sheet[f"O{row_number}"] = (
                f'{assessment["leadAssessor"]["firstName"]} {assessment["leadAssessor"]["lastName"]}'
            )
            if assessment["summaryOfResults"]:
                sheet[f"P{row_number}"] = assessment["summaryOfResults"]
            else:
                sheet[f"P{row_number}"].comment = Comment(
                    text=f"RegScale assessment #{assessment['id']} doesn't have any information in Summary of Results.",
                    author=author,
                    height=150,
                )
                sheet[f"P{row_number}"].fill = yellow_fill

            # update the counter
            update_counter += 1
            # update the progress bar
            job_progress.update(populating_controls, advance=1)

    # add the date and time to the output filename
    output_name = Path(
        os.path.join(
            file_name.parent,
            file_name.stem + get_current_datetime("_Updated_%Y%m%d_%H%M%S") + file_name.suffix,
        )
    )

    # save the updated workbook
    wb.save(output_name)

    logger.info(
        "%s has been created with %i update(s). %i row(s) were skipped because of missing controls in SSP #%i.",
        output_name.name,
        update_counter,
        skipped_counter,
        ssp_id,
    )
    # return the output path
    return output_name


def determine_assessment_result(assessment: dict) -> str:
    """
    Determine the assessment result based on the assessment data

    :param dict assessment: Assessment data
    :return: Assessment result
    :rtype: str
    """
    if assessment["assessmentResult"] == "Pass":
        return "Compliant"
    elif assessment["assessmentResult"] in ["Fail", "Partial Pass"]:
        return "Non-Compliant"
    else:
        return "Not Applicable"


def map_finish_date(
    assessment: dict,
    sheet: Any,
    row_number: int,
    author: str,
    yellow_fill: PatternFill,
) -> None:
    """
    Map the finish date of the assessment

    :param dict assessment: Assessment data
    :param Any sheet: Excel sheet object
    :param int row_number: Row number
    :param str author: Author of the comment
    :param PatternFill yellow_fill: Yellow fill object
    :rtype: None
    """
    finish_date = reformat_str_date(assessment["actualFinish"], "%d-%b-%Y") if assessment["actualFinish"] else None

    # map the control to the Excel spreadsheet
    sheet[f"M{row_number}"] = determine_assessment_result(assessment)
    if finish_date:
        sheet[f"N{row_number}"] = finish_date
    else:
        sheet[f"N{row_number}"].comment = Comment(
            text=f"Assessment #{assessment['id']} in RegScale doesn't have a finish date.",
            author=author,
            height=150,
        )
        sheet[f"N{row_number}"].fill = yellow_fill
