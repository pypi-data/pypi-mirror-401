#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""FedRAMP Scanner Integration"""

import logging
import re
from typing import Iterator, List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    import numpy

import polars as pl
from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import column_index_from_string  # type: ignore
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
from openpyxl.workbook import Workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet

from regscale.core.app.utils.app_utils import error_and_exit, get_current_datetime
from regscale.core.utils.date import date_str
from regscale.integrations.scanner.base import BaseScannerIntegration
from regscale.integrations.scanner.models import IntegrationAsset, IntegrationFinding
from regscale.integrations.scanner_integration import issue_due_date
from regscale.models import ImportValidater, IssueSeverity, Mapping, regscale_models
from regscale.validation.address import validate_ip_address, validate_mac_address

logger = logging.getLogger("regscale")

# Column name constants to avoid duplication
POAM_ID = "POAM ID"
WEAKNESS_NAME = "Weakness Name"
WEAKNESS_DESCRIPTION = "Weakness Description"
WEAKNESS_DETECTOR_SOURCE = "Weakness Detector Source"
WEAKNESS_SOURCE_IDENTIFIER = "Weakness Source Identifier"
ASSET_IDENTIFIER = "Asset Identifier"
POINT_OF_CONTACT = "Point of Contact"
RESOURCES_REQUIRED = "Resources Required"
OVERALL_REMEDIATION_PLAN = "Overall Remediation Plan"
ORIGINAL_DETECTION_DATE = "Original Detection Date"
SCHEDULED_COMPLETION_DATE = "Scheduled Completion Date"
PLANNED_MILESTONES = "Planned Milestones"
MILESTONE_CHANGES = "Milestone Changes"
STATUS_DATE = "Status Date"
ORIGINAL_RISK_RATING = "Original Risk Rating"
ADJUSTED_RISK_RATING = "Adjusted Risk Rating"
RISK_ADJUSTMENT = "Risk Adjustment"
FALSE_POSITIVE = "False Positive"
OPERATIONAL_REQUIREMENT = "Operational Requirement"
DEVIATION_RATIONALE = "Deviation Rationale"
FILE_PATH_ERROR = "File path is required."

# Pre-computed severity mapping for _determine_category_fast (avoids recreating per call)
SEVERITY_MAPPING = {
    "medium": IssueSeverity.Moderate.name,
    "high": IssueSeverity.High.name,
    "critical": IssueSeverity.High.name,
    "low": IssueSeverity.Low.name,
}
# Pre-computed lowercase severity member names (avoids list comprehension per call)
SEVERITY_MEMBERS_LOWER = frozenset(mem.lower() for mem in IssueSeverity.__members__)


class FedrampPoamIntegration(BaseScannerIntegration):
    """Integration class for FedRAMP POAM scanning using efficient batch processing."""

    # Keys set in the `set_keys` method of `ScannerIntegration`
    title = "FedRAMP"
    create_vulnerabilities = False  # POAM imports create issues only, not vulnerabilities
    file_path: str = ""
    poam_sheets: List[str] = []
    validators: dict = {}
    workbook: Optional[Workbook] = None

    asset_identifier_field = "otherTrackingNumber"
    finding_severity_map = {
        "Low": regscale_models.IssueSeverity.Low,
        "Moderate": regscale_models.IssueSeverity.Moderate,
        "High": regscale_models.IssueSeverity.High,
        "Critical": regscale_models.IssueSeverity.Critical,
    }
    poam_id_header = POAM_ID
    blank_records: int = 0
    blank_threshold: int = 3
    error_records: int = 0
    skipped_records: int = 0
    processed_assets: set[str] = set()  # Track processed assets across all methods
    # Pre-computed column indices for fast array access (populated per sheet)
    _column_indices: dict = {}

    fedramp_poam_columns = [
        POAM_ID,
        WEAKNESS_NAME,
        WEAKNESS_DESCRIPTION,
        WEAKNESS_DETECTOR_SOURCE,
        WEAKNESS_SOURCE_IDENTIFIER,
        ASSET_IDENTIFIER,
        POINT_OF_CONTACT,
        RESOURCES_REQUIRED,
        OVERALL_REMEDIATION_PLAN,
        ORIGINAL_DETECTION_DATE,
        SCHEDULED_COMPLETION_DATE,
        PLANNED_MILESTONES,
        MILESTONE_CHANGES,
        STATUS_DATE,
        ORIGINAL_RISK_RATING,
        ADJUSTED_RISK_RATING,
        RISK_ADJUSTMENT,
        FALSE_POSITIVE,
        OPERATIONAL_REQUIREMENT,
        DEVIATION_RATIONALE,
        "Comments",
    ]

    """
    Unused columns:
    # "Vendor Dependency",
    # "Last Vendor Check-in Date",
    # "Vendor Dependent Product Name",
    # "Supporting Documents",
    # "Auto-Approve",
    # "Binding Operational Directive 22-01 tracking",
    # "Binding Operational Directive 22-01 Due Date",
    # "CVE",
    # "Service Name",
    """

    def __init__(self, plan_id: int, **kwargs: dict):
        super().__init__(plan_id=plan_id)
        try:
            # Use read_only mode for memory efficiency, purposefully use kwarg index to force KeyError
            if "file_path" in kwargs:
                self.file_path = kwargs["file_path"]
            if not self.file_path:
                error_and_exit(FILE_PATH_ERROR)
            self.workbook = self.workbook or load_workbook(filename=self.file_path, data_only=True, read_only=True)
            self.poam_sheets = kwargs.get("poam_sheets") or [
                sheet for sheet in self.workbook.sheetnames if re.search("POA&M Items|Configuration Findings", sheet)
            ]
        except (FileNotFoundError, InvalidFileException, KeyError) as e:
            logger.error("Failed to load workbook: %s", e)
            return
        # Validate Here
        if not self.validators and isinstance(self.poam_sheets, list):
            for sheet in self.poam_sheets:
                ws = self.workbook[sheet]
                mapping_path = "./mappings/fedramp_poam/" + sheet
                validator = ImportValidater(
                    file_path=self.file_path,
                    disable_mapping=True,
                    required_headers=self.fedramp_poam_columns,
                    worksheet_name=sheet,
                    mapping_file_path=mapping_path,
                    prompt=True,
                    skip_rows=self.find_header_row(ws),
                    ignore_unnamed=True,
                )
                self.validators[sheet] = validator
        self.processed_assets = set()  # Reset processed assets on init

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup resources."""
        if self.workbook:
            self.workbook.close()

    def _build_column_indices(self, mapping: Mapping) -> dict:
        """
        Build a mapping of field names to column indices for fast array access.
        This eliminates repeated dict lookups during row iteration.

        :param Mapping mapping: The validator mapping object
        :return: Dictionary mapping field names to column indices
        :rtype: dict
        """
        # Get the list of column names in order (keys of the mapping dict)
        columns = list(mapping.mapping.keys())
        return {col: idx for idx, col in enumerate(columns)}

    def _load_sheet_as_polars(self, validator) -> Optional[pl.DataFrame]:
        """
        Load validator data into a Polars DataFrame for high-performance processing.
        Polars is 10-100x faster than pandas for large datasets.

        :param validator: The ImportValidater object with pandas DataFrame
        :return: Polars DataFrame or None if conversion fails
        :rtype: Optional[pl.DataFrame]
        """
        try:
            # Convert pandas DataFrame to Polars DataFrame
            if hasattr(validator, "data") and validator.data is not None:
                # Get expected column names from mapping (these are the keys we'll use to access data)
                expected_columns = list(validator.mapping.mapping.keys())

                # Convert all pandas columns to string first to avoid type inference issues
                # Excel files often have mixed types that cause conversion errors
                pandas_df = validator.data.astype(str)

                # Convert to Polars
                df = pl.from_pandas(pandas_df)

                # Rename columns to match expected field names
                # The mapping.keys() are what we use to access data (e.g., "Asset Identifier")
                if len(df.columns) >= len(expected_columns):
                    rename_map = {df.columns[i]: expected_columns[i] for i in range(len(expected_columns))}
                    df = df.rename(rename_map)
                    logger.debug(
                        "Successfully converted pandas DataFrame to Polars with %d rows, columns: %s",
                        len(df),
                        list(df.columns)[:5],
                    )
                    return df
                else:
                    logger.warning(
                        "Column count mismatch: Polars has %d columns, expected %d",
                        len(df.columns),
                        len(expected_columns),
                    )
        except Exception as e:
            logger.warning("Failed to convert to Polars DataFrame: %s. Falling back to pandas.", str(e))
        return None

    def _normalize_polars_row(self, row_dict: dict) -> dict:
        """
        Normalize a Polars row dictionary to match our expected format.
        Handles None values and converts column names to our internal keys.

        :param dict row_dict: Row dictionary from Polars iter_rows(named=True)
        :return: Normalized row values dictionary
        :rtype: dict
        """
        # Map Polars column names to our internal keys
        key_mapping = {
            POAM_ID: "poam_id",
            WEAKNESS_NAME: "weakness_name",
            WEAKNESS_DESCRIPTION: "weakness_description",
            WEAKNESS_DETECTOR_SOURCE: "weakness_detector_source",
            WEAKNESS_SOURCE_IDENTIFIER: "weakness_source_identifier",
            ASSET_IDENTIFIER: "asset_identifier",
            POINT_OF_CONTACT: "point_of_contact",
            RESOURCES_REQUIRED: "resources_required",
            OVERALL_REMEDIATION_PLAN: "overall_remediation_plan",
            ORIGINAL_DETECTION_DATE: "original_detection_date",
            SCHEDULED_COMPLETION_DATE: "scheduled_completion_date",
            PLANNED_MILESTONES: "planned_milestones",
            MILESTONE_CHANGES: "milestone_changes",
            STATUS_DATE: "status_date",
            ORIGINAL_RISK_RATING: "original_risk_rating",
            ADJUSTED_RISK_RATING: "adjusted_risk_rating",
            RISK_ADJUSTMENT: "risk_adjustment",
            FALSE_POSITIVE: "false_positive",
            OPERATIONAL_REQUIREMENT: "operational_requirement",
            DEVIATION_RATIONALE: "deviation_rationale",
            "Comments": "comments",
            "CVE": "cve",
            "Controls": "controls",
        }

        result = {}
        for polars_key, internal_key in key_mapping.items():
            val = row_dict.get(polars_key)
            # Handle None and convert to empty string for consistency
            if val is None:
                result[internal_key] = ""
            elif isinstance(val, str):
                result[internal_key] = val.strip()
            else:
                result[internal_key] = val

        return result

    def _get_row_value(self, row: tuple, field: str, indices: dict, default: str = "") -> str:
        """
        Get a value from a row using pre-computed column index.
        O(1) array access instead of O(1) dict lookup + O(1) dict lookup = faster.

        :param tuple row: The row data as a tuple/array
        :param str field: The field name to retrieve
        :param dict indices: Pre-computed column indices
        :param str default: Default value if field not found
        :return: The field value, stripped if string
        :rtype: str
        """
        idx = indices.get(field)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        if val is None:
            return default
        if isinstance(val, str):
            return val.strip()
        return val

    def _extract_all_row_values(self, row: tuple, indices: dict) -> dict:
        """
        Extract all commonly needed values from a row in a single pass.
        This replaces 25+ individual get_value calls with direct array indexing.

        :param tuple row: The row data as a tuple/array
        :param dict indices: Pre-computed column indices
        :return: Dictionary with all extracted values
        :rtype: dict
        """

        def get(field: str, default: str = "") -> str:
            return self._get_row_value(row, field, indices, default)

        return {
            "poam_id": get(POAM_ID),
            "weakness_name": get(WEAKNESS_NAME),
            "weakness_description": get(WEAKNESS_DESCRIPTION),
            "weakness_detector_source": get(WEAKNESS_DETECTOR_SOURCE),
            "weakness_source_identifier": get(WEAKNESS_SOURCE_IDENTIFIER),
            "asset_identifier": get(ASSET_IDENTIFIER),
            "point_of_contact": get(POINT_OF_CONTACT),
            "resources_required": get(RESOURCES_REQUIRED),
            "overall_remediation_plan": get(OVERALL_REMEDIATION_PLAN),
            "original_detection_date": get(ORIGINAL_DETECTION_DATE),
            "scheduled_completion_date": get(SCHEDULED_COMPLETION_DATE),
            "planned_milestones": get(PLANNED_MILESTONES),
            "milestone_changes": get(MILESTONE_CHANGES),
            "status_date": get(STATUS_DATE),
            "original_risk_rating": get(ORIGINAL_RISK_RATING),
            "adjusted_risk_rating": get(ADJUSTED_RISK_RATING),
            "risk_adjustment": get(RISK_ADJUSTMENT),
            "false_positive": get(FALSE_POSITIVE),
            "operational_requirement": get(OPERATIONAL_REQUIREMENT),
            "deviation_rationale": get(DEVIATION_RATIONALE),
            "comments": get("Comments"),
            "cve": get("CVE"),
            "controls": get("Controls"),
        }

    def _determine_category_fast(self, row_values: dict) -> str:
        """
        Determine the category of the finding from pre-extracted row values.
        OPTIMIZED: Uses module-level constants instead of recreating per call.

        :param dict row_values: Pre-extracted row values from _extract_all_row_values
        :return: The category of the finding
        :rtype: str
        """
        res = row_values.get("original_risk_rating", "")
        if not res:
            return IssueSeverity.Low.name
        if res.lower() not in SEVERITY_MEMBERS_LOWER:
            res = SEVERITY_MAPPING.get(res.lower(), IssueSeverity.Low.name)
        return res

    def _calculate_plugin_id(self, raw_plugin_id: str, fallback_id: str) -> int:
        """
        Calculate plugin ID from raw value or generate hash-based ID.

        :param str raw_plugin_id: Raw plugin ID value
        :param str fallback_id: Fallback value for hash generation
        :return: Integer plugin ID
        :rtype: int
        """
        try:
            if raw_plugin_id and str(raw_plugin_id).isdigit():
                return int(raw_plugin_id)
            return abs(hash(str(raw_plugin_id or ""))) % (10**9)
        except (ValueError, TypeError):
            return abs(hash(fallback_id)) % (10**9)

    def _validate_kwargs(self, **kwargs) -> tuple:
        """
        Validate required kwargs for finding parsing.

        :param kwargs: Keyword arguments to validate
        :return: Tuple of (status, category, index, sheet, resolve_status)
        :raises TypeError: If any required argument is missing or wrong type
        """
        status = kwargs.get("status")
        if not isinstance(status, str):
            raise TypeError("status must be a string")
        category = kwargs.get("category")
        if not isinstance(category, str):
            raise TypeError("category must be a string")
        index = kwargs.get("index")
        if not isinstance(index, int):
            raise TypeError("index must be an integer")
        sheet = kwargs.get("sheet")
        if not isinstance(sheet, str):
            raise TypeError("sheet must be a string")
        resolve_status = kwargs.get("resolve_status")
        if not isinstance(resolve_status, str):
            raise TypeError("resolve_status must be a string")
        return status, category, index, sheet, resolve_status

    def _parse_finding_fast(self, row_values: dict, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationFinding objects.
        OPTIMIZED: Uses pre-extracted row_values instead of repeated get_value() calls.
        Creates a separate finding for each asset and CVE combination.

        :param dict row_values: Pre-extracted row values from _extract_all_row_values
        :param kwargs: Arbitrary keyword arguments
        :rtype: Iterator[IntegrationFinding]
        :yields: IntegrationFinding
        """
        _, _, index, sheet, _ = self._validate_kwargs(**kwargs)

        try:
            # Direct access from pre-extracted values (no get_value calls needed)
            poam_id = row_values.get("poam_id", "")
            weakness_name = str(row_values.get("weakness_name", ""))

            # Early validation - return immediately if invalid
            if not poam_id and weakness_name in [None, "None", ""]:
                self.blank_records += 1
                return

            if not poam_id or not poam_id.upper():
                logger.debug(
                    "Invalid POAM ID on row %i, sheet %s: weakness_name=%s, poam_id=%s",
                    index,
                    sheet,
                    weakness_name,
                    poam_id,
                )
                logger.warning("Invalid POAM ID on row %i, sheet %s. Skipping.", index, sheet)
                return

            if not weakness_name:
                logger.warning("Title is required on row %i, sheet %s. Unable to import", index, sheet)
                return

            # Get and validate plugin ID
            raw_plugin_id = row_values.get("weakness_source_identifier", "")
            plugin_id_int = self._calculate_plugin_id(raw_plugin_id, poam_id)

            # Get asset identifiers
            asset_ids = row_values.get("asset_identifier", "")
            if not asset_ids:
                logger.warning("No asset identifier found on row %i, sheet %s. Skipping.", index, sheet)
                return

            # Clean asset identifiers
            asset_id_list = self.gen_asset_list(asset_ids)
            if not asset_id_list:
                logger.warning("No valid asset identifiers found on row %i, sheet %s. Skipping.", index, sheet)
                return

            # Get and validate CVEs and vendor-specific identifiers
            cves, vendor_ids = self.process_cve(row_values.get("cve", ""), index, sheet)
            cve_list = cves.split("\n") if cves else [""]
            vendor_id_list = vendor_ids.split("\n") if vendor_ids else []

            # Extract common finding data using pre-extracted row values (no more get_value calls)
            common_data = self._build_common_data_fast(row_values, **kwargs)

            # Consolidate all assets into a single finding per POAM+CVE combination
            # Use first asset as primary identifier, all assets in issue_asset_identifier_value
            primary_asset = asset_id_list[0]
            consolidated_assets = ", ".join(asset_id_list)

            # Process CVEs - ONE finding per CVE with all assets consolidated
            for cve in cve_list:
                finding = self._create_cve_finding(common_data, primary_asset, cve, plugin_id_int, consolidated_assets)
                if finding and finding.is_valid():
                    yield finding

            # Process vendor-specific identifiers - ONE finding per vendor ID with all assets consolidated
            for vendor_id in vendor_id_list:
                finding = self._create_vendor_finding(
                    common_data, primary_asset, vendor_id, plugin_id_int, consolidated_assets
                )
                if finding and finding.is_valid():
                    yield finding

        except Exception as e:
            logger.error("Error processing row %i in sheet %s: %s", index, sheet, str(e))
            self.error_records += 1

    def _parse_date_safe(self, value: str, field_name: str, default: str = "") -> str:
        """
        Safely parse a date value with error handling.

        :param str value: Raw date value to parse
        :param str field_name: Field name for logging
        :param str default: Default value if parsing fails
        :return: Parsed date string or default
        :rtype: str
        """
        invalid_values = ("", "NaT", "#REF!", None)
        if not value or str(value) in invalid_values:
            return default
        try:
            return date_str(value)
        except Exception as e:
            logger.debug("Failed to parse %s '%s': %s", field_name, value, e)
            return default

    def _build_common_data_fast(self, row_values: dict, **kwargs) -> dict:
        """
        Build common finding data from pre-extracted row values.
        OPTIMIZED: Direct dict access instead of 20+ get_value() calls.

        :param dict row_values: Pre-extracted row values
        :param kwargs: Additional keyword arguments
        :return: Dictionary containing common finding field values
        :rtype: dict
        """
        category = kwargs.get("category", "Low")
        status = kwargs.get("status", "Open")
        index = kwargs.get("index", 0)
        sheet = kwargs.get("sheet", "")
        resolve_status = kwargs.get("resolve_status", "CURRENT_DATE")
        previous_status_date = kwargs.get("previous_status_date")

        severity = getattr(IssueSeverity, category.title(), IssueSeverity.NotAssigned)

        # Parse dates using helper
        date_created = self._parse_date_safe(
            row_values.get("original_detection_date", ""), "original detection date", get_current_datetime()
        )

        due_date = self._parse_date_safe(row_values.get("scheduled_completion_date", ""), "scheduled completion date")
        if date_created and not due_date:
            due_date = issue_due_date(severity, date_created)

        status_date = self._parse_date_safe(row_values.get("status_date", ""), "status date")
        if not status_date:
            status_date = self.determine_status_date(
                index=index, sheet=sheet, resolve_status=resolve_status, previous_status_date=previous_status_date
            )

        controls = row_values.get("controls", "")

        return {
            "poam_id": row_values.get("poam_id", ""),
            "weakness_name": str(row_values.get("weakness_name", "")),
            "description": row_values.get("weakness_description", "") or "",
            "severity": severity,
            "category": category,
            "status": status,
            "date_created": date_created,
            "due_date": due_date,
            "status_date": status_date,
            "base_plugin_name": row_values.get("weakness_detector_source", "") or "",
            "observations": str(row_values.get("milestone_changes", "")) or "",
            "poam_comments": self.empty(row_values.get("comments", "")),
            "remediation": self.empty(row_values.get("overall_remediation_plan", "")),
            "basis_for_adjustment": str(self._get_basis_for_adjustment_fast(row_values)),
            "source_report": str(row_values.get("weakness_detector_source", "")),
            "point_of_contact": str(row_values.get("point_of_contact", "")),
            "milestone_changes": str(row_values.get("milestone_changes", "")),
            "planned_milestone_changes": str(row_values.get("planned_milestones", "")),
            "adjusted_risk_rating": row_values.get("adjusted_risk_rating", ""),
            "risk_adjustment": self.determine_risk_adjustment(row_values.get("risk_adjustment", "") or ""),
            "operational_requirements": str(row_values.get("operational_requirement", "")),
            "deviation_rationale": str(row_values.get("deviation_rationale", "")),
            "affected_controls": str(controls) if controls else None,
        }

    def _get_basis_for_adjustment_fast(self, row_values: dict) -> Optional[str]:
        """
        Get the basis for risk adjustment from pre-extracted row values.
        OPTIMIZED: Direct dict access instead of get_value() calls.

        :param dict row_values: Pre-extracted row values
        :return: The basis for adjustment
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(row_values.get("comments", ""))
        risk_rating = row_values.get("original_risk_rating", "")
        adjusted_risk_rating = row_values.get("adjusted_risk_rating", "")

        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    def fetch_findings(self, *args, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Fetches findings from FedRAMP POAM files.

        Uses streaming (yields immediately) to minimize memory usage for large POAM files.

        :yield: Iterator of validated integration findings
        """
        if not self.file_path:
            error_and_exit(FILE_PATH_ERROR)

        count = 0
        try:
            for sheet in self.poam_sheets:
                validator = self.validators.get(sheet)
                if not validator:
                    logger.warning("No validator found for sheet: %s", sheet)
                    continue

                sheet_kwargs = {**kwargs, "sheet": sheet}
                # Stream findings directly - yield immediately instead of accumulating
                for finding in self._process_sheet(**sheet_kwargs):
                    count += 1
                    yield finding

            self.num_findings_to_process = count

        except Exception as e:
            logger.error("Error fetching findings from POAM file: %s", str(e))
            self.num_findings_to_process = count

    def _process_row_findings(
        self,
        row_values: dict,
        status: str,
        category: str,
        index: int,
        sheet: str,
        resolve_status: str,
        previous_status_date: Optional[str],
    ) -> Iterator[tuple]:
        """
        Process a single row and yield findings with updated status date.

        :param dict row_values: Pre-extracted row values
        :param str status: Status value
        :param str category: Category value
        :param int index: Row index
        :param str sheet: Sheet name
        :param str resolve_status: Resolve status setting
        :param Optional[str] previous_status_date: Previous status date
        :yields: Tuple of (finding, updated_status_date)
        """
        for finding in self._parse_finding_fast(
            row_values=row_values,
            previous_status_date=previous_status_date,
            status=status,
            category=category,
            index=index,
            sheet=sheet,
            resolve_status=resolve_status,
        ):
            if isinstance(finding, IntegrationFinding):
                yield finding, finding.date_last_updated

    def _process_single_row(
        self,
        row_values: dict,
        index: int,
        sheet: str,
        status: str,
        category: str,
        resolve_status: str,
        previous_status_date: Optional[str],
    ) -> Iterator[tuple]:
        """
        Process a single row with validation and finding generation.

        :param dict row_values: Pre-extracted row values
        :param int index: Row index
        :param str sheet: Sheet name
        :param str status: Status value
        :param str category: Current category value
        :param str resolve_status: Resolve status setting
        :param Optional[str] previous_status_date: Previous status date
        :yields: Tuple of (finding, updated_status_date, new_category)
        """
        valid_categories = [IssueSeverity.Low.name, IssueSeverity.Moderate.name, IssueSeverity.High.name]

        # Determine category from row
        parsed_category = self._determine_category_fast(row_values)
        if parsed_category:
            category = parsed_category
            if category not in valid_categories:
                logger.warning("Invalid Original Risk Rating: %s in sheet %s. Skipping.", category, sheet)
                return

        # Validate required fields
        if not status:
            logger.warning("Status is required in sheet %s. Skipping.", sheet)
            return
        if not category:
            logger.warning("Category is required in sheet %s. Skipping.", sheet)
            return

        # Process findings
        for finding, updated_date in self._process_row_findings(
            row_values=row_values,
            status=status,
            category=category,
            index=index,
            sheet=sheet,
            resolve_status=resolve_status,
            previous_status_date=previous_status_date,
        ):
            yield finding, updated_date, category

    def _process_sheet(self, **kwargs: dict) -> Iterator[IntegrationFinding]:
        """
        Process a single sheet from the POAM workbook.
        Uses streaming (yields immediately) to minimize memory usage.
        OPTIMIZED: Uses pre-computed column indices for O(1) array access instead of dict lookups.

        :param str sheet: The sheet name
        :param **kwargs: Arbitrary keyword arguments
        :yields: IntegrationFinding objects
        :rtype: Iterator[IntegrationFinding]
        """
        if not self.workbook:
            return

        sheet = kwargs.get("sheet")
        resolve_status = kwargs.get("resolve_empty_status_date", "CURRENT_DATE")
        ws = self.workbook[sheet]
        validator = self.validators.get(sheet)

        # Initialize sheet-level state
        category = ws["C3"].value or "Low"
        if not ws["C3"].value:
            logger.warning("Category is required in cell C3. Defaulting to Low for sheet %s.", sheet)

        status = self.determine_status(sheet)
        if status is None:
            logger.warning("Unable to determine POA&M status for sheet %s. Skipping import.", sheet)
            return

        # Validate sheet setup
        start_row = self._get_sheet_start_row(validator, sheet)
        if start_row is None:
            return

        if not validator or not validator.mapping:
            logger.error("Validator mapping or validator mapping is None")
            return

        # Build column indices for NumPy path
        column_indices = self._build_column_indices(validator.mapping)

        # Try Polars for high-performance processing
        polars_df = self._load_sheet_as_polars(validator)
        use_polars = polars_df is not None

        logger.info(
            "Processing sheet: %s for findings, rows: %i (using %s)",
            sheet,
            len(validator.data),
            "Polars" if use_polars else "NumPy",
        )

        # Process rows using unified logic
        previous_status_date: str = None
        row_iterator = self._get_row_iterator(polars_df, validator, column_indices, start_row, use_polars)

        for index, row_values in row_iterator:
            try:
                for finding, updated_date, new_category in self._process_single_row(
                    row_values=row_values,
                    index=index,
                    sheet=sheet,
                    status=status,
                    category=category,
                    resolve_status=resolve_status,
                    previous_status_date=previous_status_date,
                ):
                    previous_status_date = updated_date
                    category = new_category
                    yield finding
            except Exception as e:
                logger.error("Error processing row %i in sheet %s: %s", index, sheet, str(e))
                self.error_records += 1

    def _get_sheet_start_row(self, validator, sheet: str) -> Optional[int]:
        """
        Get the starting row for sheet processing with validation.

        :param validator: The ImportValidater object
        :param str sheet: Sheet name for logging
        :return: Start row index or None if invalid
        :rtype: Optional[int]
        """
        try:
            start_row = self.find_start_row(validator.data.values)
        except IndexError:
            return None

        if start_row is None:
            logger.warning("No POAM entries found in sheet %s. Skipping.", sheet)
            return None

        return start_row

    def _get_row_iterator(
        self, polars_df, validator, column_indices: dict, start_row: int, use_polars: bool
    ) -> Iterator[tuple]:
        """
        Create a unified row iterator for both Polars and NumPy paths.

        :param polars_df: Polars DataFrame or None
        :param validator: The ImportValidater object
        :param dict column_indices: Pre-computed column indices for NumPy path
        :param int start_row: Starting row index
        :param bool use_polars: Whether to use Polars path
        :yields: Tuple of (index, row_values dict)
        """
        if use_polars:
            for index, row_dict in enumerate(polars_df.iter_rows(named=True)):
                if index >= start_row:
                    yield index, self._normalize_polars_row(row_dict)
        else:
            for index, row in enumerate(validator.data.values):
                if index >= start_row:
                    yield index, self._extract_all_row_values(row, column_indices)

    def determine_category(self, data: dict, validator: ImportValidater) -> str:
        """
        Determine the category of the finding by direct string or from a mapping.

        :param dict data: The row data
        :param ImportValidater validator: The ImportValidater object
        :return: The category of the finding
        """
        dat_map = {
            "medium": IssueSeverity.Moderate.name,
            "high": IssueSeverity.High.name,
            "critical": IssueSeverity.High.name,
            "low": IssueSeverity.Low.name,
        }
        res = validator.mapping.get_value(data, ORIGINAL_RISK_RATING)
        if res.lower() not in [mem.lower() for mem in IssueSeverity.__members__]:
            res = dat_map.get(res.lower(), IssueSeverity.Low.name)
        return res

    @staticmethod
    def is_poam(finding: IntegrationFinding) -> bool:
        """
        Determine if this finding is a POAM.

        :param IntegrationFinding finding: The finding to check
        :return: True if this is a POAM finding
        :rtype: bool
        """
        return True  # All FedRAMP findings are POAMs

    @staticmethod
    def get_issue_title(finding: IntegrationFinding) -> str:
        """
        Get the title for an issue.

        :param IntegrationFinding finding: The finding
        :return: The issue title
        :rtype: str
        """
        return finding.title[:255]  # Enforce title length limit

    def _validate_parse_kwargs(self, **kwargs) -> tuple:
        """
        Validate required kwargs for parse_finding.

        :param kwargs: Keyword arguments to validate
        :return: Tuple of (status, category, index, sheet, resolve_status)
        :raises TypeError: If any required parameter has wrong type
        """
        validations = [
            ("status", str),
            ("category", str),
            ("index", int),
            ("sheet", str),
            ("resolve_status", str),
        ]
        values = []
        for key, expected_type in validations:
            val = kwargs.get(key)
            if not isinstance(val, expected_type):
                raise TypeError("%s must be a %s" % (key, expected_type.__name__))
            values.append(val)
        return tuple(values)

    def _validate_poam_row(self, poam_id: str, weakness_name: str, index: int, sheet: str) -> bool:
        """
        Validate a POAM row for required fields.

        :param str poam_id: POAM ID value
        :param str weakness_name: Weakness name value
        :param int index: Row index
        :param str sheet: Sheet name
        :return: True if valid, False otherwise
        :rtype: bool
        """
        if not poam_id and weakness_name in (None, "None", ""):
            self.blank_records += 1
            return False

        if not poam_id or not poam_id.upper():
            logger.debug(
                "Invalid POAM ID on row %d, sheet %s: weakness_name=%s, poam_id=%s",
                index,
                sheet,
                weakness_name,
                poam_id,
            )
            logger.warning("Invalid POAM ID on row %d, sheet %s. Skipping.", index, sheet)
            return False

        if not weakness_name:
            logger.warning("Title is required on row %d, sheet %s. Unable to import", index, sheet)
            return False

        return True

    def parse_finding(self, data: dict, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationFinding objects.
        Creates a separate finding for each asset and CVE combination.
        Uses streaming (yields immediately) to minimize memory usage.

        :param dict data: The row data
        :param kwargs: Arbitrary keyword arguments
        :rtype: Iterator[IntegrationFinding]
        :yields: IntegrationFinding
        """
        _, _, index, sheet, _ = self._validate_parse_kwargs(**kwargs)
        val_mapping = kwargs.get("validator").mapping

        try:
            poam_id = val_mapping.get_value(data, self.poam_id_header)
            weakness_name = str(val_mapping.get_value(data, WEAKNESS_NAME))

            if not self._validate_poam_row(poam_id, weakness_name, index, sheet):
                return

            raw_plugin_id = val_mapping.get_value(data, WEAKNESS_SOURCE_IDENTIFIER)
            plugin_id_int = self._calculate_plugin_id(raw_plugin_id, poam_id)

            # Get and validate asset identifiers
            asset_ids = val_mapping.get_value(data, ASSET_IDENTIFIER)
            if not asset_ids:
                logger.warning("No asset identifier found on row %d, sheet %s. Skipping.", index, sheet)
                return

            asset_id_list = self.gen_asset_list(asset_ids)
            if not asset_id_list:
                logger.warning("No valid asset identifiers found on row %d, sheet %s. Skipping.", index, sheet)
                return

            # Get and validate CVEs and vendor-specific identifiers
            cves, vendor_ids = self.process_cve(val_mapping.get_value(data, "CVE"), index, sheet)
            cve_list = cves.split("\n") if cves else [""]
            vendor_id_list = vendor_ids.split("\n") if vendor_ids else []

            # Extract common finding data once per row (memory efficient)
            common_data = self._extract_common_finding_data(data, val_mapping, **kwargs)

            # Stream findings immediately - no accumulation
            for asset_id in asset_id_list:
                # Process CVEs - yield immediately
                for cve in cve_list:
                    finding = self._create_cve_finding(common_data, asset_id, cve, plugin_id_int)
                    if finding and finding.is_valid():
                        yield finding

                # Process vendor-specific identifiers - yield immediately
                for vendor_id in vendor_id_list:
                    finding = self._create_vendor_finding(common_data, asset_id, vendor_id, plugin_id_int)
                    if finding and finding.is_valid():
                        yield finding

        except Exception as e:
            logger.error(f"Error processing row {index} in sheet {sheet}: {str(e)}")
            self.error_records += 1

    def determine_status_date(self, **kwargs):
        """
        Determine the status date.

        :param kwargs: Arbitrary keyword arguments
        :return: The status date
        :rtype: str
        """
        index = kwargs.get("index")
        sheet = kwargs.get("sheet")
        resolve_status = kwargs.get("resolve_status")
        status_map = {
            "CURRENT_DATE": date_str(get_current_datetime()),
            "USE_NEIGHBOR": date_str(kwargs.get("previous_status_date")),
        }
        res = date_str(status_map.get(resolve_status), "%m-%d-%Y")
        if res:
            logger.warning(
                "Status Date missing on row %i, sheet %s, defaulting to %s: %s",
                index,
                sheet,
                resolve_status.lower().replace("_", " "),
                res,
            )
            return res
        logger.warning(
            f"Status Date missing on row {index}, sheet {sheet}. Unable to find valid neighbor, falling back to current date."
        )
        return date_str(status_map.get("CURRENT_DATE"), "%Y-%m-%d")

    def _extract_common_finding_data(self, data: dict, val_mapping: Mapping, **kwargs) -> dict:
        """
        Extract common field data shared across all findings from a row.

        :param dict data: The row data dictionary
        :param Mapping val_mapping: The validator mapping object
        :param kwargs: Additional keyword arguments including status, category, index, sheet
        :return: Dictionary containing common finding field values
        :rtype: dict
        """
        category = kwargs.get("category", "Low")
        status = kwargs.get("status", "Open")
        severity = getattr(IssueSeverity, category.title(), IssueSeverity.NotAssigned)

        # Parse dates using shared helper
        date_created = self._parse_date_safe(
            val_mapping.get_value(data, ORIGINAL_DETECTION_DATE), "original detection date", get_current_datetime()
        )

        due_date = self._parse_date_safe(val_mapping.get_value(data, SCHEDULED_COMPLETION_DATE), "scheduled completion")
        if date_created and not due_date:
            due_date = issue_due_date(severity, date_created)

        status_date = self._parse_date_safe(val_mapping.get_value(data, STATUS_DATE), "status date")
        if not status_date:
            status_date = self.determine_status_date(**kwargs)

        controls = val_mapping.get_value(data, "Controls")

        return {
            "poam_id": val_mapping.get_value(data, self.poam_id_header),
            "weakness_name": str(val_mapping.get_value(data, WEAKNESS_NAME)),
            "description": val_mapping.get_value(data, WEAKNESS_DESCRIPTION) or "",
            "severity": severity,
            "category": category,
            "status": status,
            "date_created": date_created,
            "due_date": due_date,
            "status_date": status_date,
            "base_plugin_name": val_mapping.get_value(data, WEAKNESS_DETECTOR_SOURCE) or "",
            "observations": str(val_mapping.get_value(data, MILESTONE_CHANGES)) or "",
            "poam_comments": self.empty(val_mapping.get_value(data, "Comments")),
            "remediation": self.empty(val_mapping.get_value(data, OVERALL_REMEDIATION_PLAN)),
            "basis_for_adjustment": str(self.get_basis_for_adjustment(val_mapping=val_mapping, data=data)),
            "source_report": str(val_mapping.get_value(data, WEAKNESS_DETECTOR_SOURCE)),
            "point_of_contact": str(val_mapping.get_value(data, POINT_OF_CONTACT)),
            "milestone_changes": str(val_mapping.get_value(data, MILESTONE_CHANGES)),
            "planned_milestone_changes": str(val_mapping.get_value(data, PLANNED_MILESTONES)),
            "adjusted_risk_rating": val_mapping.get_value(data, ADJUSTED_RISK_RATING),
            "risk_adjustment": self.determine_risk_adjustment(val_mapping.get_value(data, RISK_ADJUSTMENT)),
            "operational_requirements": str(val_mapping.get_value(data, OPERATIONAL_REQUIREMENT)),
            "deviation_rationale": str(val_mapping.get_value(data, DEVIATION_RATIONALE)),
            "affected_controls": str(controls) if controls else None,
        }

    def _create_cve_finding(
        self,
        common_data: dict,
        asset_id: str,
        cve: str,
        plugin_id_int: int,
        consolidated_assets: Optional[str] = None,
    ) -> Optional[IntegrationFinding]:
        """
        Create an IntegrationFinding for a CVE identifier.

        :param dict common_data: Common field data extracted from the row
        :param str asset_id: The primary asset identifier
        :param str cve: The CVE identifier (may be empty string)
        :param int plugin_id_int: Base plugin ID integer
        :param Optional[str] consolidated_assets: All assets comma-separated for issue_asset_identifier_value
        :return: IntegrationFinding object or None if status_date is missing
        :rtype: Optional[IntegrationFinding]
        """
        if not common_data.get("status_date"):
            return None

        if cve:
            unique_plugin_id = abs(hash(f"{plugin_id_int}:{cve}")) % (10**9)
            identifier = cve
        else:
            unique_plugin_id = plugin_id_int
            identifier = ""

        title = (
            f"{common_data['weakness_name'][:240]} - {identifier}" if identifier else common_data["weakness_name"][:255]
        )

        return IntegrationFinding(
            control_labels=[],
            title=title,
            category=f"FedRAMP POAM: {common_data['category']}",
            description=common_data["description"],
            severity=common_data["severity"],
            status=(
                regscale_models.IssueStatus.Closed
                if common_data["status"].lower() == "closed"
                else regscale_models.IssueStatus.Open
            ),
            asset_identifier=asset_id,
            issue_asset_identifier_value=consolidated_assets,
            external_id=f"{common_data['poam_id']}:{identifier}" if identifier else common_data["poam_id"],
            date_created=common_data["date_created"],
            date_last_updated=common_data["status_date"],
            due_date=common_data["due_date"],
            cve=identifier if identifier else None,
            plugin_name=common_data["base_plugin_name"],
            plugin_id=str(unique_plugin_id),
            observations=common_data["observations"],
            poam_comments=common_data["poam_comments"],
            remediation=common_data["remediation"],
            basis_for_adjustment=common_data["basis_for_adjustment"],
            vulnerability_type="FedRAMP",
            source_report=common_data["source_report"],
            point_of_contact=common_data["point_of_contact"],
            milestone_changes=common_data["milestone_changes"],
            planned_milestone_changes=common_data["planned_milestone_changes"],
            adjusted_risk_rating=common_data["adjusted_risk_rating"],
            risk_adjustment=common_data["risk_adjustment"],
            operational_requirements=common_data["operational_requirements"],
            deviation_rationale=common_data["deviation_rationale"],
            affected_controls=common_data["affected_controls"],
            poam_id=common_data["poam_id"],
        )

    def _create_vendor_finding(
        self,
        common_data: dict,
        asset_id: str,
        vendor_id: str,
        plugin_id_int: int,
        consolidated_assets: Optional[str] = None,
    ) -> Optional[IntegrationFinding]:
        """
        Create an IntegrationFinding for a vendor-specific identifier.

        :param dict common_data: Common field data extracted from the row
        :param str asset_id: The primary asset identifier
        :param str vendor_id: The vendor-specific identifier
        :param int plugin_id_int: Base plugin ID integer
        :param Optional[str] consolidated_assets: Comma-separated list of all affected assets
        :return: IntegrationFinding object or None if status_date is missing
        :rtype: Optional[IntegrationFinding]
        """
        if not common_data.get("status_date"):
            return None

        unique_plugin_id = abs(hash(f"{plugin_id_int}:{vendor_id}")) % (10**9)
        vendor_plugin_name = f"{vendor_id}"
        if common_data["base_plugin_name"]:
            vendor_plugin_name = f"{common_data['base_plugin_name']} - {vendor_id}"

        return IntegrationFinding(
            control_labels=[],
            title=f"{common_data['weakness_name'][:240]} - {vendor_id}",
            category=f"FedRAMP POAM: {common_data['category']}",
            description=common_data["description"],
            severity=common_data["severity"],
            status=(
                regscale_models.IssueStatus.Closed
                if common_data["status"].lower() == "closed"
                else regscale_models.IssueStatus.Open
            ),
            asset_identifier=asset_id,
            issue_asset_identifier_value=consolidated_assets,
            external_id=f"{common_data['poam_id']}:{vendor_id}",
            date_created=common_data["date_created"],
            date_last_updated=common_data["status_date"],
            due_date=common_data["due_date"],
            cve=None,
            plugin_name=vendor_plugin_name,
            plugin_id=str(unique_plugin_id),
            observations=common_data["observations"],
            poam_comments=common_data["poam_comments"],
            remediation=common_data["remediation"],
            basis_for_adjustment=common_data["basis_for_adjustment"],
            vulnerability_type="FedRAMP",
            source_report=common_data["source_report"],
            point_of_contact=common_data["point_of_contact"],
            milestone_changes=common_data["milestone_changes"],
            planned_milestone_changes=common_data["planned_milestone_changes"],
            adjusted_risk_rating=common_data["adjusted_risk_rating"],
            risk_adjustment=common_data["risk_adjustment"],
            operational_requirements=common_data["operational_requirements"],
            deviation_rationale=common_data["deviation_rationale"],
            affected_controls=common_data["affected_controls"],
            poam_id=common_data["poam_id"],
        )

    def _is_header_row(self, asset_ids_str: str) -> bool:
        """
        Check if asset identifier contains header/description text.

        :param str asset_ids_str: Lowercase asset identifier string
        :return: True if this is a header row
        :rtype: bool
        """
        header_keywords = (
            "date the weakness",
            "aka discovery",
            "permanent column",
            "date of intended",
            "last changed or closed",
            "port/protocol",
            "specified in the inventory",
        )
        return any(keyword in asset_ids_str for keyword in header_keywords)

    def parse_asset(self, row: List, validator: ImportValidater) -> List[IntegrationAsset]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationAsset objects.
        Handles multiple comma-separated asset identifiers.

        :param List row: The row data from the spreadsheet
        :param ImportValidater validator: The ImportValidater object
        :rtype: List[IntegrationAsset]
        """
        row_assets = []
        try:
            if not (validator and validator.mapping):
                logger.error("Validator mapping is None")
                return row_assets

            val_mapping = validator.mapping
            data = dict(zip(val_mapping.mapping, row))
            asset_ids = val_mapping.get_value(data, ASSET_IDENTIFIER)

            if not asset_ids:
                return row_assets

            # Skip header rows
            if self._is_header_row(str(asset_ids).lower()):
                logger.debug("Skipping row with header/description text in asset identifier: %s", str(asset_ids)[:100])
                return row_assets

            asset_id_list = self.gen_asset_list(asset_ids)
            if not asset_id_list:
                return row_assets

            # Get raw type once per row
            raw_type = self._clean_asset_str(val_mapping.get_value(data, RESOURCES_REQUIRED))

            for asset_id in asset_id_list:
                asset = self._create_integration_asset(asset_id, raw_type)
                row_assets.append(asset)

        except (KeyError, ValueError, TypeError) as kex:
            logger.error("Error parsing asset from row: %s (Exception type: %s)", str(kex), type(kex).__name__)
        except Exception as ex:
            logger.error("Unknown Error parsing asset from row: %s", str(ex))

        return row_assets

    def _parse_asset_fast(self, row_dict: dict) -> List[IntegrationAsset]:
        """
        Parse assets from a pre-normalized row dictionary (Polars optimized path).
        10-100x faster than parse_asset() by avoiding dict(zip()) and get_value() calls.

        :param dict row_dict: Pre-normalized row dictionary from Polars iter_rows(named=True)
        :return: List of parsed IntegrationAsset objects
        :rtype: List[IntegrationAsset]
        """
        row_assets = []
        try:
            # Direct dict access - no get_value() calls needed
            asset_ids = row_dict.get(ASSET_IDENTIFIER) or row_dict.get("asset_identifier", "")
            if not asset_ids:
                return row_assets

            # Skip header rows
            if self._is_header_row(str(asset_ids).lower()):
                return row_assets

            asset_id_list = self.gen_asset_list(asset_ids)
            if not asset_id_list:
                return row_assets

            # Get resources_required directly
            raw_type = row_dict.get(RESOURCES_REQUIRED) or row_dict.get("resources_required", "")
            raw_type = self._clean_asset_str(raw_type)

            for asset_id in asset_id_list:
                row_assets.append(self._create_integration_asset(asset_id, raw_type))

        except (KeyError, ValueError, TypeError) as kex:
            logger.error("Error parsing asset from row (fast path): %s", str(kex))
        except Exception as ex:
            logger.error("Unknown error parsing asset (fast path): %s", str(ex))

        return row_assets

    def _clean_asset_str(self, val: Optional[str], default: str = "") -> str:
        """
        Clean and validate string values for asset parsing (optimized helper).

        :param Optional[str] val: Value to clean
        :param str default: Default value if invalid
        :return: Cleaned string
        :rtype: str
        """
        if not val or not isinstance(val, str):
            return default

        val = str(val).strip()
        invalid_patterns = (
            "n/a",
            "none",
            "null",
            "undefined",
            "planned",
            "pending",
            "tbd",
            "remediation",
            "deviation",
            "request",
            "vulnerability",
        )
        if any(pattern in val.lower() for pattern in invalid_patterns):
            return default

        # Remove date-like strings
        if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", val):
            return default

        # Remove long descriptions
        if len(val) > 100 or "\n" in val:
            return default

        return val

    def _determine_asset_type_fast(self, asset_id: str, raw_type: str) -> str:
        """
        Determine asset type based on asset ID and raw type (optimized helper).

        :param str asset_id: The asset identifier
        :param str raw_type: Raw type from spreadsheet
        :return: Determined asset type
        :rtype: str
        """
        if raw_type and raw_type != "Other":
            return raw_type

        asset_id_lower = asset_id.lower()
        type_patterns = (
            (("docker", "container", "image", "registry"), "Container"),
            (("lambda", "function", "azure-function"), "Function"),
            (("s3", "bucket", "blob", "storage"), "Storage"),
            (("db", "database", "rds", "sql"), "Database"),
            (("ec2", "vm", "instance"), "Virtual Machine"),
        )
        for patterns, asset_type in type_patterns:
            if any(p in asset_id_lower for p in patterns):
                return asset_type
        return "Other"

    def _handle_long_asset_name(self, asset_id: str, max_length: int = 450) -> tuple[str, str]:
        """
        Handle asset names that exceed database field limits.
        Generates a hash-based identifier for long names and preserves full name in notes.

        :param str asset_id: The asset identifier
        :param int max_length: Maximum allowed length (default: 450)
        :return: Tuple of (shortened_name, notes)
        :rtype: tuple[str, str]
        """
        if len(asset_id) <= max_length:
            return asset_id, ""

        # Generate hash-based identifier
        import hashlib

        hash_suffix = hashlib.sha256(asset_id.encode()).hexdigest()[:8]
        truncated = asset_id[: max_length - 9]  # Leave room for underscore and hash
        short_name = f"{truncated}_{hash_suffix}"
        notes = f"Full identifier: {asset_id}"

        logger.debug("Asset identifier exceeds %d chars, truncated to: %s...", max_length, short_name[:100])
        return short_name, notes

    def _create_integration_asset(self, asset_id: str, raw_type: str) -> IntegrationAsset:
        """
        Create an IntegrationAsset from an asset identifier.

        :param str asset_id: The asset identifier
        :param str raw_type: Raw type from spreadsheet
        :return: IntegrationAsset object
        :rtype: IntegrationAsset
        """
        asset_name, asset_notes = self._handle_long_asset_name(asset_id)
        asset_type = self._determine_asset_type_fast(asset_id, raw_type)

        return IntegrationAsset(
            name=asset_name,
            identifier=asset_name,
            asset_type=asset_type,
            asset_category=regscale_models.AssetCategory.Hardware,
            parent_id=self.plan_id,
            parent_module=regscale_models.SecurityPlan.get_module_string(),
            status="Active (On Network)",
            ip_address=asset_id if validate_ip_address(asset_id) else "",
            fqdn=asset_id if self.is_valid_fqdn(asset_id) else "",
            mac_address=asset_id if validate_mac_address(asset_id) else "",
            notes=asset_notes,
            date_last_updated=get_current_datetime(),
        )

    def gen_asset_list(self, asset_ids: str) -> List[str]:
        """
        Generate a list of asset identifiers from a string.
        Handles multiple separator types: commas, semicolons, pipes, tabs, newlines.
        Also handles multiple spaces between identifiers and validates each identifier.

        FedRAMP-compliant asset identifiers include:
        - IP addresses (IPv4/IPv6)
        - FQDNs/hostnames
        - MAC addresses
        - URLs
        - Alphanumeric asset IDs/serial numbers

        :param str asset_ids: The asset identifier string
        :return: The list of valid asset identifiers
        :rtype: List[str]
        """
        if not asset_ids or not isinstance(asset_ids, str):
            return []

        # Remove surrounding brackets if present (handles cases like "[10.10.1.1, 10.10.1.2]")
        asset_ids = asset_ids.strip()
        if asset_ids.startswith("[") and asset_ids.endswith("]"):
            asset_ids = asset_ids[1:-1].strip()

        # Split on delimiters: commas, semicolons, pipes, tabs, newlines, and 2+ spaces
        # Pattern handles: "ip1, ip2" and "ip1  url1" (multiple spaces between values)
        raw_items = re.split(r"[,;\|\t\n\r]+|\s{2,}", asset_ids)

        # Validate and clean each identifier
        valid_identifiers = []
        for item in raw_items:
            cleaned = self._sanitize_asset_identifier(item)
            if cleaned:
                valid_identifiers.append(cleaned)

        return valid_identifiers

    def _sanitize_asset_identifier(self, identifier: str) -> Optional[str]:
        """
        Sanitize and validate a single asset identifier for FedRAMP compliance.

        Valid FedRAMP asset identifiers:
        - IP addresses (IPv4/IPv6)
        - FQDNs/hostnames (e.g., server.example.com)
        - MAC addresses
        - URLs (http/https)
        - Alphanumeric asset IDs (e.g., ASSET-001, SN12345)
        - Port/protocol combos (e.g., "10.10.1.1 ( 443 / TCP )")

        Invalid identifiers (filtered out):
        - Empty strings
        - Whitespace only
        - Pure numeric (unless IP)
        - Single characters
        - Description text

        :param str identifier: Raw identifier to sanitize
        :return: Sanitized identifier or None if invalid
        :rtype: Optional[str]
        """
        if not identifier or not isinstance(identifier, str):
            return None

        # Strip whitespace
        cleaned = identifier.strip()

        # Skip empty strings
        if not cleaned:
            return None

        # Skip very short identifiers (likely garbage)
        if len(cleaned) < 3:
            return None

        # Skip if it's just a number (not an IP)
        if cleaned.isdigit():
            return None

        # Skip common description patterns
        skip_patterns = [
            r"^(n/a|na|none|null|tbd|pending|unknown)$",  # Common placeholders
            r"^(the|this|that|and|or|for|to|from|with)$",  # Common words
            r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}",  # Date patterns
            r"^(january|february|march|april|may|june|july|august|september|october|november|december)",
        ]
        for pattern in skip_patterns:
            if re.search(pattern, cleaned.lower()):
                return None

        # Validate as one of the accepted identifier types
        if self._is_valid_asset_identifier(cleaned):
            return cleaned

        return None

    def _is_valid_asset_identifier(self, identifier: str) -> bool:
        """
        Check if identifier matches FedRAMP asset identifier patterns.

        :param str identifier: Identifier to validate
        :return: True if valid FedRAMP asset identifier
        :rtype: bool
        """
        # Check for valid IP address (IPv4 or IPv6)
        if validate_ip_address(identifier):
            return True

        # Check for IP with port/protocol (e.g., "10.10.1.1 ( 443 / TCP )")
        ip_port_pattern = r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s*\([^)]+\)$"
        if re.match(ip_port_pattern, identifier):
            return True

        # Check for valid MAC address
        if validate_mac_address(identifier):
            return True

        # Check for valid FQDN/hostname
        if self.is_valid_fqdn(identifier):
            return True

        # Check for valid URL
        url_pattern = r"^https?://[^\s]+"
        if re.match(url_pattern, identifier, re.IGNORECASE):
            return True

        # Check for alphanumeric asset ID patterns (e.g., ASSET-001, SRV-WEB-01, SN12345)
        # Must contain at least one letter and one number, or common separators
        asset_id_pattern = r"^[A-Za-z0-9][-A-Za-z0-9_.:]+[A-Za-z0-9]$"
        if re.match(asset_id_pattern, identifier):
            # Ensure it's not just a generic word - must have mix of letters/numbers or separators
            has_letter = any(c.isalpha() for c in identifier)
            has_digit = any(c.isdigit() for c in identifier)
            has_separator = any(c in "-_." for c in identifier)
            if (has_letter and has_digit) or has_separator:
                return True

        return False

    @staticmethod
    def empty(string: Optional[str]) -> Optional[str]:
        """
        Convert empty strings and "None" to None.

        :param Optional[str] string: The input string
        :return: The processed string or None
        :rtype: Optional[str]
        """
        if not isinstance(string, str):
            return None
        if string.lower() in ["none", "n/a"]:
            return None
        return string

    @staticmethod
    def determine_status(sheet: str) -> Optional[str]:
        """
        Determine the status based on sheet name.

        :param str sheet: The sheet name
        :return: The status (Open/Closed) or None
        :rtype: Optional[str]
        """
        sheet_lower = sheet.lower()
        if "closed" in sheet_lower:
            return "Closed"
        elif "open" in sheet_lower or "configuration findings" in sheet_lower:
            return "Open"
        return None

    def find_start_row(self, array: "numpy.ndarray") -> Optional[int]:
        """
        Find the first row containing POAM data.

        :param array: NumPy array containing the data
        :return: The row number where POAM entries start
        :rtype: Optional[int]
        """
        if array[0][0] == "Unique identifier for each POAM Item" and array[1][0] == "Unique Identifier":
            if array[2][0] == "V-1Example":
                return 3
            return 2

        return 0

    def get_basis_for_adjustment(self, val_mapping: Mapping, data: dict) -> Optional[str]:
        """
        Get the basis for risk adjustment.

        :param Mapping val_mapping: The mapping object
        :param dict data: The row data
        :return: The basis for adjustment
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(val_mapping.get_value(data, "Comments"))  # e.g. row 23
        risk_rating = val_mapping.get_value(data, ORIGINAL_RISK_RATING)
        adjusted_risk_rating = val_mapping.get_value(data, ADJUSTED_RISK_RATING)

        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    # Standard CVE pattern that RegScale server will accept
    CVE_PATTERN = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)

    # Vendor-specific vulnerability identifier patterns
    VENDOR_PATTERNS = (
        re.compile(r"RHSA-\d{4}:\d+", re.IGNORECASE),
        re.compile(r"GHSA-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}", re.IGNORECASE),
        re.compile(r"ALAS-\d{4}-\d+", re.IGNORECASE),
        re.compile(r"SUSE-(?:SU|RU)-\d{4}:\d+-\d+", re.IGNORECASE),
        re.compile(r"D(?:LA|SA)-\d+-\d+", re.IGNORECASE),
        re.compile(r"NSWG-ECO-\d+", re.IGNORECASE),
        re.compile(r"TEMP-\d{7}-[A-Z0-9]+", re.IGNORECASE),
    )

    # Pattern to detect potential vulnerability IDs for logging
    VULN_ID_PATTERN = re.compile(r"[A-Z].*\d|\d.*[A-Z]", re.IGNORECASE)

    def _classify_vuln_identifier(self, identifier: str) -> tuple[Optional[str], Optional[str]]:
        """
        Classify a single vulnerability identifier as CVE or vendor ID.

        :param str identifier: The vulnerability identifier to classify
        :return: Tuple of (cve, vendor_id) - one will be populated, other None
        :rtype: tuple[Optional[str], Optional[str]]
        """
        # Check for standard CVE
        cve_match = self.CVE_PATTERN.search(identifier)
        if cve_match:
            return cve_match.group(0).upper(), None

        # Check vendor patterns
        for pattern in self.VENDOR_PATTERNS:
            match = pattern.search(identifier)
            if match:
                return None, match.group(0).upper()

        return None, None

    def process_cve(self, cve: Optional[str], index: int, sheet: str) -> tuple[Optional[str], Optional[str]]:
        """
        Process and validate CVE string. Separates standard CVEs from vendor-specific identifiers.

        :param Optional[str] cve: The CVE or vulnerability identifier string
        :param int index: The row index
        :param str sheet: The sheet name
        :return: Tuple of (cve_string, vendor_id_string) - CVEs joined by newlines, vendor IDs joined by newlines
        :rtype: tuple[Optional[str], Optional[str]]
        """
        cve = self.empty(cve)
        if not cve:
            return None, None

        # Split by comma and clean
        cve_list = [c.strip() for c in cve.split(",") if c.strip()]
        if not cve_list:
            return None, None

        valid_cves = []
        vendor_ids = []
        skip_values = ("no", "n/a", "none", "null", "")

        for single_cve in cve_list:
            if single_cve.lower() in skip_values:
                continue

            cve_id, vendor_id = self._classify_vuln_identifier(single_cve)
            if cve_id:
                valid_cves.append(cve_id)
            elif vendor_id:
                vendor_ids.append(vendor_id)
            elif self.VULN_ID_PATTERN.search(single_cve):
                logger.debug(
                    "Unrecognized vulnerability identifier format: %s on row %d, sheet %s. Skipping.",
                    single_cve,
                    index,
                    sheet,
                )

        cves_str = "\n".join(valid_cves) if valid_cves else None
        vendor_str = "\n".join(vendor_ids) if vendor_ids else None

        return cves_str, vendor_str

    def is_valid_fqdn(self, hostname: str) -> bool:
        """
        Check if the hostname is valid.

        :param str hostname: The hostname string
        :return: True if the hostname is valid
        :rtype: bool
        """
        if validate_ip_address(hostname):
            return False

        if not hostname or len(hostname) > 255:
            return False

        allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-.")
        if not all(char in allowed for char in hostname):
            return False

        parts = hostname.split(".")
        if len(parts) < 2:
            return False

        if hostname[-1] == ".":
            hostname = hostname[:-1]

        return all(
            1 <= len(part) <= 63 and not part.startswith("-") and not part.endswith("-") for part in hostname.split(".")
        )

    def find_header_row(self, ws: Worksheet) -> int:
        """
        Find the header row in the POAM sheet.

        :param ws: Worksheet
        :return: The header row number
        :rtype: int
        """
        # Loop every row
        header_row = None
        for ix, row in enumerate(ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, values_only=True)):
            for cell in row:
                if cell and self.poam_id_header in str(cell):
                    header_row = ix + 1
                    break
            if header_row:
                break
        if not header_row:
            error_and_exit("Unable to find the header row in the POAM sheet.")
        return header_row

    def progress_bar(self, progress, total, width=50):
        filled = int(width * progress // total)
        bar = "=" * filled + "-" * (width - filled)
        percent = progress / total * 100
        return f"[{bar}] {percent:.1f}%"

    def fetch_assets(self, *args, **kwargs) -> Iterator[IntegrationAsset]:
        """
        Fetch assets from FedRAMP POAM files.

        Uses Polars for high-performance processing (10-100x faster than pandas).
        Falls back to pandas/numpy if Polars conversion fails.
        Streams (yields immediately) to minimize memory usage for large POAM files.

        :raises POAMProcessingError: If there's an error processing the POAM file
        :return: Iterator of parsed integration assets
        :rtype: Iterator[IntegrationAsset]
        """
        if not self.file_path:
            error_and_exit(FILE_PATH_ERROR)

        total_processed = 0

        try:
            logger.info("Starting POAM sheets processing from %s", self.file_path)

            with self.context.get_lock("processed_assets"):
                for sheet_name in self.poam_sheets:
                    try:
                        validator = self.validators.get(sheet_name)
                        if not validator:
                            logger.warning("No validator found for sheet: %s", sheet_name)
                            continue

                        data = validator.data
                        if data.empty:
                            logger.warning("Empty sheet found: %s", sheet_name)
                            continue

                        start_row = self.find_start_row(data.values)
                        rows_count = len(data.values) - start_row

                        # Try Polars for high-performance processing
                        polars_df = self._load_sheet_as_polars(validator)
                        use_polars = polars_df is not None

                        logger.info(
                            "Processing sheet '%s' with %d rows starting from row %d (using %s)",
                            sheet_name,
                            rows_count,
                            start_row,
                            "Polars" if use_polars else "pandas",
                        )

                        # Process rows using unified iterator
                        for asset, processed_count in self._process_asset_rows(
                            polars_df, validator, start_row, use_polars, sheet_name, total_processed
                        ):
                            total_processed = processed_count
                            yield asset

                    except Exception as sheet_error:
                        logger.error("Failed to process sheet '%s': %s", sheet_name, str(sheet_error), exc_info=True)
                        continue

            self.num_assets_to_process = total_processed

        except Exception as e:
            logger.error("Critical error while processing POAM file: %s", str(e), exc_info=True)
            self.num_assets_to_process = total_processed

        finally:
            logger.info("Completed processing with %d assets and %d errors", total_processed, self.error_records)

    def _process_asset_rows(
        self, polars_df, validator, start_row: int, use_polars: bool, sheet_name: str, total_processed: int
    ) -> Iterator[tuple]:
        """
        Process rows from either Polars or pandas path for asset extraction.

        :param polars_df: Polars DataFrame or None
        :param validator: ImportValidater for pandas path
        :param int start_row: Starting row index
        :param bool use_polars: Whether to use Polars path
        :param str sheet_name: Sheet name for logging
        :param int total_processed: Current count of processed assets
        :yields: Tuple of (asset, updated_total_processed)
        """
        if use_polars:
            for ix, row_dict in enumerate(polars_df.iter_rows(named=True)):
                if ix >= start_row:
                    for asset in self._safe_parse_asset_row(
                        ix, sheet_name, lambda rd=row_dict: self._parse_asset_fast(rd)
                    ):
                        total_processed += 1
                        yield asset, total_processed
        else:
            for ix, row in enumerate(validator.data.values[start_row:], start=start_row):
                for asset in self._safe_parse_asset_row(ix, sheet_name, lambda r=row: self.parse_asset(r, validator)):
                    total_processed += 1
                    yield asset, total_processed

    def _safe_parse_asset_row(self, ix: int, sheet_name: str, parse_fn) -> Iterator:
        """
        Safely parse asset row with error handling.

        :param int ix: Row index
        :param str sheet_name: Sheet name for logging
        :param parse_fn: Callable that returns parsed assets
        :yields: Parsed assets
        """
        try:
            yield from parse_fn()
        except Exception as row_error:
            logger.error("Failed to process row %d in sheet '%s': %s", ix, sheet_name, str(row_error))
            self.error_records += 1

    def find_max_row(self, start_row: int, ws: Worksheet) -> int:
        """
        A Method to find the max row in the worksheet.

        :param start_row: int
        :param ws: Worksheet
        :return: The max row number
        :rtype: int
        """
        last_row = ws.max_row
        for row in range(start_row, last_row):
            if ws.cell(row=row, column=1).value:
                continue
            else:
                return row
        return last_row

    def determine_risk_adjustment(self, param):
        """
        Determine the risk adjustment.

        Yes, No or Pending

        :param param: The parameter to check
        :return: The risk adjustment
        """
        adjustment_map = {
            "false": "No",
            "no": "No",
            "": "No",
            None: "No",
            "true": "Yes",
            "yes": "Yes",
            "pending": "Pending",
            "closed": "No",
            "n/a": "No",
        }
        # BMC Prefers this
        return adjustment_map.get(param.lower(), "No")
