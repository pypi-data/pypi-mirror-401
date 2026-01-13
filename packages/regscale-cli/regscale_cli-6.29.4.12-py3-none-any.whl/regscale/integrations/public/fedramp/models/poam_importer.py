#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""A class to import Fedramp V4 and V5 POAMs"""
import logging
import re
from collections import Counter
from typing import Optional, Union

import rich.progress
from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.compat import safe_string
from openpyxl.utils import column_index_from_string  # type: ignore
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
from pathlib import Path

from regscale.core.app.utils.app_utils import create_progress_object, get_current_datetime
from regscale.core.utils.date import date_str, datetime_str
from regscale.integrations.integration.issue import IntegrationIssue
from regscale.integrations.scanner_integration import issue_due_date
from regscale.integrations.variables import ScannerVariables
from regscale.models import IssueSeverity, regscale_models

logger = logging.getLogger("regscale")


class POAM(IntegrationIssue):
    """
    Custom Integration issue class
    """

    def __init__(self, file_path: str, module: str, module_id: int, poam_id_header: str = "POAM ID"):
        super().__init__()
        self.skipped_records = 0
        self.blank_records = 0
        self.blank_threshold = 3
        self.error_records = 0
        self.progress = create_progress_object()
        if not file_path:
            raise ValueError("File path is required")
        self.file_path = Path(file_path)
        self.module = module
        self.module_id = module_id
        self.poam_id_header = poam_id_header
        self.poam_data: dict[str, regscale_models.Issue] = {}
        data = self.import_poam()
        self.data = data

        self.create_or_update_issues(
            issues=list(self.poam_data.values()), parent_id=self.module_id, parent_module=self.module
        )
        logger.info("Finished importing POAMs..")

    def create_or_update_issues(
        self,
        issues: list[regscale_models.Issue],
        parent_id: int,
        parent_module: str,
    ):
        """
        Create issues in RegScale

        :param list[Issue] issues: list of issues to create or update
        :param int parent_id: parent id
        :param str parent_module: parent module
        """
        with self.progress as progress:
            issue_task = progress.add_task(
                "[#f8b737]Determining if issues need to be updated or created...", total=len(issues)
            )
            issue_updates = []
            issue_creations = []
            for issue in issues:
                issue.parentId = parent_id
                issue.parentModule = parent_module
                if issue.id != 0:
                    issue_updates.append(issue)
                else:
                    issue_creations.append(issue)
                progress.update(issue_task, advance=1)
            if issue_creations:
                regscale_models.Issue.batch_create(issue_creations, self.progress)
            if issue_updates:
                regscale_models.Issue.batch_update(issue_updates, self.progress)

    def pull(self):
        """
        Pull inventory from an Integration platform into RegScale
        """
        # Implement the pull method here
        pass

    def file_type(self):
        """
        A method to return the file type
        """
        file_type = None
        if self.file_path:
            file_type = self.file_path.suffix
        return file_type

    @staticmethod
    def get_index_from_column_name(column_name: str) -> int:
        """
        A method to get the index from a column name

        :param str column_name: A column name
        :return: The index of the column
        :rtype: int
        """
        return column_index_from_string(column_name) - 1

    def get_row_val(self, row: tuple, column_name: str) -> Optional[str]:
        """
        Get the value from the row

        :param tuple row: The row
        :param str column_name: The column name
        :return: The value or None
        :rtype: Optional[str]
        """
        try:
            index = self.get_index_from_column_name(column_name)
            return row[index] if index < len(row) else None
        except Exception as e:
            logger.error(f"Error getting value for column {column_name}: {str(e)}")
            return None

    def get_basis_for_adjustment(self, row: tuple) -> Optional[str]:
        """
        Get the basis for adjustment

        :param tuple row: The row
        :return: The basis for adjustment or None if adjusted risk rating is the same as risk rating
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(row[23])
        risk_rating = self.get_row_val(row, "S")
        adjusted_risk_rating = self.get_row_val(row, "T")
        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    def process_cve(self, cve: Optional[str], index: int, sheet: str) -> Optional[str]:
        """
        Process and validate CVE string.

        :param Optional[str] cve: The CVE string to process
        :param int index: The row index for logging purposes
        :param str sheet: The sheet name for logging purposes
        :return: Processed CVE string or None
        :rtype: Optional[str]
        """
        cve = self.empty(cve)
        if not cve:
            return None

        cve_pattern = r".*CVE-\d{4}-\d{4,7}.*"
        match = re.match(cve_pattern, cve, re.IGNORECASE)
        if match:
            cve_match = re.search(r"CVE-\d{4}-\d{4,7}", cve, re.IGNORECASE)
            if cve_match:
                return cve_match.group(0).upper()  # Ensure consistent formatting
            return None  # No CVE found within the matching string
        else:
            logger.warning(f"Invalid CVE format: {cve} on row {index}, sheet {sheet}. Setting to empty string.")
            return ""

    def gen_issue_from_row(
        self, row: tuple, status: str, category: str, index: int, sheet: str
    ) -> Optional[regscale_models.Issue]:
        """
        Generate an Issue object from a row in the POAM spreadsheet.

        :param tuple row: A row from the POAM spreadsheet
        :param str status: The status of the issue (Open or Closed)
        :param str category: The category of the issue
        :param int index: The index of the row in the spreadsheet
        :param str sheet: The name of the sheet being processed
        :return: An Issue object if successfully generated, None otherwise
        :rtype: Optional[Issue]
        """
        # Extract and validate key fields
        poam_id = self.get_row_val(row, "A")
        weakness_name = str(self.get_row_val(row, "C"))

        if not poam_id or not poam_id.upper():
            logger.warning(f"Invalid POAM ID on row {index}, sheet {sheet}. Skipping.")
            return None
        if not weakness_name:
            logger.warning(f"Title is required on row {index}, sheet {sheet}. Unable to import")
            return None

        # Process risk ratings and adjustments
        original_risk_rating = self.empty(self.get_row_val(row, "S"))
        adjusted_risk_rating = self.get_row_val(row, "T")
        adjusted_risk_rating = original_risk_rating or "N/A" if adjusted_risk_rating == "N/A" else adjusted_risk_rating

        # Process CVE
        cve = self.process_cve(self.get_row_val(row, "AD"), index, sheet)

        # Determine severity level
        severity_level = getattr(IssueSeverity, category.title(), IssueSeverity.NotAssigned)

        # Process dates
        date_created = date_str(self.get_row_val(row, "K"))
        date_last_updated = datetime_str(self.get_row_val(row, "O"))
        due_date = self.get_row_val(row, "L")
        if due_date == "#REF!":
            due_date = ""
        due_date = date_str(due_date) or issue_due_date(severity_level, date_created, high=30, moderate=90, low=364)
        date_completed = None
        # Create and return the Issue object
        try:
            if status == "Closed":
                date_completed = date_str(date_last_updated) or due_date or get_current_datetime()

            issue: regscale_models.Issue = regscale_models.Issue(
                integrationFindingId=poam_id,
                otherIdentifier=poam_id,
                dateCreated=date_created,
                dateLastUpdated=date_last_updated,
                title=weakness_name[:255],
                description=self.get_row_val(row, "D"),
                status=status,
                severityLevel=severity_level,
                assetIdentifier=self.get_row_val(row, "G"),
                isPoam=True,
                issueOwnerId=ScannerVariables.userId,
                securityPlanId=self.module_id if self.module == "securityplans" else 0,
                cve=cve,
                sourceReport=self.get_row_val(row, "E"),
                pluginId=str(self.get_row_val(row, "F")),
                autoApproved="No",
                dueDate=due_date,
                parentId=self.module_id,  # type: ignore
                parentModule=self.module,  # type: ignore
                basisForAdjustment=self.get_basis_for_adjustment(row),
                dateCompleted=date_completed,  # when an issue is closed it has to have a date completed cannot be null
                manualDetectionSource=self.get_row_val(row, "E"),
                manualDetectionId=str(self.get_row_val(row, "F")),
                changes=safe_string(self.get_row_val(row, "N")),
                poamComments=self.empty(self.get_row_val(row, "Z")),
                deviationRationale=self.empty(self.get_row_val(row, "X")),
                remediationDescription=self.empty(self.get_row_val(row, "J")),
                vendorDependency=self.empty(self.get_row_val(row, "P")),
                vendorLastUpdate=self.empty(date_str(self.get_row_val(row, "Q"))),
                vendorName=self.empty(self.get_row_val(row, "R")),
                adjustedRiskRating=adjusted_risk_rating,
                originalRiskRating=original_risk_rating or adjusted_risk_rating,
                falsePositive=self.set_false_positive(row),
                identification="Vulnerability Assessment",
                operationalRequirement=self.set_operational_requirement(row),
                dateFirstDetected=date_str(self.get_row_val(row, "K")),
                riskAdjustment=self.set_risk_adjustment(row),
            ).create_or_update(bulk_update=True)
            if poc := self.get_row_val(row, "H"):
                _ = regscale_models.Property(
                    key="POC",
                    value=poc,
                    parentId=issue.id,
                    parentModule="issues",
                ).create_or_update(bulk_update=True, bulk_create=True)
        except Exception as e:
            logger.error(f"Error creating Issue object on row {index}, sheet {sheet}: {str(e)}", exc_info=True)
            return None

        self.poam_data[poam_id] = issue
        return issue

    def import_poam(self) -> Optional[Workbook]:
        """
        Import POAM data from the workbook.

        :return: The processed workbook or None if import failed
        :rtype: Optional[Workbook]
        """
        try:
            workbook = load_workbook(filename=self.file_path, data_only=True, read_only=True)
        except (FileNotFoundError, InvalidFileException) as e:
            logger.error(f"Failed to load workbook: {e}")
            return None

        poam_sheets = [sheet for sheet in workbook.sheetnames if re.search("POA&M Items", sheet)]

        with self.progress as progress:
            parsing_progress = progress.add_task("[#f8b737]Parsing data from workbook...", total=len(poam_sheets))

            for sheet in poam_sheets:
                self.process_sheet(workbook[sheet], sheet, progress)
                progress.update(parsing_progress, advance=1)

        self.count_issues_by_status()
        return workbook

    def process_sheet(self, ws, sheet_name: str, progress: rich.progress.Progress):
        """
        Process a single sheet in the POAM workbook.

        :param ws: The worksheet object
        :param str sheet_name: The name of the sheet
        :param rich.progress.Progress progress: The progress object for updating task progress
        """
        category = ws["C3"].value or "Low"
        if not ws["C3"].value:
            logger.warning(f"Category is required in cell C3. Defaulting to Low import for sheet {sheet_name}.")
        if not category:
            logger.warning(f"Category is required in cell C3. Skipping import for sheet {sheet_name}.")
            return

        status = self.determine_status(sheet_name)
        if status is None:
            logger.warning(f"Unable to determine POA&M status for sheet {sheet_name}. Skipping import.")
            return

        start_row = self.find_start_row(ws)
        if start_row is None:
            logger.warning(f"No POAM entries found in sheet {sheet_name}. Skipping.")
            return

        parsing_poams = progress.add_task(
            f"[#ef5d23]Parsing '{sheet_name}' sheet for POAMs...", total=ws.max_row - start_row + 1
        )

        for index, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            try:
                self.process_row(row, status, category, index, sheet_name)
                if self.blank_records >= self.blank_threshold:
                    logger.warning("Too many empty records skipped. Stopping import.")
                    progress.update(parsing_poams, completed=ws.max_row - start_row + 1)
                    break
            except Exception as e:
                logger.error(f"Error processing row {index} in sheet {sheet_name}: {str(e)}")
                self.error_records += 1
            progress.update(parsing_poams, advance=1)
        regscale_models.Issue.bulk_save(progress_context=progress)
        regscale_models.Property.bulk_save(progress_context=progress)

    def find_start_row(self, ws) -> Optional[int]:
        """
        Find the first row with 'V-' or any identifier-number in column A.

        :param ws: The worksheet object
        :return: The row number where POAM entries start, or None if not found
        :rtype: Optional[int]
        """
        for row_index, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), 1):
            if row[0] and self.poam_id_header in str(row[0]):
                logger.info(f"Found POAM header parsing data from row {row_index + 1}")
                return row_index + 1
        return None

    @staticmethod
    def identify_id_data(value: str) -> bool:
        """
        Identify the ID

        :param str value: The value
        :return: The ID
        :rtype: bool
        """
        return bool(re.match(r".+-\d+$", value))

    @staticmethod
    def determine_status(sheet: str) -> Optional[str]:
        """
        Determine the status based on sheet name.

        :param str sheet: The name of the sheet
        :return: The status of the POA&M (Closed, Open, or None)
        :rtype: Optional[str]
        """
        # Check if the sheet name contains 'closed' (case-insensitive)
        if "closed" in sheet.lower():
            return "Closed"
        # Check if the sheet name contains 'open' (case-insensitive)
        elif "open" in sheet.lower():
            return "Open"
        # If neither 'closed' nor 'open' is found in the sheet name
        else:
            # Log a warning message
            logger.debug(f"Unable to determine POA&M status for sheet {sheet}. Skipping import.")
            # Return None to indicate that the status couldn't be determined
            return None

    def process_row(self, row: tuple, status: str, category: str, index: int, sheet: str):
        """
        Process a single row of the POAM sheet.

        :param tuple row: The row data from the POAM sheet
        :param str status: The status of the POAM (Open or Closed)
        :param str category: The category of the POAM
        :param int index: The index of the current row
        :param str sheet: The name of the current sheet
        """
        try:
            # Get the POAM ID from column A and handle empty values
            poam_id = self.empty(self.get_row_val(row, "A"))

            # Check if POAM ID is missing
            if not poam_id:
                logger.warning(f"POAM ID is required. Skipping import for row {index} in sheet {sheet}.")
                self.blank_records += 1
                return
            self.blank_records = 0
            # Check if closed POAM already exists in the data
            if status == "Closed" and poam_id in self.poam_data:
                logger.warning(
                    f"POAM {poam_id} already exists with status {status}. Skipping import for row {index} in sheet {sheet}."
                )
                self.skipped_records += 1
                return

            # Generate issue from row data
            issue = self.gen_issue_from_row(row, status, category, index, sheet)
            if issue:
                # Add the generated issue to the POAM data dictionary
                self.poam_data[poam_id] = issue
            else:
                logger.warning(f"Failed to generate issue for POAM {poam_id} from row {index} in sheet {sheet}")
                self.skipped_records += 151
        except Exception as e:
            logger.error(f"Error in process_row for row {index} in sheet {sheet}: {str(e)}")
            self.error_records += 1

    def count_issues_by_status(self):
        """
        A method to count the issues and log the counts.
        """
        status_list = [issue.status for issue in self.poam_data.values() if issue]
        status_counts = Counter(status_list)
        logger.info(
            "Found %i issues in the POAM Workbook, %i Open and %i Closed.",
            len(self.poam_data),
            status_counts["Open"],
            status_counts["Closed"],
        )
        error_msg = f"Skipped {self.skipped_records} records, {self.error_records} errors"
        if self.error_records:
            logger.error(error_msg)
        elif self.skipped_records:
            logger.warning(error_msg)
        else:
            logger.info(error_msg)

    @staticmethod
    def empty(string: Optional[str]) -> Union[str, None]:
        """
        A method to empty the data

        :param str string: A string
        :return: None if the string is 'None' or the input is not a string
        :rtype: Union[str, None]
        """
        if not isinstance(string, str):
            return None

        if string.lower() in ["none", "n/a"]:
            return None

        return string

    def set_false_positive(self, row: tuple) -> str:
        """
        Set the false positive value

        :param tuple row: The row
        :return: The false positive value
        :rtype: str
        """
        # Map lowercased values to their corresponding responses
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending Review"}

        # Get the value from the row and convert it to lowercase
        if row_value := self.get_row_val_str(row, "V"):
            row_value = row_value.lower()

        # Get the corresponding response from the map, default to 'No' if not found
        return value_map.get(row_value, "No") if row_value else "No"

    def set_operational_requirement(self, row: tuple) -> str:
        """
        Set the operational requirement value

        :param tuple row: The row
        :return: The operational requirement value
        :rtype: str
        """
        # Map lowercased values to their corresponding responses
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending"}

        # Get the value from the row and convert it to lowercase
        if row_value := self.get_row_val_str(row, "W"):
            row_value = row_value.lower()

        # Get the corresponding response from the map, default to No if not found
        return value_map.get(row_value, "No") if row_value else "No"

    def set_risk_adjustment(self, row: tuple) -> str:
        """
        Set the risk adjustment value

        :param tuple row: The row
        :return: The Risk adjustment string
        :rtype: str
        """
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending"}
        if row_value := self.get_row_val_str(row, "U"):
            row_value = row_value.lower()
        return value_map.get(row_value, "No") if row_value else "No"

    def get_row_val_str(self, row: tuple, column_name: str) -> str:
        """
        Get the safe string

        :param tuple row: The row
        :param str column_name: The column name
        :return: The safe string
        :rtype: str
        """
        return safe_string(self.get_row_val(row, column_name))
