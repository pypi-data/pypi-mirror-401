#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# pylint: disable=C0302
"""Module to allow user to make changes to certain models in an Excel
spreadsheet for user-friendly experience"""

# standard python imports
import logging
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd  # Type Checking

import math
import os
import shutil
from pathlib import Path
from typing import Optional, Union, Any
from operator import attrgetter

import click
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, Font, NamedStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic.fields import FieldInfo

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    error_and_exit,
    get_user_names,
    check_empty_nan,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.facility import Facility
from regscale.models.regscale_models.assessment import Assessment
from regscale.models.regscale_models.modules import Modules
from regscale.models.regscale_models.control import Control
from regscale.models.regscale_models.control_implementation import ControlImplementation
from regscale.models.regscale_models.issue import Issue
from regscale.models.regscale_models.asset import Asset
from regscale.models.regscale_models.component import Component, ComponentType
from regscale.models.regscale_models.risk import Risk

# Task should be included, but doesn't have a model in
# regscale.models.regscale_models yet.
# from regscale.models.regscale_models

ALL_PRE = "All_"
NEW_PRE = "New_"
OLD_PRE = "Old_"
FILE_POST = "s.xlsx"
DIFFERENCES_FILE = "differences.txt"
SELECT_PROMPT = "Please select an option from the dropdown list."
DATE_ENTRY_PROMPT = "Please enter a valid date in the following format: mm/dd/yyyy"
SELECTION_ERROR = "Your entry is not one of the available options."
INVALID_ENTRY_ERROR = "Your entry is not a valid option."
INVALID_ENTRY_TITLE = "Invalid Entry"

logger = logging.getLogger("regscale")

exclude_fields = [
    "uuid",
    "createdBy",
    "createdById",
    "lastUpdatedBy",
    "lastUpdatedById",
    "dateLastUpdated",
    "dateCreated",
]


# pylint: disable=R0902,R0903
class FieldMakeup:
    """
    This class is for holding metadata about each field in the model being processed.
    """

    def __init__(self, field_name: str, col_name: str, data_type: str):
        self.field_name = field_name
        self.column_name = col_name
        self.data_type = data_type
        self.sort_order = 0
        self.lookup_field = ""
        self.enum_values = []
        self.treat_as_date = False
        self.cell_col = ""
        self.required = False
        self.treat_enum_as_lookup = False

    def __post_init__(self):
        if self.data_type == "bool":
            self.enum_values = ["TRUE", "FALSE"]


# pylint: enable=R0902,R0903

obj_fields = []
include_fields = []

lookup_dfs = {}


@click.group(name="model")
def model():
    """
    Performs actions on CLI models Feature to update issues to RegScale.
    """


# Make Empty Spreadsheet for creating new assessments.
@model.command(name="new")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
@click.option(
    "--model",
    type=click.Choice(
        [
            "assessment",
            "issue",
            "component",
            "asset",
        ],  # , 'risk'], #, 'task'],
        case_sensitive=False,
    ),
    help="Specify the type of new bulk load file to create.",
    default="assessment",
    required=True,
)
def generate_new_file(path: Path, model: str):
    """This function will build an Excel spreadsheet for users to be
    able to create new assessments."""
    new_file(path, model)


# pylint: disable=W0612
def new_file(path: Path, obj_type: str) -> None:
    """
    Function to build Excel spreadsheet for creation of new assessments

    :param Path path: directory of file location
    :param str obj_type: type of new spreadsheet to create
    :return: None
    :rtype: None
    """
    check_file_path(path)

    # get model specified
    obj = get_obj(obj_type)  # noqa F841

    if not obj.is_new_excel_record_allowed():
        logger.warning("Creating new records for this model type in Excel spreadsheets are not allowed.")
        return
    # build workbook
    # create excel file and setting formatting

    workbook_title = get_workbook_title(obj_type, NEW_PRE, "s")
    workbook_filename = get_workbook_title(obj_type, NEW_PRE, FILE_POST)
    build_workbook(path, workbook_filename, workbook_title)  # noqa F841

    logger.info(f"Your excel workbook has been created. Please open {workbook_filename} and add new {obj_type}s.")


# pylint: enable=W0612


@model.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
@click.option(
    "--model",
    type=click.Choice(
        [
            "assessment",
            "control",
            "issue",
            "component",
            "asset",
        ],  # , 'risk'], #, 'task'],
        case_sensitive=False,
    ),
    help="Specify the type of bulk load file to generate.",
    default="assessment",
    required=True,
)
def generate(regscale_id: int, regscale_module: str, path: Path, model: str):
    """
    This function will build and populate a spreadsheet of all assessments
    with the selected RegScale Parent Id and RegScale Module for users to any necessary edits.
    """
    all_of_model(parent_id=regscale_id, parent_module=regscale_module, path=path, obj_type=model)


def all_of_model(parent_id: int, parent_module: str, path: Path, obj_type: str) -> None:
    """
    This function will pull all records of the type specified for the parent ID and module
    specified, populate an Excel spreadsheet with the data, and save it in the specified
    folder for the user to edit as appropriate.

    :param int parent_id: RegScale Parent Id
    :param str parent_module: RegScale Parent Module
    :param Path path: directory of file location
    :param str obj_type: The model type to download
    :return: None
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    app = Application()

    # get model specified
    obj = get_obj(obj_type)
    if obj.use_query():
        existing_data = get_all_by_query(obj, parent_id, parent_module, app)
    else:
        existing_data = get_all_by_parent(obj_type, parent_id, parent_module)
    logger.debug(existing_data)
    if len(existing_data) > 0:
        match_fields_to_data(existing_data)
        check_file_path(path)
        workbook_title = get_workbook_title(obj_type, "", f"({parent_id}_{parent_module}")
        workbook_filename = get_workbook_title(obj_type, ALL_PRE, FILE_POST)
        old_workbook_filename = get_workbook_title(obj_type, OLD_PRE, FILE_POST)
        build_workbook(path, workbook_filename, workbook_title)
        shutil.copy(
            os.path.join(path, workbook_filename),
            os.path.join(path, old_workbook_filename),
        )
        all_df = put_data_into_df(existing_data)
        with pd.ExcelWriter(
            os.path.join(path, workbook_filename),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_df.to_excel(
                writer,
                sheet_name=workbook_title,
                index=False,
            )
        with pd.ExcelWriter(
            os.path.join(path, old_workbook_filename),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_df.to_excel(
                writer,
                sheet_name=workbook_title,
                index=False,
            )

        workbook2 = load_workbook(os.path.join(path, old_workbook_filename))
        worksheet2 = workbook2.active
        worksheet2.protection.sheet = True
        workbook2.save(filename=os.path.join(path, old_workbook_filename))

        # Adding Data Validation to ALL_ASSESSMENTS_WB file to be adjusted internally.
        workbook = load_workbook(os.path.join(path, workbook_filename))

        workbook.save(filename=os.path.join(path, workbook_filename))
        logger.info(f"Your excel workbook has been created. Please open {workbook_filename} and add new {obj_type}s.")

    else:
        app.logger.info("Please check your selections for RegScale Id and RegScale Module and try again.")
        error_and_exit(
            "There was an error creating your workbook. No "
            + obj_type
            + " exist for the given RegScale Id and RegScale Module."
        )


@model.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of excel workbook locations.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
@click.option(
    "--model",
    type=click.Choice(
        [
            "assessment",
            "control",
            "issue",
            "component",
            "asset",
        ],  # , 'risk'], #, 'task'],
        case_sensitive=False,
    ),
    help="Specify the type of bulk load file to load.",
    default="assessment",
    required=True,
)
def load(path: Path, model: str) -> None:
    """
    This function uploads updated assessments and new assessments to
    RegScale from the Excel files that users have edited.
    """
    upload_data(path=path, obj_type=model)


# pylint: disable=R0914
def upload_data(path: Path, obj_type: str) -> None:
    """
    Function will upload assessments to RegScale if user as made edits to any
    of the assessment excel workbooks

    :param Path path: directory of file location
    :param str obj_type: The model type to download
    :return: None
    :rtype: None
    """
    import pandas as pd

    app = Application()
    api = Api()

    # get model specified - This is to populate the obj_fields structure
    obj = get_obj(obj_type)
    if not obj:
        app.logger.error("Unable to instantiate an object of type {}".format(obj_type))
    all_workbook_filename = get_workbook_title(obj_type, ALL_PRE, FILE_POST)
    old_workbook_filename = get_workbook_title(obj_type, OLD_PRE, FILE_POST)
    new_workbook_filename = get_workbook_title(obj_type, NEW_PRE, FILE_POST)
    if os.path.isfile(os.path.join(path, new_workbook_filename)):
        upload_new_data(app, path, obj_type, new_workbook_filename)
    else:
        app.logger.info("No new " + obj_type + " detected. Checking for edited " + obj_type + "s.")

    if os.path.isfile(os.path.join(path, all_workbook_filename)):
        if not os.path.isfile(os.path.join(path, old_workbook_filename)):
            return app.logger.error("Missing pre-change copy file, unable to determine if changes were made. Aborting!")

        # Get the sheet name from the Excel file
        workbook_path = os.path.join(path, all_workbook_filename)
        with pd.ExcelFile(workbook_path) as xls:
            sheet_name = xls.sheet_names[0] if xls.sheet_names else "Sheet1"

        df1 = pd.read_excel(os.path.join(path, old_workbook_filename), sheet_name=0, index_col="Id")

        df2 = pd.read_excel(workbook_path, sheet_name=0, index_col="Id")

        if df1.equals(df2):
            error_and_exit("No differences detected.")

        app.logger.info("Changes detected in workbook. Processing updates...")
        # Need to strip out any net new rows before doing this comparison
        df3 = strip_any_net_new_rows(app, df2, all_workbook_filename, obj_type, path, new_workbook_filename, sheet_name)
        try:
            changes = compare_dataframes(df1, df3)
        except ValueError:
            changes = compare_dataframes(df1, df2)
        changes.to_csv(
            os.path.join(path, DIFFERENCES_FILE),
            header=True,
            index=True,
            sep=" ",
            mode="w+",
        )
        app.logger.info(
            "Please check differences.txt file located in %s to see changes made.",
            path,
        )
        upload_existing_data(app, api, path, obj_type, all_workbook_filename)
    else:
        app.logger.info("No files found for the specified type to load to RegScale.")
    return app.logger.info(
        obj_type + " files have been uploaded. Changes made to existing files can be seen in "
        "differences.txt file. Thank you!"
    )


def compare_dataframes(df1: "pd.DataFrame", df2: "pd.DataFrame") -> "pd.DataFrame":
    """
    Compare two DataFrames and return a DataFrame with the differences.

    :param pd.DataFrame df1: The first DataFrame to compare
    :param pd.DataFrame df2: The second DataFrame to compare
    :return: A DataFrame with the differences between the two DataFrames
    :rtype: pd.DataFrame
    """
    import numpy as np
    import pandas as pd

    diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
    ne_stacked = diff_mask.stack()
    changed = ne_stacked[ne_stacked]
    changed.index.names = ["Id", "Column"]
    difference_locations = np.nonzero(diff_mask)
    changed_from = df1.values[difference_locations]
    changed_to = df2.values[difference_locations]
    return pd.DataFrame({"From": changed_from, "To": changed_to}, index=changed.index)


@model.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
@click.option(
    "--model",
    type=click.Choice(
        [
            "assessment",
            "control",
            "issue",
            "component",
            "asset",
        ],  # , 'risk'], #, 'task'],
        case_sensitive=False,
    ),
    help="Specify the type of bulk load file to delete.",
    default="assessment",
    required=True,
)
def generate_delete_file(path: Path, model: str):
    """This command will delete files used during the Assessment editing process."""
    delete_file(path, model)


def delete_file(path: Path, obj_type: str) -> int:
    """
    Deletes files used during the process

    :param Path path: directory of file location
    :param str obj_type: The model type to download
    :return: Number of files deleted
    :rtype: int
    """
    log = create_logger()
    all_workbook_filename = get_workbook_title(obj_type, ALL_PRE, FILE_POST)
    old_workbook_filename = get_workbook_title(obj_type, OLD_PRE, FILE_POST)
    new_workbook_filename = get_workbook_title(obj_type, NEW_PRE, FILE_POST)
    file_names = [
        new_workbook_filename,
        all_workbook_filename,
        old_workbook_filename,
        DIFFERENCES_FILE,
    ]
    deleted_files = []

    for file_name in file_names:
        if os.path.isfile(path / file_name):
            os.remove(path / file_name)
            deleted_files.append(file_name)
        else:
            log.warning("No %s file found. Checking for other files before exiting.", file_name)
    log.info("%i file(s) have been deleted: %s", len(deleted_files), ", ".join(deleted_files))
    return len(deleted_files)


def upload_new_data(app: Application, path: Path, obj_type: str, workbook_filename: str) -> None:
    """
    This method reads in the spreadsheet filled with new records to upload into RegScale, converts
    them into the appropriate object type, and saves them in RegScale.

    :param Application app: The Application instance
    :param Path path: The path where the Excel file can be found
    :param str obj_type: The model type to load the records as
    :param str workbook_filename: The file name of the Excel spreadsheet
    :return: None
    :rtype: None
    """
    new_files = os.path.join(path, workbook_filename)
    wb_data = map_workbook_to_dict(new_files)
    load_objs = convert_dict_to_model(wb_data, obj_type)
    post_and_save_models(app, load_objs, path, obj_type, workbook_filename)


def strip_any_net_new_rows(
    app: Application,
    df: "pd.DataFrame",
    workbook_filename: str,
    obj_type: str,
    path: Path,
    new_workbook_filename: str,
    sheet_name: Optional[str] = None,
) -> "pd.DataFrame":
    """
    This method scans the loaded workbook for any new rows and strips them out to insert separately.

    :param Application app: The Application instance
    :param pd.DataFrame df: The DataFrame from the loaded workbook
    :param str workbook_filename: The file name of the Excel spreadsheet
    :param str obj_type: The model type to load the records as
    :param Path path: The path where the Excel file can be found
    :param str new_workbook_filename: The file name of the Excel spreadsheet with new records.
    :param Optional[str] sheet_name: The name of the worksheet being processed
    :return: pd.DataFrame The updated DataFrame, minus any new rows
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    df_updates = []
    df_inserts = []
    indexes = []
    columns = list(df.columns)
    obj = get_obj(obj_type)
    for x in df.index:
        if math.isnan(x):
            data_rec = {}
            for y in columns:
                data_rec[y] = df.at[x, y]
            df_inserts.append(convert_new_record_to_model(data_rec, obj_type, path, workbook_filename, sheet_name))
        else:
            indexes.append(x)
            data_rec = []
            for y in columns:
                data_rec.append(df.at[x, y])
            df_updates.append(data_rec)
    new_df = pd.DataFrame(df_updates, index=indexes, columns=columns)
    if len(df_inserts) > 0:
        if obj.is_new_excel_record_allowed():
            # Use workbook_filename (the actual file containing the data) instead of new_workbook_filename
            post_and_save_models(app, df_inserts, path, obj_type, workbook_filename)
        else:
            app.logger.warning(
                "New rows have been found in the Excel spreadsheet being loaded. New records for this model are not allowed."
            )

    return new_df


def convert_new_record_to_model(
    data_rec: dict, obj_type: str, path: Path, workbook_filename: str, sheet_name: Optional[str] = None
) -> object:
    """
    This method takes the new record found in the Excel file of existing records, and converts it
    into a model object for inserting into the database.

    :param dict data_rec: The new record data extracted from the Excel file
    :param str obj_type: The model type to load the records as
    :param Path path: The path where the Excel file can be found
    :param str workbook_filename: The file name of the Excel spreadsheet
    :param Optional[str] sheet_name: The name of the worksheet being processed
    :return: object
    :rtype: object
    :raises ValueError:
    """
    new_obj = {}
    for cur_field in obj_fields:
        new_obj[cur_field.field_name] = get_basic_field_value(cur_field, data_rec)
        new_obj[cur_field.field_name] = format_loaded_field_value(cur_field, new_obj[cur_field.field_name])
        if len(cur_field.lookup_field) > 0:
            match_value = new_obj[cur_field.field_name]
            if (match_value is None) & cur_field.required:
                message = f"{cur_field.column_name}: No value selected in new row. Please select a value from the list."
                # logger.error(message)
                raise ValueError(message)
            workbook_path = os.path.join(path, workbook_filename)
            new_obj[cur_field.field_name] = lookup_value_in_sheet(
                workbook_path, cur_field.lookup_field, match_value, cur_field.column_name
            )
        if new_obj[cur_field.field_name] is None:
            new_obj[cur_field.field_name] = generate_default_value_for_field(cur_field.field_name, cur_field.data_type)
        elif cur_field.data_type == "str":
            if not isinstance(new_obj[cur_field.field_name], str):
                new_obj[cur_field.field_name] = str(new_obj[cur_field.field_name])

    parse_parent_data(new_obj, sheet_name)

    return cast_dict_as_model(new_obj, obj_type)


def parse_parent_data(new_obj: dict, sheet_name: str) -> None:
    """
    Parse parentId and parentModule from worksheet name.

    :param dict new_obj: The new object to parse the parent info for
    :param str sheet_name: The worksheet name to parse
    :rtype: None
    """
    # Parse parentId and parentModule from sheet name if available
    if sheet_name:
        parent_id, parent_module = parse_parent_info_from_sheet_name(sheet_name)
        if parent_id is not None:
            new_obj["parentId"] = parent_id
        if parent_module is not None:
            new_obj["parentModule"] = parent_module


def generate_default_value_for_field(field_name: str, data_type: str) -> Any:
    """
    Generate a default value for a required field.

    :param str field_name: Name of the field to generate a default value for
    :param str data_type: the data type to generate a default value for
    :return Any: the default value to use when creating a new record
    :rtype Any:
    """
    if field_name == "id":
        return 0
    if field_name.find("Id") >= 0:
        return None
    if data_type == "int":
        return 0
    if data_type == "bool":
        return False
    if data_type == "str":
        return ""
    if data_type == "float":
        return 0.0


def parse_parent_info_from_sheet_name(sheet_name: str) -> tuple[Optional[int], Optional[str]]:
    """
    Parse parentId and parentModule from worksheet name.

    Expected format: Issue(46_securityplans
    Where:
    - Issue( is the model prefix
    - 46 is the parentId
    - securityplans is the parentModule

    :param str sheet_name: The worksheet name to parse
    :return: Tuple of (parentId, parentModule), or (None, None) if pattern doesn't match
    :rtype: tuple[Optional[int], Optional[str]]
    """
    if not sheet_name or "(" not in sheet_name or "_" not in sheet_name:
        return None, None

    try:
        # Find the opening parenthesis
        paren_index = sheet_name.index("(")
        # Get the part after the parenthesis
        after_paren = sheet_name[paren_index + 1 :]

        # Split by underscore
        if "_" in after_paren:
            parts = after_paren.split("_", 1)  # Split on first underscore only
            parent_id = int(parts[0])
            parent_module = parts[1]
            return parent_id, parent_module
    except (ValueError, IndexError):
        # If parsing fails, return None values
        pass

    return None, None


# pylint: disable=E1136,R0914
def upload_existing_data(app: Application, api: Api, path: Path, obj_type: str, workbook_filename: str) -> None:
    """
    This method reads in the spreadsheet filled with existing records to update in RegScale
    using the RegScaleModel save() and bulk_save() methods.

    :param Application app: The Application instance
    :param Api api: The instance api handler
    :param Path path: The path where the Excel file can be found
    :param str obj_type: The model type to load the records as
    :param str workbook_filename: The file name of the Excel spreadsheet
    :return: None
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    obj = get_obj(obj_type)
    # Loading in differences.txt file and using Id to parse xlsx file for rows to update

    diff = pd.read_csv(os.path.join(path, DIFFERENCES_FILE), header=0, sep=" ", index_col=None)
    ids = []

    changes = []
    for _, row in diff.iterrows():
        row_chgs = {}
        ids.append(row["Id"])
        row_chgs["id"] = row["Id"]
        row_chgs["column"] = row["Column"]
        row_chgs["value"] = row["To"]
        changes.append(row_chgs)

    logger.debug(changes)
    id_df = pd.DataFrame(ids, index=None, columns=["Id"])
    id_df2 = id_df.drop_duplicates()
    logger.info(f"Found {len(id_df2)} unique {obj_type} ID(s) with changes: {id_df2['Id'].tolist()}")

    updated_files = os.path.join(path, workbook_filename)
    df3 = pd.read_excel(updated_files, sheet_name=0, index_col=None)
    logger.debug(f"Read {len(df3)} total rows from Excel file")

    updated = df3[df3["Id"].isin(id_df2["Id"])]
    logger.info(f"Filtered to {len(updated)} {obj_type}(s) matching changed IDs")

    if len(updated) == 0:
        logger.error(
            f"No {obj_type}s found in Excel file matching the IDs in differences.txt. "
            f"Expected IDs: {id_df2['Id'].tolist()}. "
            f"This usually means the Excel file doesn't contain these records."
        )
        return

    updated = map_workbook_to_dict(updated_files, updated)
    logger.debug(f"Converted to dictionary with {len(updated)} entries")
    config = app.config

    # Load existing model instances from API
    load_objs = load_model_for_id(api, updated, config["domain"] + obj.get_endpoint("get"), obj_type)

    # Apply changes to model instances and queue for bulk update
    modified_objects = []
    for cur_obj in load_objs:
        # Apply Excel changes to the model instance
        modified_obj = find_and_apply_changes(cur_obj, changes, updated)

        # Ignore change tracking to ensure all updates are saved
        modified_obj._ignore_has_changed = True
        # Queue the instance for bulk update
        modified_obj.save(bulk=True)
        modified_objects.append(modified_obj)

    # Execute bulk update using the model class
    if modified_objects:
        app.logger.info("Executing bulk update for %i %s(s)...", len(modified_objects), obj_type)
        model_class = type(modified_objects[0])
        results = model_class.bulk_save()

        updated_count = len(results.get("updated", []))
        created_count = len(results.get("created", []))

        app.logger.info(
            "Bulk operation completed: Updated %i %s(s), Created %i %s(s)",
            updated_count,
            obj_type,
            created_count,
            obj_type,
        )


# pylint: enable=E1136,R0914


def find_and_apply_changes(cur_object: object, changes: list, updates: dict) -> object:
    """
    This method looks through the changes and applies those that should be applied to
    the current model instance.

    :param object cur_object: the current model instance being updated
    :param list changes: a list of the specific changes to apply
    :param dict updates: a dictionary of updated models to be applied to the current object(s)
    :return: object the updated model instance
    :rtype: object
    """
    for cur_change in changes:
        if cur_change["id"] == cur_object.id:
            field_def = get_field_def_for_column(cur_change["column"])
            if field_def is None:
                logger.warning(
                    f"Column '{cur_change['column']}' not found in model fields for {type(cur_object).__name__} "
                    f"ID {cur_object.id}. Change will be skipped."
                )
                continue
            if len(field_def.lookup_field) > 0:
                value = check_empty_nan(extract_update_for_column(field_def.field_name, cur_change["id"], updates))
                setattr(cur_object, field_def.field_name, value)
            else:
                field_name = get_field_name_for_column(cur_change["column"])
                if not field_name:
                    logger.warning(
                        f"Could not find field name for column '{cur_change['column']}' in {type(cur_object).__name__} "
                        f"ID {cur_object.id}. Change will be skipped."
                    )
                    continue
                value = check_empty_nan(cur_change["value"])
                logger.debug(
                    f"Applying change to {type(cur_object).__name__} ID {cur_object.id}: "
                    f"{field_name} = {value} (was: {getattr(cur_object, field_name, 'N/A')})"
                )
                setattr(cur_object, field_name, value)
    return cur_object


def extract_update_for_column(field_name: str, rec_id: int, updates: dict) -> Any:
    """
    This method will look through the updated record dictionary and extract the updated
    value for the field being updated.

    :param str field_name: The name of the field to get the updated value for
    :param int rec_id: The id of the record to be updated
    :param dict updates: a dictionary of updates
    :return Any: The updated model to be applied to the database
    :rtype Any:
    """
    update_keys = updates.keys()
    for cur_key in update_keys:
        cur_update = updates[cur_key]
        if ("Id" in cur_update.keys()) & (field_name in cur_update.keys()) & (cur_update["Id"] == rec_id):
            return cur_update[field_name]
    return None


def get_field_name_for_column(column_label: str) -> str:
    """
    This method iterates through the fields and finds the matching column.

    :param str column_label: The column label to find the field name for
    :return: str the field name to use
    :rtype: str
    """
    for cur_field in obj_fields:
        if cur_field.column_name == column_label:
            return cur_field.field_name
    return ""


def get_field_def_for_column(column_label: str) -> Any:
    """
    This method iterates through the fields and finds the matching column,
    then returns the column configuration build from the model definition.

    :param str column_label: The column label to find the field definition for
    :return Any: the field definition
    :rtype Any:
    """
    for cur_field in obj_fields:
        if cur_field.column_name == column_label:
            return cur_field
    return None


# pylint: disable=R0913
def post_and_save_models(
    app: Application,
    new_models: list,
    workbook_path: Path,
    obj_type: str,
    load_file_name: str,
) -> None:
    """
    Function to post new records to RegScale and save record ids to excel workbook.
    Uses the RegScaleModel .create() method for new objects.

    :param Application app: RegScale CLI Application object
    :param list new_models: List of new records to post to RegScale
    :param Path workbook_path: Path to workbook to save assessment ids to
    :param str obj_type: the model type to upload
    :param str load_file_name: The file name of the Excel file to update with record IDs
    :return: None
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    try:
        # Create new objects using .create() method
        new_objs = []
        for cur_obj in new_models:
            # Use .create() for new objects (id=0 or None)
            cur_obj._ignore_has_changed = True
            new_obj = cur_obj.create()
            cur_obj.create_new_connecting_model(new_obj)
            new_objs.append(cur_obj)

        # Save IDs and all other fields to Excel
        if new_objs:
            # Create a list of dicts with all field values from created objects
            obj_data = []
            for obj in new_objs:
                obj_dict = {"id_number": obj.id}
                # Add all fields from obj_fields to ensure we capture API-populated fields
                for field in obj_fields:
                    field_value = getattr(obj, field.field_name, None)
                    if field_value is not None:
                        obj_dict[field.field_name] = field_value
                obj_data.append(obj_dict)

            new_objs_df = pd.DataFrame(obj_data)
            for file_name in [load_file_name]:
                with pd.ExcelWriter(
                    os.path.join(workbook_path, file_name),
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                ) as writer:
                    new_objs_df.to_excel(
                        writer,
                        sheet_name=obj_type + "_Ids",
                        index=False,
                    )
            app.logger.info("%i total %s(s) were added to RegScale.", len(new_objs), obj_type)
    except Exception as e:
        app.logger.error(e)


# pylint: enable=R0913


def map_pandas_timestamp(date_time: "pd.Timestamp") -> Optional[str]:
    """
    Function to map pandas timestamp to string

    :param pd.Timestamp date_time:
    :return: String representation of pandas timestamp
    :rtype: Optional[str]
    """
    import pandas as pd  # Optimize import performance

    if pd.isnull(date_time):
        return None
    if isinstance(date_time, float):
        return None
    if date_time is not None and not pd.isna(date_time) and not isinstance(date_time, str):
        return date_time.strftime("%Y-%m-%d %H:%M:%S")
    return date_time or None


def load_model_for_id(api: Api, wb_data: dict, url: str, obj_type: str) -> list:
    """
    This method loads the current record for the updated objects and returns model instances.

    :param Api api: the API object instance to use
    :param dict wb_data: The submitted workbook data in a dict
    :param str url: the base url to use to retrieve the model data
    :param str obj_type: The model type to cast the data to
    :return: list of model instances of the specified type
    :rtype: list
    """
    load_data = []
    failed_loads = []

    logger.info(f"Loading {len(wb_data)} {obj_type}(s) from API for update...")

    for cur_obj in wb_data:
        obj = wb_data[cur_obj]
        cur_id = int(obj["Id"])
        if cur_id > 0:
            url_to_use = url.replace("{id}", str(cur_id))
            url_to_use = check_url_for_double_slash(url_to_use)
            logger.debug(f"Fetching {obj_type} ID {cur_id} from {url_to_use}")
            result = api.get(url_to_use)
            if result.status_code == 200:
                dict_data = result.json()
                model_instance = cast_dict_as_model(dict_data, obj_type)
                load_data.append(model_instance)
                logger.debug(f"Successfully loaded {obj_type} ID {cur_id}")
            else:
                failed_loads.append((cur_id, result.status_code))
                logger.warning(
                    f"Failed to load {obj_type} ID {cur_id} from API. Status code: {result.status_code}. "
                    f"This record will not be updated."
                )

    if failed_loads:
        logger.warning(
            f"Failed to load {len(failed_loads)} {obj_type}(s) from API: "
            f"{', '.join([f'ID {id} (HTTP {code})' for id, code in failed_loads])}"
        )

    logger.info(f"Successfully loaded {len(load_data)} {obj_type}(s) from API for update.")
    return load_data


def check_url_for_double_slash(url: str) -> str:
    """
    This method checks URLs for a double slash in the wrong place.

    :param str url: the base url to be checked for double slash characters
    :return str: the url without double slashes
    :rtype str:
    """
    protocol_part = url[: url.find("://") + 3]
    remainder = url[url.find("://") + 3 :]
    remainder = remainder.replace("//", "/")
    return protocol_part + remainder


def convert_dict_to_model(wb_data: dict, obj_type: str) -> list:
    """
    This method converts the workbook dict to match what the model expects and then
    casts the resulting dict as an instance of the specified model.

    :param dict wb_data: The submitted workbook data in a dict
    :param str obj_type: The model type to instantiate
    :return: list instance of the specified model, populated with the dict
    :rtype: list
    """
    loaded_data = []
    for cur_obj in wb_data:
        cur_data = wb_data[cur_obj]
        new_obj = {}
        for cur_field in obj_fields:
            new_obj[cur_field.field_name] = get_basic_field_value(cur_field, cur_data)
            new_obj[cur_field.field_name] = format_loaded_field_value(cur_field, new_obj[cur_field.field_name])
            if new_obj[cur_field.field_name] is None:
                if cur_field.field_name == "id":
                    new_obj[cur_field.field_name] = 0
            elif cur_field.data_type == "str":
                if not isinstance(new_obj[cur_field.field_name], str):
                    new_obj[cur_field.field_name] = str(new_obj[cur_field.field_name])

        loaded_data.append(cast_dict_as_model(new_obj, obj_type))
    return loaded_data


def get_basic_field_value(cur_field: FieldMakeup, cur_wb_data: dict) -> str:
    """
    Lookup and return the basic value for the field.

    :param FieldMakeup cur_field: The current field metadata
    :param dict cur_wb_data: The collection of values from the Workbook
    :return: str
    :rtype: str
    """
    if cur_field.field_name in cur_wb_data.keys():
        return cur_wb_data[cur_field.field_name]
    if cur_field.column_name in cur_wb_data.keys():
        return cur_wb_data[cur_field.column_name]
    return ""


def format_loaded_field_value(cur_field: FieldMakeup, cur_value: Any) -> Any:
    """
    Format the current value based on the field data type

    :param FieldMakeup cur_field: The current field metadata
    :param Any cur_value: The current value of the field
    :return: Any the return value type depends on the field metadata
    :rtype: Any
    """
    if cur_field.treat_as_date:
        return map_pandas_timestamp(cur_value)
    else:
        return check_empty_nan(cur_value)


def cast_dict_as_model(obj_data: dict, obj_type: str) -> object:
    """
    This method uses the .from_dict() method on each of the respective models to
    instantiate them with the submitted data.

    :param dict obj_data: The submitted workbook data in a dict
    :param str obj_type: The model type to instantiate
    :return: object instance of the specified model, populated with the dict
    :rtype: object
    """
    rtn_obj = None
    if obj_type == "assessment":
        rtn_obj = Assessment.from_dict(obj_data)
    elif obj_type == "control":
        rtn_obj = Control.from_dict(obj_data)
    elif obj_type == "issue":
        rtn_obj = Issue.from_dict(obj_data)
    elif obj_type == "asset":
        rtn_obj = Asset.from_dict(obj_data)
    elif obj_type == "component":
        rtn_obj = Component.from_dict(obj_data)
    elif obj_type == "risk":
        rtn_obj = Risk.from_dict(obj_data)
    # elif obj_type == "task":
    #     rtn_obj = None
    return rtn_obj


def map_workbook_to_dict(file_path: str, workbook_data: Optional["pd.DataFrame"] = None) -> dict:
    """
    Function to map workbook to dictionary

    :param str file_path: Path to workbook file
    :param Optional[pd.DataFrame] workbook_data: Dataframe to map to dictionary
    :return: dict representation of workbook
    :rtype: dict
    """
    return map_workbook_to_lookups(file_path, workbook_data).T.to_dict()


def map_workbook_to_lookups(file_path: str, workbook_data: Optional["pd.DataFrame"] = None) -> "pd.DataFrame":
    """
    Function to map workbook to dictionary

    :param str file_path: Path to workbook file
    :param Optional[pd.DataFrame] workbook_data: Dataframe to map to dictionary
    :return: pd.DataFrame representation of workbook
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    if workbook_data is not None:
        wb_data = workbook_data
    else:
        wb_data = pd.read_excel(file_path)

    # Only drop rows where ALL values are NaN (completely empty rows)
    # Don't drop rows with some NaN values - those are legitimate records with optional empty fields
    wb_data = wb_data.dropna(how="all")
    for cur_row in obj_fields:
        if len(cur_row.lookup_field) > 0 and cur_row.lookup_field != "module":
            if cur_row.column_name in wb_data.columns:
                lookup_wb = pd.read_excel(file_path, sheet_name=cur_row.column_name)
                if cur_row.lookup_field == "user":
                    lookup_wb = lookup_wb.rename(
                        columns={
                            "User": cur_row.column_name,
                            "UserId": cur_row.field_name,
                        }
                    )
                else:
                    lookup_wb = lookup_wb.rename(
                        columns={
                            "name": cur_row.column_name,
                            "id": cur_row.field_name,
                        }
                    )
                    lookup_wb[cur_row.column_name] = lookup_wb[cur_row.column_name].astype(
                        str
                    )  # Ensure consistent data type
                wb_data = wb_data.merge(
                    lookup_wb,
                    how="left",
                    on=cur_row.column_name,
                    validate="many_to_many",
                )
    return wb_data


def lookup_value_in_sheet(file_path: str, lookup_field: str, match_value: str, sheet_name: str) -> Any:
    """
    This method looks up the specified value in the specified sheet to get the corresponding value.

    :param str file_path: Path to workbook file
    :param str lookup_field: The field being looked up
    :param str match_value: The value to match against the lookup sheet
    :param str sheet_name: The name of the lookup sheet to use
    :return: Any The lookup value
    :rtype: Any
    """
    import pandas as pd  # Optimize import performance

    logger.debug("Looking up value in sheet - Field = {}".format(lookup_field))
    if lookup_field == "module":
        return match_value
    match_col = "name"
    val_col = "id"
    if lookup_field == "user":
        logger.debug("Lookup user field")
        match_col = "User"
        val_col = "UserId"
    try:
        lookup_wb = pd.read_excel(file_path, sheet_name=sheet_name)
    except ValueError:
        return None
    for x in lookup_wb.index:
        lookup_val = lookup_wb.at[x, match_col]
        if lookup_val == match_value:
            logger.debug("Found Match!")
            return lookup_wb.at[x, val_col]
    if lookup_field == "user":
        logger.debug("Lookup user field returning empty string!")
        return ""
    return 0


def put_data_into_df(obj_list: list) -> "pd.DataFrame":
    """
    This method takes the passed data and loads it into a data frame for
    converting into an Excel spreadsheet.

    :param list obj_list: a list of records for loading into the data frame
    :return: pd.DataFrame
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    df = []
    headers = []
    # build a list of column headers
    for cur_field in obj_fields:
        if cur_field.sort_order >= 0:
            headers.append(cur_field.column_name)
    for cur_obj in obj_list:
        cur_row = []
        for cur_field in obj_fields:
            # If the value is a foreign key, we need to look up the display value
            if cur_field.sort_order >= 0:
                cur_row.append(get_field_lookup_value(cur_field, cur_obj))
        df.append(cur_row)
    return pd.DataFrame(df, columns=headers)


def get_field_lookup_value(cur_field: FieldMakeup, cur_obj: dict) -> str:
    """
    Looks up the lookup value for the specified field in the passed in dict of lookup values

    :param FieldMakeup cur_field: the current field metadata
    :param dict cur_obj: a dict of the lookup values
    :return: str the value to append
    :rtype: str
    """
    if cur_field.field_name in cur_obj.keys():
        if len(cur_field.lookup_field) > 0:
            return lookup_value(cur_field.lookup_field, cur_obj[cur_field.field_name])
        else:
            return cur_obj[cur_field.field_name]
    return ""


def lookup_value(lookup_field: str, lookup_value_str: str) -> str:
    """
    This method looks up the display value for a foreign key value.

    :param str lookup_field: the field name of the lookup field
    :param str lookup_value_str: the foreign key value to look up
    :return: str the display value to present to the user
    :rtype: str
    """
    if lookup_value_str is not None:
        df = lookup_dfs[lookup_field]
        if lookup_field == "user":
            lookup_col = "UserId"
            return_col = "User"
        elif lookup_field == "module":
            lookup_col = "name"
            return_col = "name"
        else:
            lookup_col = "id"
            return_col = "name"
        if len(df) > 0:
            for cur_row in df.itertuples():
                cur_row_dict = getattr(cur_row, lookup_col)
                if lookup_value_str == cur_row_dict:
                    return getattr(cur_row, return_col)
    return lookup_value_str


# pylint: disable=R0912,R0915,R0914
def build_workbook(path: str, workbook_filename: str, workbook_title: str) -> Workbook:  # noqa C901
    """
    This method creates the Excel workbook, populating it with the appropriate
    column headings and lookup sheets.

    :param str path: The folder in which to save the created workbook
    :param str workbook_filename: The filename to use for the generated workbook
    :param str workbook_title: The title to use for the primary sheet in the workbook
    :return: Workbook The generated workbook
    :rtype: Workbook
    """
    workbook_sheets = []
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = workbook_title
    column_headers = build_header_list()
    for col, val in enumerate(column_headers, start=1):
        worksheet.cell(row=1, column=col).value = val
        worksheet.cell(row=1, column=col).font = Font(bold=True)
        set_col_for_field(worksheet.cell(row=1, column=col).column_letter, val)

    # create and format reference worksheets for dropdowns
    for cur_field in obj_fields:
        if cur_field.sort_order >= 0:
            if len(cur_field.lookup_field) > 0:
                workbook.create_sheet(title=cur_field.column_name)
                workbook_sheets.append(cur_field.column_name)
            if cur_field.treat_enum_as_lookup:
                workbook.create_sheet(title=cur_field.column_name)
                workbook_sheets.append(cur_field.column_name)

    workbook.save(filename=os.path.join(path, workbook_filename))

    build_workbook_lookup_sheets(path, workbook_filename)

    return build_workbook_data_validations(path, workbook_filename, workbook_sheets)


def build_workbook_lookup_sheets(path: str, workbook_filename: str) -> None:
    """
    This method builds the workbook supporting sheets with lookup data.

    :param str path: The folder in which to save the created workbook
    :param str workbook_filename: The filename to use for the generated workbook
    :return: None
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    # Pull in reference data for drop-downs
    with pd.ExcelWriter(
        os.path.join(path, workbook_filename),
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        for cur_field in obj_fields:
            if cur_field.sort_order >= 0:
                if len(cur_field.lookup_field) > 0:
                    get_data_frame(cur_field.lookup_field).to_excel(
                        writer,
                        sheet_name=cur_field.column_name,
                        index=False,
                    )
                if cur_field.treat_enum_as_lookup:
                    create_enum_data_frame(cur_field.enum_values).to_excel(
                        writer,
                        sheet_name=cur_field.column_name,
                        index=False,
                    )


def build_workbook_data_validations(path: str, workbook_filename: str, workbook_sheets: list) -> Workbook:
    """
    This method builds the data validations for the workbook being built

    :param str path: The folder in which to save the created workbook
    :param str workbook_filename: The filename to use for the generated workbook
    :param list workbook_sheets: The list of workbook sheets to create as lookups
    :return: Workbook The generated workbook
    :rtype: Workbook
    """
    workbook = load_workbook(os.path.join(path, workbook_filename))
    worksheet = workbook.active

    for sheet in workbook_sheets:
        workbook[sheet].protection.sheet = True

    # create data validations for enum values
    data_validations_info = []
    date_cols = []
    edit_cols = []
    for cur_field in obj_fields:
        if cur_field.sort_order >= 0:
            if len(cur_field.lookup_field) > 0 or cur_field.treat_enum_as_lookup:
                dv_info = {
                    "sheet": cur_field.column_name,
                    "columns": [cur_field.cell_col],
                    "allow_blank": True,
                }
                data_validations_info.append(dv_info)
            elif len(cur_field.enum_values) > 0:
                val_str = ""
                for cur_enum in cur_field.enum_values:
                    val_str += cur_enum + ", "
                val_str = val_str[: len(val_str) - 2]
                dv_info = {
                    "formula1": '"' + val_str + '"',
                    "columns": [cur_field.cell_col],
                    "allow_blank": True,
                }
                data_validations_info.append(dv_info)
            elif cur_field.treat_as_date:
                dv_info = {
                    "type": "date",
                    "columns": [cur_field.cell_col],
                    "allow_blank": False,
                }
                data_validations_info.append(dv_info)
                date_cols.append(cur_field.cell_col)
            elif cur_field.field_name != "id":
                edit_cols.append(cur_field.cell_col)

    create_data_validations(
        data_validations_info=data_validations_info,
        workbook=workbook,
        worksheet=worksheet,
    )
    workbook.save(filename=os.path.join(path, workbook_filename))

    return set_date_style_to_workbook(path, workbook_filename, edit_cols, date_cols)


# pylint: enable=R0912,R0915,R0914
def set_date_style_to_workbook(path: str, workbook_filename: str, edit_cols: list, date_cols: list) -> Workbook:
    """
    This method sets the date style and header row freeze on the specified workbook.

    :param str path: The folder in which to save the created workbook
    :param str workbook_filename: The filename to use for the generated workbook
    :param list edit_cols: a list of columns that should be protected
    :param list date_cols: a list of columns that should be formatted as date
    :return: Workbook The generated workbook
    :rtype: Workbook
    """
    # Freezing top row and adding data style to date columns to assure validation
    workbook = load_workbook(os.path.join(path, workbook_filename))
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
    workbook.add_named_style(date_style)

    for col in date_cols:  # Columns to edit
        for cell in worksheet[col]:
            if cell.row > 1:
                cell.style = date_style

    # Adjusting width of columns
    adjust_column_widths_and_styles(worksheet, edit_cols, date_cols, date_style)

    workbook.save(filename=os.path.join(path, workbook_filename))
    return workbook


def get_maximum_rows(*, sheet_object: object) -> int:
    """
    This function finds the last row containing data in a spreadsheet

    :param object sheet_object: excel worksheet to be referenced
    :return: int representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(any(col.value is not None for col in row) for max_row, row in enumerate(sheet_object, 1))


def get_workbook_title(obj_type: str, prefix: str, postfix: str) -> str:
    """
    This method generates the name for the new workbook being generated.

    :param str obj_type: The model type to generate a workbook for
    :param str prefix: The prefix to use when creating the new workbook
    :param str postfix: The postfix to use when creating the new workbook
    :return: str The name of the new workbook being generated
    :rtype: str
    """
    return prefix + convert_property_to_column_label(obj_type) + postfix


def get_all_by_parent(obj_type: str, parent_id: int, parent_module: str) -> list:
    """
    Uses the get_all_by_parent() method on the appropriate model to get the data

    :param str obj_type: the type of model to return
    :param int parent_id: the parent id to use to retrieve the records
    :param str parent_module: the parent module to use for retrieving the records
    :return: list of records
    :rtype: list
    """
    rtn_list = []
    if obj_type == "assessment":
        rtn_list = Assessment.get_all_by_parent(parent_id, parent_module)
    elif obj_type == "control":
        rtn_list = ControlImplementation.get_all_by_parent(parent_id, parent_module)
    elif obj_type == "issue":
        rtn_list = Issue.get_all_by_parent(parent_id, parent_module)
    elif obj_type == "asset":
        rtn_list = Asset.get_all_by_parent(parent_id, parent_module)
    elif obj_type == "component":
        rtn_list = Component.get_all_by_parent(parent_id, parent_module)
    elif obj_type == "risk":
        rtn_list = Risk.get_all_by_parent(parent_id, parent_module)
    # elif obj_type == "task":
    #     return []
    return convert_all_to_dict(rtn_list)


def convert_all_to_dict(objs: list) -> list:
    """
    Converts a list of model objects to a list of dictionaries.

    :param list objs: List of objects to convert to dict
    :return list: List of dict
    :return list:
    """
    rtn_list = []
    for obj in objs:
        dict_obj = obj.dict()
        rtn_list.append(dict_obj)
    return rtn_list


def get_all_by_query(obj: object, parent_id: int, parent_module: str, app: Application) -> list:
    """
    Uses the get_export_query() method on the model to get the data

    :param object obj: the instance of the model type specified
    :param int parent_id: the parent id to use to retrieve the records
    :param str parent_module: the parent module to use for retrieving the records
    :param Application app: the application object
    :return: list the collection of records
    :rtype: list
    """
    return obj.get_export_query(app, parent_id, parent_module)


def get_obj(obj_type: str) -> object:
    """
    Returns an instance of the object type specified by the user.

    :param str obj_type: the type of model to return
    :return: object of the mode specified
    :rtype: object
    """
    object_mapping = {
        "assessment": Assessment(),
        "control": ControlImplementation(controlOwnerId="", status="", controlID=0),
        "issue": Issue(),
        "asset": Asset(name="", assetType="", status="", assetCategory=""),
        "component": Component(title="", description="", componentType=ComponentType.ComplianceArtifact),
        "risk": Risk(),
    }
    if obj_type in object_mapping:
        obj = object_mapping[obj_type]
        build_object_field_list(obj)
        return obj
    return None


def build_object_field_list(obj: object) -> None:
    """
    This method examines the instantiated model and extracts the list of fields and
    other information needed for the model processing.

    :param object obj: This should be a model object that is descended from RegScaleModel
    :return: None
    :rtype: None
    """
    # Build the list of fields for the model type
    pos_dict = obj.get_sort_position_dict()
    field_names = obj.__class__.model_fields.keys()
    extra_fields = obj.get_extra_fields()
    include_field_list = obj.get_include_fields()
    for item in include_field_list:
        include_fields.append(item)
    for cur_field in field_names:
        if cur_field not in exclude_fields:
            field_makeup = FieldMakeup(
                cur_field,
                convert_property_to_column_label(cur_field),
                get_field_data_type(obj.__class__.model_fields[cur_field]),
            )
            field_makeup.sort_order = find_sort_pos(cur_field, pos_dict)
            field_makeup.enum_values = obj.get_enum_values(cur_field)
            field_makeup.treat_enum_as_lookup = should_treat_enum_as_lookup(field_makeup.enum_values)
            field_makeup.lookup_field = obj.get_lookup_field(cur_field)
            field_makeup.treat_as_date = obj.is_date_field(cur_field)
            field_makeup.required = is_field_required(obj, cur_field)
            if field_makeup.sort_order >= 0:
                obj_fields.append(field_makeup)
    for cur_field in extra_fields:
        field_makeup = FieldMakeup(cur_field, convert_property_to_column_label(cur_field), "str")
        field_makeup.sort_order = find_sort_pos(cur_field, pos_dict)
        if field_makeup.sort_order >= 0:
            obj_fields.append(field_makeup)
    obj_fields.sort(key=attrgetter("sort_order"))


def should_treat_enum_as_lookup(enum_values: list) -> bool:
    """
    This method concatenates the list of enums together into a single string to get the total
    length and determine if it is too long to be a treated as an enum in the workbook.

    :param list enum_values: the list of enum values for this field
    :return: bool indicating if the list is too long to treat as an enum
    :rtype: bool
    """
    enum_str = ""
    for cur_val in enum_values:
        if len(enum_str) > 0:
            enum_str += ", "
        enum_str += cur_val
    if len(enum_str) > 256:
        return True
    return False


def get_field_data_type(field_info: FieldInfo) -> str:
    """
    determine the data type of the field from the field info annotation.

    :param FieldInfo field_info: The field annotation taken from the model object
    :return: str the data type to use for the field
    :rtype: str
    """
    if field_info.annotation == dict:
        return "dict"
    if field_info.annotation in (int, Union[int, None]):
        return "int"
    if field_info.annotation in (bool, Union[bool, None]):
        return "bool"
    if field_info.annotation in (str, Union[str, None], Union[str, int, None]):
        return "str"
    if field_info.annotation == Union[float, None]:
        return "float"
    return "enum"


def is_field_required(obj: object, field_name: str) -> bool:
    """
    Determine if the field is required from the annotation. If its a Union, odds are
    that one of the options is None, so assume it's not required. If it's a simple data type
    then assume its required.

    :param object obj: The object to check to determine if the field is required
    :param str field_name: The field name to be checked
    :return: bool indicating if the field is required
    :rtype: bool
    """
    field_info = obj.__class__.model_fields[field_name]
    if field_info.annotation == dict:
        return True
    if field_info.annotation == int:
        return True
    if field_info.annotation == bool:
        return True
    if field_info.annotation == str:
        return True
    return obj.is_required_field(field_name)


def convert_property_to_column_label(field_name: str) -> str:
    """
    This method takes the property name from the model and converts it into
    a column header label by capitalizing the first letter and removing "Id" from
    the end (assuming these are foreign keys)

    :param str field_name: The property name as read from the model
    :return: str a formatted column header label
    :rtype: str
    """
    rtn_field_name = field_name
    field_ext = field_name[len(field_name) - 2 :]
    if field_ext == "Id":
        rtn_field_name = field_name[: len(field_name) - 2]
    rtn_field_name = rtn_field_name[:1].upper() + rtn_field_name[1:]
    return rtn_field_name


def set_col_for_field(col: str, col_name: str) -> None:
    """
    This method updates the list of fields for the model being processed with the
    Excel column letter(s).

    :param str col: The letter(s) column
    :param str col_name: The column name
    :return: None
    :rtype: None
    """
    for cur_field in obj_fields:
        if col_name == cur_field.column_name:
            cur_field.cell_col = col


def find_sort_pos(field_name: str, pos_dict: dict) -> int:
    """
    This method is to be called from the derived classes to find and return the
    sort position in the specified dict of field names and positions.

    :param str field_name: The property name to specify the sort of
    :param dict pos_dict: a dict of field names and positions
    :return: int
    :rtype: int
    """
    # A return value of -1 supress the value, leaving it out of the generated spreadsheet
    rtn_value = -1
    if field_name in pos_dict.keys():
        rtn_value = pos_dict[field_name]
    return rtn_value


def build_header_list() -> list:
    """
    This method iterates through the list of fields and builds a list of column headers.

    :return: list of str
    :rtype: list
    """
    headers = []
    for cur_field in obj_fields:
        if cur_field.sort_order >= 0:
            logger.debug(cur_field.column_name + " - " + str(cur_field.sort_order))
            headers.append(cur_field.column_name)
    return headers


def get_data_frame(field_name: str) -> "pd.DataFrame":
    """
    This method retrieves a data frame populated with the lookup data
    for a foreign key field.

    :param str field_name: the field name of the foreign key data
    :return: pd.DataFrame the populated data frame
    :rtype: pd.DataFrame
    """
    if field_name == "user":
        df = get_user_names()
        lookup_dfs["user"] = df
        return df
    if field_name == "module":
        df = get_module_list()
        lookup_dfs["module"] = df
        return df
    df = get_field_names(field_name)
    lookup_dfs[field_name] = df
    return df


def create_enum_data_frame(enum_list: list) -> "pd.DataFrame":
    """
    This method will take a list of enum values are format them as a Data Frame.

    :param list enum_list: The list of enum values
    :return: pd.DataFrame the populated data frame
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    return pd.DataFrame(enum_list, columns=["name"])


def get_facility_list() -> "pd.DataFrame":
    """
    This method returns a list of facilities in a data frame

    :return: pd.DataFrame the populated data frame
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    fac_list = Facility.get_list()
    field_names = [[i["name"], i["id"]] for i in fac_list]
    all_names = pd.DataFrame(field_names, index=None, columns=["name", "id"])

    return all_names


def get_module_list() -> "pd.DataFrame":
    """
    This method returns a list of modules in a data frame

    :return: pd.DataFrame the populated data frame
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    list_of_modules = Modules().api_names()
    return pd.DataFrame(list_of_modules, columns=["name"])


def get_field_names(field_name: str) -> "pd.DataFrame":
    """
    This function uses GraphQL to retrieve all names of a given parent table in database

    :param str field_name: the foreign key table to retrieve
    :return: pandas dataframe with facility names
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    api = Api()

    body = """
    query {
        field_name(skip: 0, take: 50, order: {name: ASC}, ) {
            items {
                name
                id
            }
            totalCount
            pageInfo {
                hasNextPage
            }
        }
    }
    """.replace(
        "field_name", field_name
    )

    field_items = api.graph(query=body)
    names = field_items[str(field_name)]["items"]
    field_names = [[i["name"], i["id"]] for i in names]
    all_names = pd.DataFrame(field_names, index=None, columns=["name", "id"])

    return all_names


def create_data_validations(data_validations_info: list, workbook: Workbook, worksheet: Worksheet) -> None:
    """
    Function to create data validations for excel worksheet

    :param list data_validations_info: List containing dictionaries with
        information for data validations
    :param Workbook workbook: Workbook object to add data validations to
    :param Worksheet worksheet: The worksheet object to add data validations to
    :return: None
    :rtype: None
    """
    for _, dv_info in enumerate(data_validations_info, start=1):
        formula1 = dv_info.get("formula1")
        if sheet_name := dv_info.get("sheet"):
            formula1 = f"={sheet_name}!$A$2:$A${str(get_maximum_rows(sheet_object=workbook[sheet_name]))}"

        data_validation = DataValidation(
            type=dv_info.get("type", "list"),
            formula1=formula1,
            allow_blank=dv_info.get("allow_blank", True),
            showDropDown=False,
            error=(SELECTION_ERROR if dv_info.get("type", "list") == "list" else INVALID_ENTRY_ERROR),
            errorTitle=INVALID_ENTRY_TITLE,
            prompt=(SELECT_PROMPT if dv_info.get("type", "list") == "list" else DATE_ENTRY_PROMPT),
            showErrorMessage=True if dv_info.get("type", "date") else None,
            showInputMessage=True if dv_info.get("type", "date") else None,
        )

        worksheet.add_data_validation(data_validation)
        for column in dv_info["columns"]:
            data_validation.add(f"{column}2:{column}1048576")


def adjust_column_widths_and_styles(
    worksheet: Worksheet,
    editable_columns: Optional[list[str]] = None,
    date_columns: Optional[list[str]] = None,
    date_col_style: Optional[NamedStyle] = None,
) -> None:
    """
    Function to adjust column widths based on length of data in column, and apply
    styles to specific columns and rows

    :param Worksheet worksheet: Worksheet to adjust column widths for
    :param Optional[list[str]] editable_columns: List of rows to unlock for editing
    :param Optional[list[str]] date_columns: List of columns to add date style to
    :param Optional[NamedStyle] date_col_style: NamedStyle object to apply to date columns, defaults to None
    :return: None
    :rtype: None
    """
    editable_columns = editable_columns or []
    date_columns = date_columns or []
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            # Determine max length for column width
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)
            cell.protection = Protection(locked=True)

            # Set cell protection for specific columns
            if column_letter in editable_columns and cell.row > 1:
                cell.protection = Protection(locked=False)

            # Apply date style for specific columns and rows
            if column_letter in date_columns and cell.row > 1 and date_col_style:
                cell.style = date_col_style

        # Set adjusted column width
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def match_fields_to_data(model_data: list) -> None:
    """
    This method is to iterate through the list of model fields, and eliminate any that are not
    included in the data returned from RegScale. This is for those models that are using a
    graphQL query to return the data, which may not include all the fields on the model.

    :param list model_data: the data returned from the query
    :return: None
    :rtype: None
    """
    for cur_rec in model_data:
        cur_rec_keys = cur_rec.keys()
        for cur_field in obj_fields:
            if cur_field.field_name not in cur_rec_keys:
                if cur_field.field_name not in include_fields:
                    cur_field.sort_order = -1
