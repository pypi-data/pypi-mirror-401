#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Module to allow user to make changes to Assessments in an Excel
spreadsheet for user-friendly experience"""

# standard python imports
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd  # Type Checking
    from regscale.core.app.api import Api
    from regscale.core.app.application import Application

import math
import os
import shutil
from pathlib import Path
from typing import Optional, Union, Any

import click

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    error_and_exit,
    reformat_str_date,
    get_user_names,
    get_current_datetime,
    check_empty_nan,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models import Assessment
from regscale.models.regscale_models.modules import Modules

ALL_ASSESSMENTS = "all_assessments.xlsx"
NEW_ASSESSMENTS = "new_assessments.xlsx"
OLD_ASSESSMENTS = "old_assessments.xlsx"
DIFFERENCES_FILE = "differences.txt"
SELECT_PROMT = "Please select an option from the dropdown list."
DATE_ENTRY_PROMPT = "Please enter a valid date in the following format: mm/dd/yyyy"
SELECTION_ERROR = "Your entry is not one of the available options."
INVALID_ENTRY_ERROR = "Your entry is not a valid option."
INVALID_ENTRY_TITLE = "Invalid Entry"


@click.group(name="assessments")
def assessments():
    """
    Performs actions on Assessments CLI Feature to create new or update assessments to RegScale.
    """


# Make Empty Spreadsheet for creating new assessments.
@assessments.command(name="generate_new_file")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate_new_file(path: Path):
    """This function will build an Excel spreadsheet for users to be
    able to create new assessments."""
    new_assessment(path)


def new_assessment(path: Path) -> None:
    """
    Function to build Excel spreadsheet for creation of new assessments

    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    logger = create_logger()

    check_file_path(path)

    # create excel file and setting formatting

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "New_Assessments"

    column_headers = [
        "Title",
        "LeadAssessor",
        "Facility",
        "Organization",
        "AssessmentType",
        "PlannedStart",
        "PlannedFinish",
        "Status",
        "ActualFinish",
        "AssessmentResult",
        "ParentId",
        "ParentModule",
    ]
    for col, val in enumerate(column_headers, start=1):
        worksheet.cell(row=1, column=col).value = val

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        for cell in worksheet[col]:
            if cell.row == 1:
                cell.font = Font(bold=True)

    # create and format reference worksheets for dropdowns
    workbook.create_sheet(title="Facilities")
    workbook.create_sheet(title="Organizations")
    workbook.create_sheet(title="Accounts")
    workbook.create_sheet(title="Modules")
    workbook.create_sheet(title="AssessmentTypes")
    workbook.create_sheet(title="Assessment_Ids")

    workbook.save(filename=path / NEW_ASSESSMENTS)

    # pull in Facility, Organization, Module, and Account Usernames into Excel Spreadsheet to create drop downs
    list_of_modules = Modules().api_names()
    module_names = pd.DataFrame(list_of_modules, columns=["name"])
    with pd.ExcelWriter(
        path / NEW_ASSESSMENTS,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        get_field_names(field_name="facilities").to_excel(
            writer,
            sheet_name="Facilities",
            index=False,
        )
        get_field_names(field_name="organizations").to_excel(
            writer,
            sheet_name="Organizations",
            index=False,
        )
        get_user_names().to_excel(
            writer,
            sheet_name="Accounts",
            index=False,
        )
        module_names.to_excel(
            writer,
            sheet_name="Modules",
            index=False,
        )
        get_assessment_types().to_excel(
            writer,
            sheet_name="AssessmentTypes",
            index=False,
        )

    # Creating data Validation for fields
    workbook = load_workbook(os.path.join(path.absolute(), NEW_ASSESSMENTS))
    worksheet = workbook.active
    # lock worksheets containing data for dropdowns
    for sheet in [
        "Facilities",
        "Accounts",
        "Organizations",
        "AssessmentTypes",
        "Modules",
    ]:
        workbook[sheet].protection.sheet = True
    # Data structure for variable elements
    data_validations_info = [
        {"sheet": "Accounts", "columns": ["B"], "allow_blank": False},
        {"sheet": "Facilities", "columns": ["C"], "allow_blank": True},
        {"sheet": "Organizations", "columns": ["D"], "allow_blank": True},
        {"sheet": "AssessmentTypes", "columns": ["E"], "allow_blank": False},
        {"sheet": "Modules", "columns": ["L"], "allow_blank": True},
        {
            "formula1": '"Scheduled, In Progress, Complete, Cancelled"',
            "columns": ["H"],
            "allow_blank": True,
        },
        {
            "formula1": '"Pass, Fail, N/A, Partial Pass"',
            "columns": ["J"],
            "allow_blank": True,
        },
        {"type": "date", "columns": ["F", "G"], "allow_blank": False},
        {"type": "date", "columns": ["I"], "allow_blank": True},
    ]
    # Create data validations
    create_data_validations(
        data_validations_info=data_validations_info,
        workbook=workbook,
        worksheet=worksheet,
    )
    workbook.save(filename=os.path.join(path.absolute(), NEW_ASSESSMENTS))

    # Freezing top row and adding data style to date columns to assure validation

    workbook = load_workbook(os.path.join(path.absolute(), NEW_ASSESSMENTS))
    worksheet = workbook.active
    freeze_range = worksheet.cell(2, 14)
    worksheet.freeze_panes = freeze_range
    date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
    workbook.add_named_style(date_style)

    for col in ["F", "G", "I"]:  # Columns to edit
        for cell in worksheet[col]:
            if cell.row > 1:
                cell.style = date_style

    # Adjusting width of columns
    adjust_column_widths_and_styles(worksheet)

    workbook.save(filename=path / NEW_ASSESSMENTS)

    logger.info(
        "Your excel workbook has been created. Please open the new_assessments workbook and add new assessments."
    )
    return None


@assessments.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate(regscale_id: int, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all assessments
    with the selected RegScale Parent Id and RegScale Module for users to any necessary edits.
    """
    all_assessments(parent_id=regscale_id, parent_module=regscale_module, path=path)


def all_assessments(parent_id: int, parent_module: str, path: Path) -> None:
    """Function takes organizer record and module and build excel worksheet of assessments

    :param int parent_id: RegScale Parent Id
    :param str parent_module: RegScale Parent Module
    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    from regscale.core.app.application import Application

    app = Application()
    existing_assessment_data = Assessment.fetch_all_assessments_by_parent(
        app=app,
        parent_id=parent_id,
        parent_module=parent_module,
        org_and_facil=True,
    )
    if (
        existing_assessment_data["assessments"]["totalCount"] > 0
    ):  # Checking to see if assessment exists for selected RegScale Id and RegScale Module.
        check_file_path(path)
        sheet_names = ["Facilities", "Organizations", "Accounts", "AssessmentTypes"]

        all_assessments_wb = path / ALL_ASSESSMENTS
        old_assessments_wb = path / OLD_ASSESSMENTS

        # Loading data from db into two workbooks.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Assessments({parent_id}_{parent_module})"
        for worksheet in sheet_names:
            workbook.create_sheet(title=worksheet)
        workbook.save(filename=path / ALL_ASSESSMENTS)
        shutil.copy(
            all_assessments_wb,
            old_assessments_wb,
        )
        assessments_data = [
            [
                a["id"],
                a["title"],
                f"{a['leadAssessor']['lastName'].strip()}, {a['leadAssessor']['firstName'].strip()} "
                + f"({a['leadAssessor']['userName'].strip()})",
                a["facility"]["name"] if a["facility"] else None,
                a["org"]["name"] if a["org"] else None,
                a["assessmentType"],
                reformat_str_date(a["plannedStart"]),
                reformat_str_date(a["plannedFinish"]),
                a["status"],
                reformat_str_date(a["actualFinish"]) if a["actualFinish"] else None,
                a["assessmentResult"] or None,
                a["parentId"],
                a["parentModule"],
            ]
            for a in existing_assessment_data["assessments"]["items"]
        ]

        all_ass_df = pd.DataFrame(
            assessments_data,
            columns=[
                "Id",
                "Title",
                "LeadAssessor",
                "Facility",
                "Organization",
                "AssessmentType",
                "PlannedStart",
                "PlannedFinish",
                "Status",
                "ActualFinish",
                "AssessmentResult",
                "ParentId",
                "ParentModule",
            ],
        )

        with pd.ExcelWriter(
            all_assessments_wb,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments({parent_id}_{parent_module})",
                index=False,
            )
        with pd.ExcelWriter(
            old_assessments_wb,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments({parent_id}_{parent_module})",
                index=False,
            )

        # Pulling in Facility Names into Excel Spreadsheet to create dropdown.
        with pd.ExcelWriter(
            all_assessments_wb,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            get_field_names(field_name="facilities").to_excel(
                writer,
                sheet_name="Facilities",
                index=False,
            )
            get_field_names(field_name="organizations").to_excel(
                writer,
                sheet_name="Organizations",
                index=False,
            )
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )
            get_assessment_types().to_excel(
                writer,
                sheet_name="AssessmentTypes",
                index=False,
            )

        # Adding protection to OLD_ASSESSMENTS_WB file that will be used as reference.
        workbook2 = load_workbook(old_assessments_wb)
        worksheet2 = workbook2.active
        worksheet2.protection.sheet = True
        workbook2.save(filename=old_assessments_wb)

        # Adding Data Validation to ALL_ASSESSMENTS_WB file to be adjusted internally.
        workbook = load_workbook(all_assessments_wb)
        worksheet = workbook.active
        # lock worksheets containing data for dropdowns
        for sheet in sheet_names:
            workbook[sheet].protection.sheet = True

        data_validations_info = [
            {"sheet": "Accounts", "columns": ["C"], "allow_blank": False},
            {"sheet": "Facilities", "columns": ["D"], "allow_blank": True},
            {"sheet": "Organizations", "columns": ["E"], "allow_blank": True},
            {"sheet": "AssessmentTypes", "columns": ["F"], "allow_blank": False},
            {
                "formula1": '"Scheduled, In Progress, Complete, Cancelled"',
                "columns": ["I"],
                "allow_blank": True,
            },
            {
                "formula1": '"Pass, Fail, N/A, Partial Pass"',
                "columns": ["K"],
                "allow_blank": True,
            },
            {"type": "date", "columns": ["G", "H"], "allow_blank": False},
            {"type": "date", "columns": ["J"], "allow_blank": True},
        ]
        create_data_validations(
            data_validations_info=data_validations_info,
            workbook=workbook,
            worksheet=worksheet,
        )

        # Worksheet freeze top row
        freeze_range = worksheet.cell(2, 17)
        worksheet.freeze_panes = freeze_range
        date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
        workbook.add_named_style(date_style)

        # Adding Date Style to Worksheet, formatting cells, and unlocking
        # cells that can be edited in each assessment
        adjust_column_widths_and_styles(
            worksheet=worksheet,
            editable_columns=["C", "D", "E", "F", "G", "H", "I", "J", "K"],
            date_columns=["G", "H", "J"],
            date_col_style=date_style,
        )
        workbook.save(filename=all_assessments_wb)

        return app.logger.info(
            "Your data has been loaded into your excel workbook. Please open the all_assessments workbook "
            "and make your desired changes."
        )
    else:
        app.logger.info("Please check your selections for RegScale Id and RegScale Module and try again.")
        error_and_exit(
            "There was an error creating your workbook. No assessments exist for the given RegScale Id "
            "and RegScale Module."
        )


@assessments.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of excel workbook locations.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def load(path: Path) -> None:
    """
    This function uploads updated assessments and new assessments to
    RegScale from the Excel files that users have edited.
    """
    upload_data(path=path)


def upload_data(path: Path) -> None:
    """
    Function will upload assessments to RegScale if user as made edits to any
    of the assessment excel workbooks

    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    import numpy as np  # Optimize import performance
    from regscale.core.app.application import Application

    app = Application()
    new_assessments_wb = path / NEW_ASSESSMENTS
    all_assessments_wb = path / ALL_ASSESSMENTS
    old_assessments_wb = path / OLD_ASSESSMENTS

    # Checking if new assessments have been created and updating RegScale database.
    if os.path.isfile(new_assessments_wb):
        new_files = new_assessments_wb
        new = map_workbook_to_dict(new_files)
        new_assessments = [
            Assessment(
                leadAssessorId=value["LeadAssessorId"] or app.config["userId"],
                title=value["Title"],
                assessmentType=value["AssessmentType"],
                plannedStart=map_pandas_timestamp(value["PlannedStart"]),
                plannedFinish=map_pandas_timestamp(value["PlannedFinish"]),
                status=value["Status"],
                parentModule=check_empty_nan(value["ParentModule"]),
                facilityId=check_empty_nan(value.get("FacilityId")),
                orgId=check_empty_nan(value.get("OrganizationId")),
                assessmentResult=check_assessment_result(value["AssessmentResult"]),
                actualFinish=map_pandas_timestamp(value["ActualFinish"]),
                parentId=check_empty_nan(value["ParentId"]),
                lastUpdatedById=app.config["userId"],
                dateLastUpdated=get_current_datetime(),
            ).create()
            for value in new.values()
        ]
        post_and_save_assessments(
            app=app,
            new_assessments=new_assessments,
            workbook_path=path,
        )
    else:
        app.logger.info("No new assessments detected. Checking for edited assessments")

    if os.path.isfile(all_assessments_wb):
        # Checking all_assessments file for differences before updating database

        df1 = pd.read_excel(old_assessments_wb, sheet_name=0, index_col="Id")

        df2 = pd.read_excel(all_assessments_wb, sheet_name=0, index_col="Id")

        if df1.equals(df2):
            error_and_exit("No differences detected.")

        else:
            app.logger.warning("Differences found!")

        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
        ne_stacked = diff_mask.stack()
        changed = ne_stacked[ne_stacked]
        changed.index.names = ["Id", "Column"]
        difference_locations = np.where(diff_mask)
        changed_from = df1.values[difference_locations]
        changed_to = df2.values[difference_locations]
        changes = pd.DataFrame({"From": changed_from, "To": changed_to}, index=changed.index)
        changes.to_csv(
            path / DIFFERENCES_FILE,
            header=True,
            index=True,
            sep=" ",
            mode="w+",
        )
        app.logger.info(
            "Please check differences.txt file located in %s to see changes made.",
            path,
        )
        # Loading in differences.txt file and using Id to parse xlsx file for rows to update

        diff = pd.read_csv(path / DIFFERENCES_FILE, header=0, sep=" ", index_col=None)
        ids = []

        for _, row in diff.iterrows():
            ids.append(row["Id"])

        id_df = pd.DataFrame(ids, index=None, columns=["Id"])
        id_df2 = id_df.drop_duplicates()
        updated_files = all_assessments_wb
        df3 = pd.read_excel(updated_files, sheet_name=0, index_col=None)
        updated = df3[df3["Id"].isin(id_df2["Id"])]
        updated = map_workbook_to_dict(updated_files, updated)
        _ = [
            Assessment(
                leadAssessorId=value["LeadAssessorId"],
                id=value["Id"],
                title=value["Title"],
                assessmentType=value["AssessmentType"],
                plannedStart=value["PlannedStart"],
                plannedFinish=value["PlannedFinish"],
                status=value["Status"],
                parentModule=value["ParentModule"],
                facilityId=check_empty_nan(value.get("FacilityId")),
                orgId=check_empty_nan(value.get("OrganizationId")),
                assessmentResult=check_assessment_result(value["AssessmentResult"]),
                actualFinish=check_empty_nan(value["ActualFinish"]),
                parentId=value["ParentId"],
                lastUpdatedById=app.config["userId"],
                dateLastUpdated=get_current_datetime(),
            ).save(bulk=True)
            for value in updated.values()
        ]

        Assessment.bulk_save()

    else:
        app.logger.info("No Assessments exist to load to RegScale.")
    return app.logger.info(
        "Assessment files have been uploaded. Changes made to existing files can be seen in "
        "differences.txt file. Thank you!"
    )


@assessments.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
def delete_files(path: Path):
    """This command will delete files used during the Assessment editing process."""
    delete_file(path)


def delete_file(path: Path) -> int:
    """
    Deletes files used during the process

    :param Path path: directory of file location
    :return: Number of files deleted
    :rtype: int
    """
    logger = create_logger()
    file_names = [
        NEW_ASSESSMENTS,
        ALL_ASSESSMENTS,
        OLD_ASSESSMENTS,
        DIFFERENCES_FILE,
    ]
    deleted_files = []

    for file_name in file_names:
        if os.path.isfile(path / file_name):
            os.remove(path / file_name)
            deleted_files.append(file_name)
        else:
            logger.warning("No %s file found. Checking for other files before exiting.", file_name)
    logger.info("%i file(s) have been deleted: %s", len(deleted_files), ", ".join(deleted_files))
    return len(deleted_files)


def get_maximum_rows(*, sheet_object: Any) -> int:
    """
    This function finds the last row containing data in a spreadsheet

    :param Any sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(any(col.value is not None for col in row) for max_row, row in enumerate(sheet_object, 1))


def get_field_names(field_name: str) -> "pd.DataFrame":
    """
    This function uses GraphQL to retrieve all names of a given parent table in database

    :param str field_name: name of parent table to retrieve names from
    :return: pandas dataframe with facility names
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance
    from regscale.core.app.api import Api

    api = Api()

    body = """
    query {
        field_name(skip: 0, take: 50, order: {name: ASC}, ) {
            items {
                name
                id
            }
            totalCount
            pageInfo {
                hasNextPage
            }
        }
    }
    """.replace(
        "field_name", field_name
    )

    field_items = api.graph(query=body)
    names = field_items[str(field_name)]["items"]
    field_names = [[i["name"], i["id"]] for i in names]
    all_names = pd.DataFrame(field_names, index=None, columns=["name", "id"])

    return all_names


def get_assessment_types() -> "pd.DataFrame":
    """
    This function uses GraphQL to retrieve all assessment types in database

    :return: pandas dataframe with assessment types
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance
    from regscale.core.app.api import Api

    api = Api()

    body = """
        query{
          assessments (skip: 0, take: 50, order: {assessmentType: ASC}, ) {
            items {
              assessmentType
            }
            totalCount
            pageInfo {
              hasNextPage
            }
          }
        } """

    assessments_raw = api.graph(query=body)
    assessment_types = assessments_raw["assessments"]["items"]
    field_names = [i["assessmentType"] for i in assessment_types]
    all_assessment_types = pd.DataFrame(field_names, index=None, columns=["assessmentType"])
    return all_assessment_types.drop_duplicates()


def check_assessment_result(value: Any) -> Union[str, float]:
    """
    This function takes a given value for an assessment and
    checks if value is empty or NaN based on value type.

    :param Any value: A string or float object
    :return: A string value, float value. or ""
    :rtype: Union[str, float]
    """
    # this function has to be checked separate to account for API
    # only accepting empty string unlike other class params
    if isinstance(value, str) and value.strip() == "":
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def adjust_column_widths_and_styles(
    worksheet: Worksheet,
    editable_columns: Optional[list[str]] = None,
    date_columns: Optional[list[str]] = None,
    date_col_style: Optional[NamedStyle] = None,
) -> None:
    """
    Function to adjust column widths based on length of data in column, and apply
    styles to specific columns and rows

    :param Worksheet worksheet: Worksheet to adjust column widths for
    :param Optional[list[str]] editable_columns: List of rows to unlock for editing
    :param Optional[list[str]] date_columns: List of columns to add date style to
    :param Optional[NamedStyle] date_col_style: NamedStyle object to apply to date columns, defaults to None
    :rtype: None
    """
    editable_columns = editable_columns or []
    date_columns = date_columns or []
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            # Determine max length for column width
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)

            # Set cell protection for specific columns
            if column_letter in editable_columns and cell.row > 1:
                cell.protection = Protection(locked=False)

            # Apply date style for specific columns and rows
            if column_letter in date_columns and cell.row > 1 and date_col_style:
                cell.style = date_col_style

        # Set adjusted column width
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_data_validations(data_validations_info: list[dict], workbook: Workbook, worksheet: Worksheet) -> None:
    """
    Function to create data validations for excel worksheet

    :param list[dict] data_validations_info: List containing dictionaries with
        information for data validations
    :param Workbook workbook: Workbook object to add data validations to
    :param Worksheet worksheet: The worksheet object to add data validations to
    :rtype: None
    """
    for _, dv_info in enumerate(data_validations_info, start=1):
        formula1 = dv_info.get("formula1")
        if sheet_name := dv_info.get("sheet"):
            formula1 = f"={sheet_name}!$A$2:$A${str(get_maximum_rows(sheet_object=workbook[sheet_name]))}"

        data_validation = DataValidation(
            type=dv_info.get("type", "list"),
            formula1=formula1,
            allow_blank=dv_info.get("allow_blank", True),
            showDropDown=False,
            error=(SELECTION_ERROR if dv_info.get("type", "list") == "list" else INVALID_ENTRY_ERROR),
            errorTitle=INVALID_ENTRY_TITLE,
            prompt=(SELECT_PROMT if dv_info.get("type", "list") == "list" else DATE_ENTRY_PROMPT),
            showErrorMessage=True if dv_info.get("type", "date") else None,
            showInputMessage=True if dv_info.get("type", "date") else None,
        )

        worksheet.add_data_validation(data_validation)
        for column in dv_info["columns"]:
            data_validation.add(f"{column}2:{column}1048576")


def post_and_save_assessments(app: "Application", new_assessments: list[Assessment], workbook_path: Path) -> None:
    """
    Function to post new assessments to RegScale and save assessment ids to excel workbook

    :param Application app: RegScale CLI Application object
    :param list[Assessment] new_assessments: List of new assessments to post to RegScale
    :param Path workbook_path: Path to workbook to save assessment ids to
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    new_assessments_df = pd.DataFrame([assessment.id for assessment in new_assessments], columns=["id_number"])
    for file_name in [NEW_ASSESSMENTS, ALL_ASSESSMENTS]:
        with pd.ExcelWriter(
            workbook_path / file_name,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            new_assessments_df.to_excel(
                writer,
                sheet_name="Assessment_Ids",
                index=False,
            )
    app.logger.info(
        "%i total assessment(s) were added to RegScale.",
        len(new_assessments),
    )


def map_pandas_timestamp(date_time: "pd.Timestamp") -> Optional[str]:
    """
    Function to map pandas timestamp to string

    :param pd.Timestamp date_time: Timestamp to map to string
    :return: String representation of pandas timestamp
    :rtype: Optional[str]
    """
    import pandas as pd  # Optimize import performance

    if isinstance(date_time, float):
        return None
    elif date_time is not None and not pd.isna(date_time) and not isinstance(date_time, str):
        return date_time.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return date_time or None


def map_workbook_to_dict(file_path: str, workbook_data: Optional["pd.DataFrame"] = None) -> dict:
    """
    Function to map workbook to dictionary

    :param str file_path: Path to workbook file
    :param Optional[pd.DataFrame] workbook_data: Dataframe to map to dictionary
    :return: Dictionary representation of workbook
    :rtype: dict
    """
    import pandas as pd  # Optimize import performance

    if workbook_data is not None:
        wb_data = workbook_data
    else:
        wb_data = pd.read_excel(file_path)
    wb_data["Facility"] = wb_data["Facility"].astype(str).fillna("None")  # Handle missing facilities
    wb_data["Organization"] = wb_data["Organization"].astype(str).fillna("None")  # Handle missing organizations

    # Reading and preparing the 'Facilities' sheet
    facilities = pd.read_excel(file_path, sheet_name="Facilities")
    facilities = facilities.rename(columns={"name": "Facility", "id": "FacilityId"})
    facilities["Facility"] = facilities["Facility"].astype(str)  # Ensure consistent data type

    # Reading and preparing the 'Organizations' sheet
    organizations = pd.read_excel(file_path, sheet_name="Organizations")
    organizations = organizations.rename(columns={"name": "Organization", "id": "OrganizationId"})
    organizations["Organization"] = (
        organizations["Organization"].astype(str).fillna("None")
    )  # Handle missing organizations

    # Reading and preparing the 'Accounts' sheet
    accounts = pd.read_excel(file_path, sheet_name="Accounts")
    accounts = accounts.rename(columns={"User": "LeadAssessor", "UserId": "LeadAssessorId"})

    # Merging dataframes
    wb_data = wb_data.merge(accounts, how="left", on="LeadAssessor", validate="many_to_many")
    wb_data = wb_data.merge(facilities, how="left", on="Facility", validate="many_to_many")
    wb_data = wb_data.merge(organizations, how="left", on="Organization", validate="many_to_many")
    return wb_data.T.to_dict()
