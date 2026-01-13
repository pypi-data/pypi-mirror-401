#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Module to allow user to make changes to issues in an excel spreadsheet for user friendly experience"""

# standard python imports
import os
import shutil
from typing import Any

import click
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from rich.console import Console
from rich.table import Table

from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    error_and_exit,
    reformat_str_date,
    get_current_datetime,
    get_user_names,
    check_empty_nan,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.issue import Issue

logger = create_logger()


@click.group(name="issues")
def issues():
    """
    Performs actions on POAM CLI Feature to update issues to RegScale.
    """


@issues.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate_all_issues(regscale_id: int, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all issues
    with the selected RegScale Parent Id and RegScale Module for users to make any neccessary edits.

    """
    all_issues(regscale_id=regscale_id, regscale_module=regscale_module, path=path)


def all_issues(regscale_id: int, regscale_module: str, path: Path) -> None:
    """Function to build excel spreadsheet with all issues matching organizer records given.

    :param int regscale_id: RegScale Parent Id
    :param str regscale_module: RegScale Parent Module
    :param Path path: directory of file location
    :rtype: None
    """
    # see if provided RegScale Module is an accepted option in Organizer Modules list
    import pandas as pd  # Optimize import performance
    from regscale.core.app.api import Api

    api = Api()
    verify_organizer_module(regscale_module)

    body = """
            query {
                    issues (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                      items {
                       id
                       issueOwnerId
                       issueOwner {
                         firstName
                         lastName
                         userName
                       }
                       title
                       dateCreated
                       description
                       severityLevel
                       costEstimate
                       levelOfEffort
                       dueDate
                       identification
                       status
                       dateCompleted
                       activitiesObserved
                       failuresObserved
                       requirementsViolated
                       safetyImpact
                       securityImpact
                       qualityImpact
                       securityChecks
                       recommendedActions
                       parentId
                       parentModule
                      }
                      totalCount
                      pageInfo {
                        hasNextPage
                      }
                    }
                 }
                """.replace(
        "parent_module", regscale_module
    ).replace(
        "parent_id", str(regscale_id)
    )
    existing_issue_data = api.graph(query=body)

    if (
        existing_issue_data["issues"]["totalCount"] > 0
    ):  # Checking to see if assessment exists for selected RegScale Id and RegScale Module.
        check_file_path(path)

        # Loading data from db into two workbooks.

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Issues({regscale_id}_{regscale_module})"
        workbook.save(filename=os.path.join(path, "all_issues.xlsx"))
        shutil.copy(
            os.path.join(path, "all_issues.xlsx"),
            os.path.join(path, "old_issues.xlsx"),
        )

        raw_data = existing_issue_data["issues"]["items"]
        issues_data = []
        for a in raw_data:
            issue_id = a["id"]
            issue_owner = (
                (
                    str(a["issueOwner"]["lastName"]).strip()
                    + ", "
                    + str(a["issueOwner"]["firstName"]).strip()
                    + " ("
                    + str(a["issueOwner"]["userName"]).strip()
                    + ")"
                )
                if a["issueOwner"]
                else "None"
            )
            title = a["title"]
            date_created = reformat_str_date(a["dateCreated"])
            description = a["description"] if a["description"] else "None"
            severity_level = a["severityLevel"]
            cost_estimate = a["costEstimate"] if a["costEstimate"] and a["costEstimate"] != "None" else 0.00
            level_of_effort = a["levelOfEffort"] if a["levelOfEffort"] and a["levelOfEffort"] != "None" else 0
            due_date = reformat_str_date(a["dueDate"])
            identification = a["identification"] if a["identification"] else "None"
            status = a["status"] if a["status"] else "None"
            date_completed = reformat_str_date(a["dateCompleted"]) if a["dateCompleted"] else ""
            activities_observed = a["activitiesObserved"] or ""
            failures_observed = a["failuresObserved"] or ""
            requirements_violated = a["requirementsViolated"] or ""
            safety_impact = a["safetyImpact"] or ""
            security_impact = a["securityImpact"] or ""
            quality_impact = a["qualityImpact"] or ""
            security_checks = a["securityChecks"] or ""
            recommended_actions = a["recommendedActions"] or ""
            issues_data.append(
                [
                    issue_id,
                    issue_owner,
                    title,
                    date_created,
                    description,
                    severity_level,
                    cost_estimate,
                    level_of_effort,
                    due_date,
                    identification,
                    status,
                    date_completed,
                    activities_observed,
                    failures_observed,
                    requirements_violated,
                    safety_impact,
                    security_impact,
                    quality_impact,
                    security_checks,
                    recommended_actions,
                ]
            )

        all_ass_df = pd.DataFrame(
            issues_data,
            columns=[
                "Id",
                "IssueOwner",
                "Title",
                "DateCreated",
                "Description",
                "SeverityLevel",
                "CostEstimate",
                "LevelOfEffort",
                "DueDate",
                "Identification",
                "Status",
                "DateCompleted",
                "ActivitiesObserved",
                "FailuresObserved",
                "RequirementsViolated",
                "SafetyImpact",
                "SecurityImpact",
                "QualityImpact",
                "SecurityChecks",
                "RecommendedActions",
            ],
        )

        with pd.ExcelWriter(
            os.path.join(path, "all_issues.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Issues({regscale_id}_{regscale_module})",
                index=False,
            )
        with pd.ExcelWriter(
            os.path.join(path, "old_issues.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Issues({regscale_id}_{regscale_module})",
                index=False,
            )

        # Pulling in Account Users into Excel Spreasheet to create drop down.

        with pd.ExcelWriter(
            os.path.join(path, "all_issues.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )

        # Pulling in Identifications into separate Excel Spreadsheet to create drop down. (Over 256 character Limit)

        identification_list = [
            "A-123 Review",
            "Assessment/Audit (External)",
            "Assessment/Audit (Internal)",
            "Critical Control Review",
            "FDCC/USGCB",
            "GAO Audit",
            "IG Audit",
            "Incidnet Response Lessons Learned",
            "ITAR",
            "Other",
            "Penetration Test",
            "Risk Assessment",
            "Security Authorization",
            "Security Control Assessment",
            "Vulnerability Assessment",
        ]

        identifications = pd.DataFrame({"IdentificationOptions": identification_list})

        with pd.ExcelWriter(
            os.path.join(path, "all_issues.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            identifications.to_excel(
                writer,
                sheet_name="Identifications",
                index=False,
            )

        # Adding protection to "old_issues.xlsx" file that will be used as reference.

        workbook2 = load_workbook(os.path.join(path, "old_issues.xlsx"))
        worksheet2 = workbook2.active
        worksheet2.protection.sheet = True
        workbook2.save(filename=os.path.join(path, "old_issues.xlsx"))

        # Adding Data Validation to "all_issues.xlsx" file to be adjusted internally.

        workbook = load_workbook(os.path.join(path, "all_issues.xlsx"))
        worksheet = workbook.active
        accounts_worksheet = workbook["Accounts"]
        identifications_worksheet = workbook["Identifications"]

        accounts_worksheet.protection.sheet = True
        worksheet.protection.sheet = True
        identifications_worksheet.protection.sheet = True

        # Adding DropDown Menu to Columns to Match RegScale options for fields

        severitylevels = '"I - High - Signficant Deficiency, II - Moderate - Reportable Condition, III - Low - Other Weakness, IV - Not Asssigned"'
        statuses = '"Closed, Draft, Open, Pending Decommission, Supply Chain/Procurement Dependency, Vendor Dependency for Fix, Delayed, Cancelled, Exception/Waiver"'

        dv1 = DataValidation(
            type="list",
            formula1="=Accounts!$A$2:$A$" + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv2 = DataValidation(
            type="list",
            formula1=severitylevels,
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv3 = DataValidation(
            type="list",
            formula1=statuses,
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv4 = DataValidation(
            type="list",
            formula1="=Identifications!$A$2:$A$" + str(get_maximum_rows(sheet_object=workbook["Identifications"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )

        # Adding Date Style to Worksheet

        date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
        workbook.add_named_style(date_style)
        currency = '"$"* #,##0.00_);("$"* #,##0.00);"$"* #,##0.00_);'

        for col in ["D", "G", "I", "L"]:
            for cell in worksheet[col]:
                if col != "G" and cell.row > 1:
                    cell.style = date_style
                elif col == "G" and cell.row > 1:
                    cell.number_format = currency

        # Adding data validation to avoid manual error on entry columns

        dv5 = DataValidation(
            type="date",
            allow_blank=False,
            showErrorMessage=True,
            showInputMessage=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        dv6 = DataValidation(
            type="date",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        dv7 = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=False,
            showErrorMessage=True,
            showInputMessage=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid whole number denoting number of hours.",
        )
        worksheet.add_data_validation(dv1)
        worksheet.add_data_validation(dv2)
        worksheet.add_data_validation(dv3)
        worksheet.add_data_validation(dv4)
        worksheet.add_data_validation(dv5)
        worksheet.add_data_validation(dv6)
        worksheet.add_data_validation(dv7)
        dv1.add("B2:B1048576")
        dv2.add("F2:F1048576")
        dv3.add("K2:K1048576")
        dv4.add("J2:J1048576")
        dv5.add("I2:I1048576")
        dv6.add("D2:D1048576")
        dv6.add("L2:L1048576")
        dv7.add("H2:H1048576")

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        # Unlocking cells that can be edited in each Issue.

        for col in [
            "B",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
        ]:  # Columns to edit.
            for cell in worksheet[col]:
                cell.protection = Protection(locked=False)

        workbook.save(filename=os.path.join(path, "all_issues.xlsx"))

    else:
        logger.info("No Issues exist. Please check your selections for RegScale Id and RegScale Module and try again.")
        error_and_exit("There was an error creating your workbook for the given RegScale Id and RegScale Module.")

    logger.info(
        "Your data has been loaded. Please open the all_issues workbook and make your desired changes %s. " % path
    )
    return None


@issues.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of excel workbook locations.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate_upload_data(path: Path):
    """This function uploads updated issues to the RegScale from the Excel files that users have edited."""
    upload_data(path)


def upload_data(path: Path) -> None:
    """
    This function uploads updated issues to the RegScale.

    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    import numpy as np  # Optimize import performance

    # Checking all_issues file for differences before updating database
    workbook = load_workbook(os.path.join(path, "all_issues.xlsx"))
    sheet_name = workbook.sheetnames[0]
    sheet_name = sheet_name[sheet_name.find("(") + 1 : sheet_name.find(")")].split("_")

    # set the variables to the correct values
    for item in set(sheet_name):
        try:
            regscale_parent_id = int(item)
        except ValueError:
            regscale_module = item

    df1 = pd.read_excel(os.path.join(path, "old_issues.xlsx"), sheet_name=0, index_col="Id")

    df2 = pd.read_excel(os.path.join(path, "all_issues.xlsx"), sheet_name=0, index_col="Id")

    if df1.equals(df2):
        logger.info("No differences detected.")
        error_and_exit("No changes were made to the all_issues.xlsx file. Thank you!")
    else:
        logger.warning("*** WARNING *** Differences Found")

        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
        ne_stacked = diff_mask.stack()
        changed = ne_stacked[ne_stacked]
        changed.index.names = ["Id", "Column"]
        difference_locations = np.where(diff_mask)
        changed_to = df1.values[difference_locations]
        changed_from = df2.values[difference_locations]
        changes = pd.DataFrame({"From": changed_from, "To": changed_to}, index=changed.index)
        changes.to_csv(
            os.path.join(path, "differences.txt"),
            header=True,
            index=True,
            sep=" ",
            mode="a",
        )

        diff = pd.read_csv(os.path.join(path, "differences.txt"), header=0, sep=" ", index_col=None)
        ids = []
        for i, row in diff.iterrows():
            ids.append(row["Id"])

        id_df = pd.DataFrame(ids, index=None, columns=["Id"])
        id_df2 = id_df.drop_duplicates()
        updated_file = os.path.join(path, "all_issues.xlsx")

        reader = pd.read_excel(updated_file)
        updated = reader[reader["Id"].isin(id_df2["Id"])]

        accounts = pd.read_excel(updated_file, sheet_name="Accounts")
        accounts = accounts.rename(columns={"User": "IssueOwner", "UserId": "IssueOwnerId"})
        updated = updated.merge(accounts, how="left", on="IssueOwner")
        updated = updated.T.to_dict()
        updated_issues = [
            Issue(
                id=value["Id"],
                issueOwnerId=value["IssueOwnerId"],
                title=value["Title"],
                dateCreated=value["DateCreated"],
                description=value["Description"],
                severityLevel=value["SeverityLevel"],
                costEstimate=value["CostEstimate"],
                levelOfEffort=value["LevelOfEffort"],
                dueDate=value["DueDate"],
                identification=check_empty_nan(value["Identification"], "Other"),
                status=value["Status"],
                dateCompleted=check_empty_nan(value["DateCompleted"], ""),
                activitiesObserved=check_empty_nan(value["ActivitiesObserved"]),
                failuresObserved=check_empty_nan(value["FailuresObserved"]),
                requirementsViolated=check_empty_nan(value["RequirementsViolated"]),
                safetyImpact=check_empty_nan(value["SafetyImpact"]),
                securityImpact=check_empty_nan(value["SecurityImpact"]),
                qualityImpact=check_empty_nan(value["QualityImpact"]),
                securityChecks=check_empty_nan(value["SecurityChecks"]),
                recommendedActions=check_empty_nan(value["RecommendedActions"]),
                dateLastUpdated=get_current_datetime(),
                parentModule=regscale_module,
                parentId=regscale_parent_id,
            )
            for value in updated.values()
        ]
        Issue.batch_update(updated_issues)
    logger.info("Changes made to existing files can be seen in differences.txt file. Thank you! %s" % path)
    return None


@issues.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of excel workbook locations.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate_delete_file(path: Path):
    """This function deletes files used during the POAM Editor process."""
    delete_file(path)


def delete_file(path: Path) -> None:
    """
    Deletes files used during the process.

    :param Path path: Path for artifacts folder or location of excel files
    :rtype: None
    """
    os.remove(os.path.join(path, "all_issues.xlsx"))
    os.remove(os.path.join(path, "old_issues.xlsx"))
    if os.path.isfile(os.path.join(path, "differences.txt")):
        os.remove(os.path.join(path, "differences.txt"))
    else:
        pass
    logger.info("Files have been deleted. Thank you.")
    return None


def get_maximum_rows(*, sheet_object: Any) -> int:
    """This function finds the last row containing data in a spreadsheet

    :param Any sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(any(col.value is not None for col in row) for max_row, row in enumerate(sheet_object, 1))


def verify_organizer_module(module: str) -> None:
    """
    Function to check the provided module is a valid RegScale Organizer Module and will display the acceptable RegScale modules

    :param str module: desired module
    :rtype: None
    """

    # create console and table objects
    console = Console()
    table = Table("RegScale Module", "Accepted Value", title="RegScale Modules", safe_box=True)

    # list of RegScaleOrganizer Modules
    organizers = ["components", "policies", "projects", "securityplans", "supplychain"]

    # iterate through items and add them to table object
    for i in range(len(organizers)):
        table.add_row(organizers[i], organizers[i])

    if module not in organizers:
        # print the table object in console
        console.print(table)
        error_and_exit("Please provide an option from the Accepted Value column.")
