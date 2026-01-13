#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module to allow user to make changes to Control Implementations in an Excel spreadsheet for a user-friendly experience
"""

import math
import os
import shutil
import sys
from typing import Union, Any, TYPE_CHECKING

import click
from pathlib import Path

if TYPE_CHECKING:
    from regscale.core.app.api import Api


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
import warnings


from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_empty_nan,
    check_file_path,
    error_and_exit,
    get_current_datetime,
    get_user_names,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models import Control, ControlImplementation


ALL_IMPS = "all_implementations.xlsx"
OLD_IMPS = "old_implementations.xlsx"
DIFFS = "differences.txt"
ERROR_MSG = "Your entry is not one of the available options"
INVALID_MSG = "Invalid Entry"
PROMPT_MSG = "Please select from the list"


@click.group(name="control_editor")
def control_editor():
    """
    Performs actions on Control Editor Feature to edit controls to RegScale.
    """
    warnings.filterwarnings("always", category=DeprecationWarning)
    warnings.warn(
        "Control Editor is deprecated and will be removed in a future release. Use `regscale model` with the `--model control` argument instead.",
        DeprecationWarning,
        stacklevel=2,
    )


# Get data and pull into Excel worksheets.


@control_editor.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for created excel files to be saved to.",
    default=Path("./artifacts"),
    required=True,
)
def generate_data_download(regscale_id: int, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all control implementations
    with the selected RegScale Parent Id and RegScale Module.
    """
    warnings.filterwarnings("always", category=DeprecationWarning)
    warnings.warn(
        "Control Editor is deprecated and will be removed in a future release. Use `regscale model generate --model control` instead.",
        DeprecationWarning,
        stacklevel=2,
    )
    data_load(regscale_id, regscale_module, path)


def data_load(parent_id: int, parent_module: str, path: Path) -> None:
    """Function takes organizer record and module and build excel worksheet of control implementations.

    :param int parent_id: RegScale Parent Id
    :param str parent_module: RegScale Parent Module
    :param Path path: directory of file location
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    from regscale.core.app.api import Api

    api = Api()
    logger = api.logger
    all_imps_wb = path / ALL_IMPS
    old_imps_wb = path / OLD_IMPS

    # Making directory for files

    check_file_path(path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = f"Impls_PId({parent_id}_{parent_module})"
    workbook.create_sheet("Accounts")

    workbook.save(filename=all_imps_wb)
    shutil.copy(
        all_imps_wb,
        old_imps_wb,
    )

    # Loading data from RegScale database into two workbooks.

    all_imps_df = _fetch_implementations(api, parent_id, parent_module)

    with pd.ExcelWriter(
        all_imps_wb,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        all_imps_df.to_excel(
            writer,
            sheet_name=f"Impls_PId({parent_id}_{parent_module})",
            index=False,
        )
        get_user_names().to_excel(
            writer,
            sheet_name="Accounts",
            index=False,
        )

    with pd.ExcelWriter(
        old_imps_wb,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        all_imps_df.to_excel(
            writer,
            sheet_name=f"Impls_PId({parent_id}_{parent_module})",
            index=False,
        )

    # Adding Data validation to "old_implementations.xlsx" file that will be used as reference.

    workbook2 = load_workbook(old_imps_wb)
    worksheet2 = workbook2.active
    worksheet2.protection.sheet = True
    workbook2.save(filename=old_imps_wb)

    # Adding Data Validation to "all_implementations.xlsx" file to be adjusted internally by clients.

    workbook = load_workbook(all_imps_wb)
    worksheet = workbook.active
    worksheet.protection.sheet = True
    accounts_worksheet = workbook["Accounts"]
    accounts_worksheet.protection.sheet = True

    dv1 = DataValidation(
        type="list",
        formula1='"Not Implemented, Fully Implemented, In Remediation, Not Applicable, Inherited, Planned"',
        allow_blank=True,
        showDropDown=False,
        error=ERROR_MSG,
        errorTitle=INVALID_MSG,
        prompt=PROMPT_MSG,
    )
    dv2 = DataValidation(
        type="list",
        formula1='"Provider, Customer, Shared, Not Applicable"',
        allow_blank=True,
        showDropDown=False,
        error=ERROR_MSG,
        errorTitle=INVALID_MSG,
        prompt=PROMPT_MSG,
    )
    dv3 = DataValidation(type="list", formula1='"TRUE, FALSE"', allow_blank=True)
    dv4 = DataValidation(
        type="list",
        formula1="=Accounts!$A$2:$A$" + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
        allow_blank=False,
        showDropDown=False,
        error=ERROR_MSG,
        errorTitle=INVALID_MSG,
        prompt=PROMPT_MSG,
    )

    worksheet.add_data_validation(dv1)
    worksheet.add_data_validation(dv2)
    worksheet.add_data_validation(dv3)
    worksheet.add_data_validation(dv4)
    dv1.add("G2:G1048576")
    dv2.add("J2:J1048576")
    dv3.add("K2:K1048576")
    dv4.add("C2:C1048576")

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width < 50:
            worksheet.column_dimensions[column].width = adjusted_width
        else:
            worksheet.column_dimensions[column].width = 50
        check_and_format_cells(column, col)

    workbook.save(filename=all_imps_wb)

    logger.info("Successfully created the directory %s.", path)
    logger.info("All files are located within directory.")

    logger.info(
        "Your data has been loaded into your excel workbook. "
        "Please open the all_implementations workbook and make your desired changes."
    )
    return None


def check_and_format_cells(column: str, col: list[Any]) -> None:
    """
    Function to align cells in the provided column

    :param str column: Column to align
    :param list[Any] col: Column to align
    :rtype: None
    """
    if column in [
        "E",
        "F",
        "I",
    ]:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    if column in [
        "C",
        "G",
        "H",
        "I",
        "J",
        "K",
    ]:  # Check if current column is column to edit
        for cell in col:
            cell.fill = PatternFill(start_color="7C7C7C", end_color="7C7C7C", fill_type="solid")
            cell.protection = Protection(locked=False)  # Unprotect the cell


def _extract_control_owner_display(item: dict) -> str:
    """
    Extract and format control owner display name from item data.

    :param dict item: Item data containing controlOwner information
    :return: Formatted control owner display string
    :rtype: str
    """
    if not item.get("controlOwner") or item["controlOwner"] is None:
        return "Unassigned"

    control_owner = item["controlOwner"]
    last_name = str(control_owner.get("lastName", "")).strip() if control_owner.get("lastName") else ""
    first_name = str(control_owner.get("firstName", "")).strip() if control_owner.get("firstName") else ""
    user_name = str(control_owner.get("userName", "")).strip() if control_owner.get("userName") else ""

    if last_name or first_name or user_name:
        return f"{last_name}, {first_name} ({user_name})"
    return "Unassigned"


def _extract_control_data(item: dict) -> tuple:
    """
    Extract control-related data from item.

    :param dict item: Item data containing control information
    :return: Tuple of (control_id, control_title, control_description, control_weight, catalogue_id)
    :rtype: tuple
    """
    if not item.get("control") or item["control"] is None:
        return "", "", "", 0, 0

    control = item["control"]
    return (
        control.get("controlId", ""),
        control.get("title", ""),
        control.get("description", ""),
        control.get("weight", 0),
        control.get("catalogueID", 0),
    )


def _build_implementation_row(item: dict) -> list:
    """
    Build a single implementation row from item data.

    :param dict item: Item data from GraphQL response
    :return: List representing a row of implementation data
    :rtype: list
    """
    control_owner_display = _extract_control_owner_display(item)
    control_id, control_title, control_description, control_weight, catalogue_id = _extract_control_data(item)

    return [
        item.get("id", 0),
        item.get("controlID", 0),
        control_owner_display,
        control_id,
        control_title,
        control_description,
        item.get("status", ""),
        item.get("policy", ""),
        item.get("implementation", ""),
        item.get("responsibility", ""),
        item.get("inheritable", False),
        control_weight,
        catalogue_id,
    ]


def _fetch_implementations(api: "Api", parent_id: int, parent_module: str) -> "pd.DataFrame":
    """
    Function to fetch implementations from RegScale.

    :param Api api: API object to make calls to RegScale
    :param int parent_id: Parent ID of the implementation in RegScale
    :param str parent_module: Parent module of the implementation in RegScale
    :return: DataFrame of implementations
    :rtype: pd.DataFrame
    """
    import pandas as pd  # Optimize import performance

    body = f"""
                query {{
                  controlImplementations(
                    skip: 0
                    take: 50
                    where: {{
                      parentId: {{ eq: {parent_id} }}
                      parentModule: {{ eq: "{parent_module}" }}
                    }}
                  ) {{
                    items {{
                      id
                      controlID
                      controlOwner {{
                        firstName
                        lastName
                        userName
                      }}
                      control {{
                        title
                        description
                        controlId
                        weight
                        catalogueID
                      }}
                      status
                      policy
                      implementation
                      responsibility
                      inheritable
                      parentId
                      parentModule
                    }}
                    totalCount
                    pageInfo {{
                      hasNextPage
                    }}
                  }}
                }}
        """
    existing_implementation_data = api.graph(query=body)

    if existing_implementation_data["controlImplementations"]["totalCount"] <= 0:
        error_and_exit("No records exist for the given RegScale Id and RegScale Module.")

    items = existing_implementation_data.get("controlImplementations", {}).get("items", [])
    all_imps = [_build_implementation_row(item) for item in items]

    all_imps_df = pd.DataFrame(
        all_imps,
        columns=[
            "Id",
            "ControlId",
            "ControlOwner",
            "ControlName",
            "ControlTitle",
            "Description",
            "Status",
            "Policy",
            "Implementation",
            "Responsibility",
            "Inheritable",
            "Weight",
            "CatalogueId",
        ],
    )
    return all_imps_df


@control_editor.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path where excel workbooks are located.",
    default=Path("./artifacts"),
    required=True,
)
@click.option(
    "--skip_prompt",
    type=click.BOOL,
    help="To Skip (Y/N) Prompt, input True.",
    default=False,
    required=False,
)
def generate_db_update(path: Path, skip_prompt: bool):
    """
    This function will check changes made to spreadsheet and upload any changes made to RegScale.
    """
    warnings.filterwarnings("always", category=DeprecationWarning)
    warnings.warn(
        "Control Editor is deprecated and will be removed in a future release. Use `regscale model load --model control` instead.",
        DeprecationWarning,
        stacklevel=2,
    )
    db_update(path, skip_prompt)


def db_update(path: Path, skip_prompt: bool = True) -> None:
    """Function will check changes made by user and upload any changes to RegScale.

    :param Path path: directory of file location
    :param bool skip_prompt: boolean to skip prompt save message, defaults to True
    :rtype: None
    """
    import pandas as pd  # Optimize import performance
    import numpy as np  # Optimize import performance

    logger = create_logger()
    all_imps_wb = path / ALL_IMPS
    old_imps_wb = path / OLD_IMPS
    regscale_parent_id = None
    regscale_parent_module = None

    logger.info("Proceed only after you have made the necessary changes and have saved file.")

    x = "y" if skip_prompt else input("Ready to Proceed (Y/N): ").lower()

    if x[0] == "y":
        file_path = all_imps_wb
        if not os.path.exists(file_path):
            error_and_exit(f"Unable to locate the file {file_path}.")

        df = load_workbook(file_path)

        sheet_name = df.sheetnames[0]
        sheet_name = sheet_name[sheet_name.find("(") + 1 : sheet_name.find(")")].split("_")
        # set the variables to the correct values
        for item in set(sheet_name):
            try:
                regscale_parent_id = int(item)
            except ValueError:
                regscale_parent_module = item
        if not regscale_parent_id or not regscale_parent_module:
            error_and_exit("Unable to locate the RegScale Parent ID and RegScale Parent Module.")

        df1 = pd.read_excel(all_imps_wb, sheet_name=0, index_col="Id")

        df2 = pd.read_excel(old_imps_wb, sheet_name=0, index_col="Id")

        if df1.equals(df2):
            logger.warning("No differences detected.")
            sys.exit(0)

        else:
            logger.warning("*** WARNING *** Differences Found.")

            # Logs changes to txt file

            diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
            ne_stacked = diff_mask.stack()
            changed = ne_stacked[ne_stacked]
            changed.index.names = ["Id", "Column"]
            difference_locations = np.nonzero(diff_mask)
            changed_to = df1.values[difference_locations]
            changed_from = df2.values[difference_locations]
            changes = pd.DataFrame({"From": changed_from, "To": changed_to}, index=changed.index)
            changes.to_csv(
                path / DIFFS,
                header=True,
                index=True,
                sep=" ",
                mode="a",
            )

            upload_data(regscale_parent_id, regscale_parent_module, path)

    logger.info("Please check differences.txt file located in artifacts folder to see changes made.")
    return None


def upload_data(regscale_parent_id: int, regscale_parent_module: str, path: Path) -> None:
    """
    Batch uploads updated control implementation statements to the provided RegScale parent ID.

    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_parent_module: RegScale parent module
    :param Path path: file path where control spreadsheet resides
    :rtype: None
    """
    import pandas as pd  # Optimize import performance

    diff = pd.read_csv(path / DIFFS, header=0, sep=" ", index_col=None)
    ids = []
    for i, row in diff.iterrows():
        ids.append(row["Id"])

    id_df = pd.DataFrame(ids, index=None, columns=["Id"])
    id_df2 = id_df.drop_duplicates()

    reader = pd.read_excel(path / ALL_IMPS)
    accounts = pd.read_excel(path / ALL_IMPS, sheet_name="Accounts")
    accounts = accounts.rename(columns={"User": "ControlOwner", "UserId": "ControlOwnerId"})

    updates = reader[reader["Id"].isin(id_df2["Id"])]
    updates = updates.merge(accounts, how="left", on="ControlOwner")
    updates = updates.T.to_dict()

    _ = [build_implementation(i, regscale_parent_id, regscale_parent_module) for i in updates.values()]

    ControlImplementation.bulk_save()


def build_implementation(i: dict, regscale_parent_id: int, regscale_parent_module: str) -> ControlImplementation:
    """
    Builds a ControlImplementation object from a dictionary

    :param dict i: dictionary of control implementation
    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_parent_module: RegScale parent module
    :return: ControlImplementation object
    :rtype: ControlImplementation
    """
    control = Control(
        title=i["ControlTitle"],
        description=i["Description"],
        controlId=i["ControlName"],
        weight=i["Weight"],
        catalogueID=i["CatalogueId"],
    )
    control_implementation = ControlImplementation(
        id=i["Id"],
        controlOwnerId=i["ControlOwnerId"],
        control=control.dict(),
        status=i["Status"],
        implementation=check_empty_nan(i["Implementation"]),
        policy=check_empty_nan(i["Policy"]),
        controlID=i["ControlId"],
        responsibility=check_empty_nan(i["Responsibility"]),
        parentId=regscale_parent_id,
        parentModule=regscale_parent_module,
        inheritable=check_inheritable(i["Inheritable"]),
        dateLastUpdated=get_current_datetime(),
    )
    return control_implementation.save(bulk=True)


# Delete and remove files from user's system.
@control_editor.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
def generate_delete_file(path: Path):
    """This command will delete files used during the Control editing process."""
    warnings.filterwarnings("always", category=DeprecationWarning)
    warnings.warn(
        "Control Editor is deprecated and will be removed in a future release. Use `regscale model` with the `--model control` argument instead.",
        DeprecationWarning,
        stacklevel=2,
    )
    delete_file(path)


def delete_file(path: Path) -> int:
    """
    Deletes files used during the process

    :param Path path: directory of file location
    :return: Number of files deleted
    :rtype: int
    """
    logger = create_logger()
    file_names = [
        ALL_IMPS,
        OLD_IMPS,
        DIFFS,
    ]
    deleted_files = []

    for file_name in file_names:
        if os.path.isfile(path / file_name):
            os.remove(path / file_name)
            deleted_files.append(file_name)
        else:
            logger.warning("No %s file found. Checking for other files before exiting.", file_name)
    logger.info("%i files have been deleted: %s", len(deleted_files), ", ".join(deleted_files))
    return len(deleted_files)


def check_inheritable(
    value: Any,
) -> Union[
    str, float, bool
]:  # this function has to be checked separate to account for API only accepting False Boolean unlike other class params
    """This function takes a given value for an inheritable and checks if value is empty or NaN based on value type.

    :param Any value: A string or float object
    :return: A string value, float value or False
    :rtype: Union[str, float, bool]
    """
    if isinstance(value, str) and value.strip() == "":
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return value


def get_maximum_rows(*, sheet_object: Any) -> int:
    """This function finds the last row containing data in a spreadsheet

    :param Any sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(any(col.value is not None for col in row) for max_row, row in enumerate(sheet_object, 1))
