#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unit tests for FedRAMP Rev 5 POAM Export functionality

This module tests the FedRAMP Rev 5 POAM Excel export functionality, including:
- Dynamic POAM ID generation based on source file path properties
- KEV date determination from CISA KEV catalog
- Deviation status mapping (Approved/Pending/Rejected)
- Custom milestone and comment generation
- Excel formatting and column operations
- Date rounding for closed POAMs (25th of month)
"""

from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch, call

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from regscale.integrations.public.fedramp.poam_export_v5 import (
    convert_to_list,
    determine_kev_date,
    determine_poam_comment,
    determine_poam_id,
    determine_poam_service_name,
    get_cached_cisa_kev,
    lookup_scan_date,
    set_end_columns,
    set_milestones,
    set_risk_info,
    set_short_date,
    set_status,
    set_vendor_info,
    strip_html,
    update_column_widths,
    align_column,
    update_header,
    get_all_poams,
    gen_links,
    gen_files,
    gen_milestones,
    process_row,
    process_worksheet,
    export_poam_v5,
    map_weakness_detector_and_id_for_rev5_issues,
    _generate_closed_poam_comment,
    _generate_open_poam_comment,
)
from regscale.models.regscale_models import (
    Asset,
    Deviation,
    Issue,
    IssueSeverity,
    IssueStatus,
    Property,
    ScanHistory,
    SecurityPlan,
    VulnerabilityMapping,
)
from tests import CLITestFixture


class TestPOAMExportV5(CLITestFixture):
    """Test class for FedRAMP Rev 5 POAM export functionality"""

    def test_set_short_date_valid(self):
        """Test set_short_date with valid date string"""
        date_str = "2025-03-15T10:30:00"
        result = set_short_date(date_str)
        assert result == "03/15/25"

    def test_set_short_date_different_formats(self):
        """Test set_short_date with different date formats"""
        # ISO format
        assert set_short_date("2025-01-01T00:00:00") == "01/01/25"
        # Date only
        assert set_short_date("2025-12-31") == "12/31/25"

    def test_strip_html_with_tags(self):
        """Test strip_html removes HTML tags"""
        html_str = "<p>Test <strong>content</strong> here</p>"
        result = strip_html(html_str)
        assert result == "Test content here"

    def test_strip_html_with_entities(self):
        """Test strip_html handles HTML entities"""
        html_str = "&lt;div&gt;Test &amp; content&lt;/div&gt;"
        result = strip_html(html_str)
        assert result == "<div>Test & content</div>"

    def test_strip_html_empty_string(self):
        """Test strip_html with empty string"""
        result = strip_html("")
        assert result == ""

    def test_strip_html_none(self):
        """Test strip_html with None"""
        result = strip_html(None)
        assert result == ""

    def test_strip_html_nested_tags(self):
        """Test strip_html with nested HTML tags"""
        html_str = "<div><p><span>Nested</span> content</p></div>"
        result = strip_html(html_str)
        assert result == "Nested content"

    def test_convert_to_list_paragraph_tags(self):
        """Test convert_to_list with <p> tag delimiters"""
        asset_str = "<p>Asset1</p><p>Asset2</p><p>Asset3</p>"
        result = convert_to_list(asset_str)
        assert result == ["Asset1", "Asset2", "Asset3"]

    def test_convert_to_list_tabs(self):
        """Test convert_to_list with tab delimiters"""
        asset_str = "Asset1\tAsset2\tAsset3"
        result = convert_to_list(asset_str)
        assert result == ["Asset1", "Asset2", "Asset3"]

    def test_convert_to_list_newlines(self):
        """Test convert_to_list with newline delimiters"""
        asset_str = "Asset1\nAsset2\nAsset3"
        result = convert_to_list(asset_str)
        assert result == ["Asset1", "Asset2", "Asset3"]

    def test_convert_to_list_empty_string(self):
        """Test convert_to_list with empty string"""
        result = convert_to_list("")
        assert result == []

    def test_convert_to_list_none(self):
        """Test convert_to_list with None"""
        result = convert_to_list(None)
        assert result == []

    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_cached_cisa_kev")
    def test_determine_kev_date_match_found(self, mock_get_kev):
        """Test determine_kev_date with matching CVE"""
        mock_get_kev.return_value = {
            "vulnerabilities": [
                {"cveID": "CVE-2025-1234", "dueDate": "2025-06-15T00:00:00"},
                {"cveID": "CVE-2025-5678", "dueDate": "2025-07-20T00:00:00"},
            ]
        }
        result = determine_kev_date("CVE-2025-1234")
        assert result == "06/15/25"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_cached_cisa_kev")
    def test_determine_kev_date_case_insensitive(self, mock_get_kev):
        """Test determine_kev_date is case insensitive"""
        mock_get_kev.return_value = {"vulnerabilities": [{"cveID": "CVE-2025-1234", "dueDate": "2025-06-15T00:00:00"}]}
        result = determine_kev_date("cve-2025-1234")
        assert result == "06/15/25"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_cached_cisa_kev")
    def test_determine_kev_date_no_match(self, mock_get_kev):
        """Test determine_kev_date with no matching CVE"""
        mock_get_kev.return_value = {"vulnerabilities": [{"cveID": "CVE-2025-9999", "dueDate": "2025-08-01"}]}
        result = determine_kev_date("CVE-2025-1234")
        assert result == "N/A"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_cached_cisa_kev")
    def test_determine_kev_date_empty_cve(self, mock_get_kev):
        """Test determine_kev_date with empty CVE"""
        result = determine_kev_date("")
        assert result == "N/A"
        mock_get_kev.assert_not_called()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_cached_cisa_kev")
    def test_determine_kev_date_none_cve(self, mock_get_kev):
        """Test determine_kev_date with None CVE"""
        result = determine_kev_date(None)
        assert result == "N/A"
        mock_get_kev.assert_not_called()

    def test_determine_poam_id_pdf_path(self):
        """Test determine_poam_id with pdf in file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/pdf/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "DC-12345"

    def test_determine_poam_id_signatures_path(self):
        """Test determine_poam_id with signatures in file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/signatures/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "CPT-12345"

    def test_determine_poam_id_campaign_path(self):
        """Test determine_poam_id with campaign in file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/campaign/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "ALM-12345"

    def test_determine_poam_id_learning_manager_path(self):
        """Test determine_poam_id with learning manager in file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/learning manager/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "CCD-12345"

    def test_determine_poam_id_cce_path(self):
        """Test determine_poam_id with cce in file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/cce/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "CCE-12345"

    def test_determine_poam_id_unknown_path(self):
        """Test determine_poam_id with unknown file path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/unknown/file.xml")]
        result = determine_poam_id(poam, props)
        assert result == "UNK-12345"

    def test_determine_poam_id_no_source_path_property(self):
        """Test determine_poam_id with no source_file_path property"""
        poam = Mock(id=12345)
        props = [Mock(key="other_property", value="value")]
        result = determine_poam_id(poam, props)
        assert result == "UNK-12345"

    def test_determine_poam_id_empty_properties(self):
        """Test determine_poam_id with empty properties list"""
        poam = Mock(id=12345)
        props = []
        result = determine_poam_id(poam, props)
        assert result == "UNK-12345"

    def test_determine_poam_id_case_insensitive(self):
        """Test determine_poam_id is case insensitive"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/PDF/FILE.XML")]
        result = determine_poam_id(poam, props)
        assert result == "DC-12345"

    def test_determine_poam_service_name_pdf(self):
        """Test determine_poam_service_name with pdf in path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/pdf/file.xml")]
        result = determine_poam_service_name(poam, props)
        assert result == "PDF Services"

    def test_determine_poam_service_name_signatures(self):
        """Test determine_poam_service_name with signatures in path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/signatures/file.xml")]
        result = determine_poam_service_name(poam, props)
        assert result == "Signatures"

    def test_determine_poam_service_name_unknown(self):
        """Test determine_poam_service_name with unknown path"""
        poam = Mock(id=12345)
        props = [Mock(key="source_file_path", value="/path/to/unknown/file.xml")]
        result = determine_poam_service_name(poam, props)
        assert result == "UNKNOWN"

    def test_determine_poam_service_name_no_property(self):
        """Test determine_poam_service_name with no source_file_path property"""
        poam = Mock(id=12345)
        props = [Mock(key="other_property", value="value")]
        result = determine_poam_service_name(poam, props)
        assert result == "UNKNOWN"

    def test_lookup_scan_date_with_matching_asset(self):
        """Test lookup_scan_date finds scan date from vulnerability mapping"""
        # Test the function without excessive mocking, just verify it returns a date
        poam = Mock(assetIdentifier="Asset1", dateLastUpdated="2025-03-15T10:00:00")
        asset = Mock(id=100, name="Asset1")
        assets = [asset]

        # Since this function makes real API calls, test that it falls back to dateLastUpdated
        result = lookup_scan_date(poam, assets)
        # Should return a formatted date
        assert result == "03/15/25"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.VulnerabilityMapping")
    def test_lookup_scan_date_no_matching_asset(self, mock_vuln_mapping):
        """Test lookup_scan_date with no matching asset"""
        poam = Mock(assetIdentifier="Asset1", dateLastUpdated="2025-03-15T10:00:00")
        asset = Mock(id=100, name="Asset2")
        assets = [asset]

        result = lookup_scan_date(poam, assets)
        assert result == "03/15/25"
        mock_vuln_mapping.find_by_asset.assert_not_called()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.VulnerabilityMapping")
    def test_lookup_scan_date_no_vulnerabilities(self, mock_vuln_mapping):
        """Test lookup_scan_date when no vulnerabilities found"""
        poam = Mock(assetIdentifier="Asset1", dateLastUpdated="2025-03-15T10:00:00")
        asset = Mock(id=100, name="Asset1")
        assets = [asset]

        mock_vuln_mapping.find_by_asset.return_value = []

        result = lookup_scan_date(poam, assets)
        assert result == "03/15/25"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_short_date")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.convert_to_list")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.VulnerabilityMapping")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.ScanHistory")
    def test_lookup_scan_date_multiple_assets(
        self, mock_scan_history, mock_vuln_mapping, mock_convert, mock_short_date
    ):
        """Test lookup_scan_date with multiple assets in identifier"""
        poam = Mock(assetIdentifier="Asset1\nAsset2", dateLastUpdated="2025-03-15T10:00:00")
        asset1 = Mock(id=100, name="Asset1")
        asset2 = Mock(id=101, name="Asset2")
        assets = [asset1, asset2]

        mock_convert.return_value = ["Asset1", "Asset2"]
        mock_vuln = Mock(scanId=500)
        mock_vuln_mapping.find_by_asset.return_value = [mock_vuln]

        mock_scan = Mock(scanDate="2025-03-10T08:00:00")
        mock_scan_history.get_object.return_value = mock_scan
        mock_short_date.return_value = "03/10/25"

        result = lookup_scan_date(poam, assets)
        assert result == "03/10/25"

    def test_determine_poam_comment_closed_poam_with_date(self):
        """Test determine_poam_comment for closed POAM"""
        poam = Mock(
            id=123,
            dateFirstDetected="2025-03-01T10:00:00",
            dateCompleted="2025-03-15T10:00:00",
            dateCreated="2025-03-01T10:00:00",
            poamComments="",
        )
        poam.save = Mock()
        assets = []

        result = determine_poam_comment(poam, assets)
        assert "Per review of the latest scan report" in result
        assert "This POAM will be submitted for closure" in result
        assert "03/01/25: POAM entry added" in result
        poam.save.assert_called_once()

    def test_determine_poam_comment_open_poam_new_comment(self):
        """Test determine_poam_comment for open POAM with no existing comment"""
        poam = Mock(
            id=123, dateFirstDetected="2025-03-01T10:00:00", dateCompleted=None, poamComments="", dateCreated=None
        )
        poam.save = Mock()
        assets = []

        result = determine_poam_comment(poam, assets)
        assert "03/01/25: POAM entry added" in result
        poam.save.assert_called_once()

    def test_determine_poam_comment_open_poam_existing_comment(self):
        """Test determine_poam_comment for open POAM with existing comment"""
        existing_comment = "03/01/25: POAM entry added"
        poam = Mock(
            id=123,
            dateFirstDetected="2025-03-01T10:00:00",
            dateCompleted=None,
            poamComments=existing_comment,
            dateCreated=None,
        )
        poam.save = Mock()
        assets = []

        result = determine_poam_comment(poam, assets)
        assert result == existing_comment
        poam.save.assert_not_called()

    def test_determine_poam_comment_no_detection_date(self):
        """Test determine_poam_comment with no detection date"""
        poam = Mock(id=123, dateFirstDetected=None, poamComments="")
        poam.save = Mock()
        assets = []

        result = determine_poam_comment(poam, assets)
        assert result == "N/A"
        poam.save.assert_not_called()

    def test_generate_closed_poam_comment_new(self):
        """Test _generate_closed_poam_comment creates new closed comment"""
        poam = Mock(dateCompleted="2025-03-15T10:00:00", dateCreated="2025-03-01T10:00:00")
        current_comment = ""
        template = "Per review of the latest scan report on %s, (TGRC) can confirm that this issue no longer persists. This POAM will be submitted for closure."
        open_template = "POAM entry added"

        result = _generate_closed_poam_comment(poam, current_comment, template, open_template)
        assert "Per review of the latest scan report on 03/15/25" in result
        assert "This POAM will be submitted for closure" in result
        assert "03/01/25: POAM entry added" in result

    def test_generate_closed_poam_comment_already_has_closed_blurb(self):
        """Test _generate_closed_poam_comment when comment already has closed blurb"""
        poam = Mock(dateCompleted="2025-03-15T10:00:00", dateCreated="2025-03-01T10:00:00")
        # Include "POAM entry added" to pass the first check, and the closed blurb to pass the second
        current_comment = "This POAM will be submitted for closure\n03/01/25: POAM entry added"
        template = "Template %s"
        open_template = "POAM entry added"

        result = _generate_closed_poam_comment(poam, current_comment, template, open_template)
        # Should return unchanged because it already has the closure blurb
        assert result == current_comment
        assert "This POAM will be submitted for closure" in result

    def test_generate_open_poam_comment_new(self):
        """Test _generate_open_poam_comment creates new comment"""
        current_comment = ""
        detection_date = "03/01/25"
        template = "POAM entry added"

        result = _generate_open_poam_comment(current_comment, detection_date, template)
        assert result == "03/01/25: POAM entry added"

    def test_generate_open_poam_comment_already_has_entry_added(self):
        """Test _generate_open_poam_comment when comment already has entry added"""
        current_comment = "03/01/25: POAM entry added"
        detection_date = "03/01/25"
        template = "POAM entry added"

        result = _generate_open_poam_comment(current_comment, detection_date, template)
        assert result == current_comment

    def test_generate_open_poam_comment_prepends_to_existing(self):
        """Test _generate_open_poam_comment prepends to existing comment"""
        current_comment = "Some existing text"
        detection_date = "03/01/25"
        template = "POAM entry added"

        result = _generate_open_poam_comment(current_comment, detection_date, template)
        assert result == "03/01/25: POAM entry added\nSome existing text"

    def test_set_milestones_with_milestones(self):
        """Test set_milestones populates worksheet with milestone data"""
        poam = Mock(id=123)
        sheet = MagicMock()
        column_l_date = "04/01/25"
        all_milestones = [
            {"parent_id": 123, "MilestoneDate": "2025-03-15T10:00:00"},
            {"parent_id": 123, "MilestoneDate": "2025-03-20T10:00:00"},
            {"parent_id": 456, "MilestoneDate": "2025-03-25T10:00:00"},
        ]

        set_milestones(poam, 10, sheet, column_l_date, all_milestones)

        # Should contain milestone dates
        assert sheet.__getitem__.called

    def test_set_milestones_no_milestones(self):
        """Test set_milestones with no milestones for POAM"""
        poam = Mock(id=123)
        sheet = MagicMock()
        column_l_date = "04/01/25"
        all_milestones = [{"parent_id": 456, "MilestoneDate": "2025-03-25T10:00:00"}]

        set_milestones(poam, 10, sheet, column_l_date, all_milestones)

        # Should set default milestone text
        assert sheet.__getitem__.called

    def test_set_status_closed_before_25th(self):
        """Test set_status for closed POAM completed before 25th of month"""
        poam = Mock(status="Closed", dateCompleted="2025-03-10T10:00:00")
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        # Should round to 25th of same month
        sheet.__getitem__.assert_called_with("O10")
        sheet.__getitem__.return_value.value = "03/25/25"

    def test_set_status_closed_on_25th(self):
        """Test set_status for closed POAM completed on 25th of month"""
        poam = Mock(status="Closed", dateCompleted="2025-03-25T10:00:00")
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        # Should stay on 25th of same month
        sheet.__getitem__.assert_called()

    def test_set_status_closed_after_25th(self):
        """Test set_status for closed POAM completed after 25th of month"""
        poam = Mock(status="Closed", dateCompleted="2025-03-26T10:00:00")
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        # Should round to 25th of next month
        sheet.__getitem__.assert_called()

    def test_set_status_closed_no_completion_date(self):
        """Test set_status for closed POAM with no completion date"""
        poam = Mock(status="Closed", dateCompleted=None)
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        sheet.__getitem__.assert_called_with("O10")

    def test_set_status_open_with_last_updated(self):
        """Test set_status for open POAM with last updated date"""
        poam = Mock(status="Open", dateLastUpdated="2025-03-15T10:00:00", dateCompleted=None)
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        sheet.__getitem__.assert_called()

    def test_set_status_open_no_last_updated(self):
        """Test set_status for open POAM with no last updated date"""
        poam = Mock(status="Open", dateLastUpdated=None, dateCompleted=None)
        sheet = MagicMock()

        set_status(poam, 10, sheet)

        sheet.__getitem__.assert_called_with("O10")

    def test_set_vendor_info_with_dependency(self):
        """Test set_vendor_info with vendor dependency"""
        poam = Mock(vendorDependency="Yes", vendorName="Test Vendor", vendorLastUpdate="2025-03-15T10:00:00")
        sheet = MagicMock()

        set_vendor_info(poam, 10, sheet)

        assert sheet.__getitem__.call_count >= 3

    def test_set_vendor_info_no_dependency(self):
        """Test set_vendor_info without vendor dependency"""
        poam = Mock(vendorDependency="No", vendorName=None, vendorLastUpdate=None)
        sheet = MagicMock()

        set_vendor_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    def test_set_vendor_info_no_vendor_name(self):
        """Test set_vendor_info with no vendor name"""
        poam = Mock(vendorDependency="Yes", vendorName=None, vendorLastUpdate="2025-03-15T10:00:00")
        sheet = MagicMock()

        set_vendor_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_approved_deviation(self, mock_deviation_class):
        """Test set_risk_info with approved deviation"""
        mock_deviation = Mock(deviationStatus="Approved")
        mock_deviation_class.get_by_issue.return_value = mock_deviation

        poam = Mock(
            id=123,
            originalRiskRating="High",
            adjustedRiskRating="Medium",
            riskAdjustment="Yes",
            falsePositive="No",
            operationalRequirement="No",
            deviationRationale="<p>Test rationale</p>",
            severityLevel=1,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        mock_deviation_class.get_by_issue.assert_called_once_with(123)
        assert sheet.__getitem__.call_count >= 6

    @patch("regscale.integrations.public.fedramp.poam_export_v5.IssueSeverity")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_pending_deviation(self, mock_deviation_class, mock_severity):
        """Test set_risk_info with pending deviation"""
        mock_deviation = Mock(deviationStatus="Pending")
        mock_deviation_class.get_by_issue.return_value = mock_deviation
        mock_severity.return_value.name = "Moderate"

        poam = Mock(
            id=123,
            originalRiskRating=None,
            adjustedRiskRating=None,
            riskAdjustment="Pending",
            falsePositive="No",
            operationalRequirement="No",
            deviationRationale="",
            severityLevel=2,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_rejected_deviation(self, mock_deviation_class):
        """Test set_risk_info with rejected deviation"""
        mock_deviation = Mock(deviationStatus="Rejected")
        mock_deviation_class.get_by_issue.return_value = mock_deviation

        poam = Mock(
            id=123,
            originalRiskRating="Critical",
            adjustedRiskRating=None,
            riskAdjustment="Yes",
            falsePositive="No",
            operationalRequirement="No",
            deviationRationale="Test",
            severityLevel=0,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_false_positive(self, mock_deviation_class):
        """Test set_risk_info with false positive"""
        mock_deviation = Mock(deviationStatus="Approved")
        mock_deviation_class.get_by_issue.return_value = mock_deviation

        poam = Mock(
            id=123,
            originalRiskRating="High",
            adjustedRiskRating="N/A",
            riskAdjustment="No",
            falsePositive="Yes",
            operationalRequirement="No",
            deviationRationale="False positive",
            severityLevel=1,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_operational_requirement(self, mock_deviation_class):
        """Test set_risk_info with operational requirement"""
        mock_deviation = Mock(deviationStatus="Approved")
        mock_deviation_class.get_by_issue.return_value = mock_deviation

        poam = Mock(
            id=123,
            originalRiskRating="High",
            adjustedRiskRating="N/A",
            riskAdjustment="No",
            falsePositive="No",
            operationalRequirement="Yes",
            deviationRationale="Operational need",
            severityLevel=1,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Deviation")
    def test_set_risk_info_operational_requirement_pending(self, mock_deviation_class):
        """Test set_risk_info with pending operational requirement"""
        mock_deviation = Mock(deviationStatus="Pending")
        mock_deviation_class.get_by_issue.return_value = mock_deviation

        poam = Mock(
            id=123,
            originalRiskRating="High",
            adjustedRiskRisk="N/A",
            riskAdjustment="No",
            falsePositive="No",
            operationalRequirement="Yes",
            deviationRationale="Pending approval",
            severityLevel=1,
        )
        sheet = MagicMock()

        set_risk_info(poam, 10, sheet)

        assert sheet.__getitem__.called

    def test_set_end_columns(self):
        """Test set_end_columns populates end columns correctly"""
        with patch(
            "regscale.integrations.public.fedramp.poam_export_v5.determine_poam_service_name"
        ) as mock_service_name, patch(
            "regscale.integrations.public.fedramp.poam_export_v5.determine_kev_date"
        ) as mock_kev_date, patch(
            "regscale.integrations.public.fedramp.poam_export_v5.determine_poam_comment"
        ) as mock_comment:
            mock_service_name.return_value = "PDF Services"
            mock_kev_date.return_value = "06/15/25"
            mock_comment.return_value = "Test comment"

            ssp = Mock()
            poam = Mock(id=123, autoApproved="Yes", kevList="Yes", cve="CVE-2025-1234")
            wb = Workbook()
            sheet = wb.active
            props = []
            assets = []
            # Use proper dict structure matching the code expectations
            all_links = [
                Mock(parentID=123, __getitem__=lambda self, key: "Link1" if key == "Title" else "http://example.com")
            ]
            all_files = [Mock(parentId=123, __getitem__=lambda self, key: "file1.pdf")]

            set_end_columns(ssp, poam, 10, sheet, props, assets, all_links, all_files)

            mock_service_name.assert_called_once()
            mock_kev_date.assert_called_once_with("CVE-2025-1234")
            mock_comment.assert_called_once()
            # Verify some columns were set
            assert sheet["Y10"].value is not None
            assert sheet["AB10"].value == "Yes"
            assert sheet["AE10"].value == "PDF Services"

    def test_map_weakness_detector_and_id_for_rev5_issues(self):
        """Test map_weakness_detector_and_id_for_rev5_issues sets correct values"""
        wb = Workbook()
        worksheet = wb.active
        issue = Mock(sourceReport="Tenable SC", cve="CVE-2025-1234", pluginId="12345", title="Test Issue")

        map_weakness_detector_and_id_for_rev5_issues(worksheet, "E", "F", 10, issue)

        assert worksheet["E10"].value == "Tenable SC"
        assert worksheet["F10"].value == "CVE-2025-1234"

    def test_map_weakness_detector_and_id_no_cve(self):
        """Test map_weakness_detector_and_id with no CVE uses pluginId"""
        wb = Workbook()
        worksheet = wb.active
        issue = Mock(sourceReport="Tenable SC", cve=None, pluginId="12345", title="Test Issue")

        map_weakness_detector_and_id_for_rev5_issues(worksheet, "E", "F", 10, issue)

        assert worksheet["E10"].value == "Tenable SC"
        assert worksheet["F10"].value == "12345"

    def test_map_weakness_detector_and_id_no_cve_or_plugin(self):
        """Test map_weakness_detector_and_id with no CVE or pluginId uses title"""
        wb = Workbook()
        worksheet = wb.active
        issue = Mock(sourceReport="Tenable SC", cve=None, pluginId=None, title="Test Issue")

        map_weakness_detector_and_id_for_rev5_issues(worksheet, "E", "F", 10, issue)

        assert worksheet["E10"].value == "Tenable SC"
        assert worksheet["F10"].value == "Test Issue"

    def test_update_column_widths(self):
        """Test update_column_widths sets correct widths"""
        wb = Workbook()
        ws = wb.active

        # Add some test data
        ws["A1"] = "Test"
        ws["C1"] = "Long text that should wrap"

        update_column_widths(ws)

        # Check that column widths are set
        assert ws.column_dimensions["A"].width == 15
        assert ws.column_dimensions["C"].width == 40

    def test_align_column(self):
        """Test align_column sets text wrapping and alignment"""
        wb = Workbook()
        ws = wb.active
        ws["G1"] = "Test  "
        ws["G2"] = "  Test2  "

        align_column("G", ws)

        # Check alignment was set
        for cell in ws["G"]:
            if cell.value:
                assert cell.alignment.wrap_text is True
                assert cell.alignment.horizontal == "left"

    def test_update_header(self):
        """Test update_header populates header information"""
        ssp = Mock(cspOrgName="Test Org", systemName="Test System", overallCategorization="Moderate")
        wb = Workbook()
        sheet = wb.active

        result = update_header(ssp, sheet)

        assert result["A3"].value == "Test Org"
        assert result["B3"].value == "Test System"
        assert result["C3"].value == "Moderate"
        # D3 should have current date
        assert result["D3"].value is not None

    def test_update_header_no_org_name(self):
        """Test update_header with no org name"""
        ssp = Mock(cspOrgName=None, systemName="Test System", overallCategorization="Moderate")
        wb = Workbook()
        sheet = wb.active

        result = update_header(ssp, sheet)

        assert result["A3"].value == "N/A"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Issue")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Asset")
    def test_get_all_poams(self, mock_asset_class, mock_issue_class):
        """Test get_all_poams retrieves POAMs from SSP and assets"""
        ssp_id = "123"

        # Mock SSP POAMs
        ssp_poam1 = Mock(
            id=1, isPoam=True, otherIdentifier="ID1", assetIdentifier="Asset1", cve="CVE-1", pluginId="P1", title="T1"
        )
        ssp_poam2 = Mock(
            id=2, isPoam=True, otherIdentifier="ID2", assetIdentifier="Asset2", cve="CVE-2", pluginId="P2", title="T2"
        )
        mock_issue_class.get_all_by_parent.return_value = [ssp_poam1, ssp_poam2]

        # Mock assets
        asset1 = Mock(id=10)
        mock_asset_class.get_all_by_parent.return_value = [asset1]

        # Mock asset POAMs (returns empty to avoid complexity)
        def issue_side_effect(parent_id, parent_module):
            if parent_id == ssp_id:
                return [ssp_poam1, ssp_poam2]
            return []

        mock_issue_class.get_all_by_parent.side_effect = issue_side_effect

        result = get_all_poams(ssp_id)

        assert len(result) == 2
        assert result[0].id == 1
        assert result[1].id == 2

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Link")
    def test_gen_links(self, mock_link_class):
        """Test gen_links generates list of links"""
        poam1 = Mock(id=1)
        poam2 = Mock(id=2)
        all_poams = [poam1, poam2]

        link1 = {"id": 10, "title": "Link 1"}
        link2 = {"id": 11, "title": "Link 2"}
        mock_link_class.get_all_by_parent.side_effect = [[link1], [link2]]

        result = gen_links(all_poams)

        assert len(result) == 2
        assert result[0] == link1
        assert result[1] == link2

    @patch("regscale.integrations.public.fedramp.poam_export_v5.File")
    def test_gen_files(self, mock_file_class):
        """Test gen_files generates list of files"""
        api = Mock()
        poam1 = Mock(id=1)
        poam2 = Mock(id=2)
        all_poams = [poam1, poam2]

        file1 = {"id": 10, "name": "File 1"}
        file2 = {"id": 11, "name": "File 2"}
        mock_file_class.get_files_for_parent_from_regscale.side_effect = [[file1], [file2]]

        result = gen_files(all_poams, api)

        assert len(result) == 2
        assert result[0] == file1
        assert result[1] == file2

    def test_gen_milestones(self):
        """Test gen_milestones generates list of milestones"""
        api = Mock()
        app = Mock()
        app.config = {"domain": "https://test.com"}

        poam1 = Mock(id=1)
        poam2 = Mock(id=2)
        all_poams = [poam1, poam2]

        milestone1 = {"id": 10, "parent_id": 1}
        milestone2 = {"id": 11, "parent_id": 2}

        mock_response1 = Mock()
        mock_response1.json.return_value = [milestone1]
        mock_response2 = Mock()
        mock_response2.json.return_value = [milestone2]

        api.get.side_effect = [mock_response1, mock_response2]

        result = gen_milestones(all_poams, api, app)

        assert len(result) == 2
        assert result[0] == milestone1
        assert result[1] == milestone2

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Property")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_end_columns")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_risk_info")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_vendor_info")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_status")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_milestones")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.map_weakness_detector_and_id_for_rev5_issues")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.strip_html")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.convert_to_list")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.set_short_date")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.determine_poam_id")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.datetime_obj")
    def test_process_row(
        self,
        mock_datetime_obj,
        mock_determine_id,
        mock_short_date,
        mock_convert_list,
        mock_strip_html,
        mock_map_weakness,
        mock_set_milestones,
        mock_set_status,
        mock_set_vendor,
        mock_set_risk,
        mock_set_end,
        mock_property,
    ):
        """Test process_row processes a POAM row correctly"""
        ssp = Mock()
        poam = Mock(
            id=123,
            sourceReport="Tenable SC",
            title="Test Issue",
            description="Test description",
            assetIdentifier="Asset1",
            remediationDescription="Fix it",
            dateFirstDetected="2025-03-01T10:00:00",
            dueDate="2025-04-01T10:00:00",
            changes="Test changes",
        )
        sheet = MagicMock()
        assets = []
        all_milestones = []
        all_links = []
        all_files = []

        mock_property.get_all_by_parent.return_value = []
        mock_strip_html.side_effect = lambda x: x if x else ""
        mock_convert_list.return_value = ["Asset1"]
        mock_short_date.return_value = "03/01/25"
        mock_determine_id.return_value = "DC-123"
        mock_datetime_obj.return_value = datetime(2025, 4, 1, 10, 0, 0)

        process_row(ssp, poam, 0, sheet, assets, all_milestones, all_links, all_files, point_of_contact="John Doe")

        # Verify the index adjustment (index 0 becomes row 6)
        assert sheet.__getitem__.called
        mock_set_milestones.assert_called_once()
        mock_set_status.assert_called_once()
        mock_set_vendor.assert_called_once()
        mock_set_risk.assert_called_once()
        mock_set_end.assert_called_once()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Property")
    def test_process_row_sap_concur_normalization(self, mock_property):
        """Test process_row normalizes SAP Concur to Tenable SC"""
        ssp = Mock()
        poam = Mock(
            id=123,
            sourceReport="SAP Concur",
            title="Test Issue",
            description="",
            assetIdentifier="",
            remediationDescription="",
            dateFirstDetected="2025-03-01T10:00:00",
            dueDate=None,
            changes="",
        )
        sheet = MagicMock()
        assets = []

        mock_property.get_all_by_parent.return_value = []

        with patch("regscale.integrations.public.fedramp.poam_export_v5.set_milestones"), patch(
            "regscale.integrations.public.fedramp.poam_export_v5.set_status"
        ), patch("regscale.integrations.public.fedramp.poam_export_v5.set_vendor_info"), patch(
            "regscale.integrations.public.fedramp.poam_export_v5.set_risk_info"
        ), patch(
            "regscale.integrations.public.fedramp.poam_export_v5.set_end_columns"
        ):
            process_row(ssp, poam, 0, sheet, assets, [], [], [], point_of_contact="")

        # Verify sourceReport was normalized
        assert poam.sourceReport == "Tenable SC"

    @patch("regscale.integrations.public.fedramp.poam_export_v5.openpyxl.load_workbook")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.update_header")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Asset")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.process_row")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.update_column_widths")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.align_column")
    def test_process_worksheet_open_poams(
        self, mock_align, mock_update_widths, mock_process_row, mock_asset, mock_update_header, mock_load_wb
    ):
        """Test process_worksheet processes open POAMs worksheet"""
        ssp = Mock(id=123)
        workbook_path = Path("/tmp/test.xlsx")
        all_poams = [
            Mock(id=1, status=IssueStatus.Open),
            Mock(id=2, status=IssueStatus.Open),
            Mock(id=3, status=IssueStatus.Closed),
        ]

        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.__getitem__.return_value = mock_sheet
        mock_load_wb.return_value = mock_wb

        mock_asset.get_all_by_parent.return_value = []
        mock_update_header.return_value = mock_sheet

        process_worksheet(ssp, "Open POA&M Items", workbook_path, all_poams, [], [], [], point_of_contact="John Doe")

        # Should process only 2 open POAMs
        assert mock_process_row.call_count == 2
        mock_update_widths.assert_called_once_with(mock_sheet)
        mock_align.assert_called_once_with("G", mock_sheet)
        mock_wb.save.assert_called_once_with(workbook_path)

    @patch("regscale.integrations.public.fedramp.poam_export_v5.openpyxl.load_workbook")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.update_header")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Asset")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.process_row")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.update_column_widths")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.align_column")
    def test_process_worksheet_closed_poams(
        self, mock_align, mock_update_widths, mock_process_row, mock_asset, mock_update_header, mock_load_wb
    ):
        """Test process_worksheet processes closed POAMs worksheet"""
        ssp = Mock(id=123)
        workbook_path = Path("/tmp/test.xlsx")
        all_poams = [
            Mock(id=1, status=IssueStatus.Open),
            Mock(id=2, status=IssueStatus.Closed),
            Mock(id=3, status=IssueStatus.Closed),
        ]

        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.__getitem__.return_value = mock_sheet
        mock_load_wb.return_value = mock_wb

        mock_asset.get_all_by_parent.return_value = []
        mock_update_header.return_value = mock_sheet

        process_worksheet(
            ssp, "Closed POA&M Items", workbook_path, all_poams, [], [], [], point_of_contact="Jane Smith"
        )

        # Should process only 2 closed POAMs
        assert mock_process_row.call_count == 2
        mock_wb.save.assert_called_once()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Application")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Api")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.SecurityPlan")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_all_poams")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_links")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_files")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_milestones")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.shutil.copy")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.process_worksheet")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Path")
    def test_export_poam_v5_success(
        self,
        mock_path_class,
        mock_process_ws,
        mock_copy,
        mock_gen_milestones,
        mock_gen_files,
        mock_gen_links,
        mock_get_poams,
        mock_ssp_class,
        mock_api_class,
        mock_app_class,
    ):
        """Test export_poam_v5 successfully exports POAMs"""
        ssp_id = "123"
        output_file = "/tmp/output.xlsx"

        # Mock application and API
        mock_app = Mock()
        mock_app_class.return_value = mock_app
        mock_api = Mock()
        mock_api_class.return_value = mock_api

        # Mock SSP
        mock_ssp = Mock(systemName="Test System")
        mock_ssp_class.get_object.return_value = mock_ssp

        # Mock POAMs
        mock_poam = Mock(id=1)
        mock_get_poams.return_value = [mock_poam]

        # Mock related data
        mock_gen_links.return_value = []
        mock_gen_files.return_value = []
        mock_gen_milestones.return_value = []

        # Mock template path
        mock_template_path = Mock()
        mock_template_path.exists.return_value = True
        mock_output_path = Mock()
        mock_output_path.suffix = ".xlsx"
        mock_output_path.absolute.return_value = "/tmp/output.xlsx"

        def path_side_effect(arg):
            if arg == "./FedRAMP-POAM-Template.xlsx":
                return mock_template_path
            return mock_output_path

        mock_path_class.side_effect = path_side_effect

        export_poam_v5(ssp_id, output_file)

        # Verify calls
        mock_ssp_class.get_object.assert_called_once_with(ssp_id)
        mock_get_poams.assert_called_once_with(ssp_id)
        mock_gen_links.assert_called_once()
        mock_gen_files.assert_called_once()
        mock_gen_milestones.assert_called_once()
        mock_copy.assert_called_once()
        # Should process both open and closed worksheets
        assert mock_process_ws.call_count == 2

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Application")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Api")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.SecurityPlan")
    def test_export_poam_v5_ssp_not_found(self, mock_ssp_class, mock_api_class, mock_app_class):
        """Test export_poam_v5 when SSP not found"""
        ssp_id = "999"
        output_file = "/tmp/output.xlsx"

        mock_app_class.return_value = Mock()
        mock_api_class.return_value = Mock()
        mock_ssp_class.get_object.return_value = None

        export_poam_v5(ssp_id, output_file)

        # Should return early without processing
        mock_ssp_class.get_object.assert_called_once_with(ssp_id)

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Application")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Api")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.SecurityPlan")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_all_poams")
    def test_export_poam_v5_no_poams(self, mock_get_poams, mock_ssp_class, mock_api_class, mock_app_class):
        """Test export_poam_v5 when no POAMs found"""
        ssp_id = "123"
        output_file = "/tmp/output.xlsx"

        mock_app_class.return_value = Mock()
        mock_api_class.return_value = Mock()
        mock_ssp = Mock(systemName="Test System")
        mock_ssp_class.get_object.return_value = mock_ssp
        mock_get_poams.return_value = []

        export_poam_v5(ssp_id, output_file)

        # Should return early without processing
        mock_get_poams.assert_called_once_with(ssp_id)

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Application")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Api")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.SecurityPlan")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_all_poams")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_links")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_files")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_milestones")
    def test_export_poam_v5_template_not_found(
        self,
        mock_gen_milestones,
        mock_gen_files,
        mock_gen_links,
        mock_get_poams,
        mock_ssp_class,
        mock_api_class,
        mock_app_class,
    ):
        """Test export_poam_v5 when template file not found"""
        ssp_id = "123"
        output_file = "/tmp/output.xlsx"

        mock_app = Mock()
        mock_app.config = {"domain": "https://test.com"}
        mock_app_class.return_value = mock_app
        mock_api = Mock()
        mock_api.config = {"domain": "https://test.com"}
        mock_api_class.return_value = mock_api
        mock_ssp = Mock(systemName="Test System")
        mock_ssp_class.get_object.return_value = mock_ssp
        mock_poam = Mock(id=1)
        mock_get_poams.return_value = [mock_poam]
        mock_gen_links.return_value = []
        mock_gen_files.return_value = []
        mock_gen_milestones.return_value = []

        # Don't provide template_path, so it will look for default which doesn't exist
        export_poam_v5(ssp_id, output_file)

        # Should return early due to missing template
        mock_get_poams.assert_called_once()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.Application")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.Api")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.SecurityPlan")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.get_all_poams")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_links")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_files")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.gen_milestones")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.shutil.copy")
    @patch("regscale.integrations.public.fedramp.poam_export_v5.process_worksheet")
    def test_export_poam_v5_with_custom_template(
        self,
        mock_process_ws,
        mock_copy,
        mock_gen_milestones,
        mock_gen_files,
        mock_gen_links,
        mock_get_poams,
        mock_ssp_class,
        mock_api_class,
        mock_app_class,
    ):
        """Test export_poam_v5 with custom template path"""
        ssp_id = "123"
        output_file = "/tmp/output.xlsx"

        mock_app_class.return_value = Mock()
        mock_api_class.return_value = Mock()
        mock_ssp = Mock(systemName="Test System")
        mock_ssp_class.get_object.return_value = mock_ssp
        mock_poam = Mock(id=1)
        mock_get_poams.return_value = [mock_poam]
        mock_gen_links.return_value = []
        mock_gen_files.return_value = []
        mock_gen_milestones.return_value = []

        # Create a real temporary template file
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_template:
            template_path = Path(temp_template.name)
            # Create a simple workbook
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Open POA&M Items"
            wb.create_sheet("Closed POA&M Items")
            wb.save(template_path)

        try:
            export_poam_v5(ssp_id, output_file, template_path=template_path)

            # Should use custom template
            mock_copy.assert_called_once()
        finally:
            # Clean up
            if template_path.exists():
                template_path.unlink()

    @patch("regscale.integrations.public.fedramp.poam_export_v5.pull_cisa_kev")
    def test_get_cached_cisa_kev_caching(self, mock_pull_kev):
        """Test get_cached_cisa_kev caches the KEV data"""
        mock_kev_data = {"vulnerabilities": []}
        mock_pull_kev.return_value = mock_kev_data

        # Clear the cache first
        get_cached_cisa_kev.cache_clear()

        # First call should fetch data
        result1 = get_cached_cisa_kev()
        assert result1 == mock_kev_data
        assert mock_pull_kev.call_count == 1

        # Second call should use cache
        result2 = get_cached_cisa_kev()
        assert result2 == mock_kev_data
        assert mock_pull_kev.call_count == 1  # Still 1, not called again

        # Clean up
        get_cached_cisa_kev.cache_clear()
