"""Excel file header extractor."""

from datetime import datetime
from pathlib import Path

import openpyxl

from ..types import ExcelMetadata
from ..utils.file_utils import get_file_owner


class ExcelExtractor:
    """Extract headers from Excel files."""

    def __init__(self, count_rows: bool = False):
        """Initialize Excel extractor.

        Args:
            count_rows: Whether to count total rows (uses max_row which may be approximate)
        """
        self.count_rows = count_rows

    def extract_header(self, file_path: Path) -> ExcelMetadata:
        """Extract header row and metadata from Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            ExcelMetadata instance with file metadata and header information
        """
        try:
            # Load workbook in read-only mode for performance
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True, keep_links=False
            )

            # Get the active sheet
            sheet = workbook.active

            # Extract header row (first row)
            header_row = []
            for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), []):
                # Convert to string and handle None values
                header_row.append(str(cell) if cell is not None else "")

            # Get dimensions (only if count_rows is enabled)
            # Note: max_row may not be accurate for all Excel files
            rows = sheet.max_row if self.count_rows else None
            cols = sheet.max_column

            # Get file stats
            stats = file_path.stat()

            # Get file owner
            file_owner = get_file_owner(file_path)

            return ExcelMetadata(
                filename=file_path.name,
                filepath=str(file_path.absolute()),
                file_type=file_path.suffix.lower(),
                file_size=stats.st_size,
                file_owner=file_owner,
                modified_at=datetime.fromtimestamp(stats.st_mtime).isoformat(),
                headers=header_row,
                row_count=rows,
                column_count=cols,
                sheet_name=sheet.title if hasattr(sheet, "title") else "Sheet1",
                sheet_names=workbook.sheetnames,
                extracted_at=datetime.now().isoformat(),
            )

        except Exception as e:
            # Return error information
            return ExcelMetadata(
                filename=file_path.name,
                filepath=str(file_path.absolute()),
                file_type=file_path.suffix.lower(),
                error=str(e),
                extracted_at=datetime.now().isoformat(),
            )
