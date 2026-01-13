# SPDX-License-Identifier: MIT

"""Excel export utilities for Pydantic models."""

import io
from datetime import datetime
from contextlib import contextmanager

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from starlette.responses import StreamingResponse

from fastlib.logging.handlers import logger

class ExcelExporter:
    """Excel export handler with dynamic header mapping from Pydantic Schema."""

    DEFAULT_FONT = Font(name="Microsoft YaHei", size=11)
    HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
    HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def __init__(
        self,
        schema: type[BaseModel],
        file_name: str,
        sheet_name: str | None = None,
        include_timestamp: bool = True,
    ):
        self.schema = schema
        self.file_name = file_name
        self.sheet_name = sheet_name or file_name
        self.include_timestamp = include_timestamp
        
        # 核心逻辑：获取字段名到 Field title 的映射
        self.field_names = list(schema.model_fields.keys())
        self.header_mapping = {
            name: field.title if field.title else name 
            for name, field in schema.model_fields.items()
        }

    def _generate_filename(self) -> str:
        if self.include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            return f"{self.file_name}_{timestamp}.xlsx"
        return f"{self.file_name}.xlsx"

    def _prepare_dataframe(self, data_list: list[BaseModel] | None = None) -> pd.DataFrame:
        """将 Pydantic 列表转换为 DataFrame 并重命名表头。"""
        if not data_list:
            # 返回只有表头的空 DataFrame
            return pd.DataFrame(columns=self.header_mapping.values())

        # 1. 转换数据
        data_dicts = [item.model_dump() for item in data_list]
        df = pd.DataFrame(data_dicts, columns=self.field_names)
        
        # 2. 重命名列名为 Schema 中定义的 title
        return df.rename(columns=self.header_mapping)

    def _apply_styles(self, worksheet: Worksheet) -> None:
        """设置样式：表头背景色、字体、自动列宽。"""
        # 设置表头样式
        for cell in worksheet[1]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.DEFAULT_FONT
                cell.alignment = Alignment(vertical="center", horizontal="left")

        # 自动调整列宽
        for idx, column in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for cell in column:
                if cell.value:
                    # 考虑中文字符宽度，简单乘以1.5
                    val_str = str(cell.value)
                    length = len(val_str.encode('gbk')) if isinstance(val_str, str) else len(val_str)
                    max_length = max(max_length, length)
            
            adjusted_width = min(max(max_length + 2, 12), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @contextmanager
    def _excel_writer(self, stream: io.BytesIO):
        writer = pd.ExcelWriter(stream, engine="openpyxl")
        try:
            yield writer
        finally:
            writer.close()

    async def export(self, data_list: list[BaseModel] | None = None) -> StreamingResponse:
        filename = self._generate_filename()
        stream = io.BytesIO()

        try:
            df = self._prepare_dataframe(data_list)

            with self._excel_writer(stream) as writer:
                df.to_excel(writer, index=False, sheet_name=self.sheet_name)
                worksheet = writer.sheets[self.sheet_name]
                self._apply_styles(worksheet)

            stream.seek(0)
            
            # 对文件名进行编码，防止中文文件名乱码
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)

            return StreamingResponse(
                stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                },
            )
        except Exception as e:
            logger.error(f"Failed to export Excel: {e}")
            raise e