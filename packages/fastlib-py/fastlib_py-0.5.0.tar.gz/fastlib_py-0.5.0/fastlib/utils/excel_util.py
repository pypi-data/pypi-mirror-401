# SPDX-License-Identifier: MIT

"""Excel export utilities for Pydantic models."""

import asyncio
import io
import unicodedata
import urllib.parse
from datetime import datetime, timezone
from functools import partial

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from starlette.responses import StreamingResponse

from fastlib.logging.handlers import logger


class ExcelExporter:
    DEFAULT_FONT = Font(name="Microsoft YaHei", size=11)
    HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
    HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")

    def __init__(
        self,
        schema: type[BaseModel],
        file_name: str,
        sheet_name: str | None = None,
        include_timestamp: bool = True,
    ):
        self.schema = schema
        self.file_name = file_name
        self.sheet_name = sheet_name or file_name[:31]
        self.include_timestamp = include_timestamp
        self.field_names = list(schema.model_fields.keys())
        self.header_mapping = {
            name: field.title or name
            for name, field in schema.model_fields.items()
        }

    @staticmethod
    def _get_display_width(text: str) -> int:
        """计算字符串显示宽度（全角字符算2）"""
        return sum(2 if unicodedata.east_asian_width(c) in "FW" else 1 for c in text)

    def _generate_filename(self) -> str:
        if self.include_timestamp:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
            return f"{self.file_name}_{timestamp}.xlsx"
        return f"{self.file_name}.xlsx"

    def _prepare_dataframe(self, data_list: list[BaseModel] | None) -> pd.DataFrame:
        if not data_list:
            return pd.DataFrame(columns=list(self.header_mapping.values()))

        data_dicts = [item.model_dump() for item in data_list]
        df = pd.DataFrame(data_dicts, columns=self.field_names)
        return df.rename(columns=self.header_mapping)

    def _apply_styles(self, worksheet: Worksheet) -> None:
        # 表头样式
        for cell in worksheet[1]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # 数据行样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.DEFAULT_FONT
                cell.alignment = self.CELL_ALIGNMENT

        # 自动列宽
        for idx, column in enumerate(worksheet.columns, 1):
            max_width = max(
                (self._get_display_width(str(cell.value)) for cell in column if cell.value),
                default=0,
            )
            adjusted_width = min(max(max_width + 2, 12), 60)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

    def _export_sync(self, data_list: list[BaseModel] | None) -> StreamingResponse:
        """同步导出逻辑"""
        filename = self._generate_filename()
        output = io.BytesIO()

        try:
            df = self._prepare_dataframe(data_list)

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=self.sheet_name)
                self._apply_styles(writer.sheets[self.sheet_name])

            content = output.getvalue()
            encoded_filename = urllib.parse.quote(filename)

            return StreamingResponse(
                io.BytesIO(content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}",
                    "Content-Length": str(len(content)),
                    "Access-Control-Expose-Headers": "Content-Disposition, Content-Length",
                },
            )
        except Exception:
            logger.exception("Excel export failed")
            raise
        finally:
            output.close()

    async def export(self, data_list: list[BaseModel] | None = None) -> StreamingResponse:
        """异步导出，避免阻塞事件循环"""
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, partial(self._export_sync, data_list))