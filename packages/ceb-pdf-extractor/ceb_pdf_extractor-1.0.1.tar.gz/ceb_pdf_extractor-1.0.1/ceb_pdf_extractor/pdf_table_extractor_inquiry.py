import camelot
import pdfplumber
import re
import os
from openpyxl import Workbook
import pandas as pd
from collections import defaultdict
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class PDFTableExtractor_Inquiry:
    def __init__(self, pdf_path):
        self.pdf_path = os.path.abspath(pdf_path)
        self.pdf_filename = os.path.splitext(os.path.basename(self.pdf_path))[0]
        self.output_dir = os.path.join(os.path.dirname(self.pdf_path), f"光大银行pdf转excel({self.pdf_filename})")

        self.customer_data = defaultdict(list)

        self.extracted_files = []
        self.skipped_pages = []
        self.total_pages = 0

        self._create_output_dir()

    def _create_output_dir(self):
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            print(f"✅ 创建输出目录: {self.output_dir}")

    def _extract_customer_info(self, page_text):
        info = {}

        name_match = re.search(r'账户名称[：:]\s*([^\s\n]+)', page_text)
        if not name_match:
            name_match = re.search(r'客户姓名[：:]\s*([^\s\n]+)', page_text)
        if not name_match:
            name_match = re.search(r'户名[：:]\s*([^\s\n]+)', page_text)

        account_match = re.search(r'客户账号[：:]\s*([\d]+)', page_text)
        if not account_match:
            account_match = re.search(r'账号[：:]\s*([\d]+)', page_text)

        if name_match and account_match:
            info['账户名称'] = name_match.group(1).strip()
            info['客户账号'] = account_match.group(1).strip()

            type_match = re.search(r'账户类型[：:]\s*([^\s\n]+)', page_text)
            if type_match:
                info['账户类型'] = type_match.group(1).strip()
            else:
                info['账户类型'] = '未知'

            sys_account_match = re.search(r'系统账号[：:]\s*([\d]+)', page_text)
            if sys_account_match:
                info['系统账号'] = sys_account_match.group(1).strip()
            else:
                info['系统账号'] = info['客户账号']

            date_match = re.search(r'查询起止日期[：:]\s*([^\s]+)', page_text)
            if date_match:
                date_text = date_match.group(1).strip()
                date_only = re.search(r'(\d{8}-\d{8})', date_text)
                if date_only:
                    info['查询起止日期'] = date_only.group(1)
                else:
                    info['查询起止日期'] = date_text
            else:
                info['查询起止日期'] = ''

            total_match = re.search(r'交易总笔数[：:]\s*(\d+)', page_text)
            if total_match:
                info['交易总笔数'] = total_match.group(1).strip()
            else:
                info['交易总笔数'] = '0'

            out_count_match = re.search(r'转出笔数[：:]\s*(\d+)', page_text)
            if out_count_match:
                info['转出笔数'] = out_count_match.group(1).strip()
            else:
                info['转出笔数'] = '0'

            out_amount_match = re.search(r'转出金额[：:]\s*([\d,\.]+)', page_text)
            if out_amount_match:
                info['转出金额'] = out_amount_match.group(1).strip()
            else:
                info['转出金额'] = '0.00'

            in_count_match = re.search(r'存入笔数[：:]\s*(\d+)', page_text)
            if in_count_match:
                info['存入笔数'] = in_count_match.group(1).strip()
            else:
                info['存入笔数'] = '0'

            in_amount_match = re.search(r'存入金额[：:]\s*([\d,\.]+)', page_text)
            if in_amount_match:
                info['存入金额'] = in_amount_match.group(1).strip()
            else:
                info['存入金额'] = '0.00'

            return info
        return None

    def _has_no_details(self, page_text):
        return "无明细" in page_text or "不存在交易明细" in page_text

    def _extract_table_fixed_area(self, page_num):
        try:
            table_area = ['0,380,800,30']

            tables = camelot.read_pdf(
                self.pdf_path,
                flavor="stream",
                pages=str(page_num),
                table_areas=table_area,
                row_tol=30,
                strip_text='\n'
            )

            if len(tables) > 0:
                print(f"第{page_num}页: 使用固定区域成功提取表格")
                return tables[0]
            else:
                print(f"第{page_num}页: 固定区域未提取到表格，尝试全页面提取")
                return self._extract_table_full_page(page_num)

        except Exception as e:
            print(f"第{page_num}页: 固定区域提取出错: {e}")
            return self._extract_table_full_page(page_num)

    def _extract_table_full_page(self, page_num):
        try:
            tables = camelot.read_pdf(
                self.pdf_path,
                flavor="stream",
                pages=str(page_num),
                row_tol=30,
                strip_text='\n'
            )

            if len(tables) > 0:
                print(f"第{page_num}页: 全页面提取成功")
                return tables[0]
            else:
                print(f"第{page_num}页: 未提取到表格")
                return None

        except Exception as e:
            print(f"第{page_num}页: 全页面提取出错: {e}")
            return None

    def _remove_header_row(self, df, header_keywords=None):
        if df.empty:
            return df

        if header_keywords is None:
            header_keywords = ['查询卡号', '交易类型', '借贷标志', '币种', '交易金额', '交易余额', '交易时间',
                               '交易对方名称', '交易对方账号', '交易对方账号开户行', '交易摘要']

        first_row = df.iloc[0].astype(str).str.strip().tolist()
        first_row_str = ' '.join(first_row).lower()

        is_header = any(keyword in first_row_str for keyword in header_keywords)

        if is_header:
            df_cleaned = df.iloc[1:].reset_index(drop=True)
            print(f"    检测到表头行，已移除，剩余数据行数: {len(df_cleaned)}")
            return df_cleaned
        else:
            return df

    def _get_page_text(self, page_num):
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                page = pdf.pages[page_num - 1]
                page_text = page.extract_text() or ""
                return page_text
        except Exception as e:
            print(f"第{page_num}页: 提取页面文本时出错: {e}")
            return ""

    def create_excel_file(self, customer_info, filepath):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "交易明细"

            print(f"✅ 创建新的Excel文件: {os.path.basename(filepath)}")
            return wb, ws, filepath

        except Exception as e:
            print(f"❌ 创建Excel文件时出错: {e}")
            return None, None, None

    def _apply_excel_format(self, worksheet, data_rows):
        if not data_rows:
            return

        max_row = len(data_rows)
        max_col = max(len(row) for row in data_rows) if data_rows else 0

        if max_row == 0 or max_col == 0:
            return

        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )

        for col in range(1, max_col + 1):
            max_length = 0

            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    content = str(cell_value)
                    chinese_count = sum(1 for char in content if '\u4e00' <= char <= '\u9fff')
                    length = len(content) + chinese_count
                    max_length = max(max_length, length)

            if max_length > 0:
                column_letter = get_column_letter(col)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(
                    vertical='center',
                    horizontal='center',
                    wrap_text=True
                )
                cell.number_format = '@'
                cell.border = no_border

        if max_row > 0:
            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)

    def _clean_table_data(self, header, data_rows):
        if not header or not data_rows:
            return header, data_rows

        transaction_amount_col = None
        for i, col_name in enumerate(header):
            if col_name and isinstance(col_name, str):
                if "交易金额" in col_name or "金额" in col_name:
                    transaction_amount_col = i
                    break

        if transaction_amount_col is None:
            print("    警告: 未找到'交易金额'列，跳过数据合并清洗")
            cleaned_rows = []
            for row in data_rows:
                is_page_info = False
                for cell in row:
                    if isinstance(cell, str) and re.search(r'第\s*\d+\s*页\s*共\s*\d+\s*页', cell):
                        is_page_info = True
                        break
                if not is_page_info:
                    cleaned_rows.append(row)
            return header, cleaned_rows

        cleaned_rows = []
        i = 0
        while i < len(data_rows):
            current_row = data_rows[i]

            is_page_info = False
            for cell in current_row:
                if isinstance(cell, str) and re.search(r'第\s*\d+\s*页\s*共\s*\d+\s*页', cell):
                    is_page_info = True
                    break

            if is_page_info:
                i += 1
                continue

            if i < len(data_rows):
                transaction_amount = current_row[transaction_amount_col] if transaction_amount_col < len(
                    current_row) else ""

                if not transaction_amount or (isinstance(transaction_amount, str) and transaction_amount.strip() == ""):
                    if cleaned_rows:
                        last_row = cleaned_rows[-1]

                        merged_row = list(last_row)

                        for col_idx in range(len(current_row)):
                            if col_idx >= len(merged_row):
                                merged_row.extend([''] * (col_idx - len(merged_row) + 1))

                            current_cell = current_row[col_idx] if col_idx < len(current_row) else ""
                            last_cell = merged_row[col_idx] if col_idx < len(merged_row) else ""

                            if col_idx == transaction_amount_col:
                                continue

                            if current_cell and (not last_cell or last_cell.strip() == ""):
                                merged_row[col_idx] = str(current_cell)
                            elif current_cell and last_cell:
                                if str(current_cell) not in str(last_cell):
                                    merged_row[col_idx] = f"{last_cell}\n{current_cell}"

                        cleaned_rows[-1] = merged_row
                    else:
                        cleaned_rows.append(current_row)
                else:
                    cleaned_rows.append(current_row)

            i += 1

        return header, cleaned_rows

    def scan_pages(self):
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                self.total_pages = len(pdf.pages)
            print(f"📄 PDF总页数: {self.total_pages}")
        except Exception as e:
            print(f"❌ 获取PDF页数失败: {e}")
            return self.customer_data

        current_active_customer = None

        for page_num in range(1, self.total_pages + 1):
            print(f"\n处理第 {page_num}/{self.total_pages} 页...")

            page_text = self._get_page_text(page_num)

            if self._has_no_details(page_text):
                print(f"  第{page_num}页: 包含'无明细'，跳过")
                self.skipped_pages.append(page_num)
                continue

            customer_info = self._extract_customer_info(page_text)
            if customer_info and '账户名称' in customer_info and '客户账号' in customer_info:
                print(f"  第{page_num}页: 找到客户信息")
                print(f"    账户名称: {customer_info.get('账户名称', '未知')}")
                print(f"    客户账号: {customer_info.get('客户账号', '未知')}")

                table = self._extract_table_fixed_area(page_num)

                if table is not None and not table.df.empty:
                    customer_key = f"{customer_info.get('账户名称', '未知')}_{customer_info.get('客户账号', '未知')}"
                    current_active_customer = customer_key

                    self.customer_data[customer_key].append({
                        "page_num": page_num,
                        "customer_info": customer_info,
                        "table_df": table.df,
                        "is_header_page": True
                    })
                    print(f"    成功提取表头页表格数据（行数: {len(table.df)}）")
                else:
                    print(f"    未提取到表格数据")
                    self.skipped_pages.append(page_num)
            else:
                print(f"  第{page_num}页: 未找到客户信息，尝试作为续页提取")

                if not current_active_customer:
                    print(f"    无活跃客户，跳过")
                    self.skipped_pages.append(page_num)
                    continue

                table = self._extract_table_full_page(page_num)

                if table is not None and not table.df.empty:
                    clean_df = self._remove_header_row(table.df)

                    if clean_df.empty:
                        print(f"    续页移除表头后无数据，跳过")
                        self.skipped_pages.append(page_num)
                        continue

                    self.customer_data[current_active_customer].append({
                        "page_num": page_num,
                        "customer_info": None,
                        "table_df": clean_df,
                        "is_header_page": False
                    })
                    print(f"    成功提取续页表格数据（移除表头后行数: {len(clean_df)}）")
                else:
                    print(f"    未提取到表格数据")
                    self.skipped_pages.append(page_num)

        return self.customer_data

    def process_customer_data(self):
        if not self.customer_data:
            print("\n⚠️ 未找到有效的客户数据")
            return []

        excel_files_info = []

        for customer_key, page_data_list in self.customer_data.items():
            print(f"\n📝 处理账户: {customer_key}")
            print(f"   包含 {len(page_data_list)} 页数据")

            base_info = next((item for item in page_data_list if item["customer_info"]), None)
            if not base_info:
                print(f"   ❌ 无有效客户信息，跳过")
                continue

            customer_info = base_info["customer_info"]
            account_name = customer_info.get("账户名称", "未知账户")
            account_number = customer_info.get("客户账号", "未知账号")

            safe_name = re.sub(r'[\\/*?:"<>|]', "_", account_name)
            safe_account = re.sub(r'[\\/*?:"<>|]', "_", account_number)
            filename = f"{safe_name}_{safe_account}（分）.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            counter = 1
            original_filename = filename
            while os.path.exists(filepath):
                name_without_ext = os.path.splitext(original_filename)[0]
                ext = os.path.splitext(original_filename)[1]
                filename = f"{name_without_ext}_{counter}{ext}"
                filepath = os.path.join(self.output_dir, filename)
                counter += 1

            wb, ws, filepath = self.create_excel_file(customer_info, filepath)
            if not wb:
                print(f"   ❌ 创建Excel文件失败")
                continue

            total_rows = 0
            all_rows = []

            supplementary_columns = [
                '账户名称', '客户账号', '账户类型', '系统账号', '查询起止日期',
                '交易总笔数', '转出笔数', '转出金额', '存入笔数', '存入金额'
            ]

            supplementary_values = []
            for col in supplementary_columns:
                value = customer_info.get(col, '')
                if value is None:
                    value = ''
                else:
                    value = str(value)
                supplementary_values.append(value)

            for page_data in page_data_list:
                page_num = page_data["page_num"]
                is_header_page = page_data["is_header_page"]
                table_df = page_data["table_df"]

                print(f"\n   处理第{page_num}页数据:")

                data_rows = []
                for _, row in table_df.iterrows():
                    row_list = [str(val) if val is not None else '' for val in row.tolist()]
                    data_rows.append(row_list)

                if is_header_page:
                    all_rows.extend(data_rows)
                    total_rows += len(data_rows)
                    print(f"    表头页写入: {len(data_rows)}行")
                else:
                    all_rows.extend(data_rows)
                    total_rows += len(data_rows)
                    print(f"    续页写入: {len(data_rows)}行")

            if all_rows and len(all_rows) > 1:
                print(f"\n   开始清洗数据...")
                print(f"    清洗前行数: {len(all_rows)}")

                header_row = all_rows[0]
                data_rows = all_rows[1:]

                cleaned_header, cleaned_data_rows = self._clean_table_data(header_row, data_rows)

                all_rows = [cleaned_header] + cleaned_data_rows
                print(f"    清洗后行数: {len(all_rows)}")

            if all_rows and len(all_rows) > 1:
                original_header = all_rows[0]
                new_header = supplementary_columns + original_header
                ws.append(new_header)

                for i in range(1, len(all_rows)):
                    new_row = supplementary_values + all_rows[i]
                    ws.append(new_row)

                formatted_rows = [new_header] + [supplementary_values + row for row in all_rows[1:]]
                self._apply_excel_format(ws, formatted_rows)

                wb.save(filepath)

                excel_files_info.append({
                    "filename": filename,
                    "filepath": filepath,
                    "account_name": account_name,
                    "account_number": account_number,
                    "total_pages": len(page_data_list),
                    "total_rows": len(all_rows) - 1,
                    "customer_info": customer_info
                })
                self.extracted_files.append(filepath)

                print(f"✅ 保存Excel文件: {filename}")
                print(f"   账户: {account_name}, 账号: {account_number}")

        return excel_files_info

    def _generate_report(self, excel_files_info):
        if not excel_files_info:
            print("\n⚠️ 没有生成任何Excel文件")
            return

        original_pdf_name = os.path.basename(self.pdf_path)
        report_filename = f"清洗报告（{original_pdf_name}）.txt"
        report_path = os.path.join(self.output_dir, report_filename)

        unique_customers = set()
        unique_accounts = set()
        total_data_rows = 0

        for file_info in excel_files_info:
            account_name = file_info.get('account_name', '')
            account_number = file_info.get('account_number', '')
            data_rows = file_info.get('total_rows', 0)

            if account_name:
                unique_customers.add(account_name)
            if account_number:
                unique_accounts.add(account_number)

            total_data_rows += data_rows

        ten_thousands = total_data_rows / 10000

        with open(report_path, 'w', encoding='utf-8') as f:
            report_line = f"共{len(unique_customers)}个客户，{len(unique_accounts)}个账户，{ten_thousands:.2f}万条数据"
            f.write(report_line)

        print(f"\n📊 清洗报告已生成: {report_path}")
        print(f"📋 报告内容: {report_line}")

    def process(self):
        try:
            print(f"🚀 开始处理PDF文件: {os.path.basename(self.pdf_path)}")

            if not os.path.exists(self.pdf_path):
                print(f"❌ 文件不存在: {self.pdf_path}")
                return {"error": "文件不存在", "状态": "失败"}

            customer_data = self.scan_pages()

            if not customer_data:
                print("\n⚠️ 未找到有效的客户数据")
                return {"提取的文件数": 0, "跳过的页数": len(self.skipped_pages), "状态": "失败"}

            print(f"\n✅ 找到 {len(customer_data)} 个客户账户")

            excel_files_info = self.process_customer_data()

            self._generate_report(excel_files_info)

            print("\n" + "=" * 70)
            print("✅ 处理完成！")
            print(f"📁 生成Excel文件数: {len(excel_files_info)}")
            print(f"⏭️  跳过的页面数: {len(self.skipped_pages)}")
            if self.skipped_pages:
                print(f"   跳过的页码: {self.skipped_pages}")
            print(f"📂 输出目录: {self.output_dir}")
            print("=" * 70)

            return excel_files_info

        except Exception as e:
            print(f"\n❌ 处理PDF文件时出错: {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e), "状态": "失败"}