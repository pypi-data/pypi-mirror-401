import fitz
import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

class PDFTableExtractor:
    def __init__(self, pdf_path, target_size, size_tolerance,
                 row_tolerance, col_tolerance,
                 customer_info_fields, supplementary_columns, report_title="光大银行PDF对账单提取报告"):
        self.pdf_path = os.path.abspath(pdf_path)
        self.pdf_filename = os.path.splitext(os.path.basename(self.pdf_path))[0]
        self.output_dir = os.path.join(os.path.dirname(self.pdf_path), f"光大银行pdf转excel({self.pdf_filename})")
        self.doc = fitz.open(self.pdf_path)
        self.target_size = target_size
        self.size_tolerance = size_tolerance
        self.row_tolerance = row_tolerance
        self.col_tolerance = col_tolerance

        self.customer_info_fields = customer_info_fields
        self.supplementary_columns = supplementary_columns
        self.report_title = report_title

        self.customer_data = defaultdict(list)

        self.extracted_files = []
        self.skipped_pages = []

        self._create_output_dir()

    def _create_output_dir(self):
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            print(f"✅ 创建输出目录: {self.output_dir}")

    def _extract_customer_info(self, page):
        full_text = page.get_text("text")

        info = {}

        for field_name, patterns in self.customer_info_fields.items():
            value = ''
            for pattern in patterns:
                match = re.search(pattern, full_text)
                if match:
                    if match.groups():
                        value = match.group(1).strip()
                        if value:
                            break
                        else:
                            value= ' '
                            break
                    else:
                        value = match.group(0).strip()
                        break
                if not value:
                    value= ' '
            info[field_name] = value

        fields_list = list(self.customer_info_fields.keys())
        if len(fields_list) >= 2:
            key_field1 = fields_list[0]
            key_field2 = fields_list[1]

            if not info.get(key_field1) or not info.get(key_field2):
                return None

        return info

    def _has_no_transaction_details(self, page):
        full_text = page.get_text("text")
        return "不存在交易明细" in full_text \
               or "无符合条件的开户记录" in full_text \
               or "交易日期" not in full_text \
               or "无明细" in full_text

    def _extract_table_cells(self, page):
        text_dict = page.get_text("dict")
        cells = []
        for block in text_dict.get("blocks", []):
            if block["type"] == 0:
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        if abs(span["size"] - self.target_size) <= self.size_tolerance:
                            text = span["text"].strip()
                            if text:
                                cells.append({
                                    "text": text,
                                    "bbox": span["bbox"],
                                    "x_center": (span["bbox"][0] + span["bbox"][2]) / 2,
                                    "y_center": (span["bbox"][1] + span["bbox"][3]) / 2,
                                    "y_start": span["bbox"][1],
                                })
        return cells

    def _cells_to_matrix(self, cells):
        if not cells:
            return []

        rows = self._group_cells_into_rows(cells)

        if not rows:
            return []

        column_positions = self._detect_columns_from_header(rows)

        if not column_positions:
            return []

        matrix = self._assign_cells_with_header_constraint(rows, column_positions)

        return matrix

    def _group_cells_into_rows(self, cells):
        cells.sort(key=lambda c: c["y_center"])

        rows = []
        current_row = [cells[0]]
        current_y = cells[0]["y_center"]

        for cell in cells[1:]:
            if abs(cell["y_center"] - current_y) <= self.row_tolerance:
                current_row.append(cell)
            else:
                current_row.sort(key=lambda c: c["x_center"])
                rows.append(current_row)
                current_row = [cell]
                current_y = cell["y_center"]

        if current_row:
            current_row.sort(key=lambda c: c["x_center"])
            rows.append(current_row)

        return rows

    def _detect_columns_from_header(self, rows):
        if not rows:
            return []

        header_row = rows[0]
        header_x_positions = [cell["x_center"] for cell in header_row]
        header_x_positions.sort()

        column_positions = []
        for x in header_x_positions:
            if not column_positions:
                column_positions.append(x)
            else:
                min_distance = min(abs(x - pos) for pos in column_positions)
                if min_distance <= self.col_tolerance:
                    closest_idx = min(range(len(column_positions)),
                                      key=lambda i: abs(x - column_positions[i]))
                    column_positions[closest_idx] = (column_positions[closest_idx] + x) / 2
                else:
                    column_positions.append(x)

        column_positions.sort()
        return column_positions

    def _assign_cells_with_header_constraint(self, rows, column_positions):
        max_cols = len(column_positions)
        matrix = []

        for row_idx, row_cells in enumerate(rows):
            matrix_row = [""] * max_cols
            sorted_cells = sorted(row_cells, key=lambda c: c["x_center"])

            for cell in sorted_cells:
                closest_col_idx = self._find_closest_column(cell["x_center"], column_positions)

                if abs(cell["x_center"] - column_positions[closest_col_idx]) > self.col_tolerance * 2:
                    continue

                self._place_cell_in_column(cell, closest_col_idx, matrix_row, max_cols)

            if row_idx == 0:
                matrix_row = self._ensure_header_completeness(row_cells, matrix_row, column_positions)

            matrix.append(matrix_row)

        return matrix

    def _find_closest_column(self, x_center, column_positions):
        distances = [abs(x_center - pos) for pos in column_positions]
        return distances.index(min(distances))

    def _place_cell_in_column(self, cell, target_col_idx, matrix_row, max_cols):
        if not matrix_row[target_col_idx]:
            matrix_row[target_col_idx] = cell["text"]
            return

        for i in range(target_col_idx + 1, max_cols):
            if not matrix_row[i]:
                matrix_row[i] = cell["text"]
                return

        for i in range(target_col_idx - 1, -1, -1):
            if not matrix_row[i]:
                matrix_row[i] = cell["text"]
                return

        matrix_row[target_col_idx] += " " + cell["text"]

    def _ensure_header_completeness(self, header_cells, matrix_row, column_positions):
        assigned_texts = [text for text in matrix_row if text]
        all_header_texts = [cell["text"] for cell in header_cells]
        unassigned_texts = [text for text in all_header_texts if text not in assigned_texts]

        if not unassigned_texts:
            return matrix_row

        for header_text in unassigned_texts:
            header_cell = next((c for c in header_cells if c["text"] == header_text), None)
            if not header_cell:
                continue

            closest_col_idx = self._find_closest_column(header_cell["x_center"], column_positions)

            if not matrix_row[closest_col_idx]:
                matrix_row[closest_col_idx] = header_text
            else:
                empty_cols = [i for i, text in enumerate(matrix_row) if not text]
                if empty_cols:
                    distances = [abs(closest_col_idx - i) for i in empty_cols]
                    nearest_empty_idx = empty_cols[distances.index(min(distances))]
                    matrix_row[nearest_empty_idx] = header_text

        return matrix_row

    def _merge_single_cell_rows_in_matrix(self, matrix):
        if not matrix or len(matrix) < 2:
            return matrix

        rows_to_delete = []

        for row_idx in range(len(matrix) - 1, 0, -1):
            row = matrix[row_idx]
            non_empty_cells = [cell for cell in row if cell and str(cell).strip()]

            if len(non_empty_cells) == 1:
                non_empty_col = None
                non_empty_value = None
                for col_idx, cell in enumerate(row):
                    if cell and str(cell).strip():
                        non_empty_col = col_idx
                        non_empty_value = cell
                        break

                if non_empty_col is not None:
                    prev_row = matrix[row_idx - 1]
                    prev_value = prev_row[non_empty_col]

                    if prev_value and str(prev_value).strip():
                        if isinstance(prev_value, str) and isinstance(non_empty_value, str):
                            prev_row[non_empty_col] = f"{prev_value};{non_empty_value}"
                        else:
                            prev_row[non_empty_col] = f"{str(prev_value)};{str(non_empty_value)}"
                    else:
                        prev_row[non_empty_col] = non_empty_value

                    rows_to_delete.append(row_idx)

        for row_idx in sorted(rows_to_delete, reverse=True):
            matrix.pop(row_idx)

        return matrix

    def scan_pages(self):
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]

            if self._has_no_transaction_details(page):
                print(f"  第{page_num + 1}页: 不存在交易明细，跳过")
                self.skipped_pages.append(page_num + 1)
                continue

            customer_info = self._extract_customer_info(page)
            if not customer_info:
                print(f"  第{page_num + 1}页: 未找到完整客户信息，跳过")
                continue

            fields_list = list(self.customer_info_fields.keys())
            if len(fields_list) >= 2:
                customer_name = customer_info.get(fields_list[0], '未知')
                customer_account = customer_info.get(fields_list[1], '未知')
            else:
                customer_name = '未知'
                customer_account = '未知'

            print(f"  第{page_num + 1}页: {fields_list[0]} {customer_name} ({fields_list[1]} {customer_account})")

            cells = self._extract_table_cells(page)

            if not cells:
                print(f"    未找到表格数据")
                continue

            matrix = self._cells_to_matrix(cells)

            if matrix:
                customer_key = f"{customer_name}_{customer_account}"
                self.customer_data[customer_key].append({
                    "page_num": page_num + 1,
                    "customer_info": customer_info,
                    "matrix": matrix,
                    "rows": len(matrix),
                    "cols": len(matrix[0]) if matrix else 0,
                    "is_header_page": True
                })

                print(f"    提取表格: {len(matrix)}行 × {len(matrix[0])}列")

        return self.customer_data

    def create_excel_file(self, customer_info, filepath):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "对账单数据"

            print(f"✅ 创建新的Excel文件: {os.path.basename(filepath)}")
            return wb, ws, filepath

        except Exception as e:
            print(f"❌ 创建Excel文件时出错: {e}")
            return None, None, None

    def _apply_excel_format(self, worksheet, data_rows):
        if not data_rows:
            return

        max_row = len(data_rows)
        max_col = max(len(row) for row in data_rows) if data_rows else 0

        if max_row == 0 or max_col == 0:
            return

        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )

        for col in range(1, max_col + 1):
            max_length = 0

            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    content = str(cell_value)
                    chinese_count = sum(1 for char in content if '\u4e00' <= char <= '\u9fff')
                    length = len(content) + chinese_count
                    max_length = max(max_length, length)

            if max_length > 0:
                column_letter = get_column_letter(col)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(
                    vertical='center',
                    horizontal='center',
                    wrap_text=True
                )
                cell.number_format = '@'
                cell.border = no_border

        if max_row > 0:
            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)

    def process_customer_data(self):
        if not self.customer_data:
            print("\n⚠️ 未找到有效的客户数据")
            return []

        excel_files_info = []

        for customer_key, page_data_list in self.customer_data.items():
            print(f"\n📝 处理账户: {customer_key}")
            print(f"   包含 {len(page_data_list)} 页数据")

            base_info = next((item for item in page_data_list if item["customer_info"]), None)
            if not base_info:
                print(f"   ❌ 无有效客户信息，跳过")
                continue

            customer_info = base_info["customer_info"]

            fields_list = list(self.customer_info_fields.keys())
            if len(fields_list) >= 2:
                customer_name = customer_info.get(fields_list[0], "未知")
                customer_account = customer_info.get(fields_list[1], "未知")
            else:
                customer_name = "未知"
                customer_account = "未知"

            safe_name = re.sub(r'[\\/*?:"<>|]', "_", customer_name)
            safe_account = re.sub(r'[\\/*?:"<>|]', "_", customer_account)
            filename = f"{safe_name}_{safe_account}(分).xlsx"
            filepath = os.path.join(self.output_dir, filename)

            counter = 1
            original_filename = filename
            while os.path.exists(filepath):
                name_without_ext = os.path.splitext(original_filename)[0]
                ext = os.path.splitext(original_filename)[1]
                filename = f"{name_without_ext}_{counter}{ext}"
                filepath = os.path.join(self.output_dir, filename)
                counter += 1

            wb, ws, filepath = self.create_excel_file(customer_info, filepath)
            if not wb:
                print(f"   ❌ 创建Excel文件失败")
                continue

            all_rows = []

            for i, page_data in enumerate(page_data_list):
                page_num = page_data["page_num"]
                matrix = page_data["matrix"]

                print(f"\n   处理第{page_num}页数据:")

                if not matrix or len(matrix) == 0:
                    print(f"    无表格数据，跳过")
                    continue

                if len(matrix) > 1:
                    data_rows = matrix[1:]
                    merged_data_rows = self._merge_single_cell_rows_in_matrix(data_rows)

                    if i == 0:
                        merged_matrix = [matrix[0]] + merged_data_rows
                        all_rows.extend(merged_matrix)
                        print(f"    添加: 表头 + {len(merged_data_rows)}行数据")
                    else:
                        all_rows.extend(merged_data_rows)
                        print(f"    添加: {len(merged_data_rows)}行数据（跳过表头）")
                else:
                    if i == 0:
                        all_rows.extend(matrix)
                        print(f"    添加: {len(matrix)}行（可能只是表头）")
                    else:
                        print(f"    跳过: 只有1行（可能是表头）")

            print(f"\n   总计合并行数: {len(all_rows)} (表头1行 + {len(all_rows) - 1 if all_rows else 0}行数据)")

            supplementary_values = []
            for col in self.supplementary_columns:
                value = customer_info.get(col, '')
                if value is None:
                    value = ''
                else:
                    value = str(value)
                supplementary_values.append(value)

            if all_rows:
                original_header = all_rows[0]
                new_header = self.supplementary_columns + original_header
                ws.append(new_header)

                for i in range(1, len(all_rows)):
                    new_row = supplementary_values + all_rows[i]
                    ws.append(new_row)

                formatted_rows = [new_header] + [supplementary_values + row for row in all_rows[1:]]
                self._apply_excel_format(ws, formatted_rows)

                wb.save(filepath)

                excel_files_info.append({
                    "filename": filename,
                    "filepath": filepath,
                    "customer_name": customer_name,
                    "customer_account": customer_account,
                    "total_pages": len(page_data_list),
                    "total_rows": len(all_rows) - 1 if len(all_rows) > 0 else 0,
                    "customer_info": customer_info
                })
                self.extracted_files.append(filepath)

                print(f"✅ 保存Excel文件: {filename}")
                print(f"   {fields_list[0]}: {customer_name}, {fields_list[1]}: {customer_account}")

        return excel_files_info

    def _generate_report(self, excel_files_info):
        if not excel_files_info:
            print("\n⚠️ 没有生成任何Excel文件")
            return

        report_path = os.path.join(self.output_dir, "处理报告.txt")

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write(f"{self.report_title}\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"源文件: {os.path.basename(self.pdf_path)}\n")
            f.write(f"总页数: {len(self.doc)}\n")
            f.write(f"生成文件数: {len(excel_files_info)}\n")
            f.write(f"跳过页数: {len(self.skipped_pages)}\n")
            f.write(f"输出目录: {self.output_dir}\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("文件详情:\n")
            f.write("-" * 60 + "\n\n")

            fields_list = list(self.customer_info_fields.keys())
            field1_name = fields_list[0] if fields_list else "客户名称"
            field2_name = fields_list[1] if len(fields_list) > 1 else "客户账号"

            for i, file_info in enumerate(excel_files_info, 1):
                f.write(f"{i}. {file_info['filename']}\n")
                f.write(f"   {field1_name}: {file_info.get('customer_name', '未知')}\n")
                f.write(f"   {field2_name}: {file_info.get('customer_account', '未知')}\n")
                f.write(f"   包含页数: {file_info['total_pages']}\n")
                f.write(f"   总行数: {file_info['total_rows']}\n\n")

        print(f"\n📊 处理报告已生成: {report_path}")

    def _generate_clean_report(self, excel_files_info):
        if not excel_files_info:
            print("\n⚠️ 没有生成任何Excel文件")
            return

        customer_names = set()
        customer_accounts = set()
        total_data_rows = 0

        for file_info in excel_files_info:
            customer_names.add(file_info['customer_name'])
            customer_accounts.add(file_info['customer_account'])
            total_data_rows += file_info['total_rows']

        total_data_wan = round(total_data_rows / 10000, 3)

        report_content = f"共{len(customer_names)}个客户，{len(customer_accounts)}个账户，{total_data_wan}万条数据"

        report_filename = f"清洗报告（{self.pdf_filename}）.txt"
        report_path = os.path.join(self.output_dir, report_filename)

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_content)

        print(f"\n📊 清洗报告已生成: {report_path}")
        print(f"📋 报告内容: {report_content}")

    def process(self):
        try:
            print(f"🚀 开始处理PDF文件: {os.path.basename(self.pdf_path)}")

            customer_data = self.scan_pages()

            if not customer_data:
                print("\n⚠️ 未找到有效的客户数据")
                return {"提取的文件数": 0, "跳过的页数": len(self.skipped_pages), "状态": "失败"}

            print(f"\n✅ 找到 {len(customer_data)} 个客户账户")

            excel_files_info = self.process_customer_data()

            self._generate_clean_report(excel_files_info)

            self.doc.close()

            print("\n" + "=" * 70)
            print("✅ 处理完成！")
            print(f"📁 生成Excel文件数: {len(excel_files_info)}")
            print(f"⏭️  跳过的页面数: {len(self.skipped_pages)}")
            print(f"📂 输出目录: {self.output_dir}")
            print("=" * 70)

            return excel_files_info

        except Exception as e:
            print(f"\n❌ 处理PDF文件时出错: {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e), "状态": "失败"}