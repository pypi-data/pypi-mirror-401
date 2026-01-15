import asyncio
import sys
import tempfile
import json
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest

from silkworm.pipelines import (
    CallbackPipeline,
    CSVPipeline,
    RssPipeline,
    SQLitePipeline,
    XMLPipeline,
)
from silkworm.spiders import Spider

# SnowflakePipeline tests - skip if snowflake-connector-python not installed
from silkworm.pipelines import SNOWFLAKE_AVAILABLE, SnowflakePipeline  # type: ignore

# GoogleSheetsPipeline tests - skip if google-api-python-client not installed
from silkworm.pipelines import GOOGLE_SHEETS_AVAILABLE, GoogleSheetsPipeline  # type: ignore


async def test_xml_pipeline_creates_valid_xml():
    with tempfile.TemporaryDirectory() as tmpdir:
        xml_path = Path(tmpdir) / "test.xml"
        pipeline = XMLPipeline(xml_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Verify XML is valid
        tree = ET.parse(xml_path)
        root = tree.getroot()

        assert root.tag == "items"
        items = list(root)
        assert len(items) == 2

        # Check first item
        assert items[0].tag == "item"
        assert items[0].find("text").text == "Hello"
        assert items[0].find("author").text == "John"

        # Check second item
        assert items[1].tag == "item"
        assert items[1].find("text").text == "World"
        assert items[1].find("author").text == "Jane"


async def test_xml_pipeline_custom_elements():
    with tempfile.TemporaryDirectory() as tmpdir:
        xml_path = Path(tmpdir) / "test.xml"
        pipeline = XMLPipeline(xml_path, root_element="quotes", item_element="quote")
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Test"}, spider)
        await pipeline.close(spider)

        tree = ET.parse(xml_path)
        root = tree.getroot()

        assert root.tag == "quotes"
        assert root[0].tag == "quote"


async def test_xml_pipeline_handles_nested_dict():
    with tempfile.TemporaryDirectory() as tmpdir:
        xml_path = Path(tmpdir) / "test.xml"
        pipeline = XMLPipeline(xml_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"user": {"name": "Alice", "age": 30}, "active": True}, spider
        )
        await pipeline.close(spider)

        tree = ET.parse(xml_path)
        root = tree.getroot()
        item = root[0]

        assert item.find("user/name").text == "Alice"
        assert item.find("user/age").text == "30"
        assert item.find("active").text == "True"


async def test_xml_pipeline_handles_list():
    with tempfile.TemporaryDirectory() as tmpdir:
        xml_path = Path(tmpdir) / "test.xml"
        pipeline = XMLPipeline(xml_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"tags": ["python", "web", "scraping"]}, spider)
        await pipeline.close(spider)

        tree = ET.parse(xml_path)
        root = tree.getroot()
        item = root[0]

        tags_elem = item.find("tags")
        tag_items = tags_elem.findall("item")
        assert len(tag_items) == 3
        assert tag_items[0].text == "python"
        assert tag_items[1].text == "web"
        assert tag_items[2].text == "scraping"


async def test_xml_pipeline_not_opened_raises_error():
    pipeline = XMLPipeline("test.xml")
    spider = Spider()

    with pytest.raises(RuntimeError, match="XMLPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


async def test_rss_pipeline_respects_max_items():
    with tempfile.TemporaryDirectory() as tmpdir:
        rss_path = Path(tmpdir) / "feed.xml"
        pipeline = RssPipeline(
            rss_path,
            channel_title="Test Feed",
            channel_link="https://example.com",
            channel_description="Test description",
            max_items=2,
        )
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"title": "First", "link": "https://example.com/1", "description": "One"},
            spider,
        )
        await pipeline.process_item(
            {"title": "Second", "link": "https://example.com/2", "description": "Two"},
            spider,
        )
        await pipeline.process_item(
            {"title": "Third", "link": "https://example.com/3", "description": "Three"},
            spider,
        )
        await pipeline.close(spider)

        tree = ET.parse(rss_path)
        root = tree.getroot()
        channel = root.find("channel")
        items = channel.findall("item")
        assert len(items) == 2
        titles = [item.find("title").text for item in items]
        assert titles == ["Second", "Third"]


async def test_rss_pipeline_skips_items_missing_required_fields():
    with tempfile.TemporaryDirectory() as tmpdir:
        rss_path = Path(tmpdir) / "feed.xml"
        pipeline = RssPipeline(
            rss_path,
            channel_title="Test Feed",
            channel_link="https://example.com",
            channel_description="Test description",
        )
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"title": "Ok", "link": "https://example.com/1", "description": "One"},
            spider,
        )
        await pipeline.process_item(
            {"title": "Missing Description", "link": "https://example.com/2"},
            spider,
        )
        await pipeline.close(spider)

        tree = ET.parse(rss_path)
        root = tree.getroot()
        channel = root.find("channel")
        items = channel.findall("item")
        assert len(items) == 1
        assert items[0].find("title").text == "Ok"


# CallbackPipeline tests
async def test_callback_pipeline_with_sync_callback():
    processed_items = []

    def process_item(item, spider):
        processed_items.append(item)
        return item

    pipeline = CallbackPipeline(callback=process_item)
    spider = Spider()

    await pipeline.open(spider)
    result = await pipeline.process_item({"text": "Hello", "author": "Alice"}, spider)
    await pipeline.close(spider)

    assert len(processed_items) == 1
    assert processed_items[0] == {"text": "Hello", "author": "Alice"}
    assert result == {"text": "Hello", "author": "Alice"}


async def test_callback_pipeline_with_async_callback():
    processed_items = []

    async def process_item(item, spider):
        await asyncio.sleep(0.01)  # Simulate async operation
        processed_items.append(item)
        return item

    pipeline = CallbackPipeline(callback=process_item)
    spider = Spider()

    await pipeline.open(spider)
    result = await pipeline.process_item({"text": "World", "author": "Bob"}, spider)
    await pipeline.close(spider)

    assert len(processed_items) == 1
    assert processed_items[0] == {"text": "World", "author": "Bob"}
    assert result == {"text": "World", "author": "Bob"}


async def test_callback_pipeline_callback_can_modify_item():
    def add_timestamp(item, spider):
        item["processed"] = True
        item["spider_name"] = spider.name
        return item

    pipeline = CallbackPipeline(callback=add_timestamp)
    spider = Spider()
    spider.name = "test_spider"

    await pipeline.open(spider)
    result = await pipeline.process_item({"text": "Test"}, spider)
    await pipeline.close(spider)

    assert result == {"text": "Test", "processed": True, "spider_name": "test_spider"}


async def test_callback_pipeline_callback_returning_none():
    def process_item(item, spider):
        # Callback that doesn't return anything
        print(item)
        return None

    pipeline = CallbackPipeline(callback=process_item)
    spider = Spider()

    await pipeline.open(spider)
    result = await pipeline.process_item({"text": "Test"}, spider)
    await pipeline.close(spider)

    # When callback returns None, original item should be returned
    assert result == {"text": "Test"}


async def test_callback_pipeline_multiple_items():
    processed_items = []

    def process_item(item, spider):
        processed_items.append(item)
        return item

    pipeline = CallbackPipeline(callback=process_item)
    spider = Spider()

    await pipeline.open(spider)
    await pipeline.process_item({"id": 1}, spider)
    await pipeline.process_item({"id": 2}, spider)
    await pipeline.process_item({"id": 3}, spider)
    await pipeline.close(spider)

    assert len(processed_items) == 3
    assert processed_items[0] == {"id": 1}
    assert processed_items[1] == {"id": 2}
    assert processed_items[2] == {"id": 3}


def test_callback_pipeline_requires_callable():
    with pytest.raises(TypeError, match="callback must be callable"):
        CallbackPipeline(callback="not_callable")

    with pytest.raises(TypeError, match="callback must be callable"):
        CallbackPipeline(callback=123)

    with pytest.raises(TypeError, match="callback must be callable"):
        CallbackPipeline(callback=None)


async def test_callback_pipeline_with_lambda():
    pipeline = CallbackPipeline(
        callback=lambda item, spider: {**item, "processed": True}
    )
    spider = Spider()

    await pipeline.open(spider)
    result = await pipeline.process_item({"text": "Test"}, spider)
    await pipeline.close(spider)

    assert result == {"text": "Test", "processed": True}


async def test_callback_pipeline_callback_can_filter_item():
    """Test that callback can return a different item or filter it."""

    def filter_short_text(item, spider):
        if len(item.get("text", "")) < 5:
            return None  # Filter out short items
        return item

    pipeline = CallbackPipeline(callback=filter_short_text)
    spider = Spider()

    await pipeline.open(spider)

    # Short text should still return original item when callback returns None
    result1 = await pipeline.process_item({"text": "Hi"}, spider)
    assert result1 == {"text": "Hi"}

    # Long text should pass through
    result2 = await pipeline.process_item({"text": "Hello World"}, spider)
    assert result2 == {"text": "Hello World"}

    await pipeline.close(spider)


async def test_csv_pipeline_creates_valid_csv():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = Path(tmpdir) / "test.csv"
        pipeline = CSVPipeline(csv_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify CSV
        content = csv_path.read_text()
        lines = content.strip().split("\n")

        assert len(lines) == 3  # header + 2 data rows
        # Check header has both fields (order may vary)
        header_fields = set(lines[0].split(","))
        assert header_fields == {"text", "author"}
        assert "Hello" in content
        assert "John" in content
        assert "World" in content
        assert "Jane" in content


async def test_csv_pipeline_with_custom_fieldnames():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = Path(tmpdir) / "test.csv"
        pipeline = CSVPipeline(csv_path, fieldnames=["author", "text"])
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.close(spider)

        content = csv_path.read_text()
        lines = content.strip().split("\n")

        # Verify header order matches custom fieldnames
        assert lines[0] == "author,text"
        assert "John,Hello" in content


async def test_csv_pipeline_flattens_nested_dict():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = Path(tmpdir) / "test.csv"
        pipeline = CSVPipeline(csv_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"name": "Alice", "address": {"city": "NYC", "zip": "10001"}}, spider
        )
        await pipeline.close(spider)

        content = csv_path.read_text()

        assert "address_city" in content
        assert "address_zip" in content
        assert "NYC" in content
        assert "10001" in content


async def test_csv_pipeline_converts_list_to_string():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = Path(tmpdir) / "test.csv"
        pipeline = CSVPipeline(csv_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"author": "John", "tags": ["python", "web", "scraping"]}, spider
        )
        await pipeline.close(spider)

        content = csv_path.read_text()

        assert "python, web, scraping" in content


async def test_csv_pipeline_not_opened_raises_error():
    pipeline = CSVPipeline("test.csv")
    spider = Spider()

    with pytest.raises(RuntimeError, match="CSVPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


async def test_csv_pipeline_handles_extra_fields():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = Path(tmpdir) / "test.csv"
        pipeline = CSVPipeline(csv_path, fieldnames=["author"])
        spider = Spider()

        await pipeline.open(spider)
        # Item has extra field "text" that's not in fieldnames
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.close(spider)

        content = csv_path.read_text()
        lines = content.strip().split("\n")

        # Only author should be in output
        assert lines[0] == "author"
        assert "John" in content
        # text should not be in output
        assert "Hello" not in content


def test_sqlite_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        SQLitePipeline(table="invalid-table-name")

    with pytest.raises(ValueError, match="Invalid table name"):
        SQLitePipeline(table="123invalid")

    with pytest.raises(ValueError, match="Invalid table name"):
        SQLitePipeline(table="table; DROP TABLE users;")


# TaskiqPipeline tests - skip if taskiq not installed
try:
    from taskiq import InMemoryBroker  # type: ignore
    from silkworm.pipelines import TaskiqPipeline

    TASKIQ_AVAILABLE = True
except ImportError:
    TASKIQ_AVAILABLE = False


@pytest.mark.skipif(not TASKIQ_AVAILABLE, reason="taskiq not installed")
async def test_taskiq_pipeline_sends_items_to_queue():
    broker = InMemoryBroker()
    processed_items = []

    @broker.task
    async def process_item(item):
        processed_items.append(item)
        return item

    # Pass the task directly
    pipeline = TaskiqPipeline(broker, task=process_item)
    spider = Spider()

    await pipeline.open(spider)
    await pipeline.process_item({"text": "Hello", "author": "Alice"}, spider)
    await pipeline.process_item({"text": "World", "author": "Bob"}, spider)

    # Wait for InMemoryBroker to process tasks asynchronously
    await asyncio.sleep(0.1)

    await pipeline.close(spider)

    # InMemoryBroker processes tasks asynchronously
    assert len(processed_items) == 2
    assert processed_items[0] == {"text": "Hello", "author": "Alice"}
    assert processed_items[1] == {"text": "World", "author": "Bob"}


@pytest.mark.skipif(not TASKIQ_AVAILABLE, reason="taskiq not installed")
async def test_taskiq_pipeline_not_opened_raises_error():
    broker = InMemoryBroker()

    @broker.task
    async def process_item(item):
        return item

    pipeline = TaskiqPipeline(broker, task=process_item)
    spider = Spider()

    with pytest.raises(RuntimeError, match="TaskiqPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


@pytest.mark.skipif(not TASKIQ_AVAILABLE, reason="taskiq not installed")
async def test_taskiq_pipeline_invalid_task_name_raises_error():
    broker = InMemoryBroker()

    @broker.task
    async def process_item(item):
        return item

    # Use task_name parameter with an invalid name
    pipeline = TaskiqPipeline(broker, task_name="nonexistent_task")
    spider = Spider()

    with pytest.raises(ValueError, match="Task 'nonexistent_task' not found"):
        await pipeline.open(spider)


@pytest.mark.skipif(not TASKIQ_AVAILABLE, reason="taskiq not installed")
def test_taskiq_pipeline_without_task_or_name_raises_error():
    broker = InMemoryBroker()

    with pytest.raises(
        ValueError, match="Either 'task' or 'task_name' must be provided"
    ):
        TaskiqPipeline(broker)


@pytest.mark.skipif(not TASKIQ_AVAILABLE, reason="taskiq not installed")
def test_taskiq_pipeline_without_taskiq_raises_import_error():
    # This test simulates what happens when taskiq is not installed
    # We can't really test this without mocking, but we ensure the error message is correct
    from silkworm.pipelines import TASKIQ_AVAILABLE

    if not TASKIQ_AVAILABLE:
        pytest.skip("taskiq is installed, cannot test ImportError path")
    # If taskiq is available, this test is satisfied


# MsgPackPipeline tests - skip if ormsgpack not installed
# Note: We import both ormsgpack and msgpack. ormsgpack is required for the
# MsgPackPipeline to work (for writing), but we use msgpack for reading in tests
# because ormsgpack doesn't have an Unpacker class to read multiple objects from a stream.
try:
    import ormsgpack  # type: ignore  # noqa: F401
    import msgpack  # type: ignore
    from silkworm.pipelines import MsgPackPipeline

    ORMSGPACK_AVAILABLE = True
except ImportError:
    ORMSGPACK_AVAILABLE = False


@pytest.mark.skipif(not ORMSGPACK_AVAILABLE, reason="ormsgpack not installed")
async def test_msgpack_pipeline_writes_items():
    with tempfile.TemporaryDirectory() as tmpdir:
        msgpack_path = Path(tmpdir) / "test.msgpack"
        pipeline = MsgPackPipeline(msgpack_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify MsgPack data
        with open(msgpack_path, "rb") as f:
            data = f.read()

        # Unpack both items using msgpack.Unpacker
        unpacker = msgpack.Unpacker()
        unpacker.feed(data)
        items = list(unpacker)

        assert len(items) == 2
        assert items[0] == {"text": "Hello", "author": "John"}
        assert items[1] == {"text": "World", "author": "Jane"}


@pytest.mark.skipif(not ORMSGPACK_AVAILABLE, reason="ormsgpack not installed")
async def test_msgpack_pipeline_append_mode():
    with tempfile.TemporaryDirectory() as tmpdir:
        msgpack_path = Path(tmpdir) / "test.msgpack"
        spider = Spider()

        # Write first item
        pipeline1 = MsgPackPipeline(msgpack_path, mode="write")
        await pipeline1.open(spider)
        await pipeline1.process_item({"text": "First"}, spider)
        await pipeline1.close(spider)

        # Append second item
        pipeline2 = MsgPackPipeline(msgpack_path, mode="append")
        await pipeline2.open(spider)
        await pipeline2.process_item({"text": "Second"}, spider)
        await pipeline2.close(spider)

        # Read and verify both items
        with open(msgpack_path, "rb") as f:
            data = f.read()

        unpacker = msgpack.Unpacker()
        unpacker.feed(data)
        items = list(unpacker)

        assert len(items) == 2
        assert items[0] == {"text": "First"}
        assert items[1] == {"text": "Second"}


@pytest.mark.skipif(not ORMSGPACK_AVAILABLE, reason="ormsgpack not installed")
async def test_msgpack_pipeline_handles_nested_data():
    with tempfile.TemporaryDirectory() as tmpdir:
        msgpack_path = Path(tmpdir) / "test.msgpack"
        pipeline = MsgPackPipeline(msgpack_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"user": {"name": "Alice", "age": 30}, "tags": ["python", "web"]}, spider
        )
        await pipeline.close(spider)

        # Read and verify
        with open(msgpack_path, "rb") as f:
            data = f.read()

        unpacker = msgpack.Unpacker()
        unpacker.feed(data)
        items = list(unpacker)

        assert len(items) == 1
        assert items[0] == {
            "user": {"name": "Alice", "age": 30},
            "tags": ["python", "web"],
        }


@pytest.mark.skipif(not ORMSGPACK_AVAILABLE, reason="ormsgpack not installed")
async def test_msgpack_pipeline_not_opened_raises_error():
    pipeline = MsgPackPipeline("test.msgpack")
    spider = Spider()

    with pytest.raises(RuntimeError, match="MsgPackPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


@pytest.mark.skipif(not ORMSGPACK_AVAILABLE, reason="ormsgpack not installed")
def test_msgpack_pipeline_invalid_mode_raises_error():
    with pytest.raises(ValueError, match="mode must be 'write' or 'append'"):
        MsgPackPipeline("test.msgpack", mode="invalid")


# PolarsPipeline tests - skip if polars not installed
try:
    import polars as pl  # type: ignore
    from silkworm.pipelines import PolarsPipeline

    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False


@pytest.mark.skipif(not POLARS_AVAILABLE, reason="polars not installed")
async def test_polars_pipeline_writes_parquet():
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_path = Path(tmpdir) / "test.parquet"
        pipeline = PolarsPipeline(parquet_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify Parquet data
        df = pl.read_parquet(parquet_path)
        assert len(df) == 2
        assert df["text"].to_list() == ["Hello", "World"]
        assert df["author"].to_list() == ["John", "Jane"]


@pytest.mark.skipif(not POLARS_AVAILABLE, reason="polars not installed")
async def test_polars_pipeline_append_mode():
    with tempfile.TemporaryDirectory() as tmpdir:
        parquet_path = Path(tmpdir) / "test.parquet"
        spider = Spider()

        # Write first item
        pipeline1 = PolarsPipeline(parquet_path, mode="write")
        await pipeline1.open(spider)
        await pipeline1.process_item({"text": "First"}, spider)
        await pipeline1.close(spider)

        # Append second item
        pipeline2 = PolarsPipeline(parquet_path, mode="append")
        await pipeline2.open(spider)
        await pipeline2.process_item({"text": "Second"}, spider)
        await pipeline2.close(spider)

        # Read and verify both items
        df = pl.read_parquet(parquet_path)
        assert len(df) == 2
        assert df["text"].to_list() == ["First", "Second"]


@pytest.mark.skipif(not POLARS_AVAILABLE, reason="polars not installed")
def test_polars_pipeline_invalid_mode_raises_error():
    with pytest.raises(ValueError, match="mode must be 'write' or 'append'"):
        PolarsPipeline("test.parquet", mode="invalid")


# ExcelPipeline tests - skip if openpyxl not installed
try:
    import openpyxl  # type: ignore
    from silkworm.pipelines import ExcelPipeline

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_excel_pipeline_writes_xlsx():
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = Path(tmpdir) / "test.xlsx"
        pipeline = ExcelPipeline(excel_path, sheet_name="quotes")
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify Excel data
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["quotes"]

        # Check header
        header = [cell.value for cell in ws[1]]
        assert set(header) == {"text", "author"}

        # Check data
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert any(
            row[header.index("text")] == "Hello"
            and row[header.index("author")] == "John"
            for row in rows
        )
        assert any(
            row[header.index("text")] == "World"
            and row[header.index("author")] == "Jane"
            for row in rows
        )


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
async def test_excel_pipeline_flattens_nested_dict():
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = Path(tmpdir) / "test.xlsx"
        pipeline = ExcelPipeline(excel_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"name": "Alice", "address": {"city": "NYC", "zip": "10001"}}, spider
        )
        await pipeline.close(spider)

        # Read and verify
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]

        assert "address_city" in header
        assert "address_zip" in header


# YAMLPipeline tests - skip if pyyaml not installed
try:
    import yaml  # type: ignore
    from silkworm.pipelines import YAMLPipeline

    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


@pytest.mark.skipif(not YAML_AVAILABLE, reason="pyyaml not installed")
async def test_yaml_pipeline_writes_yaml():
    with tempfile.TemporaryDirectory() as tmpdir:
        yaml_path = Path(tmpdir) / "test.yaml"
        pipeline = YAMLPipeline(yaml_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify YAML data
        with open(yaml_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)

        assert len(data) == 2
        assert data[0] == {"text": "Hello", "author": "John"}
        assert data[1] == {"text": "World", "author": "Jane"}


@pytest.mark.skipif(not YAML_AVAILABLE, reason="pyyaml not installed")
async def test_yaml_pipeline_handles_nested_data():
    with tempfile.TemporaryDirectory() as tmpdir:
        yaml_path = Path(tmpdir) / "test.yaml"
        pipeline = YAMLPipeline(yaml_path)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"user": {"name": "Alice", "age": 30}, "tags": ["python", "web"]}, spider
        )
        await pipeline.close(spider)

        # Read and verify
        with open(yaml_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)

        assert len(data) == 1
        assert data[0] == {
            "user": {"name": "Alice", "age": 30},
            "tags": ["python", "web"],
        }


# AvroPipeline tests - skip if fastavro not installed
try:
    import fastavro  # type: ignore
    from silkworm.pipelines import AvroPipeline

    FASTAVRO_AVAILABLE = True
except ImportError:
    FASTAVRO_AVAILABLE = False


@pytest.mark.skipif(not FASTAVRO_AVAILABLE, reason="fastavro not installed")
async def test_avro_pipeline_writes_with_schema():
    with tempfile.TemporaryDirectory() as tmpdir:
        avro_path = Path(tmpdir) / "test.avro"
        schema = {
            "type": "record",
            "name": "Quote",
            "fields": [
                {"name": "text", "type": "string"},
                {"name": "author", "type": "string"},
            ],
        }
        pipeline = AvroPipeline(avro_path, schema=schema)
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify Avro data
        with open(avro_path, "rb") as f:
            reader = fastavro.reader(f)
            records = list(reader)

        assert len(records) == 2
        assert records[0] == {"text": "Hello", "author": "John"}
        assert records[1] == {"text": "World", "author": "Jane"}


@pytest.mark.skipif(not FASTAVRO_AVAILABLE, reason="fastavro not installed")
async def test_avro_pipeline_infers_schema():
    with tempfile.TemporaryDirectory() as tmpdir:
        avro_path = Path(tmpdir) / "test.avro"
        pipeline = AvroPipeline(avro_path)  # No schema provided
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"text": "Hello", "author": "John", "count": 5}, spider
        )
        await pipeline.close(spider)

        # Read and verify Avro data
        with open(avro_path, "rb") as f:
            reader = fastavro.reader(f)
            records = list(reader)

        assert len(records) == 1
        assert records[0]["text"] == "Hello"
        assert records[0]["author"] == "John"
        assert records[0]["count"] == 5


# ElasticsearchPipeline tests - skip if elasticsearch not installed
try:
    from elasticsearch import AsyncElasticsearch  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import ElasticsearchPipeline

    ELASTICSEARCH_AVAILABLE = True
except ImportError:
    ELASTICSEARCH_AVAILABLE = False
    ElasticsearchPipeline = None  # type: ignore


@pytest.mark.skipif(not ELASTICSEARCH_AVAILABLE, reason="elasticsearch not installed")
def test_elasticsearch_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = ElasticsearchPipeline(  # type: ignore
        hosts=["http://localhost:9200"],
        index="test_index",
    )
    assert pipeline.index == "test_index"
    assert pipeline.hosts == ["http://localhost:9200"]


# MongoDBPipeline tests - skip if motor not installed
try:
    import motor.motor_asyncio  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import MongoDBPipeline

    MOTOR_AVAILABLE = True
except ImportError:
    MOTOR_AVAILABLE = False
    MongoDBPipeline = None  # type: ignore


@pytest.mark.skipif(not MOTOR_AVAILABLE, reason="motor not installed")
def test_mongodb_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = MongoDBPipeline(  # type: ignore
        connection_string="mongodb://localhost:27017",
        database="test_db",
        collection="test_collection",
    )
    assert pipeline.database == "test_db"
    assert pipeline.collection == "test_collection"


# S3JsonLinesPipeline tests - skip if opendal not installed
try:
    import opendal  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import S3JsonLinesPipeline

    OPENDAL_AVAILABLE = True
except ImportError:
    OPENDAL_AVAILABLE = False
    S3JsonLinesPipeline = None  # type: ignore


@pytest.mark.skipif(not OPENDAL_AVAILABLE, reason="opendal not installed")
def test_s3_jsonlines_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = S3JsonLinesPipeline(  # type: ignore
        bucket="test-bucket",
        key="data/items.jl",
        region="us-east-1",
    )
    assert pipeline.bucket == "test-bucket"
    assert pipeline.key == "data/items.jl"
    assert pipeline.region == "us-east-1"


# VortexPipeline tests - skip if vortex not installed
try:
    import vortex  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import VortexPipeline

    VORTEX_AVAILABLE = True
except ImportError:
    VORTEX_AVAILABLE = False
    VortexPipeline = None  # type: ignore


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
async def test_vortex_pipeline_writes_vortex_file():
    with tempfile.TemporaryDirectory() as tmpdir:
        vortex_path = Path(tmpdir) / "test.vortex"
        pipeline = VortexPipeline(vortex_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Verify Vortex file was created and can be read
        assert vortex_path.exists()
        vortex_file = vortex.file.open(str(vortex_path))
        arrow_reader = vortex_file.to_arrow()
        table = arrow_reader.read_all()

        assert len(table) == 2
        data = table.to_pydict()
        assert data["text"] == ["Hello", "World"]
        assert data["author"] == ["John", "Jane"]


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
async def test_vortex_pipeline_handles_nested_data():
    with tempfile.TemporaryDirectory() as tmpdir:
        vortex_path = Path(tmpdir) / "test.vortex"
        pipeline = VortexPipeline(vortex_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"user": {"name": "Alice", "age": 30}, "tags": ["python", "web"]}, spider
        )
        await pipeline.close(spider)

        # Verify data can be read back
        vortex_file = vortex.file.open(str(vortex_path))
        arrow_reader = vortex_file.to_arrow()
        table = arrow_reader.read_all()

        assert len(table) == 1
        # Vortex/Arrow preserves nested structures


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
async def test_vortex_pipeline_handles_empty():
    with tempfile.TemporaryDirectory() as tmpdir:
        vortex_path = Path(tmpdir) / "test.vortex"
        pipeline = VortexPipeline(vortex_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.close(spider)

        # File should not be created when no items
        assert not vortex_path.exists()


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
def test_vortex_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = VortexPipeline("test.vortex")  # type: ignore
    assert pipeline.path == Path("test.vortex")


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
async def test_vortex_pipeline_handles_various_types():
    with tempfile.TemporaryDirectory() as tmpdir:
        vortex_path = Path(tmpdir) / "test.vortex"
        pipeline = VortexPipeline(vortex_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {
                "string": "test",
                "integer": 42,
                "float": 3.14,
                "boolean": True,
                "null": None,
                "list": [1, 2, 3],
            },
            spider,
        )
        await pipeline.close(spider)

        # Verify data types are preserved
        vortex_file = vortex.file.open(str(vortex_path))
        arrow_reader = vortex_file.to_arrow()
        table = arrow_reader.read_all()

        assert len(table) == 1
        data = table.to_pydict()
        assert data["string"] == ["test"]
        assert data["integer"] == [42]
        assert abs(data["float"][0] - 3.14) < 0.01
        assert data["boolean"] == [True]


# MySQLPipeline tests - skip if aiomysql not installed
try:
    import aiomysql  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import MySQLPipeline

    AIOMYSQL_AVAILABLE = True
except ImportError:
    AIOMYSQL_AVAILABLE = False
    MySQLPipeline = None  # type: ignore


@pytest.mark.skipif(not AIOMYSQL_AVAILABLE, reason="aiomysql not installed")
def test_mysql_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = MySQLPipeline(  # type: ignore
        host="localhost",
        port=3306,
        user="root",
        password="password",
        database="test_db",
        table="test_table",
    )
    assert pipeline.host == "localhost"
    assert pipeline.port == 3306
    assert pipeline.user == "root"
    assert pipeline.database == "test_db"
    assert pipeline.table == "test_table"


@pytest.mark.skipif(not AIOMYSQL_AVAILABLE, reason="aiomysql not installed")
def test_mysql_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        MySQLPipeline(table="invalid-table-name")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        MySQLPipeline(table="123invalid")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        MySQLPipeline(table="table; DROP TABLE users;")  # type: ignore


# PostgreSQLPipeline tests - skip if asyncpg not installed
try:
    import asyncpg  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import PostgreSQLPipeline

    ASYNCPG_AVAILABLE = True
except ImportError:
    ASYNCPG_AVAILABLE = False
    PostgreSQLPipeline = None  # type: ignore


@pytest.mark.skipif(not ASYNCPG_AVAILABLE, reason="asyncpg not installed")
def test_postgresql_pipeline_initialization():
    # Just test that we can initialize the pipeline
    pipeline = PostgreSQLPipeline(  # type: ignore
        host="localhost",
        port=5432,
        user="postgres",
        password="password",
        database="test_db",
        table="test_table",
    )
    assert pipeline.host == "localhost"
    assert pipeline.port == 5432
    assert pipeline.user == "postgres"
    assert pipeline.database == "test_db"
    assert pipeline.table == "test_table"


@pytest.mark.skipif(not ASYNCPG_AVAILABLE, reason="asyncpg not installed")
def test_postgresql_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        PostgreSQLPipeline(table="invalid-table-name")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        PostgreSQLPipeline(table="123invalid")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        PostgreSQLPipeline(table="table; DROP TABLE users;")  # type: ignore


# WebhookPipeline tests
try:
    from silkworm.pipelines import WebhookPipeline

    WEBHOOK_AVAILABLE = True
except ImportError:
    WEBHOOK_AVAILABLE = False
    WebhookPipeline = None  # type: ignore


@pytest.mark.skipif(not WEBHOOK_AVAILABLE, reason="rnet not available")
def test_webhook_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = WebhookPipeline(  # type: ignore
        url="https://webhook.site/test",
        method="POST",
        headers={"Authorization": "Bearer token"},
        batch_size=5,
    )
    assert pipeline.url == "https://webhook.site/test"
    assert pipeline.method == "POST"
    assert pipeline.headers == {"Authorization": "Bearer token"}
    assert pipeline.batch_size == 5


@pytest.mark.skipif(not WEBHOOK_AVAILABLE, reason="rnet not available")
async def test_webhook_pipeline_not_opened_raises_error():
    pipeline = WebhookPipeline("https://webhook.site/test")  # type: ignore
    spider = Spider()

    with pytest.raises(RuntimeError, match="WebhookPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


@pytest.mark.skipif(
    not GOOGLE_SHEETS_AVAILABLE, reason="google-api-python-client not installed"
)
def test_google_sheets_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = GoogleSheetsPipeline(  # type: ignore
        spreadsheet_id="test_id",
        credentials_file="test_creds.json",
        sheet_name="TestSheet",
        batch_size=50,
    )
    assert pipeline.spreadsheet_id == "test_id"
    assert pipeline.credentials_file == "test_creds.json"
    assert pipeline.sheet_name == "TestSheet"
    assert pipeline.batch_size == 50


@pytest.mark.skipif(
    not GOOGLE_SHEETS_AVAILABLE, reason="google-api-python-client not installed"
)
async def test_google_sheets_pipeline_not_opened_raises_error():
    pipeline = GoogleSheetsPipeline(  # type: ignore
        spreadsheet_id="test_id", credentials_file="test_creds.json"
    )
    spider = Spider()

    with pytest.raises(RuntimeError, match="GoogleSheetsPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


@pytest.mark.skipif(
    not SNOWFLAKE_AVAILABLE, reason="snowflake-connector-python not installed"
)
def test_snowflake_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = SnowflakePipeline(  # type: ignore
        account="test_account",
        user="test_user",
        password="test_password",
        database="test_db",
        schema="test_schema",
        warehouse="test_warehouse",
        table="test_table",
        role="test_role",
    )
    assert pipeline.account == "test_account"
    assert pipeline.user == "test_user"
    assert pipeline.database == "test_db"
    assert pipeline.schema == "test_schema"
    assert pipeline.warehouse == "test_warehouse"
    assert pipeline.table == "test_table"
    assert pipeline.role == "test_role"


@pytest.mark.skipif(
    not SNOWFLAKE_AVAILABLE, reason="snowflake-connector-python not installed"
)
def test_snowflake_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        SnowflakePipeline(  # type: ignore
            account="test",
            user="test",
            password="test",
            database="test",
            schema="test",
            warehouse="test",
            table="invalid-table-name",
        )

    with pytest.raises(ValueError, match="Invalid table name"):
        SnowflakePipeline(  # type: ignore
            account="test",
            user="test",
            password="test",
            database="test",
            schema="test",
            warehouse="test",
            table="123invalid",
        )


@pytest.mark.skipif(
    not SNOWFLAKE_AVAILABLE, reason="snowflake-connector-python not installed"
)
async def test_snowflake_pipeline_not_opened_raises_error():
    pipeline = SnowflakePipeline(  # type: ignore
        account="test_account",
        user="test_user",
        password="test_password",
        database="test_db",
        schema="test_schema",
        warehouse="test_warehouse",
    )
    spider = Spider()

    with pytest.raises(RuntimeError, match="SnowflakePipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


# FTPPipeline tests - skip if aioftp not installed
try:
    import aioftp  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import FTPPipeline

    AIOFTP_AVAILABLE = True
except ImportError:
    AIOFTP_AVAILABLE = False
    FTPPipeline = None  # type: ignore


@pytest.mark.skipif(not AIOFTP_AVAILABLE, reason="aioftp not installed")
def test_ftp_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = FTPPipeline(  # type: ignore
        host="ftp.example.com",
        user="username",
        password="password",
        remote_path="data/items.jl",
        port=21,
    )
    assert pipeline.host == "ftp.example.com"
    assert pipeline.user == "username"
    assert pipeline.password == "password"
    assert pipeline.remote_path == "data/items.jl"
    assert pipeline.port == 21


# SFTPPipeline tests - skip if asyncssh not installed
try:
    import asyncssh  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import SFTPPipeline

    ASYNCSSH_AVAILABLE = True
except ImportError:
    ASYNCSSH_AVAILABLE = False
    SFTPPipeline = None  # type: ignore


@pytest.mark.skipif(not ASYNCSSH_AVAILABLE, reason="asyncssh not installed")
def test_sftp_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = SFTPPipeline(  # type: ignore
        host="sftp.example.com",
        user="username",
        password="password",
        remote_path="data/items.jl",
        port=22,
    )
    assert pipeline.host == "sftp.example.com"
    assert pipeline.user == "username"
    assert pipeline.password == "password"
    assert pipeline.remote_path == "data/items.jl"
    assert pipeline.port == 22


@pytest.mark.skipif(not ASYNCSSH_AVAILABLE, reason="asyncssh not installed")
def test_sftp_pipeline_initialization_with_private_key():
    # Test that we can initialize the pipeline with private key
    pipeline = SFTPPipeline(  # type: ignore
        host="sftp.example.com",
        user="username",
        remote_path="data/items.jl",
        private_key="/path/to/key",
    )
    assert pipeline.host == "sftp.example.com"
    assert pipeline.user == "username"
    assert pipeline.password is None
    assert pipeline.private_key == "/path/to/key"


@pytest.mark.skipif(not ASYNCSSH_AVAILABLE, reason="asyncssh not installed")
def test_sftp_pipeline_requires_password_or_key():
    # Test that initializing without password or key raises error
    with pytest.raises(
        ValueError, match="Either password or private_key must be provided"
    ):
        SFTPPipeline(  # type: ignore
            host="sftp.example.com",
            user="username",
            remote_path="data/items.jl",
        )


# CassandraPipeline tests - skip if cassandra-driver not installed or on Windows
if sys.platform == "win32":
    CASSANDRA_AVAILABLE = False
    CassandraPipeline = None  # type: ignore
else:
    try:
        from cassandra.cluster import Cluster  # type: ignore[import-not-found]  # noqa: F401
        from silkworm.pipelines import CassandraPipeline

        CASSANDRA_AVAILABLE = True
    except ImportError:
        CASSANDRA_AVAILABLE = False
        CassandraPipeline = None  # type: ignore


@pytest.mark.skipif(not CASSANDRA_AVAILABLE, reason="cassandra-driver not installed")
@pytest.mark.skipif(
    sys.platform == "win32", reason="cassandra tests disabled on Windows"
)
def test_cassandra_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = CassandraPipeline(  # type: ignore
        hosts=["127.0.0.1"],
        keyspace="test_keyspace",
        table="test_table",
        username="cassandra",
        password="cassandra",
        port=9042,
    )
    assert pipeline.hosts == ["127.0.0.1"]
    assert pipeline.keyspace == "test_keyspace"
    assert pipeline.table == "test_table"
    assert pipeline.username == "cassandra"
    assert pipeline.password == "cassandra"
    assert pipeline.port == 9042


@pytest.mark.skipif(not CASSANDRA_AVAILABLE, reason="cassandra-driver not installed")
@pytest.mark.skipif(
    sys.platform == "win32", reason="cassandra tests disabled on Windows"
)
def test_cassandra_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        CassandraPipeline(table="invalid-table-name")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        CassandraPipeline(table="123invalid")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        CassandraPipeline(table="table; DROP TABLE users;")  # type: ignore


@pytest.mark.skipif(not CASSANDRA_AVAILABLE, reason="cassandra-driver not installed")
@pytest.mark.skipif(
    sys.platform == "win32", reason="cassandra tests disabled on Windows"
)
async def test_cassandra_pipeline_not_opened_raises_error():
    pipeline = CassandraPipeline(  # type: ignore
        hosts=["127.0.0.1"],
        keyspace="test_keyspace",
        table="test_table",
    )
    spider = Spider()

    with pytest.raises(RuntimeError, match="CassandraPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


# CouchDBPipeline tests - skip if aiocouch not installed
try:
    import aiocouch  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import CouchDBPipeline

    AIOCOUCH_AVAILABLE = True
except ImportError:
    AIOCOUCH_AVAILABLE = False
    CouchDBPipeline = None  # type: ignore


@pytest.mark.skipif(not AIOCOUCH_AVAILABLE, reason="aiocouch not installed")
def test_couchdb_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = CouchDBPipeline(  # type: ignore
        url="http://localhost:5984",
        database="test_db",
        username="admin",
        password="password",
    )
    assert pipeline.url == "http://localhost:5984"
    assert pipeline.database == "test_db"
    assert pipeline.username == "admin"
    assert pipeline.password == "password"


@pytest.mark.skipif(not AIOCOUCH_AVAILABLE, reason="aiocouch not installed")
async def test_couchdb_pipeline_not_opened_raises_error():
    pipeline = CouchDBPipeline(  # type: ignore
        url="http://localhost:5984",
        database="test_db",
    )
    spider = Spider()

    with pytest.raises(RuntimeError, match="CouchDBPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


# DynamoDBPipeline tests - skip if aioboto3 not installed
try:
    import aioboto3  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import DynamoDBPipeline

    AIOBOTO3_AVAILABLE = True
except ImportError:
    AIOBOTO3_AVAILABLE = False
    DynamoDBPipeline = None  # type: ignore


@pytest.mark.skipif(not AIOBOTO3_AVAILABLE, reason="aioboto3 not installed")
def test_dynamodb_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = DynamoDBPipeline(  # type: ignore
        table_name="test_table",
        region_name="us-west-2",
        aws_access_key_id="test_key",
        aws_secret_access_key="test_secret",
        endpoint_url="http://localhost:8000",
    )
    assert pipeline.table_name == "test_table"
    assert pipeline.region_name == "us-west-2"
    assert pipeline.aws_access_key_id == "test_key"
    assert pipeline.aws_secret_access_key == "test_secret"
    assert pipeline.endpoint_url == "http://localhost:8000"


@pytest.mark.skipif(not AIOBOTO3_AVAILABLE, reason="aioboto3 not installed")
async def test_dynamodb_pipeline_not_opened_raises_error():
    pipeline = DynamoDBPipeline(  # type: ignore
        table_name="test_table",
        region_name="us-east-1",
    )
    spider = Spider()

    with pytest.raises(RuntimeError, match="DynamoDBPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


# DuckDBPipeline tests - skip if duckdb not installed
try:
    import duckdb  # type: ignore[import-not-found]  # noqa: F401
    from silkworm.pipelines import DuckDBPipeline

    DUCKDB_AVAILABLE = True
except ImportError:
    DUCKDB_AVAILABLE = False
    DuckDBPipeline = None  # type: ignore


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
async def test_duckdb_pipeline_writes_items():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        pipeline = DuckDBPipeline(db_path, table="items")  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Hello", "author": "John"}, spider)
        await pipeline.process_item({"text": "World", "author": "Jane"}, spider)
        await pipeline.close(spider)

        # Read and verify data
        conn = duckdb.connect(str(db_path))
        result = conn.execute("SELECT spider, data FROM items ORDER BY id").fetchall()
        conn.close()

        assert len(result) == 2
        assert result[0][0] == "Spider"  # Default spider name
        assert json.loads(result[0][1]) == {"text": "Hello", "author": "John"}
        assert result[1][0] == "Spider"
        assert json.loads(result[1][1]) == {"text": "World", "author": "Jane"}


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
async def test_duckdb_pipeline_handles_nested_data():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        pipeline = DuckDBPipeline(db_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item(
            {"user": {"name": "Alice", "age": 30}, "tags": ["python", "web"]}, spider
        )
        await pipeline.close(spider)

        # Read and verify
        conn = duckdb.connect(str(db_path))
        result = conn.execute("SELECT data FROM items").fetchone()
        conn.close()

        data = json.loads(result[0])
        assert data == {"user": {"name": "Alice", "age": 30}, "tags": ["python", "web"]}


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
async def test_duckdb_pipeline_not_opened_raises_error():
    pipeline = DuckDBPipeline("test.db")  # type: ignore
    spider = Spider()

    with pytest.raises(RuntimeError, match="DuckDBPipeline not opened"):
        await pipeline.process_item({"test": "data"}, spider)


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
def test_duckdb_pipeline_invalid_table_name():
    # Test that invalid table names are rejected
    with pytest.raises(ValueError, match="Invalid table name"):
        DuckDBPipeline(table="invalid-table-name")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        DuckDBPipeline(table="123invalid")  # type: ignore

    with pytest.raises(ValueError, match="Invalid table name"):
        DuckDBPipeline(table="table; DROP TABLE users;")  # type: ignore


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
def test_duckdb_pipeline_initialization():
    # Test that we can initialize the pipeline
    pipeline = DuckDBPipeline(  # type: ignore
        database="data/test.db",
        table="test_table",
    )
    assert pipeline.database == Path("data/test.db")
    assert pipeline.table == "test_table"


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
async def test_duckdb_pipeline_creates_parent_directories():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "subdir" / "test.db"
        pipeline = DuckDBPipeline(db_path)  # type: ignore
        spider = Spider()

        await pipeline.open(spider)
        await pipeline.process_item({"text": "Test"}, spider)
        await pipeline.close(spider)

        # Verify database file was created
        assert db_path.exists()


@pytest.mark.skipif(not DUCKDB_AVAILABLE, reason="duckdb not installed")
async def test_duckdb_pipeline_persistent_across_sessions():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "test.db"
        spider = Spider()

        # First session: write items
        pipeline1 = DuckDBPipeline(db_path)  # type: ignore
        await pipeline1.open(spider)
        await pipeline1.process_item({"text": "First"}, spider)
        await pipeline1.close(spider)

        # Second session: write more items
        pipeline2 = DuckDBPipeline(db_path)  # type: ignore
        await pipeline2.open(spider)
        await pipeline2.process_item({"text": "Second"}, spider)
        await pipeline2.close(spider)

        # Verify both items are present
        conn = duckdb.connect(str(db_path))
        result = conn.execute("SELECT COUNT(*) FROM items").fetchone()
        conn.close()

        assert result[0] == 2
