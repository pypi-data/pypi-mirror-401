"""
Integration tests that run spiders with all available pipelines.

These tests verify that pipelines produce correct output files when used
with actual spider runs. Each test:
1. Creates a temporary directory for output files
2. Runs a test spider that yields sample data
3. Verifies the pipeline produces correctly formatted output files
4. Validates the content matches the expected data

This ensures all pipelines work end-to-end in a realistic scenario.
"""

import csv
import json
import platform
import sqlite3
import sys
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest

from silkworm import HTMLResponse, Request, Response, Spider
from silkworm.pipelines import (
    CSVPipeline,
    JsonLinesPipeline,
    SQLitePipeline,
    XMLPipeline,
)

# Test data that will be used across all pipeline tests
SAMPLE_QUOTES = [
    {"text": "First quote", "author": "Author One", "tags": ["tag1", "tag2"]},
    {"text": "Second quote", "author": "Author Two", "tags": ["tag3", "tag4"]},
    {"text": "Third quote", "author": "Author Three", "tags": ["tag5"]},
]


class TestSpider(Spider):
    """A simple spider that yields test data."""

    __test__ = False  # prevent pytest from treating this helper as a test class
    name = "test"
    start_urls = ("http://example.com",)

    def __init__(self, quotes_data=None, **kwargs):
        super().__init__(**kwargs)
        self.quotes_data = quotes_data or SAMPLE_QUOTES

    async def parse(self, response: Response):
        # Yield all test quotes
        for quote in self.quotes_data:
            yield quote


def create_mock_response():
    """Create a mock HTMLResponse for testing."""
    mock_html = "<html><body>Test</body></html>"
    mock_request = Request(url="http://example.com")
    return HTMLResponse(
        url=mock_request.url,
        status=200,
        headers={},
        body=mock_html.encode("utf-8"),
        request=mock_request,
    )


async def run_spider_with_pipeline(spider_cls, pipeline, **spider_kwargs):
    """Helper to run a spider with a pipeline and return the spider instance."""
    spider = spider_cls(**spider_kwargs)
    mock_response = create_mock_response()

    # Open the pipeline
    await pipeline.open(spider)

    # Process items from the spider's parse method
    async for item in spider.parse(mock_response):
        if isinstance(item, dict):
            await pipeline.process_item(item, spider)

    # Close the pipeline
    await pipeline.close(spider)

    return spider


async def test_jsonlines_pipeline_integration():
    """Test JsonLinesPipeline produces correct JSON Lines file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.jl"
        pipeline = JsonLinesPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        with open(output_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        assert len(lines) == len(SAMPLE_QUOTES)

        # Verify each line is valid JSON and matches expected data
        for i, line in enumerate(lines):
            item = json.loads(line)
            assert item == SAMPLE_QUOTES[i]


async def test_csv_pipeline_integration():
    """Test CSVPipeline produces correct CSV file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.csv"
        pipeline = CSVPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        with open(output_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == len(SAMPLE_QUOTES)

        # Verify each row matches expected data
        for i, row in enumerate(rows):
            assert row["text"] == SAMPLE_QUOTES[i]["text"]
            assert row["author"] == SAMPLE_QUOTES[i]["author"]
            # CSV converts lists to comma-separated strings
            expected_tags = ", ".join(SAMPLE_QUOTES[i]["tags"])
            assert row["tags"] == expected_tags


async def test_xml_pipeline_integration():
    """Test XMLPipeline produces correct XML file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.xml"
        pipeline = XMLPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        tree = ET.parse(output_path)
        root = tree.getroot()

        assert root.tag == "items"
        items = list(root)
        assert len(items) == len(SAMPLE_QUOTES)

        # Verify each item matches expected data
        for i, item in enumerate(items):
            assert item.tag == "item"
            assert item.find("text").text == SAMPLE_QUOTES[i]["text"]
            assert item.find("author").text == SAMPLE_QUOTES[i]["author"]

            # Verify tags list structure
            tags_elem = item.find("tags")
            tag_items = tags_elem.findall("item")
            expected_tags = SAMPLE_QUOTES[i]["tags"]
            assert len(tag_items) == len(expected_tags)
            for j, tag in enumerate(tag_items):
                assert tag.text == expected_tags[j]


async def test_sqlite_pipeline_integration():
    """Test SQLitePipeline produces correct SQLite database."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "quotes.db"
        pipeline = SQLitePipeline(db_path, table="quotes")

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the database exists and has correct content
        assert db_path.exists()

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Verify table exists
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='quotes'"
        )
        assert cursor.fetchone() is not None

        # Verify data
        cursor.execute("SELECT * FROM quotes")
        rows = cursor.fetchall()
        assert len(rows) == len(SAMPLE_QUOTES)

        # Get column names
        cursor.execute("PRAGMA table_info(quotes)")
        columns = [col[1] for col in cursor.fetchall()]

        # Verify each row matches expected data
        for row in rows:
            row_dict = dict(zip(columns, row))
            # SQLitePipeline stores items as JSON in the 'data' column
            item_data = json.loads(row_dict["data"])
            # Find matching quote in SAMPLE_QUOTES
            matching_quote = next(
                (q for q in SAMPLE_QUOTES if q["text"] == item_data["text"]), None
            )
            assert matching_quote is not None
            assert item_data["author"] == matching_quote["author"]
            assert item_data["tags"] == matching_quote["tags"]
            # Verify spider column
            assert row_dict["spider"] == "test"

        conn.close()


async def test_xml_pipeline_custom_elements():
    """Test XMLPipeline with custom root and item elements."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.xml"
        pipeline = XMLPipeline(output_path, root_element="quotes", item_element="quote")

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct structure
        assert output_path.exists()

        tree = ET.parse(output_path)
        root = tree.getroot()

        assert root.tag == "quotes"
        items = list(root)
        assert len(items) == len(SAMPLE_QUOTES)
        assert items[0].tag == "quote"


async def test_csv_pipeline_custom_fieldnames():
    """Test CSVPipeline with custom field names ordering."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.csv"
        pipeline = CSVPipeline(output_path, fieldnames=["author", "text"])

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct field order
        assert output_path.exists()

        with open(output_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            header = next(reader)
            assert header == ["author", "text"]

            rows = list(reader)
            assert len(rows) == len(SAMPLE_QUOTES)

            # Verify data is in correct order
            for i, row in enumerate(rows):
                assert row[0] == SAMPLE_QUOTES[i]["author"]
                assert row[1] == SAMPLE_QUOTES[i]["text"]


async def test_sqlite_pipeline_custom_table():
    """Test SQLitePipeline with custom table name."""
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = Path(tmpdir) / "data.db"
        pipeline = SQLitePipeline(db_path, table="custom_quotes")

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the database and table exist
        assert db_path.exists()

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='custom_quotes'"
        )
        assert cursor.fetchone() is not None

        cursor.execute("SELECT COUNT(*) FROM custom_quotes")
        count = cursor.fetchone()[0]
        assert count == len(SAMPLE_QUOTES)

        conn.close()


# Optional pipeline tests - skip if dependencies not installed

try:
    from silkworm.pipelines import MsgPackPipeline

    # Note: We use msgpack for reading because ormsgpack (used by MsgPackPipeline
    # for writing) doesn't have an Unpacker class to read multiple objects from a stream.
    # The two libraries are compatible for reading/writing MessagePack data.
    import msgpack  # For reading back the data

    MSGPACK_AVAILABLE = True
except ImportError:
    MSGPACK_AVAILABLE = False


@pytest.mark.skipif(not MSGPACK_AVAILABLE, reason="ormsgpack/msgpack not installed")
async def test_msgpack_pipeline_integration():
    """Test MsgPackPipeline produces correct MessagePack file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.msgpack"
        pipeline = MsgPackPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        with open(output_path, "rb") as f:
            data = f.read()

        # Unpack all items
        unpacker = msgpack.Unpacker()
        unpacker.feed(data)
        items = list(unpacker)

        assert len(items) == len(SAMPLE_QUOTES)

        # Verify each item matches expected data
        for i, item in enumerate(items):
            assert item == SAMPLE_QUOTES[i]


try:
    from silkworm.pipelines import PolarsPipeline
    import polars as pl

    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False


@pytest.mark.skipif(not POLARS_AVAILABLE, reason="polars not installed")
async def test_polars_pipeline_integration():
    """Test PolarsPipeline produces correct Parquet file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.parquet"
        pipeline = PolarsPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        df = pl.read_parquet(output_path)
        assert len(df) == len(SAMPLE_QUOTES)

        # Verify data using idiomatic Polars access
        text_values = df["text"].to_list()
        author_values = df["author"].to_list()

        for i in range(len(SAMPLE_QUOTES)):
            assert text_values[i] == SAMPLE_QUOTES[i]["text"]
            assert author_values[i] == SAMPLE_QUOTES[i]["author"]


try:
    from silkworm.pipelines import ExcelPipeline
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="openpyxl not installed")
async def test_excel_pipeline_integration():
    """Test ExcelPipeline produces correct Excel file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.xlsx"
        pipeline = ExcelPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Check that we have header + data rows
        assert ws.max_row == len(SAMPLE_QUOTES) + 1  # +1 for header

        # Get header
        header = [cell.value for cell in ws[1]]
        assert "text" in header
        assert "author" in header

        wb.close()


try:
    from silkworm.pipelines import YAMLPipeline
    import yaml

    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


@pytest.mark.skipif(not YAML_AVAILABLE, reason="pyyaml not installed")
async def test_yaml_pipeline_integration():
    """Test YAMLPipeline produces correct YAML file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.yaml"
        pipeline = YAMLPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        with open(output_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)

        assert len(data) == len(SAMPLE_QUOTES)

        # Verify each item matches expected data
        for i, item in enumerate(data):
            assert item == SAMPLE_QUOTES[i]


try:
    from silkworm.pipelines import AvroPipeline
    import fastavro

    AVRO_AVAILABLE = True
except ImportError:
    AVRO_AVAILABLE = False


@pytest.mark.skipif(not AVRO_AVAILABLE, reason="fastavro not installed")
async def test_avro_pipeline_integration():
    """Test AvroPipeline produces correct Avro file with inferred schema."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.avro"
        pipeline = AvroPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        with open(output_path, "rb") as f:
            reader = fastavro.reader(f)
            records = list(reader)

        assert len(records) == len(SAMPLE_QUOTES)

        # Verify each record matches expected data
        for i, record in enumerate(records):
            assert record["text"] == SAMPLE_QUOTES[i]["text"]
            assert record["author"] == SAMPLE_QUOTES[i]["author"]
            assert record["tags"] == SAMPLE_QUOTES[i]["tags"]


try:
    from silkworm.pipelines import VortexPipeline
    import vortex

    VORTEX_AVAILABLE = True
except ImportError:
    VORTEX_AVAILABLE = False


@pytest.mark.skipif(not VORTEX_AVAILABLE, reason="vortex not installed")
@pytest.mark.skipif(sys.platform == "win32", reason="vortex tests disabled on Windows")
async def test_vortex_pipeline_integration():
    """Test VortexPipeline produces correct Vortex file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "quotes.vortex"
        pipeline = VortexPipeline(output_path)

        await run_spider_with_pipeline(TestSpider, pipeline)

        # Verify the file exists and has correct content
        assert output_path.exists()

        vortex_file = vortex.file.open(str(output_path))
        arrow_reader = vortex_file.to_arrow()
        table = arrow_reader.read_all()

        assert len(table) == len(SAMPLE_QUOTES)

        data = table.to_pydict()
        assert data["text"] == [q["text"] for q in SAMPLE_QUOTES]
        assert data["author"] == [q["author"] for q in SAMPLE_QUOTES]


async def test_multiple_pipelines_simultaneously():
    """Test running a spider with multiple pipelines at once."""
    with tempfile.TemporaryDirectory() as tmpdir:
        jl_path = Path(tmpdir) / "quotes.jl"
        csv_path = Path(tmpdir) / "quotes.csv"
        xml_path = Path(tmpdir) / "quotes.xml"

        jl_pipeline = JsonLinesPipeline(jl_path)
        csv_pipeline = CSVPipeline(csv_path)
        xml_pipeline = XMLPipeline(xml_path)

        spider = TestSpider()
        mock_response = create_mock_response()

        # Open all pipelines
        await jl_pipeline.open(spider)
        await csv_pipeline.open(spider)
        await xml_pipeline.open(spider)

        # Process items through all pipelines
        async for item in spider.parse(mock_response):
            if isinstance(item, dict):
                await jl_pipeline.process_item(item, spider)
                await csv_pipeline.process_item(item, spider)
                await xml_pipeline.process_item(item, spider)

        # Close all pipelines
        await jl_pipeline.close(spider)
        await csv_pipeline.close(spider)
        await xml_pipeline.close(spider)

        # Verify all files exist
        assert jl_path.exists()
        assert csv_path.exists()
        assert xml_path.exists()

        # Quick verification of JSON Lines
        with open(jl_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        assert len(lines) == len(SAMPLE_QUOTES)

        # Quick verification of CSV
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == len(SAMPLE_QUOTES)

        # Quick verification of XML
        tree = ET.parse(xml_path)
        root = tree.getroot()
        items = list(root)
        assert len(items) == len(SAMPLE_QUOTES)


# Database pipeline tests with test containers
# Testcontainers are disabled on Windows as Docker doesn't work well on Windows CI
IS_WINDOWS = platform.system() == "Windows"

if not IS_WINDOWS:
    try:
        from silkworm.pipelines import MySQLPipeline
        from testcontainers.mysql import MySqlContainer  # noqa: F401
        import aiomysql

        MYSQL_AVAILABLE = True
    except ImportError:
        MYSQL_AVAILABLE = False
else:
    MYSQL_AVAILABLE = False


@pytest.mark.skipif(
    not MYSQL_AVAILABLE,
    reason="aiomysql or testcontainers not installed or running on Windows",
)
async def test_mysql_pipeline_integration(mysql_container):
    """Test MySQLPipeline with a real MySQL container."""
    # Extract connection details from container
    host = mysql_container.get_container_host_ip()
    port = int(mysql_container.get_exposed_port(3306))
    user = mysql_container.username
    password = mysql_container.password
    database = mysql_container.dbname

    pipeline = MySQLPipeline(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        table="test_quotes",
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to the database
    pool = await aiomysql.create_pool(
        host=host,
        port=port,
        user=user,
        password=password,
        db=database,
    )

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT COUNT(*) FROM test_quotes")
            result = await cur.fetchone()
            assert result[0] == len(SAMPLE_QUOTES)

            # Verify data content
            await cur.execute("SELECT spider, data FROM test_quotes")
            rows = await cur.fetchall()
            for row in rows:
                spider_name, data_json = row
                assert spider_name == "test"
                item_data = json.loads(data_json)
                # Verify item exists in SAMPLE_QUOTES
                matching_quote = next(
                    (q for q in SAMPLE_QUOTES if q["text"] == item_data["text"]), None
                )
                assert matching_quote is not None
                assert item_data["author"] == matching_quote["author"]

    pool.close()
    await pool.wait_closed()


if not IS_WINDOWS:
    try:
        from silkworm.pipelines import PostgreSQLPipeline
        from testcontainers.postgres import PostgresContainer  # noqa: F401
        import asyncpg

        POSTGRESQL_AVAILABLE = True
    except ImportError:
        POSTGRESQL_AVAILABLE = False
else:
    POSTGRESQL_AVAILABLE = False


@pytest.mark.skipif(
    not POSTGRESQL_AVAILABLE,
    reason="asyncpg or testcontainers not installed or running on Windows",
)
async def test_postgresql_pipeline_integration(postgres_container):
    """Test PostgreSQLPipeline with a real PostgreSQL container."""
    # Extract connection details from container
    host = postgres_container.get_container_host_ip()
    port = int(postgres_container.get_exposed_port(5432))
    user = postgres_container.username
    password = postgres_container.password
    database = postgres_container.dbname

    pipeline = PostgreSQLPipeline(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        table="test_quotes",
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to the database
    conn = await asyncpg.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
    )

    count = await conn.fetchval("SELECT COUNT(*) FROM test_quotes")
    assert count == len(SAMPLE_QUOTES)

    # Verify data content
    rows = await conn.fetch("SELECT spider, data FROM test_quotes")
    for row in rows:
        spider_name = row["spider"]
        data_json = row["data"]
        assert spider_name == "test"
        item_data = json.loads(data_json)
        # Verify item exists in SAMPLE_QUOTES
        matching_quote = next(
            (q for q in SAMPLE_QUOTES if q["text"] == item_data["text"]), None
        )
        assert matching_quote is not None
        assert item_data["author"] == matching_quote["author"]

    await conn.close()


if not IS_WINDOWS:
    try:
        from silkworm.pipelines import MongoDBPipeline
        from testcontainers.mongodb import MongoDbContainer  # noqa: F401
        import motor.motor_asyncio

        MONGODB_AVAILABLE = True
    except ImportError:
        MONGODB_AVAILABLE = False
else:
    MONGODB_AVAILABLE = False


@pytest.mark.skipif(
    not MONGODB_AVAILABLE,
    reason="motor or testcontainers not installed or running on Windows",
)
async def test_mongodb_pipeline_integration(mongodb_container):
    """Test MongoDBPipeline with a real MongoDB container."""
    # Extract connection details from container
    connection_string = mongodb_container.get_connection_url()

    pipeline = MongoDBPipeline(
        connection_string=connection_string,
        database="test_db",
        collection="test_quotes",
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to the database
    client = motor.motor_asyncio.AsyncIOMotorClient(connection_string)
    db = client["test_db"]
    collection = db["test_quotes"]

    count = await collection.count_documents({})
    assert count == len(SAMPLE_QUOTES)

    # Verify data content
    cursor = collection.find({})
    items = await cursor.to_list(length=None)
    assert len(items) == len(SAMPLE_QUOTES)

    for item in items:
        # Remove MongoDB's _id field before comparison
        item.pop("_id", None)
        # Verify item exists in SAMPLE_QUOTES
        matching_quote = next(
            (q for q in SAMPLE_QUOTES if q["text"] == item["text"]), None
        )
        assert matching_quote is not None
        assert item["author"] == matching_quote["author"]
        assert item["tags"] == matching_quote["tags"]

    client.close()


if not IS_WINDOWS:
    try:
        from silkworm.pipelines import ElasticsearchPipeline
        from testcontainers.elasticsearch import ElasticsearchContainer  # noqa: F401
        from elasticsearch import AsyncElasticsearch

        ELASTICSEARCH_AVAILABLE = True
    except ImportError:
        ELASTICSEARCH_AVAILABLE = False
else:
    ELASTICSEARCH_AVAILABLE = False


@pytest.mark.skipif(
    not ELASTICSEARCH_AVAILABLE,
    reason="elasticsearch or testcontainers not installed or running on Windows",
)
async def test_elasticsearch_pipeline_integration(elasticsearch_container):
    """Test ElasticsearchPipeline with a real Elasticsearch container."""
    # Extract connection details from container
    host = elasticsearch_container.get_container_host_ip()
    port = int(elasticsearch_container.get_exposed_port(9200))
    es_url = f"http://{host}:{port}"

    pipeline = ElasticsearchPipeline(
        hosts=[es_url],
        index="test_quotes",
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to Elasticsearch
    client = AsyncElasticsearch([es_url])

    # Refresh index to make documents available for search
    await client.indices.refresh(index="test_quotes")

    # Verify document count
    count_result = await client.count(index="test_quotes")
    assert count_result["count"] == len(SAMPLE_QUOTES)

    # Verify data content
    search_result = await client.search(
        index="test_quotes",
        query={"match_all": {}},
        size=100,
    )
    hits = search_result["hits"]["hits"]
    assert len(hits) == len(SAMPLE_QUOTES)

    for hit in hits:
        item = hit["_source"]
        # Verify item exists in SAMPLE_QUOTES
        matching_quote = next(
            (q for q in SAMPLE_QUOTES if q["text"] == item["text"]), None
        )
        assert matching_quote is not None
        assert item["author"] == matching_quote["author"]
        assert item["tags"] == matching_quote["tags"]

    await client.close()


if not IS_WINDOWS:
    try:
        from silkworm.pipelines import CassandraPipeline
        from testcontainers.cassandra import CassandraContainer  # noqa: F401
        from cassandra.cluster import Cluster

        CASSANDRA_AVAILABLE = True
    except ImportError:
        CASSANDRA_AVAILABLE = False
else:
    CASSANDRA_AVAILABLE = False


@pytest.mark.skipif(
    not CASSANDRA_AVAILABLE,
    reason="cassandra-driver or testcontainers not installed or running on Windows",
)
async def test_cassandra_pipeline_integration(cassandra_container):
    """Test CassandraPipeline with a real Cassandra container."""
    # Extract connection details from container
    host = cassandra_container.get_container_host_ip()
    port = int(cassandra_container.get_exposed_port(9042))

    pipeline = CassandraPipeline(
        hosts=[host],
        port=port,
        keyspace="test_keyspace",
        table="test_quotes",
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to Cassandra
    cluster = Cluster([host], port=port)
    session = cluster.connect("test_keyspace")

    # Verify row count
    rows = list(session.execute("SELECT COUNT(*) FROM test_quotes"))
    assert rows[0].count == len(SAMPLE_QUOTES)

    # Verify data content
    rows = list(session.execute("SELECT spider, data FROM test_quotes"))
    assert len(rows) == len(SAMPLE_QUOTES)

    for row in rows:
        spider_name = row.spider
        data_json = row.data
        assert spider_name == "test"
        item_data = json.loads(data_json)
        # Verify item exists in SAMPLE_QUOTES
        matching_quote = next(
            (q for q in SAMPLE_QUOTES if q["text"] == item_data["text"]), None
        )
        assert matching_quote is not None
        assert item_data["author"] == matching_quote["author"]
        assert item_data["tags"] == matching_quote["tags"]

    cluster.shutdown()


if not IS_WINDOWS:
    try:
        from silkworm.pipelines import CouchDBPipeline
        from testcontainers.couchdb import CouchDbContainer  # noqa: F401
        import aiocouch

        COUCHDB_AVAILABLE = True
    except ImportError:
        COUCHDB_AVAILABLE = False
else:
    COUCHDB_AVAILABLE = False


@pytest.mark.skipif(
    not COUCHDB_AVAILABLE,
    reason="aiocouch or testcontainers not installed or running on Windows",
)
async def test_couchdb_pipeline_integration(couchdb_container):
    """Test CouchDBPipeline with a real CouchDB container."""
    # Extract connection details from container
    url = couchdb_container.get_connection_url()
    username = couchdb_container.username
    password = couchdb_container.password

    pipeline = CouchDBPipeline(
        url=url,
        database="test_db",
        username=username,
        password=password,
    )

    await run_spider_with_pipeline(TestSpider, pipeline)

    # Verify data was inserted by connecting to CouchDB
    async with aiocouch.CouchDB(url, user=username, password=password) as client:
        db = await client["test_db"]

        # Get all documents
        all_docs = await db.akeys()
        # CouchDB includes design documents, so we filter for our data
        # Note: This is fine for small test datasets; for production use cases,
        # consider using a view or _all_docs with startkey/endkey parameters
        data_docs = [doc_id for doc_id in all_docs if not doc_id.startswith("_design")]
        assert len(data_docs) == len(SAMPLE_QUOTES)

        # Verify data content
        for doc_id in data_docs:
            doc = await db[doc_id]
            # CouchDB pipeline stores items with spider metadata
            assert doc["spider"] == "test"
            item = doc["data"]
            # Verify item exists in SAMPLE_QUOTES
            matching_quote = next(
                (q for q in SAMPLE_QUOTES if q["text"] == item["text"]), None
            )
            assert matching_quote is not None
            assert item["author"] == matching_quote["author"]
            assert item["tags"] == matching_quote["tags"]
