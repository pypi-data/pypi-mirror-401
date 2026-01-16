import os
from pathlib import Path

import polars as pl
import streamlit as st
from openpyxl import load_workbook
from polars_readstat import scan_readstat
from pydantic import BaseModel, Field, field_validator

from datasure.utils.duckdb_utils import duckdb_get_table, duckdb_save_table

# --- Pydantic Models for Validation ---

# VALID FILE TYPES
VALID_FILE_TYPES = {".csv", ".xlsx", ".xls", ".json", ".dta"}


class FileConfig(BaseModel):
    """Configuration for file import with validation."""

    alias: str = Field(
        ..., min_length=1, max_length=20, description="Unique alias for the file"
    )
    filename: str = Field(..., description="Full path to the file")
    sheet_name: str | None = Field(None, description="Sheet name for Excel files")
    source: str = Field(default="local storage", description="Source of the file")

    @field_validator("alias")
    def validate_alias(cls, v):
        """Validate alias to ensure it's non-empty and trimmed."""
        if not v or not v.strip():
            raise ValueError("Alias cannot be empty")
        return v.strip()

    @field_validator("filename")
    def validate_filename(cls, v):
        "Validate filename to ensure it exists and is of a supported type."
        if not v or not v.strip():
            raise ValueError("File path cannot be empty")

        file_path = Path(v.strip())
        if not file_path.exists():
            raise ValueError("File not found. Please check the file path")

        if not file_path.is_file():
            raise ValueError("Path is not a file")

        if file_path.suffix.lower() not in VALID_FILE_TYPES:
            raise ValueError(
                f"Invalid file type. Supported types: {', '.join(VALID_FILE_TYPES)}"
            )

        return str(file_path)


class ImportLogEntry(BaseModel):
    """Model for import log entries."""

    refresh: bool = True
    load: bool = True
    alias: str
    filename: str
    sheet_name: str | None = None
    source: str = "local storage"
    server: str = ""
    username: str = ""
    form_id: str = ""
    private_key: str = ""
    save_to: str = ""
    attachments: bool = False


# --- Optimized File Operations ---


def get_excel_sheet_names(file_path: str) -> list[str]:
    """Get sheet names from Excel file efficiently.

    Args:
        file_path: Path to the Excel file

    Returns
    -------
        List of sheet names
    """
    try:
        # Use read_only=True for better performance
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        return workbook.sheetnames  # noqa: TRY300
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return []


def load_data_efficiently(filename: str, sheet_name: str | None = None) -> pl.DataFrame:
    """Load data efficiently using Polars.

    Args:
        filename: Path to the file
        sheet_name: Sheet name for Excel files

    Returns
    -------
        Polars DataFrame with the loaded data
    """
    file_path = Path(filename)
    file_ext = file_path.suffix.lower()

    try:
        if file_ext == ".csv":
            # Polars CSV reader is much faster than pandas
            return pl.read_csv(
                filename,
                encoding="utf8-lossy",  # More robust encoding handling
                ignore_errors=True,  # Continue reading despite minor errors
                infer_schema_length=10000,  # Better type inference
            )

        elif file_ext in [".xlsx", ".xls"]:
            return pl.read_excel(filename, sheet_name=sheet_name)

        elif file_ext == ".json":
            # Polars has efficient JSON reading
            return pl.read_json(filename)

        elif file_ext == ".dta":
            return scan_readstat(filename).collect()

        else:
            raise ValueError(f"Unsupported file format: {file_ext}")  # noqa: TRY301

    except Exception as e:
        st.error(f"Error loading file {filename}: {e}")
        return pl.DataFrame()


# --- Streamlit Form with Enhanced Validation ---


def render_local_file_form(
    project_id: str, edit_mode: bool = False, defaults: dict | None = None
) -> None:
    """Create form for adding files from local storage with enhanced validation.

    Args:
        project_id: Project identifier
        edit_mode: Whether in edit mode
        defaults: Default values for form fields
    """
    defaults = defaults or {}
    mode = "edit" if edit_mode else "add"

    # UI Setup
    assets_dir = Path(__file__).parent.parent / "assets"
    image_path = assets_dir / "hard-disk.png"
    if image_path.exists():
        st.image(str(image_path), width=100)

    st.subheader("Add File from Local Storage")

    if edit_mode:
        st.info("You are in edit mode. Please modify the file details below.")

    # Form inputs with validation
    with st.container(border=True):
        disable_submit = True
        sheets = []
        sheet_name = None
        default_index = None
        alias = st.text_input(
            label="Alias*",
            value=defaults.get("alias", ""),
            help="Enter a unique, short, descriptive name for the file (max 20 characters)",
            disabled=edit_mode,
            max_chars=20,
            key=f"alias_input_{mode}",
        )

        file_path = st.text_input(
            label="File Path*",
            value=defaults.get("filename", None),
            help="Add full file name and path. e.g., C:/data/survey.dta",
            key=f"file_path_input_{mode}",
        )

        # Dynamic sheet selection for Excel files
        if file_path is not None:
            path_obj = Path(file_path)
            # validate file accessibility
            if not validate_file_accessibility(path_obj):
                st.error(
                    "File is not accessible. Please check the path, filename or permissions."
                )

            else:
                if path_obj.suffix.lower() in [".xlsx", ".xls"]:
                    try:
                        sheets = get_excel_sheet_names(file_path)
                        if sheets and defaults.get("sheet_name") in sheets:
                            default_index = sheets.index(defaults.get("sheet_name"))

                        disable_submit = bool(not sheets)

                    except Exception:
                        pass  # sheets will remain empty
                else:
                    disable_submit = False  # Non-Excel files
                    sheets = []

                if path_obj.suffix.lower() in [".xlsx", ".xls"]:
                    sheet_name = st.selectbox(
                        label="Sheet Name",
                        options=sheets,
                        index=default_index,
                    )

        st.markdown("**required*")

        # Submit button
        if st.button(
            "Update File" if edit_mode else "Add File",
            type="primary",
            width="stretch",
            disabled=disable_submit,
        ):
            _handle_form_submission(
                project_id=project_id,
                alias=alias,
                file_path=file_path,
                sheet_name=sheet_name or "",
                edit_mode=edit_mode,
                defaults=defaults,
            )


def _handle_form_submission(
    project_id: str,
    alias: str,
    file_path: str,
    sheet_name: str,
    edit_mode: bool,
    defaults: dict,
) -> None:
    """Handle form submission with validation."""
    try:
        # Validate using Pydantic
        file_config = FileConfig(alias=alias, filename=file_path, sheet_name=sheet_name)

        # Get existing import log
        import_log = duckdb_get_table(project_id, alias="import_log", db_name="logs")

        # Check alias uniqueness
        if not import_log.is_empty():
            existing_aliases = import_log.select("alias").to_series().to_list()
            if not edit_mode and alias in existing_aliases:
                st.error("Alias already exists. Please choose a different alias.")
                return

        if edit_mode:
            # Update existing entry
            import_log = import_log.with_columns(
                [
                    pl.when(pl.col("alias") == defaults.get("alias"))
                    .then(pl.lit(file_config.filename))
                    .otherwise(pl.col("filename"))
                    .alias("filename"),
                    pl.when(pl.col("alias") == defaults.get("alias"))
                    .then(pl.lit(file_config.sheet_name))
                    .otherwise(pl.col("sheet_name"))
                    .alias("sheet_name"),
                ]
            )
        else:
            # Create new entry
            new_entry = ImportLogEntry(
                alias=file_config.alias,
                filename=file_config.filename,
                sheet_name=file_config.sheet_name,
            )

            new_row_df = pl.DataFrame([new_entry])

            if import_log.is_empty():
                import_log = new_row_df
            else:
                import_log = pl.concat([import_log, new_row_df], how="vertical")

        # Save updated log
        duckdb_save_table(project_id, import_log, alias="import_log", db_name="logs")

        st.success(f"File {'updated' if edit_mode else 'added'} successfully!")

    except Exception as e:
        st.error(f"Validation error: {e}")


# --- Optimized Data Loading ---


def load_local_data(
    project_id: str, alias: str, filename: str, sheet_name: str | None = None
) -> None:
    """Load data from local storage with optimized performance.

    Args:
        project_id: Project identifier
        alias: Alias for the data
        filename: Path to the file
        sheet_name: Sheet name for Excel files
    """
    try:
        # Load data efficiently
        data = load_data_efficiently(filename, sheet_name)
        if data.is_empty():
            st.warning(f"No data loaded from {filename}")
            return

        # Save to DuckDB
        duckdb_save_table(project_id, data, alias=alias, db_name="raw")

        st.success(f"Data loaded successfully! Shape: {data.shape}")

    except Exception as e:
        st.error(f"Error loading data: {e}")


# --- Utility Functions ---


def validate_file_accessibility(file_path: Path) -> bool:
    """Check if file is accessible and readable."""
    try:
        path_obj = Path(file_path)
        return (
            path_obj.exists() and path_obj.is_file() and os.access(file_path, os.R_OK)
        )
    except Exception:
        return False


def get_file_info(file_path: Path) -> dict:
    """Get file information for display."""
    try:
        path_obj = Path(file_path)
        stat = path_obj.stat()
        return {
            "size": stat.st_size,
            "modified": stat.st_mtime,
            "extension": path_obj.suffix.lower(),
            "exists": path_obj.exists(),
        }
    except Exception:
        return {"exists": False}
