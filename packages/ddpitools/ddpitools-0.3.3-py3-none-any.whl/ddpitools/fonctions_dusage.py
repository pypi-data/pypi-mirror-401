# Début de section à ne pas modifier à date pour utiliser les fonctionnalités CEGC 
#Import des librairies qui permettent de faire à des fonctions prédéfinies (lire un fichier, fonctions Dataiku, gestion de fichier json...)

import requests
import json
import warnings
import base64
import time
import os
from typing import Optional
import pandas as pd
import numpy as np
from datetime import datetime as dt
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

date_formatee =  dt.now().strftime('%Y-%m-%d')

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

def rgb_to_hex(r, g, b):
    """
        Cette fonction convertie une combinaison RGB en code hexadécimal
        Params:
            r : red, intensité de rouge
            g : green, intensité de vert
            b : blue, intensité de bleu
        returns : une séquence hexadéciamle

        exemple : rgb_to_hex(255, 165, 1)
        résultat : "FFA51"
    """

    return "#0{:X}{:X}{:X}".format(int(r), 
                                int(g), 
                                int(b))


class UploadError(Exception):
    """Custom exception for upload errors."""
    pass

class EXPORTFILE:

    def __init__(self) ->None:
        self.lkp_blue = rgb_to_hex(0, 34, 93)
        self.lkp_green = rgb_to_hex(0, 136, 81)
        self.lkp_magenta = rgb_to_hex(148, 0, 113)
        self.lkp_grey = rgb_to_hex(169, 169, 169).replace("0","")
        self.lkp_comp_blue = rgb_to_hex(0, 113, 148)
        self.lkp_light_blue = rgb_to_hex(35, 95, 221)
        self.lkp_light_green = rgb_to_hex(0, 227, 166)
        self.lkp_purple = "#940071"
        self.date_formatee =  dt.now().strftime('%Y-%m-%d')


    #Fonctions développées par CEGC
    def send_email(self, base_url: str, user_id: str, subject: str, body_content: str,
                to_recipients: list, cc_recipients: list) -> None:
        url = f'{base_url}/send-email'
        headers = {'Content-Type': 'application/json'}
        data = {
            "user_id": user_id,
            "subject": subject,
            "body_content": body_content,
            "toRecipients": to_recipients,
            "ccRecipients": cc_recipients
        }
        response = requests.post(url, headers=headers, data=json.dumps(data), verify=False)
        if response.status_code == 200:
            print("✅ Email sent successfully!")
        else:
            print(f"❌ Failed to send email. Status code: {response.status_code}, Response: {response.text}")


    def validate_parameters(self,*args) -> None:
        if not all(args):
            raise ValueError("All parameters must be provided and non-empty.")


    def upload_large_file(self, url: str, hostname: str, site_name: str, file_name: str,
                        target_folder_path: str, file, file_size: int,
                        initial_chunk_size: int, max_retries: int) -> Optional[dict]:
        num_chunks = (file_size // initial_chunk_size) + (1 if file_size % initial_chunk_size > 0 else 0)
        for chunk_number in range(num_chunks):
            chunk_data = file.read(initial_chunk_size)
            encoded_chunk = base64.b64encode(chunk_data).decode('utf-8')
            payload = {
                "hostname": hostname,
                "site_name": site_name,
                "file_name": file_name,
                "target_folder_path": target_folder_path,
                "file_content": encoded_chunk
            }
            for attempt in range(max_retries):
                try:
                    response = requests.post(url, json=payload,
                                            headers={'Accept': 'application/json', 'Content-Type': 'application/json'},
                                            verify=False, timeout=30)
                    response.raise_for_status()
                    print(f"Uploaded chunk {chunk_number + 1}/{num_chunks}")
                    break
                except requests.exceptions.Timeout:
                    print(f'Timeout for chunk {chunk_number + 1}. Retrying... ({attempt + 1}/{max_retries})')
                    time.sleep(2 ** attempt)
                except requests.exceptions.RequestException as req_err:
                    print(f'Error for chunk {chunk_number + 1}: {req_err}')
                    break
        return response.json() if 'response' in locals() else None


    def upload_small_file(self, url: str, hostname: str, site_name: str, file_name: str,
                        target_folder_path: str, file, max_retries: int) -> Optional[dict]:
        file_content = file.read()
        encoded_content = base64.b64encode(file_content).decode('utf-8')
        payload = {
            "hostname": hostname,
            "site_name": site_name,
            "file_name": file_name,
            "target_folder_path": target_folder_path,
            "file_content": encoded_content
        }
        for attempt in range(max_retries):
            try:
                response = requests.post(url, json=payload,
                                        headers={'Accept': 'application/json', 'Content-Type': 'application/json'},
                                        verify=False, timeout=30)
                response.raise_for_status()
                print("✅ Uploaded small file successfully.")
                return response.json()
            except requests.exceptions.Timeout:
                print(f'Timeout for small file. Retrying... ({attempt + 1}/{max_retries})')
                time.sleep(2 ** attempt)
            except requests.exceptions.RequestException as req_err:
                print(f'Error for small file: {req_err}')
                break
        return None


    def upload_file_in_chunks_base64(self, base_url: str, hostname: str, site_name: str,
                                    file_name: str, target_folder_path: str,
                                    file_path: str, initial_chunk_size: int = 4 * 1024 * 1024,
                                    max_retries: int = 3) -> Optional[dict]:
        self.validate_parameters(hostname, site_name, file_name, target_folder_path, file_path)
        url = f'{base_url}/upload-file-to-sharepoint/'
        file_size = os.path.getsize(file_path)
        is_large_file = file_size > initial_chunk_size



        with open(file_path, 'rb') as file:
            if is_large_file:
                return self.upload_large_file(url, hostname, site_name, file_name,
                                        target_folder_path, file, file_size,
                                        initial_chunk_size, max_retries)
            else:
                return self.upload_small_file(url, hostname, site_name, file_name,
                                        target_folder_path, file, max_retries)


    def download_file(self, base_url: str, hostname: str, site_name: str, file_path: str) -> None:
        url = f'{base_url}/download-file-from-sharepoint/'
        params = {
            "hostname": hostname,
            "site_name": site_name,
            "file_path": file_path
        }

        headers = {
            'Accept': 'application/octet-stream',
            'Content-Type': 'application/json'
        }

        try:
            print(f"→ Trying to download via API Gateway: {url}")
            response = requests.get(url, params=params, headers=headers, verify=False)

            if response.status_code == 200:
                local_path = f"/tmp/{os.path.basename(file_path)}"
                with open(local_path, 'wb') as f:
                    f.write(response.content)
                print(f"✅ File downloaded successfully: {local_path}")
                return

            # Si blocage Axway détecté
            if ("MessageBlocked" in response.text or
                "soapfaults" in response.text or
                response.status_code in [400, 403, 405]):

                print("⚠️  Detected Axway block. Retrying directly on backend...")

                direct_base_url = "https://ard.bench.mycloud.intranatixis.com"
                direct_url = f'{direct_base_url}/download-file-from-sharepoint/'

                response2 = requests.get(direct_url, params=params, headers=headers, verify=False)

                if response2.status_code == 200:
                    local_path = f"/tmp/{os.path.basename(file_path)}"
                    with open(local_path, 'wb') as f:
                        f.write(response2.content)
                    print(f"✅ File downloaded successfully (direct backend): {local_path}")
                else:
                    print(f"❌ Download failed on backend too: {response2.status_code} - {response2.text}")

            else:
                print(f"❌ Error downloading file: {response.status_code} - {response.text}")

        except requests.exceptions.RequestException as e:
            print(f"⚠️ Exception occurred: {e}")


    def colnum_string(self, n):                                                                                                   
        string = ''                                                                                                         
        while n > 0:                                                                                                        
            n, remainder = divmod(n - 1, 26)                                                                                
            string       = chr(65 + remainder) + string                                                                     
        return string


    def histo_date(self):
        # Dictionnaire qui contiendra tous nos résultats
        dates_resultat = {}
        today = dt.date.today()
        dernier_jour_mois_precedent = today.replace(day=1) - dt.timedelta(days=1)

        # Ajout au dictionnaire au format souhaité
        dates_resultat['histo_courante'] = dernier_jour_mois_precedent.strftime('%Y-%m-%d')

        for i in range(1, 6):
            annee_cible = today.year - i
            dates_resultat['histo_fin_{}'.format(annee_cible)] = dt.date(annee_cible, 12, 31).strftime('%Y-%m-%d')
            dates_resultat['histo_mi_{}'.format(annee_cible)] = date_fin_juin = dt.date(annee_cible, 6, 30).strftime('%Y-%m-%d')

        return dates_resultat
    

    def formater_feuille(self, ws, df):
        """
        Applique une mise en forme standard à une feuille Excel contenant un DataFrame.
        - ws: L'objet worksheet d'openpyxl.
        - df: Le DataFrame pandas original (utilisé pour l'ajustement des colonnes).
        """
        # Définition des styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='00b377', end_color='00b377', fill_type='solid') #004C99
        header_align = Alignment(horizontal='center', vertical='center')
        
        date_format = 'DD/MM/YYYY'
        currency_format = '#,##0.00€'
        
        # Mise en forme des en-têtes (ligne 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Mise en forme des colonnes (données) et ajustement de la largeur
        for col_idx, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # Ajustement de la largeur
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 3
            ws.column_dimensions[col_letter].width = max_length
            
            # Formats spécifiques par type de colonne
            if 'date' in column.lower():
                for cell in ws[col_letter][1:]: # [1:] pour ignorer l'en-tête
                    cell.number_format = date_format
            elif 'montant' in column.lower():
                for cell in ws[col_letter][1:]:
                    cell.number_format = currency_format
    
    
    def export_table(self,nom_fichier,**tables_avec_noms):

        file_path = "/tmp/{}_{}{}{}.xlsx".format(nom_fichier, dt.now().year, dt.now().month, dt.now().day)
        with pd.ExcelWriter(file_path, engine='openpyxl', date_format='YYYY-MM-DD') as writer:
            for nom_feuille, table in tables_avec_noms.items():
                if not isinstance(table, pd.DataFrame):
                    print(f"  - Avertissement : L'élément '{nom_feuille}' n'est pas un DataFrame, il sera ignoré.")
                    continue
                print(f"  - Écriture de la feuille : '{nom_feuille}'...")

                table.to_excel(writer, sheet_name=nom_feuille, index=False)
                ws = writer.sheets[nom_feuille]
                self.formater_feuille(ws, table)
            print("Exportation terminée avec succès.")
        return file_path

