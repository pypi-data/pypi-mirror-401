import copy
import logging
import math
import os
import re
from datetime import datetime, timezone
from inspect import getsourcefile

from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PIL_Image
from PIL import UnidentifiedImageError
from rich.progress import track

from artemis_sg import app_creds, item, items
from artemis_sg.config import CFG
from artemis_sg.isbn import validate_isbn
from artemis_sg.items import Items
from artemis_sg.vendor import vendor

MODULE = os.path.splitext(os.path.basename(__file__))[0]


def get_worksheet(wb_obj: Workbook, worksheet: str | None = None) -> Worksheet:
    """
    Get openypyxl Worksheet object matching name of given `worksheet`.
    Will return first Worksheet in Workbook if lookup `worksheet` string is not given.

    :param wb_obj: openpyxl Workbook
    :param worksheet: name of worksheet to return
    """
    ws = wb_obj.worksheets[0] if not worksheet else wb_obj[worksheet]
    return ws


def get_isbns_from_sheet(vendor_code, workbook, worksheet) -> list:
    vendr = vendor(vendor_code)
    isbn_key = vendr.isbn_key
    sheet_data = get_sheet_data(workbook, worksheet)
    sheet_keys = [x for x in sheet_data.pop(0) if x]
    items = Items(sheet_keys, sheet_data, isbn_key, vendr.vendor_specific_id_col)
    isbns = [item.isbn for item in items]
    return isbns


def get_sheet_keys(ws: Worksheet) -> list[str]:
    """
    Get list of column headers from worksheet

    :param ws: openpyxl Worksheet
    :returns: list of column headers in uppercase
    """
    headers = next(ws.values)
    return [x.strip().upper() if isinstance(x, str) else x for x in headers]


def shift_col(ws: Worksheet, col_key: str, target_idx: int) -> None:
    """
    Move worksheet column with given `col_key` to `target_idx` column
    and shift remaining column over to accomodate it.

    Example:
    | GIVEN worksheet columns: ('foo', 'bar', 'ISBN-13', 'baz', 'Order')
    | AND col_key = 'ISBN-13'
    | AND target_idx = 1
    | THEN new columns are: ('ISBN-13', 'foo', 'bar', 'baz', 'Order')

    :param ws: openpyxl Worksheet
    :param col_key: column header key to move
    :param target_idx: column position to move to (1 indexed)
    """
    ws.insert_cols(target_idx)
    sheet_keys = get_sheet_keys(ws)
    sheet_key_idx = sheet_keys.index(col_key) + 1  # for openpyxl
    sheet_key_idx_ltr = get_column_letter(sheet_key_idx)
    col_delta = target_idx - sheet_key_idx
    ws.move_range(
        f"{sheet_key_idx_ltr}1:{sheet_key_idx_ltr}{ws.max_row}", rows=0, cols=col_delta
    )
    ws.delete_cols(sheet_key_idx)


def format_headers(worksheet: Worksheet) -> None:
    """
    Sets the font and highlight color of headers to be the same as first cell's.
    """
    headers = worksheet[1]
    cell_font = Font(name=headers[0].font.name, bold=headers[0].font.bold)
    cell_color = PatternFill(
        patternType=headers[0].fill.patternType,
        fgColor=headers[0].fill.fgColor,
    )
    for cell in worksheet[1]:
        cell.font = cell_font
        cell.fill = cell_color


def freeze_first_row(worksheet: Worksheet) -> None:
    """
    Freeze first row of `worksheet`

    :param ws: openpyxl Worksheet
    """
    logging.info("Freezing first row of worksheet.")
    worksheet.views.sheetView[0].topLeftCell = "A1"
    worksheet.freeze_panes = "A2"


def create_col(ws: Worksheet, col_key: str, target_idx: int) -> None:
    """
    Create worksheet column with given `col_key` at `target_idx` column.

    :param ws: openpyxl Worksheet
    :param col_key: column header key to move
    :param target_idx: column position to move to (1 indexed)
    """
    ws.insert_cols(target_idx)
    col_header = f"{get_column_letter(target_idx)}1"
    ws[col_header] = col_key.title()


def sequence_worksheet(ws: Worksheet, col_order: list[str], isbn_key: str) -> None:
    """
    Reorder the columns of the worksheet `ws` in accordance with the given
    `col_order`.  Columns not defined in the `col_order` will be appended
    to the end of the worksheet columns.

    :param ws: openpyxl Worksheet
    :param col_order: list of column keys in order of preference
    :param isbn_key: the key in `col_order` that corresponds with ISBN
    """
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(col_order):
        order_idx = i + 1  # for openpyxl
        if key_name == "ISBN":
            key_name = isbn_key  # noqa: PLW2901
        if key_name in sheet_keys:
            shift_col(ws, key_name, order_idx)
        else:
            create_col(ws, key_name, order_idx)


def size_sheet_cols(
    ws: Worksheet,
    isbn_key: str,
    fmt_opts: dict = CFG["asg"]["spreadsheet"]["sheet_image"],
) -> None:
    """
    Resize the width of the `ws` worksheet columns in accordance with the
    preferences defined in `fmt_opts`.

    :param ws: openpyxl Worksheet
    :param isbn_key: the key worksheet column headers that corresponds with ISBN
    :param fmt_opts: dictionary of formatting options
    """
    fmt_opts.setdefault("col_buffer", 1.25)
    fmt_opts.setdefault("max_col_width", 50)
    fmt_opts.setdefault("isbn_col_width", 13)
    fmt_opts.setdefault("image_col_width", 20)
    dim_holder = DimensionHolder(worksheet=ws)
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(sheet_keys):
        col_idx = i + 1  # for openpyxl
        col_idx_ltr = get_column_letter(col_idx)
        width = (
            max(len(str(cell.value)) for cell in ws[col_idx_ltr])
            * fmt_opts["col_buffer"]
        )
        width = min(width, fmt_opts["max_col_width"])
        dim_holder[col_idx_ltr] = ColumnDimension(ws, index=col_idx_ltr, width=width)
        if key_name == isbn_key:
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=math.ceil(fmt_opts["isbn_col_width"] * fmt_opts["col_buffer"]),
            )
        if key_name == "IMAGE":
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=fmt_opts["image_col_width"],
            )

    ws.column_dimensions = dim_holder


def insert_image(
    image_directory: str,
    ws: Worksheet,
    isbn: str,
    image_cell: Cell,
    image_row_height: int = CFG["asg"]["spreadsheet"]["sheet_image"][
        "image_row_height"
    ],
) -> None:
    """
    Insert an image into a Worksheet cell associated with an ISBN

    :param image_directory: path to the directory where item images are stored
    :param ws: openpyxl Worksheet
    :param isbn: ISBN number of the item
    :param image_cell: `Cell` object that image should be inserted into
    :param image_row_height: Height to set row containing image
    """
    namespace = f"{MODULE}.{insert_image.__name__}"
    # Set row height
    row_dim = ws.row_dimensions[image_cell.row]
    row_dim.height = image_row_height

    # Insert image into cell
    filename = f"{isbn}.jpg"
    filepath = os.path.join(image_directory, filename)
    logging.debug(f"{namespace}: Attempting to insert '{filepath}'.")
    if os.path.isfile(filepath):
        img = Image(filepath)
        ws.add_image(img, f"{image_cell.column_letter}{image_cell.row}")
        logging.info(f"{namespace}: Inserted '{filepath}'.")


def add_image_to_row(
    ws, items, image_directory, isbn_cell, image_cell, vendor_code_cell=None
):
    # Format to center content
    item = items.find_item(isbn_cell.value) or items.find_item(
        str(vendor_code_cell.value)
    )
    image_cell.alignment = Alignment(horizontal="center")
    if item:
        isbn = item.isbn
        insert_image(image_directory, ws, isbn, image_cell)


def get_datetime_obj(date_str):
    """
    >>> get_datetime_obj("01/14/2003")
    datetime.datetime(2003, 1, 14, 0, 0, tzinfo=datetime.timezone.utc)
    >>> get_datetime_obj("14/01/2003")
    datetime.datetime(2003, 1, 14, 0, 0, tzinfo=datetime.timezone.utc)
    >>> get_datetime_obj("14/14/2003")
    ''
    """
    us_format = "%m/%d/%Y"
    eu_format = "%d/%m/%Y"
    if isinstance(date_str, datetime):
        return date_str.replace(tzinfo=timezone.utc)
    try:
        date_obj = datetime.strptime(date_str, us_format).replace(tzinfo=timezone.utc)
    except ValueError:
        try:
            date_obj = datetime.strptime(date_str, eu_format).replace(
                tzinfo=timezone.utc
            )
        except ValueError:
            return ""
    return date_obj


def sort_rows_by_cutoff_date(vendor_code, ws, isbn_key, cutoff_date):
    """
    Moves all rows after the cutoff date to the top of the sheet.
    """
    sheet_keys = get_sheet_keys(ws)
    isbn_idx = sheet_keys.index(isbn_key.strip().upper())

    rows = (list(ws.iter_rows()))[1:]
    ws.delete_rows(2, len(rows))
    try:
        date_col_name = CFG["asg"]["spreadsheet"]["sheet_image"][
            "vendor_date_col_names"
        ][vendor_code]
        date_idx = sheet_keys.index(date_col_name.strip().upper())
    except IndexError as e:
        err = f"Cannot find date column for vendor '{vendor_code}' in config"
        raise IndexError(err) from e
    except ValueError as e:
        err = f"Cannot find date column '{date_col_name}' in headers"
        raise ValueError(err) from e

    sheet_data = [[cell.value for cell in row] for row in rows]
    isbns_and_dates = [(row[isbn_idx], row[date_idx]) for row in sheet_data]
    isbns_after, isbns_before = get_isbns_after_cutoff_date(
        isbns_and_dates, cutoff_date
    )
    row_idx = 2
    for isbn_list in (isbns_after, isbns_before):
        for isbn in isbn_list:
            for row in rows:
                if row[isbn_idx].value == isbn:
                    for col_idx, cell in enumerate(row):
                        c = ws.cell(row=row_idx, column=col_idx + 1, value=cell.value)
                        c.number_format = cell.number_format
                        c.font = copy.copy(cell.font)
                        c.fill = copy.copy(cell.fill)
                        c.border = copy.copy(cell.border)
                        c.alignment = copy.copy(cell.alignment)
                    row_idx += 1
                    break
    return isbns_after


def get_isbns_after_cutoff_date(data, cutoff_date):
    """
    Get all isbn's that have a recieve date on or after the cutoff date
    :param data: list of (isbn, date) tuples
    :param cutoff_data: str
    :returns isbns after cutoff and isbns before cutoff date.

    """
    cutoff_date_obj = get_datetime_obj(cutoff_date)
    if not cutoff_date_obj:
        err = (f"Unable to parse date: {cutoff_date}", "Format as 'DD/MM/YYYY'")
        raise ValueError(err)
    isbns_after_cutoff = []
    isbns_before_cutoff = []
    for isbn, date in data:
        date_obj = get_datetime_obj(date)
        if date_obj and date_obj >= cutoff_date_obj:
            isbns_after_cutoff.append(isbn)
        else:
            isbns_before_cutoff.append(isbn)

    return isbns_after_cutoff, isbns_before_cutoff


def add_rank_to_row(ws, items, rank_cols, isbn_cell, vendr_code_cell=None):
    """
    Adds an item's rank data to the row.
    :param ws: worksheet
    :param rank_cols: list of keys in item.data["RANK"]
    :param isbn_cell: openpxyl cell to get the isbn and row from
    :param vendr_code_cell: openpxyl cell to use instead of isbn_cell if needed
    """
    item = (
        items.find_item(isbn_cell.value) or items.find_item(str(vendr_code_cell.value))
        if vendr_code_cell
        else None
    )
    headers = get_sheet_keys(ws)
    if item:
        for rank_type in rank_cols:
            add_to_row(
                headers,
                ws[isbn_cell.row],
                f"{rank_type} RANK",
                item.data.get("RANKS", {}).get(rank_type),
            )


def add_rank_cols(ws, vendr):
    """
    Adds rank column headers to a spreadsheet.
    :param ws: spreadsheet
    :param vendr: vendor
    :param insert_idx: index to add rank columns
    :return rank_cols: list of the item keys for each rank column.
    """
    col_insert_idx = ws.max_column + 1
    if vendr.failover_scraper == "AmznUkScraper":
        rank_cols = ["AMAZON", "AMAZON_UK"]
    else:
        rank_cols = ["AMAZON"]
    ws.insert_cols(col_insert_idx, len(rank_cols))
    for cnt, rank_type in enumerate(rank_cols):
        ws.cell(row=1, column=col_insert_idx + cnt).value = f"{rank_type} Rank"
    return rank_cols


def sheet_image(
    vendor_code: str,
    workbook: str,
    worksheet: str,
    image_directory: str,
    options: dict,
    col_order: list[str] = CFG["asg"]["spreadsheet"]["sheet_image"]["col_order"],
) -> Worksheet:
    """
    Create a new spreadsheet file based on a given `workbook`/`worksheet`
    with an added column for a product image added for each ISBN in the `worksheet`.

    Images are obtained from the given `image_directory` and are expected to be named
    in the the pattern `{isbn}.jpg`.

    The spreadsheet columns are reordered in accordance with the given
    `col_order` parameter.

    The new spreadsheet is saved with the filename(path) given by the `out` parameter.

    :param vendor_code: Vendor code used to lookup isbn_key (See `vendor`)
    :param workbook: path to Excel workbook file
    :param worksheet: name of worksheet in workbook file
    :param image_directory: path to the directory where item images are stored
    :param options
        :options["out"]: path where the spreadsheet will be saved.
        options["cutoff_date"]: only add images for titles after this date if specified
    :param col_order: list of column keys in order of preference
    :returns: Worksheet updated with item images
    """
    namespace = f"{MODULE}.{sheet_image.__name__}"
    scraped_items_db = CFG["asg"]["data"]["file"]["scraped"]
    out = options.get("out")
    cutoff_date = options.get("cutoff_date")
    vendr = vendor(vendor_code)
    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")
    rank_cols = add_rank_cols(ws, vendr)
    format_headers(ws)
    sequence_worksheet(ws, col_order, isbn_key)
    isbns_after_cutoff = (
        []
        if not cutoff_date
        else sort_rows_by_cutoff_date(vendor_code, ws, isbn_key, cutoff_date)
    )
    size_sheet_cols(ws, isbn_key)

    sheet_data = get_sheet_data(workbook, worksheet)
    sheet_keys = [x for x in sheet_data.pop(0) if x]
    items = Items(sheet_keys, sheet_data, isbn_key, vendr.vendor_specific_id_col)
    items.load_scraped_data(scraped_items_db)
    sk = get_sheet_keys(ws)
    try:
        img_idx = sk.index("IMAGE") + 1
        img_idx_ltr = get_column_letter(img_idx)
    except ValueError as e:
        logging.error(f"{namespace}: Err finding 'IMAGE' column in sheet '{workbook}'.")
        logging.error("Aborting.")
        raise e
    try:
        isbn_idx = sk.index(isbn_key) + 1
        isbn_idx_ltr = get_column_letter(isbn_idx)
    except ValueError as e:
        logging.error(
            f"{namespace}: Err, no '{isbn_key}' column in sheet '{workbook}'."
        )
        logging.error("Aborting.")
        raise e
    try:
        vendr_code_idx = sk.index(vendr.vendor_specific_id_col.strip().upper()) + 1
        vendr_code_idx_ltr = get_column_letter(vendr_code_idx)
    except ValueError:
        vendr_code_idx_ltr = isbn_idx_ltr
    for i in range(1, ws.max_row):
        isbn_cell = ws[f"{isbn_idx_ltr}{i + 1}"]
        image_cell = ws[f"{img_idx_ltr}{i + 1}"]
        vendr_code_cell = ws[f"{vendr_code_idx_ltr}{i + 1}"]
        if not cutoff_date or isbn_cell.value in isbns_after_cutoff:
            add_image_to_row(
                ws, items, image_directory, isbn_cell, image_cell, vendr_code_cell
            )
            add_rank_to_row(
                ws,
                items,
                rank_cols,
                isbn_cell,
                vendr_code_cell,
            )
    freeze_first_row(ws)
    format_headers(ws)
    # Save workbook
    wb.save(out)

    return ws


def validate_qty(qty: str | int | float) -> str:
    """Cast given value into a string integer or None if unable"""
    namespace = f"{MODULE}.{validate_qty.__name__}"
    try:
        valid_qty = str(int(qty)).strip()
    except Exception as e:
        logging.error(f"{namespace}: Err reading Order qty '{qty}', err: '{e}'")
        valid_qty = None
    return valid_qty


def get_order_items(
    vendr: vendor, workbook: str, worksheet: str
) -> list[tuple[str, str]]:
    """
    Get a list of (ISBN, Quantity) pairs for items from the "Order" column of
    the given `workbook`/`worksheet`.

    :param vendr: vendor object
    :param workbook: path to Excel workbook file
    :param worksheet: name of worksheet in workbook file
    """
    namespace = f"{MODULE}.{get_order_items.__name__}"

    order_items = []
    try:
        order_col = CFG["asg"]["spreadsheet"]["order"]["order_col"].upper()
    except AttributeError:
        logging.error(f"{namespace}: No order column set in config.toml")
        order_col = ""

    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")
    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")
    sheet_keys = get_sheet_keys(ws)
    isbn_key_idx = sheet_keys.index(isbn_key)
    order_key_idx = sheet_keys.index(order_col)

    try:
        vendor_specific_id_col_name = vendr.vendor_specific_id_col
        try:
            vendor_specific_id_col_idx = sheet_keys.index(
                vendor_specific_id_col_name.strip().upper()
            )
        except ValueError:
            logging.debug(f"{vendor_specific_id_col_name} column not in spreadsheet")
            vendor_specific_id_col_idx = isbn_key_idx
    except AttributeError:
        logging.debug("No vendor specific id column")
        vendor_specific_id_col_idx = isbn_key_idx

    for row in ws.iter_rows(min_row=2):
        # Validate ISBN
        isbn = validate_isbn(row[isbn_key_idx].value)

        try:
            product_id = row[vendor_specific_id_col_idx].value
        except IndexError:
            product_id = isbn

        qty = row[order_key_idx].value

        isbn = isbn or product_id
        order_amt = row[order_key_idx].value
        # Validate Order Qty
        qty = validate_qty(order_amt)
        if not isbn or not qty:
            continue
        order_items.append((isbn, qty))

    return order_items


def _mkthumb(
    big_file: str,
    thumb_file: str,
    thumb_width: str = CFG["asg"]["spreadsheet"]["mkthumbs"]["width"],
    thumb_height: str = CFG["asg"]["spreadsheet"]["mkthumbs"]["height"],
) -> None:
    """
    Create an image file of `big_file` at path `thumb_file` of
    dimensions (`thumb_width`, `thumb_height`).

    :param big_file: path of image file to create thumbnail of
    :param thumb_file: path where image file of thumbnail should be saved
    :thumb_width: width (in pixels) that thumbnail should be
    :thumb_height: height (in pixels) that thumbnail should be
    """
    namespace = f"{MODULE}.{_mkthumb.__name__}"

    # validate big_file, delete if invalid
    try:
        fg = PIL_Image.open(big_file)
    except UnidentifiedImageError:
        logging.error(f"{namespace}: Err reading '{big_file}', deleting it.")
        os.remove(big_file)
        return
    # don't remake thumbnails
    if os.path.isfile(thumb_file):
        return
    here = os.path.dirname(getsourcefile(lambda: 0))
    data = os.path.abspath(os.path.join(here, "data"))
    logo = os.path.join(data, "artemis_logo.png")
    logging.debug(f"{namespace}: Found image for thumbnail background at '{logo}'")
    back = PIL_Image.open(logo)
    bk = back.copy()
    fg.thumbnail((thumb_width, thumb_height))
    size = (int((bk.size[0] - fg.size[0]) / 2), int((bk.size[1] - fg.size[1]) / 2))
    bk.paste(fg, size)
    logging.debug(f"{namespace}: Attempting to save thumbnail '{thumb_file}'")
    bkn = bk.convert("RGB")
    bkn.save(thumb_file)
    logging.info(f"{namespace}: Successfully created thumbnail '{thumb_file}'")


def mkthumbs(image_directory: str) -> None:
    """
    Create thumbnail image files of all images in given `image_directory`

    :param image_directory: path location of image files
    """
    namespace = f"{MODULE}.{mkthumbs.__name__}"

    sub_dir = "thumbnails"
    thumb_dir = os.path.join(image_directory, sub_dir)
    logging.debug(f"{namespace}: Defining thumbnail directory as '{thumb_dir}'")
    if not os.path.isdir(thumb_dir):
        logging.debug(f"{namespace}: Creating directory '{thumb_dir}'")
        os.mkdir(thumb_dir)
        if os.path.isdir(thumb_dir):
            logging.info(f"{namespace}: Successfully created directory '{thumb_dir}'")
        else:
            logging.error(
                f"{namespace}: Failed to create directory '{thumb_dir}'. Aborting."
            )
            raise Exception
    files = os.listdir(image_directory)

    for f in track(files, description="Creating thumbnails..."):
        # Valid files are JPG or PNG that are not supplemental images.
        image = re.match(r"^.+\.(?:jpg|png)$", f)
        if not image:
            continue
        # Supplemental images have a "-[0-9]+" suffix before the file type.
        # AND a file without that suffix exists int he image_directory.
        suffix = re.match(r"(^.+)-[0-9]+(\.(?:jpg|png))$", f)
        if suffix:
            primary = suffix.group(1) + suffix.group(2)
            primary_path = os.path.join(image_directory, primary)
            if os.path.isfile(primary_path):
                continue
        thumb_file = os.path.join(thumb_dir, f)
        big_file = os.path.join(image_directory, f)
        _mkthumb(big_file, thumb_file)


def get_sheet_data(workbook: str, worksheet: str | None = None) -> list[tuple[any]]:
    """
    Get data from a given `workbook`/`worksheet`.

    :param workbook: path to Excel workbook file
    :param worksheet: name of worksheet in workbook file, first worksheet if None
    :returns: The data is returned as a list of tuples where the first item in
    the list contains the keys for the data and the subsequent tuples contain
    the values per row corresponding to those keys.
    """
    namespace = f"{MODULE}.{get_sheet_data.__name__}"
    #########################################################################
    # Try to open sheet_id as an Excel file
    sheet_data = []
    try:
        wb = load_workbook(workbook)
        ws = get_worksheet(wb, worksheet)
        for row in ws:
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.value = datetime.strftime(cell.value, "%m/%d/%Y")
            sheet_data.append([cell.value for cell in row])
    except (FileNotFoundError, InvalidFileException):
        #########################################################################
        # Google specific stuff
        # authenticate to google sheets
        logging.info(f"{namespace}: Authenticating to google api.")
        creds = app_creds.app_creds()
        sheets_api = build("sheets", "v4", credentials=creds)
        # get sheet data
        if not worksheet:
            sheets = (
                sheets_api.spreadsheets()
                .get(spreadsheetId=workbook)
                .execute()
                .get("sheets", "")
            )
            ws = sheets.pop(0).get("properties", {}).get("title")
        else:
            ws = worksheet
        sheet_data = (
            sheets_api.spreadsheets()
            .values()
            .get(range=ws, spreadsheetId=workbook)
            .execute()
            .get("values")
        )
        #########################################################################
    return sheet_data


def convert_dimensions(dimension: str) -> tuple[str, str, str]:
    """
    Converts string dimension into width, length, & height.
    Also converts cm to inches.

    Only works for dimensions scraped for Amazon or AmazonUK

    Example Dimensions:

    "9.84 x 10.63 inches" (width x height)
    >>> convert_dimensions("9.84 x 10.63 inches")
    ('9.84', '', '10.63')

    "1.0 x 2.54 cm" cm to inch conversion
    >>> convert_dimensions("1.0 x 2.54 cm")
    ('0.39', '', '1.0')

    "5.28 x 0.87 x 7.95 inches" (width x length x height)
    >>> convert_dimensions("5.28 x 0.87 x 7.95 inches")
    ('5.28', '0.87', '7.95')

    "" (no dimension was found)
    >>> convert_dimensions("")
    ('', '', '')

    :param: dimension string value
    :returns: (width, length, height)
    """
    width, length, height = "", "", ""
    standard_dimension_count = 3
    needs_unit_conversion = False
    if re.search(r"cm", dimension):  # dimensions in cm
        dimension = re.sub(r"\s*cm\s*", "", dimension)
        needs_unit_conversion = True
    if re.search(r"inches", dimension):  # dimension in inches
        dimension = re.sub(r"\s*inches\s*", "", dimension)  # remove measure unit
    parsed_dimensions = re.split(r" x ", dimension)
    if needs_unit_conversion:  # convert cm to inches
        for idx, dim in enumerate(parsed_dimensions):
            parsed_dimensions[idx] = str(round((float(dim) / 2.54), 2))
    if len(parsed_dimensions) > 1:
        if len(parsed_dimensions) < standard_dimension_count:  # 2 dimensions, else 3
            width = parsed_dimensions[0]
            height = parsed_dimensions[1]
        else:
            width = parsed_dimensions[0]
            length = parsed_dimensions[1]
            height = parsed_dimensions[2]
    return width, length, height


def get_percentage_discount_prices(
    price: str | int | float, *discount_percentages: str
) -> dict[str, float]:
    """
    Construct a dictionary of discounted prices on a given price
    for any number of given discount percentages.

    Example:

    >>> get_percentage_discount_prices(10.00, "10%", "50%")
    {'10%': 9.0, '50%': 5.0}

    :param price: price to apply discounts to
    :param discount_percentages: discounts expressed in a form like "50%"
    :returns: dictionary of discounted prices keyed by percentage
    """
    price = float(price)
    discount_percent_dict = {}
    for discount in discount_percentages:
        f_disc = float(discount.strip("%")) / 100
        discount_percent_dict[discount] = price - (price * f_disc)
    return discount_percent_dict


def waves_set_pound_price(
    headers: list[str],
    rrp,
    book_format,
    row: tuple[Cell, ...],
    map_scheme: dict = CFG["asg"]["spreadsheet"]["sheet_waves"]["pound_pricing_map"],
    unmapped_multiplier: float = CFG["asg"]["spreadsheet"]["sheet_waves"][
        "pound_pricing_unmapped_multiplier"
    ],
) -> None:
    """
    Set 'Pound Pricing' (£) price in appropriate cell of openpyxl spreadsheet row.

    :param headers: list of spreadsheet column headers
    :param item: item.Item object that may have RRP set,
    :param row: openpyxl row of spreadsheet
    :param map_scheme: dictionary of pricing schemes
    :param unmapped_multiplier: default pricing multiplier
    """
    if rrp and "POUND PRICING" in headers:
        try:
            rrp = float(rrp)
        except ValueError:
            return
        pounds, pence = f"{rrp:.2f}".split(".")
        try:
            pound_price = map_scheme[book_format][pounds][pence]
        except KeyError:
            pound_price = rrp * unmapped_multiplier
        pound_price = f"{round(pound_price, 2):.2f}"
        add_to_row(headers, row, "POUND PRICING", pound_price)


def waves_calculate_fields(headers: list[str], row: tuple[Cell], fields: dict) -> None:
    """
    Empty cell values match with key: ""

    Any cell value matches with key: "any"

    :param headers: list of spreadsheet column keys, upper_cased.
    :param row: openpyxl row of spreadsheet
    :param fields: A nested dictionary with the following structure:
        | First level: Name of column/field to change for a given row.
        | Second level:
        |     map_from:
        |         Names of columns to access values for a given row.
        |         These values are used as keys, for indexing into map.
        |     map:
        |         Nested dict.
    """
    for field in list(fields.keys()):
        map_from = fields[field]["map_from"]
        field_map = fields[field]["map"]
        if isinstance(map_from, str):
            map_from = [map_from.strip().upper()]
        if isinstance(map_from, list):
            map_from = [val.strip().upper() for val in map_from]
        if field.strip().upper() in headers:
            current_level = field_map
            for idx, field_key in enumerate(map_from):
                try:
                    field_val = row[headers.index(field_key.strip().upper())].value
                    field_val = "" if field_val is None else field_val
                    if field_val not in current_level:
                        field_val = "any"
                    if idx == len(map_from) - 1:
                        row[headers.index(field.strip().upper())].value = current_level[
                            field_val
                        ]
                    else:
                        current_level = current_level[field_val]
                except (KeyError, ValueError):
                    break


def add_to_row(headers, row, key, value):
    """
    If the key exists in headers, it adds the value at that index.
    """
    if key in headers:
        row[headers.index(key)].value = value


def rename_columns(
    ws: Worksheet,
    headers: list[str],
    rename_fields: dict = CFG["asg"]["spreadsheet"]["sheet_waves"]["rename_fields"],
) -> None:
    """
    Rename `worksheet` columns given mapping in `rename_fields`.

    :param ws: openpyxl worksheet to update
    :param headers: list of column headers
    :rename_fields: dictionary mapping old_name to new_name column names
    """
    for old_name, new_name in rename_fields.items():
        if old_name in headers:
            col_idx = headers.index(old_name) + 1
            ws.cell(row=1, column=col_idx, value=new_name)


def add_waves_row_data(
    headers: list[str],
    item: item.Item,
    row: tuple[Cell, ...],
    gbp_to_usd: float | None,
    **kwargs: any,
) -> None:
    """
    Adds width, height, length, discounts,
    description, images, and preset fields to a row.

    :param headers: list of spreadsheet column headers
    :param item: item.Item object that may have RRP set,
    :param row: openpyxl row of spreadsheet
    :param gbp_to_usd: conversion rate
    :param **kwargs: various sheet_waves key value pairings
    """

    if item:
        # convert dimension to w, h, l
        width, length, height = convert_dimensions(item.data["DIMENSION"])
        # set new dimension columns
        add_to_row(headers, row, "HEIGHT", height)
        add_to_row(headers, row, "WIDTH", width)
        add_to_row(headers, row, "LENGTH", length)
        for idx, img in enumerate(kwargs["image_columns"]):  # add image urls
            try:
                add_to_row(
                    headers,
                    row,
                    img.strip().upper(),
                    item.image_urls[idx],
                )
            except IndexError:
                break
        add_to_row(headers, row, "DESCRIPTION", item.data.get("DESCRIPTION"))
        rrp = item.data.get("RRP")
        book_format = item.data.get("FORMAT")
    else:
        rrp = row[headers.index("RRP")].value if "RRP" in headers else None
        book_format = (
            row[headers.index("FORMAT")].value if "FORMAT" in headers else None
        )

    if rrp and gbp_to_usd:
        discount_prices = get_percentage_discount_prices(
            rrp, *kwargs["discounted_prices"]
        )
        for disc, val in discount_prices.items():
            col_key = kwargs["discount_text_map"].format(t=disc).strip().upper()
            disc_price = val * gbp_to_usd
            add_to_row(headers, row, col_key, f"{round(disc_price, 2):.2f}")

    waves_set_pound_price(headers, rrp, book_format, row)

    for field, val in kwargs["preset_fields"].items():
        add_to_row(headers, row, field.strip().upper(), val)
    fields = kwargs["calculate_fields"]
    waves_calculate_fields(headers, row, fields)


def waves_update_from_inventory(vendor_code, isbn_key, ws, inventory_filepath):
    namespace = f"{MODULE}.{sheet_waves.__name__}"
    ws_headers = get_sheet_keys(ws)
    if not inventory_filepath:  # set product_active to 1
        for row in ws.iter_rows(min_row=2):
            add_to_row(ws_headers, row, "PRODUCT_ACTIVE", 1)
        return

    ws_isnb_idx = ws_headers.index(isbn_key) + 1
    ws_product_active_idx = (
        ws_headers.index("PRODUCT_ACTIVE") + 1
        if "PRODUCT_ACTIVE" in ws_headers
        else None
    )
    inventory_workbook = load_workbook(inventory_filepath)
    inventory_ws = get_worksheet(inventory_workbook)
    try:
        vendor_qty_col_names = CFG["asg"]["spreadsheet"]["sheet_waves"][
            "qty_col_names"
        ][vendor_code]
    except KeyError as e:
        err = (
            f"'{vendor_code}' not found in config.toml under "
            "[asg][spreadsheet][sheet_waves][qty_col_names]. "
            "Cannot update quantity."
        )
        raise ValueError(err) from e

    inventory_qty_col_name = vendor_qty_col_names["inventory"].strip().upper()
    ws_qty_col_name = vendor_qty_col_names["workbook"].strip().upper()
    inventory_qty_idx = (
        get_sheet_keys(inventory_ws).index(inventory_qty_col_name.upper()) + 1
    )
    try:
        inventory_ws_isbn_idx = get_sheet_keys(inventory_ws).index(isbn_key.upper()) + 1
    except ValueError as e:
        err = f"{namespace}: ISBN key '{isbn_key}' not found in: {inventory_filepath}"
        raise ValueError(err) from e

    inventory_ws_isbn_column = [
        cell.value
        for row in inventory_ws.iter_rows(
            min_row=2, min_col=inventory_ws_isbn_idx, max_col=inventory_ws_isbn_idx
        )
        for cell in row
    ]

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        isbn = ws.cell(row=idx, column=ws_isnb_idx).value
        try:
            matched_isbn = inventory_ws_isbn_column.index(isbn)
            new_qty = inventory_ws.cell(
                row=matched_isbn + 2, column=inventory_qty_idx
            ).value
            new_qty = int(new_qty)
        except (ValueError, TypeError):
            new_qty = 0
        add_to_row(ws_headers, row, ws_qty_col_name, new_qty)
        # set product_active
        if int(new_qty) > 0 and ws_product_active_idx:
            add_to_row(ws_headers, row, "PRODUCT_ACTIVE", 1)
        elif ws_product_active_idx:
            add_to_row(ws_headers, row, "PRODUCT_ACTIVE", None)


def sheet_waves(
    vendor_code: str,
    workbook: str,
    worksheet: str,
    scraped_items_db: str,
    options: dict,
    **kwargs: any,
) -> Worksheet:
    """
    Create a new spreadsheet file based on a given `workbook`/`worksheet`
    with added columns and data based on waves preferences.

    The new spreadsheet is saved with the filename(path) given by the `out` parameter.

    :param vendor_code: Vendor code used to lookup isbn_key (See `vendor`)
    :param workbook: path to Excel workbook file
    :param worksheet: name of worksheet in workbook file
    :param scraped_items_db: path to scraped items database file
    :param options:
    {   out: path where the new spreadsheet will be saved
        gbp_to_usd: conversion rate,
        inventory: inventory list to update workbook quantities
    }
    :param **kwargs: various sheet_waves key value pairings
        (canonically in CFG["asg"]["spreadsheet"]["sheet_waves"])
    :returns: updated Worksheet
    """
    namespace = f"{MODULE}.{sheet_waves.__name__}"
    out = options.get("out")
    gbp_to_usd = options.get("gbp_to_usd")
    inventory = options.get("inventory")
    for discount in kwargs["discounted_prices"]:
        kwargs["data_columns"].append(kwargs["discount_text_map"].format(t=discount))
    for field in kwargs["preset_fields"]:
        kwargs["data_columns"].append(field)
    addl_columns = kwargs["data_columns"] + kwargs["image_columns"]
    # get vendor info from database
    vendr = vendor(vendor_code)

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Vendor: {vendr.code}, ISBN_KEY: {isbn_key}")
    sheet_data = get_sheet_data(workbook, worksheet)

    sheet_keys = [x for x in sheet_data.pop(0) if x]  # filter out None

    items_obj = items.Items(
        sheet_keys, sheet_data, vendr.isbn_key, vendr.vendor_specific_id_col
    )
    items_obj.load_scraped_data(scraped_items_db)

    # Load worksheet
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}, Worksheet is {worksheet}")
    # Append columns
    col_insert_idx = ws.max_column + 1
    ws.insert_cols(col_insert_idx, len(addl_columns))
    for i, col in enumerate(addl_columns, start=1):
        ws.cell(row=1, column=col_insert_idx + i, value=col)

    updated_keys = get_sheet_keys(ws)
    try:
        isbn_idx = updated_keys.index(isbn_key)
    except ValueError as e:
        err = f"{namespace}: ISBN key '{isbn_key}' not found in sheet keys"
        raise ValueError(err) from e
    try:
        vendr_id_idx = updated_keys.index(vendr.vendor_specific_id_col.strip().upper())
    except ValueError:
        vendr_id_idx = None
    waves_update_from_inventory(vendr.code, isbn_key, ws, inventory)
    # Insert data into cells
    for row in ws.iter_rows(min_row=2):
        # get isbn cell
        isbn = str(row[isbn_idx].value)
        # find items_obj matching isbn
        vendr_id = str(row[vendr_id_idx].value) if vendr_id_idx else None
        item = (
            items_obj.find_item(isbn) or items_obj.find_item(vendr_id)
            if vendr_id
            else None
        )

        add_waves_row_data(
            updated_keys,
            item,
            row,
            gbp_to_usd,
            **kwargs,
        )

    rename_columns(ws, sheet_keys)
    format_headers(ws)
    # Save workbook
    wb.save(out)
    return ws
