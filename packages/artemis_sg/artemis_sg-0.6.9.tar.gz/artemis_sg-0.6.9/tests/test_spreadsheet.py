# ruff: noqa: S101
import logging
import os
import shutil
from datetime import datetime
from unittest.mock import MagicMock, Mock, patch

import pytest
from openpyxl.styles import Font, PatternFill

from artemis_sg import items, spreadsheet, vendor
from artemis_sg.config import CFG


@pytest.fixture()
def sheet_output_file(target_directory):
    return os.path.join(target_directory, "sheet_image_output_file.xlsx")


@pytest.fixture()
def populated_target_directory(tmp_path_factory, image_filepath):
    path = tmp_path_factory.mktemp("data")
    shutil.copyfile(image_filepath, os.path.join(path, "9780802150578.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9780500093924.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9780802150493.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "672125069899.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9999999999990.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9999999999990-1.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "FI-1234.jpg"))
    with open(os.path.join(path, "9999999999999.jpg"), "w") as f:
        f.write("I am not an image file")
        f.close()
    yield path


@pytest.mark.parametrize("worksheet", [None, "Sheet1"])
def test_sheet_image_finds_worksheet(
    spreadsheet_filepath,
    worksheet,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet
    AND a worksheet
    WHEN sheet_image is run
    THEN the worksheet contains expected column
    """
    options = {"out": sheet_output_file}
    ws = spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        worksheet,
        populated_target_directory,
        options,
    )

    # spreadsheet_filepath["Sheet1"] columns are [foo,bar,ISBN-13,baz,Order]
    assert "ISBN-13" in [cell.value.upper() for cell in ws[1]]


def test_sheet_image_saves_output_file(
    spreadsheet_filepath,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet
    AND an output filepath
    WHEN sheet_image is run
    THEN a file is saved as the given output spreadsheet file
    """
    options = {"out": sheet_output_file}
    spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        None,
        populated_target_directory,
        options,
    )

    assert os.path.exists(sheet_output_file)


def test_sheet_image_handles_column_orderding(
    spreadsheet_filepath,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet
    AND an output filepath
    AND a column order of ["ISBN", "IMAGE", "FOO", "ORDER", "ADDL"]
    WHEN sheet_image is run
    THEN the worksheet columns begin with  ["ISBN", "IMAGE", "FOO", "ORDER", "ADDL"]
    """
    col_order = ["ISBN", "IMAGE", "FOO", "ORDER", "ADDL"]
    # Spreadsheet["Sheet1"] columns are [foo,bar,ISBN-13,baz,Order]
    expected_sheet_keys = [
        "ISBN-13",
        "IMAGE",
        "FOO",
        "ORDER",
        "ADDL",
        "BAR",
        "BAZ",
        "AMAZON RANK",
    ]
    options = {"out": sheet_output_file}
    ws = spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        None,
        populated_target_directory,
        options,
        col_order,
    )

    headers = [cell.value.upper() for cell in ws[1]]
    assert headers == expected_sheet_keys


def test_sheet_image_adds_image_data(
    spreadsheet_filepath,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet
    AND an output filepath
    WHEN sheet_image is run
    THEN the file contains an "IMAGE" column
    AND the file contains image an openpyxl image object
    AND the object contains data
    AND the object is anchored to a cell
    """
    options = {"out": sheet_output_file}
    ws = spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        None,
        populated_target_directory,
        options,
    )

    # Unfortunately, we need to inspect the internals of the openpyxl.worksheet
    # object to verify that images are there.
    # See source:
    # https://foss.heptapod.net/openpyxl/openpyxl/-/blob/branch/default/openpyxl/drawing/image.py
    # https://foss.heptapod.net/openpyxl/openpyxl/-/blob/branch/default/openpyxl/drawing/drawing.py
    assert len(ws._images) > 0
    assert len(ws._images[0]._data()) > 0
    if isinstance(ws._images[0].anchor, str):
        assert ws._images[0].anchor != "A1"
    else:
        assert ws._images[0].anchor._from.col > 0
        assert ws._images[0].anchor._from.row > 0


def test_issue_101_sheet_image_handles_isbn_float_values(
    caplog,
    spreadsheet_filepath,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet containing an ISBN of 9780802150493.09
    AND an output filepath
    WHEN sheet_image is run
    """
    # The image file without the mantissa of the float ISBN should be added.
    # See: https://gitlab.com/johnduarte/artemis_sg/-/issues/101
    clean_float_image_file = os.path.join(
        populated_target_directory, "9780802150493.jpg"
    )
    caplog.set_level(logging.INFO)
    options = {"out": sheet_output_file}
    spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        None,
        populated_target_directory,
        options,
    )

    # Nothing direcly observable from the spreadsheet.  Look at the log to ensure
    # that the float ISBN has cleaned to find the image file.
    assert (
        "root",
        logging.INFO,
        f"spreadsheet.insert_image: Inserted '{clean_float_image_file}'.",
    ) in caplog.record_tuples


def test_issue_213_sheet_image_doesnt_duplicate_order_column(
    spreadsheet_filepath,
    populated_target_directory,
    sheet_output_file,
):
    """
    GIVEN a spreadsheet with an "Order" column
    AND the "Order" column row 2 value is 42
    WHEN sheet_image is run
    THEN "ORDER" appears once in the worksheet columns
    AND the "Order" column row 2 value is 42
    """
    expected_val = 42
    options = {"out": sheet_output_file}
    ws = spreadsheet.sheet_image(
        "sample",
        spreadsheet_filepath,
        None,
        populated_target_directory,
        options,
    )

    order_headers = [cell for cell in ws[1] if cell.value.upper() == "ORDER"]
    assert len(order_headers) == 1  # "ORDER" occurs once in column headers
    assert ws[f"{order_headers[0].column_letter}2"].value == expected_val


def test_mkthumbs_deletes_corrupted_image(populated_target_directory):
    """
    GIVEN a corrupted JPEG file in an image directory
    WHEN mkthumbs is run with the image directory
    THEN mkthumbs should complete without error
    AND the corrupted file should not exist in the image directory
    """
    corrupted_file = "9999999999999.jpg"
    image_directory = str(populated_target_directory)

    spreadsheet.mkthumbs(image_directory)

    assert True
    assert corrupted_file not in os.listdir(image_directory)


def test_mkthumbs_creates_thumbnails(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory
    WHEN mkthumbs is run with the image directory
    THEN thumbnails subdirectory should be created
    AND the JPEG should exist in the subdirectory
    """
    image_file = "9999999999990.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert subdir in os.listdir(image_directory)
    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_doesnt_create_supplementals(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory with a '-1' suffix
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should not exist in the subdirectory
    """
    image_file = "9999999999990-1.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file not in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_creates_invalid_isbn(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory named with an invalid isbn
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should exist in the subdirectory
    """
    image_file = "672125069899.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_creates_item_with_hyphen(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory named with a hyphen ("FI-1234.jpg")
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should exist in the subdirectory
    """
    image_file = "FI-1234.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_ignores_no_basename(tmp_path_factory, image_filepath):
    """
    GIVEN a JPEG file in an image directory named '.jpg'
    WHEN mkthumbs is run with the image directory
    THEN the file should not exist in the subdirectory
    """
    subdir = "thumbnails"
    image_file = ".jpg"
    image_directory = tmp_path_factory.mktemp("test_no_basename")
    shutil.copyfile(image_filepath, os.path.join(image_directory, image_file))

    spreadsheet.mkthumbs(image_directory)

    assert image_file not in os.listdir(os.path.join(image_directory, subdir))


@patch("artemis_sg.spreadsheet.add_image_to_row")
def test_sheet_image_uses_vendor_col_when_no_isbn(
    mock_add_image_to_row,
    spreadsheet_filepath,
    target_directory,
    tmp_path_factory,
    sheet_output_file,
):
    """
    GIVEN a workbook, vendor code, and spreadsheet
    AND the spreadsheet has a column with the name of the vendor.vendor_specific_id_col
    AND the isbn column values are invalid
    THEN spreadsheet.add_image_to_row is called with the vendor id column values
    """
    vendr = vendor.vendor("sample")
    worksheet = "Sheet1"
    workbook = spreadsheet_filepath
    input_wb = spreadsheet.load_workbook(workbook)
    input_ws = spreadsheet.get_worksheet(input_wb, worksheet)
    headers = [cell.value for cell in input_ws[1]]
    isbn_idx = headers.index("ISBN-13") + 1  # openpxl
    vendor_specific_id_col_idx = len(headers) + 1
    input_ws.cell(
        row=1, column=vendor_specific_id_col_idx
    ).value = vendr.vendor_specific_id_col
    for idx in range(2, 5):
        input_ws.cell(row=idx, column=isbn_idx).value = "invalid isbn"
        input_ws.cell(row=idx, column=vendor_specific_id_col_idx).value = f"0000{idx}"
    input_ws.cell(
        row=4, column=vendor_specific_id_col_idx
    ).value = 4  # test with non-string values
    input_file_name = f"{target_directory}/wksht_w_tb_code_no_valid_isbn.xlsx"
    input_wb.save(input_file_name)
    image_directory = tmp_path_factory.mktemp("test_dir")
    options = {"out": sheet_output_file}

    spreadsheet.sheet_image(
        "sample", input_file_name, worksheet, image_directory, options
    )

    expected_tb_codes = ["00002", "00003", 4]

    for idx, call in enumerate(mock_add_image_to_row.call_args_list):
        call_args, _call_kwargs = call
        tb_code_cell = call_args[5]
        assert tb_code_cell.value == expected_tb_codes[idx]


def test_get_order_items(spreadsheet_filepath):
    """
    GIVEN a spreadsheet with "ISBN-13" and "Order" columns
    AND the spreadsheet contains rows with items and quantities
    WHEN get_order_items is run with a vendor_code
    AND the workbook and worksheet references for the spreadsheet
    THEN a list of order items is returned
    """
    expected_list = [
        ("9780802150578", "42"),
        ("9780500093924", "3"),
        ("9780802150493", "3"),
    ]
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"

    order_items = spreadsheet.get_order_items(
        vendor.vendor(vendor_code), workbook, worksheet
    )

    assert isinstance(order_items, list)
    assert order_items == expected_list


def test_get_order_items_w_vendor_specific_id_col(
    spreadsheet_filepath, target_directory
):
    expected_list = [
        ("00002", "42"),
        ("00003", "3"),
        ("00004", "3"),
    ]
    vendr = vendor.vendor("sample")
    vendr.vendor_specific_id_col = "test id"
    worksheet = "Sheet1"
    workbook = spreadsheet_filepath
    input_wb = spreadsheet.load_workbook(workbook)
    input_ws = spreadsheet.get_worksheet(input_wb, worksheet)
    headers = [cell.value for cell in input_ws[1]]
    isbn_idx = headers.index("ISBN-13") + 1  # openpxl
    vendor_specific_id_col_idx = len(headers) + 1
    input_ws.cell(
        row=1, column=vendor_specific_id_col_idx
    ).value = vendr.vendor_specific_id_col
    for idx in range(2, 5):
        input_ws.cell(row=idx, column=isbn_idx).value = "invalid isbn"
        input_ws.cell(row=idx, column=vendor_specific_id_col_idx).value = f"0000{idx}"
    input_file_name = f"{target_directory}/wksht_w_tb_code_no_valid_isbn.xlsx"
    input_wb.save(input_file_name)

    order_items = spreadsheet.get_order_items(vendr, input_file_name, worksheet)

    assert isinstance(order_items, list)
    assert order_items == expected_list


def test_get_sheet_data_gdoc_id_no_worksheet(monkeypatch):
    """
    Given a valid Google Doc ID
    WHEN get_sheet_data is executed with the ID
    AND no worksheet name is provided
    THEN the first worksheet is chosen
    """
    workbook = "not_a_file"
    worksheet = None
    expected_worksheet = "CoolSheet"

    mock_sheets = Mock()
    call_chain = (
        "spreadsheets.return_value."
        "get.return_value."
        "execute.return_value."
        "get.return_value."
        "pop.return_value.get.return_value.get.return_value"
    )
    config = {"name": "mock_sheets", call_chain: expected_worksheet}
    mock_sheets.configure_mock(**config)
    monkeypatch.setattr(
        spreadsheet.app_creds, "app_creds", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(spreadsheet, "build", lambda *args, **kwargs: mock_sheets)

    spreadsheet.get_sheet_data(workbook, worksheet)

    mock_sheets.spreadsheets().get.assert_called_with(spreadsheetId=workbook)
    mock_sheets.spreadsheets().values().get.assert_called_with(
        range=expected_worksheet, spreadsheetId=workbook
    )


def test_get_sheet_data_file_worksheet(monkeypatch, spreadsheet_filepath):
    """
    Given a valid spreadsheet file
    WHEN get_sheet_data is executed with the file path
    AND no worksheet name is provided
    THEN the first worksheet data is loaded
    """
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"

    data = spreadsheet.get_sheet_data(workbook, worksheet)

    assert isinstance(data, list)
    assert "ISBN-13" in data[0]


def test_get_sheet_data_file_no_worksheet(monkeypatch, spreadsheet_filepath):
    """
    Given a valid spreadsheet file
    WHEN get_sheet_data is executed with the file path
    AND no worksheet name is provided
    THEN the first worksheet data is loaded
    """
    workbook = spreadsheet_filepath
    worksheet = None

    data = spreadsheet.get_sheet_data(workbook, worksheet)

    assert isinstance(data, list)
    assert "ISBN-13" in data[0]


def test_get_sheet_data_with_datetime_objects(target_directory, spreadsheet_filepath):
    """
    GIVEN a workbook
    AND a datetime object is a cell value in the workbook
    WHEN spreadsheet.get_sheet_data is run with the workbook
    THEN any datetime objects are converted to strings in the returned data
    """
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, None)
    headers = spreadsheet.get_sheet_keys(ws)
    insert_idx = len(headers)
    date_str = "1/2/2003"
    us_format = "%m/%d/%Y"
    date_obj = datetime.strptime(date_str, us_format)  # noqa: DTZ007
    ws.cell(row=2, column=insert_idx).value = date_obj
    input_file_name = f"{target_directory}/wksht_w_tb_code_no_valid_isbn.xlsx"
    wb.save(input_file_name)

    res = spreadsheet.get_sheet_data(input_file_name)
    assert res[1][4] == "01/02/2003"
    assert res[1][4] != date_obj


def test_format_headers(
    spreadsheet_filepath,
    sheet_output_file,
    valid_datafile,
    target_directory,
    populated_target_directory,
):
    """
    Given a valid spreadsheet file
    AND the spreadsheet has the first cell bolded
    AND highlighted
    WHEN sheet_waves is run
    AND sheet_image is run
    THEN both output spreadsheets have headers bolded and highlighted
    """
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    cell_font = Font(bold=True)
    cell_fill = PatternFill(
        patternType="solid",
        fgColor="FFFF00",
    )
    input_wb = spreadsheet.load_workbook(workbook)
    input_ws = spreadsheet.get_worksheet(input_wb, None)
    row01 = input_ws[1]
    formatted_cell = row01[0]
    formatted_cell.font = cell_font
    formatted_cell.fill = cell_fill
    altered_wkbk = f"{target_directory}/sheet_with_one_cell_formatted.xlsx"
    input_wb.save(altered_wkbk)
    options = {"out": sheet_output_file, "gbp_to_usd": None, "inventory": None}
    spreadsheet.sheet_waves(
        vendor_code,
        altered_wkbk,
        None,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )

    assert os.path.exists(sheet_output_file)
    wb_wave = spreadsheet.load_workbook(sheet_output_file)
    ws_wave = spreadsheet.get_worksheet(wb_wave, None)
    row01 = ws_wave[1]
    for cell in row01:
        assert cell.font.bold == formatted_cell.font.bold
        assert cell.fill.fgColor == formatted_cell.fill.fgColor
    options = {"out": sheet_output_file}
    spreadsheet.sheet_image(
        vendor_code,
        altered_wkbk,
        None,
        populated_target_directory,
        options,
    )
    wb_sheet_image = spreadsheet.load_workbook(sheet_output_file)
    ws_sheet_image = spreadsheet.get_worksheet(wb_sheet_image, None)
    row01 = ws_sheet_image[1]
    for cell in row01:
        assert cell.font.bold == formatted_cell.font.bold
        assert cell.fill.fgColor == formatted_cell.fill.fgColor


def test_sheet_waves_output_file(
    spreadsheet_filepath, sheet_output_file, valid_datafile
):
    """
    GIVEN a spreadsheet
    AND the workbook and worksheet references for the spreadsheet
    AND an output filepath
    WHEN sheet_waves is run in debug mode
    THEN a file is saved as the given output file.
    AND the file contains a Description column
    AND the file contains a Width column
    AND the file contains a Height column
    AND the file contains a Length column
    AND the file contains a Image URL column
    AND the file contains a Image 1 column
    AND the file contains a Image 2 column
    AND the file contains a Image 3 column
    AND the file contains a Image 4 column
    AND the file contains a Image 5 column
    AND the file contains a Image 6 column
    """
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"
    options = {"out": sheet_output_file, "gbp_to_usd": None, "inventory": None}
    spreadsheet.sheet_waves(
        vendor_code,
        workbook,
        worksheet,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )

    assert os.path.exists(sheet_output_file)
    wb = spreadsheet.load_workbook(sheet_output_file)
    ws = spreadsheet.get_worksheet(wb, worksheet)
    row01 = ws[1]
    headers = [cell.value for cell in row01]
    assert "Description" in headers
    assert "Width" in headers
    assert "Height" in headers
    assert "Length" in headers
    assert "Image URL" in headers
    assert "Image 1" in headers
    assert "Image 2" in headers
    assert "Image 3" in headers
    assert "Image 4" in headers
    assert "Image 5" in headers
    assert "Image 6" in headers


def test_sheet_waves_pound_pricing(spreadsheet_filepath):
    """
    GIVEN a spreadsheet
    AND the workbook and worksheet references for the spreadsheet
    AND the spreadsheet has a RRP and FORMAT column
    WHEN waves_set_pound_price function is run
    THEN pound price is added to the row in the spreadsheet
    """
    worksheet = "Sheet1"
    test_wb = spreadsheet.load_workbook(spreadsheet_filepath)
    test_ws = spreadsheet.get_worksheet(test_wb, worksheet)
    insert_idx = len(test_ws[1])
    format_idx = insert_idx
    rrp_idx = insert_idx + 1
    pound_price_idx = insert_idx + 2
    test_ws.cell(row=1, column=format_idx, value="FORMAT")
    test_ws.cell(row=1, column=rrp_idx, value="RRP")
    test_ws.cell(row=1, column=pound_price_idx, value="Pound Pricing")
    headers = spreadsheet.get_sheet_keys(test_ws)
    test_data = [
        {"FORMAT": "af", "RRP": 3.50},
        {"FORMAT": "cf", "RRP": 3.99},
        {"FORMAT": "invalid", "RRP": 10},
        {"FORMAT": None, "RRP": None},
    ]
    pound_prices = ["1.20", "1.50", "4.00", None]
    for idx, data in enumerate(test_data):
        item_idx = idx + 2
        test_ws.cell(row=item_idx, column=format_idx, value=data["FORMAT"])
        test_ws.cell(row=item_idx, column=rrp_idx, value=data["RRP"])

        spreadsheet.waves_set_pound_price(
            headers, data["RRP"], data["FORMAT"], test_ws[item_idx]
        )
        assert (
            test_ws.cell(row=item_idx, column=pound_price_idx).value
            == pound_prices[idx]
        )


def test_waves_calculate_fields(spreadsheet_filepath):
    """
    GIVEN a row of a workbook
    AND a nested dict
    WHEN waves_calculate_fields is called
    THEN row values are set based on a mapping scheme
    """
    worksheet = "Sheet1"
    test_wb = spreadsheet.load_workbook(spreadsheet_filepath)
    test_ws = spreadsheet.get_worksheet(test_wb, worksheet)
    headers = spreadsheet.get_sheet_keys(test_ws)
    test_ws.cell(row=1, column=len(headers) + 1, value="test_col")
    test_ws.cell(row=1, column=len(headers) + 2, value="test_col2")
    test_ws.cell(row=1, column=len(headers) + 3, value="test_col3")
    headers = spreadsheet.get_sheet_keys(test_ws)
    row = test_ws[2]
    fields = {}
    fields["foo"] = {
        "map_from": "bar",
        "map": {
            row[headers.index("BAR")].value: "waldo",
        },
    }
    # test_col is used to test empty cell values as keys
    fields["Order"] = {
        "map_from": ["test_col", "bar"],
        "map": {"": {row[headers.index("BAR")].value: "foobar"}},
    }
    fields["baz"] = {
        "map_from": ["bar", "test_col"],
        "map": {row[headers.index("BAR")].value: {"": "fred"}},
    }
    fields["test_col2"] = {
        "map_from": ["bar", "foo"],
        "map": {row[headers.index("BAR")].value: {"any": "thud"}},
    }
    fields["test_col3"] = {
        "map_from": ["bar", "foo"],
        "map": {"any": {"waldo": "corge"}},
    }
    spreadsheet.waves_calculate_fields(headers, row, fields)

    assert row[headers.index("FOO")].value == "waldo"
    assert row[headers.index("ORDER")].value == "foobar"
    assert row[headers.index("BAZ")].value == "fred"
    assert row[headers.index("TEST_COL2")].value == "thud"
    assert row[headers.index("TEST_COL3")].value == "corge"


def test_waves_adds_prices_to_sheet(
    spreadsheet_filepath, sheet_output_file, valid_datafile, target_directory
):
    """
    GIVEN a spreadsheet with 'Pound Pricing', 'RRP', and 'Format' columns
    AND the workbook and worksheet references for the spreadsheet
    WHEN sheet_waves is run with the --gbp_to_usd flag
    THEN discounted prices are added to the output sheet.
    AND pound prices are added to the output sheet.
    """
    test_wb_filepath = spreadsheet_filepath
    test_ws_name = "Sheet1"
    test_wb = spreadsheet.load_workbook(test_wb_filepath)
    test_ws = spreadsheet.get_worksheet(test_wb, test_ws_name)
    vendor_code = "sample"
    insert_idx = len(test_ws[1])
    format_idx = insert_idx
    rrp_idx = insert_idx + 1
    pound_price_idx = insert_idx + 2
    test_ws.cell(row=1, column=format_idx, value="FORMAT")
    test_ws.cell(row=1, column=rrp_idx, value="RRP")
    test_ws.cell(row=1, column=pound_price_idx, value="Pound Pricing")
    test_data = [
        {"FORMAT": "af", "RRP": 3.50},
        {"FORMAT": "cf", "RRP": 3.99},
        {"FORMAT": "invalid", "RRP": 10},
        {"FORMAT": None, "RRP": None},
    ]
    pound_prices = ["1.20", "1.50", "4.00", None]
    for idx, data in enumerate(test_data):
        item_idx = idx + 2
        test_ws.cell(row=item_idx, column=format_idx, value=data["FORMAT"])
        test_ws.cell(row=item_idx, column=rrp_idx, value=data["RRP"])
    altered_sheet_filepath = f"{target_directory}/altered_sheet.xlsx"
    test_wb.save(altered_sheet_filepath)
    options = {"out": sheet_output_file, "gbp_to_usd": 1.36, "inventory": None}
    spreadsheet.sheet_waves(
        vendor_code,
        altered_sheet_filepath,
        test_ws_name,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )
    output_wb = spreadsheet.load_workbook(sheet_output_file)
    output_ws = spreadsheet.get_worksheet(output_wb, test_ws_name)

    discount_50_results = ["2.38", "2.71", "6.80", None]
    discount_60_results = ["1.90", "2.17", "5.44", None]
    headers = spreadsheet.get_sheet_keys(output_ws)
    discount_50_perc_idx = headers.index("50% OFF (USD)")
    discount_60_perc_idx = headers.index("60% OFF (USD)")
    pound_price_idx = headers.index("POUND PRICING")
    for idx in range(len(test_data)):
        item_idx = idx + 2
        assert (
            output_ws.cell(row=item_idx, column=pound_price_idx + 1).value
            == pound_prices[idx]
        )
        assert (
            output_ws.cell(row=item_idx, column=discount_50_perc_idx + 1).value
            == discount_50_results[idx]
        )
        assert (
            output_ws.cell(row=item_idx, column=discount_60_perc_idx + 1).value
            == discount_60_results[idx]
        )


def test_waves_preset_data_added_to_sheet(
    spreadsheet_filepath, sheet_output_file, target_directory, valid_datafile
):
    """
    GIVEN a valid vendor, and workbook
    AND preset_fields defined in config
    WHEN the sheet-waves command is run
    THEN the preset_field values are added to the sheet for items with valid isbns's.
    """
    vendor_code = "sample"
    worksheet = "Sheet1"
    workbook = spreadsheet_filepath
    options = {"out": sheet_output_file, "gbp_to_usd": None, "inventory": None}
    spreadsheet.sheet_waves(
        vendor_code,
        workbook,
        worksheet,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )
    wb = spreadsheet.load_workbook(sheet_output_file)
    ws = spreadsheet.get_worksheet(wb, worksheet)
    headers = [cell.value for cell in ws[1]]
    preset_fields = CFG["asg"]["spreadsheet"]["sheet_waves"]["preset_fields"]
    for idx in range(2, ws.max_row + 1):
        for key, val in preset_fields.items():
            assert ws.cell(row=idx, column=headers.index(key) + 1).value == val


def test_waves_qty_update_from_inventory(spreadsheet_filepath, target_directory):
    """
    GIVEN a valid workbook and vendor
    AND inventory spreadsheet
    WHEN waves_update_from_inventory is run
    THEN the quantities on the workbook are updated from the inventory.
    """
    vendor_code = "sample"
    test_wb_name = spreadsheet_filepath
    test_ws_name = "Sheet1"
    test_wb = spreadsheet.load_workbook(test_wb_name)
    test_ws = spreadsheet.get_worksheet(test_wb, test_ws_name)

    headers = [cell.value for cell in test_ws[1]]
    isbn_key = "ISBN-13"
    qty_idx = len(headers)
    new_quantities = [12, 13, 14]
    old_quantites = [2, 26, 1, 99]
    res = [12, 13, 14, 0]
    test_ws.cell(row=1, column=qty_idx, value="quantity")
    for idx, val in enumerate(new_quantities):
        test_ws.cell(row=idx + 2, column=qty_idx, value=val)
    inventory_list_name = f"{target_directory}/inventory.xlsx"
    test_wb.save(inventory_list_name)
    test_ws.cell(row=1, column=qty_idx, value="on_hand")
    isbn_not_in_inventory = 999999999
    test_ws.cell(row=5, column=headers.index(isbn_key) + 1, value=isbn_not_in_inventory)
    for idx, val in enumerate(old_quantites):
        test_ws.cell(row=idx + 2, column=qty_idx, value=val)
    spreadsheet.waves_update_from_inventory(
        vendor_code, isbn_key, test_ws, inventory_list_name
    )
    for idx, val in enumerate(res):
        assert test_ws.cell(row=idx + 2, column=qty_idx).value == val


def test_add_to_row(spreadsheet_filepath):
    """
    GIVEN a valid spreadsheet row
    AND the spreadsheet headers, header key, and value
    WHEN spreadsheet.add_to_row in run
    THEN the value will be added to the row at the index
    that the key is in headers.
    """
    worksheet = "Sheet1"
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, worksheet)
    headers = spreadsheet.get_sheet_keys(ws)
    row = ws[2]
    spreadsheet.add_to_row(headers, row, "ISBN-13", "test_val")
    assert ws.cell(row=2, column=headers.index("ISBN-13") + 1).value == "test_val"


@patch.object(
    items.Items,
    "find_item",
    autospec=True,
    wraps=items.Items.find_item,
)
def test_sheet_waves_uses_vendor_id_col_when_no_isbn(
    mock_find_item,
    target_directory,
    spreadsheet_filepath,
    valid_datafile,
    sheet_output_file,
):
    """
    GIVEN a workbook, worksheet, vendor, and valid datafile
    AND the worksheet has invalid isbns in the isbn column
    AND the worksheet has a vendr.vendor_specific_id_col
    WHEN sheet_waves is called
    THEN the items are searched using the vendor specific id col
    """
    options = {"out": sheet_output_file}
    test_sheet = "Sheet1"
    test_wb = spreadsheet.load_workbook(spreadsheet_filepath)
    test_ws = spreadsheet.get_worksheet(test_wb, test_sheet)
    headers = spreadsheet.get_sheet_keys(test_ws)
    isbn_idx = headers.index("ISBN-13") + 1

    test_ws.cell(row=2, column=isbn_idx, value="invalid isbn")
    test_ws.cell(row=3, column=isbn_idx, value="invalid isbn")
    test_ws.cell(row=4, column=isbn_idx, value="invalid isbn")
    vendr_ids = [12, 13, 14]
    vendr = vendor.vendor("sample")
    vendr_id_col = len(headers) + 1
    test_ws.cell(row=1, column=vendr_id_col, value=vendr.vendor_specific_id_col)
    for idx in range(len(vendr_ids)):
        test_ws.cell(row=idx + 2, column=vendr_id_col, value=vendr_ids[idx])

    input_workbook_name = f"{target_directory}/input.xlsx"
    test_wb.save(input_workbook_name)
    spreadsheet.sheet_waves(
        "sample",
        input_workbook_name,
        test_sheet,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )

    called_isbns = [call.args[1] for call in mock_find_item.call_args_list]
    assert called_isbns == [
        "invalid isbn",
        str(vendr_ids[0]),
        "invalid isbn",
        str(vendr_ids[1]),
        "invalid isbn",
        str(vendr_ids[2]),
    ]


def test_waves_product_active_col_from_qty(
    spreadsheet_filepath, sheet_output_file, valid_datafile, target_directory
):
    """
    GIVEN a workbook, vendor, and inventory
    WHEN sheet_waves is run with an inventory list
    THEN titles with a nonzero quantity have product_active as 1
    AND the rest are set to None
    """
    vendor_code = "sample"
    test_wb_name = spreadsheet_filepath
    test_ws_name = "Sheet1"

    test_wb = spreadsheet.load_workbook(test_wb_name)
    test_ws = spreadsheet.get_worksheet(test_wb, test_ws_name)
    headers = spreadsheet.get_sheet_keys(test_ws)
    isbn_idx = headers.index("ISBN-13") + 1
    qty_idx = len(headers)
    new_quantities = [12, 0, None, ""]
    old_quantites = [2, 26, 1, 99, 45]
    test_ws.cell(row=5, column=isbn_idx, value="9780802157003")
    test_ws.cell(row=1, column=qty_idx, value="QUANTITY")
    for idx, val in enumerate(new_quantities):
        test_ws.cell(row=idx + 2, column=qty_idx).value = val
    inventory_list_name = f"{target_directory}/inventory.xlsx"
    test_wb.save(inventory_list_name)
    test_ws.cell(row=1, column=qty_idx, value="ON_HAND")
    input_workbook_name = f"{target_directory}/input.xlsx"

    for idx, val in enumerate(old_quantites):
        test_ws.cell(row=idx + 2, column=qty_idx, value=val)
    test_wb.save(input_workbook_name)
    options = {
        "out": sheet_output_file,
        "gbp_to_usd": None,
        "inventory": inventory_list_name,
    }
    spreadsheet.sheet_waves(
        vendor_code,
        input_workbook_name,
        test_ws_name,
        valid_datafile,
        options,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )

    output_wb = spreadsheet.load_workbook(sheet_output_file)
    output_ws = spreadsheet.get_worksheet(output_wb, test_ws_name)
    headers = spreadsheet.get_sheet_keys(output_ws)
    product_active_idx = headers.index("PRODUCT_ACTIVE")

    assert output_ws.cell(row=2, column=product_active_idx + 1).value == 1
    assert output_ws.cell(row=3, column=product_active_idx + 1).value is None
    assert output_ws.cell(row=4, column=product_active_idx + 1).value is None
    assert output_ws.cell(row=4, column=product_active_idx + 1).value is None


def test_waves_product_active_col_from_qty_without_inventory_list(
    spreadsheet_filepath, sheet_output_file, valid_datafile, target_directory
):
    """
    GIVEN a workbook, and vendor
    WHEN sheet_waves is run without an inventory list
    THEN all rows have product_active set to 1
    """
    vendor_code = "sample"
    test_wb_name = spreadsheet_filepath
    test_ws_name = "Sheet1"

    test_wb = spreadsheet.load_workbook(test_wb_name)
    test_ws = spreadsheet.get_worksheet(test_wb, test_ws_name)
    headers = spreadsheet.get_sheet_keys(test_ws)
    isbn_idx = headers.index("ISBN-13") + 1
    qty_idx = len(headers)
    new_quantities = [12, 0, None, ""]
    old_quantites = [2, 26, 1, 99, 45]
    test_ws.cell(row=5, column=isbn_idx, value="9780802157003")
    test_ws.cell(row=1, column=qty_idx, value="QUANTITY")
    for idx, val in enumerate(new_quantities):
        test_ws.cell(row=idx + 2, column=qty_idx).value = val
    inventory_list_name = f"{target_directory}/inventory.xlsx"
    test_wb.save(inventory_list_name)
    test_ws.cell(row=1, column=qty_idx, value="ON_HAND")
    input_workbook_name = f"{target_directory}/input.xlsx"

    for idx, val in enumerate(old_quantites):
        test_ws.cell(row=idx + 2, column=qty_idx, value=val)
    test_wb.save(input_workbook_name)
    options_w_no_inventory = {
        "out": sheet_output_file,
        "gbp_to_usd": None,
        "inventory": None,
    }
    spreadsheet.sheet_waves(
        vendor_code,
        input_workbook_name,
        test_ws_name,
        valid_datafile,
        options_w_no_inventory,
        **CFG["asg"]["spreadsheet"]["sheet_waves"],
    )
    output_wb = spreadsheet.load_workbook(sheet_output_file)
    output_ws = spreadsheet.get_worksheet(output_wb, test_ws_name)
    headers = spreadsheet.get_sheet_keys(output_ws)
    product_active_idx = headers.index("PRODUCT_ACTIVE")
    assert output_ws.cell(row=2, column=product_active_idx + 1).value == 1
    assert output_ws.cell(row=3, column=product_active_idx + 1).value == 1
    assert output_ws.cell(row=4, column=product_active_idx + 1).value == 1
    assert output_ws.cell(row=4, column=product_active_idx + 1).value == 1


def test_get_isbns_after_cutoff_date():
    """
    GIVEN a list of tuples with isbn and a date as a string
    AND a cutoff-date
    WHEN get_isbns_after_cutoff_date is run
    THEN a two lists of isbns is returned
    AND one is for titles after cutoff and one is for before
    """
    data = [
        ("1", "01/28/2023"),
        ("4", "28/01/2021"),
        ("5", "invalid date"),
        ("2", "20/2/2025"),
        ("3", "02/20/2023"),
    ]
    isbns_after, isbns_before = spreadsheet.get_isbns_after_cutoff_date(
        data, "01/01/2022"
    )
    assert (isbns_after, isbns_before) == (["1", "2", "3"], ["4", "5"])


def test_sheet_image_with_invalid_cutoff_date(
    spreadsheet_filepath, populated_target_directory, sheet_output_file
):
    """
    GIVEN a spreadsheet
    AND a worksheet
    WHEN sheet_image is run with the cutoff as an invalid date string
    THEN a ValueError is raised
    """
    options = {"out": sheet_output_file, "cutoff_date": "001/10/2003"}
    with pytest.raises(ValueError):
        spreadsheet.sheet_image(
            "sample",
            spreadsheet_filepath,
            None,
            populated_target_directory,
            options,
        )


def test_sort_rows_by_cutoff_date(spreadsheet_filepath, target_directory):
    """
    GIVEN a spreadsheet, vendor code, and isbn key
    AND the vendor's recieve date column name is defined in config
    WHEN sort_rows_by_cutoff_date is run with a valid cutoff date
    THEN the top rows of sheet have titles after the cutoff date first
    """
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, None)
    headers = spreadsheet.get_sheet_keys(ws)
    start_date_col_idx = len(headers) + 1

    data = [("3", "01/28/2021"), ("1", "20/2/2022"), ("2", "02/20/2023")]
    ws.cell(row=1, column=start_date_col_idx, value="Start Date")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=start_date_col_idx).value = data[row_idx - 2][1]
    input_ws_filepath = f"{target_directory}/input_file.xlsx"
    wb.save(input_ws_filepath)
    sheet_data = spreadsheet.get_sheet_data(input_ws_filepath)
    input_wb = spreadsheet.load_workbook(input_ws_filepath)
    input_ws = spreadsheet.get_worksheet(input_wb, None)
    vendor_code = "sample"
    isbn_key = "ISBN-13"
    cutoff_date = "01/01/2022"
    spreadsheet.sort_rows_by_cutoff_date(vendor_code, input_ws, isbn_key, cutoff_date)
    input_wb.save(input_ws_filepath)
    output_sheet_data = spreadsheet.get_sheet_data(input_ws_filepath)
    assert output_sheet_data == [
        sheet_data[0],
        sheet_data[2],
        sheet_data[3],
        sheet_data[1],
    ]


@patch("artemis_sg.spreadsheet.add_image_to_row")
def test_sheet_image_w_cutoff_only_add_images_if_after_cutoff(
    mock_add_image_to_row,
    spreadsheet_filepath,
    target_directory,
    sheet_output_file,
    populated_target_directory,
):
    """
    GIVEN a spreadsheet, and vendor code
    AND the vendor's recieve date column name is defined in config
    WHEN sheet-image is run with the cutoff date flag
    THEN only titles after the cutoff date should have images
    """
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, None)
    headers = spreadsheet.get_sheet_keys(ws)
    start_date_col_idx = len(headers) + 1

    data = [("3", "01/28/2021"), ("1", "01/01/2022"), ("2", "02/20/2023")]
    ws.cell(row=1, column=start_date_col_idx, value="Start Date")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=start_date_col_idx).value = data[row_idx - 2][1]
    input_ws_filepath = f"{target_directory}/input_file.xlsx"
    wb.save(input_ws_filepath)
    vendor_code = "sample"
    cutoff_date = "01/01/2022"
    options = {"out": sheet_output_file, "cutoff_date": cutoff_date}
    spreadsheet.sheet_image(
        vendor_code, input_ws_filepath, None, populated_target_directory, options
    )
    num_isbns_after_cutoff = 2
    assert mock_add_image_to_row.call_count == num_isbns_after_cutoff


def test_add_rank_cols(spreadsheet_filepath):
    """
    GIVEN a worksheet and vendor with a failover scraper of "AmznUkScraper"
    WHEN add_rank_cols is called
    "AMAZON Rank" and "AMAZON_UK Rank" are added to the headers
    """
    vendr = Mock()
    vendr.failover_scraper = "AmznUkScraper"
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, None)
    spreadsheet.add_rank_cols(ws, vendr)
    sk = spreadsheet.get_sheet_keys(ws)
    assert "AMAZON RANK" in sk
    assert "AMAZON_UK RANK" in sk


def test_add_rank_to_row(spreadsheet_filepath):
    """
    GIVEN a spreadsheet, items object, and isbn cell
    AND a vendor with the failover_scraper set to "AmznUKScraper"
    AND a list of rank columns ["AMAZON", "AMAZON_UK"]
    THEN the ranks for each item are added to the spreadsheet
    """
    vendr = Mock()
    vendr.failover_scraper = "AmznUkScraper"
    wb = spreadsheet.load_workbook(spreadsheet_filepath)
    ws = spreadsheet.get_worksheet(wb, None)
    items_obj = Mock()
    tb_code_cell = Mock()
    isbn_cell = ws.cell(row=2, column=3)
    insert_idx = ws.max_column + 1
    ws.cell(row=1, column=insert_idx).value = "AMAZON RANK"
    ws.cell(row=1, column=insert_idx + 1).value = "AMAZON_UK RANK"
    rank_cols = ["AMAZON", "AMAZON_UK"]
    mock_item = MagicMock()  # need MagicMock to override __bool__
    mock_item.__bool__.return_value = True
    amzn_rank = 123
    amzn_uk_rank = 321
    mock_item.data = {"RANKS": {"AMAZON": amzn_rank, "AMAZON_UK": amzn_uk_rank}}
    items_obj.find_item.return_value = mock_item

    spreadsheet.add_rank_to_row(ws, items_obj, rank_cols, isbn_cell, tb_code_cell)
    assert ws.cell(row=2, column=insert_idx).value == amzn_rank
    assert ws.cell(row=2, column=insert_idx + 1).value == amzn_uk_rank
