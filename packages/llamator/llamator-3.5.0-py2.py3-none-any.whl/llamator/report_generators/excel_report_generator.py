import os

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from llamator.attack_provider.attack_registry import test_classes


def get_tests_mapping() -> dict:
    """
    Builds a mapping dictionary from registered attack classes.
    For each registered attack class, its 'info' field is extracted.
    The mapping key is 'code_name' and the value is the entire 'info' dictionary.
    """
    mapping = {}
    for cls in test_classes:
        if hasattr(cls, "info"):
            mapping[cls.info["code_name"]] = cls.info
    return mapping


def create_attack_report(attack_data: list[dict], file_path: str) -> None:
    """
    Generates a single-sheet Excel report based on the provided attack data.

    Parameters
    ----------
    attack_data : list[dict]
        Contains 'attack_text', 'response_text', and 'status'.
    file_path : str
        Path to save the resulting .xlsx file.
    """
    df = pd.DataFrame(attack_data)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attack Results")
        workbook = writer.book
        worksheet = workbook["Attack Results"]

        color_fill_mapping = {
            "broken": "FFCCCB",
            "resilient": "C1E1C1",
            "error": "FFD580",
        }

        for row_idx in range(2, len(df) + 2):
            cell = worksheet[f"C{row_idx}"]
            fill_color = color_fill_mapping.get(str(cell.value).lower())
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        worksheet.column_dimensions["A"].width = 50
        worksheet.column_dimensions["B"].width = 50
        worksheet.column_dimensions["C"].width = len("resilient")

        for row_idx in range(2, len(df) + 2):
            worksheet[f"A{row_idx}"].alignment = Alignment(wrap_text=True)
            worksheet[f"B{row_idx}"].alignment = Alignment(wrap_text=True)

        workbook.save(file_path)


def create_attack_report_from_artifacts(
    artifacts_dir: str, csv_folder_name: str = "csv_report", report_file_name: str = "attacks_report.xlsx"
) -> None:
    """
    Creates a multi-sheet Excel report from each CSV file found in the specified folder.

    Parameters
    ----------
    artifacts_dir : str
        Directory containing artifacts.
    csv_folder_name : str
        Subfolder with CSV files (default "csv_report").
    report_file_name : str
        Name for the resulting Excel report (default "attacks_report.xlsx").
    """
    csv_folder_path = os.path.join(artifacts_dir, csv_folder_name)
    output_file_path = os.path.join(artifacts_dir, report_file_name)

    color_fill_mapping = {
        "broken": "FFF0F0",  # pale red
        "resilient": "F0FFF0",  # pale green
        "error": "FFF8E7",  # pale orange
    }
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    tests_mapping = get_tests_mapping()

    with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
        for csv_file in os.listdir(csv_folder_path):
            if csv_file.endswith(".csv"):
                code_name = os.path.splitext(csv_file)[0]
                test_info = tests_mapping.get(code_name, None)
                if test_info:
                    sheet_name = test_info["name"]
                else:
                    sheet_name = code_name

                df = pd.read_csv(os.path.join(csv_folder_path, csv_file)).dropna(how="all")
                safe_sheet_name = sheet_name[:31]
                df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

                workbook = writer.book
                worksheet = workbook[safe_sheet_name]

                num_records = len(df.index)
                num_columns = len(df.columns)
                for row in range(2, num_records + 2):
                    status_cell = worksheet[f"{chr(64+num_columns)}{row}"]
                    fill_color = color_fill_mapping.get(str(status_cell.value).lower())
                    if fill_color:
                        status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                    for col_idx in range(1, worksheet.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                        cell.border = thin_border

                # Header row styling
                for col_idx in range(1, worksheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.alignment = Alignment(horizontal="center", vertical="center")
                    header_cell.border = thin_border

                for column in range(65, 64 + num_columns):
                    worksheet.column_dimensions[chr(column)].width = 150 // (num_columns - 1)
                worksheet.column_dimensions[chr(64 + num_columns)].width = len("resilient") + 5
