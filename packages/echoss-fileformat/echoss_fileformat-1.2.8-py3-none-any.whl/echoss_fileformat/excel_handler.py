import io
import pandas as pd
# for new format xlsx
from openpyxl import load_workbook
# for old format xls
import xlrd
from typing import Literal, Optional, Union
from .csv_handler import CsvHandler
from .echoss_logger import get_logger, set_logger_level

logger = get_logger('echoss_fileformat')

def find_problematic_cells(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active

    for row in sheet.iter_rows():
        for cell in row:
            try:
                cell_value = cell.value
                if isinstance(cell_value, str):
                    # 셀 값을 개별 문자로 순회하며 ord() 함수를 적용
                    for char in cell_value:
                        _ = ord(char)
            except Exception as e:
                print(f"Error in cell {cell.coordinate}: {e}")


class ExcelHandler(CsvHandler):
    """Excel file handler

    학습데이터로 Excel 파일은 전체 읽기를 기본으로 해서
    해더와 사용 컬럼 지정을 제공한다
    """
    format = "xlsx"

    def __init__(self, processing_type: str = 'array', encoding='utf-8', error_log='error.log'):
        """Excel 파일 핸들러 초기화

        Args:
            processing_type: multi-header dump 시에 index=False 가 지원되지 않아서,
                'array' 에서는 load 시 index로 추정되는 컬럼 별도 예외 drop 처리
                'object' 는 처리 없이 그대로 읽어들임
            encoding: 문서 인코딩 'utf-8' 기본값
            error_log: 에러 발생 시에 저장되는 파일 'error.log' 기본값
        """
        super().__init__(processing_type=processing_type, encoding=encoding, error_log=error_log)
        # self.engine = 'openpyxl' , 멀티헤더 처리 이슈로 분리해서 테스트 후 효과가 없었음
        self.read_engine = 'openpyxl'
        self.write_engine = 'openpyxl'

    def load(self, file_or_filename: Union[io.TextIOWrapper, io.BytesIO, io.BufferedIOBase, str],
             sheet_name=0, skiprows=0, header=0, nrows=None, usecols=None, **kwargs) -> Optional[pd.DataFrame]:
        """Excel 파일 읽기

        Args:
            file_or_filename (file-like object): file object or file name
            sheet_name: 1개의 sheet 만 지정. 0으로 시작하는 일련 번호 또는 쉬트 이름. None 이면 첫 쉬트 사용
            skiprows (Union[int, list]) : 데이터가 시작되는 row index 또는 배열 지정.
                header 보다 먼저 적용되고, header의 인덱스는 이 처리 결과의 인덱스

            header (Union[int, list]): 헤더로 사용될 1부터 시작되는 row index, 멀티헤더인 경우에는 [1, 2, 3] 형태로 사용
            nrows (int): skiprows 부터 N개의 데이터 row 만 읽을 경우 숫자 지정
            usecols (Union[int, list]): 전체 컬럼 사용시 None, 컬럼 번호나 이름의 리스트 [0, 1, 2] or ['foo', 'bar', 'baz']
        """
        mode = self._check_file_or_filename(file_or_filename)

        try:
            df = pd.read_excel(
                file_or_filename,
                sheet_name=sheet_name,
                header=header,
                skiprows=skiprows,
                nrows=nrows,
                usecols=usecols,
                parse_dates=True,
                **kwargs
            )

            if self.processing_type == CsvHandler.TYPE_ARRAY:
                if isinstance(df.columns, pd.MultiIndex):
                    df = df.drop([col for col in df.columns if 'Unnamed' in str(col)], axis=1)

                # 모든 column 값이 NaN 인 row는 제거
                df.dropna(how='all', inplace=True)
                self.pass_list.append(df)
            elif self.processing_type == CsvHandler.TYPE_OBJECT:
                return df
        except Exception as e:
            # debuging ord() expected a character bu string of length 4 found
            find_problematic_cells(file_or_filename)

            self.fail_list.append(str(file_or_filename))
            logger.error(f"{file_or_filename} load raise {e}")
            if self.processing_type == CsvHandler.TYPE_OBJECT:
                return None

    def loads(self, str_or_bytes: Union[str, bytes],
              sheet_name=0, header=0, skiprows=0, nrows=None, usecols=None, **kwargs):
        """문자열이나 bytes 에서 Excel 읽기

        Args:
            str_or_bytes (str, bytes): text 모드 string 또는 binary 모드 bytes
            sheet_name: 1개의 sheet 만 지정. 0으로 시작하는 일련 번호 또는 쉬트 이름. None 이면 첫 쉬트 사용
            header (Union[int, list]): 헤더로 사용될 1부터 시작되는 row index, 멀티헤더인 경우에는 [1, 2, 3] 형태로 사용
            skiprows (Union[int, list]) : 데이터가 시작되는 row index 또는 배열 지정.
                header 보다 먼저 적용되고, header의 인덱스는 이 처리 결과의 인덱스

            nrows (int): skiprows 부터 N개의 데이터 row 만 읽을 경우 숮자 지정
            usecols (Union[int, list]): 전체 컬럼 사용시 None, 컬럼 번호나 이름의 리스트 [0, 1, 2] or ['foo', 'bar', 'baz']
        """
        file_obj = None
        try:
            if isinstance(str_or_bytes, str):
                file_obj = io.StringIO(str_or_bytes)
            elif isinstance(str_or_bytes, bytes):
                file_obj = io.BytesIO(str_or_bytes)
            if file_obj:
                df = self.load(file_obj, sheet_name=sheet_name,
                          skiprows=skiprows, header=header, nrows=nrows, usecols=usecols, **kwargs)
                if self.processing_type == CsvHandler.TYPE_OBJECT:
                    return df
        except Exception as e:
            self.fail_list.append(str_or_bytes)
            logger.error(f"'{str_or_bytes}' loads raise {e}")
            if self.processing_type == CsvHandler.TYPE_OBJECT:
                return None

    #
    # def to_pandas() 는 data_list 에 dataframe 을 저장하는 방식이 CsvHandler 와 동일하여 따로 정의하지 않음
    #

    def dump(self, file_or_filename, sheet_name='Sheet1', data: pd.DataFrame = None, **kwargs) -> None:
        """데이터를 Excel 파일로 쓰기

        파일은 text, binary 모드 파일객체이거나 파일명 문자열
        Args:
            file_or_filename (file, str): 파일객체 또는 파일명
            sheet_name: 쉬트 이름.
            data: dataframe 으로 설정시 사용. 기존 유틸리티의 호환성을 위해서 남김
        """
        if self.processing_type == CsvHandler.TYPE_OBJECT:
            if data is None or not isinstance(data, pd.DataFrame):
                logger.error(f"processing_type '{self.processing_type}' need data parameter")
                raise TypeError(f"processing_type '{self.processing_type}' need data parameter")

        try:
            if data is None:
                df = self.to_pandas()
            else:
                df = data

            self._check_file_or_filename(file_or_filename)

            # index 처리 방법을 먼저 정함
            use_index = False
            if isinstance(df.columns, pd.MultiIndex):
                use_index = True
            elif not isinstance(df.index, pd.core.indexes.range.RangeIndex):
                use_index = True
            # elif self.processing_type == CsvHandler.TYPE_OBJECT:
            #    use_index = True

            # multi-header 문제떄문에 ExcelWriter 버전으로 대체. 효과는 없었엄. index=True 로 dump하고 후처리 방식으로 변경
            df.to_excel(
                file_or_filename,
                sheet_name=sheet_name,
                index=use_index,
                **kwargs
            )

            # write to Excel file
            # with pd.ExcelWriter(file_or_filename, engine=self.write_engine) as writer:
            #     df.to_excel(writer, sheet_name=sheet_name, index=use_index)

        except Exception as e:
            logger.error(f"'{str(file_or_filename)}' dump raise {e}")

    def dumps(self, sheet_name='Sheet1', data: pd.DataFrame = None, **kwargs) -> str:
        """데이터를 CSV 파일로 쓰기

        파일은 text, binary 모드 파일객체이거나 파일명 문자열
        Args:
            mode (str): 출력 모드 'text' 또는 'binary' 선택. 기본 'text' 는 문자열 출력
            sheet_name: 쉬트 이름.
            data: 내장 dataframe 대신 사용할 data. 기존 유틸리티의 호환성을 위해서 남김
        """
        file_obj = io.StringIO()

        try:
            self.dump(file_obj, sheet_name=sheet_name, data=data, **kwargs)
        except Exception as e:
            logger.error(f"{self} dumps raise: {e}")

        return file_obj.getvalue()
