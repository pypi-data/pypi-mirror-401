#
# GAMS - General Algebraic Modeling System Python API
#
# Copyright (c) 2017-2026 GAMS Development Corp. <support@gams.com>
# Copyright (c) 2017-2026 GAMS Software GmbH <support@gams.com>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#

import os
import warnings
from gams import transfer as gt
from gams.connect.agents._excel import ExcelAgent
from gams.connect.agents._excel import Workbook
from gams.connect.connectvalidator import ConnectValidator
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ExcelWriter(ExcelAgent):
    _index_parameter_map = {
        "cdim": "columnDimension",
        "columndimension": "columnDimension",
        "mergedcells": "mergedCells",
        "clearsheet": "clearSheet",
    }

    _EXCEL_ROW_LIMIT = 1048576
    _EXCEL_COL_LIMIT = 16384

    def __init__(self, cdb, inst, agent_index):
        super().__init__(cdb, inst, agent_index)
        self._parse_options(self._inst)
        if os.path.splitext(self._file)[1] != ".xlsx":
            self._connect_error("The ExcelWriter does support .xlsx files only.")

    def _check_openpyxl(self):
        try:
            opxl_version = list(map(int, openpyxl.__version__.split(".")))
            if opxl_version[0] < 3 or (
                opxl_version[0] == 3 and opxl_version[1] < 1
            ):  # check openpyxl<3.1.0
                warnings.warn(
                    f"The used openpyxl version is {openpyxl.__version__}. openpyxl<3.1.0 can lead to wrong data being exported. Upgrading to version 3.1.0 or later is recommended.",
                    category=Warning,
                )
        except ValueError:  # silence errors when converting to int
            pass

    def _parse_options(self, inst):
        inst["file"] = os.path.abspath(inst["file"])
        self._file = inst["file"]
        self._merged_cells = inst["mergedCells"]
        self._cdim = inst["columnDimension"]
        self._value_subs = inst["valueSubstitutions"]
        self._clear_sheet = inst["clearSheet"]
        self._trace = inst["trace"]
        self._symbols = inst["symbols"]
        self._index = inst["index"]
        self._engine = "openpyxl"

        self._toc = inst["tableOfContents"]
        if self._toc and not isinstance(self._toc, dict):
            v = ConnectValidator(
                {"tableOfContents": self._cdb.load_schema(self)["tableOfContents"]}
            )
            self._toc = {"tableOfContents": {}}
            self._toc = v.normalized(self._toc)
            self._toc = v.normalize_of_rules(self._toc)
            self._toc = self._toc["tableOfContents"]
            inst["tableOfContents"] = self._toc

    def _open(self):
        if os.path.exists(self._file):
            try:
                self._wb = Workbook(self._file, read_only=False, data_only=False)
            except PermissionError as e:
                self._connect_error(
                    str(e)
                    + "\nThe file may already be open and might need to be closed first."
                )
        else:
            if self._index:
                self._connect_error(
                    f"Workbook >{self._file}< needs to exist if used with index option."
                )
            self._wb = Workbook()

    def _write(self, df, rdim, cdim, sheet, nw_row, nw_col, merged_cells):
        row = nw_row + 1
        use_index = True if rdim > 0 else False
        use_header = True if cdim > 0 else False
        last_row_labels = [None] * rdim

        for idx, r in enumerate(
            dataframe_to_rows(df, index=use_index, header=use_header)
        ):
            r2 = list(r)
            if idx < cdim:
                if merged_cells:
                    for idx2, x in enumerate(r2[rdim:]):
                        if idx2 == 0:  # first iteration
                            merge_start = merge_end = nw_col + rdim + 1
                        elif x is None:
                            merge_end += 1
                            if idx2 == len(r2) - 1 - rdim:  # last iteration
                                if merge_start != merge_end:
                                    sheet.merge_cells(
                                        start_row=row,
                                        start_column=merge_start,
                                        end_row=row,
                                        end_column=merge_end,
                                    )
                        else:
                            if merge_start != merge_end:
                                sheet.merge_cells(
                                    start_row=row,
                                    start_column=merge_start,
                                    end_row=row,
                                    end_column=merge_end,
                                )
                            merge_start = merge_end = merge_end + 1
                else:
                    for idx2, x in enumerate(r2[rdim:]):
                        if x is None:
                            r2[idx2 + rdim] = r2[idx2 + rdim - 1]
            # remove extra empty row
            elif idx == cdim and rdim > 0:
                continue
            else:
                if merged_cells:
                    if idx == cdim + 1:  # first iteration
                        merge_start = [nw_row + cdim + 1] * rdim
                        merge_end = [nw_row + cdim + 1] * rdim
                    else:
                        for x in range(rdim):
                            if r2[x] is None:
                                merge_end[x] += 1
                                if idx == df.shape[0] + cdim:  # last iteration
                                    sheet.merge_cells(
                                        start_row=merge_start[x],
                                        start_column=x + nw_col + 1,
                                        end_row=merge_end[x],
                                        end_column=x + nw_col + 1,
                                    )
                            else:
                                if merge_start[x] != merge_end[x]:
                                    sheet.merge_cells(
                                        start_row=merge_start[x],
                                        start_column=x + nw_col + 1,
                                        end_row=merge_end[x],
                                        end_column=x + nw_col + 1,
                                    )
                                merge_start[x] = merge_end[x] = merge_end[x] + 1
                else:
                    for x in range(rdim):
                        if r2[x] is None:
                            r2[x] = last_row_labels[x]
                    last_row_labels = r2[:rdim]

            for col in range(len(r2)):
                if r2[col] is not None:
                    sheet.cell(row, col + nw_col + 1).value = r2[col]
            row += 1

    def _pivot_cdim_only(self, df, dim, value_text):
        cols = df.columns.values.tolist()
        df = df.sort_values(by=cols[:dim]).reset_index(drop=True)
        df["_first"] = value_text
        df = df.pivot(index=["_first"], columns=cols[:-1], values=[value_text])
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None], axis=0, inplace=True)  # remove index names
        df.rename_axis([None] * dim, axis=1, inplace=True)  # remove column index names
        return df

    def _pivot_rdim_only(self, df, dim, value_text):
        cols = df.columns.values.tolist()
        df["_first"] = value_text
        df = df.pivot(
            index=cols[:dim], columns=["_first"], values=[value_text]
        ).sort_index()
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None] * dim, axis=0, inplace=True)  # remove index names
        df.rename_axis([None], axis=1, inplace=True)  # remove column index names
        return df

    def _pivot_rdim_cdim(self, df, rdim, cdim, value_text):
        dim = rdim + cdim
        cols = df.columns.values.tolist()
        df = df.sort_values(by=cols[rdim:dim]).reset_index(drop=True)
        df = df.pivot(
            index=cols[:rdim], columns=cols[rdim:-1], values=[value_text]
        ).sort_index()
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None] * rdim, axis=0, inplace=True)  # remove index names
        df.rename_axis([None] * cdim, axis=1, inplace=True)  # remove column index names
        return df

    def _reshape_dataframe(self, df, dim, rdim, cdim, value):
        # turn integer column names into string column names to avoid problems reshaping if rdim=0
        df.columns = df.columns.map(str)
        if dim == 0:
            pass
        elif rdim == 0:
            df = self._pivot_cdim_only(df, dim, value)
        elif cdim == 0:
            df = self._pivot_rdim_only(df, dim, value)
        else:
            df = self._pivot_rdim_cdim(df, rdim, cdim, value)
        return df

    def _expected_excel_shape(self, df, rdim, cdim):
        nr_rows = df.shape[0] + cdim
        nr_cols = df.shape[1] + rdim
        return (nr_rows, nr_cols)

    def _validate_range(self, sym_name, df, rdim, cdim, nw_col, nw_row, se_col, se_row):
        required_rows, required_columns = self._expected_excel_shape(df, rdim, cdim)

        if required_rows + nw_row > self._EXCEL_ROW_LIMIT:
            self._connect_error(
                f"Attempting to write >{required_rows}< rows starting from row >{nw_row}< exceeds Excel's row limit of >{self._EXCEL_ROW_LIMIT}< for symbol >{sym_name}<."
            )

        if required_columns + nw_col > self._EXCEL_COL_LIMIT:
            self._connect_error(
                f"Attempting to write >{required_columns}< columns starting from column >{nw_col}< exceeds Excel's column limit of >{self._EXCEL_COL_LIMIT}< for symbol >{sym_name}<."
            )

        if se_row is not None:
            actual_rows = se_row - nw_row  # + 1
            if required_rows > actual_rows:
                self._connect_error(f"Data exceeds range for symbol >{sym_name}<.")

        if se_col is not None:
            actual_columns = se_col - nw_col  # + 1
            if required_columns > actual_columns:
                self._connect_error(f"Data exceeds range for symbol >{sym_name}<.")

    def _write_toc(self):
        if self._toc["sort"]:
            self._toc_symbols.sort(key=lambda x: x[0].name)
        sheet = self.sheet_by_name(self._toc["sheetName"], self._wb, True, True)
        self._wb.move_sheet(
            sheet, -self._wb.index(sheet)
        )  # move toc sheet to the beginning
        sheet.cell(1, 1).value = "Name"
        sheet.cell(1, 2).value = "Type"
        sheet.cell(1, 3).value = "Dimension"
        sheet.cell(1, 4).value = "Record Count"
        sheet.cell(1, 5).value = "Explanatory text"
        row = 2
        for sym, rng in self._toc_symbols:
            sheet.cell(row, 1).value = sym.name
            sheet.cell(row, 1).hyperlink = f"#{rng}"
            sheet.cell(row, 1).font = Font(underline="single", color="0563C1")
            sheet.cell(row, 2).value = (
                "Parameter" if isinstance(sym, gt.Parameter) else "Set"
            )
            sheet.cell(row, 3).value = sym.dimension
            sheet.cell(row, 4).value = sym.number_records
            sheet.cell(row, 5).value = sym.description
            row += 1

    def _write_symbols(self, symbols, validate=False):
        if validate:
            sym_schema = self._cdb.load_schema(self)["symbols"]["oneof"][1]["schema"][
                "schema"
            ]
            v = ConnectValidator(sym_schema)
        for i, sym in enumerate(symbols):
            if validate:
                sym = v.validated(sym)
                if sym is None:
                    self._connect_error(
                        f"Validation of item {i} in index failed: {v.errors}"
                    )
                sym = v.normalize_of_rules(sym)
            self._write_symbol(sym)

    def _create_symbol_instructions(self, rec):
        is_symbol = not None in (rec[1], rec[2])
        inst = {}
        if is_symbol:
            sym_type = str(rec[0]).lower().strip() if rec[0] else ""
            inst["name"] = rec[1].strip()
            inst["range"] = rec[2].strip()
            if self._trace > 1:
                self._cdb.print_log(
                    f"Parsing symbol >{inst['name']}< with >type: {sym_type}< and >range: {inst['range']}<."
                )
            if len(sym_type) > 0:
                sym_name = inst["name"]
                self._symbols_exist_cdb(sym_name, should_exist=True)
                symbol = self._cdb.container[sym_name]
                if sym_type == "par":
                    if not isinstance(symbol, gt.Parameter):
                        self._connect_error(
                            f"Type mismatch: Symbol type was specified as >{sym_type}< but the symbol is of type >{type(symbol).__name__}<"
                        )
                elif sym_type == "set":
                    if not isinstance(symbol, gt.Set):
                        self._connect_error(
                            f"Type mismatch: Symbol type was specified as >{sym_type}< but the symbol is of type >{type(symbol).__name__}<"
                        )
                else:
                    self._connect_error(
                        f"Unknown symbol type >{sym_type}<. Valid symbol types are >par< and >set<."
                    )
        return inst

    def _write_from_index(self):
        symbols = self.parse_index(self._index, self._wb, self._index_parameter_map)
        self._write_symbols(symbols, True)

    def _write_symbol(self, sym):
        sym_raw = sym.copy()
        self._update_sym_inst(sym, self._inst)

        sym["range"] = self._dict_get(sym, "range", sym["name"] + "!A1")
        sym["range"] = self.normalize_range(sym["range"])

        sym_name = sym["name"]
        self._symbols_exist_cdb(sym_name, should_exist=True)
        gt_sym = self._cdb.container[sym_name]
        dim = gt_sym.dimension
        if sym["columnDimension"] == "infer":
            sym["columnDimension"] = 1 if dim > 0 else 0

        sym_range = sym["range"]
        merged_cells = sym["mergedCells"]
        cdim = sym["columnDimension"]
        value_subs = sym["valueSubstitutions"]
        clear_sheet = sym["clearSheet"]

        if self._trace > 0:
            self._log_instructions(
                sym, sym_raw, description=f"Write symbol >{sym['name']}<:"
            )

        if self._trace > 2:
            self._cdb.print_log(
                f"Connect Container symbol >{sym_name}<:\n {gt_sym.records}\n"
            )

        if not isinstance(gt_sym, (gt.Set, gt.Parameter)):
            self._connect_error(
                f"Symbol type >{type(gt_sym)}< of symbol >{sym_name}< is not supported. Supported symbol types are set and parameter."
            )
        sym_type = "par" if isinstance(gt_sym, gt.Parameter) else "set"

        df = gt_sym.records.copy(deep=True)

        if isinstance(gt_sym, gt.Set):
            value = "element_text"
        elif isinstance(gt_sym, gt.Parameter):
            value = "value"

        if cdim > dim:
            self._connect_error(
                f"columnDimension >{cdim}< exceeds dimension of symbol >{sym_name}<."
            )
        rdim = dim - cdim

        df = self._apply_value_substitutions(
            df,
            value_subs,
            sym_type,
            sv_eps="EPS",
            sv_na="NA",
            sv_undef="UNDEF",
            sv_posinf="INF",
            sv_neginf="-INF",
        )
        if self._trace > 2:
            self._cdb.print_log(
                f"DataFrame after valueSubstitutions ({sym_name}):\n{df}\n"
            )

        if (
            value == "element_text" and rdim * cdim > 0
        ):  # replace empty element_text by Y when exporting a true table
            df.loc[df[value] == "", value] = "Y"

        sheet, nw_col, nw_row, se_col, se_row, toc_range = self.parse_range(
            sym_range, self._wb, clear_sheet, True
        )

        df = self._reshape_dataframe(df, dim, rdim, cdim, value)
        if self._trace > 2:
            self._cdb.print_log(f"DataFrame after reshaping ({sym_name}):\n{df}\n")

        self._validate_range(sym_name, df, rdim, cdim, nw_col, nw_row, se_col, se_row)

        self._toc_symbols.append((gt_sym, toc_range))

        if gt_sym.records is None or gt_sym.number_records == 0:
            return

        self._write(
            df,
            rdim,
            cdim,
            sheet,
            nw_row,
            nw_col,
            merged_cells,
        )

    def execute(self):
        if self._trace > 0:
            self._log_instructions(self._inst, self._inst_raw)
            self._describe_container(self._cdb.container, "Connect Container:")

        toc = self._inst_raw.get("tableOfContents")

        self._open()

        self._check_openpyxl()
        self._toc_symbols = []
        try:
            if self._index:
                self._write_from_index()
            elif self._symbols == "all":
                self._symbols = [
                    {"name": s[0]}
                    for s in self._cdb.container
                    if isinstance(s[1], (gt.Parameter, gt.Set))
                ]
                self._write_symbols(self._symbols, True)
            else:
                self._write_symbols(self._symbols)
            if self._toc:
                self._write_toc()
        finally:
            self._wb.close()
        self._close()

    def _close(self):
        if len(self._wb.sheetnames) == 0:
            self._cdb.print_log(f"No sheets in Excel file >{self._file}<. Skipping.")
        else:
            try:
                self._wb.save(self._file)
            except PermissionError as e:
                self._connect_error(
                    str(e)
                    + "\nThe file may already be open and might need to be closed first."
                )
