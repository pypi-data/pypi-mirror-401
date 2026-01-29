#
# GAMS - General Algebraic Modeling System Python API
#
# Copyright (c) 2017-2026 GAMS Development Corp. <support@gams.com>
# Copyright (c) 2017-2026 GAMS Software GmbH <support@gams.com>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#

import datetime
import os
from gams.connect.agents.connectagent import ConnectAgent
from gams.core.gmd import *
import gams.transfer as gt
import openpyxl
import openpyxl.cell.cell
import openpyxl.cell.read_only
import openpyxl.utils.datetime
import pandas as pd


class RawExcelReader(ConnectAgent):

    def __init__(self, cdb, inst, agent_index):
        super().__init__(cdb, inst, agent_index)
        self._parse_options(self._inst)

    def _parse_options(self, inst):
        inst["file"] = os.path.abspath(inst["file"])
        self._xlsx_file = inst["file"]
        self._read_only = not inst["mergedCells"]
        self._s_name = inst["sName"]
        self._w_name = inst["wName"]
        self._ws_name = inst["wsName"]
        self._r_name = inst["rName"]
        self._c_name = inst["cName"]
        self._vs_name = inst["vsName"]
        self._vu_name = inst["vuName"]
        self._vf_name = inst["vfName"]
        self._rowLabel = inst["rowLabel"]
        self._colLabel = inst["columnLabel"]
        self._trace = inst["trace"]
        self._sheetLabel = inst["sheetLabel"]

    def _utf8_lead_byte(self, b):
        # A UTF-8 intermediate byte starts with the bits 10xxxxxx.
        return (b & 0xC0) != 0x80

    def _utf8_byte_truncate(self, text, max_bytes):
        # If text[max_bytes] is not a lead byte, back up until a lead byte is
        # found and truncate before that character.
        utf8 = text.encode("utf8")
        if len(utf8) <= max_bytes:
            return text
        i = max_bytes
        while i > 0 and not self._utf8_lead_byte(utf8[i]):
            i -= 1
        return utf8[:i].decode("utf8")

    def _get_cell_value(self, sheet, cell):
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            return cell.value

        # "Oh no, the cell is merged!"
        for rng in sheet.merged_cells.ranges:
            if cell.coordinate in rng:
                return rng.start_cell.value

        self._connect_error("Merged cell is not in any merge range!")

    def _addUel(self, uel, keys):
        nuel = self._trunc_org.get(uel, -1)
        if nuel != -1:
            keys[3] = self._uel_list[nuel - 1]
            self._vu.append((*keys, uel))
            return True
        nuel = self._trunc_issued.get(uel, -1)
        if nuel == -1 and len(uel.encode("utf8")) < GLOBAL_UEL_IDENT_SIZE:
            if uel.find("'") != -1 and uel.find('"') != -1:
                return False
            if len(uel) and min(uel) < " ":
                return False
            self._uel_list.append(uel)
            nuel = len(self._uel_list)
            keys[3] = uel
            self._vu.append((*keys, uel))
            self._trunc_org[uel] = nuel
            self._trunc_issued[uel] = nuel
            return True
        # We need a new name for uel because it is too long or has already been issued as a trunc name
        next_trunc_nr = 0
        while True:
            next_trunc_nr += 1
            xs = "~" + str(next_trunc_nr)
            newname = (
                self._utf8_byte_truncate(uel, GLOBAL_UEL_IDENT_SIZE - len(xs) - 1) + xs
            )
            if (self._trunc_issued.get(newname, -1) == -1) and (
                self._trunc_org.get(newname, -1) == -1
            ):
                break

        if newname.find("'") != -1 and newname.find('"') != -1:
            return False
        if len(newname) and min(newname) < " ":
            return False
        self._uel_list.append(newname)
        nuel = len(self._uel_list)
        keys[3] = newname
        self._vu.append((*keys, self._utf8_byte_truncate(uel, GMS_SSSIZE - 2)))
        self._trunc_org[uel] = nuel
        self._trunc_issued[newname] = nuel
        return True

    def execute(self):
        if self._trace > 0:
            self._log_instructions(self._inst, self._inst_raw)
            self._describe_container(self._cdb.container, "Connect Container (before):")

        self._symbols_exist_cdb(self._s_name)
        self._S = self._cdb._container.addSet(
            self._s_name, "*", description="Workbook sheets"
        )

        self._symbols_exist_cdb(self._w_name)
        self._W = self._cdb._container.addSet(
            self._w_name, "*", description="Workbook sheets by name"
        )

        self._symbols_exist_cdb(self._ws_name)
        self._WS = self._cdb._container.addSet(
            self._ws_name,
            [self._S, self._W],
            description="Workbook map",
        )

        self._symbols_exist_cdb(self._r_name)
        self._R = self._cdb._container.addSet(self._r_name, "*", description="Rows")

        self._symbols_exist_cdb(self._c_name)
        self._C = self._cdb._container.addSet(self._c_name, "*", description="Columns")

        self._symbols_exist_cdb(self._vs_name)
        self._VS = self._cdb._container.addSet(
            self._vs_name,
            [self._S, self._R, self._C],
            description="Cells with explanatory text",
        )

        self._symbols_exist_cdb(self._vu_name)
        self._VU = self._cdb._container.addSet(
            self._vu_name,
            [self._S, self._R, self._C, "*"],
            description="Cells with potential GAMS label",
        )

        self._symbols_exist_cdb(self._vf_name)
        self._VF = self._cdb._container.addParameter(
            self._vf_name,
            [self._S, self._R, self._C],
            description="Cells with numerical value",
        )

        self._trunc_org = {}
        self._trunc_issued = {}
        self._uel_list = []
        self._next_trunc_nr = 1
        self._vu = []

        wb = openpyxl.load_workbook(
            filename=self._xlsx_file, read_only=self._read_only, data_only=True
        )

        if self._trace > 1:
            self._cdb.print_log(f"Number of sheets: {len(wb.sheetnames)}")

        self._S.setRecords(
            [(f"{self._sheetLabel}{s[0]+1}", s[1]) for s in enumerate(wb.sheetnames)]
        )
        try:
            max_row = max([s.max_row for s in wb.worksheets])
        except:
            for s in wb.worksheets:
                s.calculate_dimension(force=True)
            max_row = max([s.max_row for s in wb.worksheets])

        if self._trace > 1:
            self._cdb.print_log(f"Max row: {max_row}")
        self._R.setRecords([f"{self._rowLabel}{r+1}" for r in range(max_row)])
        max_col = max([s.max_column for s in wb.worksheets])
        if self._trace > 1:
            self._cdb.print_log(f"Max column: {max_col}")
        self._C.setRecords([f"{self._colLabel}{c+1}" for c in range(max_col)])
        self._W.setRecords(wb.sheetnames)
        self._WS.setRecords(
            [(f"{self._sheetLabel}{s[0]+1}", s[1]) for s in enumerate(wb.sheetnames)]
        )

        # Now go over the cells
        skeys = [None] * 4
        vf = []
        vs = []

        for s in enumerate(wb.worksheets):
            skeys[0] = f"{self._sheetLabel}{s[0]+1}"
            for r in s[1].iter_rows():
                if not len(r):
                    continue
                skeys[1] = None
                for c in r:
                    if isinstance(c, openpyxl.cell.read_only.EmptyCell):
                        continue
                    if not self._read_only:  # potentially merged cells
                        cell_value = self._get_cell_value(s[1], c)
                    else:
                        cell_value = c.value
                    if cell_value == None:
                        continue

                    if skeys[1] == None:
                        skeys[1] = f"{self._rowLabel}{c.row}"
                    skeys[2] = f"{self._colLabel}{c.column}"

                    if self._trace > 2:
                        self._cdb.print_log(
                            f"Cell {self._sheetLabel}{s[0]+1}!{self._rowLabel}={c.row} {self._colLabel}={c.column}"
                        )
                    if isinstance(cell_value, datetime.datetime):
                        if self._trace > 2:
                            self._cdb.print_log(
                                f"  Date cell: {openpyxl.utils.datetime.to_excel(cell_value)}"
                            )
                        vf.append(
                            (
                                *skeys[:-1],
                                openpyxl.utils.datetime.to_excel(cell_value),
                            )
                        )
                        self._addUel(str(cell_value), skeys)
                        continue

                    if isinstance(cell_value, str):
                        if self._trace > 2:
                            self._cdb.print_log(f"  Str cell: >{cell_value}<")
                        vs.append(
                            (
                                *skeys[:-1],
                                self._utf8_byte_truncate(cell_value, GMS_SSSIZE - 2),
                            )
                        )

                    try:
                        dval = float(cell_value)
                        if pd.isna(dval):
                            dval = gt.SpecialValues.NA
                        if self._trace > 2:
                            self._cdb.print_log(f"  Float cell: {dval}")
                    except:
                        dval = None
                        if isinstance(cell_value, str):
                            CELL_VALUE = cell_value.upper()
                            if CELL_VALUE == "EPS":
                                dval = gt.SpecialValues.EPS
                            elif CELL_VALUE == "NA":
                                dval = gt.SpecialValues.NA
                            elif CELL_VALUE == "UNDEF":
                                dval = gt.SpecialValues.UNDEF
                        if self._trace > 2 and isinstance(dval, float):
                            self._cdb.print_log(f"  Converted float: {dval}")
                    if isinstance(dval, float):
                        vf.append((*skeys[:-1], dval))

                    try:
                        sval = str(cell_value)
                        if self._trace > 2 and isinstance(dval, float):
                            self._cdb.print_log(f"  Converted string: >{sval}<")
                    except:
                        continue
                    self._addUel(sval, skeys)

        if len(vf):
            self._VF.setRecords(vf)
        if len(vs):
            self._VS.setRecords(vs)
        if len(self._vu):
            self._VU.setRecords(self._vu)

        # For symbols with None records, empty df is assigned
        self._transform_sym_none_to_empty(self._VF)
        self._transform_sym_none_to_empty(self._VS)
        self._transform_sym_none_to_empty(self._VU)

        wb.close()

        if self._trace > 0:
            self._describe_container(self._cdb.container, "Connect Container (after):")
