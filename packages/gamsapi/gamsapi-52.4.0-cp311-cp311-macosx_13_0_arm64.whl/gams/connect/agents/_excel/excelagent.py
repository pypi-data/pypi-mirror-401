#
# GAMS - General Algebraic Modeling System Python API
#
# Copyright (c) 2017-2026 GAMS Development Corp. <support@gams.com>
# Copyright (c) 2017-2026 GAMS Software GmbH <support@gams.com>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#

from abc import abstractmethod
from gams.connect.agents.connectagent import ConnectAgent
import numpy as np
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


class ExcelAgent(ConnectAgent):
    @abstractmethod
    def __init__(self, cdb, inst, agent_index):
        super().__init__(cdb, inst, agent_index)

        def _to_int_if_whole_vec(x):
            if isinstance(x, float) and x.is_integer():
                return int(x)
            return x

        self._to_int_if_whole = np.vectorize(_to_int_if_whole_vec, otypes=[object])

    @abstractmethod
    def _create_symbol_instructions(self, rec): ...

    def _key_map(self, key, index_parameter_map):
        if key is None:
            return None
        key = key.lower().strip()
        if key in index_parameter_map:
            return index_parameter_map[key]
        if self._trace > 1 and key != "":
            self._cdb.print_log(f"Ignoring unsupported option >{key}<.")
        return ""

    def _parse_key_value(self, value, index_parameter_map):
        args = value
        args = args.replace("=", " ")
        args = args.split(" ")
        args = list(
            filter(lambda x: x != " ", args)
        )  # NOTE: This part of code doesnt allow spaces when specifying key-value pairs (e.g. ignoreColumns = 1:3 raises an error  -> ignoreColumns=1:3)
        ret = {}
        for args_idx in range(0, len(args), 2):
            key = args[args_idx].strip()
            key = self._key_map(key, index_parameter_map)
            if key is not None and len(key) != 0:
                if args_idx + 1 >= len(args):
                    self._connect_error(
                        f"Error parsing >{value}< from index sheet. Option >{key}< does not have a value specified."
                    )
                v = args[args_idx + 1].strip()
                if v is not None:
                    v = self._value_map(key, v, index_parameter_map)
                    ret[key] = v
        return ret

    def _value_map(self, key, value, index_parameter_map):
        if isinstance(value, str):
            value = value.strip()
        bool_mapping = {
            # 1: True,
            True: True,
            "true": True,
            "1": True,
            # 0: False,
            False: False,
            "false": False,
            "0": False,
        }
        m = {}
        supported_keys = dict.fromkeys(index_parameter_map.values())
        for k in supported_keys:
            if k == "autoMerge":
                m["autoMerge"] = bool_mapping
            if k == "mergedCells":
                m["mergedCells"] = bool_mapping
            if k == "ignoreText":
                m["ignoreText"] = bool_mapping
            if k == "clearSheet":
                m["clearSheet"] = bool_mapping
        v = value
        if key in supported_keys:
            if key in ["autoMerge", "ignoreText", "mergedCells", "clearSheet"]:
                if isinstance(value, str):
                    value = value.lower()
                v = m[key].get(value, None)
                if v is None:
                    self._connect_error(
                        f"Invalid value >{value}< for {key}. Allowed values are {', '.join(str(x) for x in m[key].keys())}."
                    )
            elif key == "ignoreRows":
                if isinstance(value, str):
                    l = value.split(",")
                    l = [x.strip() for x in l]
                    v = []
                    # [9, "4:7", 11] -> [9, 4, 5, 6, 7, 11]
                    for x in l:
                        if x.isnumeric():
                            v.append(int(x))
                        elif ":" in x:
                            v.extend(self._parse_rows_range(x))
                        else:
                            self._connect_error(
                                f"Value >{x}< for ignoreRows in index file need to be either numeric or in range format e.g. 3:7."
                            )
                else:
                    v = [value]
            elif key == "ignoreColumns":
                if isinstance(value, str):
                    l = value.split(",")
                    l = [x.strip() for x in l]
                    v = []
                    for x in l:
                        if x.isnumeric():
                            v.append(int(x))
                        elif ":" in x:
                            v.extend(self._parse_columns_range(x))
                        else:
                            v.append(column_index_from_string(x))
                else:
                    v = [value]
            elif isinstance(value, str) and value.isnumeric():
                v = int(value)
        return v

    def _coords_to_row_col(self, coordinates):
        c = coordinate_from_string(coordinates)
        return c[1], column_index_from_string(c[0])

    def normalize_range(self, sym_range):
        if sym_range.endswith("!"):
            sym_range += "A1"
        return sym_range

    def _split_range(self, sym_range, wb):
        if "!" not in sym_range:  # named range
            rng_dict = {x.lower(): x for x in wb.defined_names}
            sym_range_lower = sym_range.lower()
            if sym_range_lower not in rng_dict:
                self._connect_error(f"Named range >{sym_range}< does not exist.")
            rng = wb.resolve_named_range(rng_dict[sym_range_lower])
            if len(rng) == 0:
                self._connect_error(
                    f"Named range >{sym_range}< with invalid reference."
                )
            if len(rng) > 1:
                self._connect_error(
                    f"Named range is not contiguous: {sym_range} -> {rng}"
                )
            rng = rng[0]
            resolved_range = "!".join(rng)
            sym_range = resolved_range
        sym_range = self.normalize_range(sym_range)
        return sym_range.split("!")

    def _range_to_coords(self, rng):
        nw, se = rng.split(":") if ":" in rng else (rng, None)
        nw_row, nw_col = self._coords_to_row_col(nw)
        se_col = se_row = None
        if se is not None:
            se_row, se_col = self._coords_to_row_col(se)
        return nw_col - 1, nw_row - 1, se_col, se_row

    def parse_range(self, sym_range, wb, clear_sheet=False, create_missing=False):
        sheet, rng = self._split_range(sym_range, wb)
        toc_range = f"'{sheet}'!{rng.split(':')[0]}"
        sheet = self.sheet_by_name(sheet, wb, clear_sheet, create_missing)
        nw_col, nw_row, se_col, se_row = self._range_to_coords(rng)
        return sheet, nw_col, nw_row, se_col, se_row, toc_range

    def parse_index(self, index, wb, index_parameter_map):
        if self._trace > 1:
            self._cdb.print_log(f"Parsing index sheet using range >{index}<")
        sheet, nw_col, nw_row, se_col, se_row, _ = self.parse_range(index, wb)
        data = wb.get_sheet_data(sheet)
        if self._engine == "xlwings":
            data = self._to_int_if_whole(data)

        data = data[nw_row:se_row, nw_col:se_col]
        header = data[0]
        for idx, h in enumerate(header):
            header[idx] = self._key_map(h, index_parameter_map)

        symbols = []
        global_par = {}
        for rec in data[1:]:
            local_par = self._create_symbol_instructions(rec)
            is_symbol = len(local_par) > 0
            for idx, h in enumerate(header[3:]):
                key = self._key_map(h, index_parameter_map)
                value = rec[idx + 3]
                if isinstance(value, str):
                    value = value.strip()
                params = {}
                if key is None:  # parse arbitrary key-value pairs
                    if value is not None:
                        params = self._parse_key_value(value, index_parameter_map)
                elif len(key) == 0:
                    continue
                elif value is not None:
                    params = {key: self._value_map(key, value, index_parameter_map)}

                for k, v in params.items():
                    if is_symbol:
                        if self._trace > 1:
                            self._cdb.print_log(
                                f"Adding option >{k}: {v}< for symbol >{local_par['name']}<."
                            )
                        local_par[k] = v
                    else:
                        if self._trace > 1:
                            self._cdb.print_log(f"\nAdding root option >{k}: {v}<.")
                        global_par[k] = v

            if is_symbol:
                sym = global_par.copy()
                sym.update(local_par)
                symbols.append(sym)
        return symbols

    def sheet_by_name(self, sheet, wb, clear_sheet=False, create_missing=False):
        for idx, s in enumerate(wb.sheetnames):
            if sheet.lower() == s.lower():
                if clear_sheet:
                    wb.delete_sheet(s)
                    return wb.create_sheet(s, idx)
                return wb.get_sheet(s)
        if create_missing:
            return wb.create_sheet(sheet)
        else:
            self._connect_error(f"Sheet >{sheet}< not found.")

    def _parse_rows_range(self, rng):
        """
        Parse a range string in the format 'start:end' and return a list of all integers in the range.
        e.g. 3:7 -> [3, 4, 5, 6, 7]
        """
        r_components = rng.split(":")
        r_components = [comp.strip() for comp in r_components]
        if len(r_components) != 2:
            self._connect_error(
                f"Invalid range in ignoreRows >{rng}<. Must be specified in the format 'start:end'"
            )

        if not all(x.isnumeric() for x in r_components):
            self._connect_error(
                f"Invalid range in ignoreRows >{rng}<. Both 'start' and 'end' must be integers in 'start:end'"
            )

        r_from_to = list(map(int, r_components))

        if r_from_to[0] > r_from_to[1]:
            self._connect_error(
                f"Invalid range in ignoreRows >{rng}<. 'start' >{r_components[0]}< must be less than or equal to 'end' >{r_components[1]}< in 'start:end'"
            )

        r_all = list(range(r_from_to[0], r_from_to[1] + 1))
        return r_all

    def _parse_columns_range(self, rng):
        """
        Parse a range string in the format 'start:end' and return a list of all integers in the range.
        e.g.    C:G -> [3, 4, 5, 6, 7] , 2 : E -> [2, 3, 4, 5] , D  :7 -> [4, 5, 6, 7]
        """
        r_components = rng.split(":")
        r_components = [comp.strip() for comp in r_components]
        if len(r_components) != 2:
            self._connect_error(
                f"Invalid range in ignoreColumns >{rng}<. Must be specified in the format 'start:end'"
            )

        if not all(x.isnumeric() or x.isalpha() for x in r_components):
            self._connect_error(
                f"Invalid range in ignoreColumns >{rng}<. Both 'start' and 'end' must be either integers or letters in 'start:end'"
            )

        r_components = [
            (column_index_from_string(comp) if comp.isalpha() else int(comp))
            for comp in r_components
        ]

        if r_components[0] > r_components[1]:
            self._connect_error(
                f"Invalid range in ignoreColumns >{rng}<. 'start' >{r_components[0]}< must be less than or equal to 'end' >{r_components[1]}< in 'start:end'"
            )

        r_from_to = list(map(int, r_components))

        r_all = list(range(r_from_to[0], r_from_to[1] + 1))
        return r_all
