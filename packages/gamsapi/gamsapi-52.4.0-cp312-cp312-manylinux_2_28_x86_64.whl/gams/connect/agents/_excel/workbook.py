#
# GAMS - General Algebraic Modeling System Python API
#
# Copyright (c) 2017-2026 GAMS Development Corp. <support@gams.com>
# Copyright (c) 2017-2026 GAMS Software GmbH <support@gams.com>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#

import sys
import numpy as np


class Workbook:
    def __init__(self, file=None, engine="openpyxl", read_only=False, data_only=False):
        self._engine = engine
        self._file = file

        if self._engine == "xlwings":
            import xlwings

            self._app = xlwings.App(visible=False, add_book=False)
            self._wb = self._app.books.open(self._file, read_only=read_only)
        elif self._engine == "openpyxl":
            import openpyxl

            if self._file is None:
                self._wb = openpyxl.Workbook()
                # remove default sheet
                self._wb.remove(self._wb.active)
            else:
                self._wb = openpyxl.load_workbook(
                    self._file, read_only=read_only, data_only=data_only
                )
        else:
            raise Exception("Invalid engine.")

    def close(self):
        if self._engine == "xlwings":
            if self._app is not None:
                self._app.quit()
        elif self._engine == "openpyxl":
            if self._wb is not None:
                self._wb.close()
        else:
            raise Exception("Invalid engine.")

    def save(self, path):
        if self._engine == "xlwings":
            self._wb.save(path)
        elif self._engine == "openpyxl":
            self._wb.save(path)
        else:
            raise Exception("Invalid engine.")

    def get_sheet(self, sheet_name):
        if self._engine == "xlwings":
            return self._wb.sheets[sheet_name]
        elif self._engine == "openpyxl":
            return self._wb[sheet_name]
        else:
            raise Exception("Invalid engine.")

    def index(self, worksheet):
        if self._engine == "xlwings":
            raise Exception("Not implemented.")
        elif self._engine == "openpyxl":
            return self._wb.index(worksheet)
        else:
            raise Exception("Invalid engine.")

    def move_sheet(self, sheet, offset=0):
        if self._engine == "xlwings":
            raise Exception("Not implemented.")
        elif self._engine == "openpyxl":
            self._wb.move_sheet(sheet, offset)
        else:
            raise Exception("Invalid engine.")

    def get_sheet_data(self, sheet):
        if self._engine == "xlwings":
            last_cell = sheet.used_range.last_cell
            full_range = sheet.range(sheet.range("A1"), last_cell)
            return full_range.options(
                np.array,
                ndim=2,
                dtype=object,
                empty=None,
                err_to_str=True,
            ).value
        elif self._engine == "openpyxl":
            return np.array(
                list(sheet.values), dtype=object
            )  # dtype=object is required to not convert int values (e.g. 1) to float automatically (e.g. 1.0)
        else:
            raise Exception("Invalid engine.")

    def delete_sheet(self, sheet_name):
        if self._engine == "xlwings":
            self._wb.sheets[sheet_name].delete()
        elif self._engine == "openpyxl":
            del self._wb[sheet_name]
        else:
            raise Exception("Invalid engine.")

    def create_sheet(self, sheet_name, index=None):
        if self._engine == "xlwings":
            raise Exception("Not imlemented.")
        elif self._engine == "openpyxl":
            return self._wb.create_sheet(sheet_name, index)
        else:
            raise Exception("Invalid engine.")

    def resolve_named_range(self, sym_range):
        if self._engine == "xlwings":
            nr = self._wb.names[sym_range].refers_to_range
            return [(nr.sheet.name, nr.address)]
        elif self._engine == "openpyxl":
            return list(self.defined_names[sym_range].destinations)
        else:
            raise Exception("Invalid engine.")

    @property
    def defined_names(self):
        if self._engine == "xlwings":
            return [x.name for x in self._wb.names]
        elif self._engine == "openpyxl":
            return self._wb.defined_names
        else:
            raise Exception("Invalid engine.")

    @property
    def sheetnames(self):
        if self._engine == "xlwings":
            return [s.name for s in self._wb.sheets]
        elif self._engine == "openpyxl":
            return self._wb.sheetnames
        else:
            raise Exception("Invalid engine.")
