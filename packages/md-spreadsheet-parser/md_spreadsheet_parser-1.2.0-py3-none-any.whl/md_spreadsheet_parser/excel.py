"""
Excel/TSV/CSV parsing module with merged cell and hierarchical header support.

This module provides functions to parse tabular data exported from Excel
(TSV/CSV format) or directly from openpyxl Worksheets, handling:
- Merged cells (forward-fill empty cells)
- Hierarchical headers (flatten to "Parent - Child" format)
"""

import csv
import io
from typing import Any, TYPE_CHECKING, Union

from .models import Table
from .schemas import ExcelParsingSchema, DEFAULT_EXCEL_SCHEMA

# --- Optional openpyxl support ---
if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# Type alias for parse_excel source parameter
# Note: Worksheet is only available at runtime if openpyxl is installed
ExcelSource = Union[str, list[list[str]], "Worksheet"]

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None  # type: ignore


def _parse_tsv(text: str, delimiter: str) -> list[list[str]]:
    """Parse TSV/CSV text into a 2D list using Python's csv module."""
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return list(reader)


def _forward_fill(row: list[str]) -> list[str]:
    """Fill empty cells with the previous non-empty value (left-to-right)."""
    result = []
    prev = ""
    for cell in row:
        if cell.strip():
            prev = cell
        result.append(prev)
    return result


def _flatten_headers(
    parent_row: list[str], child_row: list[str], separator: str
) -> list[str]:
    """
    Flatten 2-row headers into single row.
    Format: "Parent - Child" if Parent differs from Child, else just Child.
    """
    headers = []
    max_len = max(len(parent_row), len(child_row))

    for i in range(max_len):
        parent = parent_row[i] if i < len(parent_row) else ""
        child = child_row[i] if i < len(child_row) else ""

        if parent and child and parent != child:
            headers.append(f"{parent}{separator}{child}")
        else:
            headers.append(child if child else parent)

    return headers


def _safe_str(value: Any) -> str:
    """
    Convert value to string, handling None and integer-floats cleanly.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def parse_excel_text(
    rows: list[list[str]],
    schema: ExcelParsingSchema = DEFAULT_EXCEL_SCHEMA,
) -> Table:
    """
    Parse a 2D string array into a Table with merged cell and header handling.

    Args:
        rows: 2D list of strings (e.g., from csv.reader or worksheet iteration).
        schema: Configuration for header processing.

    Returns:
        Table object with processed headers and data rows.
    """
    if not rows:
        return Table(headers=None, rows=[])

    if schema.header_rows == 1:
        # Single header row
        header_row = rows[0]
        if schema.fill_merged_headers:
            header_row = _forward_fill(header_row)
        headers = header_row
        data_rows = rows[1:]

    elif schema.header_rows == 2:
        # Two header rows: Parent-Child flattening
        if len(rows) < 2:
            # Not enough rows for 2-row header
            return Table(headers=rows[0] if rows else None, rows=[])

        parent_row = rows[0]
        child_row = rows[1]

        if schema.fill_merged_headers:
            parent_row = _forward_fill(parent_row)

        headers = _flatten_headers(parent_row, child_row, schema.header_separator)
        data_rows = rows[2:]

    else:
        # Should not reach here due to schema validation
        raise ValueError(f"Invalid header_rows: {schema.header_rows}")

    # Convert data_rows to list[list[str]] ensuring all are strings
    processed_rows = [[_safe_str(cell) for cell in row] for row in data_rows]

    return Table(headers=headers, rows=processed_rows)


def parse_excel(
    source: ExcelSource,
    schema: ExcelParsingSchema = DEFAULT_EXCEL_SCHEMA,
) -> Table:
    """
    Parse Excel data from various sources.

    Args:
        source: One of:
            - openpyxl.Worksheet (if openpyxl is installed)
            - str: TSV/CSV text content
            - list[list[str]]: Pre-parsed 2D array
        schema: Configuration for parsing.

    Returns:
        Table object with processed headers and data.

    Raises:
        TypeError: If source type is not supported.
    """
    rows: list[list[str]]

    # Check for openpyxl Worksheet (duck typing via hasattr)
    if HAS_OPENPYXL and hasattr(source, "iter_rows"):
        # At runtime, source is a Worksheet with iter_rows method
        ws: Any = source
        rows = [
            [_safe_str(cell) for cell in row] for row in ws.iter_rows(values_only=True)
        ]

    # Check for string (TSV/CSV content)
    elif isinstance(source, str):
        rows = _parse_tsv(source, schema.delimiter)

    # Check for pre-parsed 2D array
    elif isinstance(source, list):
        # Assume it's already list[list[str]]
        rows = source

    else:
        supported = "openpyxl.Worksheet, str, or list[list[str]]"
        if not HAS_OPENPYXL:
            supported = (
                "str or list[list[str]] (install openpyxl for Worksheet support)"
            )
        raise TypeError(
            f"Unsupported source type: {type(source).__name__}. Expected {supported}."
        )

    return parse_excel_text(rows, schema)
