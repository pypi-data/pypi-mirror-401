from md_spreadsheet_parser import parse_excel, ExcelParsingSchema

import openpyxl
from openpyxl import Workbook


def test_openpyxl_parsing_basic():
    """Verify parsing a simple openpyxl Worksheet."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "TestSheet"

    # Header
    ws["A1"] = "Name"
    ws["B1"] = "Age"

    # Data
    ws["A2"] = "Alice"
    ws["B2"] = 30

    ws["A3"] = "Bob"
    ws["B3"] = 25

    table = parse_excel(ws)

    assert table.headers == ["Name", "Age"]
    assert table.rows == [["Alice", "30"], ["Bob", "25"]]


def test_openpyxl_parsed_merged_headers():
    """Verify parsing merged headers works with real openpyxl objects."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    # Merged Header Row: "Category" spans A-C
    # A1 = "Category", B1=Merged, C1=Merged, D1="Info"
    ws["A1"] = "Category"
    ws["D1"] = "Info"
    ws.merge_cells("A1:C1")

    # Sub-headers
    ws["A2"] = "Type"
    ws["B2"] = "Subtype"
    ws["C2"] = "Code"
    ws["D2"] = "Description"

    # Data
    ws["A3"] = "Fruit"
    ws["B3"] = "Citrus"
    ws["C3"] = "F001"
    ws["D3"] = "Lemon"

    schema = ExcelParsingSchema(header_rows=2)
    table = parse_excel(ws, schema)

    # Expected behavior: "Category" is forward-filled to A, B, C then joined with subheaders
    expected_headers = [
        "Category - Type",
        "Category - Subtype",
        "Category - Code",
        "Info - Description",
    ]

    assert table.headers == expected_headers
    assert table.rows[0] == ["Fruit", "Citrus", "F001", "Lemon"]


def test_openpyxl_empty_sheet():
    """Verify empty sheet handling."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    # No data

    table = parse_excel(ws)
    assert table.headers is None
    assert table.rows == []


def test_parse_real_excel_file():
    """
    Verify parsing of tests/fixtures/test_data.xlsx.
    The file contains merged headers, 2-row hierarchy, and mixed data types.
    Structure:
    - Row 1: #, Sales (merged 3 cols), Product Info (merged 2 cols), Status (merged vertical?)
    - Row 2: ID, Region, Year, Quarter, Category, Name, (Empty due to vertical merge)
    """
    import os

    fixture_path = os.path.join(os.path.dirname(__file__), "../fixtures/test_data.xlsx")

    wb = openpyxl.load_workbook(fixture_path, data_only=True)
    ws = wb.active
    assert ws is not None

    schema = ExcelParsingSchema(header_rows=2)
    table = parse_excel(ws, schema)

    # Expected Headers logic:
    # A: "#" (R1) vs "ID" (R2) -> "# - ID"
    # B: "Sales" (R1) vs "Region" (R2) -> "Sales - Region"
    # C: "Sales" (From R1 merge) vs "Year" (R2) -> "Sales - Year"
    # D: "Sales" (From R1 merge) vs "Quarter" (R2) -> "Sales - Quarter"
    # E: "Product Info" (R1) vs "Category" (R2) -> "Product Info - Category"
    # F: "Product Info" (From R1 merge) vs "Name" (R2) -> "Product Info - Name"
    # G: "Status" (R1) vs "" (R2 vertical merge empty) -> "Status"

    expected_headers = [
        # Note: The text provided by user suggests "#" and "ID" might be distinct.
        # If A1 is "#" and A2 is "ID", we get "# - ID".
        "# - ID",
        "Sales - Region",
        "Sales - Year",
        "Sales - Quarter",
        "Product Info - Category",
        "Product Info - Name",
        "Status",
    ]

    assert table.headers == expected_headers

    # Row 1 check (ID=1)
    # 1, North, 2024, Q1, Electronics, Laptop, In Progress
    assert table.rows[0] == [
        "1",
        "North",
        "2024",
        "Q1",
        "Electronics",
        "Laptop",
        "In Progress",
    ]

    # Row 3 check (ID=3) - South, Furniture
    assert table.rows[2] == [
        "3",
        "South",
        "2024",
        "Q1",
        "Furniture",
        "Chair",
        "Complete",
    ]

    # Row 5 check (ID=5) - East, Pen
    assert table.rows[4] == [
        "5",
        "East",
        "2023",
        "Q4",
        "Stationery",
        "Pen",
        "In Progress",
    ]
