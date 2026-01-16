"""
Taple 服务基类
"""
import csv
import tempfile
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Iterator, Callable, Tuple
try:
    import openpyxl
except ImportError:
    openpyxl = None

from ...schemas.taple import (
    Table, Sheet, Column, Row, Cell, MergedCell, TableView,
    CellUpdate
)
from ...utils.converter import timestamp_to_datetime
from .idempotent_taple_mixin import IdempotentTapleMixin
from ...errors import ValidationError


class BaseTapleService(IdempotentTapleMixin):
    """
    Taple 服务基类，提供通用的数据转换方法
    """

    def _convert_table(self, proto_table: Any) -> Table:
        """转换 Proto Table 为模型"""
        return Table(
            id=proto_table.id,
            file_id=proto_table.file_id,
            org_id=proto_table.org_id,
            user_id=proto_table.user_id,
            name=proto_table.name if proto_table.name else None,
            description=proto_table.description if proto_table.description else None,
            created_by_role=proto_table.created_by_role,
            created_by=proto_table.created_by,
            created_at=timestamp_to_datetime(proto_table.created_at),
            updated_at=timestamp_to_datetime(proto_table.updated_at),
            deleted_at=timestamp_to_datetime(proto_table.deleted_at) if proto_table.deleted_at else None,
        )

    def _convert_sheet(self, proto_sheet: Any) -> Sheet:
        """转换 Proto Sheet 为模型"""
        return Sheet(
            id=proto_sheet.id,
            table_id=proto_sheet.table_id,
            org_id=proto_sheet.org_id,
            user_id=proto_sheet.user_id,
            name=proto_sheet.name,
            description=proto_sheet.description if proto_sheet.description else None,
            position=proto_sheet.position,
            version=proto_sheet.version,
            created_by_role=proto_sheet.created_by_role,
            created_by=proto_sheet.created_by,
            created_at=timestamp_to_datetime(proto_sheet.created_at),
            updated_at=timestamp_to_datetime(proto_sheet.updated_at),
            deleted_at=timestamp_to_datetime(proto_sheet.deleted_at) if proto_sheet.deleted_at else None
        )

    def _convert_column(self, proto_column: Any) -> Column:
        """转换 Proto Column 为模型"""
        properties = None
        if proto_column.properties:
            from google.protobuf.json_format import MessageToDict
            properties = MessageToDict(proto_column.properties)

        return Column(
            id=proto_column.id,
            sheet_id=proto_column.sheet_id,
            org_id=proto_column.org_id,
            user_id=proto_column.user_id,
            column_key=proto_column.column_key,
            name=proto_column.name,
            column_type=proto_column.column_type,
            description=proto_column.description if proto_column.description else None,
            position=proto_column.position,
            width=proto_column.width if proto_column.width else None,
            hidden=proto_column.hidden if proto_column.hidden else None,
            properties=properties,
            version=proto_column.version,
            created_by_role=proto_column.created_by_role,
            created_by=proto_column.created_by,
            created_at=timestamp_to_datetime(proto_column.created_at),
            updated_at=timestamp_to_datetime(proto_column.updated_at),
            deleted_at=timestamp_to_datetime(proto_column.deleted_at) if proto_column.deleted_at else None
        )

    def _convert_row(self, proto_row: Any) -> Row:
        """转换 Proto Row 为模型"""
        return Row(
            id=proto_row.id,
            sheet_id=proto_row.sheet_id,
            org_id=proto_row.org_id,
            user_id=proto_row.user_id,
            row_key=proto_row.row_key,
            position=proto_row.position,
            height=proto_row.height if proto_row.height else None,
            hidden=proto_row.hidden if proto_row.hidden else None,
            version=proto_row.version,
            created_by_role=proto_row.created_by_role,
            created_by=proto_row.created_by,
            created_at=timestamp_to_datetime(proto_row.created_at),
            updated_at=timestamp_to_datetime(proto_row.updated_at),
            deleted_at=timestamp_to_datetime(proto_row.deleted_at) if proto_row.deleted_at else None
        )

    def _convert_cell(self, proto_cell: Any) -> Cell:
        """转换 Proto Cell 为模型"""
        styles = None
        if proto_cell.styles:
            from google.protobuf.json_format import MessageToDict
            styles = MessageToDict(proto_cell.styles)

        return Cell(
            id=proto_cell.id,
            sheet_id=proto_cell.sheet_id,
            column_id=proto_cell.column_id,
            row_id=proto_cell.row_id,
            org_id=proto_cell.org_id,
            user_id=proto_cell.user_id,
            column_key=proto_cell.column_key,
            row_key=proto_cell.row_key,
            raw_value=proto_cell.raw_value if proto_cell.raw_value else None,
            formatted_value=proto_cell.formatted_value if proto_cell.formatted_value else None,
            formula=proto_cell.formula if proto_cell.formula else None,
            styles=styles,
            data_type=proto_cell.data_type if proto_cell.data_type else None,
            version=proto_cell.version,
            created_by_role=proto_cell.created_by_role,
            created_by=proto_cell.created_by,
            created_at=timestamp_to_datetime(proto_cell.created_at),
            updated_at=timestamp_to_datetime(proto_cell.updated_at),
            deleted_at=timestamp_to_datetime(proto_cell.deleted_at) if proto_cell.deleted_at else None
        )

    def _convert_merged_cell(self, proto_merged_cell: Any) -> MergedCell:
        """转换 Proto MergedCell 为模型"""
        return MergedCell(
            id=proto_merged_cell.id,
            sheet_id=proto_merged_cell.sheet_id,
            org_id=proto_merged_cell.org_id,
            user_id=proto_merged_cell.user_id,
            start_column_id=proto_merged_cell.start_column_id,
            end_column_id=proto_merged_cell.end_column_id,
            start_row_id=proto_merged_cell.start_row_id,
            end_row_id=proto_merged_cell.end_row_id,
            created_at=timestamp_to_datetime(proto_merged_cell.created_at),
            updated_at=timestamp_to_datetime(proto_merged_cell.updated_at),
            deleted_at=timestamp_to_datetime(proto_merged_cell.deleted_at) if proto_merged_cell.deleted_at else None
        )

    def _convert_table_view(self, proto_view: Any) -> TableView:
        """转换 Proto TableView 为模型"""
        from google.protobuf.json_format import MessageToDict
        
        # 转换配置字段
        config = MessageToDict(proto_view.config) if proto_view.HasField('config') else {}
        filter_criteria = MessageToDict(proto_view.filter_criteria) if proto_view.HasField('filter_criteria') else None
        sort_criteria = MessageToDict(proto_view.sort_criteria) if proto_view.HasField('sort_criteria') else None
        if proto_view.HasField('visible_columns'):
            visible_columns_dict = MessageToDict(proto_view.visible_columns)
            # 如果服务器返回的是旧格式（包含 items 字段的结构），需要转换
            if isinstance(visible_columns_dict, dict) and 'items' in visible_columns_dict:
                # 旧格式：将列表转换为字典，默认所有列都显示
                if isinstance(visible_columns_dict['items'], list):
                    visible_columns = {col: True for col in visible_columns_dict['items']}
                else:
                    visible_columns = visible_columns_dict
            else:
                visible_columns = visible_columns_dict
        else:
            visible_columns = None
        group_criteria = MessageToDict(proto_view.group_criteria) if proto_view.HasField('group_criteria') else None

        return TableView(
            id=proto_view.id,
            table_id=proto_view.table_id,
            sheet_id=proto_view.sheet_id,
            org_id=proto_view.org_id,
            user_id=proto_view.user_id,
            file_id=proto_view.file_id,
            filter_criteria=filter_criteria,
            sort_criteria=sort_criteria,
            visible_columns=visible_columns,
            group_criteria=group_criteria,
            created_by_role=proto_view.created_by_role,
            created_by=proto_view.created_by,
            view_name=proto_view.view_name,
            view_type=proto_view.view_type,
            is_hidden=proto_view.is_hidden,
            is_default=proto_view.is_default,
            config=config,
            created_at=timestamp_to_datetime(proto_view.created_at),
            updated_at=timestamp_to_datetime(proto_view.updated_at),
            deleted_at=timestamp_to_datetime(proto_view.deleted_at) if proto_view.HasField('deleted_at') else None
        )

    def _convert_dict_to_struct(self, data: Dict[str, Any]) -> Any:
        """转换字典为 Proto Struct"""
        from google.protobuf.struct_pb2 import Struct
        from google.protobuf.json_format import ParseDict
        
        struct = Struct()
        ParseDict(data, struct)
        return struct

    def _convert_cell_updates_to_proto(self, updates: List[CellUpdate]) -> List[Any]:
        """转换 CellUpdate 列表为 Proto 格式"""
        from ...rpc.gen import taple_service_pb2
        
        proto_updates = []
        for update in updates:
            proto_update = taple_service_pb2.CellUpdate(
                column_key=update.column_key,
                row_key=update.row_key
            )
            
            if update.raw_value is not None:
                proto_update.raw_value = update.raw_value
            if update.formula is not None:
                proto_update.formula = update.formula
            if update.styles is not None:
                proto_update.styles.CopyFrom(self._convert_dict_to_struct(update.styles))
                
            proto_updates.append(proto_update)
            
        return proto_updates

    def _is_file_id(self, source: Union[str, Path]) -> bool:
        """判断source是否为file_id"""
        if isinstance(source, Path):
            return False
        # file_id 通常是UUID格式或特定格式的ID
        source_str = str(source)
        # 检查是否是文件路径
        if '/' in source_str or '\\' in source_str or os.path.exists(source_str):
            return False
        # 简单判断：如果长度合适且不包含文件扩展名，可能是file_id
        return len(source_str) > 10 and not source_str.endswith(('.csv', '.xlsx', '.xls'))

    def _detect_file_format(self, file_path: Union[str, Path]) -> str:
        """自动检测文件格式"""
        path = Path(file_path)
        extension = path.suffix.lower()
        if extension in ['.csv', '.tsv']:
            return 'csv'
        elif extension in ['.xlsx', '.xls']:
            return 'excel'
        else:
            # 尝试通过内容判断
            with open(path, 'rb') as f:
                header = f.read(8)
                if header.startswith(b'PK'):  # ZIP文件头，Excel文件
                    return 'excel'
            return 'csv'  # 默认CSV

    def _read_csv_stream(
        self, 
        file_path: Union[str, Path], 
        encoding: str = 'utf-8',
        batch_size: int = 1000,
        delimiter: str = ','
    ) -> Iterator[Dict[str, Any]]:
        """流式读取CSV文件"""
        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            # 尝试检测分隔符
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            try:
                detected_delimiter = sniffer.sniff(sample).delimiter
                delimiter = detected_delimiter
            except:
                pass  # 使用默认分隔符
            
            reader = csv.DictReader(f, delimiter=delimiter)
            columns = reader.fieldnames
            
            if not columns:
                raise ValidationError("CSV文件没有列头")
            
            # 返回列定义
            yield {'type': 'columns', 'data': columns}
            
            # 批量返回行数据
            batch = []
            row_count = 0
            for row in reader:
                batch.append(row)
                row_count += 1
                if len(batch) >= batch_size:
                    yield {'type': 'rows', 'data': batch, 'start_index': row_count - len(batch)}
                    batch = []
            
            if batch:
                yield {'type': 'rows', 'data': batch, 'start_index': row_count - len(batch)}

    def _read_excel_stream(
        self,
        file_path: Union[str, Path],
        batch_size: int = 1000,
        sheet_mapping: Optional[Dict[str, str]] = None
    ) -> Iterator[Dict[str, Any]]:
        """流式读取Excel文件"""
        if openpyxl is None:
            raise ImportError("需要安装openpyxl库来处理Excel文件: pip install openpyxl")
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            if sheet_mapping and sheet_name not in sheet_mapping:
                continue
            
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            
            # 获取第一行作为列头
            headers = None
            for row in rows:
                if row and any(cell is not None for cell in row):
                    headers = [str(cell) if cell is not None else f"Column_{i+1}" 
                              for i, cell in enumerate(row)]
                    break
            
            if not headers:
                continue
            
            # 返回工作表信息
            target_sheet_name = sheet_mapping.get(sheet_name, sheet_name) if sheet_mapping else sheet_name
            yield {
                'type': 'sheet',
                'name': target_sheet_name,
                'columns': headers,
                'original_name': sheet_name
            }
            
            # 批量返回行数据
            batch = []
            row_count = 0
            for row in rows:
                if row and any(cell is not None for cell in row):
                    row_dict = {}
                    for i, (header, value) in enumerate(zip(headers, row)):
                        if value is not None:
                            row_dict[header] = value
                    if row_dict:  # 只添加非空行
                        batch.append(row_dict)
                        row_count += 1
                        if len(batch) >= batch_size:
                            yield {'type': 'rows', 'data': batch, 'start_index': row_count - len(batch)}
                            batch = []
            
            if batch:
                yield {'type': 'rows', 'data': batch, 'start_index': row_count - len(batch)}
        
        wb.close()

    def _infer_column_type(self, values: List[Any]) -> str:
        """推断列类型"""
        if not values:
            return 'text'
        
        # 检查数值类型
        numeric_count = 0
        for value in values[:100]:  # 只检查前100个值
            if value is None or str(value).strip() == '':
                continue
            try:
                float(str(value))
                numeric_count += 1
            except:
                pass
        
        if numeric_count > len(values) * 0.8:  # 80%以上是数字
            return 'number'
        
        # 默认为文本
        return 'text'