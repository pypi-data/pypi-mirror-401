"""Excel Parser using openpyxl."""

from __future__ import annotations

import os
from typing import Optional, Any

from . import ParseOptions, ParseResult, DocumentItem


def get_cell_value(cell_value: Any) -> Optional[str]:
    """Extract the display value from a cell, handling formulas and special types."""
    if cell_value is None:
        return None

    # Handle formula results (openpyxl stores computed values directly)
    # Just convert to string
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    elif isinstance(cell_value, str):
        return cell_value if cell_value.strip() else None
    elif hasattr(cell_value, "isoformat"):  # datetime
        return cell_value.isoformat().split("T")[0]
    else:
        return str(cell_value) if cell_value else None


def parse_xlsx(file_path: str, options: Optional[ParseOptions] = None) -> ParseResult:
    """
    Parse an Excel file, extracting text per sheet.

    Args:
        file_path: Path to the Excel file
        options: Parsing options (max_items limits sheets)

    Returns:
        ParseResult with per-sheet items
    """
    filename = os.path.basename(file_path)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install with: pip install openpyxl"
        )

    try:
        workbook = load_workbook(file_path, data_only=True)  # data_only=True gets computed values
        sheet_names = workbook.sheetnames
        max_items = (options or {}).get("max_items") or len(sheet_names)

        items: list[DocumentItem] = []
        for i, sheet_name in enumerate(sheet_names[:max_items]):
            sheet = workbook[sheet_name]
            text_lines: list[str] = []

            for row in sheet.iter_rows():
                values: list[str] = []
                for cell in row:
                    value = get_cell_value(cell.value)
                    if value:
                        values.append(value)
                if values:
                    text_lines.append(" | ".join(values))

            if text_lines:
                items.append({
                    "number": i,
                    "name": sheet_name,
                    "text": "\n".join(text_lines),
                })

        workbook.close()

        return {
            "type": "xlsx",
            "filename": filename,
            "total_items": len(sheet_names),
            "items": items,
        }
    except Exception as e:
        raise RuntimeError(
            f'Failed to parse Excel file "{filename}": {e}. '
            f"Ensure the file is a valid .xlsx/.xls file."
        )
