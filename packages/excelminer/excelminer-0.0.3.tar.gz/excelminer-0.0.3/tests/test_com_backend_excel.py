from __future__ import annotations

import os
import platform
from pathlib import Path

import pytest

from excelminer.backends.base import AnalysisContext, AnalysisOptions
from excelminer.backends.com_backend import ComBackend, _extract_command_details
from excelminer.model.graph import WorkbookGraph


def _run_com_tests_enabled() -> bool:
    # COM automation can be flaky and has been observed to raise fatal Windows exceptions
    # in some environments. Keep these integration tests opt-in.
    return os.environ.get("EXCELMINER_RUN_COM_TESTS", "").strip() in {
        "1",
        "true",
        "True",
        "yes",
        "YES",
    }


def _excel_is_available() -> bool:
    if not _run_com_tests_enabled():
        return False
    if platform.system() != "Windows":
        return False

    try:
        import win32com.client  # type: ignore[import-not-found]
    except Exception:
        return False

    try:
        import pythoncom  # type: ignore[import-not-found]

        pythoncom.CoInitialize()
        com_initialized = True
    except Exception:
        pythoncom = None
        com_initialized = False

    xl = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        return True
    except Exception:
        return False
    finally:
        try:
            if xl is not None:
                xl.Quit()
        except Exception:
            pass

        try:
            if pythoncom is not None and com_initialized:
                pythoncom.CoUninitialize()
        except Exception:
            pass


@pytest.mark.integration
def test_com_backend_extracts_sheets_and_formulas_if_excel_available(
    tmp_path: Path,
) -> None:
    if not _run_com_tests_enabled():
        pytest.skip("Set EXCELMINER_RUN_COM_TESTS=1 to run Excel COM integration tests")

    if not _excel_is_available():
        pytest.skip("Excel COM automation not available")

    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        pytest.skip(f"openpyxl missing in test env: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 1
    ws["A2"].value = 2
    ws["A3"].value = "=SUM(A1:A2)"

    ws2 = wb.create_sheet("Second")
    ws2["B2"].value = "=Sheet1!A3"

    xlsx = tmp_path / "com.xlsx"
    wb.save(xlsx)

    backend = ComBackend()
    ctx = AnalysisContext(
        path=xlsx,
        options=AnalysisOptions(
            include_com=True,
            include_formulas=True,
            include_defined_names=True,
            include_connections=False,
        ),
    )

    assert backend.can_handle(ctx) is True

    g = WorkbookGraph()
    rep = backend.extract(ctx, g)

    # This test is intentionally permissive about issues (Excel may warn on links, etc)
    assert rep.backend == "com"

    sheet_nodes = [n for n in g.nodes.values() if n.kind == "sheet"]
    assert {s.attrs.get("name") for s in sheet_nodes} >= {"Sheet1", "Second"}

    formula_nodes = [n for n in g.nodes.values() if n.kind == "formula_cell"]
    assert len(formula_nodes) >= 1


def test_com_backend_is_opt_in_for_xlsx(tmp_path: Path) -> None:
    xlsx = tmp_path / "x.xlsx"
    xlsx.write_bytes(b"dummy")

    b = ComBackend()
    ctx = AnalysisContext(path=xlsx, options=AnalysisOptions(include_com=False))
    assert b.can_handle(ctx) is False


def test_com_backend_command_details_are_extracted() -> None:
    class FakeParam:
        def __init__(self, name: str, value: object) -> None:
            self.Name = name
            self.Value = value

    class FakeParams:
        def __init__(self, params: list[FakeParam]) -> None:
            self._params = params
            self.Count = len(params)

        def Item(self, idx: int) -> FakeParam:
            return self._params[idx - 1]

    class FakeOleDb:
        def __init__(self) -> None:
            self.CommandText = "SELECT * FROM Sales WHERE Region = ?"
            self.CommandType = 2
            self.Parameters = FakeParams([FakeParam("Region", "West")])

    class FakeConn:
        def __init__(self) -> None:
            self.OLEDBConnection = FakeOleDb()

    details = _extract_command_details(FakeConn())

    assert details.get("command_text") == "SELECT * FROM Sales WHERE Region = ?"
    assert details.get("command_type") == "2"
    params = details.get("command_parameters")
    assert isinstance(params, list)
    assert params[0]["name"] == "Region"
    assert params[0]["value"] == "West"
