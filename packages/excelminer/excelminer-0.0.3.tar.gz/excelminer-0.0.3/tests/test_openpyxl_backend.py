from __future__ import annotations

from pathlib import Path

import pytest

from excelminer.backends.base import AnalysisContext, AnalysisOptions
from excelminer.backends.openpyxl_backend import OpenpyxlBackend
from excelminer.model.graph import WorkbookGraph


def _write_formula_workbook(path: Path) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        pytest.skip(f"openpyxl missing in test env: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=1+1"
    ws["B2"] = "=A1*2"
    ws["D4"] = "=B2*2"
    wb.save(path)


def test_openpyxl_backend_formula_chunking(tmp_path: Path) -> None:
    path = tmp_path / "formulas.xlsx"
    _write_formula_workbook(path)

    options = AnalysisOptions(
        include_formulas=True,
        row_chunk_size=1,
        column_chunk_size=1,
        max_rows_per_sheet=3,
        max_cols_per_sheet=3,
    )
    ctx = AnalysisContext(path=path, options=options)
    graph = WorkbookGraph()

    backend = OpenpyxlBackend()
    report = backend.extract(ctx, graph)

    assert report.backend == "openpyxl"
    assert report.stats["formula_cells"] == 2

    sheet_node = graph.get_by_key("sheet", "Sheet1")
    assert sheet_node is not None

    a1 = graph.get_by_key("formula_cell", "Sheet1!A1")
    b2 = graph.get_by_key("formula_cell", "Sheet1!B2")
    d4 = graph.get_by_key("formula_cell", "Sheet1!D4")

    assert a1 is not None
    assert b2 is not None
    assert d4 is None

    contains_edges = [
        edge
        for edge in graph.edges
        if edge.kind == "contains" and edge.src == sheet_node.id
    ]
    assert {edge.dst for edge in contains_edges} == {a1.id, b2.id}
