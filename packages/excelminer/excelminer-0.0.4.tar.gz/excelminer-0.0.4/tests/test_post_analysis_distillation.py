from __future__ import annotations

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName

from excelminer import AnalysisOptions, analyze_workbook
from tests.fixtures.workbooks import build_large_formula_workbook


def _save_wb(tmp_path: Path, wb: Workbook, name: str) -> Path:
    p = tmp_path / name
    wb.save(p)
    return p


def _localname(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _inject_chart_defined_name(path: Path, defined_name: str) -> None:
    with zipfile.ZipFile(path, "a") as zf:
        chart_parts = [
            name
            for name in zf.namelist()
            if name.startswith("xl/charts/chart") and name.endswith(".xml")
        ]
        assert chart_parts

        for part in chart_parts:
            data = zf.read(part)
            root = ET.fromstring(data)
            changed = False
            for el in root.iter():
                if _localname(el.tag) != "f":
                    continue
                if not (el.text or "").strip():
                    continue
                el.text = defined_name
                changed = True
            if changed:
                xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                zf.writestr(part, xml_bytes)


def test_post_analysis_distillation_creates_formula_groups(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"].value = "=B1+C1"
    ws["A2"].value = "=B2+C2"

    path = _save_wb(tmp_path, wb, "formulas.xlsx")

    graph, reports, ctx = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=True,
            include_defined_names=False,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=True,
        ),
    )

    assert ctx.issues == []
    assert any(r.backend == "distillation" for r in reports)

    groups = [n for n in graph.nodes.values() if n.kind == "formula_group"]
    assert len(groups) == 1

    g = groups[0]
    assert g.attrs["sheet"] == "Data"
    assert g.attrs["count"] == 2
    assert "address_ranges" in g.attrs

    member_edges = [e for e in graph.edges if e.kind == "member_of" and e.dst == g.id]
    assert len(member_edges) == 2


def test_post_analysis_distillation_prunes_unused_defined_names(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Defined names: one used, one unused
    wb.defined_names.add(DefinedName("UsedName", attr_text="Data!$A$1"))
    wb.defined_names.add(DefinedName("UnusedName", attr_text="Data!$A$2"))

    # Reference UsedName in a formula so it should survive pruning.
    ws["B1"].value = "=UsedName+1"

    path = _save_wb(tmp_path, wb, "names.xlsx")

    # Without distillation: both defined names remain
    g0, _, _ = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=True,
            include_defined_names=True,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=False,
        ),
    )
    names0 = {
        str(n.attrs.get("name")) for n in g0.nodes.values() if n.kind == "defined_name"
    }
    assert {"UsedName", "UnusedName"}.issubset(names0)

    # With distillation: unused defined name is removed
    g1, reports, ctx = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=True,
            include_defined_names=True,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=True,
        ),
    )

    assert ctx.issues == []
    assert any(r.backend == "distillation" for r in reports)

    names1 = {
        str(n.attrs.get("name")) for n in g1.nodes.values() if n.kind == "defined_name"
    }
    assert "UsedName" in names1
    assert "UnusedName" not in names1


def test_post_analysis_distillation_prunes_names_with_reused_formulas(
    tmp_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    wb.defined_names.add(DefinedName("UsedName", attr_text="Data!$A$1"))
    wb.defined_names.add(DefinedName("UnusedName", attr_text="Data!$A$2"))

    for row in range(1, 6):
        ws[f"B{row}"].value = "=UsedName+1"

    path = _save_wb(tmp_path, wb, "names_reused.xlsx")

    g1, reports, ctx = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=True,
            include_defined_names=True,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=True,
        ),
    )

    assert ctx.issues == []
    assert any(r.backend == "distillation" for r in reports)

    names1 = {
        str(n.attrs.get("name")) for n in g1.nodes.values() if n.kind == "defined_name"
    }
    assert "UsedName" in names1
    assert "UnusedName" not in names1


def test_post_analysis_distillation_generates_ref_hints(
    tmp_path: Path,
) -> None:
    path = build_large_formula_workbook(tmp_path / "large_formulas.xlsx", rows=25)

    graph, reports, ctx = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=True,
            include_defined_names=False,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=True,
        ),
    )

    assert ctx.issues == []
    assert any(r.backend == "distillation" for r in reports)

    groups = [n for n in graph.nodes.values() if n.kind == "formula_group"]
    assert len(groups) == 1

    group = groups[0]
    assert group.attrs["count"] == 25
    assert group.attrs["address_ranges"] == ["A1:A25"]

    ref_hints = group.attrs["ref_hints"]
    assert ref_hints["ref1"]["kind"] == "relative"
    assert ref_hints["ref1"]["offset"] == {"dr": 0, "dc": 1}
    assert ref_hints["ref2"]["kind"] == "relative"
    assert ref_hints["ref2"]["offset"] == {"dr": 0, "dc": 2}
    assert group.attrs["ref_hints_vary"] is False


def test_post_analysis_distillation_preserves_chart_defined_names(
    tmp_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for row, value in enumerate([1, 2, 3], start=1):
        ws[f"A{row}"].value = value

    wb.defined_names.add(DefinedName("ChartRange", attr_text="Data!$A$1:$A$3"))

    chart = LineChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
    ws.add_chart(chart, "C1")

    path = _save_wb(tmp_path, wb, "chart_names.xlsx")
    _inject_chart_defined_name(path, "ChartRange")

    graph, reports, ctx = analyze_workbook(
        path,
        options=AnalysisOptions(
            include_formulas=False,
            include_defined_names=True,
            include_connections=False,
            include_cells=False,
            post_analysis_distillation=True,
        ),
    )

    assert ctx.issues == []
    assert any(r.backend == "distillation" for r in reports)

    names = {
        str(n.attrs.get("name"))
        for n in graph.nodes.values()
        if n.kind == "defined_name"
    }
    assert "ChartRange" in names
