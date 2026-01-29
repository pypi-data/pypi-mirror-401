from __future__ import annotations

import zipfile
from pathlib import Path

import pytest

from excelminer import AnalysisOptions, analyze_to_dict
from excelminer.backends.base import AnalysisContext, BackendReport
from excelminer.model.graph import WorkbookGraph


@pytest.mark.parametrize("include_formulas", [False, True])
def test_analyze_to_dict_smoke(tmp_path: Path, include_formulas: bool) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        pytest.skip(f"openpyxl missing in test env: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 1
    ws["A2"].value = 2
    ws["A3"].value = "=SUM(A1:A2)"

    ws2 = wb.create_sheet("Second")
    ws2["B2"].value = "=Sheet1!A3"

    xlsx = tmp_path / "test.xlsx"
    wb.save(xlsx)

    result = analyze_to_dict(
        xlsx,
        options=AnalysisOptions(
            include_connections=False,
            include_defined_names=False,
            include_formulas=include_formulas,
        ),
    )

    assert result["path"].endswith("test.xlsx")
    assert "graph" in result
    assert "nodes" in result["graph"]
    assert "edges" in result["graph"]

    kinds = {n["kind"] for n in result["graph"]["nodes"]}
    assert "sheet" in kinds

    formula_nodes = [n for n in result["graph"]["nodes"] if n["kind"] == "formula_cell"]
    if include_formulas:
        assert len(formula_nodes) >= 1
    else:
        assert len(formula_nodes) == 0


def test_analyze_to_dict_includes_vba_project_for_xlsm(tmp_path: Path) -> None:
    xlsm = tmp_path / "macros.xlsm"

    # Minimal OOXML workbook plus VBA payload.
    content_types = """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="xml" ContentType="application/xml"/>
</Types>
"""
    workbook_xml = """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>
"""

    with zipfile.ZipFile(xlsm, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/vbaProject.bin", b"\x00\x01")

    # Disable unrelated extractors so this remains a focused API-level check.
    result = analyze_to_dict(
        xlsm,
        options=AnalysisOptions(
            include_vba=True,
            include_connections=False,
            include_powerquery=False,
            include_pivots=False,
            include_defined_names=False,
            include_cells=False,
            include_formulas=False,
            include_com=False,
        ),
    )

    kinds = {n["kind"] for n in result["graph"]["nodes"]}
    assert "vba_project" in kinds

    result_no_vba = analyze_to_dict(
        xlsm,
        options=AnalysisOptions(
            include_vba=False,
            include_connections=False,
            include_powerquery=False,
            include_pivots=False,
            include_defined_names=False,
            include_cells=False,
            include_formulas=False,
            include_com=False,
        ),
    )
    kinds_no_vba = {n["kind"] for n in result_no_vba["graph"]["nodes"]}
    assert "vba_project" not in kinds_no_vba


def test_analyze_to_dict_serializes_structured_issues(tmp_path: Path) -> None:
    class IssueBackend:
        name = "issue_backend"

        def can_handle(self, ctx: AnalysisContext) -> bool:  # noqa: ARG002
            return True

        def extract(
            self, ctx: AnalysisContext, graph: WorkbookGraph
        ) -> BackendReport:  # noqa: ARG002
            report = BackendReport(backend=self.name)
            report.add_issue(
                "failed to parse workbook",
                kind="parse_error",
                detail="malformed xml",
            )
            return report

    xlsx = tmp_path / "issue.xlsx"
    xlsx.write_bytes(b"dummy")

    result = analyze_to_dict(xlsx, backends=[IssueBackend()])

    issue = result["reports"][0]["issues"][0]
    assert issue["kind"] == "parse_error"
    assert issue["message"] == "failed to parse workbook"
    assert issue["detail"] == "malformed xml"
