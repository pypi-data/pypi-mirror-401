from __future__ import annotations

import builtins
import sys
import types
from pathlib import Path

import pytest

from excelminer.backends.base import AnalysisContext, AnalysisOptions
from excelminer.backends.calamine_backend import CalamineBackend
from excelminer.model.graph import WorkbookGraph


def test_calamine_backend_extracts_used_range_as_cell_block(tmp_path: Path) -> None:
    # These are real deps for this test (skip rather than failing on minimal installs).
    pytest.importorskip("pandas")
    pytest.importorskip("python_calamine")

    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        pytest.skip(f"openpyxl missing in test env: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Create a sparse used-range so the backend must compute bounding box.
    ws["A1"].value = 1
    ws["B1"].value = "x"
    ws["C3"].value = 5

    xlsx = tmp_path / "calamine.xlsx"
    wb.save(xlsx)

    ctx = AnalysisContext(
        path=xlsx,
        options=AnalysisOptions(
            include_cells=True,
            include_formulas=False,
            include_connections=False,
            include_defined_names=False,
            include_com=False,
        ),
    )

    g = WorkbookGraph()
    rep = CalamineBackend().extract(ctx, g)

    assert rep.issues == []

    blocks = [n for n in g.nodes.values() if n.kind == "cell_block"]
    assert len(blocks) == 1

    block = blocks[0]
    assert block.attrs["sheet"] == "Data"
    assert block.attrs["range"] == "A1:C3"
    assert block.attrs["block_type"] == "value_block"

    stats = block.attrs.get("stats")
    assert isinstance(stats, dict)
    assert stats["rows"] == 3
    assert stats["cols"] == 3
    assert stats["non_null"] == 3

    # Sample should include top-left values.
    sample = block.attrs.get("sample")
    assert isinstance(sample, list)
    assert sample[0][0] == 1
    assert sample[0][1] == "x"


def test_calamine_backend_truncates_used_range_by_max_cells(tmp_path: Path) -> None:
    pytest.importorskip("pandas")
    pytest.importorskip("python_calamine")

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Big"

    # Fill a 10x10 used range: A1:J10
    for r in range(1, 11):
        for c in range(1, 11):
            ws.cell(row=r, column=c).value = r * 100 + c

    xlsx = tmp_path / "big.xlsx"
    wb.save(xlsx)

    ctx = AnalysisContext(
        path=xlsx,
        options=AnalysisOptions(
            include_cells=True,
            max_cells_per_sheet=20,  # with 10 cols, should truncate to 2 rows
            include_connections=False,
            include_defined_names=False,
            include_formulas=False,
        ),
    )
    g = WorkbookGraph()
    rep = CalamineBackend().extract(ctx, g)

    assert any("truncated used range" in issue.message for issue in rep.issues)

    blocks = [n for n in g.nodes.values() if n.kind == "cell_block"]
    assert len(blocks) == 1
    assert blocks[0].attrs["range"] == "A1:J2"
    stats = blocks[0].attrs["stats"]
    assert stats["rows"] == 2
    assert stats["cols"] == 10


def test_calamine_backend_reports_missing_deps(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # We simulate missing deps even if they're installed.
    xlsx = tmp_path / "x.xlsx"
    xlsx.write_bytes(b"dummy")

    ctx = AnalysisContext(path=xlsx, options=AnalysisOptions(include_cells=True))
    g = WorkbookGraph()

    real_import = builtins.__import__

    def fake_import(name: str, *args, **kwargs):  # type: ignore[no-untyped-def]
        if name == "pandas":
            raise ImportError("no pandas")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", fake_import)
    rep = CalamineBackend().extract(ctx, g)
    assert rep.issues
    assert "pandas not available" in rep.issues[0].message

    def fake_import2(name: str, *args, **kwargs):  # type: ignore[no-untyped-def]
        # CI / minimal installs may not have pandas; stub it so we can
        # deterministically test the python_calamine missing-dep path.
        if name == "pandas":
            return types.ModuleType("pandas")
        if name == "python_calamine":
            raise ImportError("no python-calamine")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", fake_import2)
    rep2 = CalamineBackend().extract(ctx, g)
    assert rep2.issues
    assert "python-calamine not available" in rep2.issues[0].message


def test_calamine_backend_reads_only_requested_sheets(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    xlsx = tmp_path / "lazy.xlsx"
    xlsx.write_bytes(b"dummy")

    parse_calls: list[str] = []

    class FakeValues:
        def __init__(self, data: list[list[bool]]) -> None:
            self._flat = [item for row in data for item in row]
            self.size = len(self._flat)

        def sum(self) -> int:
            return int(sum(1 for value in self._flat if value))

    class FakeSeries:
        def __init__(self, data: list[bool], index: list[int]) -> None:
            self._data = data
            self.index = index

        def any(self) -> bool:
            return any(self._data)

        def __getitem__(self, key):  # type: ignore[no-untyped-def]
            if isinstance(key, FakeSeries):
                filtered = [
                    (value, idx)
                    for value, idx, include in zip(self._data, self.index, key._data)
                    if include
                ]
                data = [value for value, _ in filtered]
                index = [idx for _, idx in filtered]
                return FakeSeries(data, index)
            raise TypeError("Unsupported key type for FakeSeries")

    class FakeMask:
        def __init__(self, data: list[list[bool]]) -> None:
            self._data = data
            self.values = FakeValues(data)

        def any(self, axis: int) -> FakeSeries:
            if axis == 1:
                data = [any(row) for row in self._data]
                return FakeSeries(data, list(range(len(data))))
            if axis == 0:
                cols = list(zip(*self._data)) if self._data else []
                data = [any(col) for col in cols]
                return FakeSeries(data, list(range(len(data))))
            raise ValueError("axis must be 0 or 1")

    class FakeILoc:
        def __init__(self, df: "FakeDataFrame") -> None:
            self._df = df

        def __getitem__(self, key):  # type: ignore[no-untyped-def]
            row_sel, col_sel = key
            rows = self._df._data[row_sel]
            data = [row[col_sel] for row in rows]
            return FakeDataFrame(data)

    class FakeDataFrame:
        def __init__(self, data: list[list[object | None]]) -> None:
            self._data = data
            self.iloc = FakeILoc(self)

        def notna(self) -> FakeMask:
            mask = [[value is not None for value in row] for row in self._data]
            return FakeMask(mask)

        def itertuples(self, index: bool = False, name: str | None = None):
            del index, name
            for row in self._data:
                yield tuple(row)

    class FakeExcelFile:
        def __init__(self, path: Path, engine: str) -> None:
            assert path == xlsx
            assert engine == "calamine"
            self.sheet_names = ["Sheet1", "Sheet2"]

        def parse(self, sheet_name: str, header=None, dtype=None):  # type: ignore[no-untyped-def]
            del header, dtype
            parse_calls.append(sheet_name)
            return FakeDataFrame([[1, None], [None, 2]])

        def close(self) -> None:
            return None

    fake_pandas = types.ModuleType("pandas")
    fake_pandas.ExcelFile = FakeExcelFile
    fake_pandas.isna = lambda value: value is None

    monkeypatch.setitem(sys.modules, "pandas", fake_pandas)
    monkeypatch.setitem(
        sys.modules, "python_calamine", types.ModuleType("python_calamine")
    )

    ctx = AnalysisContext(
        path=xlsx,
        options=AnalysisOptions(
            include_cells=True,
            max_sheets=1,
            include_connections=False,
            include_defined_names=False,
            include_formulas=False,
        ),
    )
    g = WorkbookGraph()
    rep = CalamineBackend().extract(ctx, g)

    assert rep.issues == []
    assert parse_calls == ["Sheet1"]
