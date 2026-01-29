from __future__ import annotations

"""Openpyxl-backed semantic extractor for formulas and cell data."""

import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from contextlib import contextmanager
from typing import Iterator
from dataclasses import dataclass

from excelminer.backends.base import AnalysisContext, BackendReport
from excelminer.model.entities import FormulaCell, Sheet
from excelminer.model.graph import WorkbookGraph

_A1_REF_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$", re.IGNORECASE)
_CELL_RANGE_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_ ]+|'\[[^]]+\][^']+'|\[[^]]+\][A-Za-z0-9_ ]+)!)?"
    r"(?P<cell>\$?[A-Z]{1,3}\$?\d+)"
    r"(?::(?P<cell2>\$?[A-Z]{1,3}\$?\d+))?",
)
_STRUCTURED_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_])(?P<structured>[A-Za-z_][A-Za-z0-9_ ]*\[[^\]]+\])",
)
_NAME_TOKEN_RE = re.compile(
    r"(?<![A-Za-z0-9_.])(?P<name>[A-Za-z_][A-Za-z0-9_.]*)\b",
)


def _normalize_sheet_name(s: str) -> str:
    """Return a normalized sheet name without surrounding quotes."""
    s = (s or "").strip()
    if s.startswith("'") and s.endswith("'") and len(s) >= 2:
        s = s[1:-1]
    return s


def _sheet_key(name: str) -> str:
    """Build a stable key for sheet entity deduplication."""
    return re.sub(r"\s+", " ", (name or "").strip())


def _chunk_bounds(limit: int, chunk_size: int) -> Iterator[tuple[int, int]]:
    """Yield inclusive start/end bounds for a chunked range."""
    for start in range(1, limit + 1, chunk_size):
        end = min(limit, start + chunk_size - 1)
        yield start, end


@contextmanager
def _formula_executor(max_workers: int | None) -> Iterator[ProcessPoolExecutor | None]:
    """Context manager that yields a process pool executor or None."""
    if max_workers is not None and max_workers <= 1:
        yield None
        return

    try:
        executor = ProcessPoolExecutor(max_workers=max_workers)
    except Exception:  # noqa: BLE001
        yield None
        return

    try:
        yield executor
    finally:
        executor.shutdown(wait=True, cancel_futures=True)


def _resolve_formula_deps(
    formula_batch: list[tuple[str, str, str]],
    *,
    executor: ProcessPoolExecutor | None,
    report: BackendReport,
) -> list[tuple[str, str, str, dict[str, object]]]:
    """Resolve formula dependencies with optional parallelism."""
    if not formula_batch:
        return []

    results: list[tuple[str, str, str, dict[str, object]] | None] = [
        None
    ] * len(formula_batch)

    if executor is None:
        for idx, (sheet_name, address, formula) in enumerate(formula_batch):
            try:
                deps = _extract_deps(formula)
            except Exception as exc:  # noqa: BLE001
                report.add_issue(
                    "dependency extraction failed",
                    kind="runtime",
                    detail=f"{sheet_name}!{address}: {exc}",
                )
                deps = {}
            results[idx] = (sheet_name, address, formula, deps)
        return [r for r in results if r is not None]

    try:
        futures = {
            executor.submit(_extract_deps, formula): (idx, sheet_name, address, formula)
            for idx, (sheet_name, address, formula) in enumerate(formula_batch)
        }
    except Exception as exc:  # noqa: BLE001
        report.add_issue(
            "dependency extraction failed",
            kind="runtime",
            detail=str(exc),
        )
        return _resolve_formula_deps(formula_batch, executor=None, report=report)

    for future in as_completed(futures):
        idx, sheet_name, address, formula = futures[future]
        try:
            deps = future.result()
        except Exception as exc:  # noqa: BLE001
            report.add_issue(
                "dependency extraction failed",
                kind="runtime",
                detail=f"{sheet_name}!{address}: {exc}",
            )
            deps = {}
        results[idx] = (sheet_name, address, formula, deps)

    return [r for r in results if r is not None]


def _extract_deps(formula: str) -> dict[str, object]:
    """Extract simple cell references from a formula string."""
    if not formula:
        return {}

    refs: list[dict[str, str]] = []
    seen: set[tuple[str, ...]] = set()
    spans: list[tuple[int, int]] = []

    def overlaps(span: tuple[int, int]) -> bool:
        for start, end in spans:
            if span[0] < end and span[1] > start:
                return True
        return False

    def add_ref(ref: dict[str, str], span: tuple[int, int]) -> None:
        key = (
            ref.get("kind", "cell"),
            ref.get("workbook", ""),
            ref.get("sheet", ""),
            ref.get("cell", ""),
            ref.get("range", ""),
            ref.get("structured", ""),
            ref.get("name", ""),
        )
        if key in seen:
            return
        seen.add(key)
        refs.append(ref)
        spans.append(span)

    for m in _STRUCTURED_REF_RE.finditer(formula):
        structured = (m.group("structured") or "").strip()
        if structured:
            add_ref({"kind": "structured", "structured": structured}, m.span())

    for m in _CELL_RANGE_RE.finditer(formula):
        sheet_token = _normalize_sheet_name(m.group("sheet") or "")
        cell = (m.group("cell") or "").replace("$", "")
        cell2 = (m.group("cell2") or "").replace("$", "")

        workbook = ""
        sheet = sheet_token
        if sheet_token.startswith("[") and "]" in sheet_token:
            workbook, sheet = sheet_token[1:].split("]", 1)

        if workbook:
            if cell2:
                add_ref(
                    {
                        "kind": "external",
                        "workbook": workbook,
                        "sheet": sheet,
                        "range": f"{cell}:{cell2}",
                    },
                    m.span(),
                )
            else:
                add_ref(
                    {
                        "kind": "external",
                        "workbook": workbook,
                        "sheet": sheet,
                        "cell": cell,
                    },
                    m.span(),
                )
            continue

        if cell2:
            add_ref(
                {
                    "kind": "range",
                    "sheet": sheet,
                    "range": f"{cell}:{cell2}",
                },
                m.span(),
            )
        else:
            add_ref({"sheet": sheet, "cell": cell}, m.span())

    for m in _NAME_TOKEN_RE.finditer(formula):
        name = (m.group("name") or "").strip()
        if not name or overlaps(m.span()):
            continue
        if _A1_REF_RE.match(name):
            continue

        rest = formula[m.end() :].lstrip()
        if rest[:1] in ("(", "!", "[", ":"):
            continue

        add_ref({"kind": "name", "name": name}, m.span())

    return {"refs": refs} if refs else {}


@dataclass(slots=True)
class OpenpyxlBackend:
    """Semantic backend: scans worksheets for formulas and basic cell stats.

    Notes:
    - openpyxl does not evaluate formulas; it reads the formula text.
    - This backend is best used after OOXMLZipBackend has registered sheets.
    """

    name: str = "openpyxl"

    def can_handle(self, ctx: AnalysisContext) -> bool:
        """Return True when openpyxl can parse the workbook and options require it."""
        p = ctx.path
        if not p.exists() or not p.is_file():
            return False
        if not (ctx.options.include_formulas or ctx.options.include_cells):
            return False
        return p.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm")

    def extract(self, ctx: AnalysisContext, graph: WorkbookGraph) -> BackendReport:
        """Scan worksheets for formulas and register them in the graph."""
        report = BackendReport(backend=self.name)

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except Exception as e:  # noqa: BLE001
            report.add_issue("openpyxl import failed", kind="dependency", detail=str(e))
            return report

        try:
            wb = load_workbook(
                filename=str(ctx.path),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as e:  # noqa: BLE001
            report.add_issue(
                "failed to load workbook", kind="parse_error", detail=str(e)
            )
            return report

        formula_count = 0
        sheets_scanned = 0

        worker_limit = ctx.options.formula_decomposition_workers
        with _formula_executor(worker_limit) as executor:
            for ws in wb.worksheets:
                sheets_scanned += 1
                if (
                    ctx.options.max_sheets is not None
                    and sheets_scanned > ctx.options.max_sheets
                ):
                    break

                # Ensure Sheet node exists (OOXML backend may have created it already).
                sheet_key = _sheet_key(ws.title)
                sheet_id = f"sheet:{sheet_key}"
                sheet_node = graph.get_by_key("sheet", sheet_key)
                if not sheet_node:
                    sheet_node = graph.upsert(
                        Sheet.make(key=sheet_key, id=sheet_id, name=ws.title)
                    )

                if not ctx.options.include_formulas:
                    continue

                scanned_cells = 0
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                if ctx.options.max_rows_per_sheet is not None:
                    max_row = min(max_row, ctx.options.max_rows_per_sheet)
                if ctx.options.max_cols_per_sheet is not None:
                    max_col = min(max_col, ctx.options.max_cols_per_sheet)
                if max_row == 0 or max_col == 0:
                    continue

                row_chunk = max(1, ctx.options.row_chunk_size)
                col_chunk = max(1, ctx.options.column_chunk_size)
                hit_cell_limit = False

                for row_start, row_end in _chunk_bounds(max_row, row_chunk):
                    if hit_cell_limit:
                        break
                    for col_start, col_end in _chunk_bounds(max_col, col_chunk):
                        if hit_cell_limit:
                            break
                        formula_batch: list[tuple[str, str, str]] = []
                        col_letters = [
                            get_column_letter(c)
                            for c in range(col_start, col_end + 1)
                        ]
                        for row_offset, row in enumerate(
                            ws.iter_rows(
                                min_row=row_start,
                                max_row=row_end,
                                min_col=col_start,
                                max_col=col_end,
                                values_only=True,
                            ),
                            start=0,
                        ):
                            row_idx = row_start + row_offset
                            for col_offset, v in enumerate(row, start=0):
                                scanned_cells += 1
                                if (
                                    ctx.options.max_cells_per_sheet is not None
                                    and scanned_cells > ctx.options.max_cells_per_sheet
                                ):
                                    hit_cell_limit = True
                                    break

                                if not isinstance(v, str) or not v.startswith("="):
                                    continue

                                addr = f"{col_letters[col_offset]}{row_idx}"
                                formula_batch.append((ws.title, addr, v))
                            if hit_cell_limit:
                                break

                        for (
                            sheet_name,
                            address,
                            formula,
                            deps,
                        ) in _resolve_formula_deps(
                            formula_batch, executor=executor, report=report
                        ):
                            key = f"{sheet_name}!{address}"
                            node = graph.upsert(
                                FormulaCell.make(
                                    key=key,
                                    id=f"formula:{key}",
                                    sheet_name=sheet_name,
                                    address=address,
                                    formula=formula,
                                    deps=deps,
                                )
                            )
                            # Use a contains edge to anchor formulas to their sheet.
                            graph.add_edge(sheet_node.id, node.id, "contains")
                            formula_count += 1

        report.stats = {
            "sheets_scanned": sheets_scanned,
            "formula_cells": formula_count,
            **graph.stats(),
        }
        return report
