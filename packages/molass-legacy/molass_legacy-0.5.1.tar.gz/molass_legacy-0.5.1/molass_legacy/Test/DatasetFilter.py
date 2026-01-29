# coding: utf-8
"""
    DatasetFilter.py

    Copyright (c) 2021, SAXS Team, KEK-PF
"""
import os
import re
from openpyxl import load_workbook
from DataUtils import cut_upper_folders

class DatasetFilter:
    def __init__(self, book_path):
        wb = load_workbook(book_path)
        ws = wb.active
        dataset_dict = {}
        for i in range(2, 500):
            row = []
            for j in [1, 2, 3]:
                c = ws.cell(row=i, column=j)
                if c.value is None:
                    break
                row.append(c.value)
            if len(row) == 0:
                break

            if row[2] != 'O':
                continue

            key = row[0][0:8]
            count = dataset_dict.get(key, 0)
            dataset_dict[key] = count+1

        self.dataset_dict = dataset_dict
        self.find_extra_info(os.path.dirname(book_path))

    def find_extra_info(self, folder):
        filter_txt = os.path.join(folder, "filter.txt")
        self.extra_dict = None
        if not os.path.exists(filter_txt):
            return

        in_folder_re = re.compile(r'(\S+)')
        self.extra_dict = {}
        fh = open(filter_txt)
        for line in fh:
            print(line)
            m = in_folder_re.search(line)
            if m:
                in_folder = m.group(1)
                nodes = in_folder.split('/')
                self.extra_dict[nodes[0][0:8]] = 1
        fh.close()

    def is_in_the_selection(self, in_folder):
        in_folder_ = cut_upper_folders(in_folder)
        nodes = in_folder_.split('/')
        key = nodes[0][0:8]
        count = self.dataset_dict.get(key, 0)
        if self.extra_dict is None:
            ret = count > 0
        else:
            ret = count > 0 and self.extra_dict.get(key, 0) == 1
        return ret
