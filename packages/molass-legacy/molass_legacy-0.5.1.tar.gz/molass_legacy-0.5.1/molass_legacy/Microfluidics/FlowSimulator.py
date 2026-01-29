# coding: utf-8
"""
    FlowSimulator.py

    Copyright (c) 2019, SAXS Team, KEK-PF
"""
import os
import numpy as np
from bisect import bisect_left
import logging
from openpyxl import load_workbook
from MethodFile import MethodFile

#                    D, F, H, L
TOTAL_RATE_COLUMNS = [3, 5, 7, 11]
RATE_B_START_COLUMN = 5
RATE_B_STOP_COLUMN = 6
SAFE_MARGIN_RATIO = 0.05

class SimInfo:
    def __init__(self, **kwargs): 
        self.__dict__.update(kwargs)

class FlowSimulator(MethodFile):
    def __init__(self, filepath, bookpath=None, num_files=None):
        MethodFile.__init__(self, filepath)
        self.logger = logging.getLogger(__name__)
        if bookpath is None:
            bookpath = filepath.replace('.mtd', '.xlsx')
        self.bookpath = bookpath
        self.num_files = num_files

        self.opt_factor = None
        self.start_duration = None
        self.start_frate_total = None
        self.start_time = None
        self.start_conc = None
        self.stop_duration = None
        self.stop_frate_total = None
        self.stop_time = None
        self.stop_conc = None
        self.excel_book_ok = False
        self.eno_interval = None

        try:
            self.read_the_book(bookpath)
            self.excel_book_ok = True
        except FileNotFoundError:
            pass
        except:
            from molass_legacy.KekLib.ExceptionTracebacker import ExceptionTracebacker
            etb = ExceptionTracebacker()
            print(etb)

    def read_the_book(self, bookpath):
        debug = True
        self.wb = load_workbook(bookpath, data_only=True)
        self.delay_factor = self.wb['Parameter']['B11'].value
        self.eno_interval = self.wb['Time vs conc.']['B3'].value
        self.time_delta = self.eno_interval/2
        if debug:
            print('delay_factor=', self.delay_factor, 'eno_interval=', self.eno_interval)

    def set_params(self, delay_factor=None, debug=False):
        if not self.excel_book_ok:
            self.set_params_without_book(delay_factor)
            return

        if delay_factor is None:
            delay_factor = self.delay_factor
        ws = self.wb['Method']
        range_ = ws['A3:O30']
        for k, row in enumerate(range_, start=1):
            values = [cell.value for cell in row]
            if debug:
                print([k], values)
            name = values[0]
            if name is None:
                break

            if name == 'const flow rate (pump A, B)':
                self.start_duration = values[-1]
            elif name == 'Total flow rate modulation':
                self.pre_start_time = values[-1]
            elif name == 'Y channel initialize':
                frate_total = np.sum([values[k] for k in TOTAL_RATE_COLUMNS])
                self.y_channel_time = values[-1] + delay_factor/frate_total
            elif name == 'Linearly changing flow rate':
                self.stop_duration = values[-1]
                self.start_frate_total = np.sum([values[k] for k in TOTAL_RATE_COLUMNS])
                self.start_time = self.start_duration + delay_factor/self.start_frate_total
                self.start_conc = values[RATE_B_START_COLUMN]/self.start_frate_total
                rate_b_stop = values[RATE_B_STOP_COLUMN]
            elif name == 'const. uL/min (pump B)':
                self.stop_frate_total = np.sum([values[k] for k in TOTAL_RATE_COLUMNS])
                self.stop_time = self.stop_duration + delay_factor/self.stop_frate_total
                self.stop_conc = rate_b_stop/self.stop_frate_total
            last_values = values

        self.end_time = last_values[-1] + delay_factor/np.sum([last_values[k] for k in TOTAL_RATE_COLUMNS])

        if self.num_files is not None:
            self.verify_eno_interval(self.num_files)

        if debug:
            for r in [  self.start_duration,
                        self.pre_start_time,
                        self.y_channel_time,
                        self.start_frate_total,
                        self.stop_duration,
                        self.stop_frate_total,
                        self.start_time,
                        self.stop_time,
                        self.start_conc,
                        self.stop_conc,
                        ]:
                print(r)

    def verify_eno_interval(self, num_files):
        ratio = num_files/self.end_time
        a_diff = abs(ratio - 1)
        ratio_xls = num_files*self.eno_interval/self.end_time
        a_diff_xls = abs(ratio_xls - 1)
        if a_diff_xls > a_diff:
            # TODO: this fix should be done earlier in self.set_params by giving num_files to it
            fix_value = 1.0
            self.logger.warning("Excel book eno_interval %g has been changed to %g due to time scale inconsistency.", self.eno_interval, fix_value)
            self.eno_interval = fix_value

    def set_params_without_book(self, delay_factor):
        from MethodFileSimulator import MethodFileSimulator
        MethodFileSimulator.set_params(self, delay_factor, default_init=True, debug=True)

    def get_y_channel_interval(self, t):
        i = bisect_left(t, self.pre_start_time)
        j = bisect_left(t, self.start_time)
        return i, j

    def get_linear_range_info(self, t, vscale=None):
        i = bisect_left(t, self.pre_start_time)
        c = bisect_left(t, self.y_channel_time)
        j = bisect_left(t, self.start_time)
        k = bisect_left(t, self.stop_time)
        print('t[[i, j, k]]=', t[[i, j, k]])
        return SimInfo(start=i, stop=None, yc_peak=c, lin_start=j, lin_end=k, factor=self.opt_factor, vscale=vscale)

    def get_y_channel_index(self, t):
        return bisect_left(t, self.y_channel_time)

    def get_optimizer_indeces(self, t):
        i = bisect_left(t, self.y_channel_time)
        j = bisect_left(t, self.start_time)
        k = bisect_left(t, self.stop_time)
        return i, j, k

    def make_simulation_data(self, delay_factor=None, debug=False):
        if delay_factor is None:
            delay_factor = self.delay_factor

        self.opt_factor = delay_factor

        self.set_params(delay_factor, debug=debug)

        tsec = 0
        tmin = None
        sm_rows = []
        for n, dur, a1, a2, b1, b2, c1, c2, d1, d2, e1, e2 in self.rows:
            tr1 = a1 + b1 + c1 + e1
            delay = 0  if n < 3 else delay_factor/tr1   # to avoid obscure appearance
            tmin = tsec/60 + delay
            if tsec > 0:
                sm_rows[-1][-2] = tmin
            tr2 = a2 + b2 + c2 + e2
            sm_rows.append([n, tr1, tmin, b1/tr1])
            sm_rows.append([n, tr2, None, b2/tr2])
            tsec += dur

        sm_rows[-1][-2] = tmin

        if debug:
            for k, row in enumerate(sm_rows):
                print([k], row)

        sm_array = np.array(sm_rows)
        t = sm_array[:,2]
        print('eno_interval=', self.eno_interval)
        x = (t - t[0]+self.time_delta)/self.eno_interval - 1      # int(x) = eno
        y = sm_array[:,3]
        return t, x, y

    def guess_range(self, num_files=None):
        safe_margin = (self.stop_time - self.start_time)*SAFE_MARGIN_RATIO
        start = int((self.start_time + safe_margin)/self.eno_interval)
        stop  = int((self.stop_time - safe_margin)/self.eno_interval)+1

        if self.excel_book_ok:
            # self.eno_interval should have been verified
            pass
        else:
            ratio = num_files/self.end_time
            if abs(ratio - 1) > 0.1:
                self.eno_interval = 1/ratio
                start = int(start*ratio)
                stop = int(stop*ratio)
                self.logger.warning("Default eno_interval has been changed to %g due to time scale inconsistency.", self.eno_interval)

        return slice(start, stop)
