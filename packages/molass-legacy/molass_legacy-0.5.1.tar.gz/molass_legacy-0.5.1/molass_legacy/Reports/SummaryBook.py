"""

    SummaryBook.py

    Copyright (c) 2017-2025, SAXS Team, KEK-PF

"""
import os
import numpy                as np
import csv
import re
import logging
from datetime               import datetime
# from openpyxl               import Workbook
from openpyxl.styles        import Alignment
from openpyxl.styles.fonts  import Font
from openpyxl.utils         import get_column_letter
from openpyxl.chart         import BarChart, Reference, Series
from openpyxl.chart.data_source     import StrRef
from molass_legacy.KekLib.OpenPyXlUtil import save_allowing_user_reply
from molass_legacy.AutorgKek.AppVersion import autoguinier_version_for_publication
from molass_legacy._MOLASS.Version import get_version_string, molass_version_for_publication
from molass_legacy.ATSAS.AtsasVersion import atsas_version_for_publication
from molass_legacy._MOLASS.SerialSettings import get_setting, SHOW_TEXT_DICT, get_beamline_name, get_xray_picking
from molass_legacy.UV.PlainCurveUtils import get_flat_wavelength
from molass_legacy.SerialAnalyzer.LinearityScore import linearity_score100, stderror_score100, FACTOR_WEIGHT
from molass_legacy.KekLib.TimeUtils import seconds_to_datetime
from molass_legacy.Reports.DefaultFont import set_default_font
set_default_font()

NUM_SCORE_FACTORS   = 5
ZX_RATIO_MAX        = 0.97

def round_3g_str( v ):
    return 'NA' if v is None else '%.3g' % v

def round_3g( values ):
    return [ None if x is None else float( '%.3g' % x ) for x in values ]

def round_3g_with_error( values, errors ):
    return [ None if x is None else '%.3g±%.3g' % ( x, e ) for x, e in zip( values, errors ) ]

def np_product_safe( x, vec ):
    assert len( x.shape ) == 1
    assert len( x ) == len( vec )
    return np.array( [ None if v is None else v*vec[i] for i, v in enumerate( x ) ] )

class SummaryBook:
    def __init__( self, wb, controller ):
        self.logger = logging.getLogger( __name__ )
        self.parent = controller
        self.doing_sec = controller.doing_sec
        self.atsas_is_available = controller.atsas_is_available
        self.datgnom = controller.datgnom

        self.j0 = j0 = controller.xr_j0
        self.applied_ranges = controller.applied_ranges
        self.report_ranges = controller.report_ranges
        self.wb = wb
        self.ws_list = []
        self.has_extrapolation_result = len( controller.zx_summary_list ) > 0

        # As of openpyxl 2.4.8,
        # using this wb.active sheet causes a problem with EXCEL 2016.
        # wb.remove_sheet(wb.active)

        if self.has_extrapolation_result:
            ws = wb.create_sheet( "Extrapolation Summary" )
            self.ws_list.append( ws )
            self.make_zx_summary( ws, controller )

        ws = wb.create_sheet( "Entire Summary" )
        self.ws_list.append( ws )
        self.make_entire_summary( ws, controller )

        if self.has_extrapolation_result:
            ws = wb.create_sheet( "Summary for Publication" )
            self.ws_list.append( ws )
            self.make_publication_summary( ws, controller )

    def append(self, ws, row_list, alignments={}, number_formats={}, fonts={}, character_fonts={}, for_all_components=False, row_height=None ):
        self.row += 1
        ws.append( row_list )

        row = self.row

        for k, v in alignments.items():
            cell = ws.cell(row=row, column=k)
            if type(v) == int:
                cell.alignment = Alignment( indent=v )
            elif type(v) == str:
                cell.alignment = Alignment( horizontal=v )
            elif type(v) == dict:
                cell.alignment = Alignment( **v )
            else:
                assert False

        for k, v in number_formats.items():
            cell = ws.cell(row=row, column=k)
            cell.number_format = v

        for k, v in fonts.items():
            cell = ws.cell(row=row, column=k)
            cell.font = v

        if for_all_components and self.num_components > 1:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=self.num_components+1)

        if row_height is not None:
            ws.row_dimensions[row].height = row_height

        if False:
            # openpyxl does not yet support these characters attr
            for k, v in character_fonts.items():
                cell = ws.cell(row=row, column=k)
                cell.characters[v[0]].font = v[1]

    def make_zx_summary( self, ws, controller  ):
        fontname = get_setting("report_default_font")

        # column lengths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 24

        array1 = np.array( controller.zx_summary_list )
        zx_summary_array = np.hstack( [ array1, np.zeros( ( array1.shape[0], 1 + NUM_SCORE_FACTORS ) ) ] )

        zx_summary_array2 = np.array( controller.zx_summary_list2 )

        min_c = zx_summary_array[:,5]
        max_c = zx_summary_array[:,6]
        zx_ratio = ( max_c - min_c ) / max_c
        zx_summary_array[:,15] = zx_ratio

        zx_summary_array[:,16] = linearity_score100( zx_summary_array[:,11] )
        zx_summary_array[:,17] = stderror_score100(  zx_summary_array[:,12] )
        zx_summary_array[:,18] = linearity_score100( zx_summary_array[:,13] )
        zx_summary_array[:,19] = stderror_score100(  zx_summary_array[:,14] )
        zx_summary_array[:,20] = round_3g( np.min( [ np.ones( zx_ratio.shape ), zx_ratio/ZX_RATIO_MAX ], axis=0 )*FACTOR_WEIGHT )

        n_peak = int( np.max( zx_summary_array[:,0] ) )
        self.zx_summary_array = zx_summary_array
        self.zx_summary_array2 = zx_summary_array2
        self.n_peak = n_peak
        # print( 'n_peak=', n_peak )

        self.row    = 0
        peak_side_titles    = []
        peak_side_cols      = []
        peak_side_cells     = []
        peak_side_formulae  = []
        score_row   = 35
        range_type = get_setting('range_type')

        for i in range( zx_summary_array.shape[0] ):
            col = get_column_letter( i + 3 )
            ws.column_dimensions[col].width = 12
            m = int( zx_summary_array[i, 0] )
            j = zx_summary_array[i, 1]
            peak_num_ranges = controller.peak_num_ranges_list[m]

            if n_peak == 0:
                paren = ''
            else:
                paren = '(%d)' % (m+1)

            if self.doing_sec:
                if peak_num_ranges == 1 and range_type < 5:
                    comp = 'Both-sides'
                else:
                    comp = 'Asc-side' if j == 0 else 'Desc-side'
            else:
                comp = 'Nat-comp' if m == 0 else 'Unf-comp'
            comp_name = '%s%s' % ( comp, paren )
            peak_side_titles.append( comp_name )
            peak_side_cols.append( col )
            peak_side_cells.append( col+'1' )
            # TODO: compute 15, 18
            peak_side_formulae.append( '=SUM(%s%d:%s%d)' % ( col, score_row, col, score_row+4 ) )

        self.append( ws, [ None, None ] + peak_side_titles )
        for cell in peak_side_cells:
            ws[cell].alignment = Alignment(horizontal="center")

        factors = [ 'Rg Extrapolation Linearity', 'Rg Extrapolation StdError', 'I(0)/C Extrapolation Linearity',  'I(0)/C Extrapolation StdError', 'Extrapolation Ratio' ]

        self.append( ws, [ 'Extrapolation Range', 'From'              ] + list( self.j0 + zx_summary_array[:,2] ) )
        self.append( ws, [ None,                  'To'                ] + list( self.j0 + zx_summary_array[:,3] ) )
        self.append( ws, [ None,                  'Number of Points'  ] + list( zx_summary_array[:,4] ) )
        self.append( ws, [] )
        self.append( ws, [ 'Concentration',       'Minimum'           ] + round_3g( min_c ) )
        self.append( ws, [ None,                  'Maximum'           ] + round_3g( max_c ) )
        self.append( ws, [] )
        self.append( ws, [ 'Guinier Analysis',    'Rg'                ] + round_3g( zx_summary_array[:,7] ) )
        guinier_start_row = self.row
        self.append( ws, [ None,                  'Rg error'          ] + round_3g( zx_summary_array[:,8] ) )
        rg_results = controller.guinier_results
        basic_qualities = [result.basic_quality for result in rg_results]
        self.append( ws, [ None,                  'Rg basic quality'  ] + round_3g(basic_qualities) )
        self.append( ws, [ None,                  'Rg quality'        ] + round_3g([result.Quality for result in rg_results]) )
        self.append( ws, [ None,                  'qmin'              ] + round_3g([result.min_q for result in rg_results]) )
        self.append( ws, [ None,                  'qmax'              ] + round_3g([result.max_q for result in rg_results]) )
        self.append( ws, [ None,                  'qRg max'           ] + round_3g([result.max_qRg for result in rg_results]) )
        self.append( ws, [ None,                  'I(0)/C'            ] + round_3g( zx_summary_array[:,9] ) )
        self.append( ws, [ None,                  'I(0)/C error'      ] + round_3g( zx_summary_array[:,10] ) )

        iz_max_c        = np_product_safe( zx_summary_array[:, 9],  max_c )
        iz_err_max_c    = np_product_safe( zx_summary_array[:,10],  max_c )
        self.append( ws, [ None,                  'I(0) (max C)'        ] + round_3g( iz_max_c ) )
        self.append( ws, [ None,                  'I(0) error (max C)'  ] + round_3g( iz_err_max_c ) )
        guinier_end_row = self.row

        guinier_warning = False
        for k, quality in enumerate(basic_qualities):
            if quality < 0.5:
                guinier_warning = True
                for i in range(guinier_start_row, guinier_end_row+1):
                    cell = ws.cell(row=i, column=3+k)
                    cell.font = Font(name=fontname, color='FF0000')

        if guinier_warning:
            row = guinier_start_row + 2
            column = 4 + len(basic_qualities)
            for i, line in enumerate([
                "For Rg basic quality < 0.5, qmax may have been intentionally widened",
                "to get a possibly better Rg value resulting in an undesired qRg max > 1.3,",
                "which violates the recommended guideline.",
                "See Section 5.1.1 of the User's Guide for details.",
                ]):
                cell = ws.cell(row=row + i, column=column)
                cell.value = line
                cell.font = Font(name=fontname, color='FF0000')

            self.logger.warning("Summary includes low quality results. Be aware.")

        if self.atsas_is_available:
            pr_rgs = round_3g( zx_summary_array2[:,0] )
            pr_rg_errors = round_3g( zx_summary_array2[:,1] ) 
            pr_qmins = round_3g( zx_summary_array2[:,5] )
            pr_qmaxes = round_3g( zx_summary_array2[:,6] ) 
            pr_izs = round_3g( zx_summary_array2[:,2] )
            pr_iz_errors = round_3g( zx_summary_array2[:,3] )
            pr_dmaxes = round_3g( zx_summary_array2[:,4] )
        else:
            pr_rgs = []
            pr_rg_errors = []
            pr_qmins = []
            pr_qmaxes = []
            pr_izs = []
            pr_iz_errors = []
            pr_dmaxes = []
        self.append( ws, [ None ])
        self.append( ws, [ 'P(r) Analysis',       'Rg'       ] + pr_rgs )
        self.append( ws, [ None,                  'Rg error' ] + pr_rg_errors )
        self.append( ws, [ None,                  'qmin'       ] + pr_qmins )
        self.append( ws, [ None,                  'qmax'       ] + pr_qmaxes )
        self.append( ws, [ None,                  'I(0)/C'       ] + pr_izs )
        self.append( ws, [ None,                  'I(0)/C error' ] + pr_iz_errors )
        self.append( ws, [ None,                  'Dmax'                ] + pr_dmaxes )
        self.append( ws, [] )
        self.append( ws, [ 'Quality Measurement', factors[0]          ] + round_3g( zx_summary_array[:,11] ) )
        self.append( ws, [ None,                  factors[1]          ] + round_3g( zx_summary_array[:,12] ) )
        self.append( ws, [ None,                  factors[2]          ] + round_3g( zx_summary_array[:,13] ) )
        self.append( ws, [ None,                  factors[3]          ] + round_3g( zx_summary_array[:,14] ) )
        self.append( ws, [ None,                  factors[4]          ] + round_3g( zx_summary_array[:,15] ) )
        self.append( ws, [] )
        self.append( ws, [ 'Quality Score',       factors[0]          ] + list( zx_summary_array[:,16] ) )
        assert self.row == score_row
        self.append( ws, [ None,                  factors[1]          ] + list( zx_summary_array[:,17] ) )
        self.append( ws, [ None,                  factors[2]          ] + list( zx_summary_array[:,18] ) )
        self.append( ws, [ None,                  factors[3]          ] + list( zx_summary_array[:,19] ) )
        self.append( ws, [ None,                  factors[4]          ] + list( zx_summary_array[:,20] ) )
        self.append( ws, [ None,                  'Total Score'         ] + peak_side_formulae  )

        for row in range(score_row,score_row+5):
            for col in peak_side_cols:
                ws[ col+str(row) ].number_format = '0.0'

        # Stacked Bar Chart
        num_cols = len(peak_side_cols)

        c_ = BarChart()
        c_.title    = "Quality Scores"
        c_.type     = "col"
        c_.style    = 11
        c_.grouping = "stacked"
        c_.overlap  = 100
        categories  = Reference(ws, min_col=3, max_col=2+num_cols, min_row=1)
        for k, factor in enumerate(factors):
            values = Reference(ws, min_col=3, max_col=2+num_cols, min_row=score_row+k)
            series = Series(values, title=factor)
            c_.series.append(series)
        c_.set_categories(categories)
        c_.y_axis.scaling.max = 100

        chart_start_pos = get_column_letter( 4 + num_cols ) + '21'
        ws.add_chart(c_, chart_start_pos)

    def make_entire_summary( self, ws, controller ):

        opt_params = controller.mapped_info.opt_params
        if False:
            self.logger.info('opt_params=' + str(opt_params))

        # column lengths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60

        self.row    = 0

        # Column Title
        self.append( ws, [ 'Category',                    'Item Name',                                'Item Value' ] )

        # Input Data
        self.append( ws, [ 'Input Data',                'Xray Scattering Data Folder',              get_setting( 'in_folder' ) ] )
        self.append( ws, [ None,                        'Xray Scattering Data Shape',               str(controller.serial_data.intensity_array.shape) ] )
        self.append( ws, [ None,                        'UV Absorbance Data Folder/File',           get_setting( 'uv_folder' ) + '/' + get_setting( 'uv_file' ) ] )
        self.append( ws, [ None,                        'UV Absorbance Data Shape',                 str(controller.serial_data.conc_array.shape) ] )
        self.append( ws, [] )

        # Output Data
        self.append( ws, [ 'Output Data',               'Analysis Result Folder/Book',              get_setting( 'analysis_folder' )+ '/' + get_setting( 'result_book' ) ] )
        self.append( ws, [ None,                        'Averaged Data Folder',                     get_setting( 'averaged_data_folder' ) ] )
        self.append( ws, [ None,                        'Extrapolated Data Folder',                 controller.zx_out_folder ] )
        self.append( ws, [] )
        self.append( ws, [ 'Input Data Pre-processing', 'Number of Elution Points for Averaging',   get_setting( 'num_curves_averaged' ) ], alignments={ 3:'left' } )
        self.append( ws, [] )

        # Concentration Factors
        self.append( ws, [ 'Concentration Factors',     'Path Length Factor',                       get_setting( 'path_length' ) ], alignments={ 3:'left' }, number_formats={3:'0.0000'} )
        self.append( ws, [ None,                        'Extinction Coefficient',                   get_setting( 'extinction' ) ],  alignments={ 3:'left' }, number_formats={3:'0.0000'} )
        self.append( ws, [] )

        # UV/Xray Mapping Settings
        self.append( ws, [ 'UV/Xr-ay Mapping Settings', 'UV Absorbance Picking Wavelength λ₁ (nm)', get_setting('absorbance_picking')], alignments={ 3:'left' }, number_formats={3:'0.0'} )
        self.append( ws, [ None,                        'UV Absorbance Baseline Wavelength λ₂ (nm)', get_flat_wavelength(controller.serial_data.lvector) ], alignments={ 3:'left' }, number_formats={3:'0.0'} )
        uv_correction = opt_params.get_uv_correction_str()
        self.append( ws, [ None,                        'UV Absorbance Baseline Correction', uv_correction] )
        self.append( ws, [ None,                        'Xray Picking Scattering Vector q (Å⁻¹)',     get_xray_picking()], alignments={ 3:'left' }, number_formats={3:'0.000'} )
        xray_correction = opt_params.get_xray_correction_str()
        self.append( ws, [ None,                        'Xray Scattering Baseline Correction', xray_correction ] )
        self.append( ws, [] )

        # Zero-Concentration Analysis Settings
        matrix_formulation = get_setting('matrix_formulation')
        use_elution_models = get_setting('use_elution_models')
        if matrix_formulation:
            with_without = "with" if use_elution_models else "without"
            formulation_text = "Matrix formulation %s elution models" % with_without
        else:
            formulation_text = "Non-matrix formulation (iteration for each Q)"

        self.append( ws, [ 'Zero-Concentration Analysis Settings',  'Optimizer Formulation',  formulation_text  ], alignments={ 3:'left' } )

        text_dict   = SHOW_TEXT_DICT['range_type']
        range_type  = get_setting( 'range_type' )
        show_text   = text_dict[range_type]
        self.append( ws, [ None,                        'Range Determination Policy',  show_text  ], alignments={ 3:'left' } )

        ranges_str = ', '.join( [ str( range_ ) for range_ in self.report_ranges ]  )
        self.append( ws, [ None,                        'Analysis Ranges',                  ranges_str  ] )
        if not matrix_formulation:
            self.append( ws, [ None,                    'Width of Regression for Q-axis',   get_setting( 'zx_num_q_points' )  ], alignments={ 3:'left' } )

            zx_boundary_method = get_setting( 'zx_boundary_method' )
            if zx_boundary_method == 'FIXED':
                show_text   = 'Fixed at %g' % ( get_setting( 'zx_boundary' ) )
            else:
                text_dict   = SHOW_TEXT_DICT['zx_boundary_method']
                show_text   = text_dict[zx_boundary_method]
            self.append( ws, [ None,                    'Regression Boundary Method',       show_text  ] )
        self.append( ws, [] )

        # Analysis Quality
        self.append( ws, [ 'Analysis Quality',   'UV/Xray Mapping Adequacy',         None  ] )
        self.append( ws, [ None,                        'Xray Scattering Baseline Fitness', None  ] )
        self.append( ws, [] )

        # Time Info
        self.append( ws, [ 'Time Info',                 'Date/Time Created',                datetime.now() ], alignments={ 3:'left' } )
        self.append( ws, [ None,                        'Basesurface Correction',           seconds_to_datetime(controller.seconds_correction)], number_formats={3:'mm:ss'}, alignments={3:'left'} )
        self.append( ws, [ None,                        'Guinier Analysis Execution',       seconds_to_datetime(controller.seconds_guinier)], number_formats={3:'mm:ss'}, alignments={3:'left'} )
        self.append( ws, [ None,                        'Extrapolation Execution',          seconds_to_datetime(controller.seconds_extrapolation)], number_formats={3:'mm:ss'}, alignments={3:'left'} )
        self.append( ws, [] )

        # Other Info
        self.append( ws, [ 'Other Info',                'Program Version',          get_version_string() ] )
        env_info = controller.env_info
        self.append( ws, [ None,                        'Excel Version',            env_info.excel_version ] )
        self.append( ws, [ None,                        'Logical Number of CPUs',   os.cpu_count() ], alignments={3:'left'} )
        if env_info.gpu_info is not None:
            self.append( ws, [ None,                    'CUDA Tools Version',       env_info.cuda_tools_ver_str ] )

    def make_publication_summary( self, ws, controller ):
        fontname = get_setting("report_default_font")
        bold_font = Font(name=fontname, bold=True)
        gray_font = Font(name=fontname, color="C0C0C0")
        red_font = Font(name=fontname, color='FF0000')

        try:
            xray_info = self.get_xray_scattering_info()
        except Exception as exc:
            print( exc )
            xray_info = {}

        # q range
        serial_data = controller.serial_data
        qvector     = serial_data.qvector
        q_range     = '%.3g - %.3g' % tuple( qvector[ [0,-1] ] )    # note that qvector has been already trimmed

        # column lengths
        ws.column_dimensions['A'].width = 60

        zx_summary_array = self.zx_summary_array
        peak_names      = []
        peak_names_align  = {}
        conc_ranges     = []
        num_frames_list = []

        # self.num_components = len(self.applied_ranges)
        self.num_components = zx_summary_array.shape[0]     # here, the ascending side and descending side are treated as separate components

        k = 0
        for p, range_ in enumerate( self.applied_ranges ):
            fromto_list = range_.get_fromto_list()
            for j, ft in enumerate(fromto_list):
                col = get_column_letter(k+2)
                ws.column_dimensions[col].width = 17
                if len(fromto_list) == 1:
                    side = "bth"
                else:
                    side = "asc" if j == 0 else "dsc"
                peak_names.append( 'Component-%d (%s)' % (p+1, side) )
                conc_ranges.append( '%.3g - %.3g' % tuple(zx_summary_array[k, [5,6]]) )
                num_frames = ft[1] - ft[0] + 1
                num_frames_list.append(num_frames)
                peak_names_align[k+3] = 'center'
                k += 1

        self.row    = 0
        self.append( ws, [ 'A. Sample details',                 ], fonts={ 1:bold_font } )
        peak_names_align[1] = 1     # add an indent
        self.append( ws, [   'Sample name(s)' ] + peak_names, alignments=peak_names_align )
        self.append( ws, [   'Organism',                        ], alignments={ 1:1 } )
        self.append( ws, [   'Source',                          ], alignments={ 1:1 } )
        self.append( ws, [   'Uniprot ID (residues in construct)',   ], alignments={ 1:1 } )
        self.append( ws, [   'Extinction coefficient (A280nm, 0.1%(w/v))', get_setting( 'extinction' )  ], alignments={ 1:1, 2:'left' }, number_formats={2:'0.0000'}, for_all_components=True )
        self.append( ws, [   'P.S.V, ν from chemical composition (cm³ g⁻¹)' ], alignments={ 1:1 } )
        self.append( ws, [   'Contrast, Δρ (1010 cm⁻²)',      ], alignments={ 1:1 } )
        self.append( ws, [   'Theoretical MW (Da) ',            ], alignments={ 1:1 } )
        self.append( ws, [   'SEC-SAXS column',                 ], alignments={ 1:1 } )
        self.append( ws, [   'Loading conc. (mg ml⁻¹)',         ], alignments={ 1:1 } )
        self.append( ws, [   'Injection volume (μL)',          ], alignments={ 1:1 } )
        self.append( ws, [   'Flow rate (ml min⁻¹)',            ], alignments={ 1:1 } )
        self.append( ws, [   'Conc. range (mg ml⁻¹)',           ] + conc_ranges, alignments={ 1:1 } )
        self.append( ws, [   'Solvent composition',             ], alignments={ 1:1 } )
        if get_setting('use_xray_conc'):
            conc_method = None
            for_all_components = None
        else:
            conc_method =  "UV-Visible spectroscopy"
            for_all_components = True
        self.append( ws, [   'Concentration method', conc_method ], alignments={ 1:1 }, for_all_components=for_all_components )
        self.append( ws, [] )

        self.append( ws, [ 'B. SAS data collection',            ], fonts={ 1:bold_font } )

        if False:
            self.append( ws, [  'Beam geometry (μm)',              xray_info.get( 'Beam geometry' )  ], alignments={ 1:1 } )
            self.append( ws, [  'Wavelength (Å)',                   xray_info.get( 'Wavelength' )  ], alignments={ 1:1, 2:'left' } )
            self.append( ws, [  'Exposure time (s)',                xray_info.get( 'Exposure time' )  ], alignments={ 1:1, 2:'left' } )
            self.append( ws, [  'Temperature (K)',                  ], alignments={ 1:1 } )

        self.append( ws, [   'Instrument', get_beamline_name(formal=True)  ], alignments={ 1:1 }, for_all_components=True )
        self.append( ws, [   'Wavelength (Å)',                  ], alignments={ 1:1 } )
        self.append( ws, [   'Beam geometry (μm)',             ], alignments={ 1:1 } )
        self.append( ws, [   'Camera length (mm)',              ], alignments={ 1:1 } )
        self.append( ws, [   'Q range (Å⁻¹)', q_range  ], alignments={ 1:1 }, character_fonts={ 1:(slice(0,1),bold_font) }, for_all_components=True )
        self.append( ws, [   'Absolute scaling method', 'Comparison with scattering from 1 mm pure H₂O'  ], alignments={ 1:1 }, for_all_components=True )
        self.append( ws, [   'Exposure time (sec), No. of frames' ] + [" , %d" % n for n in num_frames_list], alignments={ 1:1 } )
        self.append( ws, [   'Sample path length (mm)',         1, ], alignments={ 1:1, 2:'left' }, number_formats={2:'0.00'}, for_all_components=True )
        self.append( ws, [   'Sample temperature (K)',          ], alignments={ 1:1 } )
        if self.num_components > 1:
            alignments = { 1:{'indent':1, 'vertical':'center'}, 2:{'wrap_text':True} }
            row_height = 29.0
        else:
            alignments = { 1:1 }
            row_height = None
        self.append( ws, [   'Normalization', 'To incident and transmitted intensities measured with a micro ion chamber and a beam stopper with embedded photodiode'  ],
                                                                    alignments=alignments, for_all_components=True, row_height=row_height )
        self.append( ws, [   'Method for monitoring radiation damage',   'Data frame-by-frame comparison'  ], alignments={ 1:1 }, for_all_components=True )
        self.append( ws, [] )

        self.append( ws, [ 'C. Software employed for SAS data reduction, analysis and interpretation',  ], fonts={ 1:bold_font } )
        sangler_version = get_setting('sangler_version')
        sangler_version_str = "Unknown" if sangler_version is None else "SAngler " + sangler_version
        program_versions = "%s, %s" %(sangler_version_str, molass_version_for_publication())
        self.append( ws, [  'SAS data processing',              program_versions  ], alignments={ 1:1 }, for_all_components=True )
        self.append( ws, [  'Extinction coefficient estimate',          'ProtParam' ], alignments={ 1:1 }, fonts={ 2:gray_font } )
        self.append( ws, [  'Calculation of contrast and PSV values',   'MULCh'     ], alignments={ 1:1 }, fonts={ 2:gray_font } )
        if self.atsas_is_available:
            datgnom_path = self.datgnom.exe_path
            datgnom_re = re.compile(r"(datgnom\w*)")
            m = datgnom_re.search(datgnom_path)
            exename = m.group(1)    # datgnom or datgnom4
            datgnom_str = ', %s (%s)'% (exename, atsas_version_for_publication())
        else:
            datgnom_str = ''
        program_versions = "%s from %s%s" % (autoguinier_version_for_publication(), molass_version_for_publication(), datgnom_str)
        self.append( ws, [  'Basic analyses: (Guinier, P(r))',  program_versions  ],
                                                            alignments={ 1:1 }, character_fonts={ 1:(slice(26,27),bold_font) },
                                                            for_all_components=True )
        self.append( ws, [  'Shape modeling',                   ], alignments={ 1:1 } )
        self.append( ws, [  'Atomistic modeling',               ], alignments={ 1:1 } )
        self.append( ws, [  '3D graphic model representations', ], alignments={ 1:1 } )
        self.append( ws, [] )

        self.append( ws, [ 'D. Structural parameters',  ], fonts={ 1:bold_font } )
        self.append( ws, [ 'Guinier Analysis',  ], alignments={ 1:1 }, fonts={ 1:bold_font } )
        self.append( ws, [   'I(0) (cm⁻¹)',                     ] + round_3g_with_error(zx_summary_array[:,9], zx_summary_array[:,10]), alignments={ 1:2 } )
        guinier_start_row = self.row
        self.append( ws, [   'Rg (Å)',                          ] + round_3g_with_error(zx_summary_array[:,7], zx_summary_array[:,8]), alignments={ 1:2 } )

        rg_results = controller.guinier_results
        q_ranges = [ "%s - %s" % (round_3g_str(result.min_q), round_3g_str(result.max_q)) for result in rg_results]
        self.append( ws, [   'Q range (Å⁻¹)',                   ] +  q_ranges, alignments={ 1:2 } )
        qrg_ranges = [ "%s - %s" % (round_3g_str(result.min_qRg), round_3g_str(result.max_qRg)) for result in rg_results]
        self.append( ws, [   'Q×Rg range',                     ] + qrg_ranges, alignments={ 1:2 } )
        self.append( ws, [   'M from I(0) (ratio to predicted value)',    ], alignments={ 1:2 } )
        guinier_end_row = self.row

        max_qRg_warning  = False
        max_qRg_list = [result.max_qRg for result in rg_results]
        for j, max_qRg in enumerate(max_qRg_list):
            if max_qRg > 1.3:
                max_qRg_warning = True
                for i in range(guinier_start_row, guinier_end_row+1):
                    cell = ws.cell(row=i, column=2+j)
                    cell.font = red_font
        if max_qRg_warning:
            cell = ws.cell(row=guinier_start_row+3, column=3+len(max_qRg_list))
            cell.font = red_font
            cell.value = 'See "Extrapolation Summary" sheet for quality warnings.'

        zx_summary_array2 = self.zx_summary_array2 

        self.append( ws, [  'P(r) Analysis',                    ], alignments={ 1:1 }, fonts={ 1:bold_font })
        self.append( ws, [   'I(0) (cm⁻¹)',                     ] + round_3g_with_error(zx_summary_array2[:,2], zx_summary_array2[:,3]), alignments={ 1:2 } )
        self.append( ws, [   'Rg (Å)',                          ] + round_3g_with_error(zx_summary_array2[:,0], zx_summary_array2[:,1] ), alignments={ 1:2 } )
        alignments={ 1:2 }
        for k in range(self.num_components):
            alignments[k+2] = 'left'
        self.append( ws, [   'Dmax (Å)',                        ] + round_3g(zx_summary_array2[:, 4]), alignments=alignments )
        qranges = []
        if self.atsas_is_available:
            for qmin, qmax in zip(zx_summary_array2[:,5], zx_summary_array2[:,6]):
                qranges.append('%.4g - %.4g' % (qmin, qmax))
        self.append( ws, [   'Q range(Å⁻¹)',                    ] + qranges, alignments={ 1:2 } )
        self.append( ws, [   'M from I(0) (ratio to predicted value)',    ], alignments={ 1:2 } )
        self.append( ws, [   'Porod volume (Å³) (ratio to predicted)',  ], alignments={ 1:2 } )
        self.append( ws, [   'M from Bayesian Inference',       ], alignments={ 1:2 } )

        self.append( ws, [] )
        self.append( ws, [] )

        self.append( ws, [ '※This table was constructed with reference to the 2017 BioSAXS data publication guidelines.\n(https://doi.org/10.1107/S2059798317011597)', ],
                        alignments={ 1:1, 1:{'wrap_text':True} }, fonts={ 1:bold_font } )

    def get_xray_scattering_info( self ):
        ret_dict = {}
        in_folder = get_setting( 'in_folder' )
        info_file = os.path.join(in_folder, 'Xray_Scattering_Info.txt')
        if not os.path.isfile( info_file ):
            self.logger.warning( 'Xray_Scattering_Info.txt not found in %s' % in_folder )
            return ret_dict
        fh = open( info_file )
        wavelength_re   = re.compile( r'Wavelength:\s+(\d+\.\d+)' )
        pixel_size_re   = re.compile( r'PixelSize:\s+(\d+e-\d+)\s*m\s+x\s+(\d+e-\d+)\s*m' )
        expos_time_re   = re.compile( r'Exposure Time:\s+(\d+\.\d+)\s*s'  )
        for line in fh:
            m =  wavelength_re.search( line )
            if m:
                wavelength = m.group(1)
                print( 'wavelength=', wavelength )
                ret_dict['Wavelength'] = float(wavelength)
                continue

            m =  pixel_size_re.search( line )
            if m:
                h_res = m.group(1)
                v_res = m.group(2)
                geometry = '%d × %d' % tuple([ int(float(x)*1e6) for x in [ h_res, v_res ] ])
                print( 'geometry=', geometry )
                ret_dict['Beam geometry'] = geometry
                continue

            m =  expos_time_re.search( line )
            if m:
                expos_time = m.group(1)
                print( 'expos_time=', expos_time )
                ret_dict['Exposure time'] = float(expos_time)
                continue

        fh.close()
        return ret_dict

    def add_format_setting(self, book_file):
        from .SummaryExcelFormatter import SummaryArgs

        ws = self.ws_list[-1]
        if ws.title.find("Publication") < 0:
            return

        args = SummaryArgs(ws.title, book_file)

        if self.parent.more_multicore:
            self.add_format_setting_more_multicore(args)
        else:
            self.add_format_setting_less_multicore(args)

    def add_format_setting_more_multicore(self, args):
        self.parent.teller.tell('summary_book', args=args)

    def add_format_setting_less_multicore(self, args):
        from .SummaryExcelFormatter import add_summary_format_setting
        add_summary_format_setting(self.parent.excel_client, args, self.logger)

    def save( self, xlsx_file ):
        save_allowing_user_reply( self.wb, xlsx_file )

    def save_as_csv( self, path ):
        """
            learned at
            https://stackoverflow.com/questions/10802417/how-to-save-an-excel-worksheet-as-csv-from-python-unix
        """
        for i, ws in enumerate( self.ws_list ):
            with open( path.replace( '.csv', '-%d.csv' % i ), 'w', encoding='utf-8', newline='') as f:
                c = csv.writer(f)
                for r in ws.rows:
                    row = [cell.value for cell in r]
                    c.writerow(row)
