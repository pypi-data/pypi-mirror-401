# coding: utf-8
"""

    OpenPyXlUtils.py

    Copyright (c) 2019, Masatsuyo Takahashi, KEK-PF

"""
from openpyxl import Workbook, load_workbook

def merge_into_a_book( input_books, merged_book, progress_cb=None ):
    """
        it is not supported in openpyxl.
        instead, you can do as follows.

        wb = Workbook()
        wb.remove(wb.active)

        for i in range(3):
            ii = i + 1
            ws = wb.create_sheet('TestSheet%d' % ii)
            # edit ws

        wb.save(out_book)

        See the current state in the following pages.

        copy a worksheet in openpyxl
        https://stackoverflow.com/questions/34808394/copy-a-worksheet-in-openpyxl

        "I'm afraid copying worksheets is not supported because it is far from easy to do."

        Add a new sheet to a existing workbook in python
        https://stackoverflow.com/questions/40385689/add-a-new-sheet-to-a-existing-workbook-in-python
    """
    pass
