"""

    StageExtrapolation.py

    Copyright (c) 2016-2023, SAXS Team, KEK-PF

"""
import os
import numpy as np
from scipy import stats
from bisect import bisect_right
from openpyxl import Workbook
from molass_legacy.KekLib.NumpyUtils import np_savetxt, get_valid_index, simply_safe_sprintf
from molass_legacy.KekLib.BasicUtils import mkdirs_with_retry, ordinal_str 
from molass_legacy.KekLib.ExceptionTracebacker import ExceptionTracebacker, log_exception
from molass_legacy.AutorgKek.KekToolsGP import AutorgKek, ErrorResult
from molass_legacy._MOLASS.SerialSettings import get_setting
from molass_legacy.AutorgKek.AtsasTools import autorg as autorg_atsas
from molass_legacy.ATSAS.Almerge import AlmergeExecutor
from molass_legacy.ATSAS.DatGnom import DatgnomExecutor
from molass_legacy.SerialAnalyzer.ZeroExtrapolator import ZeroExtrapolator, GuinierPorodAnalyzer, NUM_EXTRAPOLATION_POINTS
from molass_legacy.Reports.ZeroExtrapolationResultBook import divide
from molass_legacy.Reports.DatgnomResultBook import make_datgnom_result_book
from molass_legacy.KekLib.ProgressInfo import put_info, put_max_info, put_error
from molass_legacy.SerialAnalyzer.ProgressInfoUtil import extimate_zero_extrapolation, STREAM_ZERO_EX, NUM_SHEET_TYPES
from molass_legacy.SerialAnalyzer.LinearityScore import LinearityScore
from molass_legacy.SerialAnalyzer.DevSettings import get_dev_setting
from molass_legacy.Reports.ReportUtils import make_zero_extrapolation_book, make_zero_extrapolation_overlay_book
from molass_legacy.Test.TesterLogger import write_to_tester_log
from molass_legacy.GuinierAnalyzer.AutorgKekAdapter import AutorgKekAdapter
from molass_legacy.Reports.DefaultFont import set_default_font
set_default_font()

USE_SIMPLIFIED_GUINIER_POROD    = False
TEST_LOG_AQ_GAP     = False
LARGE_FLOAT_VALUE   = 0     # this is temporary. fix this case

def control_extrapolation( self ):
    from molass_legacy.SysArgs import sys_args
    if sys_args is not None and sys_args.devel:
        from molass.Backward.ConcTracker import DecompositionProxy, ConcTracker
        c_vector = self.c_vector
        xr_curve = self.serial_data.get_xr_curve()

        adjusted_conc_factor = 1     # concentration factor is already applied to c_vector
        datatype = get_setting('concentration_datatype')
        self.conc_tracker = ConcTracker(DecompositionProxy(c_vector, xr_curve), adjusted_conc_factor, datatype)
    else:
        self.conc_tracker = None
    if len(self.applied_ranges) > 0:
        prepare_extrapolation( self )
        # log_peak_info( self )
        try:
            do_extrapolation(self)
            clean_tempfolders(self)
        except Exception as exc:
            etb = ExceptionTracebacker()
            self.logger.error( etb )
            put_error( self.stream )
    else:
        self.logger.warning( 'No range for zero concentration extrapolation was found.' )

    if self.conc_tracker is not None:
        savepath = os.path.join(self.work_folder, 'tracked_concentrations.png')
        self.conc_tracker.plot(savepath=savepath)

def log_peak_info( self ):
    for m, range_ in enumerate( self.applied_ranges ):
        write_to_tester_log( 'range_info=' + str( ( m,  range_) ) + '\n' )
        middle = range_.get_fromto_list()[0][1]
        """
            this middle is not appropriate for the new style PairedRange's
            when the ascending side and the descending side are made apart.
        """
        result = self.guinier_result_array[middle][1]
        # TODO: result.fit.Rg is not appropriate when restored
        write_to_tester_log( 'peak_rg_info=' + str( ( m, result.Rg, result.fit.Rg ) ) + '\n' )

def prepare_extrapolation( self ):

    self.to_extrapolate_intensity_array = []
    self.rg_vector_array = []
    self.iz_vector_array = []
    self.indeces_array = []

    to_extrapolate = {}

    for m, range_ in enumerate( self.applied_ranges ):
        self.stop_check()

        from_to_list = range_.get_fromto_list()

        indeces_pair = []

        for j, fromto in enumerate(from_to_list):
            f, t = fromto
            indeces = np.array( np.arange( f, t+1 ), dtype=int )
            # print( 'indeces[%d]=' % m, indeces )
            indeces_pair.append( indeces )
            for i in indeces:
                key_list = to_extrapolate.get(i)
                if key_list is None:
                    to_extrapolate[i] = key_list = []
                key_list.append( (m, j) )
                # key_list contais two tuples only for i=middle in old style

        self.indeces_array.append(indeces_pair)
        self.to_extrapolate_intensity_array.append( [ [], [] ] )
        self.rg_vector_array.append( [ [], [] ] )
        self.iz_vector_array.append( [ [], [] ] )

    # self.logger.info('prepare_extrapolation: to_extrapolate=%s', str(to_extrapolate))
    """
    for MCT-20190522,
    to_extrapolate=
        {0: [(0, 0), (1, 0)],
         1: [(0, 0), (1, 0)],
         2: [(0, 0), (1, 0)],
            ...
        54: [(0, 0), (1, 0)]}
    """

    self.guinier_boundaries_dict = {}
    for i, rec in enumerate( self.guinier_result_array ):
        intensity   = rec[0]
        result_kek  = rec[1]
        result_atsas = rec[2]
        assert result_atsas is not None

        key_list = to_extrapolate.get(i)
        if key_list is None:
            continue

        # key_list contais two tuples only for the peak, and one tuple for others.
        for m, j in key_list:
            self.to_extrapolate_intensity_array[m][j].append( intensity )
            Rg_stdev = 0 if result_kek.Rg_stdev is None else result_kek.Rg_stdev
            if np.isinf( Rg_stdev ):
                Rg_stdev = LARGE_FLOAT_VALUE
            self.rg_vector_array[m][j].append( [ result_kek.Rg, Rg_stdev, result_atsas.Rg, result_atsas.Rg_stdev ] )
            I0_stdev = 0 if result_kek.I0_stdev is None else result_kek.I0_stdev
            if np.isinf( I0_stdev ):
                I0_stdev = LARGE_FLOAT_VALUE
            self.iz_vector_array[m][j].append( [ result_kek.I0, I0_stdev, result_atsas.I0, result_atsas.I0_stdev ] )

            guinier_boundary = self.guinier_boundaries_dict.get( (m, j) )
            if guinier_boundary is None:
                self.guinier_boundaries_dict[(m,j)] = [ result_kek.To ]
            else:
                if result_kek.To is None:
                    pass
                else:
                    if guinier_boundary[0] is None or result_kek.To > guinier_boundary[0]:
                        guinier_boundary[0] = result_kek.To

    # self.logger.info('prepare_extrapolation: guinier_boundaries_dict=%s', str(self.guinier_boundaries_dict))

    max_dict = extimate_zero_extrapolation( self.applied_ranges )
    # self.logger.info( 'max_dict=' + str(max_dict) )
    put_max_info( STREAM_ZERO_EX, max_dict )

def do_extrapolation(self, debug=False):
    if debug:
        self.logger.info("Performing do_extrapolation...")
    qvector = self.qvector      # self.qvector has been set at the start of run_gunier_analysis
    bq_sum_start = bisect_right( qvector, 0.2 )

    zx = ZeroExtrapolator(qvector, self.preview_params, self.serial_data, self.mapped_info, self.applied_ranges, self.conc_tracker, self.known_info_list)

    if USE_SIMPLIFIED_GUINIER_POROD:
        q_slice = slice( 30, len(qvector)//4 )
        x_ = qvector[q_slice]

    if self.using_averaged_files:
        self.used_datafiles = self.averaged_datafiles
    else:
        self.used_datafiles = self.serial_data.datafiles

    if False:
        # self.logger.info("self.used_datafiles[0:5]=%s", str(self.used_datafiles[0:5]))
        print("self.used_datafiles[0:5]=%s" % str(self.used_datafiles[0:5]))

    tester_zx_save  = get_dev_setting( 'tester_zx_save' )
    if tester_zx_save == 1 or self.maintenance_mode:
        self.zx_out_folder  = self.work_folder + '/extrapolated'
        zx_save = True
    else:
        self.zx_out_folder  = None
        zx_save = False

    if zx_save and not os.path.exists( self.zx_out_folder ):
        mkdirs_with_retry( self.zx_out_folder )

    if self.atsas_is_available:
        self.almerge = AlmergeExecutor()
        self.datgnom = DatgnomExecutor()
        if self.excel_is_available:
            self.result_wb_datgnom = None
        else:
            self.result_wb_datgnom = Workbook()
            self.result_wb_datgnom.remove_sheet(self.result_wb_datgnom.active)
    else:
        self.almerge = None
        self.datgnom = None

    self.logger.info( 'extrapolating for ' + str(self.report_ranges) )
    self.peak_num_ranges_list = []

    ignore_all_bqs = get_setting('ignore_all_bqs')
    ignore_bq_list = get_setting('ignore_bq_list')

    self.guinier_results = []
    range_no = 0
    for m, range_ in enumerate( self.applied_ranges ):
        data_array = []
        boundary_indeces = []
        param_array_list = []
        fromto_list = range_.get_fromto_list()
        peak_num_ranges = len(fromto_list)
        self.peak_num_ranges_list.append(peak_num_ranges)
        ignore_bq_for_overlay_book = []
        for j, fromto in enumerate(fromto_list):
            intensities = self.to_extrapolate_intensity_array[m][j]
            log_message = 'extrapolation start for the %s peak ad(%d) with range %s' % ( ordinal_str(m), j, str(fromto) )
            self.logger.info( log_message )
            write_to_tester_log( log_message + '\n' )
            self.stop_check()

            if self.doing_sec:
                if peak_num_ranges == 1 and self.range_type != 5:
                    ad = 'both-side'
                else:
                    ad = 'asc-side' if j == 0 else 'desc-side'
            else:
                ad = 'nat-comp' if m == 0 else 'unf-comp'

            if zx_save:
                paren_no = '' if len(self.applied_ranges ) == 1 else '(%d)' % (m + 1)
                zx_out_sub_folder = self.zx_out_folder + '/' + ad + paren_no
                if not os.path.exists( zx_out_sub_folder ):
                    mkdirs_with_retry( zx_out_sub_folder )

            try:
                indeces = self.indeces_array[m][j]
            except:
                # TOODO: better control
                continue

            c_vector = self.c_vector[indeces]
            # print( 'c_vector=', c_vector )

            try:
                guinier_boundary = self.guinier_boundaries_dict.get( (m, j) )[0]
            except Exception as exc:
                # as in 20161119/Kosugi3a_Backsub
                self.logger.warning( 'failed to get guinier_boundary for (m, j)=' + str( (m, j) ) )
                for k, v in sorted( self.guinier_boundaries_dict.items() ):
                    self.logger.info( str( (k, v) ) ) 
                raise exc

            self.logger.info( 'Guinier boundary is assumed at Q[%d]=%.3g.' % ( guinier_boundary, qvector[guinier_boundary] ) )

            try:
                cx_vector, min_c, max_c = zx.make_cx_vector( j, range_, indeces, c_vector )
                zx_result, lrf_info = zx.extrapolate( m, j, range_no, intensities, c_vector,
                                                        cx_vector, max_c,
                                                        guinier_boundary=guinier_boundary,
                                                        temp_folder=self.temp_folder )

                data_matrix_c, ze_array, param_array, ab_error_array, ze_error_array, boundary_j = zx_result
            except Exception as exc:
                etb = ExceptionTracebacker()
                self.logger.error( str(etb) )
                raise exc

            data_array.append( [ data_matrix_c, ze_array ] )
            boundary_indeces.append( boundary_j )
            param_array_list.append( param_array )

            if debug:
                self.logger.info("put_info( ( %d, %d ), %d )", STREAM_ZERO_EX, m*NUM_SHEET_TYPES, 1)
            put_info( ( STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j ), 1 )
            self.logger.info( ordinal_str(m+1) + ' peak ' + ad + ' extrapolation done.' )

            izx_vector = []
            rgx_vector = []
            rg_results = []
            for k in range( ze_array.shape[1] ):
                self.stop_check()
                c = cx_vector[k]

                if USE_SIMPLIFIED_GUINIER_POROD:
                    y_ = ze_array[q_slice, k]
                    gp = GuinierPorodAnalyzer( x_, y_ )
                    G, G_stdev, Rg, Rg_stdev = gp.fit()
                else:
                    ze_intensity = np.vstack( [ qvector, ze_array[:, k], ze_error_array[:, k] ] ).T
                    try:
                        if self.use_simpleguinier == 0:
                            autorg_kek = AutorgKek( ze_intensity )
                        else:
                            autorg_kek = AutorgKekAdapter( ze_intensity )
                        result = autorg_kek.run()
                        G, G_stdev, Rg, Rg_stdev = result.I0, result.I0_stdev, result.Rg, result.Rg_stdev
                    except Exception as exc:
                        etb = ExceptionTracebacker()
                        self.logger.error( etb )
                        G, G_stdev, Rg, Rg_stdev = None, 0, None, 0
                        result = None

                    if zx_save:
                        zx_filename = make_zx_filename( self, j, indeces, k, c )
                        np_savetxt( zx_out_sub_folder + '/' +zx_filename, ze_intensity )

                rg_results.append(result)

                if G_stdev is None:
                    G_stdev = 0
                if np.isinf( G_stdev ):
                    G_stdev = LARGE_FLOAT_VALUE

                if Rg_stdev is None:
                    Rg_stdev = 0
                if np.isinf( Rg_stdev ):
                    Rg_stdev = LARGE_FLOAT_VALUE

                izx_vector.append( [ G, G_stdev ] )
                rgx_vector.append( [ Rg, Rg_stdev ] )
                if debug:
                    self.logger.info("put_info( ( %d, %d ), %d )", STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j, 2 + k)
                put_info( ( STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j ), 2 + k )

            if self.log_memory_usage == 1:
                m_info = self.process.memory_info()
                memory_usage    = " " + str( ( m_info.vms, ) )
            else:
                memory_usage    = ""

            # self.logger.info( str([k]) + 'rgx_vector=' + str(rgx_vector) )

            zx_index = 0 if j == 0 else -1    # at zero-conc
            self.guinier_results.append(rg_results[zx_index])
            zx_rg_z, zx_rg_z_error = rgx_vector[ zx_index ]
            zx_iz_z, zx_iz_z_error = izx_vector[ zx_index ]
            zx_index = -1 if j == 0 else 0    # at peak
            zx_rg_p = rgx_vector[ zx_index ][0]
            zx_conc_p = cx_vector[ zx_index ]

            rg_vector = self.rg_vector_array[m][j]
            iz_vector = self.iz_vector_array[m][j]
            c_rg_iz = [ [ c_vector,  cx_vector ],
                        [ np.array(rg_vector), np.array(rgx_vector) ],
                        [ np.array(iz_vector), np.array(izx_vector) ],
                      ]

            if self.atsas_is_available:
                make_atsas_folder( self )
                self.atsas_exz_file = make_filename( self, m, j, peak_num_ranges, '/atsas', '.dat' )
                try:
                    # self.logger.info( '%d-%d indeces=' % (m,j) + str(indeces) )
                    almerge_result = self.almerge.execute( self.c_vector, self.used_datafiles, indeces, self.atsas_exz_file )
                except:
                    etb = ExceptionTracebacker()
                    etb.log()
                    almerge_result = None
                if almerge_result is None:
                    result_atsas, result_atsas_eval = ErrorResult(), ErrorResult()
                    overlap_from_max = None
                else:
                    overlap_from_max_ = almerge_result.overlap_from_max
                    if overlap_from_max_ >= len(qvector):
                        overlap_from_max_ = len(qvector) - 1
                        write_to_tester_log( 'overlap_from_max_ was changed to %d\n' % ( overlap_from_max_ ) )
                    overlap_from_max = qvector[overlap_from_max_]

                    result_atsas, result_atsas_eval = autorg_atsas( self.atsas_exz_file )

                if result_atsas is not None:
                    self.logger.info( ordinal_str(m+1) + ' peak ' + ad
                                        + ' atsas result: Rg=' + str(result_atsas.Rg)
                                        + ' I(0)=' + str(result_atsas.I0) )

                write_to_tester_log( simply_safe_sprintf( 'compare atsas Rg: m=%d j=%d Rg=%.17g Rg=%.17g Rg_stdev=%.17g Rg_stdev=%.17g\n',
                                                            m, j, zx_rg_z, result_atsas.Rg, zx_rg_z_error, result_atsas.Rg_stdev ) )
                write_to_tester_log( simply_safe_sprintf( 'compare atsas Iz/c: m=%d j=%d Iz/c=%.17g Iz/c=%.17g Iz/c_stdev=%.17g Iz/c_stdev=%.17g\n',
                                                            m, j, zx_iz_z, divide( result_atsas.I0, zx_conc_p, None), zx_iz_z_error, divide( result_atsas.I0_stdev, zx_conc_p, 0 ) ) )
            else:
                almerge_result  = None
                result_atsas    = None
                overlap_from_max= None

            try:
                write_to_tester_log( simply_safe_sprintf( 'compare zero-peak: m=%d j=%d Rg=%g Rg=%g\n', m, j, zx_rg_z, zx_rg_p ) )
                bq_max_q = np.nan if boundary_j is None else qvector[boundary_j]
                write_to_tester_log( simply_safe_sprintf( 'bq_max=%g bq_boundary=%g\n', zx.bq_max, bq_max_q ) )

                zx_rg_str = simply_safe_sprintf( '%.3g ', zx_rg_z )
                self.logger.info( ordinal_str(m+1) + ' peak ' + ad + ' rg re-estimation done with extrapolated Rg=' + zx_rg_str + memory_usage )

                # Rg-extrapolation linearity
                lin_score = LinearityScore( cx_vector, np.array(rgx_vector)[:,0] )
                rg_slope, rg_r_value, rg_std_err = lin_score.get_params()

                # Iz-extrapolation linearity
                lin_score = LinearityScore( cx_vector, np.array(izx_vector)[:,0] )
                iz_slope, iz_r_value, iz_std_err = lin_score.get_params()

                if zx_rg_z is None or zx_iz_z is None:
                    write_to_tester_log( 'WARNING: zx_rg_z=' + str(zx_rg_z) + ', zx_iz_z=' + str(zx_iz_z) + '\n' )
                else:
                    write_to_tester_log( simply_safe_sprintf( 'extrapolation linearity: m=%d j=%d  rg_slope=%g rg_r_value=%g rg_std_err=%g iz_slope=%g iz_r_value=%g iz_std_err=%g min_c=%g max_c=%g overlap_from_max=%s\n',
                                                                m, j, rg_slope, rg_r_value, rg_std_err/zx_rg_z, iz_slope, iz_r_value, iz_std_err/zx_iz_z, min_c, max_c, str(overlap_from_max) ) )

                if TEST_LOG_AQ_GAP:
                    if boundary_j is None:
                        aq_gap = np.nan
                    else:
                        xb = qvector[boundary_j]
                        yb = []
                        for slice_ in [ slice(boundary_j-60, boundary_j), slice(boundary_j,boundary_j+60) ]:
                            x = qvector[slice_]
                            y = param_array[slice_, 0]
                            slope, intercept, _, _, _ = stats.linregress( x, y )
                            yb.append( slope*xb + intercept )
                        aq_gap = yb[1] - yb[0]

                    bq_sum = np.sum( param_array[bq_sum_start:, 1] )
                    write_to_tester_log( 'bq_sum=%g aq_gap=%g\n' % ( bq_sum, aq_gap ) )

                # TODO: pH7 error

                fromto_list = range_.get_fromto_list()
                from_, to_ = fromto_list[j]
                self.zx_summary_list.append( [ m, j, from_, to_, to_ - from_ + 1, min_c, max_c,
                                                zx_rg_z, zx_rg_z_error, zx_iz_z, zx_iz_z_error,
                                                rg_r_value, rg_std_err, iz_r_value, iz_std_err,
                                                ] )

            except Exception as exc:
                etb = ExceptionTracebacker()
                self.logger.error( etb )

            # output_smoothing has been deprecated

            param_array_out = param_array
            orig_param_array = None

            num_zx_points   = ze_array.shape[1]
            mjn = ( m, j, len(self.applied_ranges), peak_num_ranges )

            if ignore_bq_list is None:
                ignore_bq = ignore_all_bqs
            else:
                ignore_bq = ignore_bq_list[range_no]
            ignore_bq_for_overlay_book.append(ignore_bq)

            try:
                if self.excel_is_available:
                    wb = Workbook()
                    ws = wb.active
                else:
                    wb = self.result_wb
                    ws = wb.create_sheet('Extrapolation')

                book_file = self.temp_folder + '/--zero_extrapolation-temp-%d-%d.xlsx' % ( m, j )
                make_zero_extrapolation_book( self.excel_is_available, wb, ws, book_file, mjn,
                                                indeces, c_rg_iz, self.serial_data.qvector,
                                                data_matrix_c, self.xr_j0,
                                                ze_array, param_array_out, orig_param_array,
                                                guinier_boundary,
                                                almerge_result, result_atsas,
                                                boundary_j, parent=self, lrf_info=lrf_info,
                                                logger=self.logger )
                if debug:
                    self.logger.info("put_info( ( %d, %d ), %d )", STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j, 2 + num_zx_points)
                put_info( ( STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j ), 2 + num_zx_points )
                self.temp_books.append( book_file )

                if self.log_memory_usage == 1:
                    m_info = self.process.memory_info()
                    memory_usage    = " " + str( ( m_info.vms, ) )
                else:
                    memory_usage    = ""
                self.logger.info( ordinal_str(m+1) + ' peak ' + ad + ' book done.' + memory_usage )
            except Exception as exc:
                etb = ExceptionTracebacker()
                self.logger.error( etb )

            output_extrapolation_params( self, m, j, peak_num_ranges, qvector, param_array_out, ab_error_array, indeces, ignore_bq)
            range_no += 1

            if self.atsas_is_available:
                try:
                    datgnom_out_file = make_filename( self, m, j, peak_num_ranges, '/atsas', '.out' )
                    datgnom_result = self.datgnom.execute( self.zx_a_file_path, zx_rg_z, datgnom_out_file )
                    self.zx_summary_list2.append( [ datgnom_result.RgPr, datgnom_result.RgPr_Error, datgnom_result.IzPr, datgnom_result.IzPr_Error, datgnom_result.Dmax, *datgnom_result.Qrange ] )
                    write_to_tester_log( 'datgnom_result=' + str( self.zx_summary_list2[-1] ) + '\n' )
                    book_file = self.temp_folder + '/--zero_extrapolation-atsas-%d-%d.xlsx' % ( m, j )
                    try:
                        if self.excel_is_available:
                            wb = Workbook()
                            ws = wb.active
                        else:
                            wb = self.result_wb_datgnom
                            ws = wb.create_sheet('DATGNOM')

                        make_datgnom_result_book( self.excel_is_available, wb, ws,
                                        mjn, self.range_type, datgnom_out_file, book_file, self)
                        self.temp_books_atsas.append( book_file )
                        self.logger.info( ordinal_str(m+1) + ' peak ' + ad + ' atsas book done.' )
                    except Exception as exc:
                        self.logger.error( 'make_datgnom_result_book failed with ' + str(exc) )
                except Exception as exc:
                    log_exception(self.logger, "datgnom failure: ")
                    self.zx_summary_list2.append( [None] * 5 )
                if debug:
                    self.logger.info("put_info( ( %d, %d ), %d )", STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j, 3 + num_zx_points)
                put_info( ( STREAM_ZERO_EX, m*NUM_SHEET_TYPES+j ), 3 + num_zx_points )
            else:
                self.zx_summary_list2.append( [None] * 5 )

        if peak_num_ranges == 2:
            ignore_bq = np.sum(ignore_bq_for_overlay_book) == 2
            try:
                if self.excel_is_available:
                    wb = Workbook()
                    ws = wb.active
                else:
                    wb = self.result_wb
                    ws = wb.create_sheet('Extrapolation')

                book_file = self.temp_folder + '/--zero_extrapolation-temp-%d-%d.xlsx' % ( m, 2 )
                self.temp_books.append( book_file )
                make_zero_extrapolation_overlay_book( self.excel_is_available, wb, ws,
                                                        book_file, ( m, len(self.applied_ranges) ),
                                                        self.serial_data.qvector, param_array_list, boundary_indeces,
                                                        parent=self,
                                                        lrf_info=lrf_info,
                                                        logger=self.logger )
                if debug:
                    self.logger.info("put_info( ( %d, %d ), %d )", STREAM_ZERO_EX, m*NUM_SHEET_TYPES+2, 1)
                put_info( ( STREAM_ZERO_EX, m*NUM_SHEET_TYPES+2 ), 1 )

                if self.log_memory_usage == 1:
                    m_info = self.process.memory_info()
                    memory_usage    = " " + str( ( m_info.vms, ) )
                else:
                    memory_usage    = ""
                self.logger.info( ordinal_str(m+1) + ' peak overlay book done.' + memory_usage )
            except Exception as exc:
                etb = ExceptionTracebacker()
                self.logger.error( etb )

def make_atsas_folder( self ):
    self.atsas_folder = self.work_folder + '/atsas'
    if not os.path.exists( self.atsas_folder ):
        mkdirs_with_retry( self.atsas_folder )

def make_filename( self, m, ud, peak_num_ranges, sub_folder, tail ):
    if self.doing_sec:
        if peak_num_ranges == 1:
            comp = '_bth'
        else:
            comp = '_asc' if ud == 0 else '_dsc'
    else:
        comp = '_nat' if m == 0 else '_unf'
    filename = self.work_folder + sub_folder + '/pk%d' % (m+1) + comp + tail
    return filename

def make_zx_filename( self, j, indeces, k, c ):
    conc_str = ( '%.4g' % c ).replace( '.', '' )[-4:]
    if j == 0:
        if k < NUM_EXTRAPOLATION_POINTS:
            fno = None
        else:
            fno = indeces[k-NUM_EXTRAPOLATION_POINTS]
    else:
        if k < len( indeces ):
            fno = indeces[k]
        else:
            fno = None
    fno_str = '____' if fno is None else '%05d' % fno
    return 'extrapolated-%03d-%s-%s.dat' % ( k, fno_str, conc_str )

def output_extrapolation_params( self, m, ud, peak_num_ranges, qvector, param_array, ab_error_array, indeces, ignore_bq ):
    self.zx_a_file_path = make_filename( self, m, ud, peak_num_ranges, '', '_A_cn.dat' )
    self.zx_b_file_path = make_filename( self, m, ud, peak_num_ranges, '', '_B_cn.dat',  )

    # qvector_ = np.atleast_2d( qvector ).T
    filepaths = [ self.zx_a_file_path, self.zx_b_file_path ]

    num_files = 1 if ignore_bq else 2
    for j in range(num_files):
        values = param_array[:,j]
        if j > 0:
            if np.sum(values) == 0:     # task: need a more effective way to judge this
                continue

        fh = open( filepaths[j], "wb" )
        fh.write( str.encode( "# Extrapolated using the following files with []-numbering starting from 1.\n" ) )
        for k in indeces:
            fh.write( str.encode( "# [%d] %s\n" % ( k+1, self.used_datafiles[k] )  ) )
        fh.write( str.encode( "#\n# Q\tIntensity\tError\n" ) )
        fh.close()
        array = np.vstack( [ qvector, values, ab_error_array[:,j] ] ).T
        np_savetxt( filepaths[j], array, mode="a" )

def clean_tempfolders(self):
    averaged_data_folder = get_setting('averaged_data_folder')  # this should have set in AnalyzerDialogProxy
    if os.path.exists(averaged_data_folder):
        keep_tempfolder_averaged = get_setting("keep_tempfolder_averaged")
        if not keep_tempfolder_averaged:
            import shutil
            try:
                # assert False
                shutil.rmtree(averaged_data_folder)
                self.logger.info("%s has been removed.", averaged_data_folder)
            except:
                log_exception(self.logger, "averaged_data_folder cleanup failed: ")
