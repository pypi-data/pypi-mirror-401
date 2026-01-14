import logging
from typing import Any, TypedDict

import polars as pl
from openpyxl import load_workbook
from python_calamine import CalamineWorkbook, SheetVisibleEnum

log = logging.getLogger(__name__)


class FindHeaderRowOptions(TypedDict):
    max_rows: int
    expected_keywords: list[str] | None


def get_sheet_names(
    source: Any,
    *,
    visible_only: bool = False,
) -> list[str]:
    """
    Get sheet names from an excel file

    Parameters
    ----------
    source : Any
        Path or file-like object to read
    visible_only : bool, optional
        Whether to only include visible sheets, by default False

    Returns
    -------
    list[str]
        List of sheet names
    """
    if hasattr(source, "seek"):
        source.seek(0)

    sheets_meta = CalamineWorkbook.from_object(source).sheets_metadata

    result: list[str] = []
    for meta in sheets_meta:
        if visible_only and meta.visible != SheetVisibleEnum.Visible:
            continue

        result.append(meta.name)

    return result


# TODO: Add support for reading values only from visible columns
def read_visible_rows(
    source: Any,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    strip_values: bool = True,
    strip_headers: bool = True,
) -> pl.DataFrame:
    """
    Reads only the visible rows from a specified sheet or first sheet using OpenPyXL.

    Parameters
    ----------
    source : Any
        Path or file-like object (excel workbook)
    sheet_name : str | None, optional
        Name of the sheet to read, by default None. If None, reads the first sheet.
    header_row : int | None, optional
        Index of the header row (0-indexed), by default None
    strip_values : bool, optional
        Whether to strip leading and trailing whitespace from values of string type, by default True
    strip_headers : bool, optional
        Whether to strip leading and trailing whitespace from headers, by default True

    Returns
    -------
    pl.DataFrame
        Polars DataFrame containing the visible rows and header_row as the schema

    Raises
    ------
    ValueError
        If sheet_name is not found in the workbook
    ValueError
        header_row index is out of boundsf
    """
    if hasattr(source, "seek"):
        source.seek(0)

    wb = load_workbook(source, data_only=True)
    if sheet_name:
        try:
            ws = wb[sheet_name]
        except KeyError:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    else:
        ws = wb[wb.sheetnames[0]]

    visible_rows = []
    for row in ws.iter_rows():
        row_idx = row[0].row
        is_row_hidden = ws.row_dimensions[row_idx].hidden
        if is_row_hidden:
            continue

        # Get row values
        row_values = []
        for cell in row:
            val = cell.value
            if isinstance(cell.value, str) and strip_values:
                val = val.strip() or None
            row_values.append(cell.value)

        visible_rows.append(row_values)

    wb.close()

    if not visible_rows:
        return pl.DataFrame()

    idx = header_row if header_row is not None else 0
    if idx >= len(visible_rows):
        raise ValueError(f"header_row index: '{header_row}' out of bounds")

    headers = visible_rows[idx]
    data = visible_rows[idx + 1 :]

    # Handle default column names
    cleaned_headers = []
    for i, header in enumerate(headers):
        if header is None:
            header = f"column_{i}"

        header = str(header).strip() if header and strip_headers else str(header)
        cleaned_headers.append(header)

    # Handle duplicate headers
    seen = {}
    for i, header in enumerate(cleaned_headers):
        if header in seen:
            seen[header] += 1
            cleaned_headers[i] = f"{header}_{seen[header]}"
        else:
            seen[header] = 0

    # Remove empty rows
    cleaned_data = []
    for row in data:
        if not all(
            cell is None or (isinstance(cell, str) and cell.strip() == "")
            for cell in row
        ):
            cleaned_data.append(row)

    if not cleaned_data:
        return pl.DataFrame(schema=cleaned_headers, strict=False, orient="row")

    return pl.DataFrame(
        cleaned_data,
        schema=cleaned_headers,
        strict=False,
        orient="row",
        infer_schema_length=len(cleaned_data),
    )


def locate_header_row(
    source: Any,
    *,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
    max_rows: int = 200,
    expected_keywords: list[str] | None = None,
) -> int:
    """
    Finds the header row in an excel file by identifying the first row with maximum consecutive non-null values.

    Parameters
    ----------
    source : Any
        Path or file-like object to read
    sheet_id : int | None, optional
        0-based index of the sheet to read, by default None (Cannot be used with sheet_name)
    sheet_name : str | None, optional
        Name of the worksheet to read, by default None (Cannot be used with sheet_id)
    max_rows : int, optional
        Maximum number of rows to scan for header identification, by default 200
    expected_keywords : list[str] | None, optional
        List of keywords to look for in the header row. If a row contains all of these keywords, it is considered a header row, by default None

    Returns
    -------
    int
        Zero-based index of the first row with maximum consecutive non-null values.

        If expected_keywords is provided, this is the first row with all expected keywords and maximum consecutive non-null values.

    Raises
    ------
    ValueError
        If both sheet_id and sheet_name are provided.

    Notes
    -----
    - If first few rows are empty and the first non-empty row is the header, then simply use read_excel_sheet() or pl.read_excel() directly instead of finding the header row.

    - This function would be best suited when you want to read a filtered excel sheet using read_excel_sheet(visible_rows_only=True, header_row=header_row). OR If the header row is not the first non-empty row.
    """
    if sheet_id and sheet_name:
        raise ValueError("sheet_id and sheet_name cannot be both specified.")

    if hasattr(source, "seek"):
        source.seek(0)

    wb = CalamineWorkbook.from_object(source)
    if sheet_id:
        ws = wb.get_sheet_by_index(0)
    elif sheet_name:
        ws = wb.get_sheet_by_name(sheet_name)
    else:
        ws = wb.get_sheet_by_index(0)

    max_consecutive = 0
    header_row = 0
    for i, row in enumerate(ws.iter_rows()):
        if i > max_rows:
            log.debug("Reached max_rows limit")
            break

        consecutive_count = 0
        is_all_keywords_present = False

        if expected_keywords:
            # If expected keywords are provided, check if all of them are present in the row
            row_values = [str(value).strip().lower() for value in row if value]

            is_all_keywords_present = all(
                keyword.lower() in row_values for keyword in expected_keywords
            )

        # Check for non-null consecutive values
        for value in row:
            if value:
                consecutive_count += 1
            else:
                break

        if consecutive_count > max_consecutive:
            max_consecutive = consecutive_count
            header_row = i
            if expected_keywords and is_all_keywords_present:
                # This is the first row with all expected keywords, and highest consecutive non-null count, so its most likely the header row
                log.info(
                    f"Found first header row at index: '{i}' with all expected keywords and maximum consecutive non-null values"
                )
                break

    log.info(
        f"Identified header row at index: {header_row} with {max_consecutive} consecutive non-null values"
    )
    return header_row
